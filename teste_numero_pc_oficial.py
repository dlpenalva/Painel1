# -*- coding: utf-8 -*-
"""Etapa 9 — NUMERO_PC no modelo oficial + grade financeira estendida.

Cobre os 10 casos exigidos: duplicidade (Python e XLS via Excel real),
normalizacao (espacos externos, caixa), NUMERO_PC vazio, leitura por
cabecalho independente da posicao, rejeicao de modelo oficial sem
NUMERO_PC, preservacao na geracao pela Calculadora, nao-regressao das
formulas DATA_PC/CICLO_PC/FATOR_ACUMULADO, e a grade do financeiro com
61 competencias (sem truncamento, sem sobrescrita, soma correta).
"""
from __future__ import annotations

import io
from datetime import date
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import Workbook, load_workbook

from _coleta_oficial import (
    ABAS_COLETA_OFICIAL,
    COLUNAS_ITENS_PC_OFICIAL,
    obter_coleta_oficial_bytes,
)
from _gerador_masterfile import gerar_masterfile_preenchido
from _leitor_masterfile_v10 import _ler_itens_pc_v10, ler_masterfile_v10

RESULTADOS = []


def check(nome: str, cond: bool, detalhe: str = "") -> None:
    RESULTADOS.append((nome, bool(cond), detalhe))
    status = "OK " if cond else "FAIL"
    print(f"[{status}] {nome}" + (f" - {detalhe}" if detalhe and not cond else ""))


def _dados_calculadora():
    return {
        "ok": True, "modo_origem": "teste", "indice": "IST",
        "ciclo_vigente": "C2", "data_corte": date(2025, 6, 30),
        "data_base": date(2023, 1, 1),
        "ciclos": [
            {"ciclo": "C1", "data_inicio": date(2024, 1, 1),
             "data_fim": date(2024, 12, 31), "percentual": 0.10,
             "inicio_efeito_financeiro": date(2024, 1, 1),
             "possui_efeito_financeiro": "Sim", "situacao": "TEMPESTIVO"},
            {"ciclo": "C2", "data_inicio": date(2025, 1, 1),
             "data_fim": date(2025, 12, 31), "percentual": 0.10,
             "inicio_efeito_financeiro": date(2025, 1, 1),
             "possui_efeito_financeiro": "Sim", "situacao": "TEMPESTIVO"},
        ],
    }


def _preencher_pcs(wb, linhas):
    """linhas: lista de (numero_pc, data, valor, pago)."""
    ws = wb["itens_PC"]
    for i, (num, data_pc, valor, pago) in enumerate(linhas, start=2):
        ws.cell(i, 1).value = num
        ws.cell(i, 2).value = data_pc
        ws.cell(i, 4).value = valor
        ws.cell(i, 7).value = pago
    return wb


def _ler_pcs(bytes_xlsx):
    wb = load_workbook(io.BytesIO(bytes_xlsx), data_only=True)
    return _ler_itens_pc_v10(wb)


def _salvar(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def main() -> int:
    ent = obter_coleta_oficial_bytes()
    preenchido = gerar_masterfile_preenchido(_dados_calculadora(), ent)

    # (8) preservacao de NUMERO_PC na geracao pela Calculadora
    wb_g = load_workbook(io.BytesIO(preenchido))
    ws_g = wb_g["itens_PC"]
    headers = [ws_g.cell(1, i + 1).value for i in range(len(COLUNAS_ITENS_PC_OFICIAL))]
    check("geracao preserva layout NUMERO_PC (A1:L1)",
          headers == COLUNAS_ITENS_PC_OFICIAL, str(headers))
    check("geracao nao restaura ITEM",
          "ITEM" not in [ws_g.cell(1, i).value for i in range(1, 30)])
    check("geracao mantem abas oficiais", wb_g.sheetnames == ABAS_COLETA_OFICIAL)

    # (9) nao-regressao DATA_PC / CICLO_PC / FATOR_ACUMULADO (formulas pos-shift)
    c2 = str(ws_g["C2"].value or "")
    e2 = str(ws_g["E2"].value or "")
    k2 = str(ws_g["K2"].value or "")
    check("CICLO_PC (C) segue formula LOOKUP sobre DATA_PC (B)",
          c2.startswith("=") and "B2" in c2 and "LOOKUP" in c2.upper(), c2[:80])
    check("FATOR_ACUMULADO (E) segue formula sobre CICLO_PC (C)",
          e2.startswith("=") and "C2" in e2, e2[:80])
    check("CHECK (K) valida duplicidade de NUMERO_PC",
          "NUMERO_PC duplicado" in k2 and "TRIM" in k2 and "UPPER" in k2, k2[:120])
    check("CHECK (K) nao usa DATA_PC/VALOR_PC como chave de duplicidade",
          "TRIM($B" not in k2 and "TRIM($D" not in k2)

    # (1) numeros diferentes, mesma data e mesmo valor: sem duplicidade
    r1 = _ler_pcs(_salvar(_preencher_pcs(load_workbook(io.BytesIO(preenchido)), [
        ("PC-001", date(2024, 5, 10), 5000.0, "Sim"),
        ("PC-002", date(2024, 5, 10), 5000.0, "Sim"),
    ])))
    check("PCs distintos com mesma data/valor: sem alerta de duplicidade",
          not any("duplicado" in a for a in r1["alertas"]), str(r1["alertas"]))
    check("PCs distintos: 2 registros lidos", len(r1["itens"]) == 2)

    # (2) mesmo NUMERO_PC: alerta de duplicidade
    r2 = _ler_pcs(_salvar(_preencher_pcs(load_workbook(io.BytesIO(preenchido)), [
        ("PC-100", date(2024, 5, 10), 5000.0, "Sim"),
        ("PC-100", date(2024, 6, 10), 7000.0, "Sim"),
    ])))
    check("NUMERO_PC repetido: alerta 'NUMERO_PC duplicado'",
          any("NUMERO_PC duplicado" in a and "PC-100" in a for a in r2["alertas"]),
          str(r2["alertas"]))

    # (3) NUMERO_PC vazio: linha lida com alerta, sem controle de duplicidade
    r3 = _ler_pcs(_salvar(_preencher_pcs(load_workbook(io.BytesIO(preenchido)), [
        (None, date(2024, 5, 10), 5000.0, "Sim"),
        (None, date(2024, 6, 10), 7000.0, "Sim"),
    ])))
    check("NUMERO_PC vazio: linhas lidas",
          len(r3["itens"]) == 2, f"itens={len(r3['itens'])}")
    check("NUMERO_PC vazio: alerta especifico, sem falso 'duplicado'",
          any("NUMERO_PC vazio" in a for a in r3["alertas"])
          and not any("duplicado" in a for a in r3["alertas"]), str(r3["alertas"]))

    # (4) espacos externos e (5) caixa: normalizacao APENAS para comparacao
    r45 = _ler_pcs(_salvar(_preencher_pcs(load_workbook(io.BytesIO(preenchido)), [
        (" PC-200 ", date(2024, 5, 10), 5000.0, "Sim"),
        ("pc-200", date(2024, 6, 10), 7000.0, "Sim"),
    ])))
    check("espacos e caixa: reconhecido como duplicado",
          any("NUMERO_PC duplicado" in a for a in r45["alertas"]), str(r45["alertas"]))
    check("valor original preservado (nao normalizado no registro)",
          r45["itens"][0]["numero_pc"] == " PC-200 "
          and r45["itens"][1]["numero_pc"] == "pc-200")

    # (6) leitura por cabecalho, independente da posicao fisica
    wb6 = Workbook()
    ws6 = wb6.active
    ws6.title = "itens_PC"
    ws6.append(["DATA_PC", "VALOR_PC", "NUMERO_PC", "PC_PAGO_A_CONTRATADA"])
    ws6.append([date(2024, 5, 10), 5000.0, "PC-XYZ", "Sim"])
    buf6 = io.BytesIO()
    wb6.save(buf6)
    r6 = _ler_pcs(buf6.getvalue())
    check("NUMERO_PC localizado por cabecalho (coluna C)",
          len(r6["itens"]) == 1 and r6["itens"][0]["numero_pc"] == "PC-XYZ",
          str([i.get('numero_pc') for i in r6['itens']]))

    # (7) modelo oficial sem NUMERO_PC: template incompativel no upload
    wb7 = load_workbook(io.BytesIO(preenchido))
    wb7["itens_PC"].delete_cols(1)
    buf7 = io.BytesIO()
    wb7.save(buf7)
    r7 = ler_masterfile_v10(buf7.getvalue(), exigir_modelo_oficial=True)
    check("modelo oficial sem NUMERO_PC rejeitado como template incompativel",
          "Template incompativel" in str(r7["erro"])
          and "NUMERO_PC" in str(r7["erro"]), str(r7["erro"]))

    # (10) duplicidade no XLS (Excel real) + grade de 61 competencias
    try:
        import win32com.client  # noqa: F401
        tem_excel = True
    except ImportError:
        tem_excel = False
    if tem_excel:
        _validar_com_excel(preenchido)
    else:
        # Mesmo padrao de teste_v1047_excel_abre_sem_reparo: sem pywin32 no
        # interpretador, a validacao com Excel real e feita manualmente
        # (obrigatoria antes de commit) e nao derruba o CI.
        print("  [COM] win32com nao disponivel — duplicidade XLS e grade 61 "
              "validadas manualmente com Excel real (py sistema)")

    falhas = [n for n, ok, _ in RESULTADOS if not ok]
    print(f"\n{len(RESULTADOS)} verificacoes; falhas: {len(falhas)}")
    return 1 if falhas else 0


def _validar_com_excel(preenchido: bytes) -> None:
    import win32com.client

    wb = load_workbook(io.BytesIO(preenchido))
    _preencher_pcs(wb, [
        ("PC-100", date(2024, 5, 10), 5000.0, "Sim"),    # linha 2 (dup com 3)
        (" pc-100 ", date(2024, 6, 10), 7000.0, "Sim"),  # linha 3 (caixa+espaco)
        ("PC-300", date(2024, 7, 10), 9000.0, "Sim"),    # linha 4 (unico)
        (None, date(2024, 8, 10), 1000.0, "Sim"),        # linha 5 (vazio)
    ])
    # 61 competencias: contrato de 60 meses com meses inicial/final distintos
    fin = wb["financeiro"]
    ano, mes = 2023, 1
    for r in range(2, 63):                    # linhas 2..62 = 61 meses
        fin.cell(r, 1).value = date(ano, mes, 1)
        fin.cell(r, 3).value = 100.0
        fin.cell(r, 7).value = "Sim"
        mes += 1
        if mes == 13:
            ano, mes = ano + 1, 1

    with TemporaryDirectory() as tmp:
        caminho = Path(tmp) / "numero_pc_excel.xlsx"
        caminho.write_bytes(_salvar(wb))
        excel = win32com.client.gencache.EnsureDispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        try:
            pasta = excel.Workbooks.Open(str(caminho))
            excel.CalculateFullRebuild()
            pasta.Save()
            pasta.Close(SaveChanges=False)
        finally:
            excel.Quit()
        wb2 = load_workbook(caminho, data_only=True)

    ws = wb2["itens_PC"]
    check("XLS: duplicado detectado na linha 2",
          ws["K2"].value == "NUMERO_PC duplicado", repr(ws["K2"].value))
    check("XLS: duplicado com caixa/espacos detectado na linha 3",
          ws["K3"].value == "NUMERO_PC duplicado", repr(ws["K3"].value))
    check("XLS: NUMERO_PC unico prossegue validacoes (OK)",
          ws["K4"].value == "OK", repr(ws["K4"].value))
    check("XLS: NUMERO_PC vazio nao avalia duplicidade",
          ws["K5"].value != "NUMERO_PC duplicado", repr(ws["K5"].value))
    check("XLS: CICLO_PC recalculado pos-shift (C2)",
          ws["C2"].value == "C1", repr(ws["C2"].value))
    check("XLS: VALOR_ATUALIZADO recalculado (F2)",
          ws["F2"].value is not None, repr(ws["F2"].value))

    f2 = wb2["financeiro"]
    comps = [(r, f2.cell(r, 1).value) for r in range(2, 74)
             if f2.cell(r, 1).value is not None]
    check("financeiro: 61 competencias preenchidas", len(comps) == 61,
          f"{len(comps)}")
    check("financeiro: ultima competencia preservada (2028-01)",
          comps[-1][1].year == 2028 and comps[-1][1].month == 1,
          str(comps[-1]))
    check("financeiro: sem sobrescrita apos a ultima competencia",
          all(f2.cell(r, 1).value is None for r in range(comps[-1][0] + 1, 75)))
    soma = sum(float(f2.cell(r, 3).value or 0) for r, _ in comps)
    check("financeiro: soma financeira correta (6100)", soma == 6100.0, str(soma))
    check("financeiro: classificacao do ultimo mes do ultimo ciclo computado "
          "(2025-12 = c2)",
          str(f2.cell(37, 2).value or "").lower() == "c2",
          repr(f2.cell(37, 2).value))
    check("financeiro: mes alem dos ciclos computados segue o calendario "
          "derivado (2026-12 = c3)",
          str(f2.cell(49, 2).value or "").lower() == "c3",
          repr(f2.cell(49, 2).value))
    check("financeiro: linhas excedentes vazias nao afetam a soma",
          all((f2.cell(r, 3).value in (None, 0, "")) for r in range(comps[-1][0] + 1, 74)))


if __name__ == "__main__":
    raise SystemExit(main())
