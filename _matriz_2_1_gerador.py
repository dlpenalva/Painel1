
# -*- coding: utf-8 -*-
"""
Gerador experimental da Matriz 2.1 — cl8us/FARC/Streamlit

Versão: matriz21_remanescente_execucao_v2

Objetivo:
- Criar XLSX enxuto, operacional e explícito para coleta de dados críticos.
- Tratar formalmente remanescente em ciclo em execução.
- Criar validações contra dupla contagem.
- Criar aba de conciliação de referência para casos reais, como PADTEC/GMP.
- Não substitui a Matriz 2.0.

Gera:
    C:\_DesktopReal\15.ColetaUnica\ColetaReajuste_Matriz21.xlsx

Uso:
    cd /d C:\_DesktopReal\15.ColetaUnica
    py _matriz_2_1_gerador.py
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule


ROOT = Path(__file__).resolve().parent
OUT = ROOT / "ColetaReajuste_Matriz21.xlsx"

# Cores
AZUL = "1F4E79"
AMARELO = "FFF2CC"      # preenchimento manual
CINZA = "E7E6E6"        # automático/fórmula
VERDE = "D9EAD3"        # OK/total
VERMELHO = "F4CCCC"     # erro
LARANJA = "FCE4D6"      # alerta
ROXO = "EADCF8"         # blocos/premissas
BRANCO = "FFFFFF"

BORDA = Side(style="thin", color="B7B7B7")


def _fill(color):
    return PatternFill("solid", fgColor=color)


def _border():
    return Border(top=BORDA, bottom=BORDA, left=BORDA, right=BORDA)


def _title(ws, text, subtitle=None):
    ws["A1"] = text
    ws["A1"].font = Font(bold=True, size=15, color="17365D")
    ws["A1"].alignment = Alignment(vertical="center", horizontal="left", wrap_text=False)
    ws.row_dimensions[1].height = 24
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = Font(italic=True, color="666666")
        ws["A2"].alignment = Alignment(vertical="center", horizontal="left", wrap_text=False)
        ws.row_dimensions[2].height = 22


def _style_header(ws, row, c1, c2, color=AZUL):
    for col in range(c1, c2 + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _fill(color)
        cell.font = Font(color=BRANCO, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border()


def _input_cell(cell, num_format=None):
    cell.fill = _fill(AMARELO)
    cell.border = _border()
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    if num_format:
        cell.number_format = num_format
    return cell


def _auto_cell(cell, num_format=None):
    cell.fill = _fill(CINZA)
    cell.border = _border()
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    if num_format:
        cell.number_format = num_format
    return cell


def _money(cell):
    cell.number_format = 'R$ #,##0.00'
    return cell


def _pct(cell):
    cell.number_format = '0.00%'
    return cell


def _factor(cell):
    cell.number_format = '0.000000'
    return cell


def _date(cell):
    cell.number_format = 'dd/mm/yyyy'
    return cell


def _set_widths(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _add_list_validation(ws, formula1, ranges):
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
    dv.error = "Escolha uma opção da lista."
    dv.errorTitle = "Valor inválido"
    ws.add_data_validation(dv)
    for rg in ranges:
        dv.add(rg)
    return dv


def _add_hidden_list_sheet(wb):
    ws = wb.create_sheet("_LISTAS")
    ws.sheet_state = "hidden"

    listas = {
        "A": ["Ciclo", "C0", "C1", "C2", "C3", "C4"],
        "B": ["SimNao", "Sim", "Não"],
        "C": ["FormaC0", "Valor consolidado", "Financeiro mensal", "Itens", "Não disponível"],
        "D": ["FormaSaldo", "Itens", "Valor consolidado", "Não disponível"],
        "E": ["MarcoSaldo", "Após última competência financeira", "Início do ciclo em execução", "Data de corte", "Saldo atual sem vínculo claro", "Não disponível"],
        "F": ["RegraSupressao", "Computar atualizada", "Computar nominal", "Já embutida no saldo", "Não computar"],
        "G": ["TipoAditivo", "Acréscimo", "Supressão"],
        "H": ["TratamentoAditivo", "Computar atualizada", "Computar nominal", "Já embutido no saldo", "Informativo / não computar"],
        "I": ["FormaAditivo", "Qtd x VU", "Valor consolidado"],
    }

    for col, values in listas.items():
        for idx, value in enumerate(values, 1):
            ws[f"{col}{idx}"] = value

    return ws


def build():
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    ws_ent = wb.create_sheet("ENTRADAS_CRITICAS")
    ws_base = wb.create_sheet("BASE_CICLOS")
    ws_fin = wb.create_sheet("FINANCEIRO")
    ws_itens = wb.create_sheet("ITENS_SALDO")
    ws_adit = wb.create_sheet("ADITIVOS_SUPRESSOES")
    ws_mem = wb.create_sheet("MEMORIA_VTA")
    ws_conc = wb.create_sheet("CONCILIACAO_REFERENCIA")
    ws_val = wb.create_sheet("VALIDACOES")
    _add_hidden_list_sheet(wb)

    # ENTRADAS_CRITICAS
    _title(
        ws_ent,
        "Matriz 2.1 — Entradas Críticas",
        "Aba principal. Informe premissas essenciais: C0, corte, ciclo em execução, saldo remanescente e tratamento de aditivos/supressões."
    )

    headers = ["Bloco", "Campo", "Valor / Seleção", "Regra operacional", "Observação"]
    for idx, h in enumerate(headers, 1):
        ws_ent.cell(4, idx, h)
    _style_header(ws_ent, 4, 1, 5)

    rows = [
        ("CORTE", "Data início C0 / início contratual", "", "Obrigatório para montar BASE_CICLOS.", "Ex.: 01/10/2022"),
        ("CORTE", "Data de corte da apuração", "", "Obrigatório.", "Ex.: 30/04/2026"),
        ("CORTE", "Há ciclo em execução?", "Sim", "Sim/Não.", "Use Sim quando houver execução parcial no ciclo atual."),
        ("CORTE", "Ciclo em execução", "C3", "Dropdown C0-C4.", "Ex.: C3"),
        ("CORTE", "Última competência financeira informada", "", "Obrigatório se houver financeiro parcial.", "Ex.: 30/04/2026"),
        ("CORTE", "Há saldo remanescente após essa competência?", "Sim", "Sim/Não.", ""),
        ("CORTE", "Data de referência do saldo remanescente", "", "Obrigatório se houver saldo.", "Ex.: 01/05/2026"),
        ("", "", "", "", ""),
        ("C0", "Forma de informar C0", "Valor consolidado", "Valor consolidado / Financeiro mensal / Itens / Não disponível.", ""),
        ("C0", "Valor consolidado do C0", "", "Obrigatório se a forma for Valor consolidado.", "Ex.: R$ 33.792.666,16"),
        ("C0", "Fonte/observação do C0", "", "Campo livre para origem da informação.", "Ex.: Memória GMP / histórico fiscal"),
        ("", "", "", "", ""),
        ("SALDO", "Forma de informar saldo remanescente", "Valor consolidado", "Itens / Valor consolidado / Não disponível.", ""),
        ("SALDO", "Marco do saldo remanescente", "Após última competência financeira", "Define se o saldo pode ser somado ao financeiro parcial sem dupla contagem.", ""),
        ("SALDO", "Ciclo do saldo", "C3", "Dropdown C0-C4.", "Ex.: C3"),
        ("SALDO", "Valor consolidado do saldo remanescente", "", "Preencher se a forma for Valor consolidado.", ""),
        ("SALDO", "Data de referência do saldo", "", "Deve conversar com o corte/última competência.", ""),
        ("SALDO", "Observação do saldo", "", "Campo livre.", "Ex.: saldo a partir de maio/2026"),
        ("", "", "", "", ""),
        ("ADITIVOS", "Há aditivos/supressões nesta análise?", "Sim", "Sim/Não.", ""),
        ("ADITIVOS", "Regra padrão para supressões", "Computar atualizada", "Computar atualizada / Computar nominal / Já embutida no saldo / Não computar.", ""),
        ("ADITIVOS", "Observação geral de aditivos/supressões", "", "Campo livre.", ""),
    ]

    for r, vals in enumerate(rows, 5):
        for c, val in enumerate(vals, 1):
            ws_ent.cell(r, c, val)
            ws_ent.cell(r, c).border = _border()
            ws_ent.cell(r, c).alignment = Alignment(vertical="top", wrap_text=True)
        if vals[0]:
            ws_ent.cell(r, 1).fill = _fill(ROXO)
            ws_ent.cell(r, 1).font = Font(bold=True)
        _input_cell(ws_ent.cell(r, 3))
        if vals[1] in [
            "Data início C0 / início contratual",
            "Data de corte da apuração",
            "Última competência financeira informada",
            "Data de referência do saldo remanescente",
            "Data de referência do saldo",
        ]:
            _date(ws_ent.cell(r, 3))
        if vals[1] in ["Valor consolidado do C0", "Valor consolidado do saldo remanescente"]:
            _money(ws_ent.cell(r, 3))

    _add_list_validation(ws_ent, "'_LISTAS'!$B$2:$B$3", ["C7", "C10", "C24"])
    _add_list_validation(ws_ent, "'_LISTAS'!$A$2:$A$6", ["C8", "C19"])
    _add_list_validation(ws_ent, "'_LISTAS'!$C$2:$C$5", ["C13"])
    _add_list_validation(ws_ent, "'_LISTAS'!$D$2:$D$4", ["C17"])
    _add_list_validation(ws_ent, "'_LISTAS'!$E$2:$E$6", ["C18"])
    _add_list_validation(ws_ent, "'_LISTAS'!$F$2:$F$5", ["C25"])
    _set_widths(ws_ent, {"A": 16, "B": 46, "C": 32, "D": 58, "E": 52})
    ws_ent.freeze_panes = "A5"

    # BASE_CICLOS
    _title(ws_base, "Base dos Ciclos", "Tabela-verdade automática dos ciclos, percentuais e fatores. Máximo C4.")
    headers = ["Ciclo", "Período", "Data início", "Data fim", "% reajuste ciclo", "Fator próprio", "Fator acumulado", "Possui efeito financeiro?"]
    for i, h in enumerate(headers, 1):
        ws_base.cell(4, i, h)
    _style_header(ws_base, 4, 1, len(headers))

    for idx in range(5):
        r = 5 + idx
        ws_base.cell(r, 1, f"C{idx}")
        ws_base.cell(r, 2, f'=IF(C{r}="","",RIGHT("0"&MONTH(C{r}),2)&"/"&YEAR(C{r})&" a "&RIGHT("0"&MONTH(D{r}),2)&"/"&YEAR(D{r}))')
        ws_base.cell(r, 3, "=ENTRADAS_CRITICAS!$C$5" if idx == 0 else f"=EDATE($C$5,{idx*12})")
        ws_base.cell(r, 4, f'=IF(C{r}="","",EDATE(C{r},12)-1)')
        ws_base.cell(r, 5, 0 if idx == 0 else "")
        ws_base.cell(r, 6, f"=1+E{r}")
        ws_base.cell(r, 7, 1 if idx == 0 else f"=G{r-1}*F{r}")
        ws_base.cell(r, 8, "Não" if idx == 0 else "Sim")

        for c in range(1, len(headers) + 1):
            _auto_cell(ws_base.cell(r, c))
        _input_cell(ws_base.cell(r, 5))
        _input_cell(ws_base.cell(r, 8))
        _date(ws_base.cell(r, 3))
        _date(ws_base.cell(r, 4))
        _pct(ws_base.cell(r, 5))
        _factor(ws_base.cell(r, 6))
        _factor(ws_base.cell(r, 7))

    _add_list_validation(ws_base, "'_LISTAS'!$B$2:$B$3", ["H5:H9"])
    _set_widths(ws_base, {"A": 12, "B": 22, "C": 16, "D": 16, "E": 18, "F": 18, "G": 18, "H": 24})
    ws_base.freeze_panes = "A5"

    # FINANCEIRO
    _title(ws_fin, "Financeiro", "Preencher competência, valor reconhecido/medido e observação. O ciclo, fator e valor atualizado são automáticos.")
    headers = ["Competência", "Valor reconhecido/medido", "Observação", "Ciclo identificado", "Fator aplicável", "Valor atualizado", "Delta", "Efeito financeiro?"]
    for i, h in enumerate(headers, 1):
        ws_fin.cell(4, i, h)
    _style_header(ws_fin, 4, 1, len(headers))

    for r in range(5, 155):
        _input_cell(ws_fin.cell(r, 1), "dd/mm/yyyy")
        _input_cell(ws_fin.cell(r, 2), 'R$ #,##0.00')
        _input_cell(ws_fin.cell(r, 3))
        ws_fin.cell(r, 4, f'=IF(A{r}="","",INDEX(BASE_CICLOS!$A$5:$A$9,MATCH(A{r},BASE_CICLOS!$C$5:$C$9,1)))')
        ws_fin.cell(r, 5, f'=IF(D{r}="","",INDEX(BASE_CICLOS!$G$5:$G$9,MATCH(D{r},BASE_CICLOS!$A$5:$A$9,0)))')
        ws_fin.cell(r, 6, f'=IF(OR(B{r}="",E{r}=""),"",B{r}*E{r})')
        ws_fin.cell(r, 7, f'=IF(OR(F{r}="",B{r}=""),"",F{r}-B{r})')
        ws_fin.cell(r, 8, f'=IF(D{r}="","",INDEX(BASE_CICLOS!$H$5:$H$9,MATCH(D{r},BASE_CICLOS!$A$5:$A$9,0)))')
        for c in range(4, 9):
            _auto_cell(ws_fin.cell(r, c))
        _factor(ws_fin.cell(r, 5))
        _money(ws_fin.cell(r, 6))
        _money(ws_fin.cell(r, 7))

    _set_widths(ws_fin, {"A": 18, "B": 24, "C": 42, "D": 18, "E": 18, "F": 20, "G": 18, "H": 18})
    ws_fin.freeze_panes = "A5"

    # ITENS_SALDO
    _title(ws_itens, "Itens e Saldo", "Informe itens, quantidades originais, remanescentes por ciclo e/ou saldo após a última competência financeira.")
    headers = [
        "Item", "Qtd original C0", "VU C0", "Rem. início C1", "Rem. início C2",
        "Rem. início C3", "Rem. início C4", "Saldo após última competência financeira",
        "Saldo na data de corte", "Observação", "VT original", "Fator saldo",
        "VT saldo após última competência atualizado", "VT saldo data corte atualizado"
    ]
    for i, h in enumerate(headers, 1):
        ws_itens.cell(4, i, h)
    _style_header(ws_itens, 4, 1, len(headers))

    for r in range(5, 155):
        for c in range(1, 11):
            _input_cell(ws_itens.cell(r, c))
        _money(ws_itens.cell(r, 3))
        ws_itens.cell(r, 11, f'=IF(OR(B{r}="",C{r}=""),"",B{r}*C{r})')
        ws_itens.cell(r, 12, '=IF(ENTRADAS_CRITICAS!$C$19="","",INDEX(BASE_CICLOS!$G$5:$G$9,MATCH(ENTRADAS_CRITICAS!$C$19,BASE_CICLOS!$A$5:$A$9,0)))')
        ws_itens.cell(r, 13, f'=IF(OR(H{r}="",C{r}="",L{r}=""),"",H{r}*C{r}*L{r})')
        ws_itens.cell(r, 14, f'=IF(OR(I{r}="",C{r}="",L{r}=""),"",I{r}*C{r}*L{r})')
        for c in range(11, 15):
            _auto_cell(ws_itens.cell(r, c))
        _money(ws_itens.cell(r, 11))
        _factor(ws_itens.cell(r, 12))
        _money(ws_itens.cell(r, 13))
        _money(ws_itens.cell(r, 14))

    _set_widths(ws_itens, {
        "A": 18, "B": 18, "C": 16, "D": 18, "E": 18, "F": 18, "G": 18,
        "H": 30, "I": 22, "J": 42, "K": 18, "L": 16, "M": 32, "N": 26
    })
    ws_itens.freeze_panes = "A5"

    # ADITIVOS_SUPRESSOES
    _title(ws_adit, "Aditivos e Supressões", "Permite informar por Qtd × VU ou por valor consolidado. Supressão computável entra negativa.")
    headers = [
        "Identificação", "Tipo", "Forma de informar", "Data", "Ciclo/Marco",
        "Quantidade", "Valor unitário", "Valor consolidado informado",
        "Valor original", "Aplicar fator?", "Fator aplicável", "Valor atualizado",
        "Tratamento", "Valor computável", "Observação"
    ]
    for i, h in enumerate(headers, 1):
        ws_adit.cell(4, i, h)
    _style_header(ws_adit, 4, 1, len(headers))

    for r in range(5, 75):
        for c in [1, 2, 3, 4, 6, 7, 8, 10, 13, 15]:
            _input_cell(ws_adit.cell(r, c))
        _date(ws_adit.cell(r, 4))
        _money(ws_adit.cell(r, 7))
        _money(ws_adit.cell(r, 8))
        ws_adit.cell(r, 5, f'=IF(D{r}="","",INDEX(BASE_CICLOS!$A$5:$A$9,MATCH(D{r},BASE_CICLOS!$C$5:$C$9,1)))')
        ws_adit.cell(r, 9, f'=IF(C{r}="Valor consolidado",H{r},IF(OR(F{r}="",G{r}=""),"",F{r}*G{r}))')
        ws_adit.cell(r, 11, f'=IF(J{r}="Sim",INDEX(BASE_CICLOS!$G$5:$G$9,MATCH(E{r},BASE_CICLOS!$A$5:$A$9,0)),1)')
        ws_adit.cell(r, 12, f'=IF(I{r}="","",I{r}*K{r})')
        ws_adit.cell(r, 14, f'=IF(AND(A{r}="",B{r}="",C{r}="",D{r}="",F{r}="",G{r}="",H{r}=""),"",IF(OR(M{r}="Informativo / não computar",M{r}="Já embutido no saldo"),0,IF(B{r}="Supressão",-1,1)*IF(M{r}="Computar nominal",I{r},L{r})))')
        for c in [5, 9, 11, 12, 14]:
            _auto_cell(ws_adit.cell(r, c))
        _money(ws_adit.cell(r, 9))
        _factor(ws_adit.cell(r, 11))
        _money(ws_adit.cell(r, 12))
        _money(ws_adit.cell(r, 14))

    _add_list_validation(ws_adit, "'_LISTAS'!$G$2:$G$3", ["B5:B74"])
    _add_list_validation(ws_adit, "'_LISTAS'!$I$2:$I$3", ["C5:C74"])
    _add_list_validation(ws_adit, "'_LISTAS'!$B$2:$B$3", ["J5:J74"])
    _add_list_validation(ws_adit, "'_LISTAS'!$H$2:$H$5", ["M5:M74"])
    _set_widths(ws_adit, {
        "A": 24, "B": 16, "C": 20, "D": 14, "E": 14, "F": 16, "G": 18,
        "H": 24, "I": 18, "J": 16, "K": 16, "L": 18, "M": 28, "N": 18, "O": 44
    })
    ws_adit.freeze_panes = "A5"

    # MEMORIA_VTA
    _title(ws_mem, "Memória do Valor Total Atualizado", "Aba de conferência. O Python deverá recalcular esta memória no processamento final.")
    headers = ["Componente", "Fonte", "Valor", "Status", "Observação"]
    for i, h in enumerate(headers, 1):
        ws_mem.cell(4, i, h)
    _style_header(ws_mem, 4, 1, len(headers))

    memoria_rows = [
        ("C0", '=ENTRADAS_CRITICAS!$C$13', '=IF(ENTRADAS_CRITICAS!$C$13="Valor consolidado",ENTRADAS_CRITICAS!$C$14,IF(ENTRADAS_CRITICAS!$C$13="Financeiro mensal",SUMIF(FINANCEIRO!$D:$D,"C0",FINANCEIRO!$F:$F),IF(ENTRADAS_CRITICAS!$C$13="Itens",SUMPRODUCT((ITENS_SALDO!$B$5:$B$154-ITENS_SALDO!$D$5:$D$154)*ITENS_SALDO!$C$5:$C$154),0)))', '=IF(C5>0,"Incluído","Ausente")', "C0 explícito na aba ENTRADAS_CRITICAS."),
        ("C1", "Financeiro", '=SUMIF(FINANCEIRO!$D:$D,"C1",FINANCEIRO!$F:$F)', '=IF(C6>0,"Incluído","Sem financeiro")', ""),
        ("C2", "Financeiro", '=SUMIF(FINANCEIRO!$D:$D,"C2",FINANCEIRO!$F:$F)', '=IF(C7>0,"Incluído","Sem financeiro")', ""),
        ("C3 executado", "Financeiro", '=SUMIF(FINANCEIRO!$D:$D,"C3",FINANCEIRO!$F:$F)', '=IF(C8>0,"Incluído","Sem financeiro")', "Executado até a última competência financeira informada, quando aplicável."),
        ("C4 executado", "Financeiro", '=SUMIF(FINANCEIRO!$D:$D,"C4",FINANCEIRO!$F:$F)', '=IF(C9>0,"Incluído","Sem financeiro")', ""),
        ("Saldo remanescente após última competência", '=ENTRADAS_CRITICAS!$C$17&" | "&ENTRADAS_CRITICAS!$C$18', '=IF(ENTRADAS_CRITICAS!$C$17="Valor consolidado",ENTRADAS_CRITICAS!$C$20,IF(AND(ENTRADAS_CRITICAS!$C$17="Itens",ENTRADAS_CRITICAS!$C$18="Após última competência financeira"),SUM(ITENS_SALDO!$M$5:$M$154),IF(AND(ENTRADAS_CRITICAS!$C$17="Itens",ENTRADAS_CRITICAS!$C$18="Data de corte"),SUM(ITENS_SALDO!$N$5:$N$154),0)))', '=IF(C10>0,"Incluído","Ausente")', "Só deve somar com financeiro parcial se representar saldo após última competência ou data de corte equivalente."),
        ("Aditivos/supressões", "ADITIVOS_SUPRESSOES", '=SUM(ADITIVOS_SUPRESSOES!$N$5:$N$74)', '=IF(C11<>0,"Incluído","Sem valor computável")', "Supressões computáveis entram negativas."),
        ("TOTAL", "Consolidação", '=SUM(C5:C11)', "VTA", "Valor Total Atualizado do Contrato."),
    ]
    for r, vals in enumerate(memoria_rows, 5):
        for c, val in enumerate(vals, 1):
            ws_mem.cell(r, c, val)
            _auto_cell(ws_mem.cell(r, c))
        _money(ws_mem.cell(r, 3))
        if vals[0] == "TOTAL":
            for c in range(1, 6):
                ws_mem.cell(r, c).fill = _fill(VERDE)
                ws_mem.cell(r, c).font = Font(bold=True)

    _set_widths(ws_mem, {"A": 42, "B": 38, "C": 20, "D": 18, "E": 78})
    ws_mem.freeze_panes = "A5"

    # CONCILIACAO_REFERENCIA
    _title(ws_conc, "Conciliação com Referência Externa", "Use esta aba para reconciliar a memória cl8us com uma memória externa, como GMP/PADTEC.")
    headers = ["Parcela", "Valor referência externa", "Valor cl8us", "Diferença", "Status", "Observação"]
    for i, h in enumerate(headers, 1):
        ws_conc.cell(4, i, h)
    _style_header(ws_conc, 4, 1, len(headers))

    conciliacao = [
        ("C0 executado", "", '=MEMORIA_VTA!C5', '=B5-C5', '=IF(ABS(D5)<=1,"OK","DIVERGENTE")', "Ex.: referência GMP C0."),
        ("C1 atualizado", "", '=MEMORIA_VTA!C6', '=B6-C6', '=IF(ABS(D6)<=1,"OK","DIVERGENTE")', ""),
        ("C2 atualizado", "", '=MEMORIA_VTA!C7', '=B7-C7', '=IF(ABS(D7)<=1,"OK","DIVERGENTE")', ""),
        ("C3 executado até corte", "", '=MEMORIA_VTA!C8', '=B8-C8', '=IF(ABS(D8)<=1,"OK","DIVERGENTE")', ""),
        ("Residual/remanescente após corte", "", '=MEMORIA_VTA!C10', '=B9-C9', '=IF(ABS(D9)<=1,"OK","DIVERGENTE")', ""),
        ("Aditivos/supressões computáveis", "", '=MEMORIA_VTA!C11', '=B10-C10', '=IF(ABS(D10)<=1,"OK","DIVERGENTE")', ""),
        ("TOTAL", '=SUM(B5:B10)', '=MEMORIA_VTA!C12', '=B11-C11', '=IF(ABS(D11)<=1,"OK","DIVERGENTE")', "A ponte deve fechar contra a referência externa."),
    ]

    for r, vals in enumerate(conciliacao, 5):
        for c, val in enumerate(vals, 1):
            ws_conc.cell(r, c, val)
            if c == 2:
                _input_cell(ws_conc.cell(r, c), 'R$ #,##0.00')
            else:
                _auto_cell(ws_conc.cell(r, c))
        _money(ws_conc.cell(r, 2))
        _money(ws_conc.cell(r, 3))
        _money(ws_conc.cell(r, 4))
        if vals[0] == "TOTAL":
            for c in range(1, 7):
                ws_conc.cell(r, c).fill = _fill(VERDE)
                ws_conc.cell(r, c).font = Font(bold=True)

    ws_conc.conditional_formatting.add("E5:E11", FormulaRule(formula=['E5="DIVERGENTE"'], fill=_fill(VERMELHO)))
    ws_conc.conditional_formatting.add("E5:E11", FormulaRule(formula=['E5="OK"'], fill=_fill(VERDE)))
    _set_widths(ws_conc, {"A": 34, "B": 24, "C": 22, "D": 20, "E": 18, "F": 62})
    ws_conc.freeze_panes = "A5"

    # VALIDACOES
    _title(ws_val, "Validações da Matriz 2.1", "Checks operacionais. Erros críticos devem ser resolvidos antes de processar documentos.")
    headers = ["Validação", "Resultado", "Se der ERRO/ALERTA, o que fazer"]
    for i, h in enumerate(headers, 1):
        ws_val.cell(4, i, h)
    _style_header(ws_val, 4, 1, len(headers))

    checks = [
        ("C0 informado ou justificado", '=IF(OR(ENTRADAS_CRITICAS!$C$13="Não disponível",MEMORIA_VTA!$C$5>0),"OK","ERRO")', "Informar C0 consolidado, financeiro mensal ou itens."),
        ("C0 usado no VTA quando informado", '=IF(AND(ENTRADAS_CRITICAS!$C$13<>"Não disponível",MEMORIA_VTA!$C$5=0),"ERRO","OK")', "Verificar ENTRADAS_CRITICAS e MEMORIA_VTA."),
        ("Saldo remanescente tratado", '=IF(AND(ENTRADAS_CRITICAS!$C$10="Sim",MEMORIA_VTA!$C$10=0),"ERRO","OK")', "Informar saldo por valor consolidado ou itens."),
        ("Risco de dupla contagem no ciclo em execução", '=IF(AND(ENTRADAS_CRITICAS!$C$7="Sim",SUMIF(FINANCEIRO!$D:$D,ENTRADAS_CRITICAS!$C$8,FINANCEIRO!$F:$F)>0,ENTRADAS_CRITICAS!$C$18="Início do ciclo em execução",MEMORIA_VTA!$C$10>0),"ERRO","OK")', "Se há financeiro parcial no ciclo, o saldo deve ser após a última competência ou deve ser declarado que já desconta a execução."),
        ("Saldo atual sem vínculo claro", '=IF(AND(ENTRADAS_CRITICAS!$C$10="Sim",ENTRADAS_CRITICAS!$C$18="Saldo atual sem vínculo claro"),"ALERTA","OK")', "Informar marco do saldo para evitar dupla contagem."),
        ("Aditivos declarados e computados", '=IF(AND(ENTRADAS_CRITICAS!$C$24="Sim",MEMORIA_VTA!$C$11=0),"ALERTA","OK")', "Verificar ADITIVOS_SUPRESSOES e tratamento."),
        ("Supressões não positivas", '=IF(SUMIFS(ADITIVOS_SUPRESSOES!$N:$N,ADITIVOS_SUPRESSOES!$B:$B,"Supressão")>0,"ERRO","OK")', "Supressão computável deve entrar negativa."),
        ("Valor original do aditivo calculado", '=IF(COUNTIF(ADITIVOS_SUPRESSOES!$I$5:$I$74,"#VALUE!")>0,"ERRO","OK")', "Verificar forma de informar, quantidade, VU ou valor consolidado."),
        ("VTA calculado na memória", '=IF(MEMORIA_VTA!$C$12>0,"OK","ERRO")', "Verificar C0, financeiro, saldo e aditivos."),
        ("Conciliação referência preenchida", '=IF(CONCILIACAO_REFERENCIA!$B$11>0,CONCILIACAO_REFERENCIA!$E$11,"ALERTA")', "Preencher valores de referência externa se houver memória GMP/PADTEC."),
    ]

    for r, vals in enumerate(checks, 5):
        for c, val in enumerate(vals, 1):
            ws_val.cell(r, c, val)
            _auto_cell(ws_val.cell(r, c))
        ws_val.cell(r, 2).font = Font(bold=True)

    ws_val.conditional_formatting.add("B5:B30", FormulaRule(formula=['B5="ERRO"'], fill=_fill(VERMELHO)))
    ws_val.conditional_formatting.add("B5:B30", FormulaRule(formula=['B5="ALERTA"'], fill=_fill(LARANJA)))
    ws_val.conditional_formatting.add("B5:B30", FormulaRule(formula=['B5="DIVERGENTE"'], fill=_fill(VERMELHO)))
    ws_val.conditional_formatting.add("B5:B30", FormulaRule(formula=['B5="OK"'], fill=_fill(VERDE)))
    _set_widths(ws_val, {"A": 44, "B": 18, "C": 76})
    ws_val.freeze_panes = "A5"

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False

    wb.save(OUT)
    print(f"OK: Matriz 2.1 experimental gerada em: {OUT}")
    print("Ajustes desta versão:")
    print("- Campo específico para ciclo em execução.")
    print("- Marco do saldo remanescente.")
    print("- Saldo após última competência financeira em ITENS_SALDO.")
    print("- Validação contra dupla contagem.")
    print("- Aba CONCILIACAO_REFERENCIA.")


def gerar_coleta_reajuste_matriz21_bytes():
    """Retorna o XLSX da Matriz 2.1 em bytes para uso no Streamlit."""
    build()
    return OUT.read_bytes()


def gerar_coleta_matriz21_bytes():
    return gerar_coleta_reajuste_matriz21_bytes()


def gerar_coleta_reajuste_matriz21():
    return gerar_coleta_reajuste_matriz21_bytes()


if __name__ == "__main__":
    build()

# FIM - versão v3 com referências C24/C25 corrigidas.

# >>> PATCH_SAIDA_PORTAVEL_MATRIZ21_V1
# Correção de release: evita gravar a coleta em caminho fixo/local durante o download.
# A geração para Streamlit passa a usar arquivo temporário único, impedindo erro de permissão
# quando ColetaReajuste_Matriz21.xlsx estiver aberto no Excel ou preso em outra pasta.

def gerar_coleta_reajuste_matriz21_bytes(dados_admissibilidade=None):
    from pathlib import Path as _Path
    import tempfile as _tempfile
    import uuid as _uuid

    global OUT

    _out_anterior = globals().get("OUT", None)
    _tmp_path = _Path(_tempfile.gettempdir()) / f"ColetaReajuste_Matriz21_{_uuid.uuid4().hex}.xlsx"

    try:
        OUT = _tmp_path

        try:
            _retorno = build(dados_admissibilidade=dados_admissibilidade)
        except TypeError:
            _retorno = build()

        _caminho = _Path(_retorno) if _retorno else _tmp_path
        if not _caminho.exists():
            _caminho = _tmp_path
        if not _caminho.exists():
            raise FileNotFoundError(f"Arquivo Matriz 2.1 não foi gerado: {_caminho}")

        _data = _caminho.read_bytes()

        _validador = globals().get("_validar_identidade_matriz21")
        if callable(_validador):
            _validador(_data)

        return _data

    finally:
        try:
            if _tmp_path.exists():
                _tmp_path.unlink()
        except Exception:
            pass

        if _out_anterior is not None:
            OUT = _out_anterior


def gerar_coleta_matriz21_bytes(dados_admissibilidade=None):
    return gerar_coleta_reajuste_matriz21_bytes(dados_admissibilidade)


def gerar_coleta_reajuste_matriz21(dados_admissibilidade=None):
    return gerar_coleta_reajuste_matriz21_bytes(dados_admissibilidade)
# <<< PATCH_SAIDA_PORTAVEL_MATRIZ21_V1
# >>> PATCH_M21_INDICE_DADOS_ADMISSIBILIDADE_V5
# Consolida o indice da calculadora dentro do XLS Matriz 2.1 sem depender de caminho fixo.
try:
    _gerar_coleta_reajuste_matriz21_bytes_base_v5 = gerar_coleta_reajuste_matriz21_bytes

    def _m21_extrair_indice_dados_v5(dados_admissibilidade):
        if not isinstance(dados_admissibilidade, dict):
            return ""
        for chave in (
            "indice",
            "indice_utilizado",
            "indice_contratual",
            "tipo_indice",
            "indice_nome",
        ):
            valor = dados_admissibilidade.get(chave)
            if valor not in (None, ""):
                return str(valor).strip()
        ciclos = dados_admissibilidade.get("ciclos") or []
        if isinstance(ciclos, list):
            for ciclo in ciclos:
                if isinstance(ciclo, dict):
                    for chave in (
                        "indice",
                        "indice_utilizado",
                        "indice_contratual",
                        "tipo_indice",
                        "indice_nome",
                    ):
                        valor = ciclo.get(chave)
                        if valor not in (None, ""):
                            return str(valor).strip()
        return ""

    def _m21_injetar_indice_no_xlsx_v5(data, indice):
        if not indice:
            return data
        from io import BytesIO
        from openpyxl import load_workbook

        bio_in = BytesIO(data)
        wb = load_workbook(bio_in)
        try:
            if "ENTRADAS_CRITICAS" not in wb.sheetnames:
                return data
            ws = wb["ENTRADAS_CRITICAS"]
            linha_indice = None
            for row in range(1, ws.max_row + 1):
                valor_b = ws.cell(row=row, column=2).value
                if isinstance(valor_b, str) and valor_b.strip().lower() in {
                    "índice utilizado",
                    "indice utilizado",
                    "índice contratual",
                    "indice contratual",
                }:
                    linha_indice = row
                    break
            if linha_indice is None:
                linha_indice = max(5, ws.max_row + 1)
                ws.cell(row=linha_indice, column=1).value = "INDICE"
                ws.cell(row=linha_indice, column=2).value = "Índice utilizado"
                ws.cell(row=linha_indice, column=4).value = "Preenchido automaticamente pela calculadora."
                ws.cell(row=linha_indice, column=5).value = "Usado para cards e documentos."
            ws.cell(row=linha_indice, column=3).value = indice
            bio_out = BytesIO()
            wb.save(bio_out)
            return bio_out.getvalue()
        finally:
            wb.close()

    def gerar_coleta_reajuste_matriz21_bytes(dados_admissibilidade=None):
        indice = _m21_extrair_indice_dados_v5(dados_admissibilidade)
        try:
            data = _gerar_coleta_reajuste_matriz21_bytes_base_v5(dados_admissibilidade)
        except TypeError:
            data = _gerar_coleta_reajuste_matriz21_bytes_base_v5()
        return _m21_injetar_indice_no_xlsx_v5(data, indice)

    def gerar_coleta_matriz21_bytes(dados_admissibilidade=None):
        return gerar_coleta_reajuste_matriz21_bytes(dados_admissibilidade)

    def gerar_coleta_reajuste_matriz21(dados_admissibilidade=None):
        return gerar_coleta_reajuste_matriz21_bytes(dados_admissibilidade)
except Exception:
    pass
# <<< PATCH_M21_INDICE_DADOS_ADMISSIBILIDADE_V5
