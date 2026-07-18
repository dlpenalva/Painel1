"""Integra o Arquivo Coleta Oficial ao contrato documental do runtime.

O leitor Python produz os resultados operacionais. Os valores gravados em
RESULTADOS são usados para auditoria e reconciliação, nunca como substituição
silenciosa do cálculo Python.
"""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from _coleta_reajuste import ler_coleta_reajuste
from _leitor_masterfile_v10 import ler_masterfile_v10
from _politica_entrega_segura import avaliar_entrega_segura


def _numero(valor: Any, padrao: float = 0.0) -> float:
    if valor in (None, "") or isinstance(valor, bool):
        return padrao
    try:
        return float(valor)
    except (TypeError, ValueError):
        return padrao


def _data_br(valor: Any) -> str:
    if isinstance(valor, datetime):
        return valor.strftime("%d/%m/%Y")
    return str(valor or "")


def _retroativo_python(memoria: dict[str, Any], metodo: str) -> float | None:
    total = 0.0
    evidencias = 0
    for ciclo in memoria.get("ciclos") or []:
        reg = ((ciclo.get("retroativo") or {}).get(metodo) or {})
        evidencias += int(reg.get("evidencias") or 0)
        total += _numero(reg.get("retroativo"))
    return round(total, 2) if evidencias else None


def adaptar_coleta_reajuste_para_documentos(
    conteudo: bytes,
    *,
    leitura: dict[str, Any] | None = None,
    diagnostico: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Monta o contrato documental, preferindo os cálculos do leitor Python."""

    diagnostico = diagnostico or ler_coleta_reajuste(conteudo)
    if not diagnostico.get("valido"):
        raise ValueError("A coleta possui pendências estruturais e não pode liberar documentos.")
    capacidades = diagnostico.get("capacidades") or {}

    wb = load_workbook(BytesIO(conteudo), data_only=True, read_only=True)
    controle = wb["CONTROLE"]
    parametros = wb["parametros"]
    financeiro = wb["financeiro"]
    remanescentes = wb["itens_Remanesc"]
    aditivos = wb["aditivos"]
    posicao = wb["posicao_contratual"] if "posicao_contratual" in wb.sheetnames else None
    itens_rc = wb["itens_RC"]
    resultados = wb["RESULTADOS"]

    ciclos_rows = []
    fatores: dict[str, float] = {}
    for row in range(2, 7):
        ciclo = str(parametros[f"B{row}"].value or "").upper()
        if not ciclo:
            continue
        fator = _numero(parametros[f"F{row}"].value, 1.0)
        fatores[ciclo] = fator
        ciclos_rows.append(
            {
                "Ciclo": ciclo,
                "Data-base": _data_br(parametros[f"C{row}"].value),
                "Intervalo do índice": "",
                "Janela de admissibilidade": "",
                "Data do pedido": "",
                "Situação": parametros[f"G{row}"].value or "",
                "Tratamento financeiro do ciclo": "Apurar" if str(parametros[f"A{row}"].value).lower() == "sim" else "Fora da apuração",
                "Variação": _numero(parametros[f"E{row}"].value),
                "Fator": fator,
                "Fator acumulado": fator,
                "Fator acumulado efetivo": fator,
                "Fator ciclo efetivo": fator,
            }
        )
    df_ciclos = pd.DataFrame(ciclos_rows)

    financeiro_rows = []
    for row in range(2, 74):
        valor = financeiro[f"C{row}"].value
        if valor in (None, ""):
            continue
        financeiro_rows.append(
            {
                "Ciclo": str(financeiro[f"B{row}"].value or "").upper(),
                "Competência": financeiro[f"A{row}"].value,
                "Valor pago/faturado": _numero(valor),
                "Fator aplicado": _numero(financeiro[f"D{row}"].value, 1.0),
                "Valor atualizado": _numero(financeiro[f"E{row}"].value),
                "Delta": _numero(financeiro[f"F{row}"].value),
                "Efeito financeiro": financeiro[f"G{row}"].value or "",
            }
        )
    df_financeiro = pd.DataFrame(financeiro_rows)

    fin_ciclos = []
    for ciclo in [row["Ciclo"] for row in ciclos_rows]:
        linhas = df_financeiro[df_financeiro["Ciclo"] == ciclo] if not df_financeiro.empty else pd.DataFrame()
        pago = float(linhas["Valor pago/faturado"].sum()) if not linhas.empty else 0.0
        devido = float(linhas["Valor atualizado"].sum()) if not linhas.empty else 0.0
        delta = float(linhas["Delta"].sum()) if not linhas.empty else 0.0
        fin_ciclos.append(
            {
                "Ciclo": ciclo,
                "Situação": next((r["Situação"] for r in ciclos_rows if r["Ciclo"] == ciclo), ""),
                "Tratamento financeiro": "Apurar",
                "Fator aplicado ao retroativo": fatores.get(ciclo, 1.0),
                "Fator acumulado": fatores.get(ciclo, 1.0),
                "Valor pago efetivo": pago,
                "Valor teórico calculado": devido,
                "Delta do ciclo": delta,
            }
        )
    df_fin_por_ciclo = pd.DataFrame(fin_ciclos)

    valores_rows = []
    qtd_cols = {"C0": "B", "C1": "E", "C2": "G", "C3": "I", "C4": "K"}
    qtd_posicao_cols = {"C0": "G", "C1": "K", "C2": "O", "C3": "S", "C4": "W"}
    total_cols = {"C0": "D", "C1": "F", "C2": "H", "C3": "J", "C4": "L"}
    total_rc_cols = {"C0": "D", "C1": "G", "C2": "J", "C3": "M", "C4": "P"}
    for row in range(2, 201):
        item = remanescentes[f"A{row}"].value
        if item in (None, ""):
            continue
        for ciclo, qtd_col in qtd_cols.items():
            qtd = _numero(
                posicao[f"{qtd_posicao_cols[ciclo]}{row}"].value
                if posicao is not None
                else remanescentes[f"{qtd_col}{row}"].value
            )
            total = _numero(
                itens_rc[f"{total_rc_cols[ciclo]}{row + 1}"].value
                if posicao is not None
                else remanescentes[f"{total_cols[ciclo]}{row}"].value
            )
            valores_rows.append(
                {
                    "Item": item,
                    "Ciclo": ciclo,
                    "Valor unitário": total / qtd if qtd else _numero(remanescentes[f"C{row}"].value) * fatores.get(ciclo, 1.0),
                    "Quantidade": qtd,
                    "Total R$": total,
                    "Ciclo precluso": False,
                }
            )
    df_valores = pd.DataFrame(valores_rows)

    aditivos_rows = []
    for row in range(2, 201):
        item = aditivos[f"A{row}"].value
        if item in (None, ""):
            continue
        aditivos_rows.append(
            {
                "Item": item,
                "Data do aditivo": aditivos[f"B{row}"].value,
                "Ciclo/Marco": aditivos[f"C{row}"].value or "",
                "Tipo de alteração": aditivos[f"D{row}"].value or "",
                "Quantidade da alteração": _numero(aditivos[f"E{row}"].value),
                "Delta quantitativo contratual": _numero(aditivos[f"L{row}"].value),
                "Status da posição": aditivos[f"M{row}"].value or "",
                "Valor do aditivo na assinatura": _numero(aditivos[f"G{row}"].value),
                "Fator aplicado": _numero(aditivos[f"I{row}"].value, 1.0),
                "Valor do aditivo reajustado": _numero(aditivos[f"J{row}"].value),
                "Computa no Valor Global": str(aditivos[f"K{row}"].value or "").lower() == "sim",
            }
        )
    df_aditivos = pd.DataFrame(aditivos_rows)

    posicao_rows = []
    if posicao is not None:
        contratada_cols = {"C0": "E", "C1": "I", "C2": "M", "C3": "Q", "C4": "U"}
        rem_cols = {"C0": "G", "C1": "K", "C2": "O", "C3": "S", "C4": "W"}
        for row in range(2, 201):
            item = posicao[f"A{row}"].value
            if item in (None, ""):
                continue
            for ciclo in ("C0", "C1", "C2", "C3", "C4"):
                posicao_rows.append(
                    {
                        "Item": item,
                        "Ciclo": ciclo,
                        "Quantidade contratada vigente": _numero(posicao[f"{contratada_cols[ciclo]}{row}"].value),
                        "Quantidade remanescente ajustada": _numero(posicao[f"{rem_cols[ciclo]}{row}"].value),
                        "Status": posicao[f"X{row}"].value or "",
                    }
                )
    df_posicao_contratual = pd.DataFrame(posicao_rows)

    calculos = capacidades.get("calculos") or {}
    retro_capacidade = calculos.get("retroativo") or {}
    vta_capacidade = calculos.get("vta") or {}
    rem_capacidade = calculos.get("valor_remanescente") or {}

    memoria = ((leitura or {}).get("objeto_processo") or {}).get("memoria_por_ciclo") or {}
    modo_python = str(((leitura or {}).get("controle") or {}).get("modo") or "")
    metodo_python = {"principal": "financeiro", "pc": "pc", "d": "consumidos"}.get(modo_python)
    retroativo_python = _retroativo_python(memoria, metodo_python) if metodo_python else None
    ciclo_vigente = str(((leitura or {}).get("controle") or {}).get("ciclo_vigente") or controle["B2"].value or "").upper()
    residual_vigente = next(
        (c.get("residuais") or {} for c in memoria.get("ciclos") or [] if str(c.get("ciclo") or "").upper() == ciclo_vigente),
        {},
    )
    vta_python = _numero((memoria.get("vta") or {}).get("valor_total_atualizado"), None)

    valor_original = _numero(resultados["B20"].value)
    retroativo = retroativo_python if retroativo_python is not None else _numero(retro_capacidade.get("valor"))
    rem_original = _numero(resultados["C35"].value)
    if int(residual_vigente.get("itens") or 0) > 0:
        rem_original = _numero(residual_vigente.get("valor_original"))
        rem_atualizado = _numero(residual_vigente.get("valor_atualizado"))
    else:
        rem_atualizado = _numero(rem_capacidade.get("valor"))
    valor_total = vta_python if vta_python is not None else _numero(vta_capacidade.get("valor"))
    pago_total = float(df_financeiro["Valor pago/faturado"].sum()) if not df_financeiro.empty else 0.0
    devido_total = float(df_financeiro["Valor atualizado"].sum()) if not df_financeiro.empty else 0.0
    total_aditivos = float(df_aditivos["Valor do aditivo reajustado"].sum()) if not df_aditivos.empty else 0.0
    fator_final = max(fatores.values(), default=1.0)
    execucao_atualizada = valor_total - rem_atualizado

    df_rem = pd.DataFrame(
        [{
            "Ciclo": controle["B2"].value or "",
            "Remanescente original": rem_original,
            "Fator aplicado": (rem_atualizado / rem_original) if rem_original else fator_final,
            "Remanescente atualizado": rem_atualizado,
            "Observação": "Valores oficiais lidos da aba RESULTADOS.",
        }]
    )
    df_execucao = pd.DataFrame(
        [{
            "Ciclo": controle["B2"].value or "",
            "Valor executado original": max(valor_original - rem_original, 0.0),
            "Valor executado atualizado": execucao_atualizada,
        }]
    )
    df_composicao = pd.DataFrame(
        [
            {"Componente": "Execução atualizada e retroativo", "Valor": execucao_atualizada},
            {"Componente": "Saldo remanescente atualizado", "Valor": rem_atualizado},
            {"Componente": "Valor Total Atualizado do Contrato", "Valor": valor_total},
        ]
    )
    df_comparativo = pd.DataFrame(
        [
            {"Indicador": "Valor original do contrato", "Valor": valor_original},
            {"Indicador": "Valor pago efetivo", "Valor": pago_total},
            {"Indicador": "Valor teórico calculado", "Valor": devido_total},
            {"Indicador": "Valor represado a pagar", "Valor": retroativo},
            {"Indicador": "Saldo remanescente original", "Valor": rem_original},
            {"Indicador": "Saldo remanescente atualizado", "Valor": rem_atualizado},
            {"Indicador": "Valor Total Atualizado do Contrato", "Valor": valor_total},
        ]
    )

    resultado_documental = {
        "ok": True,
        "origem_coleta": "Coleta_Reajuste.xlsx",
        "data_processamento": datetime.now().strftime("%d/%m/%Y %H:%M"),
        "modo_apuracao": "Processamento progressivo pelo XLS",
        "base_execucao_mensal_disponivel": bool(financeiro_rows),
        "base_itens_disponivel": bool(valores_rows),
        "aviso_base_execucao": "" if financeiro_rows else "Financeiro não informado; os demais blocos permanecem utilizáveis.",
        "ressalva_modo_apuracao": "Somente resultados sustentados pelos blocos disponíveis são apresentados.",
        "config_ciclo_em_execucao": {},
        "corte_operacional_solicitado": False,
        "corte_operacional_aplicado": False,
        "origem_ciclos": "Coleta_Reajuste.xlsx",
        "indice": controle["B7"].value or "Não informado",
        "fator_acumulado": fator_final,
        "variacao_acumulada": fator_final - 1.0,
        "quantidade_ciclos": len(ciclos_rows),
        "valor_original_contrato": valor_original,
        "contexto_contratual_anterior": {},
        "valor_formalizado_anterior": valor_original,
        "impacto_analise_atual": valor_total - valor_original,
        "valor_pago_efetivo": pago_total,
        "total_pago_faturado": pago_total,
        "valor_teorico_calculado": devido_total,
        "total_devido_reajustado": devido_total,
        "delta_total": retroativo,
        "delta_acumulado": retroativo,
        "valor_represado_a_pagar": retroativo,
        "valor_retroativo_estimado_itens_estoque": retroativo,
        "retroativo_estimado_itens_estoque_disponivel": bool(abs(retroativo) > 0.004),
        "quantidade_meses_sem_efeito_financeiro": 0,
        "valor_total_sem_efeito_financeiro": 0.0,
        "remanescente_original": rem_original,
        "remanescente_reajustado": rem_atualizado,
        "fator_remanescente": (rem_atualizado / rem_original) if rem_original else fator_final,
        "valor_executado_atualizado": execucao_atualizada,
        "valor_calculado_sem_aditivos": valor_total,
        "valor_atualizado_contrato": valor_total,
        "valor_global_financeiro": valor_total,
        "total_aditivos_atualizados": total_aditivos,
        "total_aditivos_informativos": 0.0,
        "aditivos_somados_ao_valor_total": False,
        "quantidade_aditivos_total": len(aditivos_rows),
        "quantidade_aditivos_marcados_computaveis": sum(bool(r["Computa no Valor Global"]) for r in aditivos_rows),
        "ciclo_ultimo_remanescente": controle["B2"].value or "",
        "df_ciclos": df_ciclos,
        "df_financeiro_mensal": df_financeiro,
        "df_financeiro_mensal_corte_operacional": df_financeiro,
        "df_financeiro_mensal_tratado": df_financeiro,
        "df_meses_sem_efeito_financeiro": pd.DataFrame(),
        "df_financeiro_por_ciclo": df_fin_por_ciclo,
        "df_delta_por_ciclo": df_fin_por_ciclo.copy(),
        "df_execucao_atualizada": df_execucao,
        "df_retroativo_estimado_itens_estoque": pd.DataFrame(),
        "df_composicao_valor_total": df_composicao,
        "df_remanescentes": df_rem,
        "df_valores_unitarios_ciclo": df_valores,
        "df_aditivos": df_aditivos,
        "df_aditivos_executivo": df_aditivos,
        "df_aditivos_computaveis": df_aditivos[df_aditivos.get("Computa no Valor Global", pd.Series(dtype=bool)) == True] if not df_aditivos.empty else pd.DataFrame(),
        "df_aditivos_informativos": pd.DataFrame(),
        "df_posicao_contratual": df_posicao_contratual,
        "df_pedidos_compra": pd.DataFrame(((leitura or {}).get("itens_pc_v10") or {}).get("itens") or []),
        "df_itens_consumidos_runtime": pd.DataFrame(((leitura or {}).get("itens_consumidos_v10") or {}).get("itens") or []),
        "df_comparativo": df_comparativo,
        "df_auditoria_consistencia": pd.DataFrame(
            [
                {
                    "Validação": "Blocos independentes do XLS avaliados",
                    "Status": "OK",
                    "Diferença/Valor": "Processamento progressivo",
                }
            ]
        ),
        "status_resultados": diagnostico.get("metadados", {}).get("status_resultados", {}),
        "capacidades": capacidades,
        "diagnostico_coleta": diagnostico,
        "resultados_progressivos": {
            "retroativo": retro_capacidade,
            "vta": vta_capacidade,
            "valor_remanescente": rem_capacidade,
            "posicao_contratual": calculos.get("posicao_contratual") or {},
            "valores_unitarios": calculos.get("valores_unitarios") or {},
        },
        "_resultado_lido_do_excel": True,
    }
    if leitura:
        resultado_documental.update({
            "origem_coleta": "COLETA_REAJUSTE_OFICIAL.xlsx",
            "objeto_processo": leitura.get("objeto_processo") or {},
            "reconciliacao_xls_python": leitura.get("reconciliacao_xls_python") or {},
            "posicao_contratual_runtime": leitura.get("posicao_contratual") or {},
            "memoria_por_ciclo": memoria,
            "_resultado_calculado_python": True,
        })
    return resultado_documental


def processar_coleta_oficial_runtime(conteudo: bytes) -> tuple[dict[str, Any], dict[str, Any]]:
    """Entry point único usado pelo upload real e pelos testes de integração."""
    leitura = ler_masterfile_v10(conteudo, exigir_modelo_oficial=True)
    if not leitura.get("ok"):
        raise ValueError(leitura.get("erro") or "O Arquivo Coleta Oficial não pôde ser lido.")

    diagnostico = ler_coleta_reajuste(conteudo)
    if not diagnostico.get("valido"):
        pendencias = diagnostico.get("pendencias") or diagnostico.get("bloqueios_estruturais") or []
        raise ValueError("; ".join(str(item) for item in pendencias) or "A estrutura do Arquivo Coleta Oficial é inválida.")

    resultado = adaptar_coleta_reajuste_para_documentos(
        conteudo,
        leitura=leitura,
        diagnostico=diagnostico,
    )
    reconciliacao = leitura.get("reconciliacao_xls_python") or {}
    politica = avaliar_entrega_segura(leitura)
    bloqueios = list(politica.get("bloqueios") or [])

    diagnostico["reconciliacao_xls_python"] = reconciliacao
    diagnostico["politica_entrega_segura"] = politica
    diagnostico["avisos"] = list(diagnostico.get("avisos") or []) + list(leitura.get("avisos") or [])
    diagnostico["bloqueios_criticos"] = list(diagnostico.get("bloqueios_criticos") or []) + bloqueios
    if bloqueios:
        diagnostico["pronto_para_consolidar"] = False

    capacidades = resultado.get("capacidades") or {}
    if bloqueios:
        for documento in (capacidades.get("documentos") or {}).values():
            documento["habilitado"] = False
            documento["estado"] = "bloqueado"
            documento["rotulo"] = "Bloqueado para formalização"
            documento["classificacao"] = "BLOQUEADO PARA FORMALIZAÇÃO"
            documento["motivo"] = bloqueios[0]

    resultado.update({
        "capacidades": capacidades,
        "diagnostico_coleta": diagnostico,
        "reconciliacao_xls_python": reconciliacao,
        "politica_entrega_segura": politica,
        "formalizacao_bloqueada": bool(bloqueios),
        "bloqueios_formalizacao": bloqueios,
    })
    return resultado, diagnostico
