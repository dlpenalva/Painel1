"""Template oficial do XLS Coleta (modelo novo, com aba posicao_contratual).

Fonte de verdade estrutural: templates/COLETA_REAJUSTE_OFICIAL.xlsx
(SHA-256 8569357302273d874aefc3cf7e5e25af6d60d216bc94457f953d0dc4a9666c6f,
apos rodada de UX de 17/07/2026 via Excel real: destaque CONTROLE!B1 +
protecao, itens_PC coluna A em Texto com estilo de entrada, DATA_PC
dd/mm/aaaa, CICLO_PC com alerta de data invalida, V:AC ocultas,
posicao_contratual com cores por ciclo, MEMORIA_RESULTADOS oculta com a
matematica homologada e RESULTADOS executiva com quatro tabelas e um unico
status global. Os nomes definidos permanecem como contrato de leitura.

Correcao aprovada sobre o arquivo fornecido (Coleta_Reajuste_Nova.xlsx,
SHA-256 6353357b...8069ff3): o original possuia uma OMISSAO CRITICA — a
aba itens_PC vinha sem NUMERO_PC, eliminando o identificador documental e
o controle de duplicidade homologados. O template oficial corrigido
reintroduz NUMERO_PC na coluna A (ITEM permanece removido), desloca
formulas/validacoes/estilos/referencias e amplia o CHECK_PC_FINANCEIRO
com a regra de duplicidade de NUMERO_PC (TRIM+UPPER; vazio nao avalia).
A grade do financeiro foi estendida ate a linha 73 (72 competencias).

Este modulo NAO reconstroi o XLS celula a celula: carrega o template oficial
e apenas limpa os dados demonstrativos das celulas de entrada, preservando
todas as formulas, validacoes, estilos, merges e intervalos nomeados.
"""
from __future__ import annotations

import hashlib
import re
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from dateutil.relativedelta import relativedelta

from _capacidade_pcs import CAPACIDADE_PCS, ULTIMA_LINHA_PCS

ROOT = Path(__file__).resolve().parent
TEMPLATE_COLETA_OFICIAL = ROOT / "templates" / "COLETA_REAJUSTE_OFICIAL.xlsx"
# Nome de referencia interno do template (nao renomear o arquivo fisico).
NOME_ARQUIVO_COLETA_OFICIAL = "COLETA_REAJUSTE_OFICIAL.xlsx"
# Nome user-facing entregue nos botoes de download da interface (§15).
NOME_DOWNLOAD_COLETA = "Coleta_Reajuste.xlsx"


def nome_download_coleta(dados_admissibilidade=None):
    """Nome user-facing do download, com ciclos e indice da analise.

    Mantem o nome-base atual e acrescenta os ciclos ja apurados, ordenados e
    sem repeticao, seguidos do indice canonico: Coleta_Reajuste_C1_C2_IPCA.xlsx.
    Ciclo precluso tambem entra — o nome representa o ESCOPO da apuracao, nao
    o resultado da admissibilidade.

    Altera apenas o parametro `file_name` do download: conteudo, template,
    abas e processamento do XLSX nao dependem desta funcao. Sem ciclos
    recuperaveis com seguranca, devolve NOME_DOWNLOAD_COLETA. Sem indice
    reconhecivel, preserva o nome anterior composto apenas pelos ciclos.
    """
    from _reajuste_utils import ciclos_da_analise

    try:
        ciclos = ciclos_da_analise(dados_admissibilidade)
    except Exception:
        return NOME_DOWNLOAD_COLETA
    if not ciclos:
        return NOME_DOWNLOAD_COLETA

    base, _, extensao = NOME_DOWNLOAD_COLETA.rpartition(".")
    sufixo = "_".join(f"C{numero}" for numero in ciclos)
    indice = str((dados_admissibilidade or {}).get("indice") or "").strip().upper()
    indice_sem_hifen = indice.replace("-", "")
    if indice.startswith("IST"):
        sufixo_indice = "IST"
    elif indice.startswith("ICTI"):
        sufixo_indice = "ICTI"
    elif indice.startswith("IPCA"):
        sufixo_indice = "IPCA"
    elif indice_sem_hifen.startswith("IGPM"):
        sufixo_indice = "IGPM"
    elif indice.startswith("INPC"):
        sufixo_indice = "INPC"
    else:
        sufixo_indice = ""
    if sufixo_indice:
        sufixo = f"{sufixo}_{sufixo_indice}"
    return f"{base}_{sufixo}.{extensao}"

# Ordem oficial das abas do novo modelo. CICLO_EM_EXECUCAO e acrescentada em
# runtime (sem versionar outro binario XLSX) e as demais continuam tendo o
# template oficial como fonte de verdade.
# comparativo_VTA (aba de referencia adicionada via Excel COM) fica na 1a
# posicao fisica que o Excel produz ao inseri-la; nao aparece na web e a aba
# ativa original e preservada.
ABAS_COLETA_OFICIAL = [
    "comparativo_VTA",
    "CONTROLE",
    "parametros",
    "financeiro",
    "itens_Remanesc",
    "itens_Consumidos",
    "itens_PC",
    "aditivos",
    "posicao_referencia",
    "CICLO_EM_EXECUCAO",
    "posicao_contratual",
    "itens_RC",
    "historico_VU",
    "cobertura_temporal",
    "MEMORIA_RESULTADOS",
    "RESULTADOS",
]

ABA_POSICAO_CONTRATUAL = "posicao_contratual"

# Colunas oficiais de itens_PC (linha 1, A:L) — NUMERO_PC obrigatorio na
# coluna A como identificador documental e chave de duplicidade. ITEM nao
# existe no modelo oficial (decisao aprovada; nao restaurar).
COLUNAS_ITENS_PC_OFICIAL = [
    "NUMERO_PC",
    "DATA_PC",
    "CICLO_PC",
    "VALOR_PC",
    "FATOR_ACUMULADO",
    "VALOR_ATUALIZADO",
    "PC_PAGO_A_CONTRATADA",
    "RETROATIVO_RECONHECIDO_A_PAGAR",
    "VALOR_ATUALIZADO_EM_ANALISE",
    "DELTA_POTENCIAL",
    "CHECK_PC_FINANCEIRO",
    "EFEITO_FINANCEIRO_PC",
]

# Cabecalhos essenciais da aba posicao_contratual (linha 1, A:X)
COLUNAS_POSICAO_CONTRATUAL = (
    ["ITEM", "VU_ORIGINAL", "QTD_BASE_ORIGINAL"]
    + [
        rotulo
        for n in range(5)
        for rotulo in (
            f"DELTA_C{n}",
            f"QTD_CONTRATADA_C{n}",
            f"QTD_REM_BASE_C{n}",
            f"QTD_REM_AJUSTADA_C{n}",
        )
    ]
    + ["CHECK_POSICAO_CONTRATUAL"]
)

# Celulas de entrada com dados demonstrativos no template — limpar na geracao.
# Somente celulas de VALOR sao limpas; formulas nunca sao tocadas.
_RESIDUOS_POR_ABA: dict[str, list[str]] = {
    # B2=ciclo vigente demo, B3=data de corte demo, B7=indice, B8=data-base
    "CONTROLE": ["B2", "B3", "B7", "B8"],
}
# parametros: linhas 2-6, colunas de entrada (B=CICLO e F=formula preservados)
_RESIDUOS_POR_ABA["parametros"] = [
    f"{col}{lin}" for lin in range(2, 7) for col in ("A", "C", "D", "E", "G", "H")
]
# Etapa 4: bloco MEMORIA DE CALCULO (J2:R80) — somente valores; os cabecalhos
# J1:R1 pertencem ao template e nao sao residuos.
_RESIDUOS_POR_ABA["parametros"] += [
    f"{col}{lin}" for lin in range(2, 81)
    for col in ("J", "K", "L", "M", "N", "O", "P", "Q", "R")
]
# financeiro: competencia (A), valor pago (C) e efeito (G) — linhas 2-73
# (grade estendida para 72 competencias mensais)
_RESIDUOS_POR_ABA["financeiro"] = [
    f"{col}{lin}" for lin in range(2, 74) for col in ("A", "C", "G")
]
# itens_Remanesc: entradas do fiscal (item, base, VU, remanescentes por ciclo)
_RESIDUOS_POR_ABA["itens_Remanesc"] = [
    f"{col}{lin}" for lin in range(2, 201)
    for col in ("A", "B", "C", "E", "G", "I", "K")
]
# aditivos: eventos demonstrativos (F soh quando valor manual sobrepoe formula)
_RESIDUOS_POR_ABA["aditivos"] = [
    f"{col}{lin}" for lin in range(2, 201)
    for col in ("A", "B", "D", "E", "F", "H", "K")
]
# Etapa 26F: a unica interface manual fica na RESULTADOS executiva.
_RESIDUOS_POR_ABA["RESULTADOS"] = [
    f"{col}{lin}" for lin in range(43, 51) for col in ("C", "D", "E", "F", "G")
]


def assinatura_template_coleta(caminho: str | Path | None = None) -> str:
    """SHA-256 dos bytes do template oficial — assinatura robusta de cache.

    Muda sempre que o conteudo binario do XLS muda, mesmo com nome, versao
    da aplicacao, tamanho e timestamp identicos. Template ausente/ilegivel
    retorna "" (o consumidor decide como tratar; obter_coleta_oficial_bytes
    ja falha explicitamente nesse caso).
    """
    alvo = Path(caminho) if caminho is not None else TEMPLATE_COLETA_OFICIAL
    try:
        return hashlib.sha256(alvo.read_bytes()).hexdigest()
    except OSError:
        return ""


def eh_layout_coleta_oficial(wb) -> bool:
    """True quando o workbook segue o novo modelo (aba posicao_contratual)."""
    return ABA_POSICAO_CONTRATUAL in wb.sheetnames


def _limpar_residuos(wb) -> None:
    for aba, celulas in _RESIDUOS_POR_ABA.items():
        if aba not in wb.sheetnames:
            continue
        ws = wb[aba]
        for coord in celulas:
            cell = ws[coord]
            if isinstance(cell.value, str) and cell.value.startswith("="):
                continue  # formula do template: nunca limpar
            if cell.value is not None:
                cell.value = None


def _garantir_dropdown_metodo(wb) -> None:
    """CONTROLE!B1 sempre oferece os rotulos canonicos do metodo.

    Fonte unica: _metodo_apuracao.OPCOES_DROPDOWN. O template binario
    homologado permanece intacto; a lista da validacao de dados e
    reescrita em runtime, de modo que NOVAS Coletas ofereçam
    "Financeiro (Mensalidade)" sem exibir o rotulo legado
    "Principal (Financeiro)" — que segue aceito na LEITURA como alias.
    """
    from _metodo_apuracao import OPCOES_DROPDOWN

    if "CONTROLE" not in wb.sheetnames:
        return
    ws = wb["CONTROLE"]
    lista = '"' + ",".join(OPCOES_DROPDOWN) + '"'
    for dv in ws.data_validations.dataValidation:
        if dv.type == "list" and str(dv.sqref) == "B1":
            dv.formula1 = lista


def _garantir_formulas_cronologia_execucao(wb) -> None:
    """CICLO por cronologia fixa da execucao — ETAPA 31.

    Reescreve em runtime as formulas de enquadramento de CICLO do template
    (financeiro!B, itens_PC!C, aditivos!C): o ciclo de um fato/competencia da
    EXECUCAO e o bloco fixo de 12 meses contado do inicio cronologico de C0
    (parametros!C2), com 60 competencias no total (C0..C4). A janela de
    reajuste parametros!C:D deixa de enquadrar a execucao — ela pode conter
    intervalo intencional quando o reajuste seguinte foi retardado, e nenhum
    fato valido pode cair "Fora dos ciclos" por causa desse intervalo.
    O template binario homologado permanece intacto (mesma estrategia do
    dropdown do metodo).
    """
    ref = "parametros!$C$2"

    def _meses(celula: str) -> str:
        return (
            f"(YEAR({celula})-YEAR({ref}))*12+MONTH({celula})-MONTH({ref})"
        )

    def _bloco(celula: str, minusculo: bool) -> str:
        prefixo = '"c"' if minusculo else '"C"'
        return (
            f'IF(NOT(ISNUMBER({ref})),"Fora dos ciclos",'
            f'IF(OR({celula}<{ref},{_meses(celula)}>59),"Fora dos ciclos",'
            f'{prefixo}&INT(({_meses(celula)})/12)))'
        )

    def _reescrever(nome_aba: str, coluna: int, montar) -> None:
        if nome_aba not in wb.sheetnames:
            return
        ws = wb[nome_aba]
        for r in range(2, ws.max_row + 1):
            atual = ws.cell(r, coluna).value
            if (
                isinstance(atual, str)
                and atual.startswith("=")
                and "parametros!$C$" in atual
            ):
                ws.cell(r, coluna).value = montar(r)

    _reescrever(
        "financeiro", 2,
        lambda r: f'=IF($A{r}="","",{_bloco(f"$A{r}", True)})',
    )
    _reescrever(
        "itens_PC", 3,
        lambda r: (
            f'=IF(B{r}="","",IF(NOT(ISNUMBER(B{r})),"DATA_PC invalida",'
            f'{_bloco(f"B{r}", False)}))'
        ),
    )
    _reescrever(
        "aditivos", 3,
        lambda r: (
            f'=IF(B{r}="","",IFERROR({_bloco(f"B{r}", False)},'
            f'"Fora dos ciclos"))'
        ),
    )


def _garantir_fator_historico_desacoplado(wb) -> None:
    """Desacopla o fator DESTA analise do fator HISTORICO integral (31.1).

    CONTROLE!B11 passou a ser o fator apurado NESTA analise (1 + B10); o
    template, porem, tem tres consumidores criados quando B11 significava o
    fator HISTORICO integral. A fonte canonica visivel do historico passa a
    ser RESULTADOS!H5 (o titulo de G5 ja e "Fator histórico integral"):

      * RESULTADOS!H5 recebe a logica historica fail-closed que vivia em
        CONTROLE!B11 antes da Etapa 31 (C0=1; Cn somente com parametros!E
        completo de C1..Cn -> parametros!F do ciclo vigente; senao vazio —
        historico faltante JAMAIS e inventado);
      * RESULTADOS!H8 valida a existencia de $H$5 (nao mais de B11);
      * comparativo_VTA!B208 multiplica pelo RESULTADOS!$H$5;
      * RESULTADOS!C12 (texto auditavel) passa a citar RESULTADOS!H5.

    O template binario homologado permanece intacto (mesma estrategia do
    dropdown e da cronologia da execucao).
    """
    formula_historica = (
        '=IFERROR(IF(UPPER(CONTROLE!$B$2)="C0",1,'
        'IF(UPPER(CONTROLE!$B$2)="C1",'
        'IF(COUNT(parametros!$E$3:$E$3)=1,parametros!$F$3,""),'
        'IF(UPPER(CONTROLE!$B$2)="C2",'
        'IF(COUNT(parametros!$E$3:$E$4)=2,parametros!$F$4,""),'
        'IF(UPPER(CONTROLE!$B$2)="C3",'
        'IF(COUNT(parametros!$E$3:$E$5)=3,parametros!$F$5,""),'
        'IF(UPPER(CONTROLE!$B$2)="C4",'
        'IF(COUNT(parametros!$E$3:$E$6)=4,parametros!$F$6,""),""))))),"")'
    )
    if "RESULTADOS" in wb.sheetnames:
        res = wb["RESULTADOS"]
        atual_h5 = str(res["H5"].value or "")
        if "CONTROLE!$B$11" in atual_h5:
            res["H5"].value = formula_historica
        atual_h8 = str(res["H8"].value or "")
        if "CONTROLE!$B$11" in atual_h8:
            res["H8"].value = atual_h8.replace("CONTROLE!$B$11", "$H$5")
        atual_c12 = str(res["C12"].value or "")
        if "CONTROLE!B11" in atual_c12:
            res["C12"].value = atual_c12.replace(
                "CONTROLE!B11", "RESULTADOS!H5"
            )
    if "comparativo_VTA" in wb.sheetnames:
        cv = wb["comparativo_VTA"]
        atual_b208 = str(cv["B208"].value or "")
        if "CONTROLE!$B$11" in atual_b208:
            cv["B208"].value = atual_b208.replace(
                "CONTROLE!$B$11", "RESULTADOS!$H$5"
            )


def _garantir_completude_abertura_temporal(wb) -> None:
    """Completude da abertura pela DATA de nascimento do item — ETAPA 44.

    `MEMORIA_RESULTADOS!W41:W45` responde "a abertura de Cn esta completa?" e
    so pode exigir remanescente dos itens que JA INTEGRAVAM o contrato naquela
    data. O criterio de aplicabilidade era `posicao_contratual!Y`
    (CICLO_NASCIMENTO), derivado da QUANTIDADE: um item incluido no MEIO de Cn
    recebe Y=n e passa a ser exigido na abertura de Cn — fotografia em que ele
    ainda nao existia. O vazio de um item inexistente virava "abertura
    incompleta" e podia empurrar o VTA para um ciclo anterior.

    A troca e apenas do criterio de APLICABILIDADE, para a coluna canonica
    temporal `posicao_contratual!AL` (CICLO_NASCIMENTO_DATA, derivada da DATA
    DE EFEITO do aditivo) — a mesma ja usada por `W57`, por `itens_RC!AB`
    ("NAO APLICAVEL") e pela formatacao condicional de itens_Remanesc. Item
    comprovadamente criado depois da abertura deixa de ser exigido nela: o
    motor o trata como zero estrutural, sem escrever 0,00 em celula alguma.

    Nenhuma QUANTIDADE muda: `G/K/O/S/W`, `T21/T22/T23/T25` e `B26` (VTA_FINAL)
    permanecem com a formula e o valor originais. O template binario
    homologado permanece intacto (mesma estrategia do dropdown do metodo, da
    cronologia da execucao e do fator historico).
    """
    if "MEMORIA_RESULTADOS" not in wb.sheetnames:
        return
    if ABA_POSICAO_CONTRATUAL not in wb.sheetnames:
        return
    pc = wb[ABA_POSICAO_CONTRATUAL]
    # Fail-closed: linhagem sem a coluna temporal permanece exatamente como
    # esta (trocar Y por uma coluna inexistente zeraria a exigencia).
    if str(pc["AL1"].value or "").strip() != "CICLO_NASCIMENTO_DATA":
        return

    mem = wb["MEMORIA_RESULTADOS"]
    criterio_antigo = "posicao_contratual!$Y$2:$Y$201"
    criterio_novo = "posicao_contratual!$AL$2:$AL$201"
    for celula in ("W41", "W42", "W43", "W44", "W45"):
        atual = mem[celula].value
        if isinstance(atual, str) and criterio_antigo in atual:
            mem[celula].value = atual.replace(criterio_antigo, criterio_novo)


# Orientacao de novo item por aditivo — ETAPA 44. Reaproveita os DOIS espacos
# de orientacao que ja tratam de novo item na aba aditivos (cabecalho A1 e
# mensagem de validacao da coluna M): nenhuma secao, linha, caixa ou aba nova.
_ORIENTACAO_ADITIVOS_A1_ANTIGA = (
    "ITEM\n(NOVO ITEM: cadastrar em itens_Remanesc como N001, N002...; "
    "QTD_BASE_ORIGINAL fica 0 automaticamente; informar o VU_ORIGINAL)"
)
_ORIENTACAO_ADITIVOS_A1_NOVA = (
    "ITEM\n(NOVO ITEM: cadastrar em itens_Remanesc PRIMEIRO - N001, N002...; "
    "QTD_BASE_ORIGINAL fica 0 automaticamente; informar o VU_ORIGINAL. "
    "Depois, a data e a quantidade aqui.)"
)
_MSG_NOVO_ITEM_ANTIGA = (
    "NOVO ITEM NAO CADASTRADO - CADASTRAR EM itens_Remanesc; "
    "QTD_BASE_ORIGINAL sera 0 automaticamente; informar VU_ORIGINAL."
)
_MSG_NOVO_ITEM_NOVA = (
    "NOVO ITEM NAO CADASTRADO - CADASTRAR EM itens_Remanesc: ITEM + "
    "VU_ORIGINAL (QTD_BASE_ORIGINAL sera 0 automaticamente); "
    "depois a data e a quantidade aqui."
)


# T27 conta os itens que DEVERIAM ter base fisica de C0 e nao a tem. A forma
# abaixo e a homologada ate a Etapa 44 e serve de guarda: so ela e reescrita.
_T27_ANTIGA = (
    '=SUMPRODUCT((posicao_contratual!$A$2:$A$201<>"")'
    "*(1-ISNUMBER(posicao_contratual!$G$2:$G$201)"
    "*ISNUMBER(posicao_contratual!$K$2:$K$201)"
    "*(ISNUMBER(historico_VU!$C$2:$C$201)"
    '+(posicao_contratual!$Y$2:$Y$201<>"")'
    "*(posicao_contratual!$Y$2:$Y$201>0))))"
)
# Aplicabilidade temporal como FATOR (nunca como parcela somada): cada fator
# vale 0 ou 1, de modo que nenhuma linha pode contribuir com valor negativo.
_T27_NOVA = (
    '=SUMPRODUCT((posicao_contratual!$A$2:$A$201<>"")'
    "*(1-ISNUMBER(posicao_contratual!$AL$2:$AL$201)"
    "*(posicao_contratual!$AL$2:$AL$201>0))"
    "*(1-ISNUMBER(posicao_contratual!$G$2:$G$201)"
    "*ISNUMBER(posicao_contratual!$K$2:$K$201)"
    "*ISNUMBER(historico_VU!$C$2:$C$201)))"
)


def _garantir_base_fisica_c0_temporal(wb) -> None:
    """T27 so cobra base fisica de C0 de quem existia em C0 — ETAPA 44.1.

    `MEMORIA_RESULTADOS!T27` conta os itens que DEVERIAM ter base fisica de C0
    (G, K e VU_C0) e nao a tem; `T25` vira "CALCULO MANUAL REQUERIDO" quando
    `T27>0` sem PC de C0. A forma anterior trazia uma dispensa para novo item
    SOMADA dentro do produto:

        1 - ISNUMBER(G)*ISNUMBER(K)*(ISNUMBER(VU_C0) + (Y<>"")*(Y>0))

    Essa dispensa era inalcancavel: `G` e `IF($Y2>0,"",ROUND(E2,2))`, ou seja,
    fica vazia exatamente quando `Y>0`. O item nascido depois de C0 era
    reprovado por `ISNUMBER(G)=FALSO` antes de a dispensa ser avaliada e
    acusava falta de base fisica de uma fotografia em que nem existia.

    Trocar apenas `Y` por `AL` nao resolveria e ainda corromperia a contagem:
    com `AL>0` podendo coexistir com `ISNUMBER(G)` verdadeiro (item nascido no
    MEIO de C0), a soma `(ISNUMBER(VU_C0) + dispensa)` chega a 2 dentro do
    produto e a linha contribui com -1, mascarando pendencia legitima de outro
    item.

    A aplicabilidade temporal passa a ser um FATOR proprio, derivado da data
    real de inclusao (`posicao_contratual!AL`, CICLO_NASCIMENTO_DATA):

        (A<>"")
        * (1 - ISNUMBER(AL)*(AL>0))
        * (1 - ISNUMBER(G)*ISNUMBER(K)*ISNUMBER(VU_C0))

    Cada fator vale 0 ou 1 — nenhuma linha pode contribuir com valor negativo.
    `ISNUMBER(AL)` (e nao `AL<>""`) porque `"">0` e VERDADEIRO no Excel: o teste
    de tipo e a forma segura, e ja e a usada por `W57` e pela formatacao
    condicional de itens_Remanesc. Item sem nascimento derivavel (`AL` vazio)
    continua integralmente validado — fail-closed preservado.

    Nada mais muda: `X2`, `G/K/O/S/W`, `W41:W45`, `T21/T22/T23` e a formula de
    `B26` permanecem intactos; `T25` e `B26` mudam apenas pelo efeito natural
    de `T27`. O template binario homologado permanece intacto.
    """
    if "MEMORIA_RESULTADOS" not in wb.sheetnames:
        return
    if ABA_POSICAO_CONTRATUAL not in wb.sheetnames:
        return
    if "historico_VU" not in wb.sheetnames:
        return
    pc = wb[ABA_POSICAO_CONTRATUAL]
    if str(pc["AL1"].value or "").strip() != "CICLO_NASCIMENTO_DATA":
        return
    mem = wb["MEMORIA_RESULTADOS"]
    # Estrutura nao reconhecida (outra linhagem de template, ou ja reescrita):
    # nao inventar formula.
    if str(mem["T27"].value or "") != _T27_ANTIGA:
        return
    mem["T27"].value = _T27_NOVA


_CINZA_NAO_APLICAVEL = "FFD9D9D9"
# QTD_REM_BASE_Cn (entrada manual) -> indice do ciclo cuja abertura ela retrata.
_COLS_QTD_ABERTURA_MANUAL = (("E", 1), ("G", 2), ("I", 3), ("K", 4))


def _garantir_cinza_nao_aplicavel_visivel(wb) -> None:
    """Faz o cinza "NAO APLICAVEL" realmente pintar no Excel — ETAPA 44.

    A regra ja existia e ja era a primeira (com stopIfTrue) de cada coluna
    QTD_REM_BASE_Cn, comparando o espelho local `BI` (CICLO_NASCIMENTO_DATA)
    com o indice do ciclo. O que faltava era a RENDERIZACAO: o preenchimento
    do formato diferencial (dxf) fora gravado como
    `<patternFill patternType="solid"><fgColor rgb="FFD9D9D9"/></patternFill>`.

    Num dxf o Excel le a cor de fundo em `bgColor` — `fgColor` sozinho, com
    patternType solid, resolve para "sem preenchimento". Por isso a celula de
    um item ainda inexistente continuava com o fundo comum de entrada, sem
    nenhum sinal de NAO APLICAVEL. Os dxf que o proprio Excel escreveu neste
    mesmo arquivo usam a forma canonica `<patternFill><bgColor .../>`.

    A correcao acrescenta `bgColor` ao dxf da regra de NAO APLICAVEL,
    preservando `fgColor` (a leitura por openpyxl continua identica). Nada
    mais muda: formula, ordem, prioridade e stopIfTrue das quatro regras
    permanecem, e as outras tres regras nao sao tocadas. Nenhum texto entra na
    celula numerica e nenhum 0,00 e escrito — a celula segue VAZIA, so
    visualmente marcada como nao aplicavel.
    """
    from openpyxl.styles import PatternFill

    if "itens_Remanesc" not in wb.sheetnames:
        return
    ws = wb["itens_Remanesc"]
    for coluna, idx in _COLS_QTD_ABERTURA_MANUAL:
        faixa = f"{coluna}2:{coluna}201"
        marca = f"$BI2>{idx}"
        for formatacao in ws.conditional_formatting:
            if str(formatacao.sqref) != faixa:
                continue
            for regra in formatacao.rules:
                formula = str(regra.formula[0]) if regra.formula else ""
                if marca not in formula or "ISNUMBER($BI2)" not in formula:
                    continue  # so a regra de NAO APLICAVEL e reescrita
                if regra.dxf is None or regra.dxf.fill is None:
                    continue
                regra.dxf.fill = PatternFill(
                    patternType="solid",
                    fgColor=_CINZA_NAO_APLICAVEL,
                    bgColor=_CINZA_NAO_APLICAVEL,
                )


def _garantir_orientacao_novo_item_por_aditivo(wb) -> None:
    """Diz ao fiscal ONDE COMECAR o cadastro de item novo — ETAPA 44.

    O aditivo e o gatilho: quem inclui um item novo abre a aba `aditivos`. A
    orientacao passa a explicitar a ORDEM (itens_Remanesc primeiro: ITEM +
    VU_ORIGINAL; aditivos depois: data e quantidade), reescrevendo em runtime
    os textos que JA EXISTEM — o cabecalho `A1` e a mensagem de validacao da
    coluna `M`. Sem nova secao, sem linha nova, sem caixa e sem aba nova; o
    template binario homologado permanece intacto.
    """
    if "aditivos" not in wb.sheetnames:
        return
    ws = wb["aditivos"]
    if str(ws["A1"].value or "") == _ORIENTACAO_ADITIVOS_A1_ANTIGA:
        ws["A1"].value = _ORIENTACAO_ADITIVOS_A1_NOVA
    for row in range(2, ULTIMA_LINHA_ORIENTACAO_ADITIVOS + 1):
        celula = ws.cell(row=row, column=13)  # coluna M
        atual = celula.value
        if isinstance(atual, str) and _MSG_NOVO_ITEM_ANTIGA in atual:
            celula.value = atual.replace(
                _MSG_NOVO_ITEM_ANTIGA, _MSG_NOVO_ITEM_NOVA
            )


_VERDE_RETROATIVO_RECONHECIDO = "FFC6EFCE"
_AMBAR_RETROATIVO_POTENCIAL = "FFFFEB9C"
_NEUTRO_VALOR_EM_ANALISE = "FFF2F2F2"


def _garantir_apresentacao_retroativos_e_aditivos(wb) -> None:
    """Aplica apenas a apresentacao user-facing e a decisao tecnica de K.

    Os nomes tecnicos do template permanecem aceitos pela validacao estrutural.
    Na copia entregue, reconhecido e potencial recebem rotulos amigaveis e
    cores opostas. Em ``aditivos``, todo delta valido ja alimenta
    ``posicao_contratual`` e o remanescente; K registra automaticamente essa
    prevencao de dupla contagem e fica oculto da visao normal.
    """
    from copy import copy as _copy
    from openpyxl.styles import Font, PatternFill

    if "itens_PC" in wb.sheetnames:
        ws = wb["itens_PC"]
        for coord in ("H1", "Q1"):
            ws[coord].value = "RETROATIVO RECONHECIDO"
        for coord in ("J1", "S1"):
            ws[coord].value = "RETROATIVO POTENCIAL"

        estilos = (
            (("H", 1, 1), _VERDE_RETROATIVO_RECONHECIDO),
            (("I", 1, 1), _NEUTRO_VALOR_EM_ANALISE),
            (("J", 1, 1), _AMBAR_RETROATIVO_POTENCIAL),
            (("Q", 1, 7), _VERDE_RETROATIVO_RECONHECIDO),
            (("R", 1, 7), _NEUTRO_VALOR_EM_ANALISE),
            (("S", 1, 7), _AMBAR_RETROATIVO_POTENCIAL),
        )
        for (coluna, inicio, fim), cor in estilos:
            fill = PatternFill("solid", fgColor=cor)
            for linha in range(inicio, fim + 1):
                ws[f"{coluna}{linha}"].fill = fill
        for coord in ("H1", "I1", "J1", "Q1", "R1", "S1"):
            fonte = _copy(ws[coord].font)
            fonte.color = "FF1F1F1F"
            fonte.bold = True
            ws[coord].font = fonte

    if "aditivos" in wb.sheetnames:
        ws = wb["aditivos"]
        ws["H1"].value = "Aplicar reajuste? (Sim/Nao)"
        for linha in range(2, 201):
            if ws[f"H{linha}"].value == "Não":
                ws[f"H{linha}"].value = "Nao"
            celula = ws[f"K{linha}"]
            celula.value = (
                f'=IF(A{linha}="","",IF(M{linha}="OK","Nao",""))'
            )
            celula.fill = PatternFill("solid", fgColor=_NEUTRO_VALOR_EM_ANALISE)
            celula.font = Font(name="Calibri", size=9, color="FF666666")
        ws.column_dimensions["K"].hidden = True
        for validacao in list(ws.data_validations.dataValidation):
            faixas = [
                str(faixa)
                for faixa in validacao.sqref.ranges
                if str(faixa) != "K2:K200"
            ]
            if len(faixas) == len(validacao.sqref.ranges):
                continue
            if faixas:
                validacao.sqref = " ".join(faixas)
            else:
                ws.data_validations.dataValidation.remove(validacao)


def _validar_estrutura_itens_pc(wb) -> None:
    """Barreira contra regressao critica: itens_PC jamais pode sair esvaziada.

    Valida cabecalhos A1:L1 e a densidade de formulas da grade (C:L, ate a
    capacidade canonica de PCs — Etapa 26G). Se o template em disco for
    trocado por uma versao sem a estrutura homologada, a geracao falha
    explicitamente em vez de entregar o XLS.
    """
    ws = wb["itens_PC"]
    cabecalhos = [ws.cell(1, c).value for c in range(1, 13)]
    if cabecalhos != COLUNAS_ITENS_PC_OFICIAL:
        raise ValueError(
            "itens_PC invalida: cabecalhos A1:L1 divergem do modelo oficial "
            f"(encontrado: {cabecalhos})"
        )
    formulas = sum(
        1
        for row in ws.iter_rows(
            min_row=2, max_row=ULTIMA_LINHA_PCS, min_col=3, max_col=12
        )
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    )
    if formulas < CAPACIDADE_PCS * 7:
        raise ValueError(
            f"itens_PC invalida: apenas {formulas} formulas na grade "
            f"C2:L{ULTIMA_LINHA_PCS} (estrutura esvaziada ou truncada)"
        )
    # Visao salva rolada (ex.: topLeftCell=AD1 com V:AC ocultas e sem grade)
    # faz a aba parecer completamente vazia no Excel — causa raiz do bug de
    # 17/07/2026. A visao deve sempre abrir em A1.
    tlc = ws.sheet_view.topLeftCell
    if tlc not in (None, "A1"):
        raise ValueError(
            f"itens_PC invalida: visao salva rolada para {tlc} — a aba "
            "abriria visualmente vazia; corrija o template (topLeftCell=A1)"
        )


def obter_coleta_oficial_bytes() -> bytes:
    """Retorna o novo XLS Coleta oficial em branco (estrutura integral).

    Carrega o template oficial com data_only=False (formulas preservadas)
    e limpa apenas os dados demonstrativos das celulas de entrada.
    """
    if not TEMPLATE_COLETA_OFICIAL.exists():
        raise FileNotFoundError(
            f"Template oficial nao encontrado: {TEMPLATE_COLETA_OFICIAL}"
        )
    wb = load_workbook(TEMPLATE_COLETA_OFICIAL, data_only=False)

    # 29C.1: a nova visao operacional e criada em runtime. Isso preserva o
    # template binario homologado e impede que artefatos XLSX entrem no Git.
    from _ciclo_em_execucao import garantir_aba_ciclo_em_execucao
    garantir_aba_ciclo_em_execucao(wb, limpar_entradas=True)

    faltantes = [a for a in ABAS_COLETA_OFICIAL if a not in wb.sheetnames]
    if faltantes:
        raise ValueError(
            f"Template oficial invalido; abas ausentes: {', '.join(faltantes)}"
        )

    _limpar_residuos(wb)
    _garantir_dropdown_metodo(wb)
    _garantir_formulas_cronologia_execucao(wb)
    _garantir_fator_historico_desacoplado(wb)
    _garantir_completude_abertura_temporal(wb)
    _garantir_base_fisica_c0_temporal(wb)
    _garantir_cinza_nao_aplicavel_visivel(wb)
    _garantir_orientacao_novo_item_por_aditivo(wb)
    _validar_estrutura_itens_pc(wb)
    _garantir_apresentacao_retroativos_e_aditivos(wb)

    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    saida = BytesIO()
    wb.save(saida)
    return saida.getvalue()


def _data(valor: Any) -> date | None:
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    texto = str(valor or "").strip()
    for formato in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%Y"):
        try:
            return datetime.strptime(texto, formato).date()
        except ValueError:
            continue
    return None


def _numero_ciclo(valor: Any) -> int | None:
    match = re.search(r"\bC\s*([0-4])\b", str(valor or "").upper())
    return int(match.group(1)) if match else None


def _percentual(ciclo: dict[str, Any]) -> float | None:
    for chave in ("percentual_aplicado", "percentual_indice", "percentual", "variacao"):
        valor = ciclo.get(chave)
        if valor in (None, "") or isinstance(valor, bool):
            continue
        try:
            numero = float(valor)
        except (TypeError, ValueError):
            continue
        return numero / 100 if abs(numero) > 1 else numero
    try:
        fator = float(ciclo.get("fator"))
    except (TypeError, ValueError):
        return None
    return fator - 1 if fator >= 0.5 else fator


def normalizar_dados_calculadora(dados: dict[str, Any] | None) -> dict[str, Any]:
    """Adapta o estado atual das Calculadoras ao gerador do modelo oficial."""
    origem = dict(dados or {})
    fornecidos: dict[int, dict[str, Any]] = {}
    for bruto in origem.get("ciclos") or []:
        if not isinstance(bruto, dict):
            continue
        numero = _numero_ciclo(bruto.get("ciclo") or bruto.get("Ciclo"))
        if numero is not None and numero > 0:
            fornecidos[numero] = bruto
    if not fornecidos:
        raise ValueError("A Calculadora não informou nenhum ciclo entre C1 e C4.")

    data_base = _data(origem.get("data_base") or origem.get("data_base_original"))
    # JANELA DO REAJUSTE (parametros!C:D) — REGRA PETREA DA ETAPA 31.
    # A ancora inicial e tomada UMA unica vez (marco explicito do ciclo mais
    # antigo, senao a janela do indice desse ciclo, senao a data-base
    # original). O inicio do ciclo seguinte decorre da ancora financeira:
    #
    #   DATA_INICIO(Cn+1) = COMPETENCIA(INICIO_EFEITO_FINANCEIRO(Cn)) + 12m
    #
    # nunca ANTES do marco teorico (DATA_INICIO(Cn) + 12m) — o retardo dos
    # efeitos so POSTERGA o proximo reajuste, jamais o antecipa. Ciclo sem
    # inicio de efeito (ex.: precluso sem acordo) usa o fallback teorico
    # (+12m), sem inventar ancora financeira. TODA janela C:D tem EXATAMENTE
    # 12 competencias: DATA_FIM(Cn) = DATA_INICIO(Cn) + 12m - 1 dia, SEMPRE.
    # O retardo NUNCA alonga o proprio ciclo; ele desloca o INICIO do ciclo
    # seguinte, o que pode abrir um INTERVALO INTENCIONAL entre DATA_FIM(Cn)
    # e DATA_INICIO(Cn+1). Esse intervalo NAO e erro e NAO e preenchido: as
    # competencias nele apenas ainda nao pertencem a nenhuma janela de
    # reajuste — a execucao contratual segue a cronologia fixa de 12 meses
    # (conceito independente; ver calendario cronologico da execucao).
    # A janela do indice (data_base/periodo_inicio dos ciclos seguintes) segue
    # sem participar da materializacao. A DATA DE PAGAMENTO nunca participa.
    ancora_numero: int | None = None
    ancora_inicio: date | None = None
    for numero in sorted(fornecidos):
        inicio = _data(fornecidos[numero].get("data_inicio")
                       or fornecidos[numero].get("inicio_ciclo"))
        if inicio:
            ancora_numero, ancora_inicio = numero, inicio.replace(day=1)
            break
    if ancora_inicio is None:
        # Sem marco explicito, a ancora e a janela do indice do ciclo MAIS
        # ANTIGO informado: so nele a janela coincide com o aniversario
        # contratual. A janela dos ciclos SEGUINTES jamais e consultada.
        primeiro = min(fornecidos)
        janela = _data(fornecidos[primeiro].get("data_base")
                       or fornecidos[primeiro].get("periodo_inicio"))
        if janela is not None:
            ancora_numero = primeiro
            ancora_inicio = janela.replace(day=1) + relativedelta(months=12)
        elif data_base is not None:
            ancora_numero = primeiro
            ancora_inicio = (
                data_base.replace(day=1) + relativedelta(months=12 * primeiro)
            )
    if ancora_inicio is None:
        raise ValueError("Não foi possível identificar a data-base dos ciclos da Calculadora.")

    ultimo = max(fornecidos)

    def _efeito_competencia(numero: int) -> date | None:
        """Inicio dos efeitos financeiros do ciclo, normalizado a competencia.

        O dia exato do pedido pertence a admissibilidade; para o regime
        financeiro e para o interregno vale o PRIMEIRO DIA do mes.
        """
        bruto = fornecidos.get(numero) or {}
        efeito = _data(
            bruto.get("inicio_efeito_financeiro")
            or bruto.get("financeiro_inicio")
            or bruto.get("inicio_financeiro")
        )
        return efeito.replace(day=1) if efeito else None

    inicios: dict[int, date] = {ancora_numero: ancora_inicio}
    for numero in range(ancora_numero + 1, ultimo + 1):
        teorico = inicios[numero - 1] + relativedelta(months=12)
        efeito_anterior = _efeito_competencia(numero - 1)
        candidato = (
            efeito_anterior + relativedelta(months=12)
            if efeito_anterior is not None else teorico
        )
        inicios[numero] = max(candidato, teorico)
    for numero in range(ancora_numero - 1, -1, -1):
        # Ciclos anteriores a ancora nao carregam inicio de efeito proprio:
        # blocos teoricos de 12 meses para tras (comportamento preservado).
        inicios[numero] = inicios[numero + 1] - relativedelta(months=12)

    ciclos = []
    for numero in range(1, ultimo + 1):
        bruto = fornecidos.get(numero, {})
        inicio = inicios[numero]
        # Janela de EXATAMENTE 12 competencias; eventual retardo do reajuste
        # seguinte abre um intervalo intencional, jamais alonga esta janela.
        fim = inicio + relativedelta(months=12) - relativedelta(days=1)
        objeto_atual = bool(
            numero in fornecidos
            and bruto.get("objeto_analise_atual", not bruto.get("ciclo_ja_concedido", False))
        )
        ciclos.append({
            **bruto,
            "ciclo": f"C{numero}",
            "data_inicio": inicio,
            "data_fim": fim,
            "data_pedido": _data(bruto.get("data_pedido")),
            # ETAPA 48 — fotografia fisica exata ja decidida pela
            # Calculadora. Viaja em paralelo a cadeia mensal: preserva
            # dia/mes/ano e NUNCA sofre replace(day=1) nem realimenta
            # data_inicio/janelas. Payload sem o campo -> None (o gerador
            # aplica o fallback para data_inicio ao preencher parametros!I).
            "data_abertura_fisica_exata": _data(bruto.get("data_abertura_fisica_exata")),
            # Data final ja decidida pela Calculadora, normalizada a
            # COMPETENCIA (dia 1). O gerador apenas a propaga; nao recria
            # tempestividade, negociacao ou excecoes.
            "inicio_efeito_financeiro": _efeito_competencia(numero),
            "percentual": _percentual(bruto),
            "possui_efeito_financeiro": "Sim" if objeto_atual else "Não",
            "situacao": bruto.get("situacao_aplicada") or bruto.get("situacao") or "",
            # Apresentacao TEMPESTIVO*: flag interna (a exibicao usa o
            # asterisco; formulas e motores nunca dependem dele).
            "efeito_financeiro_retardado": bool(
                bruto.get("efeito_financeiro_retardado")
            ),
        })

    marco_inicial = data_base or inicios.get(0) or min(inicios.values())
    data_corte = _data(origem.get("data_corte")) or max(c["data_fim"] for c in ciclos)
    return {
        **origem,
        "ok": True,
        "modo_origem": origem.get("modo_origem") or origem.get("origem") or origem.get("tipo") or "Calculadora",
        "data_base": marco_inicial,
        "data_corte": data_corte,
        "ciclo_vigente": f"C{ultimo}",
        "ciclos": ciclos,
    }


# Legibilidade da mensagem fiscal de novo item (aditivos!M): dimensoes minimas
# funcionais; dimensoes maiores ja presentes no arquivo nunca sao reduzidas.
LARGURA_MINIMA_ADITIVOS_M = 45.0
ALTURA_MINIMA_LINHAS_ADITIVOS = 45.0
ULTIMA_LINHA_ORIENTACAO_ADITIVOS = 200


def garantir_formatacao_orientacao_aditivos(wb) -> None:
    """Garante a mensagem fiscal integralmente legivel em aditivos!M.

    Reaplica WrapText e dimensoes minimas no workbook FINAL gerado/migrado,
    protegendo contra linhagens de arquivo anteriores a 26H.2 em que a
    formatacao do template nao existia.
    """
    if "aditivos" not in wb.sheetnames:
        return
    ws = wb["aditivos"]
    dim = ws.column_dimensions["M"]
    if not dim.width or dim.width < LARGURA_MINIMA_ADITIVOS_M:
        dim.width = LARGURA_MINIMA_ADITIVOS_M
    from copy import copy as _copy
    for row in range(2, ULTIMA_LINHA_ORIENTACAO_ADITIVOS + 1):
        celula = ws.cell(row=row, column=13)
        if not celula.alignment.wrap_text:
            alinhamento = _copy(celula.alignment)
            alinhamento.wrapText = True
            celula.alignment = alinhamento
        altura = ws.row_dimensions[row]
        if not altura.height or altura.height < ALTURA_MINIMA_LINHAS_ADITIVOS:
            altura.height = ALTURA_MINIMA_LINHAS_ADITIVOS


def gerar_coleta_oficial_preenchida(dados_calculadora: dict[str, Any] | None) -> bytes:
    """Gera o XLS oficial; preenche marcos quando a Calculadora tem dados."""
    base = obter_coleta_oficial_bytes()
    if not (dados_calculadora or {}).get("ciclos"):
        return base
    from _gerador_masterfile import gerar_masterfile_preenchido

    return gerar_masterfile_preenchido(
        normalizar_dados_calculadora(dados_calculadora),
        base,
    )
