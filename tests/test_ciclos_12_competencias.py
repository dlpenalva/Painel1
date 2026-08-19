# -*- coding: utf-8 -*-
"""Regressao permanente do calendario de ciclos e do enquadramento temporal.

REGRA PETREA — INTERREGNO E EFEITOS (atualizada na ETAPA 31):
  * pagamento nunca e marco temporal;
  * admissibilidade usa a data exata do pedido;
  * efeito financeiro usa a COMPETENCIA mensal;
  * tempestivo no mesmo mes da data-base = efeito desde aquela competencia;
  * tempestivo em mes posterior dentro da janela = TEMPESTIVO* (apresentacao)
    e efeito desde a competencia do pedido;
  * o proximo reajuste nasce 12 meses apos o INICIO DOS EFEITOS FINANCEIROS
    do reajuste anterior — a janela do ciclo anterior NAO se alonga (12
    competencias EXATAS, sempre); o retardo abre um INTERVALO INTENCIONAL
    entre as janelas de reajuste (parametros!C:D), que e permitido;
  * cada janela de reajuste tem EXATAMENTE 12 competencias mensais
    consecutivas;
  * a EXECUCAO (financeiro, PC, aditivos) enquadra pela CRONOLOGIA FIXA de
    blocos de 12 meses (5 x 12 = 60 competencias, sem lacuna) — nenhum fato
    valido fica sem ciclo por causa do intervalo entre janelas.

Dentro de um ciclo, o INICIO_EFEITO_FINANCEIRO decide somente a partir de
qual competencia DAQUELE ciclo o reajuste produz efeitos (gate G/H da aba
financeiro); competencias anteriores pertencem ao ciclo sem efeito.

Cobertura (itens 1 a 38 do enunciado da correcao):

   1 cada ciclo com exatamente 12 competencias
   2 C1 = 01/01/2025 a 31/12/2025
   3 C2 = 01/01/2026 a 31/12/2026
   4 C3 = 01/01/2027 a 31/12/2027
   5 jan/fev de C1 pertencem a C1, efeito Nao
   6 jan/fev de C2 pertencem a C2, efeito Nao
   7 marco de C1 pertence a C1, efeito Sim
   8 marco de C2 pertence a C2, efeito Sim
   9 INICIO_EFEITO_FINANCEIRO nao altera DATA_INICIO
  10 INICIO_EFEITO_FINANCEIRO nao altera DATA_FIM
  11 INICIO_EFEITO_FINANCEIRO nao desloca o ciclo seguinte
  12 nenhuma competencia "Fora dos ciclos" entre ciclos consecutivos
  13 postergacao de 1 mes
  14 postergacao de 2 meses
  15 postergacao de 11 meses
  16 pedido no ultimo mes do ciclo (11 sem efeito, 1 com efeito)
  17 virada de ano
  18 fevereiro em ano bissexto
  19 caso sem postergacao
  20 PC em 31/12 de um ciclo
  21 PC em 01/01 do ciclo seguinte
  22 PC anterior ao inicio do efeito
  23 PC exatamente na data do inicio do efeito
  24 PC posterior a data de corte
  25 total enquadrado do Cenario B = 200.000,00
  26 total sem efeito do Cenario B = 40.000,00
  27 total com efeito do Cenario B = 160.000,00
  28 retroativo do Cenario B = 9.165,28
  29 metodo PC
  30 metodo Financeiro
  31 metodo Itens nao sofre deslocamento
  32 posicao fisica nao sofre deslocamento
  33 aditivo posterior a abertura nao retroage
  34 VTA sem dupla contagem dos PCs do ciclo atual
  35 RESULTADOS e MEMORIA_RESULTADOS apresentam a mesma verdade
  36 quatro metricas coerentes
  37 seis documentos coerentes
  38 modelo em branco permanece neutro
"""
from __future__ import annotations

import sys
from datetime import date, timedelta
from pathlib import Path

import openpyxl
import pytest

RAIZ = Path(__file__).resolve().parents[1]
if str(RAIZ) not in sys.path:
    sys.path.insert(0, str(RAIZ))
if str(RAIZ / "tests") not in sys.path:
    sys.path.insert(0, str(RAIZ / "tests"))

from _coleta_oficial import normalizar_dados_calculadora  # noqa: E402
from _efeitos_financeiros_pc import efeito_financeiro_pc  # noqa: E402
from _gerador_masterfile import (  # noqa: E402
    _completar_periodos_ciclos,
    _preencher_financeiro,
)
from _leitor_masterfile_v10 import (  # noqa: E402
    _materializar_aberturas_temporais,
    _pc_dentro_do_corte,
    _totais_canonicos_pc,
)
from _motor_composicao_vta import montar_composicao_vta  # noqa: E402
from _motor_posicao_contratual import (  # noqa: E402
    MESES_POR_CICLO,
    calendario_ciclos_contratuais,
    competencias_do_ciclo,
    divergencias_calendario_canonico,
    somar_meses,
)
from _motor_temporal import (  # noqa: E402
    ENQ_CICLO,
    ENQ_INDETERMINADO,
    ENQ_INTERVALO_PRECLUSO,
    classificar_enquadramento_pc,
)
from _objeto_processo_reajuste import _montar_contrato  # noqa: E402

from test_enquadramento_pc_vta_cenario1 import (  # noqa: E402
    POR_CICLO_CANONICO,
    POR_CICLO_LACUNA,
    VALOR_PC,
    _fator,
    _leitura_vta,
    _pc,
    _vinte_pcs,
)

TEMPLATE = RAIZ / "templates" / "COLETA_REAJUSTE_OFICIAL.xlsx"
ANCORA_C1 = date(2025, 1, 1)
CORTE_A = date(2026, 8, 5)
CORTE_B = date(2026, 8, 31)


def _calendario(ancora: date = ANCORA_C1) -> dict:
    return calendario_ciclos_contratuais(ancora)


def _por_ciclo(ancora: date, inicio_efeito: dict[str, date] | None = None) -> dict:
    """Linha temporal no formato do leitor, com fatores e inicios de efeito."""
    cal = _calendario(ancora)
    saida = {}
    for nome, reg in cal.items():
        saida[nome] = {
            "data_inicio": reg["data_inicio"],
            "data_fim": reg["data_fim"],
            "fator_acumulado": 1.0 if nome == "C0" else 1.1,
            "computar_nesta_apuracao": "Nao" if nome == "C0" else "Sim",
        }
        if inicio_efeito and nome in inicio_efeito:
            saida[nome]["inicio_efeito_financeiro"] = inicio_efeito[nome]
    return saida


# --------------------------------------------------------------------------- #
# 1-4, 17-19 — o calendario canonico
# --------------------------------------------------------------------------- #
@pytest.mark.parametrize("ciclo", ["C0", "C1", "C2", "C3", "C4"])
def test_01_cada_ciclo_tem_exatamente_12_competencias(ciclo):
    cal = _calendario()
    competencias = competencias_do_ciclo(cal[ciclo]["data_inicio"])
    assert len(competencias) == MESES_POR_CICLO == 12
    assert competencias[0] == cal[ciclo]["data_inicio"]
    assert competencias[-1] <= cal[ciclo]["data_fim"]
    assert somar_meses(competencias[-1], 1) > cal[ciclo]["data_fim"]


def test_02_c1_vai_de_01_01_2025_a_31_12_2025():
    cal = _calendario()
    assert cal["C1"]["data_inicio"] == date(2025, 1, 1)
    assert cal["C1"]["data_fim"] == date(2025, 12, 31)


def test_03_c2_vai_de_01_01_2026_a_31_12_2026():
    cal = _calendario()
    assert cal["C2"]["data_inicio"] == date(2026, 1, 1)
    assert cal["C2"]["data_fim"] == date(2026, 12, 31)
    # O periodo rejeitado pelo enunciado nao pode reaparecer.
    assert (cal["C2"]["data_inicio"], cal["C2"]["data_fim"]) != (
        date(2026, 3, 1), date(2027, 2, 28)
    )


def test_04_c3_vai_de_01_01_2027_a_31_12_2027():
    cal = _calendario()
    assert cal["C3"]["data_inicio"] == date(2027, 1, 1)
    assert cal["C3"]["data_fim"] == date(2027, 12, 31)
    assert cal["C4"]["data_inicio"] == date(2028, 1, 1)
    assert cal["C4"]["data_fim"] == date(2028, 12, 31)


def test_17_virada_de_ano_preserva_os_12_meses():
    """Ancora em dezembro: o ciclo atravessa o ano sem perder competencia."""
    cal = calendario_ciclos_contratuais(date(2025, 12, 1))
    assert cal["C1"]["data_inicio"] == date(2025, 12, 1)
    assert cal["C1"]["data_fim"] == date(2026, 11, 30)
    assert len(competencias_do_ciclo(cal["C1"]["data_inicio"])) == 12
    assert cal["C2"]["data_inicio"] == date(2026, 12, 1)
    assert divergencias_calendario_canonico(list(cal.values())) == []


def test_18_fevereiro_em_ano_bissexto():
    """Ancora em marco/2027: C1 fecha em 29/02/2028 (bissexto), 12 meses."""
    cal = calendario_ciclos_contratuais(date(2027, 3, 1))
    assert cal["C1"]["data_fim"] == date(2028, 2, 29)
    assert len(competencias_do_ciclo(cal["C1"]["data_inicio"])) == 12
    assert cal["C2"]["data_inicio"] == date(2028, 3, 1)
    assert cal["C2"]["data_fim"] == date(2029, 2, 28)
    assert divergencias_calendario_canonico(list(cal.values())) == []


def test_19_caso_sem_postergacao_mantem_12_meses_com_efeito():
    """Efeito no primeiro dia do ciclo: 12 competencias, todas com efeito."""
    por_ciclo = _por_ciclo(ANCORA_C1, {"C1": date(2025, 1, 1)})
    competencias = competencias_do_ciclo(por_ciclo["C1"]["data_inicio"])
    efeitos = [efeito_financeiro_pc(c, "C1", por_ciclo["C1"]) for c in competencias]
    assert len(competencias) == 12
    assert efeitos == ["Sim"] * 12


# --------------------------------------------------------------------------- #
# 9-11 — INICIO_EFEITO_FINANCEIRO nao mexe no calendario
# --------------------------------------------------------------------------- #
@pytest.mark.parametrize("meses", [0, 1, 2, 6, 11])
def test_09_10_11_inicio_do_efeito_nao_desloca_o_calendario(meses):
    """Itens 9, 10 e 11 de uma vez: inicio, fim e ciclo seguinte intactos."""
    base = _calendario()
    inicio_efeito = somar_meses(base["C1"]["data_inicio"], meses)
    por_ciclo = _por_ciclo(ANCORA_C1, {"C1": inicio_efeito})
    assert por_ciclo["C1"]["data_inicio"] == base["C1"]["data_inicio"]   # item 9
    assert por_ciclo["C1"]["data_fim"] == base["C1"]["data_fim"]         # item 10
    assert por_ciclo["C2"]["data_inicio"] == base["C2"]["data_inicio"]   # item 11
    assert por_ciclo["C2"]["data_fim"] == base["C2"]["data_fim"]
    assert divergencias_calendario_canonico(list(por_ciclo.values())) == []


def test_12_nenhuma_competencia_fica_fora_entre_ciclos_consecutivos():
    """Item 12: as 60 competencias de C0..C4 estao todas dentro de um ciclo."""
    por_ciclo = _por_ciclo(ANCORA_C1, {"C1": date(2025, 3, 1), "C2": date(2026, 3, 1)})
    vistas = []
    for nome in ("C0", "C1", "C2", "C3", "C4"):
        vistas.extend(competencias_do_ciclo(por_ciclo[nome]["data_inicio"]))
    assert len(vistas) == 60
    assert len(set(vistas)) == 60          # sem sobreposicao
    for competencia in vistas:
        enq = classificar_enquadramento_pc(competencia, por_ciclo)
        assert enq.tipo == ENQ_CICLO, competencia
        assert enq.ciclo is not None
    # A sequencia e continua: nenhum buraco de um mes.
    ordenadas = sorted(vistas)
    for anterior, seguinte in zip(ordenadas, ordenadas[1:]):
        assert somar_meses(anterior, 1) == seguinte


# --------------------------------------------------------------------------- #
# 13-16 — postergacao do inicio do efeito financeiro
# --------------------------------------------------------------------------- #
@pytest.mark.parametrize("postergacao", [1, 2, 11])
def test_13_14_15_postergacao_reduz_so_o_efeito(postergacao):
    """Itens 13, 14 e 15: o ciclo continua com 12 competencias."""
    inicio_efeito = somar_meses(ANCORA_C1, postergacao)
    por_ciclo = _por_ciclo(ANCORA_C1, {"C1": inicio_efeito})
    competencias = competencias_do_ciclo(por_ciclo["C1"]["data_inicio"])
    efeitos = [efeito_financeiro_pc(c, "C1", por_ciclo["C1"]) for c in competencias]
    assert len(competencias) == 12
    assert efeitos.count("Nao") == postergacao
    assert efeitos.count("Sim") == 12 - postergacao
    assert efeitos[:postergacao] == ["Nao"] * postergacao
    assert por_ciclo["C2"]["data_inicio"] == date(2026, 1, 1)


def test_16_pedido_no_ultimo_mes_do_ciclo():
    """Item 16: 11 competencias sem efeito e exatamente 1 com efeito."""
    por_ciclo = _por_ciclo(ANCORA_C1, {"C1": date(2025, 12, 1)})
    competencias = competencias_do_ciclo(por_ciclo["C1"]["data_inicio"])
    efeitos = [efeito_financeiro_pc(c, "C1", por_ciclo["C1"]) for c in competencias]
    assert efeitos.count("Nao") == 11
    assert efeitos.count("Sim") == 1
    assert efeitos[-1] == "Sim"
    assert por_ciclo["C1"]["data_fim"] == date(2025, 12, 31)
    assert por_ciclo["C2"]["data_inicio"] == date(2026, 1, 1)


# --------------------------------------------------------------------------- #
# 5-8, 20-24 — enquadramento de PCs
# --------------------------------------------------------------------------- #
@pytest.mark.parametrize("data_pc", [date(2025, 1, 15), date(2025, 2, 15)])
def test_05_jan_fev_de_c1_pertencem_a_c1_sem_efeito(data_pc):
    pc = _pc("PC-X", data_pc)
    assert pc["ciclo"] == "C1"
    assert pc["enquadramento"] == ENQ_CICLO
    assert pc["efeito_financeiro_pc"] == "Nao"
    assert pc["fator_efetivo_considerado"] == 1.0
    assert pc["retroativo_reconhecido_a_pagar"] == 0.0


@pytest.mark.parametrize("data_pc", [date(2026, 1, 15), date(2026, 2, 15)])
def test_06_jan_fev_de_c2_pertencem_a_c2_sem_efeito(data_pc):
    pc = _pc("PC-X", data_pc)
    assert pc["ciclo"] == "C2"
    assert pc["enquadramento"] == ENQ_CICLO
    assert pc["efeito_financeiro_pc"] == "Nao"
    assert pc["fator_efetivo_considerado"] == 1.0
    assert pc["retroativo_reconhecido_a_pagar"] == 0.0
    assert pc["valor_historico_considerado"] == pytest.approx(VALOR_PC, abs=0.01)


def test_07_marco_de_c1_pertence_a_c1_com_efeito():
    pc = _pc("PC-X", date(2025, 3, 15))
    assert pc["ciclo"] == "C1"
    assert pc["efeito_financeiro_pc"] == "Sim"
    assert pc["retroativo_reconhecido_a_pagar"] > 0.0


def test_08_marco_de_c2_pertence_a_c2_com_efeito():
    pc = _pc("PC-X", date(2026, 3, 15))
    assert pc["ciclo"] == "C2"
    assert pc["efeito_financeiro_pc"] == "Sim"
    assert pc["retroativo_reconhecido_a_pagar"] > 0.0


def test_20_pc_em_31_12_fecha_o_ciclo():
    assert classificar_enquadramento_pc(
        date(2025, 12, 31), POR_CICLO_CANONICO).ciclo == "C1"


def test_21_pc_em_01_01_abre_o_ciclo_seguinte():
    assert classificar_enquadramento_pc(
        date(2026, 1, 1), POR_CICLO_CANONICO).ciclo == "C2"


def test_22_pc_anterior_ao_inicio_do_efeito():
    reg = POR_CICLO_CANONICO["C2"]
    assert efeito_financeiro_pc(date(2026, 2, 28), "C2", reg) == "Nao"


def test_23_pc_exatamente_na_data_do_inicio_do_efeito():
    reg = POR_CICLO_CANONICO["C2"]
    assert efeito_financeiro_pc(date(2026, 3, 1), "C2", reg) == "Sim"


def test_24_pc_posterior_a_data_de_corte():
    assert _pc_dentro_do_corte(date(2026, 8, 15), CORTE_A) is False
    assert _pc_dentro_do_corte(date(2026, 8, 15), CORTE_B) is True
    tc = _totais_canonicos_pc(_vinte_pcs(data_corte=CORTE_A), data_corte=CORTE_A)
    assert tc["posterior_ao_corte"]["quantidade"] == 1
    assert tc["ate_o_corte"]["quantidade"] == 19
    # Permanece no inventario do arquivo.
    assert tc["informado"]["quantidade"] == 20


# --------------------------------------------------------------------------- #
# 25-28 — totais canonicos do Cenario B
# --------------------------------------------------------------------------- #
def _totais_cenario_b():
    return _totais_canonicos_pc(_vinte_pcs(data_corte=CORTE_B), data_corte=CORTE_B)


def test_25_total_enquadrado_do_cenario_b():
    assert _totais_cenario_b()["enquadrado_ciclos"]["valor_pc"] == pytest.approx(
        200_000.00, abs=0.01
    )


def test_26_total_sem_efeito_do_cenario_b():
    bloco = _totais_cenario_b()["sem_efeito_ciclo"]
    assert bloco["valor_pc"] == pytest.approx(40_000.00, abs=0.01)
    assert bloco["retroativo"] == 0.0


def test_27_total_com_efeito_do_cenario_b():
    assert _totais_cenario_b()["com_efeito"]["valor_pc"] == pytest.approx(
        160_000.00, abs=0.01
    )


def test_28_retroativo_do_cenario_b():
    esperado = round(
        10 * (round(VALOR_PC * _fator("C1"), 2) - VALOR_PC)
        + 6 * (round(VALOR_PC * _fator("C2"), 2) - VALOR_PC), 2
    )
    assert esperado == pytest.approx(9_165.28, abs=0.01)
    assert _totais_cenario_b()["ate_o_corte"]["retroativo"] == pytest.approx(
        9_165.28, abs=0.01
    )


def test_28b_retroativo_do_cenario_a():
    tc = _totais_canonicos_pc(_vinte_pcs(data_corte=CORTE_A), data_corte=CORTE_A)
    assert tc["ate_o_corte"]["retroativo"] == pytest.approx(8_386.30, abs=0.01)
    assert tc["enquadrado_ciclos"]["valor_pc"] == pytest.approx(190_000.00, abs=0.01)
    assert tc["com_efeito"]["valor_pc"] == pytest.approx(150_000.00, abs=0.01)
    assert tc["sem_efeito_ciclo"]["valor_pc"] == pytest.approx(40_000.00, abs=0.01)


# --------------------------------------------------------------------------- #
# 29-32 — os metodos
# --------------------------------------------------------------------------- #
def test_29_metodo_pc_usa_o_valor_considerado():
    comp = montar_composicao_vta(_leitura_vta(fisico=True))
    c1 = next(l for l in comp["execucao_por_ciclo"] if l["ciclo"] == "C1")
    considerado = round(2 * VALOR_PC + 10 * round(VALOR_PC * _fator("C1"), 2), 2)
    assert c1["valor_atualizado"] == pytest.approx(considerado, abs=0.02)
    assert c1["fonte"] == "pc"


def test_30_metodo_financeiro_gera_12_competencias_por_ciclo():
    """Item 30: a grade financeira materializa 12 competencias por ciclo."""
    cal = _calendario()
    ciclos = {
        nome: {
            **reg,
            "possui_efeito_financeiro": "Sim" if nome in ("C1", "C2") else "Nao",
            "inicio_efeito_financeiro": {
                "C1": date(2025, 3, 1), "C2": date(2026, 3, 1),
            }.get(nome),
        }
        for nome, reg in cal.items()
    }
    wb = openpyxl.Workbook()
    ws = wb.active
    _preencher_financeiro(
        ws, ciclos, data_corte_fallback=cal["C2"]["data_fim"],
        novo=True, marco_inicial=cal["C0"]["data_inicio"],
    )
    por_ciclo: dict[str, list] = {}
    for linha in range(2, 74):
        competencia = ws[f"A{linha}"].value
        if competencia is None:
            continue
        competencia = getattr(competencia, "date", lambda: competencia)()
        for nome, reg in cal.items():
            if reg["data_inicio"] <= competencia <= reg["data_fim"]:
                por_ciclo.setdefault(nome, []).append(
                    (competencia, ws[f"G{linha}"].value)
                )
                break
        else:  # pragma: no cover - falha explicita
            raise AssertionError(f"competencia {competencia} fora dos ciclos")
    assert [len(por_ciclo[c]) for c in ("C0", "C1", "C2")] == [12, 12, 12]
    efeitos_c2 = [e for _, e in sorted(por_ciclo["C2"])]
    assert efeitos_c2[:2] == ["Nao", "Nao"]       # jan e fev de 2026
    assert efeitos_c2[2:] == ["Sim"] * 10


def test_31_metodo_itens_nao_sofre_deslocamento():
    """Item 31: as aberturas por item sao enderecadas por ROTULO de ciclo."""
    registro = {
        "ITEM": 3, "VU_ORIGINAL": 200,
        "QTD_REM_AJUSTADA_C0": 60, "DELTA_POSTERIOR_ABERTURA_C0": 0,
        "QTD_REM_AJUSTADA_C1": 70, "DELTA_POSTERIOR_ABERTURA_C1": 20,
        "QTD_REM_AJUSTADA_C2": 52, "DELTA_POSTERIOR_ABERTURA_C2": 0,
    }
    copia = dict(registro)
    _materializar_aberturas_temporais(registro)
    _materializar_aberturas_temporais(copia)
    # A materializacao nao consome calendario algum: e invariante a datas.
    assert registro == copia
    assert registro["QTD_REM_ABERTURA_C1"] == 50
    assert registro["QTD_REM_ABERTURA_C2"] == 52


def test_32_posicao_fisica_nao_sofre_deslocamento():
    """Item 32: a posicao fisica e enderecada pelo ciclo vigente, nao por data."""
    leitura = _leitura_vta(fisico=True)
    comp = montar_composicao_vta(leitura)
    fisico = next(
        l for l in comp["execucao_por_ciclo"] if l["fonte"] == "posicao_fisica"
    )
    assert fisico["valor_atualizado"] == pytest.approx(3018.12, abs=0.01)
    assert comp["saldo_remanescente"]["valor_atualizado"] == pytest.approx(
        8192.04, abs=0.01
    )
    # Deslocar o calendario nao muda a posicao fisica informada pelo fiscal.
    leitura_deslocada = dict(leitura)
    leitura_deslocada["parametros_v10"] = {"por_ciclo": POR_CICLO_LACUNA}
    comp2 = montar_composicao_vta(leitura_deslocada)
    fisico2 = next(
        l for l in comp2["execucao_por_ciclo"] if l["fonte"] == "posicao_fisica"
    )
    assert fisico2["valor_atualizado"] == fisico["valor_atualizado"]
    assert comp2["saldo_remanescente"]["valor_atualizado"] == pytest.approx(
        8192.04, abs=0.01
    )


# --------------------------------------------------------------------------- #
# 33-34 — aditivos e dupla contagem
# --------------------------------------------------------------------------- #
def test_33_aditivo_posterior_a_abertura_nao_retroage():
    registro = {
        "QTD_REM_AJUSTADA_C1": 70, "DELTA_POSTERIOR_ABERTURA_C1": 20,
        "QTD_REM_AJUSTADA_C2": 95, "DELTA_POSTERIOR_ABERTURA_C2": -15,
    }
    _materializar_aberturas_temporais(registro)
    assert registro["QTD_REM_ABERTURA_C1"] == 50    # acrescimo nao retroage
    assert registro["QTD_REM_ABERTURA_C2"] == 110   # supressao nao retroage


def test_34_vta_sem_dupla_contagem_dos_pcs_do_ciclo_atual():
    comp = montar_composicao_vta(_leitura_vta(fisico=True))
    fontes = [l["fonte"] for l in comp["execucao_por_ciclo"]]
    assert "pc_intervalo_precluso" not in fontes
    assert not any(
        l["ciclo"] == "C2" and l["fonte"].startswith("pc")
        for l in comp["execucao_por_ciclo"]
    )
    assert any("dupla contagem" in a for a in comp["alertas"])
    soma = sum(l["valor_atualizado"] for l in comp["execucao_por_ciclo"])
    soma += comp["saldo_remanescente"]["valor_atualizado"]
    assert comp["vta_composicao"] == pytest.approx(soma, abs=0.02)


# --------------------------------------------------------------------------- #
# 35 — RESULTADOS e MEMORIA_RESULTADOS com a mesma verdade
# --------------------------------------------------------------------------- #
def test_35_resultados_e_memoria_leem_a_mesma_fonte():
    """As celulas oficiais de RESULTADOS apontam para a MEMORIA, sem recalculo."""
    wb = openpyxl.load_workbook(TEMPLATE, data_only=False)
    res, mem = wb["RESULTADOS"], wb["MEMORIA_RESULTADOS"]
    assert "MEMORIA_RESULTADOS!$W$50" in str(res["B10"].value)
    assert "MEMORIA_RESULTADOS!$W$48" in str(res["B11"].value)
    assert "MEMORIA_RESULTADOS!$W$51" in str(res["B13"].value)
    assert "MEMORIA_RESULTADOS!$W$52" in str(res["H13"].value)
    # O VTA da posicao atual soma execucao historica + posicao fisica do ciclo.
    # HOTFIX RETRO/VTA: o historico e method-aware (W66/W67); o ramo PC/Itens
    # preserva T21+T22 e o Financeiro usa financeiro!E (W61:W65).
    w50 = str(mem["W50"].value)
    assert "$W$66" in w50
    assert "CICLO_EM_EXECUCAO!F13:F211" in w50
    assert "CICLO_EM_EXECUCAO!G13:G211" in w50
    assert "$W$67" in str(mem["W48"].value)
    w66, w67 = str(mem["W66"].value), str(mem["W67"].value)
    assert "$T$21" in w66 and "$T$22" in w66 and '"Financeiro"' in w66
    # Execucao historica pelo VALOR CONSIDERADO (base + retroativo), nunca pelo
    # VALOR_ATUALIZADO com fator integral.
    for endereco, formula in (("T21", str(mem["T21"].value)),
                              ("T22", str(mem["T22"].value)),
                              ("W67", w67)):
        assert "itens_PC!$Q$" in formula, endereco
        assert "itens_PC!$P$" not in formula, endereco
    # Data de corte materializada.
    assert "CONTROLE!$B$3" in str(mem["T31"].value)
    for endereco in ("T34", "T35", "T36"):
        assert "$T$31" in str(mem[endereco].value), endereco


def test_35b_itens_pc_publica_o_valor_considerado():
    wb = openpyxl.load_workbook(TEMPLATE, data_only=False)
    ws = wb["itens_PC"]
    assert ws["U1"].value == "VALOR_CONSIDERADO"
    assert "$D2" in str(ws["U2"].value) and "$H2" in str(ws["U2"].value)


def test_35c_resultados_publica_as_doze_medidas():
    wb = openpyxl.load_workbook(TEMPLATE, data_only=False)
    res = wb["RESULTADOS"]
    rotulos = [
        str(res.cell(linha, 1).value or "")
        for linha in range(55, 67)
    ]
    esperados = [
        "Total cadastrado de PCs",
        "Total considerado ate a data de corte",
        "Total enquadrado nos ciclos",
        "Total com efeito financeiro",
        "Total sem efeito financeiro",
        "Retroativo reconhecido",
        "Execucao fisica do ciclo atual",
        "Remanescente fisico atual",
        "VTA oficial",
        "Referencias auditaveis",
        "Reconciliacao",
        "Status",
    ]
    for i, esperado in enumerate(esperados):
        assert rotulos[i] == f"{i + 1}. {esperado}", rotulos[i]


# --------------------------------------------------------------------------- #
# 36 — resultado consolidado na web
# --------------------------------------------------------------------------- #
def test_36_resultado_consolidado_preserva_payload_canonico():
    """A nova apresentação usa o payload canônico e remove os quatro boxes."""
    fonte = (RAIZ / "pages" / "03_Valor_Global.py").read_text(encoding="utf-8")
    consolidacao = (RAIZ / "_resultado_consolidado.py").read_text(encoding="utf-8")
    assert "render_resultado_consolidado(resultado, diagnostico_coleta)" in fonte
    assert "Valor Total Atualizado — VTA" in fonte
    assert "Retroativo reconhecido" in fonte
    assert "Retroativo potencial" in fonte
    assert 'resumo_indice.metric("Índice"' not in fonte
    assert 'resumo_ciclos.metric("Ciclos analisados"' not in fonte
    # O reconhecido continua lendo a medida canônica "ate_o_corte".
    assert 'totais_pc.get("ate_o_corte")' in consolidacao
    assert 'bloco_corte.get("retroativo")' in consolidacao
    # Valores dos dois cenarios.
    tc_a = _totais_canonicos_pc(_vinte_pcs(data_corte=CORTE_A), data_corte=CORTE_A)
    tc_b = _totais_cenario_b()
    assert tc_a["ate_o_corte"]["retroativo"] == pytest.approx(8_386.30, abs=0.01)
    assert tc_b["ate_o_corte"]["retroativo"] == pytest.approx(9_165.28, abs=0.01)
    # Percentual acumulado identico nos dois cenarios (nao depende do corte).
    acumulado = round((_fator("C2") - 1) * 100, 2)
    assert acumulado == pytest.approx(7.79, abs=0.005)


# --------------------------------------------------------------------------- #
# 37 — seis documentos consumindo o payload canonico
# --------------------------------------------------------------------------- #
def test_37_seis_cards_preservados():
    """Titulos, ordem e quantidade dos seis cards documentais."""
    import re

    fonte = (RAIZ / "pages" / "06_Central_Arquivos.py").read_text(encoding="utf-8")
    nomes = re.findall(r'"nome":\s*"([^"]+)"', fonte)
    assert nomes == [
        "Sumário Executivo",
        "Adequação Orçamentária",
        "Despacho Saneador",
        "Termo de Apostila",
        "Garantia Contratual",
        "DOU",
    ]


def test_37b_documentos_leem_o_retroativo_canonico():
    """O objeto documental prefere a medida canonica ao cache do XLS."""
    leitura = {
        "controle": {"modo": "PC", "ciclo_vigente": "C2"},
        "resumo": {"retroativo": 999_999.99},
        "parametros_v10": {"por_ciclo": POR_CICLO_CANONICO},
        "itens_pc_v10": {
            "totais_canonicos": _totais_canonicos_pc(
                _vinte_pcs(data_corte=CORTE_B), data_corte=CORTE_B
            )
        },
    }
    basicos = _montar_contrato(leitura)["dados_basicos"]
    assert basicos["retroativo_oficial"] == pytest.approx(9_165.28, abs=0.01)
    assert basicos["origem_retroativo_oficial"] == "totais_canonicos_pc.ate_o_corte"
    assert basicos["totais_canonicos_pc"]["enquadrado_ciclos"]["valor_pc"] == \
        pytest.approx(200_000.00, abs=0.01)


def test_37c_sem_payload_canonico_o_xls_permanece_a_fonte():
    """Sem PCs lidos, nada muda: o comportamento anterior e preservado."""
    leitura = {
        "controle": {"modo": "Financeiro"},
        "resumo": {"retroativo": 1234.56},
        "parametros_v10": {"por_ciclo": POR_CICLO_CANONICO},
        "itens_pc_v10": {"totais_canonicos": {}},
    }
    basicos = _montar_contrato(leitura)["dados_basicos"]
    assert basicos["retroativo_oficial"] == pytest.approx(1234.56, abs=0.01)
    assert basicos["origem_retroativo_oficial"] == "resultados_xls"


def test_37d_documental_expoe_o_payload_unico():
    fonte = (RAIZ / "_coleta_reajuste_documentos.py").read_text(encoding="utf-8")
    assert '"totais_canonicos_pc"' in fonte
    assert '"composicao_vta"' in fonte


# --------------------------------------------------------------------------- #
# 38 — modelo em branco neutro
# --------------------------------------------------------------------------- #
def test_38_modelo_em_branco_permanece_neutro():
    from _templates_documentos import (
        gerar_modelo_branco_despacho,
        gerar_modelo_branco_termo,
    )
    for gerar in (gerar_modelo_branco_despacho, gerar_modelo_branco_termo):
        conteudo = gerar()
        assert isinstance(conteudo, (bytes, bytearray))
        assert len(conteudo) > 0
        # Nenhum numero do Cenario 1 pode vazar para o modelo em branco.
        for proibido in (b"200.000,00", b"9.165,28", b"239.293,17", b"8.386,30"):
            assert proibido not in conteudo


# --------------------------------------------------------------------------- #
# Guardas do calendario corrompido (a regressao que esta suite impede)
# --------------------------------------------------------------------------- #
def test_calendario_com_sobreposicao_e_recusado():
    """ETAPA 31: o gap entre janelas e INTENCIONAL e permitido; sobreposicao
    e janela sem 12 competencias exatas continuam recusadas."""
    problemas = divergencias_calendario_canonico([
        {"ciclo": nome, **reg} for nome, reg in POR_CICLO_LACUNA.items()
    ])
    # C2 deslocado para marco: o gap C1->C2 e permitido (nao e problema);
    # a sobreposicao C2->C3 permanece recusada.
    assert [p.split(":")[0] for p in problemas] == ["C2->C3"]
    assert all("sobrepostas" in p for p in problemas)
    # Um periodo que nao some 12 competencias exatas tambem e recusado.
    curto = divergencias_calendario_canonico([
        {"ciclo": "C1", "data_inicio": date(2025, 1, 1),
         "data_fim": date(2025, 10, 31)},
    ])
    assert any("12 competencias consecutivas" in p for p in curto)
    # Janela ALONGADA (13 competencias) tambem e recusada: o retardo do
    # reajuste seguinte NUNCA alonga a janela do ciclo atual.
    longo = divergencias_calendario_canonico([
        {"ciclo": "C1", "data_inicio": date(2025, 1, 1),
         "data_fim": date(2026, 1, 31)},
    ])
    assert any("12 competencias consecutivas" in p for p in longo)


def test_gap_de_janelas_nao_vira_indeterminado_nem_precluso():
    """ETAPA 31: data no gap intencional das janelas enquadra pela
    CRONOLOGIA fixa da execucao — jamais INDETERMINADO ou precluso."""
    enq = classificar_enquadramento_pc(date(2026, 1, 15), POR_CICLO_LACUNA)
    assert enq.tipo == ENQ_CICLO
    assert enq.ciclo == "C2"
    assert enq.tipo != ENQ_INTERVALO_PRECLUSO
    assert enq.tipo != ENQ_INDETERMINADO


def test_gerador_da_coleta_encadeia_pelo_inicio_dos_efeitos():
    """Regra petrea do interregno (ETAPA 31): C(n+1) nasce 12 meses apos o
    INICIO DOS EFEITOS FINANCEIROS de C(n); a janela do indice continua sem
    redefinir ciclo algum; a janela do ciclo anterior NAO se alonga — o
    espaco entre elas e um intervalo intencional."""
    dados = normalizar_dados_calculadora({
        "data_base_original": "01/01/2024",
        "ciclos": [
            # C1 com efeito retardado para marco/2025 (TEMPESTIVO*):
            # C2 nasce em marco/2026; C1 fecha em 31/12/2025 (12 comp.)
            # e o intervalo 01-02/2026 e intencional.
            {"ciclo": "C1", "data_inicio": date(2025, 1, 1),
             "financeiro_inicio": date(2025, 3, 1), "percentual": 0.0449},
            # Janela do indice de C2 (data_base) segue sem participar da
            # materializacao do calendario.
            {"ciclo": "C2", "data_base": date(2025, 3, 1),
             "financeiro_inicio": date(2026, 3, 1), "percentual": 0.0316},
        ],
    })
    por_ciclo = {c["ciclo"]: c for c in dados["ciclos"]}
    assert por_ciclo["C1"]["data_inicio"] == date(2025, 1, 1)
    # Janela de 12 competencias EXATAS: o retardo nunca a alonga.
    assert por_ciclo["C1"]["data_fim"] == date(2025, 12, 31)
    assert por_ciclo["C2"]["data_inicio"] == date(2026, 3, 1)
    # C2 com efeito em 03/2026 (mesma competencia do novo inicio do ciclo):
    # janela de 12 competencias, fechando em 28/02/2027.
    assert por_ciclo["C2"]["data_fim"] == date(2027, 2, 28)
    # O inicio do efeito financeiro segue preservado (competencia, dia 1).
    assert por_ciclo["C2"]["inicio_efeito_financeiro"] == date(2026, 3, 1)
    # Intervalo intencional entre as janelas (01-02/2026): permitido e
    # nunca preenchido artificialmente.
    assert por_ciclo["C2"]["data_inicio"] > por_ciclo["C1"]["data_fim"] + timedelta(days=1)
    assert divergencias_calendario_canonico(list(por_ciclo.values())) == []


def test_gerador_sem_retardo_mantem_blocos_de_12_meses():
    """Cenario D: quando o efeito comeca na competencia prevista, a linha
    temporal permanece identica ao baseline (blocos de 12 meses)."""
    dados = normalizar_dados_calculadora({
        "data_base_original": "01/01/2024",
        "ciclos": [
            {"ciclo": "C1", "data_inicio": date(2025, 1, 1),
             "financeiro_inicio": date(2025, 1, 1), "percentual": 0.0449},
            {"ciclo": "C2", "data_base": date(2024, 1, 1),
             "financeiro_inicio": date(2026, 1, 1), "percentual": 0.0316},
        ],
    })
    por_ciclo = {c["ciclo"]: c for c in dados["ciclos"]}
    assert por_ciclo["C1"]["data_inicio"] == date(2025, 1, 1)
    assert por_ciclo["C1"]["data_fim"] == date(2025, 12, 31)
    assert por_ciclo["C2"]["data_inicio"] == date(2026, 1, 1)
    assert por_ciclo["C2"]["data_fim"] == date(2026, 12, 31)


def test_completar_periodos_preserva_a_contiguidade():
    cal = _calendario()
    completos = _completar_periodos_ciclos({
        "C1": dict(cal["C1"]), "C2": dict(cal["C2"]),
    })
    assert completos["C0"]["data_inicio"] == date(2024, 1, 1)
    assert completos["C0"]["data_fim"] == date(2024, 12, 31)
    assert divergencias_calendario_canonico([
        {"ciclo": nome, **reg} for nome, reg in completos.items()
    ]) == []
