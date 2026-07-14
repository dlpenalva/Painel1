import io
import json
import unittest
import xml.etree.ElementTree as ET
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import load_workbook

from _coleta_reajuste import CAMINHO_MODELO_COLETA, gerar_coleta_reajuste, ler_coleta_reajuste
from _coleta_reajuste_documentos import adaptar_coleta_reajuste_para_documentos


FIXTURE = Path(__file__).parent / "fixtures" / "posicao_contratual.json"
NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_REL_DOC = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
NS_REL_PKG = "http://schemas.openxmlformats.org/package/2006/relationships"


def _injetar_cache_formula(payload: bytes, aba: str, celula: str, valor: str) -> bytes:
    """Injeta apenas o valor em cache, preservando a fórmula do XLS para o teste do leitor."""

    origem = ZipFile(io.BytesIO(payload))
    workbook = ET.fromstring(origem.read("xl/workbook.xml"))
    rels = ET.fromstring(origem.read("xl/_rels/workbook.xml.rels"))
    sheet = next(
        node
        for node in workbook.findall(f".//{{{NS_MAIN}}}sheet")
        if node.attrib["name"] == aba
    )
    rel_id = sheet.attrib[f"{{{NS_REL_DOC}}}id"]
    target = next(
        node.attrib["Target"]
        for node in rels.findall(f"{{{NS_REL_PKG}}}Relationship")
        if node.attrib["Id"] == rel_id
    )
    caminho = target.lstrip("/")
    if not caminho.startswith("xl/"):
        caminho = f"xl/{caminho}"

    sheet_xml = ET.fromstring(origem.read(caminho))
    cell = next(
        node
        for node in sheet_xml.findall(f".//{{{NS_MAIN}}}c")
        if node.attrib.get("r") == celula
    )
    cell.attrib["t"] = "str"
    cached = cell.find(f"{{{NS_MAIN}}}v")
    if cached is None:
        cached = ET.SubElement(cell, f"{{{NS_MAIN}}}v")
    cached.text = valor

    destino = io.BytesIO()
    with ZipFile(destino, "w", ZIP_DEFLATED) as saida:
        for info in origem.infolist():
            conteudo = ET.tostring(sheet_xml, encoding="utf-8", xml_declaration=False) if info.filename == caminho else origem.read(info.filename)
            saida.writestr(info, conteudo)
    origem.close()
    return destino.getvalue()


class PosicaoContratualTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.fixture = json.loads(FIXTURE.read_text(encoding="utf-8"))

    def test_decisoes_de_negocio_estao_explicitas_sem_mudar_a_regra_quantitativa(self):
        wb = load_workbook(CAMINHO_MODELO_COLETA, data_only=False)
        aditivos = wb["aditivos"]
        posicao = wb["posicao_contratual"]
        remanesc = wb["itens_Remanesc"]

        self.assertEqual(aditivos["D1"].value, "TIPO DE ALTERAÇÃO FORMALIZADA")
        self.assertEqual(aditivos["K1"].value, "CONSIDERADO NO CÁLCULO FINANCEIRO? (SIM/NÃO)")
        self.assertNotIn("K", aditivos["L2"].value)
        self.assertNotIn("aditivos!$K$", posicao["L2"].value)
        self.assertIn('aditivos!$C$2:$C$200,"C2"', posicao["L2"].value)
        self.assertIn("BASE 0", aditivos["A1"].value)
        self.assertIn("0 PARA NOVO ITEM FORMALIZADO", remanesc["B1"].value)
        for cell in ("E1", "G1", "I1", "K1"):
            self.assertIn("Somente contrato original; nao somar aditivos", remanesc[cell].value)

    def test_fixture_protege_agregacao_fracoes_e_nascimento_do_item(self):
        casos = self.fixture["casos"]
        multiplo = casos["multiplos_aditivos_mesmo_ciclo"]
        self.assertAlmostEqual(sum(multiplo["eventos_c2"]), multiplo["delta_c2_esperado"])
        fracao = casos["quantidades_fracionarias"]
        self.assertAlmostEqual(fracao["base_original"] + sum(fracao["eventos_c2"]), fracao["posicao_c2_esperada"])
        novo = casos["novo_item"]
        self.assertEqual(novo["base_original"], 0.0)
        self.assertEqual(novo["posicao_c0_a_c4_esperada"], [0.0, 0.0, 5.5, 5.5, 5.5])

        wb = load_workbook(CAMINHO_MODELO_COLETA, data_only=False)
        posicao = wb["posicao_contratual"]
        self.assertIn("SUMIFS(aditivos!$L$2:$L$200", posicao["L2"].value)
        self.assertIn("I2+L2", posicao["M2"].value)
        self.assertIn("M2+P2", posicao["Q2"].value)
        self.assertIn("Q2+T2", posicao["U2"].value)
        self.assertEqual(wb["aditivos"]["L2"].number_format, "#,##0.00")

    def test_alertas_criticos_invalidam_upload_consolidacao_e_oito_documentos(self):
        base = gerar_coleta_reajuste(
            {
                "indice": "INDICE SINTETICO",
                "ciclos": [{"ciclo": "C2", "data_base": "01/01/2025", "percentual_aplicado": 0.05}],
            }
        )
        for chave in ("posicao_negativa", "remanescente_supera_posicao"):
            with self.subTest(chave=chave):
                alerta = self.fixture["casos"][chave]["alerta_esperado"]
                payload = _injetar_cache_formula(base, "posicao_contratual", "X2", alerta)
                diagnostico = ler_coleta_reajuste(payload)
                self.assertFalse(diagnostico["valido"])
                self.assertFalse(diagnostico["pronto_para_consolidar"])
                self.assertTrue(any(alerta in item for item in diagnostico["bloqueios_criticos"]))
                self.assertFalse(diagnostico["capacidades"]["estruturalmente_valido"])
                with self.assertRaisesRegex(ValueError, "não pode liberar documentos"):
                    adaptar_coleta_reajuste_para_documentos(payload)


if __name__ == "__main__":
    unittest.main()
