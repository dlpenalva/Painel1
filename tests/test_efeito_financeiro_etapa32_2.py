# -*- coding: utf-8 -*-
"""Etapa 32.2 — ciclo fixo de 12 competências x efeito financeiro por ciclo."""
from datetime import date
from io import BytesIO

import pytest
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook, load_workbook

from _coleta_oficial import gerar_coleta_oficial_preenchida
from _gerador_masterfile import (
    _defasagem_financeira_competencias,
    _preencher_financeiro,
)


def _payload_caso_real():
    return {
        "origem": "Reajustes Múltiplos",
        "tipo": "Múltiplo",
        "indice": "IPCA",
        "data_base_original": "10/10/2022",
        "ciclos": [
            {
                "ciclo": "C2",
                "data_base": "10/10/2022",
                "data_pedido": "01/11/2023",
                "financeiro_inicio": "01/11/2023",
                "percentual_aplicado": 0.05,
                "objeto_analise_atual": True,
                "situacao": "TEMPESTIVO*",
                "efeito_financeiro_retardado": True,
            },
            {
                "ciclo": "C3",
                "data_base": "01/11/2023",
                "data_pedido": "01/11/2024",
                "financeiro_inicio": "01/11/2024",
                "percentual_aplicado": 0.04,
                "objeto_analise_atual": True,
                "situacao": "TEMPESTIVO",
            },
            {
                "ciclo": "C4",
                "data_base": "01/11/2024",
                "data_pedido": "01/11/2025",
                "financeiro_inicio": "01/11/2025",
                "percentual_aplicado": 0.03,
                "objeto_analise_atual": True,
                "situacao": "TEMPESTIVO",
            },
        ],
    }


@pytest.fixture(scope="module")
def wb_caso_real():
    wb = load_workbook(
        BytesIO(gerar_coleta_oficial_preenchida(_payload_caso_real())),
        data_only=False,
    )
    yield wb
    wb.close()


def _dia(valor):
    return valor.date() if hasattr(valor, "date") else valor


def _linhas_por_ciclo(wb):
    financeiro = wb["financeiro"]
    marco_c0 = _dia(wb["parametros"]["C2"].value).replace(day=1)
    saida = {f"C{i}": [] for i in range(5)}
    for linha in range(2, 74):
        competencia = financeiro[f"A{linha}"].value
        if competencia is None:
            continue
        competencia = _dia(competencia).replace(day=1)
        meses = (
            (competencia.year - marco_c0.year) * 12
            + competencia.month - marco_c0.month
        )
        saida[f"C{meses // 12}"].append(
            (competencia, financeiro[f"G{linha}"].value, linha)
        )
    return saida


def test_caso_real_c2_perde_so_uma_competencia(wb_caso_real):
    c2 = _linhas_por_ciclo(wb_caso_real)["C2"]
    assert len(c2) == 12
    assert [(data, efeito) for data, efeito, _ in c2] == [
        (date(2023, 10, 1), "Nao"),
        *[(date(2023, 11, 1) + relativedelta(months=i), "Sim") for i in range(11)],
    ]


def test_caso_real_c3_e_c4_nao_herdam_defasagem(wb_caso_real):
    ciclos = _linhas_por_ciclo(wb_caso_real)
    for nome, primeira, ultima in (
        ("C3", date(2024, 10, 1), date(2025, 9, 1)),
        ("C4", date(2025, 10, 1), date(2026, 9, 1)),
    ):
        bloco = ciclos[nome]
        assert len(bloco) == 12
        assert bloco[0][:2] == (primeira, "Sim")
        assert bloco[-1][:2] == (ultima, "Sim")
        assert [efeito for _, efeito, _ in bloco] == ["Sim"] * 12


def test_competencias_criticas_permanecem_nos_mesmos_ciclos(wb_caso_real):
    ciclos = _linhas_por_ciclo(wb_caso_real)
    assert ciclos["C2"][0][:2] == (date(2023, 10, 1), "Nao")
    assert ciclos["C3"][0][:2] == (date(2024, 10, 1), "Sim")
    assert ciclos["C4"][0][:2] == (date(2025, 10, 1), "Sim")
    assert {nome: len(bloco) for nome, bloco in ciclos.items()} == {
        "C0": 12,
        "C1": 12,
        "C2": 12,
        "C3": 12,
        "C4": 12,
    }


def test_parametros_fatores_e_formulas_dependentes_permanecem_intactos(wb_caso_real):
    parametros = wb_caso_real["parametros"]
    financeiro = wb_caso_real["financeiro"]
    assert [parametros[f"B{linha}"].value for linha in range(2, 7)] == [
        "C0", "C1", "C2", "C3", "C4",
    ]
    assert [_dia(parametros[f"C{linha}"].value) for linha in range(2, 7)] == [
        date(2021, 10, 1),
        date(2022, 10, 1),
        date(2023, 10, 1),
        date(2024, 11, 1),
        date(2025, 11, 1),
    ]
    assert [parametros[f"E{linha}"].value for linha in range(4, 7)] == [
        0.05, 0.04, 0.03,
    ]
    assert [_dia(parametros[f"H{linha}"].value) for linha in range(4, 7)] == [
        date(2023, 11, 1), date(2024, 11, 1), date(2025, 11, 1),
    ]
    for linha in (26, 38, 50):
        assert str(financeiro[f"B{linha}"].value).startswith("=IF(")
        assert str(financeiro[f"D{linha}"].value).startswith("=IF(")
        assert f'IF(G{linha}="Sim"' in str(financeiro[f"E{linha}"].value)
        assert f'IF(G{linha}="Sim"' in str(financeiro[f"F{linha}"].value)


@pytest.mark.parametrize(
    ("data_base", "inicio_efeito", "esperado"),
    (
        (date(2022, 10, 10), date(2023, 10, 1), 0),
        (date(2022, 10, 10), date(2023, 11, 1), 1),
        (date(2022, 10, 10), date(2023, 12, 1), 2),
    ),
)
def test_defasagem_generica_zero_um_dois(data_base, inicio_efeito, esperado):
    assert _defasagem_financeira_competencias({
        "data_base": data_base,
        "inicio_efeito_financeiro": inicio_efeito,
    }) == esperado


def test_ciclo_posterior_tempestivo_nao_herda_duas_competencias_perdidas():
    ciclos = {
        "C0": {"possui_efeito_financeiro": "Não"},
        "C1": {
            "data_base": date(2022, 10, 10),
            "inicio_efeito_financeiro": date(2023, 12, 1),
            "possui_efeito_financeiro": "Sim",
        },
        "C2": {
            "data_base": date(2023, 12, 1),
            "inicio_efeito_financeiro": date(2024, 12, 1),
            "possui_efeito_financeiro": "Sim",
        },
    }
    wb = Workbook()
    ws = wb.active
    _preencher_financeiro(
        ws, ciclos, novo=True, marco_inicial=date(2022, 10, 1)
    )
    efeitos_c1 = [ws[f"G{linha}"].value for linha in range(14, 26)]
    efeitos_c2 = [ws[f"G{linha}"].value for linha in range(26, 38)]
    assert efeitos_c1 == ["Nao", "Nao"] + ["Sim"] * 10
    assert efeitos_c2 == ["Sim"] * 12
    wb.close()
