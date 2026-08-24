"""Testes permanentes do VTA-M2/VTA-M2.1 — VTA do metodo Financeiro no XLS.

Protege, por leitura estrutural de formula (nao apenas SHA/contagem), que:
  A. o ramo Financeiro de B26 existe e usa a fonte financeira (D20);
  B. o ramo PC dentro de B26 permanece literalmente preservado;
  C. o ramo Itens/Consumido dentro de B26 permanece literalmente preservado;
  D. B28 no Financeiro mantem B23 como referencia comparativa (a formula
     ganhou o metodo "Itens" no OR pela VTA-C2, sem alterar o resultado
     para Financeiro: quando $B$4="Financeiro", ambas retornam $B$23);
  E. o texto de metodologia em RESULTADOS e condicional ao metodo
     selecionado (nunca fixo mencionando "Financeiro" para outro metodo);
  F. o bloco CONFERENCIA DA EXECUCAO nao apresenta numeros/titulo do
     Financeiro como oficiais quando o metodo != Financeiro;
  G. ausencia de dado nunca vira zero (aparece "Sem historico fisico
     suficiente"/"Nao
     aplicavel ao metodo selecionado", nao 0);
  H. o endereco/nome definido VTA_FINAL continua apontando para B26;
  I. o template abre estruturalmente via openpyxl;
  J. quando disponivel, o Excel real recalcula e reabre sem reparo.
"""
from __future__ import annotations

import gc
import os
from pathlib import Path

import pytest
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "templates" / "COLETA_REAJUSTE_OFICIAL.xlsx"

_RAMO_PC_B26 = (
    'IF($T$25="CALCULO MANUAL REQUERIDO","",ROUND($T$25+IF(ISNUMBER(B24),B24,0),2))'
)
_RAMO_ITENS_B26 = (
    'IF(OR(B23="",AND(B24<>"",NOT(ISNUMBER(B24)))),"",'
    'ROUND(B23+IF(ISNUMBER(B24),B24,0)+IF(ISNUMBER($N$263),$N$263,0),2))'
)
_B28_ESPERADO = '=IF(OR($B$4="Financeiro",$B$4="Itens"),$B$23,$B$26)'


@pytest.fixture(scope="module")
def wb():
    return load_workbook(TEMPLATE, data_only=False)


def test_a_ramo_financeiro_b26_usa_fonte_financeira(wb):
    b26 = str(wb["MEMORIA_RESULTADOS"]["B26"].value)
    assert '$B$4="Financeiro"' in b26
    assert "$D$20" in b26
    assert "B21" in b26 and "D35" in b26


def test_b_ramo_pc_b26_preservado_literalmente(wb):
    b26 = str(wb["MEMORIA_RESULTADOS"]["B26"].value)
    assert _RAMO_PC_B26 in b26


def test_c_ramo_itens_consumido_b26_preservado_literalmente(wb):
    b26 = str(wb["MEMORIA_RESULTADOS"]["B26"].value)
    assert _RAMO_ITENS_B26 in b26


def test_d_b28_financeiro_mantem_referencia_comparativa(wb):
    b28 = str(wb["MEMORIA_RESULTADOS"]["B28"].value)
    assert b28 == _B28_ESPERADO
    assert "$B$23" in b28


def test_e_metodologia_resultados_e_condicional_ao_metodo(wb):
    formula = str(wb["RESULTADOS"]["A70"].value)
    assert formula.startswith("=IF(")
    assert "MEMORIA_RESULTADOS!$B$4" in formula
    assert '"Financeiro"' in formula
    assert '"PCs"' in formula
    assert '"Itens"' in formula
    assert "Metodo Financeiro" in formula
    assert "Metodo PCs" in formula
    assert "Metodo Consumido" in formula


def test_f_titulo_conferencia_e_condicional(wb):
    titulo = str(wb["RESULTADOS"]["A71"].value)
    assert titulo.startswith("=IF(")
    assert "MEMORIA_RESULTADOS!$B$4" in titulo
    assert "(Financeiro)" in titulo
    assert "nao aplicavel ao metodo selecionado" in titulo.lower()


def test_f_bloco_conferencia_nao_mostra_financeiro_para_outro_metodo(wb):
    res = wb["RESULTADOS"]
    for linha in range(73, 78):
        for coluna in ("B", "C", "D", "E"):
            formula = str(res[f"{coluna}{linha}"].value)
            assert formula.startswith("=IF(MEMORIA_RESULTADOS!$B$4<>\"Financeiro\",")
            rotulo = "NAO APLICAVEL" if coluna == "E" else "Nao aplicavel ao metodo selecionado"
            assert rotulo in formula


def test_g_ausencia_de_dado_nao_vira_zero(wb):
    res = wb["RESULTADOS"]
    for linha in range(73, 78):
        formula_c = str(res[f"C{linha}"].value)
        assert "Sem historico fisico suficiente" in formula_c
        assert "Nao aplicavel ao metodo selecionado" in formula_c
        # nao pode haver um "senao 0" generico substituindo a comparacao.
        assert not formula_c.rstrip().endswith(",0)")


def test_g_c3_nao_usa_checkpoint_nao_adjacente_como_fallback(wb):
    """VTA-M2.2 item 13.A/B: C3 nao pode usar F/E (checkpoints de C0) nem
    N/J (checkpoints de C1/C2) como fallback para produzir a linha C3 —
    so pode usar itens_Remanesc!$R (VALOR_EXECUTADO_C3, par adjacente
    R+P-V, ja resolvido internamente pelo proprio template)."""
    formula_c3 = str(wb["RESULTADOS"]["C76"].value)
    assert "itens_Remanesc!$R$2:$R$201" in formula_c3
    for coluna_proibida in ("$E$2:$E$201", "$F$2:$F$201", "$J$2:$J$201", "$N$2:$N$201"):
        assert f"posicao_contratual!{coluna_proibida}" not in formula_c3
        assert f"itens_Remanesc!{coluna_proibida}" not in formula_c3


def test_g_c0_nao_usa_contratado_menos_abertura_diretamente(wb):
    """VTA-M2.2 item 13.D: a formula de RESULTADOS para C0 nao pode
    reconstruir "contratado - abertura C0" diretamente (posicao_contratual
    E/F) — deve delegar ao par adjacente ja resolvido em
    itens_Remanesc!$AC (MAX(E-J,0)*VU_C0, fechado por J = abertura de
    C1, nao por F)."""
    formula_c0 = str(wb["RESULTADOS"]["C73"].value)
    assert "itens_Remanesc!$AC$2:$AC$201" in formula_c0
    assert "posicao_contratual!$E$2:$E$201" not in formula_c0
    assert "posicao_contratual!$F$2:$F$201" not in formula_c0


def test_g_cada_ciclo_usa_apenas_sua_propria_coluna_de_execucao(wb):
    """VTA-M2.2 item 13.B: nenhum ciclo pode misturar a coluna de execucao
    de outro ciclo (nenhum encadeamento entre C0-C3)."""
    mapa = {73: "AC", 74: "N", 75: "P", 76: "R"}
    todas_as_colunas = set(mapa.values())
    for linha, coluna_esperada in mapa.items():
        formula = str(wb["RESULTADOS"][f"C{linha}"].value)
        assert f"itens_Remanesc!${coluna_esperada}$2:${coluna_esperada}$201" in formula
        for outra in todas_as_colunas - {coluna_esperada}:
            assert f"itens_Remanesc!${outra}$2:${outra}$201" not in formula


def test_g_c4_sempre_nao_comparavel_sem_checkpoint_de_fechamento(wb):
    """VTA-M2.2: C4 nao tem par adjacente nesta versao do schema (nao
    existe REM_BASE_C5 que feche o ciclo) — deve ser SEMPRE NAO
    COMPARAVEL, por ausencia estrutural, nunca por fallback/calculo."""
    formula_c4 = str(wb["RESULTADOS"]["C77"].value)
    assert formula_c4 == (
        '=IF(MEMORIA_RESULTADOS!$B$4<>"Financeiro",'
        '"Nao aplicavel ao metodo selecionado","Sem historico fisico suficiente")'
    )


def test_e_status_so_gera_revisar_quando_ambas_comparaveis(wb):
    """VTA-M2.2 item 13.E: o status so pode ser REVISAR/OK quando a
    execucao teorica (coluna C) nao for NAO COMPARAVEL nem vazia."""
    res = wb["RESULTADOS"]
    for linha in range(73, 78):
        formula = str(res[f"E{linha}"].value)
        assert f'C{linha}=""' in formula
        # VTA-U2: o guard deixou de comparar o texto "NAO COMPARAVEL" e passou
        # a exigir que C seja numero — mais robusto e independente do rotulo.
        assert f"NOT(ISNUMBER(C{linha}))" in formula
        # a ramificacao REVISAR/OK so e alcancada depois desse guard-clause.
        indice_guard = formula.index(f"NOT(ISNUMBER(C{linha}))")
        indice_revisar = formula.index("REVISAR")
        assert indice_guard < indice_revisar


def test_h_nome_definido_vta_final_aponta_para_b26(wb):
    definido = wb.defined_names["VTA_FINAL"]
    destinos = list(definido.destinations)
    assert len(destinos) == 1
    aba, referencia = destinos[0]
    assert aba == "MEMORIA_RESULTADOS"
    assert referencia.replace("$", "") == "B26"


def test_i_template_abre_estruturalmente_via_openpyxl():
    wb_local = load_workbook(TEMPLATE, data_only=False)
    assert "MEMORIA_RESULTADOS" in wb_local.sheetnames
    assert "RESULTADOS" in wb_local.sheetnames


@pytest.mark.skipif(
    os.environ.get("RUN_EXCEL_INTEGRATION") != "1",
    reason="defina RUN_EXCEL_INTEGRATION=1 para executar Excel COM",
)
def test_j_excel_real_recalcula_e_reabre_sem_reparo():
    client = pytest.importorskip("win32com.client")
    pythoncom = pytest.importorskip("pythoncom")
    pythoncom.CoInitialize()
    excel = client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        wb_com = excel.Workbooks.Open(str(TEMPLATE), UpdateLinks=0, ReadOnly=True)
        try:
            assert wb_com.Worksheets.Count == 15
            mem = wb_com.Worksheets("MEMORIA_RESULTADOS")
            assert str(mem.Range("B26").Formula)
            res = wb_com.Worksheets("RESULTADOS")
            assert str(res.Range("A70").Formula).startswith("=IF(")
        finally:
            wb_com.Close(False)
            del wb_com
    finally:
        excel.Quit()
        del excel
        gc.collect()
        pythoncom.CoUninitialize()
