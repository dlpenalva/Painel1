from __future__ import annotations

import ast
import re
from io import BytesIO
from pathlib import Path
from zipfile import ZipFile

import openpyxl
import pandas as pd

from _seguranca_xlsx import opcoes_excel_writer_seguro
from _adequacao_ui import gerar_xlsx_projecao


ROOT = Path(__file__).resolve().parents[1]
PAGINA = ROOT / "pages" / "03_Valor_Global.py"
PAYLOADS = ["=1+1", "+1+1", "-1+1", "@SUM(1,1)", "https://example.invalid/test"]


def _carregar_gerador(nome: str):
    arvore = ast.parse(PAGINA.read_text(encoding="utf-8"), filename=str(PAGINA))
    nomes = {
        "numero_ciclo",
        "numero_seguro",
        "limpar_nan_inf_df",
        "normalizar_ciclo",
        "normalizar_texto",
        "numero_br",
        "texto_seguro",
        "formatar_data_br",
        nome,
    }
    funcoes = [
        no for no in arvore.body if isinstance(no, ast.FunctionDef) and no.name in nomes
    ]
    modulo = ast.Module(body=funcoes, type_ignores=[])
    ast.fix_missing_locations(modulo)
    espaco = {
        "BytesIO": BytesIO,
        "pd": pd,
        "re": re,
        "opcoes_excel_writer_seguro": opcoes_excel_writer_seguro,
    }
    exec(compile(modulo, str(PAGINA), "exec"), espaco)
    return espaco[nome]


def _rels_externos(conteudo: bytes) -> list[str]:
    with ZipFile(BytesIO(conteudo)) as pacote:
        return [
            nome
            for nome in pacote.namelist()
            if nome.endswith(".rels")
            and b'TargetMode="External"' in pacote.read(nome)
        ]


def test_writer_de_itens_preserva_payloads_como_texto_sem_hyperlink() -> None:
    gerar = _carregar_gerador("gerar_excel_valores_unitarios_por_ciclo")
    df = pd.DataFrame(
        {
            "Item": PAYLOADS,
            "Ciclo": ["C1"] * len(PAYLOADS),
            "Valor unitário": [1.0] * len(PAYLOADS),
            "Quantidade": [1.0] * len(PAYLOADS),
            "Total R$": [1.0] * len(PAYLOADS),
            "Ciclo precluso": [False] * len(PAYLOADS),
        }
    )
    conteudo = gerar(df, pd.DataFrame({"Ciclo": ["C1"]}))
    wb = openpyxl.load_workbook(BytesIO(conteudo), data_only=False)
    valores = [wb["VALORES_POR_CICLO"].cell(row=i, column=1) for i in range(2, 7)]
    assert set(celula.value for celula in valores) == set(PAYLOADS)
    assert all(celula.data_type == "s" for celula in valores)
    assert _rels_externos(conteudo) == []
    assert wb["MATRIZ_VALORES_CICLO"]["D8"].data_type == "f"


def test_planilha_executiva_preserva_formula_controlada_como_texto() -> None:
    gerar = _carregar_gerador("gerar_planilha_executiva")
    resultado = {
        "df_valores_unitarios_ciclo": pd.DataFrame(
            {
                "Item": ["=1+1", "https://example.invalid/test"],
                "Ciclo": ["C1", "C1"],
                "Quantidade": [1.0, 1.0],
                "Valor unitário": [1.0, 1.0],
                "Total R$": [1.0, 1.0],
            }
        )
    }
    conteudo = gerar(resultado)
    wb = openpyxl.load_workbook(BytesIO(conteudo), data_only=False)
    itens = [wb["DETALHAMENTO_ITENS"]["A4"], wb["DETALHAMENTO_ITENS"]["A5"]]
    assert [celula.value for celula in itens] == ["=1+1", "https://example.invalid/test"]
    assert all(celula.data_type == "s" for celula in itens)
    assert _rels_externos(conteudo) == []


def test_formula_deliberada_do_sistema_continua_formula() -> None:
    saida = BytesIO()
    with pd.ExcelWriter(
        saida,
        engine="xlsxwriter",
        engine_kwargs=opcoes_excel_writer_seguro(),
    ) as writer:
        pd.DataFrame({"texto": ["=1+1"]}).to_excel(writer, index=False)
        writer.sheets["Sheet1"].write_formula(2, 0, "=SUM(1,1)")
    wb = openpyxl.load_workbook(BytesIO(saida.getvalue()), data_only=False)
    assert wb["Sheet1"]["A2"].data_type == "s"
    assert wb["Sheet1"]["A3"].data_type == "f"


def test_writer_da_adequacao_preserva_texto_controlavel() -> None:
    df_ultimos = pd.DataFrame(
        {"Competência": ["=1+1"], "Valor pago/medido": [1.0]}
    )
    df_projecao = pd.DataFrame({"Fonte": ["https://example.invalid/test"]})
    conteudo = gerar_xlsx_projecao(
        df_ultimos,
        df_projecao,
        [("Observação", "@SUM(1,1)")],
    )
    wb = openpyxl.load_workbook(BytesIO(conteudo), data_only=False)
    textos = [
        celula
        for ws in wb.worksheets
        for linha in ws.iter_rows()
        for celula in linha
        if celula.value in {"=1+1", "https://example.invalid/test", "@SUM(1,1)"}
    ]
    assert {celula.value for celula in textos} == {
        "=1+1",
        "https://example.invalid/test",
        "@SUM(1,1)",
    }
    assert all(celula.data_type == "s" for celula in textos)
    assert _rels_externos(conteudo) == []


def test_todos_os_writers_derivados_da_coleta_usam_opcoes_seguras() -> None:
    pagina = PAGINA.read_text(encoding="utf-8")
    adequacao = (ROOT / "_adequacao_ui.py").read_text(encoding="utf-8")
    assert pagina.count("engine_kwargs=opcoes_excel_writer_seguro()") == 2
    assert adequacao.count("engine_kwargs=opcoes_excel_writer_seguro()") == 1
