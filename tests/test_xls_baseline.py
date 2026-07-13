import ast
import io
import re
import unittest
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

import openpyxl
import pandas as pd
from xlsxwriter.utility import xl_col_to_name


ROOT = Path(__file__).resolve().parents[1]
PAGES = (
    ROOT / "pages" / "01_Calculo_Simples.py",
    ROOT / "pages" / "02_Calculo_Represados.py",
)


def load_function(path: Path, function_name: str):
    """Carrega apenas o gerador puro, sem executar a interface Streamlit."""
    tree = ast.parse(path.read_text(encoding="utf-8"), filename=str(path))
    function = next(
        node
        for node in tree.body
        if isinstance(node, ast.FunctionDef) and node.name == function_name
    )
    module = ast.Module(body=[function], type_ignores=[])
    ast.fix_missing_locations(module)
    namespace = {
        "BytesIO": io.BytesIO,
        "ZoneInfo": ZoneInfo,
        "datetime": datetime,
        "pd": pd,
        "re": re,
        "xl_col_to_name": xl_col_to_name,
    }
    exec(compile(module, str(path), "exec"), namespace)
    return namespace[function_name]


class XlsBaselineTests(unittest.TestCase):
    def test_consumo_por_itens_gera_xlsx_real_nos_dois_fluxos(self):
        dados = {
            "indice": "IPCA",
            "data_base_original": "01/2024",
            "ciclos": [
                {
                    "ciclo": "C1",
                    "data_base": "01/2025",
                    "janela_admissibilidade": "01/2025 a 03/2025",
                    "data_pedido": "15/01/2025",
                    "financeiro_inicio": "01/2025",
                    "percentual_aplicado": 0.045,
                    "fator": 1.045,
                    "fator_acumulado": 1.045,
                    "situacao": "TEMPESTIVO",
                }
            ],
        }

        for page in PAGES:
            with self.subTest(page=page.name):
                generator = load_function(
                    page, "gerar_modelo_consumo_itens_ciclo_excel"
                )
                payload = generator(dados)
                self.assertTrue(payload.startswith(b"PK"))

                workbook = openpyxl.load_workbook(
                    io.BytesIO(payload), data_only=False
                )
                self.assertEqual(
                    workbook.sheetnames,
                    [
                        "INICIO",
                        "PARAMETROS",
                        "CICLOS_APURADOS",
                        "CONSUMO_ITENS",
                        "CICLO_EM_EXECUCAO",
                        "CALCULO_AUTOMATICO",
                        "RESUMO",
                    ],
                )
                self.assertEqual(
                    workbook["CICLOS_APURADOS"]["A6"].value, "C1"
                )
                formulas = [
                    cell.value
                    for row in workbook["CONSUMO_ITENS"].iter_rows()
                    for cell in row
                    if isinstance(cell.value, str) and cell.value.startswith("=")
                ]
                self.assertGreater(len(formulas), 100)

    def test_paginas_mantem_os_dois_geradores_de_xls(self):
        for page in PAGES:
            source = page.read_text(encoding="utf-8")
            with self.subTest(page=page.name):
                self.assertIn("def gerar_arquivo_coleta_excel", source)
                self.assertIn(
                    "def gerar_modelo_consumo_itens_ciclo_excel", source
                )


if __name__ == "__main__":
    unittest.main()
