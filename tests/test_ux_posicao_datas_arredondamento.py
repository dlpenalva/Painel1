# -*- coding: utf-8 -*-
"""Posicao fisica unificada, datas padronizadas e arredondamento itemizado.

Cobre o que nao tinha cobertura direta antes deste pacote:

  1 CICLO_EM_EXECUCAO e a unica entrada manual da posicao fisica;
  2 posicao_referencia deriva data, ciclo, completude e origem dela;
  3 data igual ao corte  -> VALIDADO;
  4 data anterior        -> ESTIMADO;
  5 data posterior       -> REVISE e nao alimenta o VTA;
  6 PC posterior ao corte permanece cadastrado sem entrar na cobertura;
  7 rotulos de data padronizados (um texto por conceito);
  8 itens_RC: titulos dos blocos e nota da data da alteracao contratual;
  9 cores das abas de preenchimento;
 10 valor itemizado = ARRED(QTD x VU_ATUALIZADO; 2), com historico_VU canonico.
"""
from __future__ import annotations

import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

RAIZ = Path(__file__).resolve().parents[1]
if str(RAIZ) not in sys.path:
    sys.path.insert(0, str(RAIZ))

from _ciclo_em_execucao import (  # noqa: E402
    ABA_CICLO_EM_EXECUCAO,
    CELULA_VALIDACAO_TEMPORAL,
    COLUNAS_VISIVEIS,
    COR_ABA_ENTRADA,
    LINHA_CABECALHO,
    garantir_aba_ciclo_em_execucao,
)

TEMPLATE = RAIZ / "templates" / "COLETA_REAJUSTE_OFICIAL.xlsx"

ABAS_DE_PREENCHIMENTO = (
    "CONTROLE", "financeiro", "itens_Remanesc", "itens_Consumidos",
    "itens_PC", "aditivos", "cobertura_temporal",
)


@pytest.fixture(scope="module")
def wb():
    livro = load_workbook(TEMPLATE, data_only=False)
    yield livro
    livro.close()


@pytest.fixture(scope="module")
def wb_gerado():
    """Template com a aba CICLO_EM_EXECUCAO, como no arquivo entregue."""
    livro = load_workbook(TEMPLATE, data_only=False)
    garantir_aba_ciclo_em_execucao(livro)
    yield livro
    livro.close()


# --------------------------------------------------------------------------- #
# 1-2. Fonte unica da posicao fisica
# --------------------------------------------------------------------------- #
def test_posicao_fisica_pedida_uma_unica_vez(wb_gerado):
    """A quantidade e digitada so em CICLO_EM_EXECUCAO!C13:C211."""
    cee = wb_gerado[ABA_CICLO_EM_EXECUCAO]
    assert cee.cell(13, 3).value is None                # celula de entrada, vazia
    assert not cee.cell(13, 3).protection.locked        # e editavel

    pr = wb_gerado["posicao_referencia"]
    b2 = str(pr["B2"].value)
    assert b2.startswith("=")                           # deixou de ser digitada
    assert 'INDIRECT("CICLO_EM_EXECUCAO!$C$13:$C$211")' in b2
    # correspondencia por ITEM, nunca pela mesma linha fisica
    assert 'MATCH($A2,INDIRECT("CICLO_EM_EXECUCAO!$A$13:$A$211"),0)' in b2
    # nunca soma as duas abas nem copia valor estatico
    assert "+" not in b2 and "SUM" not in b2
    assert bool(pr.protection.sheet)


def test_data_ciclo_completude_e_origem_vem_de_ciclo_em_execucao(wb):
    pr = wb["posicao_referencia"]
    assert 'INDIRECT("CICLO_EM_EXECUCAO!$D$5")' in str(pr["I9"].value)
    for coord in ("I1", "I2", "I5"):
        assert "$I$9" in str(pr[coord].value), coord
    assert "POSICAO FISICA INFORMADA - " in str(pr["I8"].value)


# --------------------------------------------------------------------------- #
# 3-5. Validacao temporal da posicao
# --------------------------------------------------------------------------- #
def test_status_da_referencia_fisica_tem_os_tres_estados(wb):
    formula = str(wb["posicao_referencia"]["I10"].value)
    assert '$I$9>CONTROLE!$B$3,"REVISE"' in formula          # posterior
    assert '$I$9=CONTROLE!$B$3,"VALIDADO","ESTIMADO"' in formula


def test_mensagem_de_posicao_posterior_ao_corte(wb, wb_gerado):
    esperada = "POSICAO FISICA POSTERIOR A DATA DE CORTE."
    assert esperada in str(wb["posicao_referencia"]["I11"].value)
    # a mesma advertencia aparece na aba onde o fiscal digita a data
    cee = wb_gerado[ABA_CICLO_EM_EXECUCAO]
    assert esperada in str(cee[CELULA_VALIDACAO_TEMPORAL].value)


def test_posicao_posterior_ao_corte_nao_alimenta_o_vta(wb):
    """Fotografia futura nao representa apuracao encerrada em data anterior."""
    w49 = str(wb["MEMORIA_RESULTADOS"]["W49"].value)
    assert 'INDIRECT("CICLO_EM_EXECUCAO!$D$5")' in w49
    assert "CONTROLE!$B$3" in w49
    assert ">CONTROLE!$B$3),0,1)" in w49
    # a data de corte nunca e reescrita pela fotografia
    assert str(wb["CONTROLE"]["B3"].value or "") == ""


def test_apresentacao_das_duas_referencias_do_vta(wb):
    h10 = str(wb["RESULTADOS"]["H10"].value)
    assert "UTILIZADA - POSICAO FISICA DE " in h10
    assert "ESTIMADA - POSICAO FISICA DE " in h10
    assert "NAO DISPONIVEL - POSICAO FISICA NAO INFORMADA OU INCOMPLETA" in h10
    assert "REVISE - POSICAO POSTERIOR A DATA DE CORTE" in h10
    # a abertura nunca se apresenta como posicao fisica atual
    h11 = str(wb["RESULTADOS"]["H11"].value)
    assert "REFERENCIA - ABERTURA DO CICLO C" in h11
    assert "POSICAO FISICA" not in h11
    # a igualdade entre as duas referencias nao e tratada como erro
    assert str(wb["RESULTADOS"]["B13"].value).startswith("=")


# --------------------------------------------------------------------------- #
# 6. PCs posteriores ao corte
# --------------------------------------------------------------------------- #
def test_pcs_posteriores_ao_corte_nao_definem_a_cobertura(wb):
    ws = wb["cobertura_temporal"]
    assert str(ws["A14"].value) == "ÚLTIMO PC CONSIDERADO ATÉ O CORTE — AUTO"
    assert "itens_PC!$B$2:$B$5001<=CONTROLE!$B$3" in str(ws["B14"].value)
    assert str(ws["A17"].value) == "EXISTEM PCS POSTERIORES AO CORTE?"
    b17 = str(ws["B17"].value)
    assert "SIM - ULTIMO PC CADASTRADO EM " in b17 and '"NAO"' in b17


# --------------------------------------------------------------------------- #
# 7. Datas padronizadas: um texto por conceito
# --------------------------------------------------------------------------- #
def test_rotulos_de_data_padronizados(wb, wb_gerado):
    assert str(wb["CONTROLE"]["A3"].value) == "DATA DE CORTE DA APURAÇÃO"
    assert "Última data considerada" in str(wb["CONTROLE"]["C3"].value)
    cob = wb["cobertura_temporal"]
    assert str(cob["A4"].value) == "DATA DE GERAÇÃO/ANÁLISE — AUTO"
    assert str(cob["A13"].value) == "FINANCEIRO CONFERIDO ATÉ — OPCIONAL"
    assert str(cob["A15"].value) == "PCS CONFERIDOS ATÉ — OPCIONAL"
    assert str(cob["B4"].value).startswith("=")          # automatica
    cee = wb_gerado[ABA_CICLO_EM_EXECUCAO]
    assert str(cee["A5"].value) == "DATA DA POSIÇÃO FÍSICA (PREENCHER)"


def test_sem_dois_textos_para_o_mesmo_conceito(wb):
    """Nenhum rotulo legado sobrevive pedindo a mesma coisa com outro nome."""
    proibidos = (
        "Data de corte (única p/ contrato)",
        "Data da analise (GCC, opcional)",
        "Financeiro confirmado completo ate (GCC, opcional)",
        "PC confirmado completo ate (GCC, opcional)",
        "Ultima evidencia PC (nao e completo ate)",
    )
    textos = set()
    for aba in ("CONTROLE", "cobertura_temporal"):
        ws = wb[aba]
        for row in ws.iter_rows(max_col=3):
            for c in row:
                if isinstance(c.value, str):
                    textos.add(c.value.strip())
    for proibido in proibidos:
        assert proibido not in textos, proibido


# --------------------------------------------------------------------------- #
# 8. itens_RC: titulos e orientacoes
# --------------------------------------------------------------------------- #
def test_itens_rc_titulos_e_nota(wb):
    ws = wb["itens_RC"]
    assert str(ws["B1"].value) == (
        "POSIÇÃO AJUSTADA POR CICLO "
        "(AUTO — INCLUI ALTERAÇÕES CONTRATUAIS APLICÁVEIS)"
    )
    assert str(ws["Q1"].value) == (
        "POSIÇÃO FÍSICA ATUAL (AUTO — ORIGEM: CICLO_EM_EXECUCAO)"
    )
    assert str(ws["Z2"].value) == "DATA DE EFEITO DA ALTERAÇÃO CONTRATUAL — AUTO"
    nota = str(ws["AE1"].value)
    assert "acréscimo" in nota and "supressão" in nota and "novo item" in nota
    assert "Não corresponde ao início do efeito financeiro do reajuste." in nota
    # cada coluna continua identificando o seu ciclo
    for col, ciclo in (("B", "C0"), ("E", "C1"), ("H", "C2"),
                       ("K", "C3"), ("N", "C4")):
        assert str(ws[f"{col}2"].value) == f"VU ATUALIZADO {ciclo}"


def test_itens_rc_formulas_preservadas(wb):
    """Alterar apenas titulos: a consolidacao automatica nao muda."""
    ws = wb["itens_RC"]
    assert str(ws["D3"].value) == (
        '=IF(A3="TOTAL",ROUND(SUM($D2:D$3),2),'
        'IF(OR(A3="",B3="",C3=""),"",ROUND(B3*C3,2)))'
    )


# --------------------------------------------------------------------------- #
# 9. Cores das abas
# --------------------------------------------------------------------------- #
def test_cores_das_abas_de_preenchimento(wb_gerado):
    for aba in ABAS_DE_PREENCHIMENTO + (ABA_CICLO_EM_EXECUCAO,):
        cor = wb_gerado[aba].sheet_properties.tabColor
        assert cor is not None and cor.rgb == COR_ABA_ENTRADA, aba
    # posicao_referencia e automatica: nao recebe a cor de entrada
    pr = wb_gerado["posicao_referencia"].sheet_properties.tabColor
    assert pr is None or pr.rgb != COR_ABA_ENTRADA
    # RESULTADOS preserva a cor propria
    assert wb_gerado["RESULTADOS"].sheet_properties.tabColor.rgb == "FF8A1538"


def test_cabecalho_de_quantidade_alinhado_a_data_da_posicao():
    assert COLUNAS_VISIVEIS[2] == "QTD REMANESCENTE NA DATA DA POSIÇÃO (PREENCHER)"
    assert LINHA_CABECALHO == 12


# --------------------------------------------------------------------------- #
# 10. Arredondamento itemizado
# --------------------------------------------------------------------------- #
def _sem_fator_bruto(formula: str) -> bool:
    """Nenhum valor itemizado pode usar QTD x VU_ORIGINAL x FATOR."""
    return not any(
        marca in formula
        for marca in ("*$Z$2", "*$Z$3", "*$Z$4", "*$Z$5", "*$Z$6", "*$I$7")
    )


@pytest.mark.parametrize(
    "col,qtd,vu",
    [
        ("F", "posicao_contratual!K2", "historico_VU!D2"),
        ("H", "posicao_contratual!O2", "historico_VU!E2"),
        ("J", "posicao_contratual!S2", "historico_VU!F2"),
        ("L", "posicao_contratual!W2", "historico_VU!G2"),
        ("N", "M2", "historico_VU!D2"),
        ("P", "O2", "historico_VU!E2"),
        ("R", "Q2", "historico_VU!F2"),
        ("T", "S2", "historico_VU!G2"),
        ("AC", "AB2", "historico_VU!C2"),
    ],
)
def test_itens_remanesc_usa_vu_atualizado_arredondado(wb, col, qtd, vu):
    formula = str(wb["itens_Remanesc"][f"{col}2"].value)
    assert f"ROUND({qtd}*{vu},2)" in formula
    assert _sem_fator_bruto(formula)


def test_aditivos_usam_vu_do_historico(wb):
    formula = str(wb["aditivos"]["J2"].value)
    assert "VLOOKUP($A2,historico_VU!$A:$G," in formula
    assert "F2*I2" not in formula          # QTD x VU_ORIGINAL x FATOR eliminado
    assert formula.startswith('=IF(OR(L2="",F2=""),"",ROUND(L2*')


@pytest.mark.parametrize("col,qtd", [("O", "D2"), ("P", "N2"), ("R", "L2")])
def test_posicao_referencia_usa_vu_do_historico(wb, col, qtd):
    formula = str(wb["posicao_referencia"][f"{col}2"].value)
    assert f"ROUND({qtd}*IF($I$4=" in formula
    assert "historico_VU!C2" in formula and "historico_VU!G2" in formula
    assert _sem_fator_bruto(formula)


@pytest.mark.parametrize(
    "linha,qtd,vu",
    [
        (26, "posicao_contratual!$G$2:$G$201", "historico_VU!$C$2:$C$201"),
        (27, "posicao_contratual!$K$2:$K$201", "historico_VU!$D$2:$D$201"),
        (28, "posicao_contratual!$O$2:$O$201", "historico_VU!$E$2:$E$201"),
        (29, "posicao_contratual!$S$2:$S$201", "historico_VU!$F$2:$F$201"),
        (30, "posicao_contratual!$W$2:$W$201", "historico_VU!$G$2:$G$201"),
    ],
)
def test_referencia_por_ciclo_usa_o_vu_canonico(wb, linha, qtd, vu):
    """AGREGADO por ciclo: nao e valor itemizado, mas usa o MESMO VU canonico.

    Fica registrado como fronteira consciente: a variante por item
    SUMPRODUCT(ROUND(IFERROR(qtd*vu,0),2)) foi testada no Excel real e
    retornou 0,00, entao a formula homologada foi preservada. O arredondamento
    por item continua garantido onde o item e calculado (itens_RC,
    CICLO_EM_EXECUCAO, itens_Remanesc e MEMORIA_RESULTADOS, que alimenta o VTA).
    """
    formula = str(wb["RESULTADOS"][f"C{linha}"].value)
    assert f"ROUND(SUMPRODUCT({qtd},{vu}),2)" in formula
    assert "$Z$" not in formula and "parametros!$F$" not in formula


def test_historico_vu_e_a_fonte_canonica_do_vu(wb):
    """VU_ATUALIZADO_Cn = ARRED(VU_ORIGINAL x FATOR_ACUMULADO_Cn; 2)."""
    formula = str(wb["historico_VU"]["D2"].value)
    assert "ROUND(itens_Remanesc!C2*$L$3/INDEX($L$2:$L$6," in formula


def test_metodos_pc_e_financeiro_preservados(wb):
    """Valor monetario da operacao x fator, arredondado no resultado."""
    pc = str(wb["itens_PC"]["E2"].value) + str(wb["itens_PC"]["F2"].value)
    assert "historico_VU" not in pc          # PC nao tem qtd x preco unitario
    fin = "".join(str(wb["financeiro"][f"{c}2"].value) for c in "DEF")
    assert "historico_VU" not in fin
