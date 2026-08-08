# -*- coding: utf-8 -*-
"""ETAPA 31 — invariantes permanentes do modelo temporal do XLS.

Tres conceitos independentes (nunca voltar a fundi-los):

  A) JANELA DO REAJUSTE (parametros!C:D): 12 competencias EXATAS;
     DATA_INICIO(Cn+1) = COMPETENCIA(H(Cn)) + 12m; gap intencional entre
     janelas e PERMITIDO;
  B) CRONOLOGIA DA EXECUCAO: blocos fixos de 12 meses da data-base original
     (5 x 12 = 60 competencias, sem lacuna, sem sobreposicao); H nunca muda
     o CICLO;
  C) EFEITO FINANCEIRO (parametros!H): gate SOMENTE do novo delta — nunca
     exclui valor executado, quantidade, VTA ou posicao contratual.

Caso de referencia obrigatorio (secao 33 do enunciado):
  data-base original 01/04/2022;
  parametros: C0 04/2022-03/2023 | C1 04/2023-03/2024 |
              C2 04/2024-03/2025 H=06/2024 | C3 06/2025-05/2026 H=06/2025 |
              C4 06/2026-05/2027 H=06/2026;
  financeiro: C0..C4 cronologicos 04/2022-03/2027 = 60 competencias;
  efeito: 04-05/2024, 04-05/2025 e 04-05/2026 = Nao (vermelho automatico).
"""
from __future__ import annotations

import io
import sys
from datetime import date, timedelta
from pathlib import Path

import pytest
from openpyxl import load_workbook

RAIZ = Path(__file__).resolve().parents[1]
if str(RAIZ) not in sys.path:
    sys.path.insert(0, str(RAIZ))

from _coleta_oficial import (  # noqa: E402
    gerar_coleta_oficial_preenchida,
    normalizar_dados_calculadora,
)
from _gerador_masterfile import _completar_periodos_ciclos  # noqa: E402
from _leitor_masterfile_v10 import (  # noqa: E402
    _ciclo_por_competencia,
    _ler_parcelas_sombra_financeiro,
)
from _motor_posicao_contratual import (  # noqa: E402
    calendario_execucao_por_ciclo,
    determinar_ciclo_por_data,
    divergencias_calendario_canonico,
)
from _motor_temporal import montar_motor_temporal  # noqa: E402


def _payload_caso_referencia() -> dict:
    """Analise C2+C3+C4 do caso de referencia (secao 33)."""
    return {
        "origem": "Reajustes Múltiplos",
        "indice": "IPCA",
        "data_base_original": "01/04/2022",
        "variacao_acumulada": 0.1237,
        "fator_acumulado": 1.1237,
        "ciclos": [
            {"ciclo": "C2", "data_inicio": "01/04/2024",
             "data_pedido": "04/06/2024", "percentual_aplicado": 0.05,
             "financeiro_inicio": "01/06/2024", "objeto_analise_atual": True,
             "situacao": "✅ TEMPESTIVO*", "efeito_financeiro_retardado": True},
            {"ciclo": "C3", "data_pedido": "04/06/2025",
             "percentual_aplicado": 0.04, "financeiro_inicio": "01/06/2025",
             "objeto_analise_atual": True, "situacao": "✅ TEMPESTIVO"},
            {"ciclo": "C4", "data_pedido": "04/06/2026",
             "percentual_aplicado": 0.03, "financeiro_inicio": "01/06/2026",
             "objeto_analise_atual": True, "situacao": "✅ TEMPESTIVO"},
        ],
    }


@pytest.fixture(scope="module")
def wb_referencia():
    wb = load_workbook(
        io.BytesIO(gerar_coleta_oficial_preenchida(_payload_caso_referencia())),
        data_only=False,
    )
    yield wb
    wb.close()


def _dia(valor):
    return valor.date() if hasattr(valor, "date") else valor


# --------------------------------------------------------------------------- #
# PARAMETROS — invariantes da janela do reajuste
# --------------------------------------------------------------------------- #
def test_parametros_caso_referencia_tabela_completa(wb_referencia):
    """Tabela obrigatoria da secao 5/33 do enunciado."""
    par = wb_referencia["parametros"]
    esperado = {
        2: ("C0", date(2022, 4, 1), date(2023, 3, 31), None),
        3: ("C1", date(2023, 4, 1), date(2024, 3, 31), None),
        4: ("C2", date(2024, 4, 1), date(2025, 3, 31), date(2024, 6, 1)),
        5: ("C3", date(2025, 6, 1), date(2026, 5, 31), date(2025, 6, 1)),
        6: ("C4", date(2026, 6, 1), date(2027, 5, 31), date(2026, 6, 1)),
    }
    for linha, (ciclo, ini, fim, efeito) in esperado.items():
        assert par[f"B{linha}"].value == ciclo
        assert _dia(par[f"C{linha}"].value) == ini
        assert _dia(par[f"D{linha}"].value) == fim
        assert _dia(par[f"H{linha}"].value) == efeito


def test_parametros_toda_janela_tem_12_competencias_exatas(wb_referencia):
    par = wb_referencia["parametros"]
    ciclos = [
        {"ciclo": par[f"B{r}"].value, "data_inicio": _dia(par[f"C{r}"].value),
         "data_fim": _dia(par[f"D{r}"].value)}
        for r in range(2, 7)
    ]
    assert divergencias_calendario_canonico(ciclos) == []
    for c in ciclos:
        meses = (
            (c["data_fim"].year - c["data_inicio"].year) * 12
            + c["data_fim"].month - c["data_inicio"].month + 1
        )
        assert meses == 12, f"{c['ciclo']}: {meses} competencias"


def test_parametros_gap_intencional_e_permitido(wb_referencia):
    """C2 fecha 03/2025 e C3 abre 06/2025: 04-05/2025 fora das janelas."""
    par = wb_referencia["parametros"]
    fim_c2 = _dia(par["D4"].value)
    inicio_c3 = _dia(par["C5"].value)
    assert fim_c2 == date(2025, 3, 31)
    assert inicio_c3 == date(2025, 6, 1)
    assert inicio_c3 > fim_c2 + timedelta(days=1)  # gap real
    # e o diagnostico canonico NAO acusa problema
    ciclos = [
        {"ciclo": par[f"B{r}"].value, "data_inicio": _dia(par[f"C{r}"].value),
         "data_fim": _dia(par[f"D{r}"].value)}
        for r in range(2, 7)
    ]
    assert divergencias_calendario_canonico(ciclos) == []


def test_h_nao_alonga_o_proprio_ciclo_e_desloca_o_seguinte():
    dados = normalizar_dados_calculadora(_payload_caso_referencia())
    por_ciclo = {c["ciclo"]: c for c in dados["ciclos"]}
    c2 = por_ciclo["C2"]
    # H(C2)=06/2024 NAO alonga C2 (fecha 03/2025, 12 competencias)...
    assert c2["data_fim"] == date(2025, 3, 31)
    # ...e desloca o inicio de C3 para H+12m.
    assert por_ciclo["C3"]["data_inicio"] == date(2025, 6, 1)


def test_ciclo_sem_ancora_usa_marco_teorico():
    completos = _completar_periodos_ciclos({
        "C1": {"ciclo": "C1", "data_inicio": date(2023, 4, 1),
               "data_fim": date(2024, 3, 31)},  # sem H
    })
    assert completos["C2"]["data_inicio"] == date(2024, 4, 1)
    assert completos["C2"]["data_fim"] == date(2025, 3, 31)


# --------------------------------------------------------------------------- #
# FINANCEIRO — invariantes da cronologia da execucao
# --------------------------------------------------------------------------- #
def _grade(wb) -> list[tuple[date, str]]:
    fin = wb["financeiro"]
    saida = []
    for r in range(2, 74):
        comp = fin[f"A{r}"].value
        if comp is not None:
            saida.append((_dia(comp), fin[f"G{r}"].value))
    return saida


def test_financeiro_60_competencias_12_por_ciclo(wb_referencia):
    grade = _grade(wb_referencia)
    assert len(grade) == 60
    assert grade[0][0] == date(2022, 4, 1)
    assert grade[-1][0] == date(2027, 3, 1)
    # sequencia mensal continua, sem buracos nem duplicidade
    for (a, _), (b, _) in zip(grade, grade[1:]):
        assert (b.year, b.month) == (
            (a.year + 1, 1) if a.month == 12 else (a.year, a.month + 1)
        )
    # 12 competencias por bloco cronologico C0..C4
    marco = grade[0][0]
    contagem = {}
    for comp, _ in grade:
        indice = ((comp.year - marco.year) * 12 + comp.month - marco.month) // 12
        contagem[indice] = contagem.get(indice, 0) + 1
    assert contagem == {0: 12, 1: 12, 2: 12, 3: 12, 4: 12}


@pytest.mark.parametrize("data_corte", [
    None,               # payload sem corte explicito (max data_fim)
    "31/03/2024",       # analise inicial: corte MUITO anterior aos 60 meses
    "31/12/2030",       # corte alem da cronologia
])
def test_financeiro_sempre_60_competencias_independente_do_corte(data_corte):
    """REGRA PETREA (gate final): a grade e SEMPRE a cronologia contratual
    completa — 60 competencias, 12 por ciclo, do marco de C0 ao marco+59m —
    independentemente de data_corte, ultimo ciclo analisado, efeito ou
    pedido. Corte antecipado NUNCA encurta a grade."""
    payload = _payload_caso_referencia()
    if data_corte is not None:
        payload["data_corte"] = data_corte
    wb = load_workbook(
        io.BytesIO(gerar_coleta_oficial_preenchida(payload)), data_only=False
    )
    grade = _grade(wb)
    wb.close()
    assert len(grade) == 60
    marco = grade[0][0]
    assert marco == date(2022, 4, 1)
    assert grade[-1][0] == date(2027, 3, 1)   # marco + 59 meses
    # sequencia mensal continua
    for (a, _), (b, _) in zip(grade, grade[1:]):
        assert (b.year, b.month) == (
            (a.year + 1, 1) if a.month == 12 else (a.year, a.month + 1)
        )
    # 12 competencias por bloco cronologico C0..C4
    contagem = {}
    for comp, _ in grade:
        indice = ((comp.year - marco.year) * 12 + comp.month - marco.month) // 12
        contagem[indice] = contagem.get(indice, 0) + 1
    assert contagem == {0: 12, 1: 12, 2: 12, 3: 12, 4: 12}


def test_financeiro_h_muda_somente_o_efeito_nunca_o_ciclo(wb_referencia):
    grade = dict(
        ((comp.year, comp.month), efeito) for comp, efeito in _grade(wb_referencia)
    )
    # Caso de referencia (secao 15): meses anteriores a H = Nao; desde H = Sim.
    for chave in [(2024, 4), (2024, 5), (2025, 4), (2025, 5), (2026, 4), (2026, 5)]:
        assert grade[chave] == "Nao", chave
    for chave in [(2024, 6), (2025, 3), (2025, 6), (2026, 3), (2026, 6), (2027, 3)]:
        assert grade[chave] == "Sim", chave
    # C0/C1 (fora da analise): sempre Nao.
    assert grade[(2022, 4)] == "Nao" and grade[(2024, 3)] == "Nao"


def test_financeiro_ciclo_cronologico_via_leitor(wb_referencia):
    """O espelho Python da formula B enquadra pelo bloco fixo, inclusive nos
    meses do gap das janelas (04-05/2025 -> C3 cronologico)."""
    casos = [
        (date(2022, 4, 1), "C0"), (date(2024, 4, 1), "C2"),
        (date(2025, 3, 1), "C2"), (date(2025, 4, 1), "C3"),
        (date(2025, 5, 1), "C3"), (date(2025, 6, 1), "C3"),
        (date(2026, 4, 1), "C4"), (date(2027, 3, 1), "C4"),
    ]
    for comp, esperado in casos:
        assert _ciclo_por_competencia(wb_referencia, comp) == esperado, comp


def test_formula_b_do_financeiro_usa_cronologia_e_nao_janelas(wb_referencia):
    formula = wb_referencia["financeiro"]["B2"].value
    assert formula.startswith("=")
    # bloco fixo ancorado no inicio cronologico de C0
    assert "parametros!$C$2" in formula
    # nenhuma referencia as demais janelas (C3..C6/D2..D6)
    for ref in ("$C$3", "$C$4", "$C$5", "$C$6", "$D$2", "$D$3", "$D$4", "$D$5", "$D$6"):
        assert ref not in formula, ref
    assert '"Fora dos ciclos"' in formula


def test_marcacao_visual_dos_meses_sem_efeito_ja_nasce_no_template(wb_referencia):
    """Regra petrea visual (secao 14): G="Nao" dispara o rosa FFC7CE do
    template (formatacao condicional homologada), sem acao do fiscal."""
    fin = wb_referencia["financeiro"]
    regras = []
    for cf in fin.conditional_formatting:
        for regra in cf.rules:
            formulas = list(getattr(regra, "formula", []) or [])
            cor = None
            if regra.dxf is not None and regra.dxf.fill is not None:
                cor = getattr(regra.dxf.fill.bgColor, "rgb", None)
            regras.append((" ".join(str(f) for f in formulas), cor))
    assert any(
        '$G2="Nao"' in formula and cor == "FFFFC7CE"
        for formula, cor in regras
    ), regras


# --------------------------------------------------------------------------- #
# VALOR — efeito Nao nunca exclui a base; zera somente o novo delta
# --------------------------------------------------------------------------- #
def test_leitor_financeiro_nao_exclui_base_dos_meses_sem_efeito():
    """Protecao permanente (secoes 16/35): valores em DOIS meses antes de H e
    UM mes desde H — os TRES participam da execucao; so o ultimo produz o
    novo delta."""
    wb = load_workbook(
        io.BytesIO(gerar_coleta_oficial_preenchida(_payload_caso_referencia())),
        data_only=False,
    )
    fin = wb["financeiro"]
    linhas = {}
    for r in range(2, 74):
        comp = fin[f"A{r}"].value
        if comp is None:
            continue
        chave = (_dia(comp).year, _dia(comp).month)
        if chave in ((2024, 4), (2024, 5), (2024, 6)):
            fin[f"C{r}"].value = 100.0
            linhas[chave] = r
    # delta cacheado apenas no mes com efeito (como o Excel gravaria)
    fin[f"F{linhas[(2024, 6)]}"].value = 5.0
    parcelas = _ler_parcelas_sombra_financeiro(wb)
    wb.close()
    bases = [p for p in parcelas if p["tipo_financeiro"] == "Execucao Atualizada"]
    retroativas = [
        p for p in parcelas if p["tipo_financeiro"] == "Retroativo Reconhecido"
    ]
    # os TRES meses entram na base (300,00) — efeito Nao nao exclui valor
    assert len(bases) == 3
    assert sum(float(p["valor"]) for p in bases) == pytest.approx(300.0)
    assert all("G nao exclui base" in p["justificativa_vta"] for p in bases)
    # e o novo delta existe SOMENTE para o mes com efeito Sim
    assert len(retroativas) == 1
    assert retroativas[0]["valor"] == pytest.approx(5.0)


def test_pc_antes_do_efeito_conta_no_vta_com_delta_zero():
    """Teste economico dirigido — PC (secoes 23/36)."""
    por_ciclo = {
        "C0": {"data_inicio": date(2022, 4, 1), "data_fim": date(2023, 3, 31)},
        "C1": {"data_inicio": date(2023, 4, 1), "data_fim": date(2024, 3, 31)},
        "C2": {"data_inicio": date(2024, 4, 1), "data_fim": date(2025, 3, 31),
               "computar_nesta_apuracao": "Sim", "percentual_reajuste": 0.05,
               "fator_acumulado": 1.05,
               "inicio_efeito_financeiro": date(2024, 6, 1)},
        "C3": {"data_inicio": date(2025, 6, 1), "data_fim": date(2026, 5, 31),
               "computar_nesta_apuracao": "Sim", "percentual_reajuste": 0.04,
               "fator_acumulado": 1.092,
               "inicio_efeito_financeiro": date(2025, 6, 1)},
        "C4": {"data_inicio": date(2026, 6, 1), "data_fim": date(2027, 5, 31),
               "computar_nesta_apuracao": "Sim", "percentual_reajuste": 0.03,
               "fator_acumulado": 1.1248,
               "inicio_efeito_financeiro": date(2026, 6, 1)},
    }
    itens_pc = [
        # PC no gap das janelas (C3 cronologico), ANTES de H(C3)=06/2025
        {"numero_pc": "PC-GAP", "data_pc": date(2025, 4, 20),
         "valor_pc": 1000.0, "ciclo": "", "pc_pago_a_contratada": "Sim"},
        # PC apos H(C3): novo delta normal
        {"numero_pc": "PC-EFEITO", "data_pc": date(2025, 6, 20),
         "valor_pc": 1000.0, "ciclo": "", "pc_pago_a_contratada": "Sim"},
    ]
    resultado = montar_motor_temporal({
        "ciclos": {"por_ciclo": por_ciclo},
        "itens_pc": itens_pc,
    })
    por_numero = {pc.numero_pc: pc for pc in resultado.pcs}
    antes, depois = por_numero["PC-GAP"], por_numero["PC-EFEITO"]
    # Ambos enquadrados no C3 cronologico — nenhum "fora dos ciclos".
    assert antes.ciclo_temporal == "C3"
    assert depois.ciclo_temporal == "C3"
    # PC antes de H: valor conta integralmente; novo delta/retroativo = 0.
    assert antes.efeito_financeiro_pc == "Nao"
    assert antes.valor_pc == pytest.approx(1000.0)
    assert antes.fator_aplicado == pytest.approx(1.0)   # preserva o vigente
    assert antes.retroativo == pytest.approx(0.0)
    # PC desde H: valor conta e o novo delta e calculado.
    assert depois.efeito_financeiro_pc == "Sim"
    assert depois.fator_aplicado == pytest.approx(1.092)
    assert depois.retroativo == pytest.approx(92.0)
    # O VTA/executado considera os DOIS valores.
    assert resultado.totais["valor_pc"] == pytest.approx(2000.0)


def test_aditivo_no_gap_das_janelas_continua_valido():
    """Teste dirigido — Aditivos (secoes 24/25/38): 30/04, 31/05 e 01/06 de
    2025 sao fatos validos; os dois primeiros caem no gap das janelas e
    permanecem enquadrados no C3 cronologico."""
    janelas = [
        {"ciclo": "C0", "data_inicio": date(2022, 4, 1), "data_fim": date(2023, 3, 31)},
        {"ciclo": "C1", "data_inicio": date(2023, 4, 1), "data_fim": date(2024, 3, 31)},
        {"ciclo": "C2", "data_inicio": date(2024, 4, 1), "data_fim": date(2025, 3, 31)},
        {"ciclo": "C3", "data_inicio": date(2025, 6, 1), "data_fim": date(2026, 5, 31)},
        {"ciclo": "C4", "data_inicio": date(2026, 6, 1), "data_fim": date(2027, 5, 31)},
    ]
    execucao = list(calendario_execucao_por_ciclo(janelas).values())
    for dia in (date(2025, 4, 30), date(2025, 5, 31), date(2025, 6, 1)):
        resultado = determinar_ciclo_por_data(dia, execucao)
        assert resultado.ciclo == "C3", dia


# --------------------------------------------------------------------------- #
# 31.1 — Fator DESTA analise (CONTROLE!B11) x Fator HISTORICO (RESULTADOS!H5)
# --------------------------------------------------------------------------- #
def test_fator_historico_desacoplado_do_fator_da_analise(wb_referencia):
    """Etapa 31.1: CONTROLE!B11 = fator DESTA analise; RESULTADOS!H5 = fator
    HISTORICO integral (logica fail-closed que vivia em B11 antes da etapa).
    Nenhum consumidor do historico aponta mais para CONTROLE!B11."""
    res = wb_referencia["RESULTADOS"]
    cv = wb_referencia["comparativo_VTA"]
    ctl = wb_referencia["CONTROLE"]
    # CONTROLE preservado (fator da analise = 1 + B10).
    assert ctl["B11"].value == '=IF(ISNUMBER(B10),1+B10,"")'
    # H5 = logica historica movida (ciclo vigente + cadeia parametros!E/F),
    # sem qualquer referencia ao novo B11.
    h5 = str(res["H5"].value)
    assert "CONTROLE!$B$2" in h5
    assert "parametros!$F$6" in h5 and "COUNT(parametros!$E$3:$E$6)" in h5
    assert "B$11" not in h5
    # H8 valida $H$5 (mantendo as demais condicoes) e nao mais B11.
    h8 = str(res["H8"].value)
    assert "$H$5" in h8 and "B$11" not in h8
    assert 'SEARCH("CALCULADO",MEMORIA_RESULTADOS!$E$26)' in h8
    # B208 = contrato original x FATOR HISTORICO (RESULTADOS!$H$5).
    b208 = str(cv["B208"].value)
    assert "RESULTADOS!$H$5" in b208 and "B$11" not in b208
    # Texto auditavel cita a nova fonte real.
    c12 = str(res["C12"].value)
    assert "RESULTADOS!H5" in c12 and "CONTROLE!B11" not in c12
    # D6 continua derivando de H5 (variacao historica integral).
    assert str(res["D6"].value) == '=IF($H$5="","",$H$5-1)'


def test_historico_completo_alimenta_h5_e_analise_alimenta_b10():
    """Secao 12 do gate 31.1: historico C1=5% + analise C2=10%.

    CONTROLE!B10 = 10% (so a analise); RESULTADOS!H5 depende da cadeia
    historica completa (parametros!E3:E4 preenchidos -> F4 = 1,05 x 1,10)."""
    payload = {
        "origem": "Reajustes Múltiplos", "indice": "IPCA",
        "data_base_original": "01/04/2023",
        "ciclos": [
            {"ciclo": "C1", "data_base": "01/04/2023",
             "data_pedido": "01/04/2024", "percentual_aplicado": 0.05,
             "financeiro_inicio": "01/04/2024",
             "objeto_analise_atual": False, "ciclo_ja_concedido": True,
             "situacao": "Concedido historicamente"},
            {"ciclo": "C2", "data_pedido": "01/04/2025",
             "percentual_aplicado": 0.10, "financeiro_inicio": "01/04/2025",
             "objeto_analise_atual": True, "situacao": "✅ TEMPESTIVO"},
        ],
    }
    wb = load_workbook(
        io.BytesIO(gerar_coleta_oficial_preenchida(payload)), data_only=False
    )
    ctl = wb["CONTROLE"]
    par = wb["parametros"]
    # Analise atual: somente C2 -> B10 = 10% e B11 = 1 + B10.
    assert ctl["B10"].value == pytest.approx(0.10)
    assert ctl["B11"].value == '=IF(ISNUMBER(B10),1+B10,"")'
    # Cadeia historica disponivel para H5: E3 (C1) e E4 (C2) preenchidos.
    assert par["E3"].value == pytest.approx(0.05)
    assert par["E4"].value == pytest.approx(0.10)
    # Ciclo vigente C2: H5 resolvera para parametros!F4 (1,05 x 1,10)
    # somente porque COUNT(E3:E4) = 2 — logica fail-closed preservada.
    assert ctl["B2"].value == "C2"
    assert "parametros!$F$4" in str(wb["RESULTADOS"]["H5"].value)
    wb.close()


def test_historico_ausente_deixa_h5_fail_closed(wb_referencia):
    """Secao 13 do gate 31.1: analise C2+C3+C4 sem C1 -> a cadeia historica
    esta incompleta (parametros!E3 vazio) e H5/B208 nao podem apresentar o
    fator apenas da analise como se fosse o historico integral."""
    par = wb_referencia["parametros"]
    ctl = wb_referencia["CONTROLE"]
    # C1 fora da analise: E3 vazio -> COUNT(E3:E6) < 4 -> H5 = "" no Excel.
    assert par["E3"].value in (None, "")
    # B10/B11 desta analise continuam preenchidos (nao regridem).
    assert ctl["B10"].value == pytest.approx(0.1237)
    assert ctl["B11"].value == '=IF(ISNUMBER(B10),1+B10,"")'


def test_consumidos_sem_granularidade_temporal_documentada(wb_referencia):
    """Secao 27: itens_Consumidos so tem baldes agregados por ciclo (sem
    data/competencia). O gate por H e estruturalmente inaplicavel a este
    metodo — nao ratear, nao presumir. Este teste trava a premissa
    documentada."""
    ws = wb_referencia["itens_Consumidos"]
    cabecalhos = [str(c.value or "") for c in ws[1]]
    assert any("QTD_CONS_C0" in c for c in cabecalhos)
    assert not any("DATA" in c.upper() or "COMPETENCIA" in c.upper()
                   for c in cabecalhos)
