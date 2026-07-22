"""Etapa 4 — Adequacao Orcamentaria: testes unitarios, golden-master e COM.

Unitarios cobrem as regras do motor com cenarios sinteticos. Os testes golden e
COM leem o golden normativo externo
`C:\\_DesktopReal\\23.Xls\\10.adequacao_orcamentaria_v4_.xlsx` (nao versionado):
sao pulados quando o arquivo nao esta disponivel. A paridade monetaria final
Python x Excel deve ser R$ 0,00.
"""
from __future__ import annotations

import os
from datetime import date
from pathlib import Path

import pytest

from _adequacao_orcamentaria import (
    Pedido,
    calcular_adequacao_orcamentaria as calc,
    media_pedidos_compra,
    _round2,
)

GOLDEN = Path(r"C:\_DesktopReal\23.Xls\10.adequacao_orcamentaria_v4_.xlsx")


def _pcs(*trios):
    return [Pedido(identificacao=i, data=d, valor=v) for i, d, v in trios]


# ---------------------------------------------------------------- arredondamento (AA)

def test_round2_meia_para_cima_como_excel():
    assert _round2(2.675) == 2.68
    assert _round2(0.125) == 0.13
    assert _round2(1.005) == 1.01
    assert _round2(1718.263153) == 1718.26
    assert _round2(None) == 0.0


# ---------------------------------------------------------------- media / janela

def test_media_pcs_divide_pela_janela_com_meses_zero():  # E, H
    base = media_pedidos_compra(
        _pcs(("a", date(2024, 3, 10), 200.0), ("b", date(2024, 7, 5), 400.0)),
        date(2024, 12, 1), 12)
    assert base["total_historico"] == 600.0
    assert base["media_mensal"] == 50.0
    assert base["meses_com_pedido"] == 2
    assert base["meses_sem_pedido"] == 10


def test_pc_excluido_nao_entra():  # F
    base = media_pedidos_compra(
        [Pedido("a", date(2024, 3, 10), 200.0, considerar=True),
         Pedido("b", date(2024, 4, 10), 400.0, considerar=False)],
        date(2024, 12, 1), 12)
    assert base["total_historico"] == 200.0
    assert base["pedidos_considerados"] == 1


def test_varios_pcs_no_mesmo_mes_somam():  # G
    base = media_pedidos_compra(
        _pcs(("a", date(2024, 3, 5), 100.0), ("b", date(2024, 3, 20), 50.0)),
        date(2024, 12, 1), 12)
    assert base["total_historico"] == 150.0
    assert base["meses_com_pedido"] == 1


def test_janela_1_mes():  # C
    base = media_pedidos_compra(
        _pcs(("a", date(2024, 12, 10), 300.0), ("fora", date(2024, 11, 10), 999.0)),
        date(2024, 12, 1), 1)
    assert base["total_historico"] == 300.0
    assert base["media_mensal"] == 300.0
    assert base["inicio_janela"] == date(2024, 12, 1)
    assert base["fim_janela"] == date(2024, 12, 31)


def test_janela_60_meses():  # D
    base = media_pedidos_compra(_pcs(("a", date(2022, 1, 15), 6000.0)), date(2024, 12, 1), 60)
    assert base["inicio_janela"] == date(2020, 1, 1)
    assert base["fim_janela"] == date(2024, 12, 31)
    assert base["media_mensal"] == 100.0


def test_origem_financeiro_media_simples():  # A
    r = calc(origem="Financeiro mensal", percentual=0.10,
             ultima_competencia=date(2024, 12, 1), data_fim_vigencia=date(2025, 3, 15),
             retroativo=0.0, financeiro_mensal=[100.0, 200.0, 300.0])
    assert r["media_mensal"] == 200.0
    assert r["origem"] == "financeiro"


# ---------------------------------------------------------------- cenario base PCs

def _cenario(**kw):
    d = dict(origem="Pedidos de compra", percentual=0.10,
             ultima_competencia=date(2024, 12, 1), data_fim_vigencia=date(2025, 3, 15),
             retroativo=1000.0, janela_meses=12, saldo_contratual=None,
             pedidos=_pcs(("a", date(2024, 3, 10), 600.0), ("b", date(2024, 6, 10), 600.0)))
    d.update(kw)
    return calc(**d)


def test_cenario_base_projecao_e_complemento():  # B, K, N, W
    r = _cenario()
    assert r["media_mensal"] == 100.0
    assert r["referencia_reajustada"] == 110.0
    assert r["meses_projetados"] == 3          # jan, fev, mar/2025 (mar integral)
    assert r["diferenca_futura"] == 30.0       # 3 x 10
    assert r["complemento_estimado"] == 1030.0
    assert [m["competencia"] for m in r["memoria_mensal"]] == [
        date(2025, 1, 1), date(2025, 2, 1), date(2025, 3, 1)]


def test_percentual_zero():  # J
    r = _cenario(percentual=0.0)
    assert r["referencia_reajustada"] == 100.0
    assert r["diferenca_futura"] == 0.0
    assert r["complemento_estimado"] == 1000.0  # so o retroativo


def test_retroativo_zero():  # V
    r = _cenario(retroativo=0.0)
    assert r["complemento_estimado"] == 30.0


def test_override_vazio_usa_automatico():  # O
    r = _cenario(overrides={date(2025, 1, 1): {"valor": None}})
    assert r["memoria_mensal"][0]["base_considerada"] == 100.0
    assert r["memoria_mensal"][0]["situacao"] == "Projecao automatica"


def test_override_zero_sem_execucao():  # P
    r = _cenario(overrides={date(2025, 1, 1): {"valor": 0.0}})
    m = r["memoria_mensal"][0]
    assert m["base_considerada"] == 0.0
    assert m["diferenca"] == 0.0
    assert m["situacao"] == "Valor informado pelo fiscal"
    assert r["diferenca_futura"] == 20.0


def test_override_positivo():  # Q
    r = _cenario(overrides={date(2025, 1, 1): {"valor": 200.0}})
    m = r["memoria_mensal"][0]
    assert m["base_considerada"] == 200.0
    assert m["valor_reajustado"] == 220.0
    assert m["diferenca"] == 20.0


def test_override_ja_reajustado_converte_para_base():  # R
    r = _cenario(overrides={date(2025, 1, 1): {"valor": 110.0, "ja_reajustado": True}})
    m = r["memoria_mensal"][0]
    assert abs(m["base_considerada"] - 100.0) < 1e-9
    assert m["diferenca"] == 10.0


def test_saldo_vazio_sem_cap():  # S
    r = _cenario(saldo_contratual=None)
    assert all(m["base_considerada"] == 100.0 for m in r["memoria_mensal"])


def test_saldo_suficiente():  # T
    r = _cenario(saldo_contratual=100000.0)
    assert r["diferenca_futura"] == 30.0


def test_saldo_limitante():  # U
    r = _cenario(saldo_contratual=150.0)
    bases = [m["base_considerada"] for m in r["memoria_mensal"]]
    assert bases == [100.0, 50.0, 0.0]
    assert any(m["situacao"] == "Limitado ao saldo" for m in r["memoria_mensal"])


def test_projecao_cruza_dois_exercicios():  # L, X, Y
    # PCs dentro da janela 12 (dez/2024..nov/2025) para manter media = 100
    r = _cenario(ultima_competencia=date(2025, 11, 1), data_fim_vigencia=date(2026, 2, 15),
                 pedidos=_pcs(("a", date(2025, 1, 10), 600.0), ("b", date(2025, 6, 10), 600.0)))
    assert r["meses_projetados"] == 3  # dez/2025, jan/2026, fev/2026
    progs = {p["exercicio"]: round(p["valor"], 2) for p in r["programacao_por_exercicio"]}
    assert progs == {2025: 1010.0, 2026: 20.0}
    assert abs(r["soma_programacao"] - r["complemento_estimado"]) < 0.005


def test_termino_no_mesmo_exercicio():  # M
    r = _cenario()
    assert [p["exercicio"] for p in r["programacao_por_exercicio"]] == [2025]


def test_data_final_anterior_a_projecao():  # AC
    r = _cenario(data_fim_vigencia=date(2024, 12, 15))
    assert r["meses_projetados"] == 0
    assert any("SEM MESES DE PROJECAO" in c for c in r["checks"])


def test_janela_invalida_gera_check():  # AB
    r = _cenario(janela_meses=99)
    assert any("JANELA FORA DE 1..60" in c for c in r["checks"])


def test_nenhum_pc_considerado():  # AD, AE
    r = _cenario(pedidos=[])
    assert r["media_mensal"] == 0.0
    assert any("NENHUM PC CONSIDERADO" in c for c in r["checks"])


def test_estrutura_serializavel_completa():  # estrutura minima (Secao 14)
    r = _cenario()
    for chave in ("origem", "percentual", "total_historico", "media_mensal",
                  "referencia_reajustada", "retroativo", "diferenca_futura",
                  "complemento_estimado", "programacao_por_exercicio",
                  "memoria_mensal", "status", "base_futura", "saldo_contratual"):
        assert chave in r


# ---------------------------------------------------------------- golden-master (Z)

def _inputs_do_golden():
    import openpyxl
    f = openpyxl.load_workbook(GOLDEN, data_only=False)
    R, PC = f["RESUMO"], f["PEDIDOS_COMPRA"]
    peds = []
    for row in range(5, 105):
        b = PC.cell(row, 2).value
        if b in (None, ""):
            continue
        peds.append(Pedido.de_dict({"id": PC.cell(row, 1).value, "data": b,
                                    "valor": PC.cell(row, 3).value,
                                    "considerar": PC.cell(row, 4).value}))
    return dict(
        origem="Pedidos de compra", percentual=R["B7"].value,
        ultima_competencia=R["B8"].value, data_fim_vigencia=R["B9"].value,
        retroativo=R["B10"].value, janela_meses=R["B11"].value,
        saldo_contratual=R["B12"].value, pedidos=peds,
    )


@pytest.mark.skipif(not GOLDEN.exists(), reason="golden externo ausente")
def test_golden_master():
    r = calc(**_inputs_do_golden())
    assert round(r["total_historico"], 3) == 557972.097
    assert abs(r["media_mensal"] - 14306.976846153848) < 1e-9
    assert r["referencia_reajustada"] == 16025.24
    assert round(r["diferenca_futura"], 2) == 18900.86
    assert round(r["retroativo"], 2) == 16888.59
    assert round(r["complemento_estimado"], 2) == 35789.45
    assert r["base_historica"]["pedidos_considerados"] == 45
    assert r["base_historica"]["meses_com_pedido"] == 6
    assert r["base_historica"]["meses_sem_pedido"] == 33
    assert r["meses_projetados"] == 11
    progs = {p["exercicio"]: round(p["valor"], 2) for p in r["programacao_por_exercicio"]}
    assert progs == {2026: 27198.15, 2027: 8591.30}


# ---------------------------------------------------------------- Excel COM (AF)

@pytest.mark.skipif(
    os.environ.get("RUN_EXCEL_INTEGRATION") != "1" or not GOLDEN.exists(),
    reason="defina RUN_EXCEL_INTEGRATION=1 e disponibilize o golden",
)
def test_paridade_excel_com():
    """Recalcula o golden no Excel real e compara com o motor Python (diff = 0,00)."""
    import gc
    import pythoncom
    import win32com.client

    pythoncom.CoInitialize()
    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    wb = None
    try:
        wb = excel.Workbooks.Open(str(GOLDEN.resolve()), UpdateLinks=0, ReadOnly=True)
        excel.CalculateFullRebuild()
        R = wb.Worksheets("RESUMO")
        T = wb.Worksheets("TEXTO")
        xls = {
            "dif_futura": R.Range("F6").Value,
            "retroativo": R.Range("F7").Value,
            "complemento": R.Range("F8").Value,
            "ref_reaj": R.Range("B18").Value,
            "media": R.Range("B17").Value,
            "prog1": T.Range("B10").Value,
            "prog2": T.Range("B11").Value,
        }
    finally:
        if wb is not None:
            wb.Close(SaveChanges=False)
        excel.Quit()
        gc.collect()
        pythoncom.CoUninitialize()

    r = calc(**_inputs_do_golden())
    progs = {p["exercicio"]: p["valor"] for p in r["programacao_por_exercicio"]}
    pares = [
        (r["diferenca_futura"], xls["dif_futura"]),
        (r["retroativo"], xls["retroativo"]),
        (r["complemento_estimado"], xls["complemento"]),
        (r["referencia_reajustada"], xls["ref_reaj"]),
        (r["media_mensal"], xls["media"]),
        (progs.get(2026), xls["prog1"]),
        (progs.get(2027), xls["prog2"]),
    ]
    for py, xl in pares:
        assert round(float(py), 2) == round(float(xl), 2), f"Python {py} x Excel {xl}"


# ---------------------------------------------------------------- regressao golden VERSIONADA (Secao 13)
# Snapshot numerico do cenario golden auditado (10.adequacao_orcamentaria_v4_.xlsx),
# embutido no Git para proteger a matematica na CI mesmo sem o arquivo externo.
# NAO substitui test_golden_master / test_paridade_excel_com (camada externa).
PEDIDOS_GOLDEN = [
    ("2024-03-11", 22346.28), ("2024-03-11", 14098.85), ("2024-03-11", 12000.78),
    ("2024-03-11", 12000.78), ("2024-03-11", 12000.78), ("2024-03-11", 13035.33),
    ("2024-03-15", 11704.9), ("2024-03-11", 11704.9), ("2024-03-15", 2393.95),
    ("2024-05-03", 11946.98), ("2024-05-03", 11946.98), ("2024-05-03", 14276.79),
    ("2024-05-03", 11946.98), ("2024-08-14", 14583.01), ("2024-05-03", 14276.79),
    ("2024-05-03", 14276.79), ("2024-05-03", 14276.79), ("2024-05-03", 11946.98),
    ("2024-05-03", 12000.78), ("2024-05-03", 12000.78), ("2024-06-07", 12000.78),
    ("2024-06-07", 12000.78), ("2024-05-03", 11946.98), ("2024-05-03", 14276.79),
    ("2024-05-03", 14276.79), ("2024-05-03", 14276.79), ("2024-05-03", 14276.79),
    ("2024-06-07", 12000.78), ("2024-05-03", 12000.78), ("2024-05-03", 13035.33),
    ("2024-05-03", 14276.79), ("2024-05-03", 14276.79), ("2025-04-10", 19242.63),
    ("2025-04-10", 4965.84), ("2025-04-10", 2636.03), ("2025-04-10", 3724.38),
    ("2025-04-11", 10283.427), ("2025-04-10", 14583.01), ("2025-04-10", 16759.71),
    ("2025-04-10", 16759.71), ("2025-04-10", 4965.84), ("2026-03-12", 14636.81),
    ("2026-03-13", 12000.78), ("2026-03-13", 12000.78), ("2026-03-13", 12000.78),
]


def _pedidos_golden():
    from datetime import datetime as _dt
    return [Pedido(identificacao=f"PC-{n}",
                   data=_dt.strptime(d, "%Y-%m-%d").date(), valor=v)
            for n, (d, v) in enumerate(PEDIDOS_GOLDEN)]


def _resultado_golden():
    return calc(origem="Pedidos de compra", percentual=0.1201,
                ultima_competencia=date(2026, 6, 1), data_fim_vigencia=date(2027, 5, 5),
                retroativo=16888.59, janela_meses=39, saldo_contratual=911237.89,
                pedidos=_pedidos_golden())


def test_regressao_golden_versionada():
    r = _resultado_golden()
    assert round(r["total_historico"], 3) == 557972.097
    assert abs(r["media_mensal"] - 14306.976846153848) < 1e-9
    assert r["referencia_reajustada"] == 16025.24
    assert round(r["diferenca_futura"], 2) == 18900.86
    assert round(r["complemento_estimado"], 2) == 35789.45
    assert r["base_historica"]["pedidos_considerados"] == 45
    assert r["base_historica"]["meses_com_pedido"] == 6
    assert r["base_historica"]["meses_sem_pedido"] == 33
    assert r["meses_projetados"] == 11
    progs = {p["exercicio"]: round(p["valor"], 2) for p in r["programacao_por_exercicio"]}
    assert progs == {2026: 27198.15, 2027: 8591.30}


# ---------------------------------------------------------------- itens_PC -> motor (wiring)

def test_pedidos_de_itens_pc_mapeia_sem_redigitacao():  # H, C/D
    from _adequacao_orcamentaria import pedidos_de_itens_pc
    registros = [
        {"numero_pc": "PC-1", "data_pc": date(2024, 3, 10), "valor_pc": 100.0},
        {"NUMERO_PC": "PC-2", "DATA_PC": date(2024, 4, 10), "VALOR_PC": 200.0},
    ]
    peds = pedidos_de_itens_pc(registros)
    assert [p.identificacao for p in peds] == ["PC-1", "PC-2"]
    assert [p.valor for p in peds] == [100.0, 200.0]
    assert all(p.considerar for p in peds)


def test_pedidos_de_itens_pc_exclusao_especifica():  # H
    from _adequacao_orcamentaria import pedidos_de_itens_pc
    registros = [
        {"numero_pc": "PC-1", "data_pc": date(2024, 3, 10), "valor_pc": 100.0},
        {"numero_pc": "PC-2", "data_pc": date(2024, 4, 10), "valor_pc": 200.0},
    ]
    peds = pedidos_de_itens_pc(registros, exclusoes=["PC-2"])
    assert peds[0].considerar is True
    assert peds[1].considerar is False


def test_classificar_pedidos_situacao_na_janela():  # F, G, L
    from _adequacao_orcamentaria import classificar_pedidos, Pedido as P
    peds = [P("dentro", date(2024, 6, 1), 100.0),
            P("fora", date(2020, 1, 1), 100.0),
            P("excluido", date(2024, 6, 1), 100.0, considerar=False)]
    cl = classificar_pedidos(peds, date(2024, 12, 1), 12)
    sit = {x["identificacao"]: x["situacao"] for x in cl["pedidos"]}
    assert sit == {"dentro": "Considerado", "fora": "Fora da janela", "excluido": "Excluido"}


def test_wiring_golden_via_itens_pc():  # C, O, P — end-to-end pela camada de dados
    from _adequacao_orcamentaria import pedidos_de_itens_pc
    from datetime import datetime as _dt
    registros = [{"numero_pc": f"PC-{n}", "data_pc": _dt.strptime(d, "%Y-%m-%d").date(),
                  "valor_pc": v} for n, (d, v) in enumerate(PEDIDOS_GOLDEN)]
    peds = pedidos_de_itens_pc(registros)
    r = calc(origem="Pedidos de compra", percentual=0.1201,
             ultima_competencia=date(2026, 6, 1), data_fim_vigencia=date(2027, 5, 5),
             retroativo=16888.59, janela_meses=39, saldo_contratual=911237.89, pedidos=peds)
    assert r["referencia_reajustada"] == 16025.24
    assert round(r["complemento_estimado"], 2) == 35789.45
    progs = {p["exercicio"]: round(p["valor"], 2) for p in r["programacao_por_exercicio"]}
    assert progs == {2026: 27198.15, 2027: 8591.30}


def test_todos_pcs_fora_da_janela():  # L
    from _adequacao_orcamentaria import pedidos_de_itens_pc
    registros = [{"numero_pc": "x", "data_pc": date(2019, 1, 1), "valor_pc": 500.0}]
    peds = pedidos_de_itens_pc(registros)
    r = calc(origem="Pedidos de compra", percentual=0.10, ultima_competencia=date(2024, 12, 1),
             data_fim_vigencia=date(2025, 3, 15), janela_meses=12, pedidos=peds)
    assert r["base_historica"]["pedidos_considerados"] == 0
    assert r["media_mensal"] == 0.0


# ---------------------------------------------------------------- pagina (wiring estatico, Secao 12)
# Paginas Streamlit executam ao importar; a verificacao e estatica (le a fonte),
# como nos demais testes de pagina do projeto.
PAGINA = (Path(__file__).resolve().parents[1] / "pages" / "12_Adequacao_Orcamentaria.py").read_text(encoding="utf-8")


def test_pagina_importa_o_motor_unico():  # E (motor), X (sem 2o motor)
    assert "from _adequacao_orcamentaria import" in PAGINA
    assert "_round2" in PAGINA and "media_pedidos_compra" in PAGINA
    # a pagina nao redefine a matematica do motor
    assert "def calcular_adequacao_orcamentaria" not in PAGINA
    assert "def media_pedidos_compra" not in PAGINA


def test_pagina_oferece_duas_origens():  # A, B, M, N
    assert 'st.radio(' in PAGINA and '"Financeiro"' in PAGINA and '"Pedidos de compra"' in PAGINA
    assert 'key="adequacao_v2_origem"' in PAGINA  # estado estavel na alternancia


def test_pagina_reutiliza_itens_pc_sem_redigitacao():  # C, D, H
    assert "carregar_itens_pc_da_sessao" in PAGINA
    assert "pedidos_de_itens_pc(" in PAGINA
    assert "itens_pc_v10" in PAGINA
    assert 'st.multiselect(' in PAGINA  # exclusoes explicitas (Considerar/Excluir)


def test_pagina_expoe_janela_e_resumo():  # J (janela), K resumo PCs
    assert 'st.slider("Janela histórica dos pedidos (meses)", 1, 60' in PAGINA
    for card in ("Período histórico", "PCs considerados", "Meses com PCs",
                 "Meses sem PCs", "Total histórico", "Média mensal (PCs)"):
        assert card in PAGINA


def test_pagina_trata_ausencia_de_pcs():  # K
    assert "NÃO HÁ PEDIDOS DE COMPRA DISPONÍVEIS" in PAGINA


def test_pagina_janela_soberana_sem_mover():  # L
    assert "0 PCs considerados na janela" in PAGINA
    assert "não é movida para encaixar pedidos" in PAGINA


def test_pagina_export_identifica_origem():  # Q
    assert '("Origem histórica", origem_hist_rotulo)' in PAGINA


def test_pagina_projecao_usa_referencia_da_origem():  # G/O — media_ref comum as origens
    assert "montar_base_editor(periodos, media_ref)" in PAGINA
    assert "calcular_projecao(df_editor, media_ref, fator_reajuste)" in PAGINA
