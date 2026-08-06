from __future__ import annotations

import copy
import io
from datetime import date

import pytest
from openpyxl import Workbook, load_workbook

from _ciclo_em_execucao import (
    ABA_CICLO_EM_EXECUCAO,
    calcular_posicao_ciclo_por_data,
    garantir_aba_ciclo_em_execucao,
    ler_ciclo_em_execucao,
)
from _coleta_oficial import TEMPLATE_COLETA_OFICIAL, obter_coleta_oficial_bytes
from _leitor_masterfile_v10 import ler_masterfile_v10


T0 = date(2026, 5, 1)
FIM = date(2027, 4, 30)
D = date(2026, 7, 31)


def _legado_sentela(wb: Workbook) -> dict[str, object]:
    memoria = wb.create_sheet("MEMORIA_RESULTADOS")
    memoria["B16"] = 40.0
    memoria["B26"] = 10_500.0
    memoria["C35"] = 8_000.0
    memoria["D35"] = 8_400.0
    resultados = wb.create_sheet("RESULTADOS")
    resultados["B26"] = 10_500.0
    historico = wb.create_sheet("historico_VU")
    historico["F2"] = 100.0
    return {
        "retroativo": memoria["B16"].value,
        "vta": memoria["B26"].value,
        "remanescente_base": memoria["C35"].value,
        "remanescente_atualizado": memoria["D35"].value,
        "resultado_consolidado": resultados["B26"].value,
        "vu": historico["F2"].value,
    }


def _wb_local() -> tuple[Workbook, dict[str, object]]:
    wb = Workbook()
    wb.active.title = "posicao_referencia"
    ctrl = wb.create_sheet("CONTROLE")
    ctrl["B2"] = "C3"
    parametros = wb.create_sheet("parametros")
    parametros["B5"], parametros["C5"], parametros["D5"] = "C3", T0, FIM
    legado = _legado_sentela(wb)
    garantir_aba_ciclo_em_execucao(wb)
    ws = wb[ABA_CICLO_EM_EXECUCAO]
    ws["C3"], ws["F3"], ws["H3"] = "C3", T0, FIM
    return wb, legado


def _linha_valida(ws, linha: int = 13, *, item: str = "ITEM-1", atual=65, vu=100):
    ws.cell(linha, 1).value = item
    ws.cell(linha, 2).value = 80
    ws.cell(linha, 3).value = atual
    ws.cell(linha, 4).value = None if atual in (None, "") else 80 - atual
    ws.cell(linha, 5).value = vu
    ws.cell(linha, 6).value = None if atual in (None, "") or vu is None else (80 - atual) * vu
    ws.cell(linha, 7).value = None if atual in (None, "") or vu is None else atual * vu
    ws.cell(linha, 9).value = 0
    ws.cell(linha, 10).value = None if atual in (None, "") else 0
    ws.cell(linha, 11).value = "OK"


def _snapshot_legado(wb: Workbook) -> dict[str, object]:
    return {
        "retroativo": wb["MEMORIA_RESULTADOS"]["B16"].value,
        "vta": wb["MEMORIA_RESULTADOS"]["B26"].value,
        "remanescente_base": wb["MEMORIA_RESULTADOS"]["C35"].value,
        "remanescente_atualizado": wb["MEMORIA_RESULTADOS"]["D35"].value,
        "resultado_consolidado": wb["RESULTADOS"]["B26"].value,
        "vu": wb["historico_VU"]["F2"].value,
    }


def test_a_arquivo_antigo_sem_aba_e_aceito_sem_alerta_local():
    resultado = ler_masterfile_v10(
        TEMPLATE_COLETA_OFICIAL.read_bytes(),
        exigir_modelo_oficial=True,
    )
    posicao = resultado["ciclo_em_execucao"]
    assert resultado["ok"] is True
    assert posicao["layout_tipo"] == "ausente"
    assert posicao["utilizado"] is False
    assert posicao["erros"] == []
    assert posicao["alertas"] == []
    assert "CICLO_EM_EXECUCAO" not in resultado.get("abas_ausentes", [])
    assert not any("CICLO_EM_EXECUCAO" in aviso for aviso in resultado["avisos"])


def test_b_aba_totalmente_vazia_equivale_a_nao_utilizada():
    conteudo = obter_coleta_oficial_bytes()
    wb = load_workbook(io.BytesIO(conteudo), data_only=True)
    posicao = ler_ciclo_em_execucao(wb)
    assert posicao["layout_tipo"] == "itemizado"
    assert posicao["utilizado"] is False
    assert posicao["erros"] == []
    assert posicao["alertas"] == []
    assert posicao.get("total_valor_remanescente") is None

    resultado = ler_masterfile_v10(conteudo, exigir_modelo_oficial=True)
    assert resultado["ok"] is True
    assert not any("CICLO_EM_EXECUCAO" in aviso for aviso in resultado["avisos"])


@pytest.mark.parametrize(
    ("cenario", "preparar", "valido", "erro_esperado"),
    [
        (
            "C_somente_data",
            lambda ws: (setattr(ws["D5"], "value", D), _linha_valida(ws, atual=None)),
            False,
            "REMANESCENTE_ATUAL_NAO_INFORMADO",
        ),
        (
            "D_apenas_alguns_itens",
            lambda ws: (
                setattr(ws["D5"], "value", D),
                _linha_valida(ws, 13, item="ITEM-1", atual=65),
                _linha_valida(ws, 14, item="ITEM-2", atual=None),
            ),
            False,
            "REMANESCENTE_ATUAL_NAO_INFORMADO",
        ),
        (
            "E_completa_valida",
            lambda ws: (setattr(ws["D5"], "value", D), _linha_valida(ws)),
            True,
            None,
        ),
        (
            "F_data_fora_do_ciclo",
            lambda ws: (setattr(ws["D5"], "value", date(2027, 5, 1)), _linha_valida(ws)),
            False,
            "DATA_DA_POSICAO_FORA_DO_CICLO",
        ),
        (
            "F_quantidade_superior",
            lambda ws: (setattr(ws["D5"], "value", D), _linha_valida(ws, atual=81)),
            False,
            "REMANESCENTE_ATUAL_SUPERA_DISPONIVEL",
        ),
        (
            "F_ausencia_de_vu",
            lambda ws: (setattr(ws["D5"], "value", D), _linha_valida(ws, vu=None)),
            False,
            "VU_ATUALIZADO_INVALIDO",
        ),
        (
            "F_inconsistencia_fisica",
            lambda ws: (
                setattr(ws["D5"], "value", D),
                _linha_valida(ws),
                setattr(ws["J13"], "value", 1),
            ),
            False,
            "CHECK_FISICO_DIVERGENTE",
        ),
    ],
)
def test_c_a_f_estados_locais_nao_alteram_resultados_legados(
    cenario, preparar, valido, erro_esperado
):
    wb, legado = _wb_local()
    antes = copy.deepcopy(legado)
    ws = wb[ABA_CICLO_EM_EXECUCAO]
    preparar(ws)

    posicao = ler_ciclo_em_execucao(wb)
    assert posicao["utilizado"] is True, cenario
    assert posicao["valido"] is valido, cenario
    assert _snapshot_legado(wb) == antes, cenario
    if valido:
        assert posicao["erros"] == []
        assert posicao["total_valor_remanescente"] == 6_500.0
    else:
        assert posicao["total_valor_remanescente"] is None
        assert any(erro_esperado in erro for erro in posicao["erros"]), (
            cenario,
            posicao["erros"],
        )


def test_f_evento_duplicado_bloqueia_apenas_o_motor_local():
    legado = {
        "vta": 10_500.0,
        "retroativo": 40.0,
        "remanescente": 8_400.0,
        "vu": 100.0,
        "documentos": "disponiveis",
    }
    antes = copy.deepcopy(legado)
    movimentos = [
        {
            "item": "ITEM-1",
            "data_efeito": date(2026, 6, 1),
            "delta_quantidade": 10,
            "id_evento": "AD-1",
        },
        {
            "item": "ITEM-1",
            "data_efeito": date(2026, 6, 1),
            "delta_quantidade": 10,
            "id_evento": "AD-1",
        },
    ]
    posicao = calcular_posicao_ciclo_por_data(
        ciclo="C3",
        data_inicio=T0,
        data_fim=FIM,
        data_posicao=D,
        itens=[{
            "item": "ITEM-1",
            "remanescente_inicio": 80,
            "remanescente_atual": 65,
            "vu_atualizado": 100,
        }],
        movimentos=movimentos,
    )
    assert posicao["valido"] is False
    assert posicao["total_valor_remanescente"] is None
    assert posicao["itens"][0]["alteracoes_liquidas_periodo"] == 10
    assert any("EVENTO_QUANTITATIVO_DUPLICADO" in erro for erro in posicao["erros"])
    assert legado == antes


def test_f_status_de_erro_do_excel_bloqueia_apenas_o_resultado_local():
    wb, legado = _wb_local()
    antes = copy.deepcopy(legado)
    ws = wb[ABA_CICLO_EM_EXECUCAO]
    ws["D5"] = D
    _linha_valida(ws)
    ws["K13"] = "ERRO: EVENTO QUANTITATIVO DUPLICADO"

    posicao = ler_ciclo_em_execucao(wb)

    assert posicao["utilizado"] is True
    assert posicao["valido"] is False
    assert posicao["total_valor_remanescente"] is None
    assert any("EVENTO QUANTITATIVO DUPLICADO" in erro for erro in posicao["erros"])
    assert _snapshot_legado(wb) == antes


def test_nenhuma_formula_legada_depende_da_aba_opcional():
    # A aba opcional pode ser CONSUMIDA por formulas de apresentacao, desde que
    # SEMPRE via INDIRECT (a referencia vive dentro de um literal entre aspas e
    # resolve para "" quando a aba nao existe — sem #REF!, sem reparo). O que
    # continua PROIBIDO e a referencia DURA (fora de aspas), que criaria
    # dependencia real da aba opcional. Removemos os literais e exigimos que o
    # nome nao sobre no codigo da formula.
    import re

    wb = load_workbook(io.BytesIO(obter_coleta_oficial_bytes()), data_only=False)
    referencias = []
    for ws in wb.worksheets:
        if ws.title == ABA_CICLO_EM_EXECUCAO:
            continue
        for row in ws.iter_rows():
            for cell in row:
                valor = cell.value
                if not (isinstance(valor, str) and valor.startswith("=")):
                    continue
                sem_literais = re.sub(r'"[^"]*"', "", valor)
                if ABA_CICLO_EM_EXECUCAO in sem_literais.upper():
                    referencias.append(f"{ws.title}!{cell.coordinate}:{valor}")
    assert referencias == []

    formulas_novas = [
        cell.value
        for row in wb[ABA_CICLO_EM_EXECUCAO].iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    ]
    assert formulas_novas
    assert not any(
        "VTA" in formula.upper()
        or "RESULTADOS!" in formula.upper()
        or "MEMORIA_RESULTADOS!" in formula.upper()
        for formula in formulas_novas
    )
