# -*- coding: utf-8 -*-
"""14 testes obrigatorios da Etapa VTA — 3 referencias / RESULTADOS / itens_RC.

Fonte da mutacao: `tools/aplicar_vta_posicoes_tabela1.py`.
Travas verificadas: B26/T25 homologados intactos; nenhuma formula criada para
alcancar valor; vazio nunca vira zero; aditivos nao retroagem nem contam duas
vezes; contrato integralmente reajustado permanece comparativo.

Estrategia:
* valores computados sao lidos do arquivo real INBLOKO ja mutado (fixture de
  sessao que roda o aplicador via Excel COM e le com openpyxl data_only);
* cenarios de fallback (ciclo vigente adiantado, abertura posterior incompleta)
  sao exercitados sobrescrevendo entradas via Excel COM e recalculando o motor
  REAL da planilha (nao ha mock de formula);
* comportamentos dependentes da aba fiscal CICLO_EM_EXECUCAO (ausente no arquivo
  real) sao verificados estruturalmente na formula promovida do template.
"""
from __future__ import annotations

import os
import shutil
import tempfile
from pathlib import Path

import pytest
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "templates" / "COLETA_REAJUSTE_OFICIAL.xlsx"
ARQUIVO_REAL = Path(
    r"C:\Users\danie\OneDrive\Desktop"
    r"\2. Coleta de Dados Reajuste INBLOKO_2026-07-31.xlsx"
)
VTA_ESPERADO = 13_468_851.41
TOL = 0.005


# --------------------------------------------------------------------------- #
# Fixtures                                                                     #
# --------------------------------------------------------------------------- #
@pytest.fixture(scope="session")
def real_mutado(tmp_path_factory) -> Path:
    """Aplica a mutacao ao arquivo real e devolve o caminho ja recalculado.

    A mutacao usa Excel COM; sob a suite integral (dezenas de testes COM em
    sequencia) a instancia pode ficar transitoriamente contendida. Tolera-se
    ate tres tentativas antes de falhar, garantindo que o modulo valide os
    MESMOS valores tanto isolado quanto na regressao integral.
    """
    import time

    if os.environ.get("RUN_EXCEL_INTEGRATION") != "1":
        pytest.skip("defina RUN_EXCEL_INTEGRATION=1 para executar Excel COM")
    if not ARQUIVO_REAL.is_file():
        pytest.skip(f"arquivo real ausente: {ARQUIVO_REAL}")
    from tools.aplicar_vta_posicoes_tabela1 import aplicar

    destino = tmp_path_factory.mktemp("vta_pos") / "real_mutado.xlsx"
    ultima_excecao = None
    for tentativa in range(3):
        try:
            aplicar(ARQUIVO_REAL, destino)
            return destino
        except Exception as exc:  # pragma: no cover - resiliencia COM
            ultima_excecao = exc
            if destino.exists():
                destino.unlink()
            time.sleep(5 * (tentativa + 1))
    raise RuntimeError(f"aplicar() falhou apos 3 tentativas: {ultima_excecao}")


@pytest.fixture(scope="session")
def mem(real_mutado):
    wb = load_workbook(real_mutado, data_only=True)
    return wb["MEMORIA_RESULTADOS"]


@pytest.fixture(scope="session")
def resultados(real_mutado):
    wb = load_workbook(real_mutado, data_only=True)
    return wb["RESULTADOS"]


def _com_recalcula(origem: Path, overrides: dict, leituras: list[tuple[str, str]]):
    """Sobrescreve entradas via Excel COM, recalcula e devolve valores lidos."""
    client = pytest.importorskip("win32com.client")
    pythoncom = pytest.importorskip("pythoncom")
    tmp_dir = Path(tempfile.mkdtemp(prefix="vta_cenario_"))
    tmp = tmp_dir / "cenario.xlsx"
    shutil.copyfile(origem, tmp)
    pythoncom.CoInitialize()
    excel = client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    valores = {}
    try:
        wb = excel.Workbooks.Open(str(tmp), UpdateLinks=0, ReadOnly=False, CorruptLoad=0)
        for (aba, cel), valor in overrides.items():
            ws = wb.Worksheets(aba)
            try:
                ws.Unprotect()
            except Exception:
                pass
            ws.Range(cel).Value = valor
        excel.CalculateFullRebuild()
        for aba, cel in leituras:
            valores[(aba, cel)] = wb.Worksheets(aba).Range(cel).Value
        wb.Close(SaveChanges=False)
    finally:
        excel.Quit()
        del excel
        pythoncom.CoUninitialize()
        shutil.rmtree(tmp_dir, ignore_errors=True)
    return valores


# --------------------------------------------------------------------------- #
# 1. Arquivo real: T21+T22+T23 = 13.468.851,41                                 #
# --------------------------------------------------------------------------- #
def test_01_vta_pc_oficial_arquivo_real(mem):
    t21, t22, t23 = mem["T21"].value, mem["T22"].value, mem["T23"].value
    assert abs((t21 + t22 + t23) - VTA_ESPERADO) <= TOL
    assert abs(mem["T25"].value - VTA_ESPERADO) <= TOL
    assert abs(mem["B26"].value - VTA_ESPERADO) <= TOL


# --------------------------------------------------------------------------- #
# 2. FORMA 2 (ultima abertura) reconcilia com o VTA-PC quando vigente completa #
# --------------------------------------------------------------------------- #
def test_02_forma2_igual_vta_pc_quando_abertura_vigente(mem):
    # Arquivo real: vigente C1, abertura C1 completa => W46 = T20 => W48 == T25.
    assert mem["W46"].value == mem["T20"].value
    assert abs(mem["W48"].value - mem["T25"].value) <= TOL
    assert abs(mem["W48"].value - VTA_ESPERADO) <= TOL
    # SUM(AB) na abertura vigente = remanescente vigente (T23), sem dupla contagem.
    soma_ab = sum(
        mem[f"AB{r}"].value
        for r in range(2, 202)
        if isinstance(mem[f"AB{r}"].value, (int, float))
    )
    assert abs(soma_ab - mem["T23"].value) <= TOL


# --------------------------------------------------------------------------- #
# 3. Componentes da FORMA 2 sao rastreaveis (C0 + encerrados + abertura)       #
# --------------------------------------------------------------------------- #
def test_03_forma2_componentes_rastreaveis(mem):
    t21 = mem["T21"].value  # C0 executado
    soma_ab = sum(
        mem[f"AB{r}"].value
        for r in range(2, 202)
        if isinstance(mem[f"AB{r}"].value, (int, float))
    )
    # Vigente C1: encerrados (1..W46-1) = 0; W48 = T21 + 0 + SUM(AB).
    assert abs(mem["W48"].value - round(t21 + soma_ab, 2)) <= TOL


# --------------------------------------------------------------------------- #
# 4. Fallback: vigente adiantado + aberturas posteriores incompletas           #
# --------------------------------------------------------------------------- #
def test_04_fallback_ultima_abertura_completa(real_mutado):
    # Forca ciclo vigente C3 sobre um contrato cujas aberturas C2/C3 nao existem;
    # o motor deve recuar para a ultima abertura completa (C1) e explicitar o
    # motivo, sem dupla contagem nem retroacao.
    lidos = _com_recalcula(
        real_mutado,
        overrides={("CONTROLE", "B2"): "C3"},
        leituras=[
            ("MEMORIA_RESULTADOS", "T20"),
            ("MEMORIA_RESULTADOS", "W46"),
            ("MEMORIA_RESULTADOS", "W47"),
            ("MEMORIA_RESULTADOS", "W48"),
        ],
    )
    t20 = lidos[("MEMORIA_RESULTADOS", "T20")]
    w46 = lidos[("MEMORIA_RESULTADOS", "W46")]
    w47 = lidos[("MEMORIA_RESULTADOS", "W47")]
    w48 = lidos[("MEMORIA_RESULTADOS", "W48")]
    assert t20 == 3
    assert w46 == 1, "deve recuar para a ultima abertura completa (C1)"
    assert w47 and "incompletas" in str(w47)
    # FORMA 2 permanece finita e rastreavel (COM devolve moeda como Decimal).
    assert float(w48) > 0


# --------------------------------------------------------------------------- #
# 5. Sem CICLO_EM_EXECUCAO: FORMA 1 indisponivel; aberturas A:P preservadas    #
# --------------------------------------------------------------------------- #
def test_05_sem_ciclo_execucao_forma1_indisponivel(mem, real_mutado):
    assert mem["W49"].value in (0, None)
    assert mem["W50"].value in (None, "")  # FORMA 1 vazia
    assert isinstance(mem["W48"].value, (int, float))  # FORMA 2 calculada
    # itens_RC: colunas de abertura A:P intactas; bloco Q vazio (aba ausente).
    wb = load_workbook(real_mutado, data_only=True)
    rc = wb["itens_RC"]
    assert rc["A3"].value not in (None, "")  # item preservado
    assert rc["D3"].value is not None        # total C0 preservado (col D)
    assert rc["Q3"].value in (None, "")      # posicao atual vazia sem a aba


# --------------------------------------------------------------------------- #
# 6. CICLO_EM_EXECUCAO incompleta (A9 vazio) => W49=0; guarda estrutural        #
# --------------------------------------------------------------------------- #
def test_06_ciclo_incompleto_a9_vazio_zera_disponibilidade():
    wb = load_workbook(TEMPLATE)  # formulas
    f49 = wb["MEMORIA_RESULTADOS"]["W49"].value
    assert 'INDIRECT("CICLO_EM_EXECUCAO!$A$9")' in f49
    assert "ISERROR" in f49 and '=""' in f49  # A9 vazio ou erro => 0
    # Fotografia POSTERIOR a data de corte tambem zera a disponibilidade: uma
    # posicao futura nao representa a apuracao encerrada em data anterior.
    assert 'INDIRECT("CICLO_EM_EXECUCAO!$D$5")>CONTROLE!$B$3' in f49


# --------------------------------------------------------------------------- #
# 7. itens_RC consome a MESMA fonte da aba (diferenca 0 por construcao)         #
# --------------------------------------------------------------------------- #
def test_07_itens_rc_mesma_fonte_da_aba():
    wb = load_workbook(TEMPLATE)
    rc = wb["itens_RC"]
    # Colunas Q:Y apontam para CICLO_EM_EXECUCAO por INDIRECT (fonte unica).
    assert 'INDIRECT("CICLO_EM_EXECUCAO!B"&(ROW()+10))' in rc["S3"].value
    assert 'INDIRECT("CICLO_EM_EXECUCAO!G"&(ROW()+10))' in rc["X3"].value
    assert "K" in rc["Y3"].value and "CICLO_EM_EXECUCAO" in rc["Y3"].value
    # Nenhuma digitacao manual: todas as celulas do bloco sao formula.
    for col in ("Q", "R", "S", "T", "U", "V", "W", "X", "Y"):
        v = rc[f"{col}3"].value
        assert isinstance(v, str) and v.startswith("=")


# --------------------------------------------------------------------------- #
# 8. Decomposicoes divergentes => W52 REVISE (sem adocao silenciosa)            #
# --------------------------------------------------------------------------- #
def test_08_divergencia_marca_revise():
    wb = load_workbook(TEMPLATE)
    f52 = wb["MEMORIA_RESULTADOS"]["W52"].value
    # Tolerancia D4; acima dela => REVISE; nunca adota FORMA 1 no lugar oficial.
    assert "ABS($W$51)<=$D$4" in f52
    assert "RECONCILIADO" in f52 and "REVISE" in f52
    # A reconciliacao e diferenca explicita, nao substituicao do VTA oficial.
    f51 = wb["MEMORIA_RESULTADOS"]["W51"].value
    assert "$W$50-$W$48" in f51


# --------------------------------------------------------------------------- #
# 9. Item novo em C0: abertura seguinte vazia=pendencia; zero=valida            #
# --------------------------------------------------------------------------- #
def test_09_item_novo_c0_vazio_nao_vira_zero(mem):
    # A completude (W41:W45) exige ISNUMBER da coluna de remanescente: uma
    # abertura VAZIA (texto "") nao satisfaz ISNUMBER => nao vira zero valido.
    wb = load_workbook(TEMPLATE)
    f41 = wb["MEMORIA_RESULTADOS"]["W41"].value
    assert "1-ISNUMBER(posicao_contratual!$G$2:$G$201)" in f41
    # No arquivo real, C0 completa (todas as aberturas numericas) => W41=1.
    assert mem["W41"].value == 1


# --------------------------------------------------------------------------- #
# 10. Aditivo no meio do ciclo: nao aplicavel na abertura anterior ao nascimento#
# --------------------------------------------------------------------------- #
def test_10_aditivo_nao_retroage_antes_do_nascimento():
    # A exigencia por ciclo usa CICLO_NASCIMENTO (Y): item so e exigido no ciclo
    # n quando Y<=n. Antes do nascimento a abertura fica NAO APLICAVEL (nao
    # conta como pendencia nem como zero).
    wb = load_workbook(TEMPLATE)
    for linha, limiar in (("W41", 0), ("W42", 1), ("W43", 2), ("W44", 3), ("W45", 4)):
        f = wb["MEMORIA_RESULTADOS"][linha].value
        assert f"posicao_contratual!$Y$2:$Y$201<={limiar}" in f


# --------------------------------------------------------------------------- #
# 11. AB espelha Y (mesma matematica ROUND) trocando vigente por abertura sel.  #
# --------------------------------------------------------------------------- #
def test_11_ab_espelha_y_com_abertura_selecionada():
    wb = load_workbook(TEMPLATE)
    ab = wb["MEMORIA_RESULTADOS"]["AB2"].value
    y = wb["MEMORIA_RESULTADOS"]["Y2"].value
    # AB usa $W$46 (abertura selecionada); Y usa $T$20 (vigente). Mesma estrutura.
    assert "$W$46+1" in ab and "$T$20+1" in y
    assert ab.replace("$W$46", "$T$20") == y  # identica a menos do seletor


# --------------------------------------------------------------------------- #
# 12. Execucao fisica implicita C0 (coluna X) preservada e coerente             #
# --------------------------------------------------------------------------- #
def test_12_execucao_fisica_implicita_c0(mem):
    """C0 pela movimentacao entre as ABERTURAS temporalmente corretas.

    A abertura de um ciclo e a quantidade declarada MENOS os deltas com data de
    efeito posterior a ela (DELTA_POSTERIOR_ABERTURA_Cn, colunas AB/AC): um
    aditivo do meio de C1 nao pode retroagir sobre a abertura de C1 e, por
    consequencia, encolher a execucao de C0.
    """
    wb = load_workbook(TEMPLATE)
    fx = wb["MEMORIA_RESULTADOS"]["X2"].value
    assert (
        "((posicao_contratual!$G2-N(posicao_contratual!$AB2))"
        "-(posicao_contratual!$K2-N(posicao_contratual!$AC2)))*historico_VU!$C2"
    ) in fx
    # T21 = PC C0 quando existe, senao SUM(X). No arquivo real, valor coerente.
    assert isinstance(mem["T21"].value, (int, float)) and mem["T21"].value > 0


# --------------------------------------------------------------------------- #
# 13. Contrato integralmente reajustado permanece comparativo (nunca adotado)   #
# --------------------------------------------------------------------------- #
def test_13_contrato_integral_e_apenas_comparativo(resultados, mem):
    assert resultados["A12"].value == "CONTRATO ORIGINAL INTEGRALMENTE REAJUSTADO"
    assert resultados["H12"].value == "COMPARATIVO — NAO E VTA OFICIAL"
    b12 = resultados["B12"].value
    assert isinstance(b12, (int, float)) and b12 > VTA_ESPERADO  # 19.501.217,31
    # Nao e o VTA oficial: VTA_FINAL (B26) permanece o valor-PC.
    assert abs(mem["B26"].value - VTA_ESPERADO) <= TOL


# --------------------------------------------------------------------------- #
# 14. Regressao: VTA oficial B26 inalterado; nomes definidos intactos           #
# --------------------------------------------------------------------------- #
def test_14_regressao_b26_e_nomes_intactos(real_mutado):
    wb = load_workbook(real_mutado)  # formulas
    memf = wb["MEMORIA_RESULTADOS"]
    # B26/T25 mantem a formula homologada (nao tocada).
    assert memf["B26"].value.startswith("=IF(AND(B24")
    assert "$T$25" in memf["B26"].value
    assert memf["T25"].value.startswith("=IF(OR($T$24=0")
    nomes = {str(n.split("!")[-1]) for n in wb.defined_names}
    assert "VTA_FINAL" in nomes
    # VTA_ATUALIZACAO_CHEIA repointado para B12 (contrato integral reajustado).
    alvo = wb.defined_names["VTA_ATUALIZACAO_CHEIA"].value
    assert "RESULTADOS!$B$12" in alvo


def test_14b_ciclo_em_execucao_e_opcional_sem_ref_dura():
    # Nenhuma referencia DURA a CICLO_EM_EXECUCAO: o nome so pode aparecer
    # dentro de literais entre aspas (argumento de INDIRECT ou texto
    # descritivo). Qualquer ocorrencia FORA de aspas seria referencia direta a
    # aba ausente => #REF! => reparo. Removemos os literais e exigimos ausencia.
    import re
    import zipfile
    import xml.etree.ElementTree as ET

    ns = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    with zipfile.ZipFile(TEMPLATE) as z:
        for nome in z.namelist():
            if not (nome.startswith("xl/worksheets/") and nome.endswith(".xml")):
                continue
            raiz = ET.fromstring(z.read(nome))
            for f in raiz.findall(".//m:f", ns):
                texto = f.text or ""
                if "CICLO_EM_EXECUCAO" not in texto:
                    continue
                sem_literais = re.sub(r'"[^"]*"', "", texto)
                assert "CICLO_EM_EXECUCAO" not in sem_literais, (
                    f"{nome}: referencia dura fora de INDIRECT: {texto}"
                )
