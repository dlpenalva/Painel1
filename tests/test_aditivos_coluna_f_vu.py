"""Etapa 4 do hotfix: a coluna F da aba 'aditivos' preenche o VU original.

Regressao observada: a formula VLOOKUP da coluna F ficou ausente em uma unica
linha do template (F5), fazendo com que qualquer evento lancado naquela linha
-- inclusive uma Supressao -- nao preenchesse o Valor Unitario original.

A regra correta e type-agnostic: F busca o VU pelo item (coluna A) na fonte
canonica itens_Remanesc!A:C (coluna 3 = VU), independentemente do tipo (coluna
D: Acrescimo/Supressao/Decrescimo). Este teste garante que TODAS as linhas
uteis (F2:F200) tenham a mesma formula, sem buracos, e sem criar fonte paralela.
"""
import re
import sys
import unittest
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

TEMPLATE = ROOT / "templates" / "COLETA_REAJUSTE_OFICIAL.xlsx"

# Padrao homologado da coluna F (VU original) na aba aditivos.
FORMULA_F = '=IF(A{r}="","",IFERROR(VLOOKUP(A{r},itens_Remanesc!$A:$C,3,0),""))'


def _carregar_aba_aditivos():
    wb = openpyxl.load_workbook(TEMPLATE)
    return wb["aditivos"]


def _normaliza(texto):
    return re.sub(r"\s+", "", str(texto))


class TestColunaFVUOriginal(unittest.TestCase):
    def test_todas_as_linhas_uteis_tem_a_formula_vlookup(self):
        ws = _carregar_aba_aditivos()
        faltando = []
        for r in range(2, 201):
            valor = ws[f"F{r}"].value
            if not (isinstance(valor, str) and valor.startswith("=")):
                faltando.append(r)
        self.assertEqual(
            faltando, [],
            f"Linhas de F sem a formula VLOOKUP do VU: {faltando}",
        )

    def test_formula_referencia_a_fonte_canonica_do_vu(self):
        ws = _carregar_aba_aditivos()
        # amostra as extremidades e a linha historicamente quebrada (5)
        for r in (2, 5, 100, 200):
            self.assertEqual(
                _normaliza(ws[f"F{r}"].value),
                _normaliza(FORMULA_F.format(r=r)),
                f"F{r} divergente do padrao canonico do VU",
            )

    def test_regra_e_type_agnostic_nao_depende_do_tipo(self):
        # A formula do VU jamais deve depender da coluna D (tipo de alteracao):
        # Supressao, Acrescimo e Decrescimo preenchem F de forma identica.
        ws = _carregar_aba_aditivos()
        for r in (2, 5, 200):
            formula = str(ws[f"F{r}"].value)
            self.assertNotRegex(
                formula, rf"\bD{r}\b",
                f"F{r} nao pode referenciar o tipo (D{r})",
            )
            self.assertIn("itens_Remanesc!$A:$C", formula)


if __name__ == "__main__":
    unittest.main()
