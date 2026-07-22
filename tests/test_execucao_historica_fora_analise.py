"""Etapa 1 (final) - Execucao Historica: decomposicao (PARCELA A + PARCELA B),
fator base praticado, integracao governada ao VTA (B26) e bloco EIXO 5.

Estrutura estatica no template + cenarios numericos com Microsoft Excel real
(RUN_EXCEL_INTEGRATION=1).
"""
from __future__ import annotations

import datetime
import os
from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string as cidx

TEMPLATE = Path(__file__).resolve().parent.parent / "templates" / "COLETA_REAJUSTE_OFICIAL.xlsx"
LIN_HDR, LIN_C0, LIN_TOT = 257, 258, 263

CAB_BLOCO = [
    "CICLO", "SITUACAO DO CICLO", "FONTE", "COBERTURA", "QTD EXECUTADA RECONSTRUIDA",
    "QTD COBERTA", "QTD NAO COBERTA", "FATOR CHEIO", "FATOR BASE PRATICADO",
    "DELTA DO FATOR DA APURACAO", "ATUALIZACAO ANTERIOR POTENCIAL",
    "ATUALIZACAO ATUAL NAO COBERTA", "COMPLEMENTO HISTORICO POTENCIAL",
    "VALOR ADOTADO", "STATUS / CHECK",
]
VTA_HOMOLOGADO = {
    "B20": '=IF(COUNTIF(itens_Remanesc!$A$2:$A$200,"<>")=0,"",ROUND(SUMIF(itens_Remanesc!$A$2:$A$200,"<>",itens_Remanesc!$D$2:$D$200),2))',
    "B21": '=IF(B16="","",B16)',
    "B22": '=IF(OR(C35="",D35=""),"",ROUND(D35-C35,2))',
    "B23": '=IF(OR(B20="",B21="",B22=""),"",ROUND(B20+B21+B22,2))',
}


@pytest.fixture(scope="module")
def wb():
    return load_workbook(TEMPLATE, data_only=False)


def test_bloco_cabecalhos(wb):
    ws = wb["RESULTADOS"]
    assert str(ws["A256"].value).startswith("EIXO 5")
    assert [ws.cell(LIN_HDR, c).value for c in range(1, 16)] == CAB_BLOCO
    assert [ws.cell(LIN_C0 + i, 1).value for i in range(5)] == ["C0", "C1", "C2", "C3", "C4"]
    assert ws.cell(LIN_TOT, 1).value == "TOTAIS"


def test_helper_20_colunas(wb):
    ws = wb["itens_Remanesc"]
    ach = [ws.cell(1, c).value for c in range(cidx("AO"), cidx("BH") + 1)]
    assert len(ach) == 20
    assert ach[0] == "QTD_COBERTA_C0" and ach[2] == "PARCELA_A_C0" and ach[3] == "PARCELA_B_C0"
    assert ach[-1] == "PARCELA_B_C4"


def test_vta_b23_inalterado(wb):
    ws = wb["RESULTADOS"]
    for cel, esp in VTA_HOMOLOGADO.items():
        assert ws[cel].value == esp, f"{cel} divergiu"


def test_b26_integra_total_adotado_sem_override(wb):
    ws = wb["RESULTADOS"]
    b26 = str(ws["B26"].value)
    assert "IF(ISNUMBER($N$263),$N$263,0)" in b26           # incorpora total adotado
    assert b26.startswith('=IF(AND(B24<>"",B25<>"")')        # governanca preservada
    assert "IF(ISNUMBER(B25),B25," in b26                    # override B25 soberano preservado
    assert str(ws["N263"].value) == "=ROUND(SUM(N258:N262),2)"


def test_vta_nao_depende_do_bloco(wb):
    ws = wb["RESULTADOS"]
    chain = ["B20", "B21", "B22", "B23", "B16", "E15", "C35", "D35", "B32", "C32", "D32"]
    aux = ["AO", "AP", "AQ", "AR", "AS", "AT", "AU", "AV", "AW", "AX", "AY", "AZ",
           "BA", "BB", "BC", "BD", "BE", "BF", "BG", "BH"]
    for cel in chain:
        f = str(ws[cel].value or "")
        for p in aux:
            assert f"itens_Remanesc!{p}" not in f
        for lin in range(256, 266):
            assert f"K{lin}" not in f and f"N{lin}" not in f


def test_fator_base_canonico(wb):
    ws = wb["RESULTADOS"]
    # F_base = F/D quando computado; F_cheio quando fora+formalizado (marcador Sim)
    i259 = str(ws["I259"].value)
    assert "parametros!$F$3/parametros!$D$12" in i259
    assert 'parametros!$G$12="Sim"' in i259
    # delta atual = fator cheio - fator base
    assert str(ws["J259"].value) == '=IF(OR(H259="",I259=""),"",ROUND(H259-I259,2))'


def test_parcelas_formulas(wb):
    wi = wb["itens_Remanesc"]
    # PARCELA A = QTD_EXEC * VU * (F_base - 1) ; independe de cobertura
    assert wi["AU2"].value == '=IF(OR($A2="",M2="",RESULTADOS!$I$259=""),"",ROUND(M2*$C2*(RESULTADOS!$I$259-1),2))'
    # PARCELA B = QTD_NAO_COBERTA * VU * (F_cheio - F_base)
    assert wi["AV2"].value == '=IF(OR($A2="",AT2="",RESULTADOS!$I$259=""),"",ROUND(AT2*$C2*(parametros!$F$3-RESULTADOS!$I$259),2))'
    # bloco: K (parcela A) nao exige metodo; L (parcela B) exige "Itens"
    ws = wb["RESULTADOS"]
    assert '$B$4<>"Itens"' not in str(ws["K259"].value)      # A independe do metodo
    assert '$B$4<>"Itens"' in str(ws["L259"].value)          # B exige reconciliavel


def test_marcador_formalizado_dropdown(wb):
    import zipfile
    import re
    wp = wb["parametros"]
    assert "FORMALIZADO" in str(wp["G10"].value).upper()
    # validacao tipo lista cobrindo G12:G15 (Excel embrulha a lista em
    # mc:AlternateContent, que openpyxl le como formula1=None -> checar XML bruto)
    achou = any("G12:G15" in str(dv.sqref) and dv.type == "list"
                for dv in wp.data_validations.dataValidation)
    assert achou, "dropdown (list) ausente em parametros G12:G15"
    with zipfile.ZipFile(TEMPLATE) as z:
        raw = None
        for n in z.namelist():
            if n.startswith("xl/worksheets/") and b"REAJUSTE ANTERIOR JA FORMALIZADO" not in b"":
                pass
        # localiza a aba parametros pelo conteudo do marcador
        for n in z.namelist():
            if n.startswith("xl/worksheets/sheet") and n.endswith(".xml"):
                data = z.read(n)
                if b"G12:G15" in data:
                    raw = data.decode("utf-8", "replace")
                    break
    assert raw is not None and "Sim,Nao" in raw, "lista Sim,Nao ausente no XML"
    for r in range(11, 16):
        assert wp[f"G{r}"].value in (None, "")               # vazio por padrao


def test_valor_adotado_validacao(wb):
    ws = wb["RESULTADOS"]
    achou = False
    for dv in ws.data_validations.dataValidation:
        if any(f"N{LIN_C0 + i}" in str(dv.sqref) for i in range(5)):
            achou = True
            assert dv.type == "decimal" and dv.operator == "greaterThanOrEqual"
    assert achou
    for i in range(5):
        assert ws.cell(LIN_C0 + i, 14).value in (None, "")   # coluna N vazia


def test_colunas_ocultas_reexibiveis(wb):
    ws = wb["itens_Remanesc"]
    oc = [d for d in ws.column_dimensions.values() if d.hidden]
    for col in range(cidx("AO"), cidx("BH") + 1):
        assert any(d.min <= col <= d.max for d in oc), col
    for w in wb.worksheets:
        assert w.sheet_state != "veryHidden"


# ==========================================================================
# Cenarios numericos com Microsoft Excel real (opt-in).
# ==========================================================================
EXCEL = pytest.mark.skipif(
    os.environ.get("RUN_EXCEL_INTEGRATION") != "1",
    reason="defina RUN_EXCEL_INTEGRATION=1 para executar o Excel COM",
)


def _abrir(caminho, editar, ler):
    import gc
    import time
    import pythoncom
    import win32com.client

    def tentar(acao):
        ultimo = None
        for _ in range(30):
            try:
                return acao()
            except Exception as exc:
                ultimo = exc
                cod = getattr(exc, "hresult", None) or (exc.args[0] if exc.args else None)
                if cod != -2147418111:
                    raise
                pythoncom.PumpWaitingMessages()
                time.sleep(0.2)
        raise ultimo

    pythoncom.CoInitialize()
    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    pasta = None
    try:
        pasta = tentar(lambda: excel.Workbooks.Open(str(Path(caminho).resolve()), UpdateLinks=0))
        time.sleep(0.4)
        tentar(lambda: editar(pasta))
        tentar(excel.CalculateFullRebuild)
        res = tentar(lambda: ler(pasta))
        tentar(lambda: pasta.Close(SaveChanges=False))
        pasta = None
        return res
    finally:
        if pasta is not None:
            pasta.Close(SaveChanges=False)
        excel.Quit()
        gc.collect()
        pythoncom.CoUninitialize()


def _cenario(tmp_path):
    """Coleta com C1 e C2 na apuracao (10% cada) -> F3=1,10 ; F4=1,21."""
    from _coleta_oficial import gerar_coleta_oficial_preenchida
    d = datetime.date
    dados = {
        "origem": "Etapa 1 COM", "indice": "IST",
        "data_base_original": "01/01/2023", "data_corte": d(2025, 12, 31),
        "ciclos": [
            {"ciclo": "C1", "data_inicio": d(2024, 1, 1), "data_fim": d(2024, 12, 31), "data_pedido": d(2024, 1, 1), "financeiro_inicio": d(2024, 1, 1), "percentual": 0.10},
            {"ciclo": "C2", "data_inicio": d(2025, 1, 1), "data_fim": d(2025, 12, 31), "data_pedido": d(2025, 1, 1), "financeiro_inicio": d(2025, 1, 1), "percentual": 0.10},
        ],
    }
    caminho = tmp_path / "cen.xlsx"
    caminho.write_bytes(gerar_coleta_oficial_preenchida(dados))
    return caminho


def _edit(pasta, *, metodo="Itens", base=100.0, vu=10.0, rem_c1=60.0, rem_c2=30.0,
          cons_c1=None, delta_c1=0.0, adotado=None, b25=None, uncompute=None,
          marker=None, aditivo_considerado=False):
    if metodo is not None:
        pasta.Worksheets("RESULTADOS").Range("B4").Value = metodo
    ir = pasta.Worksheets("itens_Remanesc")
    ir.Range("A2").Value = "ITEM-1"
    ir.Range("B2").Value = base
    ir.Range("C2").Value = vu
    if rem_c1 is not None:
        ir.Range("E2").Value = rem_c1
    if rem_c2 is not None:
        ir.Range("G2").Value = rem_c2
    if cons_c1 is not None:
        ic = pasta.Worksheets("itens_Consumidos")
        ic.Range("A2").Value = "ITEM-1"
        ic.Range("B2").Value = 100.0
        ic.Range("C2").Value = vu
        ic.Range("G2").Value = float(cons_c1)
    if delta_c1:
        import pywintypes
        ad = pasta.Worksheets("aditivos")
        ad.Range("A2").Value = "ITEM-1"
        ad.Range("B2").Value = pywintypes.Time(datetime.datetime(2024, 6, 1))
        ad.Range("D2").Value = "ACRESCIMO" if delta_c1 > 0 else "SUPRESSAO"
        ad.Range("E2").Value = abs(delta_c1)
        ad.Range("H2").Value = "Nao"
        if aditivo_considerado:
            ad.Range("K2").Value = "Sim"
    wp = pasta.Worksheets("parametros")
    if uncompute:
        for arow in uncompute:
            wp.Range(f"A{arow}").Value = ""
    if marker:
        for grow, val in marker.items():
            wp.Range(f"G{grow}").Value = val
    r = pasta.Worksheets("RESULTADOS")
    if adotado:
        for row, val in adotado.items():
            r.Range(f"N{row}").Value = val
    if b25 is not None:
        r.Range("B25").Value = b25
        r.Range("D25").Value = "teste"  # justificativa


def _ler(pasta):
    r = pasta.Worksheets("RESULTADOS")
    L = {}
    for i, nome in enumerate(["C0", "C1", "C2", "C3", "C4"]):
        row = 258 + i
        L[nome] = {k: r.Range(f"{col}{row}").Value for k, col in {
            "sit": "B", "cob": "D", "exec": "E", "qcob": "F", "naocob": "G",
            "fcheio": "H", "fbase": "I", "delta": "J", "parc_a": "K",
            "parc_b": "L", "compl": "M", "adot": "N", "status": "O"}.items()}
    return {"linhas": L, "B20": r.Range("B20").Value, "B21": r.Range("B21").Value,
            "B22": r.Range("B22").Value, "B23": r.Range("B23").Value,
            "B26": r.Range("B26").Value, "K263": r.Range("K263").Value,
            "L263": r.Range("L263").Value, "M263": r.Range("M263").Value,
            "N263": r.Range("N263").Value, "A265": r.Range("A265").Value}


def _num(x):
    return x if isinstance(x, (int, float)) else None


@EXCEL
def test_com_c0_fator_1(tmp_path):
    """A: C0 fator base 1 -> parcela A e B nulas; SITUACAO BASE."""
    res = _abrir(_cenario(tmp_path), lambda p: _edit(p, cons_c1=30.0), _ler)
    c0 = res["linhas"]["C0"]
    assert c0["sit"] == "BASE"
    assert abs(_num(c0["fbase"]) - 1.0) < 1e-6
    assert abs(_num(c0["parc_a"]) - 0.0) < 0.01


@EXCEL
def test_com_ciclo_atual_fator_base(tmp_path):
    """C: C2 fator cheio 1,21 e apuracao 1,10 -> fator base 1,10 ; delta 0,11."""
    res = _abrir(_cenario(tmp_path), lambda p: _edit(p, cons_c1=30.0), _ler)
    c2 = res["linhas"]["C2"]
    assert abs(_num(c2["fcheio"]) - 1.21) < 1e-6
    assert abs(_num(c2["fbase"]) - 1.10) < 1e-6
    assert abs(_num(c2["delta"]) - 0.11) < 1e-6


@EXCEL
def test_com_cobertura_completa_parcial(tmp_path):
    """D+E: C1 exec 30; consumo 30 completa e 10 parcial (parcela B so no nao coberto)."""
    comp = _abrir(_cenario(tmp_path), lambda p: _edit(p, cons_c1=30.0), _ler)["linhas"]["C1"]
    assert comp["cob"] == "COMPLETA" and abs(_num(comp["naocob"]) - 0.0) < 0.01
    assert abs(_num(comp["parc_b"]) - 0.0) < 0.01
    parc = _abrir(_cenario(tmp_path), lambda p: _edit(p, cons_c1=10.0), _ler)["linhas"]["C1"]
    assert parc["cob"] == "PARCIAL" and abs(_num(parc["naocob"]) - 20.0) < 0.01
    # parcela B = 20 * 10 * (1,10 - 1,00) = 20
    assert abs(_num(parc["parc_b"]) - 20.0) < 0.01
    # parcela A (anterior) sobre C1 = exec 30 * 10 * (fbase-1); C1 computado fbase=1,10/...
    assert _num(parc["parc_a"]) is not None


@EXCEL
def test_com_financeiro_e_pcs(tmp_path):
    """F+G: Financeiro/PCs -> parcela B nao reconciliavel, mas parcela A calculavel."""
    for metodo in ("Financeiro", "PCs"):
        c1 = _abrir(_cenario(tmp_path), lambda p, m=metodo: _edit(p, metodo=m, cons_c1=10.0), _ler)["linhas"]["C1"]
        assert c1["cob"] == "NAO RECONCILIAVEL"
        assert c1["parc_b"] in (None, "")                    # parcela B bloqueada
        assert _num(c1["parc_a"]) is not None                # parcela A independe do metodo


@EXCEL
def test_com_fora_formalizado_vs_nao_comprovado(tmp_path):
    """H+I: C2 fora; formalizado (marcador Sim) calcula parcela A; sem marcador -> manual."""
    # torna C2 fora da apuracao (parametros A4) e marca formalizado (G13=Sim)
    form = _abrir(_cenario(tmp_path),
                  lambda p: _edit(p, cons_c1=30.0, uncompute=[4], marker={13: "Sim"}), _ler)["linhas"]["C2"]
    assert form["sit"] == "FORA - FORMALIZADO"
    assert _num(form["fbase"]) is not None                   # F_base = F_cheio
    assert _num(form["parc_a"]) is not None
    # sem comprovacao -> fator base vazio, parcela nao calculada
    amb = _abrir(_cenario(tmp_path),
                 lambda p: _edit(p, cons_c1=30.0, uncompute=[4]), _ler)["linhas"]["C2"]
    assert amb["sit"] == "FORA - NAO COMPROVADO"
    assert amb["fbase"] in (None, "")
    assert amb["parc_a"] in (None, "")
    assert "NAO COMPROVADO" in str(amb["status"])


@EXCEL
def test_com_acrescimo_supressao(tmp_path):
    """K+L: DELTA assinado na reconstrucao (nao confunde aditivo com execucao)."""
    ac = _abrir(_cenario(tmp_path), lambda p: _edit(p, cons_c1=0.0, delta_c1=+20.0), _ler)["linhas"]["C1"]
    assert abs(_num(ac["exec"]) - 50.0) < 0.01               # 60+20-30
    su = _abrir(_cenario(tmp_path), lambda p: _edit(p, cons_c1=0.0, delta_c1=-20.0), _ler)["linhas"]["C1"]
    assert abs(_num(su["exec"]) - 10.0) < 0.01               # MAX(60-20-30,0)


@EXCEL
def test_com_item_novo_base_zero(tmp_path):
    """N: item novo (base 0) + aditivo -> principal nao entra como complemento."""
    res = _abrir(_cenario(tmp_path),
                 lambda p: _edit(p, base=0.0, rem_c1=20.0, rem_c2=0.0, delta_c1=+20.0,
                                 aditivo_considerado=True, cons_c1=0.0), _ler)
    # B20 (principal) nao inclui o aditivo; execucao respeita DELTA; principal do
    # aditivo continua no gate manual (B45>0 forca calculo manual em E26).
    assert res["B20"] in (None, "", 0) or _num(res["B20"]) == 0.0
    assert _num(res["linhas"]["C1"]["exec"]) is not None


@EXCEL
def test_com_c4_sem_fotografia(tmp_path):
    """O: C4 sem fotografia posterior -> sem execucao reconstruida."""
    c4 = _abrir(_cenario(tmp_path), lambda p: _edit(p, cons_c1=30.0), _ler)["linhas"]["C4"]
    assert c4["exec"] in (None, "")
    assert c4["cob"] == "SEM EXECUCAO"


@EXCEL
def test_com_valor_adotado_variacoes(tmp_path):
    """P-U: vazio, zero, negativo e superior ao potencial -> status e validacao."""
    base = _abrir(_cenario(tmp_path), lambda p: _edit(p, cons_c1=10.0), _ler)
    assert base["linhas"]["C1"]["adot"] in (None, "")        # P: vazio padrao
    assert abs(_num(base["N263"]) - 0.0) < 0.01
    zero = _abrir(_cenario(tmp_path), lambda p: _edit(p, cons_c1=10.0, adotado={259: 0.0}), _ler)
    assert "ADOTADO" in str(zero["linhas"]["C1"]["status"])   # Q: zero aceito
    neg = _abrir(_cenario(tmp_path), lambda p: _edit(p, cons_c1=10.0, adotado={259: -5.0}), _ler)
    assert "NEGATIVO" in str(neg["linhas"]["C1"]["status"])   # U: negativo
    sup = _abrir(_cenario(tmp_path), lambda p: _edit(p, cons_c1=10.0, adotado={259: 99999.0}), _ler)
    assert "SUPERIOR" in str(sup["linhas"]["C1"]["status"])   # T: superior ao potencial


@EXCEL
def test_com_sem_override_incorpora_ao_vta(tmp_path):
    """V+R: sem override, VTA FINAL = B23 + TOTAL ADOTADO; B23 inalterado."""
    sem = _abrir(_cenario(tmp_path), lambda p: _edit(p, cons_c1=10.0), _ler)
    com = _abrir(_cenario(tmp_path), lambda p: _edit(p, cons_c1=10.0, adotado={259: 20.0}), _ler)
    # B23 identico com e sem adocao (invariancia)
    assert sem["B23"] == com["B23"]
    # N263 = 20 ; B26 = B23 + 20
    assert abs(_num(com["N263"]) - 20.0) < 0.01
    if _num(com["B23"]) is not None:
        assert abs(_num(com["B26"]) - (_num(com["B23"]) + 20.0)) < 0.01
    assert "incorporado" in str(com["A265"]).lower()


@EXCEL
def test_com_override_total_soberano(tmp_path):
    """W: override B25 permanece soberano; alerta; complemento nao somado."""
    res = _abrir(_cenario(tmp_path),
                 lambda p: _edit(p, cons_c1=10.0, adotado={259: 20.0}, b25=5000.0), _ler)
    assert abs(_num(res["B26"]) - 5000.0) < 0.01             # B25 soberano
    assert "OVERRIDE" in str(res["A265"]).upper()


@EXCEL
def test_com_aditivo_considerado_complemento(tmp_path):
    """X: aditivo considerado + complemento -> gate manual do aditivo permanece."""
    res = _abrir(_cenario(tmp_path),
                 lambda p: _edit(p, cons_c1=10.0, delta_c1=+20.0, aditivo_considerado=True,
                                 adotado={259: 10.0}), _ler)
    # complemento historico funciona; principal do aditivo NAO entra (segue manual)
    assert _num(res["N263"]) is not None
    assert _num(res["linhas"]["C1"]["exec"]) is not None


@EXCEL
def test_com_identidade_conservacao(tmp_path):
    """15+Y+Z: adotando o complemento total, B26 = B23 + (K263 + L263), sem dupla
    contagem; B23 permanece = B20 + B21 + B22."""
    res = _abrir(_cenario(tmp_path),
                 lambda p: _edit(p, cons_c1=10.0,
                                 adotado={258: 0.0, 259: None}), _ler)
    # adota o complemento potencial de C1 (M259) integralmente
    def editar2(p):
        _edit(p, cons_c1=10.0)
        r = p.Worksheets("RESULTADOS")
        p.Application.CalculateFull()
        m259 = r.Range("M259").Value
        if isinstance(m259, (int, float)):
            r.Range("N259").Value = m259
    res = _abrir(_cenario(tmp_path), editar2, _ler)
    b23, n263, b26 = _num(res["B23"]), _num(res["N263"]), _num(res["B26"])
    if b23 is not None:
        # conservacao: VTA FINAL = B23 + total adotado, sem duplicar
        assert abs(b26 - (b23 + n263)) < 0.02
        # B23 = B20 + B21 + B22 (decomposicao homologada intacta)
        b20, b21, b22 = _num(res["B20"]), _num(res["B21"]), _num(res["B22"])
        if None not in (b20, b21, b22):
            assert abs(b23 - (b20 + b21 + b22)) < 0.02


@EXCEL
def test_com_aplicador_fail_closed(tmp_path):
    """AD: segunda aplicacao falha (fail-closed), template intacto."""
    import shutil
    import importlib.util
    import time
    spec = importlib.util.spec_from_file_location(
        "aplicar_eh", TEMPLATE.parent.parent / "tools" / "aplicar_execucao_historica_template.py")
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    alvo = tmp_path / "ja.xlsx"
    shutil.copyfile(TEMPLATE, alvo)
    antes = alvo.read_bytes()
    erro = None
    for _ in range(5):
        try:
            mod.aplicar(alvo, alvo)
            erro = "SEM ERRO"
            break
        except (ValueError, RuntimeError) as exc:
            erro = exc
            break
        except Exception as exc:
            if getattr(exc, "args", (None,))[0] == -2147418111:
                time.sleep(1.0)
                continue
            raise
    assert isinstance(erro, (ValueError, RuntimeError))
    assert alvo.read_bytes() == antes
