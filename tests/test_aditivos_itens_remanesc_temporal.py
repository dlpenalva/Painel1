# -*- coding: utf-8 -*-
"""Comportamento temporal de itens novos (aditivos) em itens_Remanesc.

Prova, sobre as formulas REAIS do template, as regras da Secao 6 e fixa o
comportamento homologado. Tabela de validacao dos casos N001/N002/N003:

  item | inclusao        | Y (nascimento) | abertura | aplicavel | esperado
  -----+-----------------+----------------+----------+-----------+--------------
  N001 | C0 (base)       | 0              | C1       | SIM(pend) | vazio=pendencia
  N002 | durante C1      | 1              | C0       | NAO       | "" (NAO APLIC.)
  N002 | durante C1      | 1              | C1       | SIM       | delta (1x)
  N003 | 1o dia de C1    | 1              | C1       | SIM       | delta (1x)

Regras cobertas:
* Regra 1: item anterior a abertura -> celula manual (J = itens_Remanesc!E)
  informavel; vazio = pendencia (NUNCA vira zero).
* Regra 2/3: abertura anterior ao nascimento -> "" (NAO APLICAVEL, nao zero, nao
  pendencia, nao bloqueia VTA). A abertura do ciclo de nascimento usa o delta
  UMA vez (sem dupla contagem).
* Regra 4: itens_Remanesc distingue visualmente os 4 estados (CF).

RESSALVA RESOLVIDA (Regra 2 x Regra 3): `posicao_contratual!Y` deriva da
QUANTIDADE (primeira coluna QTD_CONTRATADA_Cn>0), nao da DATA do aditivo — logo
um item nascido no 1o dia e outro nascido no meio do MESMO ciclo recebiam o
mesmo Y. A distincao temporal passou a existir em `posicao_contratual!AL`
(CICLO_NASCIMENTO_DATA), derivada da data de efeito, sem tocar em Y nem na
cadeia oficial do VTA. Este modulo continua fixando o comportamento HOMOLOGADO
da camada oficial (Y e as colunas G/K/O/S/W); a camada temporal e coberta por
`tests/test_temporalidade_aditivos_data_efeito.py`.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "templates" / "COLETA_REAJUSTE_OFICIAL.xlsx"


@pytest.fixture(scope="module")
def pc():
    wb = openpyxl.load_workbook(TEMPLATE)
    return wb["posicao_contratual"]


def test_ciclo_nascimento_por_quantidade(pc):
    y = str(pc["Y2"].value)
    for trecho in ("$E2>0", "$I2>0", "$M2>0", "$Q2>0", "$U2>0"):
        assert trecho in y
    assert '4,""' in y


def test_pre_nascimento_nao_aplicavel(pc):
    assert 'IF($Y2>0,""' in str(pc["G2"].value)
    assert 'IF($Y2>1,""' in str(pc["K2"].value)
    assert 'IF($Y2>2,""' in str(pc["O2"].value)


def test_abertura_informada_soma_delta_uma_vez(pc):
    k = str(pc["K2"].value)
    assert "IF(ISNUMBER(J2),ROUND(J2+H2,2)" in k
    assert k.rstrip().endswith('ROUND(H2,2),"")))))')


def test_nascimento_usa_delta_uma_unica_vez(pc):
    k = str(pc["K2"].value)
    assert 'IF($Y2=1,ROUND(H2,2),"")' in k
    o = str(pc["O2"].value)
    assert 'IF($Y2=2,ROUND(L2,2),"")' in o


def test_regra4_quatro_estados_visuais():
    wb = openpyxl.load_workbook(TEMPLATE)
    ws = wb["itens_Remanesc"]
    por_faixa = {str(cf.sqref): list(cf.rules) for cf in ws.conditional_formatting}
    for col, idx in (("E", 1), ("G", 2), ("I", 3), ("K", 4)):
        regras = por_faixa[f"{col}2:{col}201"]
        joined = " ".join(str(r.formula) for r in regras)
        # NAO APLICAVEL passou a olhar o nascimento POR DATA, espelhado em BI.
        assert f"$BI2>{idx}" in joined
        assert f"{col}2>0" in joined and f"{col}2=0" in joined and f'{col}2=""' in joined


def test_base_zero_automatica_novo_item():
    wb = openpyxl.load_workbook(TEMPLATE)
    ws = wb["itens_Remanesc"]
    b2 = str(ws["B2"].value)
    assert 'LEFT(TRIM($A2),1)="N"' in b2
    assert b2.rstrip().endswith(',0,"")')
