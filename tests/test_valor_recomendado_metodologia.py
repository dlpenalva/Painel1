from datetime import date

from openpyxl import Workbook

from _leitor_masterfile_v10 import _ler_parcelas_sombra_financeiro
from _motor_metodologias import MET_FIN_REMANESC, montar_motor_metodologias


def test_leitor_financeiro_materializa_apenas_ciclos_c0_a_c4():
    wb = Workbook()
    ws = wb.active
    ws.title = "financeiro"
    ws.append([
        "COMPETENCIA", "CICLO", "VALOR_PAGO", "FATOR", "VALOR_ATUALIZADO",
        "DELTA", "EFEITO_FINANCEIRO",
    ])
    ws.append([date(2024, 1, 1), "C0", 100.0, 1.0, 100.0, 0.0, "Nao"])
    ws["B74"] = "TOTAL"
    ws["C74"] = 100.0
    ws["G74"] = "Nao"

    parcelas = _ler_parcelas_sombra_financeiro(wb)

    assert any(p["linha"] == 2 for p in parcelas)
    assert not any(p["linha"] == 74 for p in parcelas)
    assert {p["ciclo"] for p in parcelas} <= {"C0", "C1", "C2", "C3", "C4"}


def test_financeiro_mais_remanescente_consumido_da_composicao_canonica():
    leitura = {
        "event_log_sombra": {
            "eventos": [
                {
                    "fonte_parcela": "Financeiro",
                    "tipo_financeiro": "Execucao Atualizada",
                    "valor": 7_300_890.27,
                },
                {
                    "fonte_parcela": "Financeiro",
                    "tipo_financeiro": "Retroativo Reconhecido",
                    "valor": 24_678.92,
                },
            ]
        },
        "composicao_vta": {
            "saldo_remanescente": {
                "valor_atualizado": 1_388_251.07,
                "fonte": "posicao_contratual",
            }
        },
        "execucao_saldo": {"itens": []},
        "itens_consumidos_v10": {"itens": []},
    }
    painel = {
        "disponivel": True,
        "situacao_financeira": {
            "resumo_oficial": {"saldo_remanescente": 0.0},
            "totais": {"retroativo": 24_678.92},
        },
        "situacao_pcs": {"pcs": []},
        "vta": {"oficial": 8_713_820.26},
        "alertas": [],
    }

    motor = montar_motor_metodologias(leitura, painel)
    evidencias = motor["evidencias"]
    resultado = motor["resultado_recomendado"]

    assert evidencias["financeiro"]["pago"] == 7_300_890.27
    assert evidencias["financeiro"]["reconhecido"] == 24_678.92
    assert evidencias["remanescentes"]["valor"] == 1_388_251.07
    assert evidencias["remanescentes"]["fonte"] == (
        "composicao_vta.saldo_remanescente.valor_atualizado"
    )
    assert resultado["metodologia"] == MET_FIN_REMANESC
    assert resultado["valor_recomendado"] == 8_689_141.34
    assert resultado["valor_recomendado"] != resultado["vta"]
