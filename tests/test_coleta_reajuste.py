import io
import unittest
from pathlib import Path
from unittest.mock import patch
from zipfile import ZipFile

from openpyxl import load_workbook

from _coleta_reajuste import (
    ABAS_CANONICAS,
    CAMINHO_MODELO_COLETA,
    NOMES_RESULTADOS_OBRIGATORIOS,
    NOME_ARQUIVO_COLETA,
    eh_coleta_reajuste,
    gerar_coleta_reajuste,
    ler_coleta_reajuste,
)


ROOT = Path(__file__).resolve().parents[1]


def formulas(workbook):
    return {
        f"{ws.title}!{cell.coordinate}": cell.value
        for ws in workbook.worksheets
        for row in ws.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    }


class ColetaReajusteTests(unittest.TestCase):
    def test_modelo_canonico_esta_limpo_e_formula_first(self):
        wb = load_workbook(CAMINHO_MODELO_COLETA, data_only=False)
        self.assertEqual(tuple(wb.sheetnames), ABAS_CANONICAS)
        self.assertNotIn("itens_Execucao_Saldo", wb.sheetnames)
        self.assertNotIn("REGRA_NEGOCIO_CLAUS", wb.sheetnames)
        self.assertEqual(wb.sheetnames[-1], "RESULTADOS")
        self.assertEqual(wb["RESULTADOS"].max_row, 263)
        self.assertEqual(wb["RESULTADOS"]["A1"].value, "RESULTADOS CONSOLIDADOS — REAJUSTE CONTRATUAL")
        self.assertEqual(
            [ws.title for ws in wb.worksheets if ws.sheet_properties.tabColor is not None],
            ["RESULTADOS"],
        )
        self.assertTrue(NOMES_RESULTADOS_OBRIGATORIOS.issubset(set(wb.defined_names)))

        resultados = wb["RESULTADOS"]
        self.assertEqual(resultados["B4"].value, "Financeiro")
        self.assertEqual(resultados["D4"].value, 0.005)
        self.assertIn('IF($B$4="Financeiro",B15', resultados["E15"].value)
        self.assertNotIn("SUM(B15:D15)", resultados["B16"].value)
        self.assertIn("B20+B21+B22", resultados["B23"].value)
        self.assertIn("CONTROLE!$B$1", resultados["B35"].value)
        self.assertIn("CÁLCULO MANUAL REQUERIDO", resultados["F36"].value)
        self.assertIn("fator-base praticado", resultados["A48"].value)

        ws = wb["parametros"]
        self.assertEqual(ws.max_column, 7)
        self.assertEqual(
            [ws.cell(1, col).value for col in range(1, 8)],
            [
                "COMPUTAR_NESTA_APURACAO",
                "CICLO",
                "DATA_INICIO",
                "DATA_FIM",
                "PERCENTUAL_DO_CICLO",
                "FATOR_ACUMULADO",
                "SITUACAO",
            ],
        )
        self.assertNotIn("PERIODO_DO_CICLO", [ws.cell(1, col).value for col in range(1, 8)])
        self.assertEqual([ws[f"C{row}"].value for row in range(2, 7)], [None] * 5)
        self.assertEqual([ws[f"E{row}"].value for row in range(3, 7)], [None] * 4)

        ws = wb["financeiro"]
        self.assertEqual(ws.max_row, 61)
        self.assertEqual(
            [ws.cell(1, col).value for col in range(1, 8)],
            [
                "COMPETENCIA",
                "CICLO",
                "VALOR_PAGO",
                "FATOR_APLICAVEL",
                "VALOR_ATUALIZADO",
                "DELTA",
                "EFEITO_FINANCEIRO",
            ],
        )
        for row in range(2, 62):
            self.assertIsNone(ws[f"A{row}"].value)
            self.assertTrue(ws[f"B{row}"].value.startswith("="))
            self.assertIsNone(ws[f"C{row}"].value)
            self.assertTrue(ws[f"D{row}"].value.startswith("="))
            self.assertTrue(ws[f"E{row}"].value.startswith("="))
            self.assertTrue(ws[f"F{row}"].value.startswith("="))

        consumidos = wb["itens_Consumidos"]
        self.assertEqual(consumidos["O1"].value, "CONS_QTD_TOTAL")
        self.assertEqual(consumidos["P1"].value, "CONS_VALOR_TOTAL")
        self.assertIn("SUM(E2,G2,I2,K2,M2)", consumidos["O2"].value)
        self.assertIn("SUM(F2,H2,J2,L2,N2)", consumidos["P2"].value)

        pcs = wb["itens_PC"]
        self.assertEqual(pcs["A1"].value, "DATA_PC")
        self.assertNotIn("NUMERO_PC", [pcs.cell(1, col).value for col in range(1, 11)])
        self.assertIn("A2", pcs["B2"].value)
        self.assertIn("COUNTA(A2:F2)", pcs["J2"].value)
        self.assertTrue(any(str(dv.sqref) == "F2:F100" for dv in pcs.data_validations.dataValidation))
        self.assertIn("itens_PC!$B$2:$B$100", resultados["C10"].value)
        self.assertIn("itens_PC!$J$2:$J$100", resultados["D44"].value)

        regras_parametros = [
            formula
            for regras in wb["parametros"].conditional_formatting._cf_rules.values()
            for regra in regras
            for formula in (regra.formula or [])
        ]
        self.assertTrue(any('$A3="Nao"' in formula and '$E3=""' in formula for formula in regras_parametros))

        remanesc = wb["itens_Remanesc"]
        self.assertIn('SUMIF($A$2:A2,"<>"', remanesc["D3"].value)
        self.assertIn('"TOTAL"', remanesc["U3"].value)
        self.assertIn("QTD_BASE_ORIGINAL", remanesc["B1"].value)
        self.assertIn("QTD_REM_BASE_SEM_ADITIVO_C2", remanesc["G1"].value)
        self.assertIn("posicao_contratual!O2", remanesc["H2"].value)
        self.assertIn("posicao_contratual!J2-posicao_contratual!N2", remanesc["M2"].value)
        self.assertIn('"TOTAL"', wb["itens_RC"]["A4"].value)
        self.assertIsNone(wb["itens_RC"]["A203"].value)
        self.assertIn("SUMIFS(itens_RC!$D$3:$D$202", resultados["D32"].value)
        self.assertIn("SUM(posicao_contratual!$O$2:$O$200)", resultados["B32"].value)
        self.assertIn("SUMPRODUCT(posicao_contratual!$O$2:$O$200", resultados["C32"].value)

        aditivos = wb["aditivos"]
        self.assertEqual(aditivos["L1"].value, "DELTA_QTD_CONTRATUAL")
        self.assertIn('LEFT(UPPER(D2),5)="ACRES"', aditivos["L2"].value)
        self.assertIn("L2*F2", aditivos["J2"].value)

        posicao = wb["posicao_contratual"]
        self.assertEqual(posicao["A1"].value, "ITEM")
        self.assertEqual(posicao["M1"].value, "QTD_CONTRATADA_C2")
        self.assertIn('aditivos!$C$2:$C$200,"C2"', posicao["L2"].value)
        self.assertIn("I2+L2", posicao["M2"].value)
        self.assertIn("N2+(M2-$C2)", posicao["O2"].value)
        self.assertIn("ITEM_DUPLICADO", posicao["X2"].value)
        self.assertIn("posicao_contratual!O2", wb["itens_RC"]["I3"].value)
        self.assertIn("posicao_contratual!M2", wb["historico_VU"]["P2"].value)

        todas_formulas = "\n".join(formulas(wb).values()).upper()
        for termo in ("#REF!", "ITENS_EXECUCAO_SALDO", "REGRA_NEGOCIO_CLAUS", "HISTORICO!", "NUMERO_PC"):
            self.assertNotIn(termo, todas_formulas)
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    self.assertIsNone(getattr(cell, "comment", None))
                    if isinstance(cell.value, str):
                        self.assertNotIn("OBSERV", cell.value.upper())

        with ZipFile(CAMINHO_MODELO_COLETA) as pacote:
            nomes = "\n".join(pacote.namelist()).lower()
        for termo in ("externallink", "comments", "vml"):
            self.assertNotIn(termo, nomes)

    def test_geracao_preserva_resultados_e_nunca_salva_workbook_data_only(self):
        dados = {
            "indice": "IPCA",
            "ciclos": [{"ciclo": "C1", "data_base": "01/01/2024", "percentual_aplicado": 0.04}],
        }
        with patch("_coleta_reajuste.load_workbook", wraps=load_workbook) as carregador:
            payload = gerar_coleta_reajuste(dados)
        self.assertTrue(carregador.call_args_list)
        self.assertTrue(all(chamada.kwargs.get("data_only") is False for chamada in carregador.call_args_list))

        entregue = load_workbook(io.BytesIO(payload), data_only=False)
        resultados = entregue["RESULTADOS"]
        self.assertEqual(resultados.sheet_state, "visible")
        self.assertEqual(resultados["A1"].value, "RESULTADOS CONSOLIDADOS — REAJUSTE CONTRATUAL")
        self.assertGreaterEqual(
            sum(
                1
                for row in resultados.iter_rows()
                for cell in row
                if isinstance(cell.value, str) and cell.value.startswith("=")
            ),
            3000,
        )
        builder = (ROOT / "tools" / "build_coleta_reajuste_template.py").read_text(encoding="utf-8")
        self.assertNotIn('create_sheet("RESULTADOS"', builder)

    def test_ciclo_unico_c2_traz_c0_c1_e_c2_sem_ciclos_futuros(self):
        dados = {
            "tipo": "Simples",
            "indice": "IPCA (433)",
            "data_base_original": "01/09/2023",
            "contexto_contratual_anterior": {
                "ultimo_ciclo_concedido": "C1",
                "percentual_ja_aplicado_pct": 4.0,
            },
            "ciclos": [
                {
                    "ciclo": "C2",
                    "data_base": "01/09/2023",
                    "financeiro_inicio": "01/10/2024",
                    "percentual_aplicado": 0.05,
                    "situacao_aplicada": "Tempestivo",
                }
            ],
        }
        payload = gerar_coleta_reajuste(dados)
        wb = load_workbook(io.BytesIO(payload), data_only=False)
        modelo = load_workbook(CAMINHO_MODELO_COLETA, data_only=False)

        self.assertEqual(formulas(wb), formulas(modelo))
        self.assertEqual(wb["CONTROLE"]["B2"].value, "C2")
        self.assertEqual(wb["CONTROLE"]["B7"].value, "IPCA (433)")
        self.assertEqual(wb["parametros"]["E3"].value, 0.04)
        self.assertEqual(wb["parametros"]["E4"].value, 0.05)
        self.assertEqual(wb["parametros"]["A3"].value, "Nao")
        self.assertEqual(wb["parametros"]["A4"].value, "Sim")
        self.assertEqual(wb["parametros"]["C3"].value.strftime("%m/%Y"), "09/2023")
        self.assertEqual(wb["parametros"]["C4"].value.strftime("%m/%Y"), "09/2024")
        self.assertEqual(wb["parametros"]["G3"].value, "Histórico fora desta apuração")

        ws = wb["financeiro"]
        self.assertEqual(ws["A2"].value.strftime("%m/%Y"), "09/2022")
        self.assertEqual(ws["A25"].value.strftime("%m/%Y"), "08/2024")
        self.assertEqual(ws["A26"].value.strftime("%m/%Y"), "09/2024")
        self.assertEqual(ws["A37"].value.strftime("%m/%Y"), "08/2025")
        self.assertIsNone(ws["A38"].value)
        self.assertEqual(ws["G26"].value, "Nao")
        self.assertEqual(ws["G27"].value, "Sim")
        self.assertEqual(ws["A26"].font.color.rgb, "FF123B63")
        self.assertTrue(ws["A26"].font.bold)
        self.assertEqual(sum(ws[f"A{row}"].value is not None for row in range(2, 62)), 36)

        diagnostico = ler_coleta_reajuste(payload)
        self.assertTrue(diagnostico["valido"])
        self.assertFalse(diagnostico["pronto_para_consolidar"])
        self.assertEqual(diagnostico["metadados"]["ciclos_em_analise"], ["C2"])

    def test_multiplos_c1_a_c4_ocupam_as_60_competencias(self):
        ciclos = []
        for numero, ano, percentual in (
            (1, 2022, 0.03),
            (2, 2023, 0.04),
            (3, 2024, 0.05),
            (4, 2025, 0.06),
        ):
            ciclos.append(
                {
                    "ciclo": f"C{numero}",
                    "data_base": f"01/09/{ano}",
                    "financeiro_inicio": f"01/09/{ano + 1}",
                    "percentual_aplicado": percentual,
                }
            )
        payload = gerar_coleta_reajuste({"tipo": "Múltiplo", "indice": "IST", "ciclos": ciclos})
        wb = load_workbook(io.BytesIO(payload), data_only=False)
        ws = wb["financeiro"]
        self.assertEqual(sum(ws[f"A{row}"].value is not None for row in range(2, 62)), 60)
        self.assertEqual(ws["A2"].value.strftime("%m/%Y"), "09/2022")
        self.assertEqual(ws["A61"].value.strftime("%m/%Y"), "08/2027")
        self.assertTrue(ws["B61"].value.startswith("=IF(A61"))
        self.assertEqual([wb["parametros"][f"A{row}"].value for row in range(3, 7)], ["Sim"] * 4)

    def test_upload_parcial_nao_inventa_consolidacao(self):
        dados = {
            "indice": "IPCA",
            "ciclos": [
                {
                    "ciclo": "C1",
                    "data_base": "01/01/2024",
                    "financeiro_inicio": "01/01/2025",
                    "percentual_aplicado": 0.045,
                }
            ],
        }
        payload = gerar_coleta_reajuste(dados)
        wb = load_workbook(io.BytesIO(payload), data_only=False)
        wb["financeiro"]["C2"] = 1000.0
        output = io.BytesIO()
        wb.save(output)
        diagnostico = ler_coleta_reajuste(output.getvalue())

        self.assertTrue(diagnostico["valido"])
        self.assertFalse(diagnostico["pronto_para_consolidar"])
        self.assertEqual(diagnostico["contagens"]["competencias_com_valor"], 1)
        self.assertNotIn("total", diagnostico)
        self.assertTrue(
            any("status de RESULTADOS" in aviso for aviso in diagnostico["avisos"]),
            diagnostico["avisos"],
        )

    def test_resultados_sem_formula_estrutural_e_bloqueado(self):
        dados = {
            "indice": "IPCA",
            "ciclos": [{"ciclo": "C1", "data_base": "01/01/2024", "percentual_aplicado": 0.04}],
        }
        wb = load_workbook(io.BytesIO(gerar_coleta_reajuste(dados)), data_only=False)
        wb["RESULTADOS"]["B16"] = None
        output = io.BytesIO()
        wb.save(output)

        diagnostico = ler_coleta_reajuste(output.getvalue())
        self.assertFalse(diagnostico["valido"])
        self.assertTrue(
            any("RESULTADOS!B16" in pendencia for pendencia in diagnostico["pendencias"]),
            diagnostico["pendencias"],
        )

    def test_coleta_com_aba_excluida_e_reconhecida_e_bloqueada(self):
        dados = {
            "indice": "IPCA",
            "ciclos": [{"ciclo": "C1", "data_base": "01/01/2024", "percentual_aplicado": 0.04}],
        }
        wb = load_workbook(io.BytesIO(gerar_coleta_reajuste(dados)), data_only=False)
        wb.remove(wb["itens_Consumidos"])
        output = io.BytesIO()
        wb.save(output)
        payload = output.getvalue()

        self.assertTrue(eh_coleta_reajuste(payload))
        diagnostico = ler_coleta_reajuste(payload)
        self.assertFalse(diagnostico["valido"])
        self.assertIn("itens_Consumidos", diagnostico["pendencias"][0])

    def test_streamlit_usa_um_unico_nome_e_o_novo_motor(self):
        self.assertEqual(NOME_ARQUIVO_COLETA, "Coleta_Reajuste.xlsx")
        simples = (ROOT / "pages" / "01_Calculo_Simples.py").read_text(encoding="utf-8")
        multiplos = (ROOT / "pages" / "02_Calculo_Represados.py").read_text(encoding="utf-8")
        valores = (ROOT / "pages" / "03_Valor_Global.py").read_text(encoding="utf-8")
        for fonte in (simples, multiplos):
            self.assertIn("gerar_coleta_reajuste", fonte)
            self.assertIn("file_name=NOME_ARQUIVO_COLETA", fonte)
        self.assertNotIn("render_botao_download_modelo_consumo(modelo_consumo)", multiplos)
        self.assertIn("ler_coleta_reajuste", valores)
        self.assertIn("render_status_apuracao", valores)
        self.assertIn("render_status_documentos", valores)
        self.assertIn('if diagnostico.get("valido"):', valores)

        adapter = (ROOT / "_coleta_reajuste_documentos.py").read_text(encoding="utf-8")
        self.assertNotIn('if not diagnostico.get("pronto_para_consolidar"):', adapter)
        self.assertIn('fator = _numero(parametros[f"F{row}"].value, 1.0)', adapter)
        self.assertIn('"Data-base": _data_br(parametros[f"C{row}"].value)', adapter)
        self.assertIn('"Situação": parametros[f"G{row}"].value or ""', adapter)
        self.assertIn('"Variação": _numero(parametros[f"E{row}"].value)', adapter)


if __name__ == "__main__":
    unittest.main()
