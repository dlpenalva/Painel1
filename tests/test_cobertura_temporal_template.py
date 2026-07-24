"""Etapa — aba de diagnostico "cobertura_temporal" no template oficial.

Hotfix: o BLOCO B separa ULTIMA EVIDENCIA (auto, MAX) de COBERTURA CONFIRMADA
COMPLETA (entrada GCC). Testes estaticos (openpyxl) validam a estrutura, o reuso
do painel homologado de posicao_referencia, as cores reutilizadas, os dois novos
campos GCC e a INVARIANCIA do VTA oficial. O teste de integracao
(RUN_EXCEL_INTEGRATION=1) recalcula a projecao fail-closed no Excel real.
"""
from __future__ import annotations

import gc
import os
import shutil
from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "templates" / "COLETA_REAJUSTE_OFICIAL.xlsx"
LEGADO = ROOT / "templates" / "Coleta_Reajuste.xlsx"
ABA = "cobertura_temporal"

pytest_com = pytest.mark.skipif(
    os.environ.get("RUN_EXCEL_INTEGRATION") != "1",
    reason="defina RUN_EXCEL_INTEGRATION=1 para executar o Excel COM",
)


def _wb():
    return load_workbook(TEMPLATE, data_only=False)


# ---------------------------------------------------------------- estaticos

def test_aba_existe_e_visivel():
    wb = _wb()
    assert ABA in wb.sheetnames
    assert wb[ABA].sheet_state == "visible"
    assert wb.sheetnames[-1] == "RESULTADOS"           # RESULTADOS continua ultima


def test_titulo_e_banners():
    ws = _wb()[ABA]
    assert "COBERTURA TEMPORAL" in str(ws["A1"].value)
    assert "NAO altera o VTA" in str(ws["A1"].value)
    assert "BLOCO A" in str(ws["A3"].value)
    assert "ULTIMA EVIDENCIA" in str(ws["A10"].value)  # BLOCO B renomeado
    assert "BLOCO C" in str(ws["A18"].value)
    assert ws["A1"].fill.fgColor.rgb == "FF1F4E79"


def test_reusa_painel_posicao_referencia():
    ws = _wb()[ABA]
    assert ws["B7"].value == "=posicao_referencia!$I$6"
    assert ws["B8"].value == '=IF(posicao_referencia!$I$2,posicao_referencia!$I$5,"")'
    assert ws["B11"].value == "=posicao_referencia!$I$5"
    assert ws["B22"].value == "=posicao_referencia!$I$5"     # observada (data)
    assert ws["B23"].value == "=posicao_referencia!$I$8"     # observada (origem)


def test_ultima_evidencia_nao_e_completo_ate():
    ws = _wb()[ABA]
    # rotulos deixam claro que MAX e ultima evidencia, nao "completo ate".
    assert "Ultima evidencia Financeiro" in str(ws["A12"].value)
    assert "completo ate" not in str(ws["A12"].value).lower().replace("nao e completo ate", "")
    assert "Ultima evidencia PC" in str(ws["A14"].value)
    assert "MAX(financeiro!$A$2:$A$200)" in ws["B12"].value
    assert "MAX(itens_PC!$B$2:$B$200)" in ws["B14"].value


def test_campos_gcc_confirmacao_completa():
    ws = _wb()[ABA]
    # dois novos campos GCC (amarelo de entrada), SEM formula.
    assert "Financeiro confirmado completo ate" in str(ws["A13"].value)
    assert "PC confirmado completo ate" in str(ws["A15"].value)
    assert ws["B13"].value in (None, "")
    assert ws["B15"].value in (None, "")
    assert ws["B13"].fill.fgColor.rgb == "FFFEF9C3"
    assert ws["B15"].fill.fgColor.rgb == "FFFEF9C3"


def test_projecao_fail_closed_usa_cobertura_confirmada():
    """Projecao ancora em MAX(fisica, confirmadas GCC B13/B15), nunca no MAX(evidencia)."""
    f = _wb()[ABA]["B16"].value
    assert "MAX($B$11,$B$13,$B$15)" in f          # fisica + confirmadas GCC
    assert "$B$12" not in f and "$B$14" not in f  # NAO usa ultima evidencia
    assert _wb()[ABA]["B16"].fill.fgColor.rgb == "FFFCE4D6"


def test_modo_temporal_seis_estados():
    f = _wb()[ABA]["B19"].value
    for estado in ("POSICAO_ATUAL", "HIBRIDO_TEMPORAL", "FINANCEIRO_POSTERIOR",
                   "PC_POSTERIOR", "POSICAO_DE_CORTE"):
        assert estado in f
    assert "posicao_referencia!$I$2" in f


def test_entradas_gcc_amarelas():
    ws = _wb()[ABA]
    amarelas = [r for r in range(4, 25) if ws.cell(r, 2).fill.fgColor.rgb == "FFFEF9C3"]
    assert amarelas == [4, 13, 15]     # data analise + 2 confirmacoes GCC


def test_projecao_categoria_nova_laranja():
    ws = _wb()[ABA]
    assert ws["B16"].fill.fgColor.rgb == "FFFCE4D6"
    assert ws["B24"].fill.fgColor.rgb == "FFFCE4D6"
    assert "nao cria retroativo" in ws["B24"].value


def test_legenda_quatro_categorias():
    ws = _wb()[ABA]
    assert "LEGENDA" in str(ws["A26"].value)
    rotulos = [str(ws.cell(r, 1).value) for r in range(27, 31)]
    assert rotulos == ["FISCAL", "GCC", "AUTOMATICO", "PROJECAO"]


def test_sem_today_now_hoje():
    ws = _wb()[ABA]
    textos = [str(c.value) for row in ws.iter_rows() for c in row if isinstance(c.value, str)]
    for f in textos:
        for proib in ("TODAY(", "NOW(", "HOJE(", "AGORA("):
            assert proib not in f.upper()


def test_vta_oficial_invariante():
    wb = _wb()
    r = wb["RESULTADOS"]
    assert r["B23"].value == '=IF(OR(B20="",B21="",B22=""),"",ROUND(B20+B21+B22,2))'
    assert "$N$263" in r["B26"].value and ABA not in str(r["B26"].value)
    assert r["B25"].value in (None, "")


def test_template_legado_intacto():
    if not LEGADO.exists():
        pytest.skip("legado ausente")
    assert ABA not in load_workbook(LEGADO, data_only=False).sheetnames


# ---------------------------------------------------------------- COM

def _recalc(caminho: Path, inspecionar):
    import time
    import pythoncom
    import win32com.client

    pythoncom.CoInitialize()
    excel = win32com.client.DispatchEx("Excel.Application")
    for a, v in (("Visible", False), ("DisplayAlerts", False)):
        try:
            setattr(excel, a, v)
        except AttributeError:
            pass
    wb = None
    try:
        wb = excel.Workbooks.Open(str(caminho.resolve()), UpdateLinks=0)
        excel.CalculateFullRebuild()
        time.sleep(0.3)
        return inspecionar(wb)
    finally:
        if wb is not None:
            try:
                wb.Close(SaveChanges=False)
            except Exception:
                pass
        try:
            excel.Quit()
        except Exception:
            pass
        gc.collect()
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass


def _montar_cenario(dest, *, gcc_pc=None):
    shutil.copy2(TEMPLATE, dest)
    wb = load_workbook(dest, data_only=False)
    p = wb["parametros"]
    for num, ini, fim in ((0, date(2023, 1, 1), date(2023, 12, 31)),
                          (1, date(2024, 1, 1), date(2024, 12, 31)),
                          (2, date(2025, 1, 1), date(2025, 12, 31)),
                          (3, date(2026, 1, 1), date(2026, 12, 31))):
        p.cell(num + 2, 3).value = ini
        p.cell(num + 2, 4).value = fim
    ir = wb["itens_Remanesc"]
    ir["A2"], ir["B2"], ir["C2"] = "ITEM-1", 100.0, 10.0
    ir["E2"] = ir["G2"] = ir["I2"] = 100.0    # fotografia C1,C2,C3
    wb["CONTROLE"]["B2"] = "C3"
    pc = wb["itens_PC"]
    pc["A2"], pc["B2"], pc["D2"] = "PC-1", date(2026, 5, 20), 2000.0
    cob = wb[ABA]
    cob["B4"] = date(2026, 6, 30)             # data da analise
    if gcc_pc is not None:
        cob["B15"] = gcc_pc                    # PC confirmado completo ate (GCC)
    wb.save(dest)
    return dest


def _ler(w):
    cv = w.Worksheets(ABA)
    r = w.Worksheets("RESULTADOS")
    return {
        "modo": cv.Range("B19").Value,
        "fisica": cv.Range("B11").Value,
        "pc_ultima": cv.Range("B14").Value,
        "pc_conf": cv.Range("B15").Value,
        "proj": cv.Range("B16").Value,
        "b23": r.Range("B23").Value,
        "b26": r.Range("B26").Value,
    }


@pytest_com
def test_com_projecao_fail_closed_sem_gcc(tmp_path):
    """PC maio sem confirmacao GCC: projecao ancora na fisica (marco), nao junho."""
    dest = _montar_cenario(tmp_path / "COLETA_REAJUSTE_OFICIAL.xlsx", gcc_pc=None)
    d = _recalc(dest, _ler)
    assert d["modo"] == "PC_POSTERIOR"
    assert (d["fisica"].year, d["fisica"].month) == (2026, 1)   # abertura C3
    assert (d["pc_ultima"].year, d["pc_ultima"].month) == (2026, 5)
    # projecao = dia seguinte a fisica (jan/2026), NAO junho por causa do PC.
    assert (d["proj"].year, d["proj"].month) == (2026, 1)


@pytest_com
def test_com_projecao_autorizada_com_gcc(tmp_path):
    """PC maio + GCC confirma ate 31/05: projecao autorizada a partir de 01/06."""
    dest = _montar_cenario(tmp_path / "COLETA_REAJUSTE_OFICIAL.xlsx",
                           gcc_pc=date(2026, 5, 31))
    d = _recalc(dest, _ler)
    assert (d["pc_conf"].year, d["pc_conf"].month, d["pc_conf"].day) == (2026, 5, 31)
    assert (d["proj"].year, d["proj"].month, d["proj"].day) == (2026, 6, 1)
