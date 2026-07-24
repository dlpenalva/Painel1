"""Wiring ponta a ponta: XLS oficial -> ler_masterfile_v10 -> montar_cobertura_temporal.

Prova que as declaracoes GCC da aba cobertura_temporal (Data da analise,
Financeiro/PC confirmado completo ate) fluem pelo leitor real ate o motor e que
as regras temporais homologadas se mantem:
  * MAX(data) continua sendo apenas ULTIMA EVIDENCIA (nunca confirmacao);
  * celula vazia => None (nunca vira data/zero);
  * projecao fail-closed usa a cobertura confirmada, nao a ultima evidencia;
  * arquivo oficial anterior SEM a aba continua legivel (compatibilidade).
"""
from __future__ import annotations

import io
import gc
import os
import shutil
from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook

from _leitor_masterfile_v10 import ler_masterfile_v10
from _motor_cobertura_temporal import montar_cobertura_temporal

ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "templates" / "COLETA_REAJUSTE_OFICIAL.xlsx"
ABA = "cobertura_temporal"

pytest_com = pytest.mark.skipif(
    os.environ.get("RUN_EXCEL_INTEGRATION") != "1",
    reason="defina RUN_EXCEL_INTEGRATION=1 para executar o Excel COM",
)


def _preencher(wb, *, gcc_fin=None, gcc_pc=None, data_analise=date(2026, 6, 30),
               com_aba=True):
    """Preenche calendario + PC + declaracoes GCC (via openpyxl, valores diretos)."""
    p = wb["parametros"]
    for num, (ini, fim) in enumerate([
            (date(2023, 1, 1), date(2023, 12, 31)),
            (date(2024, 1, 1), date(2024, 12, 31)),
            (date(2025, 1, 1), date(2025, 12, 31)),
            (date(2026, 1, 1), date(2026, 12, 31))]):
        p.cell(num + 2, 3).value = ini
        p.cell(num + 2, 4).value = fim
    c = wb["CONTROLE"]
    c["B1"] = "Pedidos de Compras"   # modo pc -> le itens_PC
    c["B2"] = "C3"
    ir = wb["itens_Remanesc"]
    ir["A2"], ir["B2"], ir["C2"] = "ITEM-1", 100.0, 10.0
    ir["E2"] = ir["G2"] = ir["I2"] = 100.0
    pc = wb["itens_PC"]
    pc["A2"], pc["B2"], pc["D2"] = "PC-1", date(2026, 5, 20), 2000.0
    if com_aba:
        cob = wb[ABA]
        cob["B4"] = data_analise
        if gcc_fin is not None:
            cob["B13"] = gcc_fin
        if gcc_pc is not None:
            cob["B15"] = gcc_pc
    else:
        del wb[ABA]                  # simula arquivo oficial anterior sem a aba


def _bytes(**kw) -> bytes:
    wb = load_workbook(TEMPLATE, data_only=False)
    _preencher(wb, **kw)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --------------------------------------------------------------------------- #
# CENARIO 1 — confirmacoes GCC preenchidas.
# --------------------------------------------------------------------------- #
def test_cenario1_confirmacoes_preenchidas():
    res = ler_masterfile_v10(_bytes(
        gcc_fin=date(2026, 5, 31), gcc_pc=date(2026, 5, 31)))
    # XLS -> leitor
    assert res["controle"]["data_analise"] == date(2026, 6, 30)
    assert res["confirmacao_gcc"]["financeiro_ate"] == date(2026, 5, 31)
    assert res["confirmacao_gcc"]["pc_ate"] == date(2026, 5, 31)
    # leitor -> motor (mesmo payload real)
    r = montar_cobertura_temporal(res)
    assert r.data_analise == date(2026, 6, 30)
    assert r.pc_cobertura_confirmada_ate == date(2026, 5, 31)
    assert r.financeiro_cobertura_confirmada_ate == date(2026, 5, 31)
    assert r.pc_cobertura_adotada_ate == date(2026, 5, 31)   # confirmacao prevalece
    assert r.projecao_autorizada_a_partir_de == date(2026, 6, 1)
    # MAX continua apenas ultima evidencia (PC informado em 20/05).
    assert r.pc_ultima_evidencia == date(2026, 5, 20)


# --------------------------------------------------------------------------- #
# CENARIO 2 — confirmacoes GCC vazias.
# --------------------------------------------------------------------------- #
def test_cenario2_confirmacoes_vazias():
    res = ler_masterfile_v10(_bytes(gcc_fin=None, gcc_pc=None))
    assert res["controle"]["data_analise"] == date(2026, 6, 30)
    assert res["confirmacao_gcc"] == {"financeiro_ate": None, "pc_ate": None}
    r = montar_cobertura_temporal(res)
    assert r.pc_ultima_evidencia == date(2026, 5, 20)         # ultima evidencia
    assert r.pc_cobertura_confirmada_ate is None
    assert r.pc_cobertura_adotada_ate is None                 # MAX nao vira confirmacao
    # projecao NAO comeca em junho por causa do PC de maio; ancora na fisica.
    assert r.projecao_autorizada_a_partir_de.month == 1       # fisica = abertura C3 (jan)
    assert r.projecao_autorizada_a_partir_de < date(2026, 6, 1)


# --------------------------------------------------------------------------- #
# CENARIO 3 — arquivo oficial anterior SEM a aba cobertura_temporal.
# --------------------------------------------------------------------------- #
def test_cenario3_compatibilidade_sem_aba():
    res = ler_masterfile_v10(_bytes(com_aba=False))
    assert res["ok"] is True and not res["erro"]              # nao quebra
    assert res["controle"]["data_analise"] is None
    assert res["confirmacao_gcc"] == {"financeiro_ate": None, "pc_ate": None}
    assert res["cobertura_temporal"]["ok"] is False           # aba ausente
    # motor continua fail-closed sem confirmacao nem data de analise.
    r = montar_cobertura_temporal(res)
    assert r.data_analise is None
    assert r.pc_cobertura_adotada_ate is None
    assert r.projecao_autorizada_a_partir_de is None          # sem analise, sem projecao


def test_vazio_nunca_vira_data():
    """Celula GCC vazia jamais e coagida a data (fail-closed)."""
    res = ler_masterfile_v10(_bytes(gcc_fin=None, gcc_pc=date(2026, 5, 31)))
    assert res["confirmacao_gcc"]["financeiro_ate"] is None   # vazio
    assert res["confirmacao_gcc"]["pc_ate"] == date(2026, 5, 31)


# --------------------------------------------------------------------------- #
# COM — copia preenchida e recalculada no Excel real (prova do fluxo completo).
# --------------------------------------------------------------------------- #
@pytest_com
def test_com_cenario1_recalc_real(tmp_path):
    import time
    import pythoncom
    import win32com.client

    dest = tmp_path / "COLETA_REAJUSTE_OFICIAL.xlsx"
    shutil.copy2(TEMPLATE, dest)
    wb = load_workbook(dest, data_only=False)
    _preencher(wb, gcc_fin=date(2026, 5, 31), gcc_pc=date(2026, 5, 31))
    wb.save(dest)

    pythoncom.CoInitialize()
    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        w = excel.Workbooks.Open(str(dest.resolve()), UpdateLinks=0)
        excel.CalculateFullRebuild()
        time.sleep(0.3)
        w.Save()
        w.Close(SaveChanges=False)
    finally:
        excel.Quit()
        gc.collect()
        pythoncom.CoUninitialize()

    res = ler_masterfile_v10(dest.read_bytes())
    assert res["controle"]["data_analise"] == date(2026, 6, 30)
    assert res["confirmacao_gcc"]["pc_ate"] == date(2026, 5, 31)
    r = montar_cobertura_temporal(res)
    assert r.projecao_autorizada_a_partir_de == date(2026, 6, 1)
