# -*- coding: utf-8 -*-
"""Etapa 26J.1 — bloqueadores da auditoria independente da 26J.

Cobertura permanente:
  A. unicidade semantica do VTA (a composicao legada B23 nunca e exposta
     como VTA; a fonte canonica e o VTA FINAL/B26);
  B. reconstrucao economica do VTA independente do motor (expectativa
     derivada das entradas, nao do objeto de composicao);
  C. acrescimo + supressao atravessando a cadeia ate o VTA (delta liquido
     x VU do ciclo), sem dupla contagem do aditivo;
  D. round-trip real de formatacao da mensagem fiscal (template degradado
     -> geracao -> XLSX final legivel);
  E. semantica de pre-vigencia (vazio antes do nascimento; "Nao informado"
     reservado a ausencia real);
  F. residual 26C: protecao anti-duplicidade aplicada nao e promovida a
     "possivel dupla contagem".
"""
from __future__ import annotations

import io
from datetime import date

from openpyxl import load_workbook

import _motor_composicao_vta as motor_vta
from _capacidades_apuracao import avaliar_capacidades_apuracao


# --------------------------------------------------------------------------- #
# A. Unicidade semantica do VTA
# --------------------------------------------------------------------------- #

_VTA_OFICIAL = 137_375_560.29
_VTA_LEGADO_B23 = 125_682_374.84


def _metadados_pc(status_vta: str = "CÁLCULO MANUAL REQUERIDO") -> dict:
    return {
        "ciclos_em_analise": ["C0", "C1"],
        "status_resultados": {
            "geral": "REVISE",
            "metodo_retroativo": "PCs",
            "vta": status_vta,
            "retroativo": "CALCULADO — CONFERIR",
            "remanescente": "CALCULADO — CONFERIR",
            "valores": {
                "retroativo_oficial": 168_372.32,
                "vta_base_contratual": 122_036_510.12,
                "vta_retroativo": 168_372.32,
                "vta_ajuste_remanescente": 3_477_492.40,
                "vta_calculado": _VTA_LEGADO_B23,
                "vta_pc_execucao_anterior": 13_973_327.58,
                "vta_pc_parcelas_intermediarias": 0,
                "vta_pc_remanescente_corte": 123_402_232.71,
                "vta_pc_total": _VTA_OFICIAL,
                "vta_ajuste_manual": None,
                "vta_manual_oficial": None,
                "vta_oficial": _VTA_OFICIAL,
                "remanescente_atualizado": 123_402_232.71,
            },
        },
    }


_CONTAGENS = {
    "competencias_com_valor": 3,
    "itens_remanescentes": 10,
    "itens_consumidos": 0,
    "pedidos_de_compra": 2118,
    "aditivos": 9,
    "posicao_contratual_itens": 10,
    "posicao_contratual_calculada": 10,
    "historico_vu_itens": 10,
    "historico_vu_calculado": 10,
}


def test_a_vta_unico_mesmo_com_status_pendente():
    """B23 (composicao legada) nunca substitui o VTA FINAL, nem com REVISE."""
    resultado = avaliar_capacidades_apuracao(_CONTAGENS, _metadados_pc())
    vta = resultado["calculos"]["vta"]
    assert vta["valor"] == _VTA_OFICIAL
    assert vta["origem"] == "RESULTADOS"
    assert vta["disponivel"] is True
    # A grandeza legada permanece registrada, mas jamais no campo VTA.
    assert vta["valor"] != _VTA_LEGADO_B23


def test_a_rastreabilidade_vta_reproduz_pelo_metodo_pc():
    resultado = avaliar_capacidades_apuracao(_CONTAGENS, _metadados_pc())
    trilha = resultado["rastreabilidade"]["resultados"]["vta"]
    assert trilha["valor"] == _VTA_OFICIAL
    assert trilha["valor_reproduzido"] == _VTA_OFICIAL
    assert trilha["reproduzivel"] is True
    assert trilha["componentes"] == {
        "execucao_anterior_ao_corte": 13_973_327.58,
        "parcelas_intermediarias": 0,
        "remanescente_no_corte": 123_402_232.71,
    }
    legado = trilha["composicao_legada_remanescentes"]
    assert legado["valor"] == _VTA_LEGADO_B23
    assert "não é o" in legado["significado"] or "nao e o" in legado["significado"]


def test_a_todos_os_campos_vta_convergem():
    """Deteccao permanente de um segundo 'VTA' numerico divergente."""
    resultado = avaliar_capacidades_apuracao(_CONTAGENS, _metadados_pc())
    valores_vta = {
        "calculos.vta": resultado["calculos"]["vta"]["valor"],
        "rastreabilidade.vta": resultado["rastreabilidade"]["resultados"]["vta"]["valor"],
    }
    assert set(valores_vta.values()) == {_VTA_OFICIAL}, valores_vta


def test_a_metodo_legado_continua_reproduzivel():
    """Fora do metodo PC, B26 = B23 (+ajustes) e a trilha legada permanece."""
    metadados = _metadados_pc()
    sr = metadados["status_resultados"]
    sr["metodo_retroativo"] = "Financeiro"
    sr["valores"]["vta_oficial"] = _VTA_LEGADO_B23
    resultado = avaliar_capacidades_apuracao(_CONTAGENS, metadados)
    trilha = resultado["rastreabilidade"]["resultados"]["vta"]
    assert trilha["valor"] == _VTA_LEGADO_B23
    assert trilha["valor_reproduzido"] == _VTA_LEGADO_B23
    assert trilha["reproduzivel"] is True


# --------------------------------------------------------------------------- #
# B/C. Testes economicos independentes (motor de composicao VTA)
# --------------------------------------------------------------------------- #

# Cenario sintetico versionavel: entradas explicitas; a expectativa e
# calculada AQUI, com aritmetica propria, nunca copiada do motor.
#   - item 1: VU_C1 = 4.554,77 (VU do cenario auditado 26J)
#   - execucao anterior ao corte: PCs de C0 ja atualizados.
_ITENS_ECON = [
    # (ITEM, VU_C0, VU_C1, QTD_REM_AJUSTADA_C0, QTD_REM_AJUSTADA_C1)
    (1, 4_418.74, 4_554.77, 12, 10),
    (2, 6_358.99, 6_554.75, 8, 8),
    (3, 8_188.70, 8_440.79, 5, 3),
]
_PCS_C0 = [37_500.00, 12_449.50]  # valores atualizados de PCs anteriores a C1


def _leitura_econ(qtd_rem_c1_item1: float) -> dict:
    posicao = []
    historico = []
    for item, vu0, vu1, q0, q1 in _ITENS_ECON:
        if item == 1:
            q1 = qtd_rem_c1_item1
        posicao.append({
            "ITEM": item, "VU_ORIGINAL": vu0,
            "QTD_REM_AJUSTADA_C0": q0, "QTD_REM_AJUSTADA_C1": q1,
            "QTD_REM_AJUSTADA_C2": 0, "QTD_REM_AJUSTADA_C3": 0,
            "QTD_REM_AJUSTADA_C4": 0,
        })
        historico.append({
            "item": item,
            "vu_ciclos": {"VU_C0": vu0, "VU_C1": vu1},
        })
    itens_pc = [
        {"ciclo": "C0", "valor_pc": v, "valor_atualizado": v,
         "entra_no_calculo": "Sim"}
        for v in _PCS_C0
    ]
    return {
        "controle": {"modo": "pc", "ciclo_vigente": "C1"},
        "parametros_v10": {"por_ciclo": {}},
        "posicao_contratual": {"itens": posicao},
        "historico_vu": {"itens": historico},
        "itens_pc_v10": {"itens": itens_pc},
    }


def _vta_esperado_independente(qtd_rem_c1_item1: float) -> float:
    """Reconstrucao externa: execucao anterior + remanescente no corte."""
    execucao = round(sum(_PCS_C0), 2)
    remanescente = 0.0
    for item, _vu0, vu1, _q0, q1 in _ITENS_ECON:
        if item == 1:
            q1 = qtd_rem_c1_item1
        remanescente += round(q1 * round(vu1, 2), 2)
    return round(execucao + round(remanescente, 2), 2)


def test_b_vta_reconstruido_independentemente():
    comp = motor_vta.montar_composicao_vta(_leitura_econ(10))
    assert comp["disponivel"] is True
    esperado = _vta_esperado_independente(10)
    # Conferencia aritmetica explicita do proprio teste:
    # execucao 49.949,50 + remanescente (10x4.554,77 + 8x6.554,75 +
    # 3x8.440,79) = 49.949,50 + 123.308,07 = 173.257,57
    assert esperado == 173_257.57
    assert comp["vta_composicao"] == esperado


def test_c_acrescimo_e_supressao_propagam_ate_o_vta():
    """Acrescimo +2 e supressao -1 no item 1 em C1: delta = +1 x VU_C1.

    O efeito fisico do aditivo ja esta consolidado na posicao contratual
    (quantidade remanescente ajustada); nenhuma parcela autonoma de aditivo
    e somada novamente ao VTA (prova de nao-dupla-contagem).
    """
    vu_c1_item1 = next(vu1 for item, _v0, vu1, _q0, _q1 in _ITENS_ECON if item == 1)
    base = motor_vta.montar_composicao_vta(_leitura_econ(10))
    com_delta = _leitura_econ(11)  # 10 + 2 (acrescimo) - 1 (supressao)
    com_delta["aditivos_visiveis"] = {
        "ok": True,
        "itens": [
            {"evento": "Acrescimo item 1", "ciclo_marco": "C1",
             "valor_assinatura": 2 * vu_c1_item1,
             "valor_atualizado": 2 * vu_c1_item1,
             "ja_refletido_em": "posicao_contratual/remanescente"},
            {"evento": "Supressao item 1", "ciclo_marco": "C1",
             "valor_assinatura": -vu_c1_item1,
             "valor_atualizado": -vu_c1_item1,
             "ja_refletido_em": "posicao_contratual/remanescente"},
        ],
    }
    alterado = motor_vta.montar_composicao_vta(com_delta)
    delta = round(alterado["vta_composicao"] - base["vta_composicao"], 2)
    assert delta == round(1 * vu_c1_item1, 2) == 4_554.77
    # Nenhum aditivo entra como parcela autonoma no metodo PC (o efeito
    # fisico ja esta na posicao contratual — vedada dupla contagem).
    assert alterado["aditivos"] == []
    assert alterado["total_aditivos_atualizados"] == 0.0
    # Reconstrucao independente tambem confirma o estado final.
    assert alterado["vta_composicao"] == _vta_esperado_independente(11)


# --------------------------------------------------------------------------- #
# D. Round-trip real da formatacao da mensagem fiscal (aditivos!M)
# --------------------------------------------------------------------------- #

def _dados_geracao() -> dict:
    return {
        "origem": "Teste 26J.1",
        "indice": "IST",
        "data_base_original": "01/01/2024",
        "data_corte": date(2025, 12, 31),
        "ciclos": [{
            "ciclo": "C1", "data_inicio": date(2025, 1, 1),
            "data_fim": date(2025, 12, 31), "data_pedido": date(2025, 1, 1),
            "financeiro_inicio": date(2025, 1, 1), "percentual": 0.10,
        }],
    }


def test_d_round_trip_reaplica_formatacao_da_mensagem_fiscal(tmp_path):
    """template degradado -> geracao runtime -> XLSX FINAL legivel."""
    from _coleta_oficial import (
        ALTURA_MINIMA_LINHAS_ADITIVOS,
        LARGURA_MINIMA_ADITIVOS_M,
        TEMPLATE_COLETA_OFICIAL,
        normalizar_dados_calculadora,
    )
    from _gerador_masterfile import gerar_masterfile_preenchido

    # Simula linhagem antiga (pre-26H.2): sem WrapText e dimensoes reduzidas.
    wb = load_workbook(TEMPLATE_COLETA_OFICIAL)
    ws = wb["aditivos"]
    ws.column_dimensions["M"].width = 30.0
    from copy import copy as _copy
    for row in range(2, 201):
        cel = ws.cell(row=row, column=13)
        alin = _copy(cel.alignment)
        alin.wrapText = False
        cel.alignment = alin
        ws.row_dimensions[row].height = 15.9
    degradado = io.BytesIO()
    wb.save(degradado)
    wb.close()

    final = gerar_masterfile_preenchido(
        normalizar_dados_calculadora(_dados_geracao()),
        degradado.getvalue(),
    )
    caminho = tmp_path / "final_26j1.xlsx"
    caminho.write_bytes(final)

    reaberto = load_workbook(caminho)
    ws_final = reaberto["aditivos"]
    assert ws_final.column_dimensions["M"].width >= LARGURA_MINIMA_ADITIVOS_M
    for row in (2, 50, 200):
        assert ws_final.cell(row=row, column=13).alignment.wrap_text is True
        assert (
            ws_final.row_dimensions[row].height or 0
        ) >= ALTURA_MINIMA_LINHAS_ADITIVOS
    reaberto.close()


def test_d_garantia_nao_reduz_dimensoes_melhores():
    from _coleta_oficial import garantir_formatacao_orientacao_aditivos
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "aditivos"
    ws.column_dimensions["M"].width = 60.0
    ws.row_dimensions[2].height = 80.0
    garantir_formatacao_orientacao_aditivos(wb)
    assert ws.column_dimensions["M"].width == 60.0
    assert ws.row_dimensions[2].height == 80.0
    assert ws.cell(row=2, column=13).alignment.wrap_text is True


# --------------------------------------------------------------------------- #
# E. Pre-vigencia (N002/N003 antes do nascimento)
# --------------------------------------------------------------------------- #

def test_e_pre_vigencia_celulas_vazias_antes_do_nascimento():
    from _sumario_executivo import _celulas_item_no_ciclo

    item = {
        "item": "N002",
        "ciclo_nascimento": 1,
        "quantidade_ciclos": {"C0": None, "C1": 945.0},
        "vu_c0": None,
        "total_c0": None,
        "vu_ciclos": {"C1": 7_031.62},
        "total_ciclos": {"C0": None, "C1": 6_644_880.90},
    }
    # Antes do nascimento: vazio, nunca zero nem "Nao informado".
    assert _celulas_item_no_ciclo(item, "C0") == ["", "", ""]
    # No ciclo de nascimento: valores reais formatados.
    qtd, vu, total = _celulas_item_no_ciclo(item, "C1")
    assert qtd == "945,00"
    assert vu == "R$ 7.031,62"
    assert total == "R$ 6.644.880,90"


def test_e_ausencia_real_continua_nao_informado():
    from _sumario_executivo import NAO_INFORMADO, _celulas_item_no_ciclo

    item = {
        "item": "77",
        "ciclo_nascimento": 0,   # vigente desde C0
        "quantidade_ciclos": {"C0": None},
        "vu_c0": None,
        "total_c0": None,
        "vu_ciclos": {},
        "total_ciclos": {},
    }
    assert _celulas_item_no_ciclo(item, "C0") == [NAO_INFORMADO] * 3


def test_e_montar_secao_itens_propaga_ciclo_nascimento():
    from _sumario_executivo import _montar_secao_itens

    memoria = {
        "vu_itens": [{
            "item": "N002",
            "quantidade_contratada": 945,
            "vu_original": 7_031.62,
            "vu_c0": None,
            "vu_ciclos": {"C1": 7_031.62},
            "vu_por_ciclo": {"C0": None, "C1": 7_031.62},
            "quantidade_por_ciclo": {"C0": None, "C1": 945.0},
            "total_por_ciclo": {"C0": None, "C1": 6_644_880.90},
            "ciclo_nascimento": 1,
        }],
    }
    parametros = {"por_ciclo": {"C1": {"fator_acumulado": 1.0307}}}
    itens = _montar_secao_itens(memoria, parametros)
    assert itens[0]["ciclo_nascimento"] == 1
    assert itens[0]["quantidade_ciclos"]["C0"] is None


# --------------------------------------------------------------------------- #
# F. Residual 26C — protecao aplicada nao vira "possivel dupla contagem"
# --------------------------------------------------------------------------- #

def test_f_protecao_temporal_nao_gera_acao_de_dupla_contagem():
    from _assistente_fiscal import ACAO_REVISAR_DUPLA, _traduzir_inconsistencias

    painel = {"alertas": [{
        "nivel": "ALERTA",
        "codigo": "AVISO_LEITOR",
        "mensagem": (
            "composicao_vta: Composicao VTA-PC: 1047 PC(s) no ciclo vigente "
            "(C1) ou posteriores nao entram na execucao (mesmo corte "
            "temporal; evita dupla contagem contra o remanescente). Seguem "
            "como diagnostico/projecao."
        ),
    }]}
    itens = _traduzir_inconsistencias(painel)
    assert len(itens) == 1
    assert itens[0]["gravidade"] == "informação"
    assert itens[0]["acao"] == ""
    assert itens[0]["descricao"].startswith("Proteção contra dupla contagem aplicada")
    assert itens[0]["acao"] != ACAO_REVISAR_DUPLA


def test_f_risco_real_continua_gerando_acao_de_revisao():
    from _assistente_fiscal import ACAO_REVISAR_DUPLA, _traduzir_inconsistencias

    painel = {"alertas": [{
        "nivel": "ALERTA",
        "codigo": "EVENT_LOG",
        "mensagem": "Parcela declarada como já refletido em financeiro (duplicidade).",
    }]}
    itens = _traduzir_inconsistencias(painel)
    assert len(itens) == 1
    assert itens[0]["gravidade"] == "atenção"
    assert itens[0]["acao"] == ACAO_REVISAR_DUPLA
