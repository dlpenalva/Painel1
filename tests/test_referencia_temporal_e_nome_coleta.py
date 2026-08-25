"""Referencia temporal anterior ao ciclo analisado + nome do download da Coleta.

Duas melhorias isoladas do fluxo da Calculadora:

1. bloco somente leitura com a janela de 12 meses-calendario anterior a data-base
   do PRIMEIRO ciclo abrangido pela analise, exibido apenas para
   C2 ou superior. NAO e o periodo de efeitos financeiros do ciclo anterior —
   o ciclo anterior pode ter sido pedido com atraso;
2. nome do arquivo no download com ciclos e indice da analise, seguido da
   data corrente no fuso de Brasilia.

Nada aqui toca motor, admissibilidade, datas apuradas, indices, fatores,
resultados ou o conteudo do XLSX.
"""
from __future__ import annotations

import contextlib
import io
import sys
import unittest
import zipfile
from datetime import datetime
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import _ui_utils  # noqa: E402
from _coleta_oficial import (  # noqa: E402
    ABAS_COLETA_OFICIAL,
    NOME_DOWNLOAD_COLETA,
    gerar_coleta_oficial_preenchida,
    nome_download_coleta,
)
from _reajuste_utils import (  # noqa: E402
    FUSO_BRASILIA,
    ciclos_da_analise,
    numero_do_ciclo,
    referencia_temporal_anterior,
)


def _partes_xlsx(conteudo):
    """{parte OOXML: bytes} do XLSX, sem os metadados volateis do pacote."""
    with zipfile.ZipFile(io.BytesIO(conteudo)) as zf:
        return {
            nome: zf.read(nome)
            for nome in sorted(zf.namelist())
            if nome != "docProps/core.xml"
        }


def _ciclo(
    rotulo, financeiro_inicio, situacao="Tempestivo", data_base=None, data_pedido=None
):
    """Entrada de ciclo no formato que a analise ja produz (dados_admissibilidade)."""
    ciclo = {
        "ciclo": rotulo,
        "financeiro_inicio": financeiro_inicio,
        "situacao": situacao,
        "percentual_aplicado": 0.04,
    }
    ciclo["data_base"] = data_base if data_base is not None else financeiro_inicio
    if data_pedido is not None:
        ciclo["data_pedido"] = data_pedido
    return ciclo


def _dados(*ciclos):
    return {"origem": "Reajuste Simples", "indice": "IPCA (433)", "ciclos": list(ciclos)}


# Cenarios manuais exigidos na validacao.
SO_C1 = _dados(_ciclo("C1", "01/02/2024"))
SO_C2 = _dados(_ciclo("C2", "01/02/2025"))
SO_C3 = _dados(_ciclo("C3", "01/02/2026"))
C1_C2 = _dados(_ciclo("C1", "01/02/2024"), _ciclo("C2", "01/02/2025"))
C2_C3 = _dados(_ciclo("C2", "01/02/2025"), _ciclo("C3", "01/02/2026"))
COM_PRECLUSO = _dados(
    _ciclo("C2", "01/02/2025", situacao="Precluso"),
    _ciclo("C3", "01/02/2026", situacao="Tempestivo"),
)
# Mesmo escopo C2+C3, com a ancora contratual que a geracao do XLSX exige.
XLSX_C2_C3 = _dados(
    _ciclo("C2", "01/02/2025", data_base="01/02/2025"),
    _ciclo("C3", "01/02/2026", data_base="01/02/2026"),
)
C2_DATA_BASE_DIFERENTE_DO_EFEITO = _dados(
    _ciclo(
        "C2",
        "01/05/2024",
        data_base="01/03/2023",
        data_pedido="01/05/2024",
    ),
    _ciclo(
        "C3",
        "01/05/2025",
        data_base="01/03/2024",
        data_pedido="01/05/2025",
    ),
)


class _StSomenteLeitura:
    """Stub de streamlit: registra markdown e recusa qualquer widget interativo."""

    def __init__(self):
        self.markdowns = []
        self.containers = []

    def container(self, **kwargs):
        self.containers.append(kwargs)
        return contextlib.nullcontext()

    def markdown(self, texto, **kwargs):
        self.markdowns.append(str(texto))

    def __getattr__(self, nome):
        raise AssertionError(f"bloco somente leitura nao pode usar st.{nome}")

    @property
    def texto(self):
        return "\n".join(self.markdowns)


def _renderizar(dados):
    st_falso = _StSomenteLeitura()
    original = _ui_utils.st
    _ui_utils.st = st_falso
    try:
        retorno = _ui_utils.render_referencia_temporal_anterior(dados)
    finally:
        _ui_utils.st = original
    return retorno, st_falso


class TestCiclosDaAnalise(unittest.TestCase):
    def test_numero_do_ciclo_aceita_rotulo_apurado_e_recusa_lixo(self):
        self.assertEqual(numero_do_ciclo("C3"), 3)
        self.assertEqual(numero_do_ciclo(" c3 "), 3)
        self.assertEqual(numero_do_ciclo("3"), 3)
        self.assertEqual(numero_do_ciclo(3), 3)
        for invalido in (None, "", "Ciclo 3", "C", "C3A", True, 3.5, -1):
            self.assertIsNone(numero_do_ciclo(invalido), invalido)

    def test_cenarios_manuais_de_escopo(self):
        self.assertEqual(ciclos_da_analise(SO_C1), (1,))
        self.assertEqual(ciclos_da_analise(SO_C2), (2,))
        self.assertEqual(ciclos_da_analise(SO_C3), (3,))
        self.assertEqual(ciclos_da_analise(C1_C2), (1, 2))
        self.assertEqual(ciclos_da_analise(C2_C3), (2, 3))

    def test_ciclo_precluso_permanece_no_escopo(self):
        # O escopo da apuracao nao e o resultado da admissibilidade.
        self.assertEqual(ciclos_da_analise(COM_PRECLUSO), (2, 3))

    def test_ordena_numericamente_e_nao_repete(self):
        dados = _dados(
            _ciclo("C3", "01/02/2026"),
            _ciclo("C1", "01/02/2024"),
            _ciclo("C2", "01/02/2025"),
            _ciclo("C2", "01/02/2025"),
        )
        self.assertEqual(ciclos_da_analise(dados), (1, 2, 3))

    def test_entrada_insegura_devolve_vazio(self):
        for entrada in (None, {}, [], "C1", {"ciclos": None}, {"ciclos": []},
                        {"ciclos": ["C1"]}, {"ciclos": [{"ciclo": "Ciclo 1"}]}):
            self.assertEqual(ciclos_da_analise(entrada), (), entrada)


class TestNomeDownloadColeta(unittest.TestCase):
    def test_nome_base_preservado(self):
        self.assertEqual(NOME_DOWNLOAD_COLETA, "Coleta_Reajuste.xlsx")

    def test_formato_exato_e_fuso_de_brasilia(self):
        with patch("_coleta_oficial.datetime") as datetime_mock:
            datetime_mock.now.return_value = datetime(2026, 8, 25, 12, 0, tzinfo=FUSO_BRASILIA)
            self.assertEqual(nome_download_coleta(), "Coleta_Reajuste_25-08-2026.xlsx")
            datetime_mock.now.assert_called_once_with(FUSO_BRASILIA)

    def test_dia_e_mes_usam_zero_a_esquerda(self):
        with patch("_coleta_oficial.datetime") as datetime_mock:
            datetime_mock.now.return_value = datetime(2026, 2, 5, 12, 0, tzinfo=FUSO_BRASILIA)
            self.assertEqual(nome_download_coleta(), "Coleta_Reajuste_05-02-2026.xlsx")

    def test_preserva_ciclos_e_indice_antes_da_data(self):
        with patch("_coleta_oficial.datetime") as datetime_mock:
            datetime_mock.now.return_value = datetime(2026, 8, 25, 12, 0, tzinfo=FUSO_BRASILIA)
            self.assertEqual(
                nome_download_coleta(SO_C1),
                "Coleta_Reajuste_C1_IPCA_25-08-2026.xlsx",
            )
            self.assertEqual(
                nome_download_coleta(C1_C2),
                "Coleta_Reajuste_C1_C2_IPCA_25-08-2026.xlsx",
            )

    def test_sufixo_canonico_dos_cinco_indices(self):
        casos = {
            "IST (Anatel)": "IST",
            "ICTI (Ipeadata)": "ICTI",
            "IPCA (433)": "IPCA",
            "IGP-M (189)": "IGPM",
            "INPC (188)": "INPC",
        }
        with patch("_coleta_oficial.datetime") as datetime_mock:
            datetime_mock.now.return_value = datetime(2026, 8, 25, 12, 0, tzinfo=FUSO_BRASILIA)
            for indice, sufixo in casos.items():
                with self.subTest(indice=indice):
                    dados = {**C1_C2, "indice": indice}
                    self.assertEqual(
                        nome_download_coleta(dados),
                        f"Coleta_Reajuste_C1_C2_{sufixo}_25-08-2026.xlsx",
                    )

    def test_igpm_sem_hifen_tambem_e_reconhecido(self):
        dados = {**SO_C1, "indice": "IGPM (189)"}
        with patch("_coleta_oficial.datetime") as datetime_mock:
            datetime_mock.now.return_value = datetime(2026, 8, 25, 12, 0, tzinfo=FUSO_BRASILIA)
            self.assertEqual(
                nome_download_coleta(dados),
                "Coleta_Reajuste_C1_IGPM_25-08-2026.xlsx",
            )

    def test_sem_indice_preserva_ciclos_e_data(self):
        dados = {**C2_C3, "indice": "Indice nao catalogado"}
        with patch("_coleta_oficial.datetime") as datetime_mock:
            datetime_mock.now.return_value = datetime(2026, 8, 25, 12, 0, tzinfo=FUSO_BRASILIA)
            self.assertEqual(
                nome_download_coleta(dados),
                "Coleta_Reajuste_C2_C3_25-08-2026.xlsx",
            )

    def test_entrada_sem_ciclos_usa_nome_e_data(self):
        entradas = (None, {}, {"ciclos": []}, {"ciclos": [{"ciclo": "-"}]}, "lixo")
        with patch("_coleta_oficial.datetime") as datetime_mock:
            datetime_mock.now.return_value = datetime(2026, 8, 25, 12, 0, tzinfo=FUSO_BRASILIA)
            for entrada in entradas:
                with self.subTest(entrada=entrada):
                    self.assertEqual(
                        nome_download_coleta(entrada),
                        "Coleta_Reajuste_25-08-2026.xlsx",
                    )

    def test_nao_altera_os_dados_da_analise(self):
        antes = repr(C2_C3)
        nome_download_coleta(C2_C3)
        self.assertEqual(repr(C2_C3), antes)


class TestReferenciaTemporalAnterior(unittest.TestCase):
    def test_c1_nao_tem_referencia_anterior(self):
        self.assertIsNone(referencia_temporal_anterior(SO_C1))
        self.assertIsNone(referencia_temporal_anterior(C1_C2))

    def test_exemplo_canonico_do_c3(self):
        ref = referencia_temporal_anterior(SO_C3)
        self.assertEqual(ref["ciclo_analisado"], "C3")
        self.assertEqual(ref["ciclo_anterior"], "C2")
        self.assertEqual(ref["periodo_inicio"], "01/02/2025")
        self.assertEqual(ref["periodo_fim"], "31/01/2026")
        self.assertEqual(ref["ultimo_dia_anterior"], "31/01/2026")
        self.assertEqual(ref["meses"], 12)

    def test_c2_usa_a_propria_data_base(self):
        ref = referencia_temporal_anterior(SO_C2)
        self.assertEqual(ref["ciclo_anterior"], "C1")
        self.assertEqual(ref["periodo_inicio"], "01/02/2024")
        self.assertEqual(ref["periodo_fim"], "31/01/2025")

    def test_exemplo_real_data_base_nao_e_deslocada_pelo_efeito(self):
        antes = repr(C2_DATA_BASE_DIFERENTE_DO_EFEITO)
        ref = referencia_temporal_anterior(C2_DATA_BASE_DIFERENTE_DO_EFEITO)
        ciclo, ciclo_posterior = C2_DATA_BASE_DIFERENTE_DO_EFEITO["ciclos"]

        self.assertEqual(ref["ciclo_analisado"], "C2")
        self.assertEqual(ref["ciclo_anterior"], "C1")
        self.assertEqual(ref["data_base"], "01/03/2023")
        self.assertEqual(ref["periodo_inicio"], "01/03/2022")
        self.assertEqual(ref["periodo_fim"], "28/02/2023")
        self.assertEqual(ref["ultimo_dia_anterior"], "28/02/2023")
        self.assertEqual(ciclo["data_base"], "01/03/2023")
        self.assertEqual(ciclo["data_pedido"], "01/05/2024")
        self.assertEqual(ciclo["financeiro_inicio"], "01/05/2024")
        self.assertEqual(ciclo_posterior["data_base"], "01/03/2024")
        self.assertEqual(ciclo_posterior["data_pedido"], "01/05/2025")
        self.assertEqual(ciclo_posterior["financeiro_inicio"], "01/05/2025")
        self.assertEqual(repr(C2_DATA_BASE_DIFERENTE_DO_EFEITO), antes)

    def test_referencia_e_do_primeiro_ciclo_abrangido(self):
        # C2+C3 e ciclo precluso: a ancora e sempre o primeiro ciclo.
        for dados in (C2_C3, COM_PRECLUSO):
            ref = referencia_temporal_anterior(dados)
            self.assertEqual(ref["ciclo_analisado"], "C2", dados)
            self.assertEqual(ref["ciclo_anterior"], "C1", dados)
            self.assertEqual(ref["periodo_inicio"], "01/02/2024", dados)
            self.assertEqual(ref["periodo_fim"], "31/01/2025", dados)

    def test_meses_calendario_e_nao_365_dias(self):
        # 01/03/2025 - 12 meses-calendario = 01/03/2024 (ano bissexto pelo meio);
        # 365 dias corridos dariam 02/03/2024.
        ref = referencia_temporal_anterior(_dados(_ciclo("C2", "01/03/2025")))
        self.assertEqual(ref["periodo_inicio"], "01/03/2024")
        self.assertEqual(ref["periodo_fim"], "28/02/2025")

    def test_inicio_no_meio_do_mes(self):
        ref = referencia_temporal_anterior(_dados(_ciclo("C3", "15/02/2026")))
        self.assertEqual(ref["periodo_inicio"], "15/02/2025")
        self.assertEqual(ref["periodo_fim"], "14/02/2026")

    def test_inicio_em_29_de_fevereiro_bissexto(self):
        ref = referencia_temporal_anterior(_dados(_ciclo("C2", "29/02/2024")))
        self.assertEqual(ref["periodo_inicio"], "28/02/2023")
        self.assertEqual(ref["periodo_fim"], "28/02/2024")

    def test_sem_data_apurada_nao_afirma_periodo(self):
        for valor in (None, "", "  ", "data invalida"):
            dados = _dados(_ciclo("C3", valor))
            self.assertIsNone(referencia_temporal_anterior(dados), valor)

    def test_sem_data_base_nao_usa_efeito_financeiro_como_fallback(self):
        dados = _dados(_ciclo("C2", "01/05/2024"))
        dados["ciclos"][0].pop("data_base")
        self.assertIsNone(referencia_temporal_anterior(dados))

    def test_entrada_insegura_nao_exibe_bloco(self):
        for entrada in (None, {}, {"ciclos": []}, "lixo"):
            self.assertIsNone(referencia_temporal_anterior(entrada), entrada)


class TestRenderizacaoDoBloco(unittest.TestCase):
    def test_c1_nao_renderiza_nada(self):
        retorno, st_falso = _renderizar(SO_C1)
        self.assertIsNone(retorno)
        self.assertEqual(st_falso.markdowns, [])
        self.assertEqual(st_falso.containers, [])

    def test_texto_do_bloco_para_c3(self):
        retorno, st_falso = _renderizar(SO_C3)
        self.assertIsNotNone(retorno)
        texto = st_falso.texto
        self.assertIn("Referência temporal anterior ao ciclo analisado", texto)
        self.assertIn("Ciclo anterior: **C2**", texto)
        self.assertIn(
            "Período anual imediatamente anterior à Data-Base do C3:", texto
        )
        self.assertIn("01/02/2025 a 31/01/2026 — 12 meses", texto)
        self.assertIn("Último dia anterior à Data-Base do C3:", texto)
        self.assertIn("31/01/2026", texto)
        self.assertNotIn("início dos efeitos", texto)

    def test_bloco_e_somente_leitura(self):
        _, st_falso = _renderizar(SO_C3)
        self.assertEqual(st_falso.containers, [{"border": True}])
        # O stub levanta AssertionError em qualquer st.<widget>; chegar aqui
        # significa que apenas container/markdown foram usados.

    def test_nao_atribui_o_periodo_aos_efeitos_do_ciclo_anterior(self):
        _, st_falso = _renderizar(SO_C3)
        texto = st_falso.texto.lower()
        for proibido in (
            "efeitos financeiros do ciclo anterior",
            "período de efeitos financeiros",
            "periodo de efeitos financeiros",
            "efeitos do c2",
        ):
            self.assertNotIn(proibido, texto)

    def test_sem_nota_explicativa_adicional(self):
        _, st_falso = _renderizar(SO_C3)
        texto = st_falso.texto.lower()
        for proibido in ("observação", "observacao", "nota:", "atenção", "atencao"):
            self.assertNotIn(proibido, texto)

    def test_bloco_do_multiciclo_ancora_no_primeiro_ciclo(self):
        _, st_falso = _renderizar(C2_C3)
        texto = st_falso.texto
        self.assertIn("Ciclo anterior: **C1**", texto)
        self.assertIn("01/02/2024 a 31/01/2025 — 12 meses", texto)
        self.assertNotIn("C3", texto)


class TestXlsxIntacto(unittest.TestCase):
    """O conteudo do XLSX nao depende do nome do download nem do bloco novo."""

    def test_conteudo_da_coleta_independe_do_nome_do_arquivo(self):
        antes = gerar_coleta_oficial_preenchida(XLSX_C2_C3)
        nome_download_coleta(XLSX_C2_C3)
        self.assertIsNotNone(referencia_temporal_anterior(XLSX_C2_C3))
        depois = gerar_coleta_oficial_preenchida(XLSX_C2_C3)
        self.assertEqual(_partes_xlsx(antes), _partes_xlsx(depois))

    def test_abas_preservadas_na_ordem_oficial(self):
        wb = load_workbook(io.BytesIO(gerar_coleta_oficial_preenchida(XLSX_C2_C3)))
        self.assertEqual(wb.sheetnames, ABAS_COLETA_OFICIAL)


class TestFiacaoNasPaginas(unittest.TestCase):
    SIMPLES = (ROOT / "pages" / "01_Calculo_Simples.py").read_text(encoding="utf-8")
    MULTIPLO = (ROOT / "pages" / "02_Calculo_Represados.py").read_text(encoding="utf-8")
    UPLOAD = (ROOT / "pages" / "03_Valor_Global.py").read_text(encoding="utf-8")
    INICIO = (ROOT / "pages" / "00_Calculadora_Reajustes.py").read_text(encoding="utf-8")
    CENTRAL = (ROOT / "pages" / "14_Central_Modelos_Ferramentas.py").read_text(encoding="utf-8")

    def test_quatro_downloads_usam_o_helper_canonico(self):
        for fonte in (self.INICIO, self.SIMPLES, self.MULTIPLO, self.CENTRAL):
            self.assertIn("nome_download_coleta(", fonte)
            self.assertNotIn("file_name=NOME_DOWNLOAD_COLETA", fonte)
        # Simples calcula o nome em variavel (padrao de _bytes_coleta_estavel).
        self.assertIn("file_name=_nome_coleta_estavel", self.SIMPLES)
        self.assertIn("file_name=nome_download_coleta(", self.MULTIPLO)
        self.assertIn("file_name=nome_download_coleta(", self.CENTRAL)

    def test_download_do_modelo_em_branco_usa_o_nome_datado(self):
        self.assertIn("file_name=nome_download_coleta()", self.INICIO)

    def test_bloco_renderizado_apos_a_analise_e_antes_do_download(self):
        for fonte in (self.SIMPLES, self.MULTIPLO):
            self.assertIn("render_referencia_temporal_anterior(", fonte)
            pos_dados = fonte.index("st.session_state['dados_admissibilidade'] = {")
            pos_bloco = fonte.index("render_referencia_temporal_anterior(")
            pos_download = fonte.index("nome_download_coleta(")
            self.assertLess(pos_dados, pos_bloco)
            self.assertLess(pos_bloco, pos_download)

    def test_pagina_de_upload_nao_exibe_o_bloco(self):
        self.assertNotIn("render_referencia_temporal_anterior", self.UPLOAD)
        self.assertNotIn("render_referencia_temporal_anterior", self.INICIO)


if __name__ == "__main__":
    unittest.main()
