# -*- coding: utf-8 -*-
"""RESULTADOS-CONTRACT-1 — contrato nominal sem mudanca economica/visual."""
from __future__ import annotations

import json
import zipfile
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.workbook.defined_name import DefinedName

from _leitor_masterfile_v10 import (
    _NOMES_AUDITORIA_XLS,
    _NOMES_RESULTADOS_PUBLICADOS_PR1,
    _NOMES_RESULTADOS_XLS,
    _ler_referencias_vta,
)

ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "templates" / "COLETA_REAJUSTE_OFICIAL.xlsx"
BASELINE = ROOT / "tests" / "baseline_resultados" / "01_financeiro_normal.json"

NOMES_PUBLICADOS = {
    "EXECUTADO_APURADO": "RESULTADOS!$B$83",
    "AJUSTES_DEVIDOS": "RESULTADOS!$B$84",
    "CONFERENCIA_FORMACAO_VTA": "RESULTADOS!$B$87",
    "PC_TOTAL_CADASTRADO": "MEMORIA_RESULTADOS!$T$33",
    "PC_TOTAL_ATE_CORTE": "MEMORIA_RESULTADOS!$T$34",
    "PC_TOTAL_COM_EFEITO": "MEMORIA_RESULTADOS!$T$36",
    "PC_TOTAL_SEM_EFEITO": "MEMORIA_RESULTADOS!$T$37",
}
NOMES_AUDITORIA = {
    "AUDITORIA_SITUACAO_ATUAL_CONTRATO": "MEMORIA_RESULTADOS!$W$50",
    "AUDITORIA_ULTIMA_REFERENCIA_ABERTURA": "MEMORIA_RESULTADOS!$W$48",
    "AUDITORIA_COMPARATIVO_INTEGRAL": "comparativo_VTA!$B$208",
    "AUDITORIA_DIFERENCA_REFERENCIAS": "MEMORIA_RESULTADOS!$W$51",
    "AUDITORIA_SITUACAO_ATUAL_STATUS": "RESULTADOS!$H$10",
    "AUDITORIA_ABERTURA_STATUS": "RESULTADOS!$H$11",
    "AUDITORIA_CONFERENCIA_STATUS": "MEMORIA_RESULTADOS!$W$52",
}
NOMES_PREEXISTENTES = {
    "AJUSTE_MANUAL_VTA": "MEMORIA_RESULTADOS!$B$24",
    "EXECUCAO_ATUALIZADA_CICLO": "RESULTADOS!$B$36",
    "JUSTIFICATIVA_RETRO": "MEMORIA_RESULTADOS!$D$5",
    "METODO_RETROATIVO": "MEMORIA_RESULTADOS!$B$4",
    "OPCOES_APLICAR_MANUAL": "RESULTADOS!$J$2:$J$3",
    "OPCOES_SIM_NAO": "parametros!$T$2:$T$3",
    "OPCOES_SIM_NAO_NA": "parametros!$T$2:$T$4",
    "QTD_REM_OFICIAL": "MEMORIA_RESULTADOS!$B$35",
    "REM_ATUALIZADO_OFICIAL": "MEMORIA_RESULTADOS!$D$35",
    "REM_BASE_OFICIAL": "MEMORIA_RESULTADOS!$C$35",
    "RETRO_FIN": "MEMORIA_RESULTADOS!$B$15",
    "RETRO_ITENS": "MEMORIA_RESULTADOS!$D$15",
    "RETRO_OFICIAL": "MEMORIA_RESULTADOS!$B$16",
    "RETRO_PC": "MEMORIA_RESULTADOS!$C$15",
    "SALDO_REMANESCENTE_ATUAL": "RESULTADOS!$B$38",
    "STATUS_RESULTADOS": "RESULTADOS!$B$3",
    "TOLERANCIA_DIVERGENCIA": "MEMORIA_RESULTADOS!$D$4",
    "VALOR_MANUAL_RETRO": "MEMORIA_RESULTADOS!$B$5",
    "VTA_ATUALIZACAO_CHEIA": "RESULTADOS!$B$12",
    "VTA_CALCULADO": "MEMORIA_RESULTADOS!$B$23",
    "VTA_FINAL": "MEMORIA_RESULTADOS!$B$26",
    "VTA_MANUAL_OFICIAL": "MEMORIA_RESULTADOS!$B$25",
}


def _names(wb) -> dict[str, str]:
    return {name: definition.value for name, definition in wb.defined_names.items()}


def _add_name(wb, name: str, target: str) -> None:
    wb.defined_names.add(DefinedName(name, attr_text=target))


def test_os_14_nomes_existem_nos_destinos_aprovados():
    wb = load_workbook(TEMPLATE, data_only=False)
    try:
        names = _names(wb)
        assert {name: names.get(name) for name in NOMES_PUBLICADOS} == NOMES_PUBLICADOS
        assert {name: names.get(name) for name in NOMES_AUDITORIA} == NOMES_AUDITORIA
        assert tuple(NOMES_PUBLICADOS) == _NOMES_RESULTADOS_PUBLICADOS_PR1
        assert set(NOMES_AUDITORIA) == set(_NOMES_AUDITORIA_XLS.values())
    finally:
        wb.close()


def test_names_antigos_preservados_t35_sem_nome_e_zero_vinculo_externo():
    wb = load_workbook(TEMPLATE, data_only=False, keep_links=True)
    try:
        names = _names(wb)
        assert {name: names.get(name) for name in NOMES_PREEXISTENTES} == \
            NOMES_PREEXISTENTES
        assert len(names) == len(NOMES_PREEXISTENTES) + 14
        assert "MEMORIA_RESULTADOS!$T$35" not in names.values()
        assert not wb._external_links
    finally:
        wb.close()
    with zipfile.ZipFile(TEMPLATE) as archive:
        assert not [name for name in archive.namelist() if "externalLink" in name]
        assert not any(
            b"[1]MEMORIA" in archive.read(name)
            for name in archive.namelist()
            if name.endswith((".xml", ".rels"))
        )


def test_vta_final_permanece_oficial_e_auditoria_nao_o_alimenta():
    wb = load_workbook(TEMPLATE, data_only=False)
    try:
        names = _names(wb)
        assert names["VTA_FINAL"] == "MEMORIA_RESULTADOS!$B$26"
        assert names["VTA_ATUALIZACAO_CHEIA"] == "RESULTADOS!$B$12"
        assert "VTA_ATUALIZACAO_CHEIA" not in _NOMES_RESULTADOS_XLS
        assert set(NOMES_AUDITORIA).isdisjoint(_NOMES_RESULTADOS_XLS)
        assert all(not name.startswith("AUDITORIA_") for name in _NOMES_RESULTADOS_XLS)
    finally:
        wb.close()


def _workbook_referencias(*, with_names: bool) -> Workbook:
    wb = Workbook()
    result = wb.active
    result.title = "RESULTADOS"
    result["B10"], result["B11"], result["B12"], result["B13"] = (
        -10.0, -11.0, -12.0, -13.0
    )
    result["H10"], result["H11"], result["H13"] = (
        "SENTINELA 10", "SENTINELA 11", "SENTINELA 13"
    )
    memory = wb.create_sheet("MEMORIA_RESULTADOS")
    for address, value in {
        "A1": 100.0, "A2": 200.0, "A3": -100.0, "A4": "CONFERIDO",
        "A5": "ATUAL OK", "A6": "ABERTURA OK",
    }.items():
        memory[address] = value
    wb.create_sheet("comparativo_VTA")["A1"] = 300.0
    wb.create_sheet("CONTROLE")["B2"] = "C1"
    wb.create_sheet("CICLO_EM_EXECUCAO")["A9"] = 1.0
    if with_names:
        targets = {
            "AUDITORIA_SITUACAO_ATUAL_CONTRATO": "MEMORIA_RESULTADOS!$A$1",
            "AUDITORIA_ULTIMA_REFERENCIA_ABERTURA": "MEMORIA_RESULTADOS!$A$2",
            "AUDITORIA_COMPARATIVO_INTEGRAL": "comparativo_VTA!$A$1",
            "AUDITORIA_DIFERENCA_REFERENCIAS": "MEMORIA_RESULTADOS!$A$3",
            "AUDITORIA_SITUACAO_ATUAL_STATUS": "MEMORIA_RESULTADOS!$A$5",
            "AUDITORIA_ABERTURA_STATUS": "MEMORIA_RESULTADOS!$A$6",
            "AUDITORIA_CONFERENCIA_STATUS": "MEMORIA_RESULTADOS!$A$4",
        }
        for name, target in targets.items():
            _add_name(wb, name, target)
    return wb


def test_leitor_novo_segue_names_e_ignora_coordenadas_sentinela():
    result = _ler_referencias_vta(_workbook_referencias(with_names=True))
    assert result["origem_leitura"] == "defined_names"
    assert result["forma1_posicao_atual"] == 100.0
    assert result["forma2_ultima_abertura"] == 200.0
    assert result["forma3_integral_reajustado"] == 300.0
    assert result["reconciliacao_valor"] == -100.0
    assert result["forma1_situacao"] == "ATUAL OK"
    assert result["forma2_situacao"] == "ABERTURA OK"
    assert result["reconciliacao_status"] == "CONFERIDO"


def test_leitor_legado_sem_names_preserva_coordenadas_antigas():
    result = _ler_referencias_vta(_workbook_referencias(with_names=False))
    assert result["origem_leitura"] == "legacy_coordinates"
    assert result["forma1_posicao_atual"] == -10.0
    assert result["forma2_ultima_abertura"] == -11.0
    assert result["forma3_integral_reajustado"] == -12.0
    assert result["reconciliacao_valor"] == -13.0
    assert result["reconciliacao_status"] == "SENTINELA 13"


def test_as_28_celulas_e_b83_b84_b87_nao_mudaram():
    baseline = json.loads(BASELINE.read_text(encoding="utf-8"))["contrato_xls"]
    expected = baseline["formulas"]
    runtime = {"A1", "B3", "B10", "B11", "B12", "B13", "H10", "H11", "H13"}
    internal = {
        f"{column}{row}"
        for row in range(43, 51)
        for column in ("C", "D", "G")
        if not (column == "D" and row >= 46)
    }
    assert len(runtime | internal) == 28
    wb = load_workbook(TEMPLATE, data_only=False)
    try:
        result = wb["RESULTADOS"]
        for address in runtime | internal | {"B83", "B84", "B87"}:
            assert result[address].value == expected.get(address)
        assert wb.sheetnames[-1] == "RESULTADOS"
        assert wb["MEMORIA_RESULTADOS"].sheet_state == "hidden"
    finally:
        wb.close()


def test_c43_g50_preserva_validacoes_e_coordenadas():
    wb = load_workbook(TEMPLATE, data_only=False)
    try:
        result = wb["RESULTADOS"]
        assert all(result.cell(row, col).value is None
                   for row in range(43, 51) for col in range(3, 8))
        validations = {
            str(item.sqref): (item.type, item.formula1)
            for item in result.data_validations.dataValidation
        }
        assert validations.get("C46:C50") == ("decimal", "0")
        assert validations.get("G43:G50") == (
            "list", "OPCOES_APLICAR_MANUAL"
        )
    finally:
        wb.close()
