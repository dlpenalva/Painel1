from __future__ import annotations

import io
import os
import tempfile
from datetime import date
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from _ciclo_em_execucao import (
    ABA_CICLO_EM_EXECUCAO,
    CELULA_DATA_POSICAO,
    CELULA_TOTAL,
    COLUNAS_VISIVEIS,
    LINHA_CABECALHO,
    MARCADOR_LAYOUT,
    PRIMEIRA_LINHA_ITEM,
    ULTIMA_LINHA_ITEM,
    VERSAO_LAYOUT_ITEMIZADO,
    calcular_posicao_ciclo_por_data,
    garantir_aba_ciclo_em_execucao,
    identificar_layout_ciclo_em_execucao,
    ler_ciclo_em_execucao,
)
from _coleta_oficial import (
    ABAS_COLETA_OFICIAL,
    TEMPLATE_COLETA_OFICIAL,
    obter_coleta_oficial_bytes,
)
from _leitor_masterfile_v10 import ler_masterfile_v10


T0 = date(2026, 5, 1)
FIM = date(2027, 4, 30)
D = date(2026, 7, 31)


def _calcular(
    *,
    inicio=80,
    atual=65,
    movimentos=(),
    vu=100,
    data_posicao=D,
    novo_item=False,
):
    return calcular_posicao_ciclo_por_data(
        ciclo="C3",
        data_inicio=T0,
        data_fim=FIM,
        data_posicao=data_posicao,
        itens=[{
            "item": "ITEM-1" if not novo_item else "N001",
            "remanescente_inicio": inicio,
            "remanescente_atual": atual,
            "vu_atualizado": vu,
            "novo_item": novo_item,
        }],
        movimentos=movimentos,
    )


def _mov(delta, quando, item="ITEM-1", ident=None):
    return {
        "item": item,
        "data_efeito": quando,
        "delta_quantidade": delta,
        "id_evento": ident,
    }


def _linha(resultado):
    return resultado["itens"][0]


def test_caso_1_sem_aditivo_consumo_15():
    resultado = _calcular()
    assert resultado["valido"] is True
    assert _linha(resultado)["quantidade_consumida"] == 15


def test_caso_2_acrescimo_durante_ciclo_consumo_25():
    resultado = _calcular(movimentos=[_mov(10, date(2026, 6, 15))])
    assert _linha(resultado)["alteracoes_liquidas_periodo"] == 10
    assert _linha(resultado)["quantidade_consumida"] == 25


def test_caso_3_supressao_durante_ciclo_consumo_10():
    resultado = _calcular(movimentos=[_mov(-5, date(2026, 6, 15))])
    assert _linha(resultado)["alteracoes_liquidas_periodo"] == -5
    assert _linha(resultado)["quantidade_consumida"] == 10


def test_caso_4_aditivo_ate_t0_nao_e_reaplicado():
    resultado = _calcular(movimentos=[
        _mov(30, date(2026, 4, 30)),
        _mov(20, T0),
    ])
    assert _linha(resultado)["alteracoes_liquidas_periodo"] == 0
    assert _linha(resultado)["quantidade_consumida"] == 15


def test_caso_5_aditivo_posterior_a_d_e_ignorado():
    resultado = _calcular(movimentos=[_mov(40, date(2026, 8, 1))])
    assert _linha(resultado)["alteracoes_liquidas_periodo"] == 0
    assert _linha(resultado)["quantidade_consumida"] == 15


def test_caso_6_novo_item_nasce_no_ciclo_com_abertura_zero():
    resultado = _calcular(
        inicio=0,
        atual=6,
        novo_item=True,
        movimentos=[_mov(10, date(2026, 6, 1), item="N001")],
    )
    linha = _linha(resultado)
    assert linha["remanescente_inicio"] == 0
    assert linha["quantidade_consumida"] == 4
    assert linha["data_nascimento"] == date(2026, 6, 1)


def test_item_novo_posterior_a_d_nao_aparece():
    resultado = _calcular(
        inicio=0,
        atual=0,
        novo_item=True,
        movimentos=[_mov(10, date(2026, 8, 1), item="N001")],
    )
    assert resultado["itens"] == []
    assert resultado["valido"] is False


def test_caso_7_supressao_integral_sem_consumo_permanece_rastreavel():
    resultado = _calcular(
        inicio=10,
        atual=0,
        movimentos=[_mov(-10, date(2026, 6, 1))],
    )
    linha = _linha(resultado)
    assert linha["quantidade_consumida"] == 0
    assert linha["suprimido_integralmente"] is True
    assert linha["check_fisico"] == 0


def test_caso_8_vazio_e_incompleto_sem_assumir_zero():
    resultado = _calcular(inicio=10, atual=None)
    linha = _linha(resultado)
    assert linha["remanescente_atual"] is None
    assert linha["quantidade_consumida"] is None
    assert linha["valor_remanescente_atualizado"] is None
    assert resultado["completo"] is False
    assert resultado["valido"] is False


def test_caso_9_zero_confirmado_calcula_consumo_integral():
    resultado = _calcular(inicio=10, atual=0)
    linha = _linha(resultado)
    assert linha["remanescente_atual"] == 0
    assert linha["quantidade_consumida"] == 10
    assert linha["valor_remanescente_atualizado"] == 0
    assert resultado["valido"] is True


def test_caso_10_arquivo_sem_aba_mantem_compatibilidade():
    wb = Workbook()
    antes = list(wb.sheetnames)
    layout = identificar_layout_ciclo_em_execucao(wb)
    leitura = ler_ciclo_em_execucao(wb)
    assert layout == {"tipo": "ausente", "versao": None}
    assert leitura["disponivel"] is False
    assert leitura["utilizado"] is False
    assert wb.sheetnames == antes


def test_caso_11_aba_agregada_legada_nao_e_interpretada_nem_alterada():
    wb = Workbook()
    ws = wb.active
    ws.title = ABA_CICLO_EM_EXECUCAO
    ws["A1"] = "CICLO ATUAL"
    ws["B5"] = "07/2026"
    antes = tuple((c.coordinate, c.value) for row in ws.iter_rows() for c in row)
    garantia = garantir_aba_ciclo_em_execucao(wb)
    leitura = ler_ciclo_em_execucao(wb)
    depois = tuple((c.coordinate, c.value) for row in ws.iter_rows() for c in row)
    assert garantia["legado_preservado"] is True
    assert leitura["legado_ignorado"] is True
    assert leitura["disponivel"] is False
    assert antes == depois


def test_caso_12_reconciliacao_monetaria_fecha_sem_criar_valor():
    resultado = _calcular(
        movimentos=[_mov(10, date(2026, 6, 15))],
        vu=123.45,
    )
    linha = _linha(resultado)
    assert linha["valor_consumido"] + linha["valor_remanescente_atualizado"] == (
        linha["valor_base_fisica_atualizada"]
    )
    assert resultado["total_valor_consumido"] + resultado["total_valor_remanescente"] == (
        resultado["total_base_fisica_atualizada"]
    )


def test_caso_13_remanescente_maior_que_disponivel_e_bloqueante():
    resultado = _calcular(inicio=80, atual=81)
    assert resultado["valido"] is False
    assert any("REMANESCENTE_ATUAL_SUPERA_DISPONIVEL" in e for e in resultado["erros"])
    assert any("CONSUMO_NEGATIVO" in e for e in resultado["erros"])


def test_caso_14_data_fora_do_ciclo_invalida_posicao():
    resultado = _calcular(data_posicao=date(2027, 5, 1))
    assert resultado["valido"] is False
    assert "DATA_DA_POSICAO_FORA_DO_CICLO" in resultado["erros"]


def test_evento_quantitativo_duplicado_e_bloqueado_sem_dupla_soma():
    movimentos = [
        _mov(10, date(2026, 6, 1), ident="AD-1"),
        _mov(10, date(2026, 6, 1), ident="AD-1"),
    ]
    resultado = _calcular(movimentos=movimentos)
    assert resultado["valido"] is False
    assert _linha(resultado)["alteracoes_liquidas_periodo"] == 10
    assert any("EVENTO_QUANTITATIVO_DUPLICADO" in e for e in resultado["erros"])


@pytest.fixture(scope="module")
def coleta_runtime():
    conteudo = obter_coleta_oficial_bytes()
    return conteudo, load_workbook(io.BytesIO(conteudo), data_only=False)


def test_aba_runtime_tem_sete_colunas_visiveis_e_marcador_oculto(coleta_runtime):
    _, wb = coleta_runtime
    ws = wb[ABA_CICLO_EM_EXECUCAO]
    assert wb.sheetnames == ABAS_COLETA_OFICIAL
    assert tuple(ws.cell(LINHA_CABECALHO, c).value for c in range(1, 8)) == COLUNAS_VISIVEIS
    assert ws["H1"].value == MARCADOR_LAYOUT
    assert ws["H2"].value == VERSAO_LAYOUT_ITEMIZADO
    assert all(ws.column_dimensions[col].hidden for col in "HIJKLMNO")
    assert MARCADOR_LAYOUT in wb.defined_names
    assert "DESCRI" not in " ".join(COLUNAS_VISIVEIS).upper()


def test_layout_visual_cores_por_natureza_e_total_executivo(coleta_runtime):
    _, wb = coleta_runtime
    ws = wb[ABA_CICLO_EM_EXECUCAO]
    assert ws["B12"].fill.fgColor.rgb == ws["C12"].fill.fgColor.rgb == "FF0F5B50"
    assert ws["B13"].fill.fgColor.rgb == "FFE8F5F1"
    assert ws["C13"].fill.fgColor.rgb == "FFC6EFCE"
    assert ws["D12"].fill.fgColor.rgb == "FFE67E22"
    assert ws["D13"].fill.fgColor.rgb == "FFFCE4D6"
    assert all(ws[f"{c}12"].fill.fgColor.rgb == "FF1F4E78" for c in "EFG")
    assert all(ws[f"{c}13"].fill.fgColor.rgb == "FFDDEBF7" for c in "EFG")
    assert ws["A7"].value == (
        "VALOR TOTAL ATUALIZADO REMANESCENTE NA DATA DA POSIÇÃO (AUTO)"
    )
    assert 'YEAR($D$5)' in ws["A8"].value
    assert 'TEXT(' not in ws["A8"].value
    assert ws[CELULA_TOTAL].number_format == "R$ #,##0.00"
    assert "A7:G7" in {str(r) for r in ws.merged_cells.ranges}
    assert "A9:G10" in {str(r) for r in ws.merged_cells.ranges}


def test_entradas_reais_protecao_formatos_e_validacoes(coleta_runtime):
    _, wb = coleta_runtime
    ws = wb[ABA_CICLO_EM_EXECUCAO]
    assert ws.protection.sheet is True
    assert ws[CELULA_DATA_POSICAO].value is None
    assert ws[CELULA_DATA_POSICAO].protection.locked is False
    assert ws[CELULA_DATA_POSICAO].number_format == "dd/mm/yyyy"
    for linha in (PRIMEIRA_LINHA_ITEM, ULTIMA_LINHA_ITEM):
        assert ws.cell(linha, 3).value is None  # vazio real, nao formula ""
        assert ws.cell(linha, 3).protection.locked is False
        assert ws.cell(linha, 2).protection.locked is True
        assert ws.cell(linha, 4).protection.locked is True
        assert ws.cell(linha, 3).number_format == "#,##0.00"
    validacoes = ws.data_validations.dataValidation
    assert len(validacoes) == 2
    assert any(str(v.sqref) == "D5" and v.type == "custom" for v in validacoes)
    assert any(
        str(v.sqref) == f"C{PRIMEIRA_LINHA_ITEM}:C{ULTIMA_LINHA_ITEM}"
        and "C13<=ROUND(B13+I13,2)" in v.formula1
        for v in validacoes
    )


def test_formulas_aplicam_janela_exata_e_nao_referenciam_vta(coleta_runtime):
    _, wb = coleta_runtime
    ws = wb[ABA_CICLO_EM_EXECUCAO]
    formula_delta = ws["I13"].value
    formula_consumo = ws["D13"].value
    formula_valor_consumido = ws["F13"].value
    formula_remanescente = ws["G13"].value
    # A janela do periodo comeca no dia SEGUINTE a abertura: o aditivo datado NO
    # dia da abertura pertence a fotografia de abertura (coluna B), uma unica vez.
    assert '">="&(INT($F$3)+1)' in formula_delta
    assert '"<="&$D$5' in formula_delta
    assert "B13+I13-C13" in formula_consumo
    assert "D13*E13" in formula_valor_consumido
    assert "C13*E13" in formula_remanescente
    formulas = [
        c.value
        for row in ws.iter_rows()
        for c in row
        if isinstance(c.value, str) and c.value.startswith("=")
    ]
    assert formulas
    assert not any("#REF!" in f.upper() for f in formulas)
    assert not any("VTA" in f.upper() or "RESULTADOS!" in f.upper() for f in formulas)


def test_novo_item_usa_data_de_nascimento_e_suprimido_fica_visivel(coleta_runtime):
    _, wb = coleta_runtime
    ws = wb[ABA_CICLO_EM_EXECUCAO]
    assert "COUNTIFS(aditivos!$A$2:$A$200" in ws["O13"].value
    assert '"<="&$D$5' in ws["O13"].value
    assert "$O13>0" in ws["A13"].value
    assert "posicao_contratual!$Z2" in ws["B13"].value
    # Fronteira por DIA: aditivo com efeito ate a abertura compoe a coluna B.
    assert '"<"&(INT($F$3)+1)' in ws["B13"].value
    regras = [str(regra.formula) for regras in ws.conditional_formatting._cf_rules.values() for regra in regras]
    assert any("ROUND($B13+$I13,2)=0" in regra for regra in regras)


def test_aba_nova_vazia_e_considerada_nao_utilizada(coleta_runtime):
    conteudo, _ = coleta_runtime
    wb_valores = load_workbook(io.BytesIO(conteudo), data_only=True)
    leitura = ler_ciclo_em_execucao(wb_valores)
    assert leitura["disponivel"] is True
    assert leitura["utilizado"] is False
    assert leitura["layout_version"] == VERSAO_LAYOUT_ITEMIZADO


def test_leitor_valida_posicao_e_divergencia_sem_substituir_declaracao():
    wb = Workbook()
    wb.active.title = "posicao_referencia"
    garantir_aba_ciclo_em_execucao(wb)
    ws = wb[ABA_CICLO_EM_EXECUCAO]
    ws["C3"] = "C3"
    ws["F3"] = T0
    ws["H3"] = FIM
    ws["D5"] = D
    ws["A13"] = "ITEM-1"
    ws["B13"] = 80
    ws["C13"] = 65
    ws["D13"] = 15
    ws["E13"] = 100
    ws["F13"] = 1500
    ws["G13"] = 6500
    ws["I13"] = 0
    ws["J13"] = 0
    ws["K13"] = "OK"
    leitura = ler_ciclo_em_execucao(
        wb,
        itens_consumidos={
            "itens": [{"item": "ITEM-1", "consumos": {"C3": {"qtd": 14}}}]
        },
    )
    assert leitura["valido"] is True
    assert leitura["itens"][0]["remanescente_atual"] == 65
    assert leitura["itens"][0]["quantidade_consumida"] == 15
    assert leitura["diagnosticos"][0]["divergencia"] == 1
    assert leitura["diagnosticos"][0]["natureza"] == "DIAGNOSTICO_NAO_SUBSTITUTIVO"


def test_modelo_oficial_anterior_sem_aba_segue_aceito_pelo_leitor():
    resultado = ler_masterfile_v10(
        TEMPLATE_COLETA_OFICIAL.read_bytes(),
        exigir_modelo_oficial=True,
    )
    assert "CICLO_EM_EXECUCAO" not in resultado.get("abas_ausentes", [])
    assert resultado["ciclo_em_execucao"]["layout_tipo"] == "ausente"


def test_template_binario_permanece_sem_nova_aba_e_runtime_a_adiciona():
    template = load_workbook(TEMPLATE_COLETA_OFICIAL, data_only=False, read_only=True)
    assert ABA_CICLO_EM_EXECUCAO not in template.sheetnames
    template.close()
    runtime = load_workbook(io.BytesIO(obter_coleta_oficial_bytes()), data_only=False)
    assert ABA_CICLO_EM_EXECUCAO in runtime.sheetnames
    runtime.close()


@pytest.mark.skipif(
    os.environ.get("RUN_EXCEL_INTEGRATION") != "1",
    reason="defina RUN_EXCEL_INTEGRATION=1 para Excel COM",
)
def test_excel_com_recalcula_sem_reparo_ref_ou_erro():
    import pythoncom
    import pywintypes
    import win32com.client

    wb = Workbook()
    wb.active.title = "posicao_referencia"
    ctrl = wb.create_sheet("CONTROLE")
    ctrl["B2"] = "C3"
    par = wb.create_sheet("parametros")
    par["B5"], par["C5"], par["D5"] = "C3", T0, FIM
    rem = wb.create_sheet("itens_Remanesc")
    rem["A2"], rem["C2"], rem["I2"] = "ITEM-1", 100, 80
    pos = wb.create_sheet("posicao_contratual")
    pos["R2"], pos["Z2"] = 80, False
    hv = wb.create_sheet("historico_VU")
    hv["F2"] = 110
    ad = wb.create_sheet("aditivos")
    ad["A2"], ad["B2"], ad["L2"] = "ITEM-1", date(2026, 6, 15), 10
    cons = wb.create_sheet("itens_Consumidos")
    cons["A2"], cons["K2"] = "ITEM-1", 25
    garantir_aba_ciclo_em_execucao(wb)
    ce = wb[ABA_CICLO_EM_EXECUCAO]
    ce["D5"], ce["C13"] = D, 65
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    with tempfile.TemporaryDirectory() as pasta:
        caminho = Path(pasta) / "ciclo_em_execucao_29c1.xlsx"
        wb.save(caminho)
        pythoncom.CoInitialize()
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        try:
            livro = excel.Workbooks.Open(str(caminho), UpdateLinks=0, ReadOnly=False)
            excel.CalculateFullRebuild()
            planilha = livro.Worksheets(ABA_CICLO_EM_EXECUCAO)
            valores = {
                cel: planilha.Range(cel).Value
                for cel in (
                    "C3", "F3", "H3", "A13", "B13", "C13", "D13", "E13",
                    "F13", "G13", "I13", "J13", "K13", "O13", "A9",
                )
            }
            erros = []
            try:
                erros.append(planilha.UsedRange.SpecialCells(-4123, 16).Address)
            except pywintypes.com_error:
                pass
            livro.Save()
            livro.Close(SaveChanges=False)
        finally:
            excel.Quit()
            pythoncom.CoUninitialize()
        assert valores.pop("F3").date() == T0
        assert valores.pop("H3").date() == FIM
        assert valores == {
            "C3": "C3",
            "A13": "ITEM-1",
            "B13": 80.0,
            "C13": 65.0,
            "D13": 25.0,
            "E13": 110.0,
            "F13": 2750.0,
            "G13": 7150.0,
            "I13": 10.0,
            "O13": 1.0,
            "A9": 7150.0,
            "J13": 0.0,
            "K13": "OK",
        }
        assert erros == []
        assert list(Path(pasta).glob("error*.xml")) == []
        reaberto = load_workbook(caminho, data_only=False)
        formulas = [
            c.value
            for ws in reaberto.worksheets
            for row in ws.iter_rows()
            for c in row
            if isinstance(c.value, str) and c.value.startswith("=")
        ]
        assert not any("#REF!" in f.upper() for f in formulas)
