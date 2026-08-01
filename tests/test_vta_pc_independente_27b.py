"""Etapa 27B — VTA por PCs independente do pagamento (coluna G).

Regra funcional definitiva:
  - G=Sim, G=Nao e G vazio produzem a MESMA participacao economica do PC
    no VTA (estrutura valida + ciclo anterior a linha de corte);
  - G classifica somente o retroativo: Sim -> reconhecido; Nao -> potencial;
    vazio -> potencial com pagamento nao confirmado;
  - G vazio nao e erro estrutural, nao rebaixa sozinho o selo do VTA e nao
    trava sozinho a consolidacao;
  - nenhum fallback fabrica VTA a partir somente de PCs pagos (fail-closed).

As expectativas economicas derivam das ENTRADAS (600 x 1,10 = 660;
retroativo 60), nunca do objeto produzido pelo motor.
"""
from __future__ import annotations

import io
import os
from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook

import _motor_composicao_vta as motor_vta
from _coleta_oficial import gerar_coleta_oficial_preenchida
from _coleta_reajuste_documentos import processar_coleta_oficial_runtime
from _leitor_masterfile_v10 import ler_masterfile_v10

VARIANTES_G = ("Sim", "Nao", None)  # None = celula vazia


def _dados_corte_c2() -> dict:
    """Linha de corte C2; janela de PC valida em C1 (10% + 10%)."""
    return {
        "origem": "Teste 27B",
        "indice": "IST",
        "data_base_original": "01/01/2023",
        "data_corte": date(2025, 12, 31),
        "ciclos": [
            {"ciclo": "C1", "data_inicio": date(2024, 1, 1),
             "data_fim": date(2024, 12, 31), "data_pedido": date(2024, 1, 1),
             "financeiro_inicio": date(2024, 1, 1), "percentual": 0.10},
            {"ciclo": "C2", "data_inicio": date(2025, 1, 1),
             "data_fim": date(2025, 12, 31), "data_pedido": date(2025, 1, 1),
             "financeiro_inicio": date(2025, 1, 1), "percentual": 0.10},
        ],
    }


def _workbook_pc(pago, *, extra_pcs=(), sem_remanescente=False) -> bytes:
    base = gerar_coleta_oficial_preenchida(_dados_corte_c2())
    wb = load_workbook(io.BytesIO(base))
    wb["CONTROLE"]["B1"] = "Pedidos de Compras"
    ws = wb["itens_PC"]
    ws["A2"] = "PC-001"
    ws["B2"] = date(2024, 1, 15)   # C1, anterior a linha de corte (C2)
    ws["D2"] = 600.0
    if pago is not None:
        ws["G2"] = pago
    linha = 3
    for numero, data_pc, valor, g in extra_pcs:
        ws[f"A{linha}"] = numero
        ws[f"B{linha}"] = data_pc
        if valor is not None:
            ws[f"D{linha}"] = valor
        if g is not None:
            ws[f"G{linha}"] = g
        linha += 1
    if not sem_remanescente:
        wb["itens_Remanesc"]["A2"] = "ITEM-1"
        wb["itens_Remanesc"]["B2"] = 10.0
        wb["itens_Remanesc"]["C2"] = 100.0
        wb["itens_Remanesc"]["E2"] = 4.0   # QTD_REM_BASE_C1
        wb["itens_Remanesc"]["G2"] = 4.0   # QTD_REM_BASE_C2 (corte vigente)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _leituras_variantes():
    return {g: ler_masterfile_v10(_workbook_pc(g)) for g in VARIANTES_G}


def _pc_registro(leitura):
    itens = (leitura.get("itens_pc_v10") or {}).get("itens") or []
    return next(i for i in itens if i.get("numero_pc") == "PC-001")


# ---- 13. Nucleo economico: VTA e execucao identicos nas tres variantes ----

def _leitura_motor(campos_pagamento: dict) -> dict:
    """Leitura minima para o motor canonico; pagamento varia por parametro."""
    return {
        "controle": {"modo": "pc", "ciclo_vigente": "C2"},
        "parametros_v10": {"por_ciclo": {}},
        "posicao_contratual": {"itens": [{
            "ITEM": 1, "VU_ORIGINAL": 100.0,
            "QTD_REM_AJUSTADA_C0": 10, "QTD_REM_AJUSTADA_C1": 10,
            "QTD_REM_AJUSTADA_C2": 6, "QTD_REM_AJUSTADA_C3": 0,
            "QTD_REM_AJUSTADA_C4": 0,
        }]},
        "historico_vu": {"itens": [{
            "item": 1,
            "vu_ciclos": {"VU_C0": 100.0, "VU_C1": 110.0, "VU_C2": 121.0,
                          "VU_C3": 121.0, "VU_C4": 121.0},
        }]},
        "itens_pc_v10": {"itens": [dict(
            {"ciclo": "C1", "valor_pc": 600.0, "valor_atualizado": 660.0,
             "entra_no_calculo": "Sim"},
            **campos_pagamento,
        )]},
    }


def test_vta_e_execucao_identicos_para_g_sim_nao_vazio():
    # Motor canonico: variar TODOS os campos de pagamento do PC nao pode
    # alterar VTA, execucao nem remanescente. Expectativa derivada da
    # entrada: 660 (600 x 1,10) + 726 (6 x 121) = 1.386,00.
    variantes = {
        "Sim": {"pc_pago_a_contratada": "Sim", "elegivel_retroativo_pc": True,
                "status_pagamento_pc": "PAGO_INTEGRAL"},
        "Nao": {"pc_pago_a_contratada": "Nao", "elegivel_retroativo_pc": False,
                "status_pagamento_pc": "NAO_PAGO"},
        "vazio": {"pc_pago_a_contratada": None,
                  "elegivel_retroativo_pc": False,
                  "status_pagamento_pc": "NAO_CONFIRMADO"},
    }
    for rotulo, campos in variantes.items():
        comp = motor_vta.montar_composicao_vta(_leitura_motor(campos))
        assert comp["disponivel"] is True, rotulo
        parcela_c1 = next(
            p for p in comp["execucao_por_ciclo"] if p["ciclo"] == "C1"
        )
        assert parcela_c1["valor_atualizado"] == pytest.approx(660.0, abs=0.01)
        assert comp["total_execucao_atualizada"] == pytest.approx(
            660.0, abs=0.01
        ), rotulo
        assert comp["saldo_remanescente"]["valor_atualizado"] == pytest.approx(
            726.0, abs=0.01
        ), rotulo
        assert comp["vta_composicao"] == pytest.approx(1386.0, abs=0.01), rotulo


def test_retroativo_classificado_pela_coluna_g():
    leituras = _leituras_variantes()
    # G=Sim -> reconhecido 60 (660-600); potencial zero.
    reg = _pc_registro(leituras["Sim"])
    assert reg["retroativo_reconhecido_a_pagar"] == pytest.approx(60.0, abs=0.01)
    assert not reg["delta_potencial"]
    assert reg["elegivel_retroativo_pc"] is True
    assert reg["status_pagamento_pc"] == "PAGO_INTEGRAL"
    # G=Nao -> potencial 60; nada reconhecido.
    reg = _pc_registro(leituras["Nao"])
    assert not reg["retroativo_reconhecido_a_pagar"]
    assert reg["delta_potencial"] == pytest.approx(60.0, abs=0.01)
    assert reg["valor_atualizado_em_analise"] == pytest.approx(660.0, abs=0.01)
    assert reg["elegivel_retroativo_pc"] is False
    assert reg["status_pagamento_pc"] == "NAO_PAGO"
    # G vazio -> potencial 60 com pagamento nao confirmado (nunca zera).
    reg = _pc_registro(leituras[None])
    assert not reg["retroativo_reconhecido_a_pagar"]
    assert reg["delta_potencial"] == pytest.approx(60.0, abs=0.01)
    assert reg["valor_atualizado_em_analise"] == pytest.approx(660.0, abs=0.01)
    assert reg["elegivel_retroativo_pc"] is False
    assert reg["status_pagamento_pc"] == "NAO_CONFIRMADO"
    avisos = (leituras[None].get("itens_pc_v10") or {}).get("alertas") or []
    assert any("pagamento nao confirmado" in a for a in avisos)


def test_computa_vta_estrutural_independente_de_g():
    for g, leitura in _leituras_variantes().items():
        campos = _pc_registro(leitura)["campos_vta"]
        assert campos["computa_vta"] == "Sim", f"G={g!r}"
        assert campos["status_consolidacao"] == "COMPUTADO"
        assert campos["tipo_financeiro"] == "Execucao Atualizada"


# ---- 14. Status: G=Nao/vazio nao sao motivo de degradacao ----

def test_status_runtime_identico_nas_tres_variantes():
    diagnosticos = {}
    for g in VARIANTES_G:
        _, diag = processar_coleta_oficial_runtime(_workbook_pc(g))
        diagnosticos[g] = diag
    ref = diagnosticos["Sim"]
    for g in ("Nao", None):
        diag = diagnosticos[g]
        assert diag["status_base"] == ref["status_base"], f"G={g!r}"
        assert diag["pronto_para_consolidar"] == ref["pronto_para_consolidar"]
        assert diag["bloqueios_estruturais"] == ref["bloqueios_estruturais"]
        assert diag["inconsistencias"] == ref["inconsistencias"]


def test_g_texto_invalido_continua_gerando_alerta():
    leitura = ler_masterfile_v10(_workbook_pc("Talvez"))
    alertas = (leitura.get("itens_pc_v10") or {}).get("alertas") or []
    assert any("PC_PAGO_A_CONTRATADA invalido" in a for a in alertas)


def test_duplicidade_data_e_valor_continuam_diagnosticados():
    conteudo = _workbook_pc(
        "Sim",
        extra_pcs=(
            ("PC-001", date(2024, 2, 10), 100.0, "Sim"),   # duplicado
            ("PC-SEM-VALOR", date(2024, 3, 10), None, "Sim"),
        ),
    )
    leitura = ler_masterfile_v10(conteudo)
    alertas = (leitura.get("itens_pc_v10") or {}).get("alertas") or []
    assert any("NUMERO_PC duplicado" in a for a in alertas)
    assert any(
        "PC-SEM-VALOR" in a and "VALOR_PC nao informado" in a for a in alertas
    )
    itens = (leitura.get("itens_pc_v10") or {}).get("itens") or []
    sem_valor = next(i for i in itens if i.get("numero_pc") == "PC-SEM-VALOR")
    assert sem_valor["campos_vta"]["computa_vta"] == "Nao"
    assert sem_valor["campos_vta"]["status_consolidacao"] == "EM_ANALISE"


# ---- 15. Linha de corte preservada, independentemente de G ----

def test_pc_do_ciclo_vigente_excluido_do_vta_para_qualquer_g():
    # Leitor: PC no ciclo vigente (C2) nao computa VTA, para qualquer G.
    for g in VARIANTES_G:
        conteudo = _workbook_pc(
            "Sim",
            extra_pcs=(("PC-VIG", date(2025, 3, 10), 5000.0, g),),
        )
        leitura = ler_masterfile_v10(conteudo)
        reg = next(
            i for i in (leitura["itens_pc_v10"]["itens"])
            if i.get("numero_pc") == "PC-VIG"
        )
        assert reg["campos_vta"]["computa_vta"] == "Nao", f"G={g!r}"
        assert "corte" in reg["campos_vta"]["justificativa_vta"]
    # Motor: com e sem o PC vigente o VTA nao muda (5500 fora da execucao).
    base = motor_vta.montar_composicao_vta(_leitura_motor({}))
    com_vigente = _leitura_motor({})
    com_vigente["itens_pc_v10"]["itens"].append(
        {"ciclo": "C2", "valor_pc": 5000.0, "valor_atualizado": 5500.0,
         "entra_no_calculo": "Sim"}
    )
    comp = motor_vta.montar_composicao_vta(com_vigente)
    assert comp["vta_composicao"] == pytest.approx(
        base["vta_composicao"], abs=0.005
    )


def test_linha_de_corte_por_vigente_no_motor():
    # Motor canonico: abertura de C2 -> somente C0/C1 na execucao historica;
    # o PC de C2 e excluido. Nenhum campo de pagamento participa.
    leitura = {
        "controle": {"modo": "pc", "ciclo_vigente": "C2"},
        "parametros_v10": {"por_ciclo": {}},
        "posicao_contratual": {"itens": [{
            "ITEM": 1, "VU_ORIGINAL": 100.0,
            "QTD_REM_AJUSTADA_C0": 10, "QTD_REM_AJUSTADA_C1": 10,
            "QTD_REM_AJUSTADA_C2": 6, "QTD_REM_AJUSTADA_C3": 0,
            "QTD_REM_AJUSTADA_C4": 0,
        }]},
        "historico_vu": {"itens": [{
            "item": 1,
            "vu_ciclos": {"VU_C0": 100.0, "VU_C1": 110.0, "VU_C2": 121.0,
                          "VU_C3": 121.0, "VU_C4": 121.0},
        }]},
        "itens_pc_v10": {"itens": [
            {"ciclo": "C1", "valor_pc": 600.0, "valor_atualizado": 660.0,
             "entra_no_calculo": "Sim"},
            {"ciclo": "C2", "valor_pc": 5000.0, "valor_atualizado": 5500.0,
             "entra_no_calculo": "Sim"},
        ]},
    }
    comp = motor_vta.montar_composicao_vta(leitura)
    assert comp["disponivel"] is True
    ciclos = [p["ciclo"] for p in comp["execucao_por_ciclo"]]
    assert "C1" in ciclos and "C2" not in ciclos
    # 660 (C1) + 6 x 121 (remanescente do corte) = 1.386,00
    assert comp["vta_composicao"] == pytest.approx(660.0 + 726.0, abs=0.01)


# ---- 16. Fallback fail-closed do metodo PC ----

def _memoria_vta(composicao: dict) -> dict:
    from _objeto_processo_reajuste import _montar_memoria_por_ciclo

    leitura = {
        "parametros_v10": {"por_ciclo": {
            "C1": {"fator_acumulado": 1.10},
            "C2": {"fator_acumulado": 1.21},
        }},
        "potencial_futuro": {"disponivel": True,
                             "valor_atualizado_vigente": 726.0},
        "composicao_vta": composicao,
    }
    pcs_pagos = [{
        "numero_pc": "PC-001", "ciclo_calculado": "C1",
        "efeito_financeiro_pc": "Sim", "valor_pago": 600.0,
        "elegivel_retroativo_pc": True,
    }]
    memoria = _montar_memoria_por_ciclo(leitura, {}, pcs_pagos, {})
    return memoria.get("vta") or {}


def test_fallback_nao_fabrica_vta_com_pcs_pagos():
    # Metodo PC com composicao canonica INDISPONIVEL e PCs pagos presentes:
    # o VTA nao pode ser fabricado a partir do pagamento (fail-closed).
    vta = _memoria_vta({"metodo": "pc", "disponivel": False})
    assert vta.get("natureza") == "INDETERMINADO"
    assert vta.get("valor_total_atualizado") is None      # nunca zero
    assert "fail-closed" in str(vta.get("motivo") or "")
    assert "PCs pagos" in str(vta.get("motivo") or "")


def test_fallback_nao_fabrica_vta_no_runtime_sem_base_itemizada():
    # End-to-end: sem itens_Remanesc a base itemizada inexiste; ha PC pago.
    leitura = ler_masterfile_v10(_workbook_pc("Sim", sem_remanescente=True))
    comp = leitura.get("composicao_vta") or {}
    assert comp.get("disponivel") is not True
    vta = (
        (leitura.get("objeto_processo") or {}).get("memoria_por_ciclo") or {}
    ).get("vta") or {}
    if vta.get("metodo") == "pc" or vta.get("natureza") == "INDETERMINADO":
        assert vta.get("valor_total_atualizado") is None
        assert vta.get("natureza") == "INDETERMINADO"


def test_fallback_nao_acionado_com_composicao_valida():
    composicao = {
        "metodo": "pc", "disponivel": True,
        "vta_composicao": 1386.0,
        "total_execucao_atualizada": 660.0,
        "saldo_remanescente": {"valor_atualizado": 726.0},
    }
    vta = _memoria_vta(composicao)
    assert vta.get("valor_total_atualizado") == pytest.approx(1386.0, abs=0.005)
    assert vta.get("natureza") != "INDETERMINADO"
    assert "canonica" in str(vta.get("criterio_selecao") or "")


# ---- 12/18. Metodologia PC estrutural (nao depende de G) ----

def test_metodologia_pc_disponivel_e_com_mesmo_valor_sem_pagamento():
    from _motor_metodologias import MET_PC, montar_motor_metodologias

    def _painel(pago: bool) -> dict:
        pcs = [{
            "numero_pc": "PC-001", "ciclo_temporal": "C1",
            "data_pc": "2024-01-15", "valor_pc": 600.0,
            "valor_devido": 660.0,
            "valor_pago": 600.0 if pago else 0.0,
            "retroativo": 60.0 if pago else 0.0,
            "natureza_delta": "reconhecido" if pago else "potencial",
            "elegivel_retroativo_pc": pago,
        }]
        return {
            "disponivel": True,
            "situacao_pcs": {"pcs": pcs},
            "situacao_financeira": {"resumo_oficial": {}, "totais": {}},
        }

    valores = {}
    for rotulo, pago in (("pago", True), ("nao_pago", False)):
        motor = montar_motor_metodologias({}, _painel(pago))
        alternativas = motor.get("alternativas") or []
        alt_pc = next(a for a in alternativas if a.get("nome") == MET_PC)
        # Etapa 27B: a metodologia PC e ESTRUTURAL — disponivel e com o
        # mesmo valor (execucao devida 660, da entrada) com ou sem pagamento.
        assert alt_pc.get("disponivel") is True, rotulo
        valores[rotulo] = alt_pc.get("valor_recomendado")
        assert valores[rotulo] == pytest.approx(660.0, abs=0.005), rotulo
    assert valores["nao_pago"] == pytest.approx(valores["pago"], abs=0.005)


# ---- 20. Excel real (RUN_EXCEL_INTEGRATION=1): XLS x Python ----

@pytest.mark.skipif(
    os.environ.get("RUN_EXCEL_INTEGRATION") != "1",
    reason="defina RUN_EXCEL_INTEGRATION=1 para executar o Excel COM",
)
def test_excel_com_vta_identico_e_retroativo_classificado(tmp_path: Path):
    import gc
    import pythoncom
    import win32com.client

    def _recalc(caminho: Path) -> None:
        pythoncom.CoInitialize()
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        pasta = None
        try:
            pasta = excel.Workbooks.Open(str(caminho.resolve()))
            excel.CalculateFullRebuild()
            pasta.Save()
            pasta.Close(SaveChanges=False)
            pasta = None
        finally:
            if pasta is not None:
                pasta.Close(SaveChanges=False)
            excel.Quit()
            excel = None
            gc.collect()
            pythoncom.CoUninitialize()

    resultados = {}
    for g in VARIANTES_G:
        caminho = tmp_path / f"pc_27b_{g or 'vazio'}.xlsx"
        caminho.write_bytes(_workbook_pc(g))
        _recalc(caminho)
        conteudo = caminho.read_bytes()
        resultado, diag = processar_coleta_oficial_runtime(conteudo)
        wb = load_workbook(io.BytesIO(conteudo), data_only=True)
        mem = wb["MEMORIA_RESULTADOS"]
        resultados[g] = {
            "vta_runtime": resultado.get("valor_atualizado_contrato"),
            "vta_xls": mem["B26"].value,
            "represado": resultado.get("valor_represado_a_pagar"),
            "check_k2": wb["itens_PC"]["K2"].value,
            "status_retro": mem["F16"].value,
            "status_vta": mem["E26"].value,
            "delta_potencial": mem["D45"].value,
            "status_base": diag.get("status_base"),
            "pronto": diag.get("pronto_para_consolidar"),
        }
    ref = resultados["Sim"]
    assert ref["vta_xls"] == pytest.approx(ref["vta_runtime"], abs=0.01)
    assert ref["represado"] == pytest.approx(60.0, abs=0.01)
    assert ref["check_k2"] == "OK"
    for g in ("Nao", None):
        atual = resultados[g]
        assert atual["vta_xls"] == pytest.approx(ref["vta_xls"], abs=0.005)
        assert atual["vta_runtime"] == pytest.approx(ref["vta_runtime"], abs=0.005)
        assert not atual["represado"]
        assert atual["delta_potencial"] == pytest.approx(60.0, abs=0.01)
        assert atual["check_k2"] == "OK", f"G={g!r} nao pode ser erro em K"
        assert atual["status_retro"] == ref["status_retro"], f"G={g!r}"
        assert atual["status_vta"] == ref["status_vta"], f"G={g!r}"
        assert atual["status_base"] == ref["status_base"], f"G={g!r}"
        assert atual["pronto"] == ref["pronto"], f"G={g!r}"
