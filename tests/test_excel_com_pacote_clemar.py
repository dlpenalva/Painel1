"""Testes Excel COM (opt-in) dos dois bloqueadores corrigidos na Etapa 23B.

Verificam o RESULTADO REAL do Excel (recalculo), nao apenas a string da
formula — a lacuna que deixou os bugs passarem. Rodar com Excel real:

    RUN_EXCEL_INTEGRATION=1 python -m pytest -q tests/test_excel_com_pacote_clemar.py

Workbooks sinteticos minimos e autocontidos (openpyxl para montar, Excel COM
para recalcular). Nao dependem do XLS real do usuario.
"""
from __future__ import annotations

import os
import tempfile
from pathlib import Path

import pytest
from openpyxl import Workbook

com = pytest.mark.skipif(
    os.environ.get("RUN_EXCEL_INTEGRATION") != "1",
    reason="defina RUN_EXCEL_INTEGRATION=1 para Excel COM",
)

# Formulas exatamente como o tool as escreve (Etapa 23B).
FIN_ULTIMA = (
    '=IFERROR(LOOKUP(2,1/(ISNUMBER(financeiro!$A$2:$A$200)'
    '*ISNUMBER(financeiro!$C$2:$C$200)),financeiro!$A$2:$A$200),"")'
)


def _recalcular(path):
    import pythoncom
    import win32com.client
    pythoncom.CoInitialize()
    xl = win32com.client.DispatchEx("Excel.Application")
    xl.Visible = False
    xl.DisplayAlerts = False
    try:
        wb = xl.Workbooks.Open(str(path), UpdateLinks=0)
        xl.CalculateFullRebuild()
        valores = {}
        for ws in wb.Worksheets:
            for addr in ("B12", "T25", "B26"):
                try:
                    valores[f"{ws.Name}!{addr}"] = ws.Range(addr).Value
                except Exception:
                    pass
        wb.Close(SaveChanges=False)
        return valores
    finally:
        xl.Quit()
        pythoncom.CoUninitialize()


@com
def test_cobertura_b12_ignora_linha_total():
    """Grade 04/2022..03/2027 sem pagamentos + linha TOTAL C=SUM=0 (A vazia).

    B12 (ultima evidencia Financeiro) deve ficar VAZIO — a linha TOTAL nao pode
    virar falsa evidencia.
    """
    import datetime as dt
    wb = Workbook()
    fin = wb.active
    fin.title = "financeiro"
    fin["A1"], fin["C1"] = "COMPETENCIA", "VALOR_PAGO"
    ano, mes = 2022, 4
    for r in range(2, 62):                      # 60 competencias, todas C vazia
        fin.cell(r, 1).value = dt.date(ano, mes, 1)
        mes += 1
        if mes > 12:
            mes, ano = 1, ano + 1
    fin["C62"] = "=SUM(C2:C61)"                 # linha TOTAL: A62 vazia, C62=0
    cob = wb.create_sheet("cobertura_temporal")
    cob["B12"] = FIN_ULTIMA
    with tempfile.TemporaryDirectory() as d:
        p = Path(d) / "cob.xlsx"
        wb.save(p)
        v = _recalcular(p)
    assert v["cobertura_temporal!B12"] in (None, "", 0) or str(v["cobertura_temporal!B12"]) == ""


@com
def test_vta_pc_b26_reproduz_golden():
    """VTA-PC materializado: com 1 item sintetico, B26 = C0 + PC + remanescente.

    C0 fisico=(10-6)*100=400; PC exec C1+C2=50+30=80; rem=ROUND(110*4,2)=440.
    VTA-PC = 920,00.
    """
    wb = Workbook()
    ctrl = wb.active
    ctrl.title = "CONTROLE"
    ctrl["A1"], ctrl["B1"] = "MODO DE LEITURA", "PC (Pedidos de Compra)"
    ctrl["A2"], ctrl["B2"] = "Ciclo vigente", "C4"
    # itens_PC: 2 PCs (C1, C2) + bloco de agregacao por ciclo (M/P).
    ipc = wb.create_sheet("itens_PC")
    ipc["C1"], ipc["F1"] = "CICLO_PC", "VALOR_ATUALIZADO"
    ipc["C2"], ipc["F2"] = "C1", 50.0
    ipc["C3"], ipc["F3"] = "C2", 30.0
    ipc["M1"], ipc["P1"] = "CICLO", "VALOR_ATUALIZADO_TOTAL"
    for i, cic in enumerate(["C0", "C1", "C2", "C3", "C4"], start=2):
        ipc.cell(i, 13).value = cic
        ipc.cell(i, 16).value = f'=SUMIF($C$2:$C$100,M{i},$F$2:$F$100)'
    # posicao_contratual: 1 item (A,VU_ORIG,QTD_BASE, G/K/O/S/W = QREM C0..C4)
    posc = wb.create_sheet("posicao_contratual")
    posc["A1"] = "ITEM"
    posc["A2"], posc["B2"], posc["C2"] = 1, 100, 10
    posc["G2"], posc["K2"], posc["O2"], posc["S2"], posc["W2"] = 10, 6, 5, 4, 4
    # historico_VU: 1 item (C..G = VU C0..C4)
    hv = wb.create_sheet("historico_VU")
    hv["A1"] = "ITEM"
    hv["A2"] = 1
    hv["C2"], hv["D2"], hv["E2"], hv["F2"], hv["G2"] = 100, 100, 100, 100, 110
    # RESULTADOS: bloco auxiliar VTA-PC (X/Y por item + T20:T25) + B26 (ramo PC).
    r = wb.create_sheet("RESULTADOS")
    r["X2"] = ('=IF(posicao_contratual!$A2="",0,'
               '(posicao_contratual!$G2-posicao_contratual!$K2)*historico_VU!$C2)')
    r["Y2"] = ('=IF(OR(posicao_contratual!$A2="",$T$20=""),0,'
               'ROUND(ROUND(CHOOSE($T$20+1,historico_VU!$C2,historico_VU!$D2,'
               'historico_VU!$E2,historico_VU!$F2,historico_VU!$G2),2)'
               '*CHOOSE($T$20+1,posicao_contratual!$G2,posicao_contratual!$K2,'
               'posicao_contratual!$O2,posicao_contratual!$S2,posicao_contratual!$W2),2))')
    r["T20"] = '=IFERROR(VALUE(MID(CONTROLE!$B$2,2,3)),"")'
    r["T21"] = '=IF(itens_PC!$P$2>0,itens_PC!$P$2,SUM($X$2:$X$201))'
    r["T22"] = ('=SUMPRODUCT((ROW(itens_PC!$P$2:$P$6)-2>=1)'
                '*(ROW(itens_PC!$P$2:$P$6)-2<$T$20)*itens_PC!$P$2:$P$6)')
    r["T23"] = '=SUM($Y$2:$Y$201)'
    r["T24"] = '=IF(COUNTIF(posicao_contratual!$A$2:$A$201,"<>")=0,0,1)'
    r["T25"] = ('=IF(OR($T$24=0,$T$20=""),"CALCULO MANUAL REQUERIDO",'
                'ROUND($T$21+$T$22+$T$23,2))')
    r["B26"] = ('=IF(CONTROLE!$B$1="PC (Pedidos de Compra)",'
                'IF($T$25="CALCULO MANUAL REQUERIDO","",ROUND($T$25,2)),"")')
    with tempfile.TemporaryDirectory() as d:
        p = Path(d) / "vtapc.xlsx"
        wb.save(p)
        v = _recalcular(p)
    assert round(float(v["RESULTADOS!T25"]), 2) == 920.00
    assert round(float(v["RESULTADOS!B26"]), 2) == 920.00


def test_comparativo_guard_source():
    """Etapa 24B (rapido, sem Excel): o Bloco 1 do comparativo_VTA protege
    todas as colunas com ISNUMBER de VU_ORIGINAL e da QTD do ciclo — nunca
    multiplica por celula vazia/string (evita #VALUE! em ciclo inexistente).
    """
    from pathlib import Path
    tool = (Path(__file__).resolve().parents[1] / "tools"
            / "aplicar_pacote_clemar_template.py").read_text(encoding="utf-8")
    assert "NOT(ISNUMBER({P}!B2))" in tool          # VU_ORIGINAL numerico
    assert "NOT(ISNUMBER({P}!{qcol}2))" in tool      # QTD_REM_AJUSTADA_Cn numerica


def _recalc_ret(path, cells):
    import pythoncom
    import win32com.client
    pythoncom.CoInitialize()
    xl = win32com.client.DispatchEx("Excel.Application")
    xl.Visible = False
    xl.DisplayAlerts = False
    try:
        wb = xl.Workbooks.Open(str(path), UpdateLinks=0)
        xl.CalculateFullRebuild()
        cv = wb.Worksheets("comparativo_VTA")
        out = {a: cv.Range(a).Value for a in cells}
        # varre erros em toda a aba
        import pywintypes
        erros = []
        for t in (-4123, 2):
            try:
                erros.append(cv.UsedRange.SpecialCells(t, 16).Address)
            except pywintypes.com_error:
                pass
        out["_erros"] = erros
        wb.Close(SaveChanges=False)
        return out
    finally:
        xl.Quit()
        pythoncom.CoUninitialize()


@com
def test_comparativo_ignora_ciclo_inexistente():
    """Bloco 1 do comparativo_VTA: ciclo inexistente -> vazio (nao #VALUE!);
    QTD=0 numerico -> R$ 0,00; ITEM vazio -> vazio. Recalculo Excel real.

    posicao_contratual (A=ITEM, B=VU_ORIGINAL, G/K/O/S/W=QREM_C0..C4):
      linha2 ITEM-1: C3-vigente (W vazio)            -> comparativo linha4
      linha3 ITEM-2: C2-vigente (S e W vazios)       -> comparativo linha5
      linha4 ITEM-3: todas QTD = 0 numerico          -> comparativo linha6
      linha5: ITEM vazio                              -> comparativo linha7
    """
    from openpyxl import Workbook
    wb = Workbook()
    p = wb.active
    p.title = "posicao_contratual"
    p["A1"] = "ITEM"
    # ITEM-1 (C3-vigente): C4 (W) vazio
    p["A2"], p["B2"], p["G2"], p["K2"], p["O2"], p["S2"] = "IT-1", 100, 10, 8, 5, 3
    # ITEM-2 (C2-vigente): C3 (S) e C4 (W) vazios
    p["A3"], p["B3"], p["G3"], p["K3"], p["O3"] = "IT-2", 50, 20, 15, 10
    # ITEM-3: QTD 0 numerico em todos os ciclos
    p["A4"], p["B4"] = "IT-3", 200
    for c in ("G", "K", "O", "S", "W"):
        p[f"{c}4"] = 0
    cv = wb.create_sheet("comparativo_VTA")
    Pn = "posicao_contratual"
    for row, pr in ((4, 2), (5, 3), (6, 4), (7, 5)):
        for col, qcol in (("C", "G"), ("D", "K"), ("E", "O"), ("F", "S"), ("G", "W")):
            cv[f"{col}{row}"] = (
                f'=IF(OR({Pn}!A{pr}="",NOT(ISNUMBER({Pn}!B{pr})),'
                f'NOT(ISNUMBER({Pn}!{qcol}{pr}))),"",ROUND({Pn}!B{pr}*{Pn}!{qcol}{pr},2))')
    with tempfile.TemporaryDirectory() as d:
        fp = Path(d) / "comp.xlsx"
        wb.save(fp)
        v = _recalc_ret(fp, ["C4", "D4", "E4", "F4", "G4",
                             "C5", "F5", "G5", "C6", "G6", "C7"])
    assert v["_erros"] == []                       # ZERO #VALUE! na aba
    assert round(float(v["C4"]), 2) == 1000.0      # ITEM-1 C0
    assert round(float(v["F4"]), 2) == 300.0       # ITEM-1 C3
    assert v["G4"] in (None, "")                    # ITEM-1 C4 inexistente -> vazio
    assert round(float(v["C5"]), 2) == 1000.0      # ITEM-2 C0
    assert v["F5"] in (None, "") and v["G5"] in (None, "")  # ITEM-2 C3/C4 vazios
    assert round(float(v["C6"]), 2) == 0.0         # ITEM-3 QTD 0 -> R$ 0,00
    assert round(float(v["G6"]), 2) == 0.0         # zero real, nao vazio
    assert v["C7"] in (None, "")                    # ITEM vazio -> vazio
