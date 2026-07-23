from __future__ import annotations

import io
import os
import re
import time
import zipfile
from datetime import date, datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook

from _coleta_oficial import gerar_coleta_oficial_preenchida
from _coleta_reajuste_documentos import processar_coleta_oficial_runtime


pytestmark = pytest.mark.skipif(
    os.environ.get("RUN_EXCEL_INTEGRATION") != "1",
    reason="defina RUN_EXCEL_INTEGRATION=1 para executar o Excel COM",
)


def _dados(*, longo: bool = False) -> dict:
    ciclos = [
        {"ciclo": "C1", "data_inicio": date(2024, 1, 1), "data_fim": date(2024, 12, 31), "data_pedido": date(2024, 1, 1), "financeiro_inicio": date(2024, 1, 1), "percentual": 0.10},
    ]
    if longo:
        ciclos.extend([
            {"ciclo": "C2", "data_inicio": date(2025, 1, 1), "data_fim": date(2025, 12, 31), "data_pedido": date(2025, 1, 1), "financeiro_inicio": date(2025, 1, 1), "percentual": 0.10},
            {"ciclo": "C3", "data_inicio": date(2026, 1, 1), "data_fim": date(2026, 12, 31), "data_pedido": date(2026, 1, 1), "financeiro_inicio": date(2026, 1, 1), "percentual": 0.10},
            {"ciclo": "C4", "data_inicio": date(2027, 1, 1), "data_fim": date(2028, 12, 31), "data_pedido": date(2027, 1, 1), "financeiro_inicio": date(2027, 1, 1), "percentual": 0.10},
        ])
    return {
        "origem": "Teste Excel COM",
        "indice": "IST",
        "data_base_original": "01/01/2023",
        "data_corte": date(2028, 12, 31) if longo else date(2024, 12, 31),
        "ciclos": ciclos,
    }


def _dados_efeitos() -> dict:
    return {
        "origem": "Teste efeitos financeiros Excel COM",
        "indice": "IST",
        "data_base_original": "01/02/2023",
        "data_corte": date(2025, 1, 31),
        "ciclos": [{
            "ciclo": "C1",
            "data_inicio": date(2024, 2, 1),
            "data_fim": date(2025, 1, 31),
            "data_pedido": date(2024, 3, 10),
            "financeiro_inicio": date(2024, 4, 18),
            "percentual_aplicado": 0.10,
            "objeto_analise_atual": True,
        }],
    }


def _linha_competencia(ws, ano: int, mes: int) -> int:
    for row in range(2, 74):
        valor = ws[f"A{row}"].value
        if valor and (valor.year, valor.month) == (ano, mes):
            return row
    raise AssertionError(f"competencia {mes:02d}/{ano} ausente")


def _cor_excel(rgb: str) -> int:
    rgb = rgb[-6:]
    r, g, b = (int(rgb[i:i + 2], 16) for i in (0, 2, 4))
    return r + (g << 8) + (b << 16)


def _salvar(wb, caminho: Path) -> None:
    buf = io.BytesIO()
    wb.save(buf)
    caminho.write_bytes(buf.getvalue())


def _recalcular_excel(caminho: Path) -> None:
    import gc
    import pythoncom
    import win32com.client

    pythoncom.CoInitialize()
    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    pasta = None
    try:
        pasta = excel.Workbooks.Open(str(caminho.resolve()))
        excel.CalculateFullRebuild()
        pasta.Save()
        pasta.Close(SaveChanges=False)
        pasta = None
    finally:
        if pasta is not None:
            pasta.Close(SaveChanges=False)
            pasta = None
        excel.Quit()
        excel = None
        gc.collect()
        pythoncom.CoUninitialize()


def _excel_editar_e_inspecionar(caminho: Path, editar, inspecionar):
    import gc
    import pythoncom
    import win32com.client

    def tentar(acao):
        ultimo = None
        for _ in range(30):
            try:
                return acao()
            except Exception as exc:
                ultimo = exc
                codigo = getattr(exc, "hresult", None)
                if codigo is None and getattr(exc, "args", ()):
                    codigo = exc.args[0]
                if codigo != -2147418111:
                    raise
                pythoncom.PumpWaitingMessages()
                time.sleep(0.2)
        raise ultimo

    pythoncom.CoInitialize()
    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    pasta = None
    try:
        pasta = tentar(lambda: excel.Workbooks.Open(
            str(caminho.resolve()), UpdateLinks=0, ReadOnly=False, CorruptLoad=0
        ))
        time.sleep(0.5)
        tentar(lambda: editar(pasta))
        tentar(excel.CalculateFullRebuild)
        resultado = tentar(lambda: inspecionar(pasta))
        tentar(pasta.Save)
        tentar(lambda: pasta.Close(SaveChanges=False))
        pasta = None
        return resultado
    finally:
        if pasta is not None:
            pasta.Close(SaveChanges=False)
        excel.Quit()
        excel = None
        gc.collect()
        pythoncom.CoUninitialize()


def _arquivo_sheet_por_nome(caminho: Path, nome_aba: str) -> str:
    """Resolve xl/worksheets/sheetN.xml da aba pelo nome (§26: sem hard-code).

    A numeracao sheetN.xml segue a ordem interna de criacao, nao a ordem das
    guias; com posicao_referencia adicionada, RESULTADOS deixou de ser sheet11.
    """
    with zipfile.ZipFile(caminho, "r") as z:
        wbxml = z.read("xl/workbook.xml").decode("utf-8")
        rels = z.read("xl/_rels/workbook.xml.rels").decode("utf-8")
    rid = None
    for nome, ident in re.findall(r'<sheet name="([^"]+)"[^>]*r:id="([^"]+)"', wbxml):
        if nome == nome_aba:
            rid = ident
            break
    assert rid is not None, f"aba {nome_aba} nao encontrada em workbook.xml"
    alvo = dict(re.findall(r'Id="([^"]+)"[^>]*Target="([^"]+)"', rels))[rid]
    return alvo if alvo.startswith("xl/") else "xl/" + alvo.lstrip("/")


def _alterar_cache_resultados(caminho: Path, celula: str, valor: float) -> None:
    arquivo_resultados = _arquivo_sheet_por_nome(caminho, "RESULTADOS")
    temporario = caminho.with_suffix(".tmp.xlsx")
    with zipfile.ZipFile(caminho, "r") as origem, zipfile.ZipFile(
        temporario, "w", zipfile.ZIP_DEFLATED
    ) as destino:
        for item in origem.infolist():
            dados = origem.read(item.filename)
            if item.filename == arquivo_resultados:
                texto = dados.decode("utf-8")
                padrao = re.compile(
                    rf'(<c r="{re.escape(celula)}"[^>]*>.*?<v>)([^<]*)(</v>)',
                    re.DOTALL,
                )
                texto, quantidade = padrao.subn(rf"\g<1>{valor}\g<3>", texto, count=1)
                assert quantidade == 1
                dados = texto.encode("utf-8")
            destino.writestr(item, dados)
    temporario.replace(caminho)


def test_excel_com_linha_73_entra_em_resultados(tmp_path: Path) -> None:
    wb = load_workbook(io.BytesIO(gerar_coleta_oficial_preenchida(_dados(longo=True))))
    wb["financeiro"]["C73"] = 100.0
    wb["financeiro"]["G73"] = "Sim"
    caminho = tmp_path / "linha_73.xlsx"
    _salvar(wb, caminho)
    _recalcular_excel(caminho)

    valores = load_workbook(caminho, data_only=True)
    assert str(valores["financeiro"]["B73"].value).lower() == "c4"
    assert valores["financeiro"]["F73"].value == pytest.approx(46.41, abs=0.01)
    assert valores["RESULTADOS"]["B14"].value == pytest.approx(46.41, abs=0.01)


def test_excel_com_efeitos_itens_pc_data_exata_cores_e_reabertura(tmp_path: Path) -> None:
    caminho = tmp_path / "efeitos_itens_pc.xlsx"
    caminho.write_bytes(gerar_coleta_oficial_preenchida(_dados_efeitos()))

    def editar(pasta):
        ws = pasta.Worksheets("itens_PC")
        casos = (
            (2, "PC-ANTES", datetime(2024, 4, 10), "Nao"),
            (3, "PC-EXATO", datetime(2024, 4, 18), "Nao"),
            (4, "PC-DEPOIS", datetime(2024, 4, 25), "Sim"),
        )
        for linha, numero, data_pc, pago in casos:
            ws.Cells(linha, 1).Value = numero
            ws.Cells(linha, 2).Value = data_pc
            ws.Cells(linha, 4).Value = 100.0
            ws.Cells(linha, 7).Value = pago

    def inspecionar(pasta):
        ws = pasta.Worksheets("itens_PC")
        valores = {
            linha: tuple(ws.Cells(linha, col).Value for col in (3, 5, 6, 8, 9, 10, 11, 12))
            for linha in (2, 3, 4)
        }
        cores = {linha: ws.Cells(linha, 1).DisplayFormat.Interior.Color for linha in (2, 3, 4)}
        links = pasta.LinkSources(1)
        return valores, cores, ws.Cells(2, 2).NumberFormat, links

    valores, cores, formato_data, links = _excel_editar_e_inspecionar(
        caminho, editar, inspecionar
    )
    # tupla: C, E, F, H, I, J, K, L
    assert valores[2][:6] == ("C1", 1.0, 100.0, 0.0, 100.0, 0.0)
    assert valores[2][7] == "Nao" and valores[2][6] == "OK"
    assert valores[3][:6] == ("C1", 1.1, 110.0, 0.0, 110.0, 10.0)
    assert valores[3][7] == "Sim" and valores[3][6] == "OK"
    assert valores[4][:6] == ("C1", 1.1, 110.0, 10.0, 0.0, 0.0)
    assert valores[4][7] == "Sim" and valores[4][6] == "OK"
    assert cores[2] == _cor_excel("FFF4CCCC")
    assert cores[3] != _cor_excel("FFF4CCCC")
    assert cores[4] != _cor_excel("FFF4CCCC")
    assert "dd" in str(formato_data).lower()
    assert links in (None, ())

    reaberto = load_workbook(caminho, data_only=True)
    assert reaberto["itens_PC"]["L2"].value == "Nao"
    assert reaberto["itens_PC"]["L3"].value == "Sim"
    assert reaberto["itens_PC"]["L4"].value == "Sim"
    assert not any(
        isinstance(cell.value, str) and cell.value.upper() in {
            "#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#N/A", "#NUM!", "#NULL!"
        }
        for aba in reaberto.worksheets
        for row in aba.iter_rows()
        for cell in row
    )

    impeditivo = tmp_path / "efeitos_itens_pc_sem_inicio.xlsx"
    wb_imp = load_workbook(io.BytesIO(gerar_coleta_oficial_preenchida(_dados_efeitos())))
    wb_imp["parametros"]["H3"] = None
    _salvar(wb_imp, impeditivo)

    def editar_imp(pasta):
        ws = pasta.Worksheets("itens_PC")
        ws.Cells(2, 1).Value = "PC-SEM-INICIO"
        ws.Cells(2, 2).Value = datetime(2024, 4, 18)
        ws.Cells(2, 4).Value = 100.0
        ws.Cells(2, 7).Value = "Sim"

    def inspecionar_imp(pasta):
        ws = pasta.Worksheets("itens_PC")
        return (
            ws.Cells(2, 12).Value,
            ws.Cells(2, 11).Value,
            ws.Cells(2, 1).DisplayFormat.Interior.Color,
        )

    efeito, check, cor = _excel_editar_e_inspecionar(
        impeditivo, editar_imp, inspecionar_imp
    )
    assert efeito in (None, "")
    assert "PC-SEM-INICIO" in check and "C1" in check
    assert cor == _cor_excel("FFFFE699")


def test_excel_com_runtime_financeiro_pc_reconciliacao_e_bloqueio(tmp_path: Path) -> None:
    base = gerar_coleta_oficial_preenchida(_dados())

    financeiro = load_workbook(io.BytesIO(base))
    financeiro["financeiro"]["C14"] = 600.0
    financeiro["financeiro"]["G14"] = "Sim"
    financeiro["itens_Remanesc"]["A2"] = "ITEM-1"
    financeiro["itens_Remanesc"]["B2"] = 10.0
    financeiro["itens_Remanesc"]["C2"] = 100.0
    financeiro["itens_Remanesc"]["E2"] = 4.0
    caminho_fin = tmp_path / "financeiro.xlsx"
    _salvar(financeiro, caminho_fin)
    _recalcular_excel(caminho_fin)
    resultado_fin, _ = processar_coleta_oficial_runtime(caminho_fin.read_bytes())
    assert resultado_fin["valor_represado_a_pagar"] == pytest.approx(60.0, abs=0.01)
    assert resultado_fin["remanescente_reajustado"] == pytest.approx(440.0, abs=0.01)
    assert resultado_fin["valor_atualizado_contrato"] == pytest.approx(1100.0, abs=0.01)
    assert resultado_fin["reconciliacao_xls_python"]["status_geral"] == "CONCILIADO"
    assert not resultado_fin["formalizacao_bloqueada"]

    pc = load_workbook(io.BytesIO(base))
    pc["CONTROLE"]["B1"] = "Pedidos de Compras"
    pc["RESULTADOS"]["B4"] = "PCs"
    pc["itens_PC"]["A2"] = "PC-001"
    pc["itens_PC"]["B2"] = date(2024, 1, 15)
    pc["itens_PC"]["D2"] = 600.0
    pc["itens_PC"]["G2"] = "Sim"
    pc["itens_Remanesc"]["A2"] = "ITEM-1"
    pc["itens_Remanesc"]["B2"] = 10.0
    pc["itens_Remanesc"]["C2"] = 100.0
    pc["itens_Remanesc"]["E2"] = 4.0
    caminho_pc = tmp_path / "pc.xlsx"
    _salvar(pc, caminho_pc)
    _recalcular_excel(caminho_pc)
    resultado_pc, _ = processar_coleta_oficial_runtime(caminho_pc.read_bytes())
    assert resultado_pc["df_pedidos_compra"].iloc[0]["numero_pc"] == "PC-001"
    assert resultado_pc["valor_represado_a_pagar"] == pytest.approx(60.0, abs=0.01)
    assert resultado_pc["reconciliacao_xls_python"]["status_geral"] == "CONCILIADO"
    assert not resultado_pc["formalizacao_bloqueada"]

    _alterar_cache_resultados(caminho_pc, "C15", 9999.0)
    divergente, _ = processar_coleta_oficial_runtime(caminho_pc.read_bytes())
    assert divergente["reconciliacao_xls_python"]["divergencias_relevantes"]
    # Formalizacao permanece bloqueada pela divergencia XLS x Python...
    assert divergente["formalizacao_bloqueada"]
    documentos = (divergente.get("capacidades") or {}).get("documentos", {})
    # ...mas §7: os 3 documentos diagnosticos seguem DISPONIVEIS (disponibilidade
    # documental != aptidao para formalizar). Os demais formais ficam bloqueados.
    for chave in ("sumario_executivo", "despacho_saneador", "termo_apostila"):
        assert documentos[chave]["habilitado"], chave
    for chave in ("garantia_contratual", "dou", "relatorio_executivo"):
        assert not documentos[chave]["habilitado"], chave


def test_excel_com_pcs_multiciclo_ignora_fator_historico_fora_do_objeto(tmp_path: Path) -> None:
    dados = {
        "origem": "Teste PC multiciclo",
        "indice": "IST",
        "data_base_original": "01/01/2023",
        "data_corte": date(2026, 12, 31),
        "ciclos": [
            {
                "ciclo": "C1", "data_inicio": date(2024, 1, 1),
                "data_fim": date(2024, 12, 31), "percentual": 0.05,
                "financeiro_inicio": date(2024, 1, 1),
                "ciclo_ja_concedido": True,
                "situacao": "Histórico fora do objeto atual",
            },
            {
                "ciclo": "C2", "data_inicio": date(2025, 1, 1),
                "data_fim": date(2025, 12, 31), "percentual": 0.10,
                "financeiro_inicio": date(2025, 1, 1),
            },
            {
                "ciclo": "C3", "data_inicio": date(2026, 1, 1),
                "data_fim": date(2026, 12, 31), "percentual": 0.08,
                "financeiro_inicio": date(2026, 1, 1),
            },
        ],
    }
    wb = load_workbook(io.BytesIO(gerar_coleta_oficial_preenchida(dados)))
    wb["CONTROLE"]["B1"] = "Pedidos de Compras"
    wb["RESULTADOS"]["B4"] = "PCs"
    for row, numero, data_pc, valor in (
        (2, "PC-2001", date(2025, 2, 15), 600.0),
        (3, "PC-3001", date(2026, 3, 20), 800.0),
    ):
        wb["itens_PC"][f"A{row}"] = numero
        wb["itens_PC"][f"B{row}"] = data_pc
        wb["itens_PC"][f"D{row}"] = valor
        wb["itens_PC"][f"G{row}"] = "Sim"
    caminho = tmp_path / "pc_multiciclo_historico.xlsx"
    _salvar(wb, caminho)
    _recalcular_excel(caminho)

    resultado, _ = processar_coleta_oficial_runtime(caminho.read_bytes())
    assert resultado["valor_represado_a_pagar"] == pytest.approx(210.40, abs=0.01)
    assert resultado["reconciliacao_xls_python"]["status_geral"] == "CONCILIADO"
    assert not resultado["reconciliacao_xls_python"]["divergencias_relevantes"]
    assert not any(
        "Divergência relevante XLS × Python" in bloqueio
        for bloqueio in resultado.get("bloqueios_formalizacao", [])
    )


def _arquivo_efeitos(tmp_path: Path, nome: str) -> tuple[Path, int]:
    payload = gerar_coleta_oficial_preenchida(_dados_efeitos())
    wb = load_workbook(io.BytesIO(payload), data_only=False)
    row = _linha_competencia(wb["financeiro"], 2024, 4)
    caminho = tmp_path / nome
    _salvar(wb, caminho)
    return caminho, row


def test_excel_com_efeito_a_nao_preserva_nominal_delta_zero_e_vermelho(tmp_path: Path) -> None:
    caminho, row = _arquivo_efeitos(tmp_path, "efeito_a_nao.xlsx")

    def editar(wb):
        ws = wb.Worksheets("financeiro")
        ws.Range(f"C{row}").Value = 100.0
        ws.Range(f"G{row}").Value = "Nao"

    def ler(wb):
        ws = wb.Worksheets("financeiro")
        return ws.Range(f"E{row}").Value, ws.Range(f"F{row}").Value, ws.Range(f"A{row}").DisplayFormat.Interior.Color

    atualizado, delta, cor = _excel_editar_e_inspecionar(caminho, editar, ler)
    assert atualizado == pytest.approx(100.0, abs=0.01)
    assert delta == pytest.approx(0.0, abs=0.01)
    assert cor == _cor_excel("FFC7CE")


def test_excel_com_efeito_b_sim_aplica_fator_sem_vermelho(tmp_path: Path) -> None:
    caminho, row = _arquivo_efeitos(tmp_path, "efeito_b_sim.xlsx")

    def editar(wb):
        ws = wb.Worksheets("financeiro")
        ws.Range(f"C{row}").Value = 100.0
        ws.Range(f"G{row}").Value = "Sim"

    def ler(wb):
        ws = wb.Worksheets("financeiro")
        return ws.Range(f"E{row}").Value, ws.Range(f"F{row}").Value, ws.Range(f"A{row}").DisplayFormat.Interior.Color

    atualizado, delta, cor = _excel_editar_e_inspecionar(caminho, editar, ler)
    assert atualizado == pytest.approx(110.0, abs=0.01)
    assert delta == pytest.approx(10.0, abs=0.01)
    assert cor not in (_cor_excel("FFC7CE"), _cor_excel("FCE4D6"))


def test_excel_com_efeito_c_override_persiste_vta_nominal_sem_duplicar_aviso(tmp_path: Path) -> None:
    caminho, row = _arquivo_efeitos(tmp_path, "efeito_c_override.xlsx")

    def editar(wb):
        ws = wb.Worksheets("financeiro")
        assert ws.Range(f"G{row}").Value == "Sim"
        ws.Range(f"C{row}").Value = 100.0
        ws.Range(f"G{row}").Value = "Nao"

    def ler(wb):
        ws = wb.Worksheets("financeiro")
        return ws.Range(f"E{row}").Value, ws.Range(f"F{row}").Value, ws.Range(f"G{row}").Value

    atualizado, delta, efeito = _excel_editar_e_inspecionar(caminho, editar, ler)
    assert (atualizado, delta, efeito) == (100, 0, "Nao")
    persistido = load_workbook(caminho, data_only=True, read_only=True)
    assert persistido["financeiro"][f"G{row}"].value == "Nao"
    assert persistido["financeiro"][f"E{row}"].value == pytest.approx(100.0, abs=0.01)
    assert persistido["financeiro"][f"F{row}"].value == pytest.approx(0.0, abs=0.01)
    persistido.close()
    resultado, diagnostico = processar_coleta_oficial_runtime(caminho.read_bytes())
    avisos = [a for a in diagnostico["avisos"] if "ajustada manualmente" in a]
    assert resultado["valor_represado_a_pagar"] == pytest.approx(0.0, abs=0.01)
    conferencia = next(
        item for item in resultado["memoria_por_ciclo"]["conferencias_metodologicas"]
        if item["metodo"] == "financeiro"
    )
    assert conferencia["executado_atualizado"] == pytest.approx(100.0, abs=0.01)
    assert len(avisos) == 1


def test_excel_com_efeito_d_vazio_bloqueia_e_f_e_documentos(tmp_path: Path) -> None:
    caminho, row = _arquivo_efeitos(tmp_path, "efeito_d_vazio.xlsx")

    def editar(wb):
        ws = wb.Worksheets("financeiro")
        ws.Range(f"C{row}").Value = 100.0
        ws.Range(f"G{row}").ClearContents()

    def ler(wb):
        ws = wb.Worksheets("financeiro")
        return ws.Range(f"E{row}").Value, ws.Range(f"F{row}").Value, ws.Range(f"A{row}").DisplayFormat.Interior.Color

    atualizado, delta, cor = _excel_editar_e_inspecionar(caminho, editar, ler)
    assert atualizado in (None, "")
    assert delta in (None, "")
    assert cor == _cor_excel("FCE4D6")
    with pytest.raises(ValueError, match="Efeito financeiro nao informado.*04/2024"):
        processar_coleta_oficial_runtime(caminho.read_bytes())


def test_excel_com_efeito_e_dropdown_termina_em_g73(tmp_path: Path) -> None:
    caminho, _ = _arquivo_efeitos(tmp_path, "efeito_e_dropdown.xlsx")

    def validacao(celula):
        try:
            return celula.Validation.Type, celula.Validation.Formula1
        except Exception:
            return None, None

    def ler(wb):
        ws = wb.Worksheets("financeiro")
        return validacao(ws.Range("G2")), validacao(ws.Range("G73")), validacao(ws.Range("G74"))

    g2, g73, g74 = _excel_editar_e_inspecionar(caminho, lambda wb: None, ler)
    assert g2[0] == 3 and g2[1].replace(";", ",") == "Sim,Nao"
    assert g73[0] == 3 and g73[1].replace(";", ",") == "Sim,Nao"
    assert g74 == (None, None)


def test_excel_com_efeito_f_abre_sem_reparo_e_preserva_estrutura(tmp_path: Path) -> None:
    caminho, _ = _arquivo_efeitos(tmp_path, "efeito_f_sem_reparo.xlsx")
    antes = {p.name for p in tmp_path.iterdir()}

    def ler(wb):
        nomes = [wb.Worksheets(i).Name for i in range(1, wb.Worksheets.Count + 1)]
        return nomes, wb.Worksheets("financeiro").Range("E2").Formula, wb.Worksheets("financeiro").Range("F73").Formula

    nomes, formula_e2, formula_f73 = _excel_editar_e_inspecionar(caminho, lambda wb: None, ler)
    depois = {p.name for p in tmp_path.iterdir()}
    novos_logs = [n for n in depois - antes if "repair" in n.lower() or "recover" in n.lower()]
    assert nomes == ["CONTROLE", "parametros", "financeiro", "itens_Remanesc", "itens_Consumidos", "itens_PC", "aditivos", "posicao_contratual", "itens_RC", "historico_VU", "RESULTADOS"]
    assert 'G2="Sim"' in formula_e2
    assert 'G73="Sim"' in formula_f73
    assert not novos_logs
