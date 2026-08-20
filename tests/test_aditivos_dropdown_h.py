"""Regressão cirúrgica das validações e fórmulas de aditivos."""

import re
import zipfile
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook

import _coleta_oficial
from _coleta_oficial import (
    TEMPLATE_COLETA_OFICIAL,
    _garantir_apresentacao_retroativos_e_aditivos,
    _validar_validacoes_aditivos_criticas,
    assinatura_codigo_coleta,
    obter_coleta_oficial_bytes,
)


@pytest.fixture(scope="module")
def workbooks():
    base = load_workbook(TEMPLATE_COLETA_OFICIAL, data_only=False)
    runtime = load_workbook(BytesIO(obter_coleta_oficial_bytes()), data_only=False)
    yield base, runtime
    base.close()
    runtime.close()


def _validacoes_por_faixa(wb):
    validacoes = {}
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            assinatura = (
                dv.type,
                dv.formula1,
                dv.formula2,
                dv.allowBlank,
                dv.showDropDown,
                dv.errorStyle,
                dv.error,
                dv.errorTitle,
                dv.prompt,
                dv.promptTitle,
            )
            for faixa in dv.sqref.ranges:
                validacoes[(ws.title, str(faixa))] = assinatura
    return validacoes


def _formulas(ws):
    return {
        cell.coordinate: cell.value
        for row in ws.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    }


def test_a_aditivos_h_tem_lista_sim_nao(workbooks):
    _, runtime = workbooks
    ws = runtime["aditivos"]
    validacoes = [
        dv for dv in ws.data_validations.dataValidation
        if "H2:H200" in str(dv.sqref)
    ]
    assert ws["H1"].value == "Aplicar reajuste? (Sim/Nao)"
    assert len(validacoes) == 1
    assert validacoes[0].type == "list"
    assert validacoes[0].formula1 == '"Sim,Nao"'
    # Faixa combinada do template oficial: a mesma lista atende H e K.
    assert str(validacoes[0].sqref) == "H2:H200 K2:K200"


def test_b_aditivos_k_preserva_dropdown_e_visibilidade_do_template(workbooks):
    """Determinacao anterior (K automatico e oculto) REVOGADA: K volta a ser
    entrada do usuario, exatamente como o template oficial a define."""
    base, runtime = workbooks
    ws = runtime["aditivos"]
    assert ws.column_dimensions["K"].hidden is base["aditivos"].column_dimensions["K"].hidden
    assert ws.column_dimensions["K"].hidden is False
    assert ws["K2"].value is None
    assert any("K2:K200" in str(dv.sqref) for dv in ws.data_validations.dataValidation)


def test_c_aditivos_d_preserva_validacao_existente(workbooks):
    base, runtime = workbooks
    assert _validacoes_por_faixa(runtime)[("aditivos", "D2:D200")] == (
        _validacoes_por_faixa(base)[("aditivos", "D2:D200")]
    )


def test_d_todas_as_demais_validacoes_do_template_sao_preservadas(workbooks):
    base, runtime = workbooks
    esperadas = _validacoes_por_faixa(base)
    encontradas = _validacoes_por_faixa(runtime)
    assert not (set(esperadas) - set(encontradas))
    assert {
        chave: encontradas[chave][0]
        for chave in esperadas
    } == {
        chave: assinatura[0]
        for chave, assinatura in esperadas.items()
    }


def test_e_f_j_consumindo_h_preserva_os_dois_ramos(workbooks):
    base, runtime = workbooks
    formula = runtime["aditivos"]["J2"].value
    assert formula == base["aditivos"]["J2"].value
    assert 'UPPER(H2)="SIM"' in formula.upper()
    assert formula.endswith(",F2),2))")


def test_g_h_nao_altera_delta_de_quantidade_l(workbooks):
    base, runtime = workbooks
    formula = runtime["aditivos"]["L2"].value
    assert formula == base["aditivos"]["L2"].value
    assert "H2" not in formula.upper()


def test_h_funcao_nao_altera_formulas_de_resultados_memoria_e_vta():
    wb = load_workbook(TEMPLATE_COLETA_OFICIAL, data_only=False)
    antes = {
        aba: _formulas(wb[aba])
        for aba in ("RESULTADOS", "MEMORIA_RESULTADOS", "comparativo_VTA")
    }
    _garantir_apresentacao_retroativos_e_aditivos(wb)
    for aba in ("RESULTADOS", "MEMORIA_RESULTADOS", "comparativo_VTA"):
        assert _formulas(wb[aba]) == antes[aba]
    wb.close()


def test_normaliza_nao_com_til_sem_tocar_sim_e_nao():
    wb = load_workbook(TEMPLATE_COLETA_OFICIAL, data_only=False)
    ws = wb["aditivos"]
    ws["H2"] = "Não"
    ws["H3"] = "Sim"
    ws["H4"] = "Nao"
    _garantir_apresentacao_retroativos_e_aditivos(wb)
    assert [ws[f"H{linha}"].value for linha in range(2, 5)] == ["Nao", "Sim", "Nao"]
    wb.close()


def _xml_da_aba(dados_zip: bytes, nome_aba: str) -> str:
    """Le o XML bruto de uma aba a partir dos BYTES REAIS do XLSX entregue.

    Resolve o relacionamento nome->r:id->arquivo via workbook.xml e
    workbook.xml.rels (em vez de supor sheet1.xml/sheet2.xml por ordem),
    para provar que a validacao foi de fato SERIALIZADA no XML, e nao
    apenas presente no modelo de objetos do openpyxl em memoria.
    """
    with zipfile.ZipFile(BytesIO(dados_zip)) as zf:
        workbook_xml = zf.read("xl/workbook.xml").decode("utf-8")
        rels_xml = zf.read("xl/_rels/workbook.xml.rels").decode("utf-8")

        tags = re.findall(r"<sheet\b[^>]*/>", workbook_xml)
        tag = next((t for t in tags if f'name="{nome_aba}"' in t), None)
        assert tag is not None, f"aba {nome_aba!r} nao encontrada em workbook.xml"
        r_id_match = re.search(r'r:id="(rId\d+)"', tag)
        assert r_id_match, f"aba {nome_aba!r} sem r:id em workbook.xml"
        r_id = r_id_match.group(1)

        rel_tags = re.findall(r"<Relationship\b[^>]*/>", rels_xml)
        rel_tag = next((t for t in rel_tags if f'Id="{r_id}"' in t), None)
        assert rel_tag is not None, f"relationship {r_id} nao encontrado"
        alvo = re.search(r'Target="([^"]+)"', rel_tag).group(1)
        caminho = f"xl/{alvo}" if not alvo.startswith("/") else alvo.lstrip("/")
        return zf.read(caminho).decode("utf-8")


def _validacoes_lista_do_xml(xml_aba: str) -> dict[str, str | None]:
    blocos = re.findall(
        r"<dataValidation\b[^>]*>.*?</dataValidation>|<dataValidation\b[^>]*/>",
        xml_aba,
        re.DOTALL,
    )
    por_sqref: dict[str, str | None] = {}
    for bloco in blocos:
        sqref_m = re.search(r'sqref="([^"]+)"', bloco)
        if not sqref_m:
            continue
        formula_m = re.search(r"<formula1>(.*?)</formula1>", bloco, re.DOTALL)
        por_sqref[sqref_m.group(1)] = formula_m.group(1) if formula_m else None
    return por_sqref


def test_i_xml_bruto_do_xlsx_real_persiste_h_e_d():
    """REGRA PERMANENTE: prova nos BYTES REAIS entregues (nao so no modelo
    de objetos openpyxl) que aditivos!H2:H200 e D2:D200 tem a lista suspensa
    persistida no XML da planilha. Qualquer perda futura — cache de
    processo servindo bytes antigos, bug de serializacao, template trocado —
    derruba este teste."""
    xml_aditivos = _xml_da_aba(obter_coleta_oficial_bytes(), "aditivos")
    validacoes = _validacoes_lista_do_xml(xml_aditivos)
    assert validacoes.get("H2:H200 K2:K200") == '"Sim,Nao"'
    assert validacoes.get("D2:D200") == '"Acrescimo,Supressao"'


def test_j_guarda_estrutural_nao_barra_geracao_correta():
    obter_coleta_oficial_bytes()  # nao deve levantar ValueError


def test_k_guarda_estrutural_detecta_regressao_de_dropdown():
    """Prova que a barreira de _validar_validacoes_aditivos_criticas
    efetivamente falha (CI vermelho) se H2:H200 sumir do workbook, do mesmo
    jeito que _validar_estrutura_itens_pc ja protege itens_PC."""
    wb = load_workbook(TEMPLATE_COLETA_OFICIAL, data_only=False)
    _garantir_apresentacao_retroativos_e_aditivos(wb)
    _validar_validacoes_aditivos_criticas(wb)  # fluxo correto: nao levanta

    ws = wb["aditivos"]
    for dv in list(ws.data_validations.dataValidation):
        if "H2:H200" in str(dv.sqref):
            ws.data_validations.dataValidation.remove(dv)
    with pytest.raises(ValueError, match="aditivos invalida"):
        _validar_validacoes_aditivos_criticas(wb)
    wb.close()


def test_l_assinatura_codigo_coleta_estavel_e_sensivel_ao_conteudo(tmp_path):
    """Prova o mecanismo da segunda chave de cache: uma correcao neste
    modulo (mesmo sem tocar no template) muda a assinatura, o que invalida
    o cache do download em producao no proximo deploy — sem depender de
    reboot manual (causa raiz da recorrencia investigada nesta tarefa)."""
    a = assinatura_codigo_coleta()
    b = assinatura_codigo_coleta()
    assert a == b
    assert len(a) == 64

    modulo_original = Path(_coleta_oficial.__file__)
    copia = tmp_path / "_coleta_oficial_copia.py"
    conteudo = bytearray(modulo_original.read_bytes())
    conteudo.append(ord("\n"))
    copia.write_bytes(bytes(conteudo))

    original_file = _coleta_oficial.__file__
    try:
        _coleta_oficial.__file__ = str(copia)
        assert assinatura_codigo_coleta() != a
    finally:
        _coleta_oficial.__file__ = original_file
