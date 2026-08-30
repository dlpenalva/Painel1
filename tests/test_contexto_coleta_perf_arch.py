"""PERF-ARCH-1 — contrato arquitetural do contexto de leitura da Coleta.

O fluxo oficial abria QUATRO vezes o mesmo XLSX por upload. Estes testes
travam as propriedades que sustentam a reducao para DUAS aberturas — uma por
representacao semantica — sem afrouxar a fronteira XSEC-09:

  * o fluxo oficial materializa cada representacao no maximo uma vez;
  * os leitores recebem o MESMO objeto, e nao copias;
  * leitor chamado isoladamente (sem contexto) continua abrindo os proprios
    bytes, como sempre fez;
  * o workbook compartilhado nao e mutado por nenhum leitor;
  * a geometria e validada ANTES de qualquer acesso a celula;
  * a representacao data_only=False continua disponivel e continua sendo
    formula (o contexto nao colapsa as duas semanticas);
  * o caso "arquivo sem cache do Excel" continua sendo detectado;
  * nada e compartilhado ENTRE execucoes.
"""
from __future__ import annotations

import io
from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook

import _coleta_reajuste
import _coleta_reajuste_documentos
import _contexto_coleta
import _leitor_masterfile_v10
from _coleta_reajuste_documentos import processar_coleta_oficial_runtime
from _contexto_coleta import ContextoColeta
from _leitor_masterfile_v10 import ler_masterfile_v10
from _seguranca_xlsx import ErroSegurancaXlsx, XlsxInvalidoError

ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "templates" / "COLETA_REAJUSTE_OFICIAL.xlsx"

MODULOS_COM_LOAD = (
    _contexto_coleta,
    _leitor_masterfile_v10,
    _coleta_reajuste,
    _coleta_reajuste_documentos,
)


def _coleta_oficial() -> bytes:
    """Coleta oficial minima e valida, montada a partir do template real."""
    wb = load_workbook(TEMPLATE, data_only=False)
    parametros = wb["parametros"]
    janelas = [
        (date(2023, 1, 1), date(2023, 12, 31)),
        (date(2024, 1, 1), date(2024, 12, 31)),
        (date(2025, 1, 1), date(2025, 12, 31)),
        (date(2026, 1, 1), date(2026, 12, 31)),
    ]
    for indice, (inicio, fim) in enumerate(janelas):
        parametros.cell(indice + 2, 3).value = inicio
        parametros.cell(indice + 2, 4).value = fim
    controle = wb["CONTROLE"]
    controle["B1"] = "Pedidos de Compras"
    controle["B2"] = "C3"
    remanescentes = wb["itens_Remanesc"]
    remanescentes.cell(2, 1).value = "ITEM-1"
    remanescentes.cell(2, 2).value = 100.0
    remanescentes.cell(2, 3).value = 10.0
    for coluna in (5, 7, 9):
        remanescentes.cell(2, coluna).value = 100.0
    pedidos = wb["itens_PC"]
    pedidos.cell(2, 1).value = "PC-1"
    pedidos.cell(2, 2).value = date(2026, 5, 20)
    pedidos.cell(2, 4).value = 2000.0
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


@pytest.fixture(scope="module")
def coleta() -> bytes:
    return _coleta_oficial()


class _RegistroAberturas:
    """Conta as chamadas reais a load_workbook nos modulos do fluxo."""

    def __init__(self) -> None:
        self.chamadas: list[dict] = []

    def instalar(self, monkeypatch) -> "_RegistroAberturas":
        for modulo in MODULOS_COM_LOAD:
            original = modulo.load_workbook

            def espiao(*args, _original=original, _modulo=modulo.__name__, **kwargs):
                self.chamadas.append(
                    {
                        "modulo": _modulo,
                        "data_only": bool(kwargs.get("data_only", False)),
                        "read_only": bool(kwargs.get("read_only", False)),
                    }
                )
                return _original(*args, **kwargs)

            monkeypatch.setattr(modulo, "load_workbook", espiao)
        return self

    @property
    def total(self) -> int:
        return len(self.chamadas)

    def por_representacao(self) -> dict[bool, int]:
        contagem: dict[bool, int] = {}
        for chamada in self.chamadas:
            contagem[chamada["data_only"]] = contagem.get(chamada["data_only"], 0) + 1
        return contagem


# --------------------------------------------------------------------------- #
# 1. Quantas materializacoes o fluxo oficial faz.
# --------------------------------------------------------------------------- #
def test_fluxo_oficial_abre_uma_vez_por_representacao(monkeypatch, coleta):
    registro = _RegistroAberturas().instalar(monkeypatch)

    processar_coleta_oficial_runtime(coleta)

    # Duas aberturas: uma data_only=True (valores em cache) e uma
    # data_only=False (formulas). Semanticas incompativeis continuam separadas;
    # o que sumiu foi a REPETICAO de cada uma delas.
    assert registro.total == 2, registro.chamadas
    assert registro.por_representacao() == {True: 1, False: 1}
    assert {c["modulo"] for c in registro.chamadas} == {"_contexto_coleta"}


# --------------------------------------------------------------------------- #
# 2. O contexto entrega o MESMO objeto, e nao copias.
# --------------------------------------------------------------------------- #
def test_contexto_reaproveita_o_mesmo_objeto(coleta):
    with ContextoColeta(coleta) as contexto:
        assert contexto.workbook_valores is contexto.workbook_valores
        assert contexto.workbook_formulas is contexto.workbook_formulas
        assert contexto.workbook_valores is not contexto.workbook_formulas


def test_leitores_recebem_o_workbook_do_contexto(monkeypatch, coleta):
    vistos: list[int] = []
    original = _leitor_masterfile_v10._ler_posicao_contratual

    def espiao(wb, *args, **kwargs):
        vistos.append(id(wb))
        return original(wb, *args, **kwargs)

    monkeypatch.setattr(_leitor_masterfile_v10, "_ler_posicao_contratual", espiao)

    with ContextoColeta(coleta) as contexto:
        esperado = id(contexto.workbook_valores)
        ler_masterfile_v10(coleta, exigir_modelo_oficial=True, contexto=contexto)

    assert vistos == [esperado]


# --------------------------------------------------------------------------- #
# 3. Chamada isolada continua abrindo os proprios bytes.
# --------------------------------------------------------------------------- #
def test_leitor_isolado_sem_contexto_abre_os_proprios_bytes(monkeypatch, coleta):
    registro = _RegistroAberturas().instalar(monkeypatch)

    resultado = ler_masterfile_v10(coleta, exigir_modelo_oficial=True)

    assert resultado["ok"]
    assert registro.total >= 1
    assert all(c["modulo"] != "_contexto_coleta" for c in registro.chamadas)


def test_ler_coleta_reajuste_isolado_continua_valido(coleta):
    diagnostico = _coleta_reajuste.ler_coleta_reajuste(coleta)
    explicito = _coleta_reajuste.ler_coleta_reajuste(coleta, contexto=None)
    assert diagnostico["valido"] == explicito["valido"]
    assert diagnostico["contagens"] == explicito["contagens"]


def test_varredura_de_erros_do_excel_independe_do_contexto(coleta):
    """A varredura de #REF!/#VALUE! trocou de workbook: prova a equivalencia.

    Antes, os erros salvos pelo Excel eram procurados em um workbook proprio,
    aberto em read_only e percorrido por streaming. Agora sao procurados no
    workbook data_only compartilhado. Nenhuma Coleta real do corpus carrega
    erro de calculo, entao este e o unico teste que exercita o ramo.
    """
    wb = load_workbook(io.BytesIO(coleta), data_only=False)
    wb["itens_Consumidos"]["A199"] = "#REF!"
    buffer = io.BytesIO()
    wb.save(buffer)
    com_erro = buffer.getvalue()

    isolado = _coleta_reajuste.ler_coleta_reajuste(com_erro)
    with ContextoColeta(com_erro) as contexto:
        compartilhado = _coleta_reajuste.ler_coleta_reajuste(
            com_erro, contexto=contexto
        )

    assert any(
        "O Excel salvou erros de cálculo" in item
        for item in isolado["lacunas_apuracao"]
    ), isolado["lacunas_apuracao"]
    assert isolado["lacunas_apuracao"] == compartilhado["lacunas_apuracao"]
    assert isolado["bloqueios_criticos"] == compartilhado["bloqueios_criticos"]
    assert isolado["contagens"] == compartilhado["contagens"]
    assert isolado["metadados"] == compartilhado["metadados"]


# --------------------------------------------------------------------------- #
# 4. O workbook compartilhado nao e mutado.
# --------------------------------------------------------------------------- #
def test_workbook_compartilhado_nao_e_mutado(monkeypatch, coleta):
    from openpyxl.cell.cell import Cell

    from _coleta_reajuste import ler_coleta_reajuste
    from _coleta_reajuste_documentos import adaptar_coleta_reajuste_para_documentos

    escritas: list[str] = []

    with ContextoColeta(coleta) as contexto:
        # A vigilancia so e armada DEPOIS das aberturas: o proprio parser do
        # openpyxl monta as abas e as celulas do workbook, e essa construcao
        # nao e mutacao pelos leitores.
        valores, formulas = contexto.workbook_valores, contexto.workbook_formulas
        abas_antes = (list(valores.sheetnames), list(formulas.sheetnames))
        propriedade = Cell.value
        setter_original = propriedade.fset

        def setter_vigiado(self, valor):
            escritas.append(f"{getattr(self.parent, 'title', '?')}!{self.coordinate}")
            setter_original(self, valor)

        monkeypatch.setattr(Cell, "value", property(propriedade.fget, setter_vigiado))

        leitura = ler_masterfile_v10(
            coleta, exigir_modelo_oficial=True, contexto=contexto
        )
        diagnostico = ler_coleta_reajuste(coleta, contexto=contexto)
        adaptar_coleta_reajuste_para_documentos(
            coleta, leitura=leitura, diagnostico=diagnostico, contexto=contexto
        )

        assert escritas == []
        assert (list(valores.sheetnames), list(formulas.sheetnames)) == abas_antes


# --------------------------------------------------------------------------- #
# 5. XSEC-09: geometria validada antes de qualquer acesso a celula.
# --------------------------------------------------------------------------- #
def test_geometria_validada_antes_do_primeiro_acesso(monkeypatch, coleta):
    from openpyxl.workbook.workbook import Workbook

    eventos: list[str] = []
    gate_original = _contexto_coleta.validar_geometria_workbook
    getitem_original = Workbook.__getitem__

    def gate(wb):
        eventos.append("gate")
        return gate_original(wb)

    def getitem(self, chave):
        eventos.append("acesso")
        return getitem_original(self, chave)

    monkeypatch.setattr(_contexto_coleta, "validar_geometria_workbook", gate)
    monkeypatch.setattr(Workbook, "__getitem__", getitem)

    processar_coleta_oficial_runtime(coleta)

    assert eventos, "nenhum evento registrado"
    # O primeiro evento do fluxo e o gate, nunca um acesso a aba.
    assert eventos[0] == "gate"
    # Um gate por representacao materializada.
    assert eventos.count("gate") == 2


def test_contexto_rejeita_geometria_fora_do_orcamento(monkeypatch, coleta):
    def gate_reprovado(_wb):
        raise ErroSegurancaXlsx("O arquivo excede os limites de segurança permitidos.")

    monkeypatch.setattr(_contexto_coleta, "validar_geometria_workbook", gate_reprovado)

    contexto = ContextoColeta(coleta)
    try:
        with pytest.raises(ErroSegurancaXlsx):
            contexto.workbook_valores
    finally:
        contexto.fechar()


def test_runtime_preserva_fronteira_para_bytes_invalidos():
    with pytest.raises(XlsxInvalidoError):
        processar_coleta_oficial_runtime(b"nao sou um xlsx")


# --------------------------------------------------------------------------- #
# 6. data_only e semantica: o contexto nao colapsa as duas representacoes.
# --------------------------------------------------------------------------- #
def test_representacoes_permanecem_distintas(coleta):
    with ContextoColeta(coleta) as contexto:
        formula = contexto.workbook_formulas["itens_Remanesc"]["F2"].value
        cache = contexto.workbook_valores["itens_Remanesc"]["F2"].value
    assert isinstance(formula, str) and formula.startswith("=")
    assert not (isinstance(cache, str) and cache.startswith("="))


# --------------------------------------------------------------------------- #
# 7. Caso "sem cache do Excel" continua detectado.
# --------------------------------------------------------------------------- #
def test_cache_ausente_continua_detectado(coleta):
    # A Coleta do fixture nunca passou pelo Excel: as formulas de
    # posicao_contratual nao tem valor calculado.
    isolada = ler_masterfile_v10(coleta, exigir_modelo_oficial=True)
    with ContextoColeta(coleta) as contexto:
        com_contexto = ler_masterfile_v10(
            coleta, exigir_modelo_oficial=True, contexto=contexto
        )
    assert com_contexto["posicao_contratual"]["cache_ausente"] is True
    assert (
        com_contexto["posicao_contratual"]["cache_ausente"]
        == isolada["posicao_contratual"]["cache_ausente"]
    )


# --------------------------------------------------------------------------- #
# 8. Nada e compartilhado entre execucoes.
# --------------------------------------------------------------------------- #
def test_contexto_nao_sobrevive_entre_execucoes(coleta):
    primeiro = ContextoColeta(coleta)
    identidade = id(primeiro.workbook_valores)
    primeiro.fechar()
    segundo = ContextoColeta(coleta)
    try:
        assert id(segundo.workbook_valores) != identidade
    finally:
        segundo.fechar()


def test_fechar_e_idempotente(coleta):
    contexto = ContextoColeta(coleta)
    contexto.workbook_valores
    contexto.workbook_formulas
    contexto.fechar()
    contexto.fechar()


def test_execucoes_repetidas_produzem_o_mesmo_resultado(coleta):
    primeiro, _ = processar_coleta_oficial_runtime(coleta)
    segundo, _ = processar_coleta_oficial_runtime(coleta)
    for chave in (
        "valor_atualizado_contrato",
        "valor_represado_a_pagar",
        "remanescente_reajustado",
        "ciclo_ultimo_remanescente",
        "fator_acumulado",
    ):
        assert primeiro[chave] == segundo[chave], chave
