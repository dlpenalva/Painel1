"""Testes para a fórmula nested-IF de enquadramento de ciclo no template.

Item 2.1 — substitui LOOKUP por nested-IF robusto em financeiro!B e itens_PC!C.

Estrutura:
- Testes estáticos (sem COM): verificam texto da fórmula no template.
- Testes de integração COM (RUN_EXCEL_INTEGRATION=1): 8 cenários funcionais.

Cenários COM (espelho da decisão do usuário):
1. C1+C2 somente (C3/C4 vazios) — o bug original: C2 não pode ter 14 meses
2. C1-C3 (C4 vazio)
3. C1-C4 completo
4. Fronteira exata de início (competência == DATA_INICIO)
5. Fronteira exata de fim (competência == DATA_FIM)
6. Gap entre ciclos (competência fora de todos os intervalos)
7. Fora de todos os ciclos (anterior a C0)
8. itens_PC — CICLO_PC com DATA_PC em maiúsculas
"""
from __future__ import annotations

import gc
import shutil
from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "templates" / "COLETA_REAJUSTE_OFICIAL.xlsx"

pytestmark_com = pytest.mark.skipif(
    __import__("os").environ.get("RUN_EXCEL_INTEGRATION") != "1",
    reason="defina RUN_EXCEL_INTEGRATION=1 para executar o Excel COM",
)


# ------------------------------------------------------------------
# Testes estáticos — sem Excel COM
# ------------------------------------------------------------------

class TestFormulaEstatica:
    """Verifica o texto da fórmula diretamente no template (sem recalcular)."""

    def _formula_b(self, row: int = 2) -> str:
        wb = load_workbook(TEMPLATE, data_only=False)
        return wb["financeiro"][f"B{row}"].value or ""

    def _formula_c_pc(self, row: int = 2) -> str:
        wb = load_workbook(TEMPLATE, data_only=False)
        return wb["itens_PC"][f"C{row}"].value or ""

    def test_financeiro_b2_nao_usa_lookup(self):
        assert "LOOKUP" not in self._formula_b().upper()

    def test_financeiro_b2_usa_isnumber(self):
        assert "ISNUMBER" in self._formula_b().upper()

    def test_financeiro_b2_cobre_cinco_ciclos(self):
        formula = self._formula_b().upper()
        for i in range(5):
            assert f"$C${i + 2}" in formula, f"C{i} ausente na formula"

    def test_financeiro_b2_retorna_minusculas(self):
        formula = self._formula_b()
        for label in ('"c0"', '"c1"', '"c2"', '"c3"', '"c4"'):
            assert label in formula, f"{label} ausente na formula"

    def test_financeiro_b2_retorna_fora_dos_ciclos(self):
        assert '"Fora dos ciclos"' in self._formula_b()

    def test_financeiro_b73_mesma_estrutura(self):
        formula = self._formula_b(73)
        assert "ISNUMBER" in formula.upper()
        assert "LOOKUP" not in formula.upper()
        assert "A73" in formula

    def test_financeiro_b74_total_valor(self):
        wb = load_workbook(TEMPLATE, data_only=False)
        assert wb["financeiro"]["B74"].value == "TOTAL"

    def test_financeiro_c74_sum(self):
        wb = load_workbook(TEMPLATE, data_only=False)
        formula = wb["financeiro"]["C74"].value or ""
        assert "SUM" in formula.upper()
        assert "C2:C73" in formula

    def test_financeiro_e74_sum(self):
        wb = load_workbook(TEMPLATE, data_only=False)
        formula = wb["financeiro"]["E74"].value or ""
        assert "SUM" in formula.upper()
        assert "E2:E73" in formula

    def test_financeiro_f74_sum(self):
        wb = load_workbook(TEMPLATE, data_only=False)
        formula = wb["financeiro"]["F74"].value or ""
        assert "SUM" in formula.upper()
        assert "F2:F73" in formula

    def test_financeiro_d74_vazio(self):
        wb = load_workbook(TEMPLATE, data_only=False)
        val = wb["financeiro"]["D74"].value
        assert val in (None, ""), f"D74 nao deveria ter valor: {repr(val)}"

    def test_itens_pc_c2_nao_usa_lookup(self):
        assert "LOOKUP" not in self._formula_c_pc().upper()

    def test_itens_pc_c2_usa_isnumber(self):
        assert "ISNUMBER" in self._formula_c_pc().upper()

    def test_itens_pc_c2_retorna_maiusculas(self):
        formula = self._formula_c_pc()
        for label in ('"C0"', '"C1"', '"C2"', '"C3"', '"C4"'):
            assert label in formula, f"{label} ausente na formula de itens_PC"

    def test_itens_pc_c2_detecta_data_invalida(self):
        assert "DATA_PC invalida" in self._formula_c_pc()

    def test_itens_pc_c100_mesma_estrutura(self):
        formula = self._formula_c_pc(100)
        assert "ISNUMBER" in formula.upper()
        assert "LOOKUP" not in formula.upper()
        assert "B100" in formula


# ------------------------------------------------------------------
# Helpers para testes COM
# ------------------------------------------------------------------

def _recalcular_e_ler(caminho: Path, inspecionar):
    import time
    import pythoncom
    import win32com.client

    RPC_E_CALL_REJECTED = -2147418111

    def _tentar(fn, tentativas: int = 5, espera: float = 0.4):
        for i in range(tentativas):
            try:
                return fn()
            except Exception as exc:
                codigo = getattr(exc, "hresult", None)
                if codigo == RPC_E_CALL_REJECTED and i < tentativas - 1:
                    time.sleep(espera * (i + 1))
                    continue
                raise

    pythoncom.CoInitialize()
    excel = win32com.client.DispatchEx("Excel.Application")
    for attr, val in (("Visible", False), ("DisplayAlerts", False)):
        try:
            setattr(excel, attr, val)
        except AttributeError:
            pass

    pasta = None
    try:
        pasta = _tentar(lambda: excel.Workbooks.Open(str(caminho.resolve()), UpdateLinks=0))
        _tentar(excel.CalculateFullRebuild)
        time.sleep(0.3)
        resultado = _tentar(lambda: inspecionar(pasta))
        pasta.Close(SaveChanges=False)
        pasta = None
        return resultado
    finally:
        if pasta is not None:
            try:
                pasta.Close(SaveChanges=False)
            except Exception:
                pass
        try:
            excel.Quit()
        except Exception:
            pass
        gc.collect()
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass


def _preparar_template(tmp_path: Path) -> Path:
    destino = tmp_path / "COLETA_REAJUSTE_OFICIAL.xlsx"
    shutil.copy2(TEMPLATE, destino)
    return destino


def _escrever_parametros_e_competencias(
    caminho: Path,
    ciclos: list[tuple],
    competencias: list[date],
) -> None:
    """Escreve ciclos em parametros (linhas 2-6) e competências em financeiro!A."""
    wb = load_workbook(caminho, data_only=False)
    ws_p = wb["parametros"]
    ws_f = wb["financeiro"]
    for num, inicio, fim in ciclos:
        lin = num + 2
        ws_p.cell(lin, 1).value = "Sim" if num > 0 else "Base"
        ws_p.cell(lin, 3).value = inicio
        ws_p.cell(lin, 4).value = fim
    for idx, comp in enumerate(competencias, start=2):
        ws_f.cell(idx, 1).value = comp
    wb.save(caminho)


# ------------------------------------------------------------------
# Cenário 1: C1+C2 somente — bug original
# ------------------------------------------------------------------

@pytestmark_com
def test_cenario1_c1_c2_somente_sem_bleed(tmp_path: Path) -> None:
    caminho = _preparar_template(tmp_path)
    ciclos = [
        (0, date(2023, 1, 1), date(2023, 12, 31)),
        (1, date(2024, 1, 1), date(2024, 12, 31)),
        (2, date(2025, 1, 1), date(2025, 12, 31)),
    ]
    competencias = (
        [date(2024, m, 1) for m in range(1, 13)]
        + [date(2025, m, 1) for m in range(1, 13)]
    )
    _escrever_parametros_e_competencias(caminho, ciclos, competencias)

    def inspecionar(pasta):
        ws = pasta.Worksheets("financeiro")
        return [ws.Cells(row, 2).Value for row in range(2, 26)]

    resultados = _recalcular_e_ler(caminho, inspecionar)
    assert resultados.count("c1") == 12, f"C1 deve ter 12: {resultados}"
    assert resultados.count("c2") == 12, f"C2 deve ter 12: {resultados}"
    assert "Fora dos ciclos" not in resultados


# ------------------------------------------------------------------
# Cenário 2: C1-C3 (C4 vazio)
# ------------------------------------------------------------------

@pytestmark_com
def test_cenario2_c1_c3_c4_vazio(tmp_path: Path) -> None:
    caminho = _preparar_template(tmp_path)
    ciclos = [
        (0, date(2023, 1, 1), date(2023, 12, 31)),
        (1, date(2024, 1, 1), date(2024, 12, 31)),
        (2, date(2025, 1, 1), date(2025, 12, 31)),
        (3, date(2026, 1, 1), date(2026, 12, 31)),
    ]
    competencias = [
        date(2024, m, 1) for m in range(1, 13)
    ] + [
        date(2025, m, 1) for m in range(1, 13)
    ] + [
        date(2026, m, 1) for m in range(1, 13)
    ]
    _escrever_parametros_e_competencias(caminho, ciclos, competencias)

    def inspecionar(pasta):
        ws = pasta.Worksheets("financeiro")
        return [ws.Cells(row, 2).Value for row in range(2, 38)]

    resultados = _recalcular_e_ler(caminho, inspecionar)
    assert resultados.count("c1") == 12
    assert resultados.count("c2") == 12
    assert resultados.count("c3") == 12
    assert resultados.count("c4") == 0
    assert "Fora dos ciclos" not in resultados


# ------------------------------------------------------------------
# Cenário 3: C1-C4 completo
# ------------------------------------------------------------------

@pytestmark_com
def test_cenario3_c1_c4_completo(tmp_path: Path) -> None:
    caminho = _preparar_template(tmp_path)
    ciclos = [
        (0, date(2023, 1, 1), date(2023, 12, 31)),
        (1, date(2024, 1, 1), date(2024, 12, 31)),
        (2, date(2025, 1, 1), date(2025, 12, 31)),
        (3, date(2026, 1, 1), date(2026, 12, 31)),
        (4, date(2027, 1, 1), date(2027, 12, 31)),
    ]
    competencias = [date(2024 + (m // 12), (m % 12) + 1, 1) for m in range(48)]
    _escrever_parametros_e_competencias(caminho, ciclos, competencias)

    def inspecionar(pasta):
        ws = pasta.Worksheets("financeiro")
        return [ws.Cells(row, 2).Value for row in range(2, 50)]

    resultados = _recalcular_e_ler(caminho, inspecionar)
    assert resultados.count("c1") == 12
    assert resultados.count("c2") == 12
    assert resultados.count("c3") == 12
    assert resultados.count("c4") == 12
    assert "Fora dos ciclos" not in resultados


# ------------------------------------------------------------------
# Cenário 4: Fronteira exata de início
# ------------------------------------------------------------------

@pytestmark_com
def test_cenario4_fronteira_exata_inicio(tmp_path: Path) -> None:
    caminho = _preparar_template(tmp_path)
    inicio_c1 = date(2024, 3, 1)
    ciclos = [
        (0, date(2023, 3, 1), date(2024, 2, 28)),
        (1, inicio_c1, date(2025, 2, 28)),
    ]
    _escrever_parametros_e_competencias(caminho, ciclos, [inicio_c1])

    def inspecionar(pasta):
        return pasta.Worksheets("financeiro").Cells(2, 2).Value

    assert _recalcular_e_ler(caminho, inspecionar) == "c1"


# ------------------------------------------------------------------
# Cenário 5: Fronteira exata de fim
# ------------------------------------------------------------------

@pytestmark_com
def test_cenario5_fronteira_exata_fim(tmp_path: Path) -> None:
    caminho = _preparar_template(tmp_path)
    fim_c1 = date(2025, 2, 28)
    ciclos = [
        (0, date(2023, 3, 1), date(2024, 2, 28)),
        (1, date(2024, 3, 1), fim_c1),
    ]
    _escrever_parametros_e_competencias(caminho, ciclos, [fim_c1])

    def inspecionar(pasta):
        return pasta.Worksheets("financeiro").Cells(2, 2).Value

    assert _recalcular_e_ler(caminho, inspecionar) == "c1"


# ------------------------------------------------------------------
# Cenário 6: Gap entre ciclos
# ------------------------------------------------------------------

@pytestmark_com
def test_cenario6_gap_entre_ciclos(tmp_path: Path) -> None:
    caminho = _preparar_template(tmp_path)
    ciclos = [
        (0, date(2023, 1, 1), date(2023, 12, 31)),
        (1, date(2024, 1, 1), date(2024, 6, 30)),
        (2, date(2024, 9, 1), date(2025, 8, 31)),
    ]
    competencias = [date(2024, 7, 1), date(2024, 8, 1)]
    _escrever_parametros_e_competencias(caminho, ciclos, competencias)

    def inspecionar(pasta):
        ws = pasta.Worksheets("financeiro")
        return [ws.Cells(row, 2).Value for row in (2, 3)]

    resultados = _recalcular_e_ler(caminho, inspecionar)
    assert all(v == "Fora dos ciclos" for v in resultados), f"Gap deve ser 'Fora dos ciclos': {resultados}"


# ------------------------------------------------------------------
# Cenário 7: Anterior a todos os ciclos
# ------------------------------------------------------------------

@pytestmark_com
def test_cenario7_anterior_a_c0(tmp_path: Path) -> None:
    caminho = _preparar_template(tmp_path)
    ciclos = [
        (0, date(2023, 1, 1), date(2023, 12, 31)),
        (1, date(2024, 1, 1), date(2024, 12, 31)),
    ]
    _escrever_parametros_e_competencias(caminho, ciclos, [date(2022, 6, 1)])

    def inspecionar(pasta):
        return pasta.Worksheets("financeiro").Cells(2, 2).Value

    assert _recalcular_e_ler(caminho, inspecionar) == "Fora dos ciclos"


# ------------------------------------------------------------------
# Cenário 8: itens_PC — CICLO_PC em maiúsculas
# ------------------------------------------------------------------

@pytestmark_com
def test_cenario8_itens_pc_ciclo_maiusculas(tmp_path: Path) -> None:
    caminho = _preparar_template(tmp_path)
    ciclos = [
        (0, date(2023, 1, 1), date(2023, 12, 31)),
        (1, date(2024, 1, 1), date(2024, 12, 31)),
        (2, date(2025, 1, 1), date(2025, 12, 31)),
    ]
    wb = load_workbook(caminho, data_only=False)
    ws_p = wb["parametros"]
    ws_ipc = wb["itens_PC"]
    for num, inicio, fim in ciclos:
        lin = num + 2
        ws_p.cell(lin, 1).value = "Sim" if num > 0 else "Base"
        ws_p.cell(lin, 3).value = inicio
        ws_p.cell(lin, 4).value = fim
    ws_ipc.cell(2, 1).value = "PC-001"
    ws_ipc.cell(2, 2).value = date(2024, 6, 1)   # dentro de C1
    ws_ipc.cell(3, 1).value = "PC-002"
    ws_ipc.cell(3, 2).value = date(2025, 3, 1)   # dentro de C2
    ws_ipc.cell(4, 1).value = "PC-003"
    ws_ipc.cell(4, 2).value = date(2022, 1, 1)   # fora de todos
    wb.save(caminho)

    def inspecionar(pasta):
        ws = pasta.Worksheets("itens_PC")
        return [ws.Cells(row, 3).Value for row in (2, 3, 4)]

    resultados = _recalcular_e_ler(caminho, inspecionar)
    assert resultados[0] == "C1", f"DATA_PC em C1 deve retornar 'C1': {resultados}"
    assert resultados[1] == "C2", f"DATA_PC em C2 deve retornar 'C2': {resultados}"
    assert resultados[2] == "Fora dos ciclos", f"DATA_PC anterior a C0: {resultados}"
