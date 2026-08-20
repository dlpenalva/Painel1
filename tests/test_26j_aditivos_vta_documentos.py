"""Etapa 26J: propagacao de aditivos, VTA canonico e apresentacao humana."""
from __future__ import annotations

from datetime import date
from io import BytesIO

import pytest
from openpyxl import load_workbook

from _coleta_oficial import TEMPLATE_COLETA_OFICIAL, obter_coleta_oficial_bytes
from _leitor_masterfile_v10 import _col, _mapear_colunas_por_cabecalho
from _motor_composicao_vta import _aditivos
from _objeto_processo_reajuste import _montar_memoria_por_ciclo
from _sumario_executivo import _montar_secao_aditivos, _montar_secao_itens
from _templates_documentos import (
    _adicionar_tabela,
    _composicao_didatica_vta,
    _sintese_aditivos_por_ciclo,
    montar_historico_vu_documental,
)
from docx import Document


@pytest.fixture(scope="module")
def wb_template():
    wb = load_workbook(TEMPLATE_COLETA_OFICIAL, data_only=False)
    yield wb
    wb.close()


@pytest.fixture(scope="module")
def wb_runtime():
    wb = load_workbook(BytesIO(obter_coleta_oficial_bytes()), data_only=False)
    yield wb
    wb.close()


def test_itens_pc_apresenta_reconhecido_e_potencial_com_cores_opostas(wb_runtime):
    ws = wb_runtime["itens_PC"]
    assert ws["Q1"].value == "RETROATIVO RECONHECIDO"
    assert ws["R1"].value == "VALOR_ATUALIZADO_EM_ANALISE"
    assert ws["S1"].value == "RETROATIVO POTENCIAL"
    assert ws["Q2"].fill.fgColor.rgb == "FFC6EFCE"
    assert ws["R2"].fill.fgColor.rgb == "FFF2F2F2"
    assert ws["S2"].fill.fgColor.rgb == "FFFFEB9C"


def test_leitor_aceita_rotulos_tecnicos_antigos_e_novos(wb_template, wb_runtime):
    mapa_antigo = _mapear_colunas_por_cabecalho(wb_template["itens_PC"])
    mapa_novo = _mapear_colunas_por_cabecalho(wb_runtime["itens_PC"])
    for mapa in (mapa_antigo, mapa_novo):
        assert _col(
            mapa, "RETROATIVO_RECONHECIDO_A_PAGAR", "RETROATIVO RECONHECIDO"
        ) == 8
        assert _col(mapa, "DELTA_POTENCIAL", "RETROATIVO POTENCIAL") == 10


def test_aditivos_k_preserva_entrada_e_dropdown_do_template(wb_runtime):
    ws = wb_runtime["aditivos"]
    assert ws.column_dimensions["K"].hidden is False
    assert ws["K2"].value is None
    assert any("K2" in str(dv.sqref) for dv in ws.data_validations.dataValidation)


def test_template_integra_delta_no_remanescente_uma_vez(wb_template):
    pos = wb_template["posicao_contratual"]
    assert '$Y2>0,""' in pos["G2"].value
    assert "ROUND(J2+H2,2)" in pos["K2"].value
    assert 'IF($Y2=1,ROUND(H2,2),"")' in pos["K2"].value
    assert "ROUND(N2+L2,2)" in pos["O2"].value
    assert "ROUND(R2+P2,2)" in pos["S2"].value
    assert "ROUND(V2+T2,2)" in pos["W2"].value


def test_template_pre_nascimento_vazio_e_consumidores_canonicos(wb_template):
    hist = wb_template["historico_VU"]
    rc = wb_template["itens_RC"]
    assert "posicao_contratual!$Y2>0" in hist["N2"].value
    assert "historico_VU!C2" in rc["B3"].value
    assert "historico_VU!D2" in rc["E3"].value
    assert 'OR(A3="",B3="",C3="")' in rc["D3"].value


def test_historico_documental_remove_descricao_e_preserva_vu_pre_nascimento():
    dados = {
        "historico_vu": {
            "ciclos": ["C0", "C1"],
            "ultimo_ciclo": "C1",
            "itens": [{
                "item": "N002",
                "descricao": "nao deve aparecer",
                "vus": {"C0": None, "C1": 7031.62},
            }],
        }
    }
    quadro = montar_historico_vu_documental(dados)
    assert quadro["cabecalhos"] == ["Item", "VU_C0", "VU_C1"]
    assert quadro["linhas"] == [["N002", "", "R$ 7.031,62"]]


def test_itens_por_ciclo_nao_inventam_quantidade_vu_ou_total():
    memoria = {
        "vu_itens": [{
            "item": "N002",
            "descricao": None,
            "quantidade_contratada": 945,
            "vu_original": 7031.62,
            "vu_por_ciclo": {"C0": None, "C1": 7031.62},
            "quantidade_por_ciclo": {"C0": None, "C1": 945},
            "total_por_ciclo": {"C0": None, "C1": 6644880.90},
            "vu_ciclos": {"C1": 7031.62},
        }]
    }
    parametros = {
        "por_ciclo": {
            "C0": {"fator_acumulado": 1.0},
            "C1": {"fator_acumulado": 1.03},
        }
    }
    item = _montar_secao_itens(memoria, parametros)[0]
    assert item["quantidade_ciclos"]["C0"] is None
    assert item["vu_c0"] is None
    assert item["total_c0"] is None
    assert item["quantidade_ciclos"]["C1"] == 945
    assert item["vu_ciclos"]["C1"] == 7031.62
    assert item["total_ciclos"]["C1"] == 6644880.90


def test_aditivo_ja_materializado_no_remanescente_nao_soma_novamente():
    leitura = {
        "posicao_contratual": {
            "ok": True, "cache_ausente": False, "itens": [{"ITEM": "N002"}]
        },
        "vta_sombra": {"parcelas_computadas": [{
            "fonte_parcela": "Aditivo",
            "identificador": "aditivos:C1:6",
            "ciclo": "C1",
            "valor": 6849445.64,
        }]},
    }
    computados, fora = _aditivos(leitura, [])
    assert computados == []
    assert len(fora) == 1
    assert fora[0]["ja_refletido_em"] == "posicao_contratual/remanescente"
    assert "nao soma novamente" in fora[0]["motivo"]


def test_aditivo_nao_refletido_preserva_computo_existente():
    leitura = {"aditivos_visiveis": {"ok": True, "itens": [{
        "evento": "Aditivo computavel",
        "ciclo_marco": "C1",
        "valor_assinatura": 100.0,
        "fator_acumulado": 1.05,
        "valor_atualizado": 105.0,
        "ja_refletido_em": "Nao",
    }]}}
    computados, fora = _aditivos(leitura, [])
    assert fora == []
    assert computados[0]["valor_atualizado"] == 105.0


def test_documentos_usam_componentes_canonicos_sem_repetir_aditivos():
    dados = {
        "vta": None,
        "vta_previa": 137375560.29,
        "vta_execucao_atualizada": 13973327.58,
        "vta_saldo_remanescente_atualizado": 123402232.71,
        "parcelas_vta": [{
            "fonte_parcela": "Aditivo",
            "ciclo": "C1",
            "valor": 9614154.78,
        }],
    }
    componentes = _composicao_didatica_vta(dados)
    assert componentes == [
        ("Execução atualizada anterior ao corte", 13973327.58),
        ("Saldo remanescente atualizado no corte", 123402232.71),
    ]
    assert round(sum(valor for _, valor in componentes), 2) == 137375560.29


def test_retroativo_potencial_nao_altera_composicao_do_vta():
    base = {
        "vta": 1000.0,
        "vta_execucao_atualizada": 100.0,
        "vta_saldo_remanescente_atualizado": 900.0,
    }
    com_potencial = dict(
        base,
        situacao_retroativos_pc={
            "reconhecido": 10.0, "em_analise": 110.0, "potencial": 100.0
        },
    )
    assert _composicao_didatica_vta(com_potencial) == _composicao_didatica_vta(base)


def test_tabela_docx_marca_cabecalho_para_repeticao_em_todas_as_paginas():
    doc = Document()
    tabela = _adicionar_tabela(doc, ["Item", "VU_C0"], [["1", "R$ 1,00"]])
    xml = tabela.rows[0]._tr.xml
    assert "w:tblHeader" in xml
    assert 'w:val="true"' in xml


def test_memoria_vta_pc_reusa_composicao_do_mesmo_corte():
    leitura = {
        "controle": {"modo": "PC", "ciclo_vigente": "C1"},
        "parametros_v10": {
            "por_ciclo": {
                "C0": {"fator_acumulado": 1.0},
                "C1": {"fator_acumulado": 1.03},
            }
        },
        "vta_sombra": {"parcelas_computadas": []},
        "itens_consumidos_v10": {"itens": []},
        "execucao_saldo": {"fotografias_ciclo": []},
        "posicao_contratual": {"ok": False, "itens": []},
        "historico_vu": {"itens": []},
        "itens_contrato": {"itens": []},
        "potencial_futuro": {},
        "composicao_vta": {
            "disponivel": True,
            "metodo": "pc",
            "total_execucao_atualizada": 100.0,
            "saldo_remanescente": {"valor_atualizado": 900.0},
            "vta_composicao": 1000.0,
        },
    }
    memoria = _montar_memoria_por_ciclo(leitura, {}, [], {})
    assert memoria["vta"]["metodo"] == "pc"
    assert memoria["vta"]["executado_atualizado"] == 100.0
    assert memoria["vta"]["potencial_restante_atualizado"] == 900.0
    assert memoria["vta"]["valor_total_atualizado"] == 1000.0


def test_rotulo_humano_separa_identificador_interno():
    dados_op = {"vta_sombra": {"parcelas_computadas": [{
        "fonte_parcela": "Aditivo",
        "identificador": "aditivos:C1:6",
        "ciclo": "C1",
        "linha": 6,
        "item": "N002",
        "tipo_alteracao": "Acrescimo",
        "data_aditivo": date(2026, 6, 12),
        "quantidade": 945,
        "valor_original": 6644880.90,
        "valor": 6849445.64,
    }]}}
    item = _montar_secao_aditivos(dados_op, {})["itens"][0]
    assert item["identificador_interno"] == "aditivos:C1:6"
    assert item["rotulo_documental"] == "C1 — Acréscimo — Item N002"
    assert "aditivos:" not in item["rotulo_documental"]
    assert item["data_alteracao"] == "12/06/2026"


def test_sintese_documental_agrega_por_ciclo_sem_chave_tecnica():
    itens = [
        {
            "identificador_interno": "aditivos:C0:2",
            "ciclo": "C0", "tipo_alteracao": "Acréscimo",
            "valor_atualizado": 100.0,
        },
        {
            "identificador_interno": "aditivos:C0:3",
            "ciclo": "C0", "tipo_alteracao": "Supressão",
            "valor_atualizado": -20.0,
        },
    ]
    texto = " ".join(_sintese_aditivos_por_ciclo(itens))
    assert "C0 — 1 acréscimo e 1 supressão" in texto
    assert "R$ 80,00" in texto
    assert "aditivos:C0:" not in texto
