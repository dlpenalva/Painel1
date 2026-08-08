# -*- coding: utf-8 -*-
"""Etapas 30/31 — interregno pelos efeitos financeiros + metodo.

A ETAPA 31 corrigiu a premissa temporal da Etapa 30: NAO existe uma unica
linha temporal. Tres conceitos independentes:

  * JANELA DO REAJUSTE (parametros!C:D): EXATAMENTE 12 competencias por
    ciclo; DATA_INICIO(Cn+1) = COMPETENCIA(H(Cn)) + 12m (fallback teorico
    +12m sem ancora); o retardo NUNCA alonga o proprio ciclo — ele abre um
    INTERVALO INTENCIONAL entre D(Cn) e C(Cn+1), que e PERMITIDO;
  * CRONOLOGIA DA EXECUCAO: blocos fixos de 12 meses da data-base original
    (5 x 12 = 60 competencias, sem lacuna) — enquadra financeiro, PC e
    aditivos; H nunca muda o CICLO;
  * EFEITO FINANCEIRO (parametros!H): gate SOMENTE do novo delta — nunca
    exclui valor executado, quantidade, VTA ou posicao contratual.

Cenarios cobertos:
  A) tempestivo no mesmo mes da data-base -> efeito desde a competencia da
     data-base; proximo reajuste 12 meses depois;
  B) tempestivo em competencia posterior (TEMPESTIVO*) -> efeito desde a
     competencia do pedido; o PROXIMO ciclo nasce 12 meses apos o efeito e
     a janela atual NAO se alonga;
  C) caso real multiciclo: C2 (04/2026..03/2027) com efeito em 06/2026 ->
     C3 = 06/2027..05/2028 e C4 = 06/2028..05/2029; gap intencional
     04-05/2027 nas janelas; a execucao segue enquadrando pela cronologia;
  D) sem retardo -> linha temporal identica ao baseline.

REGRA PETREA — INTERREGNO E EFEITOS (registro permanente):
  * pagamento nunca e marco;
  * admissibilidade usa data exata;
  * efeito financeiro usa competencia mensal;
  * tempestivo no mesmo mes = efeito desde aquela competencia;
  * tempestivo em mes posterior dentro da janela = TEMPESTIVO* e efeito
    desde a competencia do pedido;
  * proximo reajuste = 12 meses apos o inicio dos efeitos financeiros do
    anterior;
  * nenhuma execucao valida fica sem enquadramento: a cronologia da
    execucao nao tem lacuna, ainda que as janelas de reajuste tenham.

REGRA DE COMPATIBILIDADE DO METODO:
  * rotulo novo = "Financeiro (Mensalidade)";
  * codigo interno = "principal";
  * "Principal (Financeiro)" permanece alias legado aceito.

SEPARACAO REGIME x PERIODO DO INDICE:
  * PERIODO DE APURACAO DO INDICE = exatamente 12 competencias, ancorado no
    inicio dos efeitos financeiros do reajuste anterior (ancora..ancora+11m);
  * nenhum motor de indice consome parametros!C:D.
"""
from __future__ import annotations

import inspect
import io
import sys
from datetime import date, timedelta
from pathlib import Path

import pytest
from openpyxl import load_workbook

RAIZ = Path(__file__).resolve().parents[1]
if str(RAIZ) not in sys.path:
    sys.path.insert(0, str(RAIZ))

import _metodo_apuracao as M  # noqa: E402
from _coleta_oficial import (  # noqa: E402
    gerar_coleta_oficial_preenchida,
    normalizar_dados_calculadora,
    obter_coleta_oficial_bytes,
)
from _efeitos_financeiros_pc import efeito_financeiro_pc  # noqa: E402
from _gerador_masterfile import LEGENDA_TEMPESTIVO_RETARDADO  # noqa: E402
from _motor_posicao_contratual import (  # noqa: E402
    determinar_ciclo_por_data,
    divergencias_calendario_canonico,
)
from _motor_temporal import ENQ_CICLO, classificar_enquadramento_pc  # noqa: E402


# --------------------------------------------------------------------------- #
# Cenario A — tempestivo no mesmo mes da data-base
# --------------------------------------------------------------------------- #
def _payload_ciclo_unico(financeiro_inicio: str, retardado: bool = False) -> dict:
    """Data-base do reajuste C1 = 05/03/2025 (ancora 05/03/2024)."""
    return {
        "origem": "Reajuste Simples",
        "indice": "IPCA",
        "data_base_original": "05/03/2024",
        "ciclos": [{
            "ciclo": "C1",
            "data_base": "05/03/2024",
            "data_pedido": "30/03/2025" if not retardado else "01/04/2025",
            "percentual_aplicado": 0.05,
            "financeiro_inicio": financeiro_inicio,
            "objeto_analise_atual": True,
            "situacao": "✅ TEMPESTIVO*" if retardado else "✅ TEMPESTIVO",
            "efeito_financeiro_retardado": retardado,
        }],
    }


def test_cenario_a_mesmo_mes_efeito_na_competencia_da_data_base():
    dados = normalizar_dados_calculadora(_payload_ciclo_unico("01/03/2025"))
    c1 = dados["ciclos"][0]
    assert c1["data_inicio"] == date(2025, 3, 1)
    assert c1["inicio_efeito_financeiro"] == date(2025, 3, 1)
    # Proxima data-base = 12 meses apos o inicio dos efeitos: C2 em 03/2026.
    assert c1["data_fim"] == date(2026, 2, 28)


def test_cenario_b_retardado_efeito_na_competencia_do_pedido():
    dados = normalizar_dados_calculadora(
        _payload_ciclo_unico("01/04/2025", retardado=True)
    )
    c1 = dados["ciclos"][0]
    assert c1["data_inicio"] == date(2025, 3, 1)
    # marco/2025 pertence ao ciclo, mas sem efeito (gate G/H do financeiro).
    assert c1["inicio_efeito_financeiro"] == date(2025, 4, 1)
    assert c1["efeito_financeiro_retardado"] is True
    # ETAPA 31: a janela NUNCA se alonga — 12 competencias exatas
    # (03/2025..02/2026). O retardo desloca o INICIO do proximo ciclo
    # (04/2026), abrindo intervalo intencional em 03/2026.
    assert c1["data_fim"] == date(2026, 2, 28)
    reg = {
        "computar_nesta_apuracao": "Sim",
        "inicio_efeito_financeiro": c1["inicio_efeito_financeiro"],
    }
    assert efeito_financeiro_pc(date(2025, 3, 31), "C1", reg) == "Nao"
    assert efeito_financeiro_pc(date(2025, 4, 1), "C1", reg) == "Sim"


def test_cenario_b_normaliza_dia_exato_do_pedido_para_competencia():
    # Mesmo que o payload traga o dia exato (regra antiga), a fonte canonica
    # normaliza para o primeiro dia do mes.
    dados = normalizar_dados_calculadora(_payload_ciclo_unico("18/04/2025", True))
    assert dados["ciclos"][0]["inicio_efeito_financeiro"] == date(2025, 4, 1)


# --------------------------------------------------------------------------- #
# Cenario C — caso real multiciclo (item 8 do enunciado)
# --------------------------------------------------------------------------- #
def _payload_caso_real() -> dict:
    return {
        "origem": "Reajustes Múltiplos",
        "indice": "IPCA",
        "data_base_original": "04/04/2024",
        "ciclos": [
            {
                "ciclo": "C1",
                "data_base": "04/04/2024",
                "data_pedido": "04/04/2025",
                "percentual_aplicado": 0.05,
                "financeiro_inicio": "04/04/2025",
                "objeto_analise_atual": True,
                "situacao": "✅ TEMPESTIVO",
            },
            {
                "ciclo": "C2",
                "data_base": "04/04/2025",
                "data_pedido": "04/06/2026",
                "percentual_aplicado": 0.04,
                "financeiro_inicio": "04/06/2026",
                "objeto_analise_atual": True,
                "situacao": "✅ TEMPESTIVO*",
                "efeito_financeiro_retardado": True,
            },
        ],
    }


@pytest.fixture(scope="module")
def wb_caso_real():
    wb = load_workbook(
        io.BytesIO(gerar_coleta_oficial_preenchida(_payload_caso_real())),
        data_only=False,
    )
    yield wb
    wb.close()


def _dia(valor):
    return valor.date() if hasattr(valor, "date") else valor


def test_cenario_c_parametros_projeta_c3_e_c4_pela_nova_ancora(wb_caso_real):
    par = wb_caso_real["parametros"]
    assert _dia(par["C4"].value) == date(2026, 4, 1)    # C2 inicia 04/2026
    assert _dia(par["D4"].value) == date(2027, 3, 31)   # 12 competencias EXATAS
    assert _dia(par["H4"].value) == date(2026, 6, 1)    # efeito na competencia
    # Gap intencional 04-05/2027: C3 nasce 12 meses apos o efeito do C2.
    assert _dia(par["C5"].value) == date(2027, 6, 1)    # C3 = 06/2027
    assert _dia(par["D5"].value) == date(2028, 5, 31)   # ...a 05/2028
    assert _dia(par["C6"].value) == date(2028, 6, 1)    # C4 = 06/2028
    assert _dia(par["D6"].value) == date(2029, 5, 31)   # ...a 05/2029


def test_cenario_c_status_tempestivo_asterisco_e_legenda(wb_caso_real):
    par = wb_caso_real["parametros"]
    assert "TEMPESTIVO*" in str(par["G4"].value)
    assert par["A7"].value == LEGENDA_TEMPESTIVO_RETARDADO


def test_cenario_c_grade_cronologica_completa_60_meses(wb_caso_real):
    """Grade do financeiro: SEMPRE a cronologia contratual completa.

    ETAPA 31 (regra petrea): 60 competencias (5 x 12) do marco de C0
    (04/2024) a 03/2029, independentemente de data_corte ou do ultimo ciclo
    analisado. 04/2026 e 05/2026 pertencem ao C2 cronologico sem efeito
    (Nao); competencias futuras (C3/C4 cronologicos, ainda sem reajuste
    valido) permanecem na grade com efeito Nao — futuro nao significa
    ausencia de ciclo. O gap 04-05/2027 existe apenas nas JANELAS de
    parametros, jamais na grade da execucao.
    """
    fin = wb_caso_real["financeiro"]
    grade = {}
    for row in range(2, 74):
        comp = fin[f"A{row}"].value
        if comp is not None:
            grade[(comp.year, comp.month)] = fin[f"G{row}"].value
    assert len(grade) == 60
    assert min(grade) == (2024, 4) and max(grade) == (2029, 3)
    assert grade[(2026, 4)] == "Nao"   # antes do efeito do C2
    assert grade[(2026, 5)] == "Nao"
    assert grade[(2026, 6)] == "Sim"   # efeito inicia 06/2026
    assert grade[(2027, 3)] == "Sim"   # ultima competencia do C2 cronologico
    # Futuro dentro dos 60 meses: pertence ao ciclo cronologico (C3/C4),
    # sem reajuste valido ainda -> efeito Nao (nada inventado).
    assert grade[(2027, 4)] == "Nao"
    assert grade[(2027, 6)] == "Nao"
    assert grade[(2029, 3)] == "Nao"
    # Sequencia mensal continua, sem buracos.
    chaves = sorted(grade)
    for (a, b) in zip(chaves, chaves[1:]):
        seguinte = (a[0] + 1, 1) if a[1] == 12 else (a[0], a[1] + 1)
        assert seguinte == b


def test_cenario_c_calendario_sem_divergencias(wb_caso_real):
    par = wb_caso_real["parametros"]
    ciclos = [
        {
            "ciclo": par[f"B{row}"].value,
            "data_inicio": _dia(par[f"C{row}"].value),
            "data_fim": _dia(par[f"D{row}"].value),
        }
        for row in range(2, 7)
    ]
    assert divergencias_calendario_canonico(ciclos) == []


def _por_ciclo_caso_real(wb) -> dict:
    par = wb["parametros"]
    return {
        par[f"B{row}"].value: {
            "data_inicio": _dia(par[f"C{row}"].value),
            "data_fim": _dia(par[f"D{row}"].value),
        }
        for row in range(2, 7)
    }


def test_fronteira_pc_na_cronologia_da_execucao(wb_caso_real):
    """PC enquadra pela CRONOLOGIA fixa: 31/03/2027 -> C2; 01/04/2027 -> C3.

    ETAPA 31: datas dentro do gap intencional das janelas de parametros
    (04-05/2027) NAO ficam indeterminadas — pertencem ao C3 cronologico.
    """
    por_ciclo = _por_ciclo_caso_real(wb_caso_real)
    antes = classificar_enquadramento_pc(date(2027, 3, 31), por_ciclo)
    depois = classificar_enquadramento_pc(date(2027, 4, 1), por_ciclo)
    no_gap = classificar_enquadramento_pc(date(2027, 5, 31), por_ciclo)
    apos_gap = classificar_enquadramento_pc(date(2027, 6, 1), por_ciclo)
    assert antes.tipo == ENQ_CICLO and antes.ciclo == "C2"
    assert depois.tipo == ENQ_CICLO and depois.ciclo == "C3"
    assert no_gap.tipo == ENQ_CICLO and no_gap.ciclo == "C3"
    assert apos_gap.tipo == ENQ_CICLO and apos_gap.ciclo == "C3"


def test_fronteira_aditivo_na_cronologia_da_execucao(wb_caso_real):
    """Aditivos usam a MESMA cronologia da execucao (sem regra exclusiva)."""
    from _motor_posicao_contratual import calendario_execucao_por_ciclo

    por_ciclo = _por_ciclo_caso_real(wb_caso_real)
    execucao = calendario_execucao_por_ciclo([
        {"ciclo": nome, **reg} for nome, reg in por_ciclo.items()
    ])
    linha = list(execucao.values())
    assert determinar_ciclo_por_data(date(2027, 3, 31), linha).ciclo == "C2"
    assert determinar_ciclo_por_data(date(2027, 4, 1), linha).ciclo == "C3"
    # Aditivo dentro do gap das janelas de reajuste: fato contratual valido,
    # enquadrado no C3 cronologico — jamais desaparece.
    assert determinar_ciclo_por_data(date(2027, 5, 31), linha).ciclo == "C3"


# --------------------------------------------------------------------------- #
# Cenario D — sem retardo: baseline preservado
# --------------------------------------------------------------------------- #
def test_cenario_d_sem_retardo_linha_temporal_identica_ao_baseline():
    payload = _payload_caso_real()
    payload["ciclos"][1]["data_pedido"] = "04/04/2026"
    payload["ciclos"][1]["financeiro_inicio"] = "04/04/2026"
    payload["ciclos"][1]["situacao"] = "✅ TEMPESTIVO"
    payload["ciclos"][1]["efeito_financeiro_retardado"] = False
    wb = load_workbook(
        io.BytesIO(gerar_coleta_oficial_preenchida(payload)), data_only=False
    )
    par = wb["parametros"]
    assert _dia(par["C4"].value) == date(2026, 4, 1)
    assert _dia(par["D4"].value) == date(2027, 3, 31)   # 12 competencias
    assert _dia(par["C5"].value) == date(2027, 4, 1)
    assert _dia(par["D5"].value) == date(2028, 3, 31)
    assert par["A7"].value is None                       # sem legenda
    assert "TEMPESTIVO*" not in str(par["G4"].value)
    wb.close()


def test_retardo_nunca_antecipa_o_proximo_ciclo():
    """Efeito anterior ao inicio teorico do proprio ciclo (dado inconsistente
    ou legado) nao encurta o interregno."""
    dados = normalizar_dados_calculadora({
        "data_base_original": "01/10/2022",
        "ciclos": [
            {"ciclo": "C1", "data_base": "01/10/2023",
             "financeiro_inicio": "01/10/2023", "percentual_aplicado": 0.06,
             "objeto_analise_atual": True},
            {"ciclo": "C2", "data_base": "01/10/2024",
             "financeiro_inicio": "01/10/2024", "percentual_aplicado": 0.04,
             "objeto_analise_atual": True},
        ],
    })
    por_ciclo = {c["ciclo"]: c for c in dados["ciclos"]}
    assert por_ciclo["C1"]["data_inicio"] == date(2024, 10, 1)
    assert por_ciclo["C2"]["data_inicio"] == date(2025, 10, 1)


def test_precluso_sem_efeito_usa_fallback_teorico():
    """Item 17: precluso nao inventa INICIO_EFEITO; o ciclo seguinte nasce do
    marco teorico (+12 meses)."""
    dados = normalizar_dados_calculadora({
        "data_base_original": "05/03/2024",
        "ciclos": [
            {"ciclo": "C1", "data_base": "05/03/2024",
             "percentual_aplicado": 0.0, "objeto_analise_atual": True,
             "situacao": "❌ PRECLUSO"},
            {"ciclo": "C2", "data_base": "05/03/2025",
             "financeiro_inicio": "01/03/2026", "percentual_aplicado": 0.04,
             "objeto_analise_atual": True},
        ],
    })
    por_ciclo = {c["ciclo"]: c for c in dados["ciclos"]}
    assert por_ciclo["C1"]["data_inicio"] == date(2025, 3, 1)
    assert por_ciclo["C1"]["inicio_efeito_financeiro"] is None
    assert por_ciclo["C2"]["data_inicio"] == date(2026, 3, 1)


# --------------------------------------------------------------------------- #
# Interregno — casos do item 25
# --------------------------------------------------------------------------- #
@pytest.mark.parametrize("efeito,proximo", [
    ("01/03/2025", date(2026, 3, 1)),
    ("01/04/2025", date(2026, 4, 1)),
])
def test_interregno_12_meses_apos_o_efeito(efeito, proximo):
    """O PROXIMO ciclo nasce 12 meses apos o inicio dos efeitos (H+12m);
    a janela do ciclo atual permanece com 12 competencias exatas."""
    from _gerador_masterfile import _completar_periodos_ciclos

    retardado = efeito != "01/03/2025"
    dados = normalizar_dados_calculadora(_payload_ciclo_unico(efeito, retardado))
    c1 = dados["ciclos"][0]
    # Janela do C1 sempre com 12 competencias exatas (03/2025..02/2026).
    assert (c1["data_inicio"], c1["data_fim"]) == (date(2025, 3, 1), date(2026, 2, 28))
    completos = _completar_periodos_ciclos({"C1": c1})
    assert completos["C2"]["data_inicio"] == proximo


def test_interregno_efeito_junho_2026_gera_junho_2027():
    from _gerador_masterfile import _completar_periodos_ciclos

    dados = normalizar_dados_calculadora(_payload_caso_real())
    c2 = next(c for c in dados["ciclos"] if c["ciclo"] == "C2")
    # Janela do C2: 12 competencias exatas (04/2026..03/2027).
    assert c2["data_fim"] == date(2027, 3, 31)
    # O C3 nasce 12 meses apos o inicio dos efeitos do C2 (gap intencional
    # 04-05/2027 entre as janelas).
    completos = _completar_periodos_ciclos({c["ciclo"]: c for c in dados["ciclos"]})
    assert completos["C3"]["data_inicio"] == date(2027, 6, 1)


# --------------------------------------------------------------------------- #
# Metodo — Financeiro (Mensalidade) com alias legado
# --------------------------------------------------------------------------- #
def test_rotulo_novo_do_metodo_financeiro():
    assert M.ROTULO_PRINCIPAL == "Financeiro (Mensalidade)"
    assert M.OPCOES_DROPDOWN == [
        "SELECIONE AQUI",
        "Financeiro (Mensalidade)",
        "PC (Pedidos de Compra)",
        "Itens Consumidos",
    ]
    assert M.rotulo_metodo("principal") == "Financeiro (Mensalidade)"


@pytest.mark.parametrize("alias", [
    "Financeiro (Mensalidade)",
    "Principal (Financeiro)",       # alias legado — arquivos antigos
    "PRINCIPAL (FINANCEIRO)",
    "Principal",
    "Financeiro",
])
def test_aliases_do_metodo_normalizam_para_principal(alias):
    assert M.normalizar_metodo(alias) == "principal"
    assert M.metodo_selecionado(alias) is True


def test_leitor_aceita_coleta_legada_com_rotulo_antigo():
    from _leitor_masterfile_v10 import _normalizar_modo
    assert _normalizar_modo("Principal (Financeiro)") == "principal"
    assert _normalizar_modo("Financeiro (Mensalidade)") == "principal"


def test_dropdown_do_controle_oferece_o_rotulo_novo(wb_caso_real):
    ws = wb_caso_real["CONTROLE"]
    lista = next(
        dv.formula1
        for dv in ws.data_validations.dataValidation
        if dv.type == "list" and str(dv.sqref) == "B1"
    )
    assert lista == (
        '"SELECIONE AQUI,Financeiro (Mensalidade),'
        'PC (Pedidos de Compra),Itens Consumidos"'
    )
    assert "Principal (Financeiro)" not in lista


def test_modelo_em_branco_tambem_oferece_o_rotulo_novo():
    wb = load_workbook(io.BytesIO(obter_coleta_oficial_bytes()), data_only=False)
    ws = wb["CONTROLE"]
    lista = next(
        dv.formula1
        for dv in ws.data_validations.dataValidation
        if dv.type == "list" and str(dv.sqref) == "B1"
    )
    assert "Financeiro (Mensalidade)" in lista
    assert "Principal (Financeiro)" not in lista
    wb.close()


# --------------------------------------------------------------------------- #
# Fluxo REAL da Calculadora Simples (AppTest) — preclusao no gate do PR #28
# --------------------------------------------------------------------------- #
def _executar_calculadora_simples_preclusa():
    """Pagina real: ancora C1 = 05/03/2024, pedido 15/07/2025 -> PRECLUSO."""
    from streamlit.testing.v1 import AppTest

    at = AppTest.from_file(
        str(RAIZ / "pages" / "01_Calculo_Simples.py"), default_timeout=300
    )
    at.run()
    at.date_input[0].set_value(date(2024, 3, 5))   # Data-base/ancora do C1
    at.date_input[1].set_value(date(2025, 7, 15))  # Pedido fora dos 90 dias
    next(b for b in at.button if "Processar" in str(b.label)).click()
    at.run()
    return at


def test_fluxo_real_precluso_sem_acordo_nao_cria_ancora_financeira():
    """PRECLUSO sem acordo: payload sem datas de efeito, H vazio e o ciclo
    seguinte no marco teorico (+12 meses) — nunca pedido precluso + 12."""
    at = _executar_calculadora_simples_preclusa()
    adm = at.session_state["dados_admissibilidade"]
    ciclo = adm["ciclos"][0]
    assert "PRECLUSO" in str(ciclo["situacao"]).upper()
    assert ciclo["financeiro_inicio"] == ""
    assert ciclo["financeiro_fim"] == ""

    norm = normalizar_dados_calculadora(adm)
    c1 = next(c for c in norm["ciclos"] if c["ciclo"] == "C1")
    assert c1["inicio_efeito_financeiro"] is None
    assert (c1["data_inicio"], c1["data_fim"]) == (date(2025, 3, 1), date(2026, 2, 28))

    wb = load_workbook(
        io.BytesIO(gerar_coleta_oficial_preenchida(adm)), data_only=False
    )
    par = wb["parametros"]
    assert par["H3"].value is None                       # H do C1 vazio
    assert _dia(par["C3"].value) == date(2025, 3, 1)
    assert _dia(par["D3"].value) == date(2026, 2, 28)
    assert _dia(par["C4"].value) == date(2026, 3, 1)     # C2 teorico
    assert "C1=" not in str(wb.properties.keywords or "")
    fin = wb["financeiro"]
    assert all(fin[f"G{r}"].value != "Sim" for r in range(2, 74))
    wb.close()


def test_fluxo_real_precluso_com_acordo_ancora_na_competencia_pactuada():
    """Acordo negocial (18/08/2025): H = 01/08/2025 e proximo reajuste em
    08/2026 — unico caso em que o ciclo precluso cria a nova ancora."""
    at = _executar_calculadora_simples_preclusa()
    at.checkbox(key="superacao_negocial_simples_c1").check()
    at.run()
    at.date_input(key="inicio_negocial_simples_c1").set_value(date(2025, 8, 18))
    at.run()

    adm = at.session_state["dados_admissibilidade"]
    ciclo = adm["ciclos"][0]
    assert bool(ciclo["superacao_negocial"]) is True
    assert ciclo["financeiro_inicio"] == "01/08/2025"

    norm = normalizar_dados_calculadora(adm)
    c1 = next(c for c in norm["ciclos"] if c["ciclo"] == "C1")
    assert c1["inicio_efeito_financeiro"] == date(2025, 8, 1)
    # ETAPA 31: a janela do C1 NAO se alonga (12 competencias exatas:
    # 03/2025..02/2026); a ancora do acordo desloca apenas o proximo ciclo.
    assert c1["data_fim"] == date(2026, 2, 28)

    wb = load_workbook(
        io.BytesIO(gerar_coleta_oficial_preenchida(adm)), data_only=False
    )
    par = wb["parametros"]
    assert _dia(par["H3"].value) == date(2025, 8, 1)
    assert _dia(par["C4"].value) == date(2026, 8, 1)     # proximo reajuste
    wb.close()


# --------------------------------------------------------------------------- #
# Separacao REGIME (parametros C:D) x PERIODO DE APURACAO DO INDICE
# --------------------------------------------------------------------------- #
def _total_meses(ini: date, fim: date) -> int:
    return (fim.year * 12 + fim.month) - (ini.year * 12 + ini.month) + 1


def test_janela_do_reajuste_nao_altera_o_periodo_do_indice():
    """Caso obrigatorio do gate: C2 teorico 04/2026, efeito 06/2026.

    ETAPA 31 — JANELA C2 = 04/2026..03/2027 (12 competencias EXATAS; o
    retardo nao a alonga) e a janela do INDICE do C3 = 06/2026..05/2027
    (12 competencias, ancorada no inicio dos efeitos de C2). Sao dimensoes
    distintas: nenhuma contamina a outra, e o C3 nasce 12 meses apos o
    efeito do C2 (06/2027), abrindo gap intencional 04-05/2027.
    """
    from dateutil.relativedelta import relativedelta

    from _gerador_masterfile import _completar_periodos_ciclos

    dados = normalizar_dados_calculadora(_payload_caso_real())
    c2 = next(c for c in dados["ciclos"] if c["ciclo"] == "C2")
    # JANELA DO REAJUSTE: 12 competencias exatas, sempre.
    assert (c2["data_inicio"], c2["data_fim"]) == (date(2026, 4, 1), date(2027, 3, 31))
    assert _total_meses(c2["data_inicio"], c2["data_fim"]) == 12
    # PERIODO DO INDICE do proximo reajuste: ancora = competencia do inicio
    # dos efeitos do anterior; janela = ancora..ancora+11m = 12 competencias.
    ancora_indice_c3 = c2["inicio_efeito_financeiro"]
    fim_indice_c3 = ancora_indice_c3 + relativedelta(months=11)
    assert ancora_indice_c3 == date(2026, 6, 1)
    assert fim_indice_c3 == date(2027, 5, 1)
    assert _total_meses(ancora_indice_c3, fim_indice_c3) == 12
    # JANELA de aplicacao do C3 comeca 12 meses apos o efeito do C2.
    completos = _completar_periodos_ciclos({c["ciclo"]: c for c in dados["ciclos"]})
    assert completos["C3"]["data_inicio"] == date(2027, 6, 1)


def test_paginas_ancoram_o_indice_em_12_competencias():
    """As Calculadoras enviam aos motores de indice SEMPRE ancora..+11 meses
    (12 competencias) e encadeiam a ancora seguinte pelo inicio dos efeitos.
    parametros!C:D (regime) nunca alimenta o calculo do indice."""
    simples = (RAIZ / "pages" / "01_Calculo_Simples.py").read_text(encoding="utf-8")
    multi = (RAIZ / "pages" / "02_Calculo_Represados.py").read_text(encoding="utf-8")
    assert "dt_fim_ap = dt_base + relativedelta(months=11)" in simples
    assert "d_fim = data_atual + relativedelta(months=11)" in multi
    assert "data_base_proximo_ciclo = inicio_efeito_financeiro" in multi
    # Chamadas reais dos motores: recebem a janela da pagina, nao C:D do XLS.
    assert "get_index_data(serie_sgs_do_indice(tipo_idx), dt_base, dt_fim_ap)" in simples
    assert (
        "get_data_rep(serie_sgs_do_indice(idx_sel), data_atual, d_fim" in multi
    )


def test_memoria_de_calculo_mensal_registra_12_competencias():
    """Indice mensal na janela 06/2026..05/2027: 12 linhas MES + RESULTADO."""
    import pandas as pd
    from _memoria_calculo import normalizar_memoria_calculo

    serie = pd.DataFrame({
        "data": pd.date_range("2026-06-01", periods=12, freq="MS"),
        "valor": [0.4] * 12,
    })
    registros = normalizar_memoria_calculo(
        {"dados": serie, "metodo": "Produtório"}, 1.049, 0.049
    )
    tipos = [r["tipo"] for r in registros]
    assert tipos.count("MES") == 12
    assert tipos.count("RESULTADO") == 1
    assert registros[0]["competencia"].startswith("2026-06")
    assert registros[11]["competencia"].startswith("2027-05")


# --------------------------------------------------------------------------- #
# Pagamento nunca e marco temporal
# --------------------------------------------------------------------------- #
def test_fontes_temporais_nao_usam_data_de_pagamento():
    """Nenhuma fonte temporal consulta pagamento para ciclo/efeito/interregno.

    ``pc_pago_a_contratada``/``valor_pago`` participam apenas de Q5-Q9 do
    motor temporal (valor pago/retroativo), nunca do enquadramento (Q4).
    """
    from _coleta_oficial import normalizar_dados_calculadora as fonte_calendario

    funcoes_temporais = (
        fonte_calendario,
        efeito_financeiro_pc,
        determinar_ciclo_por_data,
        divergencias_calendario_canonico,
        classificar_enquadramento_pc,
    )
    for funcao in funcoes_temporais:
        codigo = inspect.getsource(funcao).lower().replace(
            "a data de pagamento nunca participa", ""
        )
        assert "pago" not in codigo, funcao.__name__
        assert "pagamento" not in codigo, funcao.__name__
