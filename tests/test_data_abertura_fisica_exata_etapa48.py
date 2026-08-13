# -*- coding: utf-8 -*-
"""ETAPA 48 — data fisica exata da fotografia de itens_Remanesc.

Duas datas passam a coexistir no XLS oficial:

  parametros!C2:C6  cadeia MENSAL/financeira (intocada — ex.: 01/04/2025)
  parametros!I2:I6  DATA_ABERTURA_FISICA_EXATA (nova — ex.: 20/04/2025)

O campo `data_abertura_fisica_exata` do payload apenas TRANSPORTA a decisao
que a Calculadora ja tomava (`referencia_exata_efeito`): pedido apto no
TEMPESTIVO/TEMPESTIVO*, referencia apta no ADIANTADO, data pactuada no acordo
negocial. Payload legado (sem o campo) usa data_inicio como fallback em I.

Somente as 4 familias fisicas do template leem I:
  itens_Remanesc!E1:T1, posicao_contratual!AB:AF (apenas a fronteira INT),
  posicao_referencia!I6, cobertura_temporal!B6.
"""
from __future__ import annotations

from datetime import date
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from _coleta_oficial import (
    TEMPLATE_COLETA_OFICIAL,
    gerar_coleta_oficial_preenchida,
    normalizar_dados_calculadora,
)

ROOT = Path(__file__).resolve().parents[1]


def _ciclo(numero: int, data_base: str, pedido: str, exata: str | None,
           financeiro: str, situacao: str = "TEMPESTIVO", **extra):
    reg = {
        "ciclo": f"C{numero}",
        "data_base": data_base,
        "data_pedido": pedido,
        "situacao": situacao,
        "situacao_aplicada": situacao,
        "objeto_analise_atual": True,
        "ciclo_ja_concedido": False,
        "percentual_aplicado": 0.05,
        "variacao": 0.05,
        "fator": 1.05,
        "ciclo_calculado": True,
        "efeito_financeiro_retardado": False,
        "financeiro_inicio": financeiro,
    }
    if exata is not None:
        reg["data_abertura_fisica_exata"] = exata
    reg.update(extra)
    return reg


def _payload(*ciclos):
    return {
        "origem": "Reajustes Múltiplos",
        "indice": "IPCA",
        "data_base_original": ciclos[0]["data_base"],
        "fator": 1.05,
        "fator_acumulado": 1.05,
        "variacao_acumulada": 0.05,
        "ciclos": list(ciclos),
    }


def _parametros(payload):
    wb = load_workbook(BytesIO(gerar_coleta_oficial_preenchida(payload)), data_only=False)
    return wb, wb["parametros"]


def _d(celula):
    v = celula.value
    return v.date() if hasattr(v, "date") else v


# ------------------------------------------------------------------ situacoes
def test_a_tempestivo_separa_cadeia_mensal_da_fisica():
    wb, par = _parametros(_payload(
        _ciclo(1, "20/04/2023", "20/04/2024", "20/04/2024", "01/04/2024"),
    ))
    assert par["I1"].value == "DATA_ABERTURA_FISICA_EXATA"
    assert _d(par["C3"]) == date(2024, 4, 1)      # mensal preservado
    assert _d(par["I3"]) == date(2024, 4, 20)     # fisica exata
    assert _d(par["H3"]) == date(2024, 4, 1)      # efeito financeiro intacto
    assert par["I3"].number_format == "dd/mm/yyyy"
    # itens_Remanesc exibe a fotografia a partir de I, nunca mais de C
    rem = wb["itens_Remanesc"]
    assert "parametros!$I$3" in str(rem["E1"].value)
    assert "parametros!$C$" not in str(rem["E1"].value)


def test_b_tempestivo_retardado_usa_pedido_aplicavel_exato():
    wb, par = _parametros(_payload(
        _ciclo(1, "20/04/2023", "15/06/2024", "15/06/2024", "01/06/2024",
               efeito_financeiro_retardado=True),
    ))
    assert _d(par["C3"]) == date(2024, 4, 1)      # janela ancorada no aniversario
    assert _d(par["I3"]) == date(2024, 6, 15)     # dia exato preservado
    assert _d(par["H3"]) == date(2024, 6, 1)      # competencia do efeito


def test_c_adiantado_nao_antecipa_fotografia_pela_data_do_pedido():
    # Pedido 01/08 anterior a referencia apta 23/08: o motor transporta a
    # referencia apta (referencia_exata_efeito = d_aniv), nunca o pedido.
    wb, par = _parametros(_payload(
        _ciclo(1, "23/08/2024", "01/08/2025", "23/08/2025", "01/08/2025",
               situacao="ADIANTADO"),
    ))
    assert _d(par["I3"]) == date(2025, 8, 23)
    assert _d(par["I3"]) != date(2025, 8, 1)


def test_d_acordo_negocial_grava_data_pactuada_exata():
    wb, par = _parametros(_payload(
        _ciclo(1, "18/08/2024", "10/12/2025", "18/08/2025", "01/08/2025",
               situacao="CICLO ADMITIDO POR NEGOCIACAO ENTRE AS PARTES",
               superacao_negocial=True),
    ))
    assert _d(par["C3"]) == date(2025, 8, 1)
    assert _d(par["I3"]) == date(2025, 8, 18)
    assert _d(par["H3"]) == date(2025, 8, 1)


def test_e_analise_comecando_em_c2_preserva_dia_exato():
    wb, par = _parametros(_payload(
        _ciclo(2, "20/04/2024", "20/04/2025", "20/04/2025", "01/04/2025"),
    ))
    assert _d(par["C4"]) == date(2025, 4, 1)
    assert _d(par["I4"]) == date(2025, 4, 20)


def test_f_analise_comecando_em_c3_preserva_dia_exato():
    wb, par = _parametros(_payload(
        _ciclo(3, "07/09/2025", "07/09/2026", "07/09/2026", "01/09/2026"),
    ))
    assert _d(par["C5"]) == date(2026, 9, 1)
    assert _d(par["I5"]) == date(2026, 9, 7)


def test_g_payload_legado_sem_campo_usa_fallback_data_inicio():
    payload = _payload(_ciclo(1, "20/04/2023", "20/04/2024", None, "01/04/2024"))
    assert "data_abertura_fisica_exata" not in payload["ciclos"][0]
    wb, par = _parametros(payload)          # XLS continua sendo gerado
    assert _d(par["C3"]) == date(2024, 4, 1)
    assert _d(par["I3"]) == date(2024, 4, 1)   # fallback = data_inicio mensal


# ------------------------------------------------------- transporte/normalizador
def test_normalizador_preserva_dia_e_nao_realimenta_cadeia_mensal():
    com = normalizar_dados_calculadora(_payload(
        _ciclo(1, "20/04/2023", "20/04/2024", "20/04/2024", "01/04/2024"),
        _ciclo(2, "01/04/2024", "20/04/2025", "20/04/2025", "01/04/2025"),
    ))
    sem = normalizar_dados_calculadora(_payload(
        _ciclo(1, "20/04/2023", "20/04/2024", None, "01/04/2024"),
        _ciclo(2, "01/04/2024", "20/04/2025", None, "01/04/2025"),
    ))
    for c_com, c_sem in zip(com["ciclos"], sem["ciclos"]):
        # cadeia mensal identica com e sem o campo novo (nada realimentado)
        assert c_com["data_inicio"] == c_sem["data_inicio"]
        assert c_com["data_fim"] == c_sem["data_fim"]
        assert c_com["inicio_efeito_financeiro"] == c_sem["inicio_efeito_financeiro"]
    assert com["ciclos"][0]["data_abertura_fisica_exata"] == date(2024, 4, 20)
    assert com["ciclos"][1]["data_abertura_fisica_exata"] == date(2025, 4, 20)
    assert sem["ciclos"][0]["data_abertura_fisica_exata"] is None


def test_pagina_apenas_transporta_referencia_exata_ja_decidida():
    fonte = (ROOT / "pages/02_Calculo_Represados.py").read_text(encoding="utf-8")
    linha = next(l for l in fonte.splitlines() if "'data_abertura_fisica_exata'" in l)
    # transporte puro: le referencia_exata_efeito do proprio ciclo, sem calculo
    assert "referencia_exata_efeito" in linha
    assert ".replace(day=1)" not in linha
    assert fonte.count("'data_abertura_fisica_exata'") == 1


# ------------------------------------------------------------ protecao financeira
def test_protecao_cadeias_mensais_identicas_com_e_sem_campo_novo():
    xls_com = gerar_coleta_oficial_preenchida(_payload(
        _ciclo(1, "20/04/2023", "20/04/2024", "20/04/2024", "01/04/2024"),
    ))
    xls_sem = gerar_coleta_oficial_preenchida(_payload(
        _ciclo(1, "20/04/2023", "20/04/2024", None, "01/04/2024"),
    ))
    wa, wd = load_workbook(BytesIO(xls_com)), load_workbook(BytesIO(xls_sem))
    pa, pd_ = wa["parametros"], wd["parametros"]
    for col in ("C", "D", "H"):
        for r in range(2, 7):
            assert pa[f"{col}{r}"].value == pd_[f"{col}{r}"].value, f"parametros!{col}{r}"
    for aba, faixa in (("financeiro", ("A", "B", "G")), ("aditivos", ("C",)), ("itens_PC", ("C",))):
        for col in faixa:
            for r in (2, 3, 30):
                assert wa[aba][f"{col}{r}"].value == wd[aba][f"{col}{r}"].value, f"{aba}!{col}{r}"
    assert wa["CONTROLE"]["B13"].value == wd["CONTROLE"]["B13"].value


def test_template_familias_migradas_e_consumidores_mensais_intactos():
    wb = load_workbook(TEMPLATE_COLETA_OFICIAL, data_only=False)
    par = wb["parametros"]
    # superficie runtime: template nasce com I vazio
    assert all(par[f"I{r}"].value in (None, "") for r in range(1, 8))
    # familias fisicas -> I
    rem = wb["itens_Remanesc"]
    esperado = {"E1": 3, "F1": 3, "G1": 4, "H1": 4, "I1": 5, "J1": 5, "K1": 6,
                "L1": 6, "M1": 3, "N1": 3, "O1": 4, "P1": 4, "Q1": 5, "R1": 5,
                "S1": 6, "T1": 6}
    for coord, linha in esperado.items():
        f = str(rem[coord].value)
        assert f"parametros!$I${linha}" in f and "parametros!$C$" not in f, coord
    pc = wb["posicao_contratual"]
    for col, n in zip(("AB", "AC", "AD", "AE", "AF"), range(2, 7)):
        f = str(pc[f"{col}2"].value)
        assert f"INT(parametros!$I${n})+1" in f, col
        # 48.3: guarda mensal fica e passa a reconhecer tambem I vazio
        assert f'IF(OR(parametros!$C${n}="",parametros!$I${n}=""),0,' in f, col
        assert "INT(parametros!$C$" not in f, col
    assert "parametros!$I$" in str(wb["posicao_referencia"]["I6"].value)
    assert "parametros!$I$" in str(wb["cobertura_temporal"]["B6"].value)
    # consumidores mensais seguem em C
    assert "parametros!$C$2" in str(wb["financeiro"]["B2"].value)
    assert "parametros!$I$" not in str(wb["financeiro"]["B2"].value)
    assert "parametros!$C$2" in str(wb["itens_PC"]["C2"].value)
    assert "parametros!$C$2" in str(wb["aditivos"]["C2"].value)
    ctl = str(wb["CONTROLE"]["B13"].value)
    assert "parametros!C6" in ctl and "parametros!I" not in ctl
    assert "parametros!$I$" not in str(wb["posicao_referencia"]["I1"].value)
    assert "parametros!$I$" not in str(wb["posicao_referencia"]["I3"].value)


def test_correspondencia_ciclos_coluna_i_segue_coluna_c():
    wb, par = _parametros(_payload(
        _ciclo(1, "20/04/2023", "20/04/2024", "20/04/2024", "01/04/2024"),
        _ciclo(2, "01/04/2024", "20/04/2025", "20/04/2025", "01/04/2025"),
    ))
    # linha 2=C0 ... linha 6=C4, exatamente como a coluna C
    assert [par[f"B{r}"].value for r in range(2, 7)] == ["C0", "C1", "C2", "C3", "C4"]
    assert _d(par["I3"]) == date(2024, 4, 20)   # C1
    assert _d(par["I4"]) == date(2025, 4, 20)   # C2
    # C0 derivado (sem campo proprio): fallback = data_inicio derivada
    assert _d(par["I2"]) == _d(par["C2"])
