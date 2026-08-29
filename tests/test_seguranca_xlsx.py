from __future__ import annotations

import random
from io import BytesIO
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from openpyxl import Workbook

import _seguranca_xlsx as seguranca
from _coleta_oficial import gerar_coleta_oficial_preenchida
from _coleta_reajuste import ler_coleta_reajuste
from _coleta_reajuste_documentos import processar_coleta_oficial_runtime


ROOT = Path(__file__).resolve().parents[1]
COMPONENTES = {
    "[Content_Types].xml": "<Types/>",
    "_rels/.rels": "<Relationships/>",
    "xl/workbook.xml": "<workbook/>",
}


def _zip(membros: list[tuple[str, bytes | str]]) -> bytes:
    saida = BytesIO()
    with ZipFile(saida, "w", ZIP_DEFLATED) as pacote:
        for nome, conteudo in membros:
            pacote.writestr(nome, conteudo)
    return saida.getvalue()


def _xlsx_minimo_valido() -> bytes:
    saida = BytesIO()
    Workbook().save(saida)
    return saida.getvalue()


@pytest.mark.parametrize("conteudo", [b"", b"texto renomeado para xlsx"])
def test_rejeita_conteudo_que_nao_e_zip(conteudo: bytes) -> None:
    with pytest.raises(seguranca.XlsxInvalidoError, match="não é um XLSX válido"):
        seguranca.validar_xlsx_antes_do_parser(conteudo)


def test_rejeita_zip_generico_e_mensagem_nao_vaza_detalhe_interno() -> None:
    with pytest.raises(seguranca.XlsxEstruturaError) as erro:
        seguranca.validar_xlsx_antes_do_parser(_zip([("nota.txt", "seguro")]))
    assert str(erro.value) == seguranca.MENSAGEM_ESTRUTURA_XLSX
    assert "Content_Types" not in str(erro.value)


def test_rejeita_zip_truncado() -> None:
    truncado = _xlsx_minimo_valido()[:-20]
    with pytest.raises(seguranca.XlsxInvalidoError):
        seguranca.validar_xlsx_antes_do_parser(truncado)


def test_aceita_xlsx_minimo_real() -> None:
    conteudo = _xlsx_minimo_valido()
    assert bytes(seguranca.validar_xlsx_antes_do_parser(conteudo)) == conteudo


@pytest.mark.parametrize(
    "arquivo",
    ["Coleta_Reajuste.xlsx", "COLETA_REAJUSTE_OFICIAL.xlsx"],
)
def test_aceita_modelos_oficiais(arquivo: str) -> None:
    conteudo = (ROOT / "templates" / arquivo).read_bytes()
    assert bytes(seguranca.validar_xlsx_antes_do_parser(conteudo)) == conteudo


def test_aceita_modelo_oficial_preenchido_gerado() -> None:
    conteudo = gerar_coleta_oficial_preenchida({})
    assert bytes(seguranca.validar_xlsx_antes_do_parser(conteudo)) == conteudo


def test_rejeita_tamanho_compactado_excessivo(monkeypatch) -> None:
    conteudo = _xlsx_minimo_valido()
    monkeypatch.setattr(seguranca, "TAMANHO_MAXIMO_ARQUIVO", len(conteudo) - 1)
    with pytest.raises(seguranca.XlsxLimiteError):
        seguranca.validar_xlsx_antes_do_parser(conteudo)


def test_rejeita_total_descompactado_excessivo(monkeypatch) -> None:
    conteudo = _zip(list(COMPONENTES.items()) + [("xl/dados.xml", b"A" * 100)])
    monkeypatch.setattr(seguranca, "TAMANHO_MAXIMO_DESCOMPACTADO", 99)
    with pytest.raises(seguranca.XlsxLimiteError):
        seguranca.validar_xlsx_antes_do_parser(conteudo)


def test_rejeita_maior_membro_excessivo(monkeypatch) -> None:
    conteudo = _zip(list(COMPONENTES.items()) + [("xl/dados.xml", b"A" * 100)])
    monkeypatch.setattr(seguranca, "TAMANHO_MAXIMO_MEMBRO", 99)
    with pytest.raises(seguranca.XlsxLimiteError):
        seguranca.validar_xlsx_antes_do_parser(conteudo)


def test_rejeita_quantidade_excessiva_de_membros(monkeypatch) -> None:
    conteudo = _zip(list(COMPONENTES.items()) + [("xl/extra.xml", b"x")])
    monkeypatch.setattr(seguranca, "QUANTIDADE_MAXIMA_MEMBROS", 3)
    with pytest.raises(seguranca.XlsxLimiteError):
        seguranca.validar_xlsx_antes_do_parser(conteudo)


def test_rejeita_razao_de_compressao_excessiva(monkeypatch) -> None:
    conteudo = _zip(list(COMPONENTES.items()) + [("xl/dados.xml", b"A" * 2_000)])
    monkeypatch.setattr(seguranca, "RAZAO_MAXIMA_COMPRESSAO_MEMBRO", 2)
    with pytest.raises(seguranca.XlsxLimiteError):
        seguranca.validar_xlsx_antes_do_parser(conteudo)


def test_rejeita_entrada_duplicada() -> None:
    membros = list(COMPONENTES.items()) + [("xl/workbook.xml", "duplicado")]
    with pytest.warns(UserWarning, match="Duplicate name"):
        conteudo = _zip(membros)
    with pytest.raises(seguranca.XlsxEstruturaError):
        seguranca.validar_xlsx_antes_do_parser(conteudo)


@pytest.mark.parametrize("nome", ["../segredo.xml", "/absoluto.xml", "C:/absoluto.xml"])
def test_rejeita_traversal_e_caminhos_absolutos(nome: str) -> None:
    conteudo = _zip(list(COMPONENTES.items()) + [(nome, "x")])
    with pytest.raises(seguranca.XlsxEstruturaError):
        seguranca.validar_xlsx_antes_do_parser(conteudo)


@pytest.mark.parametrize("ausente", ["[Content_Types].xml", "xl/workbook.xml"])
def test_rejeita_componente_ooxml_minimo_ausente(ausente: str) -> None:
    conteudo = _zip([(n, c) for n, c in COMPONENTES.items() if n != ausente])
    with pytest.raises(seguranca.XlsxEstruturaError):
        seguranca.validar_xlsx_antes_do_parser(conteudo)


def test_leitor_rejeita_antes_de_openpyxl(monkeypatch) -> None:
    def nao_pode_abrir(*_args, **_kwargs):
        raise AssertionError("openpyxl.load_workbook não deveria ser chamado")

    monkeypatch.setattr("_coleta_reajuste.load_workbook", nao_pode_abrir)
    with pytest.raises(seguranca.XlsxInvalidoError):
        ler_coleta_reajuste(b"nao e XLSX")


def test_runtime_rejeita_antes_de_acionar_leitores(monkeypatch) -> None:
    def nao_pode_ler(*_args, **_kwargs):
        raise AssertionError("leitor não deveria ser chamado")

    monkeypatch.setattr("_coleta_reajuste_documentos.ler_masterfile_v10", nao_pode_ler)
    with pytest.raises(seguranca.XlsxInvalidoError):
        processar_coleta_oficial_runtime(b"nao e XLSX")


def test_runtime_nao_propaga_erro_tecnico_do_parser(monkeypatch) -> None:
    monkeypatch.setattr(
        "_coleta_reajuste_documentos.ler_masterfile_v10", lambda *_a, **_k: {"ok": True}
    )
    monkeypatch.setattr(
        "_coleta_reajuste_documentos.ler_coleta_reajuste",
        lambda *_a, **_k: {"valido": True},
    )

    def falha_tecnica(*_args, **_kwargs):
        raise ValueError("detalhe interno do parser")

    monkeypatch.setattr(
        "_coleta_reajuste_documentos.adaptar_coleta_reajuste_para_documentos",
        falha_tecnica,
    )
    with pytest.raises(seguranca.XlsxInvalidoError) as erro:
        processar_coleta_oficial_runtime(_xlsx_minimo_valido())
    assert str(erro.value) == seguranca.MENSAGEM_XLSX_INVALIDO
    assert "detalhe interno" not in str(erro.value)


def test_marcador_evitaria_cinco_inspecoes_zip(monkeypatch) -> None:
    validado = seguranca.validar_xlsx_antes_do_parser(_xlsx_minimo_valido())

    def nao_pode_reabrir(*_args, **_kwargs):
        raise AssertionError("ZIP já validado não deve ser reinspecionado")

    monkeypatch.setattr(seguranca, "ZipFile", nao_pode_reabrir)
    for _ in range(5):
        assert seguranca.garantir_xlsx_validado(validado) is validado


# --------------------------------------------- XSEC-08.2: orcamento pre-parser
#
# Maximos medidos no corpus legitimo (204 workbooks distintos por SHA-256, com
# abas dentro da allowlist; os dez maiores conferidos um a um contra
# validar_geometria_workbook e todos aceitos). Sao a base da calibracao: se um
# dia um desses tetos descer abaixo destes numeros, arquivo legitimo passa a ser
# recusado.
CORPUS_MAIOR_XML_WORKSHEET = 16_738_211
CORPUS_MAIOR_SOMA_WORKSHEETS = 24_531_289


def _bytes_incompressiveis(tamanho: int) -> bytes:
    """Recheio deterministico que nao dispara a razao de compressao.

    Zeros seriam mais baratos, mas comprimem muito acima de 100:1 e o pacote
    seria recusado pelo gate de razao — mascarando o teto sob teste.
    """
    return random.Random(20260829).randbytes(tamanho)


def _ooxml_com_worksheets(tamanhos: list[int], extras: list[str] | None = None) -> bytes:
    """Pacote OOXML minimo com worksheets de tamanho declarado controlado."""
    membros: list[tuple[str, bytes | str]] = list(COMPONENTES.items())
    for indice, tamanho in enumerate(tamanhos, start=1):
        membros.append(
            (f"xl/worksheets/sheet{indice}.xml", _bytes_incompressiveis(tamanho))
        )
    for nome in extras or []:
        membros.append((nome, _bytes_incompressiveis(4096)))
    return _zip(membros)


def test_limites_de_xml_de_worksheet_sao_os_calibrados() -> None:
    assert seguranca.MAX_BYTES_XML_WORKSHEET == 32 * 1024 * 1024
    assert seguranca.MAX_BYTES_XML_WORKSHEETS == 48 * 1024 * 1024
    # A folga sobre o corpus e deliberada e precisa continuar existindo.
    assert seguranca.MAX_BYTES_XML_WORKSHEET > CORPUS_MAIOR_XML_WORKSHEET
    assert seguranca.MAX_BYTES_XML_WORKSHEETS > CORPUS_MAIOR_SOMA_WORKSHEETS


def test_worksheet_isolado_acima_do_teto_e_rejeitado(monkeypatch) -> None:
    monkeypatch.setattr(seguranca, "MAX_BYTES_XML_WORKSHEET", 8_192)
    with pytest.raises(seguranca.XlsxLimiteError) as erro:
        seguranca.validar_xlsx_antes_do_parser(_ooxml_com_worksheets([8_193]))
    assert str(erro.value) == seguranca.MENSAGEM_LIMITE_XLSX
    assert "worksheet" not in str(erro.value)


def test_soma_dos_worksheets_acima_do_teto_e_rejeitada(monkeypatch) -> None:
    """Cada aba cabe sozinha; o workbook inteiro nao cabe."""
    monkeypatch.setattr(seguranca, "MAX_BYTES_XML_WORKSHEET", 8_192)
    monkeypatch.setattr(seguranca, "MAX_BYTES_XML_WORKSHEETS", 16_384)
    with pytest.raises(seguranca.XlsxLimiteError):
        seguranca.validar_xlsx_antes_do_parser(
            _ooxml_com_worksheets([8_000, 8_000, 8_000])
        )


def test_worksheets_dentro_do_teto_passam(monkeypatch) -> None:
    monkeypatch.setattr(seguranca, "MAX_BYTES_XML_WORKSHEET", 8_192)
    monkeypatch.setattr(seguranca, "MAX_BYTES_XML_WORKSHEETS", 16_384)
    conteudo = _ooxml_com_worksheets([8_000, 8_000])
    assert bytes(seguranca.validar_xlsx_antes_do_parser(conteudo)) == conteudo


def test_orcamento_ignora_partes_que_nao_carregam_celula(monkeypatch) -> None:
    """xl/worksheets/_rels/* mora no mesmo diretorio e nao tem celula nenhuma."""
    assert seguranca._eh_xml_de_worksheet("xl/worksheets/sheet1.xml") is True
    assert seguranca._eh_xml_de_worksheet("xl/worksheets/_rels/sheet1.xml.rels") is False
    assert seguranca._eh_xml_de_worksheet("xl/workbook.xml") is False
    assert seguranca._eh_xml_de_worksheet("xl/worksheets/sheet1.bin") is False

    monkeypatch.setattr(seguranca, "MAX_BYTES_XML_WORKSHEETS", 8_192)
    conteudo = _ooxml_com_worksheets(
        [8_000], extras=["xl/worksheets/_rels/sheet1.xml.rels"]
    )
    assert bytes(seguranca.validar_xlsx_antes_do_parser(conteudo)) == conteudo


@pytest.mark.parametrize(
    "arquivo",
    ["Coleta_Reajuste.xlsx", "COLETA_REAJUSTE_OFICIAL.xlsx"],
)
def test_modelos_oficiais_cabem_no_orcamento(arquivo: str) -> None:
    with ZipFile(BytesIO((ROOT / "templates" / arquivo).read_bytes())) as pacote:
        worksheets = [
            membro.file_size
            for membro in pacote.infolist()
            if seguranca._eh_xml_de_worksheet(membro.filename)
        ]
    assert worksheets
    assert max(worksheets) <= seguranca.MAX_BYTES_XML_WORKSHEET
    assert sum(worksheets) <= seguranca.MAX_BYTES_XML_WORKSHEETS


def test_orcamento_de_xml_e_cobrado_antes_de_openpyxl(monkeypatch) -> None:
    """O ganho do XSEC-08.2 esta em rejeitar SEM abrir o workbook."""
    import _leitor_masterfile_v10 as leitor

    def nao_pode_abrir(*_args, **_kwargs):
        raise AssertionError("load_workbook foi alcancado apesar do orcamento estourado")

    monkeypatch.setattr(seguranca, "MAX_BYTES_XML_WORKSHEET", 8_192)
    monkeypatch.setattr(leitor, "load_workbook", nao_pode_abrir)
    resultado = leitor.ler_masterfile_v10(
        _ooxml_com_worksheets([8_193]), exigir_modelo_oficial=True
    )
    assert resultado["erro"] == seguranca.MENSAGEM_LIMITE_XLSX
