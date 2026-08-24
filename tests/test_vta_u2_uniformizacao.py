# -*- coding: utf-8 -*-
"""VTA-U2 — invariantes permanentes da uniformizacao natural do VTA.

O objetivo destes testes NAO e provar que XLS e Python exibem o mesmo numero:
e provar que os dois CALCULAM a mesma realidade economica por caminhos
independentes, e que uma alteracao errada em apenas um lado produz divergencia
detectavel (nunca substituicao silenciosa).

Cobertura:
  * identidade economica por metodo (executado + ajustes + remanescente = VTA);
  * lado XLS: a formula de VTA_FINAL reproduz a identidade sozinha;
  * lado Python: o motor reproduz a identidade sozinho, sem ler VTA_FINAL;
  * trava anti-dupla-contagem do registro agregador TOTAL;
  * divergencia proposital: o sistema detecta e nao copia um lado para o outro;
  * referencia fisica (B10/B11) nunca vira VTA;
  * bloco didatico da RESULTADOS reconcilia com o VTA Oficial;
  * Sumario Executivo sem o comparativo interno.
"""
from __future__ import annotations

import sys
import warnings
from pathlib import Path

import openpyxl
import pytest

RAIZ = Path(__file__).resolve().parents[1]
if str(RAIZ) not in sys.path:
    sys.path.insert(0, str(RAIZ))

from _motor_composicao_vta import montar_composicao_vta  # noqa: E402
from _reconciliacao_xls_python import (  # noqa: E402
    STATUS_CONCILIADO,
    STATUS_RELEVANTE,
    reconciliar_xls_python,
)
from _resultado_consolidado import montar_resultado_consolidado  # noqa: E402

TEMPLATE = RAIZ / "templates" / "COLETA_REAJUSTE_OFICIAL.xlsx"

# Caso real de referencia (arquivo do usuario, metodo Financeiro, ciclo C3).
EXECUTADO = 7_300_890.27
AJUSTES = 24_678.92
REMANESCENTE_BASE = 1_349_258.38
REMANESCENTE = 1_388_251.07
VTA = 8_713_820.26


@pytest.fixture(scope="module")
def wb():
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        return openpyxl.load_workbook(TEMPLATE, data_only=False)


# ---------------------------------------------------------------- fixtures

def _leitura_financeiro(*, com_agregador_total: bool = True) -> dict:
    """Leitura sintetica do metodo Financeiro com os numeros do caso real.

    O ciclo C3 tem pago 853.961,00 e considerado 878.639,92 — a diferenca de
    24.678,92 e o reajuste ja reconhecido e ainda nao pago. A linha TOTAL
    reproduz o agregador que existe na aba financeiro do arquivo real.
    """
    parcelas = [
        {"fonte_parcela": "Financeiro", "ciclo": "C0",
         "identificador": "financeiro:c0:base:7",
         "valor": 4_739_007.27, "valor_atualizado": 4_739_007.27},
        {"fonte_parcela": "Financeiro", "ciclo": "C1",
         "identificador": "financeiro:c1:base:21",
         "valor": 853_961.00, "valor_atualizado": 853_961.00},
        {"fonte_parcela": "Financeiro", "ciclo": "C2",
         "identificador": "financeiro:c2:base:33",
         "valor": 853_961.00, "valor_atualizado": 853_961.00},
        {"fonte_parcela": "Financeiro", "ciclo": "C3",
         "identificador": "financeiro:c3:base:44",
         "valor": 853_961.00, "valor_atualizado": 878_639.92},
        # Parcela-delta: o retroativo da competencia ja esta dentro do
        # valor_atualizado da parcela-base acima. Somar as duas dobraria.
        {"fonte_parcela": "Financeiro", "ciclo": "C3",
         "identificador": "financeiro:c3:delta:44",
         "valor": 24_678.92, "valor_atualizado": None},
    ]
    if com_agregador_total:
        parcelas.append({
            "fonte_parcela": "Financeiro", "ciclo": "TOTAL",
            "identificador": "financeiro:TOTAL:base:74",
            "valor": EXECUTADO, "valor_atualizado": EXECUTADO,
        })
    return {
        "controle": {"modo": "principal", "ciclo_vigente": "C3"},
        "parametros_v10": {"por_ciclo": {
            "C0": {"fator_acumulado": 1.0},
            "C1": {"fator_acumulado": 1.0},
            "C2": {"fator_acumulado": 1.0},
            "C3": {"fator_acumulado": 1.028899355437093},
        }},
        "vta_sombra": {"parcelas_computadas": parcelas},
        "potencial_futuro": {"disponivel": False},
        "posicao_contratual": {
            "ok": True, "cache_ausente": False,
            "itens": [{
                "ITEM": "1",
                "VU_ORIGINAL": REMANESCENTE_BASE,
                "QTD_REM_AJUSTADA_C3": 1.0,
            }],
        },
        "historico_vu": {"itens": [
            {"item": "1", "vu_ciclos": {"VU_C3": REMANESCENTE}},
        ]},
        "reconciliacao": {"registros": []},
        "itens_contrato": {"itens": []},
    }


def _leitura_consumido() -> dict:
    """Caso controlado homologado do metodo Consumido: 284 + 0 + 126 = 410."""
    return {
        "controle": {"modo": "d", "ciclo_vigente": "C1"},
        "parametros_v10": {"por_ciclo": {
            "C0": {"fator_acumulado": 1.0},
            "C1": {"fator_acumulado": 1.0},
        }},
        "reconciliacao": {"registros": [
            {"ciclo": "C1", "valor_computado": 284.0,
             "fonte_principal": "consumidos", "metodo_apuracao": "CONSUMO"},
        ]},
        "potencial_futuro": {
            "disponivel": True,
            "saldo_remanescente_base": 126.0,
            "fator_vigente": 1.0,
            "valor_atualizado_vigente": 126.0,
            "ciclo_vigente": "C1",
        },
        "vta_sombra": {"parcelas_computadas": []},
        "itens_contrato": {"itens": []},
    }


# ------------------------------------------------- A. lado XLS (independente)

def test_a1_xls_vta_final_financeiro_reproduz_a_identidade_sozinho(wb):
    """A formula de VTA_FINAL soma desembolsado + retroativo + remanescente.

    D20 = SOMA(financeiro!C) (pago informado), B21 = retroativo oficial,
    D35 = remanescente atualizado oficial. Nenhum termo vem do Python.
    """
    formula = str(wb["MEMORIA_RESULTADOS"]["B26"].value)
    ramo = formula[formula.index('IF($B$4="Financeiro"'):]
    assert "$D$20+B21+D35" in ramo.replace(" ", "")
    for proibido in ("python", "PYTHON", "W50"):
        assert proibido not in ramo


def test_a2_xls_d20_le_o_pago_informado_e_nao_a_reconstrucao_quantitativa(wb):
    """O executado do Financeiro e o desembolso informado, nunca a
    reconstrucao pelo quantitativo (que e so conferencia, bloco 8)."""
    d20 = str(wb["MEMORIA_RESULTADOS"]["D20"].value)
    assert "financeiro!$C$2:$C$73" in d20
    assert "itens_Remanesc" not in d20
    assert "posicao_contratual" not in d20


def test_a3_card_vta_oficial_aponta_para_a_saida_canonica(wb):
    """RESULTADOS!C5 (card VTA OFICIAL) e a medida 9 do bloco 6 usam
    VTA_FINAL, nunca a referencia fisica B10/W50, e sem hardcode."""
    res = wb["RESULTADOS"]
    for celula in ("C5", "B63"):
        formula = str(res[celula].value)
        assert "VTA_FINAL" in formula
        assert "$B$10" not in formula
        assert "W50" not in formula
        # nenhum numero literal: o valor tem de mudar com as entradas.
        assert not any(ch.isdigit() for ch in formula.replace("VTA_FINAL", ""))


def test_a4_referencias_fisicas_preservadas_mas_nao_apresentadas(wb):
    """UX final: B10-B13 continuam no arquivo, nos mesmos enderecos que o
    leitor Python usa (`_ler_referencias_vta`), mas as linhas ficam OCULTAS —
    nenhuma referencia fisica concorre visualmente com o VTA."""
    res = wb["RESULTADOS"]
    # formulas tecnicas preservadas nas mesmas celulas
    assert "MEMORIA_RESULTADOS!$W$50" in str(res["B10"].value)
    assert "MEMORIA_RESULTADOS!$W$48" in str(res["B11"].value)
    assert "comparativo_VTA!$B$208" in str(res["B12"].value)
    assert "MEMORIA_RESULTADOS!$W$51" in str(res["B13"].value)
    # e a apresentacao retirada
    for linha in range(10, 14):
        assert res.row_dimensions[linha].hidden is True
        assert str(res[f"A{linha}"].value).startswith("[AUDITORIA INTERNA]")


def _textos_visiveis(res) -> list[str]:
    ocultas = {n for n, d in res.row_dimensions.items() if d.hidden}
    return [
        str(celula.value)
        for linha in res.iter_rows(min_row=1, max_row=90, max_col=9)
        for celula in linha
        if celula.value is not None
        and celula.row not in ocultas
        and isinstance(celula.value, str)
    ]


def test_a5_area_principal_nao_tem_vta_alternativo(wb):
    """Aceite visual B/C: nada visivel nas linhas 1-14 chama posicao fisica de
    VTA nem a apresenta como referencia concorrente ao lado do VTA."""
    res = wb["RESULTADOS"]
    ocultas = {n for n, d in res.row_dimensions.items() if d.hidden}
    principal = [
        str(celula.value)
        for linha in res.iter_rows(min_row=1, max_row=14, max_col=9)
        for celula in linha
        if isinstance(celula.value, str) and celula.row not in ocultas
    ]
    assert principal
    for texto in principal:
        alto = texto.upper()
        assert "VTA OFICIAL — POSIÇÃO FÍSICA" not in alto
        assert "VTA OFICIAL - POSICAO FISICA" not in alto
        assert "REFERÊNCIA AUDITÁVEL" not in alto
        assert "POSIÇÃO FÍSICA ATUAL" not in alto
    # o card principal fala de uma unica grandeza
    assert str(res["A4"].value) == "VTA OFICIAL"
    assert res["A6"].value in (None, "")


def test_a6_card_principal_tem_titulo_valor_e_status_do_vta(wb):
    """Aceite visual A: o valor em destaque e exclusivamente VTA_FINAL."""
    res = wb["RESULTADOS"]
    assert str(res["A4"].value) == "VTA OFICIAL"
    assert "VTA_FINAL" in str(res["C5"].value)
    assert str(res["C4"].value) == "=$H$8"          # status especifico do VTA
    assert "VTA_FINAL" in str(res["H8"].value)


def test_a7_titulo_da_secao_1_enuncia_a_identidade_canonica(wb):
    titulo = str(wb["RESULTADOS"]["A9"].value)
    assert "Executado apurado" in titulo
    assert "Ajustes ainda devidos" in titulo
    assert "Remanescente atualizado" in titulo
    assert "VTA Oficial" in titulo
    assert "POSIÇÃO FÍSICA" not in titulo.upper()
    assert "POSICAO FISICA" not in titulo.upper()


# ------------------------------------- B. bloco didatico "COMO E FORMADO O VTA"

def test_b1_bloco_didatico_existe_com_as_quatro_parcelas(wb):
    res = wb["RESULTADOS"]
    assert str(res["A79"].value).startswith("9. COMO E FORMADO O VTA")
    assert "ja executado" in str(res["A80"].value)
    assert "VTA Oficial = Executado apurado" in str(res["A81"].value)
    assert [str(res[f"A{linha}"].value) for linha in range(83, 87)] == [
        "Executado apurado",
        "(+) Ajustes ainda devidos",
        "(+) Remanescente atualizado",
        "(=) VTA Oficial",
    ]


def test_b2_parcelas_derivam_das_fontes_reais_sem_digitacao(wb):
    """Nenhuma das quatro linhas pode ser valor digitado."""
    res = wb["RESULTADOS"]
    for linha in range(83, 87):
        assert str(res[f"B{linha}"].value).startswith("=")
    assert "$D$20" in str(res["B83"].value)      # Financeiro: pago informado
    assert "$F$20" in str(res["B83"].value)      # Itens: execucao atualizada
    assert "$T$21" in str(res["B83"].value)      # PCs: valor considerado
    assert "$B$21" in str(res["B84"].value)      # ajustes = retroativo oficial
    assert "$D$35" in str(res["B85"].value)      # remanescente atualizado
    assert "VTA_FINAL" in str(res["B86"].value)


def test_b3_consumido_e_pc_nao_recontam_o_ajuste_ja_embutido(wb):
    """Item 7 da tarefa: no Consumido e no PC o reajuste ja esta dentro da
    execucao — a linha de ajustes tem de ser zero, nunca o retroativo."""
    formula = str(wb["RESULTADOS"]["B84"].value)
    assert (
        'IF(OR(MEMORIA_RESULTADOS!$B$4="Itens",'
        'MEMORIA_RESULTADOS!$B$4="PCs"),0,"")'
    ) in formula
    fonte = str(wb["RESULTADOS"]["C84"].value)
    assert "ja esta dentro da execucao atualizada" in fonte
    assert "ja esta dentro do valor considerado dos PCs" in fonte


def test_b4_bloco_tem_linha_de_conferencia_contra_o_vta_oficial(wb):
    conf = str(wb["RESULTADOS"]["B87"].value)
    assert "$B$86-($B$83+N($B$84)+$B$85)" in conf.replace(" ", "")


# ------------------------------------------- C. bloco 8 (conferencia) didatico

def test_c1_conferencia_tem_rotulos_didaticos(wb):
    res = wb["RESULTADOS"]
    assert str(res["C72"].value) == "Execucao estimada pelo quantitativo"
    assert str(res["D72"].value) == "Diferenca para o Financeiro informado"
    assert str(res["E72"].value) == "Conferencia"
    assert "nao altera o VTA Oficial" in str(res["A78"].value)


def test_c2_sem_base_a_diferenca_fica_vazia_e_o_status_explica(wb):
    res = wb["RESULTADOS"]
    for linha in range(73, 78):
        diferenca = str(res[f"D{linha}"].value)
        assert f"NOT(ISNUMBER(C{linha}))" in diferenca
        assert '),"",' in diferenca  # sem base -> vazio, nunca zero
        assert "Sem dados quantitativos para comparar" in str(res[f"E{linha}"].value)


def test_c3_conferencia_nao_alimenta_o_vta_oficial(wb):
    """As celulas do bloco 8 nao podem ser referenciadas por VTA_FINAL."""
    b26 = str(wb["MEMORIA_RESULTADOS"]["B26"].value)
    for linha in range(71, 79):
        assert f"RESULTADOS!$B${linha}" not in b26
        assert f"RESULTADOS!$C${linha}" not in b26


# ------------------------------------------ D. lado Python (independente)

def test_d1_python_compoe_o_financeiro_pela_identidade_economica():
    composicao = montar_composicao_vta(_leitura_financeiro())

    assert composicao["disponivel"] is True
    assert composicao["metodo"] == "financeiro"
    assert composicao["total_execucao_atualizada"] == EXECUTADO
    assert composicao["retroativo_implicito"] == AJUSTES
    assert composicao["saldo_remanescente"]["valor_atualizado"] == REMANESCENTE
    assert composicao["saldo_remanescente"]["valor_base"] == REMANESCENTE_BASE
    assert composicao["vta_composicao"] == VTA


def test_d2_as_parcelas_exibidas_somam_exatamente_o_vta():
    """Item 12: a soma do bloco visivel tem de fechar com o VTA Oficial."""
    composicao = montar_composicao_vta(_leitura_financeiro())
    linhas = composicao["linhas"]

    assert [linha["descricao"] for linha in linhas] == [
        "Executado apurado", "Ajustes ainda devidos", "Remanescente atualizado",
    ]
    soma = round(sum(linha["valor_atualizado"] for linha in linhas), 2)
    assert soma == composicao["vta_composicao"] == VTA


def test_d3_python_nao_le_o_resultado_do_xls_para_compor_o_vta():
    """Prova de independencia: o motor nao toca em resultados_xls/VTA_FINAL."""
    fonte = (RAIZ / "_motor_composicao_vta.py").read_text(encoding="utf-8")
    for proibido in ("resultados_xls", "VTA_FINAL", "reconciliacao_xls_python"):
        assert proibido not in fonte


# -------------------------------- E. trava anti-dupla-contagem do agregador

def test_e1_registro_total_nao_soma_junto_das_parcelas_que_agrega():
    """A linha TOTAL da aba financeiro consolida C0..C4. Soma-la junto delas
    produzia 14.651.851,59 — quase o dobro da execucao real."""
    com_total = montar_composicao_vta(_leitura_financeiro(com_agregador_total=True))
    sem_total = montar_composicao_vta(_leitura_financeiro(com_agregador_total=False))

    assert com_total["vta_composicao"] == sem_total["vta_composicao"] == VTA
    assert com_total["total_execucao_atualizada"] == EXECUTADO
    assert com_total["vta_composicao"] != pytest.approx(14_651_851.59)


def test_e2_nenhuma_parcela_visivel_totaliza_o_valor_da_dupla_contagem():
    composicao = montar_composicao_vta(_leitura_financeiro())
    for linha in composicao["linhas"]:
        assert linha["valor_atualizado"] < VTA


def test_e3_parcela_delta_nao_conta_o_retroativo_duas_vezes():
    """O retroativo ja esta no valor_atualizado da parcela-base; a parcela
    ':delta:' da mesma competencia nao pode somar de novo."""
    composicao = montar_composicao_vta(_leitura_financeiro())
    assert composicao["retroativo_implicito"] == AJUSTES
    assert composicao["retroativo_implicito"] != round(AJUSTES * 2, 2)


def test_e4_fator_nao_e_reaplicado_sobre_valor_ja_atualizado():
    """C3 vale 878.639,92 (ja atualizado). Multiplicar pelo fator do ciclo
    (1,0288...) produziria 904.032,05 e inflaria a execucao."""
    composicao = montar_composicao_vta(_leitura_financeiro())
    detalhe = {d["ciclo"]: d for d in composicao["detalhamento_por_ciclo"]}
    assert detalhe["C3"]["valor_considerado"] == 878_639.92
    assert detalhe["C3"]["ajuste_devido"] == AJUSTES


def test_e5_detalhamento_por_ciclo_nao_gera_segundo_total():
    """Item 1.4: o detalhamento e camada de auditoria; nao pode aparecer
    como parcela somavel da composicao economica."""
    composicao = montar_composicao_vta(_leitura_financeiro())
    ciclos_nas_linhas = [
        linha.get("ciclo") for linha in composicao["linhas"]
        if linha.get("tipo") == "execucao"
    ]
    assert ciclos_nas_linhas == [""]
    assert len(composicao["detalhamento_por_ciclo"]) == 4


# ---------------------------------------------------- F. Consumido e PC

def test_f1_consumido_controlado_284_mais_126_igual_410():
    composicao = montar_composicao_vta(_leitura_consumido())

    assert composicao["total_execucao_atualizada"] == 284.0
    assert composicao["saldo_remanescente"]["valor_atualizado"] == 126.0
    assert composicao["vta_composicao"] == 410.0
    # o ajuste tecnico ja contido na execucao nao reaparece como parcela.
    assert composicao["retroativo_implicito"] == 0.0


def test_f2_consumido_e_pc_nao_passam_pela_composicao_do_financeiro():
    """Item 1.3: PC e Consumido nao podem ser alterados para acomodar o
    Financeiro — nem sequer entram no ramo dedicado dele."""
    consumido = montar_composicao_vta(_leitura_consumido())
    assert consumido.get("metodo") != "financeiro"
    assert "ajustes_devidos" not in consumido

    leitura_pc = dict(
        _leitura_consumido(), controle={"modo": "pc", "ciclo_vigente": "C1"}
    )
    composicao_pc = montar_composicao_vta(leitura_pc)
    assert composicao_pc["metodo"] == "pc"
    assert "ajustes_devidos" not in composicao_pc


# --------------------------------- G. divergencia proposital (sem copia)

def _leitura_para_reconciliacao(vta_xls: float, vta_python: float) -> dict:
    return {
        "resultados_xls": {
            "disponivel": True,
            "cache_ausente": False,
            "nomes_presentes": ["VTA_FINAL"],
            "valores": {"VTA_FINAL": vta_xls},
        },
        "controle": {"modo": "principal"},
        "objeto_processo": {"memoria_por_ciclo": {
            "ciclos": [],
            "vta": {"valor_total_atualizado": vta_python},
        }},
    }


def test_g1_valores_iguais_conciliam_naturalmente():
    resultado = reconciliar_xls_python(_leitura_para_reconciliacao(VTA, VTA))
    campo = next(c for c in resultado["campos"] if c["campo"] == "VTA_FINAL")

    assert campo["status"] == STATUS_CONCILIADO
    assert campo["xls"] == campo["python"] == VTA


def test_g2_divergencia_proposital_e_detectada_e_nenhum_lado_e_copiado():
    """Caso D da tarefa: um lado errado nao pode ser silenciosamente
    substituido pelo outro nem declarado validado."""
    vta_xls_errado = round(VTA + 1_000.00, 2)
    leitura = _leitura_para_reconciliacao(vta_xls_errado, VTA)
    resultado = reconciliar_xls_python(leitura)
    campo = next(c for c in resultado["campos"] if c["campo"] == "VTA_FINAL")

    assert resultado["status_geral"] == STATUS_RELEVANTE
    assert campo["status"] == STATUS_RELEVANTE
    # os DOIS valores continuam visiveis, sem escolha silenciosa.
    assert campo["xls"] == vta_xls_errado
    assert campo["python"] == VTA
    assert campo in resultado["divergencias_relevantes"]


def test_g3_mudar_so_o_python_tambem_produz_divergencia():
    """Simetria: o teste tem de pegar erro em qualquer um dos lados."""
    resultado = reconciliar_xls_python(
        _leitura_para_reconciliacao(VTA, round(VTA - 500.00, 2))
    )
    assert resultado["status_geral"] == STATUS_RELEVANTE


# -------------------------- H. referencia fisica nunca substitui o VTA

def _resultado_consolidado_base(vta_canonico, forma1, forma2) -> dict:
    return {
        "valor_atualizado_contrato": vta_canonico,
        "valor_represado_a_pagar": AJUSTES,
        "controle": {"modo": "principal", "ciclo_vigente": "C3"},
        "memoria_por_ciclo": {"vta": {"metodo": "financeiro"}},
        "referencias_vta": {
            "posicao_atual_disponivel": forma1 is not None,
            "forma1_posicao_atual": forma1,
            "forma2_ultima_abertura": forma2,
        },
        "composicao_vta": {
            "disponivel": True, "bloqueia_formalizacao": False,
            "linhas": [{"descricao": "Executado apurado",
                        "valor_atualizado": EXECUTADO}],
            "alertas": [],
        },
        "politica_entrega_segura": {
            "status": "PRONTO_PARA_VALIDACAO_FISCAL", "pendencias": [],
            "retroativo": {"metodo": "financeiro"},
        },
        "reconciliacao_xls_python": {"status_geral": "CONCILIADO"},
    }


def test_h1_vta_canonico_prevalece_sobre_a_referencia_fisica():
    consolidado = montar_resultado_consolidado(
        _resultado_consolidado_base(VTA, 7_835_180.34, 7_835_180.34)
    )
    assert consolidado["vta"] == VTA
    assert consolidado["vta"] != 7_835_180.34


def test_h2_sem_vta_canonico_o_resultado_e_fail_closed():
    consolidado = montar_resultado_consolidado(
        _resultado_consolidado_base(None, None, 7_835_180.34)
    )
    assert consolidado["vta"] is None
    assert consolidado["vta_origem"] == "indisponivel"
    assert consolidado["referencias_auditaveis"][
        "ultima_abertura_disponivel"] == 7_835_180.34


def test_h3_o_codigo_nao_tem_mais_fallback_de_vta_por_referencia_fisica():
    fonte = (RAIZ / "_resultado_consolidado.py").read_text(encoding="utf-8")
    assert "vta = vta_ultima_posicao" not in fonte
    assert "forma2_ultima_abertura" in fonte  # segue exposta como referencia


# ------------------------------------------------ I. mensagens e documentos

def test_i1_mensagem_tecnica_do_fator_vira_linguagem_didatica():
    from _resultado_consolidado import _traduzir_alerta_composicao

    tecnica = (
        "Composicao VTA: ciclo TOTAL sem fator acumulado parametrizado; "
        "execucao composta pela base, sem atualizacao."
    )
    didatica = _traduzir_alerta_composicao(tecnica)

    assert didatica != tecnica
    assert "Isso não impede o cálculo do VTA." in didatica
    for jargao in ("fator acumulado parametrizado", "Composicao VTA:", "TOTAL"):
        assert jargao not in didatica


def test_i2_sumario_executivo_sem_o_comparativo_interno():
    """Item 11: o comparativo do contrato integralmente reajustado e
    informacao interna e nao sai no Sumario Executivo."""
    fonte = (RAIZ / "_sumario_executivo.py").read_text(encoding="utf-8")

    # VTA-U2 (regra global): o quadro inteiro de referencias saiu do Sumario —
    # com ele, o comparativo e os dois "VTA pela posicao ..." que podiam ser
    # lidos como um segundo VTA.
    assert "def _bloco_referencias_vta_pdf" not in fonte
    assert "_bloco_referencias_vta_pdf(historia" not in fonte
    assert "Contrato original integralmente reajustado" not in fonte
    assert "COMPARATIVO" not in fonte
    # o VTA oficial continua sendo apresentado pela sintese.
    assert 'vta_txt = formatar_moeda(sintese.get("vta"))' in fonte


def test_i3_o_dado_interno_continua_disponivel_para_auditoria():
    """A remocao e de APRESENTACAO: o leitor segue expondo o comparativo."""
    leitor = (RAIZ / "_leitor_masterfile_v10.py").read_text(encoding="utf-8")
    assert "forma3_integral_reajustado" in leitor
