"""Fronteira de geometria e allowlist de abas (XLSX-SEC-2B, achados P0).

XSEC-01: a dimensao de uma aba e declarada pelo arquivo. Uma unica celula
remota infla o retangulo sem inflar o pacote, e a varredura de formulas
materializa o retangulo inteiro.

XSEC-02: sem allowlist, abas extras (inclusive veryHidden) entravam em
silencio e multiplicavam esse custo.

XSEC-03: o limite de transporte precisa espelhar o limite interno.

Os workbooks adversariais deste modulo sao sinteticos e minimos: nenhum teste
aqui percorre retangulo grande — o proprio gate rejeita antes.
"""
from __future__ import annotations

import tomllib
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

import _coleta_reajuste
import _leitor_masterfile_v10 as _leitor
import _seguranca_xlsx as seguranca
from _seguranca_xlsx import (
    ABAS_PERMITIDAS,
    TAMANHO_MAXIMO_ARQUIVO,
    XlsxEstruturaError,
    XlsxLimiteError,
    validar_geometria_workbook,
)

ROOT = Path(__file__).resolve().parents[1]

# Abas que a logica de negocio exige; as cinco restantes da allowlist sao
# opcionais de compatibilidade.
OBRIGATORIAS = (
    "CONTROLE",
    "parametros",
    "financeiro",
    "itens_Remanesc",
    "itens_Consumidos",
    "itens_PC",
    "aditivos",
    "posicao_contratual",
    "itens_RC",
    "historico_VU",
    "RESULTADOS",
)
OPCIONAIS = (
    "comparativo_VTA",
    "posicao_referencia",
    "cobertura_temporal",
    "MEMORIA_RESULTADOS",
    "CICLO_EM_EXECUCAO",
)


def _wb(abas, celula=None, estados=None):
    """Workbook sintetico: uma celula por aba, opcionalmente uma remota.

    ``celula`` = (aba, linha, coluna) posiciona um unico valor distante, que e
    exatamente como a geometria adversarial e construida no mundo real.
    """
    wb = Workbook()
    wb.remove(wb.active)
    for nome in abas:
        ws = wb.create_sheet(nome)
        ws["A1"] = nome
        if estados and nome in estados:
            ws.sheet_state = estados[nome]
    if celula:
        aba, linha, coluna = celula
        wb[aba].cell(row=linha, column=coluna).value = "x"
    return wb


# --------------------------------------------------------------- geometria


def test_ataque_20000_x_500_e_rejeitado():
    wb = _wb(OBRIGATORIAS, celula=("itens_Remanesc", 20_000, 500))
    with pytest.raises(XlsxLimiteError):
        validar_geometria_workbook(wb)


def test_ataque_extremo_xfd1048576_e_rejeitado():
    """Ultima celula da grade do Excel: 1.048.576 x 16.384."""
    wb = _wb(OBRIGATORIAS, celula=("itens_Remanesc", 1_048_576, 16_384))
    with pytest.raises(XlsxLimiteError):
        validar_geometria_workbook(wb)


def test_linha_e_coluna_validas_mas_area_excessiva_e_rejeitada():
    """9.999 <= 10.000 e 99 <= 100, mas o produto e 989.901."""
    assert 9_999 <= seguranca.MAX_LINHAS_POR_ABA
    assert 99 <= seguranca.MAX_COLUNAS_POR_ABA
    assert 9_999 * 99 > seguranca.MAX_AREA_POR_ABA
    wb = _wb(OBRIGATORIAS, celula=("itens_PC", 9_999, 99))
    with pytest.raises(XlsxLimiteError):
        validar_geometria_workbook(wb)


def test_abas_validas_isoladamente_mas_area_acumulada_excessiva():
    """Cada aba fica sob MAX_AREA_POR_ABA; a soma estoura o workbook."""
    abas = OBRIGATORIAS + OPCIONAIS
    assert 2_000 * 29 <= seguranca.MAX_AREA_POR_ABA
    assert len(abas) * 2_000 * 29 > seguranca.MAX_AREA_TOTAL_WORKBOOK
    wb = _wb(abas)
    for nome in abas:
        wb[nome].cell(row=2_000, column=29).value = "x"  # 58.000 por aba
    with pytest.raises(XlsxLimiteError):
        validar_geometria_workbook(wb)


def test_dezessete_abas_sao_rejeitadas():
    wb = _wb(OBRIGATORIAS + OPCIONAIS + ("EXTRA",))
    assert len(wb.sheetnames) == 17
    with pytest.raises(XlsxLimiteError):
        validar_geometria_workbook(wb)


# --------------------------------------------------------------- allowlist


@pytest.mark.parametrize("estado", ["visible", "hidden", "veryHidden"])
def test_aba_extra_e_rejeitada_em_qualquer_estado(estado):
    """A allowlist e por nome: esconder a aba nao a torna aceitavel."""
    wb = _wb(OBRIGATORIAS + ("EXTRA_1",), estados={"EXTRA_1": estado})
    assert wb["EXTRA_1"].sheet_state == estado
    with pytest.raises(XlsxEstruturaError):
        validar_geometria_workbook(wb)


@pytest.mark.parametrize("estado", ["visible", "hidden", "veryHidden"])
def test_aba_legitima_e_aceita_em_qualquer_estado(estado):
    """financeiro, itens_Consumidos, itens_PC e aditivos aparecem ora
    visible, ora hidden nos arquivos reais de producao."""
    wb = _wb(OBRIGATORIAS, estados={aba: estado for aba in OBRIGATORIAS})
    validar_geometria_workbook(wb)


def test_opcionais_aceitas_quando_presentes():
    validar_geometria_workbook(_wb(OBRIGATORIAS + OPCIONAIS))


def test_opcionais_podem_estar_ausentes():
    """Compatibilidade retroativa: este gate nao exige as cinco opcionais."""
    validar_geometria_workbook(_wb(OBRIGATORIAS))


def test_allowlist_codificada_tem_exatamente_as_dezesseis_abas():
    assert set(ABAS_PERMITIDAS) == set(OBRIGATORIAS) | set(OPCIONAIS)
    assert len(ABAS_PERMITIDAS) == 16
    assert "MEMORIA_RESULTADOS" in ABAS_PERMITIDAS
    assert "comparativo_VTA" in ABAS_PERMITIDAS


def test_limites_codificados_sao_os_aprovados():
    assert seguranca.MAX_LINHAS_POR_ABA == 10_000
    assert seguranca.MAX_COLUNAS_POR_ABA == 100
    assert seguranca.MAX_AREA_POR_ABA == 300_000
    assert seguranca.MAX_AREA_TOTAL_WORKBOOK == 500_000
    assert seguranca.MAX_ABAS_WORKBOOK == 16


# ------------------------------------------------- ordem: gate antes de varrer


def test_geometria_invalida_nao_alcanca_formulas(monkeypatch):
    """Regressao estrutural do XSEC-01.

    O valor do gate esta na ORDEM: se um dia a chamada migrar para depois da
    varredura, o custo adversarial volta inteiro. A sentinela prova que
    _formulas() nao chega a ser executada.
    """

    def _sentinela(*_args, **_kwargs):
        raise AssertionError("_formulas() foi alcancada apesar da geometria invalida")

    monkeypatch.setattr(_coleta_reajuste, "_formulas", _sentinela)

    buffer = BytesIO()
    _wb(OBRIGATORIAS, celula=("itens_Remanesc", 20_000, 500)).save(buffer)
    with pytest.raises(XlsxLimiteError):
        _coleta_reajuste.ler_coleta_reajuste(buffer.getvalue())


def test_aba_nao_permitida_nao_alcanca_formulas(monkeypatch):
    """Mesma protecao de ordem para o XSEC-02."""

    def _sentinela(*_args, **_kwargs):
        raise AssertionError("_formulas() foi alcancada apesar da aba nao permitida")

    monkeypatch.setattr(_coleta_reajuste, "_formulas", _sentinela)

    buffer = BytesIO()
    _wb(OBRIGATORIAS + ("EXTRA_1",), estados={"EXTRA_1": "veryHidden"}).save(buffer)
    with pytest.raises(XlsxEstruturaError):
        _coleta_reajuste.ler_coleta_reajuste(buffer.getvalue())


def test_mensagens_permanecem_genericas():
    """O usuario nao recebe coordenada ofensora, limite interno nem memoria."""
    wb = _wb(OBRIGATORIAS, celula=("itens_Remanesc", 20_000, 500))
    with pytest.raises(XlsxLimiteError) as erro:
        validar_geometria_workbook(wb)
    mensagem = str(erro.value)
    assert mensagem == seguranca.MENSAGEM_LIMITE_XLSX
    for vazamento in ("20000", "itens_Remanesc", "500", "10000", "300000"):
        assert vazamento not in mensagem


# --------------------------------------------------------------- legitimos


@pytest.mark.parametrize(
    "modelo", ["COLETA_REAJUSTE_OFICIAL.xlsx", "Coleta_Reajuste.xlsx"]
)
def test_templates_oficiais_passam(modelo):
    caminho = ROOT / "templates" / modelo
    wb = load_workbook(BytesIO(caminho.read_bytes()), data_only=False)
    try:
        validar_geometria_workbook(wb)
    finally:
        wb.close()


# --------------------------------------------------------------- XSEC-03


def test_max_upload_size_espelha_o_limite_interno():
    config = tomllib.loads(
        (ROOT / ".streamlit" / "config.toml").read_text(encoding="utf-8")
    )
    assert config["server"]["maxUploadSize"] == 20
    # O Streamlit multiplica por 1024*1024, entao a unidade e MiB — a mesma de
    # TAMANHO_MAXIMO_ARQUIVO. As duas fronteiras precisam coincidir byte a byte.
    assert config["server"]["maxUploadSize"] == TAMANHO_MAXIMO_ARQUIVO // (1024 * 1024)


# ------------------- XSEC-09: o gate tambem no PRIMEIRO parser vivo do upload
#
# O gate ja existia, correto, em ler_coleta_reajuste — mas esse e o TERCEIRO
# load_workbook do fluxo de upload. ler_masterfile_v10 abre o mesmo arquivo
# antes dele e percorre ~20 lacos "range(2, ws.max_row + 1)" com ws.cell(), que
# materializa a celula ausente. Sem gate ali, um max_row inflado virava custo
# real de memoria antes de qualquer veredito.


def _bytes_wb(abas, celula=None, estados=None) -> bytes:
    buffer = BytesIO()
    _wb(abas, celula=celula, estados=estados).save(buffer)
    return buffer.getvalue()


def _proibir_materializacao_de_celula(monkeypatch):
    """Arma a sentinela DEPOIS de o workbook hostil ja estar em bytes.

    Montar o proprio caso de teste usa ws.cell(); a sentinela precisa valer so
    a partir do momento em que o leitor recebe os bytes.
    """
    from openpyxl.worksheet.worksheet import Worksheet

    def _sentinela(self, *_args, **_kwargs):
        raise AssertionError(
            f"ws.cell() foi alcancado em {self.title!r} antes do gate de geometria"
        )

    monkeypatch.setattr(Worksheet, "cell", _sentinela)


def test_leitor_rejeita_geometria_hostil_sem_materializar_celula(monkeypatch):
    """Regressao estrutural do XSEC-08.1."""
    hostil = _bytes_wb(OBRIGATORIAS, celula=("itens_Remanesc", 20_000, 500))
    _proibir_materializacao_de_celula(monkeypatch)
    resultado = _leitor.ler_masterfile_v10(hostil, exigir_modelo_oficial=True)
    assert resultado["erro"] == seguranca.MENSAGEM_LIMITE_XLSX
    assert resultado["ok"] is False


def test_leitor_rejeita_aba_nao_permitida_sem_materializar_celula(monkeypatch):
    """A allowlist (XSEC-02) tambem chegava tarde demais."""
    hostil = _bytes_wb(
        OBRIGATORIAS + ("EXTRA_1",), estados={"EXTRA_1": "veryHidden"}
    )
    _proibir_materializacao_de_celula(monkeypatch)
    resultado = _leitor.ler_masterfile_v10(hostil, exigir_modelo_oficial=True)
    assert resultado["erro"] == seguranca.MENSAGEM_ESTRUTURA_XLSX


def test_leitor_rejeita_geometria_antes_de_detectar_versao(monkeypatch):
    """Prova de ORDEM contra a primeira instrucao que o leitor executa apos o load."""

    def _sentinela(*_args, **_kwargs):
        raise AssertionError("_detectar_versao() rodou apesar da geometria invalida")

    monkeypatch.setattr(_leitor, "_detectar_versao", _sentinela)
    resultado = _leitor.ler_masterfile_v10(
        _bytes_wb(OBRIGATORIAS, celula=("itens_Remanesc", 20_000, 500)),
        exigir_modelo_oficial=True,
    )
    assert resultado["erro"] == seguranca.MENSAGEM_LIMITE_XLSX


def test_caminho_legado_do_leitor_nao_sofre_a_allowlist_da_coleta():
    """A allowlist descreve a Coleta, nao o Masterfile legado.

    Aplica-la fora de exigir_modelo_oficial recusaria "historico"/"validacoes",
    que sao abas legitimas do layout antigo lido por consumidores internos.
    """
    legado = (
        "CONTROLE", "parametros", "financeiro", "itens_Remanesc",
        "itens_Consumidos", "aditivos", "historico", "itens_RC", "historico_VU",
    )
    resultado = _leitor.ler_masterfile_v10(_bytes_wb(legado))
    assert resultado["erro"] not in (
        seguranca.MENSAGEM_ESTRUTURA_XLSX,
        seguranca.MENSAGEM_LIMITE_XLSX,
    )


def test_modelo_oficial_real_continua_sendo_lido():
    """Zero falso positivo: o gate antecipado nao pode recusar arquivo legitimo."""
    conteudo = (ROOT / "templates" / "COLETA_REAJUSTE_OFICIAL.xlsx").read_bytes()
    resultado = _leitor.ler_masterfile_v10(conteudo, exigir_modelo_oficial=True)
    assert resultado["ok"] is True
    assert not resultado["erro"]
