# -*- coding: utf-8 -*-
"""RESULTADOS-ROLLBACK-1 — trava focal da apresentacao anterior da aba.

A frente RESULTADOS-UX2 (PRs #136/#137) acrescentou uma "camada humana" em
RESULTADOS!90:166, ocultou o motor tecnico inteiro (1:89) e trocou o print
setup. Ela foi APOSENTADA: a aba voltou a apresentacao do checkpoint

    f8296f7c2962352716edd22044ed9573f5eeee8a

sem que a main perdesse nada do que veio depois. Esta trava existe para que a
apresentacao descartada nao volte por engano — nem pela reaplicacao de um
gerador, nem por um novo template colado por cima.

DIFERENCA PERMITIDA CONTRA O DOADOR (allowlist explicita, item 4 do gate):

  RESULTADOS!H5/H8/C12  fator historico canonico do PR #139;
  defined names         os 14 do PR #135 + RETROATIVO_POTENCIAL_PC, todos
                        invisiveis e apontando para celulas tecnicas;
  camada PC-UX-1       substitui somente os merges de titulos/secoes e amplia
                        C54:H66 para explicacoes visiveis; o motor e seus pinos
                        permanecem nas coordenadas historicas;
  pageSetup/@scale      cache derivado: com fitToPage ligado o Excel recalcula
                        esse atributo a cada gravacao (68 no doador, 65 aqui) e
                        o ignora ao imprimir. O que vale — fitToPage, ajuste a
                        1 pagina de largura, area, orientacao e margens — e
                        conferido abaixo.

Nada mais pode divergir: valores, formatos, fontes, preenchimentos,
alinhamentos, bordas, merges, formatacao condicional, validacoes e
visibilidade de 1:87 foram provados identicos ao doador.
"""
from __future__ import annotations

import io
import sys
from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook

RAIZ = Path(__file__).resolve().parents[1]
if str(RAIZ) not in sys.path:
    sys.path.insert(0, str(RAIZ))

TEMPLATE = RAIZ / "templates" / "COLETA_REAJUSTE_OFICIAL.xlsx"

# ------------------------------------------------------------------ contrato
# Todos os valores abaixo foram extraidos do template em f8296f7.
PRIMEIRA_LINHA_UX2 = 90
ULTIMA_LINHA_UX2 = 166
ULTIMA_LINHA_TECNICA = 87

LINHAS_OCULTAS = {10, 11, 12, 13, 31, 40, 51}
# RESULTADOS-FINAL-1: o bloco "5. AJUSTES MANUAIS" (41:50) passou a ser
# OCULTO por decisao de apresentacao. Ocultar nao e apagar — o conteudo e
# as validacoes de C43:G50 continuam nas mesmas coordenadas, e quem prova
# isso e tests/test_resultados_final_1.py.
LINHAS_OCULTAS_AJUSTES_MANUAIS = set(range(41, 51))
LINHAS_OCULTAS |= LINHAS_OCULTAS_AJUSTES_MANUAIS
# openpyxl expoe o grupo <col min="10" max="14" hidden="1"/> sob a chave J.
GRUPO_OCULTO = ("J", 10, 14)                  # J:N; I volta a ser visivel

MERGES = [
    "A1:F1", "A2:F2", "A39:H39", "A53:H53", "A68:H68", "A69:H69",
    "A71:H71", "A79:H79", "A80:H80", "A81:H81", "A9:H9", "B7:H7",
    "C10:E10", "C11:E11", "C12:E12", "C13:E13", "C34:D34", "C35:D35",
    "C36:D36", "C37:D37", "C38:D38", "D5:E5", "E16:H21", "E23:H23",
    "E25:H25", "E34:H34", "E35:H38", "F10:G10", "F11:G11", "F12:G12",
    "F13:G13", "G1:H1", "G2:H2", "G3:H3",
]
# XLS-PC-VTA-ALIGN-1: o bloco 6 passou a ter 13 medidas (55:67) — as
# medidas 7 e 8 decompoem o retroativo em reconhecido / POTENCIAL /
# considerado no VTA. A linha 67 estava vazia e o bloco 7 segue na 68.
MERGES += [f"C{linha}:H{linha}" for linha in range(54, 68)]
# RESULTADOS-FINAL-1: o retroativo potencial ocupa E22:F22 (rotulo) e
# G22:H22 (valor) — faixa que estava vazia, entre os merges E16:H21 e
# E23:H23, que seguem intactos.
MERGES = sorted(MERGES + ["E22:F22", "G22:H22"])
CF_SQREFS = [
    "A7:H7", "B38", "C4", "D4:D5 D6:E6", "E4", "F4:G4 F5:H6", "G1", "H24",
    "H33", "H4", "H43:H50",
]
# RESULTADOS-FINAL-1: uma regra nova pinta E22:H22 de ambar somente quando
# ha retroativo potencial a mostrar (=$G$22<>""); fora do metodo PCs a
# faixa some em vez de exibir um bloco colorido vazio.
CF_SQREFS = sorted(CF_SQREFS + ["E22:H22"])
# VTA-POT-1: duas regras novas, ambas dentro do quadro 9 e ambas condicionadas
# a METODO_RETROATIVO="PCs" com potencial diferente de zero — pintam de
# amarelo-palha (#FFF4CC) so a parcela POTENCIAL (A84:C84) e a demonstracao
# "sem potencial + potencial = VTA" (C86). Financeiro e Itens nao ganham cor.
CF_SQREFS = sorted(CF_SQREFS + ["A84:C84", "C86"])
# XLS-PC-VTA-ALIGN-1: tres regras novas, todas condicionadas a haver
# parcela a mostrar (o gate esta na propria formula da celula, que sai
# vazia fora do metodo PC): A8:C8 e A61:B61 pintam de amarelo-palha
# (#FFF4CC) so o que e especificamente POTENCIAL; D8:E8 recebe o cinza
# claro dos cards, porque "considerado no VTA" NAO e so potencial.
CF_SQREFS = sorted(CF_SQREFS + ["A8:C8", "A61:B61", "D8:E8"])
# 143 do motor tecnico + E22/G22, as duas unicas formulas de apresentacao
# acrescentadas pela RESULTADOS-FINAL-1.
# 145 do checkpoint + B64 da sintese + B15/C15 condicionais do PC-UX-1 +
# a linha metodologica A70 restaurada. A71 continua formula, como no checkpoint.
# VTA-POT-1: +2 (148 -> 150) — A84 e C86 viram formulas de APRESENTACAO
# dentro do quadro 9. A aba segue terminando na linha 87 e 90:166 continua
# vazia: nenhuma linha, merge ou altura foi criada.
# XLS-PC-VTA-ALIGN-1: +8 (150 -> 158) — A4, A6, A8, C8, D8, E8, A81 e a
# 13a medida do bloco 6. Todas de APRESENTACAO e todas method-aware.
TOTAL_FORMULAS = 158

PRINT_AREA = "'RESULTADOS'!$A$1:$H$50"
MARGEM_LATERAL = 0.511811024
MARGEM_VERTICAL = 0.787401575
MARGEM_CABECALHO = 0.31496062

# Pinos do motor tecnico (item E do gate).
PINOS = {
    "B3": '=IF(OR($H$8="REVISE",$H$14="REVISE",$H$24="REVISE",$H$33="REVISE",'
          'COUNTIF($H$43:$H$50,"REVISE")>0),"REVISE",'
          'IF($H$33="ESTIMADO","ESTIMADO","VALIDADO"))',
    "B22": '=IF(COUNT(B16:B20)=0,"",ROUND(SUM(B16:B20),2))',
    "B37": '=IFERROR(INDEX($C$26:$C$30,MATCH(UPPER(CONTROLE!$B$2),$A$26:$A$30,0)),"")',
    "B38": '=IF(OR(B36="",B37=""),"",ROUND(B37-B36,2))',
    "B87": '=IF(OR($B$83="",$B$85="",$B$86=""),"",'
           'ROUND($B$86-($B$83+N($B$84)+$B$85),2))',
    "C5": '=IF(VTA_FINAL="","",VTA_FINAL)',
    "D5": "=$D$22",
    "D22": '=IFERROR(RETRO_OFICIAL,"")',
    "B86": '=IF(VTA_FINAL="","",VTA_FINAL)',
}

ROTULOS_C43_G50 = [
    "Retroativo manual oficial", "Ajuste do VTA", "VTA manual substitutivo",
    "Complemento histórico", "Complemento histórico", "Complemento histórico",
    "Complemento histórico", "Complemento histórico",
]


@pytest.fixture(scope="module")
def wb():
    return load_workbook(TEMPLATE, data_only=False)


@pytest.fixture(scope="module")
def ws(wb):
    return wb["RESULTADOS"]


# ------------------------------------------------------------------------ A
def test_a_aba_voltou_a_visibilidade_da_apresentacao_anterior(ws):
    """O motor tecnico volta a ser a propria aba, nao um anexo escondido."""
    ocultas = {
        r for r in range(1, ULTIMA_LINHA_TECNICA + 1)
        if ws.row_dimensions[r].hidden
    }
    assert ocultas == LINHAS_OCULTAS, (
        "a UX2 ocultava 1:89 inteiro; a apresentacao anterior oculta so "
        f"{sorted(LINHAS_OCULTAS)}"
    )
    ocultos = {c: (dim.min, dim.max)
               for c, dim in ws.column_dimensions.items() if dim.hidden}
    letra, minimo, maximo = GRUPO_OCULTO
    assert set(ocultos) == {letra}, (
        "a coluna I so era oculta para esconder os helpers I100/I125 da UX2; "
        f"ocultas agora: {sorted(ocultos)}"
    )
    assert ocultos[letra] == (minimo, maximo), "o grupo J:N foi dividido"


# ------------------------------------------------------------------------ B
def test_b_nao_existe_conteudo_ativo_na_faixa_da_camada_ux2(ws):
    """Nem valor, nem formula, nem rotulo sobrando em 90:166."""
    assert ws.max_row == ULTIMA_LINHA_TECNICA
    sobras = [
        cel.coordinate
        for linha in ws.iter_rows(min_row=PRIMEIRA_LINHA_UX2,
                                  max_row=ULTIMA_LINHA_UX2 + 40)
        for cel in linha
        if cel.value is not None
    ]
    assert sobras == [], f"restos da camada UX2 em {sobras[:10]}"


def test_b2_alturas_de_linha_da_camada_ux2_sumiram(ws):
    alturas = {
        r: ws.row_dimensions[r].height
        for r in range(PRIMEIRA_LINHA_UX2, ULTIMA_LINHA_UX2 + 1)
        if ws.row_dimensions[r].height is not None
    }
    assert alturas == {}, f"alturas remanescentes da UX2: {alturas}"


# ------------------------------------------------------------------------ C
def test_c_merges_e_formatacao_condicional_sao_os_do_checkpoint(ws):
    assert sorted(str(m) for m in ws.merged_cells.ranges) == MERGES
    assert sorted(str(r.sqref) for r in ws.conditional_formatting) == CF_SQREFS


# ------------------------------------------------------------------------ D
def test_d_print_setup_e_o_da_apresentacao_anterior(ws):
    setup = ws.page_setup
    area = ws.print_area
    assert (area if isinstance(area, list) else [area]) == [PRINT_AREA]
    assert setup.orientation == "landscape"
    assert setup.paperSize == 9
    # fit-to-page: 1 pagina de largura, altura livre (o doador nao imprimia
    # o motor tecnico numa segunda pagina).
    assert ws.sheet_properties.pageSetUpPr.fitToPage is True
    assert setup.fitToHeight == 0
    assert setup.fitToWidth in (None, 1)
    # a UX2 imprimia 2 paginas com quebra manual apos a 116.
    assert list(ws.row_breaks.brk) == []
    assert list(ws.col_breaks.brk) == []
    margens = ws.page_margins
    assert round(margens.left, 6) == round(MARGEM_LATERAL, 6)
    assert round(margens.right, 6) == round(MARGEM_LATERAL, 6)
    assert round(margens.top, 6) == round(MARGEM_VERTICAL, 6)
    assert round(margens.bottom, 6) == round(MARGEM_VERTICAL, 6)
    assert round(margens.header, 6) == round(MARGEM_CABECALHO, 6)
    assert round(margens.footer, 6) == round(MARGEM_CABECALHO, 6)


def test_d2_a_aba_abre_no_topo_e_nao_e_a_aba_ativa(ws, wb):
    """A UX2 abria a aba rolada em A90 e a deixava selecionada."""
    vista = ws.sheet_view
    assert vista.topLeftCell in (None, "A1")
    assert vista.tabSelected in (None, False)
    assert vista.showGridLines is False
    assert [s.activeCell for s in vista.selection] == ["D5"]
    assert wb.active.title != "RESULTADOS"


# ------------------------------------------------------------------------ E
def test_e_motor_tecnico_e_pinos_seguem_presentes(ws):
    for celula, formula in PINOS.items():
        assert ws[celula].value == formula, f"pino {celula} alterado"
    total = sum(
        1
        for linha in ws.iter_rows()
        for cel in linha
        if isinstance(cel.value, str) and cel.value.startswith("=")
    )
    assert total == TOTAL_FORMULAS, (
        f"a aba tem {total} formulas; a apresentacao anterior tem "
        f"{TOTAL_FORMULAS} (282 era a contagem com a camada UX2)"
    )


def test_e2_fator_historico_do_pr139_foi_preservado(ws):
    """Melhoria posterior ao checkpoint: NAO volta ao doador."""
    h5 = str(ws["H5"].value or "")
    assert h5.startswith("=IFERROR(IF(UPPER(CONTROLE!$B$2)")
    assert "parametros!$F$3" in h5
    assert h5 != '=IF(CONTROLE!$B$11="","",CONTROLE!$B$11)'
    assert "$H$5" in str(ws["H8"].value or "")
    assert "RESULTADOS!H5" in str(ws["C12"].value or "")


# ------------------------------------------------------------------------ F
def test_f_ajustes_manuais_c43_g50_intactos(ws):
    assert [ws[f"A{r}"].value for r in range(43, 51)] == ROTULOS_C43_G50
    validacoes = {
        (dv.type, str(dv.sqref)) for dv in ws.data_validations.dataValidation
    }
    assert ("list", "G43:G50") in validacoes
    assert ("decimal", "C46:C50") in validacoes
    # bloco de entrada do fiscal: continua sem formula em C43:G50
    formulas = [
        cel.coordinate
        for linha in ws.iter_rows(min_row=43, max_row=50, min_col=3, max_col=7)
        for cel in linha
        if isinstance(cel.value, str) and cel.value.startswith("=")
    ]
    assert formulas == []


# ------------------------------------------------------------------------ G
def test_g_vta_final_continua_canonico(wb, ws):
    nomes = {n: d.value for n, d in wb.defined_names.items()}
    assert nomes["VTA_FINAL"] == "MEMORIA_RESULTADOS!$B$26"
    assert ws["C5"].value == '=IF(VTA_FINAL="","",VTA_FINAL)'
    assert ws["B86"].value == '=IF(VTA_FINAL="","",VTA_FINAL)'


def test_g2_defined_names_invisiveis_foram_preservados(wb):
    """Os 14 names do PR #135 e o 15o (#136) continuam publicados."""
    nomes = {n: d.value for n, d in wb.defined_names.items()}
    assert nomes["EXECUTADO_APURADO"] == "RESULTADOS!$B$83"
    assert nomes["AJUSTES_DEVIDOS"] == "RESULTADOS!$B$84"
    assert nomes["CONFERENCIA_FORMACAO_VTA"] == "RESULTADOS!$B$87"
    assert nomes["RETROATIVO_POTENCIAL_PC"] == "MEMORIA_RESULTADOS!$T$38"
    mem = wb["MEMORIA_RESULTADOS"]
    assert mem["T38"].value == (
        '=ROUND(SUMIFS(itens_PC!$J$2:$J$5001,'
        'itens_PC!$B$2:$B$5001,"<="&$T$31),2)'
    )


# ------------------------------------------------------------------------ H
def test_h_b3_continua_sendo_a_fonte_canonica_do_status(ws):
    """A web le RESULTADOS!B3; a formula tem de existir e ficar visivel."""
    b3 = ws["B3"].value
    assert isinstance(b3, str) and b3.startswith("=")
    assert b3 == PINOS["B3"]
    assert ws.row_dimensions[3].hidden in (None, False), (
        "a UX2 escondia a linha 3 junto com todo o motor tecnico"
    )


# ------------------------------------------------------------------------ I/J
# Cenario aprovado do PR #140: data-base 05/02/2025, pedido C1 15/04/2026.
DATA_BASE = "05/02/2025"
PEDIDO_C1 = date(2026, 4, 15)


def _ciclo(numero, base, pedido, exata, financeiro):
    return {
        "ciclo": "C%d" % numero, "data_base": base, "data_pedido": pedido,
        "data_abertura_fisica_exata": exata, "financeiro_inicio": financeiro,
        "situacao": "TEMPESTIVO*", "situacao_aplicada": "TEMPESTIVO*",
        "objeto_analise_atual": True, "percentual_aplicado": 0.05,
        "fator": 1.05, "variacao": 0.05, "ciclo_calculado": True,
    }


def test_i_fotografia_de_c1_continua_na_data_exata():
    """PR #140 preservado: a Coleta gerada grava 15/04/2026, nao 01/02/2026."""
    from _coleta_oficial import gerar_coleta_oficial_preenchida

    payload = {
        "origem": "Reajustes Múltiplos", "tipo": "Represados", "indice": "IST",
        "data_base_original": DATA_BASE, "fator": 1.05, "fator_acumulado": 1.05,
        "variacao": 0.05,
        "ciclos": [_ciclo(1, DATA_BASE, "15/04/2026", "15/04/2026", "01/04/2026")],
    }
    par = load_workbook(
        io.BytesIO(gerar_coleta_oficial_preenchida(payload))
    )["parametros"]
    exata = par["I3"].value
    assert (exata.date() if hasattr(exata, "date") else exata) == PEDIDO_C1
    competencia = par["H3"].value
    competencia = competencia.date() if hasattr(competencia, "date") else competencia
    assert competencia == date(2026, 4, 1)      # cadeia financeira, sem pro-rata


def test_j_propagacao_juridica_preserva_o_dia_do_pedido():
    from _reajuste_utils import referencia_exata_pedido_subsequente

    assert referencia_exata_pedido_subsequente(PEDIDO_C1) == date(2027, 4, 15)


# ------------------------------------------------- aplicadores da UX2 sumiram
def test_aplicadores_da_apresentacao_descartada_nao_existem_mais():
    """Nenhuma ferramenta ativa pode reaplicar a camada aposentada."""
    for relativo in (
        "tools/aplicar_resultados_ux2_pr2.ps1",
        "tools/gerar_plano_resultados_ux2.py",
        "tools/resultados_ux2_plano.json",
    ):
        assert not (RAIZ / relativo).exists(), f"{relativo} voltou ao repo"
