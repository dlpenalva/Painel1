"""Entrada oficial única e amarração do estado ao arquivo que o originou.

Os testes exercitam a guarda de `_estado_upload` e o contrato do fluxo oficial
descrito em `pages/03_Valor_Global.py`. A tela do Streamlit não é executada: o
que se prova aqui é a regra que ela aplica — um arquivo, um resultado; arquivo
recusado não vira zero, vira diagnóstico.
"""

import io
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from _coleta_reajuste import eh_coleta_reajuste, gerar_coleta_reajuste, ler_coleta_reajuste
from _coleta_reajuste_documentos import adaptar_coleta_reajuste_para_documentos
from _estado_upload import (
    CHAVES_DERIVADAS_DO_UPLOAD,
    ORIGEM_COLETA_OFICIAL,
    ORIGEM_NAO_RECONHECIDA,
    VERSAO_CONTRATO_DADOS,
    documentos_liberados,
    limpar_estados_derivados,
    procedencia_registrada,
    registrar_upload,
    sha256_do_arquivo,
    upload_ja_processado,
)


ROOT = Path(__file__).resolve().parents[1]
DOCUMENTOS = (ROOT / "pages" / "03_Valor_Global.py").read_text(encoding="utf-8")


def coleta_oficial(valor_financeiro: float = 1000.0, indice: str = "IPCA") -> bytes:
    """Um Coleta_Reajuste.xlsx oficial, preenchido pelo caminho canônico."""

    dados = {
        "indice": indice,
        "ciclos": [
            {
                "ciclo": "C1",
                "data_base": "01/01/2024",
                "financeiro_inicio": "01/01/2025",
                "percentual_aplicado": 0.045,
            }
        ],
    }
    wb = load_workbook(io.BytesIO(gerar_coleta_reajuste(dados)), data_only=False)
    wb["financeiro"]["C2"] = valor_financeiro
    saida = io.BytesIO()
    wb.save(saida)
    return saida.getvalue()


def xlsx_nao_oficial() -> bytes:
    """Um .xlsx qualquer: extensão certa, estrutura alheia à Coleta."""

    wb = Workbook()
    ws = wb.active
    ws.title = "Planilha1"
    ws["A1"] = "Contrato"
    ws["B1"] = "Valor"
    ws["A2"] = "Fornecedor X"
    ws["B2"] = 999999.99
    saida = io.BytesIO()
    wb.save(saida)
    return saida.getvalue()


class ReconhecimentoDaEntradaOficialTests(unittest.TestCase):
    def test_coleta_oficial_e_reconhecida_e_produz_resultado(self):
        payload = coleta_oficial()
        self.assertTrue(eh_coleta_reajuste(payload))
        self.assertTrue(ler_coleta_reajuste(payload)["valido"])
        self.assertTrue(adaptar_coleta_reajuste_para_documentos(payload)["ok"])

    def test_xlsx_nao_oficial_nao_e_reconhecido_como_coleta(self):
        self.assertFalse(eh_coleta_reajuste(xlsx_nao_oficial()))

    def test_arquivo_corrompido_nao_e_reconhecido_e_nao_levanta(self):
        self.assertFalse(eh_coleta_reajuste(b"isto nao e um xlsx"))


class FluxoOficialDaPaginaTests(unittest.TestCase):
    """O texto da página é o contrato do fluxo: o ramo legado saiu dele."""

    def test_ramo_legado_nao_participa_do_fluxo_de_upload(self):
        self.assertIn("def processar_arquivo_coleta(bytes_arquivo):", DOCUMENTOS)
        # Definida, porém sem chamadores: nenhuma invocação da função.
        self.assertNotIn("= processar_arquivo_coleta(", DOCUMENTOS)
        self.assertNotIn("processar_arquivo_coleta(conteudo)", DOCUMENTOS)
        self.assertNotIn("Arquivo legado processado", DOCUMENTOS)

    def test_upload_amarra_o_estado_ao_hash_antes_de_decidir(self):
        self.assertIn("sha256 = sha256_do_arquivo(conteudo)", DOCUMENTOS)
        self.assertIn("if not upload_ja_processado(st.session_state, sha256):", DOCUMENTOS)
        self.assertIn("limpar_estados_derivados(st.session_state)", DOCUMENTOS)
        self.assertLess(
            DOCUMENTOS.index("sha256 = sha256_do_arquivo(conteudo)"),
            DOCUMENTOS.index("if not eh_coleta_reajuste(conteudo):"),
        )

    def test_arquivo_nao_reconhecido_e_rejeitado_sem_resultado(self):
        self.assertIn("origem=ORIGEM_NAO_RECONHECIDA", DOCUMENTOS)
        self.assertIn("MENSAGEM_ARQUIVO_NAO_RECONHECIDO", DOCUMENTOS)
        self.assertIn("DETALHES_ARQUIVO_NAO_RECONHECIDO", DOCUMENTOS)
        # A rejeição interrompe a página antes de qualquer documento.
        rejeicao = DOCUMENTOS.index('if procedencia and procedencia.get("origem") == ORIGEM_NAO_RECONHECIDA:')
        self.assertLess(rejeicao, DOCUMENTOS.index("Ações sobre os documentos"))
        self.assertIn("st.stop()", DOCUMENTOS[rejeicao:])

    def test_apuracao_oficial_continua_vindo_do_xls(self):
        self.assertIn("adaptar_coleta_reajuste_para_documentos(conteudo)", DOCUMENTOS)
        self.assertIn("origem=ORIGEM_COLETA_OFICIAL", DOCUMENTOS)


class HashDoArquivoTests(unittest.TestCase):
    def test_hash_e_integral_sobre_os_bytes(self):
        payload = coleta_oficial()
        self.assertEqual(sha256_do_arquivo(payload), sha256_do_arquivo(payload))
        self.assertEqual(len(sha256_do_arquivo(payload)), 64)

    def test_arquivos_diferentes_produzem_hashes_diferentes(self):
        self.assertNotEqual(
            sha256_do_arquivo(coleta_oficial(valor_financeiro=1000.0)),
            sha256_do_arquivo(coleta_oficial(valor_financeiro=2000.0)),
        )


class ProtecaoDoEstadoPorArquivoTests(unittest.TestCase):
    def estado_de_arquivo_valido(self, sha256: str = "hash-valido") -> dict:
        estado = {
            "resultado_valor_global": {"ok": True, "valor_atualizado_contrato": 1234.56},
            "diagnostico_coleta_v2": {"valido": True},
            "resultado_garantia": {"garantia_original": 100.0},
            "arquivo_relatorio_executivo_pdf": b"%PDF-anterior",
            "arquivo_garantia_pdf": b"%PDF-garantia-anterior",
            "saneador_texto": "texto do contrato anterior",
        }
        registrar_upload(estado, sha256=sha256, origem=ORIGEM_COLETA_OFICIAL, aceito=True)
        return estado

    def test_mesmo_hash_reaproveita_o_resultado_sem_reprocessar(self):
        estado = self.estado_de_arquivo_valido()
        self.assertTrue(upload_ja_processado(estado, "hash-valido"))

    def test_hash_diferente_nao_reaproveita_resultado_anterior(self):
        estado = self.estado_de_arquivo_valido()
        self.assertFalse(upload_ja_processado(estado, "outro-hash"))

    def test_versao_de_contrato_diferente_invalida_o_reaproveitamento(self):
        estado = self.estado_de_arquivo_valido()
        estado["procedencia_upload"]["versao_contrato"] = "v3.legado"
        self.assertFalse(upload_ja_processado(estado, "hash-valido"))

    def test_procedencia_sem_resultado_nao_e_reaproveitada(self):
        estado = self.estado_de_arquivo_valido()
        estado.pop("resultado_valor_global")
        self.assertFalse(upload_ja_processado(estado, "hash-valido"))

    def test_procedencia_registra_origem_versao_e_hash(self):
        estado = self.estado_de_arquivo_valido(sha256="abc123")
        procedencia = procedencia_registrada(estado)
        self.assertEqual(procedencia["sha256"], "abc123")
        self.assertEqual(procedencia["origem"], ORIGEM_COLETA_OFICIAL)
        self.assertEqual(procedencia["versao_contrato"], VERSAO_CONTRATO_DADOS)
        self.assertTrue(procedencia["aceito"])

    def test_troca_de_arquivo_apaga_resultado_confirmacoes_e_documentos(self):
        estado = self.estado_de_arquivo_valido()
        limpar_estados_derivados(estado)
        for chave in (
            "resultado_valor_global",
            "diagnostico_coleta_v2",
            "resultado_garantia",
            "arquivo_relatorio_executivo_pdf",
            "arquivo_garantia_pdf",
            "saneador_texto",
        ):
            self.assertNotIn(chave, estado, f"{chave} sobreviveu à troca de arquivo")

    def test_limpeza_nao_inventa_zero_nem_dataframe_vazio(self):
        estado = self.estado_de_arquivo_valido()
        limpar_estados_derivados(estado)
        # Ausência é ausência: a chave some, não vira 0 nem estrutura vazia.
        self.assertIsNone(estado.get("resultado_valor_global"))
        self.assertEqual([v for v in estado.values() if v == 0], [])

    def test_limpeza_preserva_dados_que_nao_derivam_do_upload(self):
        estado = self.estado_de_arquivo_valido()
        estado["dados_admissibilidade"] = {"indice": "IPCA"}
        estado["checklist_processual"] = "tabela digitada pelo usuário"
        estado["infos_previas_df"] = "tabela digitada pelo usuário"
        limpar_estados_derivados(estado)
        self.assertEqual(estado["dados_admissibilidade"], {"indice": "IPCA"})
        self.assertIn("checklist_processual", estado)
        self.assertIn("infos_previas_df", estado)

    def test_todo_buffer_documental_esta_coberto_pela_limpeza(self):
        buffers = [c for c in CHAVES_DERIVADAS_DO_UPLOAD if c.startswith("arquivo_")]
        self.assertIn("arquivo_planilha_executiva_xlsx", buffers)
        self.assertIn("arquivo_dou_docx", buffers)
        self.assertIn("arquivo_minuta_apostilamento_docx", buffers)
        self.assertGreaterEqual(len(buffers), 12)


class BloqueioDocumentalTests(unittest.TestCase):
    def test_arquivo_incompativel_nao_libera_documentos(self):
        estado = {}
        registrar_upload(
            estado,
            sha256=sha256_do_arquivo(xlsx_nao_oficial()),
            origem=ORIGEM_NAO_RECONHECIDA,
            aceito=False,
            motivo="não reconhecido",
        )
        self.assertFalse(documentos_liberados(estado))
        self.assertIsNone(estado.get("resultado_valor_global"))

    def test_coleta_oficial_reprovada_no_diagnostico_nao_libera_documentos(self):
        estado = {"diagnostico_coleta_v2": {"valido": False}}
        registrar_upload(
            estado,
            sha256="hash",
            origem=ORIGEM_COLETA_OFICIAL,
            aceito=False,
            motivo="reprovada",
        )
        self.assertFalse(documentos_liberados(estado))

    def test_sem_upload_algum_nao_ha_documento_liberado(self):
        self.assertFalse(documentos_liberados({}))

    def test_coleta_oficial_aceita_libera_documentos(self):
        estado = {"resultado_valor_global": {"ok": True}}
        registrar_upload(estado, sha256="hash", origem=ORIGEM_COLETA_OFICIAL, aceito=True)
        self.assertTrue(documentos_liberados(estado))


class SequenciasDeUploadTests(unittest.TestCase):
    """As trocas exigidas: válido→inválido, inválido→válido, válido→válido."""

    def simular_upload(self, estado: dict, conteudo: bytes) -> None:
        """Reproduz a decisão do fluxo oficial de `03_Valor_Global.py`."""

        sha256 = sha256_do_arquivo(conteudo)
        if not upload_ja_processado(estado, sha256):
            limpar_estados_derivados(estado)
        if not eh_coleta_reajuste(conteudo):
            registrar_upload(
                estado, sha256=sha256, origem=ORIGEM_NAO_RECONHECIDA, aceito=False, motivo="não reconhecido"
            )
            return
        diagnostico = ler_coleta_reajuste(conteudo)
        estado["diagnostico_coleta_v2"] = diagnostico
        if diagnostico.get("valido"):
            estado["resultado_valor_global"] = adaptar_coleta_reajuste_para_documentos(conteudo)
            registrar_upload(estado, sha256=sha256, origem=ORIGEM_COLETA_OFICIAL, aceito=True)
        else:
            registrar_upload(
                estado, sha256=sha256, origem=ORIGEM_COLETA_OFICIAL, aceito=False, motivo="reprovada"
            )

    def test_upload_de_coleta_oficial_valida_produz_resultado_e_libera_documentos(self):
        estado = {}
        self.simular_upload(estado, coleta_oficial())
        self.assertTrue(estado["resultado_valor_global"]["ok"])
        self.assertTrue(documentos_liberados(estado))

    def test_arquivo_invalido_apos_valido_apaga_resultado_e_bloqueia_documentos(self):
        estado = {}
        self.simular_upload(estado, coleta_oficial())
        estado["arquivo_relatorio_executivo_pdf"] = b"%PDF-do-contrato-anterior"
        estado["resultado_garantia"] = {"garantia_original": 500.0}

        self.simular_upload(estado, xlsx_nao_oficial())

        self.assertIsNone(estado.get("resultado_valor_global"))
        self.assertIsNone(estado.get("arquivo_relatorio_executivo_pdf"))
        self.assertIsNone(estado.get("resultado_garantia"))
        self.assertFalse(documentos_liberados(estado))
        self.assertEqual(procedencia_registrada(estado)["origem"], ORIGEM_NAO_RECONHECIDA)

    def test_arquivo_valido_apos_invalido_recupera_o_fluxo(self):
        estado = {}
        self.simular_upload(estado, xlsx_nao_oficial())
        self.assertFalse(documentos_liberados(estado))

        self.simular_upload(estado, coleta_oficial())

        self.assertTrue(estado["resultado_valor_global"]["ok"])
        self.assertTrue(documentos_liberados(estado))
        self.assertEqual(procedencia_registrada(estado)["origem"], ORIGEM_COLETA_OFICIAL)

    def test_troca_entre_dois_arquivos_validos_nao_mistura_contratos(self):
        primeiro = coleta_oficial(valor_financeiro=1000.0)
        segundo = coleta_oficial(valor_financeiro=7777.0)
        estado = {}

        self.simular_upload(estado, primeiro)
        estado["arquivo_garantia_pdf"] = b"%PDF-do-primeiro-contrato"
        procedencia_primeiro = dict(procedencia_registrada(estado))

        self.simular_upload(estado, segundo)

        self.assertIsNone(estado.get("arquivo_garantia_pdf"), "documento do primeiro contrato sobreviveu")
        self.assertNotEqual(procedencia_registrada(estado)["sha256"], procedencia_primeiro["sha256"])
        self.assertEqual(procedencia_registrada(estado)["sha256"], sha256_do_arquivo(segundo))

    def test_repetir_o_mesmo_arquivo_preserva_o_estado_ja_apurado(self):
        payload = coleta_oficial()
        estado = {}
        self.simular_upload(estado, payload)
        estado["arquivo_garantia_pdf"] = b"%PDF-deste-mesmo-contrato"

        self.simular_upload(estado, payload)

        # Mesmo hash: o documento já emitido para este arquivo continua válido.
        self.assertEqual(estado["arquivo_garantia_pdf"], b"%PDF-deste-mesmo-contrato")
        self.assertTrue(documentos_liberados(estado))


class ContratoDeDadosTests(unittest.TestCase):
    def test_contrato_oficial_preserva_as_68_chaves(self):
        resultado = adaptar_coleta_reajuste_para_documentos(coleta_oficial())
        self.assertEqual(len(resultado), 68)

    def test_valor_atualizado_contrato_integra_obrigatoriamente_o_contrato(self):
        resultado = adaptar_coleta_reajuste_para_documentos(coleta_oficial())
        self.assertIn("valor_atualizado_contrato", resultado)
        self.assertIsNotNone(resultado["valor_atualizado_contrato"])

    def test_chaves_inexistentes_nao_foram_criadas(self):
        resultado = adaptar_coleta_reajuste_para_documentos(coleta_oficial())
        self.assertNotIn("valor_global_estoque", resultado)
        self.assertNotIn("valor_global_contrato", resultado)


if __name__ == "__main__":
    unittest.main()
