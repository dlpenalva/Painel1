# -*- coding: utf-8 -*-
"""EF-G1 — `financeiro!G` como fonte canonica do efeito financeiro.

REGRA PETREA COBERTA AQUI:

  G = "Sim"  -> a competencia recebe o reajuste; a diferenca pode compor o
                retroativo; o VTA considera o valor ATUALIZADO.
  G = "Nao"  -> a competencia NAO gera diferenca de reajuste; retroativo zero;
                a execucao NAO desaparece: seu valor SEM reajuste continua
                compondo integralmente o VTA.
  G vazio    -> inconsistencia de preenchimento, UMA unica mensagem.
  G invalido -> inconsistencia de preenchimento, UMA unica mensagem.

E a decisao de produto: a ferramenta NAO tenta descobrir se "Sim"/"Nao" foi
marcado automaticamente ou manualmente. Nenhum aviso de "ajuste manual"
sobrevive em lugar nenhum do fluxo.
"""
from __future__ import annotations

import importlib.util
import io
import sys
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

RAIZ = Path(__file__).resolve().parents[1]
if str(RAIZ) not in sys.path:
    sys.path.insert(0, str(RAIZ))

from _coleta_oficial import gerar_coleta_oficial_preenchida  # noqa: E402
from _coleta_reajuste import ler_coleta_reajuste  # noqa: E402
from _coleta_reajuste_documentos import (  # noqa: E402
    adaptar_coleta_reajuste_para_documentos,
    processar_coleta_oficial_runtime,
)
from _leitor_masterfile_v10 import _ler_parcelas_sombra_financeiro  # noqa: E402
from _motor_composicao_vta import _parcelas_financeiro_por_ciclo  # noqa: E402

FATOR = 1.10
NOMINAL = 100.0


# ---------------------------------------------------------------------------
# infraestrutura minima (mesmo padrao dos testes de efeito financeiro ja
# homologados: gerar a Coleta oficial e materializar competencias reais)
# ---------------------------------------------------------------------------
def _dados(inicio_efeito: str = "18/04/2024") -> dict:
    return {
        "origem": "Reajuste Simples",
        "indice": "IST",
        "data_base_original": "01/02/2023",
        "data_corte": "31/01/2025",
        "ciclos": [{
            "ciclo": "C1",
            "data_inicio": "01/02/2024",
            "data_fim": "31/01/2025",
            "data_pedido": "10/03/2024",
            "financeiro_inicio": inicio_efeito,
            "percentual_aplicado": 0.10,
            "objeto_analise_atual": True,
            "situacao": "TEMPESTIVO",
        }],
    }


def _wb(dados: dict | None = None):
    return load_workbook(
        io.BytesIO(gerar_coleta_oficial_preenchida(dados or _dados())),
        data_only=False,
    )


def _bytes(wb) -> bytes:
    saida = io.BytesIO()
    wb.save(saida)
    return saida.getvalue()


def _linha(ws, ano: int, mes: int) -> int:
    for row in range(2, 74):
        valor = ws[f"A{row}"].value
        if valor and (valor.year, valor.month) == (ano, mes):
            return row
    raise AssertionError(f"competencia {mes:02d}/{ano} ausente")


def _materializar(wb, ano, mes, *, nominal=NOMINAL, efeito="Nao", fator=FATOR) -> int:
    """Escreve a competencia como o Excel a deixaria: E e F derivados de G."""
    ws = wb["financeiro"]
    row = _linha(ws, ano, mes)
    ws[f"B{row}"] = "c1"
    ws[f"C{row}"] = nominal
    ws[f"D{row}"] = fator
    ws[f"G{row}"] = efeito
    ws[f"E{row}"] = round(nominal * fator, 2) if efeito == "Sim" else nominal
    ws[f"F{row}"] = round(nominal * fator - nominal, 2) if efeito == "Sim" else 0.0
    return row


def _textos_do_diagnostico(diagnostico) -> list[str]:
    chaves = (
        "avisos", "pendencias", "bloqueios_criticos", "inconsistencias",
        "bloqueios_estruturais", "lacunas_apuracao",
    )
    return [
        str(item)
        for chave in chaves
        for item in (diagnostico.get(chave) or ())
    ]


def _avisos_de_autoria(diagnostico) -> list[str]:
    return [
        texto for texto in _textos_do_diagnostico(diagnostico)
        if "ajustado manualmente" in texto or "ajustada manualmente" in texto
    ]


def _parcela_base(parcelas, row):
    return next(
        p for p in parcelas
        if p["linha"] == row and ":base:" in str(p["identificador"])
    )


def _parcelas_delta(parcelas, row):
    return [
        p for p in parcelas
        if p["linha"] == row and ":delta:" in str(p["identificador"])
    ]


# ---------------------------------------------------------------------------
# TESTE A — G = "Nao"
# ---------------------------------------------------------------------------
def test_a_g_nao_mantem_valor_base_sem_delta():
    wb = _wb()
    row = _materializar(wb, 2024, 4, efeito="Nao")
    ws = wb["financeiro"]

    # o proprio XLS ja decide por G: sem reajuste, sem delta.
    assert ws[f"E{row}"].value == NOMINAL
    assert ws[f"F{row}"].value == 0.0

    parcelas = [p for p in _ler_parcelas_sombra_financeiro(wb) if p["linha"] == row]
    base = _parcela_base(parcelas, row)
    assert base["valor"] == NOMINAL
    # valor operacional considerado = valor-base: a execucao NAO desaparece.
    assert base["valor_atualizado"] == NOMINAL
    # nenhuma parcela de retroativo para esta competencia.
    assert _parcelas_delta(parcelas, row) == []


# ---------------------------------------------------------------------------
# TESTE B — G = "Sim"
# ---------------------------------------------------------------------------
def test_b_g_sim_aplica_reajuste_sem_dupla_contagem():
    wb = _wb()
    row = _materializar(wb, 2024, 4, efeito="Sim")
    ws = wb["financeiro"]
    assert ws[f"E{row}"].value == pytest.approx(110.0, abs=0.01)
    assert ws[f"F{row}"].value == pytest.approx(10.0, abs=0.01)

    parcelas = [p for p in _ler_parcelas_sombra_financeiro(wb) if p["linha"] == row]
    base = _parcela_base(parcelas, row)
    deltas = _parcelas_delta(parcelas, row)
    assert base["valor"] == NOMINAL
    assert base["valor_atualizado"] == pytest.approx(110.0, abs=0.01)
    assert len(deltas) == 1
    assert deltas[0]["valor"] == pytest.approx(10.0, abs=0.01)

    # A composicao do VTA soma UMA vez: 110,00 — nunca 100 + 10 + 10.
    agregado = _parcelas_financeiro_por_ciclo(
        {"vta_sombra": {"parcelas_computadas": parcelas}}
    )
    assert agregado["C1"]["pago"] == pytest.approx(100.0, abs=0.01)
    assert agregado["C1"]["considerado"] == pytest.approx(110.0, abs=0.01)


# ---------------------------------------------------------------------------
# TESTE C — "Sim" -> "Nao"
# ---------------------------------------------------------------------------
def test_c_sim_para_nao_zera_retroativo_e_preserva_execucao_no_vta():
    wb = _wb()
    row = _materializar(wb, 2024, 4, efeito="Sim")
    antes = [p for p in _ler_parcelas_sombra_financeiro(wb) if p["linha"] == row]
    assert _parcelas_delta(antes, row)          # havia retroativo

    _materializar(wb, 2024, 4, efeito="Nao")    # o fiscal troca a marcacao
    depois = [p for p in _ler_parcelas_sombra_financeiro(wb) if p["linha"] == row]

    diagnostico = ler_coleta_reajuste(_bytes(wb))
    assert _avisos_de_autoria(diagnostico) == []          # nenhum aviso de autoria
    assert diagnostico["bloqueios_criticos"] == []        # decisao respeitada
    assert _parcelas_delta(depois, row) == []             # retroativo = 0
    assert _parcela_base(depois, row)["valor_atualizado"] == NOMINAL  # segue no VTA


# ---------------------------------------------------------------------------
# TESTE D — "Nao" -> "Sim"
# ---------------------------------------------------------------------------
def test_d_nao_para_sim_passa_a_reajustar_sem_dupla_contagem():
    wb = _wb()
    row = _materializar(wb, 2024, 4, efeito="Nao")
    assert _parcelas_delta(_ler_parcelas_sombra_financeiro(wb), row) == []

    _materializar(wb, 2024, 4, efeito="Sim")
    parcelas = [p for p in _ler_parcelas_sombra_financeiro(wb) if p["linha"] == row]

    diagnostico = ler_coleta_reajuste(_bytes(wb))
    assert _avisos_de_autoria(diagnostico) == []
    assert diagnostico["bloqueios_criticos"] == []
    assert _parcelas_delta(parcelas, row)[0]["valor"] == pytest.approx(10.0, abs=0.01)

    agregado = _parcelas_financeiro_por_ciclo(
        {"vta_sombra": {"parcelas_computadas": parcelas}}
    )
    assert agregado["C1"]["considerado"] == pytest.approx(110.0, abs=0.01)


# ---------------------------------------------------------------------------
# TESTE E — G vazio
# ---------------------------------------------------------------------------
def test_e_g_vazio_bloqueia_com_mensagem_unica_e_identificavel():
    wb = _wb()
    ws = wb["financeiro"]
    for mes in (4, 5, 6):
        row = _linha(ws, 2024, mes)
        ws[f"C{row}"] = NOMINAL
        ws[f"G{row}"] = None

    diagnostico = ler_coleta_reajuste(_bytes(wb))
    mensagens = [
        item for item in diagnostico["bloqueios_criticos"]
        if "não preenchido" in item
    ]
    # UMA unica mensagem para as tres competencias.
    assert len(mensagens) == 1
    for mes in (4, 5, 6):
        assert f"{mes:02d}/2024" in mensagens[0]
    assert "coluna G" in mensagens[0]
    # inconsistencia real continua barrando a formalizacao.
    assert diagnostico["pronto_para_consolidar"] is False
    assert diagnostico["status_base"] == "ANALISE_COM_INCONSISTENCIAS"


# ---------------------------------------------------------------------------
# TESTE F — G invalido
# ---------------------------------------------------------------------------
@pytest.mark.parametrize("invalido", ["Talvez", "X", "1"])
def test_f_g_invalido_bloqueia_com_mensagem_unica(invalido):
    wb = _wb()
    ws = wb["financeiro"]
    for mes in (4, 5):
        row = _linha(ws, 2024, mes)
        ws[f"C{row}"] = NOMINAL
        ws[f"G{row}"] = invalido

    diagnostico = ler_coleta_reajuste(_bytes(wb))
    mensagens = [
        item for item in diagnostico["bloqueios_criticos"]
        if "valor inválido" in item and "Efeito financeiro" in item
    ]
    assert len(mensagens) == 1
    assert "Sim ou Nao" in mensagens[0]
    assert "04/2024" in mensagens[0] and "05/2024" in mensagens[0]
    assert diagnostico["pronto_para_consolidar"] is False


# ---------------------------------------------------------------------------
# TESTE G — falsos positivos conhecidos
# ---------------------------------------------------------------------------
def _dados_multiciclo() -> dict:
    return {
        "origem": "Reajuste com Represados",
        "indice": "IST",
        "data_base_original": "01/02/2022",
        "data_corte": "31/01/2027",
        "ciclos": [
            {"ciclo": "C1", "data_inicio": "01/02/2023", "data_fim": "31/01/2024",
             "data_pedido": "10/03/2023", "financeiro_inicio": "01/04/2023",
             "percentual_aplicado": 0.08, "situacao": "TEMPESTIVO"},
            {"ciclo": "C2", "data_inicio": "01/02/2024", "data_fim": "31/01/2025",
             "data_pedido": "10/03/2024", "financeiro_inicio": "01/04/2024",
             "percentual_aplicado": 0.06, "situacao": "TEMPESTIVO"},
            {"ciclo": "C3", "data_inicio": "01/02/2025", "data_fim": "31/01/2026",
             "data_pedido": "10/03/2025", "financeiro_inicio": "01/04/2025",
             "percentual_aplicado": 0.05, "situacao": "TEMPESTIVO"},
            {"ciclo": "C4", "data_inicio": "01/02/2026", "data_fim": "31/01/2027",
             "data_pedido": "10/03/2026", "financeiro_inicio": "01/04/2026",
             "percentual_aplicado": 0.04, "situacao": "TEMPESTIVO",
             "objeto_analise_atual": True},
        ],
    }


@pytest.mark.parametrize("ano,mes", [(2024, 2), (2025, 2), (2026, 2)])
def test_g_competencias_de_fronteira_nao_geram_nenhum_aviso(ano, mes):
    """Os falsos positivos relatados (02/2024, 02/2025, 02/2026) sumiram."""
    wb = _wb(_dados_multiciclo())
    ws = wb["financeiro"]
    row = _linha(ws, ano, mes)
    ws[f"C{row}"] = NOMINAL                    # competencia paga, G intocado

    diagnostico = ler_coleta_reajuste(_bytes(wb))
    assert _avisos_de_autoria(diagnostico) == []
    texto = " | ".join(_textos_do_diagnostico(diagnostico))
    assert f"{mes:02d}/{ano}" not in texto


def test_g_arquivo_recem_gerado_nao_acusa_nada():
    diagnostico = ler_coleta_reajuste(
        gerar_coleta_oficial_preenchida(_dados_multiciclo())
    )
    assert _avisos_de_autoria(diagnostico) == []


# ---------------------------------------------------------------------------
# TESTE H — o aviso nao existe mais na web
# ---------------------------------------------------------------------------
def test_h_nenhum_modulo_do_fluxo_renderiza_o_aviso_removido():
    proibidos = (
        "Marcacao de efeito financeiro ajustada manualmente",
        "Efeito financeiro ajustado manualmente",
        "render_avisos_override_efeito_financeiro",
        "PREFIXO_AVISO_OVERRIDE_EFEITO_FINANCEIRO",
    )
    alvos = [
        RAIZ / "_ui_utils.py",
        RAIZ / "_coleta_reajuste.py",
        RAIZ / "_coleta_reajuste_documentos.py",
        RAIZ / "pages" / "03_Valor_Global.py",
    ]
    for alvo in alvos:
        fonte = alvo.read_text(encoding="utf-8")
        for termo in proibidos:
            assert termo not in fonte, f"{termo} ainda presente em {alvo.name}"


# ---------------------------------------------------------------------------
# TESTE I — regressao do VTA (mistura Sim/Nao)
# ---------------------------------------------------------------------------
def test_i_vta_soma_base_dos_nao_e_atualizado_dos_sim():
    wb = _wb()
    _materializar(wb, 2024, 2, nominal=100.0, efeito="Nao")
    _materializar(wb, 2024, 3, nominal=200.0, efeito="Nao")
    _materializar(wb, 2024, 4, nominal=300.0, efeito="Sim")
    _materializar(wb, 2024, 5, nominal=400.0, efeito="Sim")

    parcelas = _ler_parcelas_sombra_financeiro(wb)
    agregado = _parcelas_financeiro_por_ciclo(
        {"vta_sombra": {"parcelas_computadas": parcelas}}
    )
    esperado_pago = 100.0 + 200.0 + 300.0 + 400.0
    esperado_considerado = 100.0 + 200.0 + 330.0 + 440.0

    total_pago = sum(reg["pago"] for reg in agregado.values())
    total_considerado = sum(reg["considerado"] for reg in agregado.values())
    # nenhuma execucao se perde...
    assert total_pago == pytest.approx(esperado_pago, abs=0.01)
    # ...e o reajuste entra UMA unica vez (nao 1.070 + 70).
    assert total_considerado == pytest.approx(esperado_considerado, abs=0.01)
    assert total_considerado - total_pago == pytest.approx(70.0, abs=0.01)


# ---------------------------------------------------------------------------
# TESTE J — regressao do retroativo
# ---------------------------------------------------------------------------
def test_j_somente_g_sim_gera_retroativo():
    wb = _wb()
    _materializar(wb, 2024, 2, nominal=100.0, efeito="Nao")
    _materializar(wb, 2024, 4, nominal=300.0, efeito="Sim")
    payload = _bytes(wb)

    diagnostico = ler_coleta_reajuste(payload)
    resultado = adaptar_coleta_reajuste_para_documentos(payload, diagnostico=diagnostico)

    # competencias sem efeito: demonstradas, com valor-base integral.
    df = resultado["df_meses_sem_efeito_financeiro"]
    assert resultado["quantidade_meses_sem_efeito_financeiro"] == 1
    assert resultado["valor_total_sem_efeito_financeiro"] == pytest.approx(100.0, abs=0.01)
    assert set(df["Efeito financeiro"]) == {"Nao"}

    # o pago total continua contendo as duas competencias (nada some).
    assert resultado["total_pago_faturado"] == pytest.approx(400.0, abs=0.01)
    # devido = 100 (sem efeito) + 330 (com efeito).
    assert resultado["total_devido_reajustado"] == pytest.approx(430.0, abs=0.01)


def test_j_runtime_completo_respeita_g_sem_aviso_de_autoria():
    wb = _wb()
    _materializar(wb, 2024, 4, nominal=100.0, efeito="Nao")
    resultado, diagnostico = processar_coleta_oficial_runtime(_bytes(wb))
    assert _avisos_de_autoria(diagnostico) == []
    assert resultado["valor_represado_a_pagar"] == pytest.approx(0.0, abs=0.01)


# ---------------------------------------------------------------------------
# Caminho legado da pagina 03 (nao executado pelo upload real, mas mantido
# alinhado a regra petrea: G decide, e so ele).
# ---------------------------------------------------------------------------
@pytest.fixture(scope="module")
def pagina():
    spec = importlib.util.spec_from_file_location(
        "pagina_valor_global_ef_g1", RAIZ / "pages" / "03_Valor_Global.py"
    )
    modulo = importlib.util.module_from_spec(spec)
    try:
        spec.loader.exec_module(modulo)
    except BaseException:          # a secao imperativa do Streamlit para aqui
        pass
    assert hasattr(modulo, "financeiro_com_efeito_financeiro")
    return modulo


def _ciclos_c1():
    return pd.DataFrame([{
        "Ciclo": "C1",
        "Início financeiro": "04/2024",
        "Situação": "TEMPESTIVO",
        "Tratamento financeiro do ciclo": "Apurar",
        "Fator ciclo efetivo": FATOR,
        "Fator acumulado efetivo": FATOR,
        "Fator acumulado": FATOR,
        "Fator": FATOR,
        "Variação": 0.10,
    }])


@pytest.mark.parametrize("efeito,teorico,delta", [("Sim", 110.0, 10.0), ("Nao", 100.0, 0.0)])
def test_legado_pagina_respeita_a_coluna_g(pagina, efeito, teorico, delta):
    """Mesmo na competencia 02/2024 (anterior ao inicio financeiro), G manda."""
    df = pd.DataFrame([{
        "Ciclo": "C1",
        "Competência": "02/2024",
        "Valor pago/faturado": NOMINAL,
        "Efeito financeiro": efeito,
    }])
    saida = pagina.financeiro_com_efeito_financeiro(df, _ciclos_c1())
    assert saida.loc[0, "Valor teórico calculado"] == pytest.approx(teorico, abs=0.01)
    assert saida.loc[0, "Delta computável"] == pytest.approx(delta, abs=0.01)
    # a execucao permanece integralmente, com efeito ou sem ele.
    assert saida.loc[0, "Valor pago/faturado"] == pytest.approx(NOMINAL, abs=0.01)


def test_legado_pagina_sem_coluna_canonica_mantem_reconstrucao_historica(pagina):
    """Base legada sem EFEITO_FINANCEIRO: comportamento anterior, sem regressao."""
    df = pd.DataFrame([{
        "Ciclo": "C1",
        "Competência": "02/2024",
        "Valor pago/faturado": NOMINAL,
    }])
    saida = pagina.financeiro_com_efeito_financeiro(df, _ciclos_c1())
    assert saida.loc[0, "Competência sem efeito financeiro?"] == "Sim"
    assert saida.loc[0, "Delta computável"] == pytest.approx(0.0, abs=0.01)


def test_legado_pagina_le_a_coluna_g_do_xls(pagina):
    wb = _wb()
    _materializar(wb, 2024, 2, nominal=100.0, efeito="Nao")
    _materializar(wb, 2024, 4, nominal=300.0, efeito="Sim")
    payload = _bytes(wb)
    xls = pd.ExcelFile(io.BytesIO(payload))
    df = pagina.ler_financeiro(payload, xls, pd.DataFrame())
    assert "Efeito financeiro" in df.columns
    assert set(df["Efeito financeiro"]) == {"Nao", "Sim"}
