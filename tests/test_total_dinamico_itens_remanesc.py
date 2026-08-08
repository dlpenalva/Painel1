# -*- coding: utf-8 -*-
"""REGRA PERMANENTE — linha dinamica TOTAL de itens_Remanesc.

A primeira linha vazia apos o ultimo ITEM em itens_Remanesc e a linha dinamica
TOTAL. Todas as colunas de VALOR financeiro aplicaveis devem totalizar os
itens ate essa linha. D nao pode ser a unica coluna que totaliza. A regra deve
funcionar independentemente da quantidade de itens. VU_ORIGINAL nao e
totalizavel.

Este bug ja ocorreu duas vezes: as colunas F/H/J/L/N/P/R/AC ficaram com a
totalizacao desativada por IF(FALSE,...) e T (VALOR_EXECUTADO_C4) ficou sem
formula. Este teste falha se qualquer logica equivalente voltar.

Camadas:
  1. estrutura (openpyxl, sempre roda): nenhuma IF(FALSE,...) nas colunas
     financeiras; deteccao dinamica presente em TODAS elas (3:200); fallback
     de lotacao maxima (201) com guarda de coluna vazia;
  2. cenarios com 1, 3 e 20 itens: a linha TOTAL e localizada
     programaticamente (primeira linha vazia apos o ultimo item) e a formula
     daquela linha, em cada coluna financeira, totaliza exatamente os itens
     existentes;
  3. valores reais (Excel COM, opt-in RUN_EXCEL_INTEGRATION=1): caso real de
     3 itens — D=2.372.260,00 F=2.106.660,00 H=1.563.535,00 J=1.344.654,25
     N=543.125,00 P=274.355,00 AC=265.600,00; colunas integralmente sem dados
     (L/R/T) permanecem vazias, sem zero financeiro inventado.
"""
from __future__ import annotations

import os
import tempfile
from pathlib import Path

import pytest
from openpyxl import load_workbook

RAIZ = Path(__file__).resolve().parents[1]
TEMPLATE = RAIZ / "templates" / "COLETA_REAJUSTE_OFICIAL.xlsx"

# Colunas de VALOR financeiro da tabela (VU_ORIGINAL = C fica de fora: preco
# unitario nao e coluna de total financeiro).
COLUNAS_FINANCEIRAS = ("D", "F", "H", "J", "L", "N", "P", "R", "T", "AC")
# As que ja falharam por IF(FALSE,...)/ausencia de formula (D e a referencia).
COLUNAS_CORRIGIDAS = tuple(c for c in COLUNAS_FINANCEIRAS if c != "D")

FIM = 200

com = pytest.mark.skipif(
    os.environ.get("RUN_EXCEL_INTEGRATION") != "1",
    reason="defina RUN_EXCEL_INTEGRATION=1 para Excel COM",
)


@pytest.fixture(scope="module")
def ws():
    wb = load_workbook(TEMPLATE)
    yield wb["itens_Remanesc"]
    wb.close()


def _formula(ws, col: str, linha: int) -> str:
    return str(ws[f"{col}{linha}"].value or "")


def _linha_total(ws, itens: int) -> int:
    """Localiza programaticamente a linha TOTAL: primeira A vazia apos o
    ultimo item (itens comecam na linha 2)."""
    linha = 2
    while ws[f"A{linha}"].value not in (None, ""):
        linha += 1
    assert linha == itens + 2
    return linha


# ---------------------------------------------------------------------------
# 1. Estrutura permanente do template
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("col", COLUNAS_FINANCEIRAS)
def test_nenhuma_totalizacao_desativada_por_if_false(ws, col):
    """IF(FALSE,...) desabilitando a totalizacao NAO pode voltar (3:201).

    (D2 legada mantem um IF(FALSE inofensivo — a linha 2 nunca e a TOTAL —
    por isso a varredura obrigatoria comeca na linha 3.)
    """
    for linha in range(3, FIM + 2):
        assert "IF(FALSE" not in _formula(ws, col, linha), f"{col}{linha}"


@pytest.mark.parametrize("col", COLUNAS_CORRIGIDAS)
def test_colunas_corrigidas_sem_if_false_nem_na_linha_2(ws, col):
    assert "IF(FALSE" not in _formula(ws, col, 2), f"{col}2"


@pytest.mark.parametrize("col", COLUNAS_FINANCEIRAS)
def test_toda_coluna_financeira_tem_formula_em_2_a_201(ws, col):
    """T ficou sem formula em 2:200 uma vez; nenhuma coluna pode regredir."""
    for linha in range(2, FIM + 2):
        formula = _formula(ws, col, linha)
        assert formula.startswith("="), f"{col}{linha} sem formula"


@pytest.mark.parametrize("col", COLUNAS_FINANCEIRAS)
def test_deteccao_dinamica_em_todas_as_linhas_candidatas(ws, col):
    """Toda linha 3:200 detecta 'sou a primeira vazia apos o ultimo item'."""
    for linha in range(3, FIM + 1):
        formula = _formula(ws, col, linha)
        assert f'IF(AND(A{linha}="",A{linha - 1}<>""' in formula, f"{col}{linha}"
        assert "ROUND(SUMIF(" in formula, f"{col}{linha}"


@pytest.mark.parametrize("col", COLUNAS_CORRIGIDAS)
def test_coluna_vazia_permanece_vazia_no_total(ws, col):
    """Guarda COUNT: coluna integralmente nao aplicavel nunca vira 0,00."""
    for linha in (3, 100, FIM):
        assert f"IF(COUNT({col}$2:{col}{linha - 1})=0," in _formula(ws, col, linha)
    assert f"IF(COUNT({col}$2:{col}${FIM})=0," in _formula(ws, col, FIM + 1)


@pytest.mark.parametrize("col", COLUNAS_FINANCEIRAS)
def test_fallback_de_lotacao_maxima_na_linha_201(ws, col):
    """Com 199 itens (2:200 lotada) o TOTAL fixo da linha 201 assume."""
    formula = _formula(ws, col, FIM + 1)
    assert f'IF($A${FIM}<>""' in formula
    assert f'SUMIF($A$2:$A${FIM},"<>"' in formula


def test_vu_original_nao_e_totalizavel(ws):
    """C = VU_ORIGINAL: preco unitario nunca entra na linha TOTAL."""
    for linha in range(2, FIM + 2):
        formula = _formula(ws, "C", linha)
        assert "SUMIF" not in formula, f"C{linha}"


# ---------------------------------------------------------------------------
# 2. Cenarios: 1, 3 e 20 itens — a linha TOTAL desloca e totaliza certo
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("itens", [1, 3, 20])
def test_linha_total_desloca_e_totaliza_todas_as_colunas(itens):
    wb = load_workbook(TEMPLATE)
    ws = wb["itens_Remanesc"]
    for i in range(itens):
        ws[f"A{2 + i}"] = f"I{i + 1:03d}"
    total = _linha_total(ws, itens)
    ultimo_item = total - 1

    # O rotulo dinamico TOTAL esta na coluna U da mesma linha.
    assert '"TOTAL"' in _formula(ws, "U", total)

    for col in COLUNAS_FINANCEIRAS:
        formula = _formula(ws, col, total)
        # A linha TOTAL detecta a si mesma...
        assert f'IF(AND(A{total}="",A{ultimo_item}<>""' in formula, col
        # ...e soma exatamente os itens existentes (2 ate o ultimo item).
        assert f'SUMIF($A$2:A{ultimo_item},"<>"' in formula, col
        assert "IF(FALSE" not in formula, col
        if col != "D":
            # guarda de aplicabilidade sobre o MESMO intervalo dos itens
            assert f"IF(COUNT({col}$2:{col}{ultimo_item})=0," in formula, col
    wb.close()


# ---------------------------------------------------------------------------
# 3. Valores reais no Excel (opt-in) — caso real de 3 itens
# ---------------------------------------------------------------------------

CASO_REAL_3_ITENS = {
    # col: (valores dos itens 2:4, total esperado na linha 5)
    "D": ([1_000_000.00, 1_000_000.00, 372_260.00], 2_372_260.00),
    "F": ([1_000_000.00, 1_000_000.00, 106_660.00], 2_106_660.00),
    "H": ([563_535.00, 500_000.00, 500_000.00], 1_563_535.00),
    "J": ([344_654.25, 500_000.00, 500_000.00], 1_344_654.25),
    "N": ([143_125.00, 200_000.00, 200_000.00], 543_125.00),
    "P": ([74_355.00, 100_000.00, 100_000.00], 274_355.00),
    "AC": ([65_600.00, 100_000.00, 100_000.00], 265_600.00),
}
COLUNAS_SEM_DADOS = ("L", "R", "T")   # integralmente vazias neste cenario


@com
def test_excel_real_totaliza_o_caso_de_3_itens():
    wb = load_workbook(TEMPLATE)
    ws = wb["itens_Remanesc"]
    for i, item in enumerate(("I001", "I002", "I003")):
        ws[f"A{2 + i}"] = item
    # Os valores por item substituem as formulas itemizadas SO nesta copia
    # descartavel: o alvo do teste e a formula da linha TOTAL (linha 5).
    for col, (valores, _total) in CASO_REAL_3_ITENS.items():
        for i, valor in enumerate(valores):
            ws[f"{col}{2 + i}"] = valor
    caminho = Path(tempfile.mkdtemp(prefix="cl8us_total_3itens_")) / "caso.xlsx"
    wb.save(caminho)
    wb.close()

    import pythoncom
    import win32com.client
    pythoncom.CoInitialize()
    xl = win32com.client.DispatchEx("Excel.Application")
    xl.Visible = False
    xl.DisplayAlerts = False
    try:
        livro = xl.Workbooks.Open(str(caminho), UpdateLinks=0)
        xl.CalculateFullRebuild()
        aba = livro.Worksheets("itens_Remanesc")
        assert str(aba.Range("U5").Value or "") == "TOTAL"
        for col, (_valores, total) in CASO_REAL_3_ITENS.items():
            assert aba.Range(f"{col}5").Value == pytest.approx(total), col
        for col in COLUNAS_SEM_DADOS:
            valor = aba.Range(f"{col}5").Value
            assert valor in (None, ""), (
                f"{col}5 deveria permanecer vazia (sem dados), veio {valor!r}"
            )
        livro.Close(SaveChanges=False)
    finally:
        xl.Quit()
        pythoncom.CoUninitialize()
