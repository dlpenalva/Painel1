# -*- coding: utf-8 -*-
"""Frente integrada — UX da RESULTADOS, falso positivo do efeito financeiro,
comunicacoes e composicao canonica do VTA no Sumario Executivo.

Cobre as seis frentes da tarefa, na ordem dos testes exigidos:

  A. RESULTADOS: "Ciclos apurados — C1, C2 e C3" (somente os computados);
  B. RESULTADOS: "Posição em DD/MM/AAAA" no lugar do saldo do ciclo;
  C. fronteira juridica x fisica: nenhum aviso de ajuste manual (EF-G1);
  D. EF-G1: a marcacao de `financeiro!G` e respeitada em silencio;
  E. Comunicacao a Contratada com variacao/fator acumulados canonicos;
  F. Comunicado Interno de conferencia;
  G-K. composicao canonica do VTA no PDF (Financeiro, PC, Consumido, metodo
       sem composicao segura e o PDF real);
  L. regressao numerica: nenhuma grandeza financeira muda.

Regra que orienta todos eles: nada aqui pode passar a existir por calculo novo.
Cada numero exibido tem de vir de uma fonte canonica ja existente, e a soma dos
componentes do VTA tem de fechar EXATAMENTE com o VTA oficial.
"""
from __future__ import annotations

import ast
import io
import os
import shutil
import sys
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

RAIZ = Path(__file__).resolve().parents[1]
if str(RAIZ) not in sys.path:
    sys.path.insert(0, str(RAIZ))
if str(Path(__file__).resolve().parent) not in sys.path:
    sys.path.insert(0, str(Path(__file__).resolve().parent))

from _coleta_oficial import gerar_coleta_oficial_preenchida  # noqa: E402
from _coleta_reajuste import (  # noqa: E402
    _ciclo_cronologico_financeiro,
    _ciclo_por_competencia_financeira,
    _marco_grade_financeira,
    _tabela_ciclos_financeiros,
    ler_coleta_reajuste,
)
from _motor_composicao_vta import montar_composicao_vta  # noqa: E402
from _sumario_executivo import (  # noqa: E402
    ROTULO_TOTAL_VTA,
    _montar_composicao_vta,
    formatar_moeda,
    gerar_sumario_executivo_pdf,
)

TEMPLATE = RAIZ / "templates" / "COLETA_REAJUSTE_OFICIAL.xlsx"
PAGINA = (RAIZ / "pages" / "03_Valor_Global.py").read_text(encoding="utf-8")

com_excel = pytest.mark.skipif(
    os.environ.get("RUN_EXCEL_INTEGRATION") != "1",
    reason="defina RUN_EXCEL_INTEGRATION=1 para executar o Excel COM",
)


@pytest.fixture(scope="module")
def resultados_ws():
    wb = openpyxl.load_workbook(TEMPLATE, data_only=False)
    return wb["RESULTADOS"]


# ===========================================================================
# TESTE A — RESULTADOS: quais ciclos foram apurados
# ===========================================================================

def test_a1_rodape_do_retroativo_le_a_fonte_canonica_dos_ciclos(resultados_ws):
    """A lista sai de parametros!A3:A6 (COMPUTAR_NESTA_APURACAO) + B3:B6.

    E a MESMA fonte que RESULTADOS!H14 ja usa para validar o card. Nenhuma
    regra nova decide o que foi apurado, e C0 (base contratual) fica de fora.
    """
    ws = resultados_ws
    assert "Ciclos apurados — " in ws["E6"].value
    assert "$J$11" in ws["E6"].value
    assert ws["J9"].value == '=COUNTIF(parametros!$A$3:$A$6,"Sim")'
    lista = ws["J10"].value
    assert "TEXTJOIN" in lista
    for linha in (3, 4, 5, 6):
        assert f'IF(parametros!$A${linha}="Sim",parametros!$B${linha},"")' in lista
    # C0 e a base: a linha 2 de parametros nunca entra na lista.
    assert "parametros!$A$2" not in lista
    assert "parametros!$B$2" not in lista
    # O rodape nao pode inventar valor nem virar ancora numerica.
    assert "R$" not in ws["E6"].value


def test_a2_conjuncao_natural_antes_do_ultimo_ciclo(resultados_ws):
    """"C1, C2 e C3": virgula entre os primeiros e " e " antes do ultimo."""
    formula = resultados_ws["J11"].value.replace(" ", "")
    assert "$J$9=1" in formula                      # um ciclo: sem conjuncao
    assert '"e"' in formula                          # conjuncao (sem espacos)
    assert 'SUBSTITUTE($J$10,",","|",$J$9-1)' in formula


def test_a3_valor_e_formula_do_retroativo_preservados(resultados_ws):
    """Frente 1 e apresentacao: a cadeia do retroativo nao pode ter mudado."""
    ws = resultados_ws
    assert ws["D5"].value == "=$D$22"
    assert ws["D22"].value == '=IFERROR(RETRO_OFICIAL,"")'


@com_excel
def test_a4_com_lista_os_ciclos_realmente_computados(tmp_path):
    """Prova de TEXTO em Excel real, nos cinco cenarios exigidos.

    Inclui o cenario "ciclo existente mas nao computado": marcar C0 como "Sim"
    ou deixar C2/C3/C4 em "Nao" nao pode fazer o ciclo aparecer na lista.
    """
    import gc
    from datetime import datetime

    import pythoncom
    import win32com.client

    destino = tmp_path / "ciclos_apurados.xlsx"
    shutil.copyfile(TEMPLATE, destino)
    pythoncom.CoInitialize()
    xl = win32com.client.DispatchEx("Excel.Application")
    xl.Visible = False
    xl.DisplayAlerts = False
    try:
        wb = xl.Workbooks.Open(str(destino.resolve()), UpdateLinks=0, CorruptLoad=0)
        res, ctrl = wb.Worksheets("RESULTADOS"), wb.Worksheets("CONTROLE")
        par, fin = wb.Worksheets("parametros"), wb.Worksheets("financeiro")
        for aba in (res, ctrl, par, fin):
            try:
                aba.Unprotect()
            except Exception:
                pass

        ctrl.Range("B1").Value = "Financeiro (Mensalidade)"
        ctrl.Range("B2").Value = "C4"
        ctrl.Range("B3").Value = datetime(2026, 8, 25)
        # base do metodo, para o card sair do estado vazio
        fin.Range("A14").Value = datetime(2026, 1, 1)
        fin.Range("B14").Value = "c1"
        fin.Range("C14").Value = 100.0
        fin.Range("G14").Value = "Sim"

        cenarios = [
            (["A3"], "Ciclos apurados — C1"),
            (["A3", "A4"], "Ciclos apurados — C1 e C2"),
            (["A3", "A4", "A5"], "Ciclos apurados — C1, C2 e C3"),
            (["A3", "A4", "A5", "A6"], "Ciclos apurados — C1, C2, C3 e C4"),
            # ciclos existentes, mas nao computados: so C1 entra.
            (["A2", "A3"], "Ciclos apurados — C1"),
        ]
        for marcados, esperado in cenarios:
            for linha in range(2, 7):
                par.Range(f"A{linha}").Value = "Nao"
                par.Range(f"E{linha}").Value = 0.05 if linha > 2 else 0.0
            for celula in marcados:
                par.Range(celula).Value = "Sim"
            xl.CalculateFull()
            assert res.Range("E6").Text == esperado, marcados

        wb.Close(False)
    finally:
        xl.Quit()
        gc.collect()
        pythoncom.CoUninitialize()


# ===========================================================================
# TESTE B — RESULTADOS: posicao do remanescente
# ===========================================================================

def test_b1_rodape_do_remanescente_usa_a_data_de_corte_canonica(resultados_ws):
    """A data vem de CONTROLE!B3, a mesma ancorada pelo proprio card em F5."""
    formula = resultados_ws["F6"].value
    assert "Posição em " in formula
    assert "CONTROLE!$B$3" in formula
    # composicao por DAY/MONTH/YEAR: TEXT(...) levaria codigo de formato
    # dependente de locale para dentro do arquivo.
    assert "TEXT(" not in formula
    assert "TODAY()" not in formula and "NOW()" not in formula
    # sem data de corte, declara a ausencia — nunca inventa uma data.
    assert "Data de corte não informada" in formula


def test_b2_saldo_do_ciclo_em_execucao_saiu_do_card(resultados_ws):
    """O texto removido nao pode sobrar em nenhuma celula visivel da aba."""
    ws = resultados_ws
    ocultas = {n for n, d in ws.row_dimensions.items() if d.hidden}
    visiveis = [
        celula.value
        for linha in ws.iter_rows(min_row=1, max_row=90, max_col=10)
        for celula in linha
        if isinstance(celula.value, str) and celula.row not in ocultas
    ]
    assert visiveis
    assert not any("Saldo do ciclo em execução" in texto for texto in visiveis)


def test_b3_valor_e_formula_do_remanescente_preservados(resultados_ws):
    ws = resultados_ws
    assert ws["G5"].value == "=$B$38"
    assert ws["B38"].value == '=IF(OR(B36="",B37=""),"",ROUND(B37-B36,2))'


@com_excel
def test_b4_com_posicao_em_ddmmaaaa(tmp_path):
    import gc
    from datetime import datetime

    import pythoncom
    import win32com.client

    destino = tmp_path / "posicao.xlsx"
    shutil.copyfile(TEMPLATE, destino)
    pythoncom.CoInitialize()
    xl = win32com.client.DispatchEx("Excel.Application")
    xl.Visible = False
    xl.DisplayAlerts = False
    try:
        wb = xl.Workbooks.Open(str(destino.resolve()), UpdateLinks=0, CorruptLoad=0)
        res, ctrl = wb.Worksheets("RESULTADOS"), wb.Worksheets("CONTROLE")
        for aba in (res, ctrl):
            try:
                aba.Unprotect()
            except Exception:
                pass
        ctrl.Range("B3").Value = datetime(2026, 8, 25)
        xl.CalculateFull()
        assert res.Range("F6").Text == "Posição em 25/08/2026"
        ctrl.Range("B3").ClearContents()
        xl.CalculateFull()
        assert res.Range("F6").Text == "Data de corte não informada"
        wb.Close(False)
    finally:
        xl.Quit()
        gc.collect()
        pythoncom.CoUninitialize()


# ===========================================================================
# TESTES C e D — aviso de ajuste manual do efeito financeiro
# ===========================================================================

# Caso real relatado: C1 TEMPESTIVO* desloca as janelas juridicas de C2..C4
# para marco, enquanto a grade financeira permanece em blocos fixos de 12
# competencias a partir de 02/2022.
DADOS_FRONTEIRA = {
    "origem": "Reajuste Multiciclo",
    "indice": "IST",
    "data_base_original": "01/02/2022",
    "data_corte": "25/08/2026",
    "ciclos": [
        {"ciclo": "C1", "data_inicio": "01/02/2023", "data_fim": "31/01/2024",
         "data_pedido": "10/03/2023", "financeiro_inicio": "01/03/2023",
         "percentual_aplicado": 0.0577, "objeto_analise_atual": True,
         "situacao": "TEMPESTIVO"},
        {"ciclo": "C2", "data_inicio": "01/03/2024", "data_fim": "28/02/2025",
         "data_pedido": "05/03/2024", "financeiro_inicio": "01/03/2024",
         "percentual_aplicado": 0.0450, "objeto_analise_atual": True,
         "situacao": "TEMPESTIVO"},
        {"ciclo": "C3", "data_inicio": "01/03/2025", "data_fim": "28/02/2026",
         "data_pedido": "", "financeiro_inicio": "01/03/2025",
         "percentual_aplicado": 0.0, "objeto_analise_atual": True,
         "situacao": "PRECLUSO"},
        {"ciclo": "C4", "data_inicio": "01/03/2026", "data_fim": "28/02/2027",
         "data_pedido": "10/01/2026", "financeiro_inicio": "01/03/2026",
         "percentual_aplicado": 0.0381, "objeto_analise_atual": True,
         "situacao": "ADIANTADO"},
    ],
}


@pytest.fixture(scope="module")
def coleta_fronteira() -> bytes:
    return gerar_coleta_oficial_preenchida(DADOS_FRONTEIRA)


def _linha_da_competencia(ws, ano: int, mes: int) -> int:
    for row in range(2, 74):
        valor = ws[f"A{row}"].value
        if valor and (valor.year, valor.month) == (ano, mes):
            return row
    raise AssertionError(f"competencia {mes:02d}/{ano} ausente")


def _avisos_de_override(diagnostico) -> list[str]:
    """EF-G1: qualquer texto do diagnostico que insinue autoria da marcacao."""
    textos: list[str] = []
    for chave in ("avisos", "pendencias", "bloqueios_criticos",
                  "inconsistencias", "bloqueios_estruturais", "lacunas_apuracao"):
        textos.extend(str(item) for item in (diagnostico.get(chave) or ()))
    return [
        texto for texto in textos
        if "ajustado manualmente" in texto or "ajustada manualmente" in texto
    ]


def test_c1_janelas_deslocadas_e_grade_fisica_fixa(coleta_fronteira):
    """O cenario e realmente o de fronteira: janela juridica != bloco fisico."""
    wb = openpyxl.load_workbook(io.BytesIO(coleta_fronteira), data_only=False)
    par = wb["parametros"]
    periodos = {
        str(par[f"B{r}"].value): (par[f"C{r}"].value, par[f"D{r}"].value)
        for r in range(2, 7)
    }
    assert periodos["C2"][0].month == 3 and periodos["C2"][1].month == 2
    assert periodos["C3"][0].month == 3 and periodos["C3"][1].month == 2
    assert periodos["C4"][0].month == 3

    fin = wb["financeiro"]
    marco = _marco_grade_financeira(wb)
    assert (marco.year, marco.month) == (2022, 2)
    # 02/2025 e 02/2026 sao as competencias de fronteira do caso relatado.
    for (ano, mes), fisico, juridico in (
        ((2025, 2), "C3", "C2"),
        ((2026, 2), "C4", "C3"),
    ):
        competencia = fin[f"A{_linha_da_competencia(fin, ano, mes)}"].value
        assert _ciclo_cronologico_financeiro(marco, competencia) == fisico
        assert _ciclo_por_competencia_financeira(
            wb, competencia, _tabela_ciclos_financeiros(wb)
        ) == juridico


def test_c2_arquivo_recem_gerado_nao_acusa_ajuste_manual(coleta_fronteira):
    """Nenhum aviso: o arquivo nunca foi editado por ninguem."""
    diagnostico = ler_coleta_reajuste(coleta_fronteira)
    assert _avisos_de_override(diagnostico) == []
    assert diagnostico["bloqueios_criticos"] == []


def test_c3_as_duas_competencias_de_fronteira_nao_geram_aviso(coleta_fronteira):
    """C2 — 02/2025 e C3 — 02/2026 eram os falsos positivos relatados."""
    diagnostico = ler_coleta_reajuste(coleta_fronteira)
    texto = " | ".join(diagnostico["avisos"])
    assert "02/2025" not in texto
    assert "02/2026" not in texto


def test_c4_a_reconstrucao_da_marcacao_esperada_foi_removida():
    """EF-G1: nao existe mais logica que reconstrua o valor "esperado" de G.

    A causa raiz dos falsos positivos era comparar `financeiro!G` com uma
    marcacao reconstruida por metadado/janela. Essa comparacao foi eliminada:
    os dois helpers que a sustentavam nao existem mais no leitor.
    """
    fonte = (RAIZ / "_coleta_reajuste.py").read_text(encoding="utf-8")
    assert "_inicio_efeito_definido" not in fonte
    assert "_tem_metadado_inicio_efeito" not in fonte
    assert "divergencias_manuais" not in fonte


def test_d1_override_real_sim_para_nao_nao_gera_aviso(coleta_fronteira):
    """O fiscal troca Sim->Nao: decisao respeitada, em silencio."""
    wb = openpyxl.load_workbook(io.BytesIO(coleta_fronteira), data_only=False)
    fin = wb["financeiro"]
    linha = _linha_da_competencia(fin, 2024, 6)   # bloco fisico de C2
    assert fin[f"G{linha}"].value == "Sim"
    fin[f"G{linha}"] = "Nao"
    saida = io.BytesIO()
    wb.save(saida)

    diagnostico = ler_coleta_reajuste(saida.getvalue())
    assert _avisos_de_override(diagnostico) == []
    # a alteracao e respeitada e nao vira bloqueio.
    assert diagnostico["bloqueios_criticos"] == []
    assert diagnostico.get("valido") is not False


def test_d2_override_em_ciclos_diferentes_segue_sem_qualquer_aviso(
    coleta_fronteira,
):
    wb = openpyxl.load_workbook(io.BytesIO(coleta_fronteira), data_only=False)
    fin = wb["financeiro"]
    for ano, mes in ((2023, 6), (2025, 6)):
        fin[f"G{_linha_da_competencia(fin, ano, mes)}"] = "Nao"
    saida = io.BytesIO()
    wb.save(saida)

    diagnostico = ler_coleta_reajuste(saida.getvalue())
    assert _avisos_de_override(diagnostico) == []
    assert diagnostico["bloqueios_criticos"] == []


def test_d3_o_aviso_de_ajuste_manual_nao_sobrevive_no_fluxo():
    alvos = ["_coleta_reajuste.py", "_ui_utils.py", "pages/03_Valor_Global.py"]
    for modulo in alvos:
        fonte = (RAIZ / modulo).read_text(encoding="utf-8")
        assert "Marcacao de efeito financeiro ajustada manualmente" not in fonte
        assert "Efeito financeiro ajustado manualmente" not in fonte
    ui = (RAIZ / "_ui_utils.py").read_text(encoding="utf-8")
    assert "render_avisos_override_efeito_financeiro" not in ui
    assert "PREFIXO_AVISO_OVERRIDE_EFEITO_FINANCEIRO" not in ui


# ===========================================================================
# TESTES E e F — comunicacoes
# ===========================================================================

def _funcao_pura_da_pagina(nome: str):
    """Compila apenas helpers puros da pagina, sem subir o Streamlit.

    A pagina e um script Streamlit; importa-la executaria a interface. Aqui so
    a funcao pedida e suas dependencias puras sao compiladas.
    """
    arvore = ast.parse(PAGINA)
    desejadas = {nome, "normalizar_competencia_periodo", "periodo_para_label_br"}
    corpo = [
        no for no in arvore.body
        if isinstance(no, ast.FunctionDef) and no.name in desejadas
    ]
    assert len(corpo) == len(desejadas), "helper ausente na pagina"
    import datetime as _dt
    import re as _re

    import pandas as _pd

    espaco: dict = {
        "pd": _pd, "re": _re, "datetime": _dt.datetime, "date": _dt.date,
    }
    exec(compile(ast.Module(body=corpo, type_ignores=[]), "<pagina>", "exec"), espaco)
    return espaco[nome]


def test_e1_variacao_acumulada_vem_do_fator_canonico_e_nao_da_soma():
    """A frase le `resultado["fator_acumulado"]` (parametros!F do ciclo
    vigente) — a mesma fonte do card "Percentual acumulado" — e deriva a
    variacao dele. Somar 5,77% + 4,50% + 3,81% daria 14,08%; o fator composto
    da 14,74%."""
    trecho = PAGINA[PAGINA.index("def _acumulado_canonico"):]
    trecho = trecho[:trecho.index("def _frase_variacao_acumulada")]
    assert 'resultado.get("fator_acumulado")' in trecho
    assert "float(fator) - 1.0" in trecho
    # nenhuma soma de percentuais individuais entra na conta.
    assert "df_ciclos" not in trecho
    assert "Variação" not in trecho
    assert "sum(" not in trecho


def test_comunicados_financeiros_consumem_a_classificacao_canonica():
    classificar = _funcao_pura_da_pagina(
        "_financeiro_por_ciclo_para_validacao_contratada"
    )
    resultado = {
        "df_ciclos": pd.DataFrame([
            {"Ciclo": "C1", "Variação": 0.0, "Fator acumulado efetivo": 1.0},
            {"Ciclo": "C2", "Variação": 0.0, "Fator acumulado efetivo": 1.0289},
        ]),
        "df_financeiro_mensal_tratado": pd.DataFrame([
            {"Ciclo": "C1", "Competência": "10/2024", "Efeito financeiro": "Nao"},
            {"Ciclo": "C2", "Competência": "08/2026", "Efeito financeiro": "Nao"},
            {"Ciclo": "C2", "Competência": "09/2026", "Efeito financeiro": "Sim"},
        ]),
        "df_meses_sem_efeito_financeiro": pd.DataFrame([
            {"Ciclo": "C1", "Competência": "10/2024", "Valor pago/faturado": 1000.0},
            {"Ciclo": "C2", "Competência": "08/2026", "Valor pago/faturado": 1000.0},
        ]),
    }

    financeiro = classificar(resultado, {"medidas_pc_aplicaveis": False})

    assert financeiro["C1"]["competencias_sem_efeito"] == []
    assert financeiro["C2"]["competencias_sem_efeito"] == ["08/2026"]
    assert financeiro["C2"]["inicio_efeito"] == "09/2026"


def test_comunicados_nao_transformam_efeito_vazio_em_sem_efeito():
    classificar = _funcao_pura_da_pagina(
        "_financeiro_por_ciclo_para_validacao_contratada"
    )
    resultado = {
        "df_ciclos": pd.DataFrame([{"Ciclo": "C1", "Variação": 0.05}]),
        "df_financeiro_mensal_tratado": pd.DataFrame([
            {"Ciclo": "C1", "Competência": "01/2025", "Efeito financeiro": ""},
        ]),
        "df_meses_sem_efeito_financeiro": pd.DataFrame(),
    }

    financeiro = classificar(resultado, {"medidas_pc_aplicaveis": False})

    assert financeiro["C1"]["inicio_efeito"] is None
    assert financeiro["C1"]["competencias_sem_efeito"] == []


def test_e1b_variacao_e_derivada_do_fator_ao_centavo_do_percentual():
    acumulado = _funcao_pura_da_pagina("_acumulado_canonico")
    variacao, fator = acumulado({"fator_acumulado": 1.1474})
    assert fator == 1.1474
    assert round(variacao, 6) == 0.1474
    # ausencia de fator canonico nao vira zero nem numero inventado.
    assert acumulado({}) == (None, None)
    assert acumulado({"fator_acumulado": None}) == (None, None)
    assert acumulado({"fator_acumulado": True}) == (None, None)


def test_e2_frase_da_variacao_acumulada_fica_logo_apos_a_tabela_dos_ciclos():
    trecho = PAGINA[PAGINA.index("def gerar_texto_validacao_contratada"):]
    trecho = trecho[:trecho.index("def render_validacao_contratada")]
    tabela = trecho.index('"CICLO | PERÍODO | ÍNDICE | VARIAÇÃO | EFEITO FINANCEIRO"')
    frase = trecho.index("frase_acumulada = _frase_variacao_acumulada(resultado)")
    competencias = trecho.index("competências sem efeito financeiro")
    assert tabela < frase < competencias


def test_e3_texto_da_frase_conforme_aprovado():
    trecho = PAGINA[PAGINA.index("def _frase_variacao_acumulada"):]
    trecho = trecho[:trecho.index("def _competencias_compactadas")]
    assert "Considerados os reajustes aplicáveis aos ciclos acima, a variação " in trecho
    assert "acumulada do contrato corresponde a " in trecho
    assert "equivalente ao fator acumulado de " in trecho


def test_e4_estrutura_do_comunicado_a_contratada_preservada():
    """Os blocos que ja existiam continuam todos no texto, na mesma ordem."""
    trecho = PAGINA[PAGINA.index("def gerar_texto_validacao_contratada"):]
    trecho = trecho[:trecho.index("def render_validacao_contratada")]
    marcos = [
        "VALIDAÇÃO DOS VALORES APURADOS PARA O REAJUSTE CONTRATUAL",
        "2. Foram considerados os seguintes ciclos:",
        "frase_acumulada",
        "competências sem efeito financeiro",
        "CICLO | PERÍODO/ANO | RETROATIVO RECONHECIDO",
        "TOTAL RETROATIVO RECONHECIDO:",
        "TOTAL RETROATIVO POTENCIAL:",
        "confirmação de concordância para prosseguimento da formalização",
    ]
    posicoes = [trecho.index(m) for m in marcos]
    assert posicoes == sorted(posicoes)


def test_f1_comunicado_interno_tem_titulo_e_encerramento_aprovados():
    trecho = PAGINA[PAGINA.index("# >>> COMUNICADO_INTERNO_CONFERENCIA_V1"):]
    trecho = trecho[:trecho.index("# <<< COMUNICADO_INTERNO_CONFERENCIA_V1")]
    assert 'TITULO_COMUNICADO_INTERNO = "COMUNICADO INTERNO"' in trecho
    assert (
        'SUBTITULO_COMUNICADO_INTERNO = "CONFERÊNCIA DA APURAÇÃO DO REAJUSTE"'
    ) in trecho
    assert '"Olá,"' in trecho
    assert "CICLO | PERÍODO | VARIAÇÃO | EFEITO FINANCEIRO" in trecho
    assert "CICLO | RETROATIVO RECONHECIDO" in trecho
    assert "TOTAL RETROATIVO RECONHECIDO:" in trecho
    assert "A variação acumulada apurada é de " in trecho
    assert "correspondente ao fator acumulado de " in trecho
    assert "Competências sem efeito financeiro:" in trecho
    assert (
        "Peço uma conferência dos dados utilizados na apuração, principalmente quanto "
    ) in trecho
    assert (
        "Se identificar algum ajuste, basta me sinalizar objetivamente o dado ou valor "
    ) in trecho
    assert "Não havendo ajustes, seguimos com a formalização do reajuste." in trecho


def test_f2_comunicado_interno_nao_faz_segunda_leitura_nem_recalcula():
    """Mesma apuracao, mesmas fontes canonicas do comunicado a contratada."""
    trecho = PAGINA[PAGINA.index("def gerar_texto_comunicado_interno"):]
    trecho = trecho[:trecho.index("def render_comunicado_interno")]
    for fonte in (
        "_ciclos_para_validacao_contratada",
        "_financeiro_por_ciclo_para_validacao_contratada",
        "_retroativo_por_ciclo_para_validacao_contratada",
        "_acumulado_canonico",
    ):
        assert fonte in trecho, fonte
    for proibido in ("load_workbook", "openpyxl", "ler_coleta_reajuste", "round("):
        assert proibido not in trecho, proibido


def test_f3_comunicado_interno_e_entregue_ao_lado_do_da_contratada():
    ordem = [
        PAGINA.index("render_validacao_contratada(resultado, diagnostico_coleta)"),
        PAGINA.index("render_comunicado_interno(resultado, diagnostico_coleta)"),
        PAGINA.index(
            "st.stop()",
            PAGINA.index("render_comunicado_interno(resultado, diagnostico_coleta)"),
        ),
    ]
    assert ordem == sorted(ordem)
    trecho = PAGINA[PAGINA.index("def render_comunicado_interno"):]
    trecho = trecho[:trecho.index("# <<< COMUNICADO_INTERNO_CONFERENCIA_V1")]
    assert 'st.markdown("### Comunicado interno")' in trecho
    assert 'with st.expander("Visualizar comunicado interno")' in trecho
    assert '"Baixar TXT (interno)"' in trecho
    assert trecho.count("texto_interno = gerar_texto_comunicado_interno") == 1
    # mesma classe de entrega do comunicado a contratada: sem widget com key
    # propria, portanto sem o bug de congelamento do st.text_area.
    assert "st.text_area" not in trecho


def test_f4_competencias_contiguas_sao_compactadas():
    """Um intervalo continuo vira "mm/aaaa a mm/aaaa"; meses soltos ficam."""
    compactar = _funcao_pura_da_pagina("_competencias_compactadas")
    continuo = [f"{m:02d}/2025" for m in range(2, 13)] + ["01/2026"]
    assert compactar(continuo) == ["02/2025 a 01/2026"]
    assert compactar(["02/2023"]) == ["02/2023"]
    assert compactar(["02/2023", "05/2023", "06/2023"]) == [
        "02/2023", "05/2023 a 06/2023",
    ]
    assert compactar([]) == []
    # entrada fora do padrao nao e perdida nem inventada.
    assert compactar(["competência indefinida"]) == ["competência indefinida"]


# ===========================================================================
# TESTES G a K — composicao canonica do VTA no Sumario Executivo
# ===========================================================================

def _apresentacao(leitura: dict) -> tuple[dict, dict]:
    composicao = montar_composicao_vta(leitura)
    vta = composicao.get("vta_composicao")
    return composicao, _montar_composicao_vta(
        {"composicao_vta": composicao}, {"vta": vta}
    )


def _soma(apresentacao: dict) -> float:
    return round(sum(c["valor"] for c in apresentacao["componentes"]), 2)


def test_g_financeiro_fecha_exatamente_no_vta_canonico():
    """Caso real de referencia: 7.300.890,27 + 24.678,92 + 1.388.251,07."""
    from test_vta_u2_uniformizacao import (
        AJUSTES,
        EXECUTADO,
        REMANESCENTE,
        VTA,
        _leitura_financeiro,
    )

    composicao, apresentacao = _apresentacao(_leitura_financeiro())
    assert apresentacao["exibivel"] is True
    assert apresentacao["metodo"] == "financeiro"
    assert [(c["descricao"], c["valor"]) for c in apresentacao["componentes"]] == [
        ("Executado apurado", EXECUTADO),
        ("Ajustes ainda devidos", AJUSTES),
        ("Remanescente atualizado", REMANESCENTE),
    ]
    assert _soma(apresentacao) == apresentacao["total"] == composicao["vta_composicao"]
    assert apresentacao["total"] == VTA


def test_h_pc_usa_os_componentes_reais_do_metodo_e_fecha_exato():
    """O PC nao tem "Ajustes ainda devidos": o retroativo ja esta dentro do
    valor considerado dos PCs. A tabela reflete a semantica real."""
    from test_vta_pc_composicao import _leitura

    composicao, apresentacao = _apresentacao(_leitura())
    assert apresentacao["exibivel"] is True
    assert apresentacao["metodo"] == "pc"
    descricoes = [c["descricao"] for c in apresentacao["componentes"]]
    assert not any("Ajustes ainda devidos" in d for d in descricoes)
    assert any("remanescente" in d.lower() for d in descricoes)
    assert _soma(apresentacao) == apresentacao["total"] == composicao["vta_composicao"]


def test_i_consumido_fecha_exato_e_sem_dupla_contagem():
    """Caso controlado homologado: 284 + 126 = 410, sem o ajuste tecnico
    reaparecendo como parcela autonoma."""
    from test_vta_u2_uniformizacao import _leitura_consumido

    composicao, apresentacao = _apresentacao(_leitura_consumido())
    assert apresentacao["exibivel"] is True
    assert [(c["descricao"], c["valor"]) for c in apresentacao["componentes"]] == [
        ("C1 executado", 284.0),
        ("Saldo remanescente atualizado no corte", 126.0),
    ]
    assert _soma(apresentacao) == apresentacao["total"] == 410.0
    assert composicao["retroativo_implicito"] == 0.0


def test_j1_metodo_sem_composicao_segura_mostra_so_o_vta_total():
    """PC sem base itemizada: nenhuma decomposicao e fabricada."""
    from test_vta_u2_uniformizacao import _leitura_consumido

    leitura = dict(_leitura_consumido(), controle={"modo": "pc", "ciclo_vigente": "C1"})
    composicao, apresentacao = _apresentacao(leitura)
    assert composicao["disponivel"] is False
    assert apresentacao["exibivel"] is False
    assert apresentacao["componentes"] == []


def test_j2_soma_divergente_do_vta_apresentado_nao_exibe_tabela():
    """Guarda fail-closed: um centavo de diferenca ja retira a tabela."""
    from test_vta_u2_uniformizacao import _leitura_financeiro

    composicao = montar_composicao_vta(_leitura_financeiro())
    divergente = _montar_composicao_vta(
        {"composicao_vta": composicao},
        {"vta": round(composicao["vta_composicao"] + 0.01, 2)},
    )
    assert divergente["exibivel"] is False
    assert divergente["componentes"] == []


def test_j3_vta_indisponivel_ou_previa_nao_exibe_tabela():
    """Sem VTA definitivo (PREVIA / indisponivel), a tabela nao aparece."""
    from test_vta_u2_uniformizacao import _leitura_financeiro

    composicao = montar_composicao_vta(_leitura_financeiro())
    for sintese in ({"vta": None}, {"vta": None, "vta_previa": 1.0}, {}):
        assert _montar_composicao_vta(
            {"composicao_vta": composicao}, sintese
        )["exibivel"] is False


def test_j4_composicao_bloqueada_para_formalizacao_nao_e_apresentada():
    from test_vta_u2_uniformizacao import _leitura_financeiro

    composicao = dict(
        montar_composicao_vta(_leitura_financeiro()), bloqueia_formalizacao=True
    )
    apresentacao = _montar_composicao_vta(
        {"composicao_vta": composicao}, {"vta": composicao["vta_composicao"]}
    )
    assert apresentacao["exibivel"] is False


def _dados_pdf(composicao: dict, vta: float) -> dict:
    return {
        "disponivel": True,
        "identificacao": {
            "indice": "IST", "metodo": "Financeiro",
            "ciclo_vigente": "C3", "data_corte": "25/08/2026",
        },
        "sintese": {
            "metodo_vta": "Financeiro", "vta": vta,
            "variacao_acumulada": 0.1474, "retroativo_total": 24678.92,
        },
        "composicao_vta": _montar_composicao_vta(
            {"composicao_vta": composicao}, {"vta": vta}
        ),
        "ciclos": [], "financeiro": {}, "itens": [], "historico_vu": {},
        "memoria_calculo": [], "aditivos": {}, "observacoes": [],
        "campos_nao_confiaveis": [], "referencias_vta": {},
    }


def _texto_pdf(pdf: bytes) -> str:
    fitz = pytest.importorskip("fitz")
    documento = fitz.open(stream=pdf, filetype="pdf")
    return "\n".join(pagina.get_text() for pagina in documento)


def test_k1_pdf_abre_e_mostra_a_composicao_uma_unica_vez():
    from test_vta_u2_uniformizacao import (
        AJUSTES,
        EXECUTADO,
        REMANESCENTE,
        VTA,
        _leitura_financeiro,
    )

    composicao = montar_composicao_vta(_leitura_financeiro())
    pdf = gerar_sumario_executivo_pdf(_dados_pdf(composicao, VTA))
    assert pdf.startswith(b"%PDF-")

    texto = _texto_pdf(pdf)
    assert "Composição do Valor Total Atualizado — VTA" in texto
    for rotulo, valor in (
        ("Executado apurado", EXECUTADO),
        ("Ajustes ainda devidos", AJUSTES),
        ("Remanescente atualizado", REMANESCENTE),
    ):
        assert rotulo in texto
        assert formatar_moeda(valor) in texto
    assert ROTULO_TOTAL_VTA in texto
    # o VTA aparece uma unica vez: como total do quadro que o compoe.
    assert texto.count(formatar_moeda(VTA)) == 1
    assert "Valor total atualizado (VTA)" not in texto


def test_k2_sem_composicao_o_pdf_segue_exibindo_o_vta_como_antes():
    dados = _dados_pdf({"disponivel": False}, 8_713_820.26)
    texto = _texto_pdf(gerar_sumario_executivo_pdf(dados))
    assert "Valor total atualizado (VTA)" in texto
    assert "Composição do Valor Total Atualizado — VTA" not in texto


def test_k3_pdf_nao_traz_vta_alternativo_nem_fallback_fisico():
    from test_vta_u2_uniformizacao import VTA, _leitura_financeiro

    composicao = montar_composicao_vta(_leitura_financeiro())
    texto = _texto_pdf(gerar_sumario_executivo_pdf(_dados_pdf(composicao, VTA)))
    for proibido in (
        "posição física", "Posição física", "última posição",
        "integralmente reajustado", "COMPARATIVO", "PRÉVIA",
    ):
        assert proibido not in texto, proibido


def test_k4_ultima_linha_do_quadro_e_destacada():
    """Destaque discreto: negrito e faixa, sem exagero grafico."""
    fonte = (RAIZ / "_sumario_executivo.py").read_text(encoding="utf-8")
    assert "linha_total=True" in fonte
    assert '"total_dir" if j in direita else "total"' in fonte
    assert '("LINEABOVE", (0, ultima), (-1, ultima)' in fonte


def test_k5_o_pdf_nao_conhece_formula_de_vta():
    """Arquitetura: o renderer so formata o que o motor ja compos."""
    fonte = (RAIZ / "_sumario_executivo.py").read_text(encoding="utf-8")
    bloco = fonte[fonte.index("def _bloco_sintese"):]
    bloco = bloco[:bloco.index("def _bloco_ciclos")]
    for proibido in ("vta_composicao", "valor_atualizado", "sum(", "round("):
        assert proibido not in bloco, proibido


# ===========================================================================
# TESTE L — regressao numerica
# ===========================================================================

def test_l1_nenhuma_grandeza_financeira_do_motor_mudou():
    """Golden numerico dos tres metodos, ancorado nos casos ja homologados."""
    from test_vta_pc_composicao import _leitura as _leitura_pc
    from test_vta_u2_uniformizacao import (
        AJUSTES,
        EXECUTADO,
        REMANESCENTE,
        VTA,
        _leitura_consumido,
        _leitura_financeiro,
    )

    financeiro = montar_composicao_vta(_leitura_financeiro())
    assert financeiro["total_execucao_atualizada"] == EXECUTADO
    assert financeiro["retroativo_implicito"] == AJUSTES
    assert financeiro["saldo_remanescente"]["valor_atualizado"] == REMANESCENTE
    assert financeiro["vta_composicao"] == VTA

    consumido = montar_composicao_vta(_leitura_consumido())
    assert consumido["total_execucao_atualizada"] == 284.0
    assert consumido["saldo_remanescente"]["valor_atualizado"] == 126.0
    assert consumido["vta_composicao"] == 410.0

    pc = montar_composicao_vta(_leitura_pc())
    assert pc["vta_composicao"] == round(
        pc["total_execucao_atualizada"]
        + pc["saldo_remanescente"]["valor_atualizado"], 2
    )


def test_l2_o_motor_de_composicao_nao_foi_tocado_por_esta_frente():
    """A Frente 6 e apresentacao: o motor canonico permanece intacto e segue
    sem ler o resultado do XLS para compor o VTA."""
    fonte = (RAIZ / "_motor_composicao_vta.py").read_text(encoding="utf-8")
    for proibido in ("resultados_xls", "VTA_FINAL", "reconciliacao_xls_python"):
        assert proibido not in fonte


def test_l3_a_frente_nao_altera_formula_financeira_do_template():
    """Somente os dois rodapes e as auxiliares da coluna oculta mudaram."""
    ws = openpyxl.load_workbook(TEMPLATE, data_only=False)["RESULTADOS"]
    # cadeias financeiras preservadas
    assert ws["C5"].value == '=IF(VTA_FINAL="","",VTA_FINAL)'
    assert ws["D22"].value == '=IFERROR(RETRO_OFICIAL,"")'
    assert ws["B86"].value == '=IF(VTA_FINAL="","",VTA_FINAL)'
    assert "$B$86-($B$83+N($B$84)+$B$85)" in ws["B87"].value.replace(" ", "")
    # as auxiliares novas vivem na coluna oculta de apoio
    assert ws.column_dimensions["J"].hidden is True
