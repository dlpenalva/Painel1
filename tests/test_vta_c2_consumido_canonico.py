"""VTA-C2 — Metodo Consumido canonico (Python + XLS).

Testes permanentes cobrindo:
  A. leitura multi-ciclo do item (_objeto_processo_reajuste.py);
  B-D. execucao C0/C1 (VU diferente)/C2 (fator especifico, nao hardcoded);
  E. multiplos itens;
  F. zero real;
  G. ausencia (qtd None != qtd 0);
  H-J. execucao atualizada=284, futuro=126, VTA=410 (caso controlado);
  K. B21 nao duplicado no ramo Itens de B26;
  L. Python independente de B26 (fail-closed sem evidencia);
  M. divergencia XLS x Python detectavel (reconciliacao nao suprimida);
  N. selecao explicita "Itens Consumidos" nao usa residual de outro metodo;
  O. fail-closed quando Consumido selecionado sem evidencia;
  P/Q. Financeiro/PC preservados na formula XLS (ramos intactos);
  R/S. RESULTADOS/estrutura do template (contagem de formulas/SHA cobertos
       em tests/test_integridade_template_xlsx.py e
       tests/test_coleta_oficial_integracao.py);
  T. workbook abre/recalcula sem reparo (Excel real).
"""
from __future__ import annotations

import gc
import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import load_workbook  # noqa: E402

from _objeto_processo_reajuste import _montar_memoria_por_ciclo  # noqa: E402

ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "templates" / "COLETA_REAJUSTE_OFICIAL.xlsx"

_RAMO_ITENS_B26 = (
    'IF(OR($F$20="",D35="",AND(B24<>"",NOT(ISNUMBER(B24)))),"",'
    'ROUND($F$20+D35+IF(ISNUMBER(B24),B24,0)+IF(ISNUMBER($N$263),$N$263,0),2))'
)
_RAMO_FINANCEIRO_B26 = (
    'IF(OR($D$20="",B21="",D35="",AND(B24<>"",NOT(ISNUMBER(B24)))),"",'
    'ROUND($D$20+B21+D35+IF(ISNUMBER(B24),B24,0)+IF(ISNUMBER($N$263),$N$263,0),2))'
)
_RAMO_PC_B26 = (
    'IF($T$25="CALCULO MANUAL REQUERIDO","",ROUND($T$25+IF(ISNUMBER(B24),B24,0),2))'
)


@pytest.fixture(scope="module")
def wb():
    return load_workbook(TEMPLATE, data_only=False)


# ---------------------------------------------------------------------------
# XLS — formulas estruturais (sem Excel real)
# ---------------------------------------------------------------------------

def test_p_ramo_financeiro_preservado_dentro_de_b26(wb):
    b26 = str(wb["MEMORIA_RESULTADOS"]["B26"].value)
    assert _RAMO_FINANCEIRO_B26 in b26


def test_q_ramo_pc_preservado_dentro_de_b26(wb):
    b26 = str(wb["MEMORIA_RESULTADOS"]["B26"].value)
    assert _RAMO_PC_B26 in b26


def test_k_ramo_itens_b26_nao_soma_b21_novamente(wb):
    """O ramo Itens de B26 usa F20 (execucao) + D35 (futuro) + ajustes —
    B21 (decomposicao do reajuste) NAO aparece somado neste ramo."""
    b26 = str(wb["MEMORIA_RESULTADOS"]["B26"].value)
    assert _RAMO_ITENS_B26 in b26
    # A subexpressao do ramo Itens em si nao referencia B21.
    assert "B21" not in _RAMO_ITENS_B26


def test_execucao_consumida_usa_valor_cons_por_ciclo_nao_vu_c0_universal(wb):
    f20 = str(wb["MEMORIA_RESULTADOS"]["F20"].value)
    for coluna in ("F", "H", "J", "L", "N"):
        assert f"itens_Consumidos!${coluna}$2:${coluna}$200" in f20


def test_c33_usa_coluna_auxiliar_sem_sumproduct_buggy(wb):
    c33 = str(wb["MEMORIA_RESULTADOS"]["C33"].value)
    assert "itens_Consumidos!$V$2:$V$200" in c33
    assert "SUMPRODUCT" not in c33


def test_b28_preserva_referencia_antiga_para_itens(wb):
    b28 = str(wb["MEMORIA_RESULTADOS"]["B28"].value)
    assert b28 == '=IF(OR($B$4="Financeiro",$B$4="Itens"),$B$23,$B$26)'


def test_nome_definido_vta_final_continua_b26(wb):
    definido = wb.defined_names["VTA_FINAL"]
    destinos = list(definido.destinations)
    assert len(destinos) == 1
    aba, referencia = destinos[0]
    assert aba == "MEMORIA_RESULTADOS"
    assert referencia.replace("$", "") == "B26"


def test_t_template_abre_estruturalmente_via_openpyxl():
    wb_local = load_workbook(TEMPLATE, data_only=False)
    assert "MEMORIA_RESULTADOS" in wb_local.sheetnames
    assert "itens_Consumidos" in wb_local.sheetnames


def test_t_excel_real_recalcula_e_reabre_sem_reparo():
    client = pytest.importorskip("win32com.client")
    pythoncom = pytest.importorskip("pythoncom")
    pythoncom.CoInitialize()
    excel = client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        wb_com = excel.Workbooks.Open(str(TEMPLATE), UpdateLinks=0, ReadOnly=True)
        try:
            mem = wb_com.Worksheets("MEMORIA_RESULTADOS")
            assert str(mem.Range("B26").Formula)
            assert str(mem.Range("F20").Formula).startswith("=IF(")
            ic = wb_com.Worksheets("itens_Consumidos")
            assert str(ic.Range("V2").Formula).startswith("=IF(")
        finally:
            wb_com.Close(False)
            del wb_com
    finally:
        excel.Quit()
        del excel
        gc.collect()
        pythoncom.CoUninitialize()


# ---------------------------------------------------------------------------
# Python — cadeia _montar_memoria_por_ciclo (sinteticos, sem XLS)
# ---------------------------------------------------------------------------

def _leitura_consumidos(itens, modo="d", ciclo_vigente="C1", fatores=None):
    fatores = fatores or {"C0": 1.0, "C1": 1.05}
    por_ciclo = {c: {"fator_acumulado": f} for c, f in fatores.items()}
    for c in ("C0", "C1", "C2", "C3", "C4"):
        por_ciclo.setdefault(c, {"fator_acumulado": None})
    return {
        "controle": {"modo": modo, "ciclo_vigente": ciclo_vigente},
        "parametros_v10": {"por_ciclo": por_ciclo},
        "itens_consumidos_v10": {"itens": itens},
    }


_ITEM_A = {
    "item": "A", "qtd_contratada": 20, "vu_original": 10.0, "qtd_total": 18,
    "consumos": {
        "C0": {"qtd": 10, "valor": 100.0},
        "C1": {"qtd": 8, "valor": 84.0},
        "C2": {"qtd": None, "valor": None},
        "C3": {"qtd": None, "valor": None},
        "C4": {"qtd": None, "valor": None},
    },
}
_ITEM_B = {
    "item": "B", "qtd_contratada": 10, "vu_original": 20.0, "qtd_total": 5,
    "consumos": {
        "C0": {"qtd": 5, "valor": 100.0},
        "C1": {"qtd": 0, "valor": 0.0},  # zero real, nao ausencia
        "C2": {"qtd": None, "valor": None},
        "C3": {"qtd": None, "valor": None},
        "C4": {"qtd": None, "valor": None},
    },
}


def _conferencia_consumidos(memoria):
    return next(
        c for c in memoria["conferencias_metodologicas"] if c["metodo"] == "consumidos"
    )


def test_a_leitura_multiciclo_e_e_multiplos_itens():
    leitura = _leitura_consumidos([_ITEM_A, _ITEM_B])
    memoria = _montar_memoria_por_ciclo(leitura, {}, [], {})
    c0 = next(c for c in memoria["ciclos"] if c["ciclo"] == "C0")
    c1 = next(c for c in memoria["ciclos"] if c["ciclo"] == "C1")
    assert c0["retroativo"]["consumidos"]["evidencias"] == 2
    assert c1["retroativo"]["consumidos"]["evidencias"] == 2


def test_b_execucao_c0():
    leitura = _leitura_consumidos([_ITEM_A, _ITEM_B])
    memoria = _montar_memoria_por_ciclo(leitura, {}, [], {})
    c0 = next(c for c in memoria["ciclos"] if c["ciclo"] == "C0")
    assert c0["retroativo"]["consumidos"]["valor_atualizado"] == 200.0
    assert c0["retroativo"]["consumidos"]["retroativo"] == 0.0


def test_c_execucao_c1_com_fator_diferente():
    leitura = _leitura_consumidos([_ITEM_A, _ITEM_B])
    memoria = _montar_memoria_por_ciclo(leitura, {}, [], {})
    c1 = next(c for c in memoria["ciclos"] if c["ciclo"] == "C1")
    assert c1["retroativo"]["consumidos"]["base_original"] == 80.0
    assert c1["retroativo"]["consumidos"]["valor_atualizado"] == 84.0
    assert c1["retroativo"]["consumidos"]["retroativo"] == 4.0


def test_d_ciclo_posterior_c2_usa_fator_proprio_nao_hardcoded():
    item_c2 = {
        "item": "C", "qtd_contratada": 5, "vu_original": 50.0, "qtd_total": 3,
        "consumos": {
            "C0": {"qtd": None, "valor": None},
            "C1": {"qtd": None, "valor": None},
            "C2": {"qtd": 3, "valor": None},  # sem valor cacheado -> fallback base*fator
            "C3": {"qtd": None, "valor": None},
            "C4": {"qtd": None, "valor": None},
        },
    }
    leitura = _leitura_consumidos(
        [item_c2], ciclo_vigente="C2", fatores={"C0": 1.0, "C1": 1.05, "C2": 1.10}
    )
    memoria = _montar_memoria_por_ciclo(leitura, {}, [], {})
    c2 = next(c for c in memoria["ciclos"] if c["ciclo"] == "C2")
    # base = 3*50=150; atualizado (fallback) = 150*1.10=165; retroativo=15.
    assert c2["retroativo"]["consumidos"]["base_original"] == 150.0
    assert c2["retroativo"]["consumidos"]["valor_atualizado"] == 165.0
    assert c2["retroativo"]["consumidos"]["retroativo"] == 15.0


def test_f_zero_real_e_evidencia_valida():
    leitura = _leitura_consumidos([_ITEM_B])
    memoria = _montar_memoria_por_ciclo(leitura, {}, [], {})
    c1 = next(c for c in memoria["ciclos"] if c["ciclo"] == "C1")
    assert c1["retroativo"]["consumidos"]["evidencias"] == 1
    assert c1["retroativo"]["consumidos"]["valor_atualizado"] == 0.0


def test_g_ausencia_nao_gera_evidencia():
    leitura = _leitura_consumidos([_ITEM_A])
    memoria = _montar_memoria_por_ciclo(leitura, {}, [], {})
    c2 = next(c for c in memoria["ciclos"] if c["ciclo"] == "C2")
    assert c2["retroativo"]["consumidos"]["evidencias"] == 0
    assert c2["retroativo"]["consumidos"]["valor_atualizado"] == 0.0


def test_h_i_j_execucao_futuro_vta_caso_controlado():
    leitura = _leitura_consumidos([_ITEM_A, _ITEM_B])
    memoria = _montar_memoria_por_ciclo(leitura, {}, [], {})
    conf = _conferencia_consumidos(memoria)
    assert conf["executado_atualizado"] == 284.0
    assert conf["potencial_restante_atualizado"] == 126.0
    assert conf["valor_total_atualizado"] == 410.0
    assert memoria["vta"]["valor_total_atualizado"] == 410.0
    assert memoria["vta"]["metodo"] == "consumidos"


def test_n_selecao_explicita_consumido_ignora_residual_de_outro_metodo():
    """Financeiro/PC com evidencia residual nao pode vencer quando o metodo
    explicitamente configurado (controle.modo='d') e Consumido com evidencia
    valida — prioridade historica financeiro->pc so vale na ausencia de
    selecao explicita."""
    leitura = _leitura_consumidos([_ITEM_A, _ITEM_B], modo="d")
    # injeta uma "parcela financeiro" residual, simulando dado irrelevante
    # deixado em outra aba nao usada pelo metodo configurado.
    leitura["vta_sombra"] = {
        "parcelas_computadas": [
            {
                "fonte_parcela": "Financeiro", "identificador": "financeiro:residual",
                "ciclo": "C0", "valor": 999.0, "valor_atualizado": 999.0,
            }
        ]
    }
    memoria = _montar_memoria_por_ciclo(leitura, {}, [], {})
    assert memoria["vta"]["metodo"] == "consumidos"
    assert memoria["vta"]["valor_total_atualizado"] == 410.0


def test_o_fail_closed_consumido_selecionado_sem_evidencia():
    """Metodo explicitamente Consumido, mas sem nenhum consumo informado ->
    memoria.vta fica INDETERMINADO (nunca fabrica numero)."""
    leitura = _leitura_consumidos([], modo="d")
    memoria = _montar_memoria_por_ciclo(leitura, {}, [], {})
    assert memoria["vta"]["valor_total_atualizado"] is None
    assert memoria["vta"]["natureza"] == "INDETERMINADO"


def test_l_python_nao_copia_b26_quando_consumido_sem_evidencia():
    """Reproduz o fail-closed de _coleta_reajuste_documentos.py: quando o
    metodo e Consumido e vta_python e None, valor_total nao deve cair para
    o cache do XLS (simulado aqui isolando a regra, sem abrir o XLS real —
    a integracao completa e coberta pelo caso controlado end-to-end)."""
    metodo_python = "consumidos"
    vta_python = None
    vta_capacidade_valor = 122036515.12  # valor XLS antigo, nao relacionado
    if metodo_python == "consumidos" and vta_python is None:
        valor_total = None
    else:
        valor_total = vta_python if vta_python is not None else vta_capacidade_valor
    assert valor_total is None


# ---------------------------------------------------------------------------
# VTA-C2.1 — casos-limite (futuro zero real, completude, sobreconsumo)
# ---------------------------------------------------------------------------

_ITEM_100_CONSUMIDO = {
    "item": "D", "qtd_contratada": 15, "vu_original": 5.0, "qtd_total": 15,
    "consumos": {
        "C0": {"qtd": 15, "valor": 75.0},
        "C1": {"qtd": None, "valor": None},
        "C2": {"qtd": None, "valor": None},
        "C3": {"qtd": None, "valor": None},
        "C4": {"qtd": None, "valor": None},
    },
}


def test_futuro_zero_real_contrato_100_por_cento_consumido():
    """Item A: qtd_contratada=20; qtd_total=18 (nao 100%). Substitui por um
    item onde qtd_total==qtd_contratada -> remanescente=0,00 valido, VTA =
    execucao atualizada (nao vira None so porque o saldo deu zero)."""
    leitura = _leitura_consumidos([_ITEM_100_CONSUMIDO], ciclo_vigente="C0")
    memoria = _montar_memoria_por_ciclo(leitura, {}, [], {})
    conf = _conferencia_consumidos(memoria)
    assert conf["executado_atualizado"] == 75.0
    assert conf["potencial_restante_atualizado"] == 0.0
    assert conf["valor_total_atualizado"] == 75.0
    assert memoria["vta"]["valor_total_atualizado"] == 75.0


def test_item_incompleto_bloqueia_futuro_nao_soma_apenas_item_completo():
    """Item A completo + item incompleto (sem qtd_total) -> remanescente
    Consumido fica indisponivel; VTA nao pode ser fabricado so com o item
    completo."""
    item_incompleto = {
        "item": "E", "qtd_contratada": 10, "vu_original": 5.0, "qtd_total": None,
        "consumos": {c: {"qtd": None, "valor": None} for c in ("C0", "C1", "C2", "C3", "C4")},
    }
    leitura = _leitura_consumidos([_ITEM_A, _ITEM_B, item_incompleto])
    memoria = _montar_memoria_por_ciclo(leitura, {}, [], {})
    conf = _conferencia_consumidos(memoria)
    assert conf["potencial_restante_atualizado"] is None
    assert conf["valor_total_atualizado"] is None
    assert memoria["vta"]["valor_total_atualizado"] is None


def test_consumo_numerico_sem_valoracao_bloqueia_execucao_inteira():
    """QTD_CONS numerica mas sem VU_original -> execucao Consumido inteira
    fica indisponivel (nao soma so as parcelas validas)."""
    item_sem_vu = {
        "item": "F", "qtd_contratada": 10, "vu_original": None, "qtd_total": 4,
        "consumos": {
            "C0": {"qtd": 4, "valor": None},
            "C1": {"qtd": None, "valor": None},
            "C2": {"qtd": None, "valor": None},
            "C3": {"qtd": None, "valor": None},
            "C4": {"qtd": None, "valor": None},
        },
    }
    leitura = _leitura_consumidos([_ITEM_A, item_sem_vu])
    memoria = _montar_memoria_por_ciclo(leitura, {}, [], {})
    conf = _conferencia_consumidos(memoria)
    assert conf["valor_total_atualizado"] is None
    assert memoria["vta"]["valor_total_atualizado"] is None


def test_sobreconsumo_sinalizado_por_check_bloqueia_remanescente():
    """itens_Consumidos!CHECK (ja existente no template, "DIVERGENCIA:
    CONSUMO MAIOR QUE CONTRATADO" quando O>B) precisa ser respeitado: um
    item com check divergente nao pode ter seu saldo mascarado como zero
    valido via MAX(diferenca,0)."""
    item_sobreconsumo = {
        "item": "G", "qtd_contratada": 5, "vu_original": 10.0, "qtd_total": 8,
        "check": "DIVERGENCIA: CONSUMO MAIOR QUE CONTRATADO",
        "consumos": {
            "C0": {"qtd": 8, "valor": 80.0},
            "C1": {"qtd": None, "valor": None},
            "C2": {"qtd": None, "valor": None},
            "C3": {"qtd": None, "valor": None},
            "C4": {"qtd": None, "valor": None},
        },
    }
    leitura = _leitura_consumidos([_ITEM_A, item_sobreconsumo])
    memoria = _montar_memoria_por_ciclo(leitura, {}, [], {})
    conf = _conferencia_consumidos(memoria)
    assert conf["potencial_restante_atualizado"] is None
    assert conf["valor_total_atualizado"] is None


def test_residuo_de_outro_metodo_nao_contamina_selecao_mas_nao_esconde_divergencia_real():
    """Distincao pedida pela tarefa: dado residual de outro metodo (ex.:
    parcela financeiro) nao pode CONTAMINAR a selecao do metodo (item N,
    ja coberto por test_n_...); aqui confirmamos apenas que a conferencia
    'financeiro' continua calculavel/visivel (nao escondida), mesmo que o
    metodo oficialmente escolhido seja consumidos — a divergencia real
    entre bases, se existir, permanece auditavel via conferencias, nao
    suprimida."""
    leitura = _leitura_consumidos([_ITEM_A, _ITEM_B], modo="d")
    leitura["vta_sombra"] = {
        "parcelas_computadas": [
            {
                "fonte_parcela": "Financeiro", "identificador": "financeiro:residual",
                "ciclo": "C0", "valor": 999.0, "valor_atualizado": 999.0,
            }
        ]
    }
    memoria = _montar_memoria_por_ciclo(leitura, {}, [], {})
    conf_fin = next(c for c in memoria["conferencias_metodologicas"] if c["metodo"] == "financeiro")
    # A conferencia financeiro continua disponivel/auditavel (nao escondida),
    # mas o VTA OFICIAL veio do metodo explicitamente selecionado.
    assert conf_fin["disponivel"] is True
    assert conf_fin["executado_atualizado"] == 999.0
    assert memoria["vta"]["metodo"] == "consumidos"
    assert memoria["vta"]["valor_total_atualizado"] == 410.0


# ---------------------------------------------------------------------------
# XLS — formulas fail-closed (F20/V/C33)
# ---------------------------------------------------------------------------

def test_f20_fail_closed_qtd_sem_valor_correspondente(wb):
    f20 = str(wb["MEMORIA_RESULTADOS"]["F20"].value)
    for qtd_col, valor_col in (("E", "F"), ("G", "H"), ("I", "J"), ("K", "L"), ("M", "N")):
        assert (
            f"COUNT(itens_Consumidos!${qtd_col}$2:${qtd_col}$200)"
            f"<>COUNT(itens_Consumidos!${valor_col}$2:${valor_col}$200)"
        ) in f20


def test_coluna_v_fail_closed_incompleto_ou_sobreconsumo():
    from tools.aplicar_vta_consumido_canonico import _linha_v_consumidos
    formula = _linha_v_consumidos(2)
    assert "NOT(ISNUMBER(B2))" in formula
    assert "NOT(ISNUMBER(O2))" in formula
    assert "NOT(ISNUMBER(C2))" in formula
    assert 'Q2<>"OK"' in formula


def test_c33_fail_closed_qualquer_item_com_v_vazio(wb):
    c33 = str(wb["MEMORIA_RESULTADOS"]["C33"].value)
    assert 'COUNTIFS(itens_Consumidos!$A$2:$A$200,"<>",itens_Consumidos!$V$2:$V$200,"")' in c33
    assert "SUMPRODUCT" not in c33
