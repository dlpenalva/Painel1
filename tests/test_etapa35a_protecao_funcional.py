"""Etapa 35A: contratos de protecao antes das correcoes funcionais.

Os testes deste arquivo descrevem o comportamento desejado ja decidido para
as etapas 35B--35E. GREEN significa comportamento ja correto e protegido.
XFAIL STRICT significa bug confirmado que ainda aguarda sua etapa corretora.
Nenhuma etapa intermediaria deve manter RED intencional: os defeitos conhecidos
ficam em XFAIL STRICT. Se uma correcao produzir XPASS antes da retirada consciente
do marcador na etapa correspondente, o modo estrito transforma isso em erro.

Nenhum teste abaixo autoriza ``session_state.clear()`` nem transforma revisao
em bloqueio de documento. C0=1,0 tambem continua sendo um valor legitimo.
"""

from __future__ import annotations

import io
import zipfile
from datetime import date
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from _coleta_oficial import gerar_coleta_oficial_preenchida
from _coleta_reajuste_documentos import (
    adaptar_coleta_reajuste_para_documentos,
    aplicar_bloqueio_documental,
)
from _templates_documentos import (
    ROTULO_PREVIA,
    gerar_despacho_saneador,
    gerar_termo_apostila,
)
from test_26h_novos_itens import _leitura_sintetica_26h
from test_sumario_executivo import leitura_ausencias


ROOT = Path(__file__).resolve().parents[1]


def _dados_coleta_multiciclo() -> dict:
    """Coleta minima com C1 historico e C2 como ciclo vigente."""
    return {
        "origem": "Protecao funcional Etapa 35A",
        "indice": "IST (Anatel)",
        "data_base_original": "01/02/2024",
        "data_corte": date(2027, 1, 31),
        "ciclos": [
            {
                "ciclo": "C1",
                "data_inicio": date(2025, 2, 1),
                "data_fim": date(2026, 1, 31),
                "data_pedido": date(2025, 2, 1),
                "financeiro_inicio": date(2025, 2, 1),
                "percentual_aplicado": 0.10,
                "situacao": "TEMPESTIVO",
                "objeto_analise_atual": False,
            },
            {
                "ciclo": "C2",
                "data_inicio": date(2026, 2, 1),
                "data_fim": date(2027, 1, 31),
                "data_pedido": date(2026, 2, 1),
                "financeiro_inicio": date(2026, 2, 1),
                "percentual_aplicado": 0.02,
                "situacao": "TEMPESTIVO",
                "objeto_analise_atual": True,
            },
        ],
    }


@pytest.fixture(scope="module")
def coleta_base_multiciclo() -> bytes:
    return gerar_coleta_oficial_preenchida(_dados_coleta_multiciclo())


def _mutar_coleta(conteudo: bytes, mutacao) -> bytes:
    wb = load_workbook(io.BytesIO(conteudo), data_only=False)
    mutacao(wb)
    destino = io.BytesIO()
    wb.save(destino)
    wb.close()
    return destino.getvalue()


def _adaptar(conteudo: bytes, ciclo_vigente: str) -> dict:
    # O teste isola o contrato do adaptador. A validacao estrutural e os
    # motores completos ja possuem suites proprias.
    diagnostico = {
        "valido": True,
        "capacidades": {"calculos": {}, "documentos": {}},
    }
    return adaptar_coleta_reajuste_para_documentos(
        conteudo,
        leitura={"controle": {"ciclo_vigente": ciclo_vigente}},
        diagnostico=diagnostico,
    )


def _xml_docx(conteudo: bytes) -> str:
    with zipfile.ZipFile(io.BytesIO(conteudo)) as pacote:
        return pacote.read("word/document.xml").decode("utf-8")


def _texto_dou(resultado: dict) -> str:
    from streamlit.testing.v1 import AppTest

    at = AppTest.from_file(str(ROOT / "pages" / "13_DOU.py"), default_timeout=90)
    at.session_state["resultado_valor_global"] = resultado
    at.run()
    assert not at.exception, str(getattr(at, "exception", ""))
    return at.text_area(key="previa_dou").value


def _acionar_troca_de_caso(estado: dict, nova_assinatura: str) -> None:
    """Adaptador temporario do teste para o futuro ponto de entrada da 35C.

    Nome, modulo, assinatura e estrategia de invalidacao NAO fazem parte do
    contrato funcional. Este adaptador pode ser ajustado na 35C sem alterar as
    expectativas abaixo: derivados antigos ficam invalidos e globais ficam.
    """
    try:
        from _estado_apuracao_upload import invalidar_estado_caso
    except ImportError:
        pytest.fail(
            "ainda nao existe um ponto central para invalidar o estado do caso",
            pytrace=False,
        )
    invalidar_estado_caso(estado, nova_assinatura)


# ---------------------------------------------------------------------------
# TESTE 1 - troca de upload / novo caso (GREEN desde a 35C)
# ---------------------------------------------------------------------------


def test_novo_hash_invalida_somente_estado_derivado_do_caso() -> None:
    """Contrato funcional; nao fixa nome, modulo ou assinatura da solucao."""
    estado = {
        "assinatura_upload_docs": "hash-caso-a",
        "assinatura_processada_upload_docs": "hash-caso-a",
        "resultado_valor_global": {"caso": "A"},
        "diagnostico_coleta_v2": {"caso": "A"},
        "dados_admissibilidade": {"caso": "A"},
        "arquivo_sumario_executivo_pdf": b"caso-a",
        "arquivo_despacho_saneador_docx": b"caso-a",
        "arquivo_termo_apostila_docx": b"caso-a",
        "arquivo_garantia_pdf": b"caso-a",
        "resultado_garantia": {"caso": "A"},
        "adequacao_v3_retroativo": 123.45,
        "adequacao_v3_exclusoes_pc": {"PC-A"},
        "previa_dou": "DOU do caso A",
        "arquivo_dou_docx": b"caso-a",
        "infos_previas_df": pd.DataFrame([{"caso": "A"}]),
        "avaliacao_aditivos_eventos": [{"caso": "A"}],
        # Preferencias globais nao pertencem ao contrato analisado.
        "tema_interface": "claro",
        "navegacao_global": "inicio",
    }

    _acionar_troca_de_caso(estado, "hash-caso-b")

    assert estado["assinatura_upload_docs"] == "hash-caso-b"
    for chave in (
        "assinatura_processada_upload_docs",
        "resultado_valor_global",
        "diagnostico_coleta_v2",
        "dados_admissibilidade",
        "arquivo_sumario_executivo_pdf",
        "arquivo_despacho_saneador_docx",
        "arquivo_termo_apostila_docx",
        "arquivo_garantia_pdf",
        "resultado_garantia",
        "adequacao_v3_retroativo",
        "adequacao_v3_exclusoes_pc",
        "previa_dou",
        "arquivo_dou_docx",
        "infos_previas_df",
        "avaliacao_aditivos_eventos",
    ):
        assert chave not in estado or estado[chave] is None, (
            f"estado do caso anterior ainda valido: {chave}"
        )
    assert estado["tema_interface"] == "claro"
    assert estado["navegacao_global"] == "inicio"


# ---------------------------------------------------------------------------
# TESTES 2 e 3 - fator acumulado (GREEN desde a 35B)
# ---------------------------------------------------------------------------


def test_ciclo_computado_sem_percentual_nao_recebe_fator_um(
    coleta_base_multiciclo: bytes,
) -> None:
    def mutacao(wb) -> None:
        wb["CONTROLE"]["B2"] = "C1"
        parametros = wb["parametros"]
        parametros["A3"] = "Sim"
        parametros["E3"] = None
        parametros["F3"] = None

    resultado = _adaptar(_mutar_coleta(coleta_base_multiciclo, mutacao), "C1")
    ciclo = resultado["df_ciclos"].loc[
        resultado["df_ciclos"]["Ciclo"] == "C1"
    ].iloc[0]

    assert pd.isna(ciclo["Fator acumulado"]), (
        "percentual/fator ausente em ciclo computado deve permanecer nao calculavel; "
        f"recebido={ciclo['Fator acumulado']!r}"
    )


def test_fator_apresentado_e_do_ciclo_vigente_e_nao_o_maximo_historico(
    coleta_base_multiciclo: bytes,
) -> None:
    def mutacao(wb) -> None:
        wb["CONTROLE"]["B2"] = "C2"
        parametros = wb["parametros"]
        parametros["A3"] = "Nao"  # C1 historico, fora da apuracao corrente
        parametros["E3"] = 0.10
        parametros["F3"] = None
        parametros["A4"] = "Sim"  # C2 vigente
        parametros["E4"] = -0.05
        parametros["F4"] = None

    resultado = _adaptar(_mutar_coleta(coleta_base_multiciclo, mutacao), "C2")

    # Cadeia canonica: C1=1,10; C2 vigente=1,10*0,95=1,045.
    assert resultado["fator_acumulado"] == pytest.approx(1.045)
    assert resultado["variacao_acumulada"] == pytest.approx(0.045)


def test_inicio_em_c2_preserva_c1_fora_e_falha_fechada_sem_cadeia(
    coleta_base_multiciclo: bytes,
) -> None:
    """C1 vazio e fora da apuracao nao e inventado para fechar a cadeia de C2."""

    def mutacao(wb) -> None:
        wb["CONTROLE"]["B2"] = "C2"
        parametros = wb["parametros"]
        parametros["A3"] = "Nao"
        parametros["E3"] = None
        parametros["F3"] = None
        parametros["A4"] = "Sim"
        parametros["E4"] = 0.05
        parametros["F4"] = None

    conteudo = _mutar_coleta(coleta_base_multiciclo, mutacao)
    wb = load_workbook(io.BytesIO(conteudo), data_only=False)
    try:
        assert wb["parametros"]["A3"].value == "Nao"
        assert wb["parametros"]["E3"].value is None
        assert wb["parametros"]["F3"].value is None
    finally:
        wb.close()

    resultado = _adaptar(conteudo, "C2")
    ciclos = resultado["df_ciclos"].set_index("Ciclo")

    assert ciclos.loc["C1", "Tratamento financeiro do ciclo"] == "Fora da apuração"
    assert ciclos.loc["C2", "Tratamento financeiro do ciclo"] == "Apurar"
    assert pd.isna(ciclos.loc["C2", "Fator acumulado"])
    assert resultado["fator_acumulado"] is None
    assert resultado["variacao_acumulada"] is None

    # A insuficiencia permanece revisavel; nao impede gerar os documentos.
    assert gerar_despacho_saneador(resultado)[:2] == b"PK"
    assert gerar_termo_apostila(resultado)[:2] == b"PK"


def test_inicio_em_c2_ate_c4_consolida_fator_do_c4_sem_influencia_do_c1(
    coleta_base_multiciclo: bytes,
) -> None:
    """O maior fator historico e o fallback de C1 nao substituem o C4 vigente."""

    def mutacao(wb) -> None:
        wb["CONTROLE"]["B2"] = "C4"
        parametros = wb["parametros"]
        parametros["A3"] = "Nao"
        parametros["E3"] = None
        parametros["F3"] = None
        parametros["A4"] = "Sim"
        parametros["E4"] = 0.08
        parametros["F4"] = 1.08
        parametros["A5"] = "Sim"
        parametros["E5"] = 0.04
        parametros["F5"] = 1.1232
        parametros["A6"] = "Sim"
        parametros["E6"] = -0.03
        parametros["F6"] = 1.089504

    resultado = _adaptar(_mutar_coleta(coleta_base_multiciclo, mutacao), "C4")
    ciclos = resultado["df_ciclos"].set_index("Ciclo")

    assert ciclos.loc["C1", "Tratamento financeiro do ciclo"] == "Fora da apuração"
    for ciclo in ("C2", "C3", "C4"):
        assert ciclos.loc[ciclo, "Tratamento financeiro do ciclo"] == "Apurar"
    assert ciclos.loc["C1", "Fator acumulado"] == pytest.approx(1.0)
    assert ciclos.loc["C3", "Fator acumulado"] == pytest.approx(1.1232)
    assert ciclos.loc["C4", "Fator acumulado"] == pytest.approx(1.089504)
    assert resultado["fator_acumulado"] == pytest.approx(1.089504)
    assert resultado["variacao_acumulada"] == pytest.approx(0.089504)

    # Os documentos continuam geraveis e nao adotam o fallback de C1 nem o
    # maximo de C3 como percentual acumulado. Quando a camada documental nao
    # possui evidencia suficiente para imprimir o acumulado, ela segue com o
    # placeholder fail-closed ja homologado.
    for gerar in (gerar_despacho_saneador, gerar_termo_apostila):
        conteudo = gerar(resultado)
        xml = _xml_docx(conteudo)
        assert conteudo[:2] == b"PK"
        assert "12,32%" not in xml
        assert "0,00%" not in xml


# ---------------------------------------------------------------------------
# TESTES 4 e 5 - minuta DOU (XFAIL STRICT ate a 35D)
# ---------------------------------------------------------------------------


@pytest.mark.xfail(
    reason="Etapa 35D - minuta DOU deve usar somente ciclos computados",
    strict=True,
)
def test_dou_exibe_somente_ciclos_computados_na_apuracao_corrente() -> None:
    resultado = {
        "df_ciclos": pd.DataFrame(
            [
                {
                    "Ciclo": "C1",
                    "Percentual aplicado": 0.05,
                    "Fator acumulado": 1.05,
                    "Tratamento financeiro do ciclo": "Apurar",
                },
                {
                    "Ciclo": "C2",
                    "Percentual aplicado": 0.09,
                    "Fator acumulado": 1.1445,
                    "Tratamento financeiro do ciclo": "Fora da apuracao",
                },
            ]
        ),
        "valor_atualizado_contrato": 1000.0,
    }

    texto = _texto_dou(resultado)
    assert "C1: 5,00%" in texto
    assert "C2:" not in texto


@pytest.mark.xfail(
    reason="Etapa 35D - percentual ausente no DOU nao pode virar 0,00%",
    strict=True,
)
def test_dou_percentual_ausente_permanece_placeholder() -> None:
    resultado = {
        "df_ciclos": pd.DataFrame(
            [
                {
                    "Ciclo": "C1",
                    "Percentual aplicado": None,
                    "Fator acumulado": None,
                    "Tratamento financeiro do ciclo": "Apurar",
                }
            ]
        ),
        "valor_atualizado_contrato": 1000.0,
    }

    texto = _texto_dou(resultado)
    assert "C1: 0,00%" not in texto
    assert "C1: [preencher campo]" in texto


# ---------------------------------------------------------------------------
# TESTE 6 - documento disponivel com revisao (GREEN esperado)
# ---------------------------------------------------------------------------


def test_revisao_nao_bloqueia_apostila_nem_saneador() -> None:
    capacidades = {
        "documentos": {
            "despacho_saneador": {"habilitado": True, "estado": "disponivel"},
            "termo_apostila": {"habilitado": True, "estado": "disponivel"},
            "dou": {"habilitado": True, "estado": "disponivel"},
        }
    }

    aplicar_bloqueio_documental(
        capacidades,
        ["Valor sujeito a revisao antes da formalizacao."],
    )

    assert capacidades["documentos"]["despacho_saneador"]["habilitado"] is True
    assert capacidades["documentos"]["termo_apostila"]["habilitado"] is True
    # A disponibilidade futura do DOU editavel nao faz parte deste contrato.


@pytest.mark.parametrize(
    "gerador",
    [gerar_despacho_saneador, gerar_termo_apostila],
)
def test_documento_em_revisao_mantem_selo_previa_verde(gerador) -> None:
    conteudo = gerador(_leitura_sintetica_26h())
    xml = _xml_docx(conteudo)

    assert conteudo[:2] == b"PK"
    assert ROTULO_PREVIA in xml
    assert '<w:highlight w:val="green"' in xml


@pytest.mark.parametrize(
    "gerador",
    [gerar_despacho_saneador, gerar_termo_apostila],
)
def test_documento_sem_valor_confiavel_mantem_placeholder_amarelo(gerador) -> None:
    conteudo = gerador(leitura_ausencias())
    xml = _xml_docx(conteudo)

    assert conteudo[:2] == b"PK"
    assert "[PREENCHER:" in xml
    assert '<w:highlight w:val="yellow"' in xml
