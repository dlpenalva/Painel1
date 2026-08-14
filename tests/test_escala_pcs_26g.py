# -*- coding: utf-8 -*-
"""Etapa 26G — testes de fronteira da capacidade canonica de PCs.

Cobrem: grade completa de itens_PC (formulas/DV/CF) nas linhas de fronteira,
resumo lateral, cobertura temporal, MEMORIA/RESULTADOS sem faixas truncadas,
fail-closed do VTA (guards), duplicidade e bloqueio de capacidade no upload,
e o estilo da linha 101 de itens_Remanesc.
"""
from __future__ import annotations

import io
import re
import sys
from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from _capacidade_pcs import CAPACIDADE_PCS, ULTIMA_LINHA_PCS
from _coleta_oficial import TEMPLATE_COLETA_OFICIAL, obter_coleta_oficial_bytes
from _coleta_reajuste import ler_coleta_reajuste

CAP = ULTIMA_LINHA_PCS
FRONTEIRAS = (99, 100, 101, 200, 201, 2119, CAP)


@pytest.fixture(scope="module")
def wb_template():
    return load_workbook(TEMPLATE_COLETA_OFICIAL, data_only=False)


def test_grade_itens_pc_completa_nas_fronteiras(wb_template):
    ws = wb_template["itens_PC"]
    for r in FRONTEIRAS:
        for col in "CEFHIJKL":
            valor = ws[f"{col}{r}"].value
            assert isinstance(valor, str) and valor.startswith("="), (
                f"itens_PC!{col}{r} sem formula"
            )
        assert f"COUNTIF($A$2:$A${CAP},A{r})>1" in ws[f"K{r}"].value
        assert "EFEITO" not in str(ws[f"L{r}"].value)[:3]  # formula, nao texto


def test_duplicidade_formula_sem_sumproduct_quadratico(wb_template):
    ws = wb_template["itens_PC"]
    assert "SUMPRODUCT(--(UPPER(TRIM(" not in str(ws["K2"].value)
    assert f"COUNTIF($A$2:$A${CAP},A2)>1" in str(ws["K2"].value)


def test_data_validation_e_condicional_em_toda_grade(wb_template):
    ws = wb_template["itens_PC"]
    validacoes = [
        (dv.type, str(dv.sqref)) for dv in ws.data_validations.dataValidation
    ]
    assert validacoes == [("list", f"G2:G{CAP}")]
    faixas = [str(rng.sqref) for rng in ws.conditional_formatting._cf_rules]
    assert faixas == [f"A2:L{CAP}"]
    regras = list(ws.conditional_formatting._cf_rules.values())[0]
    assert len(regras) == 2


def test_formatos_visuais_apos_linha_100(wb_template):
    ws = wb_template["itens_PC"]
    ref = {c: ws[f"{c}99"] for c in "ABDFKL"}
    for r in (100, 101, 200, 201, 2119, CAP):
        for c in "ABDFKL":
            cel = ws[f"{c}{r}"]
            assert cel.font.name == ref[c].font.name, f"{c}{r} fonte"
            assert cel.number_format == ref[c].number_format, f"{c}{r} formato"
            assert cel.fill.fgColor.rgb == ref[c].fill.fgColor.rgb, f"{c}{r} fill"
            assert (
                cel.border.left.style == ref[c].border.left.style
                and cel.border.bottom.style == ref[c].border.bottom.style
            ), f"{c}{r} borda"


def test_resumo_lateral_cobre_toda_faixa(wb_template):
    ws = wb_template["itens_PC"]
    assert ws["N2"].value == f"=COUNTIF($C$2:$C${CAP},M2)"
    assert ws["O2"].value == f"=SUMIF($C$2:$C${CAP},M2,$D$2:$D${CAP})"
    assert ws["P2"].value == f"=SUMIF($C$2:$C${CAP},M2,$F$2:$F${CAP})"
    assert ws["Q6"].value == f"=SUMIF($C$2:$C${CAP},M6,$H$2:$H${CAP})"
    assert ws["R6"].value == f"=SUMIF($C$2:$C${CAP},M6,$I$2:$I${CAP})"
    assert ws["S6"].value == f"=SUMIF($C$2:$C${CAP},M6,$J$2:$J${CAP})"
    assert ws["N7"].value == "=SUM(N2:N6)"


def test_cobertura_b14_cobre_toda_faixa(wb_template):
    # Etapa "posicao unica, datas e arredondamento" (PR #22, owner:
    # tools/aplicar_ux_posicao_datas_arredondamento.py): B14 deixou de ser o
    # MAX cru e virou o ULTIMO PC CONSIDERADO ATE O CORTE (CONTROLE!B3), com
    # fallback MAX quando nao ha corte e vazio quando nenhum PC e anterior ao
    # corte. A faixa segue cobrindo a capacidade canonica ($B$2:$B$5001).
    b14 = wb_template["cobertura_temporal"]["B14"].value
    assert b14 == (
        f'=IFERROR(IF(COUNT(itens_PC!$B$2:$B${CAP})=0,"",'
        f'IF(NOT(ISNUMBER(CONTROLE!$B$3)),MAX(itens_PC!$B$2:$B${CAP}),'
        f'IF(COUNTIFS(itens_PC!$B$2:$B${CAP},"<="&CONTROLE!$B$3,'
        f'itens_PC!$B$2:$B${CAP},">0")=0,"",'
        f'SUMPRODUCT(MAX((itens_PC!$B$2:$B${CAP}<=CONTROLE!$B$3)'
        f'*(itens_PC!$B$2:$B${CAP}>0)*itens_PC!$B$2:$B${CAP}))))),"")'
    )


def test_sem_faixas_truncadas_de_itens_pc(wb_template):
    padrao = re.compile(r"itens_PC!\$[A-Z]{1,2}\$2:\$[A-Z]{1,2}\$(100|200)\b")
    residuais = [
        f"{ws.title}!{cell.coordinate}"
        for ws in wb_template.worksheets
        for row in ws.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and padrao.search(cell.value)
    ]
    assert residuais == []


def test_memoria_fail_closed_vta(wb_template):
    mem = wb_template["MEMORIA_RESULTADOS"]
    # Guards por item: ausencia vira "" (nunca 0, nunca aritmetica com vazio).
    assert "ISNUMBER(posicao_contratual!$G2)" in mem["X2"].value
    assert "ISNUMBER(CHOOSE($T$20+1" in mem["Y2"].value
    # Contadores de completude e indisponibilidade fail-closed.
    assert "$T$26>0" in mem["T25"].value
    assert "CALCULO MANUAL REQUERIDO" in mem["T25"].value
    assert "1-ISNUMBER(CHOOSE($T$20+1" in mem["T26"].value
    assert "historico_VU!$C$2:$C$201" in mem["T27"].value
    assert 'IF($T$26>0,""' in mem["C32"].value
    # Nao ha coercao vazio->zero nos guards.
    for coord in ("X2", "Y2", "T25", "C32"):
        assert "IF(vazio;0" not in str(mem[coord].value)


def test_memoria_helpers_pcs_sem_efeito(wb_template):
    mem = wb_template["MEMORIA_RESULTADOS"]
    t28 = mem["T28"].value
    assert f'itens_PC!$L$2:$L${CAP},"Nao"' in t28
    assert '(parametros!$A$3="Sim")' in t28  # C1 condicionado a apuracao
    assert '"C0"' not in t28  # C0 jamais entra (sem efeito por definicao)
    assert f"itens_PC!$D$2:$D${CAP}" in mem["T29"].value
    assert f"itens_PC!$F$2:$F${CAP}" in mem["T30"].value


def test_resultados_linha_executiva_pcs_sem_efeito(wb_template):
    res = wb_template["RESULTADOS"]
    assert "PCs sem efeito financeiro" in res["A23"].value
    assert "MEMORIA_RESULTADOS!$T$29" in res["B23"].value
    assert "MEMORIA_RESULTADOS!$T$30" in res["D23"].value
    assert "EFEITO_FINANCEIRO_PC = Nao" in res["E23"].value
    # 26G.1: so aparece no metodo PCs e quando ha PCs sem efeito (T28>0);
    # caso contrario a RESULTADOS permanece limpa.
    for coord in ("A23", "B23", "C23", "D23", "E23"):
        assert res[coord].value.startswith(
            '=IF(OR(MEMORIA_RESULTADOS!$B$4<>"PCs",MEMORIA_RESULTADOS!$T$28=0),'
        )


def test_resultados_tabelas_cobrem_toda_faixa(wb_template):
    res = wb_template["RESULTADOS"]
    for coord in ("B16", "B17", "B18", "B19", "B20", "B36"):
        formula = res[coord].value
        assert f"itens_PC!$C$2:$C${CAP}" in formula, coord


def test_itens_remanesc_linha_101_igual_vizinhas(wb_template):
    ws = wb_template["itens_Remanesc"]
    for col in "ABCEGIK":
        ref = ws[f"{col}100"]
        alvo = ws[f"{col}101"]
        depois = ws[f"{col}102"]
        assert alvo.font.name == ref.font.name == depois.font.name
        assert alvo.font.b == ref.font.b
        assert alvo.font.sz == ref.font.sz
        assert alvo.alignment.horizontal == ref.alignment.horizontal
        assert alvo.fill.fgColor.rgb == ref.fill.fgColor.rgb
        assert alvo.number_format == ref.number_format
    for col in "DFHJ":
        assert (
            ws[f"{col}101"].alignment.horizontal
            == ws[f"{col}100"].alignment.horizontal
        )


# ---------------------------------------------------------------------------
# Upload (validacao Python em toda a base)
# ---------------------------------------------------------------------------

def _coleta_com_pcs(pcs, extras=None):
    """Gera a coleta oficial em branco e insere PCs (linha, numero, data, valor)."""
    wb = load_workbook(io.BytesIO(obter_coleta_oficial_bytes()), data_only=False)
    par = wb["parametros"]
    dados_par = {
        2: ("Nao", "C0", datetime(2025, 3, 1), datetime(2026, 1, 31), None, None),
        3: ("Sim", "C1", datetime(2026, 2, 1), datetime(2027, 1, 31), 0.05,
            datetime(2026, 4, 1)),
    }
    for linha, (a, b, c, d, e, h) in dados_par.items():
        par[f"A{linha}"] = a
        par[f"B{linha}"] = b
        par[f"C{linha}"] = c
        par[f"D{linha}"] = d
        par[f"E{linha}"] = e
        if h is not None:
            par[f"H{linha}"] = h
    ws = wb["itens_PC"]
    for linha, numero, data, valor in pcs:
        ws[f"A{linha}"] = numero
        ws[f"B{linha}"] = data
        ws[f"D{linha}"] = valor
        ws[f"G{linha}"] = "Sim"
    if extras:
        extras(wb)
    saida = io.BytesIO()
    wb.save(saida)
    return saida.getvalue()


@pytest.fixture(scope="module")
def resultado_fronteiras():
    pcs = [
        (linha, f"PC{linha}", datetime(2026, 5, 10), 100.0)
        for linha in (99, 100, 101, 200, 201, 2050, CAP)
    ]
    return ler_coleta_reajuste(_coleta_com_pcs(pcs))


def test_upload_conta_pcs_alem_da_linha_100(resultado_fronteiras):
    assert resultado_fronteiras["contagens"]["pedidos_de_compra"] == 7
    assert not any(
        "duplicado" in p for p in resultado_fronteiras["bloqueios_criticos"]
    )


def test_upload_nao_bloqueia_fronteiras_validas(resultado_fronteiras):
    assert resultado_fronteiras["bloqueios_estruturais"] == []


def test_upload_detecta_duplicidade_apos_linha_100():
    pcs = [
        (2, "4100000001", datetime(2026, 5, 10), 100.0),
        (150, "4100000001", datetime(2026, 5, 11), 200.0),
    ]
    resultado = ler_coleta_reajuste(_coleta_com_pcs(pcs))
    assert any(
        "NUMERO_PC duplicado: 4100000001" in p and "150" in p
        for p in resultado["bloqueios_criticos"]
    )


def test_upload_detecta_duplicidade_apos_linha_2000():
    pcs = [
        (50, "PC-X ", datetime(2026, 5, 10), 100.0),
        (2100, "pc-x", datetime(2026, 5, 11), 200.0),  # TRIM+UPPER
    ]
    resultado = ler_coleta_reajuste(_coleta_com_pcs(pcs))
    assert any(
        "NUMERO_PC duplicado: PC-X" in p for p in resultado["bloqueios_criticos"]
    )


def test_upload_bloqueia_capacidade_excedida():
    pcs = [(2, "PC1", datetime(2026, 5, 10), 100.0)]

    def alem(wb):
        wb["itens_PC"][f"A{CAP + 1}"] = "PC-ALEM"
        wb["itens_PC"][f"B{CAP + 1}"] = datetime(2026, 5, 12)
        wb["itens_PC"][f"D{CAP + 1}"] = 10.0

    resultado = ler_coleta_reajuste(_coleta_com_pcs(pcs, extras=alem))
    assert any(
        "Capacidade de PCs excedida" in p
        for p in resultado["bloqueios_estruturais"]
    )
    assert resultado["valido"] is False


def test_capacidade_suporta_caso_real():
    assert CAPACIDADE_PCS >= 2118 * 2, "margem minima de 2x sobre o caso real"
