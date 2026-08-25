# -*- coding: utf-8 -*-
"""Etapa 51C — ajustes visuais/estruturais das abas da Coleta oficial.

Protege os quatro ajustes de APRESENTACAO (nenhuma formula/valor/nome muda):
parametros!I integrada ao quadro; financeiro TOTAL (ancora na linha 74)
destacado e aproximado da grade por ocultacao das linhas de capacidade nao
usadas (runtime); aditivos linha 1 com altura suficiente para o texto de
orientacao; cobertura_temporal padronizada preservando os fills funcionais.
"""
from io import BytesIO
from pathlib import Path

import openpyxl
import pytest

ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "templates" / "COLETA_REAJUSTE_OFICIAL.xlsx"


@pytest.fixture(scope="module")
def wb():
    return openpyxl.load_workbook(TEMPLATE, data_only=False)


# ------------------------------------------------------------- parametros!I

def test_parametros_coluna_i_integrada_ao_quadro(wb):
    ws = wb["parametros"]
    # Endereco e conteudo intactos: a coluna continua vazia no template
    # (o runtime da Etapa 48 escreve cabecalho e datas).
    assert all(ws[f"I{r}"].value is None for r in range(1, 7))
    # Cabecalho com a MESMA linguagem visual de A1:H1.
    i1, h1 = ws["I1"], ws["H1"]
    assert i1.fill.fgColor.rgb == h1.fill.fgColor.rgb == "FF1F4E79"
    assert i1.font.bold and i1.font.size == 10
    assert i1.font.color.rgb == "FFFFFFFF"
    assert i1.border.left.style == "thin" and i1.border.bottom.style == "thin"
    # Corpo com o padrao do quadro; H é competência e I preserva data exata.
    assert ws["I3"].fill.fgColor.rgb == "FFEDEDED"
    assert ws["H3"].number_format == "mm/yyyy;@"
    assert ws["I3"].number_format == "dd/mm/yyyy;@"
    assert ws["I2"].border.left.style == "thin"


# ------------------------------------------------------------- financeiro TOTAL

def test_financeiro_total_ancora_intacta_e_destacada(wb):
    ws = wb["financeiro"]
    # Ancora canonica: NADA se move (linhas 62:73 sao capacidade estrutural).
    assert ws["B74"].value == "TOTAL"
    assert ws["C74"].value == "=SUM(C2:C73)"
    assert ws["E74"].value == "=SUM(E2:E73)"
    assert ws["F74"].value == "=SUM(F2:F73)"
    assert str(ws["B64"].value).startswith("=IF(A64")   # grade viva na linha 64
    # Destaque visual: linha inteira em negrito com fill institucional.
    for col in range(1, 8):
        cel = ws.cell(row=74, column=col)
        assert cel.font.bold, f"coluna {col} do TOTAL sem negrito"
        assert cel.fill.fgColor.rgb == "FFD6E4F0"


def test_financeiro_gerado_oculta_capacidade_nao_usada():
    """Runtime: 60 competencias escritas -> linhas 62:73 ocultas e o TOTAL
    (linha 74) aparece visualmente junto da grade. Valor do TOTAL intacto."""
    from _coleta_oficial import gerar_coleta_oficial_preenchida
    from test_coleta_oficial_integracao import _dados_calculadora

    conteudo = gerar_coleta_oficial_preenchida(_dados_calculadora())
    ws = openpyxl.load_workbook(BytesIO(conteudo))["financeiro"]
    ativas = [r for r in range(2, 62) if ws.cell(row=r, column=1).value is not None]
    assert len(ativas) == 60
    assert all(not ws.row_dimensions[r].hidden for r in range(2, 62))
    assert all(ws.row_dimensions[r].hidden for r in range(62, 74))
    assert not ws.row_dimensions[74].hidden
    assert ws["B74"].value == "TOTAL"
    assert ws["C74"].value == "=SUM(C2:C73)"


# ------------------------------------------------------------- aditivos linha 1

def test_aditivos_linha1_altura_suficiente(wb):
    ws = wb["aditivos"]
    assert (ws.row_dimensions[1].height or 0) >= 90
    a1 = ws["A1"]
    assert a1.alignment.wrap_text is True
    assert str(a1.value).startswith("ITEM")


# ------------------------------------------------------------- cobertura_temporal

def test_cobertura_fontes_padronizadas(wb):
    ws = wb["cobertura_temporal"]
    for endereco in ("B4", "B13", "B15", "C8", "C14", "C17",
                     "A26", "A27", "B26", "B27", "B28", "B29"):
        f = ws[endereco].font
        assert f.name == "Calibri" and f.size == 10, endereco


def test_cobertura_fills_funcionais_preservados(wb):
    ws = wb["cobertura_temporal"]
    esperados = {
        "B13": "FFFEF9C3", "B15": "FFFEF9C3",   # entrada GCC
        "B16": "FFFCE4D6", "B23": "FFFCE4D6",   # projecao
        "A26": "FFFEF9C3", "A27": "FFFEF9C3",   # swatches da legenda
        "A28": "FFEBF3FB", "A29": "FFFCE4D6",
    }
    for endereco, rgb in esperados.items():
        assert ws[endereco].fill.fgColor.rgb == rgb, endereco


def test_cobertura_formatos_residuais_saneados(wb):
    ws = wb["cobertura_temporal"]
    for endereco in ("B2", "B5", "B19", "B20", "B22", "B23",
                     "A26", "A27", "A28", "A29"):
        assert ws[endereco].number_format == "General", endereco


def test_cobertura_legenda_emoldurada_e_formulas_intactas(wb):
    ws = wb["cobertura_temporal"]
    assert "LEGENDA" in str(ws["A25"].value)
    assert ws["A25"].fill.fgColor.rgb == "FF1F4E79"
    assert ws["A25"].border.bottom.style == "thin"
    assert ws["B27"].border.right.style == "thin"
    assert [ws[f"A{r}"].value for r in range(26, 30)] == [
        "FISCAL", "GCC", "AUTOMATICO", "PROJECAO"]
    formulas = sum(1 for row in ws.iter_rows()
                   for cel in row
                   if isinstance(cel.value, str) and cel.value.startswith("="))
    assert formulas == 16
