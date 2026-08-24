# -*- coding: utf-8 -*-
"""Etapa 44.1 — T27 so cobra base fisica de C0 de quem existia em C0.

`MEMORIA_RESULTADOS!T27` conta os itens que DEVERIAM ter base fisica de C0 e
nao a tem; `T25` vira "CALCULO MANUAL REQUERIDO" quando `T27>0` sem PC de C0.

A forma anterior dispensava o novo item SOMANDO um termo dentro do produto —
dispensa inalcancavel, porque `G` fica vazia exatamente quando `Y>0`. Trocar
so `Y` por `AL` nao resolveria e ainda faria a linha de um item nascido no
MEIO de C0 contribuir com -1, mascarando pendencia legitima. A aplicabilidade
temporal passou a ser um FATOR proprio derivado de `AL`
(CICLO_NASCIMENTO_DATA), com todos os fatores em {0, 1}.

Correcao aplicada em RUNTIME; o template binario homologado segue intacto.
"""
from __future__ import annotations

import copy
import io
import os
import sys
from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from _coleta_oficial import (  # noqa: E402
    TEMPLATE_COLETA_OFICIAL,
    _T27_ANTIGA,
    _T27_NOVA,
    _garantir_base_fisica_c0_temporal,
    obter_coleta_oficial_bytes,
)


@pytest.fixture(scope="module")
def wb_runtime():
    return load_workbook(io.BytesIO(obter_coleta_oficial_bytes()), data_only=False)


@pytest.fixture(scope="module")
def wb_template():
    return load_workbook(TEMPLATE_COLETA_OFICIAL, data_only=False)


# --------------------------------------------------------------------------- #
# Estrutura da formula                                                         #
# --------------------------------------------------------------------------- #
def test_t27_usa_a_fonte_temporal(wb_runtime):
    t27 = str(wb_runtime["MEMORIA_RESULTADOS"]["T27"].value)
    assert t27 == _T27_NOVA
    assert "posicao_contratual!$AL$2:$AL$201" in t27
    assert "posicao_contratual!$Y$2:$Y$201" not in t27
    assert t27.count("(") == t27.count(")")


def test_t27_nao_pode_gerar_termo_negativo(wb_runtime):
    """A aplicabilidade e FATOR, nunca parcela somada.

    O defeito da substituicao ingenua era `(ISNUMBER(VU_C0) + dispensa)`
    chegando a 2 dentro do produto, o que faz `1 - 2 = -1` na linha. Aqui o
    corpo do SUMPRODUCT nao pode conter nenhuma soma.
    """
    corpo = str(wb_runtime["MEMORIA_RESULTADOS"]["T27"].value)
    corpo = corpo[len("=SUMPRODUCT("):-1]
    assert "+" not in corpo, f"soma dentro do produto reintroduz termo negativo: {corpo}"
    assert "(1-ISNUMBER(posicao_contratual!$AL$2:$AL$201)" in corpo
    assert "(1-ISNUMBER(posicao_contratual!$G$2:$G$201)" in corpo


def test_t27_usa_teste_de_tipo_e_nao_comparacao_com_texto(wb_runtime):
    """`"">0` e VERDADEIRO no Excel; a aplicabilidade usa ISNUMBER(AL)."""
    t27 = str(wb_runtime["MEMORIA_RESULTADOS"]["T27"].value)
    assert "ISNUMBER(posicao_contratual!$AL$2:$AL$201)" in t27
    assert 'posicao_contratual!$AL$2:$AL$201<>""' not in t27


def test_guarda_fail_closed_estrutura_nao_reconhecida(wb_runtime):
    """Sem AL1 canonico ou com T27 fora do padrao, nada e reescrito."""
    wb = copy.deepcopy(wb_runtime)
    mem, pc = wb["MEMORIA_RESULTADOS"], wb["posicao_contratual"]

    # (a) T27 ja reescrita: idempotente, nao reaplica.
    _garantir_base_fisica_c0_temporal(wb)
    assert mem["T27"].value == _T27_NOVA

    # (b) cabecalho temporal ausente: preserva a formula que estiver la.
    mem["T27"].value = _T27_ANTIGA
    pc["AL1"].value = "OUTRA_COISA"
    _garantir_base_fisica_c0_temporal(wb)
    assert mem["T27"].value == _T27_ANTIGA

    # (c) T27 de outra linhagem: nao inventa formula.
    pc["AL1"].value = "CICLO_NASCIMENTO_DATA"
    mem["T27"].value = "=SUMPRODUCT(1)"
    _garantir_base_fisica_c0_temporal(wb)
    assert mem["T27"].value == "=SUMPRODUCT(1)"


# --------------------------------------------------------------------------- #
# CASO F — nada fora de T27 mudou                                              #
# --------------------------------------------------------------------------- #
def test_caso_f_nada_fora_de_t27_mudou(wb_runtime, wb_template):
    runtime = wb_runtime["MEMORIA_RESULTADOS"]
    template = wb_template["MEMORIA_RESULTADOS"]
    for celula in ("X2", "T20", "T21", "T22", "T23", "T24", "T25", "T26",
                   "W46", "W47", "W48", "W53", "W54", "W55", "W56", "W57", "B26"):
        assert runtime[celula].value == template[celula].value, (
            f"{celula} mudou — fora do escopo da Etapa 44.1"
        )
    # W41:W45 seguem com o criterio temporal entregue pela Etapa 44.
    for celula, limiar in (("W41", 0), ("W42", 1), ("W43", 2), ("W44", 3), ("W45", 4)):
        assert (
            f"posicao_contratual!$AL$2:$AL$201<={limiar}"
            in str(runtime[celula].value)
        )
    # Quantidades intactas.
    rt, tp = wb_runtime["posicao_contratual"], wb_template["posicao_contratual"]
    for coluna in ("E", "G", "I", "K", "M", "O", "Q", "S", "U", "W", "Y", "AL"):
        for linha in (2, 50, 200):
            assert rt[f"{coluna}{linha}"].value == tp[f"{coluna}{linha}"].value


def test_x2_permanece_registrado_e_nao_corrigido(wb_runtime):
    """Achado adjacente: X2 ainda usa Y. Registrado, deliberadamente intacto."""
    assert "posicao_contratual!$Y2" in str(wb_runtime["MEMORIA_RESULTADOS"]["X2"].value)


# --------------------------------------------------------------------------- #
# CASOS A a E — comportamento numerico no motor real da planilha               #
# --------------------------------------------------------------------------- #
def _montar_cenario(destino: Path, com_item_sem_nascimento: bool) -> Path:
    """ITEM-01 em C0; N001 no meio de C0; N002/N003 no meio de C1."""
    destino.write_bytes(obter_coleta_oficial_bytes())
    wb = load_workbook(destino)
    ctl = wb["CONTROLE"]
    ctl["B1"], ctl["B2"], ctl["B3"] = "PCs", "C1", date(2026, 8, 1)
    par = wb["parametros"]
    par["A2"], par["C2"], par["D2"] = "Nao", date(2025, 2, 1), date(2026, 1, 31)
    par["A3"], par["C3"], par["D3"] = "Sim", date(2026, 2, 1), date(2027, 1, 31)
    par["E3"] = 0.05
    # Etapa 48: a ancora temporal dos deltas migrou de parametros!C para
    # parametros!I (DATA_ABERTURA_FISICA_EXATA), com guarda fail-closed 48.3
    # (OR(C="",I="")) que desliga o recorte temporal quando I esta vazia. O
    # runtime SEMPRE grava I para os ciclos da analise; este cenario precisa
    # fazer o mesmo, senao AL degrada para o criterio por quantidade e o
    # aditivo do meio do C0 (N001, 250 x VU 5,00 = 1.250) vira execucao
    # fisica implicita (+750 em vez de -500), inflando W48 para 22.790.
    par["I2"], par["I3"] = date(2025, 2, 1), date(2026, 2, 1)
    itens = [
        ("ITEM-01", 1000, 10.0, 800),   # existia em C0 (AL=0)
        ("N001", None, 5.0, 100),       # nasceu no MEIO de C0 (Y=0, AL=1)
        ("N002", None, 7.0, None),      # nasceu no MEIO de C1 (Y=1, AL=2)
        ("N003", None, 9.0, None),      # idem
    ]
    if com_item_sem_nascimento:
        itens.append(("ITEM-99", 0, 4.0, None))  # sem aditivo => AL vazio
    rem = wb["itens_Remanesc"]
    for i, (item, base, vu, c1) in enumerate(itens, start=2):
        rem[f"A{i}"] = item
        if base is not None:
            rem[f"B{i}"] = base
        rem[f"C{i}"] = vu
        if c1 is not None:
            rem[f"E{i}"] = c1
    adi = wb["aditivos"]
    for i, (item, dia, qtd) in enumerate(
        [("N001", date(2025, 12, 15), 250),
         ("N002", date(2026, 6, 12), 945),
         ("N003", date(2026, 6, 12), 500)], start=2
    ):
        adi[f"A{i}"] = item
        adi[f"B{i}"] = dia
        adi[f"D{i}"] = "ACRESCIMO"
        adi[f"E{i}"] = qtd
        adi[f"H{i}"] = "Nao"
        adi[f"K{i}"] = "SIM"
    wb.save(destino)
    return destino


@pytest.fixture(scope="module")
def excel():
    client = pytest.importorskip("win32com.client")
    pythoncom = pytest.importorskip("pythoncom")
    pythoncom.CoInitialize()
    app = client.DispatchEx("Excel.Application")
    app.Visible = False
    app.DisplayAlerts = False
    yield app
    app.Quit()
    del app
    pythoncom.CoUninitialize()


def _abrir(excel, caminho: Path):
    book = excel.Workbooks.Open(
        str(caminho), UpdateLinks=0, ReadOnly=False, CorruptLoad=0
    )
    excel.CalculateFullRebuild()
    return book


@pytest.mark.skipif(
    os.environ.get("RUN_EXCEL_INTEGRATION") != "1",
    reason="defina RUN_EXCEL_INTEGRATION=1 para executar Excel COM",
)
def test_casos_a_b_c_e_sem_pcs(excel, tmp_path):
    """A/B/C/E — sem PC de C0, os itens temporais deixam de bloquear o VTA."""
    book = _abrir(excel, _montar_cenario(tmp_path / "c44_1.xlsx", False))
    try:
        mem = book.Worksheets("MEMORIA_RESULTADOS")
        pc = book.Worksheets("posicao_contratual")
        assert book.Worksheets("itens_PC").Range("P2").Value in (0, 0.0, None)

        # CASO A — ITEM-01 existia em C0 e tem base completa.
        assert pc.Range("AL2").Value == 0
        # CASO B — N001 (Y=0, AL=1) sai da contagem, sem contribuir com -1.
        assert pc.Range("Y3").Value == 0 and pc.Range("AL3").Value == 1
        # CASO C — N002/N003 (nascidos em C1) tambem saem.
        assert pc.Range("AL4").Value == 2 and pc.Range("AL5").Value == 2
        assert float(mem.Range("T27").Value) == 0.0

        # CASO E — T25 e B26 liberados.
        assert float(mem.Range("T26").Value) == 0.0
        assert mem.Range("T25").Value != "CALCULO MANUAL REQUERIDO"
        assert float(mem.Range("T25").Value) == pytest.approx(21540.0, abs=0.005)
        assert float(mem.Range("B26").Value) == pytest.approx(21540.0, abs=0.005)
        # CASO F numerico — VTA e completude intactos.
        assert float(mem.Range("T23").Value) == pytest.approx(20040.0, abs=0.005)
        assert float(mem.Range("W48").Value) == pytest.approx(21540.0, abs=0.005)
        for celula, esperado in (("W41", 1), ("W42", 1), ("W46", 1)):
            assert float(mem.Range(celula).Value) == esperado
    finally:
        book.Close(SaveChanges=False)


@pytest.mark.skipif(
    os.environ.get("RUN_EXCEL_INTEGRATION") != "1",
    reason="defina RUN_EXCEL_INTEGRATION=1 para executar Excel COM",
)
def test_caso_d_nascimento_indeterminado_continua_fail_closed(excel, tmp_path):
    """D — item sem AL derivavel NAO recebe dispensa automatica."""
    book = _abrir(excel, _montar_cenario(tmp_path / "c44_1d.xlsx", True))
    try:
        mem = book.Worksheets("MEMORIA_RESULTADOS")
        pc = book.Worksheets("posicao_contratual")
        assert pc.Range("A6").Value == "ITEM-99"
        assert pc.Range("AL6").Value in ("", None)
        # Continua contado: a dispensa exige nascimento COMPROVADO.
        assert float(mem.Range("T27").Value) == 1.0
        assert mem.Range("T25").Value == "CALCULO MANUAL REQUERIDO"
        # O VTA auditavel segue calculavel e igual ao cenario anterior.
        assert float(mem.Range("W48").Value) == pytest.approx(21540.0, abs=0.005)
    finally:
        book.Close(SaveChanges=False)
