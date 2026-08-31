"""RESULTADOS-BASELINE (PR 0) — construtor deterministico dos 12 cenarios.

POR QUE ISSO EXISTE
-------------------
A aba RESULTADOS sera refatorada (leiaute + contrato de leitura). Antes de
mexer em uma linha de apresentacao e preciso CONGELAR o comportamento atual:
metodologia, matematica, VTA, retroativo, remanescente, ciclos, percentuais,
fatores, efeito financeiro, status, web e documentos.

Este modulo NAO calcula nada e NAO conhece regra de negocio. Ele apenas monta
as ENTRADAS (bytes de XLS oficiais) dos cenarios. Todo numero do baseline nasce
das cadeias de producao, jamais daqui.

POR QUE NAO REUSAR OS CONSTRUTORES EXISTENTES
---------------------------------------------
Os construtores de cenario que ja existem sao locais e parciais: o `_xls` de
test_cobertura_temporal_runtime cobre so o eixo temporal; o `_dados` de
test_coleta_oficial_runtime_excel exige RUN_EXCEL_INTEGRATION=1; e
`_fabrica_coleta` memoriza os bytes de UMA Coleta, sem variacao de cenario.
Nenhum deles cobre os 12 casos exigidos nem e importavel como fixture comum.
O que da para reaproveitar — o template oficial e a disciplina de cachear so
BYTES — esta reaproveitado aqui.

DETERMINISMO
------------
Nenhuma data "hoje", nenhum aleatorio, nenhum acesso a rede (o conftest ja
desliga a Anatel). Os mesmos bytes saem de qualquer maquina, em qualquer dia —
exceto CONTROLE!B14 (=NOW(), formula sem cache) e o docProps do pacote, que
nunca sao fotografados.

CUSTO
-----
Cada cenario carrega o template oficial (~16 MB de XML). Os bytes de cada
cenario sao gerados UMA vez por sessao e memorizados; quem precisa mutar recebe
uma copia dos bytes, nunca um objeto openpyxl compartilhado.
"""
from __future__ import annotations

import io
from datetime import date
from pathlib import Path
from typing import Any, Callable

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "templates" / "COLETA_REAJUSTE_OFICIAL.xlsx"

# Ordem oficial dos 12 cenarios minimos exigidos pelo PR 0.
ORDEM_CENARIOS = (
    "01_financeiro_normal",
    "02_pc",
    "03_itens_consumidos",
    "04_multiciclo",
    "05_reajuste_negativo_aplicado",
    "06_reajuste_negativo_neutralizado",
    "07_sem_ciclo_em_execucao",
    "08_situacao_atual_posterior_ao_corte",
    "09_referencia_anterior_ao_corte_estimado",
    "10_sem_recalculo_do_excel",
    "11_pcs_sem_efeito_financeiro",
    "12_aditivo_no_meio_do_ciclo",
)

_CACHE: dict[str, bytes] = {}


# --------------------------------------------------------------------------- #
# Blocos de preenchimento — cada um mexe SO na sua aba de entrada.
# --------------------------------------------------------------------------- #
def _parametros(wb, ciclos: list[dict[str, Any]]) -> None:
    """parametros!A:H — a janela de cada ciclo e a decisao de computar."""
    ws = wb["parametros"]
    for indice, ciclo in enumerate(ciclos):
        linha = indice + 2                      # C0 na linha 2, C1 na 3, ...
        ws.cell(linha, 1).value = "Sim" if ciclo.get("computar") else "Nao"
        ws.cell(linha, 3).value = ciclo["inicio"]
        ws.cell(linha, 4).value = ciclo["fim"]
        if ciclo.get("percentual") is not None:
            ws.cell(linha, 5).value = ciclo["percentual"]
        if ciclo.get("situacao"):
            ws.cell(linha, 7).value = ciclo["situacao"]
        if ciclo.get("inicio_efeito"):
            ws.cell(linha, 8).value = ciclo["inicio_efeito"]
        if ciclo.get("data_pedido"):
            ws.cell(linha, 21).value = ciclo["data_pedido"]   # U = DATA_PEDIDO


def _controle(wb, *, metodo: str, ciclo_vigente: str, data_corte: date,
              indice: str, data_base: date) -> None:
    ws = wb["CONTROLE"]
    ws["B1"] = metodo
    ws["B2"] = ciclo_vigente
    ws["B3"] = data_corte
    ws["B7"] = indice
    ws["B8"] = data_base


def _financeiro(wb, competencias: list[tuple[date, float]], *,
                efeito: str = "Sim") -> None:
    """financeiro!A (competencia), C (valor pago), G (efeito financeiro)."""
    ws = wb["financeiro"]
    for indice, (competencia, valor) in enumerate(competencias):
        linha = indice + 2
        ws.cell(linha, 1).value = competencia
        ws.cell(linha, 3).value = valor
        ws.cell(linha, 7).value = efeito


def _itens_remanesc(wb, itens: list[dict[str, Any]]) -> None:
    """itens_Remanesc!A/B/C + quantitativos por ciclo (E/G/I/K/M)."""
    ws = wb["itens_Remanesc"]
    for indice, item in enumerate(itens):
        linha = indice + 2
        ws.cell(linha, 1).value = item["codigo"]
        ws.cell(linha, 2).value = item["quantidade"]
        ws.cell(linha, 3).value = item["valor_unitario"]
        for coluna, quantidade in zip((5, 7, 9, 11, 13), item.get("por_ciclo", ())):
            if quantidade is not None:
                ws.cell(linha, coluna).value = quantidade


def _posicao_referencia(wb, quantidades: dict[str, float]) -> None:
    """posicao_referencia!B — quantitativo restante por item."""
    ws = wb["posicao_referencia"]
    linhas: dict[str, int] = {}
    origem = wb["itens_Remanesc"]
    for linha in range(2, 202):
        codigo = origem.cell(linha, 1).value
        if codigo:
            linhas[str(codigo)] = linha
    for codigo, quantidade in quantidades.items():
        if codigo in linhas:
            ws.cell(linhas[codigo], 2).value = quantidade


def _itens_pc(wb, pedidos: list[dict[str, Any]]) -> None:
    """itens_PC!A (numero), B (DATA_PC), D (valor), G (pago a contratada).

    Somente as quatro colunas de ENTRADA. C (ciclo), E (fator), F, H, I, J, K
    e L (EFEITO_FINANCEIRO_PC) sao formulas estruturais do modelo: sobrescrever
    qualquer uma delas quebra o gate de integridade do template. O efeito
    financeiro de um PC nao se declara — ele decorre de DATA_PC contra o
    INICIO_EFEITO_FINANCEIRO do ciclo (parametros!H).
    """
    ws = wb["itens_PC"]
    for indice, pedido in enumerate(pedidos):
        linha = indice + 2
        ws.cell(linha, 1).value = pedido["numero"]
        ws.cell(linha, 2).value = pedido["data"]
        ws.cell(linha, 4).value = pedido["valor"]
        ws.cell(linha, 7).value = pedido.get("pago", "Nao")


def _itens_consumidos(wb, itens: list[dict[str, Any]]) -> None:
    """itens_Consumidos!A (codigo), B (quantidade), C (valor unitario)."""
    ws = wb["itens_Consumidos"]
    for indice, item in enumerate(itens):
        linha = indice + 2
        ws.cell(linha, 1).value = item["codigo"]
        ws.cell(linha, 2).value = item["quantidade"]
        ws.cell(linha, 3).value = item["valor_unitario"]


def _aditivos(wb, registros: list[dict[str, Any]]) -> None:
    """aditivos!A (identificacao), C (data), F (VU), H/K (marcadores)."""
    ws = wb["aditivos"]
    for indice, registro in enumerate(registros):
        linha = indice + 2
        ws.cell(linha, 1).value = registro["identificacao"]
        ws.cell(linha, 3).value = registro["data"]
        if registro.get("valor_unitario") is not None:
            ws.cell(linha, 6).value = registro["valor_unitario"]
        ws.cell(linha, 8).value = registro.get("computar", "Sim")
        if registro.get("novo_item"):
            ws.cell(linha, 11).value = registro["novo_item"]


def _ciclo_em_execucao(wb, *, data: date, linhas: list[tuple[str, float]]) -> None:
    """Cria a aba opcional CICLO_EM_EXECUCAO (fotografia fisica do fiscal).

    Leiaute lido pelo modelo oficial: A9 marca a posicao, D5 carrega a data e
    F13:F211 os valores por item.

    A aba entra ANTES da RESULTADOS: `_coleta_reajuste` exige que RESULTADOS
    permaneca sendo a ultima aba do arquivo, e `create_sheet` sem `index` a
    empurraria do fim, produzindo um aviso que nao existe no arquivo real.
    """
    ws = wb.create_sheet("CICLO_EM_EXECUCAO", wb.sheetnames.index("RESULTADOS"))
    ws["A9"] = "POSICAO FISICA INFORMADA"
    ws["D5"] = data
    for indice, (codigo, valor) in enumerate(linhas):
        linha = 13 + indice
        ws.cell(linha, 1).value = codigo
        ws.cell(linha, 6).value = valor


def _remover_ciclo_em_execucao(wb) -> None:
    if "CICLO_EM_EXECUCAO" in wb.sheetnames:
        del wb["CICLO_EM_EXECUCAO"]


def _cobertura(wb, *, financeiro_ate: date | None = None,
               pcs_ate: date | None = None) -> None:
    ws = wb["cobertura_temporal"]
    if financeiro_ate is not None:
        ws["B13"] = financeiro_ate
    if pcs_ate is not None:
        ws["B15"] = pcs_ate


# --------------------------------------------------------------------------- #
# Massas comuns — reutilizadas por varios cenarios para que a UNICA diferenca
# entre eles seja a caracteristica que o cenario existe para exercitar.
# --------------------------------------------------------------------------- #
def _competencias(inicio: date, meses: int, valor: float) -> list[tuple[date, float]]:
    saida: list[tuple[date, float]] = []
    ano, mes = inicio.year, inicio.month
    for _ in range(meses):
        saida.append((date(ano, mes, 1), valor))
        mes += 1
        if mes > 12:
            ano, mes = ano + 1, 1
    return saida


ITENS_PADRAO = [
    {"codigo": "ITEM-001", "quantidade": 120.0, "valor_unitario": 250.00,
     "por_ciclo": (120.0, 120.0, 120.0, 120.0, 120.0)},
    {"codigo": "ITEM-002", "quantidade": 80.0, "valor_unitario": 1_500.00,
     "por_ciclo": (80.0, 80.0, 80.0, 80.0, 80.0)},
    {"codigo": "ITEM-003", "quantidade": 40.0, "valor_unitario": 3_200.00,
     "por_ciclo": (40.0, 40.0, 40.0, 40.0, 40.0)},
]

RESTANTE_PADRAO = {"ITEM-001": 45.0, "ITEM-002": 30.0, "ITEM-003": 12.0}

EXECUCAO_FISICA_PADRAO = [
    ("ITEM-001", 18_750.00),
    ("ITEM-002", 75_000.00),
    ("ITEM-003", 89_600.00),
]

# Ciclos C0..C4. TODA janela C:D e preenchida — inclusive a dos ciclos nao
# computados: janela ausente e lida como inconsistencia temporal e bloqueia o
# enquadramento, o que descaracterizaria o cenario. Quem decide o que entra na
# apuracao e a coluna A (COMPUTAR_NESTA_APURACAO), nunca a ausencia de datas.
# Janelas de 12 competencias exatas, conforme a regra petrea da Etapa 31.
CICLOS_BASE = [
    {"inicio": date(2023, 1, 1), "fim": date(2023, 12, 31), "computar": False,
     "percentual": None},
    {"inicio": date(2024, 1, 1), "fim": date(2024, 12, 31), "computar": True,
     "percentual": 0.0512, "situacao": "TEMPESTIVO",
     "inicio_efeito": date(2024, 1, 1), "data_pedido": date(2023, 12, 10)},
    {"inicio": date(2025, 1, 1), "fim": date(2025, 12, 31), "computar": True,
     "percentual": 0.0374, "situacao": "TEMPESTIVO",
     "inicio_efeito": date(2025, 1, 1), "data_pedido": date(2024, 12, 12)},
    {"inicio": date(2026, 1, 1), "fim": date(2026, 12, 31), "computar": True,
     "percentual": 0.0289, "situacao": "TEMPESTIVO",
     "inicio_efeito": date(2026, 1, 1), "data_pedido": date(2025, 12, 15)},
    {"inicio": date(2027, 1, 1), "fim": date(2027, 12, 31), "computar": False,
     "percentual": None},
]


def _ate(numero: int) -> list[dict[str, Any]]:
    """Todas as cinco janelas; computados apenas C1..C`numero`."""
    ciclos = [dict(ciclo) for ciclo in CICLOS_BASE]
    for indice, ciclo in enumerate(ciclos):
        ciclo["computar"] = 1 <= indice <= numero
    return ciclos


def _um_ciclo() -> list[dict[str, Any]]:
    """Cinco janelas preenchidas; apenas C1 computado."""
    return _ate(1)


def _base(wb, *, ciclos, metodo, ciclo_vigente, data_corte,
          indice="IST (Anatel)", data_base=date(2023, 1, 1)) -> None:
    _parametros(wb, ciclos)
    _controle(wb, metodo=metodo, ciclo_vigente=ciclo_vigente,
              data_corte=data_corte, indice=indice, data_base=data_base)
    _itens_remanesc(wb, ITENS_PADRAO)
    _posicao_referencia(wb, RESTANTE_PADRAO)


# --------------------------------------------------------------------------- #
# Os 12 cenarios.
# --------------------------------------------------------------------------- #
def _c01_financeiro_normal(wb) -> None:
    """Financeiro normal: um ciclo tempestivo, competencias pagas, execucao."""
    _base(wb, ciclos=_um_ciclo(), metodo="Financeiro (Mensalidade)",
          ciclo_vigente="C1", data_corte=date(2024, 12, 31))
    _financeiro(wb, _competencias(date(2023, 1, 1), 24, 42_500.00))
    _ciclo_em_execucao(wb, data=date(2024, 12, 31), linhas=EXECUCAO_FISICA_PADRAO)
    _cobertura(wb, financeiro_ate=date(2024, 12, 31))


def _c02_pc(wb) -> None:
    """Metodo PC: pedidos dentro do corte, todos com efeito financeiro."""
    _base(wb, ciclos=_um_ciclo(), metodo="PC (Pedidos de Compra)",
          ciclo_vigente="C1", data_corte=date(2024, 12, 31))
    # Todos posteriores ao INICIO_EFEITO_FINANCEIRO de C1 (01/01/2024): a
    # formula de itens_PC!L os classifica com efeito financeiro.
    _itens_pc(wb, [
        {"numero": "PC-2024-001", "data": date(2024, 2, 15), "valor": 180_000.00,
         "pago": "Sim"},
        {"numero": "PC-2024-002", "data": date(2024, 5, 20), "valor": 96_500.00,
         "pago": "Sim"},
        {"numero": "PC-2024-003", "data": date(2024, 9, 30), "valor": 145_250.00,
         "pago": "Nao"},
    ])
    _ciclo_em_execucao(wb, data=date(2024, 12, 31), linhas=EXECUCAO_FISICA_PADRAO)
    _cobertura(wb, pcs_ate=date(2024, 12, 31))


def _c03_itens_consumidos(wb) -> None:
    """Metodo Itens/Consumido: consumo declarado, sem financeiro nem PC."""
    _base(wb, ciclos=_um_ciclo(), metodo="Itens Consumidos",
          ciclo_vigente="C1", data_corte=date(2024, 12, 31))
    _itens_consumidos(wb, [
        {"codigo": "ITEM-001", "quantidade": 75.0, "valor_unitario": 250.00},
        {"codigo": "ITEM-002", "quantidade": 50.0, "valor_unitario": 1_500.00},
        {"codigo": "ITEM-003", "quantidade": 28.0, "valor_unitario": 3_200.00},
    ])
    _ciclo_em_execucao(wb, data=date(2024, 12, 31), linhas=EXECUCAO_FISICA_PADRAO)


def _c04_multiciclo(wb) -> None:
    """Tres ciclos computados (C1, C2 e C3) com fatores encadeados."""
    _base(wb, ciclos=_ate(3),
          metodo="Financeiro (Mensalidade)", ciclo_vigente="C3",
          data_corte=date(2026, 12, 31))
    _financeiro(wb, _competencias(date(2023, 1, 1), 48, 42_500.00))
    _ciclo_em_execucao(wb, data=date(2026, 12, 31), linhas=EXECUCAO_FISICA_PADRAO)
    _cobertura(wb, financeiro_ate=date(2026, 12, 31))


def _c05_reajuste_negativo_aplicado(wb) -> None:
    """Variacao negativa APLICADA: percentual < 0 desce ao fator do ciclo."""
    ciclos = _um_ciclo()
    ciclos[1]["percentual"] = -0.0218
    ciclos[1]["situacao"] = "TEMPESTIVO — VARIAÇÃO NEGATIVA"
    _base(wb, ciclos=ciclos, metodo="Financeiro (Mensalidade)",
          ciclo_vigente="C1", data_corte=date(2024, 12, 31))
    _financeiro(wb, _competencias(date(2023, 1, 1), 24, 42_500.00))
    _ciclo_em_execucao(wb, data=date(2024, 12, 31), linhas=EXECUCAO_FISICA_PADRAO)


def _c06_reajuste_negativo_neutralizado(wb) -> None:
    """Variacao negativa NEUTRALIZADA: percentual aplicado 0,00%."""
    ciclos = _um_ciclo()
    ciclos[1]["percentual"] = 0.0
    ciclos[1]["situacao"] = "TEMPESTIVO — VARIAÇÃO NEGATIVA NEUTRALIZADA EM 0,00%"
    _base(wb, ciclos=ciclos, metodo="Financeiro (Mensalidade)",
          ciclo_vigente="C1", data_corte=date(2024, 12, 31))
    _financeiro(wb, _competencias(date(2023, 1, 1), 24, 42_500.00))
    _ciclo_em_execucao(wb, data=date(2024, 12, 31), linhas=EXECUCAO_FISICA_PADRAO)


def _c07_sem_ciclo_em_execucao(wb) -> None:
    """Sem CICLO_EM_EXECUCAO: a posicao fisica do fiscal nao foi informada."""
    _base(wb, ciclos=_um_ciclo(), metodo="Financeiro (Mensalidade)",
          ciclo_vigente="C1", data_corte=date(2024, 12, 31))
    _financeiro(wb, _competencias(date(2023, 1, 1), 24, 42_500.00))
    _remover_ciclo_em_execucao(wb)


def _c08_situacao_atual_posterior_ao_corte(wb) -> None:
    """Fotografia do contrato POSTERIOR a data de corte da apuracao."""
    _base(wb, ciclos=_um_ciclo(), metodo="Financeiro (Mensalidade)",
          ciclo_vigente="C1", data_corte=date(2024, 6, 30))
    _financeiro(wb, _competencias(date(2023, 1, 1), 24, 42_500.00))
    # A posicao fisica e de DEPOIS do corte: a apuracao nao pode adota-la.
    _ciclo_em_execucao(wb, data=date(2024, 12, 31), linhas=EXECUCAO_FISICA_PADRAO)


def _c09_referencia_anterior_ao_corte_estimado(wb) -> None:
    """Referencia ANTERIOR ao corte: a posicao precisa ser projetada."""
    _base(wb, ciclos=_um_ciclo(), metodo="Financeiro (Mensalidade)",
          ciclo_vigente="C1", data_corte=date(2024, 12, 31))
    _financeiro(wb, _competencias(date(2023, 1, 1), 18, 42_500.00))
    _ciclo_em_execucao(wb, data=date(2024, 6, 30), linhas=EXECUCAO_FISICA_PADRAO)
    _cobertura(wb, financeiro_ate=date(2024, 6, 30))


def _c10_sem_recalculo_do_excel(wb) -> None:
    """Identico ao cenario 1 nas ENTRADAS.

    Todo XLS montado por openpyxl ja nasce sem cache de formulas — e exatamente
    o arquivo que o fiscal reenvia sem ter aberto no Excel. O cenario existe
    para fixar esse estado como caso proprio do baseline; a fotografia registra
    `cache_ausente` e a indisponibilidade que dele decorre.
    """
    _c01_financeiro_normal(wb)


def _c11_pcs_sem_efeito_financeiro(wb) -> None:
    """Metodo PC com pedidos SEM efeito financeiro (nao geram retroativo).

    O efeito nao e declarado: C1 tem INICIO_EFEITO_FINANCEIRO em 01/07/2024, e
    os dois primeiros PCs sao anteriores a essa data — competencias do ciclo
    que ainda nao carregam o reajuste. O terceiro e posterior a data de corte,
    exercitando tambem o bloco "fora do corte".
    """
    ciclos = _um_ciclo()
    ciclos[1]["inicio_efeito"] = date(2024, 7, 1)
    _base(wb, ciclos=ciclos, metodo="PC (Pedidos de Compra)",
          ciclo_vigente="C1", data_corte=date(2024, 12, 31))
    _itens_pc(wb, [
        {"numero": "PC-2024-001", "data": date(2024, 2, 15), "valor": 180_000.00,
         "pago": "Sim"},
        {"numero": "PC-2024-002", "data": date(2024, 5, 20), "valor": 96_500.00,
         "pago": "Sim"},
        {"numero": "PC-2025-003", "data": date(2025, 3, 10), "valor": 145_250.00,
         "pago": "Nao"},                       # posterior ao corte
    ])
    _ciclo_em_execucao(wb, data=date(2024, 12, 31), linhas=EXECUCAO_FISICA_PADRAO)


def _c12_aditivo_no_meio_do_ciclo(wb) -> None:
    """Aditivo assinado no meio do ciclo vigente (novo item + acrescimo)."""
    _base(wb, ciclos=_um_ciclo(), metodo="Financeiro (Mensalidade)",
          ciclo_vigente="C1", data_corte=date(2024, 12, 31))
    _financeiro(wb, _competencias(date(2023, 1, 1), 24, 42_500.00))
    _aditivos(wb, [
        {"identificacao": "TA-01/2024", "data": date(2024, 6, 15),
         "valor_unitario": 4_100.00, "computar": "Sim", "novo_item": "Sim"},
    ])
    _ciclo_em_execucao(wb, data=date(2024, 12, 31), linhas=EXECUCAO_FISICA_PADRAO)


_CONSTRUTORES: dict[str, Callable[[Any], None]] = {
    "01_financeiro_normal": _c01_financeiro_normal,
    "02_pc": _c02_pc,
    "03_itens_consumidos": _c03_itens_consumidos,
    "04_multiciclo": _c04_multiciclo,
    "05_reajuste_negativo_aplicado": _c05_reajuste_negativo_aplicado,
    "06_reajuste_negativo_neutralizado": _c06_reajuste_negativo_neutralizado,
    "07_sem_ciclo_em_execucao": _c07_sem_ciclo_em_execucao,
    "08_situacao_atual_posterior_ao_corte": _c08_situacao_atual_posterior_ao_corte,
    "09_referencia_anterior_ao_corte_estimado": _c09_referencia_anterior_ao_corte_estimado,
    "10_sem_recalculo_do_excel": _c10_sem_recalculo_do_excel,
    "11_pcs_sem_efeito_financeiro": _c11_pcs_sem_efeito_financeiro,
    "12_aditivo_no_meio_do_ciclo": _c12_aditivo_no_meio_do_ciclo,
}

DESCRICOES = {
    "01_financeiro_normal": "Financeiro normal — um ciclo tempestivo com competencias pagas",
    "02_pc": "Metodo PC — pedidos de compra dentro do corte, com efeito financeiro",
    "03_itens_consumidos": "Metodo Itens/Consumido — consumo declarado por item",
    "04_multiciclo": "Multiciclo — C1, C2 e C3 computados, fatores encadeados",
    "05_reajuste_negativo_aplicado": "Reajuste negativo APLICADO (percentual < 0)",
    "06_reajuste_negativo_neutralizado": "Reajuste negativo NEUTRALIZADO em 0,00%",
    "07_sem_ciclo_em_execucao": "Ausencia de CICLO_EM_EXECUCAO — VTA indisponivel",
    "08_situacao_atual_posterior_ao_corte": "Situacao atual do contrato posterior a data de corte",
    "09_referencia_anterior_ao_corte_estimado": "Referencia anterior ao corte — estado ESTIMADO",
    "10_sem_recalculo_do_excel": "Arquivo sem recalculo do Excel — cache de formulas ausente",
    "11_pcs_sem_efeito_financeiro": "PCs sem efeito financeiro (e PC posterior ao corte)",
    "12_aditivo_no_meio_do_ciclo": "Aditivo assinado no meio do ciclo vigente",
}


def bytes_cenario(nome: str) -> bytes:
    """Bytes do XLS oficial do cenario, gerados uma unica vez por sessao."""
    if nome not in _CONSTRUTORES:
        raise KeyError(f"cenario desconhecido: {nome}")
    if nome not in _CACHE:
        wb = load_workbook(TEMPLATE, data_only=False)
        _CONSTRUTORES[nome](wb)
        buffer = io.BytesIO()
        wb.save(buffer)
        _CACHE[nome] = buffer.getvalue()
    return _CACHE[nome]


def workbook_cenario(nome: str, *, data_only: bool = False):
    """Workbook NOVO e independente, montado a partir dos bytes do cenario."""
    return load_workbook(io.BytesIO(bytes_cenario(nome)), data_only=data_only)
