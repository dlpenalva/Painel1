# -*- coding: utf-8 -*-
"""Fotografia quantitativa de itens_Remanesc na DATA EXATA — fluxo unico.

Regra desta frente:

  A fotografia quantitativa de itens_Remanesc usa a REFERENCIA FISICA EXATA
  do ciclo (dia preservado), a mesma ja gravada em parametros!I. O primeiro
  dia da competencia continua valendo EXCLUSIVAMENTE para a cadeia financeira
  (parametros!H), sem pro-rata.

Cenario canonico:

    data-base 05/02/2025
    pedido C1 15/04/2026            -> fotografia C1 = 15/04/2026
    efeito financeiro 04/2026       -> parametros!H3 = 01/04/2026
    referencia exata C2             = 15/04/2027 (propagacao juridica intacta)

Ate aqui o fluxo SIMPLES nao transportava a referencia exata e o XLS caia no
fallback mensal (01/02/2026): os dois fluxos divergiam para o mesmo caso.
"""
from __future__ import annotations

import io
import sys
from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook

RAIZ = Path(__file__).resolve().parents[1]
if str(RAIZ) not in sys.path:
    sys.path.insert(0, str(RAIZ))

from _coleta_oficial import (  # noqa: E402
    TEMPLATE_COLETA_OFICIAL,
    gerar_coleta_oficial_preenchida,
)
from _reajuste_utils import referencia_exata_pedido_subsequente  # noqa: E402

DATA_BASE = date(2025, 2, 5)
PEDIDO_C1 = date(2026, 4, 15)
APTO_C1 = date(2026, 2, 5)
FOTOGRAFIA_C1 = date(2026, 4, 15)     # exata, com o dia
COMPETENCIA_C1 = date(2026, 4, 1)     # so a cadeia financeira
MENSAL_C1 = date(2026, 2, 1)          # cadeia juridica mensal (parametros!C)


def _d(celula):
    valor = celula.value
    return valor.date() if hasattr(valor, "date") else valor


def _parametros(payload):
    wb = load_workbook(io.BytesIO(gerar_coleta_oficial_preenchida(payload)))
    return wb, wb["parametros"]


def _pagina(arquivo, preparar):
    """Roda a pagina real e devolve o payload da admissibilidade."""
    pytest.importorskip("streamlit")
    from streamlit.testing.v1 import AppTest

    at = AppTest.from_file(str(RAIZ / "pages" / arquivo), default_timeout=600)
    at.run()
    assert not at.exception
    preparar(at)
    for botao in at.button:
        if "Processar Análise" in str(botao.label):
            botao.click()
            break
    at.run()
    assert not at.exception
    try:
        return at.session_state["dados_admissibilidade"]
    except KeyError:  # pragma: no cover - indice sem cobertura no periodo
        pytest.skip("payload indisponivel (indice sem cobertura no periodo)")


@pytest.fixture(scope="module")
def payload_simples():
    def preparar(at):
        for campo in at.date_input:
            rotulo = str(campo.label)
            if rotulo.startswith("Data-base"):
                campo.set_value(DATA_BASE)
            elif rotulo.startswith("Data do Pedido"):
                campo.set_value(PEDIDO_C1)
        at.run()
    return _pagina("01_Calculo_Simples.py", preparar)


@pytest.fixture(scope="module")
def payload_composto():
    def preparar(at):
        for campo in at.date_input:
            if "Data-base de referência" in str(campo.label):
                campo.set_value(DATA_BASE)
                break
        at.run()
        # analise de um unico ciclo: C2 fica fora e nao pode ganhar fotografia
        at.selectbox(key="rep_ciclo_final_analise").select("C1")
        at.run()
        at.date_input(key="p1_" + APTO_C1.strftime("%Y%m%d")).set_value(PEDIDO_C1)
        at.run()
    return _pagina("02_Calculo_Represados.py", preparar)


# --------------------------------------------------------------- unificacao
def test_fluxo_simples_transporta_a_referencia_exata(payload_simples):
    ciclo = payload_simples["ciclos"][0]
    assert ciclo["data_abertura_fisica_exata"] == "15/04/2026"
    # a cadeia financeira segue mensal, sem pro-rata
    assert ciclo["financeiro_inicio"] == "01/04/2026"


def test_fluxo_composto_mantem_a_referencia_exata(payload_composto):
    ciclo = payload_composto["ciclos"][0]
    assert ciclo["data_abertura_fisica_exata"] == "15/04/2026"
    assert ciclo["financeiro_inicio"] == "01/04/2026"


def test_os_dois_fluxos_gravam_a_mesma_fotografia(payload_simples, payload_composto):
    _, simples = _parametros(payload_simples)
    _, composto = _parametros(payload_composto)
    assert _d(simples["I3"]) == FOTOGRAFIA_C1
    assert _d(composto["I3"]) == FOTOGRAFIA_C1
    # antes desta frente o simples gravava a data mensal
    assert _d(simples["I3"]) != MENSAL_C1


def test_cadeias_mensal_e_financeira_nao_se_moveram(payload_simples, payload_composto):
    for payload in (payload_simples, payload_composto):
        _, par = _parametros(payload)
        assert _d(par["C3"]) == MENSAL_C1          # cadeia juridica mensal
        assert _d(par["H3"]) == COMPETENCIA_C1     # competencia do efeito
        assert _d(par["H3"]).day == 1              # mes fechado, sem pro-rata
        assert _d(par["U3"]) == PEDIDO_C1          # data do pedido intacta


# ------------------------------------------------- propagacao juridica intacta
def test_propagacao_juridica_preserva_o_dia_do_pedido():
    assert referencia_exata_pedido_subsequente(PEDIDO_C1) == date(2027, 4, 15)


def test_cadeia_exata_do_ciclo_seguinte_nao_vem_da_competencia(payload_composto):
    assert referencia_exata_pedido_subsequente(PEDIDO_C1) != COMPETENCIA_C1
    assert payload_composto["ciclos"][0]["data_pedido"] == "15/04/2026"


# ------------------------------------------------------ ciclos sem referencia
def test_ciclos_posteriores_ficam_sem_data_projetada(payload_composto):
    _, par = _parametros(payload_composto)
    assert par["I4"].value is None
    assert par["I5"].value is None
    assert par["I6"].value is None


def test_ciclo_posterior_com_efeito_proprio_usa_a_propria_data_exata():
    """C2 apto em abril e pedido em maio: fotografia = o dia do pedido."""
    def ciclo(numero, base, pedido, exata, financeiro):
        return {
            "ciclo": "C%d" % numero, "data_base": base, "data_pedido": pedido,
            "data_abertura_fisica_exata": exata, "financeiro_inicio": financeiro,
            "situacao": "TEMPESTIVO*", "situacao_aplicada": "TEMPESTIVO*",
            "objeto_analise_atual": True, "percentual_aplicado": 0.05,
            "fator": 1.05, "variacao": 0.05, "ciclo_calculado": True,
        }
    payload = {
        "origem": "Reajustes Múltiplos", "tipo": "Represados", "indice": "IST",
        "data_base_original": "05/02/2025", "fator": 1.05, "fator_acumulado": 1.1,
        "variacao": 0.1,
        "ciclos": [
            ciclo(1, "05/02/2025", "15/04/2026", "15/04/2026", "01/04/2026"),
            ciclo(2, "01/04/2026", "10/05/2027", "10/05/2027", "01/05/2027"),
        ],
    }
    _, par = _parametros(payload)
    assert _d(par["I4"]) == date(2027, 5, 10)   # fotografia: dia exato
    assert _d(par["H4"]) == date(2027, 5, 1)    # financeiro: competencia
    assert par["I5"].value is None              # C3 fora da apuracao
    assert par["I6"].value is None


# ------------------------------------------------------------------ cabecalhos
@pytest.fixture(scope="module")
def remanesc():
    return load_workbook(TEMPLATE_COLETA_OFICIAL, data_only=False)["itens_Remanesc"]


def test_coluna_manual_pede_a_quantidade_na_data_exata(remanesc):
    for celula, ciclo, linha in (("E1", "C1", 3), ("G1", "C2", 4),
                                 ("I1", "C3", 5), ("K1", "C4", 6)):
        formula = str(remanesc[celula].value)
        assert "QTD. REMANESCENTE - " + ciclo in formula
        assert "Informe a quantidade existente em " in formula
        assert "parametros!$I$%d" % linha in formula
        assert "Inicio: " not in formula          # rotulo tecnico aposentado
        assert "QTD_REM_BASE" not in formula
        assert "parametros!$H$" not in formula    # H e so da cadeia financeira
        assert "parametros!$C$" not in formula


def test_colunas_automaticas_nao_pedem_preenchimento(remanesc):
    for celula, rotulo in (("F1", "VALOR REMANESCENTE - C1"),
                           ("L1", "VALOR REMANESCENTE - C4"),
                           ("M1", "QTD. EXECUTADA - C1"),
                           ("T1", "VALOR EXECUTADO - C4")):
        formula = str(remanesc[celula].value)
        assert rotulo in formula
        assert "automaticamente" in formula
        assert "Informe" not in formula
        assert "Inicio: " not in formula


def test_cabecalho_sem_data_nao_projeta_nem_orienta(payload_composto):
    """C2 sem referencia materializada: so o rotulo, sem data e sem orientacao."""
    wb = load_workbook(io.BytesIO(gerar_coleta_oficial_preenchida(payload_composto)))
    assert wb["parametros"]["I4"].value is None
    formula = str(wb["itens_Remanesc"]["G1"].value)
    # o ramo verdadeiro do IF e exatamente o rotulo puro
    assert formula.startswith('=IF(parametros!$I$4="","QTD. REMANESCENTE - C2"')


def test_o_bloco_de_cabecalhos_continua_com_16_formulas(remanesc):
    celulas = ["%s1" % c for c in "EFGHIJKLMNOPQRST"]
    assert all(str(remanesc[c].value or "").startswith("=") for c in celulas)


# --------------------------------------------- fronteira de aditivos intocada
def test_fronteira_de_aditivos_continua_ancorada_em_parametros_i():
    pos = load_workbook(TEMPLATE_COLETA_OFICIAL, data_only=False)["posicao_contratual"]
    formula = str(pos["AB2"].value)
    assert "INT(parametros!$I$2)+1" in formula
    assert 'OR(parametros!$C$2="",parametros!$I$2="")' in formula
    assert "parametros!$H$" not in formula
