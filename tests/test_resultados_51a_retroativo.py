# -*- coding: utf-8 -*-
"""Etapa 51A — protecao da cadeia do Retroativo e acabamento da RESULTADOS.

O diagnostico da 51A comprovou que NAO havia defeito funcional no card do
retroativo: RETRO_OFICIAL (MEMORIA_RESULTADOS!B16) -> RESULTADOS!D22 ->
card E5 funcionava; nos arquivos reais a base do metodo simplesmente nao
havia sido informada (categoria F — valor oficial ausente a montante).

HOTFIX RETRO/VTA (pos-51): o VALOR do card migrou de E5 para D5 (mesclada
D5:E5) para ficar exatamente sob o rotulo "RETROATIVO TOTAL A PAGAR"; a
ancora tecnica do CICLO ATUAL migrou de D5 para J8 (coluna J oculta), com o
unico consumidor (C3) repontado.

Este modulo protege esse contrato:
  * cadeia D5 = D22 = RETRO_OFICIAL (formula e VALOR calculado em Excel real);
  * J8 e a ancora do CICLO ATUAL — o retroativo nunca foi ancora;
  * estado vazio do card orienta o usuario sem inventar R$ 0,00;
  * G3:H3 mesclada (fim do ##### na VARIACAO ACUMULADA);
  * formato-lixo '\\Pyyd\\ryy\\o' (Etapa 50, NumberFormatLocal "Padrão")
    erradicado e proibido de voltar;
  * larguras controladas e ShrinkToFit nos valores dos cards;
  * contraste do Quadro 1 (1. COMPOSICAO DO VTA).
"""
from pathlib import Path

import openpyxl
import pytest

ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "templates" / "COLETA_REAJUSTE_OFICIAL.xlsx"

FORMULA_ESTADO_VAZIO_E6 = (
    '=IF($D$22<>"",'
    'IF($J$11="","Ciclos apurados","Ciclos apurados — "&$J$11),'
    'IF($J$6=0,"SELECIONE O MÉTODO NA ABA CONTROLE",'
    '"INFORME A BASE DO MÉTODO"))'
)

# largura minima e maxima aplicadas pelo aplicador 51A; +0.75 de folga para o
# arredondamento interno do Excel ao persistir ColumnWidth.
LARGURAS_51A = {
    "A": (32.27, 36.75),
    "B": (23.27, 26.75),
    "C": (20.0, 22.75),
    "D": (35.27, 38.75),
    "E": (23.27, 26.75),
    "F": (15.27, 18.75),
    "G": (19.27, 22.75),
    "H": (24.0, 28.75),
}

ANCORAS_OCULTAS = (
    "B3", "B5", "J8", "F5", "H5", "B6", "D6", "H8", "H14",
    "A23", "B23", "C23", "D23", "E23", "A39",
)


@pytest.fixture(scope="module")
def resultados():
    wb = openpyxl.load_workbook(TEMPLATE, data_only=False)
    return wb["RESULTADOS"], wb


def test_cadeia_do_retroativo_intacta(resultados):
    """RETRO_OFICIAL -> D22 -> D5 (mesclada D5:E5): o card ESPELHA o oficial."""
    ws, wb = resultados
    assert ws["D5"].value == "=$D$22"
    assert ws["E5"].value is None  # parte da mescla D5:E5
    assert "D5:E5" in {str(m) for m in ws.merged_cells.ranges}
    assert ws["D22"].value == '=IFERROR(RETRO_OFICIAL,"")'
    assert wb.defined_names["RETRO_OFICIAL"].value == "MEMORIA_RESULTADOS!$B$16"
    memoria = wb["MEMORIA_RESULTADOS"]
    assert memoria["B16"].value == '=IF(ISNUMBER($B$5),$B$5,E15)'
    # O rotulo do card fica em D4; o VALOR ocupa toda a base (D5:E5).
    assert ws["D4"].value == "RETROATIVO TOTAL A PAGAR"


def test_j8_e_a_ancora_do_ciclo_atual(resultados):
    """A ancora do CICLO ATUAL vive em J8 (oculta); C3 e o unico consumidor."""
    ws, _ = resultados
    assert ws["J8"].value == "=UPPER(CONTROLE!$B$2)"
    assert ws["J8"].number_format == ";;;"
    assert ws["C3"].value == '=IF($J$8="","—",UPPER($J$8))'
    # D5 deixou de ser ancora: exibe o retroativo com formato de moeda.
    assert ws["D5"].number_format != ";;;"


def test_estado_vazio_do_card_e6(resultados):
    """E6 orienta o usuario sem criar regra de negocio nem R$ 0,00 fantasma."""
    ws, _ = resultados
    assert ws["E6"].value == FORMULA_ESTADO_VAZIO_E6
    assert ws["E6"].number_format == "General"
    # Nao pode haver fallback numerico: ausencia de base nunca vira zero.
    assert "R$" not in ws["E6"].value


def test_variacao_acumulada_mesclada_sem_transbordo(resultados):
    """G3:H3 mesclada: a faixa VARIACAO ACUMULADA nao exibe ##### com dados."""
    ws, _ = resultados
    assert "G3:H3" in {str(m) for m in ws.merged_cells.ranges}
    assert ws["G3"].value == "=$D$6"
    assert ws["G3"].number_format == (
        '"VARIAÇÃO ACUMULADA   "0.00%;;;"VARIAÇÃO ACUMULADA   —"'
    )
    assert ws["H3"].value is None


def test_formato_lixo_da_etapa_50_erradicado(resultados):
    """NumberFormatLocal "Padrão" gravou '\\Pyyd\\ryy\\o'; nunca mais."""
    ws, _ = resultados
    contaminadas = [
        ws.cell(row=r, column=c).coordinate
        for r in range(1, 71)
        for c in range(1, 11)
        if "Pyyd" in ws.cell(row=r, column=c).number_format
    ]
    assert contaminadas == []


def test_ancoras_ocultas_preservam_formato(resultados):
    """O saneamento de formatos jamais pode tocar as ancoras ';;;'."""
    ws, _ = resultados
    violadas = [e for e in ANCORAS_OCULTAS if ws[e].number_format != ";;;"]
    assert violadas == []


def test_larguras_controladas(resultados):
    """min(max(conteudo, minimo), maximo): nada encolhe, nada deforma."""
    ws, _ = resultados
    for coluna, (minimo, maximo) in LARGURAS_51A.items():
        largura = ws.column_dimensions[coluna].width
        assert largura is not None, f"coluna {coluna} sem largura definida"
        assert minimo <= largura <= maximo, f"coluna {coluna}: {largura}"


def test_valores_dos_cards_com_shrink_to_fit(resultados):
    """D5/G5 encolhem; C5 exibe em 15pt reais (coluna C larga o bastante)."""
    ws, _ = resultados
    for endereco in ("D5", "G5"):
        assert ws[endereco].alignment.shrink_to_fit is True, endereco
    assert not ws["C5"].alignment.shrink_to_fit, "C5 nao deve encolher"


def test_contraste_do_quadro_1(resultados):
    """1. COMPOSICAO DO VTA legivel: cabecalho branco, linhas escurecidas."""
    ws, _ = resultados
    for endereco in ("B9", "C9", "F9", "H9"):
        assert ws[endereco].font.color.rgb == "FFFFFFFF", endereco
    for linha in (11, 12, 13):
        assert ws[f"A{linha}"].font.color.rgb == "FF1F4E78", f"A{linha}"
        assert ws[f"B{linha}"].font.color.rgb == "FF404040", f"B{linha}"
    for linha in (10, 11, 12, 13):
        for coluna in ("C", "F", "H"):
            assert ws[f"{coluna}{linha}"].font.color.rgb == "FF595959", (
                f"{coluna}{linha}"
            )
    # O titulo A9 permanece branco sobre azul institucional (Etapa 50).
    assert ws["A9"].font.color.rgb == "FFFFFFFF"
    assert ws["A9"].fill.fgColor.rgb == "FF1F4E78"


# ================================================================ Excel COM
import os
import shutil

com = pytest.mark.skipif(
    os.environ.get("RUN_EXCEL_INTEGRATION") != "1",
    reason="defina RUN_EXCEL_INTEGRATION=1 para executar o Excel COM",
)


@com
def test_com_51a_retroativo_calculado_igual_ao_card(tmp_path):
    """Prova de VALOR (nao de formula): RETRO_OFICIAL == D22 == card D5.

    Cenarios: base ausente (sem metodo e com metodo), Financeiro com base,
    PCs com base. Em todos, G3 nao pode exibir #####.
    """
    import gc
    from datetime import datetime

    import pythoncom
    import win32com.client

    dest = tmp_path / "cenario_51a.xlsx"
    shutil.copyfile(TEMPLATE, dest)
    pythoncom.CoInitialize()
    xl = win32com.client.DispatchEx("Excel.Application")
    xl.Visible = False
    xl.DisplayAlerts = False
    try:
        wb = xl.Workbooks.Open(str(dest.resolve()), UpdateLinks=0, CorruptLoad=0)
        res = wb.Worksheets("RESULTADOS")
        mem = wb.Worksheets("MEMORIA_RESULTADOS")
        ctrl = wb.Worksheets("CONTROLE")
        par = wb.Worksheets("parametros")
        fin = wb.Worksheets("financeiro")
        pc = wb.Worksheets("itens_PC")
        for aba in (res, ctrl, par, fin, pc):
            try:
                aba.Unprotect()
            except Exception:
                pass

        # Cenario de ciclos: C2 vigente e computada (mesma matriz da 26F).
        ctrl.Range("B2").Value = "C2"
        par.Range("A4").Value = "Sim"
        par.Range("E2").Value = 0.0
        par.Range("E3").Value = 0.0
        par.Range("E4").Value = 0.10
        xl.CalculateFull()

        # 1) Sem metodo e sem base: card em travessao e rodape orientando.
        assert res.Range("D5").Text == "—"
        assert mem.Range("B16").Value in (None, "")
        assert res.Range("E6").Text == "SELECIONE O MÉTODO NA ABA CONTROLE"
        assert "#" not in res.Range("G3").Text

        # 2) Metodo selecionado, base ainda ausente.
        ctrl.Range("B1").Value = "Financeiro (Mensalidade)"
        xl.CalculateFull()
        assert res.Range("D5").Text == "—"
        assert res.Range("E6").Text == "INFORME A BASE DO MÉTODO"

        # 3) Financeiro com base: valor pago de c2 com efeito financeiro.
        fin.Range("A14").Value = datetime(2026, 1, 1)
        fin.Range("B14").Value = "c2"
        fin.Range("C14").Value = 100.0
        fin.Range("G14").Value = "Sim"
        xl.CalculateFull()
        retro = float(mem.Range("B16").Value)
        assert retro == pytest.approx(10.0)
        assert float(res.Range("D22").Value) == retro
        assert float(res.Range("D5").Value) == retro
        assert "#" not in res.Range("D5").Text
        assert res.Range("E6").Text == "Ciclos apurados — C2"

        # 4) PCs com base: retroativo reconhecido por PC pago.
        fin.Range("C14").ClearContents()
        pc.Range("A2").Value = "PC-001"
        pc.Range("B2").Value = datetime(2026, 1, 1)
        pc.Range("C2").Value = "C2"       # coluna reescrita pelo runtime
        pc.Range("D2").Value = 5000.0
        pc.Range("G2").Value = "Sim"
        pc.Range("H2").Value = 123.45     # retroativo reconhecido do PC
        ctrl.Range("B1").Value = "PC (Pedidos de Compra)"
        xl.CalculateFull()
        retro_pc = float(mem.Range("B16").Value)
        assert retro_pc == pytest.approx(123.45)
        assert float(res.Range("D22").Value) == retro_pc
        assert float(res.Range("D5").Value) == retro_pc
        assert res.Range("E6").Text == "Ciclos apurados — C2"
        assert "#" not in res.Range("G3").Text

        wb.Close(False)
    finally:
        xl.Quit()
        gc.collect()
        pythoncom.CoUninitialize()
