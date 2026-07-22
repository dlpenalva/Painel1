"""Etapa 3 — Posicao de Referencia do Contrato (revisao).

A posicao de referencia usa a posicao ATUAL quando o fiscal a informa de forma
completa; caso contrario faz FALLBACK automatico para a ultima fotografia
historica valida (abertura do ciclo). Testes estaticos (openpyxl) validam a
estrutura; testes de integracao (RUN_EXCEL_INTEGRATION=1) recalculam no Excel
real e verificam fallback, referencia, ausencia de posicao hibrida, conservacao
e a invariancia de B23/B26.
"""
from __future__ import annotations

import gc
import os
import shutil
from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "templates" / "COLETA_REAJUSTE_OFICIAL.xlsx"
LEGADO = ROOT / "templates" / "Coleta_Reajuste.xlsx"
ABA = "posicao_referencia"

pytest_com = pytest.mark.skipif(
    os.environ.get("RUN_EXCEL_INTEGRATION") != "1",
    reason="defina RUN_EXCEL_INTEGRATION=1 para executar o Excel COM",
)


def _wb():
    return load_workbook(TEMPLATE, data_only=False)


# ---------------------------------------------------------------- estaticos

def test_aba_existe_visivel_e_nome_referencia():
    wb = _wb()
    assert ABA in wb.sheetnames
    assert "remanescente_atual" not in wb.sheetnames  # renomeada
    assert wb[ABA].sheet_state == "visible"


def test_cabecalhos_visiveis():
    ws = _wb()[ABA]
    assert [ws.cell(1, i).value for i in range(1, 7)] == [
        "ITEM", "QTD_REM_ATUAL", "QTD_CONTRATADA_NA_DATA",
        "QTD_REM_REFERENCIA", "ORIGEM / SITUACAO", "CHECK",
    ]


def test_unico_campo_manual():
    ws = _wb()[ABA]
    for col in ("A", "C", "D", "E", "F"):
        assert str(ws[f"{col}2"].value).startswith("=")
    assert ws["B2"].value in (None, "")


def test_qtd_contratada_usa_data_do_aditivo_e_data_referencia():
    f = _wb()[ABA]["C2"].value
    assert "SUMIFS(aditivos!$L$2:$L$200" in f
    assert 'aditivos!$B$2:$B$200,"<="&$I$5' in f  # <= DATA_POSICAO_REFERENCIA


def test_ciclo_reutiliza_calendario_parametros():
    f = _wb()[ABA]["I1"].value  # CICLO da posicao atual
    for i in range(2, 7):
        assert f"parametros!$C${i}" in f and f"parametros!$D${i}" in f
    assert "F4" not in f and "RESULTADOS!" not in f


def test_fallback_nao_usa_max_nem_today():
    """Fallback = ultimo ciclo com fotografia valida; sem TODAY/HOJE/NOW/MAX de valores."""
    ws = _wb()[ABA]
    alvo = " ".join(str(ws[c].value) for c in ("I3", "I4", "I5", "I6"))
    for proib in ("TODAY", "HOJE", "NOW", "AGORA"):
        assert proib not in alvo.upper()
    i3 = ws["I3"].value  # fallback percorre C4->C0 exigindo QTD_REM_AJUSTADA presente
    for col in ("$W$", "$S$", "$O$", "$K$", "$G$"):
        assert f"posicao_contratual!{col}" in i3


def test_sem_today_em_toda_a_aba_e_bloco():
    ws = _wb()[ABA]
    textos = [str(c.value) for row in ws.iter_rows() for c in row if isinstance(c.value, str)]
    r = _wb()["RESULTADOS"]
    textos += [str(r.cell(x, 2).value) for x in range(267, 278)]
    for f in textos:
        for proib in ("TODAY(", "HOJE(", "AGORA("):
            assert proib not in f.upper()


def test_execucao_nao_infere_por_tempo():
    f = _wb()[ABA]["N2"].value
    assert "MAX(" in f
    for proib in ("MONTH", "DATEDIF", "YEARFRAC", "/12"):
        assert proib not in f


def test_colunas_tecnicas_ocultas():
    ws = _wb()[ABA]
    ocultos = {(cd.min, cd.max) for cd in ws.column_dimensions.values() if cd.hidden}
    assert (11, 18) in ocultos  # K:R


def test_controle_b3_desbloqueada_formato_dia():
    wb = _wb()
    assert wb["CONTROLE"]["B3"].protection.locked is False
    assert wb["CONTROLE"].protection.sheet is True
    assert wb["CONTROLE"]["B3"].number_format != "mm/yyyy"


def test_bloco_resultados_referencia_sem_vta():
    ws = _wb()["RESULTADOS"]
    assert "POSICAO DE REFERENCIA DO CONTRATO" in str(ws["A267"].value)
    rot = [str(ws.cell(r, 1).value) for r in range(268, 277)]
    assert any("Origem da posicao de referencia" in x for x in rot)
    assert any("Valor remanescente atualizado na posicao de referencia" in x for x in rot)
    for x in rot:
        assert "VTA" not in x.upper()  # nenhum rotulo do bloco chama o resultado de VTA


def test_b23_b26_inalterados():
    wb = _wb()
    assert wb["RESULTADOS"]["B23"].value == '=IF(OR(B20="",B21="",B22=""),"",ROUND(B20+B21+B22,2))'
    b26 = wb["RESULTADOS"]["B26"].value
    assert "$N$263" in b26 and ABA not in b26
    assert wb["RESULTADOS"]["B25"].value in (None, "")


def test_template_legado_intacto():
    if not LEGADO.exists():
        pytest.skip("legado ausente")
    assert ABA not in load_workbook(LEGADO, data_only=False).sheetnames


def test_aba_registrada_na_geracao_oficial():
    from _coleta_oficial import ABAS_COLETA_OFICIAL
    assert ABA in ABAS_COLETA_OFICIAL
    assert "remanescente_atual" not in ABAS_COLETA_OFICIAL


# ---------------------------------------------------------------- COM

def _recalc(caminho: Path, inspecionar):
    import time
    import pythoncom
    import win32com.client

    RPC = -2147418111

    def tentar(fn, n=6, espera=0.4):
        for i in range(n):
            try:
                return fn()
            except Exception as exc:
                if getattr(exc, "hresult", None) == RPC and i < n - 1:
                    time.sleep(espera * (i + 1))
                    continue
                raise

    pythoncom.CoInitialize()
    excel = win32com.client.DispatchEx("Excel.Application")
    for a, v in (("Visible", False), ("DisplayAlerts", False)):
        try:
            setattr(excel, a, v)
        except AttributeError:
            pass
    wb = None
    try:
        wb = tentar(lambda: excel.Workbooks.Open(str(caminho.resolve()), UpdateLinks=0))
        tentar(excel.CalculateFullRebuild)
        time.sleep(0.3)
        return tentar(lambda: inspecionar(wb))
    finally:
        if wb is not None:
            try:
                wb.Close(SaveChanges=False)
            except Exception:
                pass
        try:
            excel.Quit()
        except Exception:
            pass
        gc.collect()
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass


def _cenario(tmp_path, *, data_pos=None, itens=None, manual=None, aditivos=None,
             ciclos=None, fotos_ate=2):
    """Monta cenario e devolve o caminho.

    itens: lista (item, base, vu). manual: dict item->qty (posicao_referencia!B).
    fotos_ate: preenche QTD_REM_BASE dos ciclos C1..C{fotos_ate}.
    """
    tmp_path.mkdir(parents=True, exist_ok=True)
    dest = tmp_path / "COLETA_REAJUSTE_OFICIAL.xlsx"
    shutil.copy2(TEMPLATE, dest)
    if ciclos is None:
        ciclos = [
            (0, date(2023, 1, 1), date(2023, 12, 31), None),
            (1, date(2024, 1, 1), date(2024, 12, 31), 0.10),
            (2, date(2025, 1, 1), date(2025, 12, 31), 0.10),
        ]
    if itens is None:
        itens = [("ITEM-1", 100.0, 10.0)]
    wb = load_workbook(dest, data_only=False)
    p = wb["parametros"]
    for num, ini, fim, pct in ciclos:
        lin = num + 2
        p.cell(lin, 1).value = "Sim" if num > 0 else "Base"
        p.cell(lin, 3).value = ini
        p.cell(lin, 4).value = fim
        if pct is not None:
            p.cell(lin, 5).value = pct
    ir = wb["itens_Remanesc"]
    foto_col = {1: 5, 2: 7, 3: 9, 4: 11}  # E=C1,G=C2,I=C3,K=C4
    for r, (item, base, vu) in enumerate(itens, start=2):
        ir.cell(r, 1).value = item
        ir.cell(r, 2).value = base
        ir.cell(r, 3).value = vu
        for cic in range(1, fotos_ate + 1):
            ir.cell(r, foto_col[cic]).value = base
    if aditivos:
        ad = wb["aditivos"]
        for idx, (item, dt, tipo, qtd) in enumerate(aditivos, start=2):
            ad.cell(idx, 1).value = item
            ad.cell(idx, 2).value = dt
            ad.cell(idx, 4).value = tipo
            ad.cell(idx, 5).value = qtd
    wb["CONTROLE"]["B3"] = data_pos
    if manual:
        pr = wb[ABA]
        rowmap = {item: r for r, (item, *_ ) in enumerate(itens, start=2)}
        for item, q in manual.items():
            pr.cell(rowmap[item], 2).value = q
    wb.save(dest)
    return dest


def _dmy(v):
    return (v.year, v.month, v.day) if hasattr(v, "year") else v


def _ler(wb, linhas=(2,)):
    pr = wb.Worksheets(ABA)
    r = wb.Worksheets("RESULTADOS")

    def g(ws, a):
        return ws.Range(a).Value

    d = {
        "cic_atual": g(pr, "I1"), "completa": g(pr, "I2"), "fallback": g(pr, "I3"),
        "ciclo_ref": g(pr, "I4"), "data_ref": _dmy(g(pr, "I5")), "origem": g(pr, "I8"),
        "res_data": _dmy(g(r, "B268")), "res_origem": g(r, "B269"), "res_ciclo": g(r, "B270"),
        "res_valor_ref": g(r, "B272"), "res_status": g(r, "B275"), "res_conserv": g(r, "B276"),
        "b23": g(r, "B23"), "b26": g(r, "B26"),
    }
    for ln in linhas:
        d[f"contratada{ln}"] = g(pr, f"C{ln}")
        d[f"qtd_ref{ln}"] = g(pr, f"D{ln}")
        d[f"exec{ln}"] = g(pr, f"N{ln}")
        d[f"origem_item{ln}"] = g(pr, f"E{ln}")
        d[f"check{ln}"] = g(pr, f"F{ln}")
    return d


# 1 + 2 + 11 + 12 + 13 + 10: fallback sem posicao atual -> abertura do ultimo ciclo
@pytest_com
def test_com_fallback_sem_posicao_atual(tmp_path):
    dest = _cenario(tmp_path, data_pos=None, manual=None, fotos_ate=2)
    d = _recalc(dest, _ler)
    assert d["completa"] in (False, 0)
    assert d["ciclo_ref"] == "C2"                 # ultimo com fotografia
    assert d["data_ref"] == (2025, 1, 1)          # abertura C2
    assert d["qtd_ref2"] == 100.0                 # fotografia C2
    assert d["exec2"] == 0.0                       # execucao 0 no fallback
    assert "ABERTURA DO CICLO C2" in str(d["origem"])
    assert "ATUAL INFORMADA" not in str(d["origem"])
    assert d["origem_item2"] == "FOTOGRAFIA C2"


# 3 + 4: ciclo mais recente sem fotografia -> fallback para fotografia anterior
@pytest_com
def test_com_fallback_ciclo_recente_sem_fotografia(tmp_path):
    ciclos = [
        (0, date(2023, 1, 1), date(2023, 12, 31), None),
        (1, date(2024, 1, 1), date(2024, 12, 31), 0.10),
        (2, date(2025, 1, 1), date(2025, 12, 31), 0.10),
        (3, date(2026, 1, 1), date(2026, 12, 31), 0.10),  # existe, sem fotografia
    ]
    dest = _cenario(tmp_path, data_pos=None, ciclos=ciclos, fotos_ate=2)  # so C1,C2
    d = _recalc(dest, _ler)
    assert d["ciclo_ref"] == "C2"        # C3 ignorado (sem fotografia)
    assert d["data_ref"] == (2025, 1, 1)


# 5 + 14: posicao atual completa -> usa atual; manual prevalece
@pytest_com
def test_com_posicao_atual_completa(tmp_path):
    dest = _cenario(tmp_path, data_pos=date(2025, 7, 1), manual={"ITEM-1": 63.0})
    d = _recalc(dest, _ler)
    assert d["completa"] in (True, -1)
    assert d["ciclo_ref"] == "C2"
    assert d["data_ref"] == (2025, 7, 1)
    assert d["qtd_ref2"] == 63.0                  # manual prevalece
    assert d["exec2"] == 37.0                      # 100 - 63
    assert "POSICAO ATUAL INFORMADA" in str(d["origem"])
    assert d["origem_item2"] == "ATUAL INFORMADA"


# 6 + 7 + 9 + 10: posicao atual parcial -> fallback, sem hibrido
@pytest_com
def test_com_posicao_parcial_nao_cria_hibrido(tmp_path):
    itens = [("ITEM-1", 100.0, 10.0), ("ITEM-2", 200.0, 5.0)]
    dest = _cenario(tmp_path, data_pos=date(2025, 7, 1), itens=itens,
                    manual={"ITEM-1": 60.0}, fotos_ate=2)  # ITEM-2 sem manual
    d2 = _recalc(dest, lambda wb: _ler(wb, linhas=(2, 3)))
    assert d2["completa"] in (False, 0)
    assert d2["ciclo_ref"] == "C2"                 # fallback
    # nenhum item usa o manual: ambos usam fotografia C2 (100 e 200)
    assert d2["qtd_ref2"] == 100.0
    assert d2["qtd_ref3"] == 200.0
    assert "INCOMPLETA" in str(d2["origem"])
    assert "ATUAL IGNORADA" in str(d2["origem_item2"])   # manual preservado p/ conferencia


# 8: quantidades sem data -> incompleta -> fallback
@pytest_com
def test_com_quantidade_sem_data(tmp_path):
    dest = _cenario(tmp_path, data_pos=None, manual={"ITEM-1": 50.0})
    d = _recalc(dest, _ler)
    assert d["completa"] in (False, 0)
    assert "INCOMPLETA" in str(d["origem"])
    assert d["ciclo_ref"] == "C2"
    assert d["qtd_ref2"] == 100.0                  # fotografia, nao o manual 50


# 19: aditivo anterior a referencia entra; posterior nao
@pytest_com
def test_com_aditivo_anterior_e_posterior(tmp_path):
    dest = _cenario(
        tmp_path, data_pos=date(2025, 7, 1), manual={"ITEM-1": 63.0},
        aditivos=[("ITEM-1", date(2025, 3, 1), "Acrescimo", 20.0),
                  ("ITEM-1", date(2025, 10, 1), "Acrescimo", 50.0)],
    )
    d = _recalc(dest, _ler)
    assert d["contratada2"] == 120.0   # base 100 + 20 (o de out/2025 fica de fora)


# conservacao quantitativa
@pytest_com
def test_com_conservacao(tmp_path):
    dest = _cenario(tmp_path, data_pos=date(2025, 7, 1), manual={"ITEM-1": 63.0})
    d = _recalc(dest, _ler)
    assert str(d["res_conserv"]).startswith("OK")


# 15 + 16: B23/B26 invariantes entre fallback e posicao atual
@pytest_com
def test_com_b23_b26_invariantes(tmp_path):
    fb = _cenario(tmp_path / "fb", data_pos=None, manual=None)
    d0 = _recalc(fb, _ler)
    at = _cenario(tmp_path / "at", data_pos=date(2025, 7, 1), manual={"ITEM-1": 63.0})
    d1 = _recalc(at, _ler)
    assert d0["b23"] == d1["b23"]
    assert d0["b26"] == d1["b26"]


# 20: resultado da posicao nunca apresentado como VTA
@pytest_com
def test_com_resultado_nao_e_vta(tmp_path):
    dest = _cenario(tmp_path, data_pos=None, manual=None)
    d = _recalc(dest, _ler)
    assert isinstance(d["res_valor_ref"], (int, float))
    assert "REFERENCIA HISTORICA" in str(d["res_status"])
