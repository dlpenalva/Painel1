# -*- coding: utf-8 -*-
"""Regressao focal: referencia exata separada da competencia mensal."""
from __future__ import annotations

import io
import sys
from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook

RAIZ = Path(__file__).resolve().parents[1]
if str(RAIZ) not in sys.path:
    sys.path.insert(0, str(RAIZ))

from _coleta_oficial import gerar_coleta_oficial_preenchida  # noqa: E402
from _reajuste_utils import (  # noqa: E402
    classificar_pedido_por_data_exata,
    referencia_exata_pedido_subsequente,
)


REFERENCIA = date(2025, 8, 23)
LIMITE = date(2025, 11, 21)


@pytest.mark.parametrize(
    "pedido",
    [
        date(2025, 1, 10),
        date(2025, 3, 15),
        date(2025, 5, 12),
        date(2025, 7, 31),
        date(2025, 8, 1),
        date(2025, 8, 22),
    ],
)
def test_todo_pedido_anterior_a_referencia_exata_e_adiantado(pedido):
    assert classificar_pedido_por_data_exata(pedido, REFERENCIA, LIMITE) == "ADIANTADO"


@pytest.mark.parametrize("pedido", [date(2025, 8, 23), date(2025, 8, 30)])
def test_data_igual_ou_posterior_na_janela_nao_e_adiantado(pedido):
    assert classificar_pedido_por_data_exata(pedido, REFERENCIA, LIMITE) == "TEMPESTIVO"


def test_precluso_preserva_janela_existente():
    assert (
        classificar_pedido_por_data_exata(date(2025, 11, 22), REFERENCIA, LIMITE)
        == "PRECLUSO"
    )


@pytest.mark.parametrize(
    ("pedido", "esperado"),
    [
        (date(2024, 8, 23), date(2025, 8, 23)),
        (date(2025, 5, 12), date(2026, 5, 12)),
        (date(2024, 2, 29), date(2025, 2, 28)),
    ],
)
def test_proxima_referencia_preserva_dia_com_convencao_calendaria(pedido, esperado):
    assert referencia_exata_pedido_subsequente(pedido) == esperado


def test_cadeia_reportada_propaga_pedido_tempestivo_e_depois_adiantado():
    """Tempestivo propaga o pedido; adiantado propaga a referencia atual.

    O pedido antecipado e recebido e computado normalmente, mas nao pode
    antecipar o nascimento da anualidade seguinte: o C4 nasce de 23/08/2025
    (referencia do C3), nao de 01/08/2025 (data do pedido antecipado).
    """
    referencia_c2 = referencia_exata_pedido_subsequente(date(2023, 8, 15))
    pedido_c2 = date(2024, 8, 23)
    referencia_c3 = referencia_exata_pedido_subsequente(pedido_c2)
    pedido_c3 = date(2025, 8, 1)
    referencia_c4 = referencia_exata_pedido_subsequente(referencia_c3)

    assert referencia_c2 == date(2024, 8, 15)
    assert classificar_pedido_por_data_exata(pedido_c2, referencia_c2, date(2024, 11, 13)) == "TEMPESTIVO"
    assert referencia_c3 == date(2025, 8, 23)
    assert classificar_pedido_por_data_exata(pedido_c3, referencia_c3, LIMITE) == "ADIANTADO"
    assert referencia_c4 == date(2026, 8, 23)


def test_fluxo_real_multiciclo_exibe_e_propaga_as_referencias_exatas():
    from streamlit.testing.v1 import AppTest

    pagina = RAIZ / "pages" / "02_Calculo_Represados.py"
    at = AppTest.from_file(str(pagina), default_timeout=300)
    at.run()
    assert not at.exception

    at.date_input[2].set_value(date(2023, 8, 15))
    at.selectbox(key="rep_ciclo_inicial_analise").select("C2")
    at.run()
    at.selectbox(key="rep_ciclo_final_analise").select("C4")
    at.run()

    at.date_input(key="p2_20240815").set_value(date(2024, 8, 23))
    at.run()
    at.date_input(key="p3_20250823").set_value(date(2025, 8, 1))
    at.run()

    assert not at.exception
    # O C3 foi pedido de forma antecipada (01/08/2025): o C4 continua nascendo
    # da referencia do C3 (23/08/2025 + 12m), sem antecipacao da anualidade.
    assert at.date_input(key="p4_20260823").value == date(2026, 8, 23)
    resumo = at.dataframe[0].value.to_dict("records")
    assert [linha["Referência exata"] for linha in resumo] == [
        "15/08/2024",
        "23/08/2025",
        "23/08/2026",
    ]
    assert resumo[1]["Situação preliminar"] == "⚠️ ADIANTADO"
    assert [linha["Início financeiro"] for linha in resumo] == [
        "01/08/2024",
        "01/08/2025",
        "01/08/2026",
    ]


@pytest.mark.parametrize(
    ("data_lateral", "primeiro_ciclo"),
    [
        (date(2025, 8, 23), "C1"),
        (date(2024, 8, 15), "C2"),
        (date(2023, 6, 1), "C3"),
    ],
)
def test_bloco_do_primeiro_ciclo_exibe_a_elegibilidade_e_a_data_base(
    data_lateral, primeiro_ciclo
):
    """ETAPA 45 — o destaque do bloco e a ELEGIBILIDADE (lateral + 12 meses).

    Limpeza de leiaute: a data-base fica apenas na lateral; o bloco do ciclo
    exibe somente a elegibilidade (em duas linhas), sem repetir a data-base.
    """
    from dateutil.relativedelta import relativedelta
    from streamlit.testing.v1 import AppTest

    pagina = RAIZ / "pages" / "02_Calculo_Represados.py"
    at = AppTest.from_file(str(pagina), default_timeout=300)
    at.run()
    assert not at.exception

    for campo in at.date_input:
        if "Data-base de referência" in str(campo.label):
            campo.set_value(data_lateral)
            break
    at.run()
    at.selectbox(key="rep_ciclo_inicial_analise").select(primeiro_ciclo)
    at.run()
    assert not at.exception

    blocos = [
        str(elemento.value)
        for elemento in at.markdown
        if "fica apto (elegibilidade)" in str(elemento.value)
    ]
    assert blocos, "o bloco do ciclo nao exibiu a elegibilidade"
    apto = (data_lateral + relativedelta(months=12)).strftime("%d/%m/%Y")
    assert apto in blocos[0]
    assert data_lateral.strftime("%d/%m/%Y") not in blocos[0]

    captions = [str(c.value) for c in at.caption]
    assert not any(
        "Data-base de referência (início do interregno anual)" in c
        for c in captions
    )
    assert "Referência exata para o pedido" not in "".join(
        str(elemento.value) for elemento in at.markdown
    )


def _datas_aptas_exibidas(at):
    marca = "**fica apto (elegibilidade):**"
    return [
        str(elemento.value).split(marca)[1].strip()
        for elemento in at.markdown
        if marca in str(elemento.value)
    ]


def _datas_base_exibidas(at):
    # Limpeza de leiaute: a data-base deixou de ser exibida no bloco do ciclo.
    # A cadeia exata continua verificavel pela elegibilidade visivel, que e
    # SEMPRE a data-base exata + 12 meses-calendario (invariante do motor).
    from datetime import datetime
    from dateutil.relativedelta import relativedelta

    return [
        (
            datetime.strptime(apta, "%d/%m/%Y").date()
            - relativedelta(months=12)
        ).strftime("%d/%m/%Y")
        for apta in _datas_aptas_exibidas(at)
    ]


def _rodar_dois_ciclos(data_lateral, pedido, chave_pedido, com_aptas=False):
    """Roda a pagina com dois ciclos e devolve (datas-base exibidas, resumo).

    Com ``com_aptas``, devolve tambem as elegibilidades exibidas nos blocos.
    """
    from streamlit.testing.v1 import AppTest

    pagina = RAIZ / "pages" / "02_Calculo_Represados.py"
    at = AppTest.from_file(str(pagina), default_timeout=300)
    at.run()
    assert not at.exception

    for campo in at.date_input:
        if "Data-base de referência" in str(campo.label):
            campo.set_value(data_lateral)
            break
    at.run()
    at.selectbox(key="rep_ciclo_final_analise").select("C2")
    at.run()

    at.date_input(key=chave_pedido).set_value(pedido)
    at.run()
    assert not at.exception
    resumo = at.dataframe[0].value.to_dict("records")
    if com_aptas:
        return _datas_base_exibidas(at), resumo, _datas_aptas_exibidas(at)
    return _datas_base_exibidas(at), resumo


@pytest.mark.parametrize(
    ("data_lateral", "pedido", "chave_pedido", "situacao", "inicio_financeiro", "data_base_c2"),
    [
        # Tempestivo: a data exata do pedido vira a data-base do ciclo seguinte,
        # enquanto o efeito financeiro fica na competencia (dia 1).
        (date(2022, 10, 10), date(2023, 10, 20), "p1_20231010",
         "✅ TEMPESTIVO", "01/10/2023", "20/10/2023"),
        (date(2024, 8, 23), date(2025, 8, 30), "p1_20250823",
         "✅ TEMPESTIVO", "01/08/2025", "30/08/2025"),
        # Tempestivo*: idem, sem contaminar o retardo dos efeitos financeiros.
        (date(2024, 8, 23), date(2025, 10, 15), "p1_20250823",
         "✅ TEMPESTIVO*", "01/10/2025", "15/10/2025"),
        # Adiantado: a cadeia exata NAO e antecipada — preserva a referencia.
        (date(2024, 8, 23), date(2025, 8, 1), "p1_20250823",
         "⚠️ ADIANTADO", "01/08/2025", "23/08/2025"),
    ],
)
def test_data_base_do_ciclo_seguinte_vem_da_cadeia_exata(
    data_lateral, pedido, chave_pedido, situacao, inicio_financeiro, data_base_c2
):
    blocos, resumo = _rodar_dois_ciclos(data_lateral, pedido, chave_pedido)

    assert blocos[0] == data_lateral.strftime("%d/%m/%Y")
    assert blocos[1] == data_base_c2
    assert resumo[0]["Situação preliminar"] == situacao
    assert resumo[0]["Início financeiro"] == inicio_financeiro
    # A data-base exata jamais e o primeiro dia da competencia financeira.
    if situacao != "⚠️ ADIANTADO":
        assert blocos[1] != inicio_financeiro


def test_adiantado_nao_antecipa_a_referencia_apta_seguinte():
    """Complementa o TESTE 4: a referencia apta do C2 continua em 23/08/2026."""
    _, resumo = _rodar_dois_ciclos(date(2024, 8, 23), date(2025, 8, 1), "p1_20250823")
    assert resumo[0]["Situação preliminar"] == "⚠️ ADIANTADO"
    assert resumo[1]["Referência exata"] == "23/08/2026"


def _resumo_referencia_23_08_2025(pedido):
    """Roda a pagina real com referencia exata 23/08/2025 (C1) e C2 visivel."""
    from streamlit.testing.v1 import AppTest

    pagina = RAIZ / "pages" / "02_Calculo_Represados.py"
    at = AppTest.from_file(str(pagina), default_timeout=300)
    at.run()
    assert not at.exception

    for campo in at.date_input:
        if "Data-base de referência" in str(campo.label):
            campo.set_value(date(2024, 8, 23))
            break
    at.run()
    at.selectbox(key="rep_ciclo_final_analise").select("C2")
    at.run()

    at.date_input(key="p1_20250823").set_value(pedido)
    at.run()
    assert not at.exception
    return at.dataframe[0].value.to_dict("records")


@pytest.mark.parametrize(
    ("pedido", "situacao", "inicio_financeiro", "proxima_referencia"),
    [
        # Adiantado (mes anterior, mesmo mes e vespera): a antecipacao do
        # pedido nunca antecipa a referencia do ciclo seguinte.
        (date(2025, 7, 15), "⚠️ ADIANTADO", "01/08/2025", "23/08/2026"),
        (date(2025, 8, 1), "⚠️ ADIANTADO", "01/08/2025", "23/08/2026"),
        (date(2025, 8, 22), "⚠️ ADIANTADO", "01/08/2025", "23/08/2026"),
        # Tempestivo: a data efetiva do pedido alimenta o ciclo seguinte.
        (date(2025, 8, 23), "✅ TEMPESTIVO", "01/08/2025", "23/08/2026"),
        (date(2025, 8, 30), "✅ TEMPESTIVO", "01/08/2025", "30/08/2026"),
        # Tempestivo*: idem, preservando o retardo dos efeitos financeiros.
        (date(2025, 9, 1), "✅ TEMPESTIVO*", "01/09/2025", "01/09/2026"),
        (date(2025, 10, 15), "✅ TEMPESTIVO*", "01/10/2025", "15/10/2026"),
    ],
)
def test_referencia_seguinte_so_e_alimentada_por_pedido_tempestivo(
    pedido, situacao, inicio_financeiro, proxima_referencia
):
    resumo = _resumo_referencia_23_08_2025(pedido)
    assert resumo[0]["Referência exata"] == "23/08/2025"
    assert resumo[0]["Situação preliminar"] == situacao
    assert resumo[0]["Início financeiro"] == inicio_financeiro
    assert resumo[1]["Referência exata"] == proxima_referencia


def test_pedido_adiantado_nao_faz_o_ciclo_seguinte_nascer_no_mes_antecipado():
    """Sem sobreposicao: pedido de julho nao puxa o C2 para julho/2026."""
    resumo = _resumo_referencia_23_08_2025(date(2025, 7, 15))
    assert resumo[0]["Situação preliminar"] == "⚠️ ADIANTADO"
    assert resumo[1]["Referência exata"] == "23/08/2026"


# --------------------------------------------------- etapa 45: cenarios A a D
def test_etapa45_a_lateral_05_05_2025_exibe_elegibilidade_05_05_2026():
    """A — a data digitada na lateral produz a elegibilidade 12 meses adiante."""
    bases, _, aptas = _rodar_dois_ciclos(
        date(2025, 5, 5), date(2026, 5, 5), "p1_20260505", com_aptas=True
    )
    assert aptas[0] == "05/05/2026"
    assert bases[0] == "05/05/2025"


def test_etapa45_b_adiantado_nao_antecipa_a_elegibilidade_do_c2():
    """B — pedido em 20/04/2026 e ADIANTADO; C2 segue apto em 05/05/2027."""
    bases, resumo, aptas = _rodar_dois_ciclos(
        date(2025, 5, 5), date(2026, 4, 20), "p1_20260505", com_aptas=True
    )
    assert resumo[0]["Situação preliminar"] == "⚠️ ADIANTADO"
    assert aptas[0] == "05/05/2026"
    # a antecipacao nao move a data-base nem a elegibilidade do ciclo seguinte
    assert bases[1] == "05/05/2026"
    assert aptas[1] == "05/05/2027"
    assert resumo[1]["Referência exata"] == "05/05/2027"
    # o efeito financeiro permanece na competencia da elegibilidade
    assert resumo[0]["Início financeiro"] == "01/05/2026"


def test_etapa45_b2_adiantado_no_mesmo_mes_continua_adiantado():
    """B (variante) — 01/05/2026 antecede 05/05/2026: comparacao por data exata."""
    _, resumo, aptas = _rodar_dois_ciclos(
        date(2025, 5, 5), date(2026, 5, 1), "p1_20260505", com_aptas=True
    )
    assert resumo[0]["Situação preliminar"] == "⚠️ ADIANTADO"
    assert aptas[1] == "05/05/2027"


def test_etapa45_c_tempestivo_asterisco_separa_data_exata_e_competencia():
    """C — pedido 15/06/2026: data exata, competencia 06/2026 e proxima 15/06/2027."""
    bases, resumo, aptas = _rodar_dois_ciclos(
        date(2025, 5, 5), date(2026, 6, 15), "p1_20260505", com_aptas=True
    )
    assert resumo[0]["Situação preliminar"] == "✅ TEMPESTIVO*"
    # data juridica exata preservada como data-base do ciclo seguinte...
    assert bases[1] == "15/06/2026"
    # ...competencia financeira mensalizada em 06/2026...
    assert resumo[0]["Início financeiro"] == "01/06/2026"
    # ...e a proxima elegibilidade nasce 12 meses apos a data exata.
    assert aptas[1] == "15/06/2027"
    assert resumo[1]["Referência exata"] == "15/06/2027"


def test_etapa45_d_precluso_nao_redefine_sozinho_a_proxima_referencia():
    """D — precluso sem acordo preserva a referencia teorica: 05/05/2027."""
    bases, resumo, aptas = _rodar_dois_ciclos(
        date(2025, 5, 5), date(2026, 9, 30), "p1_20260505", com_aptas=True
    )
    assert resumo[0]["Situação preliminar"] == "❌ PRECLUSO"
    assert resumo[0]["Início financeiro"] == "Sem efeitos financeiros automáticos"
    # a data do pedido precluso nao entra na cadeia exata
    assert bases[1] == "05/05/2026"
    assert aptas[1] == "05/05/2027"
    assert resumo[1]["Referência exata"] == "05/05/2027"
    assert not resumo[1]["Referência exata"].startswith("15/07")


# ------------------------------------------- etapa 46: acordo negocial (A a C)
def _rodar_precluso_com_acordo(data_lateral, pedido, chave_pedido, data_pactuada=None):
    """Roda a pagina com C1 PRECLUSO admitido por acordo negocial.

    Quando ``data_pactuada`` e None, o campo "Inicio dos efeitos financeiros por
    acordo" e deixado com o valor que ja vem preenchido (a data do pedido) — o
    caso B do enunciado, em que o usuario nao toca no campo.
    Devolve (datas-base exibidas, resumo, elegibilidades exibidas).
    """
    from streamlit.testing.v1 import AppTest

    pagina = RAIZ / "pages" / "02_Calculo_Represados.py"
    at = AppTest.from_file(str(pagina), default_timeout=300)
    at.run()
    assert not at.exception

    for campo in at.date_input:
        if "Data-base de referência" in str(campo.label):
            campo.set_value(data_lateral)
            break
    at.run()
    at.selectbox(key="rep_ciclo_final_analise").select("C2")
    at.run()

    at.date_input(key=chave_pedido).set_value(pedido)
    at.run()
    at.checkbox(key="superacao_negocial_c1").check()
    at.run()
    if data_pactuada is not None:
        at.date_input(key="inicio_negocial_c1").set_value(data_pactuada)
        at.run()
    assert not at.exception
    resumo = at.dataframe[0].value.to_dict("records")
    return _datas_base_exibidas(at), resumo, _datas_aptas_exibidas(at)


def test_etapa46_a_acordo_preserva_o_dia_pactuado_na_elegibilidade_seguinte():
    """A — pactuado 18/08/2023: C2 apto em 18/08/2024, financeiro em 01/08/2023.

    A mensalizacao pertence exclusivamente a cadeia FINANCEIRA. A cadeia EXATA
    conserva o dia pactuado, de modo que a elegibilidade do ciclo seguinte nao
    retrocede para o primeiro dia da competencia.
    """
    bases, resumo, aptas = _rodar_precluso_com_acordo(
        date(2022, 5, 5), date(2023, 8, 10), "p1_20230505", date(2023, 8, 18)
    )
    assert resumo[0]["Situação preliminar"] == "❌ PRECLUSO"
    # cadeia financeira: mensalizada, exatamente como antes
    assert resumo[0]["Início financeiro"] == "01/08/2023"
    # cadeia exata: dia pactuado preservado no ciclo seguinte
    assert bases[1] == "18/08/2023"
    assert aptas[1] == "18/08/2024"
    assert resumo[1]["Referência exata"] == "18/08/2024"
    # jamais o primeiro dia da competencia financeira
    assert aptas[1] != "01/08/2024"


def test_etapa46_b_campo_nao_alterado_tambem_preserva_o_dia_exato():
    """B — o valor ja preenchido (12/04/2023) vale como data pactuada.

    Nao ter editado o campo manualmente nao elimina o dia exato: o C2 fica apto
    em 12/04/2024, e nao em 01/04/2024.
    """
    bases, resumo, aptas = _rodar_precluso_com_acordo(
        date(2022, 1, 1), date(2023, 4, 12), "p1_20230101"
    )
    assert resumo[0]["Situação preliminar"] == "❌ PRECLUSO"
    assert resumo[0]["Início financeiro"] == "01/04/2023"
    assert bases[1] == "12/04/2023"
    assert aptas[1] == "12/04/2024"
    assert resumo[1]["Referência exata"] == "12/04/2024"
    assert aptas[1] != "01/04/2024"


def test_etapa46_c_precluso_sem_acordo_permanece_identico():
    """C — sem acordo nada muda: cadeia exata preservada e sem efeito financeiro."""
    bases, resumo, aptas = _rodar_dois_ciclos(
        date(2022, 5, 5), date(2023, 8, 10), "p1_20230505", com_aptas=True
    )
    assert resumo[0]["Situação preliminar"] == "❌ PRECLUSO"
    assert resumo[0]["Início financeiro"] == "Sem efeitos financeiros automáticos"
    assert bases[1] == "05/05/2023"
    assert aptas[1] == "05/05/2024"
    assert resumo[1]["Referência exata"] == "05/05/2024"


def test_etapa46_acordo_nao_altera_classificacao_nem_competencia():
    """O acordo move apenas a cadeia exata do ciclo seguinte.

    A classificacao do proprio C1 e todos os demais campos do resumo continuam
    identicos; o unico campo que muda e o inicio financeiro, que passa a existir
    pela admissao — e o faz na COMPETENCIA (dia 1), como antes da etapa 46.
    """
    _, sem_acordo, _ = _rodar_dois_ciclos(
        date(2022, 5, 5), date(2023, 8, 10), "p1_20230505", com_aptas=True
    )
    _, com_acordo, _ = _rodar_precluso_com_acordo(
        date(2022, 5, 5), date(2023, 8, 10), "p1_20230505", date(2023, 8, 18)
    )
    campos_do_c1 = set(sem_acordo[0]) - {"Início financeiro"}
    assert campos_do_c1
    for campo in campos_do_c1:
        assert com_acordo[0][campo] == sem_acordo[0][campo], campo
    # a competencia do C1 nasce da data pactuada e permanece mensalizada
    assert com_acordo[0]["Início financeiro"] == "01/08/2023"


def test_etapa46_percentual_do_acordo_continua_vindo_do_campo_negocial():
    """O percentual aplicado ao ciclo admitido nao foi tocado pela etapa 46.

    A correcao alcanca apenas a cadeia exata; a origem do percentual (o campo
    "Percentual aplicado por acordo") permanece a mesma linha de sempre.
    """
    fonte = (RAIZ / "pages" / "02_Calculo_Represados.py").read_text(encoding="utf-8")
    assert "percentual_aplicado = percentual_negocial" in fonte
    # a mensalizacao financeira do acordo tambem segue intacta
    assert (
        "inicio_efeito_financeiro = data_inicio_efeito_negocial.replace(day=1)" in fonte
    )
    # e a ancora MENSAL do ciclo seguinte continua nascendo da competencia
    assert "data_base_proximo_ciclo = inicio_efeito_financeiro" in fonte


def test_etapa46_mensagem_do_acordo_exibe_a_data_exata():
    """A mensagem visual cita a data pactuada e a elegibilidade exata."""
    from streamlit.testing.v1 import AppTest

    pagina = RAIZ / "pages" / "02_Calculo_Represados.py"
    at = AppTest.from_file(str(pagina), default_timeout=300)
    at.run()
    for campo in at.date_input:
        if "Data-base de referência" in str(campo.label):
            campo.set_value(date(2022, 5, 5))
            break
    at.run()
    at.date_input(key="p1_20230505").set_value(date(2023, 8, 10))
    at.run()
    at.checkbox(key="superacao_negocial_c1").check()
    at.run()
    at.date_input(key="inicio_negocial_c1").set_value(date(2023, 8, 18))
    at.run()
    assert not at.exception

    avisos = [str(e.value) for e in at.info if "admissão negocial" in str(e.value)]
    assert avisos, "a mensagem do acordo nao foi exibida"
    assert "18/08/2023" in avisos[0]
    assert "18/08/2024" in avisos[0]
    assert "01/08/2024" not in avisos[0]
    # a competencia financeira continua informada, sem virar elegibilidade
    assert "08/2023" in avisos[0]


def test_tempestivo_asterisco_preserva_competencias_sem_efeito_no_xls():
    """Pedido 15/10/2025 sobre referencia 23/08/2025: 08 e 09/2025 sem efeito."""
    payload = _payload_xls("✅ TEMPESTIVO*")
    payload["ciclos"][0]["data_pedido"] = "15/10/2025"
    payload["ciclos"][0]["financeiro_inicio"] = "01/10/2025"
    payload["ciclos"][0]["efeito_financeiro_retardado"] = True
    wb = load_workbook(
        io.BytesIO(gerar_coleta_oficial_preenchida(payload)), data_only=False
    )
    try:
        parametros = wb["parametros"]
        assert "TEMPESTIVO*" in str(parametros["G3"].value)
        assert parametros["A3"].value == "Sim"
        assert parametros["E3"].value == pytest.approx(0.05)
        assert str(parametros["F3"].value).startswith("=")
        inicio = parametros["H3"].value
        assert (inicio.date() if hasattr(inicio, "date") else inicio) == date(2025, 10, 1)

        financeiro = wb["financeiro"]
        grade = {}
        for linha in range(2, financeiro.max_row + 1):
            competencia = financeiro[f"A{linha}"].value
            if competencia is not None:
                grade[(competencia.year, competencia.month)] = financeiro[f"G{linha}"].value
        assert grade[(2025, 8)] == "Nao"
        assert grade[(2025, 9)] == "Nao"
        assert grade[(2025, 10)] == "Sim"
        assert grade[(2025, 11)] == "Sim"
    finally:
        wb.close()


def test_as_duas_calculadoras_usam_a_classificacao_unica_de_adiantado():
    simples = (RAIZ / "pages" / "01_Calculo_Simples.py").read_text(encoding="utf-8")
    multiplos = (RAIZ / "pages" / "02_Calculo_Represados.py").read_text(encoding="utf-8")

    for fonte in (simples, multiplos):
        assert "classificar_pedido_por_data_exata(" in fonte
        assert "ADMISSÍVEL - RESSALVA" not in fonte


def _payload_xls(situacao: str) -> dict:
    return {
        "origem": "Reajustes Múltiplos",
        "indice": "IPCA",
        "data_base_original": "23/08/2024",
        "ciclos": [
            {
                "ciclo": "C1",
                "data_base": "23/08/2024",
                "data_pedido": "01/08/2025",
                "percentual_aplicado": 0.05,
                "fator": 1.05,
                "fator_acumulado": 1.05,
                "financeiro_inicio": "01/08/2025",
                "objeto_analise_atual": True,
                "situacao": situacao,
            }
        ],
    }


def _validacoes(ws):
    return tuple(
        (dv.type, str(dv.sqref), dv.formula1, dv.formula2)
        for dv in ws.data_validations.dataValidation
    )


def test_adiantado_permanece_computavel_e_nao_altera_xls_ou_financeiro():
    wb_tempestivo = load_workbook(
        io.BytesIO(gerar_coleta_oficial_preenchida(_payload_xls("✅ TEMPESTIVO"))),
        data_only=False,
    )
    wb_adiantado = load_workbook(
        io.BytesIO(gerar_coleta_oficial_preenchida(_payload_xls("⚠️ ADIANTADO"))),
        data_only=False,
    )
    try:
        assert wb_tempestivo.sheetnames == wb_adiantado.sheetnames
        diferencas = []
        for nome in wb_tempestivo.sheetnames:
            antes = wb_tempestivo[nome]
            depois = wb_adiantado[nome]
            assert antes.max_row == depois.max_row
            assert antes.max_column == depois.max_column
            assert tuple(antes.merged_cells.ranges) == tuple(depois.merged_cells.ranges)
            assert _validacoes(antes) == _validacoes(depois)
            assert antes.protection.sheet == depois.protection.sheet
            for linha in antes.iter_rows():
                for celula_antes in linha:
                    celula_depois = depois[celula_antes.coordinate]
                    assert celula_antes.style_id == celula_depois.style_id
                    if celula_antes.value != celula_depois.value:
                        diferencas.append((nome, celula_antes.coordinate))

        assert diferencas == [("parametros", "G3")]
        parametros = wb_adiantado["parametros"]
        assert parametros["A3"].value == "Sim"
        assert parametros["E3"].value == pytest.approx(0.05)
        assert str(parametros["F3"].value).startswith("=")
        assert "ADIANTADO" in str(parametros["G3"].value)
        assert parametros["B12"].value == "=A3"
        assert parametros["D12"].value == '=IF(C12="","",1+C12)'
        assert parametros["E12"].value == '=IF(B12="Sim",E11*D12,E11)'

        for nome in ("financeiro", "CONTROLE"):
            antes = wb_tempestivo[nome]
            depois = wb_adiantado[nome]
            assert [c.value for row in antes.iter_rows() for c in row] == [
                c.value for row in depois.iter_rows() for c in row
            ]
    finally:
        wb_tempestivo.close()
        wb_adiantado.close()
