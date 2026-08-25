import ast
import html
import io
import re
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

import openpyxl
import pandas as pd
import pytest
from dateutil.relativedelta import relativedelta
from openpyxl.utils.datetime import to_excel
from xlsxwriter.utility import xl_col_to_name

from _reajuste_utils import _competencias_mensais


RAIZ = Path(__file__).resolve().parents[1]
PAGINAS = (
    RAIZ / "pages" / "01_Calculo_Simples.py",
    RAIZ / "pages" / "02_Calculo_Represados.py",
)


def _carregar_gerador(path: Path):
    arvore = ast.parse(path.read_text(encoding="utf-8"), filename=str(path))
    funcoes = {
        no.name: no for no in arvore.body if isinstance(no, ast.FunctionDef)
    }
    necessarias = set()
    pendentes = ["gerar_arquivo_coleta_excel"]
    while pendentes:
        nome = pendentes.pop()
        if nome in necessarias:
            continue
        necessarias.add(nome)
        no = funcoes[nome]
        pendentes.extend(
            chamado.func.id
            for chamado in ast.walk(no)
            if isinstance(chamado, ast.Call)
            and isinstance(chamado.func, ast.Name)
            and chamado.func.id in funcoes
        )
    modulo = ast.Module(body=[funcoes[n] for n in necessarias], type_ignores=[])
    ast.fix_missing_locations(modulo)
    namespace = {
        "BytesIO": io.BytesIO,
        "io": io,
        "ZoneInfo": ZoneInfo,
        "datetime": datetime,
        "pd": pd,
        "re": re,
        "relativedelta": relativedelta,
        "xl_col_to_name": xl_col_to_name,
        "_competencias_mensais": _competencias_mensais,
    }
    exec(compile(modulo, str(path), "exec"), namespace)
    return namespace["gerar_arquivo_coleta_excel"]


@pytest.mark.parametrize("pagina", PAGINAS, ids=lambda p: p.stem)
def test_ciclos_preserva_data_serial_e_exibe_inicio_financeiro_como_competencia(pagina):
    payload = {
        "indice": "IPCA",
        "data_base_original": "17/01/2024",
        "ciclos": [{
            "ciclo": "C1",
            "data_base": "17/01/2025",
            "data_pedido": "22/01/2025",
            "financeiro_inicio": "01/01/2025",
            "financeiro_fim": "31/12/2025",
            "percentual_aplicado": 0.045,
            "fator": 1.045,
            "fator_acumulado": 1.045,
            "situacao": "TEMPESTIVO",
        }],
    }
    arquivo = _carregar_gerador(pagina)(payload)
    wb = openpyxl.load_workbook(io.BytesIO(arquivo), data_only=False)
    ws = wb["CICLOS"]

    assert ws["B2"].value == datetime(2025, 1, 17)
    assert ws["B2"].number_format == "dd/mm/yyyy"
    assert ws["F2"].value == datetime(2025, 1, 1)
    assert to_excel(ws["F2"].value) == 45658
    assert ws["F2"].number_format == "mm/yyyy"


def test_timeline_mensaliza_somente_evento_de_efeito_financeiro():
    pagina = RAIZ / "pages" / "03_Valor_Global.py"
    arvore = ast.parse(pagina.read_text(encoding="utf-8"), filename=str(pagina))
    funcao = next(
        no for no in arvore.body
        if isinstance(no, ast.FunctionDef) and no.name == "render_linha_tempo_contrato"
    )
    modulo = ast.Module(body=[funcao], type_ignores=[])
    ast.fix_missing_locations(modulo)

    class StreamlitFalso:
        @classmethod
        def info(cls, _mensagem):
            raise AssertionError("A timeline de prova não pode estar vazia")

        @classmethod
        def expander(cls, _rotulo):
            class Contexto:
                def __enter__(self):
                    return self

                def __exit__(self, *_args):
                    return False

            return Contexto()

        @classmethod
        def dataframe(cls, *_args, **_kwargs):
            return None

    class ComponentesFalsos:
        conteudo = ""

        @classmethod
        def html(cls, conteudo, **_kwargs):
            cls.conteudo = conteudo

    eventos = pd.DataFrame([
        {"Data": "2025-04-01", "Tipo": "Efeito financeiro", "Evento": "Efeito", "Ciclo": "C1", "Detalhe": "Competência", "Valor": None},
        {"Data": "2025-04-17", "Tipo": "Pedido de reajuste", "Evento": "Pedido", "Ciclo": "C1", "Detalhe": "Data exata", "Valor": None},
        {"Data": "2025-04-22", "Tipo": "Aditivo", "Evento": "Aditivo", "Ciclo": "C1", "Detalhe": "Data exata", "Valor": None},
    ])
    namespace = {
        "pd": pd,
        "html": html,
        "st": StreamlitFalso,
        "components": ComponentesFalsos,
        "montar_eventos_linha_tempo": lambda _resultado: eventos,
        "_limpar_marcadores_timeline": lambda valor: str(valor or ""),
        "_cor_tipo_timeline": lambda _tipo: "#000000",
        "numero_seguro": lambda valor, padrao=0.0: padrao if valor is None else float(valor),
        "moeda": lambda valor: str(valor),
        "formatar_data_br": lambda valor: pd.to_datetime(valor).strftime("%d/%m/%Y"),
    }
    exec(compile(modulo, str(pagina), "exec"), namespace)
    namespace["render_linha_tempo_contrato"]({})

    assert "04/2025" in ComponentesFalsos.conteudo
    assert "17/04/2025" in ComponentesFalsos.conteudo
    assert "22/04/2025" in ComponentesFalsos.conteudo
    assert "01/04/2025" not in ComponentesFalsos.conteudo
