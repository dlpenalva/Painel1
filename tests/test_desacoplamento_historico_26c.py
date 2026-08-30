# -*- coding: utf-8 -*-
"""ETAPA 26C — desacoplamento VALOR HISTORICO x EFEITO FINANCEIRO (itens_PC).

Regra petrea testada aqui:
  - EFEITO_FINANCEIRO_PC="Nao" zera RETROATIVO, nunca a atualizacao historica;
  - E (FATOR_ACUMULADO) = fator historico da memoria parametros!A11:E15,
    independente de L;
  - ciclo fora da apuracao/C0 continuam com fator 1 pela propria memoria.

Cobre: cenario critico C3 pre-efeito (10/04/2025) e pos-efeito (11/04/2025),
C1/C2/C4 (sem hardcode de ciclo), compatibilidade com XLS antigo (cache
L=Nao/E=1/F nominal), template (formula E, N/A em G12, B259, S:T ocultas,
comparativo executivo por referencia) e help removido no Streamlit.

O teste Excel COM (opt-in, RUN_EXCEL_INTEGRATION=1) percorre a cadeia completa
DATA_PC -> CICLO -> EFEITO -> FATOR -> VALOR_ATUALIZADO -> RETROATIVO ->
SUMARIO itens_PC -> RESULTADOS -> VTA no template real.
"""
from __future__ import annotations

import datetime as dt
import os
import re
import shutil
import tempfile
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook

from _leitor_masterfile_v10 import ler_masterfile_v10
from _motor_composicao_vta import montar_composicao_vta

ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "templates" / "COLETA_REAJUSTE_OFICIAL.xlsx"
PAGINA_REPRESADOS = ROOT / "pages" / "02_Calculo_Represados.py"

com = pytest.mark.skipif(
    os.environ.get("RUN_EXCEL_INTEGRATION") != "1",
    reason="defina RUN_EXCEL_INTEGRATION=1 para Excel COM",
)

# Janelas anuais C0..C4 e parametros do cenario critico (secao 13/14 da etapa):
# C1 fora da apuracao; C2/C3/C4 computados; INICIO_EFEITO C3 = 11/04/2025.
JANELAS = {
    "C0": (dt.date(2022, 4, 1), dt.date(2023, 3, 31)),
    "C1": (dt.date(2023, 4, 1), dt.date(2024, 3, 31)),
    "C2": (dt.date(2024, 4, 1), dt.date(2025, 3, 31)),
    "C3": (dt.date(2025, 4, 1), dt.date(2026, 3, 31)),
    "C4": (dt.date(2026, 4, 1), dt.date(2027, 3, 31)),
}
COMPUTAR = {"C0": "Nao", "C1": "Nao", "C2": "Sim", "C3": "Sim", "C4": "Sim"}
PERCENTUAL = {"C2": 0.02, "C3": 0.055, "C4": 0.04}
INICIO_EFEITO = {
    "C2": dt.date(2024, 4, 4),
    "C3": dt.date(2025, 4, 11),
    "C4": dt.date(2026, 4, 4),
}
FATOR_C2 = 1.02
FATOR_C3 = round(1.02 * 1.055, 12)          # 1.0761
FATOR_C4 = round(1.02 * 1.055 * 1.04, 12)   # 1.119144

PCS = [
    # (numero, data, valor, ciclo, efeito, fator, vatual, retroativo)
    ("PC-C1", dt.date(2024, 3, 11), 1000.0, "C1", "Nao", 1.0, 1000.00, 0.0),
    ("PC-C2", dt.date(2024, 5, 3), 1000.0, "C2", "Sim", FATOR_C2, 1020.00, 20.0),
    ("PC-C3-PRE", dt.date(2025, 4, 10), 1000.0, "C3", "Nao", FATOR_C3, 1076.10, 0.0),
    ("PC-C3-POS", dt.date(2025, 4, 11), 1000.0, "C3", "Sim", FATOR_C3, 1076.10, 76.10),
    ("PC-C4", dt.date(2026, 5, 1), 1000.0, "C4", "Sim", FATOR_C4, 1119.14, 119.14),
]


def _preencher_parametros(wb) -> None:
    ctrl = wb["CONTROLE"]
    ctrl["B1"] = "PC (Pedidos de Compra)"
    ctrl["B2"] = "C4"
    par = wb["parametros"]
    for linha, ciclo in enumerate(("C0", "C1", "C2", "C3", "C4"), start=2):
        par[f"A{linha}"] = COMPUTAR[ciclo]
        par[f"C{linha}"] = JANELAS[ciclo][0]
        par[f"D{linha}"] = JANELAS[ciclo][1]
        if ciclo in PERCENTUAL:
            par[f"E{linha}"] = PERCENTUAL[ciclo]
        if ciclo in INICIO_EFEITO:
            par[f"H{linha}"] = INICIO_EFEITO[ciclo]
    wb.properties.keywords = "CL8US_INICIO_EFEITO:" + ",".join(
        f"{c}={d.isoformat()}" for c, d in sorted(INICIO_EFEITO.items())
    )


def _fixture_pcs(compat_cache_antigo: bool = False) -> BytesIO:
    wb = load_workbook(TEMPLATE, data_only=False)
    _preencher_parametros(wb)
    ws = wb["itens_PC"]
    for i, (numero, data, valor, *_resto) in enumerate(PCS, start=2):
        ws[f"A{i}"] = numero
        ws[f"B{i}"] = data
        ws[f"D{i}"] = valor
        ws[f"G{i}"] = "Sim"
        if compat_cache_antigo:
            # XLS antigo em cache: L="Nao" congelado, E=1, F=valor nominal.
            ws[f"E{i}"] = 1.0
            ws[f"F{i}"] = valor
            ws[f"L{i}"] = "Nao"
    saida = BytesIO()
    wb.save(saida)
    saida.seek(0)
    return saida


def _itens_por_numero(res):
    itens = (res.get("itens_pc_v10") or {}).get("itens") or []
    return {str(i.get("numero_pc")): i for i in itens}


# ---- Cenarios criticos (secoes 13/14/15): ciclo, efeito, fator, valor, retro --
def test_cadeia_python_c3_pre_e_pos_efeito_e_demais_ciclos():
    res = ler_masterfile_v10(_fixture_pcs())
    por = _itens_por_numero(res)
    assert len(por) == len(PCS)
    for numero, _data, valor, ciclo, efeito, fator, vatual, retro in PCS:
        item = por[numero]
        assert item["ciclo"] == ciclo, numero
        assert item["efeito_financeiro_pc"] == efeito, numero
        assert item["fator_acumulado"] == pytest.approx(fator, abs=1e-9), numero
        assert item["valor_atualizado"] == pytest.approx(vatual, abs=0.01), numero
        assert item["retroativo_reconhecido_a_pagar"] == pytest.approx(
            retro, abs=0.01
        ), numero
    # A diferenca 10/04 x 11/04 esta APENAS no retroativo, nunca no historico.
    pre, pos = por["PC-C3-PRE"], por["PC-C3-POS"]
    assert pre["valor_atualizado"] == pos["valor_atualizado"]
    assert pre["fator_acumulado"] == pos["fator_acumulado"]
    assert pre["retroativo_reconhecido_a_pagar"] == 0.0
    assert pos["retroativo_reconhecido_a_pagar"] == pytest.approx(76.10, abs=0.01)


def test_vta_inclui_historico_do_c3_pre_efeito():
    wb = load_workbook(TEMPLATE, data_only=False)
    _preencher_parametros(wb)
    ws = wb["itens_PC"]
    for i, (numero, data, valor, *_r) in enumerate(PCS, start=2):
        ws[f"A{i}"], ws[f"B{i}"], ws[f"D{i}"], ws[f"G{i}"] = numero, data, valor, "Sim"
    ctrl = wb["CONTROLE"]
    ctrl["B1"] = "PC (Pedidos de Compra)"
    ctrl["B2"] = "C4"
    posc = wb["posicao_contratual"]
    posc["A2"], posc["B2"], posc["C2"] = 1, 100, 10
    posc["G2"], posc["K2"], posc["O2"], posc["S2"], posc["W2"] = 10, 6, 5, 4, 4
    hv = wb["historico_VU"]
    hv["A2"] = 1
    hv["C2"], hv["D2"], hv["E2"], hv["F2"], hv["G2"] = 100, 100, 100, 100, 110
    saida = BytesIO()
    wb.save(saida)
    saida.seek(0)

    res = ler_masterfile_v10(saida)
    comp = montar_composicao_vta(res)
    por_ciclo = {l["ciclo"]: l for l in comp["execucao_por_ciclo"]}
    # A execucao do VTA usa o VALOR HISTORICO CONSIDERADO, nao o
    # VALOR_ATUALIZADO: o PC anterior ao inicio do efeito financeiro pertence a
    # C3 e permanece no valor original (fator efetivo 1, retroativo zero).
    # C3 = PRE 1000,00 + POS 1076,10 = 2076,10.
    assert por_ciclo["C3"]["valor_atualizado"] == pytest.approx(2076.10, abs=0.01)
    # C1 fora da apuracao permanece nominal (fator 1 pela memoria).
    assert por_ciclo["C1"]["valor_atualizado"] == pytest.approx(1000.00, abs=0.01)
    # VTA = C0 fisico 400 + PCs C1..C3 (1000+1020+2076,10) + rem 440.
    assert comp["vta_composicao"] == pytest.approx(4936.10, abs=0.01)


def test_compatibilidade_xls_antigo_recomputa_e_avisa():
    res = ler_masterfile_v10(_fixture_pcs(compat_cache_antigo=True))
    por = _itens_por_numero(res)
    pre = por["PC-C3-PRE"]
    # Cache antigo trazia F=1000 (nominal); runtime recomputa o historico.
    assert pre["valor_atualizado"] == pytest.approx(1076.10, abs=0.01)
    assert pre["fator_acumulado"] == pytest.approx(FATOR_C3, abs=1e-9)
    alertas = (res.get("itens_pc_v10") or {}).get("alertas") or []
    assert any(
        "PC-C3-PRE" in a and "recomputado" in a for a in alertas
    ), alertas
    # Efeito continua com semantica propria (Nao), sem virar fator.
    assert pre["efeito_financeiro_pc"] == "Nao"


# ---- Template: formula E desacoplada, N/A, B259, comparativo, cor ------------
def test_template_formula_e_desacoplada_de_l():
    ws = load_workbook(TEMPLATE, data_only=False)["itens_PC"]
    for r in (2, 50, 100):
        formula = str(ws[f"E{r}"].value)
        assert f"VLOOKUP(C{r},parametros!$A$11:$E$15,5,0)" in formula, r
        assert f"L{r}" not in formula, r
    # H/J continuam governados por L (retroativo), F continua D x E.
    assert 'L2="Sim",ROUND(F2-D2,2),0' in ws["H2"].value
    assert 'L2="Sim",ROUND(F2-D2,2),0' in ws["J2"].value
    assert "ROUND(D2*E2,2)" in ws["F2"].value
    # Cor vermelha continua na regra canonica do efeito, nao no ciclo.
    formulas_cf = [
        regra.formula[0]
        for _rng, regras in ws.conditional_formatting._cf_rules.items()
        for regra in regras if regra.formula
    ]
    assert '$L2="Nao"' in formulas_cf


def test_template_g12_aceita_na_e_g13_g15_continuam_sim_nao():
    wb = load_workbook(TEMPLATE, data_only=False)
    par = wb["parametros"]
    assert par["T2"].value == "Sim"
    assert par["T3"].value == "Nao"
    assert par["T4"].value == "N/A"
    assert "OPCOES_SIM_NAO_NA" in wb.defined_names
    dvs = {str(dv.sqref): dv.formula1 for dv in par.data_validations.dataValidation}
    assert dvs.get("G12") == "OPCOES_SIM_NAO_NA"
    assert dvs.get("G13:G15") == "OPCOES_SIM_NAO"


def test_template_b259_na_tem_semantica_propria():
    r = load_workbook(TEMPLATE, data_only=False)["MEMORIA_RESULTADOS"]
    b259 = str(r["B259"].value)
    assert 'parametros!$G$12="N/A"' in b259
    assert "NAO SE APLICA" in b259
    assert "FORA - NAO COMPROVADO" in b259  # vazio segue nao comprovado
    # N/A NAO e evidencia de formalizacao: I259 continua testando so "Sim".
    i259 = str(r["I259"].value)
    assert 'parametros!$G$12="Sim"' in i259
    assert "N/A" not in i259


# --- comparativo executivo de referencia (B28/B29) -------------------------
#
# Invariante que a 26C criou em MEMORIA_RESULTADOS!A28:C29: a comparacao
# executiva e feita POR REFERENCIA, "sem matematica nova" — B28 apenas aponta
# para valores que a propria aba ja consolidou e B29 aponta para o comparativo
# do contrato integralmente reajustado.
#
# A formula literal de B28 evoluiu depois da 26C: VTA-M2 e VTA-C2 deram a
# Financeiro e a Itens um ramo proprio dentro do VTA FINAL (B26), e a
# decomposicao generica B20+B21+B22 (B23) passou a ser SO comparativa para
# esses dois metodos. O literal exato ja e travado pelas frentes que o
# definiram (tests/test_vta_m2_financeiro_condicional.py e
# tests/test_vta_c2_consumido_canonico.py); aqui protege-se a SEMANTICA que a
# 26C introduziu, que nenhuma dessas evolucoes pode quebrar.

# Metodo oficial em MEMORIA_RESULTADOS!B4 -> celula canonica que B28 exibe.
# Financeiro/Itens tem ramo proprio em B26, entao o comparativo mostra a
# decomposicao generica B23; PC (e metodo ainda nao selecionado) continuam
# mostrando o proprio VTA FINAL, exatamente como na 26C.
METODO_FONTE_COMPARATIVA = {
    "Financeiro": "$B$23",
    "Itens": "$B$23",
    "PCs": "$B$26",
    "SELECIONE O METODO EM CONTROLE!B1": "$B$26",
}
CELULAS_PERMITIDAS_B28 = {"$B$4", "$B$23", "$B$26"}
SIMBOLOS_MATEMATICOS = set("+-*/^%&")
FUNCOES_PROIBIDAS_B28 = (
    "SUMPRODUCT", "SUMIFS", "SUMIF", "SUM(", "ROUND", "COUNT", "AVERAGE",
)


def _dividir_argumentos(texto: str) -> list[str]:
    """Divide os argumentos de uma funcao Excel no nivel superior."""
    partes: list[str] = []
    atual: list[str] = []
    nivel = 0
    aspas = False
    for ch in texto:
        if ch == '"':
            aspas = not aspas
        elif not aspas:
            if ch == "(":
                nivel += 1
            elif ch == ")":
                nivel -= 1
            elif ch == "," and nivel == 0:
                partes.append("".join(atual))
                atual = []
                continue
        atual.append(ch)
    partes.append("".join(atual))
    return [p.strip() for p in partes]


def _sem_literais(formula: str) -> str:
    """Remove o conteudo das strings para nao confundir texto com referencia."""
    return re.sub(r'"[^"]*"', '""', formula)


def _celulas_referenciadas(formula: str) -> set[str]:
    return set(re.findall(r"\$?[A-Z]{1,3}\$?\d+", _sem_literais(formula)))


def _condicao_b28(expr: str, metodo: str) -> bool:
    """Avalia a condicao de B28 para um metodo, so com a gramatica permitida."""
    expr = expr.strip()
    caixa_alta = expr.upper()
    for prefixo, agregador in (("OR(", any), ("AND(", all)):
        if caixa_alta.startswith(prefixo) and expr.endswith(")"):
            interno = expr[len(prefixo):-1]
            return agregador(
                _condicao_b28(arg, metodo)
                for arg in _dividir_argumentos(interno)
            )
    if caixa_alta.startswith("NOT(") and expr.endswith(")"):
        return not _condicao_b28(expr[4:-1], metodo)
    for operador in ("<>", "="):
        if operador in expr:
            esquerda, direita = (p.strip() for p in expr.split(operador, 1))
            assert esquerda == "$B$4", (
                "B28 so pode decidir pelo metodo oficial $B$4; "
                f"achado: {esquerda!r}"
            )
            assert direita.startswith('"') and direita.endswith('"'), (
                f"B28 so compara $B$4 com literal de metodo; achado: {direita!r}"
            )
            literal = direita[1:-1]
            return metodo == literal if operador == "=" else metodo != literal
    raise AssertionError(f"condicao nao reconhecida em B28: {expr!r}")


def _referencia_exibida_por_b28(expr: str, metodo: str) -> str:
    """Resolve a celula que B28 exibe quando MEMORIA_RESULTADOS!B4 = metodo."""
    expr = expr.strip()
    if expr.upper().startswith("IF(") and expr.endswith(")"):
        argumentos = _dividir_argumentos(expr[3:-1])
        assert len(argumentos) == 3, f"IF malformado em B28: {expr!r}"
        condicao, ramo_sim, ramo_nao = argumentos
        escolhido = ramo_sim if _condicao_b28(condicao, metodo) else ramo_nao
        return _referencia_exibida_por_b28(escolhido, metodo)
    return expr


def _validar_b28(formula: object) -> None:
    """Contrato de B28: referencia pura, method-aware, sem matematica nova."""
    assert isinstance(formula, str) and formula.startswith("="), (
        f"B28 precisa continuar sendo formula; achado: {formula!r}"
    )
    corpo = formula[1:]
    # (a) le somente celulas canonicas desta aba: nada de outra aba, nada do
    #     bloco historico que a 26C desacoplou.
    assert "!" not in _sem_literais(corpo), (
        f"B28 nao pode referenciar outra aba: {formula!r}"
    )
    assert _celulas_referenciadas(corpo) == CELULAS_PERMITIDAS_B28, (
        "B28 deve ler exatamente o metodo oficial e as duas fontes canonicas "
        f"{sorted(CELULAS_PERMITIDAS_B28)}; achado: {formula!r}"
    )
    # (b) "sem matematica nova": B28 seleciona, nunca recalcula.
    assert not SIMBOLOS_MATEMATICOS & set(_sem_literais(corpo)), (
        f"B28 nao pode conter aritmetica: {formula!r}"
    )
    for proibida in FUNCOES_PROIBIDAS_B28:
        assert proibida not in corpo.upper(), (
            f"B28 nao pode agregar/recalcular ({proibida}): {formula!r}"
        )
    # (c) semantica por metodo: cada metodo exibe a sua fonte comparativa.
    for metodo, esperada in METODO_FONTE_COMPARATIVA.items():
        obtida = _referencia_exibida_por_b28(corpo, metodo)
        assert obtida == esperada, (
            f"B28 com metodo {metodo!r} deveria exibir {esperada}, "
            f"exibe {obtida}: {formula!r}"
        )


def test_template_comparativo_executivo_por_referencia():
    r = load_workbook(TEMPLATE, data_only=False)["MEMORIA_RESULTADOS"]
    _validar_b28(r["B28"].value)
    assert r["B29"].value == "=comparativo_VTA!$B$208"
    assert "COMPARACAO DE REFERENCIA" in str(r["A28"].value)
    assert "integralmente reajustado" in str(r["A29"].value)
    assert "comparativo" in str(r["C28"].value)
    # Sem matematica duplicada: as celulas sao referencias puras.
    assert "SUMPRODUCT" not in str(r["B28"].value) + str(r["B29"].value)


# Formulas que NAO podem passar no contrato acima. Existem para provar que a
# validacao semantica de B28 continua detectando regressao real, em vez de
# virar um assert que aceita qualquer coisa.
B28_REGRESSOES = (
    # perde o method-aware: Financeiro/Itens voltariam a exibir o proprio VTA
    "=$B$26",
    # inverte os metodos: PC passaria a exibir a decomposicao generica
    '=IF($B$4="PCs",$B$23,$B$26)',
    # perde a fonte canonica do VTA FINAL
    '=IF(OR($B$4="Financeiro",$B$4="Itens"),$B$23,$B$20)',
    # reintroduz matematica na celula meramente comparativa
    '=IF(OR($B$4="Financeiro",$B$4="Itens"),$B$23,$B$20+$B$21+$B$22)',
    # passa a depender de outra aba em vez da fonte consolidada local
    '=IF(OR($B$4="Financeiro",$B$4="Itens"),$B$23,comparativo_VTA!$B$208)',
    # decide por celula que nao e o metodo oficial
    '=IF(OR($C$4="Financeiro",$C$4="Itens"),$B$23,$B$26)',
    # deixa de ser formula
    0.0,
)


@pytest.mark.parametrize("formula", B28_REGRESSOES)
def test_contrato_de_b28_rejeita_regressao(formula):
    with pytest.raises(AssertionError):
        _validar_b28(formula)


def test_contrato_de_b28_independe_da_escrita_da_formula():
    """A validacao e semantica: reescrever preservando o contrato passa."""
    _validar_b28('=IF(OR($B$4="Itens",$B$4="Financeiro"),$B$23,$B$26)')
    _validar_b28('=IF($B$4="Financeiro",$B$23,IF($B$4="Itens",$B$23,$B$26))')


def test_help_removido_somente_do_marco_temporal():
    fonte = PAGINA_REPRESADOS.read_text(encoding="utf-8")
    assert "Marco temporal do último ciclo concedido/formalizado" in fonte
    assert "Este marco será usado para calcular a admissibilidade" not in fonte
    # Outros helps preservados (ex.: seletor do primeiro ciclo da analise).
    assert "Primeiro ciclo real objeto desta apuração" in fonte


# ---- Excel COM: cadeia completa no template real (secao 26) ------------------
COR_VERMELHO_BGR = 204 * 65536 + 204 * 256 + 244  # RGB(244,204,204)


@com
def test_com_cadeia_completa_c3_pre_e_pos_efeito():
    import pythoncom
    import win32com.client

    tmp_dir = Path(tempfile.mkdtemp(prefix="cl8us_26c_com_"))
    tmp = tmp_dir / "cadeia.xlsx"
    shutil.copyfile(TEMPLATE, tmp)
    pythoncom.CoInitialize()
    xl = win32com.client.DispatchEx("Excel.Application")
    xl.Visible = False
    xl.DisplayAlerts = False
    try:
        wb = xl.Workbooks.Open(str(tmp), UpdateLinks=0)
        # Abas de entrada podem estar protegidas (sem senha) no template.
        for nome in ("CONTROLE", "parametros", "itens_PC",
                     "posicao_contratual", "historico_VU"):
            try:
                wb.Worksheets(nome).Unprotect()
            except Exception:
                pass
        par = wb.Worksheets("parametros")
        for linha, ciclo in enumerate(("C0", "C1", "C2", "C3", "C4"), start=2):
            par.Cells(linha, 1).Value = COMPUTAR[ciclo]
            par.Cells(linha, 3).Value = dt.datetime.combine(
                JANELAS[ciclo][0], dt.time())
            par.Cells(linha, 4).Value = dt.datetime.combine(
                JANELAS[ciclo][1], dt.time())
            if ciclo in PERCENTUAL:
                par.Cells(linha, 5).Value = PERCENTUAL[ciclo]
            if ciclo in INICIO_EFEITO:
                par.Cells(linha, 8).Value = dt.datetime.combine(
                    INICIO_EFEITO[ciclo], dt.time())
        ipc = wb.Worksheets("itens_PC")
        for i, (numero, data, valor, *_r) in enumerate(PCS, start=2):
            ipc.Cells(i, 1).Value = numero
            ipc.Cells(i, 2).Value = dt.datetime.combine(data, dt.time())
            ipc.Cells(i, 4).Value = valor
            ipc.Cells(i, 7).Value = "Sim"
        ctrl = wb.Worksheets("CONTROLE")
        ctrl.Range("B1").Value = "PC (Pedidos de Compra)"
        ctrl.Range("B2").Value = "C4"
        posc = wb.Worksheets("posicao_contratual")
        posc.Range("A2").Value = 1
        posc.Range("B2").Value = 100
        posc.Range("C2").Value = 10
        for col, qtd in (("G", 10), ("K", 6), ("O", 5), ("S", 4), ("W", 4)):
            posc.Range(f"{col}2").Value = qtd
        hv = wb.Worksheets("historico_VU")
        hv.Range("A2").Value = 1
        for col, vu in (("C", 100), ("D", 100), ("E", 100), ("F", 100), ("G", 110)):
            hv.Range(f"{col}2").Value = vu
        par.Range("G12").Value = "N/A"

        xl.CalculateFullRebuild()

        # Cadeia por linha: CICLO -> EFEITO -> FATOR -> VALOR -> RETROATIVO.
        for i, (numero, _d, _v, ciclo, efeito, fator, vatual, retro) in enumerate(
            PCS, start=2
        ):
            assert ipc.Cells(i, 3).Value == ciclo, numero
            assert ipc.Cells(i, 12).Value == efeito, numero
            assert float(ipc.Cells(i, 5).Value) == pytest.approx(
                fator, abs=1e-9), numero
            assert float(ipc.Cells(i, 6).Value) == pytest.approx(
                vatual, abs=0.01), numero
            assert float(ipc.Cells(i, 8).Value) == pytest.approx(
                retro, abs=0.01), numero
        # Cor: pre-efeito (efeito Nao) vermelho; pos-efeito nao.
        cor_pre = int(ipc.Range("A4").DisplayFormat.Interior.Color)
        cor_pos = int(ipc.Range("A5").DisplayFormat.Interior.Color)
        assert cor_pre == COR_VERMELHO_BGR
        assert cor_pos != COR_VERMELHO_BGR
        # Sumario por ciclo (P) e VTA preservado na memoria tecnica.
        assert float(ipc.Range("P5").Value) == pytest.approx(2152.20, abs=0.01)
        res = wb.Worksheets("MEMORIA_RESULTADOS")
        assert float(res.Range("T22").Value) == pytest.approx(4172.20, abs=0.01)
        assert float(res.Range("T25").Value) == pytest.approx(5012.20, abs=0.01)
        assert float(res.Range("B26").Value) == pytest.approx(5012.20, abs=0.01)
        assert float(res.Range("B28").Value) == pytest.approx(5012.20, abs=0.01)
        # N/A em G12 com semantica propria no Eixo 5 (C1 fora da apuracao).
        assert str(res.Range("B259").Value) == "NAO SE APLICA - ANTERIOR E C0"
        assert res.Range("I259").Value in (None, "")  # N/A nao vira fator base
        wb.Close(SaveChanges=False)
    finally:
        xl.Quit()
        pythoncom.CoUninitialize()
        shutil.rmtree(tmp_dir, ignore_errors=True)
