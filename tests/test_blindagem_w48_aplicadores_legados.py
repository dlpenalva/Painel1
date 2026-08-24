# -*- coding: utf-8 -*-
"""W48.1 — trava contra reaplicacao de migradores historicos no template atual."""
from __future__ import annotations

import ast
import os
import sys
import types
from pathlib import Path

import pytest
from openpyxl import load_workbook

# Os aplicadores importam COM no topo, mas a blindagem exercitada aqui deve
# ocorrer antes de qualquer chamada COM e precisa continuar testavel em CI
# Linux. Os stubs so tornam o modulo importavel quando pywin32 nao existe.
try:  # pragma: no cover - depende do sistema operacional
    import pythoncom  # noqa: F401
except ModuleNotFoundError:  # pragma: no cover - exercitado no CI Linux
    sys.modules["pythoncom"] = types.ModuleType("pythoncom")

try:  # pragma: no cover - depende do sistema operacional
    import win32com.client  # noqa: F401
except ModuleNotFoundError:  # pragma: no cover - exercitado no CI Linux
    win32com = types.ModuleType("win32com")
    win32com_client = types.ModuleType("win32com.client")
    win32com.client = win32com_client
    sys.modules["win32com"] = win32com
    sys.modules["win32com.client"] = win32com_client

from tools import aplicar_hotfix_resultados_retro_vta as hotfix
from tools import aplicar_valor_considerado_vta as valor_considerado
from tools import aplicar_vta_posicoes_tabela1 as posicoes


ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "templates" / "COLETA_REAJUSTE_OFICIAL.xlsx"
FORMULA_CANONICA_W48 = (
    '=IF(OR($W$46="",$W$67=""),"",ROUND($W$67+$W$53+$W$54,2))'
)
ESCRITORES_ESPERADOS = {
    "tools/aplicar_hotfix_resultados_retro_vta.py",
    "tools/aplicar_valor_considerado_vta.py",
    "tools/aplicar_vta_posicoes_tabela1.py",
}


class _Range:
    def __init__(self, *, formula="", value=""):
        self.Formula = formula
        self.Value = value


class _Worksheet:
    def __init__(self, name: str, cells: dict[str, _Range] | None = None):
        self.Name = name
        self._cells = cells or {}

    def Range(self, address: str) -> _Range:
        return self._cells.get(address, _Range(formula="=1", value=""))


class _Worksheets:
    def __init__(self, sheets: list[_Worksheet]):
        self._sheets = {sheet.Name: sheet for sheet in sheets}

    def __iter__(self):
        return iter(self._sheets.values())

    def __call__(self, name: str) -> _Worksheet:
        return self._sheets[name]


class _Name:
    Name = "VTA_FINAL"


class _Workbook:
    def __init__(self, w48_formula: str = FORMULA_CANONICA_W48):
        memoria = _Worksheet(
            "MEMORIA_RESULTADOS",
            {"W48": _Range(formula=w48_formula)},
        )
        self.Worksheets = _Worksheets(
            [
                memoria,
                _Worksheet("RESULTADOS"),
                _Worksheet("itens_RC"),
                _Worksheet("posicao_contratual"),
                _Worksheet("historico_VU"),
                _Worksheet("itens_PC", {"O1": _Range(value="VALOR_PC_TOTAL")}),
                _Worksheet("comparativo_VTA"),
                _Worksheet("CONTROLE"),
            ]
        )
        self.Names = [_Name()]


def _escritores_w48() -> set[str]:
    escritores = set()
    for caminho in (ROOT / "tools").glob("*.py"):
        tree = ast.parse(caminho.read_text(encoding="utf-8"))
        for node in ast.walk(tree):
            if not isinstance(node, (ast.Assign, ast.AnnAssign)):
                continue
            targets = node.targets if isinstance(node, ast.Assign) else [node.target]
            for target in targets:
                if not isinstance(target, ast.Attribute) or target.attr != "Formula":
                    continue
                call = target.value
                if not isinstance(call, ast.Call) or not call.args:
                    continue
                arg = call.args[0]
                if isinstance(arg, ast.Constant) and arg.value == "W48":
                    escritores.add(caminho.relative_to(ROOT).as_posix())
    return escritores


def _formula_w48(caminho: Path):
    wb = load_workbook(caminho, data_only=False, read_only=True)
    try:
        return wb["MEMORIA_RESULTADOS"]["W48"].value
    finally:
        wb.close()


def test_template_e_proprietario_atual_usam_formula_canonica():
    assert _formula_w48(TEMPLATE) == FORMULA_CANONICA_W48
    assert hotfix.W48_NOVA == FORMULA_CANONICA_W48
    assert posicoes.W48_CANONICA_ATUAL == FORMULA_CANONICA_W48
    assert valor_considerado.W48_CANONICA_ATUAL == FORMULA_CANONICA_W48


def test_inventario_de_escritores_w48_permanece_explicito():
    assert _escritores_w48() == ESCRITORES_ESPERADOS


def test_posicoes_recusa_template_atual_antes_da_mutacao():
    with pytest.raises(ValueError, match="historico incompatível.*W48.*preservada"):
        posicoes._validar_origem(_Workbook())


def test_valor_considerado_recusa_template_atual_antes_da_mutacao():
    with pytest.raises(RuntimeError, match="historico incompatível.*W48.*preservada"):
        valor_considerado._validar_origem(_Workbook(), escreve_w48=True)


def test_valor_considerado_modo_somente_estilo_nao_e_bloqueado():
    valor_considerado._validar_origem(_Workbook(), escreve_w48=False)


def test_semantica_da_migracao_historica_permanece_reproduzivel():
    assert valor_considerado.FORMULA_W48 == hotfix.W48_ANTIGA
    assert valor_considerado.FORMULA_W48 != FORMULA_CANONICA_W48
    origem_historica = _Workbook(w48_formula=hotfix.W48_ANTIGA)
    posicoes._validar_origem(origem_historica)
    valor_considerado._validar_origem(origem_historica, escreve_w48=True)


@pytest.mark.parametrize(
    ("aplicador", "erro"),
    (
        (posicoes.aplicar, ValueError),
        (valor_considerado.aplicar, RuntimeError),
    ),
)
def test_copia_real_do_template_atual_e_recusada_sem_promocao(
    tmp_path: Path, aplicador, erro
):
    origem = tmp_path / "origem.xlsx"
    destino = tmp_path / "destino.xlsx"
    origem.write_bytes(TEMPLATE.read_bytes())
    formula_antes = _formula_w48(origem)

    with pytest.raises(erro, match="historico incompatível.*W48.*preservada"):
        aplicador(origem, destino)

    assert not destino.exists()
    formula_depois = _formula_w48(origem)
    assert formula_depois == formula_antes == FORMULA_CANONICA_W48


@pytest.mark.skipif(
    os.environ.get("RUN_EXCEL_INTEGRATION") != "1",
    reason="defina RUN_EXCEL_INTEGRATION=1 para executar Excel COM",
)
def test_copia_real_modo_somente_estilo_preserva_w48_e_abre(tmp_path: Path):
    origem = tmp_path / "origem_estilo.xlsx"
    destino = tmp_path / "destino_estilo.xlsx"
    origem.write_bytes(TEMPLATE.read_bytes())

    valor_considerado.aplicar(origem, destino, somente_estilo_u=True)

    assert destino.is_file()
    assert _formula_w48(destino) == FORMULA_CANONICA_W48
    wb = load_workbook(destino, data_only=False, read_only=True)
    try:
        assert "MEMORIA_RESULTADOS" in wb.sheetnames
        assert "itens_PC" in wb.sheetnames
    finally:
        wb.close()
