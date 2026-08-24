# -*- coding: utf-8 -*-
"""Regra temporal definitiva dos aditivos: DATA_EFEITO x DATA DE ABERTURA.

Fonte da mutacao: `tools/aplicar_temporalidade_aditivos.py`.

CAUSA DA RESSALVA ANTERIOR (Regra 2)
------------------------------------
`posicao_contratual!Y` (CICLO_NASCIMENTO) deriva da QUANTIDADE — o primeiro
ciclo com QTD_CONTRATADA_Cn > 0. Como `DELTA_Cn` agrega TODOS os aditivos do
ciclo por ROTULO de ciclo, um item incluido no primeiro dia e outro incluido no
meio do MESMO ciclo recebiam a mesma classificacao e a mesma abertura.

SOLUCAO
-------
Coluna nova `posicao_contratual!AL` (CICLO_NASCIMENTO_DATA), derivada de
`QTD_CONTRATUAL_ABERTURA_Cn = QTD_CONTRATADA_Cn - DELTA_POSTERIOR_ABERTURA_Cn`,
onde `DELTA_POSTERIOR_ABERTURA_Cn` sao os deltas do proprio ciclo com data
ESTRITAMENTE posterior a `parametros!C{n+2}` (abertura do ciclo).

Nenhuma formula da cadeia oficial (`T21`/`T22`/`T23`/`T25`/`B26`) e tocada: a
correcao vive na camada auditavel (FORMA 2 redecomposta) e na apresentacao.

Cenario sintetico (arquivo real: C0 = 01/02/2025..31/01/2026;
C1 = 01/02/2026..31/01/2027; vigente C1, portanto ABERTURA DO C1 = 01/02/2026):

  item | data de efeito | posicao          | AL esperado
  -----+----------------+------------------+------------
  N001 | 15/06/2025     | meio do C0       | 1 (nao existia na abertura do C0;
       |                |                  |    ja existia na abertura do C1)
  N002 | 12/06/2026     | meio do C1       | 2
  N003 | 01/02/2026     | 1o dia do C1     | 1
  N004 | 31/01/2026     | 1 dia antes      | 1
  N005 | 02/02/2026     | 1 dia depois     | 2
  N006 | 01/02 + 15/07  | 2 aditivos em C1 | 1 (abertura so com o do 1o dia)
  N007 | 01/02 + 10/09  | acresc + supres. | 1 (supressao nao retroage)
  N008 | 05/03 + 20/09  | criado e suprim. | "" (nunca existiu em abertura)
  N009 | 15/06/2027     | meio do C2       | 3
  N010 | 15/06/2028     | meio do C3       | 4
  N011 | 15/06/2029     | meio do C4       | "" (nao ha abertura C5)
"""
from __future__ import annotations

import datetime as dt
import os
import shutil
import tempfile
import time
from pathlib import Path

import pytest
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "templates" / "COLETA_REAJUSTE_OFICIAL.xlsx"
ARQUIVO_REAL = Path(
    r"C:\Users\danie\OneDrive\Desktop"
    r"\2. Coleta de Dados Reajuste INBLOKO_2026-07-31.xlsx"
)
VTA_ESPERADO = 13_468_851.41
TOL = 0.005

# Primeira linha livre de itens_Remanesc no arquivo real (21 itens: 2..22).
LINHA_ITEM = 23
ABERTURA_C1 = dt.datetime(2026, 2, 1)

# QTD_REM_BASE_C1 informada pelo fiscal (itens_Remanesc!E). Somente itens que
# ja existiam na abertura do C1 recebem confirmacao; os demais ficam vazios —
# vazio de item NAO APLICAVEL nunca e pendencia (Regra C).
REMANESCENTE_C1_INFORMADO = {"N001": 10.0, "N004": 40.0}

# (sufixo, [(data, tipo, qtd), ...], AL esperado)
CENARIOS = (
    ("N001", [(dt.datetime(2025, 6, 15), "Acrescimo", 10)], 1),
    ("N002", [(dt.datetime(2026, 6, 12), "Acrescimo", 20)], 2),
    ("N003", [(ABERTURA_C1, "Acrescimo", 30)], 1),
    ("N004", [(dt.datetime(2026, 1, 31), "Acrescimo", 40)], 1),
    ("N005", [(dt.datetime(2026, 2, 2), "Acrescimo", 50)], 2),
    ("N006", [(ABERTURA_C1, "Acrescimo", 10),
              (dt.datetime(2026, 7, 15), "Acrescimo", 5)], 1),
    ("N007", [(ABERTURA_C1, "Acrescimo", 40),
              (dt.datetime(2026, 9, 10), "Supressao", 15)], 1),
    ("N008", [(dt.datetime(2026, 3, 5), "Acrescimo", 12),
              (dt.datetime(2026, 9, 20), "Supressao", 12)], ""),
    ("N009", [(dt.datetime(2027, 6, 15), "Acrescimo", 7)], 3),
    ("N010", [(dt.datetime(2028, 6, 15), "Acrescimo", 8)], 4),
    ("N011", [(dt.datetime(2029, 6, 15), "Acrescimo", 9)], ""),
)

# QTD contratual esperada na ABERTURA do C1 (coluna AH) por item.
ABERTURA_C1_ESPERADA = {
    "N001": 10.0,   # aditivo do meio do C0 ja estava vigente em 01/02/2026
    "N002": 0.0,    # efeito posterior a abertura
    "N003": 30.0,   # 1o dia: entra uma unica vez
    "N004": 40.0,   # vespera da abertura
    "N005": 0.0,    # dia seguinte a abertura
    "N006": 10.0,   # so o aditivo do 1o dia
    "N007": 40.0,   # supressao de 10/09 nao retroage a abertura
    "N008": 0.0,
}


def _serial(data: dt.datetime) -> int:
    """Data -> numero de serie do Excel (epoca 30/12/1899)."""
    return (data - dt.datetime(1899, 12, 30)).days


def _aplicar_com_retentativa(funcao, origem: Path, destino: Path) -> Path:
    """Excel COM fica contendido sob a suite integral; tolera 3 tentativas."""
    ultima = None
    for tentativa in range(3):
        try:
            funcao(origem, destino)
            return destino
        except Exception as exc:  # pragma: no cover - resiliencia COM
            ultima = exc
            if destino.exists():
                destino.unlink()
            time.sleep(5 * (tentativa + 1))
    raise RuntimeError(f"{funcao.__module__} falhou 3x: {ultima}")


@pytest.fixture(scope="session")
def real_temporal(tmp_path_factory) -> Path:
    """Arquivo real com as duas mutacoes da etapa aplicadas (sem dados novos)."""
    if os.environ.get("RUN_EXCEL_INTEGRATION") != "1":
        pytest.skip("defina RUN_EXCEL_INTEGRATION=1 para executar Excel COM")
    if not ARQUIVO_REAL.is_file():
        pytest.skip(f"arquivo real ausente: {ARQUIVO_REAL}")
    pytest.importorskip("win32com.client")
    from tools.aplicar_temporalidade_aditivos import aplicar as aplicar_temporal
    from tools.aplicar_vta_posicoes_tabela1 import aplicar as aplicar_posicoes

    base = tmp_path_factory.mktemp("temporal")
    passo1 = _aplicar_com_retentativa(
        aplicar_posicoes, ARQUIVO_REAL, base / "passo1.xlsx"
    )
    return _aplicar_com_retentativa(
        aplicar_temporal, passo1, base / "real_temporal.xlsx"
    )


@pytest.fixture(scope="session")
def cenario(real_temporal) -> dict:
    """Injeta os itens sinteticos via Excel COM e devolve as celulas lidas."""
    client = pytest.importorskip("win32com.client")
    pythoncom = pytest.importorskip("pythoncom")

    tmp_dir = Path(tempfile.mkdtemp(prefix="temporal_cenario_"))
    tmp = tmp_dir / "cenario.xlsx"
    shutil.copyfile(real_temporal, tmp)

    pythoncom.CoInitialize()
    excel = client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    lidos: dict = {}
    try:
        wb = excel.Workbooks.Open(
            str(tmp), UpdateLinks=0, ReadOnly=False, CorruptLoad=0
        )
        rem = wb.Worksheets("itens_Remanesc")
        adi = wb.Worksheets("aditivos")
        for aba in (rem, adi):
            try:
                aba.Unprotect()
            except Exception:  # pragma: no cover - aba ja desprotegida
                pass

        # Item original ja existente que recebera um aditivo no MEIO do C1.
        item_original = str(rem.Range("A2").Value or "").strip()
        lidos["item_original"] = item_original

        def escrever_aditivo(linha: int, item: str, data: dt.datetime,
                             tipo: str, qtd: float) -> None:
            adi.Range(f"A{linha}").Value = item
            # Serial do Excel: gravar `datetime` por COM desloca o horario
            # (fuso do pywin32) e contaminaria a fronteira do dia de abertura.
            adi.Range(f"B{linha}").Value = float(_serial(data))
            adi.Range(f"B{linha}").NumberFormat = "dd/mm/yyyy;@"
            adi.Range(f"D{linha}").Value = tipo
            adi.Range(f"E{linha}").Value = qtd
            adi.Range(f"H{linha}").Value = "Sim"

        linha_adi = 2
        for indice, (nome, movimentos, _) in enumerate(CENARIOS):
            linha = LINHA_ITEM + indice
            rem.Range(f"A{linha}").Value = nome
            rem.Range(f"C{linha}").Value = 100.0  # VU_ORIGINAL
            if nome in REMANESCENTE_C1_INFORMADO:
                rem.Range(f"E{linha}").Value = REMANESCENTE_C1_INFORMADO[nome]
            for data, tipo, qtd in movimentos:
                escrever_aditivo(linha_adi, nome, data, tipo, qtd)
                linha_adi += 1
        # Item ORIGINAL com acrescimo no meio do C1 (Regra A + componente proprio).
        escrever_aditivo(
            linha_adi, item_original, dt.datetime(2026, 6, 12), "Acrescimo", 5
        )

        excel.CalculateFullRebuild()

        pc = wb.Worksheets("posicao_contratual")
        rc = wb.Worksheets("itens_RC")
        mem = wb.Worksheets("MEMORIA_RESULTADOS")
        for indice, (nome, _, _) in enumerate(CENARIOS):
            linha = LINHA_ITEM + indice
            lidos[nome] = {
                "AA": pc.Range(f"AA{linha}").Value,
                "AL": pc.Range(f"AL{linha}").Value,
                "Y": pc.Range(f"Y{linha}").Value,
                "AG": pc.Range(f"AG{linha}").Value,
                "AH": pc.Range(f"AH{linha}").Value,
                "AI": pc.Range(f"AI{linha}").Value,
                "K": pc.Range(f"K{linha}").Value,
                "AC_delta_post_c1": pc.Range(f"AC{linha}").Value,
                "rem_E": rem.Range(f"E{linha}").Value,
                "rem_F": rem.Range(f"F{linha}").Value,
                "rc_ciclo": rc.Range(f"AA{linha + 1}").Value,
                "rc_aplicavel": rc.Range(f"AB{linha + 1}").Value,
                "rc_qtd_abertura": rc.Range(f"AC{linha + 1}").Value,
            }
        lidos["original"] = {
            "K": pc.Range("K2").Value,
            "I_contratada_c1": pc.Range("I2").Value,
            "AH": pc.Range("AH2").Value,
            "AC_delta_post_c1": pc.Range("AC2").Value,
            "AL": pc.Range("AL2").Value,
            "rc_aplicavel": rc.Range("AB3").Value,
        }
        for endereco in (
            "T20", "T21", "T22", "T23", "T25", "W46", "W48",
            "W53", "W54", "W55", "W56", "W57",
        ):
            lidos[endereco] = mem.Range(endereco).Value
        lidos["B26"] = mem.Range("B26").Value
        wb.Close(SaveChanges=False)
    finally:
        excel.Quit()
        del excel
        pythoncom.CoUninitialize()
        shutil.rmtree(tmp_dir, ignore_errors=True)
    return lidos


# --------------------------------------------------------------------------- #
# Trava: VTA oficial inalterado                                                #
# --------------------------------------------------------------------------- #
def test_trava_b26_t25_intactos_no_arquivo_real(real_temporal):
    """As duas mutacoes nao alteram o VTA oficial do arquivo real."""
    wb = load_workbook(real_temporal, data_only=True)
    mem = wb["MEMORIA_RESULTADOS"]
    assert abs(mem["T25"].value - VTA_ESPERADO) <= TOL
    assert abs(mem["B26"].value - VTA_ESPERADO) <= TOL
    assert abs(
        (mem["T21"].value + mem["T22"].value + mem["T23"].value) - VTA_ESPERADO
    ) <= TOL


def test_trava_formulas_oficiais_nao_tocadas():
    """T21/T22/T23/T25/B26 e as colunas G/K/O/S/W seguem com a formula homologada."""
    wb = load_workbook(TEMPLATE)
    mem = wb["MEMORIA_RESULTADOS"]
    assert mem["T23"].value == "=SUM($Y$2:$Y$201)"
    assert mem["T25"].value.endswith("ROUND($T$21+$T$22+$T$23,2))")
    assert "$T$25" in mem["B26"].value
    pc = wb["posicao_contratual"]
    # A abertura OFICIAL segue somando o delta do ciclo por rotulo (homologada).
    assert pc["K2"].value == (
        '=IF(A2="","",IF($Y2="","",IF($Y2>1,"",'
        'IF(ISNUMBER(J2),ROUND(J2+H2,2),IF($Y2=1,ROUND(H2,2),"")))))'
    )
    assert pc["Y2"].value.startswith('=IF($A2="","",IF($E2>0,0,')
    # Nenhuma coluna nova alimenta a cadeia oficial.
    for celula in ("T21", "T22", "T23", "T25", "B26"):
        formula = str(mem[celula].value)
        for nova in ("$AC$", "$AD$", "$AL", "$AH", "W53", "W54"):
            assert nova not in formula, f"{celula} passou a depender de {nova}"


# --------------------------------------------------------------------------- #
# Camada temporal no template                                                  #
# --------------------------------------------------------------------------- #
def test_coluna_nascimento_por_data_existe_e_usa_abertura():
    wb = load_workbook(TEMPLATE)
    pc = wb["posicao_contratual"]
    assert pc["AL1"].value == "CICLO_NASCIMENTO_DATA"
    assert pc["AH1"].value == "QTD_CONTRATUAL_ABERTURA_C1"
    assert pc["AC1"].value == "DELTA_POSTERIOR_ABERTURA_C1"
    # O delta posterior recorta por DATA, nao apenas pelo rotulo do ciclo.
    ac2 = str(pc["AC2"].value)
    assert 'aditivos!$C$2:$C$200,"C1"' in ac2
    # Fronteira por DIA: aditivo datado NO dia da abertura fica do lado da
    # abertura mesmo que a celula carregue componente de hora. ETAPA 48: a
    # fronteira fisica passou a ler parametros!I (data exata); a guarda e a
    # condicao de janela permanecem na cadeia mensal (parametros!C).
    # ETAPA 48.3: a guarda tambem reconhece I vazio — ciclo futuro sem
    # fotografia fisica nao calcula DELTA_POSTERIOR (INT de celula vazia
    # viraria fronteira ficticia 01/01/1900).
    assert 'aditivos!$B$2:$B$200,">="&(INT(parametros!$I$3)+1)' in ac2
    assert 'IF(OR(parametros!$C$3="",parametros!$I$3=""),0,' in ac2
    # A qtd na abertura e a contratada MENOS o que so passou a valer depois.
    assert str(pc["AH2"].value).endswith("ROUND($I2-$AC2,2)))")
    assert str(pc["AL2"].value).count("N($A") == 5


def test_aplicador_canonico_sincronizado_com_o_template_homologado():
    """ETAPA 48.3 — o aplicador nao pode regredir a fronteira fisica.

    `tools/aplicar_temporalidade_aditivos.py` e o dono canonico de
    posicao_contratual!AB2:AF200. Se ele voltar a gerar a fronteira em
    parametros!C (pre-Etapa 48) ou perder a guarda de I vazio (Etapa 48.3),
    uma reexecucao futura destruiria o template homologado.
    """
    from tools.aplicar_temporalidade_aditivos import (
        COLS_DELTA_POSTERIOR, PC_FIM, PC_INI, _formula_delta_posterior,
    )

    for ciclo in range(5):
        formula = _formula_delta_posterior(2, ciclo)
        n = ciclo + 2                                   # parametros!C2..C6 / I2..I6
        # fronteira FISICA em parametros!I (nunca mais na cadeia mensal)
        assert f"INT(parametros!$I${n})+1" in formula, ciclo
        assert "INT(parametros!$C$" not in formula, ciclo
        # guarda: ciclo inexistente (C vazio) OU sem fotografia fisica (I vazio)
        assert f'IF(OR(parametros!$C${n}="",parametros!$I${n}=""),0,' in formula, ciclo
        # classificacao mensal do aditivo intacta
        assert f'aditivos!$C$2:$C$200,"C{ciclo}"' in formula, ciclo
        assert "SUMIFS(aditivos!$L$2:$L$200" in formula, ciclo

    # sincronizacao exata: o que o tool gera == o que esta homologado no template
    wb = load_workbook(TEMPLATE)
    pc = wb["posicao_contratual"]
    for r in range(PC_INI, PC_FIM + 1):
        for ciclo in range(5):
            coord = f"{COLS_DELTA_POSTERIOR[ciclo]}{r}"
            assert str(pc[coord].value) == _formula_delta_posterior(r, ciclo), coord


def test_formatacao_condicional_usa_nascimento_por_data():
    wb = load_workbook(TEMPLATE)
    ws = wb["itens_Remanesc"]
    por_faixa = {str(cf.sqref): list(cf.rules) for cf in ws.conditional_formatting}
    for col, idx in (("E", 1), ("G", 2), ("I", 3), ("K", 4)):
        regras = por_faixa[f"{col}2:{col}201"]
        assert len(regras) == 4
        juntas = " ".join(str(r.formula) for r in regras)
        # Referencia LOCAL (espelho BI): regra de CF que aponta para outra aba e
        # migrada pelo Excel para a extensao x14, invisivel ao openpyxl.
        assert f"$BI2>{idx}" in juntas
        assert "posicao_contratual!" not in juntas
    espelho = str(ws["BI2"].value)
    assert "posicao_contratual!$AL2" in espelho


def test_valor_de_abertura_vazio_para_item_posterior():
    """Regra C: sem quantidade na abertura, tambem nao ha VALOR de abertura."""
    wb = load_workbook(TEMPLATE)
    ws = wb["itens_Remanesc"]
    for col, idx in (("F", 1), ("H", 2), ("J", 3), ("L", 4)):
        formula = str(ws[f"{col}2"].value)
        assert (
            f"AND(ISNUMBER(posicao_contratual!$AL2),"
            f"posicao_contratual!$AL2>{idx})" in formula
        )


def test_ciclo_em_execucao_soma_delta_da_abertura_uma_vez():
    """Regra B na aba fiscal: aditivo do 1o dia entra na abertura, nao no periodo."""
    from _ciclo_em_execucao import _formula_abertura

    formula = _formula_abertura(13, 2)
    assert 'aditivos!$B$2:$B$200,"<"&(INT($F$3)+1)' in formula
    assert "aditivos!$C$2:$C$200,$C$3" in formula
    # A coluna I (alteracoes do periodo) comeca no dia SEGUINTE a abertura, de
    # modo que o delta da abertura nunca e contado duas vezes.
    assert '">"&$F$3' not in formula
    fonte = (ROOT / "_ciclo_em_execucao.py").read_text(encoding="utf-8")
    assert 'aditivos!$B$2:$B$200,">="&(INT($F$3)+1),' in fonte


def test_motor_puro_nao_reaplica_delta_da_abertura():
    """O delta do 1o dia entra UMA vez — na coluna B da aba, nao no motor puro.

    O motor recebe `remanescente_inicio` ja ajustado (a aba soma o delta da
    abertura a QTD_REM_BASE); reaplica-lo aqui seria dupla contagem.
    """
    from _ciclo_em_execucao import calcular_posicao_ciclo_por_data

    resultado = calcular_posicao_ciclo_por_data(
        ciclo="C1",
        data_inicio=dt.date(2026, 2, 1),
        data_fim=dt.date(2027, 1, 31),
        data_posicao=dt.date(2026, 8, 31),
        itens=[{
            "item": "N003",
            "remanescente_inicio": 30.0,
            "remanescente_atual": 30.0,
            "vu_atualizado": 100.0,
            "novo_item": True,
        }],
        movimentos=[{
            "item": "N003",
            "data_efeito": dt.date(2026, 2, 1),
            "delta": 30.0,
        }],
    )
    linha = resultado["itens"][0]
    assert linha["remanescente_inicio"] == 30.0
    assert linha["alteracoes_liquidas_periodo"] == 0.0, "delta nao vai ao periodo"
    assert linha["quantidade_consumida"] == 0.0, "delta nao e reaplicado"
    assert resultado["valido"] is True


def test_motor_puro_ignora_item_nascido_apos_a_posicao():
    """Regra C: item incluido depois da data da posicao nao compoe a fotografia."""
    from _ciclo_em_execucao import calcular_posicao_ciclo_por_data

    resultado = calcular_posicao_ciclo_por_data(
        ciclo="C1",
        data_inicio=dt.date(2026, 2, 1),
        data_fim=dt.date(2027, 1, 31),
        data_posicao=dt.date(2026, 5, 31),
        itens=[{
            "item": "N002",
            "remanescente_inicio": None,
            "remanescente_atual": None,
            "vu_atualizado": 100.0,
            "novo_item": True,
        }],
        movimentos=[{
            "item": "N002",
            "data_efeito": dt.date(2026, 6, 12),
            "delta": 20.0,
        }],
    )
    assert resultado["itens"] == []


# --------------------------------------------------------------------------- #
# Casos obrigatorios (Secao 6)                                                 #
# --------------------------------------------------------------------------- #
@pytest.mark.parametrize("nome,_movs,esperado", CENARIOS)
def test_classificacao_temporal_por_item(cenario, nome, _movs, esperado):
    lido = cenario[nome]
    obtido = lido["AL"]
    if esperado == "":
        assert obtido in (None, ""), f"{nome}: AL={obtido!r}"
    else:
        assert obtido == esperado, f"{nome}: AL={obtido!r}, esperado {esperado}"


@pytest.mark.parametrize("nome,esperado", sorted(ABERTURA_C1_ESPERADA.items()))
def test_qtd_contratual_na_abertura_do_c1(cenario, nome, esperado):
    obtido = cenario[nome]["AH"]
    assert obtido is not None and abs(float(obtido) - esperado) <= TOL, (
        f"{nome}: abertura C1 = {obtido!r}, esperado {esperado}"
    )


def test_n001_incluido_em_c0_e_aplicavel_na_abertura_do_c1(cenario):
    """Regra A: item anterior a abertura do C1 — o fiscal informa o saldo."""
    lido = cenario["N001"]
    assert lido["AL"] == 1, "nao existia na abertura do C0; existe na do C1"
    assert lido["rc_aplicavel"] == "SIM"
    assert abs(float(lido["rem_E"]) - 10.0) <= TOL
    assert abs(float(lido["AH"]) - 10.0) <= TOL
    # Data de efeito exposta e coerente com o aditivo (15/06/2025).
    assert lido["AA"] is not None
    assert (lido["AA"].year, lido["AA"].month, lido["AA"].day) == (2025, 6, 15)


def test_n002_incluido_no_meio_do_c1_nao_aplicavel_na_abertura(cenario):
    lido = cenario["N002"]
    assert lido["AL"] == 2, "nasce so na abertura do C2"
    assert lido["rc_aplicavel"] == "NAO APLICAVEL"
    assert abs(float(lido["AH"])) <= TOL, "abertura C1 sem quantidade"
    assert lido["rem_F"] in (None, ""), "valor de abertura deve ficar vazio"
    # Nao bloqueia: a abertura do ciclo vigente segue completa e a FORMA 2 sai.
    assert cenario["W46"] == 1
    assert cenario["W48"] not in (None, "") and float(cenario["W48"]) > 0
    # Reaparece na abertura do C2, onde ha saldo.
    assert float(lido["AI"]) > 0


def test_n003_incluido_no_primeiro_dia_entra_uma_unica_vez(cenario):
    lido = cenario["N003"]
    assert lido["AL"] == 1
    assert lido["rc_aplicavel"] == "SIM"
    assert abs(float(lido["AH"]) - 30.0) <= TOL
    # Delta da abertura nao e reaplicado como alteracao posterior.
    assert abs(float(lido["AC_delta_post_c1"])) <= TOL


def test_vespera_e_dia_seguinte_da_abertura(cenario):
    assert cenario["N004"]["AL"] == 1, "1 dia antes ja existe na abertura"
    assert cenario["N005"]["AL"] == 2, "1 dia depois nao existe na abertura"


def test_dois_aditivos_no_mesmo_ciclo(cenario):
    lido = cenario["N006"]
    assert abs(float(lido["AH"]) - 10.0) <= TOL, "so o do 1o dia na abertura"
    assert abs(float(lido["AI"]) - 15.0) <= TOL, "os dois na abertura do C2"


def test_acrescimo_seguido_de_supressao_nao_retroage(cenario):
    lido = cenario["N007"]
    assert abs(float(lido["AH"]) - 40.0) <= TOL
    assert abs(float(lido["AI"]) - 25.0) <= TOL


def test_item_totalmente_suprimido_nunca_entra_em_abertura(cenario):
    lido = cenario["N008"]
    assert lido["AL"] in (None, "")
    assert abs(float(lido["AH"])) <= TOL
    assert abs(float(lido["AI"])) <= TOL


def test_inclusao_no_meio_de_c2_c3_e_c4(cenario):
    assert cenario["N009"]["AL"] == 3
    assert cenario["N010"]["AL"] == 4
    assert cenario["N011"]["AL"] in (None, "")


def test_item_original_com_aditivo_no_meio_do_ciclo(cenario):
    """Regra A/D: a abertura oficial nao muda; o temporal isola o posterior."""
    lido = cenario["original"]
    assert lido["AL"] == 0, "item original nasce no C0"
    assert lido["rc_aplicavel"] == "SIM"
    # O acrescimo de 12/06/2026 e integralmente posterior a abertura do C1.
    assert abs(float(lido["AC_delta_post_c1"]) - 5.0) <= TOL
    # A qtd CONTRATUAL na abertura exclui o acrescimo; a coluna oficial K
    # (remanescente ajustado) permanece com a formula homologada, que o inclui.
    assert abs(
        float(lido["AH"]) - (float(lido["I_contratada_c1"]) - 5.0)
    ) <= TOL
    # O acrescimo entra na FORMA 2 por componente proprio, uma unica vez.
    assert float(cenario["W54"]) > 0


# --------------------------------------------------------------------------- #
# FORMA 2 redecomposta e reconciliacao (Secoes 4 e 7)                          #
# --------------------------------------------------------------------------- #
def test_forma2_tem_componente_proprio_para_aditivos_posteriores(cenario):
    assert cenario["W53"] is not None and cenario["W54"] is not None
    # Ha aditivos posteriores a abertura no cenario -> componente proprio > 0.
    assert float(cenario["W54"]) > 0
    assert float(cenario["W57"]) >= 2, "N002/N005 incluidos apos a abertura"


def test_trava_anti_dupla_contagem(cenario):
    assert abs(float(cenario["W55"])) <= TOL
    assert cenario["W56"] == "SEM DUPLA CONTAGEM"


def test_forma2_total_preservado_pela_redecomposicao(cenario):
    """A redecomposicao muda a LEITURA temporal, nao o total da FORMA 2."""
    esperado = (
        float(cenario["T21"]) + float(cenario["T22"]) + float(cenario["T23"])
    )
    assert abs(float(cenario["W48"]) - esperado) <= TOL


def test_forma2_reconcilia_com_o_vta_pc_no_arquivo_real(real_temporal):
    wb = load_workbook(real_temporal, data_only=True)
    mem = wb["MEMORIA_RESULTADOS"]
    assert mem["W46"].value == mem["T20"].value
    assert abs(mem["W48"].value - VTA_ESPERADO) <= TOL
    # Sem aditivos, todo o remanescente da abertura e temporalmente valido.
    assert abs(mem["W53"].value - mem["T23"].value) <= TOL
    assert abs(mem["W54"].value) <= TOL
    assert abs(mem["W55"].value) <= TOL
