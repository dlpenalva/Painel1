# -*- coding: utf-8 -*-
"""Hotfix RESULTADOS — retroativo visivel + VTA method-aware + contraste.

Fonte da mutacao: `tools/aplicar_hotfix_resultados_retro_vta.py`.

Bugs corrigidos (comprovados no arquivo real do enunciado):
  A. O card "RETROATIVO TOTAL A PAGAR" tinha o valor deslocado para E5; sob o
     rotulo (D5) vivia a ancora oculta do ciclo. O valor passa a ocupar D5:E5.
  B. MEMORIA_RESULTADOS!W48/W50 compunham a execucao historica SOMENTE por
     itens_PC!O+Q: no metodo Financeiro os ciclos encerrados C0..C(n-1)
     viravam zero e B10/B11 mostravam apenas a fotografia fisica do vigente.

Cenario Financeiro real reconstruido (valores do enunciado):
  financeiro!E por ciclo: C0=4.739.007,27; C1=853.961,00; C2=853.961,00;
  C3 (vigente)=878.639,92 — presente e NUNCA somado ao historico.
  Retroativo oficial: 24.678,92. Posicao fisica C3: 1.388.251,07.
  Esperado: B10 = B11 = 6.446.929,27 + 1.388.251,07 = 7.835.180,34; B13 = 0.
"""
from __future__ import annotations

import sys
from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

TEMPLATE = ROOT / "templates" / "COLETA_REAJUSTE_OFICIAL.xlsx"
TOL = 0.005

F1 = 0.04944229095799391
F2 = 0.01961153826825468
FIN = {"c0": 4_739_007.27, "c1": 853_961.00, "c2": 853_961.00,
       "c3": 878_639.92}
HISTORICO = 6_446_929.27          # C0 + C1 + C2 (financeiro!E)
RETRO = 24_678.92                 # delta exato de uma competencia C1 com efeito
RETRO_C, RETRO_E = 499_145.87, 523_824.79
FOTO_FISICA = 1_388_251.07
VTA_ESPERADO = 7_835_180.34
# 138.825.107 cents e primo: qty 1,00 com VU_C3 exato fecha a foto fisica.
# Cadeia em degraus: 1.297.402,45 -> 1.361.549,00 -> 1.388.251,07 (C3 = 0%).
VU0 = 1_297_402.45


# --------------------------------------------------------------------------- #
# Estrutura do template                                                        #
# --------------------------------------------------------------------------- #
@pytest.fixture(scope="module")
def wb_template():
    return load_workbook(TEMPLATE, data_only=False)


def test_w48_w50_consomem_o_historico_method_aware(wb_template):
    mem = wb_template["MEMORIA_RESULTADOS"]
    assert mem["W48"].value == (
        '=IF(OR($W$46="",$W$67=""),"",ROUND($W$67+$W$53+$W$54,2))'
    )
    w50 = str(mem["W50"].value)
    assert "$W$66" in w50 and "$W$49=0" in w50
    assert "CICLO_EM_EXECUCAO!F13:F211" in w50
    assert "CICLO_EM_EXECUCAO!G13:G211" in w50


def test_bloco_de_apoio_financeiro_por_ciclo(wb_template):
    mem = wb_template["MEMORIA_RESULTADOS"]
    for n in range(5):
        formula = str(mem[f"W{61 + n}"].value)
        assert formula == (
            f'=ROUND(SUMIF(financeiro!$B$2:$B$73,"c{n}",'
            f'financeiro!$E$2:$E$73),2)'
        )
    # SUMIF nao filtra EFEITO_FINANCEIRO: preclusos (E=C) seguem compondo.
    assert "G$2" not in str(mem["W61"].value)


def test_ramo_nao_financeiro_preserva_regra_pc(wb_template):
    """PC/Itens: exatamente as expressoes homologadas (T21+T22; O+Q ate W46)."""
    mem = wb_template["MEMORIA_RESULTADOS"]
    w66, w67 = str(mem["W66"].value), str(mem["W67"].value)
    assert '$T$21+$T$22' in w66
    assert (
        '$T$21+SUMPRODUCT((ROW(itens_PC!$O$2:$O$6)-2>=1)'
        '*(ROW(itens_PC!$O$2:$O$6)-2<$W$46)'
        '*(itens_PC!$O$2:$O$6+itens_PC!$Q$2:$Q$6))'
    ) in w67
    # VALOR CONSIDERADO (O+Q), nunca VALOR_ATUALIZADO (P).
    assert "itens_PC!$P$" not in w67
    # O vigente NUNCA entra no historico (indice estritamente menor).
    assert "<$T$20" in w66 and "<$W$46" in w67


def test_travas_fora_da_cadeia_intactas(wb_template):
    mem = wb_template["MEMORIA_RESULTADOS"]
    assert mem["T21"].value == (
        '=IF(itens_PC!$O$2>0,ROUND(itens_PC!$O$2+itens_PC!$Q$2,2),'
        'SUM($X$2:$X$201))'
    )
    assert mem["B26"].value.startswith("=IF(AND(B24")
    assert "$T$25" in mem["B26"].value
    assert mem["W51"].value == '=IF(OR($W$50="",$W$48=""),"",ROUND($W$50-$W$48,2))'


def test_textos_auditaveis_method_aware(wb_template):
    res = wb_template["RESULTADOS"]
    c10, c11 = str(res["C10"].value), str(res["C11"].value)
    assert "MEMORIA!W66" in c10 and "financeiro!E" in c10
    assert "MEMORIA!W67" in c11 and "financeiro!E" in c11
    # O ramo nao-Financeiro segue descrevendo T21/T22.
    assert "MEMORIA!T21" in c10 and "MEMORIA!T21" in c11


def test_card_retroativo_valor_sob_o_rotulo(wb_template):
    res = wb_template["RESULTADOS"]
    assert res["D4"].value == "RETROATIVO TOTAL A PAGAR"
    assert res["D5"].value == "=$D$22"
    assert "D5:E5" in {str(m) for m in res.merged_cells.ranges}
    assert res["J8"].value == "=UPPER(CONTROLE!$B$2)"
    assert res["J8"].number_format == ";;;"
    assert res["C3"].value == '=IF($J$8="","—",UPPER($J$8))'


def test_contraste_sem_residuo_8497b0(wb_template):
    res = wb_template["RESULTADOS"]
    residuais = [
        c.coordinate
        for row in res.iter_rows(min_row=1, max_row=66, max_col=10)
        for c in row
        if c.value is not None and c.font and c.font.color
        and c.font.color.rgb and "8497B0" in str(c.font.color.rgb)
    ]
    assert residuais == []
    # Titulos de card em azul institucional; chip do card 1 (fundo escuro)
    # em branco; auxiliares em cinza 595959.
    assert res["D4"].font.color.rgb == "FF1F4E78"
    assert res["F4"].font.color.rgb == "FF1F4E78"
    assert res["C4"].font.color.rgb == "FFFFFFFF"
    for endereco in ("E4", "H4", "E6", "F6"):
        assert res[endereco].font.color.rgb == "FF595959", endereco


# --------------------------------------------------------------------------- #
# Cenario Financeiro real reconstruido (motor real da planilha, Excel COM)     #
# --------------------------------------------------------------------------- #
def _montar_cenario_financeiro(destino: Path, *, com_posicao_fisica: bool) -> Path:
    from _ciclo_em_execucao import garantir_aba_ciclo_em_execucao

    wb = load_workbook(TEMPLATE)
    ctl = wb["CONTROLE"]
    ctl["B1"], ctl["B2"], ctl["B3"] = (
        "Financeiro (Mensalidade)", "C3", date(2026, 1, 15)
    )
    par = wb["parametros"]
    par["A2"], par["C2"], par["D2"] = "Nao", date(2022, 2, 1), date(2023, 1, 31)
    par["A3"], par["C3"], par["D3"], par["E3"] = (
        "Sim", date(2023, 2, 1), date(2024, 1, 31), F1
    )
    par["A4"], par["C4"], par["D4"], par["E4"] = (
        "Sim", date(2024, 2, 1), date(2025, 1, 31), F2
    )
    par["A5"], par["C5"], par["D5"], par["E5"] = (
        "Sim", date(2025, 3, 1), date(2026, 2, 28), 0.0
    )
    fin = wb["financeiro"]
    linhas = [
        (date(2022, 6, 15), FIN["c0"], "Nao"),          # C0 precluso: E = C
        (date(2023, 6, 15), RETRO_C, "Sim"),            # C1 com efeito
        (date(2023, 7, 15), round(FIN["c1"] - RETRO_E, 2), "Nao"),
        (date(2024, 6, 15), FIN["c2"], "Nao"),          # C2 precluso: E = C
        (date(2025, 6, 15), FIN["c3"], "Nao"),          # C3 VIGENTE
    ]
    for i, (dia, valor, efeito) in enumerate(linhas, start=2):
        fin[f"A{i}"], fin[f"C{i}"], fin[f"G{i}"] = dia, valor, efeito
    rem = wb["itens_Remanesc"]
    rem["A2"], rem["B2"], rem["C2"] = "ITEM-01", 1.0, VU0
    rem["E2"] = rem["G2"] = rem["I2"] = 1.0   # aberturas C1/C2/C3 sem consumo
    garantir_aba_ciclo_em_execucao(wb)
    if com_posicao_fisica:
        cex = wb["CICLO_EM_EXECUCAO"]
        cex["D5"] = date(2026, 1, 15)
        cex["C13"] = 1.0                       # remanescente declarado
    wb.save(destino)
    return destino


@pytest.fixture(scope="module")
def excel():
    client = pytest.importorskip("win32com.client")
    pythoncom = pytest.importorskip("pythoncom")
    pythoncom.CoInitialize()
    app = client.DispatchEx("Excel.Application")
    app.Visible = False
    app.DisplayAlerts = False
    yield app
    app.Quit()
    del app
    pythoncom.CoUninitialize()


def _abrir(excel, caminho: Path):
    book = excel.Workbooks.Open(
        str(caminho), UpdateLinks=0, ReadOnly=False, CorruptLoad=0
    )
    excel.CalculateFullRebuild()
    return book


def test_cenario_real_financeiro_b10_b11_e_retroativo(excel, tmp_path):
    destino = _montar_cenario_financeiro(
        tmp_path / "fin_real.xlsx", com_posicao_fisica=True
    )
    book = _abrir(excel, destino)
    try:
        mem = book.Worksheets("MEMORIA_RESULTADOS")
        res = book.Worksheets("RESULTADOS")
        # Execucao considerada por ciclo (inclui preclusos; E=C sem efeito).
        assert float(mem.Range("W61").Value) == pytest.approx(FIN["c0"], abs=TOL)
        assert float(mem.Range("W62").Value) == pytest.approx(FIN["c1"], abs=TOL)
        assert float(mem.Range("W63").Value) == pytest.approx(FIN["c2"], abs=TOL)
        assert float(mem.Range("W64").Value) == pytest.approx(FIN["c3"], abs=TOL)
        # Historico = C0+C1+C2; o vigente C3 NAO entra (sem dupla contagem).
        assert float(mem.Range("W66").Value) == pytest.approx(HISTORICO, abs=TOL)
        assert float(mem.Range("W67").Value) == pytest.approx(HISTORICO, abs=TOL)
        # FORMA 1 e FORMA 2 completas; reconciliacao zero.
        assert float(mem.Range("W49").Value) == 1
        assert float(res.Range("B10").Value) == pytest.approx(VTA_ESPERADO, abs=TOL)
        assert float(res.Range("B11").Value) == pytest.approx(VTA_ESPERADO, abs=TOL)
        assert float(res.Range("B13").Value) == pytest.approx(0.0, abs=TOL)
        # Retroativo oficial calculado e VISIVEL no card (D5, sob o rotulo).
        assert float(mem.Range("B16").Value) == pytest.approx(RETRO, abs=TOL)
        assert float(res.Range("D22").Value) == pytest.approx(RETRO, abs=TOL)
        assert float(res.Range("D5").Value) == pytest.approx(RETRO, abs=TOL)
        assert "24.678,92" in str(res.Range("D5").Text)
        # A faixa CICLO ATUAL segue funcional via ancora J8.
        assert str(res.Range("J8").Value) == "C3"
        assert "C3" in str(res.Range("C3").Text)
    finally:
        book.Close(SaveChanges=False)


def test_fallback_sem_posicao_fisica_preservado(excel, tmp_path):
    """Sem CICLO_EM_EXECUCAO utilizavel: FORMA 1 vazia; FORMA 2 completa."""
    destino = _montar_cenario_financeiro(
        tmp_path / "fin_sem_foto.xlsx", com_posicao_fisica=False
    )
    book = _abrir(excel, destino)
    try:
        mem = book.Worksheets("MEMORIA_RESULTADOS")
        res = book.Worksheets("RESULTADOS")
        assert float(mem.Range("W49").Value) == 0
        assert mem.Range("W50").Value in (None, "")
        assert res.Range("B10").Value in (None, "")
        # FORMA 2 continua com o historico completo + abertura do vigente.
        assert float(res.Range("B11").Value) == pytest.approx(VTA_ESPERADO, abs=TOL)
    finally:
        book.Close(SaveChanges=False)
