import json
import os
import shutil
import tempfile
import unittest
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from _coleta_reajuste import CAMINHO_MODELO_COLETA, ler_coleta_reajuste
from _coleta_reajuste_documentos import adaptar_coleta_reajuste_para_documentos


FIXTURE = Path(__file__).parent / "fixtures" / "posicao_contratual.json"


@unittest.skipUnless(
    os.environ.get("RUN_EXCEL_INTEGRATION") == "1" and os.name == "nt",
    "Integração permanente opt-in: requer Microsoft Excel e RUN_EXCEL_INTEGRATION=1.",
)
class PosicaoContratualExcelTests(unittest.TestCase):
    @staticmethod
    def _preencher_parametros(wb):
        parametros = wb.Worksheets("parametros")
        for row, ano in zip(range(2, 7), range(2023, 2028)):
            parametros.Range(f"A{row}").Value = "Sim"
            parametros.Range(f"C{row}").Value = datetime(ano, 1, 1)
            parametros.Range(f"D{row}").Value = datetime(ano, 12, 31)
            parametros.Range(f"E{row}").Value = 0.0

    def test_multiplos_aditivos_fracoes_e_novo_item_no_excel_real(self):
        import win32com.client

        casos = json.loads(FIXTURE.read_text(encoding="utf-8"))["casos"]
        with tempfile.TemporaryDirectory() as temp:
            caminho = Path(temp) / "homologacao_posicao.xlsx"
            shutil.copy2(CAMINHO_MODELO_COLETA, caminho)
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            try:
                wb = excel.Workbooks.Open(str(caminho.resolve()), UpdateLinks=0, ReadOnly=False)
                self._preencher_parametros(wb)

                itens = wb.Worksheets("itens_Remanesc")
                itens.Range("A2").Value = casos["multiplos_aditivos_mesmo_ciclo"]["item"]
                itens.Range("B2").Value = 100.0
                itens.Range("C2").Value = 10.0
                for cell in ("E2", "G2", "I2", "K2"):
                    itens.Range(cell).Value = 100.0
                itens.Range("A3").Value = casos["novo_item"]["item"]
                itens.Range("B3").Value = 0.0
                itens.Range("C3").Value = 20.0
                for cell in ("E3", "G3", "I3", "K3"):
                    itens.Range(cell).Value = 0.0

                aditivos = wb.Worksheets("aditivos")
                for row, (tipo, quantidade, considerado) in enumerate(
                    (("Acrescimo", 40.0, "Nao"), ("Acrescimo", 10.0, "Sim"), ("Supressao", 5.0, "Sim")),
                    2,
                ):
                    aditivos.Range(f"A{row}").Value = casos["multiplos_aditivos_mesmo_ciclo"]["item"]
                    aditivos.Range(f"B{row}").Value = datetime(2025, 6, 1)
                    aditivos.Range(f"D{row}").Value = tipo
                    aditivos.Range(f"E{row}").Value = quantidade
                    aditivos.Range(f"K{row}").Value = considerado
                aditivos.Range("A5").Value = casos["novo_item"]["item"]
                aditivos.Range("B5").Value = datetime(2025, 6, 1)
                aditivos.Range("D5").Value = "Acrescimo"
                aditivos.Range("E5").Value = 5.5
                aditivos.Range("K5").Value = "Nao"

                excel.CalculateFullRebuild()
                wb.Save()
                wb.Close(SaveChanges=True)
            finally:
                excel.Quit()

            calculado = load_workbook(caminho, data_only=True, read_only=True)
            posicao = calculado["posicao_contratual"]
            self.assertEqual([posicao[f"{col}2"].value for col in ("E", "I", "M", "Q", "U")], [100, 100, 145, 145, 145])
            self.assertEqual([posicao[f"{col}3"].value for col in ("E", "I", "M", "Q", "U")], [0, 0, 5.5, 5.5, 5.5])
            self.assertEqual(posicao["L2"].value, 45)
            self.assertEqual(posicao["L3"].value, 5.5)
            self.assertEqual(calculado["itens_RC"]["I3"].value, 145)
            self.assertEqual(calculado["itens_RC"]["I4"].value, 5.5)
            self.assertEqual(calculado["historico_VU"]["P2"].value, 145)
            self.assertEqual(calculado["historico_VU"]["P3"].value, 5.5)
            calculado.close()

    def test_alertas_calculados_no_excel_bloqueiam_upload_e_documentos(self):
        import win32com.client

        casos = json.loads(FIXTURE.read_text(encoding="utf-8"))["casos"]
        for chave in ("posicao_negativa", "remanescente_supera_posicao"):
            with self.subTest(chave=chave), tempfile.TemporaryDirectory() as temp:
                caminho = Path(temp) / f"{chave}.xlsx"
                shutil.copy2(CAMINHO_MODELO_COLETA, caminho)
                excel = win32com.client.DispatchEx("Excel.Application")
                excel.Visible = False
                excel.DisplayAlerts = False
                try:
                    wb = excel.Workbooks.Open(str(caminho.resolve()), UpdateLinks=0, ReadOnly=False)
                    self._preencher_parametros(wb)
                    itens = wb.Worksheets("itens_Remanesc")
                    itens.Range("A2").Value = casos[chave]["item"]
                    itens.Range("B2").Value = 100.0
                    itens.Range("C2").Value = 10.0
                    for cell in ("E2", "G2", "I2", "K2"):
                        itens.Range(cell).Value = 100.0
                    if chave == "posicao_negativa":
                        aditivos = wb.Worksheets("aditivos")
                        aditivos.Range("A2").Value = casos[chave]["item"]
                        aditivos.Range("B2").Value = datetime(2025, 6, 1)
                        aditivos.Range("D2").Value = "Supressao"
                        aditivos.Range("E2").Value = 150.0
                    else:
                        itens.Range("G2").Value = 110.0
                    excel.CalculateFullRebuild()
                    wb.Save()
                    wb.Close(SaveChanges=True)
                finally:
                    excel.Quit()

                calculado = load_workbook(caminho, data_only=True, read_only=True)
                self.assertEqual(calculado["posicao_contratual"]["X2"].value, casos[chave]["alerta_esperado"])
                calculado.close()
                payload = caminho.read_bytes()
                diagnostico = ler_coleta_reajuste(payload)
                self.assertFalse(diagnostico["valido"])
                self.assertFalse(diagnostico["pronto_para_consolidar"])
                with self.assertRaises(ValueError):
                    adaptar_coleta_reajuste_para_documentos(payload)


if __name__ == "__main__":
    unittest.main()
