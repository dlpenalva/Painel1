"""Integracao WEB: XLS -> processar_coleta_oficial_runtime -> motor -> UI.

Prova o CAMINHO DA APLICACAO (nao apenas motor sobre dict sintetico):
  * o runtime real executa `montar_cobertura_temporal` sobre o MESMO payload
    do leitor e persiste em `diagnostico['cobertura_temporal']`;
  * falha isolada do motor NAO derruba o restante da coleta (fail-safe);
  * arquivo sem a aba continua aceito;
  * o estado temporal vive dentro de `diagnostico` (limpo pelos resets);
  * `resumo_cobertura_temporal` formata os cenarios funcionais (dd/mm/aaaa,
    "Nao informado/confirmado/aplicavel", evidencia != cobertura).
"""
from __future__ import annotations

import io
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from _coleta_reajuste_documentos import processar_coleta_oficial_runtime
from _leitor_masterfile_v10 import ler_masterfile_v10
from _motor_cobertura_temporal import montar_cobertura_temporal
from _ui_utils import resumo_cobertura_temporal

ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "templates" / "COLETA_REAJUSTE_OFICIAL.xlsx"

POR_CICLO = {
    "C0": {"data_inicio": date(2023, 1, 1), "data_fim": date(2023, 12, 31), "fator_acumulado": 1.0},
    "C1": {"data_inicio": date(2024, 1, 1), "data_fim": date(2024, 12, 31), "fator_acumulado": 1.10},
    "C2": {"data_inicio": date(2025, 1, 1), "data_fim": date(2025, 12, 31), "fator_acumulado": 1.21},
    "C3": {"data_inicio": date(2026, 3, 1), "data_fim": date(2027, 2, 28), "fator_acumulado": 1.33},
}


# --------------------------------------------------------------------------- #
# Helpers de construcao do XLS oficial preenchido.
# --------------------------------------------------------------------------- #
def _xls(*, pc=None, gcc_pc=None, financeiro=None, com_aba=True,
         data_corte=None, remanescente=None, itens=None) -> bytes:
    wb = load_workbook(TEMPLATE, data_only=False)
    p = wb["parametros"]
    for n, (i, f) in enumerate([
            (date(2023, 1, 1), date(2023, 12, 31)),
            (date(2024, 1, 1), date(2024, 12, 31)),
            (date(2025, 1, 1), date(2025, 12, 31)),
            (date(2026, 1, 1), date(2026, 12, 31))]):
        p.cell(n + 2, 3).value = i
        p.cell(n + 2, 4).value = f
    c = wb["CONTROLE"]
    c["B1"], c["B2"] = "Pedidos de Compras", "C3"
    ir = wb["itens_Remanesc"]
    for idx, cod in enumerate(itens or ["ITEM-1"], start=2):
        ir.cell(idx, 1).value = cod
        ir.cell(idx, 2).value = 100.0
        ir.cell(idx, 3).value = 10.0
        ir.cell(idx, 5).value = ir.cell(idx, 7).value = ir.cell(idx, 9).value = 100.0
    if data_corte is not None:
        wb["CONTROLE"]["B3"] = data_corte
    if remanescente:
        pr = wb["posicao_referencia"]
        rowmap = {cod: i for i, cod in enumerate(itens or ["ITEM-1"], start=2)}
        for cod, q in remanescente.items():
            pr.cell(rowmap[cod], 2).value = q
    if pc:
        w = wb["itens_PC"]
        for i, (num, dt, val) in enumerate(pc, start=2):
            w.cell(i, 1).value, w.cell(i, 2).value, w.cell(i, 4).value = num, dt, val
    if financeiro:
        w = wb["financeiro"]
        for i, (comp, val) in enumerate(financeiro, start=2):
            w.cell(i, 1).value, w.cell(i, 3).value = comp, val
    if com_aba:
        if gcc_pc is not None:
            wb["cobertura_temporal"]["B15"] = gcc_pc
        wb["cobertura_temporal"]["B4"] = date(2026, 6, 30)
    else:
        del wb["cobertura_temporal"]
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --------------------------------------------------------------------------- #
# A + B — o runtime executa o motor sobre o MESMO payload do leitor.
# --------------------------------------------------------------------------- #
def test_runtime_executa_motor_mesmo_payload():
    data = _xls(pc=[("PC-1", date(2026, 5, 20), 2000.0)], gcc_pc=date(2026, 5, 31))
    resultado, diagnostico = processar_coleta_oficial_runtime(data)
    ct = diagnostico.get("cobertura_temporal")
    assert isinstance(ct, dict) and "bloco_a_marcos" in ct        # motor rodou no runtime
    esperado = montar_cobertura_temporal(ler_masterfile_v10(data)).to_dict()
    assert ct == esperado                                         # mesma fonte/payload
    # o estado vive dentro de `diagnostico` (chave limpa pelos resets do upload).
    assert resultado["diagnostico_coleta"]["cobertura_temporal"] == ct


# --------------------------------------------------------------------------- #
# E — fail-safe: erro do motor NAO derruba a coleta nem os documentos.
# --------------------------------------------------------------------------- #
def test_runtime_fail_safe_nao_derruba_coleta(monkeypatch):
    def _explode(*a, **k):
        raise RuntimeError("falha simulada do motor temporal")
    monkeypatch.setattr("_motor_cobertura_temporal.montar_cobertura_temporal", _explode)
    data = _xls(pc=[("PC-1", date(2026, 5, 20), 2000.0)])
    resultado, diagnostico = processar_coleta_oficial_runtime(data)
    assert diagnostico["cobertura_temporal"] == {
        "ok": False, "erro": "falha simulada do motor temporal"}
    # coleta e documentos continuam disponiveis (nao houve hard-reject).
    assert resultado is not None
    assert "capacidades" in resultado
    assert resumo_cobertura_temporal(diagnostico["cobertura_temporal"])["disponivel"] is False


# --------------------------------------------------------------------------- #
# D — arquivo sem a aba cobertura_temporal continua aceito.
# --------------------------------------------------------------------------- #
def test_runtime_compat_sem_aba():
    data = _xls(pc=[("PC-1", date(2026, 5, 20), 2000.0)], com_aba=False)
    resultado, diagnostico = processar_coleta_oficial_runtime(data)
    ct = diagnostico.get("cobertura_temporal")
    assert "bloco_a_marcos" in ct                                 # motor roda sem a aba
    assert ct["bloco_a_marcos"]["data_analise"] is None           # sem aba => sem data
    assert resultado is not None


# --------------------------------------------------------------------------- #
# C — sem estado fantasma: arquivo B nao herda dados do arquivo A.
# --------------------------------------------------------------------------- #
def test_runtime_sem_estado_fantasma():
    data_a = _xls(pc=[("PC-1", date(2026, 5, 20), 2000.0)])       # com PC
    _, diag_a = processar_coleta_oficial_runtime(data_a)
    data_b = _xls(pc=None, com_aba=False)                          # sem PC, sem aba
    _, diag_b = processar_coleta_oficial_runtime(data_b)
    assert diag_a["cobertura_temporal"]["bloco_b_cobertura"]["pc_ultima_evidencia"] == "2026-05-20"
    assert diag_b["cobertura_temporal"]["bloco_b_cobertura"]["pc_ultima_evidencia"] is None
    assert diag_a["cobertura_temporal"] != diag_b["cobertura_temporal"]


# --------------------------------------------------------------------------- #
# WEB6 runtime — fotografia fisica recente COMPLETA flui pelo runtime real.
# --------------------------------------------------------------------------- #
def test_runtime_fotografia_recente_completa():
    data = _xls(data_corte=date(2026, 5, 31), remanescente={"ITEM-1": 40.0})
    _, diag = processar_coleta_oficial_runtime(data)
    r = resumo_cobertura_temporal(diag.get("cobertura_temporal"))
    assert r["posicao_fisica_utilizada"] == "Fotografia atual informada"
    assert r["data_posicao_fisica"] == "31/05/2026"


def test_runtime_fotografia_parcial_fallback():
    data = _xls(itens=["ITEM-1", "ITEM-2"], data_corte=date(2026, 5, 31),
               remanescente={"ITEM-1": 40.0})   # ITEM-2 sem qtd -> incompleta
    _, diag = processar_coleta_oficial_runtime(data)
    r = resumo_cobertura_temporal(diag.get("cobertura_temporal"))
    assert "incompleta" in r["posicao_fisica_utilizada"].lower()   # fallback global


# --------------------------------------------------------------------------- #
# Cenarios funcionais (Secao 18) via resumo de exibicao (dd/mm/aaaa, terminologia).
# --------------------------------------------------------------------------- #
def _res(**kw):
    return {
        "controle": {"data_corte": kw.get("data_corte"), "ciclo_vigente": "C3",
                     "data_analise": kw.get("data_analise", date(2026, 6, 30))},
        "por_ciclo": POR_CICLO,
        "itens_base": list(kw.get("itens_base", ("ITEM-1",))),
        "remanescente_atual": dict(kw.get("remanescente", {})),
        "fotografias_ciclo": list(kw.get("fotografias", ("C1", "C2", "C3"))),
        "financeiro": list(kw.get("financeiro", [])),
        "itens_pc": list(kw.get("itens_pc", [])),
        "confirmacao_gcc": dict(kw.get("confirmacao_gcc", {})),
    }


def _resumo(**kw):
    return resumo_cobertura_temporal(montar_cobertura_temporal(_res(**kw)).to_dict())


def test_web1_somente_abertura_c3():
    r = _resumo()
    assert r["posicao_fisica_utilizada"] == "Abertura do C3"
    assert r["data_posicao_fisica"] == "01/03/2026"       # nao inventa junho
    assert r["modo_temporal"] == "POSICAO_DE_CORTE"


def test_web2_pc_sem_confirmacao():
    r = _resumo(itens_pc=[{"numero_pc": "PC", "data_pc": date(2026, 5, 20), "valor_pc": 1.0}])
    assert r["pc_ultima_evidencia"] == "20/05/2026"
    assert r["pc_cobertura_confirmada"] == "Não confirmado"   # maio nao vira cobertura
    assert r["data_posicao_fisica"] == "01/03/2026"


def test_web3_pc_com_confirmacao():
    r = _resumo(itens_pc=[{"numero_pc": "PC", "data_pc": date(2026, 5, 20), "valor_pc": 1.0}],
                confirmacao_gcc={"pc_ate": date(2026, 5, 31)})
    assert r["pc_cobertura_confirmada"] == "31/05/2026"
    assert r["projecao_a_partir_de"] == "01/06/2026"


def test_web4_financeiro_zero():
    fin = [{"competencia": date(2026, 3, 1), "valor": 10000.0},
           {"competencia": date(2026, 4, 1), "valor": 0.0},
           {"competencia": date(2026, 5, 1), "valor": 12000.0}]
    r = _resumo(financeiro=fin)
    assert r["financeiro_ultima_competencia"] == "05/2026"
    assert r["financeiro_cobertura_adotada"] == "31/05/2026"
    assert r["financeiro_origem_cobertura"] == "Inferida pela sequência mensal"
    assert r["projecao_a_partir_de"] == "01/06/2026"


def test_web5_financeiro_lacuna():
    fin = [{"competencia": date(2026, 3, 1), "valor": 10000.0},
           {"competencia": date(2026, 4, 1), "valor": None},
           {"competencia": date(2026, 5, 1), "valor": 12000.0}]
    r = _resumo(financeiro=fin)
    assert r["financeiro_ultima_competencia"] == "05/2026"
    assert r["financeiro_cobertura_adotada"] == "31/03/2026"   # lacuna corta
    assert r["projecao_a_partir_de"] == "01/04/2026"


def test_web6_fotografia_atual_completa():
    r = _resumo(data_corte=date(2026, 5, 31), remanescente={"ITEM-1": 40.0})
    assert r["posicao_fisica_utilizada"] == "Fotografia atual informada"
    assert r["data_posicao_fisica"] == "31/05/2026"


def test_web7_fotografia_parcial_fallback():
    r = _resumo(data_corte=date(2026, 5, 31), itens_base=("ITEM-1", "ITEM-2"),
                remanescente={"ITEM-1": 40.0})   # ITEM-2 sem qtd -> incompleta
    assert "incompleta" in r["posicao_fisica_utilizada"].lower()
    assert r["data_posicao_fisica"] == "01/03/2026"           # fallback global


def test_web8_multifonte_conferencia():
    r = _resumo(
        itens_pc=[{"numero_pc": "PC", "data_pc": date(2026, 4, 10), "valor_pc": 5000.0}],
        financeiro=[{"competencia": date(2026, 4, 1), "valor": 5000.0}])
    assert r["fonte_principal"] == "Financeiro"
    assert r["fontes_conferencia"] == "PC"                    # sem dupla contagem


def test_resumo_nunca_exibe_data_falsa():
    """Ausencias viram texto, nunca 0/01/01/1900/None/nan."""
    r = resumo_cobertura_temporal({"ok": False, "erro": "x"})
    assert r["disponivel"] is False
    vazio = _resumo()  # sem financeiro/PC
    for campo in ("financeiro_ultima_competencia", "pc_ultima_evidencia"):
        assert vazio[campo] == "Não informado"
    assert vazio["pc_cobertura_confirmada"] == "Não confirmado"
