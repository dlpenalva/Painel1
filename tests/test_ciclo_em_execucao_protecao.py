# -*- coding: utf-8 -*-
"""Protecao de CICLO_EM_EXECUCAO: campos manuais editaveis, formulas bloqueadas.

Regressao da correcao de editabilidade: o gerador definia
selectUnlockedCells=True, que em OOXML DESABILITA a selecao das celulas
desbloqueadas — impedindo o fiscal de selecionar/editar D5 e C13:C211 embora
estivessem unlocked. Correcao: selectUnlockedCells=False.

Requisitos verificados:
1. D5 selecionavel e editavel;
2. C13:C211 selecionavel e editavel;
3. formulas/celulas automaticas continuam protegidas (locked);
4. validacoes de dados permanecem;
5. protecao da planilha nao e removida integralmente;
6. edicao de formula bloqueada (Excel real recusa gravar em celula locked).
"""
from __future__ import annotations

import datetime as _dt
import os
import shutil
import tempfile
from pathlib import Path

import pytest
from openpyxl import load_workbook

from _ciclo_em_execucao import (
    ABA_CICLO_EM_EXECUCAO,
    PRIMEIRA_LINHA_ITEM,
    ULTIMA_LINHA_ITEM,
    garantir_aba_ciclo_em_execucao,
)

ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "templates" / "COLETA_REAJUSTE_OFICIAL.xlsx"


def _gerar_com_aba(destino: Path) -> Path:
    wb = load_workbook(TEMPLATE)
    garantir_aba_ciclo_em_execucao(wb)
    wb.save(destino)
    return destino


@pytest.fixture(scope="module")
def arquivo_com_aba(tmp_path_factory) -> Path:
    destino = tmp_path_factory.mktemp("ciclo_prot") / "coleta_ciclo.xlsx"
    return _gerar_com_aba(destino)


# --------------------------------------------------------------------------- #
# Configuracao estatica da protecao (openpyxl)                                 #
# --------------------------------------------------------------------------- #
def test_protecao_campos_manuais_desbloqueados(arquivo_com_aba):
    ws = load_workbook(arquivo_com_aba)[ABA_CICLO_EM_EXECUCAO]
    # Planilha protegida, mas com selecao de celulas desbloqueadas HABILITADA.
    assert ws.protection.sheet is True
    assert ws.protection.selectUnlockedCells is False, (
        "selectUnlockedCells=True desabilitaria a selecao dos campos manuais"
    )
    # D5 e toda a coluna C de itens desbloqueadas (editaveis).
    assert ws["D5"].protection.locked is False
    assert ws[f"C{PRIMEIRA_LINHA_ITEM}"].protection.locked is False
    assert ws[f"C{ULTIMA_LINHA_ITEM}"].protection.locked is False


def test_formulas_permanecem_bloqueadas(arquivo_com_aba):
    ws = load_workbook(arquivo_com_aba)[ABA_CICLO_EM_EXECUCAO]
    # Celulas automaticas (formula) continuam locked (default True).
    for coord in ("A9", f"B{PRIMEIRA_LINHA_ITEM}", f"D{PRIMEIRA_LINHA_ITEM}",
                  f"G{PRIMEIRA_LINHA_ITEM}"):
        assert ws[coord].protection.locked in (True, None)


def test_validacoes_de_dados_permanecem(arquivo_com_aba):
    ws = load_workbook(arquivo_com_aba)[ABA_CICLO_EM_EXECUCAO]
    sqrefs = {str(dv.sqref) for dv in ws.data_validations.dataValidation}
    # Data da posicao (D5) e quantidade remanescente (C13:C211).
    assert any("D5" in s for s in sqrefs)
    assert any(
        f"C{PRIMEIRA_LINHA_ITEM}" in s and f"C{ULTIMA_LINHA_ITEM}" in s
        for s in sqrefs
    )


# --------------------------------------------------------------------------- #
# Teste no Excel real: preencher, salvar, fechar, reabrir, conferir            #
# --------------------------------------------------------------------------- #
@pytest.mark.skipif(
    os.environ.get("RUN_EXCEL_INTEGRATION") != "1",
    reason="defina RUN_EXCEL_INTEGRATION=1 para executar Excel COM",
)
def test_edicao_real_excel_preserva_dados_e_bloqueia_formula(arquivo_com_aba):
    client = pytest.importorskip("win32com.client")
    pythoncom = pytest.importorskip("pythoncom")
    import pywintypes

    tmp_dir = Path(tempfile.mkdtemp(prefix="ciclo_edit_"))
    tmp = tmp_dir / "coleta.xlsx"
    shutil.copyfile(arquivo_com_aba, tmp)

    linha = PRIMEIRA_LINHA_ITEM
    data_valor = _dt.datetime(2027, 6, 15)  # datetime real -> Excel grava data
    qtd = 12.5

    pythoncom.CoInitialize()
    excel = client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        wb = excel.Workbooks.Open(str(tmp), UpdateLinks=0, ReadOnly=False, CorruptLoad=0)
        ws = wb.Worksheets(ABA_CICLO_EM_EXECUCAO)
        # (5) protecao NAO removida integralmente.
        assert bool(ws.ProtectContents) is True
        # (1)(2) campos manuais editaveis sob protecao (locked=False grava OK).
        assert bool(ws.Range("D5").Locked) is False
        assert bool(ws.Range(f"C{linha}").Locked) is False
        ws.Range("D5").Value = data_valor
        ws.Range(f"C{linha}").Value = qtd
        # (3)(6) formula continua bloqueada: gravar em celula locked deve falhar.
        assert bool(ws.Range(f"G{linha}").Locked) is True
        travou = False
        try:
            ws.Range(f"G{linha}").Value = 999
        except pywintypes.com_error:
            travou = True
        assert travou, "celula de formula deveria estar protegida contra edicao"
        wb.Save()
        wb.Close(SaveChanges=False)

        # Reabrir sem reparo e conferir persistencia.
        wb2 = excel.Workbooks.Open(str(tmp), UpdateLinks=0, ReadOnly=True, CorruptLoad=0)
        ws2 = wb2.Worksheets(ABA_CICLO_EM_EXECUCAO)
        assert ws2.Range(f"C{linha}").Value == qtd
        d5 = ws2.Range("D5").Value
        # Excel devolve a data como datetime COM.
        assert getattr(d5, "day", None) == 15 and getattr(d5, "month", None) == 6
        assert bool(ws2.ProtectContents) is True
        wb2.Close(SaveChanges=False)
    finally:
        excel.Quit()
        del excel
        pythoncom.CoUninitialize()
        shutil.rmtree(tmp_dir, ignore_errors=True)
