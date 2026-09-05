# -*- coding: utf-8 -*-
"""XLS-PC-VTA-ALIGN-1 — a regra do VTA-POT-1 em TODA area visivel do metodo PC.

    VTA OFICIAL = VTA SEM POTENCIAL + RETROATIVO POTENCIAL INCORPORADO
    RETROATIVO CONSIDERADO NO VTA = RECONHECIDO + POTENCIAL INCORPORADO

O VTA-POT-1 (PRs #144/#145) ja provava a regra no motor e no quadro 9. Esta
trava cobre a PROPAGACAO: itens_PC (cabecalho de R e o fechamento da linha
19), o bloco superior da RESULTADOS (cards) e o bloco 6 (medidas 6/7/8).

Invariantes que sustentam a auditoria:

* a parcela potencial e SOMADA ao VTA uma unica vez, em
  ``MEMORIA_RESULTADOS!T25``; todas as demais ocorrencias apenas LEEM o
  name canonico ``RETROATIVO_POTENCIAL_VTA``;
* nenhuma area visivel recalcula o potencial por conta propria;
* piso prudencial intacto (``T39 = MAX(T41,0)``): potencial negativo nao
  reduz o VTA;
* fora do metodo PC toda celula nova devolve "" — Financeiro e Itens
  Consumidos nao ganham linha, cor nem valor.
"""
from __future__ import annotations

import io
import os
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

RAIZ = Path(__file__).resolve().parents[1]
if str(RAIZ) not in sys.path:
    sys.path.insert(0, str(RAIZ))

from _coleta_oficial import (  # noqa: E402
    COLUNAS_ITENS_PC_OFICIAL,
    obter_coleta_oficial_bytes,
)

TEMPLATE = RAIZ / "templates" / "COLETA_REAJUSTE_OFICIAL.xlsx"

R2_ESPERADO = "VALOR EM ANÁLISE - NÃO PAGOS (ÁREA GEST.)"
R11_ESPERADO = "Valor em análise — não pagos (regra vigente)"

# Toda celula do produto que le a parcela potencial incorporada ao VTA.
# MEMORIA_RESULTADOS!T25 e a UNICA que a SOMA ao VTA; as demais apresentam.
REFERENCIAS_POTENCIAL_VTA = {
    "MEMORIA_RESULTADOS!T25",       # o proprio VTA-PC (unica soma)
    "itens_PC!M19", "itens_PC!S19",
    "RESULTADOS!A6", "RESULTADOS!C8", "RESULTADOS!E8",
    "RESULTADOS!B61", "RESULTADOS!B62",
    "RESULTADOS!B84", "RESULTADOS!C86",
}

ROTULOS_BLOCO6 = [
    "1. Total cadastrado de Pedidos de Compra",
    "2. Total até a data de corte",
    "3. Total distribuído nos ciclos",
    "4. Total com efeito financeiro",
    "5. Total sem efeito financeiro",
    "6. Retroativo reconhecido",
    "7. Retroativo potencial — POTENCIAL",
    "8. Retroativo considerado no VTA",
    "9. Execução do ciclo atual",
    "10. Saldo remanescente atual",
    "11. VTA oficial",
    "12. Diferença entre as formas de cálculo",
    "13. Resultado da apuração",
]


@pytest.fixture(scope="module")
def wb():
    livro = load_workbook(TEMPLATE, data_only=False)
    yield livro
    livro.close()


@pytest.fixture(scope="module")
def wb_runtime():
    livro = load_workbook(io.BytesIO(obter_coleta_oficial_bytes()),
                          data_only=False)
    yield livro
    livro.close()


# ------------------------------------------------------------------ itens_PC
def test_cabecalho_de_r_diz_nao_pagos_nos_dois_quadros(wb, wb_runtime):
    """itens_PC!I zera quando o PC foi pago: R e o valor dos NAO pagos."""
    for livro in (wb, wb_runtime):
        ws = livro["itens_PC"]
        assert ws["R2"].value == R2_ESPERADO
        assert ws["R11"].value == R11_ESPERADO
    assert '=IF(G2="Sim",0,' in str(wb["itens_PC"]["I2"].value)


def test_faixa_operacional_a_l_permanece_intacta(wb):
    ws = wb["itens_PC"]
    assert [ws.cell(1, c).value for c in range(1, 13)] == COLUNAS_ITENS_PC_OFICIAL


def test_linha_19_fecha_o_retroativo_sem_inserir_linha(wb):
    """A linha 19 ja existia vazia, logo antes de COMO OS PCs SAO TRATADOS."""
    ws = wb["itens_PC"]
    assert ws["M20"].value == "COMO OS PCs SÃO TRATADOS"
    merges = {str(m) for m in ws.merged_cells.ranges if m.min_row == 19}
    assert merges == {"M19:R19", "S19:T19"}

    rotulo = str(ws["M19"].value)
    assert rotulo.startswith('="RETROATIVO CONSIDERADO NO VTA = RECONHECIDO')
    assert "N($Q$18)" in rotulo                      # reconhecido (Quadro 2)
    assert "MEMORIA_RESULTADOS!$T$39" in rotulo      # potencial incorporado
    assert "POTENCIAL INCORPORADO" in rotulo

    # O fechamento soma o potencial CANONICO, nunca S18: S18 e o potencial de
    # TODOS os ciclos ate o corte (inclusive o vigente e o residual), e o VTA
    # so incorpora o dos ciclos ja encerrados. Fechar por S18 publicaria um
    # total que nao existe em lugar nenhum do VTA.
    assert ws["S19"].value == (
        "=ROUND(N($Q$18)+N(MEMORIA_RESULTADOS!$T$39),2)"
    )
    assert ws["Q18"].value == "=SUM(Q12:Q17)"
    assert ws["S18"].value == "=SUM(S12:S17)"


def test_colunas_tecnicas_v_ac_seguem_ocultas(wb, wb_runtime):
    """Template e copia entregue: V:AC em UM grupo, oculto e sem divisao."""
    for livro in (wb, wb_runtime):
        ocultas = {
            letra: (dim.min, dim.max)
            for letra, dim in livro["itens_PC"].column_dimensions.items()
            if dim.hidden
        }
        assert ocultas == {"V": (22, 29)}


# ------------------------------------------- RESULTADOS: bloco superior (3:8)
def test_card_do_vta_declara_que_o_valor_inclui_o_potencial(wb):
    res = wb["RESULTADOS"]
    a4 = str(res["A4"].value)
    assert a4.startswith('=IF(MEMORIA_RESULTADOS!$B$4="PCs",')
    assert (
        "VTA OFICIAL — inclui retroativo reconhecido + retroativo potencial"
        in a4
    )
    assert '"VTA OFICIAL")' in a4                     # demais metodos intactos
    # O valor do card continua sendo o VTA canonico, sem recalculo.
    assert res["C5"].value == '=IF(VTA_FINAL="","",VTA_FINAL)'
    assert "RETROATIVO_POTENCIAL_VTA" in str(res["A6"].value)


def test_card_do_retroativo_nao_publica_potencial_como_valor_a_pagar(wb):
    """Reconhecido + potencial JAMAIS sob um rotulo de obrigacao constituida."""
    res = wb["RESULTADOS"]
    assert res["D4"].value == "RETROATIVO RECONHECIDO A PAGAR"
    assert res["D5"].value == "=$D$22"                # so o reconhecido


def test_linha_8_traz_o_card_ambar_e_o_fechamento(wb):
    res = wb["RESULTADOS"]
    assert res["A8"].value == (
        '=IF(MEMORIA_RESULTADOS!$B$4<>"PCs","",'
        '"RETROATIVO POTENCIAL — POTENCIAL (incorporado ao VTA)")'
    )
    assert res["C8"].value == (
        '=IF(MEMORIA_RESULTADOS!$B$4<>"PCs","",'
        'IFERROR(RETROATIVO_POTENCIAL_VTA,""))'
    )
    assert res["D8"].value == (
        '=IF(MEMORIA_RESULTADOS!$B$4<>"PCs","",'
        '"RETROATIVO CONSIDERADO NO VTA")'
    )
    assert res["E8"].value == (
        '=IF(OR(MEMORIA_RESULTADOS!$B$4<>"PCs",$D$22=""),"",'
        'ROUND(N($D$22)+N(RETROATIVO_POTENCIAL_VTA),2))'
    )
    # H8 (selo tecnico do VTA) segue na mesma celula, sem ser tocado.
    assert str(res["H8"].value).startswith("=IF(OR(VTA_FINAL")
    # A faixa nova nao criou merge nem deslocou nada.
    assert not [m for m in res.merged_cells.ranges if m.min_row <= 8 <= m.max_row]


def test_ambar_so_na_parcela_potencial(wb):
    """O VTA nunca fica ambar; o fechamento tambem nao (nao e so potencial)."""
    sqrefs = {
        str(regra.sqref) for regra in wb["RESULTADOS"].conditional_formatting
    }
    assert "A8:C8" in sqrefs         # potencial (bloco superior)
    assert "A61:B61" in sqrefs       # medida 7 do bloco 6
    assert "A84:C84" in sqrefs       # quadro 9 (VTA-POT-1)
    assert "C5" not in sqrefs and "B86" not in sqrefs


# ----------------------------------------------- RESULTADOS: bloco 6 (55:67)
def test_bloco6_demonstra_reconhecido_potencial_e_considerado(wb):
    res = wb["RESULTADOS"]
    assert [res["A%d" % linha].value
            for linha in range(55, 68)] == ROTULOS_BLOCO6
    assert res["B60"].value == "=$D$22"
    assert res["B61"].value == (
        '=IF($B$5<>"PCs","",IFERROR(RETROATIVO_POTENCIAL_VTA,""))'
    )
    assert res["B62"].value == (
        '=IF(OR($B$5<>"PCs",$D$22=""),"",'
        'ROUND(N($D$22)+N(RETROATIVO_POTENCIAL_VTA),2))'
    )
    assert res["B65"].value == '=IF(VTA_FINAL="","",VTA_FINAL)'
    assert res["B67"].value == "=$B$3"
    assert "C67:H67" in {str(m) for m in res.merged_cells.ranges}


def test_bloco6_nao_ultrapassa_a_linha_67(wb):
    """O bloco 7 comeca na 68 e a aba continua terminando na 87."""
    res = wb["RESULTADOS"]
    assert res["A68"].value == "7. METODOLOGIA DO VTA"
    assert res.max_row == 87
    assert not [
        celula.coordinate
        for linha in res.iter_rows(min_row=88, max_row=200)
        for celula in linha
        if celula.value is not None
    ]


# ------------------------------------------------------------ bloco 9 e regra
def test_bloco9_nomeia_as_parcelas_sem_mover_ancoras(wb):
    res = wb["RESULTADOS"]
    a81 = str(res["A81"].value)
    assert a81.startswith('=IF(MEMORIA_RESULTADOS!$B$4="PCs",')
    assert "já inclui o retroativo reconhecido" in a81
    assert "retroativo POTENCIAL incorporado" in a81
    assert "ja com o retroativo reconhecido" in str(res["C83"].value)
    # Ancoras publicadas: mesmas coordenadas de sempre.
    nomes = {n: d.value for n, d in wb.defined_names.items()}
    assert nomes["EXECUTADO_APURADO"] == "RESULTADOS!$B$83"
    assert nomes["AJUSTES_DEVIDOS"] == "RESULTADOS!$B$84"
    assert nomes["CONFERENCIA_FORMACAO_VTA"] == "RESULTADOS!$B$87"
    assert nomes["VTA_FINAL"] == "MEMORIA_RESULTADOS!$B$26"
    assert nomes["RETROATIVO_POTENCIAL_VTA"] == "MEMORIA_RESULTADOS!$T$39"
    assert nomes["VTA_SEM_POTENCIAL"] == "MEMORIA_RESULTADOS!$T$40"


def test_o_potencial_e_somado_ao_vta_exatamente_uma_vez(wb):
    """Fonte unica: so T25 SOMA; o resto LE o name canonico."""
    encontradas = {
        "%s!%s" % (ws.title, celula.coordinate)
        for ws in wb.worksheets
        for linha in ws.iter_rows()
        for celula in linha
        if isinstance(celula.value, str) and celula.value.startswith("=")
        and ("$T$39" in celula.value
             or "RETROATIVO_POTENCIAL_VTA" in celula.value)
    }
    assert encontradas == REFERENCIAS_POTENCIAL_VTA
    t25 = str(wb["MEMORIA_RESULTADOS"]["T25"].value)
    assert t25.count("$T$39") == 1
    assert t25.endswith("ROUND($T$21+$T$22+$T$23+$T$39,2))")
    # A conferencia do quadro 9 continua fechando pela soma das parcelas.
    assert wb["RESULTADOS"]["B87"].value == (
        '=IF(OR($B$83="",$B$85="",$B$86=""),"",'
        'ROUND($B$86-($B$83+N($B$84)+$B$85),2))'
    )


def test_piso_prudencial_do_potencial_negativo_permanece(wb):
    mem = wb["MEMORIA_RESULTADOS"]
    assert mem["T39"].value == "=MAX($T$41,0)"
    assert str(mem["T41"].value).startswith("=IF($T$20=")


def test_financeiro_e_consumidos_nao_herdam_a_parcela(wb):
    """Nenhuma celula das abas economicas le a parcela potencial."""
    for aba in ("financeiro", "itens_Consumidos"):
        assert not [
            celula.coordinate
            for linha in wb[aba].iter_rows()
            for celula in linha
            if isinstance(celula.value, str)
            and ("$T$39" in celula.value
                 or "POTENCIAL" in celula.value.upper())
        ]
    # Toda celula nova da RESULTADOS sai vazia fora do metodo PC.
    res = wb["RESULTADOS"]
    for endereco in ("A6", "A8", "C8", "D8", "E8", "B61", "B62"):
        assert '<>"PCs"' in str(res[endereco].value), endereco


# ------------------------------------------------- Excel real (opt-in, focal)
@pytest.mark.skipif(
    os.environ.get("RUN_EXCEL_INTEGRATION") != "1",
    reason="defina RUN_EXCEL_INTEGRATION=1 para executar Excel COM",
)
def test_excel_real_fecha_a_regra_com_potencial_positivo_e_negativo(tmp_path):
    """Um XLS PC por cenario: recalculo total, Save, Close e Reopen."""
    import shutil
    from datetime import date

    client = pytest.importorskip("win32com.client")
    pythoncom = pytest.importorskip("pythoncom")

    def montar(destino, variacao_c1):
        shutil.copy2(TEMPLATE, destino)
        livro = load_workbook(destino)
        ctl = livro["CONTROLE"]
        ctl["B1"], ctl["B2"], ctl["B3"] = "PCs", "C2", date(2027, 6, 30)
        par = livro["parametros"]
        par["A2"], par["C2"], par["D2"] = ("Nao", date(2025, 2, 1),
                                           date(2026, 1, 31))
        par["I2"] = date(2025, 2, 1)
        par["A3"], par["C3"], par["D3"] = ("Sim", date(2026, 2, 1),
                                           date(2027, 1, 31))
        par["E3"] = variacao_c1
        par["H3"] = par["I3"] = date(2026, 2, 1)
        par["A4"], par["C4"], par["D4"] = ("Sim", date(2027, 2, 1),
                                           date(2028, 1, 31))
        par["E4"] = 0.05
        par["H4"] = par["I4"] = date(2027, 2, 1)
        rem = livro["itens_Remanesc"]
        rem["A2"], rem["B2"], rem["C2"] = "ITEM-01", 1000, 10.0
        rem["E2"], rem["G2"] = 800, 700
        pcs = livro["itens_PC"]
        for linha, (numero, dia, valor, pago) in enumerate([
            ("PC-C1-PAGO", date(2026, 3, 1), 1000.0, "Sim"),
            ("PC-C1-ANALISE", date(2026, 4, 1), 2000.0, "Nao"),
            ("PC-C2-ANALISE", date(2027, 3, 1), 4000.0, "Nao"),
        ], start=2):
            pcs["A%d" % linha] = numero
            pcs["B%d" % linha] = dia
            pcs["D%d" % linha] = valor
            pcs["G%d" % linha] = pago
        livro.save(destino)
        return destino

    def medir(caminho):
        pythoncom.CoInitialize()
        excel = client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        livro = None
        try:
            livro = excel.Workbooks.Open(str(caminho), UpdateLinks=0,
                                         ReadOnly=False, CorruptLoad=0)
            excel.CalculateFullRebuild()
            livro.Save()
            livro.Close(False)
            livro = excel.Workbooks.Open(str(caminho), UpdateLinks=0,
                                         ReadOnly=True, CorruptLoad=0)
            res = livro.Worksheets("RESULTADOS")
            mem = livro.Worksheets("MEMORIA_RESULTADOS")
            pcs = livro.Worksheets("itens_PC")
            enderecos = {
                "q18": (pcs, "Q18"), "s18": (pcs, "S18"), "s19": (pcs, "S19"),
                "t39": (mem, "T39"), "t40": (mem, "T40"), "t41": (mem, "T41"),
                "vta": (mem, "B26"), "c5": (res, "C5"), "c8": (res, "C8"),
                "e8": (res, "E8"), "d22": (res, "D22"), "b61": (res, "B61"),
                "b62": (res, "B62"), "b65": (res, "B65"), "b84": (res, "B84"),
                "b86": (res, "B86"), "b87": (res, "B87"),
            }
            return {nome: float(aba.Range(endereco).Value)
                    for nome, (aba, endereco) in enderecos.items()}
        finally:
            if livro is not None:
                try:
                    livro.Close(False)
                except Exception:  # noqa: BLE001 - encerramento best-effort
                    pass
            excel.Quit()
            pythoncom.CoUninitialize()

    positivo = medir(montar(tmp_path / "pc_potencial_positivo.xlsx", 0.05))
    # A regra do VTA, no proprio Excel.
    assert positivo["vta"] == pytest.approx(positivo["t40"] + positivo["t39"])
    assert positivo["t39"] == pytest.approx(100.0)
    # Um unico VTA em toda area visivel.
    for chave in ("c5", "b65", "b86"):
        assert positivo[chave] == pytest.approx(positivo["vta"]), chave
    # Somado uma unica vez: a conferencia do quadro 9 fecha em zero.
    assert positivo["b87"] == pytest.approx(0.0)
    assert positivo["b84"] == pytest.approx(positivo["t39"])
    # O fechamento "considerado no VTA" bate nas tres areas visiveis.
    considerado = positivo["d22"] + positivo["t39"]
    for chave in ("e8", "b62", "s19"):
        assert positivo[chave] == pytest.approx(considerado), chave
    assert positivo["c8"] == pytest.approx(positivo["t39"])
    assert positivo["b61"] == pytest.approx(positivo["t39"])
    # S18 (potencial de TODOS os ciclos ate o corte) e maior: por isso o
    # fechamento da linha 19 nao pode ser feito por ele.
    assert positivo["s18"] > positivo["t39"]

    negativo = medir(montar(tmp_path / "pc_potencial_negativo.xlsx", -0.05))
    assert negativo["t41"] < 0
    assert negativo["t39"] == pytest.approx(0.0)
    assert negativo["vta"] == pytest.approx(negativo["t40"])   # nao reduz
    assert negativo["c8"] == pytest.approx(0.0)
    assert negativo["b87"] == pytest.approx(0.0)
