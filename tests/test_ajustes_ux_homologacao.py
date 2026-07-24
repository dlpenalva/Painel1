"""Rodada de homologacao XLS — ajustes finais de UX e seguranca.

Testes estaticos (openpyxl) protegem a estrutura autoral; testes de integracao
(RUN_EXCEL_INTEGRATION=1) provam no Excel real o dropdown de duas opcoes, a
reabertura sem reparo e a propagacao de itens novos de aditivos.
"""
from __future__ import annotations

import gc
import os
import shutil
from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "templates" / "COLETA_REAJUSTE_OFICIAL.xlsx"
MAROON = "FF8A1538"

com = pytest.mark.skipif(os.environ.get("RUN_EXCEL_INTEGRATION") != "1",
                         reason="defina RUN_EXCEL_INTEGRATION=1 para Excel COM")


def _wb():
    return load_workbook(TEMPLATE, data_only=False)


# ---- A/B CONTROLE ----
def test_controle_b1_funcional():
    wb = _wb()
    c = wb["CONTROLE"]
    assert c["B1"].value == "Principal"
    dvs = [(str(d.sqref), d.formula1) for d in c.data_validations.dataValidation]
    assert any("B1" in s for s, _ in dvs)                 # dropdown mantido
    assert "CONTROLE!$B$1" in wb["RESULTADOS"]["B35"].value  # consumido pela logica


def test_controle_c1_vazia():
    assert _wb()["CONTROLE"]["C1"].value in (None, "")


# ---- C/D/E parametros dropdown ----
def test_parametros_g_tem_validacao():
    p = _wb()["parametros"]
    dv = [(str(d.sqref), d.type, d.formula1) for d in p.data_validations.dataValidation
          if "G12" in str(d.sqref)]
    assert dv and dv[0][1] == "list"


def test_dropdown_duas_opcoes_via_nome():
    wb = _wb()
    p = wb["parametros"]
    assert "OPCOES_SIM_NAO" in list(wb.defined_names)
    assert (p["T2"].value, p["T3"].value) == ("Sim", "Nao")     # duas opcoes distintas
    dv = next(d for d in p.data_validations.dataValidation if "G12" in str(d.sqref))
    assert "OPCOES_SIM_NAO" in (dv.formula1 or "")              # nao literal "Sim,Nao"
    assert '"Sim,Nao"' not in (dv.formula1 or "")


def test_dropdown_permite_vazio():
    p = _wb()["parametros"]
    dv = next(d for d in p.data_validations.dataValidation if "G12" in str(d.sqref))
    assert dv.allowBlank


# ---- F itens_Consumidos Q ----
def test_itens_consumidos_q_sem_fill_estatico():
    ws = _wb()["itens_Consumidos"]
    assert ws["Q2"].fill.patternType in (None,)
    assert ws["Q100"].fill.patternType in (None,)


# ---- G itens_PC A border ----
def test_itens_pc_coluna_a_border_coerente():
    ws = _wb()["itens_PC"]
    a, b = ws["A2"].border, ws["B2"].border
    assert a.left.style == b.left.style and a.right.style == b.right.style


# ---- H/I aditivos ----
def test_aditivos_check_item_ausente_explicito():
    f = _wb()["aditivos"]["M2"].value
    assert "NOVO ITEM NAO CADASTRADO" in f
    assert "itens_Remanesc" in f
    assert "N001" in f                                       # identificador de novo item
    assert "COUNTIF(itens_Remanesc!$A$2:$A$200,A2)=0" in f   # check ja existente


def test_orientacao_novo_item_n001():
    wb = _wb()
    assert "N001" in str(wb["itens_Remanesc"]["A1"].value)   # orientacao discreta
    assert str(wb["itens_Remanesc"]["A1"].value).startswith("ITEM")  # chave do leitor
    assert "N001" in str(wb["aditivos"]["A1"].value)
    assert str(wb["aditivos"]["A1"].value).startswith("ITEM")


def test_aditivos_vu_vem_de_itens_remanesc():
    # VU do aditivo nao e digitado: vem de itens_Remanesc via VLOOKUP (F).
    f = _wb()["aditivos"]["F2"].value
    assert "VLOOKUP" in f and "itens_Remanesc" in f


# ---- N painel ----
def test_painel_posicao_referencia_estilizado():
    ws = _wb()["posicao_referencia"]
    assert ws["H1"].fill.patternType == "solid"
    assert ws["I1"].fill.patternType == "solid"


# ---- O B4 ----
def test_b4_destaque_isolado():
    r = _wb()["RESULTADOS"]
    assert r["B4"].fill.fgColor.rgb == "FFF7E7B2" and r["B4"].font.bold
    assert r["A4"].fill.fgColor.rgb == "FFC6D9E8"   # A4 inalterada
    assert r["C4"].fill.fgColor.rgb == "FFC6D9E8"   # C4 inalterada


# ---- P/Q EIXOS ----
def test_eixos_padronizados_maroon():
    r = _wb()["RESULTADOS"]
    for cel in ("A8", "A17", "A19", "A30", "A256", "A267"):
        assert r[cel].fill.fgColor.rgb == MAROON, f"{cel} nao padronizado"
        assert r[cel].font.color.rgb in ("FFFFFFFF", "00FFFFFF")


def test_eixo2_sem_referencia_linha_62():
    assert "linha 62" not in str(_wb()["RESULTADOS"]["A17"].value)


# ---- R notas + AC/AD ancoras ----
def test_notas_tecnicas_movidas_sem_deslocar():
    r = _wb()["RESULTADOS"]
    assert r["A48"].value in (None, "")
    assert r["A49"].value in (None, "")
    assert r["A280"].value == "NOTAS TECNICAS"
    assert "retroativo por itens" in str(r["A281"].value)
    assert "VTA" in str(r["A282"].value)
    assert str(r["A256"].value).startswith("EIXO 5")     # EIXO 5 continua em A256
    assert "POSICAO DE REFERENCIA" in str(r["A267"].value)  # POSICAO continua em A267


# ---- S invariantes ----
def test_vta_invariante():
    r = _wb()["RESULTADOS"]
    assert r["B23"].value == '=IF(OR(B20="",B21="",B22=""),"",ROUND(B20+B21+B22,2))'
    assert "$N$263" in r["B26"].value and "posicao_referencia" not in r["B26"].value
    assert r["B25"].value in (None, "")


# ---- T integridade basica ----
def test_treze_abas():
    # 12 abas + cobertura_temporal (diagnostico da etapa de cobertura temporal).
    assert len(_wb().sheetnames) == 13


# ================================================================ COM

def _recalc(caminho, inspecionar):
    import time
    import pythoncom
    import win32com.client
    RPC = -2147418111
    pythoncom.CoInitialize()
    xl = win32com.client.DispatchEx("Excel.Application")
    xl.Visible = False
    xl.DisplayAlerts = False
    wb = None
    try:
        for i in range(6):
            try:
                wb = xl.Workbooks.Open(str(Path(caminho).resolve()), UpdateLinks=0)
                break
            except Exception as e:
                if getattr(e, "hresult", None) == RPC and i < 5:
                    time.sleep(1.5 * (i + 1)); continue
                raise
        xl.CalculateFullRebuild()
        time.sleep(0.3)
        return inspecionar(wb)
    finally:
        if wb is not None:
            wb.Close(SaveChanges=False)
        xl.Quit(); gc.collect(); pythoncom.CoUninitialize()


@com
def test_com_dropdown_duas_opcoes_reais():
    def insp(wb):
        v = wb.Worksheets("parametros").Range("G12").Validation
        nm = wb.Names("OPCOES_SIM_NAO").RefersToRange
        opc = [nm.Cells(i, 1).Value for i in range(1, nm.Cells.Count + 1)]
        return v.Type, opc
    tipo, opc = _recalc(TEMPLATE, insp)
    assert tipo == 3          # xlValidateList
    assert opc == ["Sim", "Nao"]   # duas entradas distintas


@com
def test_com_reabertura_sem_reparo(tmp_path):
    import time
    import pythoncom
    import win32com.client
    dest = tmp_path / "t.xlsx"
    shutil.copy2(TEMPLATE, dest)
    for rodada in range(2):
        pythoncom.CoInitialize()
        xl = win32com.client.DispatchEx("Excel.Application")
        xl.Visible = False; xl.DisplayAlerts = True
        try:
            wb = xl.Workbooks.Open(str(dest.resolve()), UpdateLinks=0, CorruptLoad=0)
            assert wb.Sheets.Count == 13, f"rodada {rodada}"  # +cobertura_temporal
            wb.Close(False)
        finally:
            xl.Quit(); gc.collect(); pythoncom.CoUninitialize()
            time.sleep(1)


@com
def test_com_aditivos_itens_novos_propagam(tmp_path):
    """Dois itens novos (base 0, VUs distintos) + item ausente: itens separados,
    propagam para posicao_contratual/posicao_referencia; ausente gera alerta."""
    dest = tmp_path / "adit.xlsx"
    shutil.copy2(TEMPLATE, dest)
    wb = load_workbook(dest, data_only=False)
    p = wb["parametros"]
    for num, ini, fim in [(0, date(2023, 1, 1), date(2023, 12, 31)),
                          (1, date(2024, 1, 1), date(2024, 12, 31)),
                          (2, date(2025, 1, 1), date(2025, 12, 31))]:
        p.cell(num + 2, 1).value = "Sim" if num > 0 else "Base"
        p.cell(num + 2, 3).value = ini
        p.cell(num + 2, 4).value = fim
    ir = wb["itens_Remanesc"]
    ir["A2"], ir["B2"], ir["C2"] = "N001", 0, 500.0       # item novo, base 0, VU 500
    ir["A3"], ir["B3"], ir["C3"] = "N002", 0, 730.0       # item novo, base 0, VU 730
    ad = wb["aditivos"]
    ad["A2"], ad["B2"], ad["D2"], ad["E2"] = "N001", date(2024, 6, 1), "Acrescimo", 10
    ad["A3"], ad["B3"], ad["D3"], ad["E3"] = "N002", date(2024, 6, 1), "Acrescimo", 20
    ad["A4"], ad["B4"], ad["D4"], ad["E4"] = "AUSENTE", date(2024, 6, 1), "Acrescimo", 5
    wb.save(dest)

    def insp(wb):
        pc = wb.Worksheets("posicao_contratual")
        pr = wb.Worksheets("posicao_referencia")
        adc = wb.Worksheets("aditivos")
        return {
            "pc_A_item": pc.Range("A2").Value, "pc_A_vu": pc.Range("B2").Value,
            "pc_A_contr_c1": pc.Range("I2").Value,   # QTD_CONTRATADA_C1 = base+delta
            "pc_B_item": pc.Range("A3").Value, "pc_B_vu": pc.Range("B3").Value,
            "pc_B_contr_c1": pc.Range("I3").Value,
            "pr_A": pr.Range("A2").Value, "pr_B": pr.Range("A3").Value,
            "check_ausente": adc.Range("M4").Value,
        }
    d = _recalc(dest, insp)
    # itens separados, cada um com seu VU
    assert d["pc_A_item"] == "N001" and d["pc_A_vu"] == 500.0
    assert d["pc_B_item"] == "N002" and d["pc_B_vu"] == 730.0
    # base 0 + acrescimo -> QTD_CONTRATADA_C1 = 10 (N001) e 20 (N002); nao se confundem
    assert d["pc_A_contr_c1"] == 10.0
    assert d["pc_B_contr_c1"] == 20.0
    # propagacao: os dois itens aparecem em posicao_referencia
    assert d["pr_A"] == "N001" and d["pr_B"] == "N002"
    # item ausente de itens_Remanesc -> alerta explicito, nao silencioso
    assert "NOVO ITEM NAO CADASTRADO" in str(d["check_ausente"])
