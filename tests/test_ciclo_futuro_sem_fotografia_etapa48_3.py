# -*- coding: utf-8 -*-
"""ETAPA 48.3 — ciclo futuro sintetico nao recebe fotografia fisica.

Regra estrutural:
  - ciclo efetivamente presente na analise -> parametros!I recebe a data
    fisica (exata transportada ou fallback data_inicio, como na Etapa 48);
  - ciclo POSTERIOR ao ultimo ciclo efetivamente presente na analise (criado
    somente por _completar_periodos_ciclos) -> parametros!I fica VAZIO;
  - ciclos ANTERIORES necessarios a estrutura (ex.: C0 quando a analise
    comeca em C2) preservam o fallback anterior.
parametros!C segue sendo a cadeia mensal e pode conter a janela do ciclo
futuro; a guarda fisica de posicao_contratual!AB:AF reconhece I vazio para
nao criar fronteira ficticia via INT de celula vazia.

Ajuste 2 — a aba parametros usa a MESMA guia laranja (tabColor FFFFC000) das
demais abas laranjas oficiais.
"""
from __future__ import annotations

from datetime import date
from io import BytesIO

from openpyxl import load_workbook

from _coleta_oficial import TEMPLATE_COLETA_OFICIAL, gerar_coleta_oficial_preenchida

LARANJA_OFICIAL = "FFFFC000"
ABAS_LARANJAS = (
    "CONTROLE", "financeiro", "itens_Remanesc", "itens_Consumidos",
    "itens_PC", "aditivos", "cobertura_temporal",
)


def _ciclo(numero: int, data_base: str, pedido: str, exata: str | None,
           financeiro: str, situacao: str = "TEMPESTIVO", **extra):
    reg = {
        "ciclo": f"C{numero}",
        "data_base": data_base,
        "data_pedido": pedido,
        "situacao": situacao,
        "situacao_aplicada": situacao,
        "objeto_analise_atual": True,
        "ciclo_ja_concedido": False,
        "percentual_aplicado": 0.05,
        "variacao": 0.05,
        "fator": 1.05,
        "ciclo_calculado": True,
        "efeito_financeiro_retardado": False,
        "financeiro_inicio": financeiro,
    }
    if exata is not None:
        reg["data_abertura_fisica_exata"] = exata
    reg.update(extra)
    return reg


def _payload(*ciclos):
    return {
        "origem": "Reajustes Múltiplos",
        "indice": "IPCA",
        "data_base_original": ciclos[0]["data_base"],
        "fator": 1.05,
        "fator_acumulado": 1.05,
        "variacao_acumulada": 0.05,
        "ciclos": list(ciclos),
    }


def _parametros(payload):
    wb = load_workbook(BytesIO(gerar_coleta_oficial_preenchida(payload)), data_only=False)
    return wb, wb["parametros"]


def _d(celula):
    v = celula.value
    return v.date() if hasattr(v, "date") else v


def _c1():
    return _ciclo(1, "22/11/2022", "22/11/2023", "22/11/2023", "01/11/2023")


def _c2():
    return _ciclo(2, "22/11/2023", "22/11/2024", "22/11/2024", "01/11/2024")


def _c3():
    # Cenario real observado: referencia exata 22/11/2025, pedido 02/01/2026,
    # fotografia fisica = 02/01/2026 (TEMPESTIVO*).
    return _ciclo(3, "22/11/2024", "02/01/2026", "02/01/2026", "01/01/2026",
                  efeito_financeiro_retardado=True)


def _c4():
    return _ciclo(4, "22/11/2025", "22/11/2026", "22/11/2026", "01/11/2026")


# --------------------------------------------------------------- TESTE 1
def test_c1_c2_c3_analisados_c4_futuro_sem_fotografia():
    wb, par = _parametros(_payload(_c1(), _c2(), _c3()))
    # ciclos realmente analisados: fotografia fisica exata preservada
    assert _d(par["I3"]) == date(2023, 11, 22)
    assert _d(par["I4"]) == date(2024, 11, 22)
    assert _d(par["I5"]) == date(2026, 1, 2)
    # C4 e apenas estrutural/futuro: SEM fotografia fisica
    assert par["I6"].value is None
    # a cadeia mensal do ciclo futuro permanece (comportamento atual):
    # competencia(efeito C3)+12m = 01/01/2027 — exatamente a data que ANTES
    # vazava indevidamente para I6.
    assert _d(par["C6"]) == date(2027, 1, 1)
    assert par["D6"].value is not None


# --------------------------------------------------------------- TESTE 2
def test_c1_c2_analisados_c3_c4_futuros_sem_fotografia():
    wb, par = _parametros(_payload(_c1(), _c2()))
    assert _d(par["I3"]) == date(2023, 11, 22)
    assert _d(par["I4"]) == date(2024, 11, 22)
    assert par["I5"].value is None
    assert par["I6"].value is None
    # janela mensal estrutural dos futuros segue existindo
    assert par["C5"].value is not None
    assert par["C6"].value is not None


# --------------------------------------------------------------- TESTE 3
def test_c1_a_c4_analisados_c4_recebe_data_exata_normalmente():
    wb, par = _parametros(_payload(_c1(), _c2(), _c3(), _c4()))
    assert _d(par["I6"]) == date(2026, 11, 22)


# --------------------------------------------------------------- TESTE 4
def test_analise_comecando_em_c2_nao_apaga_ciclos_anteriores():
    wb, par = _parametros(_payload(
        _ciclo(2, "20/04/2024", "20/04/2025", "20/04/2025", "01/04/2025"),
    ))
    # ciclo real: data exata
    assert _d(par["I4"]) == date(2025, 4, 20)
    # anteriores estruturais (C0/C1) preservam o fallback da Etapa 48
    assert par["I2"].value is not None
    assert par["I3"].value is not None
    assert _d(par["I2"]) == _d(par["C2"])
    assert _d(par["I3"]) == _d(par["C3"])
    # somente os POSTERIORES ao ultimo ciclo real ficam sem fotografia
    assert par["I5"].value is None
    assert par["I6"].value is None


# --------------------------------------------------------------- TESTE 5
def test_analise_comecando_em_c3_nao_apaga_ciclos_anteriores():
    wb, par = _parametros(_payload(
        _ciclo(3, "07/09/2025", "07/09/2026", "07/09/2026", "01/09/2026"),
    ))
    assert _d(par["I5"]) == date(2026, 9, 7)
    for linha in (2, 3, 4):            # C0/C1/C2 anteriores: fallback mantido
        assert par[f"I{linha}"].value is not None
        assert _d(par[f"I{linha}"]) == _d(par[f"C{linha}"])
    assert par["I6"].value is None     # somente C4 e futuro


# --------------------------------------------------------------- TESTE 6
def test_payload_legado_fallback_para_ciclo_real_e_vazio_para_futuro():
    payload = _payload(_ciclo(1, "20/04/2023", "20/04/2024", None, "01/04/2024"))
    assert "data_abertura_fisica_exata" not in payload["ciclos"][0]
    wb, par = _parametros(payload)
    # ciclo REAL sem o campo novo: fallback legado intacto (= data_inicio)
    assert _d(par["I3"]) == date(2024, 4, 1)
    assert _d(par["I3"]) == _d(par["C3"])
    # ciclos futuros apenas sintetizados: I vazio, sem fallback
    assert par["I4"].value is None
    assert par["I5"].value is None
    assert par["I6"].value is None


# ------------------------------------------------- guarda fisica com I vazio
def test_guarda_fisica_de_ab_af_reconhece_i_vazio_no_arquivo_gerado():
    wb, par = _parametros(_payload(_c1(), _c2(), _c3()))
    pc = wb["posicao_contratual"]
    for col, n in zip(("AB", "AC", "AD", "AE", "AF"), range(2, 7)):
        for r in (2, 100, 200):
            f = str(pc[f"{col}{r}"].value)
            # I vazio -> 0 ANTES de qualquer INT(parametros!$I$n): nenhuma
            # fronteira ficticia 00/01/1900 e criada para o ciclo futuro.
            assert f'IF(OR(parametros!$C${n}="",parametros!$I${n}=""),0,' in f, (col, r)
            assert f"INT(parametros!$I${n})+1" in f, (col, r)
    # o ciclo futuro C4 tem C6 preenchido e I6 vazio: a guarda antiga (so C)
    # executaria INT(celula vazia); a nova zera o delta sem avaliar a data.
    assert par["C6"].value is not None and par["I6"].value is None


# --------------------------------------------------------------- cor da aba
def test_template_aba_parametros_com_laranja_oficial():
    wb = load_workbook(TEMPLATE_COLETA_OFICIAL)
    tab = wb["parametros"].sheet_properties.tabColor
    assert tab is not None and tab.rgb == LARANJA_OFICIAL
    for aba in ABAS_LARANJAS:
        ref = wb[aba].sheet_properties.tabColor
        assert ref is not None and ref.rgb == LARANJA_OFICIAL, aba
    # nenhuma outra aba mudou de cor
    res = wb["RESULTADOS"].sheet_properties.tabColor
    assert res is not None and res.rgb == "FF8A1538"
    sem_cor = ("comparativo_VTA", "posicao_referencia", "posicao_contratual",
               "itens_RC", "historico_VU", "MEMORIA_RESULTADOS")
    for aba in sem_cor:
        assert wb[aba].sheet_properties.tabColor is None, aba


def test_xls_gerado_preserva_laranja_da_aba_parametros():
    wb, par = _parametros(_payload(_c1(), _c2(), _c3()))
    tab = wb["parametros"].sheet_properties.tabColor
    assert tab is not None and tab.rgb == LARANJA_OFICIAL
