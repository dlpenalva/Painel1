from __future__ import annotations

from datetime import date
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook

from _coleta_reajuste import ler_coleta_reajuste
from _coleta_oficial import gerar_coleta_oficial_preenchida
from _efeitos_financeiros_pc import (
    efeito_financeiro_pc,
    reconciliar_inicios_efeito,
)
from _motor_temporal import montar_motor_temporal
from _motor_vta_sombra import _valor_parcela_pc
from _leitor_masterfile_v10 import _ler_itens_pc_v10


ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "templates" / "COLETA_REAJUSTE_OFICIAL.xlsx"


def _por_ciclo(inicio_c1: date | None = date(2024, 4, 18)):
    limites = {
        "C0": (date(2022, 10, 1), date(2023, 9, 30)),
        "C1": (date(2023, 10, 1), date(2024, 9, 30)),
        "C2": (date(2024, 10, 1), date(2025, 9, 30)),
        "C3": (date(2025, 10, 1), date(2026, 9, 30)),
        "C4": (date(2026, 10, 1), date(2027, 9, 30)),
    }
    resultado = {}
    for ciclo, (data_inicio, data_fim) in limites.items():
        ativo = ciclo == "C1"
        resultado[ciclo] = {
            "ciclo": ciclo,
            "data_inicio": data_inicio,
            "data_fim": data_fim,
            "computar_nesta_apuracao": "Sim" if ativo else "Nao",
            "percentual_reajuste": 0.10 if ativo else 0.0,
            "fator_acumulado": 1.10 if ativo else 1.0,
            "inicio_efeito_financeiro": inicio_c1 if ativo else None,
        }
    return resultado


def _pc(numero: str, data_pc: date | None, pago: str = "Nao", ciclo="C1"):
    return {
        "numero_pc": numero,
        "data_pc": data_pc,
        "ciclo": ciclo,
        "valor_pc": 100.0,
        "pc_pago_a_contratada": pago,
        "campos_vta": {
            "pc_pago_a_contratada": pago,
            "elegivel_retroativo_pc": pago == "Sim",
        },
    }


def _classificar(item, por_ciclo=None):
    por = por_ciclo or _por_ciclo()
    resultado = montar_motor_temporal(
        {"ciclos": {"por_ciclo": por}, "itens_pc": {"itens": [item]}, "resumo": {}}
    )
    return resultado.pcs[0]


@pytest.mark.parametrize(
    ("data_pc", "esperado"),
    [
        (date(2024, 4, 10), "Nao"),
        (date(2024, 4, 18), "Sim"),
        (date(2024, 4, 25), "Sim"),
    ],
)
def test_comparacao_usa_dia_exato(data_pc, esperado):
    reg = _por_ciclo()["C1"]
    assert efeito_financeiro_pc(data_pc, "C1", reg) == esperado


def test_pc_anterior_preserva_ciclo_nominal_e_zeros():
    pc = _classificar(_pc("PC-ANTES", date(2024, 4, 10), "Nao"))
    assert pc.ciclo_temporal == "C1"
    assert pc.efeito_financeiro_pc == "Nao"
    assert pc.fator_aplicado == 1.0
    assert pc.valor_devido == 100.0
    assert pc.delta == 0.0
    assert pc.retroativo == 0.0


def test_pc_na_data_e_posterior_aplicam_fator_sem_mudar_ciclo():
    for data_pc in (date(2024, 4, 18), date(2024, 4, 25)):
        pc = _classificar(_pc("PC-EFEITO", data_pc, "Nao"))
        assert pc.ciclo_temporal == "C1"
        assert pc.efeito_financeiro_pc == "Sim"
        assert pc.fator_aplicado == pytest.approx(1.10)
        assert pc.valor_devido == 110.0
        assert pc.delta == 10.0


def test_pago_reconhece_somente_incremento_depois_dos_efeitos():
    antes = _classificar(_pc("PC-PAGO-A", date(2024, 4, 10), "Sim"))
    depois = _classificar(_pc("PC-PAGO-D", date(2024, 4, 18), "Sim"))
    assert antes.retroativo == 0.0
    assert depois.retroativo == 10.0


def test_c0_e_ciclo_fora_da_apuracao_usam_fator_neutro():
    c0 = _classificar(_pc("PC-C0", date(2023, 4, 10), "Nao", "C0"))
    assert c0.ciclo_temporal == "C0"
    assert c0.efeito_financeiro_pc == "Nao"
    assert c0.fator_aplicado == 1.0

    por = _por_ciclo()
    por["C2"]["fator_acumulado"] = 1.25  # historico preservado
    fora = _classificar(_pc("PC-C2", date(2025, 4, 10), "Nao", "C2"), por)
    assert por["C2"]["fator_acumulado"] == 1.25
    assert fora.efeito_financeiro_pc == "Nao"
    assert fora.fator_aplicado == 1.0
    assert fora.delta == 0.0


def test_ciclo_ativo_sem_inicio_bloqueia_sem_retroativo():
    pc = _classificar(_pc("PC-SEM-DATA-CANONICA", date(2024, 4, 18), "Sim"), _por_ciclo(None))
    assert pc.efeito_financeiro_pc is None
    assert pc.fator_aplicado is None
    assert pc.valor_devido is None
    assert pc.retroativo is None
    erros = [a for a in pc.alertas if a["nivel"] == "ERRO_GRAVE"]
    assert any("PC-SEM-DATA-CANONICA" in a["mensagem"] and "C1" in a["mensagem"] for a in erros)


def test_data_pc_invalida_nao_causa_crash_e_fica_indeterminada():
    pc = _classificar(_pc("PC-INVALIDO", None, "Sim"))
    assert pc.ciclo_temporal is None
    assert pc.efeito_financeiro_pc is None
    assert pc.fator_aplicado is None
    assert any(a["codigo"] == "PC_SEM_DATA" for a in pc.alertas)


def test_vta_sombra_pc_sem_efeito_nominal_e_com_efeito_atualizado():
    base = {"valor_pc": 100.0, "fator_acumulado": 1.10, "valor_atualizado": 110.0}
    assert _valor_parcela_pc({**base, "efeito_financeiro_pc": "Nao"}) == 100.0
    assert _valor_parcela_pc({**base, "efeito_financeiro_pc": "Sim"}) == 110.0
    assert _valor_parcela_pc(base) == 0.0  # legado sem fonte: nao presume Sim


def test_template_usa_l_sem_deslocar_resumos_metadados_e_limite():
    wb = load_workbook(TEMPLATE, data_only=False)
    ws = wb["itens_PC"]
    assert ws["L1"].value == "EFEITO_FINANCEIRO_PC"
    assert [ws.cell(1, c).value for c in range(13, 21)] == [
        "CICLO", "QTD_PC", "VALOR_PC_TOTAL", "VALOR_ATUALIZADO_TOTAL",
        "RETROATIVO_RECONHECIDO_A_PAGAR", "VALOR_ATUALIZADO_EM_ANALISE",
        "DELTA_POTENCIAL", "QTD_COM_CHECK",
    ]
    assert ws["V1"].value == "COMPUTA_VTA"
    assert ws["AC1"].value == "JUSTIFICATIVA_VTA"
    assert wb["parametros"]["H1"].value == "INICIO_EFEITO_FINANCEIRO"
    assert all(str(ws[f"L{r}"].value).startswith("=") for r in range(2, 101))
    assert ws["L101"].value is None and not ws["L101"].has_style
    assert [str(dv.sqref) for dv in ws.data_validations.dataValidation] == ["G2:G100"]
    regras = list(ws.conditional_formatting._cf_rules.items())
    assert len(regras) == 1 and str(regras[0][0].sqref) == "A2:L100"
    assert regras[0][1][0].stopIfTrue is True


def test_formulas_pc_separam_nominal_reconhecido_analise_e_delta():
    ws = load_workbook(TEMPLATE, data_only=False)["itens_PC"]
    assert 'L2="Nao",1' in ws["E2"].value
    assert "ROUND(D2*E2,2)" in ws["F2"].value
    assert 'L2="Sim",ROUND(F2-D2,2),0' in ws["H2"].value
    assert 'G2="Nao"' in ws["I2"].value and "F2" in ws["I2"].value
    assert 'L2="Sim",ROUND(F2-D2,2),0' in ws["J2"].value
    assert "INICIO_EFEITO ausente: PC" in ws["K2"].value
    assert "B2>=" in ws["L2"].value
    # M:T: Q soma H; R soma I; S soma J. Valor integral nao vira retroativo.
    assert "$H$2:$H$100" in ws["Q2"].value
    assert "$I$2:$I$100" in ws["R2"].value
    assert "$J$2:$J$100" in ws["S2"].value


def _workbook_upload(inicio_h: date | None, inicio_meta: date | None) -> bytes:
    wb = load_workbook(TEMPLATE, data_only=False)
    ws = wb["itens_PC"]
    ws["A2"] = "PC-UPLOAD"
    ws["B2"] = date(2024, 4, 18)
    ws["D2"] = 100.0
    ws["G2"] = "Sim"
    # O template oficial e limpo; o upload de teste precisa da janela do C1.
    wb["parametros"]["C3"] = date(2023, 10, 1)
    wb["parametros"]["D3"] = date(2024, 9, 30)
    wb["parametros"]["A3"] = "Sim"
    wb["parametros"]["H3"] = inicio_h
    partes = [] if inicio_meta is None else [f"C1={inicio_meta.isoformat()}"]
    wb.properties.keywords = "CL8US_INICIO_EFEITO:" + ",".join(partes)
    saida = BytesIO()
    wb.save(saida)
    return saida.getvalue()


def test_upload_bloqueia_fonte_ausente_e_inconsistente_com_pc_e_ciclo():
    ausente = ler_coleta_reajuste(_workbook_upload(None, None))
    assert any(
        "PC-UPLOAD" in msg and "C1" in msg
        for msg in ausente["bloqueios_criticos"]
    )
    divergente = ler_coleta_reajuste(
        _workbook_upload(date(2024, 4, 18), date(2024, 4, 19))
    )
    assert any(
        "inconsistente" in msg.lower() and "C1" in msg
        for msg in divergente["bloqueios_criticos"]
    )


def test_fontes_parametros_e_metadado_reconciliam_sem_terceira_fonte():
    wb = load_workbook(BytesIO(_workbook_upload(date(2024, 4, 18), date(2024, 4, 18))))
    inicios, erros, tem_visivel, tem_meta = reconciliar_inicios_efeito(wb)
    assert erros == []
    assert tem_visivel and tem_meta
    assert inicios["C1"] == date(2024, 4, 18)


def test_gerador_persiste_mesma_data_em_parametros_e_metadado():
    dados = {
        "origem": "Reajuste Simples",
        "indice": "IST",
        "data_base_original": "01/02/2023",
        "data_corte": "31/01/2025",
        "ciclos": [{
            "ciclo": "C1",
            "data_inicio": "01/02/2024",
            "data_fim": "31/01/2025",
            "financeiro_inicio": "18/04/2024",
            "percentual_aplicado": 0.10,
            "objeto_analise_atual": True,
            "situacao": "TEMPESTIVO",
        }],
    }
    wb = load_workbook(BytesIO(gerar_coleta_oficial_preenchida(dados)), data_only=False)
    assert wb["parametros"]["H3"].value.date() == date(2024, 4, 18)
    assert wb["parametros"]["H3"].number_format == "dd/mm/yyyy"
    assert "CL8US_INICIO_EFEITO:C1=2024-04-18" in wb.properties.keywords


def test_leitor_recompoe_xls_sem_cache_e_coincide_com_regra_python():
    payload = gerar_coleta_oficial_preenchida({
        "origem": "Reajuste Simples",
        "indice": "IST",
        "data_base_original": "01/02/2023",
        "data_corte": "31/01/2025",
        "ciclos": [{
            "ciclo": "C1",
            "data_inicio": "01/02/2024",
            "data_fim": "31/01/2025",
            "financeiro_inicio": "18/04/2024",
            "percentual_aplicado": 0.10,
            "objeto_analise_atual": True,
        }],
    })
    wb = load_workbook(BytesIO(payload), data_only=False)
    ws = wb["itens_PC"]
    for linha, numero, data_pc in (
        (2, "PC-ANTES", date(2024, 4, 10)),
        (3, "PC-DEPOIS", date(2024, 4, 18)),
    ):
        ws[f"A{linha}"] = numero
        ws[f"B{linha}"] = data_pc
        ws[f"D{linha}"] = 100.0
        ws[f"G{linha}"] = "Nao"
    saida = BytesIO()
    wb.save(saida)

    wb_valores = load_workbook(BytesIO(saida.getvalue()), data_only=True)
    leitura = _ler_itens_pc_v10(wb_valores)
    itens = {i["numero_pc"]: i for i in leitura["itens"]}
    assert itens["PC-ANTES"]["ciclo"] == "C1"
    assert itens["PC-ANTES"]["efeito_financeiro_pc"] == "Nao"
    assert itens["PC-ANTES"]["fator_acumulado"] == 1.0
    assert itens["PC-ANTES"]["valor_atualizado"] == 100.0
    assert itens["PC-ANTES"]["delta_potencial"] == 0.0
    assert itens["PC-DEPOIS"]["ciclo"] == "C1"
    assert itens["PC-DEPOIS"]["efeito_financeiro_pc"] == "Sim"
    assert itens["PC-DEPOIS"]["fator_acumulado"] == pytest.approx(1.10)
    assert itens["PC-DEPOIS"]["valor_atualizado"] == 110.0
    assert itens["PC-DEPOIS"]["valor_atualizado_em_analise"] == 110.0
    assert itens["PC-DEPOIS"]["delta_potencial"] == 10.0


def test_regressao_50_pcs_preserva_identidade_ciclo_e_efeito():
    payload = gerar_coleta_oficial_preenchida({
        "origem": "Regressao 50 PCs",
        "indice": "IST",
        "data_base_original": "01/02/2023",
        "data_corte": "31/01/2025",
        "ciclos": [{
            "ciclo": "C1",
            "data_inicio": "01/02/2024",
            "data_fim": "31/01/2025",
            "financeiro_inicio": "18/04/2024",
            "percentual_aplicado": 0.10,
            "objeto_analise_atual": True,
        }],
    })
    wb = load_workbook(BytesIO(payload), data_only=False)
    ws = wb["itens_PC"]
    for indice in range(50):
        linha = indice + 2
        ws[f"A{linha}"] = f"PC-{indice + 1:03d}"
        ws[f"B{linha}"] = date(2024, 4, 10 if indice % 2 == 0 else 18)
        ws[f"D{linha}"] = 100.0
        ws[f"G{linha}"] = "Nao"
    saida = BytesIO()
    wb.save(saida)
    leitura = _ler_itens_pc_v10(
        load_workbook(BytesIO(saida.getvalue()), data_only=True)
    )
    assert len(leitura["itens"]) == 50
    assert len({i["numero_pc"] for i in leitura["itens"]}) == 50
    assert {i["ciclo"] for i in leitura["itens"]} == {"C1"}
    assert sum(i["efeito_financeiro_pc"] == "Nao" for i in leitura["itens"]) == 25
    assert sum(i["efeito_financeiro_pc"] == "Sim" for i in leitura["itens"]) == 25
    assert leitura["totais"]["total_original"] == 5000.0
    assert leitura["totais"]["total_atualizado"] == 5250.0
