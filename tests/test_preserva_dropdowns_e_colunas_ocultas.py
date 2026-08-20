"""Regressao focal: o XLSX entregue nao pode perder dropdown nem ocultacao.

Duas clausulas petreas, provadas nos BYTES REAIS da Coleta Oficial (XML
serializado, nao apenas no modelo de objetos do openpyxl):

1. ``aditivos!K2:K200`` sai com a lista suspensa original do template;
2. ``itens_PC!V:AC`` saem ocultas.

Complemento das duas regras gerais que as sustentam: nenhuma validacao do
template pode desaparecer e nenhuma coluna oculta no template pode ser
reexibida na copia entregue.
"""

import re
import zipfile
from io import BytesIO

import pytest
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from _coleta_oficial import TEMPLATE_COLETA_OFICIAL, obter_coleta_oficial_bytes


COLUNAS_TECNICAS_ITENS_PC = ("V", "W", "X", "Y", "Z", "AA", "AB", "AC")


@pytest.fixture(scope="module")
def entrega():
    dados = obter_coleta_oficial_bytes()
    base = load_workbook(TEMPLATE_COLETA_OFICIAL, data_only=False)
    runtime = load_workbook(BytesIO(dados), data_only=False)
    yield dados, base, runtime
    base.close()
    runtime.close()


def _xml_da_aba(dados_zip: bytes, nome_aba: str) -> str:
    with zipfile.ZipFile(BytesIO(dados_zip)) as zf:
        workbook_xml = zf.read("xl/workbook.xml").decode("utf-8")
        rels_xml = zf.read("xl/_rels/workbook.xml.rels").decode("utf-8")
        tag = next(
            t for t in re.findall(r"<sheet\b[^>]*/>", workbook_xml)
            if f'name="{nome_aba}"' in t
        )
        r_id = re.search(r'r:id="(rId\d+)"', tag).group(1)
        rel = next(
            t for t in re.findall(r"<Relationship\b[^>]*/>", rels_xml)
            if f'Id="{r_id}"' in t
        )
        alvo = re.search(r'Target="([^"]+)"', rel).group(1)
        caminho = f"xl/{alvo}" if not alvo.startswith("/") else alvo.lstrip("/")
        return zf.read(caminho).decode("utf-8")


def _faixas_de_validacao(wb) -> dict[tuple[str, str], tuple]:
    faixas = {}
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            for faixa in dv.sqref.ranges:
                faixas[(ws.title, str(faixa))] = (dv.type, dv.formula1)
    return faixas


def _colunas_ocultas(wb) -> set[tuple[str, int]]:
    ocultas = set()
    for ws in wb.worksheets:
        for dimensao in ws.column_dimensions.values():
            if not dimensao.hidden or dimensao.min is None or dimensao.max is None:
                continue
            for coluna in range(dimensao.min, dimensao.max + 1):
                ocultas.add((ws.title, coluna))
    return ocultas


def test_aditivos_k_sai_com_o_dropdown_original_do_template(entrega):
    """AJUSTE 1: K2:K200 volta a ter a lista suspensa, com a configuracao
    EXATA do template (mesma faixa combinada com H, mesmo tipo e lista)."""
    dados, base, runtime = entrega

    def lista_de_k(wb):
        ws = wb["aditivos"]
        return [
            (str(dv.sqref), dv.type, dv.formula1, dv.allowBlank)
            for dv in ws.data_validations.dataValidation
            if any(str(f) == "K2:K200" for f in dv.sqref.ranges)
        ]

    assert lista_de_k(runtime) == lista_de_k(base)
    assert lista_de_k(runtime) == [("H2:H200 K2:K200", "list", '"Sim,Nao"', True)]

    xml = _xml_da_aba(dados, "aditivos")
    blocos = re.findall(
        r"<dataValidation\b[^>]*>.*?</dataValidation>", xml, re.DOTALL
    )
    com_k = [b for b in blocos if "K2:K200" in b]
    assert len(com_k) == 1
    assert "<formula1>\"Sim,Nao\"</formula1>" in com_k[0]

    ws = runtime["aditivos"]
    assert ws.column_dimensions["K"].hidden is False
    assert ws.column_dimensions["K"].width == base["aditivos"].column_dimensions["K"].width
    assert [ws[f"K{linha}"].value for linha in range(2, 201)] == [None] * 199


def test_itens_pc_v_ate_ac_saem_ocultas(entrega):
    """AJUSTE 2: V:AC ocultas no XML entregue, sem ocultar U nem AD."""
    dados, _base, runtime = entrega
    ws = runtime["itens_PC"]
    ocultas = {
        coluna
        for dimensao in ws.column_dimensions.values()
        if dimensao.hidden and dimensao.min is not None
        for coluna in range(dimensao.min, dimensao.max + 1)
    }
    for letra in COLUNAS_TECNICAS_ITENS_PC:
        assert column_index_from_string(letra) in ocultas, letra
    assert column_index_from_string("U") not in ocultas
    assert column_index_from_string("AD") not in ocultas

    xml = _xml_da_aba(dados, "itens_PC")
    cobertas: set[int] = set()
    for col in re.findall(r"<col\b[^>]*/>", xml):
        if 'hidden="1"' not in col:
            continue
        minimo = int(re.search(r'min="(\d+)"', col).group(1))
        maximo = int(re.search(r'max="(\d+)"', col).group(1))
        cobertas.update(range(minimo, maximo + 1))
    assert cobertas == {
        column_index_from_string(letra) for letra in COLUNAS_TECNICAS_ITENS_PC
    }


def test_nenhuma_validacao_do_template_desaparece(entrega):
    _dados, base, runtime = entrega
    esperadas = _faixas_de_validacao(base)
    encontradas = _faixas_de_validacao(runtime)
    assert not (set(esperadas) - set(encontradas))
    assert {chave: encontradas[chave][0] for chave in esperadas} == {
        chave: assinatura[0] for chave, assinatura in esperadas.items()
    }


def test_nenhuma_coluna_oculta_do_template_e_reexibida(entrega):
    _dados, base, runtime = entrega
    assert not (_colunas_ocultas(base) - _colunas_ocultas(runtime))
