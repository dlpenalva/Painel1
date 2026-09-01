# -*- coding: utf-8 -*-
"""PR3B-1 - o template oficial nao precisa ser consertado a cada geracao.

A geracao da Coleta aplica migracoes de compatibilidade sobre o template. Ate
esta frente, quatro delas existiam apenas para corrigir o PROPRIO template
oficial versionado, que carregava a versao historica da formula:

  A. fator historico     - RESULTADOS!H5/H8/C12 e comparativo_VTA!B208
  B. completude abertura - MEMORIA_RESULTADOS!W41:W45
  C. base fisica C0      - MEMORIA_RESULTADOS!T27
  D. orientacao de item  - aditivos!A1 e M2:M200

As formulas canonicas foram gravadas no template. A meta NAO e "nenhuma
migracao runtime": as migracoes seguem valendo para arquivos antigos. A meta e
que o template oficial ATUAL atravesse a geracao sem conserto estrutural.

Fora do escopo desta frente, e por isso ainda migrados em runtime:
  E. cronologia da execucao (financeiro!B, itens_PC!C, aditivos!C) - a guarda
     casa tambem com a formula canonica, entao o reescrever e inerente;
  F. apresentacao da copia entregue (rotulos de itens_PC/aditivos) - deliberada.
"""
from __future__ import annotations

import io
from pathlib import Path

from openpyxl import load_workbook

import _coleta_oficial as CO
import _coleta_reajuste as CR


ROOT = Path(__file__).resolve().parents[1]

# Superficies promovidas ao template nesta frente.
GRUPOS_PROMOVIDOS: dict[str, tuple[str, ...]] = {
    "RESULTADOS": ("H5", "H8", "C12"),
    "comparativo_VTA": ("B208",),
    "MEMORIA_RESULTADOS": ("W41", "W42", "W43", "W44", "W45", "T27"),
    "aditivos": ("A1",) + tuple(f"M{linha}" for linha in range(2, 201)),
}

# Marcas da linhagem historica que o template nao pode mais carregar.
LINHAGEM_HISTORICA = {
    ("RESULTADOS", "H5"): "CONTROLE!$B$11",
    ("RESULTADOS", "H8"): "CONTROLE!$B$11",
    ("RESULTADOS", "C12"): "CONTROLE!B11",
    ("comparativo_VTA", "B208"): "CONTROLE!$B$11",
    ("MEMORIA_RESULTADOS", "W41"): "posicao_contratual!$Y$2:$Y$201",
    ("MEMORIA_RESULTADOS", "W45"): "posicao_contratual!$Y$2:$Y$201",
}


def _template():
    return load_workbook(CO.TEMPLATE_COLETA_OFICIAL, data_only=False)


def _coleta_gerada():
    return load_workbook(io.BytesIO(CO.obter_coleta_oficial_bytes()), data_only=False)


def test_geracao_nao_conserta_os_grupos_promovidos() -> None:
    """A prova principal: template e Coleta entregue coincidem nesses pontos."""
    tpl, gerada = _template(), _coleta_gerada()
    divergentes = [
        f"{aba}!{cel}"
        for aba, celulas in GRUPOS_PROMOVIDOS.items()
        for cel in celulas
        if tpl[aba][cel].value != gerada[aba][cel].value
    ]
    assert divergentes == [], (
        "o runtime ainda esta reescrevendo pontos que deveriam ja estar "
        f"canonicos no template: {divergentes[:8]}"
    )


def test_template_nao_carrega_mais_a_linhagem_historica() -> None:
    tpl = _template()
    residuos = [
        f"{aba}!{cel}"
        for (aba, cel), marca in LINHAGEM_HISTORICA.items()
        if marca in str(tpl[aba][cel].value or "")
    ]
    assert residuos == [], f"linhagem historica ainda presente em {residuos}"


def test_fator_historico_do_template_e_fail_closed_por_parametros() -> None:
    """H5 passa a derivar de parametros!E/F, nao de CONTROLE!B11."""
    h5 = str(_template()["RESULTADOS"]["H5"].value or "")
    assert h5.startswith("=IFERROR(")
    assert "CONTROLE!$B$2" in h5
    assert "parametros!$F$" in h5
    # historico incompleto nunca e inventado: sem COUNT completo, devolve vazio
    assert "COUNT(parametros!$E$3:$E$3)=1" in h5


def test_migracao_de_compatibilidade_segue_valendo_para_arquivo_antigo() -> None:
    """Arquivo com a linhagem antiga continua sendo corrigido na geracao."""
    antigo = _template()
    antigo["RESULTADOS"]["H5"] = '=IF(CONTROLE!$B$11="","",CONTROLE!$B$11)'
    antigo["comparativo_VTA"]["B208"] = (
        "=IFERROR(ROUND(SUMPRODUCT(posicao_contratual!$B$2:$B$201,"
        'posicao_contratual!$C$2:$C$201)*CONTROLE!$B$11,2),"")'
    )

    CO._garantir_fator_historico_desacoplado(antigo)

    h5 = str(antigo["RESULTADOS"]["H5"].value or "")
    b208 = str(antigo["comparativo_VTA"]["B208"].value or "")
    assert "CONTROLE!$B$11" not in h5
    assert "parametros!$F$" in h5
    assert "RESULTADOS!$H$5" in b208


def test_migracao_de_abertura_temporal_segue_valendo_para_arquivo_antigo() -> None:
    antigo = _template()
    antigo["MEMORIA_RESULTADOS"]["W41"] = (
        '=IF(COUNTIF(posicao_contratual!$A$2:$A$201,"<>")=0,0,'
        'IF(SUMPRODUCT((posicao_contratual!$A$2:$A$201<>"")*'
        "(posicao_contratual!$Y$2:$Y$201<=0)*"
        "(1-ISNUMBER(posicao_contratual!$G$2:$G$201)))>0,0,1))"
    )

    CO._garantir_completude_abertura_temporal(antigo)

    w41 = str(antigo["MEMORIA_RESULTADOS"]["W41"].value or "")
    assert "posicao_contratual!$Y$2:$Y$201" not in w41
    assert "posicao_contratual!$AL$2:$AL$201" in w41


def test_regra_de_cores_aceita_a_identidade_visual_oficial() -> None:
    """A Coleta entregue tem 10 guias coloridas e nao pode gerar aviso."""
    leitura = CR.ler_coleta_reajuste(CO.obter_coleta_oficial_bytes())
    avisos = " ".join(leitura.get("avisos") or [])
    assert "cor de aba" not in avisos
    assert "fora da identidade visual" not in avisos


def test_paleta_oficial_cobre_as_guias_do_template_e_do_ciclo_em_execucao() -> None:
    import _ciclo_em_execucao as CE

    assert CE.COR_ABA_ENTRADA in CR.PALETA_ABAS_OFICIAL
    coloridas = {
        ws.title: str(ws.sheet_properties.tabColor.rgb or "").upper()
        for ws in _template().worksheets
        if ws.sheet_properties.tabColor is not None
    }
    assert len(coloridas) == 9
    assert coloridas["RESULTADOS"] == CR.COR_ABA_RESULTADOS
    assert set(coloridas.values()) <= set(CR.PALETA_ABAS_OFICIAL)


def test_regra_de_cores_ainda_sinaliza_cor_fora_da_paleta() -> None:
    """A validacao foi relaxada, nao removida."""
    wb = _coleta_gerada()
    wb["financeiro"].sheet_properties.tabColor = "FF00FF00"
    buffer = io.BytesIO()
    wb.save(buffer)

    leitura = CR.ler_coleta_reajuste(buffer.getvalue())
    avisos = [a for a in (leitura.get("avisos") or []) if "identidade visual" in a]
    assert avisos, "cor fora da paleta oficial deixou de ser sinalizada"
    assert "financeiro" in avisos[0]
