"""Contrato permanente da RESULTADOS executiva criada na Etapa 26F."""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "templates" / "COLETA_REAJUSTE_OFICIAL.xlsx"


@pytest.fixture(scope="module")
def wb():
    pasta = load_workbook(TEMPLATE, data_only=False)
    yield pasta
    pasta.close()


def test_resultados_executiva_e_memoria_separadas(wb):
    assert wb.sheetnames[-2:] == ["MEMORIA_RESULTADOS", "RESULTADOS"]
    assert wb["MEMORIA_RESULTADOS"].sheet_state == "hidden"
    assert wb["RESULTADOS"].sheet_state == "visible"
    assert wb["comparativo_VTA"].sheet_state == "hidden"
    assert wb["RESULTADOS"]["A1"].value == (
        "RESULTADOS CONSOLIDADOS — REAJUSTE CONTRATUAL"
    )


def test_quatro_tabelas_e_um_status_global(wb):
    ws = wb["RESULTADOS"]
    assert ws["A3"].value == "STATUS GLOBAL"
    formula_status = str(ws["B3"].value)
    for estado in ("VALIDADO", "ESTIMADO", "REVISE"):
        assert estado in formula_status
    assert "COUNTIF($H$43:$H$50" in formula_status
    assert ws["A8"].value.startswith("1. VALOR TOTAL")
    assert ws["A14"].value.startswith("2. VALOR RETROATIVO")
    assert ws["A24"].value.startswith("3. VALOR REMANESCENTE")
    assert ws["A33"].value.startswith("4. CICLO ATUAL")


def test_tabelas_usam_fontes_homologadas_sem_ref_quebrada(wb):
    ws = wb["RESULTADOS"]
    # Etapa VTA-posicoes: Tabela 1 passou a expor 3 referencias do VTA.
    # FORMA 1 (posicao atual) em B10, FORMA 2 (ultima abertura) em B11 e o
    # contrato integralmente reajustado (antigo B11) migrou para B12.
    assert ws["B10"].value == '=IF(MEMORIA_RESULTADOS!$W$50="","",MEMORIA_RESULTADOS!$W$50)'
    assert ws["B11"].value == '=IF(MEMORIA_RESULTADOS!$W$48="","",MEMORIA_RESULTADOS!$W$48)'
    assert ws["B12"].value == '=IFERROR(comparativo_VTA!$B$208,"")'
    assert "RETRO_OFICIAL" in ws["D22"].value
    assert "posicao_contratual" in ws["C28"].value
    assert "historico_VU" in ws["C28"].value
    assert "financeiro!$E$2:$E$73" in ws["B36"].value
    assert "$B$35" in ws["B36"].value
    assert "$B$36" not in ws["B36"].value
    quebradas = [
        f"{aba.title}!{celula.coordinate}"
        for aba in wb.worksheets
        for linha in aba.iter_rows()
        for celula in linha
        if isinstance(celula.value, str) and "#REF!" in celula.value.upper()
    ]
    assert quebradas == []


def test_nomes_definidos_apontam_para_memoria_absoluta(wb):
    esperados = {
        "RETRO_OFICIAL": "MEMORIA_RESULTADOS!$B$16",
        "VTA_FINAL": "MEMORIA_RESULTADOS!$B$26",
        "REM_BASE_OFICIAL": "MEMORIA_RESULTADOS!$C$35",
        "REM_ATUALIZADO_OFICIAL": "MEMORIA_RESULTADOS!$D$35",
    }
    for nome, referencia in esperados.items():
        assert wb.defined_names[nome].attr_text == referencia
    assert wb.defined_names["STATUS_RESULTADOS"].attr_text == "RESULTADOS!$B$3"


def test_tabela_manual_unica_governa_entradas_da_memoria(wb):
    ws = wb["RESULTADOS"]
    assert [ws.cell(42, col).value for col in range(1, 9)] == [
        "Tipo",
        "Ciclo",
        "Valor (+/-)",
        "Justificativa",
        "Responsável",
        "Data",
        "Aplicar?",
        "Situação",
    ]
    assert str(next(
        dv for dv in ws.data_validations.dataValidation
        if "G43:G50" in str(dv.sqref)
    ).formula1) == "OPCOES_APLICAR_MANUAL"
    memoria = wb["MEMORIA_RESULTADOS"]
    assert memoria["B5"].value == (
        '=IF(AND(RESULTADOS!$G$43="Sim",RESULTADOS!$C$43<>""),'
        'RESULTADOS!$C$43,"")'
    )
    assert memoria["B24"].value == (
        '=IF(AND(RESULTADOS!$G$44="Sim",RESULTADOS!$C$44<>""),'
        'RESULTADOS!$C$44,"")'
    )
    assert memoria["B25"].value == (
        '=IF(AND(RESULTADOS!$G$45="Sim",RESULTADOS!$C$45<>""),'
        'RESULTADOS!$C$45,"")'
    )
    assert memoria["N262"].value == (
        '=IF(AND(RESULTADOS!$G$50="Sim",ISNUMBER(RESULTADOS!$C$50),'
        'RESULTADOS!$C$50>=0),RESULTADOS!$C$50,"")'
    )


def test_achado_a_override_nao_materializa_zero(wb):
    """ZERO != VAZIO: Aplicar?=Sim com valor vazio nunca vira override 0."""
    memoria = wb["MEMORIA_RESULTADOS"]
    overrides_texto = {
        "B5": "$C$43", "D5": "$D$43",
        "B24": "$C$44", "D24": "$D$44",
        "B25": "$C$45", "D25": "$D$45",
    }
    for celula, valor in overrides_texto.items():
        formula = str(memoria[celula].value)
        assert f'RESULTADOS!{valor}<>""' in formula, f"{celula} sem guarda"
        assert formula.startswith("=IF(AND("), f"{celula} sem AND fail-closed"
    # 26F.1: complementos historicos herdam a regra do baseline (>= 0) e a
    # guarda numerica — vazio nao vira 0 e negativo nao flui para o VTA.
    complementos = {
        "N258": "$C$46", "N259": "$C$47",
        "N260": "$C$48", "N261": "$C$49", "N262": "$C$50",
    }
    for celula, valor in complementos.items():
        formula = str(memoria[celula].value)
        assert f'ISNUMBER(RESULTADOS!{valor})' in formula, f"{celula} sem ISNUMBER"
        assert f'RESULTADOS!{valor}>=0' in formula, f"{celula} sem >=0"
        assert formula.startswith("=IF(AND("), f"{celula} sem AND fail-closed"


def test_achado_b_tabela3_nao_fabrica_zero_para_ciclo_futuro(wb):
    ws = wb["RESULTADOS"]
    for linha, indice in zip(range(26, 31), range(5)):
        formula = str(ws[f"B{linha}"].value)
        assert f'VALUE(MID(UPPER(CONTROLE!$B$2),2,1)),-1)<{indice}' in formula
        assert 'COUNTIF(itens_Remanesc!$A$2:$A$200,"<>")=0,""' in formula


def test_achado_c_selos_por_tabela_e_premissa(wb):
    ws = wb["RESULTADOS"]
    assert 'SEARCH("CALCULADO",MEMORIA_RESULTADOS!$E$26)' in str(ws["H8"].value)
    assert 'SEARCH("MANUAL VALIDADO",MEMORIA_RESULTADOS!$F$16)' in str(
        ws["H14"].value
    )
    assert 'SEARCH("CONFERIR",MEMORIA_RESULTADOS!$F$36)' in str(ws["H24"].value)
    assert '"ESTIMADO"' in str(ws["H33"].value)
    assert "posicao_referencia!$I$2" in str(ws["H33"].value)
    global_formula = str(ws["B3"].value)
    for selo in ("$H$8", "$H$14", "$H$24", "$H$33"):
        assert selo in global_formula
    assert "Estimativa" in str(ws["A39"].value)
    assert "consumo fisico" in str(ws["A39"].value)


def test_26f1_h14_completude_por_ciclo_obrigatorio(wb):
    """Ciclo com COMPUTAR_NESTA_APURACAO=Sim sem estado conhecido -> REVISE.

    26F.2: H14 consome a mesma semantica tri-state da coluna de valor da
    Tabela 2 (B16:B20) — zero confirmado e numerico e conta como conhecido.
    MANUAL VALIDADO encerra o retroativo no total oficial (MEMORIA!A6).
    """
    formula = str(wb["RESULTADOS"]["H14"].value)
    assert formula.startswith(
        '=IF(ISNUMBER(SEARCH("MANUAL VALIDADO",MEMORIA_RESULTADOS!$F$16)),'
        '"VALIDADO",'
    )
    for lin_par, lin_res in zip(range(2, 7), range(16, 21)):
        assert f'parametros!$A${lin_par}="Sim"' in formula
        assert f'NOT(ISNUMBER($B${lin_res}))' in formula
    # Ciclo fora da apuracao (A<>"Sim") nao entra como obrigatorio: a condicao
    # e sempre um AND com o proprio marcador canonico de escopo.
    assert formula.count('="Sim",NOT(ISNUMBER(') == 5
    assert 'SEARCH("CALCULADO",MEMORIA_RESULTADOS!$F$16)' in formula


def test_26f2_tabela2_tristate_estrutura(wb):
    """Movimento -> valor; zero confirmado (cobertura GCC) -> 0; ausente -> ''."""
    ws = wb["RESULTADOS"]
    for linha, par_linha in zip(range(16, 21), range(2, 7)):
        formula = str(ws[f"B{linha}"].value)
        # Completude canonica: confirmado GCC (B13/B15), nunca a ultima
        # evidencia (B12/B14).
        assert "cobertura_temporal!$B$13" in formula
        assert "cobertura_temporal!$B$15" in formula
        assert "cobertura_temporal!$B$12" not in formula
        assert "cobertura_temporal!$B$14" not in formula
        # Janela necessaria: MIN(fim nominal do ciclo, corte da analise).
        assert (
            f'MIN(parametros!$D${par_linha},CONTROLE!$B$3)' in formula
        )
        # Zero confirmado materializa 0 numerico (ISNUMBER=TRUE); ausencia
        # permanece "".
        assert formula.count(",0,\"\")") == 2  # ramos Financeiro e PCs
        # Itens: sem prova fisica canonica de consumo zero -> vazio.
        assert 'IF($B$5="Itens"' in formula


def test_26f1_b35_data_ausente_permanece_vazia(wb):
    formula = str(wb["RESULTADOS"]["B35"].value)
    assert 'IF(CONTROLE!$B$3="","",CONTROLE!$B$3)' in formula
    assert 'IF(cobertura_temporal!$B$12<>"",cobertura_temporal!$B$12,"")' in formula
    assert 'IF(cobertura_temporal!$B$14<>"",cobertura_temporal!$B$14,"")' in formula
    # Nenhum fallback cru remanescente: toda ocorrencia de CONTROLE!$B$3 na
    # formula esta dentro do IF guardado (2 ocorrencias: teste e retorno).
    assert formula.count("CONTROLE!$B$3") == 2
    assert 'IF(CONTROLE!$B$3="","",CONTROLE!$B$3)' in formula


def test_26f1_validacao_complemento_historico_nao_negativo(wb):
    ws = wb["RESULTADOS"]
    dv = next(
        d for d in ws.data_validations.dataValidation
        if "C46:C50" in str(d.sqref)
    )
    assert dv.type == "decimal"
    assert dv.operator == "greaterThanOrEqual"
    assert str(dv.formula1) == "0"
    # Situacao das linhas de complemento historico rejeita negativo mesmo se a
    # Data Validation for contornada (colagem/programacao).
    for linha in range(46, 51):
        assert f"C{linha}>=0" in str(ws[f"H{linha}"].value), f"H{linha}"
    # Retroativo manual e ajustes de VTA (43-45) nao ganham a restricao.
    for linha in range(43, 46):
        assert ">=0" not in str(ws[f"H{linha}"].value), f"H{linha}"


def test_26f4_resultados_sem_painel_congelado(wb):
    """Homologacao humana: rolagem livre desde a linha 1 (sem FreezePanes)."""
    pane = wb["RESULTADOS"].sheet_view.pane
    assert pane is None or (
        getattr(pane, "state", None) != "frozen"
        and not getattr(pane, "ySplit", None)
        and not getattr(pane, "xSplit", None)
    )


def test_achado_d_builder_canonico_possui_fator_historico():
    builder = (ROOT / "tools" / "build_coleta_reajuste_template.py").read_text(
        encoding="utf-8"
    )
    assert "Fator histórico integral (até o ciclo vigente)" in builder
    assert 'COUNT(parametros!$E$3:$E$6)=4' in builder
    aplicador = (
        ROOT / "tools" / "aplicar_resultados_consolidados_26f.py"
    ).read_text(encoding="utf-8")
    assert 'COUNT(parametros!$E$3:$E$6)=4' in aplicador


# ================================================================ Excel COM
import os
import shutil

com = pytest.mark.skipif(
    os.environ.get("RUN_EXCEL_INTEGRATION") != "1",
    reason="defina RUN_EXCEL_INTEGRATION=1 para executar o Excel COM",
)


@com
def test_com_26f1_fail_closed_no_excel_real(tmp_path):
    """Prova comportamental 26F.1: H14 obrigatorio, complemento >=0, B35."""
    import gc

    import pythoncom
    import win32com.client
    from datetime import datetime

    dest = tmp_path / "cenario_26f1.xlsx"
    shutil.copyfile(TEMPLATE, dest)
    pythoncom.CoInitialize()
    xl = win32com.client.DispatchEx("Excel.Application")
    xl.Visible = False
    xl.DisplayAlerts = False
    try:
        wb = xl.Workbooks.Open(str(dest.resolve()), UpdateLinks=0, CorruptLoad=0)
        r = wb.Worksheets("RESULTADOS")
        mem = wb.Worksheets("MEMORIA_RESULTADOS")
        ctrl = wb.Worksheets("CONTROLE")
        par = wb.Worksheets("parametros")
        irem = wb.Worksheets("itens_Remanesc")
        fin = wb.Worksheets("financeiro")
        # O cenario escreve tambem em celulas que o runtime preenche (ex.:
        # parametros!E) e que ficam bloqueadas para o usuario final.
        for aba in (r, ctrl, par, irem, fin):
            try:
                aba.Unprotect()
            except Exception:
                pass

        # Cenario: C2 vigente e computada na apuracao; C0/C1 na base;
        # C3/C4 fora da apuracao (contraprova de falso REVISE).
        ctrl.Range("B1").Value = "Principal"
        ctrl.Range("B2").Value = "C2"
        par.Range("A4").Value = "Sim"
        par.Range("E2").Value = 0.0
        par.Range("E3").Value = 0.0
        par.Range("E4").Value = 0.10
        irem.Range("A2").Value = "ITEM-1"
        irem.Range("B2").Value = 10.0
        irem.Range("C2").Value = 100.0
        irem.Range("G2").Value = 4.0  # QTD_REM_BASE_C2
        xl.CalculateFull()

        # 1) Ciclo obrigatorio (C2) sem fonte oficial: linha visivel, vazia,
        #    Tabela 2 REVISE e global REVISE (precedencia).
        assert str(r.Range("A18").Value) == "C2"
        assert r.Range("B18").Value in (None, "")
        assert str(r.Range("H14").Value) == "REVISE"
        assert str(r.Range("B3").Value) == "REVISE"
        # 8/10) Sem marco de cobertura e CONTROLE!B3 vazio: B35 permanece
        #    vazio (nunca serial 0 / 00-01-1900) e H33 falha fechado.
        assert r.Range("B35").Value in (None, "")
        assert str(r.Range("H33").Value) == "REVISE"
        # B11 fail-closed com historico completo ate C2.
        assert abs(float(ctrl.Range("B11").Value) - 1.10) < 1e-9

        # 2) Preenche a fonte obrigatoria de C2: REVISE da Tabela 2 cede.
        fin.Range("A14").Value = datetime(2026, 1, 1)
        fin.Range("B14").Value = "c2"
        fin.Range("C14").Value = 100.0
        fin.Range("G14").Value = "Sim"
        xl.CalculateFull()
        assert str(r.Range("H14").Value) == "VALIDADO"
        # Contraprova: C3/C4 continuam sem dados e nao geram REVISE.
        assert r.Range("B19").Value in (None, "")
        assert r.Range("B20").Value in (None, "")
        # 9) B35 passa a refletir a ultima evidencia financeira (data valida).
        b35 = r.Range("B35").Value
        assert b35 not in (None, "") and getattr(b35, "year", None) == 2026

        # 3-7) Complemento historico: vazio / zero / positivo / negativo.
        r.Unprotect()
        assert mem.Range("N258").Value in (None, "")
        r.Range("G46").Value = "Sim"
        xl.CalculateFull()
        assert mem.Range("N258").Value in (None, "")  # vazio nunca vira 0
        assert str(r.Range("H46").Value) == "REVISE"
        r.Range("D46").Value = "Justificativa"
        r.Range("E46").Value = "GCC"
        r.Range("F46").Value = datetime(2026, 7, 1)
        r.Range("C46").Value = 0.0
        xl.CalculateFull()
        assert float(mem.Range("N258").Value) == 0.0  # zero explicito valido
        assert str(r.Range("H46").Value) == "VALIDADO"
        r.Range("C46").Value = 7.5
        xl.CalculateFull()
        assert float(mem.Range("N258").Value) == 7.5
        n263_antes = mem.Range("N263").Value
        # Negativo inserido programaticamente (contorna a Data Validation):
        r.Range("C46").Value = -5.0
        xl.CalculateFull()
        assert str(r.Range("H46").Value) == "REVISE"
        assert mem.Range("N258").Value in (None, "")  # nao reduz o VTA
        assert float(mem.Range("N263").Value or 0) <= float(n263_antes or 0)

        wb.Close(False)
    finally:
        xl.Quit()
        gc.collect()
        pythoncom.CoUninitialize()


@com
def test_com_26f2_zero_confirmado_vs_fonte_ausente(tmp_path):
    """Matriz F1-F4 / P1-P4 / I1-I3: zero confirmado != fonte ausente."""
    import gc

    import pythoncom
    import win32com.client
    from datetime import datetime

    dest = tmp_path / "cenario_26f2.xlsx"
    shutil.copyfile(TEMPLATE, dest)
    pythoncom.CoInitialize()
    xl = win32com.client.DispatchEx("Excel.Application")
    xl.Visible = False
    xl.DisplayAlerts = False
    try:
        wb = xl.Workbooks.Open(str(dest.resolve()), UpdateLinks=0, CorruptLoad=0)
        r = wb.Worksheets("RESULTADOS")
        ctrl = wb.Worksheets("CONTROLE")
        par = wb.Worksheets("parametros")
        irem = wb.Worksheets("itens_Remanesc")
        fin = wb.Worksheets("financeiro")
        pc = wb.Worksheets("itens_PC")
        cob = wb.Worksheets("cobertura_temporal")
        cons = wb.Worksheets("itens_Consumidos")
        for aba in (r, ctrl, par, irem, fin, pc, cob, cons):
            try:
                aba.Unprotect()
            except Exception:
                pass

        # Base: C1 e C2 na apuracao (A="Sim"); C1 com movimento; corte da
        # analise ANTERIOR ao fim nominal de C2 (prova da janela MIN).
        ctrl.Range("B1").Value = "Principal"
        ctrl.Range("B2").Value = "C2"
        ctrl.Range("B3").Value = datetime(2026, 6, 30)
        par.Range("A3").Value = "Sim"
        par.Range("A4").Value = "Sim"
        par.Range("E2").Value = 0.0
        par.Range("E3").Value = 0.05
        par.Range("E4").Value = 0.10
        par.Range("C3").Value = datetime(2024, 11, 1)
        par.Range("D3").Value = datetime(2025, 10, 31)
        par.Range("C4").Value = datetime(2025, 11, 1)
        par.Range("D4").Value = datetime(2026, 10, 31)
        irem.Range("A2").Value = "ITEM-1"
        irem.Range("B2").Value = 10.0
        irem.Range("C2").Value = 100.0
        fin.Range("A10").Value = datetime(2025, 1, 1)
        fin.Range("B10").Value = "c1"
        fin.Range("C10").Value = 200.0
        fin.Range("G10").Value = "Sim"
        xl.CalculateFull()

        # F1: C2 obrigatoria + cobertura incompleta + nenhuma linha -> vazio
        # + REVISE. F3: C1 com movimento -> valor.
        assert float(r.Range("B17").Value) == 200.0
        assert r.Range("B18").Value in (None, "")
        assert str(r.Range("H14").Value) == "REVISE"
        assert str(r.Range("B3").Value) == "REVISE"

        # F2: cobertura Financeiro confirmada ate o corte (30/06/2026 >= MIN
        # (31/10/2026, 30/06/2026)) + nenhum movimento em C2 -> 0,00 conhecido
        # e nao gera REVISE por ausencia.
        cob.Range("B13").Value = datetime(2026, 6, 30)
        xl.CalculateFull()
        assert float(r.Range("B18").Value) == 0.0
        assert str(r.Range("H14").Value) == "VALIDADO"
        # F4: C3/C4 fora do escopo seguem sem falso REVISE (vazias).
        assert r.Range("B19").Value in (None, "")
        assert r.Range("B20").Value in (None, "")

        # P1-P4: metodo PCs. P1: cobertura PC incompleta + nenhum PC -> vazio
        # + REVISE.
        ctrl.Range("B1").Value = "Pedidos de Compras"
        xl.CalculateFull()
        assert r.Range("B18").Value in (None, "")
        assert str(r.Range("H14").Value) == "REVISE"
        # P3: PC aplicavel em C1 -> valor normal.
        pc.Range("A2").Value = "PC-001"
        pc.Range("B2").Value = datetime(2025, 2, 10)
        pc.Range("D2").Value = 300.0
        pc.Range("G2").Value = "Sim"
        # P4: PC de C3 (fora da janela de C2) nao altera o estado de C2.
        pc.Range("A3").Value = "PC-002"
        pc.Range("B3").Value = datetime(2026, 12, 1)
        pc.Range("D3").Value = 999.0
        pc.Range("G3").Value = "Sim"
        xl.CalculateFull()
        assert float(r.Range("B17").Value) == 300.0
        assert r.Range("B18").Value in (None, "")
        # P2: cobertura PC confirmada ate o corte -> C2 vira 0,00 conhecido.
        cob.Range("B15").Value = datetime(2026, 6, 30)
        xl.CalculateFull()
        assert float(r.Range("B18").Value) == 0.0
        assert float(r.Range("B19").Value or 0) == 999.0 or r.Range("B19").Value in (None, "")

        # I1/I3: metodo Itens — sem prova fisica de consumo zero, ausencia
        # permanece vazia (limitacao documentada); movimento -> valor.
        ctrl.Range("B1").Value = "Itens Consumidos"
        xl.CalculateFull()
        assert r.Range("B18").Value in (None, "")
        assert str(r.Range("H14").Value) == "REVISE"
        cons.Range("A2").Value = "ITEM-1"
        cons.Range("C2").Value = 100.0
        cons.Range("I2").Value = 2.0  # consumo em C2
        xl.CalculateFull()
        assert float(r.Range("B18").Value) == 200.0

        wb.Close(False)
    finally:
        xl.Quit()
        gc.collect()
        pythoncom.CoUninitialize()


def test_fator_historico_fail_closed_e_rotulos_complementares(wb):
    controle = wb["CONTROLE"]
    formula = str(controle["B11"].value)
    assert controle["A11"].value == "Fator histórico integral (até o ciclo vigente)"
    assert 'UPPER($B$2)="C4"' in formula
    assert "COUNT(parametros!$E$3:$E$6)=4" in formula
    assert 'parametros!$F$6,""' in formula
    cobertura = wb["cobertura_temporal"]
    assert "ATUAL confirmada" in cobertura["A8"].value
    assert "QTD_REM_ATUAL" in cobertura["C8"].value
    assert wb["financeiro"]["D2"].number_format == "0.0000"
