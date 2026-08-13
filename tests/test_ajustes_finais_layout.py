# -*- coding: utf-8 -*-
"""Ajustes finais de layout (Etapa VTA) — legibilidade e limpeza estrutural.

Fonte da mutacao: `tools/aplicar_ajustes_finais_layout.py` (aplicada ao template).
Cobre:
* RESULTADOS Tabela 1 (8:13): quebra de texto ligada, coluna H alargada, alturas
  ajustadas ao conteudo (nao mais 20 pt fixo que cortava a composicao);
* cobertura_temporal: linha "Fonte temporal de conferencia" removida, leitura por
  rotulo preservada, nenhuma referencia orfa;
* itens_Remanesc: formatacao condicional dos 4 estados de remanescente (Regra 4);
* CICLO_EM_EXECUCAO: sem congelamento de paineis (gerador);
* travas B26/T25 intactas.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "templates" / "COLETA_REAJUSTE_OFICIAL.xlsx"


@pytest.fixture(scope="module")
def wb():
    return openpyxl.load_workbook(TEMPLATE)


# --------------------------------------------------------------------------- #
# 1. Tabela 1 legivel: wrap ligado e alturas ajustadas ao conteudo             #
# --------------------------------------------------------------------------- #
def test_tabela1_wrap_e_alturas(wb):
    ws = wb["RESULTADOS"]
    for coord in ("A10", "A11", "A12", "A13", "B10", "H10", "C10"):
        assert ws[coord].alignment.wrap_text is True, coord
    for r in (10, 11, 12, 13):
        h = ws.row_dimensions[r].height
        assert h is not None and h >= 45.0, f"linha {r} altura {h}"
    # Leiaute final (50.2): a linha 9 virou a banda combinada titulo+cabecalho
    # da secao 1, homologada com 26pt.
    assert (ws.row_dimensions[9].height or 0) >= 26.0


def test_tabela1_coluna_situacao_alargada(wb):
    ws = wb["RESULTADOS"]
    assert (ws.column_dimensions["H"].width or 0) >= 24.0


def test_tabela1_larguras_sem_desproporcao(wb):
    ws = wb["RESULTADOS"]
    for col in ("A", "B", "H"):
        w = ws.column_dimensions[col].width
        if w is not None:
            assert 9.0 <= w <= 60.0, f"coluna {col} largura {w}"


# --------------------------------------------------------------------------- #
# 2. cobertura_temporal: conferencia removida, sem orfaos, leitura por rotulo   #
# --------------------------------------------------------------------------- #
def _sem_acento(texto: str) -> str:
    """Mesma normalizacao do leitor oficial (_leitor_masterfile_v10._norm)."""
    import unicodedata
    t = unicodedata.normalize("NFKD", texto)
    return "".join(ch for ch in t if not unicodedata.combining(ch))


def test_cobertura_conferencia_removida(wb):
    ws = wb["cobertura_temporal"]
    rotulos = [str(ws.cell(r, 1).value or "").strip().lower() for r in range(1, 30)]
    assert not any(r == "fonte temporal de conferencia" for r in rotulos)
    assert any(r == "fonte temporal principal" for r in rotulos)
    # Datas padronizadas: um rotulo por conceito (nada de dois textos para a
    # mesma pergunta). Os prefixos abaixo sao os lidos pelo leitor oficial.
    for rot in ("data de geracao/analise", "financeiro conferido ate",
                "pcs conferidos ate"):
        assert any(_sem_acento(r).startswith(rot) for r in rotulos), rot


def test_cobertura_sem_referencia_orfa(wb):
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and "#REF!" in c.value:
                    raise AssertionError(f"{ws.title}!{c.coordinate}: {c.value}")


# --------------------------------------------------------------------------- #
# 3. itens_Remanesc: Regra 4 — 4 estados visuais por coluna de abertura         #
# --------------------------------------------------------------------------- #
def test_remanesc_formatacao_condicional_quatro_estados(wb):
    ws = wb["itens_Remanesc"]
    por_faixa = {str(cf.sqref): list(cf.rules) for cf in ws.conditional_formatting}
    for col in ("E", "G", "I", "K"):
        faixa = f"{col}2:{col}201"
        assert faixa in por_faixa, f"sem CF em {faixa}"
        regras = por_faixa[faixa]
        assert len(regras) == 4, f"{faixa}: {len(regras)} regras"
        formulas = " ".join(str(rr.formula) for rr in regras)
        # NAO APLICAVEL passou a olhar CICLO_NASCIMENTO_DATA (por data de efeito),
        # espelhado localmente em itens_Remanesc!BI — regra de CF que referencia
        # outra aba e migrada pelo Excel para a extensao x14, invisivel ao openpyxl.
        assert "$BI" in formulas
        assert ">0" in formulas and "=0" in formulas
        assert any(getattr(rr, "stopIfTrue", False) for rr in regras)


# --------------------------------------------------------------------------- #
# 4. CICLO_EM_EXECUCAO gerado sem congelamento de paineis                        #
# --------------------------------------------------------------------------- #
def test_ciclo_sem_congelamento():
    from openpyxl import Workbook
    import _ciclo_em_execucao as ce

    wbx = Workbook()
    if "posicao_referencia" not in wbx.sheetnames:
        wbx.create_sheet("posicao_referencia")
    ce._criar_aba_itemizada(wbx)
    ws = wbx[ce.ABA_CICLO_EM_EXECUCAO]
    assert ws.freeze_panes in (None, "A1"), f"congelamento presente: {ws.freeze_panes}"


# --------------------------------------------------------------------------- #
# 5. Travas B26/T25 permanecem homologadas                                      #
# --------------------------------------------------------------------------- #
def test_travas_b26_t25_intactas(wb):
    mem = wb["MEMORIA_RESULTADOS"]
    assert str(mem["B26"].value).startswith("=IF(AND(B24")
    assert "$T$25" in str(mem["B26"].value)
    assert str(mem["T25"].value).startswith("=IF(OR($T$24=0")
    nomes = {str(n.split("!")[-1]) for n in wb.defined_names}
    assert "VTA_FINAL" in nomes
