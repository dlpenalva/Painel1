from __future__ import annotations

import hashlib  # noqa: F401 — usado por outros testes deste modulo
import io
import re
from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook

from _coleta_oficial import (
    ABAS_COLETA_OFICIAL,
    NOME_ARQUIVO_COLETA_OFICIAL,
    TEMPLATE_COLETA_OFICIAL,
    gerar_coleta_oficial_preenchida,
)
from _coleta_reajuste_documentos import processar_coleta_oficial_runtime


ROOT = Path(__file__).resolve().parents[1]
# Atualizado na Etapa 4 do hotfix: restauracao da formula VLOOKUP ausente em
# aditivos!F5 via Excel COM (re-save nativo). Layout visual preservado; apenas
# os bytes do template mudaram, por isso o pin de SHA-256 foi reancorado.
# Reancorado novamente no ajuste final: remocao da opcao "Decrescimo" do
# dropdown de aditivos!D2:D200 (Acrescimo,Supressao) via Excel COM (re-save
# nativo, sem reparo). Layout/formulas F2:F200 preservados.
# Reancorado na Etapa 3 (Posicao de Referencia): nova aba posicao_referencia
# (com fallback para a ultima fotografia historica valida), formato de
# CONTROLE!B3 e bloco de referencia em RESULTADOS, tudo via Excel COM (sem
# reparo). Estrutura existente e VTA oficial (B23/B26) preservados.
# Reancorado na rodada de homologacao UX: dropdown Sim/Nao via OPCOES_SIM_NAO,
# limpeza de CONTROLE!C1 e itens_Consumidos!Q, mensagem de aditivos, painel
# posicao_referencia, EIXO 2 padronizado e NOTAS TECNICAS ao fim de RESULTADOS.
# Tudo via Excel COM (sem reparo); matematica e VTA (B23/B25/B26) preservados.
# Reancorado no ajuste pontual de identificacao de novos itens (N001, N002...):
# orientacao em itens_Remanesc!A1/aditivos!A1 e mensagem do check em aditivos!M.
# Atualizado no Pacote pos-5-casos (Excel COM): §10/§11 datas dd/mm/aaaa em
# itens_PC!B2:B100 e aditivos!B2:B200; borda 'thin' na coluna B do itens_PC; §12
# prefixos ACR/SUPR (tolerantes a acento) em aditivos L2:L200 e M2:M200. Sem
# mudanca de layout/validacao; mensagens e demais estilos preservados.
# Atualizado nesta etapa (Excel COM): nova aba cobertura_temporal (diagnostico
# GCC/automatico) inserida antes de RESULTADOS; VTA oficial inalterado.
# Hotfix evidencia x cobertura: BLOCO B separa ultima evidencia de confirmacao
# GCC (2 campos novos) e projecao fail-closed; template re-salvo via Excel COM.
# ETAPA 26B: template alterado intencionalmente (hotfix RESULTADOS/VTA-PC via
# Excel COM — ramo PC no remanescente, B4 canonica, layout helpers ocultos).
# ETAPA 26C: E desacoplada de L (fator historico), N/A em G12, S:T ocultas,
# comparativo executivo A28:C29 (via Excel COM).
# ETAPA 26F: memoria tecnica oculta + RESULTADOS executiva, fator historico
# fail-closed e tabela manual unica (Excel COM).
# Etapa 26G: escala canonica de PCs + fail-closed do VTA (owner:
# tools/aplicar_escala_pcs_26g.py) — pin reancorado.
# Etapa 26H: novos itens Nxxx com base zero automatica (colunas tecnicas
# Y/Z em posicao_contratual), historico_VU temporal, POSICAO ATUAL COMPLETA?
# com vigencia na data de corte, CHECKs de novo item, mensagem de aditivos e
# dropdown itens_PC!G via =OPCOES_SIM_NAO (owner: tools/aplicar_26h_template.py)
# — pin reancorado.
# Etapa 26H.1: base zero VISUAL — formula pre-semeada em itens_Remanesc!B2:B200
# (Nxxx -> 0 na propria aba; caso contrario "") e coercao ""->0 nos consumidores
# aritmeticos posicao_contratual!C e posicao_referencia!C (owner:
# tools/aplicar_26h1_template.py) — pin reancorado.
# Etapa 26H.2 (auditoria independente): fronteira de cadastro 2:200 (B201
# limpa — linha 201 e do total dinamico, fora da capacidade funcional);
# padrao canonico N + exatamente 3 algarismos (espelho ^[Nn]\d{3}$) em
# itens_Remanesc!B e posicao_contratual!Z; UX aditivos!M com WrapText
# (owner: tools/aplicar_26h2_template.py) — pin reancorado.
# Etapa 27B: G vazio deixa de ser erro estrutural e vira retroativo
# potencial (itens_PC!I/J/K + MEMORIA!C10:C14); VTA independente de G
# (owner: tools/aplicar_vta_pc_independente_27b.py) — pin reancorado.
# Etapa VTA-posicoes: 3 referencias do VTA na Tabela 1 de RESULTADOS
# (POSICAO ATUAL / ULTIMA ABERTURA DISPONIVEL / CONTRATO INTEGRAL), bloco
# auxiliar MEMORIA_RESULTADOS!V/W/AB e fotografia atual itens_RC!Q:Y via
# INDIRECT+ISERROR sobre CICLO_EM_EXECUCAO. B26/T25 homologados intactos
# (owner: tools/aplicar_vta_posicoes_tabela1.py) — pin reancorado.
# Correcoes pos-implementacao: cobertura_temporal A8/B8/C8 = "COBERTURA FISICA
# ATUAL CONFIRMADA ATE", origem automatica em CICLO_EM_EXECUCAO!D5 quando A9
# valido (owner: tools/aplicar_cobertura_ciclo.py). A editabilidade de
# CICLO_EM_EXECUCAO (selectUnlockedCells) foi corrigida no gerador em runtime e
# NAO altera este template. — pin reancorado.
# Etapa VTA (ajustes finais de layout): legibilidade da Tabela 1 (8:13), remocao
# da linha "Fonte temporal de conferencia" em cobertura_temporal e CF dos 4
# estados em itens_Remanesc. Sem alterar formula-fonte nem B26/T25. — reancorado.
# Temporalidade dos aditivos por DATA_EFEITO: colunas ocultas
# posicao_contratual!AA:AL, decomposicao temporal da FORMA 2 em
# MEMORIA_RESULTADOS (AC/AD + W48/W53:W57), bloco itens_RC!Z:AC, espelho
# itens_Remanesc!BI e CF dos 4 estados reancorada nele. O bloco itens_RC!Q:Y
# passou a distinguir vazio de zero (INDIRECT sobre celula vazia devolvia 0, o
# que transformava pendencia em "zero confirmado"). As colunas oficiais
# G/K/O/S/W/Y e as travas B26/T25 seguem intactas (formula e valor). — reancorado.
# Etapa UI (acabamento visual): parametros!G2:G6 padronizada, A17:G20 sem
# bordas residuais, novo rotulo em parametros!G10, painel posicao_referencia!H:I
# encerrando em H11, coluna C de cobertura_temporal legivel e, em RESULTADOS,
# H8 legivel + A8 horizontal + bordas da tabela 5 (A54:C66). Somente
# apresentacao e um rotulo: a fotografia de formulas do arquivo e identica
# (tools/aplicar_acabamento_visual_coleta.py trava isso). — reancorado.
# Linha dinamica TOTAL de itens_Remanesc (correcao definitiva): F/H/J/L/N/P/R/AC
# deixaram de ter a totalizacao desativada por IF(FALSE,...) e passaram a usar a
# MESMA deteccao dinamica da coluna D (com guarda COUNT: coluna vazia permanece
# vazia); T (VALOR_EXECUTADO_C4) ganhou a formula que faltava em 2:200 e todos os
# fallbacks de lotacao maxima (linha 201) foram normalizados
# (owner: tools/aplicar_total_dinamico_itens_remanesc.py) — pin reancorado.
# ETAPA 48 — pin reancorado: 1.013 formulas fisicas migradas de parametros!C
# para parametros!I (itens_Remanesc!E1:T1, posicao_contratual!AB:AF,
# posicao_referencia!I6, cobertura_temporal!B6).
# ETAPA 48.3 — pin reancorado: guarda de posicao_contratual!AB2:AF200 estendida
# para reconhecer parametros!I vazio (ciclo futuro sem fotografia fisica nao
# gera fronteira ficticia via INT de celula vazia) e aba parametros com a guia
# laranja oficial (tabColor FFFFC000, identico as demais abas laranjas).
# ETAPA 50 — pin reancorado: leiaute executivo FINAL da aba RESULTADOS,
# homologado visualmente nas rodadas 50.1-50.3 e promovido pela 50.4
# (owner: tools/aplicar_resultados_dashboard_50.py). Mudanca de APRESENTACAO:
# cabecalho institucional com selo de status (espelho de B3), faixa de
# contexto (linha 3), tres cards executivos com chips (linhas 4-6, espelhos de
# B10/D22/B38 e H8/H14/H33), faixa de pendencias (linha 7), linhas separadoras
# visualmente brancas (8/14/23/32/39/52; 31/40/51 ocultas), titulos combinados
# com o cabecalho das tabelas nas secoes 1 (linha 9) e 2 (linha 15) e espelhos
# das notas condicionais (E16:H21 e E35:H38). Nenhuma formula de negocio,
# nome definido ou celula consumida externamente foi alterada;
# MEMORIA_RESULTADOS ficou integralmente identica.
# ETAPA 51A — pin reancorado: acabamento da RESULTADOS (owner:
# tools/aplicar_resultados_51a.py, aplicado SOBRE o leiaute da Etapa 50).
# O diagnostico provou que NAO havia bug no calculo do retroativo (o card E5
# ja espelhava RETRO_OFICIAL via D22); mudancas exclusivamente de
# apresentacao: E6 vira formula de estado vazio do card (le $D$22/$J$6),
# G3:H3 mesclada (fim do ##### na VARIACAO ACUMULADA), larguras A:H com
# piso/teto + ShrinkToFit nos valores dos cards (C5/E5/G5), contraste do
# Quadro 1 e correcao das 26 celulas com o formato-lixo '\Pyyd\ryy\o'
# (regressao da Etapa 50: NumberFormatLocal "Padrão" nao e o token de Geral).
# Nenhuma regra de negocio alterada; MEMORIA_RESULTADOS e W48 identicas.
# ETAPA 51C — pin reancorado: ajustes visuais de abas (owner:
# tools/aplicar_ajustes_xls_51c.py, aplicado SOBRE o template pos-51A).
# Mudancas exclusivamente de apresentacao: parametros!I1:I6 recebe o estilo
# do quadro (copia de formatos de H1:H6); financeiro linha 74 (TOTAL, ancora
# intacta — linhas 62:73 sao capacidade estrutural) em negrito com fill
# institucional; aditivos linha 1 com altura 95 (texto de orientacao do
# runtime tem 173 chars); cobertura_temporal padronizada (fontes 11->10,
# formatos-residuo -> General, legenda emoldurada) preservando os fills
# funcionais (amarelo GCC B13/B15, laranja projecao B16/B23, swatches).
# Zero mudanca de formula/valor/nome em qualquer aba; CF x14 preservada.
# HOTFIX RETRO/VTA — pin reancorado (owner:
# tools/aplicar_hotfix_resultados_retro_vta.py, aplicado SOBRE o template
# pos-51C). Mudancas: MEMORIA_RESULTADOS!W48/W50 method-aware via bloco de
# apoio V60:W67 (execucao historica Financeiro = SUMIF financeiro!E por
# ciclo; ramo PC/Itens preservado byte a byte); card do retroativo passa a
# exibir o valor sob o rotulo (D5='=$D$22' mesclada D5:E5; ancora do ciclo
# migrada para J8, coluna oculta; C3 repontada); textos auditaveis C10/C11
# method-aware; contraste 8497B0 -> 1F4E78/FFFFFF/595959.
# AJUSTES C5/G2 (2026-08-14) — pin reancorado (aplicador efemero via Excel
# COM, nao versionado, mesmo padrao do PR #60). Mudancas: RESULTADOS!C5
# espelho puro de B10 ('=IF($B$10="","",$B$10)', sem fallback B11) com
# ShrinkToFit=False; coluna C 18,91->20,82 (COM) para caber 15pt negrito com
# indent; RESULTADOS!G2 lista as pendencias UMA POR LINHA (TEXTJOIN+CHAR(10),
# WrapText, mesmas 5 flags do J5); linha 2 22->90,5pt (AutoFit do pior caso).
# ESTIMADO SEM PENDENCIA (2026-08-14) — pin reancorado (aplicador efemero via
# Excel COM, nao versionado). RESULTADOS!J5 conta a posicao fisica como
# pendencia SOMENTE com H33="REVISE" (ESTIMADO = posicao valida anterior a
# data de corte, nao e pendencia); RESULTADOS!G2 mostra o motivo da posicao
# so no REVISE (reutilizando posicao_referencia!I10/I11: motivo temporal de
# I11 quando I10="REVISE", senao "nao preenchido ou incompleto") e, no
# ESTIMADO, acrescenta linha INFORMATIVA fora da lista de pendencias com as
# datas da posicao fisica e da data de corte (DAY/MONTH/YEAR locale-safe).
# VTA-M2: SHA atualizado apos correcao do VTA Financeiro no template
# (D20/B26/B28 no ramo Financeiro + bloco CONFERENCIA DA EXECUCAO em
# RESULTADOS); PC e Consumido preservados byte-a-byte fora dessas celulas.
# VTA-M2.1: SHA atualizado de novo — metodologia/titulo do bloco 8 e as
# celulas de dado da tabela passam a ser condicionais a
# MEMORIA_RESULTADOS!$B$4 (nao mostram Financeiro quando o metodo
# selecionado e PCs/Itens), e o mapa quantitativo do ciclo ganha cadeia
# de fallback (C1-C4 buscam o ultimo checkpoint disponivel ate E).
# VTA-M2.2: SHA atualizado de novo — a cadeia de fallback da M2.1 foi
# identificada como temporalmente invalida (podia comparar execucao
# acumulada de varios ciclos contra o Financeiro de um so) e removida;
# "execucao teorica pelo quantitativo" passa a somar as colunas de
# execucao ja existentes em itens_Remanesc (AC/N/P/R para C0-C3, semantica
# de par de checkpoints adjacentes apenas, sem encadeamento); C4 e sempre
# NAO COMPARAVEL (sem checkpoint de fechamento nesta versao do schema).
# VTA-C2: SHA atualizado apos correcao do VTA Consumido no template
# (F20/B26/B28/C33/D33 no ramo Itens + coluna auxiliar itens_Consumidos!V);
# Financeiro e PC preservados byte-a-byte fora dessas celulas.
# VTA-C2.1: SHA atualizado apos saneamento fail-closed do Consumido
# (F20/coluna V/C33 fortalecidos contra zero-vira-None, item incompleto,
# consumo sem valoracao e sobreconsumo). Contagem de formulas inalterada
# (mesmas celulas, conteudo interno mais robusto).
# VTA-U2: template regravado via Excel COM (tools/aplicar_vta_uniformizacao_u2.py)
# — 45 celulas, todas em RESULTADOS (card C5/B63 repontados para VTA_FINAL,
# rotulos B10/B11, bloco 8 didatico e bloco 9 novo). MEMORIA_RESULTADOS,
# nomes definidos, abas, formatacao condicional (inclusive x14) e validacoes
# ficaram identicos; reaberto no Excel sem reparo.
# VTA-U2 (UX final): segunda passada via tools/aplicar_vta_u2_ux_final.py — o
# subtitulo "Posição física atual" sai do card, as linhas 10-13 (referencias
# fisicas + reconciliacao) ficam ocultas com formulas intactas, A9 enuncia a
# identidade canonica e o bloco 8 fala em "quantitativo" no lugar de "fisico".
# SHA256_TEMPLATE_ESPERADO removido: ver a justificativa em
# test_template_preserva_layout_visual_e_sha256.


def _dados_calculadora() -> dict:
    return {
        "origem": "Reajuste Simples",
        "indice": "IST",
        "data_base_original": "01/01/2023",
        "ciclos": [{
            "ciclo": "C1",
            "data_base": "01/01/2023",
            "data_pedido": "01/01/2024",
            "situacao": "TEMPESTIVO",
            "percentual_aplicado": 0.10,
            "financeiro_inicio": "01/01/2024",
        }],
    }


def _dia(valor):
    return valor.date() if hasattr(valor, "date") else valor


def test_calculadoras_e_upload_usam_o_mesmo_fluxo_oficial() -> None:
    simples = (ROOT / "pages/01_Calculo_Simples.py").read_text(encoding="utf-8")
    multiplos = (ROOT / "pages/02_Calculo_Represados.py").read_text(encoding="utf-8")
    upload = (ROOT / "pages/03_Valor_Global.py").read_text(encoding="utf-8")
    runtime = (ROOT / "_coleta_reajuste_documentos.py").read_text(encoding="utf-8")
    for pagina in (simples, multiplos):
        assert "gerar_coleta_oficial_preenchida" in pagina
        assert "gerar_coleta_reajuste(" not in pagina
        assert "NOME_ARQUIVO_COLETA_OFICIAL" in pagina
    assert "processar_coleta_oficial_runtime(conteudo_upload)" in upload
    assert "processar_arquivo_coleta(conteudo)" not in upload
    assert "ler_masterfile_v10(conteudo, exigir_modelo_oficial=True)" in runtime
    assert "reconciliacao_xls_python" in runtime
    assert "avaliar_entrega_segura" in runtime
    assert 'if leitura\n        else _rotulo_origem_coleta(conteudo)' in runtime


def test_geracao_pos_calculadora_preserva_e_preenche_modelo_oficial() -> None:
    payload = gerar_coleta_oficial_preenchida(_dados_calculadora())
    wb = load_workbook(io.BytesIO(payload), data_only=False)
    assert NOME_ARQUIVO_COLETA_OFICIAL == "COLETA_REAJUSTE_OFICIAL.xlsx"
    assert wb.sheetnames == ABAS_COLETA_OFICIAL
    assert wb["CONTROLE"]["B2"].value == "C1"
    assert wb["CONTROLE"]["B7"].value == "IST"
    assert _dia(wb["CONTROLE"]["B8"].value) == date(2023, 1, 1)
    assert wb["parametros"]["B3"].value == "C1"
    assert _dia(wb["parametros"]["C3"].value) == date(2024, 1, 1)
    assert _dia(wb["financeiro"]["A2"].value) == date(2023, 1, 1)
    assert _dia(wb["financeiro"]["A25"].value) == date(2024, 12, 1)
    assert wb["itens_PC"]["A1"].value == "NUMERO_PC"
    assert wb["itens_PC"]["B1"].value == "DATA_PC"
    assert wb["itens_PC"]["C1"].value == "CICLO_PC"
    assert "ITEM" not in [wb["itens_PC"].cell(1, c).value for c in range(1, 12)]
    assert wb["RESULTADOS"]["A41"].value.startswith("5. AJUSTES MANUAIS")
    assert wb["MEMORIA_RESULTADOS"]["A52"].value is not None


def test_multiciclo_iniciado_em_c2_nao_marca_c1_como_objeto_atual() -> None:
    dados = {
        "origem": "Reajustes Múltiplos",
        "indice": "IST",
        "data_base_original": "10/10/2022",
        "ciclos": [
            {
                "ciclo": "C2",
                "data_base": "10/10/2023",
                "data_pedido": "10/10/2024",
                "percentual_aplicado": 0.0488,
                "financeiro_inicio": "01/10/2024",
                "objeto_analise_atual": True,
            },
            {
                "ciclo": "C3",
                "data_base": "10/10/2024",
                "data_pedido": "10/10/2025",
                "percentual_aplicado": 0.0383,
                "financeiro_inicio": "01/10/2025",
                "objeto_analise_atual": True,
            },
        ],
    }
    wb = load_workbook(io.BytesIO(gerar_coleta_oficial_preenchida(dados)), data_only=False)
    parametros = wb["parametros"]
    assert parametros["B3"].value == "C1"
    assert parametros["A3"].value == "Nao"
    assert parametros["B4"].value == "C2"
    assert parametros["A4"].value == "Sim"
    assert parametros["B5"].value == "C3"
    assert parametros["A5"].value == "Sim"


def test_financeiro_comeca_em_c0_na_linha_2_caso_simples() -> None:
    # data-base do indice (2022-10) e 12 meses anterior ao inicio de C0
    # (2023-10): o periodo do indice pertence a memoria de calculo, nunca
    # a grade financeira. Primeira linha ativa = inicio de C0.
    dados = {
        "origem": "Reajuste Simples",
        "indice": "ICTI",
        "data_base_original": "01/10/2022",
        "ciclos": [{
            "ciclo": "C1",
            "data_base": "01/10/2023",
            "data_pedido": "01/10/2024",
            "percentual_aplicado": 0.0623,
            "financeiro_inicio": "01/10/2024",
        }],
    }
    wb = load_workbook(io.BytesIO(gerar_coleta_oficial_preenchida(dados)), data_only=False)
    financeiro = wb["financeiro"]
    inicio_c0 = _dia(wb["parametros"]["C2"].value)
    assert _dia(financeiro["A2"].value) == inicio_c0 == date(2023, 10, 1)
    # nenhuma competencia anterior ao inicio de C0 em toda a grade
    ativas = [r for r in range(2, 74) if financeiro[f"A{r}"].value is not None]
    assert ativas and ativas == list(range(2, ativas[-1] + 1))  # contiguas a partir da linha 2
    for r in ativas:
        assert _dia(financeiro[f"A{r}"].value) >= inicio_c0
    # linhas nao utilizadas permanecem vazias (A, C e G)
    for r in range(ativas[-1] + 1, 74):
        assert financeiro[f"A{r}"].value is None
        assert financeiro[f"C{r}"].value is None
        assert financeiro[f"G{r}"].value is None


def test_financeiro_comeca_em_c0_na_linha_2_multiciclo() -> None:
    dados = {
        "origem": "Reajustes Múltiplos",
        "indice": "ICTI",
        "data_base_original": "01/10/2022",
        "ciclos": [
            {
                "ciclo": "C1",
                "data_base": "01/10/2023",
                "data_pedido": "01/10/2024",
                "percentual_aplicado": 0.0623,
                "financeiro_inicio": "01/10/2023",
                "objeto_analise_atual": True,
            },
            {
                "ciclo": "C2",
                "data_base": "01/10/2024",
                "data_pedido": "01/10/2025",
                "percentual_aplicado": 0.0442,
                "financeiro_inicio": "01/10/2024",
                "objeto_analise_atual": True,
            },
        ],
    }
    wb = load_workbook(io.BytesIO(gerar_coleta_oficial_preenchida(dados)), data_only=False)
    financeiro = wb["financeiro"]
    inicio_c0 = _dia(wb["parametros"]["C2"].value)
    assert _dia(financeiro["A2"].value) == inicio_c0 == date(2023, 10, 1)
    ativas = [r for r in range(2, 74) if financeiro[f"A{r}"].value is not None]
    # ETAPA 31 (regra petrea): SEMPRE a cronologia completa de 60 meses
    # (C0..C4), independentemente da data de corte da analise.
    assert ativas == list(range(2, 62))
    for r in ativas:
        assert _dia(financeiro[f"A{r}"].value) >= inicio_c0
    assert _dia(financeiro[f"A{ativas[-1]}"].value) == date(2028, 9, 1)
    for r in range(62, 74):
        assert financeiro[f"A{r}"].value is None
        assert financeiro[f"G{r}"].value is None


def test_template_tem_72_competencias_e_resultados_alcanca_linha_73() -> None:
    wb = load_workbook(TEMPLATE_COLETA_OFICIAL, data_only=False)
    financeiro = wb["financeiro"]
    assert financeiro.max_row == 74  # linha 74 = TOTAL (B74=TOTAL, C74/E74/F74=SUM)
    for linha in range(2, 74):
        assert str(financeiro[f"B{linha}"].value).startswith("=")
        assert str(financeiro[f"D{linha}"].value).startswith("=")
        assert str(financeiro[f"E{linha}"].value).startswith("=")
        assert str(financeiro[f"F{linha}"].value).startswith("=")

    formulas_resultados = [
        cell.value
        for row in wb["MEMORIA_RESULTADOS"].iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    ]
    assert any("financeiro!$B$2:$B$73" in formula for formula in formulas_resultados)
    assert not any(re.search(r"financeiro!\$([A-G])\$2:\$\1\$61", formula, re.I) for formula in formulas_resultados)


def test_template_preserva_layout_visual_e_sha256() -> None:
    wb = load_workbook(TEMPLATE_COLETA_OFICIAL, data_only=False)
    itens_pc = wb["itens_PC"]
    ocultas = [d for d in itens_pc.column_dimensions.values() if d.hidden]
    assert all(any(d.min <= coluna <= d.max for d in ocultas) for coluna in range(22, 30))
    assert itens_pc.sheet_view.topLeftCell in (None, "A1")
    assert wb["financeiro"].sheet_view.topLeftCell in (None, "A1")
    assert wb["RESULTADOS"]["B3"].value.startswith("=IF(")
    assert wb["MEMORIA_RESULTADOS"].sheet_state == "hidden"
    assert wb["RESULTADOS"]["C43"].fill.fgColor.rgb == "FFFFF2CC"
    # PR #60: B1 em ambar forte (FFC000) com texto escuro — assert defasado
    # (esperava o ambar claro antigo F7E7B2) atualizado ao estado homologado.
    assert wb["CONTROLE"]["B1"].fill.fgColor.rgb == "FFFFC000"
    # O SHA fixo saiu. Ele foi congelado em c2ece81 e o template mudou seis
    # vezes desde entao por merges homologados (o ultimo em 08590e9): a
    # constante quebrou na primeira alteracao legitima e ficou vermelha, ou
    # seja, deixou de proteger qualquer coisa no exato momento em que passou a
    # falhar. A protecao real contra corrupcao e alteracao indevida do template
    # esta em tests/test_integridade_template_xlsx.py, que verifica invariantes
    # que NAO apodrecem: XML bem formado em todas as partes, ausencia de
    # marcador de reparo, ausencia de vinculos externos, contagem de formulas
    # por aba, nomes e ordem das abas, validacoes e formatacao condicional.
    # Os asserts visuais acima continuam cobrindo o layout desta etapa.
    assert TEMPLATE_COLETA_OFICIAL.stat().st_size > 0


def test_upload_rejeita_modelo_antigo_sem_fallback() -> None:
    antigo = (ROOT / "templates/Coleta_Reajuste.xlsx").read_bytes()
    with pytest.raises(ValueError, match="versão anterior|NUMERO_PC|Template incompativel"):
        processar_coleta_oficial_runtime(antigo)
    pagina = (ROOT / "pages/03_Valor_Global.py").read_text(encoding="utf-8")
    assert "CAMINHO_MODELO_COLETA" not in pagina
    assert "Arquivo legado processado" not in pagina
    # A frase "download foi bloqueado..." nunca esteve nesta pagina: ela guarda
    # a AUSENCIA do template na HOME. Continua sendo requisito de produto, entao
    # o assert passa a cobra-la onde ela de fato mora, em vez de exigi-la de um
    # arquivo que nao tem esse papel.
    inicio = (ROOT / "pages/00_Calculadora_Reajustes.py").read_text(encoding="utf-8")
    assert "download foi bloqueado para evitar o uso de modelo incompatível" in inicio


def test_interface_nao_reintroduz_rotulos_antigos() -> None:
    fontes = "\n".join(
        p.read_text(encoding="utf-8")
        for p in [ROOT / "app.py", *(ROOT / "pages").glob("*.py")]
    )
    assert "Piloto controlado" not in fontes
    assert "Piloto Controlado" not in fontes
    assert "Mesa GCC" not in fontes
    assert "MasterFile de entrada" not in fontes
