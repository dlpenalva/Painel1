"""Etapa UI — acabamento visual da Coleta e destaque do seletor de ciclo.

Protege exatamente o que a etapa alterou, sem ampliar escopo:

  * na analise de UM ciclo, CONTROLE!A10/B10 apresentam a variacao apurada
    (valor percentual numerico canonico), C10 declara a ausencia de historico
    e C11 fica vazia; o multiciclo continua com a formula do template;
  * CONTROLE!B11 (fator historico integral) permanece formula nos dois casos —
    ela alimenta comparativo_VTA!B208 e RESULTADOS!H5/H8;
  * nova semantica de parametros!G10, sem terceira opcao e sem mexer nas
    validacoes Sim/Nao;
  * padronizacao de parametros!G2:G6 e limpeza de A17:G20;
  * acabamento do painel posicao_referencia!H:I encerrando em H11;
  * CICLO_EM_EXECUCAO com laranja no campo de preenchimento (C) e verde no
    derivado (D);
  * legibilidade da coluna C de cobertura_temporal;
  * RESULTADOS!H8 legivel, A8 horizontal sem mesclagem e bordas da tabela 5;
  * preservacao de TODAS as formulas do template fora de CONTROLE!B10;
  * CSS do seletor de ciclo restrito a classe do proprio widget.
"""
from __future__ import annotations

import io
import re
from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook

from _coleta_oficial import gerar_coleta_oficial_preenchida, obter_coleta_oficial_bytes
from _gerador_masterfile import (
    AVISO_HISTORICO_CICLO_UNICO,
    ROTULO_VARIACAO_CICLO_UNICO,
)

RAIZ = Path(__file__).resolve().parents[1]
PAGINA_CICLO_UNICO = (RAIZ / "pages" / "01_Calculo_Simples.py").read_text(
    encoding="utf-8"
)

CICLO_C3 = {
    "ciclo": "C3",
    "data_inicio": date(2027, 1, 1),
    "financeiro_inicio": date(2027, 3, 1),
    "percentual": 0.0308,
    "objeto_analise_atual": True,
    "situacao": "Aplicado nesta apuracao",
}
DADOS_UNICO = {
    "data_base_original": "01/01/2024",
    "indice": "IPCA",
    "ciclos": [dict(CICLO_C3)],
}
DADOS_MULTI = {
    "data_base_original": "01/01/2024",
    "indice": "IPCA",
    "ciclos": [
        {"ciclo": "C1", "data_inicio": date(2025, 1, 1),
         "financeiro_inicio": date(2025, 3, 1), "percentual": 0.0449,
         "objeto_analise_atual": True, "situacao": "Aplicado nesta apuracao"},
        {"ciclo": "C2", "financeiro_inicio": date(2026, 3, 1),
         "percentual": 0.0316, "objeto_analise_atual": True,
         "situacao": "Aplicado nesta apuracao"},
        dict(CICLO_C3),
    ],
}


def _formulas(wb) -> dict[str, str]:
    return {
        f"{ws.title}!{cell.coordinate}": cell.value
        for ws in wb.worksheets
        for row in ws.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    }


def _estilo(borda, lado):
    return getattr(getattr(borda, lado, None), "style", None)


def _abrir(dados):
    return load_workbook(io.BytesIO(gerar_coleta_oficial_preenchida(dados)))


@pytest.fixture(scope="module")
def em_branco():
    wb = load_workbook(io.BytesIO(obter_coleta_oficial_bytes()))
    yield wb
    wb.close()


@pytest.fixture(scope="module")
def unico():
    wb = _abrir(DADOS_UNICO)
    yield wb
    wb.close()


@pytest.fixture(scope="module")
def multiciclo():
    wb = _abrir(DADOS_MULTI)
    yield wb
    wb.close()


# ---------------------------------------------------------------------------
# 1. Interface web
# ---------------------------------------------------------------------------

def test_destaque_do_seletor_de_ciclo_e_restrito_ao_proprio_campo():
    assert '"Qual ciclo deseja analisar?"' in PAGINA_CICLO_UNICO
    assert 'key="sim_ciclo_analise"' in PAGINA_CICLO_UNICO
    # O seletor recebe destaque proprio via a classe que o Streamlit publica
    # para widgets com key — nunca por um seletor global de selectbox.
    # O campo interno e alcancado por data-testid + role. data-baseweb="select"
    # nao existe mais no Streamlit de producao (1.59.2) e nao pode voltar a ser
    # a ancora do destaque: com ele, so o rotulo era realcado.
    assert '.st-key-sim_ciclo_analise [data-testid="stSelectbox"] div[role="group"]' in (
        PAGINA_CICLO_UNICO
    )
    blocos = [
        bloco
        for bloco in re.findall(r"<style>(.*?)</style>", PAGINA_CICLO_UNICO, re.S)
        if "st-key-sim_ciclo_analise" in bloco
    ]
    assert len(blocos) == 1
    seletores = [
        regra.split("{", 1)[0].strip()
        for regra in blocos[0].split("}")
        if regra.split("{", 1)[0].strip()
    ]
    assert seletores
    for seletor in seletores:
        assert seletor.startswith(".st-key-sim_ciclo_analise"), seletor
    # Regressao de 07/08/2026: a regra dependia de data-baseweb="select", que o
    # Streamlit 1.59.2 nao emite. O campo ficava sem fundo/borda/acento.
    assert "data-baseweb" not in blocos[0]
    # O destaque tem de alcancar o CAMPO, nao apenas o rotulo.
    alvos_de_campo = [s for s in seletores if '[data-testid="stSelectbox"]' in s]
    assert len(alvos_de_campo) >= 1, seletores
    assert any("label" in s for s in seletores), seletores


# ---------------------------------------------------------------------------
# 2. CONTROLE!B10 / B11
# ---------------------------------------------------------------------------

def test_ciclo_unico_apresenta_variacao_apurada_em_a10_b10(unico):
    """A10/B10 apresentam a variacao canonica DESTA analise (ETAPA 31)."""
    from _gerador_masterfile import ROTULO_VARIACAO_ANALISE

    ctl = unico["CONTROLE"]
    assert ctl["A10"].value == ROTULO_VARIACAO_ANALISE
    assert ROTULO_VARIACAO_ANALISE == "Variação apurada nesta análise"
    # B10 = valor percentual NUMERICO, o mesmo que o gerador gravou em
    # parametros!E5 (percentual canonico do C3 vindo da Calculadora).
    assert ctl["B10"].value == pytest.approx(0.0308)
    assert unico["parametros"]["E5"].value == pytest.approx(0.0308)
    assert ctl["B10"].number_format == "0.00%"


def test_ciclo_unico_declara_ausencia_de_historico_em_c10(unico):
    ctl = unico["CONTROLE"]
    assert ctl["C10"].value == AVISO_HISTORICO_CICLO_UNICO
    assert (
        AVISO_HISTORICO_CICLO_UNICO
        == "Histórico anterior não incluído nesta análise."
    )
    assert "N/A" not in str(ctl["C10"].value)
    assert ctl["C10"].number_format == "@"


def test_ciclo_unico_deixa_c11_vazia(unico):
    """A nota textual anterior saiu de C11: A10/B10 ja apresentam a variacao."""
    assert unico["CONTROLE"]["C11"].value is None


def test_multiciclo_apresenta_variacao_total_composta_em_b10(multiciclo):
    """ETAPA 31: B10 nao fica vazio por falta de historico — apresenta a
    variacao TOTAL composta DESTA analise (fonte canonica)."""
    from _gerador_masterfile import ROTULO_VARIACAO_ANALISE

    ctl = multiciclo["CONTROLE"]
    assert ctl["A10"].value == ROTULO_VARIACAO_ANALISE
    assert ctl["B10"].value == pytest.approx(1.0449 * 1.0316 * 1.0308 - 1)
    assert ctl["B10"].number_format == "0.00%"
    assert ctl["C10"].value is None
    assert ctl["C11"].value is None


@pytest.mark.parametrize("fixture", ["unico", "multiciclo"])
def test_b11_e_o_fator_desta_analise(fixture, request, em_branco):
    """ETAPA 31: A11/B11 apresentam o fator desta analise (B11 = 1 + B10,
    sobre o decimal integral). B11 segue alimentando comparativo_VTA!B208 e
    RESULTADOS!H5/H8 — a formula downstream permanece intacta."""
    from _gerador_masterfile import ROTULO_FATOR_ANALISE

    wb = request.getfixturevalue(fixture)
    assert wb["CONTROLE"]["A11"].value == ROTULO_FATOR_ANALISE
    assert wb["CONTROLE"]["B11"].value == '=IF(ISNUMBER(B10),1+B10,"")'
    assert wb["comparativo_VTA"]["B208"].value == (
        em_branco["comparativo_VTA"]["B208"].value
    )


def test_nenhuma_outra_formula_do_template_muda(unico, multiciclo, em_branco):
    substituiveis = {"CONTROLE!B10", "CONTROLE!B11"}
    base = _formulas(em_branco)
    esperado = {k: v for k, v in base.items() if k not in substituiveis}
    for wb in (unico, multiciclo):
        atual = _formulas(wb)
        for coord in substituiveis:
            atual.pop(coord, None)
        assert atual == esperado


# ---------------------------------------------------------------------------
# 3. parametros
# ---------------------------------------------------------------------------

def test_situacao_g2_g6_segue_o_padrao_da_tabela(em_branco):
    par = em_branco["parametros"]
    for linha in range(2, 7):
        alvo, referencia = par.cell(linha, 7), par.cell(linha, 6)
        assert alvo.font.b == referencia.font.b is False
        assert alvo.font.sz == referencia.font.sz
        assert alvo.font.color.rgb == referencia.font.color.rgb == "FF595959"
        assert alvo.fill.fgColor.rgb == referencia.fill.fgColor.rgb == "FFEDEDED"
        assert _estilo(alvo.border, "left") == "thin"


def test_faixa_a17_g20_fica_visualmente_limpa(em_branco):
    par = em_branco["parametros"]
    for linha in range(17, 21):
        for coluna in range(1, 8):
            celula = par.cell(linha, coluna)
            assert celula.value is None
            assert not any(
                _estilo(celula.border, lado)
                for lado in ("left", "right", "top", "bottom")
            ), celula.coordinate
    # A tabela MEMORIA DO FATOR continua fechada na linha 16.
    assert _estilo(par["A16"].border, "bottom") == "thin"
    assert par.sheet_view.showGridLines is False


def test_g10_pergunta_pela_existencia_do_reajuste_anterior(em_branco):
    par = em_branco["parametros"]
    assert par["G10"].value == (
        "EXISTE REAJUSTE ANTERIOR FORMALIZADO? (Sim/Não; vazio=não comprovado)"
    )
    validacoes = {
        str(dv.sqref): dv.formula1 for dv in par.data_validations.dataValidation
    }
    assert validacoes["G13:G15"] == "OPCOES_SIM_NAO"
    assert validacoes["G12"] == "OPCOES_SIM_NAO_NA"
    # Sem terceira opcao: a lista de origem segue Sim/Nao (+N/A ja existente).
    assert [par.cell(linha, 20).value for linha in range(2, 5)] == ["Sim", "Nao", "N/A"]


# ---------------------------------------------------------------------------
# 4. posicao_referencia
# ---------------------------------------------------------------------------

def test_painel_de_marcos_termina_em_h11(em_branco):
    pos = em_branco["posicao_referencia"]
    modelo_rotulo, modelo_valor = pos["H2"], pos["I2"]
    for linha in range(1, 12):
        assert pos.cell(linha, 8).fill.fgColor.rgb == modelo_rotulo.fill.fgColor.rgb
        assert pos.cell(linha, 9).fill.fgColor.rgb == modelo_valor.fill.fgColor.rgb
        assert _estilo(pos.cell(linha, 8).border, "left") == "thin"
        assert _estilo(pos.cell(linha, 9).border, "right") == "thin"
    assert _estilo(pos["H8"].border, "bottom") is None
    assert _estilo(pos["H11"].border, "bottom") == "thin"
    assert _estilo(pos["I11"].border, "bottom") == "thin"
    # Nada foi acrescentado abaixo do painel.
    assert pos["H12"].value is None
    assert pos["I9"].number_format == "mm-dd-yy"   # data preservada


# ---------------------------------------------------------------------------
# 5. CICLO_EM_EXECUCAO
# ---------------------------------------------------------------------------

def test_coluna_c_e_preenchimento_e_coluna_d_e_automatica(unico):
    ws = unico["CICLO_EM_EXECUCAO"]
    for linha in (13, 100, 211):
        entrada, derivada = ws.cell(linha, 3), ws.cell(linha, 4)
        assert entrada.fill.fgColor.rgb == "FFFCE4D6"       # laranja claro
        assert _estilo(entrada.border, "left") == "medium"
        assert entrada.border.left.color.rgb == "FFE67E22"  # laranja
        assert entrada.protection.locked is False
        assert derivada.fill.fgColor.rgb == "FFE8F5F1"      # verde suave
        assert derivada.value.startswith("=")
    assert ws["C12"].fill.fgColor.rgb == "FFE67E22"
    assert ws["D12"].fill.fgColor.rgb == "FF0F5B50"
    # A area vizinha nao foi repintada.
    assert ws["A13"].fill.fgColor.rgb == "FFF2F2F2"
    assert ws["E13"].fill.fgColor.rgb == "FFDDEBF7"


# ---------------------------------------------------------------------------
# 6. cobertura_temporal
# ---------------------------------------------------------------------------

def test_notas_da_coluna_c_cabem_na_propria_celula(em_branco):
    cob = em_branco["cobertura_temporal"]
    assert 55.0 < cob.column_dimensions["C"].width <= 70.0
    for linha in (8, 14, 17):
        celula = cob.cell(linha, 3)
        assert celula.alignment.wrap_text is True
        linhas = max(1, -(-len(str(celula.value)) // 60))
        assert cob.row_dimensions[linha].height >= linhas * 14.5


# ---------------------------------------------------------------------------
# 7. RESULTADOS
# ---------------------------------------------------------------------------

def test_status_da_tabela1_fica_legivel(em_branco):
    """Leiaute final (50.2/50.3): H8 e H14 vivem nas linhas separadoras
    brancas — a formula do selo permanece integra e calculando (alimenta B3,
    J5 e os chips dos cards), mas nada e renderizado (";;;")."""
    res = em_branco["RESULTADOS"]
    modelo = res["H14"].font
    assert res["H8"].font.color.rgb == modelo.color.rgb == "FFFFFFFF"
    assert (res["H8"].font.name, res["H8"].font.sz, res["H8"].font.b) == (
        modelo.name,
        modelo.sz,
        modelo.b,
    )
    assert res["H8"].number_format == res["H14"].number_format == ";;;"
    assert str(res["H8"].value).startswith("=IF(OR(VTA_FINAL")


def test_titulo_da_tabela1_ocupa_a8_d8_sem_mesclagem(em_branco):
    res = em_branco["RESULTADOS"]
    assert res["A8"].alignment.wrap_text in (False, None)
    assert res["B8"].value is None
    assert not [m for m in res.merged_cells.ranges if m.min_row <= 8 <= m.max_row]
    # Leiaute final (50.2): a linha 8 e o separador entre o topo e o bloco 1;
    # o titulo coabita a linha de cabecalho da tabela (A9).
    # XLS-PC-VTA-ALIGN-1: a linha 8 passou a hospedar o card ambar do
    # retroativo POTENCIAL e o fechamento "RETROATIVO CONSIDERADO NO VTA".
    # Fora do metodo PC toda a faixa devolve "" e o separador continua
    # visualmente vazio — e por isso que cada formula comeca pelo gate.
    for endereco in ("A8", "C8", "D8", "E8"):
        formula = str(res[endereco].value or "")
        assert formula.startswith('=IF(MEMORIA_RESULTADOS!$B$4<>"PCs",""') or             formula.startswith('=IF(OR(MEMORIA_RESULTADOS!$B$4<>"PCs"'), endereco
    # PC-UX-1 renomeou o titulo do bloco 1.
    assert res["A9"].value == "1. COMO O VTA FOI CALCULADO"


def test_tabela_da_linha_53_tem_bordas(em_branco):
    res = em_branco["RESULTADOS"]
    # Etapa 50: a numeracao 5 passou a ser a dos AJUSTES MANUAIS e o anexo
    # tecnico das medidas de PCs virou a secao 6.
    # RESULTADOS-FINAL-1: o titulo deixou de anunciar "medidas com nomes
    # claros" (linguagem de desenvolvimento) e passou a nomear o que a
    # tabela entrega ao fiscal. As bordas conferidas abaixo continuam
    # sendo o objeto deste teste.
    assert str(res["A53"].value).startswith("6. TOTAIS E INDICADORES DE CONFER")
    # XLS-PC-VTA-ALIGN-1: a tabela 6 passou a ter 13 medidas (55:67).
    for linha in range(54, 68):
        for coluna in range(1, 4):
            celula = res.cell(linha, coluna)
            for lado in ("left", "right", "top", "bottom"):
                assert _estilo(celula.border, lado) == "thin", celula.coordinate
            assert celula.border.left.color.rgb == "FFB0C4D8"
    assert res.sheet_view.showGridLines is False
