from __future__ import annotations

import io
import inspect
from datetime import date

import pytest
from openpyxl import load_workbook

from _coleta_oficial import (
    TEMPLATE_COLETA_OFICIAL,
    gerar_coleta_oficial_preenchida,
    normalizar_dados_calculadora,
)
from _coleta_reajuste import ler_coleta_reajuste
from _coleta_reajuste_documentos import (
    adaptar_coleta_reajuste_para_documentos,
    processar_coleta_oficial_runtime,
)
from _gerador_masterfile import _capacidade_financeiro
from _leitor_masterfile_v10 import (
    _ler_parcelas_sombra_financeiro,
)


def _dados(inicio_efeito="18/04/2024"):
    return {
        "origem": "Reajuste Simples",
        "indice": "IST",
        "data_base_original": "01/02/2023",
        "data_corte": "31/01/2025",
        "ciclos": [{
            "ciclo": "C1",
            "data_inicio": "01/02/2024",
            "data_fim": "31/01/2025",
            "data_pedido": "10/03/2024",
            "financeiro_inicio": inicio_efeito,
            "percentual_aplicado": 0.10,
            "objeto_analise_atual": True,
            "situacao": "TEMPESTIVO",
        }],
    }


def _wb(dados=None):
    return load_workbook(
        io.BytesIO(gerar_coleta_oficial_preenchida(dados or _dados())),
        data_only=False,
    )


def _linha_competencia(ws, ano, mes):
    for row in range(2, 74):
        valor = ws[f"A{row}"].value
        if valor and (valor.year, valor.month) == (ano, mes):
            return row
    raise AssertionError(f"competencia {mes:02d}/{ano} ausente")


def _bytes(wb):
    saida = io.BytesIO()
    wb.save(saida)
    return saida.getvalue()


def _materializar_linha(wb, ano, mes, *, nominal=100.0, efeito="Nao", fator=1.1):
    row = _linha_competencia(wb["financeiro"], ano, mes)
    ws = wb["financeiro"]
    ws[f"B{row}"] = "c1"
    ws[f"C{row}"] = nominal
    ws[f"D{row}"] = fator
    ws[f"G{row}"] = efeito
    ws[f"E{row}"] = nominal * fator if efeito == "Sim" else nominal
    ws[f"F{row}"] = ws[f"E{row}"].value - nominal if efeito == "Sim" else 0.0
    return row


def test_gerador_usa_inicio_canonico_com_granularidade_mensal():
    ws = _wb()["financeiro"]
    assert ws[f"G{_linha_competencia(ws, 2024, 2)}"].value == "Nao"
    assert ws[f"G{_linha_competencia(ws, 2024, 3)}"].value == "Nao"
    assert ws[f"G{_linha_competencia(ws, 2024, 4)}"].value == "Sim"
    assert ws[f"G{_linha_competencia(ws, 2024, 5)}"].value == "Sim"


def test_c0_e_ciclo_fora_da_apuracao_ficam_sem_efeito_novo():
    wb = _wb()
    ws = wb["financeiro"]
    assert ws[f"G{_linha_competencia(ws, 2023, 2)}"].value == "Nao"
    assert wb["parametros"]["A2"].value == "Nao"


def test_todas_com_efeito_e_todas_sem_efeito():
    todas_sim = _wb(_dados("01/02/2024"))["financeiro"]
    todas_nao = _wb(_dados("01/02/2026"))["financeiro"]
    linhas_c1 = range(
        _linha_competencia(todas_sim, 2024, 2),
        _linha_competencia(todas_sim, 2025, 1) + 1,
    )
    assert all(todas_sim[f"G{row}"].value == "Sim" for row in linhas_c1)
    assert all(todas_nao[f"G{row}"].value == "Nao" for row in linhas_c1)


def test_multiplos_ciclos_usam_datas_independentes():
    dados = _dados()
    dados["data_corte"] = "31/01/2026"
    dados["ciclos"].append({
        "ciclo": "C2",
        "data_inicio": "01/02/2025",
        "data_fim": "31/01/2026",
        "financeiro_inicio": "20/05/2025",
        "percentual_aplicado": 0.05,
        "objeto_analise_atual": True,
    })
    ws = _wb(dados)["financeiro"]
    assert ws[f"G{_linha_competencia(ws, 2025, 4)}"].value == "Nao"
    assert ws[f"G{_linha_competencia(ws, 2025, 5)}"].value == "Sim"


def test_formulas_e_f_respeitam_g_sem_mascarar_vazio():
    ws = load_workbook(TEMPLATE_COLETA_OFICIAL, data_only=False)["financeiro"]
    assert 'IF(G2="",""' in ws["E2"].value
    assert 'G2="Sim"' in ws["E2"].value
    assert "ROUND(C2*D2,2)" in ws["E2"].value
    assert 'IF(G2="",""' in ws["F2"].value
    assert 'G2="Sim"' in ws["F2"].value
    assert "ROUND(E2-C2,2)" in ws["F2"].value


def test_exemplo_fevereiro_abril_xls_python_100_100_110():
    wb = _wb()
    ws = wb["financeiro"]
    linhas = [_linha_competencia(ws, 2024, mes) for mes in (2, 3, 4)]
    for row, efeito, atualizado, delta in zip(
        linhas, ("Nao", "Nao", "Sim"), (100.0, 100.0, 110.0), (0.0, 0.0, 10.0)
    ):
        ws[f"C{row}"] = 100.0
        ws[f"G{row}"] = efeito
        ws[f"E{row}"] = atualizado
        ws[f"F{row}"] = delta
    parcelas = _ler_parcelas_sombra_financeiro(wb)
    bases = [p for p in parcelas if p["identificador"].endswith(tuple(f":base:{r}" for r in linhas))]
    deltas = [p for p in parcelas if p["identificador"].endswith(tuple(f":delta:{r}" for r in linhas))]
    assert sum(p["valor_atualizado"] for p in bases) == 310.0
    assert sum(p["valor"] for p in deltas) == 10.0


def test_alteracao_manual_e_respeitada_e_gera_aviso_sem_bloqueio():
    wb = _wb()
    row = _linha_competencia(wb["financeiro"], 2024, 4)
    assert wb["financeiro"][f"G{row}"].value == "Sim"
    wb["financeiro"][f"G{row}"] = "Nao"
    wb["financeiro"][f"C{row}"] = 100.0
    diagnostico = ler_coleta_reajuste(_bytes(wb))
    assert any("C1 - 04/2024" in aviso for aviso in diagnostico["avisos"])
    assert not any("efeito financeiro nao informado" in b.lower() for b in diagnostico["bloqueios_criticos"])


def test_linha_ativa_sem_decisao_bloqueia_e_identifica_competencia():
    wb = _wb()
    row = _linha_competencia(wb["financeiro"], 2024, 4)
    wb["financeiro"][f"C{row}"] = 100.0
    wb["financeiro"][f"G{row}"] = None
    diagnostico = ler_coleta_reajuste(_bytes(wb))
    assert any("04/2024" in item and "nao informado" in item.lower() for item in diagnostico["bloqueios_criticos"])


def test_linha_nao_utilizada_vazia_nao_bloqueia():
    wb = _wb()
    wb["financeiro"]["A73"] = None
    wb["financeiro"]["C73"] = None
    wb["financeiro"]["G73"] = None
    diagnostico = ler_coleta_reajuste(_bytes(wb))
    assert not any("competencia nao informada" in item.lower() for item in diagnostico["bloqueios_criticos"])


def test_valor_invalido_bloqueia():
    wb = _wb()
    row = _linha_competencia(wb["financeiro"], 2024, 4)
    wb["financeiro"][f"G{row}"] = "Talvez"
    diagnostico = ler_coleta_reajuste(_bytes(wb))
    assert any("efeito financeiro invalido" in item.lower() for item in diagnostico["bloqueios_criticos"])


def test_dropdown_e_formatacao_condicional_sao_dinamicos():
    ws = load_workbook(TEMPLATE_COLETA_OFICIAL, data_only=False)["financeiro"]
    assert any(str(dv.sqref) == "G2:G73" and dv.formula1 == '"Sim,Nao"' for dv in ws.data_validations.dataValidation)
    regras = [
        (str(intervalo), formula)
        for intervalo, itens in ws.conditional_formatting._cf_rules.items()
        for regra in itens for formula in (regra.formula or [])
    ]
    assert all(intervalo == "<ConditionalFormatting A2:G73>" for intervalo, _ in regras)
    assert regras[0][1] == 'AND($C2<>"",$G2="")'
    assert regras[1][1] == 'AND($A2<>"",$G2="Nao")'
    assert ws["A2"].fill.fgColor.rgb not in ("FFFFC7CE", "FFFCE4D6")


def test_data_canonica_fica_auditavel_em_metadado_do_xls():
    wb = _wb()
    assert "CL8US_INICIO_EFEITO:C1=2024-04-18" in wb.properties.keywords


def test_limites_operacionais_convergem_na_linha_73():
    wb = load_workbook(TEMPLATE_COLETA_OFICIAL, data_only=False)
    ws = wb["financeiro"]
    dv = ws.data_validations.dataValidation
    assert len(dv) == 1
    assert str(dv[0].sqref) == "G2:G73"
    assert dv[0].formula1 == '"Sim,Nao"'
    assert _capacidade_financeiro(ws) == 72
    assert "for row in range(2, 74)" in inspect.getsource(ler_coleta_reajuste)
    assert all(str(ws[f"{col}{row}"].value).startswith("=") for row in range(2, 74) for col in "BDEF")
    # Linha 74 = TOTAL: B=label, C/E/F=SUM, A/D/G=vazios
    assert ws["B74"].value == "TOTAL"
    assert all(str(ws[f"{col}74"].value).upper().startswith("=SUM") for col in "CEF")
    assert all(ws[f"{col}74"].value is None for col in "ADG")
    assert all(str(intervalo.sqref) == "A2:G73" for intervalo in ws.conditional_formatting._cf_rules)
    formulas_resultados = [
        cell.value for row in wb["RESULTADOS"].iter_rows() for cell in row
        if isinstance(cell.value, str) and "financeiro!" in cell.value
    ]
    assert formulas_resultados
    assert all("$73" in formula and "$74" not in formula for formula in formulas_resultados)


@pytest.mark.parametrize("ciclo_ano_mes,efeito", [
    ((2023, 2), "Sim"), ((2023, 2), "Nao"),
    ((2024, 4), "Sim"), ((2024, 4), "Nao"),
])
def test_legado_sem_metadado_nao_emite_override(ciclo_ano_mes, efeito):
    wb = _wb()
    wb.properties.keywords = ""
    ano, mes = ciclo_ano_mes
    row = _linha_competencia(wb["financeiro"], ano, mes)
    wb["financeiro"][f"G{row}"] = efeito
    diagnostico = ler_coleta_reajuste(_bytes(wb))
    assert not any("ajustada manualmente" in aviso for aviso in diagnostico["avisos"])


def test_legado_ciclo_nao_computado_nao_emite_override():
    wb = _wb()
    wb.properties.keywords = ""
    wb["parametros"]["A3"] = "Nao"
    row = _linha_competencia(wb["financeiro"], 2024, 4)
    wb["financeiro"][f"G{row}"] = "Sim"
    diagnostico = ler_coleta_reajuste(_bytes(wb))
    assert not any("ajustada manualmente" in aviso for aviso in diagnostico["avisos"])


def test_legado_respeita_g_no_calculo_sem_aviso():
    wb = _wb()
    wb.properties.keywords = ""
    row = _materializar_linha(wb, 2024, 4, efeito="Nao")
    parcelas = [p for p in _ler_parcelas_sombra_financeiro(wb) if p["linha"] == row]
    diagnostico = ler_coleta_reajuste(_bytes(wb))
    assert len(parcelas) == 1
    assert parcelas[0]["valor_atualizado"] == 100.0
    assert not any("ajustada manualmente" in aviso for aviso in diagnostico["avisos"])


def test_runtime_propaga_cada_override_uma_unica_vez():
    wb = _wb()
    for mes in (4, 5):
        row = _linha_competencia(wb["financeiro"], 2024, mes)
        wb["financeiro"][f"G{row}"] = "Nao"
    _, diagnostico = processar_coleta_oficial_runtime(_bytes(wb))
    avisos = [a for a in diagnostico["avisos"] if "ajustada manualmente" in a]
    assert len(avisos) == 2
    assert len(set(avisos)) == 2
    assert any("C1 - 04/2024" in aviso for aviso in avisos)
    assert any("C1 - 05/2024" in aviso for aviso in avisos)


def test_competencia_malformada_nao_causa_crash_e_bloqueia():
    wb = _wb()
    row = _linha_competencia(wb["financeiro"], 2024, 4)
    wb["financeiro"][f"A{row}"] = "abril de 2024"
    wb["financeiro"][f"C{row}"] = 100.0
    diagnostico = ler_coleta_reajuste(_bytes(wb))
    assert any("competencia invalida" in item.lower() and f"linha {row}" in item.lower() for item in diagnostico["bloqueios_criticos"])


@pytest.mark.parametrize("efeito", ["sim", "nao", "Não", "SIM", "NAO", " Sim ", " Nao "])
def test_variantes_nao_literais_de_g_sao_bloqueadas(efeito):
    wb = _wb()
    row = _linha_competencia(wb["financeiro"], 2024, 4)
    wb["financeiro"][f"C{row}"] = 100.0
    wb["financeiro"][f"G{row}"] = efeito
    diagnostico = ler_coleta_reajuste(_bytes(wb))
    assert any("use o dropdown" in item.lower() for item in diagnostico["bloqueios_criticos"])


def test_fronteiras_reais_simples_e_multiciclo_chegam_ao_gerador():
    simples = normalizar_dados_calculadora(_dados())
    assert simples["ciclos"][0]["inicio_efeito_financeiro"] == date(2024, 4, 18)
    wb_simples = _wb(_dados())
    assert "CL8US_INICIO_EFEITO:C1=2024-04-18" in wb_simples.properties.keywords
    assert wb_simples["financeiro"][f"G{_linha_competencia(wb_simples['financeiro'], 2024, 4)}"].value == "Sim"

    multi = _dados()
    multi["data_corte"] = "31/01/2026"
    multi["ciclos"].append({
        "ciclo": "C2", "data_inicio": "01/02/2025", "data_fim": "31/01/2026",
        "financeiro_inicio": "20/05/2025", "percentual_aplicado": 0.05,
        "objeto_analise_atual": True,
    })
    wb_multi = _wb(multi)
    assert "C1=2024-04-18,C2=2025-05-20" in wb_multi.properties.keywords
    assert wb_multi["financeiro"][f"G{_linha_competencia(wb_multi['financeiro'], 2025, 4)}"].value == "Nao"
    assert wb_multi["financeiro"][f"G{_linha_competencia(wb_multi['financeiro'], 2025, 5)}"].value == "Sim"


@pytest.mark.parametrize("efeitos,esperado_qtd,esperado_total", [
    ([], 0, 0.0),
    ([(2024, 2, 100.0)], 1, 100.0),
    ([(2024, 2, 100.0), (2024, 3, 125.0)], 2, 225.0),
])
def test_objeto_documental_transporta_meses_sem_efeito(efeitos, esperado_qtd, esperado_total):
    wb = _wb()
    for ano, mes, nominal in efeitos:
        _materializar_linha(wb, ano, mes, nominal=nominal, efeito="Nao")
    if efeitos:
        _materializar_linha(wb, 2024, 4, nominal=200.0, efeito="Sim")
    payload = _bytes(wb)
    resultado = adaptar_coleta_reajuste_para_documentos(
        payload, diagnostico=ler_coleta_reajuste(payload)
    )
    df = resultado["df_meses_sem_efeito_financeiro"]
    assert resultado["quantidade_meses_sem_efeito_financeiro"] == esperado_qtd
    assert resultado["valor_total_sem_efeito_financeiro"] == esperado_total
    assert len(df) == esperado_qtd
    assert df.empty or set(df["Efeito financeiro"]) == {"Nao"}


def test_objeto_documental_legado_sem_metadado_respeita_g():
    wb = _wb()
    wb.properties.keywords = ""
    _materializar_linha(wb, 2024, 2, nominal=100.0, efeito="Nao")
    payload = _bytes(wb)
    diagnostico = ler_coleta_reajuste(payload)
    resultado = adaptar_coleta_reajuste_para_documentos(payload, diagnostico=diagnostico)
    assert resultado["quantidade_meses_sem_efeito_financeiro"] == 1
    assert resultado["valor_total_sem_efeito_financeiro"] == 100.0
    assert not any("ajustada manualmente" in aviso for aviso in diagnostico["avisos"])


def test_ciclo_nao_computado_preserva_nominal_fator_historico_e_vta():
    dados = _dados()
    dados["ciclos"][0]["objeto_analise_atual"] = False
    dados["ciclos"][0]["ciclo_ja_concedido"] = True
    wb = _wb(dados)
    row = _materializar_linha(wb, 2024, 4, efeito="Nao", fator=1.10)
    parcelas = [p for p in _ler_parcelas_sombra_financeiro(wb) if p["linha"] == row]
    assert wb["financeiro"][f"G{row}"].value == "Nao"
    assert wb["parametros"]["A3"].value == "Nao"
    assert wb["parametros"]["E3"].value == 0.10
    assert len(parcelas) == 1
    assert parcelas[0]["valor"] == 100.0
    assert parcelas[0]["valor_atualizado"] == 100.0
    assert wb["financeiro"][f"F{row}"].value == 0.0
