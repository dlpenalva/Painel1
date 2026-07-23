"""Testes focais do Pacote pós-teste dos 5 casos.

Cobre os itens pure-Python: §18 (última competência IPCA/IGP-M, IST/ICTI intactos),
§12 (Acréscimo/Supressão com e sem acento na classificação Python) e a matriz
HARD × SOFT block do upload tolerante (§3/§4/§8/§13/§14).
"""
import sys
import unittest
from datetime import date
from pathlib import Path
from unittest import mock

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

import _indice_utils as iu
from _motor_posicao_contratual import normalizar_tipo_movimento


class _Resp:
    def __init__(self, payload):
        self._payload = payload

    def raise_for_status(self):
        pass

    def json(self):
        return self._payload


class TestUltimaCompetenciaSGS(unittest.TestCase):
    """§18 / §27 Q,R,S,T."""

    def test_ipca_433_retorna_competencia_mm_aaaa(self):
        payload = [{"data": "01/06/2026", "valor": "0.21"}]
        with mock.patch.object(iu.requests, "get", return_value=_Resp(payload)):
            r = iu.obter_ultima_competencia_sgs(iu.SGS_IPCA)
        self.assertEqual(r["descricao"], "06/2026")
        self.assertEqual(r["serie"], 433)
        self.assertEqual((r["data"].year, r["data"].month, r["data"].day), (2026, 6, 1))

    def test_igpm_189_retorna_competencia_mm_aaaa(self):
        payload = [{"data": "01/05/2026", "valor": "0.40"}]
        with mock.patch.object(iu.requests, "get", return_value=_Resp(payload)):
            r = iu.obter_ultima_competencia_sgs(iu.SGS_IGPM)
        self.assertEqual(r["descricao"], "05/2026")
        self.assertEqual(r["serie"], 189)

    def test_falha_de_rede_nao_interrompe_e_nao_inventa_data(self):
        # §18.2: fonte indisponível -> aviso discreto, sem substituir por data falsa.
        import _ui_utils as ui
        ui._obter_ultima_competencia_sgs_cache.clear()
        with mock.patch.object(
            iu.requests, "get", side_effect=OSError("sem rede")
        ):
            texto = ui._texto_ultima_competencia_sgs(433)
        self.assertIn("não foi possível consultar", texto)
        self.assertNotIn("/20", texto)  # nenhuma competência inventada

    def test_ist_e_icti_nao_sao_afetados_por_este_helper(self):
        # §18.1: o helper SGS é exclusivo de IPCA/IGP-M; IST/ICTI têm fontes próprias.
        self.assertTrue(hasattr(iu, "obter_ultima_competencia_icti_ipeadata"))
        import _ui_utils as ui
        self.assertTrue(hasattr(ui, "obter_ultima_competencia_ist"))
        self.assertTrue(hasattr(ui, "render_alerta_icti_ipeadata"))


class TestTipoMovimentoAcento(unittest.TestCase):
    """§12 / §27 I,J,K — classificação Python tolerante a acento."""

    def test_acrescimo_com_acento(self):
        self.assertEqual(normalizar_tipo_movimento("Acréscimo"), "ACRESCIMO")

    def test_acrescimo_sem_acento_legado(self):
        self.assertEqual(normalizar_tipo_movimento("Acrescimo"), "ACRESCIMO")

    def test_supressao_com_acento(self):
        self.assertEqual(normalizar_tipo_movimento("Supressão"), "SUPRESSAO")

    def test_supressao_sem_acento_legado(self):
        self.assertEqual(normalizar_tipo_movimento("Supressao"), "SUPRESSAO")


class TestFormulaTemplateAcentoPrefixos(unittest.TestCase):
    """§12.1 — o gerador do template usa prefixos tolerantes a acento."""

    def test_gerador_usa_prefixos_curtos_antes_do_acento(self):
        builder = (ROOT / "tools" / "build_coleta_reajuste_template.py").read_text(encoding="utf-8")
        self.assertIn('LEFT(UPPER(D{r}),3)="ACR"', builder)
        self.assertIn('LEFT(UPPER(D{r}),4)="SUPR"', builder)
        # Não deve mais depender do prefixo de 5 chars que quebrava com "Acréscimo".
        self.assertNotIn('LEFT(UPPER(D{r}),5)="ACRES"', builder)
        # §12.2: dropdown permanece sem acento (validacao legivel por openpyxl);
        # a ENTRADA acentuada e aceita pelas formulas tolerantes (prefixos ACR/SUPR).
        self.assertIn('"D2:D200", "Acrescimo,Supressao"', builder)

    def test_gerador_nao_restaura_decrescimo_no_dropdown(self):
        builder = (ROOT / "tools" / "build_coleta_reajuste_template.py").read_text(encoding="utf-8")
        self.assertNotIn("Decréscimo,", builder)
        self.assertNotIn("Decrescimo,", builder)


class TestUploadTolerantePobreEHardBlock(unittest.TestCase):
    """§20/§21/§22 — matriz HARD × SOFT no diagnostico do upload."""

    @staticmethod
    def _base(ciclo="C1", ano=2024):
        from _coleta_reajuste import gerar_coleta_reajuste
        return gerar_coleta_reajuste(
            {
                "indice": "IPCA",
                "ciclos": [{
                    "ciclo": ciclo,
                    "data_base": f"01/01/{ano}",
                    "financeiro_inicio": f"01/01/{ano + 1}",
                    "percentual_aplicado": 0.045,
                }],
            }
        )

    def test_21_1_template_quase_vazio_aceito_sem_inventar(self):
        from _coleta_reajuste import ler_coleta_reajuste
        diag = ler_coleta_reajuste(self._base())
        self.assertTrue(diag["valido"])                       # aceito p/ diagnostico
        self.assertFalse(diag["pronto_para_consolidar"])      # nada definitivo
        self.assertEqual(diag["contagens"]["competencias_com_valor"], 0)
        self.assertNotIn("total", diag)                       # nenhum valor inventado
        self.assertIn(
            diag["status_base"],
            {"ANALISE_PARCIAL_INFORMACOES_INSUFICIENTES", "APTO_PARA_ANALISE"},
        )

    def test_21_3_apenas_financeiro_nao_gera_valueerror_global(self):
        import io
        from openpyxl import load_workbook
        from _coleta_reajuste import ler_coleta_reajuste
        wb = load_workbook(io.BytesIO(self._base()), data_only=False)
        for row, valor in ((2, 1000.0), (3, 1000.0), (4, 1000.0)):
            wb["financeiro"][f"C{row}"] = valor
        out = io.BytesIO()
        wb.save(out)
        diag = ler_coleta_reajuste(out.getvalue())            # nao deve levantar
        self.assertTrue(diag["valido"])
        self.assertGreaterEqual(diag["contagens"]["competencias_com_valor"], 1)

    def test_22_hard_block_aba_essencial_ausente_rejeita(self):
        import io
        from openpyxl import load_workbook
        from _coleta_reajuste import ler_coleta_reajuste
        wb = load_workbook(io.BytesIO(self._base()), data_only=False)
        wb.remove(wb["itens_Remanesc"])                       # aba estrutural essencial
        out = io.BytesIO()
        wb.save(out)
        diag = ler_coleta_reajuste(out.getvalue())
        self.assertFalse(diag["valido"])                      # HARD BLOCK
        self.assertEqual(diag["status_base"], "ARQUIVO_ESTRUTURALMENTE_INVALIDO")


class TestTemplateRealDatasEFormulas(unittest.TestCase):
    """§10/§11/§12.1 — estado do template oficial real (leitura openpyxl)."""

    @classmethod
    def setUpClass(cls):
        from openpyxl import load_workbook
        cls.wb = load_workbook(ROOT / "templates" / "COLETA_REAJUSTE_OFICIAL.xlsx")

    def test_itens_pc_data_nao_usa_literal_yyyy_escapado(self):
        # §10: o bug era o formato 'dd/mm/\\y\\y\\y\\y' (ano vira texto literal).
        fmt = self.wb["itens_PC"]["B2"].number_format
        self.assertNotIn("\\y", fmt)
        self.assertNotIn("yyyy", fmt.replace("\\", ""))

    def test_aditivos_data_nao_usa_literal_yyyy_escapado(self):
        # §11: aditivos!B passou a data completa (dd/mm/aaaa), sem literal.
        fmt = self.wb["aditivos"]["B2"].number_format
        self.assertNotIn("\\y", fmt)

    def test_template_formulas_toleram_acento(self):
        # §12.1: cadeia DELTA (L) + CHECK (M) com prefixos ACR/SUPR.
        l2 = self.wb["aditivos"]["L2"].value
        m2 = self.wb["aditivos"]["M2"].value
        self.assertIn('LEFT(UPPER(D2),3)="ACR"', l2)
        self.assertIn('LEFT(UPPER(D2),4)="SUPR"', l2)
        self.assertIn('LEFT(UPPER(D2),3)<>"ACR"', m2)
        self.assertNotIn('="ACRES"', l2)


class TestAppTestStatusEDocumentos(unittest.TestCase):
    """§28/§5/§7/§27 U-W — prova end-to-end no runtime Streamlit (AppTest)."""

    CAMINHO_PAGINA = str(ROOT / "pages" / "03_Valor_Global.py")

    def _seed_soft_block(self, at, status):
        docs = {
            chave: {
                "nome": titulo, "estado": "completo", "rotulo": "Disponível com ressalvas",
                "classificacao": "", "motivo": "", "habilitado": True,
            }
            for chave, titulo in (
                ("sumario_executivo", "Sumário Executivo"),
                ("adequacao_orcamentaria", "Adequação Orçamentária"),
                ("despacho_saneador", "Despacho Saneador"),
                ("termo_apostila", "Termo de Apostila"),
                ("garantia_contratual", "Garantia Contratual"),
                ("dou", "DOU"),
            )
        }
        diagnostico = {
            "status_base": status,
            "pronto_para_consolidar": False,
            "metadados": {"indice": "IPCA", "ciclos_em_analise": ["C2"]},
        }
        at.session_state["assinatura_processada_upload_docs"] = "sig"
        at.session_state["assinatura_upload_docs"] = "sig"
        at.session_state["resultado_valor_global"] = {
            "capacidades": {"documentos": docs},
            "diagnostico_coleta": diagnostico,
            "variacao_acumulada": 0.05,
        }
        at.session_state["diagnostico_coleta_v2"] = diagnostico

    def test_inconsistencia_mostra_banner_e_mantem_tres_documentos(self):
        from streamlit.testing.v1 import AppTest

        at = AppTest.from_file(self.CAMINHO_PAGINA, default_timeout=90)
        self._seed_soft_block(at, "ANALISE_COM_INCONSISTENCIAS")
        at.run()
        self.assertFalse(at.exception, msg=str(getattr(at, "exception", "")))
        avisos = " \n ".join(w.value for w in at.warning)
        self.assertIn("inconsistências que exigem revisão", avisos)
        markdowns = " \n ".join(m.value for m in at.markdown)
        for titulo in ("Sumário Executivo", "Despacho Saneador", "Termo de Apostila"):
            self.assertIn(titulo, markdowns)

    def test_insuficiencia_mostra_banner_proprio(self):
        from streamlit.testing.v1 import AppTest

        at = AppTest.from_file(self.CAMINHO_PAGINA, default_timeout=90)
        self._seed_soft_block(at, "ANALISE_PARCIAL_INFORMACOES_INSUFICIENTES")
        at.run()
        self.assertFalse(at.exception, msg=str(getattr(at, "exception", "")))
        avisos = " \n ".join(w.value for w in at.warning)
        self.assertIn("informações insuficientes", avisos)


if __name__ == "__main__":
    unittest.main()
