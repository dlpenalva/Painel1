"""ETAPA 26B — testes focais do hotfix RESULTADOS/VTA-PC pos-homologacao CLEMAR.

Cobrem as regressoes confirmadas:
  - chave canonica de metodo (RESULTADOS/Python consomem a MESMA decisao);
  - decoupling RETROATIVO x EXECUCAO HISTORICA no motor de composicao do VTA
    (um PC fora do retroativo continua compondo o VTA);
  - fix das formulas da aba RESULTADOS no template (ramo PC no remanescente,
    B4 normalizador canonico, VTA-PC ligada, sem "PCs — SEM QUANTIDADE"),
    validado estaticamente (sem COM) para blindar a regressao;
  - itens_PC vermelho continua na regra canonica ($L2="Nao"), nao no ciclo.

O arquivo real do usuario nunca e commitado; o golden numerico do motor esta
em test_vta_pc_composicao.py. Aqui blindamos as decisoes de metodo/decoupling e
as formulas do template.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl

import _metodo_apuracao as M
import _motor_composicao_vta as C

ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "templates" / "COLETA_REAJUSTE_OFICIAL.xlsx"


# ---- E. Chave canonica de metodo: rotulos antigos e novos convergem ---------
def test_e_chave_canonica_pc_todos_os_rotulos():
    alvo = M.normalizar_metodo(M.ROTULO_PC)
    assert alvo == "pc"
    for rotulo in (
        "PC (Pedidos de Compra)", "Pedidos de Compras", "Pedidos de Compra",
        "PCs", "pc", "  pc  ",
    ):
        assert M.normalizar_metodo(rotulo) == "pc", rotulo


def test_e_chave_canonica_financeiro_e_consumidos():
    fin = M.normalizar_metodo(M.ROTULO_PRINCIPAL)
    for rotulo in ("Principal (Financeiro)", "Principal", "Financeiro"):
        assert M.normalizar_metodo(rotulo) == fin, rotulo
    cons = M.normalizar_metodo(M.ROTULO_CONSUMIDOS)
    assert M.normalizar_metodo("Itens Consumidos") == cons
    assert fin != "pc" and cons != "pc"


# ---- A/B. Decoupling: PC fora do retroativo continua no VTA ------------------
def _leitura_pc(entra_c1="Sim"):
    # Fixture minimo: C0 fisico + C1/C2/C3 por PC; vigente C4 remanescente.
    itens = [
        (1, 15022, 17408.84, 95, 0, 0),
        (2, 2600, 3013.11, 95, 95, 45),
        (3, 6400, 7416.89, 95, 95, 54),
        (4, 2548, 2952.85, 18, 11, 0),
        (5, 2314, 2681.67, 8, 5, 0),
        (6, 4800, 5562.67, 40, 24, 10),
        (7, 2600, 3013.11, 16, 6, 0),
        (8, 3600, 4172, 108, 71, 63),
    ]
    posicao = [{
        "ITEM": it, "VU_ORIGINAL": vu0,
        "QTD_REM_AJUSTADA_C0": q0, "QTD_REM_AJUSTADA_C1": q1,
        "QTD_REM_AJUSTADA_C2": 0, "QTD_REM_AJUSTADA_C3": 0,
        "QTD_REM_AJUSTADA_C4": q4,
    } for (it, vu0, vu4, q0, q1, q4) in itens]
    historico = [{
        "item": it,
        "vu_ciclos": {"VU_C0": vu0, "VU_C1": vu0, "VU_C2": vu0,
                      "VU_C3": vu0, "VU_C4": vu4},
    } for (it, vu0, vu4, q0, q1, q4) in itens]
    # C1 marcado como fora do retroativo (efeito "Nao") mas entra_no_calculo=Sim:
    # continua na execucao historica do VTA.
    itens_pc = [
        {"ciclo": "C1", "valor_pc": 111286.55, "valor_atualizado": 111286.55,
         "efeito_financeiro_pc": "Nao", "entra_no_calculo": entra_c1},
        {"ciclo": "C2", "valor_pc": 302125.82, "valor_atualizado": 308057.35,
         "efeito_financeiro_pc": "Sim", "entra_no_calculo": "Sim"},
        {"ciclo": "C3", "valor_pc": 144559.73, "valor_atualizado": 155516.79,
         "efeito_financeiro_pc": "Sim", "entra_no_calculo": "Sim"},
    ]
    return {
        "controle": {"modo": "pc", "ciclo_vigente": "C4"},
        "parametros_v10": {"por_ciclo": {}},
        "posicao_contratual": {"itens": posicao},
        "historico_vu": {"itens": historico},
        "itens_pc_v10": {"itens": itens_pc},
    }


def test_ab_pc_efeito_nao_continua_no_vta():
    """C1 com efeito financeiro 'Nao' (retroativo zero) segue compondo o VTA."""
    comp = C.montar_composicao_vta(_leitura_pc(entra_c1="Sim"))
    por_ciclo = {l["ciclo"]: l["valor_atualizado"]
                 for l in comp["execucao_por_ciclo"]}
    assert por_ciclo.get("C1") == 111286.55  # presente no VTA apesar de efeito Nao
    assert comp["vta_composicao"] == 3117293.40


def test_ab_entra_no_calculo_nao_exclui_do_vta():
    """entra_no_calculo='Nao' (decisao explicita) remove a parcela do VTA;
    'Sim' mantem. A cor/efeito financeiro NAO governam a inclusao no VTA."""
    incluido = C.montar_composicao_vta(_leitura_pc(entra_c1="Sim"))
    excluido = C.montar_composicao_vta(_leitura_pc(entra_c1="Nao"))
    assert excluido["vta_composicao"] == incluido["vta_composicao"] - 111286.55


# ---- F/G/H (template estatico). Ramo PC no remanescente + B4 canonica -------
def _resultados():
    wb = openpyxl.load_workbook(TEMPLATE, data_only=False)
    return wb["MEMORIA_RESULTADOS"]


def test_f_b4_normalizador_canonico():
    b4 = _resultados()["B4"].value
    # B4 mapeia rotulos (antigos e novos) para o token canonico via SEARCH.
    assert 'SEARCH("Pedido"' in b4 and 'SEARCH("PC"' in b4
    assert 'SEARCH("Financeiro"' in b4 and 'SEARCH("Principal"' in b4
    assert 'SEARCH("Itens"' in b4


def test_g_remanescente_tem_ramo_pc():
    ws = _resultados()
    for cel, fonte in (("B35", "B32"), ("C35", "C32")):
        f = ws[cel].value
        assert '$B$4="PCs"' in f and '$B$4="Financeiro"' in f, cel
        assert fonte in f, cel
    # valor com reajuste do PC vem do bloco VTA-PC item a item (T23), nao de D32.
    assert '$B$4="PCs"' in ws["D35"].value and "$T$23" in ws["D35"].value
    # F35 nao pode mais rotular PC como "SEM QUANTIDADE".
    assert "SEM QUANTIDADE" not in ws["F35"].value
    assert '"INFORMADO"' in ws["F35"].value


def test_h_vta_final_liga_bloco_pc():
    b26 = _resultados()["B26"].value
    assert '$B$4="PCs"' in b26 and "$T$25" in b26
    # F36 (status) nao pode forcar CALCULO MANUAL para PC via rotulo antigo.
    assert "Pedidos de Compras" not in _resultados()["F36"].value


# ---- I. Helpers tecnicos ocultos, resultados visiveis ----------------------
def _col_oculta(ws, letra):
    # openpyxl guarda faixas (<col min max hidden>); lookup por letra nao resolve
    # spans. Verifica cobertura por qualquer dimensao oculta que contenha a coluna.
    from openpyxl.utils import column_index_from_string as _ci
    idx = _ci(letra)
    return any(d.hidden and d.min <= idx <= d.max
              for d in ws.column_dimensions.values())


def test_i_helpers_ocultos():
    ws = _resultados()
    # Helpers tecnicos: G:O e X:Y (26B) + S:T (bloco VTA-PC, Etapa 26C).
    for col in ("G", "H", "I", "J", "K", "L", "M", "N", "O", "S", "T", "X", "Y"):
        assert _col_oculta(ws, col), f"coluna {col} deveria estar oculta"


# ---- itens_PC vermelho continua na regra canonica ($L2="Nao") --------------
def test_itens_pc_vermelho_regra_canonica():
    wb = openpyxl.load_workbook(TEMPLATE, data_only=False)
    ws = wb["itens_PC"]
    formulas = []
    for rng, regras in ws.conditional_formatting._cf_rules.items():
        for r in regras:
            if r.formula:
                formulas.append(r.formula[0])
    assert any(f == '$L2="Nao"' for f in formulas)  # efeito financeiro, nao ciclo
    assert not any("CICLO" in f.upper() and "C1" in f for f in formulas)
