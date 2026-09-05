"""Testes permanentes de integridade OOXML do template oficial da Coleta.

Protegem contra a regressao de corrupcao identificada na Etapa 3:
mc:Ignorable com prefixos nao declarados, marcador repairLoad, perda de
formulas/estilos e descaracterizacao da aba financeiro.
"""
from __future__ import annotations

import gc
import os
import re
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

import pytest
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "templates" / "COLETA_REAJUSTE_OFICIAL.xlsx"

NS = {
    "m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
}

FORMULAS_POR_ABA = {
    # comparativo_VTA: aba de referencia adicionada via Excel COM (1a posicao
    # fisica). cobertura_temporal 14->15 (+ linha Metodo de apuracao).
    # Etapa 26F: calculos antigos preservados em MEMORIA_RESULTADOS e nova
    # RESULTADOS executiva com formulas de apresentacao.
    "comparativo_VTA": 1407,
    "CONTROLE": 6,
    "parametros": 32,
    "financeiro": 291,
    # 26H.1/26H.2: +199 formulas pre-semeadas de base zero visual em
    # B2:B200. FRONTEIRA FUNCIONAL: linha 200 = ultima linha de cadastro
    # contratual (A2:A200 em toda a cadeia); linha 201 = linha extra do
    # total dinamico de D/F — FORA da capacidade funcional e sem automacao
    # de input (o off-by-one B201 da 26H.1 foi apontado pela auditoria e
    # removido na 26H.2).
    # Temporalidade por DATA_EFEITO: +199 do espelho local BI2:BI200
    # (CICLO_NASCIMENTO_DATA). A formatacao condicional dos 4 estados precisa
    # referenciar a PROPRIA aba: o Excel migra para a extensao x14 (invisivel ao
    # openpyxl) qualquer regra que aponte para outra planilha.
    # Linha dinamica TOTAL (correcao definitiva): +200 da coluna T
    # (VALOR_EXECUTADO_C4), que estava sem formula em 2:200 e sem o fallback
    # de lotacao maxima em T201 (owner:
    # tools/aplicar_total_dinamico_itens_remanesc.py).
    "itens_Remanesc": 9793,
    # VTA-C2: +199 (V2:V200, coluna auxiliar "remanescente atualizado
    # (base) do item" por linha — evita bug real de SUMPRODUCT com mascara
    # booleana usado antes em C33/D33 do metodo Itens).
    "itens_Consumidos": 2005,
    # Etapa 26G: grade escalada para a capacidade canonica (5.000 PCs
    # x 8 colunas de formula) + resumo lateral N2:T6.
    # 45042 = 40042 anteriores + 5000 da coluna U (VALOR_CONSIDERADO, U2:U5001).
    # PC-UX-1: sete formulas residuais fecham os universos ate/apos o corte.
    # Reauditoria NOVO-02: +7 formulas para o residual do universo integral
    # no Quadro 1; o TOTAL apenas muda da linha 8 para a 9.
    "itens_PC": 45098,
    "aditivos": 1393,
    # Fonte unica da posicao fisica: +199 de QTD_REM_ATUAL (B2:B200), que deixou
    # de ser digitada e passou a buscar CICLO_EM_EXECUCAO por ITEM, e +3 do
    # painel (I9 data da posicao fisica, I10 status, I11 validacao temporal).
    "posicao_referencia": 2797,
    # 26H: +398 formulas das colunas tecnicas ocultas Y (CICLO_NASCIMENTO)
    # e Z (EH_NOVO_ITEM), linhas 2:200.
    # Temporalidade por DATA_EFEITO: +2388 das colunas ocultas AA:AL
    # (AA data de efeito, AB:AF delta posterior a abertura por ciclo,
    # AG:AK qtd contratual na abertura por ciclo, AL ciclo de nascimento
    # POR DATA), 12 colunas x 199 linhas 2:200.
    "posicao_contratual": 7562,
    # Etapa VTA-posicoes: +1800 do bloco POSICAO ATUAL (AUTO) Q:Y (9 colunas
    # x 200 linhas 3:202) via INDIRECT+ISERROR sobre CICLO_EM_EXECUCAO.
    # Temporalidade por DATA_EFEITO: +800 do bloco Z:AC (aplicabilidade
    # temporal na abertura adotada), 4 colunas x 200 linhas 3:202.
    "itens_RC": 5800,
    "historico_VU": 3592,
    # Etapa VTA: -1 (linha "Fonte temporal de conferencia" removida; sem consumidor).
    # Padronizacao das datas: +2 (B4 data de geracao/analise, agora automatica,
    # e B17 EXISTEM PCS POSTERIORES AO CORTE?).
    "cobertura_temporal": 16,
    # 3762 anteriores + 11 referencias para a tabela manual unica.
    # 26G: +5 (T26/T27 completude do remanescente; T28:T30 PCs sem efeito).
    # Etapa VTA-posicoes: +212 do bloco auxiliar das 3 referencias
    # (W41:W52 = 12 formulas + AB2:AB201 = 200 formulas). B26/T25 intactos.
    # Temporalidade por DATA_EFEITO: +405 da decomposicao temporal da FORMA 2
    # (AC2:AC201 abertura temporal + AD2:AD201 alteracoes posteriores = 400,
    # mais W53:W57 = 5; W48 e reescrita, nao acrescentada). B26/T25 intactos.
    # Correcao dos ciclos: +6 auxiliares (T31 limite da data de corte e
    # T33:T37, as cinco medidas canonicas). T21/T22/W48/X2:X201 sao reescritas
    # (fonte trocada para o VALOR CONSIDERADO e para as aberturas temporais),
    # nao acrescentadas. B26/T25 intactos.
    # HOTFIX RETRO/VTA: +7 do bloco de apoio method-aware (W61:W65 execucao
    # Financeiro considerada por ciclo via SUMIF financeiro!E; W66/W67
    # historico pelo metodo oficial ate o vigente/abertura). W48/W50 sao
    # reescritas (passam a consumir W67/W66), nao acrescentadas. B26/T25
    # intactos.
    # VTA-M2: +1 (D20 = "Desembolsado (Financeiro)", SOMA de financeiro!C
    # informado, unica formula nova; B26/B28 reescritos apenas no ramo
    # Financeiro; T21-T25/PC e Consumido intactos).
    # VTA-C2: +1 (F20 = "Execucao Consumida Atualizada", soma de
    # itens_Consumidos!VALOR_CONS_C0..C4 com gate de presenca; unica
    # formula nova. B26/B28/C33/D33 sao reescritas no ramo Itens, nao
    # acrescentadas. Financeiro/PC intactos).
    # PR 2: +1 formula publicada (name RETROATIVO_POTENCIAL_PC).
    # VTA-POT-1: +3 (4411 -> 4414) — T41 (retroativo potencial APURADO, que
    # pode ser negativo), T39 = MAX(T41,0) (piso prudencial: o potencial
    # elegivel ao VTA,
    # espelho do motor Python) e T40 (VTA-PC antes da parcela potencial, o
    # subtotal demonstrado em RESULTADOS!C86). T25 e REESCRITA (passa a somar
    # T39), nao acrescentada; Financeiro/Itens e todo o resto ficam intactos.
    "MEMORIA_RESULTADOS": 4414,
    # 57 do prototipo + 4 selos por tabela + 1 premissa da estimativa - 1
    # helper J4 removido (status global agora agrega os selos H8/H14/H24/H33).
    # 26G: +5 (linha executiva A23:E23 dos PCs sem efeito financeiro).
    # Etapa VTA-posicoes: +7 liquidas na Tabela 1 (3 referencias +
    # reconciliacao: B10/C10/H10, B11/C11/H11, B12, B13/H13 = 9 novas,
    # menos as 2 antigas B10/B11 substituidas). H8 preservado.
    # Correcao dos ciclos: +12 da secao 5 (as doze medidas canonicas com nomes
    # claros, B55:B66). B16:B20 e B36 sao reescritas — passam a respeitar a
    # data de corte e a posicao fisica —, nao acrescentadas.
    # Etapa 50 (leiaute final homologado 50.1-50.3): +23 formulas
    # EXCLUSIVAMENTE de apresentacao (85 -> 108): contexto A3/C3/E3/G3
    # (metodo humano, ciclo, corte, variacao), selo e contagem G1/G2,
    # pendencias A7/B7, auxiliares ocultos J5/J6, titulo dinamico do ciclo
    # A33, chips dos cards C4/E4/H4 (espelhos de H8/H14/H33), valores dos
    # cards C5/E5/G5 (espelhos de B10/D22/B38), avisos "Aguardando dado"
    # C35:C38 e espelhos das notas condicionais E16/E35 (linhas separadoras
    # 23 e 39 ficam visualmente brancas). Nenhuma delas calcula negocio.
    # Etapa 51A: +1 formula de apresentacao (108 -> 109): E6 (rodape do card
    # do retroativo) vira o estado vazio que orienta selecionar o metodo /
    # informar a base, lendo apenas $D$22 e $J$6 ja existentes.
    # HOTFIX RETRO/VTA: total inalterado (109): E5 perde a formula (mesclada
    # em D5, que passa a exibir '=$D$22' sob o rotulo do card) e J8 ganha a
    # ancora do ciclo ('=UPPER(CONTROLE!$B$2)', coluna J oculta). C3/C10/C11
    # sao reescritas (repontagem e textos method-aware), nao acrescentadas.
    # VTA-M2: +20 (bloco novo CONFERENCIA DA EXECUCAO, linhas 71-77:
    # Desembolsado informado, Execucao teorica por ciclo, Diferenca e
    # Status para C0-C4, mais texto de metodologia nas linhas 68-70).
    # VTA-M2.1: +2 (A70 metodologia e A71 titulo do bloco 8 deixam de ser
    # texto fixo e passam a FORMULA condicional a MEMORIA_RESULTADOS!$B$4,
    # para RESULTADOS nao explicar o metodo Financeiro quando o metodo
    # selecionado for PCs ou Itens/Consumido).
    # VTA-U2: +8 (bloco novo "9. COMO E FORMADO O VTA?", linhas 79-87 —
    # B83/C83 executado apurado, B84/C84 ajustes ainda devidos, B85
    # remanescente atualizado, B86 VTA Oficial e B87/C87 a conferencia que
    # prova a reconciliacao das parcelas com o VTA. C5 e B63 sao REPONTADAS
    # para VTA_FINAL (ja eram formulas, nao somam), e o bloco 8 tem suas
    # formulas reescritas com o novo rotulo de ausencia de base (nao somam).
    # RESULTADOS-ROLLBACK-1: a camada de apresentacao das linhas 90-166
    # (PR 2) foi aposentada e a aba voltou a apresentacao anterior. Restam
    # as 143 formulas do motor tecnico — o mesmo numero do checkpoint
    # f8296f7. A expectativa de 139 ja estava defasada antes do PR 2.
    # RESULTADOS-FINAL-1: +2 (143 -> 145), ambas de APRESENTACAO e ambas na
    # faixa E22:H22, que estava vazia e fora de merge/CF/validacao: G22 exibe
    # o retroativo potencial lendo so o nome definido ja existente
    # RETROATIVO_POTENCIAL_PC, e E22 e o rotulo/observacao que depende de G22.
    # Nenhuma formula economica foi criada, movida ou recalculada — A70 e B64
    # foram apenas reescritas (continuam sendo formulas, nao somam).
    # PC-UX-1 libera A70 para criar o respiro entre os blocos 7 e 8.
    # PC-UX-1: A70 e B15/C15 voltam a ser condicionais; A71 segue condicional.
    # VTA-POT-1: +2 (148 -> 150), ambas de APRESENTACAO e ambas dentro do
    # quadro 9 (a aba continua terminando na linha 87 — o contrato do rollback
    # da camada UX2 segue intacto). A84 deixa de ser texto fixo e vira formula
    # method-aware: a parcela "(+)" passa a ser o retroativo POTENCIAL no
    # metodo PC. C86 deixa de ser texto fixo e demonstra, so no metodo PC com
    # potencial, "VTA antes da parcela potencial + POTENCIAL = VTA". B84/C84
    # sao REESCRITAS, nunca acrescentadas; as ancoras EXECUTADO_APURADO (B83),
    # AJUSTES_DEVIDOS (B84), VTA_FINAL (B86) e CONFERENCIA_FORMACAO_VTA (B87)
    # ficam nas mesmas coordenadas.
    "RESULTADOS": 150,
}


def _partes_xml(z: zipfile.ZipFile) -> list[str]:
    return [n for n in z.namelist() if n.endswith((".xml", ".rels"))]


def _abas_e_partes(z: zipfile.ZipFile) -> list[tuple[str, str]]:
    wb = ET.fromstring(z.read("xl/workbook.xml"))
    rels = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
    rid2t = {
        rel.get("Id"): rel.get("Target")
        for rel in rels.findall("rel:Relationship", NS)
    }
    resultado = []
    for aba in wb.findall("m:sheets/m:sheet", NS):
        alvo = rid2t[aba.get("{%s}id" % NS["r"])].lstrip("/")
        if not alvo.startswith("xl/"):
            alvo = "xl/" + alvo
        resultado.append((aba.get("name"), alvo))
    return resultado


def test_template_existe():
    assert TEMPLATE.is_file()


def test_xml_bem_formado_em_todas_as_partes():
    with zipfile.ZipFile(TEMPLATE) as z:
        for nome in _partes_xml(z):
            ET.fromstring(z.read(nome))


def test_mc_ignorable_somente_com_prefixos_declarados():
    padrao = re.compile(rb'mc:Ignorable="([^"]*)"')
    with zipfile.ZipFile(TEMPLATE) as z:
        for nome in _partes_xml(z):
            dados = z.read(nome)
            encontrado = padrao.search(dados)
            if not encontrado:
                continue
            for prefixo in encontrado.group(1).decode().split():
                declaracao = f'xmlns:{prefixo}='.encode()
                assert declaracao in dados, (
                    f"{nome}: mc:Ignorable referencia prefixo nao "
                    f"declarado {prefixo!r}"
                )


def test_sem_marcador_repairload():
    with zipfile.ZipFile(TEMPLATE) as z:
        contaminadas = [n for n in z.namelist() if b"repairLoad" in z.read(n)]
    assert contaminadas == []


def test_sem_vinculos_externos():
    with zipfile.ZipFile(TEMPLATE) as z:
        assert [n for n in z.namelist() if "externalLink" in n] == []
        rels = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
        externos = [
            rel.get("Target")
            for rel in rels.findall("rel:Relationship", NS)
            if rel.get("TargetMode") == "External"
        ]
        assert externos == []


def test_contagem_de_formulas_por_aba():
    with zipfile.ZipFile(TEMPLATE) as z:
        abas = _abas_e_partes(z)
        assert [nome for nome, _ in abas] == list(FORMULAS_POR_ABA)
        for nome, parte in abas:
            raiz = ET.fromstring(z.read(parte))
            quantidade = len(raiz.findall(".//m:f", NS))
            assert quantidade == FORMULAS_POR_ABA[nome], (
                f"{nome}: esperava {FORMULAS_POR_ABA[nome]} formulas, "
                f"encontrei {quantidade}"
            )


def test_limites_minimos_de_estilos():
    with zipfile.ZipFile(TEMPLATE) as z:
        estilos = ET.fromstring(z.read("xl/styles.xml"))

    def contar(tag: str) -> int:
        elemento = estilos.find("m:" + tag, NS)
        return len(elemento) if elemento is not None else 0

    assert contar("cellXfs") >= 200
    assert contar("numFmts") >= 10
    assert contar("dxfs") >= 17


def test_financeiro_preservada():
    wb = load_workbook(TEMPLATE)
    ws = wb["financeiro"]
    formulas = sum(
        1
        for linha in ws.iter_rows()
        for celula in linha
        if isinstance(celula.value, str) and celula.value.startswith("=")
    )
    assert formulas == 291  # 72 fórmulas B + 216 DEF linhas 2-73 + 3 SUM em C74/E74/F74
    validacoes = [
        (dv.type, str(dv.sqref)) for dv in ws.data_validations.dataValidation
    ]
    assert validacoes == [("list", "G2:G73")]
    condicionais = sorted(str(rng.sqref) for rng in ws.conditional_formatting)
    assert condicionais == ["A2:G73"]


def test_itens_pc_efeito_financeiro_aplicado():
    wb = load_workbook(TEMPLATE)
    ws = wb["itens_PC"]
    assert ws["L1"].value == "EFEITO_FINANCEIRO_PC"
    assert isinstance(ws["L2"].value, str) and ws["L2"].value.startswith("=IF(")
    # 26G: grade ate a capacidade canonica — L101 tem formula.
    assert str(ws["L101"].value).startswith("=")
    validacoes = [
        (dv.type, str(dv.sqref)) for dv in ws.data_validations.dataValidation
    ]
    assert validacoes == [("list", "G2:G5001")]
    par = wb["parametros"]
    assert par["H1"].value == "INICIO_EFEITO_FINANCEIRO"
    assert {par.cell(r, 8).number_format for r in range(2, 7)} == {
        "mm/yyyy;@"
    }


def test_aditivos_dropdown_tipo_alteracao_sem_decrescimo():
    """Ajuste final: aditivos!D2:D200 lista apenas Acrescimo/Supressao.

    O dropdown de TIPO DE ALTERACAO FORMALIZADA nao pode mais oferecer
    "Decrescimo"; deve conter exclusivamente Acrescimo e Supressao, cobrindo
    todo o intervalo D2:D200.
    """
    wb = load_workbook(TEMPLATE)
    ws = wb["aditivos"]
    dvs_d = [
        dv for dv in ws.data_validations.dataValidation
        if dv.type == "list" and "D2:D200" in str(dv.sqref)
    ]
    assert len(dvs_d) == 1, "esperada uma validacao de lista cobrindo D2:D200"
    dv = dvs_d[0]
    assert str(dv.sqref) == "D2:D200"
    itens = [t.strip() for t in dv.formula1.strip('"').split(",")]
    assert itens == ["Acrescimo", "Supressao"]
    assert not any("Decr" in i for i in itens)


@pytest.mark.skipif(
    os.environ.get("RUN_EXCEL_INTEGRATION") != "1",
    reason="defina RUN_EXCEL_INTEGRATION=1 para executar Excel COM",
)
def test_abertura_e_reabertura_sem_reparo_no_excel_real():
    client = pytest.importorskip("win32com.client")
    pythoncom = pytest.importorskip("pythoncom")
    pythoncom.CoInitialize()
    excel = client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = True
    try:
        for rodada in range(2):
            wb = excel.Workbooks.Open(str(TEMPLATE), UpdateLinks=0, ReadOnly=True)
            assert wb.Worksheets.Count == 15, f"rodada {rodada}"
            wb.Close(False)
            del wb
    finally:
        excel.Quit()
        del excel
        gc.collect()
        pythoncom.CoUninitialize()
