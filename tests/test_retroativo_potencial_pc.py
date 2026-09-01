# -*- coding: utf-8 -*-
"""RETROATIVO_POTENCIAL_PC (contrato do name e paridade com a web).

RESULTADOS-ROLLBACK-1: os testes de leiaute da camada humana das linhas
90-166 foram removidos junto com a apresentacao que eles protegiam. O que
permanece aqui independe dela: o contrato do defined name, a paridade
economica com a grandeza que a web ja publicava antes do PR 2 e o status
oficial lido de RESULTADOS!B3.

Grandeza EXCLUSIVAMENTE informativa: expoe no XLS o mesmo numero que a web ja
publica como `retroativo_potencial`, sem entrar em nenhuma soma oficial.

Cadeia canonica que este name reproduz:

    itens_PC!J (DELTA_POTENCIAL, por PC)
      -> leitor: col_delta = _col(mapa, "DELTA_POTENCIAL", "RETROATIVO POTENCIAL")
      -> registro["delta_potencial"]
      -> _totais_canonicos_pc -> blocos["ate_o_corte"]["delta_potencial"]
      -> _resultado_consolidado -> consolidado["retroativo_potencial"]
      -> pages/03_Valor_Global.py e pages/12_Adequacao_Orcamentaria.py

Os outros dois filtros do Python sao INERTES na grade oficial e por isso nao
aparecem na formula: `itens_PC` nao possui coluna ENTRA_NO_CALCULO (o default
e "Sim") e DESCARTADO_DUPLICIDADE so e atribuido pela via fiscal
(STATUS_PAGAMENTO_PC / VALOR_EFETIVAMENTE_PAGO), colunas ausentes da grade.
"""
from __future__ import annotations

import datetime as dt
import io
from pathlib import Path

import pytest
from openpyxl import load_workbook

from _leitor_masterfile_v10 import ler_masterfile_v10

ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "templates" / "COLETA_REAJUSTE_OFICIAL.xlsx"

NOME = "RETROATIVO_POTENCIAL_PC"
DESTINO = "MEMORIA_RESULTADOS!$T$38"
FORMULA = (
    '=ROUND(SUMIFS(itens_PC!$J$2:$J$5001,'
    'itens_PC!$B$2:$B$5001,"<="&$T$31),2)'
)

# Os 14 names publicados pelo PR #135 precisam sobreviver intactos.
NOMES_PR1 = {
    "EXECUTADO_APURADO": "RESULTADOS!$B$83",
    "AJUSTES_DEVIDOS": "RESULTADOS!$B$84",
    "CONFERENCIA_FORMACAO_VTA": "RESULTADOS!$B$87",
    "PC_TOTAL_CADASTRADO": "MEMORIA_RESULTADOS!$T$33",
    "PC_TOTAL_ATE_CORTE": "MEMORIA_RESULTADOS!$T$34",
    "PC_TOTAL_COM_EFEITO": "MEMORIA_RESULTADOS!$T$36",
    "PC_TOTAL_SEM_EFEITO": "MEMORIA_RESULTADOS!$T$37",
    "AUDITORIA_SITUACAO_ATUAL_CONTRATO": "MEMORIA_RESULTADOS!$W$50",
    "AUDITORIA_ULTIMA_REFERENCIA_ABERTURA": "MEMORIA_RESULTADOS!$W$48",
    "AUDITORIA_COMPARATIVO_INTEGRAL": "comparativo_VTA!$B$208",
    "AUDITORIA_DIFERENCA_REFERENCIAS": "MEMORIA_RESULTADOS!$W$51",
    "AUDITORIA_SITUACAO_ATUAL_STATUS": "RESULTADOS!$H$10",
    "AUDITORIA_ABERTURA_STATUS": "RESULTADOS!$H$11",
    "AUDITORIA_CONFERENCIA_STATUS": "MEMORIA_RESULTADOS!$W$52",
}


@pytest.fixture(scope="module")
def wb():
    return load_workbook(TEMPLATE)


# --------------------------------------------------------------- FASE 1
def test_name_publicado_no_destino_aprovado(wb):
    assert NOME in wb.defined_names
    assert wb.defined_names[NOME].value == DESTINO


def test_formula_de_t38_e_a_definicao_canonica(wb):
    assert wb["MEMORIA_RESULTADOS"]["T38"].value == FORMULA


def test_rotulo_de_apoio_em_s38(wb):
    rotulo = wb["MEMORIA_RESULTADOS"]["S38"].value
    assert isinstance(rotulo, str) and "potencial" in rotulo.lower()


def test_d45_permanece_com_a_semantica_anterior(wb):
    """D45 NAO pode ser alterada: continua somando a coluna inteira."""
    assert wb["MEMORIA_RESULTADOS"]["D45"].value == (
        "=ROUND(SUM(itens_PC!$J$2:$J$5001),2)"
    )


def test_names_do_pr1_sobrevivem_e_t35_segue_sem_nome(wb):
    for nome, destino in NOMES_PR1.items():
        assert wb.defined_names[nome].value == destino, nome
    assert not any("$T$35" in str(d.value) for d in wb.defined_names.values())
    assert not [n for n, d in wb.defined_names.items() if "[" in str(d.value)]


def test_vta_atualizacao_cheia_continua_fora_da_cadeia_oficial(wb):
    assert "VTA_ATUALIZACAO_CHEIA" in wb.defined_names


# --------------------------------------------------------------- FASE 3
# RESULTADOS-ROLLBACK-1: a camada humana das linhas 90-166 foi aposentada e
# com ela sumiram os unicos consumidores do name no workbook. O name segue
# publicado (grandeza terminal, disponivel para leitura), mas agora NENHUMA
# formula pode cita-lo — se voltar a aparecer, ou a apresentacao descartada
# foi reintroduzida, ou o potencial entrou numa cadeia de calculo.


def test_potencial_nao_e_citado_por_nenhuma_formula(wb):
    """O name existe para ser lido, nunca somado nem exibido pela aba."""
    citacoes = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cel in row:
                v = cel.value
                if isinstance(v, str) and v.startswith("=") and NOME in v:
                    citacoes.append(f"{ws.title}!{cel.coordinate}")
    assert citacoes == [], f"{NOME} citado por {citacoes}"


def test_t38_nao_e_citada_por_nenhuma_formula(wb):
    """Ninguem pode somar T38: e saida terminal, nao insumo."""
    citam = []
    ws_mem = wb["MEMORIA_RESULTADOS"]
    for row in ws_mem.iter_rows():
        for cel in row:
            v = cel.value
            if not (isinstance(v, str) and v.startswith("=")):
                continue
            if cel.coordinate == "T38":
                continue
            if "$T$38" in v or "T38" in v.replace("$", ""):
                citam.append(cel.coordinate)
    assert citam == [], f"T38 consumida por {citam}"


def test_celulas_canonicas_do_vta_e_do_retro_intactas(wb):
    """As formulas que produzem VTA e retroativo nao foram tocadas."""
    ws = wb["RESULTADOS"]
    assert ws["C5"].value == '=IF(VTA_FINAL="","",VTA_FINAL)'
    assert ws["D5"].value == "=$D$22"
    assert ws["D22"].value == '=IFERROR(RETRO_OFICIAL,"")'
    assert ws["B86"].value == '=IF(VTA_FINAL="","",VTA_FINAL)'
    mem = wb["MEMORIA_RESULTADOS"]
    assert mem["T31"].value == (
        '=IF(ISNUMBER(CONTROLE!$B$3),CONTROLE!$B$3,DATE(9999,12,31))'
    )


def test_ajustes_manuais_c43_g50_intactos(wb):
    ws = wb["RESULTADOS"]
    assert [ws[f"A{r}"].value for r in range(43, 51)] == [
        "Retroativo manual oficial", "Ajuste do VTA", "VTA manual substitutivo",
        "Complemento histórico", "Complemento histórico",
        "Complemento histórico", "Complemento histórico", "Complemento histórico",
    ]
    validacoes = {(dv.type, str(dv.sqref)) for dv in ws.data_validations.dataValidation}
    assert ("list", "G43:G50") in validacoes
    assert ("decimal", "C46:C50") in validacoes


# --------------------------------------------------------------- FASE 2 e 4
#
# T38 e um SUMIFS. O contrato que ele encerra e:
#
#     soma de itens_PC!J onde itens_PC!B <= T31 (data de corte)
#         ==
#     leitor -> totais_canonicos.ate_o_corte.delta_potencial
#
# As fixtures escrevem VALORES LITERAIS em itens_PC (inclusive na coluna J),
# porque um workbook montado por openpyxl nao carrega cache do Excel: sem
# literal, `J` seria formula sem valor e os dois lados leriam coisas
# diferentes por construcao, nao por divergencia semantica.

COL = {"A": 1, "B": 2, "C": 3, "D": 4, "F": 6, "G": 7, "J": 10, "L": 12}


def _monta(pcs, data_corte=dt.date(2024, 12, 31)):
    """Workbook derivado do template com PCs escritos como literais."""
    wb = load_workbook(TEMPLATE)
    wb["CONTROLE"]["B1"] = "PCs"
    wb["CONTROLE"]["B2"] = "C2"
    wb["CONTROLE"]["B3"] = data_corte
    ws = wb["itens_PC"]
    for i, pc in enumerate(pcs):
        r = 2 + i
        ws.cell(r, COL["A"]).value = pc["numero"]
        ws.cell(r, COL["B"]).value = pc["data"]
        ws.cell(r, COL["C"]).value = pc.get("ciclo", "C1")
        ws.cell(r, COL["D"]).value = pc["valor"]
        ws.cell(r, COL["F"]).value = pc.get("atualizado", pc["valor"])
        ws.cell(r, COL["G"]).value = pc["pago"]
        ws.cell(r, COL["J"]).value = pc["delta"]
        ws.cell(r, COL["L"]).value = pc.get("efeito", "Sim")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _t38(raw, data_corte):
    """Avalia a semantica do SUMIFS de T38 sobre o workbook."""
    ws = load_workbook(io.BytesIO(raw), data_only=True)["itens_PC"]
    limite = data_corte or dt.date(9999, 12, 31)
    total = 0.0
    for r in range(2, 5002):
        j = ws.cell(r, COL["J"]).value
        if not isinstance(j, (int, float)) or isinstance(j, bool):
            continue
        b = ws.cell(r, COL["B"]).value
        if isinstance(b, dt.datetime):
            b = b.date()
        if isinstance(b, dt.date) and b <= limite:
            total += j
    return round(total, 2)


def _web(raw):
    res = ler_masterfile_v10(io.BytesIO(raw))
    totais = (res.get("itens_pc") or {}).get("totais_canonicos") or {}
    bloco = totais.get("ate_o_corte") or {}
    return round(float(bloco.get("delta_potencial") or 0.0), 2)


CORTE = dt.date(2024, 12, 31)

CENARIOS_POTENCIAL = {
    # nome                        PCs                                            esperado
    "pc_com_potencial": ([
        dict(numero="PC-1", data=dt.date(2024, 3, 1), valor=100000.0,
             atualizado=107436.8, pago="Nao", delta=7436.8),
    ], 7436.8),
    "pc_sem_potencial": ([
        dict(numero="PC-1", data=dt.date(2024, 3, 1), valor=100000.0,
             atualizado=100000.0, pago="Nao", delta=0.0),
    ], 0.0),
    "pc_validada_paga": ([
        # PC pago: itens_PC!J e 0 por definicao — reconhecido, nao potencial.
        dict(numero="PC-1", data=dt.date(2024, 3, 1), valor=100000.0,
             atualizado=107436.8, pago="Sim", delta=0.0),
    ], 0.0),
    "pc_em_analise": ([
        dict(numero="PC-1", data=dt.date(2024, 4, 1), valor=80000.0,
             atualizado=83500.0, pago="", delta=3500.0),
    ], 3500.0),
    "pc_sem_efeito_financeiro": ([
        dict(numero="PC-1", data=dt.date(2024, 5, 1), valor=60000.0,
             atualizado=61200.0, pago="Nao", delta=0.0, efeito="Nao"),
    ], 0.0),
    "multiplas_pcs": ([
        dict(numero="PC-1", data=dt.date(2024, 2, 1), valor=50000.0,
             atualizado=51000.0, pago="Nao", delta=1000.0),
        dict(numero="PC-2", data=dt.date(2024, 6, 1), valor=70000.0,
             atualizado=72500.0, pago="Nao", delta=2500.0),
        dict(numero="PC-3", data=dt.date(2024, 9, 1), valor=30000.0,
             atualizado=30000.0, pago="Sim", delta=0.0),
    ], 3500.0),
    "pc_posterior_ao_corte_nao_conta": ([
        dict(numero="PC-1", data=dt.date(2024, 3, 1), valor=50000.0,
             atualizado=51000.0, pago="Nao", delta=1000.0),
        # depois do corte: entra em posterior_ao_corte, nunca no potencial
        dict(numero="PC-2", data=dt.date(2025, 3, 1), valor=90000.0,
             atualizado=99000.0, pago="Nao", delta=9000.0),
    ], 1000.0),
}


@pytest.mark.parametrize("nome", sorted(CENARIOS_POTENCIAL))
def test_paridade_t38_com_a_grandeza_canonica_da_web(nome):
    pcs, esperado = CENARIOS_POTENCIAL[nome]
    raw = _monta(pcs, CORTE)
    xls, web = _t38(raw, CORTE), _web(raw)
    assert xls == esperado, f"{nome}: T38 deu {xls}, esperado {esperado}"
    assert xls == web, f"{nome}: XLS {xls} != web canonico {web}"


def test_pc_posterior_ao_corte_discrimina_t38_de_d45():
    """Prova de que a formula filtrada e necessaria: D45 erraria aqui."""
    pcs, esperado = CENARIOS_POTENCIAL["pc_posterior_ao_corte_nao_conta"]
    raw = _monta(pcs, CORTE)
    ws = load_workbook(io.BytesIO(raw), data_only=True)["itens_PC"]
    d45 = round(sum(
        ws.cell(r, COL["J"]).value
        for r in range(2, 5002)
        if isinstance(ws.cell(r, COL["J"]).value, (int, float))
    ), 2)
    assert _t38(raw, CORTE) == esperado == _web(raw)
    assert d45 == 10000.0
    assert d45 != esperado, "cenario nao discrimina; revise a fixture"


def test_potencial_nao_contamina_retro_oficial_nem_vta(monkeypatch):
    """Mesmos dados com e sem potencial produzem o MESMO retro/VTA oficiais."""
    base = [dict(numero="PC-1", data=dt.date(2024, 3, 1), valor=100000.0,
                 atualizado=100000.0, pago="Sim", delta=0.0)]
    com_potencial = base + [
        dict(numero="PC-2", data=dt.date(2024, 7, 1), valor=80000.0,
             atualizado=88000.0, pago="Nao", delta=8000.0),
    ]
    r_sem = ler_masterfile_v10(io.BytesIO(_monta(base, CORTE)))
    r_com = ler_masterfile_v10(io.BytesIO(_monta(com_potencial, CORTE)))

    def oficiais(res):
        rx = res.get("resultados_xls") or {}
        valores = rx.get("valores") or {}
        return valores.get("RETRO_OFICIAL"), valores.get("VTA_FINAL")

    assert oficiais(r_sem) == oficiais(r_com), (
        "o potencial alterou RETRO_OFICIAL/VTA_FINAL"
    )
    pot_sem = ((r_sem.get("itens_pc") or {}).get("totais_canonicos") or {}
               ).get("ate_o_corte", {}).get("delta_potencial")
    pot_com = ((r_com.get("itens_pc") or {}).get("totais_canonicos") or {}
               ).get("ate_o_corte", {}).get("delta_potencial")
    assert round(float(pot_sem or 0), 2) == 0.0
    assert round(float(pot_com or 0), 2) == 8000.0


# ------------------------------------------------- UX2.1: status oficial
#
# O status do painel vem de RESULTADOS!B3, lido do cache do XLS:
#   RESULTADOS!B3 -> _coleta_reajuste.ler_coleta_reajuste
#   -> metadados["status_resultados"]["geral"]
#   -> _resultado_consolidado._status_oficial_resultados -> web
# Nunca e fabricado a partir do VTA ou do retroativo.

import os  # noqa: E402

from _coleta_reajuste import ler_coleta_reajuste  # noqa: E402
from _resultado_consolidado import (  # noqa: E402
    _status_oficial_resultados,
)

GOLDEN_FINANCEIRO = Path(
    os.environ.get("CL8US_GOLDENS_DIR", r"C:\Users\danie\Downloads\anthropic-skills")
) / "Coleta_Reajuste_C3_ICTI_25agosto2026.xlsx"


def _status_do_arquivo(caminho_ou_bytes):
    raw = (
        caminho_ou_bytes
        if isinstance(caminho_ou_bytes, bytes)
        else Path(caminho_ou_bytes).read_bytes()
    )
    diag = ler_coleta_reajuste(raw)
    bloco = (diag.get("metadados") or {}).get("status_resultados") or {}
    return bloco, _status_oficial_resultados(bloco)


@pytest.mark.skipif(
    not GOLDEN_FINANCEIRO.exists(), reason="golden Financeiro indisponivel"
)
def test_golden_financeiro_continua_validado():
    """O caso homologado de 8,7 milhoes nao pode virar PENDENTE."""
    bloco, oficial = _status_do_arquivo(GOLDEN_FINANCEIRO)
    assert bloco.get("geral") == "VALIDADO"
    assert oficial["codigo"] == "VALIDADO"
    assert oficial["disponivel"] is True
    assert oficial["conclusivo"] is True
    assert oficial["origem"] == "resultados_xls"


def test_status_realmente_ausente_continua_fail_closed():
    """Sem cache do Excel o status nao pode ser inventado.

    O template versionado JA vem recalculado (B3 em cache, "REVISE"), entao
    ele nao serve de caso negativo. Reabrir e regravar por openpyxl descarta
    o cache das formulas e reproduz exatamente o arquivo que volta do fiscal
    sem ter passado pelo Excel — o caso em que a web deve seguir pedindo
    confirmacao.
    """
    buffer = io.BytesIO()
    load_workbook(TEMPLATE).save(buffer)      # regrava sem cache de formula
    bloco, oficial = _status_do_arquivo(buffer.getvalue())

    assert not bloco.get("geral"), "sem cache o status tem de ficar ausente"
    assert oficial["disponivel"] is False
    assert oficial["codigo"] != "VALIDADO"


def test_status_nunca_e_fabricado_a_partir_de_vta_ou_retroativo():
    """Status ausente segue ausente, haja ou nao numeros no bloco."""
    buffer = io.BytesIO()
    load_workbook(TEMPLATE).save(buffer)
    bloco, oficial = _status_do_arquivo(buffer.getvalue())
    assert oficial["disponivel"] is False
    assert oficial["codigo"] != "VALIDADO", bloco.get("valores")
