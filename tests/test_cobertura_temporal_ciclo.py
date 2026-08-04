# -*- coding: utf-8 -*-
"""cobertura_temporal!B8 — cobertura fisica ATUAL confirmada (origem CICLO).

B8 = data de CICLO_EM_EXECUCAO!D5 SOMENTE quando a posicao itemizada esta
completa e valida (sinal: CICLO_EM_EXECUCAO!A9 nao vazio — o total so aparece
com data + itens + zero erros/incompletos). Caso contrario, vazio. Via
INDIRECT+ISERROR (compat. com arquivos antigos sem a aba).

Os cenarios de validade (vazia / data+itens incompletos / posicao invalida /
completa e valida) sao caracterizados pelo sinal A9: A9 vazio para qualquer
estado nao-confirmado, A9 numerico apenas para completa e valida. A propria
formula de A9 (que discrimina esses estados) e homologada e testada em
test_ciclo_em_execucao_itemizado_29c1 / _opcionalidade_29c2.
"""
from __future__ import annotations

import datetime as _dt
import shutil
import tempfile
from pathlib import Path

import pytest
from openpyxl import load_workbook

from _ciclo_em_execucao import ABA_CICLO_EM_EXECUCAO, garantir_aba_ciclo_em_execucao

ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "templates" / "COLETA_REAJUSTE_OFICIAL.xlsx"

B8_ESPERADA = (
    '=IF(ISERROR(INDIRECT("CICLO_EM_EXECUCAO!A9")),"",'
    'IF(INDIRECT("CICLO_EM_EXECUCAO!A9")="","",'
    'INDIRECT("CICLO_EM_EXECUCAO!D5")))'
)


# --------------------------------------------------------------------------- #
# Estrutural (openpyxl)                                                        #
# --------------------------------------------------------------------------- #
def test_rotulo_ajuda_e_formula_automatica():
    ws = load_workbook(TEMPLATE)["cobertura_temporal"]
    assert str(ws["A8"].value).upper() == "COBERTURA FISICA ATUAL CONFIRMADA ATE"
    assert ws["B8"].value == B8_ESPERADA
    ajuda = str(ws["C8"].value)
    assert "CICLO_EM_EXECUCAO" in ajuda
    # Instrucoes antigas removidas.
    assert "QTD_REM_ATUAL" not in ajuda
    assert "posicao_referencia" not in ajuda
    assert "CONTROLE!B3" not in ajuda


def test_compat_arquivo_sem_aba_valor_vazio_data_only():
    # Template promovido (sem CICLO_EM_EXECUCAO): B8 recalculado = vazio.
    ws = load_workbook(TEMPLATE, data_only=True)["cobertura_temporal"]
    assert ws["B8"].value in (None, "")


# --------------------------------------------------------------------------- #
# Cenarios de valor (Excel COM)                                               #
# --------------------------------------------------------------------------- #
def _b8_com(gerar_aba: bool, a9=None, d5=None):
    """Recalcula cobertura!B8 no Excel real para um estado de CICLO."""
    client = pytest.importorskip("win32com.client")
    pythoncom = pytest.importorskip("pythoncom")
    tmp_dir = Path(tempfile.mkdtemp(prefix="cob_ciclo_"))
    tmp = tmp_dir / "coleta.xlsx"
    if gerar_aba:
        wb = load_workbook(TEMPLATE)
        garantir_aba_ciclo_em_execucao(wb)
        wb.save(tmp)
    else:
        shutil.copyfile(TEMPLATE, tmp)

    pythoncom.CoInitialize()
    excel = client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        wbx = excel.Workbooks.Open(str(tmp), UpdateLinks=0, ReadOnly=False, CorruptLoad=0)
        if gerar_aba and (a9 is not None or d5 is not None):
            ciclo = wbx.Worksheets(ABA_CICLO_EM_EXECUCAO)
            try:
                ciclo.Unprotect()
            except Exception:
                pass
            if d5 is not None:
                ciclo.Range("D5").Value = d5
            if a9 is not None:
                ciclo.Range("A9").Value = a9
        excel.CalculateFullRebuild()
        b8 = wbx.Worksheets("cobertura_temporal").Range("B8").Value
        wbx.Close(SaveChanges=False)
        return b8
    finally:
        excel.Quit()
        del excel
        pythoncom.CoUninitialize()
        shutil.rmtree(tmp_dir, ignore_errors=True)


def test_cenario_aba_ausente_vazio():
    assert _b8_com(gerar_aba=False) in (None, "")


def test_cenario_aba_vazia_vazio():
    # Aba gerada, sem data e sem itens (template vazio) => A9 "" => B8 "".
    assert _b8_com(gerar_aba=True) in (None, "")


def test_cenario_data_preenchida_itens_incompletos_ou_invalidos_vazio():
    # D5 preenchida, porem posicao nao confirmada (A9 vazio: qualquer estado
    # incompleto/invalido) => B8 permanece vazio (nao adota data nao confirmada).
    assert _b8_com(gerar_aba=True, a9="", d5=_dt.datetime(2027, 6, 15)) in (None, "")


def test_cenario_posicao_completa_e_valida_mostra_data():
    # Posicao confirmada (A9 numerico) + data => B8 = data de CICLO!D5.
    b8 = _b8_com(gerar_aba=True, a9=1000.0, d5=_dt.datetime(2027, 6, 15))
    assert getattr(b8, "day", None) == 15 and getattr(b8, "month", None) == 6
    assert getattr(b8, "year", None) == 2027
