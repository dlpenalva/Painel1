"""Teste NORMATIVO ATUAL da posição contratual — COLETA_REAJUSTE_OFICIAL.xlsx (MODELO B).

Esta é a referência normativa vigente do Claus para NOVAS coletas. Usa
exclusivamente `templates/COLETA_REAJUSTE_OFICIAL.xlsx` (via
`_coleta_oficial.TEMPLATE_COLETA_OFICIAL`). A cobertura do layout antigo
(Modelo A) vive, isolada, em `tests/test_posicao_contratual_legado.py` e NÃO
define a regra atual.

Semântica normativa (MODELO B)
------------------------------
* ``QTD_REM_BASE_Cn``   — remanescente **real** informado pelo fiscal.
* ``QTD_CONTRATADA_Cn`` — quantidade contratual vigente = base + Σ DELTAS
  assinados dos aditivos (cumulativo).
* ``QTD_REM_AJUSTADA_Cn`` — remanescente informado, normalizado/arredondado e
  validado (passthrough: C0 = contratada de nascimento; C1..C4 = ROUND do
  informado). Não reaplica o aditivo (já está em ``QTD_CONTRATADA_Cn``).
* Validação ``QTD_REM_AJUSTADA_Cn <= QTD_CONTRATADA_Cn``.

Reconstrução de execução (DELTA assinado)
-----------------------------------------
* ``QTD_EXECUTADA_C0 = MAX(QTD_CONTRATADA_C0 - QTD_REM_BASE_C1, 0)``
* ``QTD_EXECUTADA_Cn = MAX(QTD_REM_BASE_Cn + DELTA_Cn - QTD_REM_BASE_C(n+1), 0)``
"""

import unittest

from openpyxl import load_workbook

from _coleta_oficial import TEMPLATE_COLETA_OFICIAL


# --------------------------------------------------------------------------- #
# Especificação executável (Python puro) da semântica MODELO B.
# Reimplementa as fórmulas do template oficial; documenta e protege a regra.
# --------------------------------------------------------------------------- #
def delta_por_ciclo(eventos):
    d = [0.0] * 5
    for ci, q in eventos:
        d[ci] += q
    return [round(x, 2) for x in d]


def contratada_vigente(base, d):
    c = [round(base + d[0], 2)]
    for n in range(1, 5):
        c.append(round(c[-1] + d[n], 2))
    return c


def execucao_reconstruida(base, d, rem_base):
    """rem_base: dict {1..4} = remanescente real informado no início de cada
    ciclo (None se o item não existia). Retorna [exec_C0..exec_C3]."""
    c = contratada_vigente(base, d)
    ex = [None] * 4
    if rem_base.get(1) is not None:
        ex[0] = round(max(c[0] - rem_base[1], 0.0), 2)
    for n in range(1, 4):
        if rem_base.get(n) is not None and rem_base.get(n + 1) is not None:
            ex[n] = round(max(rem_base[n] + d[n] - rem_base[n + 1], 0.0), 2)
    return ex


def _rem_a_partir_do_consumo(base, d, consumo):
    """Constrói as fotografias MODELO B a partir de um consumo planejado por
    ciclo, de forma auto-consistente (fiscal informa o remanescente real)."""
    c = contratada_vigente(base, d)
    rem = {1: round(c[0] - consumo[0], 2)}
    for n in range(1, 4):
        rem[n + 1] = round(rem[n] + d[n] - consumo[n], 2)
    return rem


def status_posicao(contratada, rem_ajustada):
    """Replica CHECK_POSICAO_CONTRATUAL para um item (sem duplicidade)."""
    if any(r is not None and r > c for c, r in zip(contratada, rem_ajustada)):
        return "ALERTA: REMANESCENTE_SUPERA_POSICAO"
    if min([v for v in list(contratada) + [r for r in rem_ajustada if r is not None]]) < 0:
        return "ALERTA: POSICAO_NEGATIVA"
    return "OK"


class ReconstrucaoExecucaoOficial(unittest.TestCase):
    """Casos 1–6, 10: aditivo (acréscimo/supressão/novo item) nunca vira consumo."""

    def _checa_recuperacao(self, base, eventos, consumo):
        d = delta_por_ciclo(eventos)
        rem = _rem_a_partir_do_consumo(base, d, consumo)
        ex = execucao_reconstruida(base, d, rem)
        self.assertEqual(ex, [round(x, 2) for x in consumo])

    def test_1_sem_aditivo(self):
        self._checa_recuperacao(100.0, [], [20.0, 30.0, 0.0, 20.0])

    def test_2_acrescimo(self):
        d = delta_por_ciclo([(2, 40.0)])
        self.assertEqual(contratada_vigente(100.0, d), [100.0, 100.0, 140.0, 140.0, 140.0])
        self._checa_recuperacao(100.0, [(2, 40.0)], [20.0, 30.0, 25.0, 10.0])

    def test_3_supressao(self):
        d = delta_por_ciclo([(2, -30.0)])
        self.assertEqual(contratada_vigente(100.0, d), [100.0, 100.0, 70.0, 70.0, 70.0])
        self._checa_recuperacao(100.0, [(2, -30.0)], [10.0, 20.0, 5.0, 10.0])

    def test_4_acrescimo_e_supressao_em_ciclos_distintos(self):
        d = delta_por_ciclo([(1, 40.0), (3, -20.0)])
        self.assertEqual(contratada_vigente(100.0, d), [100.0, 140.0, 140.0, 120.0, 120.0])
        self._checa_recuperacao(100.0, [(1, 40.0), (3, -20.0)], [10.0, 15.0, 20.0, 5.0])

    def test_5_dois_aditivos_mesmo_ciclo(self):
        d = delta_por_ciclo([(2, 40.0), (2, 10.0), (2, -5.0)])
        self.assertEqual(d[2], 45.0)
        self.assertEqual(contratada_vigente(100.0, d), [100.0, 100.0, 145.0, 145.0, 145.0])
        self._checa_recuperacao(100.0, [(2, 40.0), (2, 10.0), (2, -5.0)], [0.0, 0.0, 30.0, 15.0])

    def test_6_novo_item_base_zero(self):
        d = delta_por_ciclo([(2, 5.5)])
        self.assertEqual(contratada_vigente(0.0, d), [0.0, 0.0, 5.5, 5.5, 5.5])
        rem = {1: None, 2: 0.0, 3: 3.5, 4: 3.5}  # nasce em C2 (+5.5), consome 2.0
        ex = execucao_reconstruida(0.0, d, rem)
        self.assertIsNone(ex[0])
        self.assertIsNone(ex[1])
        self.assertEqual(ex[2], 2.0)

    def test_10_execucao_usa_delta_assinado(self):
        # acréscimo e supressão no mesmo item, execução recuperada exatamente.
        self._checa_recuperacao(100.0, [(1, 20.0), (2, -10.0)], [5.0, 10.0, 8.0, 12.0])


class ValidacaoPosicaoOficial(unittest.TestCase):
    """Casos 7–9: remanescente menor/igual/maior que a contratada."""

    def test_7_remanescente_menor_que_contratada_ok(self):
        self.assertEqual(status_posicao([100, 100, 100, 100, 100], [100, 100, 90, None, None]), "OK")

    def test_8_remanescente_igual_a_contratada_ok(self):
        self.assertEqual(status_posicao([100, 100, 100, 100, 100], [100, 100, 100, None, None]), "OK")

    def test_9_remanescente_maior_que_contratada_alerta(self):
        self.assertEqual(
            status_posicao([100, 100, 100, 100, 100], [100, 100, 110, None, None]),
            "ALERTA: REMANESCENTE_SUPERA_POSICAO",
        )


class FormulasOficialModeloB(unittest.TestCase):
    """Trava do texto das fórmulas do template oficial (MODELO B)."""

    @classmethod
    def setUpClass(cls):
        cls.wb = load_workbook(TEMPLATE_COLETA_OFICIAL, data_only=False)
        cls.pos = cls.wb["posicao_contratual"]
        cls.rem = cls.wb["itens_Remanesc"]
        cls.rc = cls.wb["itens_RC"]
        cls.hv = cls.wb["historico_VU"]
        cls.ad = cls.wb["aditivos"]

    def test_cabecalho_e_remanescente_real_informado(self):
        # MODELO B: o cabeçalho pede o remanescente real (não "sem aditivo").
        self.assertIn("QTD_REM_BASE_C1", self.rem["E1"].value)
        self.assertNotIn("SEM_ADITIVO", self.rem["E1"].value)

    def test_delta_assinado(self):
        f = self.ad["L2"].value
        self.assertIn("ROUND(ABS(E2),2)", f)
        self.assertIn("ROUND(-ABS(E2),2)", f)

    def test_contratada_cumulativa(self):
        self.assertIn("SUMIFS(aditivos!$L$2:$L$200", self.pos["D2"].value)
        self.assertIn("ROUND(C2+D2,2)", self.pos["E2"].value)
        self.assertIn("ROUND(E2+H2,2)", self.pos["I2"].value)
        self.assertIn("ROUND(I2+L2,2)", self.pos["M2"].value)
        self.assertIn("ROUND(M2+P2,2)", self.pos["Q2"].value)
        self.assertIn("ROUND(Q2+T2,2)", self.pos["U2"].value)

    def test_rem_ajustada_passthrough_sem_reaplicar_aditivo(self):
        self.assertEqual(self.pos["G2"].value, '=IF(A2="","",E2)')  # C0 = contratada
        for cel, ref in (("K2", "J2"), ("O2", "N2"), ("S2", "R2"), ("W2", "V2")):
            f = self.pos[cel].value
            self.assertIn(f"ROUND({ref},2)", f)
            self.assertNotIn("aditivos", f)

    def test_check_valida_remanescente_le_contratada(self):
        f = self.pos["X2"].value
        self.assertIn("ITEM_DUPLICADO", f)
        self.assertIn("POSICAO_NEGATIVA", f)
        self.assertIn("REMANESCENTE_SUPERA_POSICAO", f)
        self.assertIn("G2>E2", f)
        self.assertIn("W2>U2", f)

    def test_execucao_reconstruida_com_delta_assinado(self):
        self.assertIn("MAX(posicao_contratual!E2-posicao_contratual!J2,0)", self.rem["AB2"].value)
        self.assertIn(
            "MAX(posicao_contratual!J2+posicao_contratual!H2-posicao_contratual!N2,0)", self.rem["M2"].value
        )
        self.assertIn(
            "MAX(posicao_contratual!N2+posicao_contratual!L2-posicao_contratual!R2,0)", self.rem["O2"].value
        )
        self.assertIn(
            "MAX(posicao_contratual!R2+posicao_contratual!P2-posicao_contratual!V2,0)", self.rem["Q2"].value
        )


class PropagacaoOficial(unittest.TestCase):
    """Casos 11–12: itens_RC e historico_VU consomem posicao_contratual."""

    @classmethod
    def setUpClass(cls):
        cls.wb = load_workbook(TEMPLATE_COLETA_OFICIAL, data_only=False)
        cls.rc = cls.wb["itens_RC"]
        cls.hv = cls.wb["historico_VU"]

    def test_11_itens_rc_recebe_rem_ajustada(self):
        self.assertEqual(self.rc["C3"].value, '=IF(posicao_contratual!A2="","",posicao_contratual!G2)')
        self.assertEqual(self.rc["F3"].value, '=IF(posicao_contratual!A2="","",posicao_contratual!K2)')
        self.assertEqual(self.rc["I3"].value, '=IF(posicao_contratual!A2="","",posicao_contratual!O2)')
        self.assertEqual(self.rc["L3"].value, '=IF(posicao_contratual!A2="","",posicao_contratual!S2)')
        self.assertEqual(self.rc["O3"].value, '=IF(posicao_contratual!A2="","",posicao_contratual!W2)')

    def test_12_historico_vu_espelha_vigente_e_rem_ajustada(self):
        for cel, ref in (("N2", "E2"), ("O2", "I2"), ("P2", "M2"), ("Q2", "Q2"), ("R2", "U2")):
            self.assertEqual(self.hv[cel].value, f'=IF(A2="","",posicao_contratual!{ref})')
        for cel, ref in (("S2", "G2"), ("T2", "K2"), ("U2", "O2"), ("V2", "S2"), ("W2", "W2")):
            self.assertEqual(self.hv[cel].value, f'=IF(A2="","",posicao_contratual!{ref})')


if __name__ == "__main__":
    unittest.main()
