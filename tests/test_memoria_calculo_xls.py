"""Etapa 4 — memoria de calculo persistida em parametros!J2:R80.

Cobre normalizacao canonica (ICTI/SGS/IST), gravacao pelo gerador oficial,
fidelidade celula a celula com a memoria exibida, compatibilidade legada,
capacidade/overflow e leitura opcional no leitor v10.
"""
from __future__ import annotations

import io
from datetime import date, datetime

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from _coleta_oficial import _RESIDUOS_POR_ABA, gerar_coleta_oficial_preenchida
from _leitor_masterfile_v10 import ler_masterfile_v10
from _memoria_calculo import (
    CABECALHOS_MEMORIA_CALCULO,
    CAPACIDADE_MEMORIA_CALCULO,
    escrever_memoria_calculo,
    ler_memoria_calculo,
    normalizar_memoria_calculo,
)


def _res_icti(meses: int = 12, inicio: str = "2023-11-01") -> dict:
    datas = pd.date_range(inicio, periods=meses, freq="MS")
    taxas = [0.45 + 0.01 * i for i in range(meses)]
    fatores = [1 + t / 100 for t in taxas]
    acum = pd.Series(fatores).cumprod()
    dados = pd.DataFrame({
        "data": datas,
        "valor": taxas,
        "fator_mensal": fatores,
        "fator_acumulado_progressivo": acum,
    })
    variacao = float(acum.iloc[-1]) - 1
    return {
        "variacao": variacao,
        "metodo": "ICTI/Ipeadata: produtorio das taxas mensais",
        "dados": dados,
        "sercodigo": "DIMAC_ICTI2",
    }


def _res_sgs(meses: int = 12, inicio: str = "2023-11-01") -> dict:
    datas = pd.date_range(inicio, periods=meses, freq="MS")
    dados = pd.DataFrame({"data": datas, "valor": ["0.45"] * meses})
    return {
        "variacao": (1 + 0.0045) ** meses - 1,
        "metodo": "Produtorio de taxas mensais (SGS/BCB)",
        "dados": dados,
    }


def _res_ist() -> dict:
    return {
        "variacao": (172.4 / 165.1) - 1,
        "i_ini": 165.1,
        "i_fim": 172.4,
        "metodo": "Divisao de Numero-Indice (Serie Local)",
        "dados": pd.DataFrame({
            "data": [pd.Timestamp(2023, 10, 1), pd.Timestamp(2024, 10, 1)],
            "indice": [165.1, 172.4],
        }),
    }


def _payload(ciclos: list[dict], indice: str = "ICTI") -> dict:
    return {
        "origem": "Reajuste Simples",
        "indice": indice,
        "data_base_original": "01/10/2022",
        "ciclos": ciclos,
    }


def _ciclo(nome: str, memoria, percentual: float = 0.056, **extras) -> dict:
    base = {
        "ciclo": nome,
        "data_base": f"01/10/{2022 + int(nome[1])}",
        "data_pedido": f"01/10/{2023 + int(nome[1])}",
        "percentual_aplicado": percentual,
        "financeiro_inicio": f"01/10/{2023 + int(nome[1])}",
        "memoria_calculo": memoria,
    }
    base.update(extras)
    return base


def _bloco(payload_bytes: bytes) -> list[tuple]:
    ws = load_workbook(io.BytesIO(payload_bytes))["parametros"]
    linhas = []
    for r in range(2, 81):
        valores = [ws.cell(r, c).value for c in range(10, 19)]
        if any(v is not None for v in valores):
            linhas.append((r, *valores))
    return linhas


# ---------------------------------------------------------------- normalizador

def test_normalizador_icti_12_meses_mais_resultado():
    res = _res_icti()
    memoria = normalizar_memoria_calculo(res, 1 + res["variacao"], res["variacao"])
    assert len(memoria) == 13
    assert [m["tipo"] for m in memoria] == ["MES"] * 12 + ["RESULTADO"]
    assert [m["ordem"] for m in memoria] == list(range(1, 14))
    primeiro = memoria[0]
    assert primeiro["competencia"] == "2023-11-01"
    assert primeiro["valor_indice"] == pytest.approx(0.0045)  # 0,45% -> 0.0045
    assert primeiro["fator_mensal"] == pytest.approx(1.0045)
    assert primeiro["fator_acumulado"] == pytest.approx(1.0045)
    resultado = memoria[-1]
    assert resultado["competencia"] is None
    assert resultado["valor_indice"] is None
    assert resultado["fator_acumulado"] == pytest.approx(1 + res["variacao"])
    assert resultado["variacao_final"] == pytest.approx(res["variacao"])
    assert "ICTI" in resultado["metodo_fonte"] and "DIMAC_ICTI2" in resultado["metodo_fonte"]


def test_normalizador_sgs_sem_fatores_mensais_inventados():
    res = _res_sgs()
    memoria = normalizar_memoria_calculo(res, 1 + res["variacao"], res["variacao"])
    assert len(memoria) == 13
    for registro in memoria[:-1]:
        assert registro["tipo"] == "MES"
        assert registro["valor_indice"] == pytest.approx(0.0045)
        assert registro["fator_mensal"] is None
        assert registro["fator_acumulado"] is None
    assert memoria[-1]["fator_acumulado"] == pytest.approx(1 + res["variacao"])
    assert memoria[-1]["variacao_final"] == pytest.approx(res["variacao"])


def test_normalizador_ist_duas_linhas_indice_mais_resultado():
    res = _res_ist()
    memoria = normalizar_memoria_calculo(res, 1 + res["variacao"], res["variacao"])
    assert [m["tipo"] for m in memoria] == ["INDICE", "INDICE", "RESULTADO"]
    assert [m["ordem"] for m in memoria] == [1, 2, 3]
    assert memoria[0]["competencia"] == "2023-10-01"
    assert memoria[0]["valor_indice"] == pytest.approx(165.1)
    assert memoria[1]["valor_indice"] == pytest.approx(172.4)
    assert memoria[0]["fator_mensal"] is None and memoria[0]["fator_acumulado"] is None
    assert memoria[2]["variacao_final"] == pytest.approx(res["variacao"])


def test_normalizador_preserva_precisao_sem_arredondamento():
    res = _res_icti()
    res["dados"].loc[0, "valor"] = 0.456789123
    memoria = normalizar_memoria_calculo(res, 1.05, 0.05)
    assert memoria[0]["valor_indice"] == 0.456789123 / 100


def test_normalizador_sem_memoria_retorna_none():
    assert normalizar_memoria_calculo(None, 1.0, 0.0) is None
    assert normalizar_memoria_calculo({"metodo": "x"}, 1.0, 0.0) is None
    assert normalizar_memoria_calculo(
        {"dados": pd.DataFrame({"data": [], "valor": []})}, 1.0, 0.0
    ) is None


# ------------------------------------------------------------------- gravacao

def test_geracao_simples_icti_grava_bloco_fiel_celula_a_celula():
    res = _res_icti()
    memoria = normalizar_memoria_calculo(res, 1 + res["variacao"], res["variacao"])
    payload = gerar_coleta_oficial_preenchida(_payload([_ciclo("C1", memoria)]))
    linhas = _bloco(payload)
    assert len(linhas) == 13
    for i, registro in enumerate(memoria):
        r, ciclo, tipo, ordem, comp, valor, fator, acum, variacao, fonte = linhas[i]
        assert r == 2 + i and ciclo == "C1"
        assert tipo == registro["tipo"] and ordem == registro["ordem"]
        if registro["competencia"]:
            assert isinstance(comp, datetime)
            assert comp.date().isoformat() == registro["competencia"]
        else:
            assert comp is None
        for gravado, canonico in (
            (valor, registro["valor_indice"]),
            (fator, registro["fator_mensal"]),
            (acum, registro["fator_acumulado"]),
            (variacao, registro["variacao_final"]),
        ):
            if canonico is None:
                assert gravado is None
            else:
                assert gravado == pytest.approx(canonico)
        assert fonte == registro["metodo_fonte"]


def test_geracao_simples_grava_datas_reais_com_formato_mm_aaaa():
    res = _res_icti()
    memoria = normalizar_memoria_calculo(res, 1 + res["variacao"], res["variacao"])
    payload = gerar_coleta_oficial_preenchida(_payload([_ciclo("C1", memoria)]))
    ws = load_workbook(io.BytesIO(payload))["parametros"]
    assert ws["M2"].value == datetime(2023, 11, 1)
    assert ws["M2"].number_format == "mm/yyyy"
    assert ws["N2"].number_format == "0.0000%"
    assert ws["Q14"].number_format == "0.0000%"


def test_geracao_simples_sgs_e_ist():
    res_sgs = _res_sgs()
    memoria_sgs = normalizar_memoria_calculo(res_sgs, 1 + res_sgs["variacao"], res_sgs["variacao"])
    linhas = _bloco(gerar_coleta_oficial_preenchida(
        _payload([_ciclo("C1", memoria_sgs)], indice="IPCA")
    ))
    assert len(linhas) == 13
    assert {l[2] for l in linhas[:-1]} == {"MES"}
    assert all(l[6] is None and l[7] is None for l in linhas[:-1])  # sem O/P inventados

    res_ist = _res_ist()
    memoria_ist = normalizar_memoria_calculo(res_ist, 1 + res_ist["variacao"], res_ist["variacao"])
    linhas = _bloco(gerar_coleta_oficial_preenchida(
        _payload([_ciclo("C1", memoria_ist)], indice="IST")
    ))
    assert len(linhas) == 3  # 2 INDICE + RESULTADO, sem 12 meses ficticios
    assert [l[2] for l in linhas] == ["INDICE", "INDICE", "RESULTADO"]


def test_geracao_multiciclo_fontes_diferentes_sequencial():
    res1, res2 = _res_icti(), _res_ist()
    m1 = normalizar_memoria_calculo(res1, 1 + res1["variacao"], res1["variacao"])
    m2 = normalizar_memoria_calculo(res2, 1 + res2["variacao"], res2["variacao"])
    payload = gerar_coleta_oficial_preenchida(_payload([
        _ciclo("C1", m1), _ciclo("C2", m2),
    ]))
    linhas = _bloco(payload)
    assert len(linhas) == 16
    assert [l[1] for l in linhas] == ["C1"] * 13 + ["C2"] * 3
    assert [l[0] for l in linhas] == list(range(2, 18))  # sequencial, sem buracos


def test_geracao_2_a_5_ciclos_e_resultado_por_ciclo():
    ciclos = []
    for n in range(1, 5):
        res = _res_icti()
        memoria = normalizar_memoria_calculo(res, 1 + res["variacao"], res["variacao"])
        ciclos.append(_ciclo(f"C{n}", memoria))
    linhas = _bloco(gerar_coleta_oficial_preenchida(_payload(ciclos)))
    assert len(linhas) == 4 * 13
    resultados = [l for l in linhas if l[2] == "RESULTADO"]
    assert [l[1] for l in resultados] == ["C1", "C2", "C3", "C4"]
    assert all(l[3] == 13 for l in resultados)  # ORDEM = registros anteriores + 1


def test_ciclo_sem_memoria_e_omitido_sem_bloqueio():
    res = _res_icti()
    memoria = normalizar_memoria_calculo(res, 1 + res["variacao"], res["variacao"])
    payload = gerar_coleta_oficial_preenchida(_payload([
        _ciclo("C1", None),
        _ciclo("C2", memoria),
    ]))
    linhas = _bloco(payload)
    assert {l[1] for l in linhas} == {"C2"}
    assert len(linhas) == 13


def test_payload_sem_memoria_gera_bloco_vazio():
    payload = gerar_coleta_oficial_preenchida(_payload([_ciclo("C1", None)]))
    assert _bloco(payload) == []
    ws = load_workbook(io.BytesIO(payload))["parametros"]
    assert [ws.cell(1, c).value for c in range(10, 19)] == list(CABECALHOS_MEMORIA_CALCULO)


def test_regeracao_nao_duplica_linhas():
    res = _res_icti()
    memoria = normalizar_memoria_calculo(res, 1 + res["variacao"], res["variacao"])
    primeira = gerar_coleta_oficial_preenchida(_payload([_ciclo("C1", memoria)]))
    from _gerador_masterfile import gerar_masterfile_preenchido
    from _coleta_oficial import normalizar_dados_calculadora
    segunda = gerar_masterfile_preenchido(
        normalizar_dados_calculadora(_payload([_ciclo("C1", memoria)])), primeira
    )
    assert len(_bloco(segunda)) == 13


def test_regeracao_com_memoria_menor_limpa_residuos_do_bloco():
    res = _res_icti()
    memoria = normalizar_memoria_calculo(res, 1 + res["variacao"], res["variacao"])
    cheio = gerar_coleta_oficial_preenchida(_payload([
        _ciclo("C1", memoria), _ciclo("C2", memoria),
    ]))
    res_ist = _res_ist()
    memoria_ist = normalizar_memoria_calculo(res_ist, 1 + res_ist["variacao"], res_ist["variacao"])
    from _gerador_masterfile import gerar_masterfile_preenchido
    from _coleta_oficial import normalizar_dados_calculadora
    menor = gerar_masterfile_preenchido(
        normalizar_dados_calculadora(_payload([_ciclo("C1", memoria_ist)], indice="IST")),
        cheio,
    )
    linhas = _bloco(menor)
    assert len(linhas) == 3 and {l[1] for l in linhas} == {"C1"}


def test_overflow_gera_erro_controlado():
    res = _res_icti(meses=CAPACIDADE_MEMORIA_CALCULO)  # 79 MES + RESULTADO = 80
    memoria = normalizar_memoria_calculo(res, 1.05, 0.05)
    with pytest.raises(ValueError, match="excede a capacidade"):
        gerar_coleta_oficial_preenchida(_payload([_ciclo("C1", memoria)]))


def test_limite_exato_de_capacidade_79_linhas():
    res = _res_icti(meses=CAPACIDADE_MEMORIA_CALCULO - 1)  # 78 MES + RESULTADO = 79
    memoria = normalizar_memoria_calculo(res, 1.05, 0.05)
    linhas = _bloco(gerar_coleta_oficial_preenchida(_payload([_ciclo("C1", memoria)])))
    assert len(linhas) == CAPACIDADE_MEMORIA_CALCULO
    assert linhas[-1][0] == 80  # ultima linha do bloco reservado


def test_residuos_incluem_bloco_memoria():
    assert "J2" in _RESIDUOS_POR_ABA["parametros"]
    assert "R80" in _RESIDUOS_POR_ABA["parametros"]
    assert "J81" not in _RESIDUOS_POR_ABA["parametros"]
    assert "I2" not in _RESIDUOS_POR_ABA["parametros"]  # separacao visual intocada


# -------------------------------------------------------- compatibilidade/leitura

def test_template_legado_sem_cabecalho_nao_bloqueia_gravacao_nem_leitura():
    wb = Workbook()
    ws = wb.active
    ws.title = "parametros"
    escrever_memoria_calculo(ws, {"C1": {"memoria_calculo": [{"tipo": "MES", "ordem": 1}]}})
    assert ws["J2"].value is None  # legado ignorado sem erro
    assert ler_memoria_calculo(ws) == {}


def test_leitor_v10_expoe_memoria_estruturada():
    res = _res_icti()
    memoria = normalizar_memoria_calculo(res, 1 + res["variacao"], res["variacao"])
    payload = gerar_coleta_oficial_preenchida(_payload([_ciclo("C1", memoria)]))
    leitura = ler_masterfile_v10(payload)
    lida = leitura["memoria_calculo"]
    assert set(lida) == {"C1"}
    assert len(lida["C1"]) == 13
    assert lida["C1"][0]["tipo"] == "MES"
    assert lida["C1"][0]["competencia"] == "2023-11-01"
    assert lida["C1"][0]["valor_indice"] == pytest.approx(0.0045)
    assert lida["C1"][-1]["tipo"] == "RESULTADO"
    assert lida["C1"][-1]["variacao_final"] == pytest.approx(res["variacao"])


def test_leitor_v10_arquivo_sem_bloco_retorna_vazio_sem_erro():
    payload = gerar_coleta_oficial_preenchida(_payload([_ciclo("C1", None)]))
    leitura = ler_masterfile_v10(payload)
    assert leitura["memoria_calculo"] == {}
    assert "memoria" not in str(leitura.get("erro", "")).lower()
