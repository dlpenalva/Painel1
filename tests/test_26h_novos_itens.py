# -*- coding: utf-8 -*-
"""Etapa 26H — novos itens Nxxx (base zero automatica), temporalidade,
dropdown G, indice no Sumario, PREVIA documental e limpeza de UI.

Cobertura complementar a prova em Excel real (Excel COM na copia ALL TECH,
documentada no Gate): aqui ficam as ancoras rapidas e regressionaveis —
helpers canonicos, formulas/DVs do template (openpyxl), camada documental
sintetica e wiring de pagina/leitor.
"""
from __future__ import annotations

import io
import re
import sys
import zipfile
from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook

RAIZ = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(RAIZ))

from _reajuste_utils import eh_novo_item, qtd_base_efetiva  # noqa: E402

TEMPLATE = RAIZ / "templates" / "COLETA_REAJUSTE_OFICIAL.xlsx"


# ---------------------------------------------------------------------------
# 1. Padrao canonico Nxxx e base zero automatica (helpers Python)
# ---------------------------------------------------------------------------

# 26H.2: nomenclatura funcional aprovada = "N" + EXATAMENTE 3 algarismos.
@pytest.mark.parametrize("item", ["N001", "N002", "N003", "N010", "N099",
                                  "N100", "N999", "n001", " N001 "])
def test_padrao_canonico_reconhece_novo_item(item):
    assert eh_novo_item(item) is True


@pytest.mark.parametrize("item", ["N", "NOVO", "N1", "N01", "N0001", "N1000",
                                  "N1A", "N1.5", "N-1", "N 1", "N 001",
                                  "N1E3x", 12, "12", "", None, "AN001"])
def test_padrao_canonico_rejeita_item_normal(item):
    assert eh_novo_item(item) is False


def test_base_zero_automatica_somente_para_novo_item():
    # Novo item com base vazia -> 0 automatico (fiscal nao digita o zero).
    assert qtd_base_efetiva("N001", None) == 0.0
    assert qtd_base_efetiva("N001", "") == 0.0
    # Item normal com base vazia NUNCA vira zero.
    assert qtd_base_efetiva("12", None) is None
    assert qtd_base_efetiva("12", "") is None
    # Base informada prevalece (inconsistencia e papel dos CHECKs, nao daqui).
    assert qtd_base_efetiva("N001", 1) == 1.0
    assert qtd_base_efetiva("12", "5") == 5.0


# ---------------------------------------------------------------------------
# 2. Template oficial — formulas e validacoes 26H (openpyxl)
# ---------------------------------------------------------------------------

@pytest.fixture(scope="module")
def wb_template():
    wb = load_workbook(TEMPLATE, data_only=False)
    yield wb
    wb.close()


def test_colunas_tecnicas_posicao_contratual(wb_template):
    ws = wb_template["posicao_contratual"]
    assert ws["Y1"].value == "CICLO_NASCIMENTO"
    assert ws["Z1"].value == "EH_NOVO_ITEM"
    # Ocultas (nunca veryHidden). O Excel grava a dimensao como faixa
    # (min=25..max=26), entao a checagem cobre as colunas 25 (Y) e 26 (Z).
    def _oculta(indice_coluna: int) -> bool:
        return any(
            dim.hidden and dim.min <= indice_coluna <= dim.max
            for dim in ws.column_dimensions.values()
        )

    assert _oculta(25) and _oculta(26)
    z2 = str(ws["Z2"].value)
    # 26H.2: espelho Excel do padrao ^[Nn]\d{3}$ (N + exatamente 3 digitos).
    assert 'LEFT(TRIM($A2),1)="N"' in z2
    assert "LEN(TRIM($A2))=4" in z2
    assert "MID(TRIM($A2),2,3)" in z2
    y2 = str(ws["Y2"].value)
    for trecho in ("$E2>0", "$I2>0", "$M2>0", "$Q2>0", "$U2>0"):
        assert trecho in y2


def test_base_zero_automatica_na_posicao_contratual(wb_template):
    ws = wb_template["posicao_contratual"]
    # F (QTD_REM_BASE_C0): novo item com base vazia -> 0; normal -> "".
    assert "IF($Z2,0," in str(ws["F2"].value)
    # CHECK X: acusa Nxxx com base != 0 e VU ausente; nao corrige input.
    x2 = str(ws["X2"].value)
    assert "NOVO ITEM COM BASE DIFERENTE DE 0 - AJUSTAR itens_Remanesc" in x2
    assert "NOVO ITEM SEM VU_ORIGINAL - INFORMAR EM itens_Remanesc" in x2
    assert x2.count("(") == x2.count(")")


def test_historico_vu_temporal(wb_template):
    ws = wb_template["historico_VU"]
    # VU_C0 vazio para item nascido apos C0 (nunca VU inventado).
    assert "posicao_contratual!$Y2" in str(ws["C2"].value)
    for coluna, ciclo in (("D", 1), ("E", 2), ("F", 3), ("G", 4)):
        f = str(ws[f"{coluna}2"].value)
        assert "posicao_contratual!$Y2" in f
        assert f"IF(posicao_contratual!$Y2>{ciclo}," in f.replace(" ", "")
        assert "INDEX($L$2:$L$6" in f
        assert f.count("(") == f.count(")")


def test_posicao_referencia_temporal(wb_template):
    ws = wb_template["posicao_referencia"]
    i2 = str(ws["I2"].value)
    # POSICAO ATUAL COMPLETA?: exige QTD_REM_ATUAL apenas dos itens vigentes
    # na data de corte (nascimento <= ciclo da data); zero explicito valido.
    assert "SUMPRODUCT" in i2
    assert "posicao_contratual!$Y$2:$Y$200" in i2
    assert "ISNUMBER($B$2:$B$200)" in i2
    assert "COUNT($B$2:$B$200)>=" not in i2  # regra antiga (sem temporalidade)
    f2 = str(ws["F2"].value)
    assert "ITEM POSTERIOR A REFERENCIA" in f2
    assert f2.count("(") == f2.count(")")


def test_aditivos_mensagem_novo_item(wb_template):
    ws = wb_template["aditivos"]
    m2 = str(ws["M2"].value)
    assert ("NOVO ITEM NAO CADASTRADO - CADASTRAR EM itens_Remanesc; "
            "QTD_BASE_ORIGINAL sera 0 automaticamente; "
            "informar VU_ORIGINAL.") in m2
    # Validacoes preexistentes preservadas (ciclo, delta, tipo, duplicidade).
    for alerta in ("CICLO_INVALIDO", "QTD_INVALIDA", "TIPO_INVALIDO",
                   "ITEM_DUPLICADO"):
        assert alerta in m2


def test_mensagens_ao_fiscal_sao_literais_e_nao_ambiguas(wb_template):
    """26H.2: clareza operacional, nao apenas equivalencia da formula."""
    ir = wb_template["itens_Remanesc"]
    ad = wb_template["aditivos"]

    assert "use N001, N002..." in str(ir["A1"].value)
    assert "QTD_BASE_ORIGINAL fica 0 automaticamente" in str(ir["A1"].value)
    assert "informe o VU_ORIGINAL" in str(ir["A1"].value)
    assert "deixar vazio - assume 0 automaticamente" in str(ir["B1"].value)

    mensagem_cadastro = (
        "NOVO ITEM NAO CADASTRADO - CADASTRAR EM itens_Remanesc; "
        "QTD_BASE_ORIGINAL sera 0 automaticamente; informar VU_ORIGINAL."
    )
    assert mensagem_cadastro in str(ad["M2"].value)
    assert "cadastrar em itens_Remanesc" in str(ad["A1"].value)
    assert "QTD_BASE_ORIGINAL fica 0 automaticamente" in str(ad["A1"].value)
    assert "informar o VU_ORIGINAL" in str(ad["A1"].value)

    textos_fiscais = (
        str(ir["A1"].value),
        str(ir["B1"].value),
        str(ad["A1"].value),
        *(str(ad[f"M{linha}"].value) for linha in range(2, 201)),
    )
    assert all("BASE 0 + VU" not in texto.upper() for texto in textos_fiscais)


def test_memoria_sem_fail_closed_indevido_para_item_pos_c0(wb_template):
    ws = wb_template["MEMORIA_RESULTADOS"]
    x2 = str(ws["X2"].value)
    assert "posicao_contratual!$Y2>0),0," in x2.replace(" ", "")
    t27 = str(ws["T27"].value)
    assert "posicao_contratual!$Y$2:$Y$201" in t27


def test_dropdown_g_usa_fonte_robusta(wb_template):
    ws = wb_template["itens_PC"]
    dv_g = [dv for dv in ws.data_validations.dataValidation
            if "G2" in str(dv.sqref)]
    assert dv_g, "DV de itens_PC!G2:G5001 ausente"
    assert "OPCOES_SIM_NAO" in str(dv_g[0].formula1)
    assert "G2:G5001" in str(dv_g[0].sqref)


def test_dropdown_g_sobrevive_a_geracao_runtime():
    """O bug 26H: a lista literal aplicada via COM nao sobrevivia ao
    round-trip openpyxl da geracao da Coleta (formula1 vazia)."""
    from _coleta_oficial import obter_coleta_oficial_bytes
    wb = load_workbook(io.BytesIO(obter_coleta_oficial_bytes()), data_only=False)
    try:
        ws = wb["itens_PC"]
        dv_g = [dv for dv in ws.data_validations.dataValidation
                if "G2" in str(dv.sqref)]
        assert dv_g and "OPCOES_SIM_NAO" in str(dv_g[0].formula1)
    finally:
        wb.close()


def test_base_zero_visual_pre_semeada_em_itens_remanesc(wb_template):
    """26H.1: ao digitar N001 em A, o fiscal VE 0 em B na propria aba.

    Formula pre-semeada em B2:B200 (mesmo espelho Excel do padrao Nxxx da
    coluna Z); item normal/linha vazia -> "" (visual identico ao vazio).
    Prova funcional no Excel real: .codex_work/26h2/prova_26h2.py.
    """
    ws = wb_template["itens_Remanesc"]
    for coord in ("B2", "B100", "B200"):
        f = str(ws[coord].value)
        assert f.startswith("="), f"{coord} sem a formula 26H.1"
        assert 'LEFT(TRIM($A' in f and "MID(TRIM($A" in f
        assert f.endswith(',0,"")'), f"{coord} sem o ramo 0/vazio"
        assert f.count("(") == f.count(")")
    # Consumidores aritmeticos coagem ""->0 (formula-"" nao e celula vazia).
    pc = wb_template["posicao_contratual"]
    assert 'IF(itens_Remanesc!B2="",0,itens_Remanesc!B2)' in str(pc["C2"].value)
    pr = wb_template["posicao_referencia"]
    assert 'IF(itens_Remanesc!B2="",0,itens_Remanesc!B2)+SUMIFS(' in str(pr["C2"].value)


def test_invariante_fronteira_cadastro_200_201(wb_template):
    """26H.2 (auditoria): a capacidade funcional de cadastro termina na
    linha 200; a linha 201 pertence ao total dinamico e NAO pode oferecer
    automacao/visual de linha de input (off-by-one B201).

    Este teste FALHA se uma alteracao futura voltar a oferecer linha de
    input alem da capacidade funcional.
    """
    ir = wb_template["itens_Remanesc"]
    # Automacao presente em toda a faixa valida...
    for coord in ("B2", "B100", "B199", "B200"):
        assert str(ir[coord].value).startswith("="), f"{coord} sem automacao"
    # ...e AUSENTE na linha 201 (nao e linha de cadastro).
    assert ir["B201"].value in (None, ""), "B201 voltou a ter automacao de input"
    # A linha 201 e somente total quando A200 esta ocupada. Nenhuma formula
    # dessa linha pode interpretar A201 como novo cadastro.
    for coord in ("D201", "F201", "H201", "J201", "L201", "N201", "P201",
                  "R201", "U201"):
        assert "A201" not in str(ir[coord].value), f"{coord} ainda consome A201"
        assert "$A$200" in str(ir[coord].value), f"{coord} sem fronteira A200"

    # Alinhamento das faixas da cadeia na fronteira 200/201.
    pc = wb_template["posicao_contratual"]
    for col in ("C", "X", "Y", "Z"):
        assert str(pc[f"{col}200"].value).startswith("="), f"posicao_contratual!{col}200"
        assert pc[f"{col}201"].value in (None, ""), f"posicao_contratual!{col}201 fora da faixa"
    pr = wb_template["posicao_referencia"]
    assert str(pr["C200"].value).startswith("=")
    assert pr["C201"].value in (None, "")
    hv = wb_template["historico_VU"]
    assert str(hv["C200"].value).startswith("=")
    assert hv["C201"].value in (None, "")
    # Consumidores de fechamento ignoram integralmente um valor forcado em
    # A201; isso evita cadastro parcial e o antigo #VALOR! em itens_RC!D202.
    rc = wb_template["itens_RC"]
    assert "A201" not in str(rc["A202"].value)
    assert 'itens_Remanesc!$A$200<>""' in str(rc["A202"].value)
    for coord in ("B202", "C202", "E202", "F202", "H202", "I202", "K202",
                  "L202", "N202", "O202"):
        assert rc[coord].value == '=""', f"{coord} ainda oferece linha parcial"
    memoria = wb_template["MEMORIA_RESULTADOS"]
    for coord in ("J253", "K253", "L253", "M253", "N253"):
        assert memoria[coord].value == '=""', f"{coord} ainda consome A201"
    ad = wb_template["aditivos"]
    assert "COUNTIF(itens_Remanesc!$A$2:$A$200,A2)" in str(ad["M2"].value)
    # Padrao N+3 digitos: N200 e novo item valido; a fronteira nao o afeta.
    assert eh_novo_item("N200") is True


def test_base_zero_visual_sobrevive_a_geracao_runtime():
    """A geracao oficial preserva a formula de B (_limpar_residuos pula '=')."""
    from _coleta_oficial import obter_coleta_oficial_bytes
    wb = load_workbook(io.BytesIO(obter_coleta_oficial_bytes()), data_only=False)
    try:
        ws = wb["itens_Remanesc"]
        for coord in ("B2", "B200"):
            assert str(ws[coord].value).startswith("="), f"{coord} perdeu a formula"
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# 3. Cadeia documental — indice, PREVIA, historico de VUs temporal
# ---------------------------------------------------------------------------

def _leitura_sintetica_26h(com_vta=False):
    por_ciclo = {
        "C0": {"ciclo": "C0", "computar_nesta_apuracao": None,
               "data_inicio": date(2025, 2, 1), "data_fim": date(2026, 1, 31),
               "percentual_reajuste": None, "fator_acumulado": 1.0,
               "situacao": "Base", "inicio_efeito_financeiro": None},
        "C1": {"ciclo": "C1", "computar_nesta_apuracao": "Sim",
               "data_inicio": date(2026, 2, 1), "data_fim": date(2027, 1, 31),
               "percentual_reajuste": 0.0307853139440224,
               "fator_acumulado": 1.0307853139440224,
               "situacao": "Computado",
               "inicio_efeito_financeiro": date(2026, 4, 1)},
    }
    leitura = {
        "ok": True,
        "controle": {"modo": "PC", "versao": "V10",
                     "ciclo_vigente": "C1", "data_corte": date(2026, 7, 1),
                     "indice": "IST (Anatel)"},
        "resumo": {},
        "parametros_v10": {"ok": True, "por_ciclo": por_ciclo,
                           "ciclos": list(por_ciclo.values())},
        "memoria_calculo": {"C1": [
            {"tipo": "RESULTADO", "ordem": 1, "fator_acumulado": 1.0307853,
             "variacao_final": 0.0307853,
             "metodo_fonte": "Divisao de Numero-Indice (IST/Anatel)"},
        ]},
        # Item 1 nasce em C0 (legado, sem vu_ciclos); N002 nasce em C1:
        # VU_C0 legitimamente ausente, VU_C1 = VU_ORIGINAL.
        "historico_vu": {"itens": [
            {"item": "1", "descricao": "Servico A", "vu_original": 10.0,
             "vu_ciclos": {}, "vu_vigente": 10.31},
            {"item": "N002", "descricao": None, "vu_original": 7031.62,
             "vu_ciclos": {"VU_C0": None, "VU_C1": 7031.62},
             "vu_vigente": 7031.62},
        ]},
        "itens_contrato": {"itens": [
            {"item": "1", "descricao": "Servico A", "qtd_contratada": 100.0,
             "vu_original": 10.0, "valor_total_original": 1000.0},
        ]},
        "resultados_xls": {
            "disponivel": True,
            "valores": {"VTA_FINAL": 127975842.65,
                        "RETRO_OFICIAL": 168372.32},
        },
    }
    if com_vta:
        leitura.pop("resultados_xls")
    return leitura


def test_indice_contratual_prioriza_controle_b7():
    from _sumario_executivo import montar_dados_sumario_executivo
    dados = montar_dados_sumario_executivo(_leitura_sintetica_26h())
    assert dados["identificacao"]["indice"] == "IST (Anatel)"


def test_vta_previa_com_numero_xls_oficial_sem_resultado_definitivo():
    from _sumario_executivo import montar_dados_sumario_executivo
    dados = montar_dados_sumario_executivo(_leitura_sintetica_26h())
    sintese = dados["sintese"]
    assert sintese.get("vta") is None
    assert sintese.get("vta_previa") == 127975842.65


def test_sem_previa_quando_nao_ha_numero_xls():
    from _sumario_executivo import montar_dados_sumario_executivo
    dados = montar_dados_sumario_executivo(_leitura_sintetica_26h(com_vta=True))
    assert dados["sintese"].get("vta_previa") is None


def test_historico_vu_nao_inventa_vu_pre_nascimento():
    from _sumario_executivo import montar_dados_sumario_executivo
    dados = montar_dados_sumario_executivo(_leitura_sintetica_26h())
    itens = {i["item"]: i for i in dados["historico_vu"]["itens"]}
    # Item legado sem vu_ciclos: fallback C0 = vu_original (compatibilidade).
    assert itens["1"]["vus"]["C0"] == 10.0
    # Novo item nascido em C1: C0 VAZIO (nunca zero/VU inventado); C1 = VU.
    assert itens["N002"]["vus"]["C0"] is None
    assert itens["N002"]["vus"]["C1"] == 7031.62


def test_pdf_sumario_sem_estado_situacao_conclusao_e_notas():
    from _sumario_executivo import (
        gerar_sumario_executivo_pdf, montar_dados_sumario_executivo,
    )
    fonte = (RAIZ / "_sumario_executivo.py").read_text(encoding="utf-8")
    assert "Estado do retroativo" not in fonte
    assert "Situação do processo" not in fonte
    assert 'estilos["nota"]' not in fonte  # notas em italico retiradas
    # PDF gera sem excecao com a PREVIA aplicada.
    pdf = gerar_sumario_executivo_pdf(
        montar_dados_sumario_executivo(_leitura_sintetica_26h())
    )
    assert pdf[:4] == b"%PDF"


def _xml_docx(docx_bytes: bytes) -> str:
    with zipfile.ZipFile(io.BytesIO(docx_bytes)) as z:
        return z.read("word/document.xml").decode("utf-8")


def test_saneador_mostra_vta_previa_com_highlight_verde():
    from _templates_documentos import gerar_despacho_saneador
    xml = _xml_docx(gerar_despacho_saneador(_leitura_sintetica_26h()))
    assert "PRÉVIA" in xml
    assert "127.975.842,65" in xml
    # Highlight verde aplicado somente na palavra PREVIA (run dedicado).
    assert re.search(
        r'<w:highlight w:val="green"/></w:rPr><w:t[^>]*>PRÉVIA</w:t>', xml
    ), "run PRÉVIA sem highlight verde dedicado"


def test_termo_apostila_reusa_quadro_de_vus_existente():
    from _templates_documentos import gerar_termo_apostila
    xml = _xml_docx(gerar_termo_apostila(_leitura_sintetica_26h()))
    # Quadro unico Item | VU_C0 | VU_C1 com o novo item presente.
    assert "N002" in xml
    assert "7.031,62" in xml


# ---------------------------------------------------------------------------
# 4. Wiring — leitor (teto canonico + B7) e limpeza da interface
# ---------------------------------------------------------------------------

def test_leitor_usa_capacidade_canonica_e_le_indice():
    fonte = (RAIZ / "_leitor_masterfile_v10.py").read_text(encoding="utf-8")
    assert "min(ws.max_row, ULTIMA_LINHA_PCS)" in fonte
    assert "min(ws.max_row, 100)" not in fonte  # teto legado eliminado
    assert '_achar_valor(c, "indice utilizado")' in fonte


def test_runtime_propaga_chaves_canonicas_da_leitura():
    fonte = (RAIZ / "_coleta_reajuste_documentos.py").read_text(encoding="utf-8")
    for chave in ('"controle"', '"historico_vu"', '"memoria_calculo"',
                  '"resultados_xls"'):
        assert f'{chave}: leitura.get({chave}' in fonte.replace("\n", "")


def test_pagina_sem_aviso_e_sem_expander_confirmado():
    fonte = (RAIZ / "pages" / "03_Valor_Global.py").read_text(encoding="utf-8")
    assert "render_status_base_coleta(resultado" not in fonte
    assert "render_cobertura_temporal(" not in fonte
