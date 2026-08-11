# -*- coding: utf-8 -*-
"""Etapa 44 — novos itens por aditivo: NAO APLICAVEL, orientacao e VTA.

Tres pontos protegidos, todos aplicados em RUNTIME (o template binario
homologado permanece intacto):

1. itens_Remanesc distingue TRES estados na abertura de cada ciclo — zero real,
   vazio real (pendencia) e NAO APLICAVEL (cinza, item ainda inexistente).
2. A orientacao de novo item por aditivo diz onde COMECAR o cadastro, sem
   secao/linha/caixa/aba nova.
3. A completude da abertura passa a usar o nascimento POR DATA
   (posicao_contratual!AL, CICLO_NASCIMENTO_DATA) no lugar do nascimento por
   QUANTIDADE (posicao_contratual!Y): item criado depois da abertura nao a
   torna incompleta e nao provoca recuo do VTA para um ciclo anterior.
"""
from __future__ import annotations

import io
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from _coleta_oficial import (  # noqa: E402
    TEMPLATE_COLETA_OFICIAL,
    obter_coleta_oficial_bytes,
)

CINZA_NAO_APLICAVEL = "FFD9D9D9"
VERDE_POSITIVO = "FFC6EFCE"
AZUL_ZERO = "FFDDEBF7"
AMARELO_PENDENCIA = "FFFFF2CC"

# QTD_REM_BASE_Cn manual em itens_Remanesc -> indice do ciclo de abertura.
COLS_ABERTURA = (("E", 1), ("G", 2), ("I", 3), ("K", 4))
# VALOR_REM_INICIO_Cn correspondente.
COLS_VALOR = (("F", 1), ("H", 2), ("J", 3), ("L", 4))


@pytest.fixture(scope="module")
def wb_runtime():
    """Coleta gerada pelo caminho oficial (a mesma que o fiscal baixa)."""
    return load_workbook(io.BytesIO(obter_coleta_oficial_bytes()), data_only=False)


@pytest.fixture(scope="module")
def wb_template():
    return load_workbook(TEMPLATE_COLETA_OFICIAL, data_only=False)


def _regras_cf(ws, faixa: str) -> list[tuple[str, str]]:
    """(formula, cor de preenchimento) das regras de uma faixa, na ordem."""
    saida: list[tuple[str, str]] = []
    for cf in ws.conditional_formatting:
        if str(cf.sqref) != faixa:
            continue
        for regra in cf.rules:
            cor = None
            if regra.dxf is not None and regra.dxf.fill is not None:
                cor = regra.dxf.fill.fgColor.rgb
            saida.append((str(regra.formula[0]), str(cor)))
    return saida


# --------------------------------------------------------------------------- #
# CASO A — item anterior a abertura: comportamento normal                      #
# --------------------------------------------------------------------------- #
def test_caso_a_item_anterior_a_abertura_continua_exigido(wb_runtime):
    """N001 (nascido antes de C1) segue exigido na abertura de C1.

    O criterio temporal e `AL <= n`: um item cujo nascimento POR DATA e C1
    permanece dentro da exigencia da abertura de C1 — a correcao nao afrouxa
    a completude, apenas a recorta pela data.
    """
    mem = wb_runtime["MEMORIA_RESULTADOS"]
    for celula, limiar in (("W41", 0), ("W42", 1), ("W43", 2), ("W44", 3), ("W45", 4)):
        formula = str(mem[celula].value)
        assert f"posicao_contratual!$AL$2:$AL$201<={limiar}" in formula
        # A exigencia continua sendo "remanescente numerico" (vazio != zero).
        assert "1-ISNUMBER(posicao_contratual!$" in formula
        assert formula.count("(") == formula.count(")")


def test_caso_a_nascimento_por_data_deriva_da_data_de_efeito(wb_runtime):
    """AL e derivado das aberturas temporais AG:AK, nao da quantidade."""
    pc = wb_runtime["posicao_contratual"]
    assert str(pc["AL1"].value).strip() == "CICLO_NASCIMENTO_DATA"
    formula = str(pc["AL2"].value)
    for coluna in ("$AG2", "$AH2", "$AI2", "$AJ2", "$AK2"):
        assert coluna in formula
    # AG:AK descontam os aditivos com efeito POSTERIOR a abertura do ciclo.
    assert "$AB2" in str(pc["AG2"].value) and "$E2" in str(pc["AG2"].value)
    assert "$AC2" in str(pc["AH2"].value) and "$I2" in str(pc["AH2"].value)


# --------------------------------------------------------------------------- #
# CASO B — item posterior a abertura: vazio + cinza, abertura segue completa   #
# --------------------------------------------------------------------------- #
@pytest.mark.parametrize("coluna,idx", COLS_ABERTURA)
def test_caso_b_nao_aplicavel_e_cinza_e_tem_prioridade(wb_runtime, coluna, idx):
    """Abertura anterior ao nascimento POR DATA: cinza discreto, nunca vermelho.

    A regra NAO APLICAVEL vem primeiro e com stopIfTrue, de modo que a celula
    jamais e pintada como pendencia; e usa o espelho local BI (a formatacao
    condicional precisa ser da propria aba para nao migrar para a extensao x14,
    invisivel ao openpyxl que gera cada Coleta).
    """
    ws = wb_runtime["itens_Remanesc"]
    regras = _regras_cf(ws, f"{coluna}2:{coluna}201")
    assert len(regras) == 4, f"CF de {coluna} incompleta: {regras}"
    formula_na, cor_na = regras[0]
    assert f"$BI2>{idx}" in formula_na and "ISNUMBER($BI2)" in formula_na
    assert cor_na == CINZA_NAO_APLICAVEL
    cores = {cor for _, cor in regras}
    assert "FFFF0000" not in cores and "FFFFC7CE" not in cores  # nada de vermelho


@pytest.mark.parametrize("coluna,idx", COLS_ABERTURA)
def test_caso_b_cinza_renderiza_no_excel(wb_runtime, wb_template, coluna, idx):
    """O dxf do NAO APLICAVEL precisa de `bgColor` para o Excel pintar.

    Num formato diferencial o Excel le a cor de fundo em `bgColor`; um
    `patternType="solid"` com apenas `fgColor` resolve para "sem
    preenchimento" — foi por isso que o cinza nunca apareceu, embora a regra
    estivesse correta. O template ainda carrega a forma nao renderizavel; a
    correcao e aplicada em runtime.
    """
    faixa = f"{coluna}2:{coluna}201"
    regra_runtime = next(
        r for cf in wb_runtime["itens_Remanesc"].conditional_formatting
        if str(cf.sqref) == faixa
        for r in cf.rules
        if f"$BI2>{idx}" in str(r.formula[0])
    )
    fill = regra_runtime.dxf.fill
    assert fill.fgColor.rgb == CINZA_NAO_APLICAVEL
    assert fill.bgColor.rgb == CINZA_NAO_APLICAVEL, (
        "sem bgColor o Excel nao pinta o dxf — o cinza fica invisivel"
    )
    # Nenhuma outra regra da faixa foi tocada (verde/azul/amarelo intactos).
    outras_rt = [
        str(r.dxf.fill.fgColor.rgb)
        for cf in wb_runtime["itens_Remanesc"].conditional_formatting
        if str(cf.sqref) == faixa
        for r in cf.rules
        if f"$BI2>{idx}" not in str(r.formula[0])
    ]
    outras_tp = [
        str(r.dxf.fill.fgColor.rgb)
        for cf in wb_template["itens_Remanesc"].conditional_formatting
        if str(cf.sqref) == faixa
        for r in cf.rules
        if f"$BI2>{idx}" not in str(r.formula[0])
    ]
    assert outras_rt == outras_tp


@pytest.mark.parametrize("coluna,idx", COLS_ABERTURA)
def test_caso_b_nao_aplicavel_nao_escreve_zero(wb_runtime, coluna, idx):
    """A celula de QTD_REM_BASE_Cn permanece de ENTRADA e visualmente vazia."""
    ws = wb_runtime["itens_Remanesc"]
    for linha in (2, 100, 200):
        assert ws[f"{coluna}{linha}"].value is None, (
            f"{coluna}{linha} deveria estar vazia (entrada do fiscal)"
        )


@pytest.mark.parametrize("coluna,idx", COLS_VALOR)
def test_caso_b_valor_de_abertura_vazio_quando_item_nao_existia(wb_runtime, coluna, idx):
    """VALOR_REM_INICIO_Cn acompanha o estado NAO APLICAVEL (vazio, nao 0,00)."""
    ws = wb_runtime["itens_Remanesc"]
    formula = str(ws[f"{coluna}2"].value)
    assert (
        f"AND(ISNUMBER(posicao_contratual!$AL2),posicao_contratual!$AL2>{idx})"
        in formula
    )
    assert formula.startswith("=IF(OR(") and ',"",' in formula


def test_caso_b_abertura_nao_fica_incompleta_por_item_inexistente(wb_runtime):
    """Nenhuma linha de completude exige item nascido depois da abertura."""
    mem = wb_runtime["MEMORIA_RESULTADOS"]
    for celula in ("W41", "W42", "W43", "W44", "W45"):
        formula = str(mem[celula].value)
        assert "posicao_contratual!$Y$2:$Y$201" not in formula, (
            f"{celula} ainda usa o nascimento por QUANTIDADE (criterio antigo)"
        )


# --------------------------------------------------------------------------- #
# CASO C — zero real continua ZERO; CASO D — vazio real continua PENDENCIA     #
# --------------------------------------------------------------------------- #
@pytest.mark.parametrize("coluna,idx", COLS_ABERTURA)
def test_caso_c_zero_real_tem_estado_proprio(wb_runtime, coluna, idx):
    """0,00 informado e um estado distinto do NAO APLICAVEL (cor propria)."""
    regras = _regras_cf(wb_runtime["itens_Remanesc"], f"{coluna}2:{coluna}201")
    formula_zero, cor_zero = regras[2]
    assert f"{coluna}2=0" in formula_zero
    assert cor_zero == AZUL_ZERO != CINZA_NAO_APLICAVEL
    _formula_pos, cor_pos = regras[1]
    assert cor_pos == VERDE_POSITIVO


@pytest.mark.parametrize("coluna,idx", COLS_ABERTURA)
def test_caso_d_vazio_real_continua_pendencia(wb_runtime, coluna, idx):
    """Item que JA existia e sem quantidade segue como informacao faltante."""
    regras = _regras_cf(wb_runtime["itens_Remanesc"], f"{coluna}2:{coluna}201")
    formula_pend, cor_pend = regras[3]
    assert f'{coluna}2=""' in formula_pend
    assert f"OR(NOT(ISNUMBER($BI2)),$BI2<={idx})" in formula_pend
    assert cor_pend == AMARELO_PENDENCIA


def test_caso_d_vazio_nunca_vira_zero_estrutural_na_celula(wb_runtime):
    """A completude segue exigindo ISNUMBER: vazio nao satisfaz a exigencia."""
    mem = wb_runtime["MEMORIA_RESULTADOS"]
    colunas = {"W41": "$G$", "W42": "$K$", "W43": "$O$", "W44": "$S$", "W45": "$W$"}
    for celula, coluna in colunas.items():
        assert f"1-ISNUMBER(posicao_contratual!{coluna}2:" in str(mem[celula].value)


# --------------------------------------------------------------------------- #
# CASO E — VTA: sem recuo indevido e sem tocar o valor oficial                 #
# --------------------------------------------------------------------------- #
def test_caso_e_vta_oficial_intacto(wb_runtime, wb_template):
    """B26/T25 e a FORMA 2 permanecem exatamente como no template homologado."""
    runtime = wb_runtime["MEMORIA_RESULTADOS"]
    template = wb_template["MEMORIA_RESULTADOS"]
    for celula in ("B26", "T21", "T22", "T23", "T24", "T25", "T26", "T27",
                   "W46", "W47", "W48", "W53", "W54", "W55", "W56", "W57"):
        assert runtime[celula].value == template[celula].value, (
            f"{celula} foi alterada — fora do escopo da Etapa 44"
        )


def test_caso_e_fallback_continua_pela_ultima_abertura_completa(wb_runtime):
    """W46 segue escolhendo a ultima abertura completa <= ciclo vigente."""
    formula = str(wb_runtime["MEMORIA_RESULTADOS"]["W46"].value)
    assert "$W$41:$W$45" in formula and "$T$20" in formula


def test_caso_e_quantidades_da_abertura_nao_foram_tocadas(wb_runtime, wb_template):
    """G/K/O/S/W (remanescente ajustado por ciclo) permanecem identicas.

    O zero estrutural do item inexistente e interpretacao da COMPLETUDE; nunca
    reescreve a quantidade que alimenta o VTA.
    """
    runtime = wb_runtime["posicao_contratual"]
    template = wb_template["posicao_contratual"]
    for coluna in ("G", "K", "O", "S", "W", "Y", "AL"):
        for linha in (2, 50, 200):
            assert (
                runtime[f"{coluna}{linha}"].value
                == template[f"{coluna}{linha}"].value
            )


# --------------------------------------------------------------------------- #
# CASO F — CICLO_EM_EXECUCAO preservada                                        #
# --------------------------------------------------------------------------- #
def test_caso_f_quantidade_atual_continua_manual(wb_runtime):
    """A coluna de posicao atual segue PARA PREENCHER: nada e presumido.

    Um acrescimo de 945 unidades NAO vira remanescente atual de 945 — pode ter
    havido consumo. A celula continua sem formula: vazio = nao informado.
    """
    ws = wb_runtime["CICLO_EM_EXECUCAO"]
    assert "PREENCHER" in str(ws["C12"].value)
    for linha in (13, 100, 211):
        assert ws[f"C{linha}"].value is None, (
            f"C{linha} deixou de ser entrada manual do fiscal"
        )


def test_caso_f_guarda_temporal_da_posicao_preservada(wb_runtime, wb_template):
    """W49 (disponibilidade da posicao fisica) nao foi tocada."""
    runtime = str(wb_runtime["MEMORIA_RESULTADOS"]["W49"].value)
    assert runtime == str(wb_template["MEMORIA_RESULTADOS"]["W49"].value)
    assert 'INDIRECT("CICLO_EM_EXECUCAO!$D$5")>CONTROLE!$B$3' in runtime


# --------------------------------------------------------------------------- #
# Orientacao de UX — discreta, reaproveitando espaco existente                 #
# --------------------------------------------------------------------------- #
def test_orientacao_indica_comecar_por_itens_remanesc(wb_runtime):
    a1 = str(wb_runtime["aditivos"]["A1"].value)
    assert a1.startswith("ITEM")          # chave do leitor preservada
    assert "itens_Remanesc PRIMEIRO" in a1
    assert "Depois, a data e a quantidade aqui." in a1
    m2 = str(wb_runtime["aditivos"]["M2"].value)
    assert "CADASTRAR EM itens_Remanesc: ITEM + VU_ORIGINAL" in m2
    assert "depois a data e a quantidade aqui." in m2


def test_orientacao_nao_desloca_layout(wb_runtime, wb_template):
    """Sem aba nova, sem linha nova e sem aumento de altura das areas."""
    novas = set(wb_runtime.sheetnames) - set(wb_template.sheetnames)
    assert novas == {"CICLO_EM_EXECUCAO"}  # ja criada em runtime desde a 29C.1
    for aba in ("aditivos", "itens_Remanesc"):
        runtime, template = wb_runtime[aba], wb_template[aba]
        assert runtime.max_row == template.max_row
        assert runtime.max_column == template.max_column
        altura_rt = getattr(runtime.row_dimensions.get(1), "height", None)
        altura_tp = getattr(template.row_dimensions.get(1), "height", None)
        assert altura_rt == altura_tp
