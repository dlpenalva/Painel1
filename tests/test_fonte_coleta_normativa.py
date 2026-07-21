"""Fonte normativa única da Coleta: NOVA coleta = COLETA_REAJUSTE_OFICIAL.xlsx (MODELO B).

Consolida por teste a regra de fonte de verdade:

* toda nova coleta é gerada a partir do template oficial (Modelo B);
* o template legado `Coleta_Reajuste.xlsx` (Modelo A) nunca gera nova coleta;
* se o template oficial estiver ausente, o sistema falha de forma explícita —
  NÃO há fallback silencioso para o Modelo A.
"""

import io
import unittest
from pathlib import Path

from openpyxl import load_workbook

import _coleta_oficial
from _coleta_oficial import NOME_ARQUIVO_COLETA_OFICIAL, obter_coleta_oficial_bytes


class FonteColetaNormativa(unittest.TestCase):
    def test_nome_do_arquivo_e_o_oficial(self):
        self.assertEqual(NOME_ARQUIVO_COLETA_OFICIAL, "COLETA_REAJUSTE_OFICIAL.xlsx")

    def test_nova_coleta_carrega_layout_modelo_b(self):
        wb = load_workbook(io.BytesIO(obter_coleta_oficial_bytes()), data_only=False)
        self.assertIn("posicao_contratual", wb.sheetnames)
        # marcador do MODELO B: remanescente real (não "sem aditivo" do legado).
        cabecalho = wb["itens_Remanesc"]["E1"].value or ""
        self.assertIn("QTD_REM_BASE_C1", cabecalho)
        self.assertNotIn("SEM_ADITIVO", cabecalho)

    def test_template_oficial_ausente_gera_erro_sem_fallback(self):
        original = _coleta_oficial.TEMPLATE_COLETA_OFICIAL
        try:
            _coleta_oficial.TEMPLATE_COLETA_OFICIAL = Path("___template_oficial_inexistente___.xlsx")
            # Deve falhar explicitamente; nunca retornar bytes do Modelo A.
            with self.assertRaises(FileNotFoundError):
                obter_coleta_oficial_bytes()
        finally:
            _coleta_oficial.TEMPLATE_COLETA_OFICIAL = original


if __name__ == "__main__":
    unittest.main()
