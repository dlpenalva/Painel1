# -*- coding: utf-8 -*-
"""RESULTADOS-FINAL-1: acabamento da aba RESULTADOS restaurada.

Owner da mudanca no template: tools/aplicar_resultados_final_1.py.

Este teste cobre exatamente o que a frente prometeu e, principalmente, o que
ela prometeu NAO fazer. A regra que orienta cada assert e a mesma: mudar
apresentacao e rotulo pode; mexer em coordenada, formula economica,
metodologia ou nome definido nao pode.
"""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "templates" / "COLETA_REAJUSTE_OFICIAL.xlsx"

# Faixa dos AJUSTES MANUAIS: fica oculta, mas nao pode perder uma celula.
PRIMEIRA_LINHA_AJUSTES = 41
ULTIMA_LINHA_AJUSTES = 50

# Pinos tecnicos que a frente jurou preservar nas mesmas coordenadas.
FORMULAS_PRESERVADAS = {
    "B3": '=IF(OR($H$8="REVISE",$H$14="REVISE",$H$24="REVISE",$H$33="REVISE",'
          'COUNTIF($H$43:$H$50,"REVISE")>0),"REVISE",'
          'IF($H$33="ESTIMADO","ESTIMADO","VALIDADO"))',
    "B22": '=IF(COUNT(B16:B20)=0,"",ROUND(SUM(B16:B20),2))',
    "B38": '=IF(OR(B36="",B37=""),"",ROUND(B37-B36,2))',
    "B86": '=IF(VTA_FINAL="","",VTA_FINAL)',
}

NOMES_DEFINIDOS_OBRIGATORIOS = {
    "STATUS_RESULTADOS": "RESULTADOS!$B$3",
    "RETROATIVO_POTENCIAL_PC": "MEMORIA_RESULTADOS!$T$38",
    "VTA_FINAL": "MEMORIA_RESULTADOS!$B$26",
    "RETRO_OFICIAL": "MEMORIA_RESULTADOS!$B$16",
    "EXECUTADO_APURADO": "RESULTADOS!$B$83",
    "AJUSTES_DEVIDOS": "RESULTADOS!$B$84",
    "CONFERENCIA_FORMACAO_VTA": "RESULTADOS!$B$87",
}

MOEDA_CANONICA = '"R$"\\ #,##0.00;\\-"R$"\\ #,##0.00;"R$"\\ 0.00;"—"'
AZUL_ESCURO = "FF1F4E78"


@pytest.fixture(scope="module")
def wb():
    return load_workbook(TEMPLATE, data_only=False)


@pytest.fixture(scope="module")
def res(wb):
    return wb["RESULTADOS"]


# --------------------------------------------------- 1. AJUSTES MANUAIS
def test_ajustes_manuais_ficam_ocultos_sem_perder_conteudo(res):
    """Ocultar nao e apagar: as 10 linhas somem da vista e continuam la."""
    for linha in range(PRIMEIRA_LINHA_AJUSTES, ULTIMA_LINHA_AJUSTES + 1):
        assert res.row_dimensions[linha].hidden is True, linha

    assert res["A41"].value == "5. AJUSTES MANUAIS"
    assert res["A42"].value == "Tipo"
    for linha in range(43, 51):
        assert str(res["H%d" % linha].value).startswith("=IF(COUNTA("), linha


def test_c43_g50_permanece_nas_mesmas_coordenadas(res):
    """A area de entrada do fiscal segue vazia, intacta e enderecavel."""
    for linha in range(43, 51):
        for coluna in "CDEFG":
            endereco = "%s%d" % (coluna, linha)
            assert res[endereco].value in (None, ""), endereco


def test_validacoes_dos_ajustes_manuais_sobrevivem(res):
    faixas = {str(dv.sqref) for dv in res.data_validations.dataValidation}
    assert {"G43:G50", "C46:C50"} <= faixas


# --------------------------------------------------- 2. TABELA 6
def test_tabela_6_titulo_e_cabecalho(res):
    assert res["A53"].value == "6. TOTAIS E INDICADORES DE CONFERÊNCIA"
    assert res["A53"].fill.fgColor.rgb == AZUL_ESCURO
    assert res["A53"].font.color.rgb == "FFFFFFFF"
    assert res["A53"].font.b is True
    assert res["C54"].value == "REFERÊNCIA PARA CONFERÊNCIA"


def test_tabela_6_nao_expoe_nomes_tecnicos(res):
    """O fiscal nao deve ler abas, celulas nem nomes definidos."""
    proibidos = ("MEMORIA", "itens_PC", "CONTROLE!", "VTA_FINAL", "!B", "!T",
                 "posicao_contratual", "CICLO_EM_EXECUCAO")
    for linha in range(55, 67):
        texto = str(res["C%d" % linha].value or "")
        for termo in proibidos:
            assert termo not in texto, "C%d expoe %r: %r" % (linha, termo, texto)
    assert "Linhas 10 a 13" not in str(res["B64"].value)


def test_tabela_6_valores_em_moeda_e_destaques_em_negrito(res):
    for linha in (55, 56, 57, 58, 59, 60, 61, 62, 63, 65):
        assert res["B%d" % linha].number_format == MOEDA_CANONICA, linha
    # B64 e B66 sao textuais: mascara monetaria neles seria mentira visual.
    assert res["B64"].number_format != MOEDA_CANONICA
    assert res["B66"].number_format != MOEDA_CANONICA
    assert res["B63"].font.b is True   # VTA Oficial
    assert res["B66"].font.b is True   # Status


# --------------------------------------------------- 3. TABELA 7
def test_tabela_7_bloco_explicativo(res):
    assert res["A68"].fill.fgColor.rgb == AZUL_ESCURO
    assert str(res["A69"].value).startswith("COMPOSIÇÃO ADOTADA")
    assert "ajustes ainda devidos" in str(res["A69"].value)
    assert "acertos ainda devidos" not in str(res["A69"].value)
    assert "saldo remanescente atualizado" in str(res["A69"].value)
    assert str(res["A70"].value).startswith('="MÉTODO APLICADO — "&')


# --------------------------------------------------- 4. TABELA 8
def test_tabela_8_padronizada(res):
    assert res["A71"].fill.fgColor.rgb == AZUL_ESCURO
    assert res["A72"].font.b is True
    for linha in range(73, 78):
        for coluna in "BCD":
            endereco = "%s%d" % (coluna, linha)
            assert res[endereco].number_format == MOEDA_CANONICA, endereco
    assert str(res["A78"].value).startswith("Esta conferencia compara")


# --------------------------------------------------- 5. TABELA 9
def test_tabela_9_vta_oficial_tem_a_maior_hierarquia(res):
    assert res["A79"].fill.fgColor.rgb == AZUL_ESCURO
    assert res["A79"].font.sz == 12          # destaque sobre as tabelas 6-8
    total = res["A86"]
    assert total.fill.fgColor.rgb == AZUL_ESCURO
    assert total.font.color.rgb == "FFFFFFFF"
    assert total.font.b is True
    # Verde e reservado a semantica de validacao; o VTA nunca o usa.
    for coluna in "ABC":
        assert res["%s86" % coluna].fill.fgColor.rgb != "FF00B050"
    conferencia = res["A87"]
    assert conferencia.fill.fgColor.rgb == "FFF2F2F2"
    assert conferencia.font.b is not True


# --------------------------------------------------- 6. RETROATIVO POTENCIAL
def test_retroativo_potencial_vem_do_nome_canonico_existente(res):
    """Sem calculo novo: G22 le RETROATIVO_POTENCIAL_PC e nada mais."""
    formula = str(res["G22"].value)
    assert "RETROATIVO_POTENCIAL_PC" in formula
    assert "T38" not in formula          # le pelo nome, nao pela coordenada
    assert "SUM" not in formula.upper()  # nao reimplementa a grandeza
    assert res["G22"].number_format == MOEDA_CANONICA


def test_retroativo_potencial_se_declara_em_analise(res):
    rotulo = str(res["E22"].value)
    assert "Retroativo potencial" in rotulo
    assert "em análise" in rotulo
    assert "Não integra o retroativo reconhecido" in rotulo
    assert res["E22"].alignment.wrapText is True


def test_retroativo_potencial_fica_ao_lado_do_reconhecido_sem_deslocar(res):
    """E22:H22 e a linha do TOTAL do retroativo reconhecido (D22)."""
    assert res["D22"].value == '=IFERROR(RETRO_OFICIAL,"")'
    mesclas = {str(m) for m in res.merged_cells.ranges}
    assert {"E22:F22", "G22:H22"} <= mesclas
    # As mesclas vizinhas nao foram tocadas nem absorvidas.
    assert {"E16:H21", "E23:H23"} <= mesclas


# --------------------------------------------------- 7. CONTRATOS
def test_pinos_tecnicos_intactos(res):
    for coordenada, formula in FORMULAS_PRESERVADAS.items():
        assert res[coordenada].value == formula, coordenada


def test_nomes_definidos_intactos(wb):
    for nome, destino in NOMES_DEFINIDOS_OBRIGATORIOS.items():
        assert nome in wb.defined_names, nome
        assert wb.defined_names[nome].value == destino, nome


def test_b83_b87_seguem_compondo_o_vta(res):
    assert str(res["B83"].value).startswith("=IF(MEMORIA_RESULTADOS!$B$4=")
    assert res["B85"].value == (
        '=IF(MEMORIA_RESULTADOS!$D$35="","",MEMORIA_RESULTADOS!$D$35)'
    )
    assert "$B$86-($B$83+N($B$84)+$B$85)" in res["B87"].value.replace(" ", "")


def test_memoria_resultados_permanece_oculta_e_intocada(wb):
    memoria = wb["MEMORIA_RESULTADOS"]
    assert memoria.sheet_state == "hidden"
    assert memoria["T38"].value == (
        '=ROUND(SUMIFS(itens_PC!$J$2:$J$5001,'
        'itens_PC!$B$2:$B$5001,"<="&$T$31),2)'
    )


def test_a_aba_nao_ganhou_camada_nova(res):
    """A UX2 vivia em 90:166; a frente nao pode ter reaberto essa porta."""
    assert res.max_row == 87
