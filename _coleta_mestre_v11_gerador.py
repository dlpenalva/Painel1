"""
_coleta_mestre_v11_gerador.py
------------------------------
Gerador da ColetaMestre v11 — Base Analítica (cl8us 2.0).

Estrutura vertical: financeiro por competência e itens por ciclo
em linhas independentes, facilitando o preenchimento e o diagnóstico.

Abas produzidas:
    INICIO              — orientação de uso (visível)
    FINANCEIRO_COMP     — financeiro mensal por competência (visível, fiscal preenche)
    ITENS_CICLO         — itens por ciclo em linhas verticais (visível, GCC/fiscal preenche)
    ADITIVOS_SUP        — aditivos e supressões (visível, GCC preenche)
    PARAMETROS_CICLOS   — parâmetros e ciclos (oculta, sistema lê)
    CONFERENCIA         — diagnóstico automático (oculta, sistema lê)

Uso:
    from _coleta_mestre_v11_gerador import gerar_coleta_mestre_v11
    bytes_xlsx = gerar_coleta_mestre_v11()
"""

from io import BytesIO

import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ── Paleta de cores ──────────────────────────────────────────────
_P = {
    "sistema":   "0F766E",  # teal escuro — títulos
    "sistema_c": "CCFBF1",  # teal claro
    "calc":      "1E3A8A",  # azul escuro — calculadora
    "calc_c":    "DBEAFE",  # azul claro
    "gcc":       "7C3AED",  # roxo — GCC
    "gcc_c":     "EDE9FE",  # roxo claro
    "fisc":      "B45309",  # âmbar escuro — fiscal
    "fisc_c":    "FFF3CD",  # âmbar claro
    "auto":      "374151",  # cinza escuro — automático
    "auto_c":    "F1F5F9",  # cinza claro
    "alrt":      "991B1B",  # vermelho — alertas
    "alrt_c":    "FEE2E2",  # vermelho claro
    "ok":        "065F46",  # verde escuro — totais/OK
    "ok_c":      "D1FAE5",  # verde claro
    "borda":     "CBD5E1",
    "branco":    "FFFFFF",
    "texto":     "0F172A",
}

_MOEDA  = '"R$" #,##0.00'
_PERC   = '0.00%'
_FATOR  = '0.000000'
_QTD    = '#,##0.##'
_DATA   = 'MM/YYYY'


# ── Helpers de estilo ────────────────────────────────────────────

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="0F172A", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Arial")

def _aln(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _brd():
    s = Side(style="thin", color=_P["borda"])
    return Border(left=s, right=s, top=s, bottom=s)

def _cel(ws, r, c, val="", bg="FFFFFF", fg="0F172A", bold=False,
         fmt=None, h="center", wrap=False, italic=False, size=10):
    cc = ws.cell(r, c, val)
    cc.fill = _fill(bg)
    cc.font = _font(bold, fg, size, italic)
    cc.alignment = _aln(h, "center", wrap)
    cc.border = _brd()
    if fmt:
        cc.number_format = fmt
    return cc

def _titulo(ws, row, c1, c2, txt, size=11):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    c = ws.cell(row, c1, txt)
    c.fill = _fill(_P["sistema"])
    c.font = _font(True, "FFFFFF", size)
    c.alignment = _aln("left", "center")
    c.border = _brd()
    ws.row_dimensions[row].height = 26

def _secao(ws, row, c1, c2, txt, bg, fg="FFFFFF"):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    c = ws.cell(row, c1, f"  {txt}")
    c.fill = _fill(bg)
    c.font = _font(True, fg, 9)
    c.alignment = _aln("left", "center")
    c.border = _brd()
    ws.row_dimensions[row].height = 20

def _nota(ws, row, c1, c2, txt, bg=None, fg=None):
    bg = bg or _P["auto_c"]
    fg = fg or _P["auto"]
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    c = ws.cell(row, c1, txt)
    c.fill = _fill(bg)
    c.font = _font(italic=True, color=fg, size=9)
    c.alignment = _aln("left", "center", True)
    c.border = _brd()
    ws.row_dimensions[row].height = 20

def _hdr(ws, row, col, txt, bg, w=None):
    c = ws.cell(row, col, txt)
    c.fill = _fill(bg)
    c.font = _font(True, "FFFFFF", 9)
    c.alignment = _aln("center", "center", True)
    c.border = _brd()
    ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 0, 46)
    if w:
        ws.column_dimensions[get_column_letter(col)].width = w
    return c

def _dv(ws, formula, ref):
    d = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(d)
    d.add(ref)
    return d


# ── Leitura do session_state ─────────────────────────────────────

def _dados_sessao():
    try:
        adm = st.session_state.get("dados_admissibilidade") or {}
    except Exception:
        adm = {}
    ciclos_raw = adm.get("ciclos", []) or []
    return {
        "indice":     str(adm.get("indice", "")),
        "data_base":  str(adm.get("data_base_original", "")),
        "valor_orig": adm.get("valor_original", 0) or 0,
        "ciclos":     ciclos_raw,
        "tem_dados":  bool(ciclos_raw),
    }


# ── Aba INICIO ───────────────────────────────────────────────────

def _construir_inicio(wb, dados):
    ws = wb.create_sheet("INICIO")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = _P["sistema"]
    ws.freeze_panes = "B7"

    for col, w in [("A", 4), ("B", 28), ("C", 48), ("D", 22)]:
        ws.column_dimensions[col].width = w

    _titulo(ws, 1, 1, 4, "cl8us · ColetaMestre v11 — Base Analítica", size=13)
    _nota(ws, 2, 1, 4,
          "  Versão de estrutura vertical. Cada linha representa uma competência ou um item/ciclo. "
          "Simplifica o preenchimento e reduz erros de dupla contagem.",
          _P["sistema_c"], _P["sistema"])

    # Esquema de preenchimento
    _secao(ws, 4, 1, 4, "Responsabilidade por cor", _P["sistema"])
    cores = [
        ("Azul", "Calculadora pré-preenche — não alterar"),
        ("Roxo",  "GCC preenche — campos de controle e equalização"),
        ("Âmbar", "Fiscal preenche — valores e quantidades"),
        ("Cinza", "Automático — calculado pela planilha"),
    ]
    bgs = [_P["calc_c"], _P["gcc_c"], _P["fisc_c"], _P["auto_c"]]
    fgs = [_P["calc"], _P["gcc"], _P["fisc"], _P["auto"]]
    for i, ((label, desc), bg, fg) in enumerate(zip(cores, bgs, fgs)):
        r = 5 + i
        _cel(ws, r, 1, "", bg)
        _cel(ws, r, 2, label, bg, fg, bold=True, h="left")
        _cel(ws, r, 3, desc, _P["auto_c"], _P["auto"], h="left", size=9)
        _cel(ws, r, 4, "", _P["auto_c"])
        ws.row_dimensions[r].height = 20

    # Abas
    _secao(ws, 10, 1, 4, "Abas desta planilha", _P["sistema"])
    abas_info = [
        ("FINANCEIRO_COMP",   "Fiscal + GCC", "Valor final reconhecido por competência, mês a mês"),
        ("ITENS_CICLO",       "GCC + Fiscal", "Itens por ciclo: C0, remanescentes, consumo"),
        ("ADITIVOS_SUP",      "GCC",          "Aditivos e supressões — tratamento no VTA"),
        ("PARAMETROS_CICLOS", "Sistema",      "Parâmetros e ciclos — leitura automática (oculta)"),
        ("CONFERENCIA",       "Sistema",      "Diagnóstico automático da base (oculta)"),
    ]
    _cel(ws, 11, 1, "", _P["sistema"])
    _cel(ws, 11, 2, "Aba", _P["sistema"], "FFFFFF", bold=True)
    _cel(ws, 11, 3, "Responsável", _P["sistema"], "FFFFFF", bold=True)
    _cel(ws, 11, 4, "Finalidade", _P["sistema"], "FFFFFF", bold=True)
    ws.row_dimensions[11].height = 20
    for i, (aba, resp, desc) in enumerate(abas_info):
        r = 12 + i
        _cel(ws, r, 1, "", _P["auto_c"])
        _cel(ws, r, 2, aba, _P["auto_c"], _P["auto"], bold=True, h="left", size=9)
        _cel(ws, r, 3, resp, _P["gcc_c"], _P["gcc"], h="left", size=9)
        _cel(ws, r, 4, desc, _P["auto_c"], _P["auto"], h="left", size=9, wrap=True)
        ws.row_dimensions[r].height = 20

    # Regra de ouro
    _secao(ws, 18, 1, 4, "Regra de ouro", _P["fisc"])
    _nota(ws, 19, 1, 4,
          "  Valor financeiro = valor final reconhecido para pagamento na competência de referência, "
          "já com glosas/descontos contratuais. Não é o valor pago pelo banco.",
          _P["fisc_c"], _P["fisc"])


# ── Aba FINANCEIRO_COMP ──────────────────────────────────────────

def _construir_financeiro_comp(wb, dados):
    ws = wb.create_sheet("FINANCEIRO_COMP")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = _P["fisc"]
    ws.freeze_panes = "A5"

    for col, w in [("A", 14), ("B", 16), ("C", 24), ("D", 18), ("E", 14), ("F", 36)]:
        ws.column_dimensions[col].width = w

    _titulo(ws, 1, 1, 6, "cl8us · FINANCEIRO_COMP — Valores por Competência")
    _nota(ws, 2, 1, 6,
          "  FISCAL: preencha SOMENTE col B (competência mm/aaaa) e col C (valor reconhecido). "
          "Col A (ciclo) e col D (efeito) são pré-preenchidos.",
          _P["fisc_c"], _P["fisc"])
    _nota(ws, 3, 1, 6,
          "  Valor reconhecido = valor final aprovado/faturado na competência, já com glosas. "
          "Não informar retenções ou líquido bancário.",
          _P["auto_c"], _P["auto"])

    # Cabeçalho
    ws.row_dimensions[4].height = 46
    hdrs = [
        ("[CALC] Ciclo",           _P["calc"],  "A"),
        ("[FISCAL] Competência\n(mm/aaaa)", _P["fisc"], "B"),
        ("[FISCAL] Valor reconhecido\n(R$)", _P["fisc"], "C"),
        ("[GCC] Tem efeito\nfinanceiro?", _P["gcc"],  "D"),
        ("[AUTO] Fator\naplicável",  _P["auto"], "E"),
        ("[GCC/AJUSTE] Observação GCC", _P["gcc"],  "F"),
    ]
    for ci, (txt, bg, _) in enumerate(hdrs, start=1):
        _hdr(ws, 4, ci, txt, bg)

    # Linhas de ciclos não-objeto (consolidados) — C0 sempre
    ciclos = dados.get("ciclos", [])
    nao_objeto = [{"ciclo": "C0", "label": "C0 — Executado consolidado (pré-reajuste)", "fat": 1.0}]
    nomes_nao_obj = {"C0"}
    for c in ciclos:
        nome_c = str(c.get("ciclo", "")).strip().upper()
        if nome_c == "C0":
            continue
        obj = c.get("objeto_analise_atual")
        obj_str = str(c.get("objeto_analise", "sim")).strip().lower()
        ja = c.get("ciclo_ja_concedido", False)
        ja_str = str(c.get("ja_formalizado", "não")).strip().lower()
        if obj is not None:
            eh_nao_obj = not bool(obj) or bool(ja)
        else:
            eh_nao_obj = (obj_str in ("não", "nao") or ja_str in ("sim",))
        if eh_nao_obj:
            fat = float(c.get("fator_acumulado", 1) or 1)
            nao_objeto.append({"ciclo": nome_c, "label": f"{nome_c} — Consolidado anterior", "fat": fat})
            nomes_nao_obj.add(nome_c)

    r = 5
    for cn in nao_objeto:
        ws.row_dimensions[r].height = 22
        _cel(ws, r, 1, cn["ciclo"], _P["auto_c"], _P["auto"], bold=True)
        _cel(ws, r, 2, cn["label"], _P["auto_c"], _P["auto"], italic=True, h="left")
        _cel(ws, r, 3, 0, _P["gcc_c"], _P["gcc"], fmt=_MOEDA)
        _cel(ws, r, 4, "Não", _P["gcc_c"], _P["gcc"])
        _cel(ws, r, 5, cn["fat"], _P["auto_c"], _P["auto"], fmt=_FATOR)
        _cel(ws, r, 6, "Linha consolidada — GCC informa o total executado.", _P["auto_c"], _P["auto"], italic=True, h="left", size=9)
        r += 1

    # Separador
    ws.merge_cells(f"A{r}:F{r}")
    c = ws.cell(r, 1, "  ▼  Linhas mensais — ciclos objeto da análise de reajuste")
    c.fill = _fill(_P["calc"])
    c.font = _font(True, "FFFFFF", 8)
    c.alignment = _aln("left", "center")
    c.border = _brd()
    ws.row_dimensions[r].height = 12
    r_inicio_mensal = r + 1
    r += 1  # avançar após linha do separador

    # Ciclos objeto — linhas mensais em branco
    ciclos_objeto = [c for c in ciclos if str(c.get("ciclo","")).strip().upper() not in nomes_nao_obj]
    if not ciclos_objeto:
        ciclos_objeto = ciclos  # fallback

    for ciclo in ciclos_objeto:
        nome = str(ciclo.get("ciclo", ""))
        fat_c = float(ciclo.get("fator_acumulado", 1) or 1)
        for _ in range(12):
            if r > 220:
                break
            ws.row_dimensions[r].height = 18
            _cel(ws, r, 1, nome, _P["calc_c"], _P["calc"])
            _cel(ws, r, 2, "", _P["fisc_c"], _P["fisc"], fmt=_DATA)
            _cel(ws, r, 3, "", _P["fisc_c"], _P["fisc"], fmt=_MOEDA)
            _cel(ws, r, 4, "Sim", _P["gcc_c"], _P["gcc"])
            _cel(ws, r, 5, fat_c, _P["auto_c"], _P["auto"], fmt=_FATOR)
            _cel(ws, r, 6, "", _P["auto_c"], _P["auto"], h="left", size=9)
            r += 1

    # Linhas em branco restantes
    while r <= 220:
        ws.row_dimensions[r].height = 18
        _cel(ws, r, 1, "", _P["auto_c"], _P["auto"])
        _cel(ws, r, 2, "", _P["fisc_c"], _P["fisc"], fmt=_DATA)
        _cel(ws, r, 3, "", _P["fisc_c"], _P["fisc"], fmt=_MOEDA)
        _cel(ws, r, 4, "Sim", _P["gcc_c"], _P["gcc"])
        _cel(ws, r, 5, 1.0, _P["auto_c"], _P["auto"], fmt=_FATOR)
        _cel(ws, r, 6, "", _P["auto_c"], _P["auto"], h="left", size=9)
        r += 1

    # Total
    ws.row_dimensions[221].height = 24
    ws.merge_cells("A221:B221")
    tc = ws.cell(221, 1, "TOTAL — linhas com efeito Sim")
    tc.fill = _fill(_P["ok"])
    tc.font = _font(True, "FFFFFF", 9)
    tc.alignment = _aln("right")
    tc.border = _brd()
    ws.cell(221, 2).fill = _fill(_P["ok"])
    ws.cell(221, 2).border = _brd()
    tv = ws.cell(221, 3, f'=ROUND(SUMIF(D{r_inicio_mensal}:D220,"Sim",C{r_inicio_mensal}:C220),2)')
    tv.fill = _fill(_P["ok"])
    tv.font = _font(True, "FFFFFF", 11)
    tv.alignment = _aln()
    tv.border = _brd()
    tv.number_format = _MOEDA
    for ci in [4, 5, 6]:
        ws.cell(221, ci).fill = _fill(_P["ok"])
        ws.cell(221, ci).border = _brd()

    # Secao ciclo em execucao — destaque ambar ao final da aba
    r_exec = 223
    ws.row_dimensions[r_exec - 1].height = 8
    _secao(ws, r_exec, 1, 6,
           "Ciclo em Execucao (corte operacional) — preencher SOMENTE se houver dados parciais do ciclo atual",
           _P["fisc"], "FFFFFF")
    _nota(ws, r_exec + 1, 1, 6,
          "  Use esta secao quando o fiscal informar dados ate uma data de corte dentro do ciclo ainda em andamento.",
          _P["fisc_c"], _P["fisc"])

    campos_exec = [
        ("Aplicar corte operacional?",         "Nao",  '"Sim,Nao"'),
        ("Ciclo em execucao",                  "",     None),
        ("Competencia de corte (mm/aaaa)",     "",     None),
        ("Valor executado C0 manual (R$)",     0,      None),
        ("Saldo remanescente ORIGINAL (R$)",   0,      None),
        ("Saldo remanescente ATUALIZADO (R$)", 0,      None),
        ("Saldo ja inclui aditivos?",          "Nao",  '"Sim,Nao"'),
    ]
    for i, (label, val, dv_formula) in enumerate(campos_exec):
        r = r_exec + 2 + i
        ws.row_dimensions[r].height = 20
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        cl = ws.cell(r, 1, label)
        cl.fill = _fill(_P["fisc_c"]); cl.font = _font(False, _P["fisc"], 9)
        cl.alignment = _aln("right", "center"); cl.border = _brd()
        for ci in [2, 3]:
            ws.cell(r, ci).fill = _fill(_P["fisc_c"]); ws.cell(r, ci).border = _brd()
        fmt = _MOEDA if "R$" in label else None
        vi = ws.cell(r, 4, val)
        vi.fill = _fill(_P["gcc_c"]); vi.font = _font(color=_P["gcc"])
        vi.alignment = _aln("left"); vi.border = _brd()
        if fmt: vi.number_format = fmt
        for ci in [5, 6]:
            ws.cell(r, ci).fill = _fill(_P["gcc_c"]); ws.cell(r, ci).border = _brd()
        if dv_formula:
            _dv(ws, dv_formula, f"D{r}")

    _dv(ws, '"Sim,Nao"', f"D5:D220")


# ── Aba ITENS_CICLO ──────────────────────────────────────────────

def _construir_itens_ciclo(wb, dados):
    """
    Aba ITENS_CICLO v11 — estrutura matricial (uma linha por item).
    Fiscal preenche: Item, Unidade, Qtd C0, VU C0, Rem C1, Rem C2... Rem corte.
    GCC preenche: fator do ciclo de corte.
    AUTO calcula: VT C0, VT Rem original, VT Rem atualizado.
    """
    ws = wb.create_sheet("ITENS_CICLO")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = _P["gcc"]
    ws.freeze_panes = "A6"

    ciclos = dados.get("ciclos", [])
    # Ciclos objeto (excluindo C0)
    nomes_ciclos = [str(c.get("ciclo","")).strip().upper()
                    for c in ciclos
                    if str(c.get("ciclo","")).strip().upper() != "C0"]
    # Fatores por ciclo
    fatores = {str(c.get("ciclo","")).strip().upper():
               float(c.get("fator_acumulado", 1) or 1)
               for c in ciclos}

    # Numero de colunas de remanescente = numero de ciclos objeto
    n_rem = len(nomes_ciclos)  # uma coluna por ciclo (rem inicio + corte = ultima)

    # Larguras base
    col_widths = [14, 10, 14, 16, 16]  # A=item, B=unidade, C=qtd C0, D=VU C0, E=VT C0(auto)
    # VT C0 ja esta em E (col 5)
    for _ in nomes_ciclos[:-1]:        # remanescentes intermediarios (C1, C2...)
        col_widths.append(14)
    col_widths.append(14)              # Rem corte (ultima)
    col_widths.append(14)              # VT Rem original (auto)
    col_widths.append(14)              # Fator ciclo corte (GCC)
    col_widths.append(16)              # VT Rem atualizado (auto)
    col_widths.append(30)              # Observacao

    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Linha 1 — titulo
    n_cols = len(col_widths)
    _titulo(ws, 1, 1, n_cols, "cl8us · ITENS_CICLO v11 — Base de Itens (uma linha por item)")

    # Linha 2 — nota fiscal
    _nota(ws, 2, 1, n_cols,
          "  FISCAL: preencha Item (col A), Unidade (col C), Qtd C0 (col D), VU C0 (col E) "
          "e quantidades remanescentes nas colunas verdes. Nao altere colunas cinzas.",
          _P["fisc_c"], _P["fisc"])

    # Linha 3 — nota GCC
    _nota(ws, 3, 1, n_cols,
          "  GCC: confirme o fator do ciclo de corte na penultima coluna (roxo). "
          "Colunas azuis sao pre-preenchidas pela Calculadora.",
          _P["gcc_c"], _P["gcc"])

    # Linha 4 — grupos de colunas
    ws.row_dimensions[4].height = 20
    # Base C0
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=5)
    c = ws.cell(4, 1, "  Base Contratual C0")
    c.fill = _fill(_P["fisc"]); c.font = _font(True, "FFFFFF", 9)
    c.alignment = _aln("left", "center"); c.border = _brd()
    for ci in range(2, 6): ws.cell(4, ci).fill = _fill(_P["fisc"]); ws.cell(4, ci).border = _brd()

    # Remanescentes
    col_rem_start = 6
    col_rem_end   = 6 + n_rem + 1  # +1 para coluna de corte
    if n_rem > 0:
        ws.merge_cells(start_row=4, start_column=col_rem_start,
                       end_row=4, end_column=col_rem_end)
        c = ws.cell(4, col_rem_start, "  Remanescentes por Ciclo  (fiscal preenche quantidade)")
        c.fill = _fill(_P["fisc"]); c.font = _font(True, "FFFFFF", 9)
        c.alignment = _aln("left", "center"); c.border = _brd()
        for ci in range(col_rem_start+1, col_rem_end+1):
            ws.cell(4, ci).fill = _fill(_P["fisc"]); ws.cell(4, ci).border = _brd()

    # Atualizacao
    col_atu_start = col_rem_end + 1
    ws.merge_cells(start_row=4, start_column=col_atu_start,
                   end_row=4, end_column=n_cols)
    c = ws.cell(4, col_atu_start, "  Atualizacao  (GCC + AUTO)")
    c.fill = _fill(_P["gcc"]); c.font = _font(True, "FFFFFF", 9)
    c.alignment = _aln("left", "center"); c.border = _brd()
    for ci in range(col_atu_start+1, n_cols+1):
        ws.cell(4, ci).fill = _fill(_P["gcc"]); ws.cell(4, ci).border = _brd()

    # Linha 5 — cabecalhos individuais
    ws.row_dimensions[5].height = 46
    hdrs_base = [
        ("[GCC] Item",        _P["gcc"]),
        ("[FISCAL] Unidade",  _P["fisc"]),
        ("[FISCAL] Qtd C0",   _P["fisc"]),
        ("[GCC] VU C0 (R$)",  _P["gcc"]),
        ("[AUTO] VT C0 (R$)", _P["auto"]),
    ]
    for ci, (txt, bg) in enumerate(hdrs_base, start=1):
        _hdr(ws, 5, ci, txt, bg)

    # Colunas de remanescente intermediario (C1, C2... exceto ultima)
    ci = 7
    for nome in nomes_ciclos[:-1]:
        fat = fatores.get(nome, 1.0)
        _hdr(ws, 5, ci, "[FISCAL]\nRem " + nome + "\n(qtd)", _P["fisc"])
        ci += 1

    # Coluna corte (ultima — remanescente atual)
    nome_corte = nomes_ciclos[-1] if nomes_ciclos else "Corte"
    fat_corte  = fatores.get(nome_corte, 1.0)
    _hdr(ws, 5, ci, "[FISCAL]\nRem " + nome_corte + "\n(corte/qtd)", _P["fisc"]); ci += 1

    # Atualizacao
    _hdr(ws, 5, ci, "[AUTO]\nVT Rem\noriginal (R$)", _P["auto"]); ci += 1
    _hdr(ws, 5, ci, "[GCC]\nFator\n" + nome_corte, _P["gcc"]); ci_fator = ci; ci += 1
    _hdr(ws, 5, ci, "[AUTO]\nVT Rem\nAtualiz. (R$)", _P["auto"]); ci += 1
    _hdr(ws, 5, ci, "[GCC] Observacao", _P["auto"])

    # Dados — linhas 6 a 205
    col_qtd_c0 = 3
    col_vu_c0  = 4
    col_vt_c0  = 5
    col_rem_ini = 6
    col_rem_corte = 6 + n_rem  # ultima coluna de remanescente
    col_vt_rem_orig  = col_rem_corte + 1
    col_fator_ciclo  = col_rem_corte + 2
    col_vt_rem_atu   = col_rem_corte + 3
    col_obs          = col_rem_corte + 4

    for r in range(6, 206):
        ws.row_dimensions[r].height = 18

        # A: item GCC
        _cel(ws, r, 1, "", _P["gcc_c"], _P["gcc"])
        # B: unidade fiscal
        _cel(ws, r, 2, "", _P["fisc_c"], _P["fisc"])
        # C: qtd C0 fiscal
        _cel(ws, r, 3, 0, _P["fisc_c"], _P["fisc"], fmt=_QTD)
        # D: VU C0 gcc
        _cel(ws, r, 4, 0, _P["gcc_c"], _P["gcc"], fmt=_MOEDA)
        # E: VT C0 auto
        ws.cell(r, 5, f"=IF(OR(C{r}=0,D{r}=0),0,ROUND(C{r}*D{r},2))").number_format = _MOEDA
        ws.cell(r, 5).fill = _fill(_P["auto_c"])
        ws.cell(r, 5).font = _font(color=_P["auto"])
        ws.cell(r, 5).alignment = _aln()
        ws.cell(r, 5).border = _brd()

        # Remanescentes intermediarios (C1, C2... exceto corte)
        for ci_r in range(col_rem_ini, col_rem_corte):
            _cel(ws, r, ci_r, 0, _P["fisc_c"], _P["fisc"], fmt=_QTD)

        # Rem corte (fiscal)
        _cel(ws, r, col_rem_corte, 0, _P["fisc_c"], _P["fisc"], fmt=_QTD)

        # VT Rem original = qtd_corte * VU C0
        cr = col_rem_corte
        ws.cell(r, col_vt_rem_orig,
                f"=IF(OR({get_column_letter(cr)}{r}=0,D{r}=0),0,"
                f"ROUND({get_column_letter(cr)}{r}*D{r},2))").number_format = _MOEDA
        ws.cell(r, col_vt_rem_orig).fill = _fill(_P["auto_c"])
        ws.cell(r, col_vt_rem_orig).font = _font(color=_P["auto"])
        ws.cell(r, col_vt_rem_orig).alignment = _aln()
        ws.cell(r, col_vt_rem_orig).border = _brd()

        # Fator ciclo corte (GCC pre-preenchido)
        _cel(ws, r, col_fator_ciclo, fat_corte, _P["gcc_c"], _P["gcc"], fmt=_FATOR)

        # VT Rem atualizado = VT Rem orig * fator
        vr = get_column_letter(col_vt_rem_orig)
        fc = get_column_letter(col_fator_ciclo)
        ws.cell(r, col_vt_rem_atu,
                f"=IF({vr}{r}=0,0,ROUND({vr}{r}*{fc}{r},2))").number_format = _MOEDA
        ws.cell(r, col_vt_rem_atu).fill = _fill(_P["auto_c"])
        ws.cell(r, col_vt_rem_atu).font = _font(color=_P["auto"])
        ws.cell(r, col_vt_rem_atu).alignment = _aln()
        ws.cell(r, col_vt_rem_atu).border = _brd()

        # Obs
        _cel(ws, r, col_obs, "", _P["auto_c"], _P["auto"], h="left", size=9)

    # Linha de total
    ws.row_dimensions[206].height = 24
    # Linha de total — sem merge para evitar conflito com celulas de formula
    ws.cell(206, 1, "TOTAL").fill = _fill(_P["ok"])
    ws.cell(206, 1).font = _font(True, "FFFFFF", 9)
    ws.cell(206, 1).alignment = _aln("right")
    ws.cell(206, 1).border = _brd()
    for ci in range(2, col_vt_rem_orig):
        ws.cell(206, ci).fill = _fill(_P["ok"]); ws.cell(206, ci).border = _brd()

    for ci_t, col_l in [
        (col_vt_c0, get_column_letter(col_vt_c0)),
        (col_rem_corte,   get_column_letter(col_rem_corte)),
        (col_vt_rem_orig, get_column_letter(col_vt_rem_orig)),
        (col_vt_rem_atu,  get_column_letter(col_vt_rem_atu)),
    ]:
        t = ws.cell(206, ci_t, f"=ROUND(SUM({col_l}6:{col_l}205),2)")
        t.fill = _fill(_P["ok"]); t.font = _font(True, "FFFFFF")
        t.alignment = _aln(); t.border = _brd()
        t.number_format = _MOEDA if ci_t != col_rem_corte else _QTD

    ws.cell(206, col_obs).fill = _fill(_P["ok"]); ws.cell(206, col_obs).border = _brd()
    ws.cell(206, col_fator_ciclo).fill = _fill(_P["ok"]); ws.cell(206, col_fator_ciclo).border = _brd()


def _construir_aditivos_sup(wb):
    ws = wb.create_sheet("ADITIVOS_SUP")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = _P["fisc"]
    ws.freeze_panes = "A5"

    for col, w in [("A", 16), ("B", 14), ("C", 18), ("D", 14), ("E", 16), ("F", 18), ("G", 18), ("H", 32)]:
        ws.column_dimensions[col].width = w

    _titulo(ws, 1, 1, 8, "cl8us · ADITIVOS_SUP — Aditivos e Supressões")
    _nota(ws, 2, 1, 8,
          "  GCC preenche. Informe SOMENTE aditivos com efeito no Valor Total Atualizado. "
          "O campo Tratamento define se entra ou não no cálculo.",
          _P["gcc_c"], _P["gcc"])
    _nota(ws, 3, 1, 8,
          "  Regra: aditivo não soma automaticamente. Só entra no VTA se GCC indicar "
          '"Computar à parte no VTA". Em caso de dúvida, selecione "Não sei".',
          _P["alrt_c"], _P["alrt"])

    # Cabeçalho
    ws.row_dimensions[4].height = 52
    hdrs = [
        ("[GCC] Data",                  _P["gcc"]),
        ("[GCC] Ciclo",                 _P["gcc"]),
        ("[GCC] Tipo",                  _P["gcc"]),
        ("[GCC] Item",                  _P["gcc"]),
        ("[GCC] Quantidade",            _P["gcc"]),
        ("[GCC] Valor unit. original",  _P["gcc"]),
        ("[AUTO] Valor original",       _P["auto"]),
        ("[TRAT] Tratamento no VTA",    _P["fisc"]),
    ]
    for ci, (txt, bg) in enumerate(hdrs, start=1):
        _hdr(ws, 4, ci, txt, bg)

    _dv(ws, '"Acréscimo,Supressão"', "C5:C54")
    _dv(ws,
        '"Já refletido no financeiro/itens — não somar novamente,Computar à parte no VTA,Reduzir do VTA,Apenas informativo,Não sei"',
        "H5:H54")

    for r in range(5, 55):
        ws.row_dimensions[r].height = 20
        _cel(ws, r, 1, "", _P["gcc_c"], _P["gcc"])
        _cel(ws, r, 2, "", _P["gcc_c"], _P["gcc"])
        _cel(ws, r, 3, "Acréscimo", _P["gcc_c"], _P["gcc"])
        _cel(ws, r, 4, "", _P["gcc_c"], _P["gcc"])
        _cel(ws, r, 5, 0, _P["gcc_c"], _P["gcc"], fmt=_QTD)
        _cel(ws, r, 6, 0, _P["gcc_c"], _P["gcc"], fmt=_MOEDA)
        _cel(ws, r, 7, f"=IF(OR(E{r}=0,F{r}=0),0,ROUND(E{r}*F{r},2))",
             _P["auto_c"], _P["auto"], fmt=_MOEDA)
        _cel(ws, r, 8, "Não sei", _P["fisc_c"], _P["fisc"])

    # Total computável
    ws.row_dimensions[55].height = 24
    ws.merge_cells("A55:F55")
    tc = ws.cell(55, 1, "TOTAL COMPUTÁVEL À PARTE NO VTA")
    tc.fill = _fill(_P["ok"])
    tc.font = _font(True, "FFFFFF", 9)
    tc.alignment = _aln("right")
    tc.border = _brd()
    for ci in range(2, 7):
        ws.cell(55, ci).fill = _fill(_P["ok"])
        ws.cell(55, ci).border = _brd()
    tv = ws.cell(55, 7, '=ROUND(SUMIF(H5:H54,"Computar à parte no VTA",G5:G54),2)')
    tv.fill = _fill(_P["ok"])
    tv.font = _font(True, "FFFFFF", 11)
    tv.alignment = _aln()
    tv.border = _brd()
    tv.number_format = _MOEDA
    ws.cell(55, 8).fill = _fill(_P["ok"])
    ws.cell(55, 8).border = _brd()


# ── Aba PARAMETROS_CICLOS ────────────────────────────────────────

def _construir_parametros_ciclos(wb, dados):
    ws = wb.create_sheet("PARAMETROS_CICLOS")
    ws.sheet_view.showGridLines = False
    ws.sheet_state = "hidden"

    for col, w in [("A", 8), ("B", 16), ("C", 16), ("D", 16), ("E", 14), ("F", 14), ("G", 16), ("H", 18), ("I", 16), ("J", 16), ("K", 20)]:
        ws.column_dimensions[col].width = w

    _titulo(ws, 1, 1, 11, "cl8us · PARAMETROS_CICLOS — leitura automática pelo sistema")
    _nota(ws, 2, 1, 11, "  Aba técnica. Dados pré-preenchidos pela Calculadora. Não alterar.", _P["auto_c"], _P["auto"])

    # Parâmetros gerais
    params = [
        ("Índice contratual",          dados["indice"]),
        ("Data-base original",         dados["data_base"]),
        ("Valor original do contrato", dados["valor_orig"]),
        ("Vigência inicial",           ""),
        ("Vigência final",             ""),
    ]
    for i, (label, val) in enumerate(params):
        r = 3 + i
        ws.row_dimensions[r].height = 20
        _cel(ws, r, 1, label, _P["auto_c"], _P["auto"], h="right", size=9)
        fmt = _MOEDA if "valor" in label.lower() else None
        _cel(ws, r, 2, val, _P["calc_c"], _P["calc"], fmt=fmt)
        for ci in range(3, 12):
            _cel(ws, r, ci, "", _P["auto_c"])

    # Cabeçalho ciclos
    ws.row_dimensions[9].height = 42
    hdrs_c = ["Ciclo", "Data-base", "Início fin.", "Fim fin.", "Percentual",
              "Fator ciclo", "Fator acumulado", "Situação", "É objeto?", "Tem efeito?", "Observação"]
    for ci, txt in enumerate(hdrs_c, start=1):
        _hdr(ws, 9, ci, txt, _P["calc"])

    # C0
    ws.row_dimensions[10].height = 20
    c0 = ["C0", dados["data_base"], "", "", 0.0, 1.0, 1.0, "Base sem reajuste", "Não", "Não", "Ciclo-base"]
    for ci, val in enumerate(c0, start=1):
        fmt = _PERC if ci == 5 else (_FATOR if ci in [6, 7] else None)
        _cel(ws, 10, ci, val, _P["calc_c"], _P["calc"], fmt=fmt)

    # Ciclos calculados
    for i, ciclo in enumerate(dados["ciclos"]):
        r = 11 + i
        ws.row_dimensions[r].height = 20
        nome   = str(ciclo.get("ciclo", f"C{i+1}"))
        db     = str(ciclo.get("data_base_original", ciclo.get("data_base", "")))
        ini    = str(ciclo.get("inicio_financeiro", ciclo.get("data_inicio_financeiro", "")))
        fim    = str(ciclo.get("fim_financeiro", ciclo.get("data_fim_financeiro", "")))
        pct    = float(ciclo.get("percentual_apurado", ciclo.get("percentual", 0)) or 0)
        fat    = float(ciclo.get("fator_ciclo", ciclo.get("fator", 1)) or 1)
        fat_ac = float(ciclo.get("fator_acumulado", 1) or 1)
        sit    = str(ciclo.get("situacao_aplicada", ciclo.get("situacao", "")) or "")
        linha  = [nome, db, ini, fim, pct, fat, fat_ac, sit, "Sim", "Sim", ""]
        for ci, val in enumerate(linha, start=1):
            fmt = _PERC if ci == 5 else (_FATOR if ci in [6, 7] else None)
            _cel(ws, r, ci, val, _P["calc_c"], _P["calc"], fmt=fmt)


# ── Aba CONFERENCIA ──────────────────────────────────────────────

def _construir_conferencia(wb):
    ws = wb.create_sheet("CONFERENCIA")
    ws.sheet_view.showGridLines = False
    ws.sheet_state = "hidden"

    for col, w in [("A", 46), ("B", 32), ("C", 36)]:
        ws.column_dimensions[col].width = w

    _titulo(ws, 1, 1, 3, "cl8us · CONFERENCIA — Diagnóstico automático")
    _nota(ws, 2, 1, 3, "  Aba técnica de leitura automática. Não alterar.", _P["auto_c"], _P["auto"])

    _secao(ws, 4, 1, 3, "Indicadores de base", _P["calc"])
    indicadores = [
        ("Linhas financeiras mensais preenchidas",
         '=COUNTIF(FINANCEIRO_COMP!C5:C220,">0")'),
        ("Total financeiro com efeito Sim",
         '=ROUND(SUMIF(FINANCEIRO_COMP!D5:D220,"Sim",FINANCEIRO_COMP!C5:C220),2)'),
        ("Linhas de itens preenchidas",
         '=COUNTA(ITENS_CICLO!C5:C200)-COUNTBLANK(ITENS_CICLO!C5:C200)'),
        ("Total valor atualizado dos itens",
         '=IFERROR(ITENS_CICLO!I201,0)'),
        ("Aditivos computáveis à parte",
         '=COUNTIF(ADITIVOS_SUP!H5:H54,"Computar à parte no VTA")'),
        ("Valor aditivos computáveis",
         '=IFERROR(ADITIVOS_SUP!G55,0)'),
    ]
    for i, (desc, formula) in enumerate(indicadores):
        r = 5 + i
        ws.row_dimensions[r].height = 20
        _cel(ws, r, 1, desc, _P["auto_c"], _P["auto"], bold=True, h="right", size=9)
        v = ws.cell(r, 2, formula)
        v.fill = _fill(_P["calc_c"])
        v.font = _font(True, _P["calc"])
        v.alignment = _aln("left")
        v.border = _brd()
        ws.cell(r, 3).fill = _fill(_P["auto_c"])
        ws.cell(r, 3).border = _brd()

    # Modo detectado
    _secao(ws, 12, 1, 3, "Modo detectado", _P["ok"])
    ws.row_dimensions[13].height = 28
    ws.merge_cells("A13:B13")
    modo = ws.cell(13, 1,
        '=IF(AND(B5>0,B3>0),"Completo: Financeiro + Itens",'
        'IF(B5>0,"Financeiro histórico",'
        'IF(B3>0,"Itemizado",'
        '"Base insuficiente")))')
    modo.fill = _fill(_P["ok"])
    modo.font = _font(True, "FFFFFF", 11)
    modo.alignment = _aln("left")
    modo.border = _brd()
    ws.cell(13, 3).fill = _fill(_P["ok"])
    ws.cell(13, 3).border = _brd()


# ── Função pública ───────────────────────────────────────────────

def gerar_coleta_mestre_v11():
    """
    Gera a ColetaMestre v11 — Base Analítica.
    Retorna: bytes do XLSX.
    """
    dados = _dados_sessao()

    wb = Workbook()
    wb.remove(wb.active)

    _construir_inicio(wb, dados)
    _construir_financeiro_comp(wb, dados)
    _construir_itens_ciclo(wb, dados)
    _construir_aditivos_sup(wb)
    _construir_parametros_ciclos(wb, dados)
    _construir_conferencia(wb)

    # Garantir INICIO como aba ativa
    try:
        wb.active = wb.index(wb["INICIO"])
    except Exception:
        pass

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
