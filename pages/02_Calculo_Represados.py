import streamlit as st
import pandas as pd
import requests
from datetime import datetime
from zoneinfo import ZoneInfo
from dateutil.relativedelta import relativedelta
from io import BytesIO

st.set_page_config(page_title="Cálculo de Represados", layout="wide")


def get_data_rep(serie, d_ini, d_fim, is_ist):
    try:
        if is_ist:
            df = pd.read_csv('ist.csv', sep=';', decimal=',', encoding='utf-8-sig')
            df.columns = [str(col).strip().lower() for col in df.columns]
            df['data'] = pd.to_datetime(df['data'], dayfirst=True, errors='coerce').dt.normalize()
            df = df.dropna(subset=['data'])

            r_ini = pd.to_datetime((d_ini - relativedelta(months=1)).replace(day=1)).normalize()
            r_fim = pd.to_datetime(d_fim.replace(day=1)).normalize()

            df_detalhado = df[(df['data'] >= r_ini) & (df['data'] <= r_fim)].copy()
            v_ini_rows = df[df['data'] == r_ini]
            v_fim_rows = df[df['data'] == r_fim]

            if v_ini_rows.empty or v_fim_rows.empty:
                return None

            v_ini = v_ini_rows['indice'].values[0]
            v_fim = v_fim_rows['indice'].values[0]

            return {
                'var': (v_fim / v_ini) - 1,
                'i_ini': v_ini,
                'i_fim': v_fim,
                'p_ini': r_ini,
                'p_fim': r_fim,
                'metodo': "Divisão de Número-Índice (IST)",
                'dados': df_detalhado[['data', 'indice']]
            }
        else:
            url = (
                f"https://api.bcb.gov.br/dados/serie/bcdata.sgs.{serie}/dados"
                f"?formato=json&dataInicial={d_ini.strftime('%d/%m/%Y')}"
                f"&dataFinal={d_fim.strftime('%d/%m/%Y')}"
            )
            response = requests.get(url, timeout=10)
            df_t = pd.DataFrame(response.json())

            if df_t.empty:
                return None

            df_t['v'] = df_t['valor'].astype(float) / 100
            df_t['data'] = pd.to_datetime(df_t['data'], dayfirst=True)

            return {
                'var': (1 + df_t['v']).prod() - 1,
                'metodo': "Produtório de taxas mensais (SGS/BCB)",
                'p_ini': d_ini,
                'p_fim': d_fim,
                'dados': df_t[['data', 'valor']]
            }
    except Exception:
        return None


def _formatar_data(valor):
    try:
        return pd.to_datetime(valor).strftime('%d/%m/%Y')
    except Exception:
        return ""


def _competencias_mensais(data_inicio, data_fim):
    inicio = pd.to_datetime(data_inicio).replace(day=1)
    fim = pd.to_datetime(data_fim).replace(day=1)
    if fim < inicio:
        return []
    return [d.strftime('%m/%Y') for d in pd.date_range(inicio, fim, freq='MS')]


def _ajustar_larguras(writer, nomes_abas):
    from openpyxl.utils import get_column_letter

    for nome_aba in nomes_abas:
        ws = writer.book[nome_aba]
        # ITENS_REMANESCENTES usa duas linhas de cabeçalho: linha 1 para datas de referência e linha 2 para títulos.
        ws.freeze_panes = "A3" if nome_aba == 'ITENS_REMANESCENTES' else "A2"
        for column_cells in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                try:
                    max_length = max(max_length, len(str(cell.value)) if cell.value is not None else 0)
                except Exception:
                    pass
            ws.column_dimensions[column_letter].width = min(max(max_length + 2, 14), 45)


def _aplicar_estilos_coleta(writer, ciclos):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill('solid', fgColor='1F4E78')
    header_font = Font(color='FFFFFF', bold=True)
    input_fill = PatternFill('solid', fgColor='FFF2CC')
    date_fill = PatternFill('solid', fgColor='D9EAD3')
    light_fill = PatternFill('solid', fgColor='EAF2F8')
    thin = Side(style='thin', color='D9E2F3')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    money_fmt = 'R$ #,##0.00'
    number_fmt = '#,##0.0000'

    for ws in writer.book.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical='center')

        header_row = 2 if ws.title == 'ITENS_REMANESCENTES' else 1
        for cell in ws[header_row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    if 'FINANCEIRO_MENSAL' in writer.book.sheetnames:
        ws = writer.book['FINANCEIRO_MENSAL']
        # Coluna C é a coluna de preenchimento pelo fiscal.
        for row in range(1, ws.max_row + 1):
            ws.cell(row=row, column=3).fill = input_fill
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='right' if row > 1 else 'center', vertical='center')
            if row > 1:
                ws.cell(row=row, column=3).number_format = money_fmt
        ws['C1'].value = 'Valor pago/faturado (preencher)'
        ws['C1'].font = header_font
        ws['C1'].fill = header_fill

    if 'ITENS_REMANESCENTES' in writer.book.sheetnames:
        ws = writer.book['ITENS_REMANESCENTES']
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
        ws.cell(row=1, column=1).value = 'Dados do item e valor original'
        ws.cell(row=1, column=1).fill = light_fill
        ws.cell(row=1, column=1).font = Font(bold=True)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')

        # Colunas: A Item | B Quantidade contratada | C Valor unitário original | D Valor total | E... remanescentes
        for row in range(3, ws.max_row + 1):
            ws.cell(row=row, column=3).number_format = money_fmt
            ws.cell(row=row, column=4).number_format = money_fmt
            ws.cell(row=row, column=4).value = f'=IF(OR(B{row}="",C{row}=""),"",B{row}*C{row})'

        # Destacar colunas de remanescente para preenchimento pelo fiscal e colocar a data de referência na linha 1.
        first_rem_col = 5
        for idx, ciclo in enumerate(ciclos):
            col = first_rem_col + idx
            letra = get_column_letter(col)
            ws.cell(row=1, column=col).value = ciclo.get('data_base', '')
            ws.cell(row=1, column=col).fill = date_fill
            ws.cell(row=1, column=col).font = Font(bold=True)
            ws.cell(row=1, column=col).alignment = Alignment(horizontal='center', vertical='center')
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col).fill = input_fill if row > 2 else header_fill
                if row == 2:
                    ws.cell(row=row, column=col).font = header_font
                    ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.column_dimensions[letra].width = 24

    if 'CICLOS' in writer.book.sheetnames:
        ws = writer.book['CICLOS']
        for row in range(2, ws.max_row + 1):
            # Fator e Fator acumulado.
            if ws.max_column >= 8:
                ws.cell(row=row, column=8).number_format = number_fmt
            if ws.max_column >= 9:
                ws.cell(row=row, column=9).number_format = number_fmt


def gerar_arquivo_coleta_excel(dados_admissibilidade):
    ciclos = dados_admissibilidade.get('ciclos', [])

    parametros = pd.DataFrame([
        {'Campo': 'Origem da análise', 'Valor': dados_admissibilidade.get('origem', '')},
        {'Campo': 'Índice utilizado', 'Valor': dados_admissibilidade.get('indice', '')},
        {'Campo': 'Data-base original', 'Valor': dados_admissibilidade.get('data_base_original', '')},
        {'Campo': 'Quantidade de ciclos', 'Valor': len(ciclos)},
        {'Campo': 'Fator acumulado final', 'Valor': round(float(dados_admissibilidade.get('fator_acumulado', 1.0)), 4)},
        {'Campo': 'Variação acumulada final', 'Valor': dados_admissibilidade.get('variacao_acumulada_formatada', '')},
        {'Campo': 'Data de geração do arquivo', 'Valor': datetime.now(ZoneInfo('America/Sao_Paulo')).strftime('%d/%m/%Y %H:%M')},
    ])

    df_ciclos = pd.DataFrame([
        {
            'Ciclo': c.get('ciclo', ''),
            'Data-base': c.get('data_base', ''),
            'Intervalo do índice': c.get('intervalo_indice', ''),
            'Janela de admissibilidade': c.get('janela_admissibilidade', ''),
            'Data do pedido': c.get('data_pedido', ''),
            'Situação': c.get('situacao', ''),
            'Variação': c.get('variacao_formatada', ''),
            'Fator': round(float(c.get('fator', 1.0)), 4),
            'Fator acumulado': round(float(c.get('fator_acumulado', 1.0)), 4),
        }
        for c in ciclos
    ])

    # Aba financeira simplificada: somente o que o fiscal precisa preencher.
    linhas_financeiro = [
        {
            'Ciclo': 'C0',
            'Competência': 'TOTAL C0',
            'Valor pago/faturado': None,
        }
    ]
    for c in ciclos:
        for competencia in _competencias_mensais(c.get('periodo_inicio'), c.get('periodo_fim')):
            linhas_financeiro.append({
                'Ciclo': c.get('ciclo', ''),
                'Competência': competencia,
                'Valor pago/faturado': None,
            })
    df_financeiro = pd.DataFrame(linhas_financeiro, columns=[
        'Ciclo', 'Competência', 'Valor pago/faturado'
    ])

    # Aba de remanescentes simplificada. A linha 1 recebe a data-base de referência de cada ciclo.
    colunas_remanescentes = [f"Remanescente início {c.get('ciclo', '')}" for c in ciclos]
    colunas_itens = [
        'Item', 'Quantidade contratada', 'Valor unitário original', 'Valor total',
        *colunas_remanescentes
    ]
    linhas_itens = [{col: None for col in colunas_itens} for _ in range(30)]
    df_itens = pd.DataFrame(linhas_itens, columns=colunas_itens)

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        parametros.to_excel(writer, sheet_name='PARAMETROS_REAJUSTE', index=False)
        df_ciclos.to_excel(writer, sheet_name='CICLOS', index=False)
        df_financeiro.to_excel(writer, sheet_name='FINANCEIRO_MENSAL', index=False)
        df_itens.to_excel(writer, sheet_name='ITENS_REMANESCENTES', index=False, startrow=1)
        _aplicar_estilos_coleta(writer, ciclos)
        _ajustar_larguras(writer, [
            'PARAMETROS_REAJUSTE', 'CICLOS', 'FINANCEIRO_MENSAL', 'ITENS_REMANESCENTES'
        ])
    output.seek(0)
    return output


st.image("https://www.telebras.com.br/wp-content/uploads/2019/06/Telebras_Logo_AzulProfundo.png", width=250)
st.title("Reajustes Múltiplos")

with st.sidebar:
    dt_base_original = st.date_input("Data-Base Original:", value=datetime(2022, 10, 10), format="DD/MM/YYYY")
    qtd_ciclos = st.number_input("Ciclos:", min_value=1, max_value=5, value=2)
    idx_sel = st.selectbox("Índice:", ["IST (Série Local)", "IPCA (433)", "IGP-M (189)"])

# data_atual representa a Data-Base do ciclo em análise.
# Para C1, é a Data-Base Original.
# Para ciclos seguintes, será atualizada conforme a regra contratual:
# - ciclo concedido/tempestivo: próxima Data-Base = data do pedido anterior;
# - pedido no mesmo mês, antes do aniversário: próxima Data-Base = aniversário contratual;
# - pedido adiantado/precluso: próxima Data-Base = aniversário contratual.
data_atual = dt_base_original
fator_acum = 1.0
historico = []
historico_coleta = []

for i in range(1, int(qtd_ciclos) + 1):
    st.markdown(f"### Ciclo {i}")

    d_fim = data_atual + relativedelta(months=11)
    d_aniv = data_atual + relativedelta(years=1)
    d_lim = d_aniv + relativedelta(days=90)

    col_a, col_b = st.columns(2)
    with col_a:
        st.write(f"**Data-Base do Ciclo:** {data_atual.strftime('%d/%m/%Y')}")
    with col_b:
        dt_ped = st.date_input(f"Data do Pedido C{i}:", value=d_aniv, key=f"p{i}", format="DD/MM/YYYY")

    # Lógica de Admissibilidade
    if dt_ped < d_aniv:
        if dt_ped.year == d_aniv.year and dt_ped.month == d_aniv.month:
            sit_emoji = "🟡 ADMISSÍVEL - RESSALVA"
            situacao_limpa = "ADMISSÍVEL - RESSALVA"
        else:
            sit_emoji = "⚠️ ADIANTADO"
            situacao_limpa = "ADIANTADO"
    elif dt_ped <= d_lim:
        sit_emoji = "✅ TEMPESTIVO"
        situacao_limpa = "TEMPESTIVO"
    else:
        sit_emoji = "❌ PRECLUSO"
        situacao_limpa = "PRECLUSO"

    # Regra de ancoragem do próximo ciclo.
    if situacao_limpa == "TEMPESTIVO":
        data_base_proximo_ciclo = dt_ped
    else:
        data_base_proximo_ciclo = d_aniv

    res_c = get_data_rep("433" if "IPCA" in idx_sel else "189", data_atual, d_fim, "IST" in idx_sel)

    # Intervalo exibido independentemente de haver dados disponíveis para o índice.
    if res_c:
        periodo_inicio = res_c['p_ini']
        periodo_fim = res_c['p_fim']
        janela_ciclo = f"{res_c['p_ini'].strftime('%m/%Y')} a {res_c['p_fim'].strftime('%m/%Y')}"
    else:
        if "IST" in idx_sel:
            periodo_inicio = (data_atual - relativedelta(months=1)).replace(day=1)
            periodo_fim = d_fim.replace(day=1)
        else:
            periodo_inicio = data_atual
            periodo_fim = d_fim
        janela_ciclo = f"{periodo_inicio.strftime('%m/%Y')} a {periodo_fim.strftime('%m/%Y')}"

    janela_adm = f"{d_aniv.strftime('%d/%m/%Y')} a {d_lim.strftime('%d/%m/%Y')}"

    st.markdown(f"""
    **Dados do Ciclo {i}:**
    - Intervalo do C{i}: {janela_ciclo}
    - Janela de Admissibilidade: {janela_adm}
    - Situação: {sit_emoji}
    """)

    v_fmt = "Indisponível"
    v_acum_parcial = f"{(fator_acum - 1) * 100:,.2f}%".replace('.', ',')
    fator_ciclo = 1.0
    ciclo_calculado = False

    if res_c:
        fator_ciclo = 1 + res_c['var']
        fator_acum *= fator_ciclo
        v_fmt = f"{res_c['var'] * 100:,.2f}%".replace('.', ',')
        v_acum_parcial = f"{(fator_acum - 1) * 100:,.2f}%".replace('.', ',')
        ciclo_calculado = True

        st.markdown(f"- Variação do Ciclo: **{v_fmt}**")

        with st.expander(f"🔍 Memória de Cálculo Detalhada - Ciclo {i}"):
            st.write(f"**Metodologia:** {res_c['metodo']}")
            st.dataframe(res_c['dados'], use_container_width=True)

        historico.append({
            "Ciclo": i,
            "Variação": v_fmt,
            "Acumulada": v_acum_parcial,
            "Situação": sit_emoji,
            "Pedido": dt_ped.strftime('%d/%m/%Y'),
            "Janela": janela_ciclo,
            "JanelaAdm": janela_adm
        })
    else:
        st.warning(
            "Não há dados disponíveis para o índice selecionado no intervalo de apuração deste ciclo. "
            "O ciclo foi exibido para controle, mas não foi incluído no cálculo acumulado."
        )

    historico_coleta.append({
        'ciclo': f'C{i}',
        'data_base': data_atual.strftime('%d/%m/%Y'),
        'intervalo_indice': janela_ciclo,
        'janela_admissibilidade': janela_adm,
        'data_pedido': dt_ped.strftime('%d/%m/%Y'),
        'situacao': sit_emoji,
        'variacao': float(res_c['var']) if res_c else 0.0,
        'variacao_formatada': v_fmt,
        'fator': float(fator_ciclo),
        'fator_acumulado': float(fator_acum),
        'ciclo_calculado': ciclo_calculado,
        'periodo_inicio': _formatar_data(periodo_inicio),
        'periodo_fim': _formatar_data(periodo_fim),
    })

    # A progressão do ciclo deve ocorrer sempre, mesmo quando não houver índice disponível.
    data_atual = data_base_proximo_ciclo

if historico:
    st.divider()
    res_final = f"{(fator_acum - 1) * 100:,.2f}%".replace('.', ',')
    st.metric("Variação Acumulada Final", res_final)

    st.subheader("Relatório de Apuração")
    corpo_relatorio = ""
    for h in historico:
        corpo_relatorio += f"""
        **C{h['Ciclo']}:** Pedido em {h['Pedido']}. Intervalo do C{h['Ciclo']}: {h['Janela']}.  
        Janela de Admissibilidade: {h['JanelaAdm']}.  
        Resultado: {h['Situação']}. Variação: {h['Variação']}.  
        Índice {idx_sel}.
        \n\n"""
    st.info(corpo_relatorio)

if historico_coleta:
    variacao_acumulada = fator_acum - 1
    st.session_state['dados_admissibilidade'] = {
        'origem': 'Reajustes Múltiplos',
        'tipo': 'Múltiplo',
        'indice': idx_sel,
        'data_base_original': dt_base_original.strftime('%d/%m/%Y'),
        'fator': float(fator_acum),
        'fator_acumulado': float(fator_acum),
        'variacao_acumulada': float(variacao_acumulada),
        'variacao_acumulada_formatada': f"{variacao_acumulada * 100:,.2f}%".replace('.', ','),
        'ciclos': historico_coleta,
    }

    st.download_button(
        label="📥 Gerar Arquivo de Coleta",
        data=gerar_arquivo_coleta_excel(st.session_state['dados_admissibilidade']),
        file_name="Coleta_Reajustes_Multiplos.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=False,
    )
