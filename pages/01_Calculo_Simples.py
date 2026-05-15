import streamlit as st
import json
import re
import pandas as pd
import requests
from datetime import datetime
from zoneinfo import ZoneInfo
from dateutil.relativedelta import relativedelta
from io import BytesIO

if not st.session_state.get("_calculadora_reajustes_embedded", False):
    st.set_page_config(page_title="Análises de Reajustes - Reajuste Simples", layout="wide")


from _ui_utils import render_indice_contrato_selectbox, render_marca_topo
from _indice_utils import calcular_ist_numero_indice, coletar_sgs_produtorio
from _reajuste_utils import _competencias_mensais, _formatar_data, _formatar_moeda_br, _formatar_moeda_br_md, _parse_moeda_br

def get_index_data(serie_codigo, data_inicio, data_fim):
    try:
        return coletar_sgs_produtorio(serie_codigo, data_inicio, data_fim, timeout=15)
    except Exception:
        return None


def get_ist_local(data_inicio, data_fim):
    try:
        return calcular_ist_numero_indice(data_inicio)
    except Exception:
        return None


def _formatar_mes_ano(valor):
    try:
        return pd.to_datetime(valor).strftime('%m/%Y')
    except Exception:
        return ""


def _data_para_date_segura(valor):
    """Converte date/datetime/Timestamp em date para validações de corte temporal."""
    try:
        if isinstance(valor, datetime):
            return valor.date()
    except Exception:
        pass
    try:
        if hasattr(valor, "date") and not isinstance(valor, str):
            return valor.date()
    except Exception:
        pass
    try:
        dt = pd.to_datetime(valor, dayfirst=True, errors="coerce")
        if pd.notna(dt):
            return dt.date()
    except Exception:
        pass
    return None


def _competencias_esperadas_indice(data_inicio, data_fim):
    """Lista competências mensais esperadas entre data_inicio e data_fim, inclusive."""
    try:
        inicio = pd.Timestamp(data_inicio).to_period("M")
        fim = pd.Timestamp(data_fim).to_period("M")
        if fim < inicio:
            return []
        return [p.strftime("%m/%Y") for p in pd.period_range(inicio, fim, freq="M")]
    except Exception:
        return []


def _competencia_de_valor(valor):
    """Extrai competência mm/aaaa de datas ou textos comuns."""
    if valor is None:
        return None
    try:
        if pd.isna(valor):
            return None
    except Exception:
        pass
    if isinstance(valor, (datetime, pd.Timestamp)):
        return pd.Timestamp(valor).to_period("M").strftime("%m/%Y")
    texto = str(valor).strip()
    if not texto:
        return None

    # dd/mm/aaaa
    m = re.search(r"\b(\d{1,2})/(\d{1,2})/(\d{4})\b", texto)
    if m:
        dia, mes, ano = m.groups()
        try:
            return pd.Timestamp(int(ano), int(mes), 1).to_period("M").strftime("%m/%Y")
        except Exception:
            return None

    # mm/aaaa
    m = re.search(r"\b(\d{1,2})/(\d{4})\b", texto)
    if m:
        mes, ano = m.groups()
        try:
            return pd.Timestamp(int(ano), int(mes), 1).to_period("M").strftime("%m/%Y")
        except Exception:
            return None

    # aaaa-mm-dd ou aaaa-mm
    m = re.search(r"\b(\d{4})-(\d{1,2})(?:-\d{1,2})?\b", texto)
    if m:
        ano, mes = m.groups()
        try:
            return pd.Timestamp(int(ano), int(mes), 1).to_period("M").strftime("%m/%Y")
        except Exception:
            return None

    try:
        dt = pd.to_datetime(texto, dayfirst=True, errors="coerce")
        if pd.notna(dt):
            return pd.Timestamp(dt).to_period("M").strftime("%m/%Y")
    except Exception:
        pass
    return None


def _competencias_encontradas_no_resultado(res):
    """Extrai competências existentes no retorno do índice, especialmente para IPCA/IGP-M."""
    comps = set()
    if not res:
        return comps

    dados = res.get("dados") if isinstance(res, dict) else None

    def adicionar_de_obj(obj):
        if isinstance(obj, dict):
            for v in obj.values():
                adicionar_de_obj(v)
        elif isinstance(obj, (list, tuple, set)):
            for v in obj:
                adicionar_de_obj(v)
        else:
            comp = _competencia_de_valor(obj)
            if comp:
                comps.add(comp)

    if isinstance(dados, pd.DataFrame):
        # Dá preferência a colunas com nomes de data/competência.
        colunas_prioritarias = [
            c for c in dados.columns
            if str(c).strip().lower() in ["data", "dt", "competencia", "competência", "mes", "mês", "periodo", "período"]
        ]
        colunas = colunas_prioritarias or list(dados.columns)
        for col in colunas:
            for valor in dados[col].dropna().tolist():
                comp = _competencia_de_valor(valor)
                if comp:
                    comps.add(comp)
    elif dados is not None:
        adicionar_de_obj(dados)

    return comps


def _validar_indice_disponivel(res, data_inicio, data_fim, indice_nome):
    """Valida se o índice possui base suficiente para processar o ciclo.

    Para IPCA/IGP-M, exige todas as competências mensais do intervalo.
    Para IST, mantém a validação por retorno existente, pois o cálculo usa número-índice.
    """
    esperadas = _competencias_esperadas_indice(data_inicio, data_fim)

    if not res:
        return {
            "ok": False,
            "motivo": "sem_retorno",
            "esperadas": esperadas,
            "encontradas": [],
            "faltantes": esperadas,
        }

    if "IST" in str(indice_nome).upper():
        return {
            "ok": True,
            "motivo": "",
            "esperadas": esperadas,
            "encontradas": esperadas,
            "faltantes": [],
        }

    encontradas = sorted(_competencias_encontradas_no_resultado(res))
    faltantes = [c for c in esperadas if c not in set(encontradas)]

    return {
        "ok": len(faltantes) == 0,
        "motivo": "competencias_ausentes" if faltantes else "",
        "esperadas": esperadas,
        "encontradas": encontradas,
        "faltantes": faltantes,
    }


def _render_alerta_indice_ausente(validacao, ciclo_label="C1"):
    faltantes = validacao.get("faltantes", []) or []
    esperadas = validacao.get("esperadas", []) or []
    st.error(
        f"Processamento inviável: falta pelo menos um mês do intervalo de apuração do índice no {ciclo_label}."
    )
    st.warning(
        "Não foi possível concluir a apuração porque há competência ausente no intervalo do índice. "
        "Atualize a base de índices ou confira o período de apuração antes de prosseguir."
    )
    dados = []
    if faltantes:
        dados.append({"Item": "Competências faltantes", "Competências": ", ".join(faltantes)})
    if esperadas:
        dados.append({"Item": "Intervalo esperado", "Competências": ", ".join(esperadas)})
    if dados:
        st.dataframe(pd.DataFrame(dados), use_container_width=True, hide_index=True)


def _render_card_contexto_contrato():
    """Exibe cabeçalho executivo do bloco de contexto do contrato."""
    st.markdown(
        """
        <div style="background:#F3F6FA;border:1px solid #D9E2EC;border-left:5px solid #1F4E78;border-radius:12px;padding:14px 16px;margin:10px 0 8px 0;">
            <div style="color:#123B63;font-weight:800;font-size:1.02rem;margin-bottom:4px;">Contexto do Contrato</div>
            <div style="color:#475569;font-size:0.92rem;line-height:1.45;">
                Use este bloco quando o contrato já possuir reajustes, repactuações, aditivos ou supressões formalizados antes desta análise.
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )



def _ciclo_para_numero(valor):
    texto = str(valor or "").strip().upper()
    if texto.startswith("C"):
        try:
            return int(texto[1:].split()[0])
        except Exception:
            return 0
    return 0

def _primeiro_ciclo_analise(contexto):
    return max(_ciclo_para_numero((contexto or {}).get('ultimo_ciclo_concedido', '')) + 1, 1)

def _data_contexto_para_datetime(valor):
    if not valor:
        return None
    try:
        dt = pd.to_datetime(valor, dayfirst=True, errors='coerce')
        if pd.notna(dt):
            return dt.to_pydatetime()
    except Exception:
        pass
    return None

def _data_base_inicial_pelo_contexto(contexto, fallback):
    contexto = contexto or {}
    ultimo = _ciclo_para_numero(contexto.get('ultimo_ciclo_concedido', ''))
    if ultimo > 0:
        # Regra adotada: no fluxo ordinário, a data do pedido do último ciclo concedido
        # corresponde ao início dos efeitos financeiros desse reajuste.
        # Mantém-se fallback legado para arquivos antigos que ainda tenham o campo de efeitos.
        data_pedido_ultimo = _data_contexto_para_datetime(contexto.get('data_pedido_ultimo_ciclo'))
        if data_pedido_ultimo:
            return data_pedido_ultimo
        data_efeitos_legado = _data_contexto_para_datetime(contexto.get('data_efeitos_ultimo_reajuste'))
        if data_efeitos_legado:
            return data_efeitos_legado
    return fallback


def _evento_historico_principal_contexto(contexto):
    """Monta registro automático do último ciclo formalizado para exportação.

    Esse registro evita que o ciclo histórico pareça ter sumido do XLS quando
    a aba CICLOS contém apenas o objeto da análise atual.
    """
    contexto = contexto or {}
    ciclo = str(contexto.get('ultimo_ciclo_concedido', '') or '').strip()
    if not ciclo or ciclo.upper().startswith('C0') or ciclo.lower() in ['nenhum', 'c0 / nenhum']:
        return None
    data_base = str(contexto.get('data_base_ultimo_ciclo', '') or '').strip()
    data_pedido = str(contexto.get('data_pedido_ultimo_ciclo', '') or '').strip()
    valor_formalizado = contexto.get('valor_formalizado_anterior', '')
    try:
        valor_formatado = _formatar_moeda_br(float(valor_formalizado or 0.0)) if float(valor_formalizado or 0.0) > 0 else ''
    except Exception:
        valor_formatado = str(contexto.get('valor_formalizado_anterior_texto', '') or '').strip()
    percentual = contexto.get('percentual_ja_aplicado_pct', 0.0)
    try:
        percentual_txt = f"{float(percentual):.2f}%".replace('.', ',')
    except Exception:
        percentual_txt = ''
    ref = str(contexto.get('referencia_documental_historico', '') or '').strip()
    obs_base = str(contexto.get('observacao_historico', '') or '').strip()
    obs_partes = []
    if data_base:
        obs_partes.append(f"Data-base: {data_base}")
    if data_pedido:
        obs_partes.append(f"Data do pedido/efeitos financeiros: {data_pedido}")
    if percentual_txt and percentual_txt != '0,00%':
        obs_partes.append(f"Percentual aplicado: {percentual_txt}")
    if ref:
        obs_partes.append(f"Referência: {ref}")
    if obs_base:
        obs_partes.append(obs_base)
    obs_partes.append("Registro gerado automaticamente pelo Contexto do Contrato.")
    return {
        'Tipo de evento': 'Reajuste',
        'Ciclo': ciclo,
        'Data': data_pedido or data_base,
        'Valor formalizado/impacto': valor_formatado,
        'Incorporado ao valor formalizado?': 'Sim',
        'Observação': '; '.join(obs_partes),
    }


def _eventos_historicos_para_exportacao(contexto):
    """Combina o evento principal do contexto com eventos adicionais informados manualmente."""
    contexto = contexto or {}
    eventos = []
    auto = _evento_historico_principal_contexto(contexto)
    if auto:
        eventos.append(auto)
    for ev in contexto.get('eventos_historicos_anteriores') or []:
        if not isinstance(ev, dict):
            continue
        # Evita duplicação do registro automático em caso de sessões antigas.
        obs = str(ev.get('Observação', '') or '')
        if 'Registro gerado automaticamente pelo Contexto do Contrato' in obs:
            continue
        possui_conteudo = any(str(ev.get(k, '') or '').strip() for k in ['Tipo de evento', 'Data', 'Valor formalizado/impacto', 'Observação'])
        if possui_conteudo:
            eventos.append(ev)
    return eventos

def _render_contexto_contratual_anterior():
    """Coleta contexto do contrato para uso pelo Valor Global.

    Este bloco é opcional. Quando preenchido, registra memória formal anterior
    para relatórios e governança, sem alterar automaticamente o Valor Total Atualizado.
    """
    contexto_salvo = st.session_state.get('contexto_contratual_anterior', {}) or {}
    _render_card_contexto_contrato()
    with st.expander('Preencher/editar Contexto do Contrato', expanded=False):
        st.caption(
            'Preencha apenas quando o contrato já possuir eventos formalizados antes desta análise. '
            'Este contexto é memória processual e de governança; não altera automaticamente o Valor Total Atualizado, que permanece calculado por execução atualizada + saldo remanescente atualizado.'
        )
        col_ctx1, col_ctx2 = st.columns(2)
        with col_ctx1:
            valor_original_txt = st.text_input(
                'Valor original do contrato',
                value=contexto_salvo.get('valor_original_contrato_texto', ''),
                placeholder='Ex.: 20.000.000,00',
                key='ctx_valor_original_contrato',
            )
            opcoes_ciclo_concedido = ['C0 / Nenhum', 'C1', 'C2', 'C3', 'C4', 'Outro']
            ciclo_salvo = str(contexto_salvo.get('ultimo_ciclo_concedido', '') or '').strip()
            if ciclo_salvo in opcoes_ciclo_concedido:
                indice_ciclo_salvo = opcoes_ciclo_concedido.index(ciclo_salvo)
            elif ciclo_salvo == '' or ciclo_salvo.upper() == 'C0' or ciclo_salvo == 'Nenhum / C0':
                indice_ciclo_salvo = 0
            elif ciclo_salvo.upper() in opcoes_ciclo_concedido:
                indice_ciclo_salvo = opcoes_ciclo_concedido.index(ciclo_salvo.upper())
            else:
                indice_ciclo_salvo = len(opcoes_ciclo_concedido) - 1
            ultimo_ciclo = st.selectbox(
                'Último ciclo já concedido/formalizado',
                options=opcoes_ciclo_concedido,
                index=indice_ciclo_salvo,
                key='ctx_ultimo_ciclo_concedido',
            )
            data_base_ultimo = st.text_input(
                'Data-base do último ciclo concedido/formalizado',
                value=contexto_salvo.get('data_base_ultimo_ciclo', ''),
                placeholder='Ex.: 23/09/2023',
                key='ctx_data_base_ultimo_ciclo',
            )
            data_pedido_ultimo = st.text_input(
                'Data do pedido do último ciclo concedido/formalizado',
                value=contexto_salvo.get('data_pedido_ultimo_ciclo', ''),
                placeholder='Ex.: 23/09/2024',
                key='ctx_data_pedido_ultimo_ciclo',
            )
        with col_ctx2:
            modo_valor_formalizado = st.radio(
                'Como deseja informar o valor formalizado antes desta análise?',
                options=['Informar valor diretamente', 'Calcular pelo valor original + percentual já aplicado'],
                index=1 if contexto_salvo.get('modo_valor_formalizado') == 'Calcular pelo valor original + percentual já aplicado' else 0,
                key='ctx_modo_valor_formalizado',
            )
            percentual_ja_aplicado_pct = st.number_input(
                'Percentual já aplicado antes desta análise (%)',
                min_value=0.0,
                max_value=1000.0,
                value=float(contexto_salvo.get('percentual_ja_aplicado_pct', 0.0) or 0.0),
                step=0.01,
                format='%.2f',
                key='ctx_percentual_ja_aplicado_pct',
                disabled=(modo_valor_formalizado == 'Informar valor diretamente'),
            )
            valor_original_previo = _parse_moeda_br(valor_original_txt)
            valor_calculado_formalizado = round(valor_original_previo * (1 + percentual_ja_aplicado_pct / 100), 2) if valor_original_previo > 0 else 0.0

            if modo_valor_formalizado == 'Calcular pelo valor original + percentual já aplicado':
                fator_aplicado_previo = 1 + percentual_ja_aplicado_pct / 100
                valor_calculado_txt = _formatar_moeda_br(valor_calculado_formalizado)
                fator_txt = f"{fator_aplicado_previo:.6f}".replace('.', ',')
                st.info(
                    f"Memória do valor formalizado anterior: {_formatar_moeda_br(valor_original_previo)} × {fator_txt} = {valor_calculado_txt}. "
                    "O campo abaixo permanece editável para ajuste manual, se houver aditivo, supressão, repactuação ou arredondamento formal."
                )

                chave_calculo_formalizado = f"{valor_original_previo:.2f}|{percentual_ja_aplicado_pct:.6f}|auto"
                if st.session_state.get('ctx_valor_formalizado_auto_hash') != chave_calculo_formalizado:
                    st.session_state['ctx_valor_formalizado_anterior'] = valor_calculado_txt
                    st.session_state['ctx_valor_formalizado_auto_hash'] = chave_calculo_formalizado
                valor_formalizado_default = st.session_state.get('ctx_valor_formalizado_anterior', valor_calculado_txt)
            else:
                valor_formalizado_default = contexto_salvo.get('valor_formalizado_anterior_texto', '')

            valor_formalizado_txt = st.text_input(
                'Valor contratual formalizado antes desta análise',
                value=valor_formalizado_default,
                placeholder='Ex.: 22.800.000,00',
                key='ctx_valor_formalizado_anterior',
                help='Valor editável. Ajuste manualmente se houver aditivos, supressões, repactuações ou arredondamento formal.',
            )
            referencia_documental_historico = st.text_input(
                'Referência documental do histórico anterior',
                value=contexto_salvo.get('referencia_documental_historico', ''),
                placeholder='Ex.: TLB-TDA-2025/00001, Parecer, Despacho ou Ata',
                key='ctx_referencia_documental_historico',
            )
            observacao = st.text_area(
                'Observação sobre o histórico anterior',
                value=contexto_salvo.get('observacao_historico', ''),
                placeholder='Ex.: C1 e C2 já concedidos; valor inclui aditivo anterior formalizado.',
                key='ctx_observacao_historico',
                height=82,
            )

        eventos_limpos = []
        with st.expander("Eventos históricos adicionais (opcional)", expanded=False):
            st.caption(
                "Use apenas para registrar outros eventos formalizados anteriores que não estejam cobertos pelo bloco principal acima, "
                "como aditivo, supressão, repactuação, apostila anterior ou acordo negocial. "
                "O último ciclo concedido já é exportado automaticamente pelo Contexto do Contrato."
            )
            eventos_salvos = contexto_salvo.get('eventos_historicos_anteriores') or []
            # Não reapresenta na tabela registros automáticos gerados em versões anteriores.
            eventos_salvos = [
                ev for ev in eventos_salvos
                if not (isinstance(ev, dict) and 'Registro gerado automaticamente pelo Contexto do Contrato' in str(ev.get('Observação', '') or ''))
            ]
            if not eventos_salvos:
                eventos_salvos = [
                    {
                        'Tipo de evento': '',
                        'Ciclo': 'C0 / Nenhum',
                        'Data': '',
                        'Valor formalizado/impacto': '',
                        'Incorporado ao valor formalizado?': 'Sim',
                        'Observação': '',
                    }
                ]
            df_eventos_ctx = pd.DataFrame(eventos_salvos)
            if 'Valor formalizado/impacto' not in df_eventos_ctx.columns:
                if 'Valor atualizado/formalizado' in df_eventos_ctx.columns:
                    df_eventos_ctx['Valor formalizado/impacto'] = df_eventos_ctx['Valor atualizado/formalizado']
                elif 'Valor original' in df_eventos_ctx.columns:
                    df_eventos_ctx['Valor formalizado/impacto'] = df_eventos_ctx['Valor original']
                else:
                    df_eventos_ctx['Valor formalizado/impacto'] = ''
            colunas_eventos = [
                'Tipo de evento', 'Ciclo', 'Data', 'Valor formalizado/impacto',
                'Incorporado ao valor formalizado?', 'Observação'
            ]
            for col_evento in colunas_eventos:
                if col_evento not in df_eventos_ctx.columns:
                    df_eventos_ctx[col_evento] = ''
            df_eventos_ctx = df_eventos_ctx[colunas_eventos]
            eventos_editados = st.data_editor(
                df_eventos_ctx,
                hide_index=True,
                use_container_width=True,
                num_rows='dynamic',
                key='ctx_eventos_historicos_anteriores',
                column_config={
                    'Tipo de evento': st.column_config.SelectboxColumn(
                        'Tipo de evento',
                        options=['', 'Reajuste', 'Repactuação', 'Aditivo', 'Supressão', 'Apostila anterior', 'Acordo negocial', 'Outro'],
                    ),
                    'Ciclo': st.column_config.SelectboxColumn(
                        'Ciclo',
                        options=['C0 / Nenhum', 'C1', 'C2', 'C3', 'C4', 'Outro'],
                    ),
                    'Incorporado ao valor formalizado?': st.column_config.SelectboxColumn(
                        'Incorporado ao valor formalizado?',
                        options=['Sim', 'Não'],
                    ),
                },
            )
            for _, evento_row in eventos_editados.iterrows():
                evento = {col: str(evento_row.get(col, '') or '').strip() for col in colunas_eventos}
                possui_conteudo_relevante = any([
                    evento.get('Tipo de evento', ''),
                    evento.get('Data', ''),
                    evento.get('Valor formalizado/impacto', ''),
                    evento.get('Observação', ''),
                ])
                if possui_conteudo_relevante:
                    eventos_limpos.append(evento)

        valor_original = _parse_moeda_br(valor_original_txt)
        valor_formalizado = _parse_moeda_br(valor_formalizado_txt)
        contexto = {
            'valor_original_contrato': valor_original,
            'valor_original_contrato_texto': valor_original_txt,
            'valor_formalizado_anterior': valor_formalizado,
            'valor_formalizado_anterior_texto': valor_formalizado_txt,
            'modo_valor_formalizado': modo_valor_formalizado,
            'percentual_ja_aplicado_pct': float(percentual_ja_aplicado_pct),
            'valor_formalizado_calculado': float(valor_calculado_formalizado),
            'ultimo_ciclo_concedido': ultimo_ciclo.strip(),
            'data_base_ultimo_ciclo': data_base_ultimo.strip(),
            'data_pedido_ultimo_ciclo': data_pedido_ultimo.strip(),
            'referencia_documental_historico': referencia_documental_historico.strip(),
            'observacao_historico': observacao.strip(),
            'eventos_historicos_anteriores': eventos_limpos,
        }
        st.session_state['contexto_contratual_anterior'] = contexto

        if valor_original > 0 or valor_formalizado > 0 or ultimo_ciclo.strip() or observacao.strip():
            st.info(
                f"Contexto informado: valor original { _formatar_moeda_br_md(valor_original) }; "
                f"valor formalizado antes desta análise { _formatar_moeda_br_md(valor_formalizado) }."
            )
    return st.session_state.get('contexto_contratual_anterior', {})

def _render_equacao_ist(i_ini, i_fim, variacao):
    equacao_html = f"""
    <div style="background:#F4F6F8;border:1px solid #E1E6EB;border-radius:10px;padding:14px 18px;margin-top:10px;">
        <div style="font-family:Consolas, Monaco, monospace;font-size:1.15rem;line-height:1.8;color:#334155;">
            <span style="color:#0F766E;">({i_fim:.3f}</span>
            <span style="color:#94A3B8;"> / </span>
            <span style="color:#0F766E;">{i_ini:.3f}</span>
            <span style="color:#94A3B8;">) - 1 = </span>
            <span style="color:#B45309;font-weight:600;">{variacao*100:.4f}%</span>
        </div>
    </div>
    """
    st.markdown(equacao_html, unsafe_allow_html=True)


def _ajustar_larguras(writer, nomes_abas):
    from openpyxl.utils import get_column_letter

    for nome_aba in nomes_abas:
        ws = writer.book[nome_aba]
        # ITENS_REMANESCENTES usa duas linhas de cabeçalho: linha 1 para datas de referência e linha 2 para títulos.
        ws.freeze_panes = "A3" if nome_aba == 'ITENS_REMANESCENTES' else "A2"
        for column_cells in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                try:
                    max_length = max(max_length, len(str(cell.value)) if cell.value is not None else 0)
                except Exception:
                    pass
            ws.column_dimensions[column_letter].width = min(max(max_length + 2, 14), 45)


def _aplicar_estilos_coleta(writer, ciclos):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill('solid', fgColor='1F4E78')
    header_font = Font(color='FFFFFF', bold=True)
    input_fill = PatternFill('solid', fgColor='FFF2CC')
    date_fill = PatternFill('solid', fgColor='D9EAD3')
    light_fill = PatternFill('solid', fgColor='EAF2F8')
    thin = Side(style='thin', color='D9E2F3')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    money_fmt = 'R$ #,##0.00'
    number_fmt = '#,##0.0000'

    for ws in writer.book.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical='center')

        header_row = 2 if ws.title == 'ITENS_REMANESCENTES' else 1
        for cell in ws[header_row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    if 'FINANCEIRO_MENSAL' in writer.book.sheetnames:
        ws = writer.book['FINANCEIRO_MENSAL']
        # Coluna C é a coluna de preenchimento pelo fiscal.
        for row in range(1, ws.max_row + 1):
            ws.cell(row=row, column=3).fill = input_fill
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='right' if row > 1 else 'center', vertical='center')
            if row > 1:
                ws.cell(row=row, column=3).number_format = money_fmt
        ws['C1'].value = 'Valor pago/faturado (preencher)'
        ws['C1'].font = header_font
        ws['C1'].fill = header_fill

    if 'ITENS_REMANESCENTES' in writer.book.sheetnames:
        ws = writer.book['ITENS_REMANESCENTES']
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
        ws.cell(row=1, column=1).value = 'Dados do item e valor original'
        ws.cell(row=1, column=1).fill = light_fill
        ws.cell(row=1, column=1).font = Font(bold=True)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')

        # Colunas: A Item | B Quantidade contratada | C Valor unitário original | D Valor total | E... remanescentes
        for row in range(3, ws.max_row + 1):
            ws.cell(row=row, column=3).number_format = money_fmt
            ws.cell(row=row, column=4).number_format = money_fmt
            ws.cell(row=row, column=4).value = f'=IF(OR(B{row}="",C{row}=""),"",B{row}*C{row})'

        # Destacar colunas de remanescente para preenchimento pelo fiscal e colocar a data de referência na linha 1.
        first_rem_col = 5
        for idx, ciclo in enumerate(ciclos):
            col = first_rem_col + idx
            letra = get_column_letter(col)
            ws.cell(row=1, column=col).value = ciclo.get('data_base', '')
            ws.cell(row=1, column=col).fill = date_fill
            ws.cell(row=1, column=col).font = Font(bold=True)
            ws.cell(row=1, column=col).alignment = Alignment(horizontal='center', vertical='center')
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col).fill = input_fill if row > 2 else header_fill
                if row == 2:
                    ws.cell(row=row, column=col).font = header_font
                    ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.column_dimensions[letra].width = 24

    if 'CICLOS' in writer.book.sheetnames:
        ws = writer.book['CICLOS']
        for row in range(2, ws.max_row + 1):
            # Fator e Fator acumulado.
            if ws.max_column >= 8:
                ws.cell(row=row, column=8).number_format = number_fmt
            if ws.max_column >= 9:
                ws.cell(row=row, column=9).number_format = number_fmt

    if 'ADITIVOS_QUANTITATIVOS' in writer.book.sheetnames:
        ws = writer.book['ADITIVOS_QUANTITATIVOS']
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=4).number_format = money_fmt
            ws.cell(row=row, column=4).fill = input_fill


def gerar_arquivo_coleta_excel(dados_admissibilidade):
    """Gera o Arquivo de Coleta para as fases de Valor Global e Relatório.

    Regras da planilha:
    - 200 linhas pré-formatadas nas abas de preenchimento.
    - Linha 201 reservada para somatórios.
    - Aditivos por item/lançamento, com acréscimo ou supressão.
    """
    output = BytesIO()
    ciclos = dados_admissibilidade.get('ciclos', [])
    data_geracao = datetime.now(ZoneInfo('America/Sao_Paulo')).strftime('%d/%m/%Y %H:%M')

    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        workbook = writer.book

        fmt_header = workbook.add_format({
            'bold': True, 'font_color': 'white', 'bg_color': '#1F4E79',
            'border': 1, 'align': 'center', 'valign': 'vcenter', 'text_wrap': True,
        })
        fmt_subheader = workbook.add_format({
            'bold': True, 'bg_color': '#D9EAD3', 'border': 1,
            'align': 'center', 'valign': 'vcenter', 'text_wrap': True,
        })
        fmt_input = workbook.add_format({'bg_color': '#FFF2CC', 'border': 1})
        fmt_input_date = workbook.add_format({'num_format': 'dd/mm/yyyy', 'bg_color': '#FFF2CC', 'border': 1})
        fmt_date = workbook.add_format({'num_format': 'dd/mm/yyyy', 'border': 1})
        fmt_date_no_border = workbook.add_format({'num_format': 'dd/mm/yyyy'})
        fmt_input_num = workbook.add_format({'num_format': '#,##0.00', 'bg_color': '#FFF2CC', 'border': 1})
        fmt_input_money = workbook.add_format({'num_format': 'R$ #,##0.00', 'bg_color': '#FFF2CC', 'border': 1})
        fmt_money = workbook.add_format({'num_format': 'R$ #,##0.00', 'border': 1})
        fmt_money_no_border = workbook.add_format({'num_format': 'R$ #,##0.00'})
        fmt_money_auto = workbook.add_format({'num_format': 'R$ #,##0.00', 'bg_color': '#EDEDED', 'border': 1})
        fmt_number = workbook.add_format({'num_format': '#,##0.00', 'border': 1})
        fmt_text = workbook.add_format({'border': 1})
        fmt_auto = workbook.add_format({'bg_color': '#EDEDED', 'border': 1})
        fmt_total = workbook.add_format({'bold': True, 'bg_color': '#E2F0D9', 'border': 1})
        fmt_total_money = workbook.add_format({'bold': True, 'bg_color': '#E2F0D9', 'border': 1, 'num_format': 'R$ #,##0.00'})
        fmt_total_num = workbook.add_format({'bold': True, 'bg_color': '#E2F0D9', 'border': 1, 'num_format': '#,##0.00'})
        fmt_decrease_red = workbook.add_format({'font_color': '#C00000'})
        fmt_percent = workbook.add_format({'num_format': '0.00%', 'border': 1})
        fmt_factor = workbook.add_format({'num_format': '0.0000', 'border': 1})
        fmt_factor_auto = workbook.add_format({'num_format': '0.0000', 'bg_color': '#EDEDED', 'border': 1})
        fmt_no_border = workbook.add_format({})
        fmt_int_no_border = workbook.add_format({'num_format': '0'})
        fmt_percent_no_border = workbook.add_format({'num_format': '0.00%'})
        fmt_factor_no_border = workbook.add_format({'num_format': '0.0000'})
        fmt_text_left_no_border = workbook.add_format({'align': 'left'})
        fmt_money_left_no_border = workbook.add_format({'num_format': 'R$ #,##0.00', 'align': 'left'})
        fmt_text_wrap_no_border = workbook.add_format({'align': 'left', 'valign': 'top', 'text_wrap': True})
        fmt_header_no_border = workbook.add_format({'bold': True, 'font_color': 'white', 'bg_color': '#1F4E79', 'align': 'center', 'valign': 'vcenter', 'text_wrap': True})
        fmt_text_wrap = workbook.add_format({'border': 1, 'align': 'left', 'valign': 'top', 'text_wrap': True})
        fmt_int_left_no_border = workbook.add_format({'num_format': '0', 'align': 'left'})
        fmt_percent_left_no_border = workbook.add_format({'num_format': '0.00%', 'align': 'left'})
        fmt_factor_left_no_border = workbook.add_format({'num_format': '0.0000', 'align': 'left'})
        cores_ciclos_col_c = ['#EAF2F8', '#E2F0D9', '#FFF2CC', '#FCE4D6', '#E4DFEC', '#DDEBF7', '#F4CCCC', '#D9EAD3']
        fmt_ciclos_col_c = [workbook.add_format({'bg_color': cor, 'border': 1}) for cor in cores_ciclos_col_c]
        cor_aba_automatica = '#D9EAF7'

        # PARAMETROS_REAJUSTE
        parametros = pd.DataFrame([
            ['Origem da análise', dados_admissibilidade.get('origem', dados_admissibilidade.get('tipo', ''))],
            ['Índice utilizado', dados_admissibilidade.get('indice', '')],
            ['Data-base original', dados_admissibilidade.get('data_base_original', '')],
            ['Quantidade de ciclos', len(ciclos)],
            ['Variação acumulada final', dados_admissibilidade.get('variacao_acumulada', 0.0)],
            ['Fator acumulado total', dados_admissibilidade.get('fator_acumulado', dados_admissibilidade.get('fator', 1.0))],
            ['Valor original do contrato (contexto)', dados_admissibilidade.get('contexto_contratual_anterior', {}).get('valor_original_contrato', '')],
            ['Valor contratual formalizado antes desta análise', dados_admissibilidade.get('contexto_contratual_anterior', {}).get('valor_formalizado_anterior', '')],
            ['Último ciclo já concedido/formalizado', dados_admissibilidade.get('contexto_contratual_anterior', {}).get('ultimo_ciclo_concedido', '')],
            ['Data-base do último ciclo concedido/formalizado', dados_admissibilidade.get('contexto_contratual_anterior', {}).get('data_base_ultimo_ciclo', '')],
            ['Data do pedido do último ciclo concedido/formalizado', dados_admissibilidade.get('contexto_contratual_anterior', {}).get('data_pedido_ultimo_ciclo', '')],
            ['Percentual já aplicado antes desta análise', dados_admissibilidade.get('contexto_contratual_anterior', {}).get('percentual_ja_aplicado_pct', '')],
            ['Referência documental do histórico anterior', dados_admissibilidade.get('contexto_contratual_anterior', {}).get('referencia_documental_historico', '')],
            ['Observação sobre histórico anterior', dados_admissibilidade.get('contexto_contratual_anterior', {}).get('observacao_historico', '')],
            ['Eventos históricos anteriores', 'Ver aba EVENTOS_HISTORICOS_ANTERIORES' if dados_admissibilidade.get('contexto_contratual_anterior', {}).get('eventos_historicos_anteriores', []) else ''],
            ['Data de geração do arquivo', data_geracao],
        ], columns=['Campo', 'Valor'])
        parametros.to_excel(writer, sheet_name='PARAMETROS_REAJUSTE', index=False)
        ws = writer.sheets['PARAMETROS_REAJUSTE']
        ws.set_column('A:A', 34)
        ws.set_column('B:B', 36)
        ws.write(0, 0, 'Campo', fmt_header)
        ws.write(0, 1, 'Valor', fmt_header)
        # Linhas 5, 6 e 7: alinhamento à esquerda para leitura limpa do bloco de parâmetros.
        ws.write(4, 0, 'Quantidade de ciclos', fmt_text_left_no_border)
        ws.write_number(4, 1, len(ciclos), fmt_int_left_no_border)
        ws.write(5, 0, 'Variação acumulada final', fmt_text_left_no_border)
        ws.write_number(5, 1, float(dados_admissibilidade.get('variacao_acumulada', 0.0)), fmt_percent_left_no_border)
        ws.write(6, 0, 'Fator acumulado total', fmt_text_left_no_border)
        ws.write_number(6, 1, float(dados_admissibilidade.get('fator_acumulado', dados_admissibilidade.get('fator', 1.0))), fmt_factor_left_no_border)
        ws.set_tab_color(cor_aba_automatica)
        contexto_excel = dados_admissibilidade.get('contexto_contratual_anterior', {}) or {}
        valor_original_contexto_txt = str(contexto_excel.get('valor_original_contrato_texto') or '').strip()
        valor_formalizado_contexto_txt = str(contexto_excel.get('valor_formalizado_anterior_texto') or '').strip()
        valor_original_contexto = float(contexto_excel.get('valor_original_contrato') or 0.0)
        valor_formalizado_contexto = float(contexto_excel.get('valor_formalizado_anterior') or 0.0)
        if valor_original_contexto_txt or valor_original_contexto > 0:
            ws.write_number(7, 1, valor_original_contexto, fmt_money_no_border)
        else:
            ws.write_blank(7, 1, None, fmt_no_border)
        if valor_formalizado_contexto_txt or valor_formalizado_contexto > 0:
            ws.write_number(8, 1, valor_formalizado_contexto, fmt_money_no_border)
        else:
            ws.write_blank(8, 1, None, fmt_no_border)

        # Formatação final da aba PARAMETROS_REAJUSTE: coluna B alinhada à esquerda,
        # com preservação dos formatos numéricos relevantes.
        for row_idx in range(1, len(parametros) + 1):
            valor_param = parametros.iloc[row_idx - 1, 1]
            ws.write(row_idx, 1, valor_param, fmt_text_left_no_border)
        ws.write_number(4, 1, len(ciclos), fmt_int_left_no_border)
        ws.write_number(5, 1, float(dados_admissibilidade.get('variacao_acumulada', 0.0)), fmt_percent_left_no_border)
        ws.write_number(6, 1, float(dados_admissibilidade.get('fator_acumulado', dados_admissibilidade.get('fator', 1.0))), fmt_factor_left_no_border)
        if valor_original_contexto_txt or valor_original_contexto > 0:
            ws.write_number(7, 1, valor_original_contexto, fmt_money_left_no_border)
        else:
            ws.write_blank(7, 1, None, fmt_text_left_no_border)
        if valor_formalizado_contexto_txt or valor_formalizado_contexto > 0:
            ws.write_number(8, 1, valor_formalizado_contexto, fmt_money_left_no_border)
        else:
            ws.write_blank(8, 1, None, fmt_text_left_no_border)
        percentual_contexto = contexto_excel.get('percentual_ja_aplicado_pct', '')
        try:
            if str(percentual_contexto).strip() != '':
                ws.write_number(12, 1, float(percentual_contexto) / 100, fmt_percent_left_no_border)
        except Exception:
            ws.write(12, 1, percentual_contexto, fmt_text_left_no_border)


        # Eventos históricos anteriores em aba própria, para evitar JSON longo em B12.
        eventos_historicos_excel = _eventos_historicos_para_exportacao(contexto_excel)
        if eventos_historicos_excel:
            ws_ev = workbook.add_worksheet('EVENTOS_HISTORICOS_ANTERIORES')
            writer.sheets['EVENTOS_HISTORICOS_ANTERIORES'] = ws_ev
            ev_headers = ['Tipo de evento', 'Ciclo', 'Data', 'Valor formalizado/impacto', 'Incorporado ao valor formalizado?', 'Observação']
            for col, title in enumerate(ev_headers):
                ws_ev.write(0, col, title, fmt_header_no_border)
            for row_idx, evento in enumerate(eventos_historicos_excel, start=1):
                valor_evento = evento.get('Valor formalizado/impacto', evento.get('Valor atualizado/formalizado', evento.get('Valor original', '')))
                ws_ev.write(row_idx, 0, evento.get('Tipo de evento', ''), fmt_text_left_no_border)
                ws_ev.write(row_idx, 1, evento.get('Ciclo', ''), fmt_text_left_no_border)
                ws_ev.write(row_idx, 2, evento.get('Data', ''), fmt_text_left_no_border)
                ws_ev.write(row_idx, 3, valor_evento, fmt_text_left_no_border)
                ws_ev.write(row_idx, 4, evento.get('Incorporado ao valor formalizado?', ''), fmt_text_left_no_border)
                ws_ev.write(row_idx, 5, evento.get('Observação', ''), fmt_text_wrap_no_border)
            ws_ev.set_column('A:A', 24, fmt_text_left_no_border)
            ws_ev.set_column('B:B', 14, fmt_text_left_no_border)
            ws_ev.set_column('C:C', 18, fmt_text_left_no_border)
            ws_ev.set_column('D:E', 28, fmt_text_left_no_border)
            ws_ev.set_column('F:F', 80, fmt_text_wrap_no_border)
            ws_ev.set_tab_color(cor_aba_automatica)

        # CICLOS
        ciclos_rows = []
        for ciclo in ciclos:
            situacao = str(ciclo.get('situacao', ''))
            tratamento = 'Precluso' if 'PRECLUSO' in situacao.upper() else 'A apurar'
            ciclos_rows.append({
                'Ciclo': ciclo.get('ciclo', ''),
                'Data-base': ciclo.get('data_base', ''),
                'Intervalo do índice': ciclo.get('intervalo_indice', ciclo.get('Janela', '')),
                'Janela de admissibilidade': ciclo.get('janela_admissibilidade', ciclo.get('JanelaAdm', '')),
                'Data do pedido': ciclo.get('data_pedido', ciclo.get('Pedido', '')),
                'Início financeiro': ciclo.get('financeiro_inicio', ''),
                'Fim financeiro': ciclo.get('financeiro_fim', ''),
                'Situação': situacao,
                'Variação': ciclo.get('variacao', 0.0),
                'Fator': ciclo.get('fator', 1.0),
                'Fator acumulado': ciclo.get('fator_acumulado', 1.0),
                'Tratamento financeiro do ciclo': tratamento,
                'Situação automática': ciclo.get('situacao_automatica', ciclo.get('situacao', '')),
                'Acordo negocial': 'Sim' if ciclo.get('superacao_negocial', False) else 'Não',
                'Situação aplicada': ciclo.get('situacao_aplicada', ciclo.get('situacao', '')),
                'Percentual apurado pelo índice': ciclo.get('percentual_indice', ciclo.get('variacao', 0.0)),
                'Percentual aplicado': ciclo.get('percentual_aplicado', ciclo.get('variacao', 0.0)),
                'Justificativa negocial': ciclo.get('justificativa_negocial', ''),
                'Referência documental': ciclo.get('referencia_documental', ''),
            })
        df_ciclos = pd.DataFrame(ciclos_rows)
        df_ciclos.to_excel(writer, sheet_name='CICLOS', index=False)
        ws = writer.sheets['CICLOS']
        ws.set_tab_color(cor_aba_automatica)
        ws.set_column('A:A', 12)
        ws.set_column('B:H', 24)
        ws.set_column('I:I', 12)
        ws.set_column('J:K', 14)
        ws.set_column('L:L', 34)
        ws.set_column('M:Q', 26)
        ws.set_column('R:S', 16)
        ws.set_column('T:U', 42, fmt_text_wrap)
        for col, title in enumerate(df_ciclos.columns):
            ws.write(0, col, title, fmt_header)
        for row in range(1, len(df_ciclos) + 1):
            data_base_excel = pd.to_datetime(df_ciclos.iloc[row-1]['Data-base'], dayfirst=True, errors='coerce')
            if pd.notna(data_base_excel):
                ws.write_datetime(row, 1, data_base_excel.to_pydatetime(), fmt_date_no_border if row in [1, 2, 3] else fmt_date)
            ws.write(row, 8, df_ciclos.iloc[row-1]['Variação'], workbook.add_format({'num_format': '0.00%'}))
            ws.write(row, 9, df_ciclos.iloc[row-1]['Fator'], workbook.add_format({'num_format': '0.0000'}))
            ws.write(row, 10, df_ciclos.iloc[row-1]['Fator acumulado'], workbook.add_format({'num_format': '0.0000'}))
            ws.write(row, 11, df_ciclos.iloc[row-1]['Tratamento financeiro do ciclo'], workbook.add_format({}))
            fmt_percentual_ciclos = fmt_percent
            if 'Percentual apurado pelo índice' in df_ciclos.columns:
                col_pct_indice = df_ciclos.columns.get_loc('Percentual apurado pelo índice')
                ws.write(row, col_pct_indice, df_ciclos.iloc[row-1]['Percentual apurado pelo índice'], fmt_percentual_ciclos)
            if 'Percentual aplicado' in df_ciclos.columns:
                col_pct_aplicado = df_ciclos.columns.get_loc('Percentual aplicado')
                ws.write(row, col_pct_aplicado, df_ciclos.iloc[row-1]['Percentual aplicado'], fmt_percentual_ciclos)
            if 'Justificativa negocial' in df_ciclos.columns:
                col_justificativa = df_ciclos.columns.get_loc('Justificativa negocial')
                ws.write(row, col_justificativa, df_ciclos.iloc[row-1]['Justificativa negocial'], fmt_text_wrap)
            if 'Referência documental' in df_ciclos.columns:
                col_referencia = df_ciclos.columns.get_loc('Referência documental')
                ws.write(row, col_referencia, df_ciclos.iloc[row-1]['Referência documental'], fmt_text_wrap)

        # BASE_EXECUCAO_MENSAL
        financeiro_rows = []
        for ciclo in ciclos:
            ciclo_nome = ciclo.get('ciclo', '')
            for competencia in _competencias_mensais(ciclo.get('financeiro_inicio', ''), ciclo.get('financeiro_fim', '')):
                financeiro_rows.append({'Ciclo': ciclo_nome, 'Competência': competencia, 'Valor bruto medido/aprovado por competência': ''})
        if not financeiro_rows:
            financeiro_rows.append({'Ciclo': '', 'Competência': '', 'Valor bruto medido/aprovado por competência': ''})
        df_fin = pd.DataFrame(financeiro_rows)
        df_fin.to_excel(writer, sheet_name='BASE_EXECUCAO_MENSAL', index=False)
        ws = writer.sheets['BASE_EXECUCAO_MENSAL']
        ws.set_column('A:A', 12)
        ws.set_column('B:B', 18)
        ws.set_column('C:C', 24)
        for col, title in enumerate(df_fin.columns):
            ws.write(0, col, title, fmt_header)
        for row in range(1, len(df_fin) + 1):
            ws.write(row, 2, '', fmt_input_money)
        total_row_fin = len(df_fin) + 1
        ultima_linha_fin_excel = len(df_fin) + 1
        ws.write(total_row_fin, 0, 'TOTAL', fmt_total)
        ws.write(total_row_fin, 1, '', fmt_total)
        ws.write_formula(total_row_fin, 2, f'=ROUND(SUM(C2:C{ultima_linha_fin_excel}),2)', fmt_total_money)
        orientacao_row_fin = total_row_fin + 2
        ws.write(orientacao_row_fin, 0, 'Orientação', fmt_total)
        ws.merge_range(orientacao_row_fin, 1, orientacao_row_fin, 2, 'Preencha a base mensal por competência com o valor bruto demandado, medido ou aprovado. Se o valor vier de consumo itemizado, informe a competência de execução/faturamento e confirme, na observação do processo, se não houve glosas, descontos, retenções, notas substituídas ou divergências entre consumo, medição e faturamento. A diferença entre estoque inicial e remanescente pode apoiar a apuração física, mas não substitui automaticamente esta base mensal.', fmt_text_wrap)
        ws.set_row(orientacao_row_fin, 78)

        # ITENS_REMANESCENTES
        ws_it = workbook.add_worksheet('ITENS_REMANESCENTES')
        writer.sheets['ITENS_REMANESCENTES'] = ws_it
        rem_cols = []
        for ciclo in ciclos:
            rem_cols.append((f"Remanescente início {ciclo.get('ciclo', '')}", ciclo.get('data_base', '')))
        base_headers = ['Item', 'Quantidade contratada', 'Valor unitário original', 'Valor total']
        headers = base_headers + [c[0] for c in rem_cols]
        ws_it.merge_range(0, 0, 0, 3, 'Dados do item e valor original', fmt_subheader)
        for idx, (_, data_ref) in enumerate(rem_cols, start=4):
            ws_it.write(0, idx, data_ref, fmt_subheader)
        for col, title in enumerate(headers):
            ws_it.write(1, col, title, fmt_header)
        ws_it.set_column(0, 0, 12)
        ws_it.set_column(1, 1, 24)
        ws_it.set_column(2, 3, 22)
        if rem_cols:
            ws_it.set_column(4, 4 + len(rem_cols) - 1, 25)
        for row in range(2, 200):
            ws_it.write(row, 0, '', fmt_text)       # A sem preenchimento
            ws_it.write(row, 1, '', fmt_number)     # B sem preenchimento
            ws_it.write(row, 2, '', fmt_money)      # C sem preenchimento
            ws_it.write_formula(row, 3, f'=IF(OR(B{row+1}="",C{row+1}=""),"",ROUND(B{row+1}*C{row+1},2))', fmt_money_auto)
            for col in range(4, 4 + len(rem_cols)):
                ws_it.write(row, col, '', fmt_input_num)
        ws_it.write(200, 0, 'TOTAL', fmt_total)
        ws_it.write(200, 1, '', fmt_total)
        ws_it.write(200, 2, '', fmt_total)
        ws_it.write_formula(200, 3, '=ROUND(SUM(D3:D200),2)', fmt_total_money)
        for col in range(4, 4 + len(rem_cols)):
            col_letter = chr(ord('A') + col)
            ws_it.write_formula(200, col, f'=SUM({col_letter}3:{col_letter}200)', fmt_total_num)

        # ADITIVOS_QUANTITATIVOS
        ws_ad = workbook.add_worksheet('ADITIVOS_QUANTITATIVOS')
        writer.sheets['ADITIVOS_QUANTITATIVOS'] = ws_ad
        ad_headers = [
            'Item', 'Data do aditivo', 'Ciclo/Marco', 'Tipo de alteração',
            'Quantidade acrescida/suprimida', 'Valor unitário original',
            'Valor original da alteração', 'Aplicar reajuste acumulado? (Sim/Não)',
            'Fator acumulado aplicável', 'Valor atualizado da alteração',
            'Tratamento do aditivo'
        ]
        for col, title in enumerate(ad_headers):
            ws_ad.write(0, col, title, fmt_header)
        ws_ad.set_column('A:A', 12)
        ws_ad.set_column('B:B', 18)
        ws_ad.set_column('C:C', 16)
        ws_ad.set_column('D:D', 20)
        ws_ad.set_column('E:E', 28)
        ws_ad.set_column('F:G', 24)
        ws_ad.set_column('H:H', 32)
        ws_ad.set_column('I:I', 22)
        ws_ad.set_column('J:J', 26)
        ws_ad.set_column('K:K', 42)
        ws_ad.write(202, 0, "Orientação", fmt_total)
        ws_ad.write(202, 1, "Use 'Computar nesta análise' para aditivos/supressões que devem impactar o Valor Global atual. Use 'Informativo - já incluído no valor formalizado' quando o lançamento já estiver contemplado no campo Valor contratual formalizado antes desta análise.", fmt_text)
        ciclo_range = f'CICLOS!$A$2:$K${len(ciclos)+1}' if ciclos else 'CICLOS!$A$2:$K$2'
        data_range = f'CICLOS!$B$2:$B${len(ciclos)+1}' if ciclos else 'CICLOS!$B$2:$B$2'
        ciclo_nome_range = f'CICLOS!$A$2:$A${len(ciclos)+1}' if ciclos else 'CICLOS!$A$2:$A$2'
        for row in range(1, 200):
            excel_row = row + 1
            ws_ad.write(row, 0, '', fmt_text)
            ws_ad.write(row, 1, '', fmt_input_date)
            # Coluna C: identifica automaticamente o ciclo pela data do aditivo.
            # Regra: último ciclo cuja Data-base seja menor ou igual à Data do Aditivo.
            if len(ciclos) > 0:
                formula_ciclo = (
                    f'=IF(B{excel_row}="","",'
                    f'IF(B{excel_row}<MIN(CICLOS!$B$2:$B${len(ciclos)+1}),"C0",'
                    f'IFERROR(LOOKUP(2,1/(CICLOS!$B$2:$B${len(ciclos)+1}<=B{excel_row}),'
                    f'CICLOS!$A$2:$A${len(ciclos)+1}),"Fora de Ciclo")))'
                )
                ws_ad.write_formula(row, 2, formula_ciclo, fmt_auto)
            else:
                ws_ad.write(row, 2, '', fmt_auto)
            ws_ad.write(row, 3, 'Acréscimo', fmt_input)
            ws_ad.write(row, 4, '', fmt_input_num)
            ws_ad.write(row, 5, '', fmt_input_money)
            ws_ad.write_formula(row, 6, f'=IF(OR(E{excel_row}="",F{excel_row}=""),"",ROUND(E{excel_row}*F{excel_row},2))', fmt_money_auto)
            ws_ad.write(row, 7, 'Sim', fmt_input)
            ws_ad.write_formula(row, 8, f'=IF(C{excel_row}="","",IFERROR(VLOOKUP(C{excel_row},{ciclo_range},11,FALSE),1))', fmt_factor_auto)
            ws_ad.write_formula(row, 9, f'=IF(G{excel_row}="","",ROUND(IF(OR(UPPER(D{excel_row})="DECRÉSCIMO",UPPER(D{excel_row})="DECRESCIMO",UPPER(D{excel_row})="SUPRESSÃO",UPPER(D{excel_row})="SUPRESSAO"),-ABS(G{excel_row}),ABS(G{excel_row}))*IF(OR(UPPER(H{excel_row})="NÃO",UPPER(H{excel_row})="NAO"),1,I{excel_row}),2))', fmt_money_auto)
            ws_ad.write(row, 10, 'Computar nesta análise', fmt_input)
        ws_ad.data_validation(1, 3, 199, 3, {'validate': 'list', 'source': ['Acréscimo', 'Decréscimo']})
        ws_ad.data_validation(1, 7, 199, 7, {'validate': 'list', 'source': ['Sim', 'Não']})
        ws_ad.data_validation(1, 10, 199, 10, {'validate': 'list', 'source': ['Computar nesta análise', 'Informativo - já incluído no valor formalizado']})
        # Coluna C: cor discreta por ciclo para facilitar a leitura quando houver mudança de marco.
        ws_ad.conditional_format(1, 2, 199, 2, {
            'type': 'formula',
            'criteria': '=$C2="C0"',
            'format': fmt_ciclos_col_c[0],
        })
        for idx_ciclo, ciclo_ref in enumerate(ciclos or []):
            ciclo_nome = str(ciclo_ref.get('ciclo', '') or '').strip()
            if ciclo_nome:
                ws_ad.conditional_format(1, 2, 199, 2, {
                    'type': 'formula',
                    'criteria': f'=$C2="{ciclo_nome}"',
                    'format': fmt_ciclos_col_c[(idx_ciclo + 1) % len(fmt_ciclos_col_c)],
                })
        # Se o usuário selecionar Decréscimo/Supressão, destacar toda a linha em fonte vermelha.
        ws_ad.conditional_format(1, 0, 199, 10, {
            'type': 'formula',
            # Fórmula robusta: usa início do texto para evitar falhas por acento/localidade.
            'criteria': '=OR(LEFT(UPPER($D2),4)="DECR",LEFT(UPPER($D2),6)="SUPRES")',
            'format': fmt_decrease_red,
        })
        ws_ad.write(200, 0, 'TOTAL', fmt_total)
        for col in range(1, 6):
            ws_ad.write(200, col, '', fmt_total)
        ws_ad.write_formula(200, 6, '=ROUND(SUM(G2:G200),2)', fmt_total_money)
        ws_ad.write(200, 7, '', fmt_total)
        ws_ad.write(200, 8, '', fmt_total)
        ws_ad.write_formula(200, 9, '=ROUND(SUM(J2:J200),2)', fmt_total_money)
        ws_ad.write(200, 10, '', fmt_total)

    output.seek(0)
    return output.getvalue()
if not st.session_state.get("_calculadora_reajustes_embedded", False):
    render_marca_topo()
if not st.session_state.get("_calculadora_reajustes_embedded", False):
    st.title("Único")

contexto_contratual = _render_contexto_contratual_anterior()

primeiro_ciclo_num = _primeiro_ciclo_analise(contexto_contratual)
ciclo_label = f"C{primeiro_ciclo_num}"
default_dt_base = _data_base_inicial_pelo_contexto(contexto_contratual, datetime(2023, 8, 2))

if _ciclo_para_numero(contexto_contratual.get('ultimo_ciclo_concedido', '')) > 0 and not contexto_contratual.get('data_pedido_ultimo_ciclo'):
    st.error(
        "Processamento inviável: há ciclo anterior concedido/formalizado, mas a data do pedido do último ciclo não foi informada."
    )
    st.warning(
        "Informe a data do pedido do último ciclo no Contexto do Contrato. Esse dado será usado como âncora do próximo ciclo, pois corresponde ao início dos efeitos financeiros do último reajuste concedido/formalizado."
    )
    st.stop()

col1, col2 = st.columns(2)
with col1:
    dt_base = st.date_input(f"Data-base/âncora do {ciclo_label}:", value=default_dt_base, format="DD/MM/YYYY")
    dt_solic = st.date_input(f"Data do Pedido {ciclo_label}:", value=datetime(2024, 4, 9), format="DD/MM/YYYY")
with col2:
    tipo_idx = render_indice_contrato_selectbox(key="indice_fluxo_unico")

chave_analise_simples = (
    ciclo_label,
    dt_base.isoformat(),
    dt_solic.isoformat(),
    tipo_idx,
    str(contexto_contratual.get('valor_original_contrato', 0.0)),
    str(contexto_contratual.get('valor_formalizado_anterior', 0.0)),
    contexto_contratual.get('ultimo_ciclo_concedido', ''),
    contexto_contratual.get('data_pedido_ultimo_ciclo', ''),
    contexto_contratual.get('modo_valor_formalizado', ''),
    str(contexto_contratual.get('percentual_ja_aplicado_pct', 0.0)),
    contexto_contratual.get('observacao_historico', ''),
)


if st.button("Processar Análise", type="primary", use_container_width=False):
    st.session_state["chave_analise_simples_processada"] = chave_analise_simples

if st.session_state.get("chave_analise_simples_processada") != chave_analise_simples:
    st.stop()

data_hoje = datetime.now(ZoneInfo("America/Sao_Paulo")).date()
data_pedido_cmp = _data_para_date_segura(dt_solic)
if data_pedido_cmp and data_pedido_cmp > data_hoje:
    st.error(f"Processamento inviável: a data do pedido informada no {ciclo_label} é futura em relação à data atual.")
    st.warning("A Calculadora processa apenas pedidos já apresentados. Revise a data do pedido antes de prosseguir.")
    st.stop()

# Definição de datas do ciclo
dt_fim_ap = dt_base + relativedelta(months=11)
dt_aniv = dt_base + relativedelta(years=1)
dt_limite = dt_aniv + relativedelta(days=90)

# Regra de Admissibilidade
if dt_solic < dt_aniv:
    if dt_solic.year == dt_aniv.year and dt_solic.month == dt_aniv.month:
        status_ped = "🟡 ADMISSÍVEL - RESSALVA"
    else:
        status_ped = "⚠️ ADIANTADO (ANTES DOS 12 MESES)"
elif dt_solic <= dt_limite:
    status_ped = "✅ TEMPESTIVO"
else:
    status_ped = "❌ PRECLUSO"

res = get_ist_local(dt_base, dt_fim_ap) if "IST" in tipo_idx else get_index_data("433" if "IPCA" in tipo_idx else "189", dt_base, dt_fim_ap)

validacao_indice = _validar_indice_disponivel(res, dt_base, dt_fim_ap, tipo_idx)
if not validacao_indice.get("ok", False):
    _render_alerta_indice_ausente(validacao_indice, ciclo_label)
    st.stop()

if res:
    v_fmt = f"{res['variacao']*100:,.2f}%".replace('.', ',')
    fator_ciclo = 1 + res['variacao']
    st.metric("Variação Apurada", v_fmt)

    st.markdown("### Dados do Ciclo")
    # Ajuste solicitado de nomenclatura e nova linha de janela
    periodo_inicio = res['d_ini'] if 'd_ini' in res else dt_base
    periodo_fim = res['d_fim'] if 'd_fim' in res else dt_fim_ap
    janela_str = f"{pd.to_datetime(periodo_inicio).strftime('%m/%Y')} a {pd.to_datetime(periodo_fim).strftime('%m/%Y')}"
    janela_adm_str = f"{dt_aniv.strftime('%d/%m/%Y')} a {dt_limite.strftime('%d/%m/%Y')}"

    st.write(f"- **Intervalo do {ciclo_label}:** {janela_str}")
    st.write(f"- **Janela de Admissibilidade:** {janela_adm_str}")
    st.write(f"- **Situação:** {status_ped}")

    # Acordo negocial: preserva o diagnóstico automático, mas permite aplicar percentual manual.
    superacao_negocial = False
    percentual_indice = float(res['variacao'])
    ciclo_negativo = percentual_indice < 0
    percentual_aplicado = 0.0 if ciclo_negativo else percentual_indice
    fator_ciclo_efetivo = 1.0 if ciclo_negativo else float(fator_ciclo)
    situacao_automatica = status_ped
    situacao_aplicada = status_ped
    tratamento_negativo = "Ciclo negativo - percentual aplicado 0,00% no acumulado" if ciclo_negativo else ""
    justificativa_negocial = ""
    referencia_documental = ""
    data_inicio_efeito_negocial = None

    if ciclo_negativo:
        st.warning(
            "A variação final apurada para o ciclo foi negativa. Para fins de composição acumulada, "
            "o percentual aplicado neste ciclo será tratado como 0,00%. Meses negativos isolados dentro "
            "do ciclo não são zerados; a regra somente se aplica quando o resultado final do ciclo é negativo."
        )
        situacao_aplicada = f"{status_ped} | 🔻 CICLO NEGATIVO (APLICADO 0,00%)"

    if "PRECLUSO" in status_ped.upper():
        st.warning(
            "O ciclo foi classificado automaticamente como precluso. "
            "Caso exista decisão negocial fundamentada na cláusula contratual aplicável, "
            "é possível registrar a admissão do reajuste por acordo entre as partes, sem apagar o diagnóstico automático."
        )
        with st.expander("Acordo negocial de admissão de reajuste", expanded=False):
            superacao_negocial = st.checkbox(
                "Ciclo admitido por negociação entre as partes",
                value=False,
                key="superacao_negocial_simples_c1",
            )
            if superacao_negocial:
                col_neg1, col_neg2 = st.columns(2)
                with col_neg1:
                    percentual_manual_pct = st.number_input(
                        "Percentual aplicado por acordo (%)",
                        min_value=0.0,
                        max_value=100.0,
                        value=round(percentual_indice * 100, 2),
                        step=0.01,
                        format="%.2f",
                        key="percentual_negocial_simples_c1",
                    )
                    data_inicio_efeito_negocial = st.date_input(
                        "Início dos efeitos financeiros por acordo",
                        value=dt_solic,
                        format="DD/MM/YYYY",
                        key="inicio_negocial_simples_c1",
                    )
                with col_neg2:
                    referencia_documental = st.text_input(
                        "Referência documental, se houver",
                        placeholder="Ex.: Despacho, Ata, Ofício ou Nota Técnica",
                        key="referencia_negocial_simples_c1",
                    )
                justificativa_negocial = st.text_area(
                    "Justificativa técnica/negocial",
                    placeholder="Registre a fundamentação da concessão por acordo negocial.",
                    key="justificativa_negocial_simples_c1",
                    height=90,
                )
                if not justificativa_negocial.strip():
                    st.info("A justificativa deve ser preenchida para fins de memória processual antes da instrução final.")
                percentual_aplicado = float(percentual_manual_pct) / 100
                fator_ciclo_efetivo = 1 + percentual_aplicado
                situacao_aplicada = "🟣 CICLO ADMITIDO POR NEGOCIAÇÃO ENTRE AS PARTES"
        if not superacao_negocial:
            percentual_aplicado = 0.0
            fator_ciclo_efetivo = 1.0
            if ciclo_negativo:
                situacao_aplicada = f"{status_ped} | 🔻 CICLO NEGATIVO (APLICADO 0,00%)"

    with st.expander("🔍 Memória de Cálculo Detalhada"):
        st.write(f"**Metodologia:** {res['metodo']}")
        if "IST" in tipo_idx:
            _render_equacao_ist(float(res['i_ini']), float(res['i_fim']), float(res['variacao']))
        else:
            st.dataframe(res['dados'])

    st.subheader("Relatório de Apuração")
    percentual_aplicado_fmt = f"{percentual_aplicado * 100:,.2f}%".replace('.', ',')
    observacao_ciclo_negativo = (
        "\n        Observação: a variação final do ciclo foi negativa; para composição acumulada, o percentual aplicado foi limitado a 0,00%."
        if ciclo_negativo and not superacao_negocial else ""
    )
    if superacao_negocial:
        inicio_negocial_txt = data_inicio_efeito_negocial.strftime('%d/%m/%Y') if data_inicio_efeito_negocial else dt_solic.strftime('%d/%m/%Y')
        relatorio_simples = f"""
        **{ciclo_label}:** Pedido realizado em {dt_solic.strftime('%d/%m/%Y')}. Intervalo do {ciclo_label}: {janela_str}.  
        Janela de Admissibilidade: {janela_adm_str}.  
        **Resultado automático:** {situacao_automatica}.  
        **Tratamento aplicado:** (*) ciclo admitido por negociação entre as partes.  
        Variação apurada pelo índice: {v_fmt}.  
        Percentual aplicado por acordo: {percentual_aplicado_fmt}.  
        Índice {tipo_idx}.  
        Data de início dos efeitos financeiros por acordo: {inicio_negocial_txt}.

        (*) O diagnóstico automático de preclusão foi preservado. O ciclo foi considerado aplicável por decisão negocial registrada pelo usuário.
        """
    else:
        relatorio_simples = f"""
        **{ciclo_label}:** Pedido realizado em {dt_solic.strftime('%d/%m/%Y')}. Intervalo do {ciclo_label}: {janela_str}.  
        Janela de Admissibilidade: {janela_adm_str}.  
        Resultado: {situacao_aplicada}.  
        Variação apurada pelo índice: {v_fmt}. Percentual aplicado no acumulado: {percentual_aplicado_fmt}.  
        Variação acumulada: {percentual_aplicado_fmt}.  
        Índice {tipo_idx}.  
        Data de início dos efeitos financeiros: {dt_solic.strftime('%d/%m/%Y')}.{observacao_ciclo_negativo}
        """
    st.info(relatorio_simples)

    inicio_efeito_financeiro = data_inicio_efeito_negocial if superacao_negocial and data_inicio_efeito_negocial else (dt_solic if dt_solic >= dt_aniv else dt_aniv)
    fim_efeito_financeiro = inicio_efeito_financeiro + relativedelta(months=11)

    ciclo_unico = {
        'ciclo': ciclo_label,
        'data_base': dt_base.strftime('%d/%m/%Y'),
        'intervalo_indice': janela_str,
        'janela_admissibilidade': janela_adm_str,
        'data_pedido': dt_solic.strftime('%d/%m/%Y'),
        'situacao': situacao_aplicada,
        'situacao_automatica': situacao_automatica,
        'situacao_aplicada': situacao_aplicada,
        'superacao_negocial': bool(superacao_negocial),
        'percentual_indice': float(percentual_indice),
        'percentual_aplicado': float(percentual_aplicado),
        'justificativa_negocial': justificativa_negocial.strip(),
        'referencia_documental': referencia_documental.strip(),
        'ciclo_negativo': bool(ciclo_negativo),
        'tratamento_ciclo_negativo': tratamento_negativo,
        'variacao': float(percentual_aplicado),
        'variacao_formatada': f"{percentual_aplicado*100:,.2f}%".replace('.', ','),
        'fator': float(fator_ciclo_efetivo),
        'fator_acumulado': float(fator_ciclo_efetivo),
        'periodo_inicio': _formatar_data(periodo_inicio),
        'periodo_fim': _formatar_data(periodo_fim),
        'financeiro_inicio': _formatar_data(inicio_efeito_financeiro),
        'financeiro_fim': _formatar_data(fim_efeito_financeiro),
    }

    st.session_state['dados_admissibilidade'] = {
        'origem': 'Reajuste Simples',
        'tipo': 'Simples',
        'indice': tipo_idx,
        'data_base_original': dt_base.strftime('%d/%m/%Y'),
        'contexto_contratual_anterior': contexto_contratual,
        'fator': float(fator_ciclo_efetivo),
        'fator_acumulado': float(fator_ciclo_efetivo),
        'variacao': float(percentual_aplicado),
        'variacao_acumulada': float(percentual_aplicado),
        'variacao_acumulada_formatada': f"{percentual_aplicado*100:,.2f}%".replace('.', ','),
        'ciclos': [ciclo_unico],
    }

    arquivo_coleta = gerar_arquivo_coleta_excel(st.session_state['dados_admissibilidade'])
    st.download_button(
        label="📥 Gerar Arquivo de Coleta",
        type="primary",
        data=arquivo_coleta,
        file_name="Coleta_Reajuste_Simples.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=False,
    )
