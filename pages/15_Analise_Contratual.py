"""
15_Analise_Contratual.py
------------------------
Página unificada de Análise Contratual — cl8us v2.0.

Consolida o fluxo de:
  - Upload e processamento da Coleta Única
  - Exibição dos resultados essenciais (4 métricas)
  - Geração de documentos (PDF, DOCX, XLSX)
  - Detalhes técnicos em expander (sob demanda)

Usa as funções de 03_Valor_Global.py e 04_Relatorio_Global.py
via exec parcial até o marcador # ── INICIO_UI ──, sem
executar o código de interface desses arquivos.
"""

from pathlib import Path
import sys
import html

import pandas as pd
import streamlit as st

from _ui_utils import render_marca_topo, render_aviso_privacidade
from _ponte_coleta_valores import processar_coleta, MODOS_COLETA
from _middle_layer_coleta import (
    normalizar_resultado,
    classificar_alertas,
    avaliar_qualidade,
    montar_df_valores_unitarios,
)

try:
    from _leitor_matriz_2_1 import ler_matriz_2_1 as _ler_m21, eh_matriz_2_1
    from _matriz21_resultado_adapter import adaptar_resultado_matriz21
    _TEM_LEITOR_M21 = True
except Exception:
    adaptar_resultado_matriz21 = None
    _TEM_LEITOR_M21 = False

try:
    from _leitor_coleta_mestre import ler_coleta_mestre as _ler_v10
    _TEM_LEITOR_V10 = True
except ImportError:
    _TEM_LEITOR_V10 = False

try:
    from _leitor_matriz_2_0 import ler_matriz_2_0, eh_matriz_2_0
    _TEM_LEITOR_M20 = True
except Exception:
    _TEM_LEITOR_M20 = False
    ler_matriz_2_0 = None
    def eh_matriz_2_0(_abas):
        return False

try:
    from _relatorio_sintese import render_relatorio_sintese
except Exception:
    render_relatorio_sintese = None


# Fluxo legado desativado nesta etapa experimental: a página Análise Contratual passa a aceitar apenas ColetaMestre v10.
_TEM_LEITOR_LEGADO = False


# ─────────────────────────────────────────────────────────────────────
# Carregamento das funções dos módulos existentes
# ─────────────────────────────────────────────────────────────────────

def _carregar_funcoes_modulo(caminho_relativo: str) -> dict:
    """
    Executa apenas a parte de funções de um módulo (até o marcador INICIO_UI),
    retornando o namespace resultante. Não executa código de UI.
    """
    caminho = Path(__file__).resolve().parent / caminho_relativo
    if not caminho.exists():
        return {}
    try:
        with open(caminho, "r", encoding="utf-8") as f:
            fonte = f.read()
        # Corta tudo a partir do marcador de UI
        partes = fonte.split("# ── INICIO_UI ──")
        fonte_funcoes = partes[0]
        env = {}
        exec(compile(fonte_funcoes, str(caminho), "exec"), env)
        return env
    except Exception as exc:
        st.warning(f"Não foi possível carregar funções de {caminho_relativo}: {exc}")
        return {}


@st.cache_resource(show_spinner=False)
def _env_valores():
    return _carregar_funcoes_modulo("03_Valor_Global.py")


@st.cache_resource(show_spinner=False)
def _env_relatorio():
    return _carregar_funcoes_modulo("04_Relatorio_Global.py")


# ─────────────────────────────────────────────────────────────────────
# Helpers de formatação
# ─────────────────────────────────────────────────────────────────────

def _moeda(valor):
    try:
        v = round(float(valor), 2)
        return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "—"


def _fator_fmt(valor):
    try:
        return f"{float(valor):,.4f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "—"


def _df_visual(df, moeda_cols=None):
    """Formata DataFrame para exibição — formata colunas monetárias."""
    if not isinstance(df, pd.DataFrame) or df.empty:
        return df
    df = df.copy()
    for col in (moeda_cols or []):
        if col in df.columns:
            def _fmt_moeda(v):
                try:
                    if v is None or (hasattr(v, "__class__") and v.__class__.__name__ == "NaTType"):
                        return "—"
                    f = float(v)
                    if f != f:  # NaN check
                        return "—"
                    return _moeda(f)
                except (TypeError, ValueError):
                    return str(v) if v is not None else "—"
            df[col] = df[col].apply(_fmt_moeda)
    return df


# ─────────────────────────────────────────────────────────────────────
# UI principal
# ─────────────────────────────────────────────────────────────────────

render_marca_topo(titulo_pagina="Análise Contratual", subtitulo_pagina="Análise e documentos")
render_aviso_privacidade(tem_upload=True, tem_download=True)

adm = st.session_state.get("dados_admissibilidade")
res = st.session_state.get("resultado_valor_global")

# ── Contexto da Admissibilidade ───────────────────────────────────────
if adm:
    _origem = adm.get("origem") or adm.get("tipo", "")
    _indice = adm.get("indice", "")
    _ciclos = len(adm.get("ciclos", []))
    _partes = [p for p in [_origem, _indice, f"{_ciclos} ciclo(s)" if _ciclos else ""] if p]
    st.caption("  ·  ".join(_partes))

# ── Upload e processamento ─────────────────────────────────────────────

st.caption(
    "Fluxo de análise: envie Coleta Matriz 2.1, Matriz 2.0 ou ColetaMestre v10 preenchida. "
    "A Matriz 2.1 é priorizada; Matriz 2.0 e v10 permanecem como compatibilidade."
)





# >>> INICIO_MATRIZ21_EXPERIMENTAL
try:
    from _matriz21_exp_ui import render_matriz21_experimental
    # Matriz 2.1 experimental ocultada no fluxo principal da release.
    # render_matriz21_experimental()
except Exception as _erro_m21_exp:
    import streamlit as _st_m21_exp
    _st_m21_exp.error(f"Erro no bloco experimental Matriz 2.1: {_erro_m21_exp}")
# <<< FIM_MATRIZ21_EXPERIMENTAL

arquivo_cu = st.file_uploader(
    "Coleta de Reajuste preenchida (.xlsx)",
    type=["xlsx"],
    key="ac_upload_cu_v10",
)

if arquivo_cu:
    bytes_cu = arquivo_cu.getvalue()
    _diag = None
    _abas = set()
    _eh_v10 = False

    try:
        import pandas as _pd
        from io import BytesIO as _BytesIO

        _xls = _pd.ExcelFile(_BytesIO(bytes_cu))
        _abas = set(_xls.sheet_names)

        # Cadeia oficial de detecção: Matriz 2.1 → Matriz 2.0 → v10/v11
        _eh_m21 = False
        _eh_m20 = False
        _eh_v10_assinatura = False
        _origem_upload = None

        if _TEM_LEITOR_M21 and eh_matriz_2_1(_abas):
            _eh_m21 = True
            _origem_upload = "matriz_2_1"
        else:
            _eh_m20 = bool((not _eh_m21) and _TEM_LEITOR_M20 and eh_matriz_2_0(_abas))
            if _eh_m20:
                _origem_upload = "matriz_2_0"
            else:
                _abas_obrigatorias = {"PARAMETROS_CONTRATO", "CICLOS"}
                _tem_base_financeira_ou_itens = bool(
                    {"FINANCEIRO", "FINANCEIRO_HISTORICO", "ITENS", "ITENS_CICLOS"} & _abas
                )
                _eh_v11 = bool({"FINANCEIRO_COMP", "ITENS_CICLO", "PARAMETROS_CICLOS"} & _abas)
                _eh_v10_assinatura = bool(
                    _eh_v11 or
                    (_abas_obrigatorias.issubset(_abas) and _tem_base_financeira_ou_itens)
                )
                if _eh_v10_assinatura:
                    _origem_upload = "v10"

        _nome_modelo_upload = "Matriz 2.1" if _eh_m21 else ("Matriz 2.0" if _eh_m20 else "ColetaMestre v10")
        _eh_v10 = bool(_origem_upload)

        if _eh_m21 and _TEM_LEITOR_M21:
            import tempfile, os
            _tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            _tmp.write(bytes_cu); _tmp.flush(); _tmp.close()
            try:
                _diag = _ler_m21(_tmp.name)
            finally:
                try: os.unlink(_tmp.name)
                except Exception: pass
        elif _eh_m20 and _TEM_LEITOR_M20:
            _diag = _ler_m20(bytes_cu)
        elif _eh_v10_assinatura and _TEM_LEITOR_V10:
            _diag = _ler_v10(bytes_cu)
    except Exception as exc:
        st.error(f"Não foi possível ler o arquivo XLSX: {exc}")
        _eh_v10 = False

    if not _eh_v10:
        st.error(
            "O arquivo enviado não foi reconhecido como Matriz 2.1, Matriz 2.0 ou ColetaMestre v10. "
            "Use o modelo oficial da coleta ou confira se o XLSX não foi alterado estruturalmente."
        )
    else:
        # ── Classificar alertas via middle layer ──────────────────
        _als              = classificar_alertas(_diag) if (_diag and _diag.get("ok")) else {"criticos": [], "atencao": [], "informativos": []}
        _alertas_criticos = _als["criticos"]
        _alertas_atencao  = _als["atencao"]
        _alertas_info     = _als["informativos"]
        _ajustes_gcc      = (_diag.get("ajustes_gcc", 0) if (_diag and _diag.get("ok")) else 0)

        # ── Linha de resumo v10 ───────────────────────────────────
        if _diag and _diag.get("ok"):
            _fin = _diag.get("financeiro", {})
            _its = _diag.get("itens", {})
            _n_fin = _fin.get("linhas_mensais_preenchidas", _fin.get("linhas_preenchidas", 0))
            _v_fin = _fin.get("total_com_efeito", _fin.get("total_pago", 0))
            _n_its = _its.get("itens_cadastrados", _its.get("linhas_preenchidas", 0))
            _n_cic = len(_diag.get("ciclos", []))

            st.markdown(
                f'<div style="font-size:0.88rem;color:#475569;margin:6px 0 2px 0;">'
                f'<strong style="color:#0F172A">{_nome_modelo_upload}</strong>'
                f' &nbsp;|&nbsp; <strong>Financeiro:</strong> {_n_fin} linhas — {_moeda(_v_fin)}'
                f' &nbsp;|&nbsp; <strong>Itens:</strong> {_n_its}'
                f' &nbsp;|&nbsp; <strong>Ciclos:</strong> {_n_cic}</div>',
                unsafe_allow_html=True,
            )
        else:
            st.info(
                "Arquivo com assinatura v10 reconhecida. O diagnóstico preliminar não foi carregado, "
                "mas o processamento principal poderá ser tentado."
            )

        # ── Alertas por nível ─────────────────────────────────────
        for _rv in _alertas_criticos:
            _txt_alerta = str(_rv).replace("⚠️", "").replace("❌", "").strip()
            st.info(_txt_alerta, icon="⚠️")
        for _rv in _alertas_atencao:
            _txt_alerta = str(_rv).replace("⚠️", "").replace("❌", "").strip()
            st.info(_txt_alerta, icon="ℹ️")



        # ── Roteiro das Informações dos Fiscais ───────────────────
        _roteiro_diag_info = (_diag or {}).get("roteiro_info_fiscais", {}) if isinstance(_diag, dict) else {}
        if _roteiro_diag_info.get("ok"):
            _roteiro_resumo = _roteiro_diag_info.get("resumo", {}) or {}
            _roteiro_modelo = _roteiro_resumo.get("modelo_sugerido") or _roteiro_diag_info.get("modelo_sugerido") or "Não informado"
            _roteiro_base = _roteiro_resumo.get("base_esperada") or _roteiro_diag_info.get("base_esperada") or "Não informada"
            _roteiro_obs = _roteiro_resumo.get("observacao_qualidade") or _roteiro_diag_info.get("observacao_qualidade") or "Roteiro identificado para apoiar o diagnóstico da base fiscal."
            st.markdown(
                f'<div style="background:#F0FDFA;border:1px solid #99F6E4;border-radius:14px;padding:11px 13px;margin:8px 0 10px 0;">'
                f'<div style="font-size:0.82rem;font-weight:800;color:#115E59;margin-bottom:4px;">Roteiro das Informações dos Fiscais identificado no XLS</div>'
                f'<div style="font-size:0.78rem;color:#134E4A;line-height:1.45;">'
                f'<strong>Modelo sugerido:</strong> {_roteiro_modelo}<br>'
                f'<strong>Base esperada:</strong> {_roteiro_base}<br>'
                f'<strong>Observação:</strong> {_roteiro_obs}'
                f'</div></div>',
                unsafe_allow_html=True,
            )

        # ── Modo declarado + seletor secundário ───────────────────
        _modo_declarado = ""
        _modo_detectado = ""

        if _diag and _diag.get("ok"):
            _modo_declarado = (_diag.get("parametros") or {}).get("modo_declarado", "")
            _modo_detectado = _diag.get("modo_detectado") or _diag.get("modo_preliminar", "")

        _modo_base = _modo_detectado or _modo_declarado
        _idx = MODOS_COLETA.index(_modo_base) if _modo_base in MODOS_COLETA else 0

        # Modo: apenas detecção automática — sem dropdown
        modo_confirmado = _modo_base or MODOS_COLETA[0]
        if _modo_detectado or _modo_declarado:
            st.markdown(
                f'<div style="margin:4px 0 10px 0;font-size:0.82rem;color:#64748B;">'
                f'Modo detectado automaticamente: '
                f'<strong style="color:#1E3A8A">{_modo_detectado or _modo_declarado}</strong>'
                f'</div>',
                unsafe_allow_html=True,
            )

        if st.button("Processar Planilha Master Reajuste", type="primary", key="ac_btn_cu_v10"):
            if _eh_m21:
                with st.spinner("Processando Matriz 2.1..."):
                    try:
                        if adaptar_resultado_matriz21 is None:
                            raise RuntimeError("Adaptador Matriz 2.1 indisponível (_matriz21_resultado_adapter.py).")
                        if False and (not _diag or not isinstance(_diag, dict) or not _diag.get("ok")):
                            _erro_m21 = (_diag or {}).get("erro") if isinstance(_diag, dict) else "diagnóstico inválido"
                            st.error(f"Matriz 2.1 reconhecida, mas o diagnóstico não está OK: {_erro_m21}")
                        else:
                            resultado = adaptar_resultado_matriz21(_diag)
                            resultado["_alertas_info"] = _alertas_info
                            resultado["_ajustes_gcc"] = _ajustes_gcc
                            resultado["origem_coleta"] = "matriz_2_1"
                            st.session_state["resultado_valor_global"] = resultado
                            st.session_state["modo_coleta_usado"] = resultado.get("modo_detectado", "Matriz 2.1")
                            st.session_state["resultado_matriz21_experimental_bruto"] = _diag
                            res = resultado
                            st.success("Matriz 2.1 processada e promovida para Análise e Documentos.")
                            st.rerun()
                    except Exception as exc:
                        st.error(f"Erro no processamento da Matriz 2.1: {exc}")
            else:
                # Bifurcacao: Matriz 2.0 tem processamento proprio; v10 usa processar_coleta
                if _eh_m20 if "_eh_m20" in dir() else False:
                    with st.spinner("Processando Matriz 2.0..."):
                        try:
                            # Matriz 2.0: resultado construido diretamente do _diag
                            _m20 = _diag or {}
                            resultado = {
                                "ok":                        _m20.get("ok", True),
                                "versao":                    "matriz_2_0",
                                "origem_coleta":             "matriz_2_0",
                                "modo_detectado":            _m20.get("modo_detectado",""),
                                "ciclos":                    _m20.get("ciclos", []),
                                "parametros":                _m20.get("parametros", {}),
                                "financeiro":                _m20.get("financeiro", {}),
                                "itens":                     _m20.get("itens", {}),
                                "aditivos":                  _m20.get("aditivos", {}),
                                "historico":                 _m20.get("historico", []),
                                "alertas":                   _m20.get("alertas", []),
                                "valor_atualizado_contrato": _m20.get("valor_atualizado_contrato", 0),
                                "valor_represado_a_pagar":   _m20.get("financeiro",{}).get("total_delta", 0),
                                "df_composicao_valor_total": _m20.get("df_composicao_valor_total"),
                                "memoria_vta_m20": _m20.get("memoria_vta_m20", []),
                                "fator_acumulado":           max((c.get("fator_acumulado",1) for c in _m20.get("ciclos",[])), default=1),
                                "variacao_acumulada":        0,
                                "_alertas_info":             _alertas_info,
                                "_ajustes_gcc":              _ajustes_gcc,
                            }
                            # Calcular variacao a partir do maior fator
                            fat = resultado["fator_acumulado"]
                            resultado["variacao_acumulada"] = round(fat - 1.0, 6)
                            st.session_state["resultado_valor_global"] = resultado
                            st.session_state["modo_coleta_usado"] = resultado["modo_detectado"]
                            res = resultado
                            st.success(f"Matriz 2.0 processada — modo: **{resultado['modo_detectado']}**")
                            st.rerun()
                        except Exception as exc:
                            st.error(f"Erro no processamento da Matriz 2.0: {exc}")
                else:
                    with st.spinner("Processando ColetaMestre v10..."):
                        try:
                            resultado = processar_coleta(bytes_cu, modo_confirmado)
                            if not resultado.get("ok"):
                                st.error(f"Erro: {resultado.get('erro', '')}")
                            else:
                                resultado["_alertas_info"] = _alertas_info
                                resultado["_ajustes_gcc"] = _ajustes_gcc
                                resultado["origem_coleta"] = "v10"
                                st.session_state["resultado_valor_global"] = resultado
                                st.session_state["modo_coleta_usado"] = modo_confirmado
                                res = resultado
                                st.success(f"Planilha processada — modo: **{modo_confirmado}**")
                                st.rerun()
                        except Exception as exc:
                            st.error(f"Erro no processamento da ColetaMestre v10: {exc}")


# ── Resultados ────────────────────────────────────────────────────────
if not res:
    st.divider()
    st.info("Carregue e processe uma Coleta Única para ver os resultados.")
    st.stop()

st.divider()

# 4 métricas principais
modo_ui = res.get("modo_apuracao", "Completo")
modo_reduzido  = str(modo_ui).lower() in ("reduzido por itens/estoque", "itens/estoque")
modo_consumo   = str(modo_ui).lower() in ("consumo por itens/ciclo",)

# Badge de modo + qualidade da base
_cor_modo = {"Completo": "#CCFBF1", "Financeiro Histórico": "#DBEAFE",
             "Reduzido por Itens/Estoque": "#EDE9FE", "Consumo por Itens/Ciclo": "#F6F3EE"}.get(modo_ui, "#F1F5F9")
_txt_modo = {"Completo": "#065F46", "Financeiro Histórico": "#1E3A8A",
             "Reduzido por Itens/Estoque": "#4C1D95", "Consumo por Itens/Ciclo": "#3F4F35"}.get(modo_ui, "#334155")
_score    = avaliar_qualidade(res, _diag if 'arquivo_cu' in dir() and arquivo_cu else None)
_dots_html = "".join(
    f'<span style="display:inline-block;width:8px;height:8px;border-radius:50%;'
    f'background:{"#1D9E75" if i < _score else "#D3D1C7"};margin-left:3px;"></span>'
    for i in range(5)
)
st.markdown(
    f'<div style="display:flex;align-items:center;gap:12px;margin-bottom:16px;">'
    f'<div style="background:{_cor_modo};color:{_txt_modo};border-radius:20px;'
    f'padding:5px 16px;font-weight:600;font-size:0.78rem;letter-spacing:0.02em;">Modo: {modo_ui}</div>'
    f'<div style="display:flex;align-items:center;gap:3px;">{_dots_html}</div>'
    f'</div>',
    unsafe_allow_html=True,
)

val_total     = res.get("valor_atualizado_contrato", res.get("valor_global_estoque", 0))
val_retro     = res.get("valor_represado_a_pagar", res.get("valor_retroativo_consumo", 0))

# >>> RELATORIO_SINTESE_M20_RENDER_V2
try:
    if render_relatorio_sintese is not None and isinstance(res, dict) and res:
        with st.expander("📄 Relatório Síntese da Apuração", expanded=False):
            try:
                _diag_rel_sintese = locals().get("_diag", None)
            except Exception:
                _diag_rel_sintese = None
            render_relatorio_sintese(res, _diag_rel_sintese)
    elif render_relatorio_sintese is None:
        st.info("Relatório Síntese indisponível: módulo _relatorio_sintese.py não carregado.")
except Exception as _e_rel_sintese:
    st.warning(f"Relatório Síntese indisponível: {_e_rel_sintese}")
# <<< RELATORIO_SINTESE_M20_RENDER_V2

fator_acum    = res.get("fator_acumulado", 1.0)
indice_res    = res.get("indice", adm.get("indice", "—") if adm else "—")

def _metric_card(col, label, value, destaque=False):
    """Card visual compacto — design refinado, sem overflow."""
    bg     = "#ECFDF5" if destaque else "#F8FAFC"
    borda  = "#6EE7B7" if destaque else "#E2E8F0"
    cor_lb = "#059669" if destaque else "#64748B"
    cor_vl = "#065F46" if destaque else "#0F172A"
    label_safe = html.escape(str(label))
    value_safe = html.escape(str(value))
    with col:
        st.markdown(
            f"""
            <div style="background:{bg}; border:1.5px solid {borda}; border-radius:16px;
                        padding:14px 16px; min-height:90px; overflow:hidden; box-sizing:border-box;">
                <div style="font-size:0.70rem; color:{cor_lb}; font-weight:600;
                            letter-spacing:0.04em; text-transform:uppercase;
                            line-height:1.2; margin-bottom:8px; white-space:nowrap;
                            overflow:hidden; text-overflow:ellipsis;">
                    {label_safe}
                </div>
                <div style="font-size:clamp(0.88rem, 1.6vw, 1.35rem); color:{cor_vl};
                            font-weight:700; letter-spacing:-0.02em; line-height:1.15;
                            word-break:break-word; overflow-wrap:anywhere;
                            hyphens:auto; overflow:hidden;"
                     title="{value_safe}">
                    {value_safe}
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

st.markdown(
    '<p style="font-size:0.72rem;font-weight:600;letter-spacing:0.08em;'
    'text-transform:uppercase;color:#94A3B8;margin:8px 0 6px 0;">Resultado consolidado</p>',
    unsafe_allow_html=True
)
m1, m2, m3, m4 = st.columns(4)
_metric_card(m1, "Valor Total Atualizado", _moeda(val_total), destaque=True)
_metric_card(m2, "Retroativo", _moeda(val_retro))
_metric_card(m3, "% Acumulado", f"{round((fator_acum - 1) * 100, 2):.2f}%".replace(".",","))
_metric_card(m4, "Índice", indice_res)

if modo_reduzido:
    st.warning("Modo Reduzido: retroativo estimado por itens/estoque. "
               "Não substitui apuração financeira definitiva.", icon="⚠️")
elif modo_consumo:
    st.warning("Modo Consumo: retroativo apurado por quantitativos consumidos/ciclo. "
               "Válido quando a fiscalização confirma equivalência com faturamento devido.", icon="⚠️")

# ── Gerar Documentos ──────────────────────────────────────────────────
st.divider()
st.markdown(
    '<p style="font-size:0.72rem;font-weight:600;letter-spacing:0.08em;'
    'text-transform:uppercase;color:#94A3B8;margin-bottom:0px;">Documentos</p>',
    unsafe_allow_html=True
)

# Vigência — lida da ColetaMestre ou solicitada dentro do botão Adequação
_params_v10  = res.get("params_v10") or res.get("params") or {}
_vig_raw     = _params_v10.get("vigencia_final", "")
_vig_default = None
if _vig_raw:
    import pandas as _pd_vig
    try:
        _vig_default = _pd_vig.to_datetime(str(_vig_raw)).date()
    except Exception:
        pass
if _vig_default:
    st.caption(f"Vigência final: {_vig_default.strftime('%d/%m/%Y')} (lida da ColetaMestre)")
_vig_input = _vig_default  # será sobrescrito pelo form da Adequação se necessário

st.markdown("""
<style>
/* ── Botões de ação (gerar documento) ── */
div[data-testid="stButton"] > button {
    border-radius: 14px !important;
    border: 1.5px solid #E2E8F0 !important;
    background: #FFFFFF !important;
    color: #1E293B !important;
    font-size: 0.88rem !important;
    font-weight: 500 !important;
    padding: 10px 14px !important;
    line-height: 1.4 !important;
    transition: background 0.12s, border-color 0.12s !important;
    height: auto !important;
    min-height: 56px !important;
    white-space: normal !important;
    text-align: left !important;
}
div[data-testid="stButton"] > button:hover {
    background: #EFF6FF !important;
    border-color: #93C5FD !important;
    color: #1E3A8A !important;
}
/* ── Botão de download — verde escuro ── */
div[data-testid="stDownloadButton"] > button {
    background-color: #1a5c38 !important;
    color: #ffffff !important;
    border: none !important;
    border-radius: 14px !important;
    font-weight: 600 !important;
    font-size: 0.88rem !important;
}
div[data-testid="stDownloadButton"] > button:hover {
    background-color: #14472b !important;
}
/* ── Sub-botões de forms (Orçamentária e Garantia) — verde ── */
div[data-testid="stForm"] div[data-testid="stButton"] > button[kind="primaryFormSubmit"],
div[data-testid="stForm"] button[kind="primaryFormSubmit"] {
    background-color: #1a5c38 !important;
    color: #ffffff !important;
    border: none !important;
    border-radius: 14px !important;
    font-weight: 600 !important;
}
/* ── Botão primário (processar) ── */
div[data-testid="stButton"] > button[kind="primary"] {
    background: #1a5c38 !important;
    color: #ffffff !important;
    border: none !important;
    border-radius: 14px !important;
}
/* ── Espaçamento entre seções ── */
div[data-testid="stVerticalBlock"] > div { gap: 0.5rem; }
/* ── Caption discreta ── */
div[data-testid="stCaptionContainer"] p {
    font-size: 0.78rem !important;
    color: #94A3B8 !important;
}
</style>
""", unsafe_allow_html=True)

_env_rel = _env_relatorio()
res_doc  = normalizar_resultado(res)

_bc1, _bc2, _bc3, _bc4 = st.columns(4)
_btn_pdf   = _bc1.button("Relatório Executivo (PDF)",       key="ac_btn_pdf",   use_container_width=True)
_btn_docx  = _bc2.button("Minuta de Apostilamento (DOCX)",  key="ac_btn_docx",  use_container_width=True)
_btn_sum   = _bc3.button("Sumário Executivo (XLSX)",         key="ac_btn_sum",   use_container_width=True)
_btn_itens = _bc4.button("Itens por Ciclo (XLSX)",           key="ac_btn_itens", use_container_width=True)

_bd1, _bd2, _bd3, _bd4 = st.columns(4)
_btn_saneador  = _bd1.button("Despacho Saneador (DOCX)",    key="ac_btn_saneador",  use_container_width=True)
_btn_orc       = _bd2.button("Previsão Orçamentária (DOCX)", key="ac_btn_orc",       use_container_width=True)
_btn_garantia  = _bd3.button("Garantia Contratual (XLSX)",   key="ac_btn_garantia",  use_container_width=True)
_btn_dou       = _bd4.button("Extrato DOU (DOCX)",           key="ac_btn_dou",       use_container_width=True)

# ── Botão ZIP — diferenciado visualmente ────────────────────────────────
st.markdown("""
<style>
div[data-testid="stButton"] button[kind="primary"].zip-btn,
div.zip-row div[data-testid="stButton"] > button {
    background: #1F4E78 !important;
    color: #ffffff !important;
    border: none !important;
    border-radius: 14px !important;
    font-weight: 600 !important;
    font-size: 0.92rem !important;
    letter-spacing: 0.02em !important;
}
</style>
""", unsafe_allow_html=True)

st.markdown('<div class="zip-row">', unsafe_allow_html=True)
_bz1, _bz2, _bz3 = st.columns([2, 1, 1])
_btn_zip = _bz1.button(
    "Baixar todos os documentos (ZIP)",
    key="ac_btn_zip",
    use_container_width=True,
    type="primary",
    help="Gera e compacta os 7 documentos em um único arquivo ZIP para download."
)
st.markdown('</div>', unsafe_allow_html=True)

# Limpar painel ativo quando qualquer botão de documento for clicado
_botoes_doc = [_btn_pdf,_btn_docx,_btn_sum,_btn_itens,_btn_saneador,_btn_dou,_btn_zip]
if any(_botoes_doc):
    st.session_state["_painel_ativo"] = None

if _btn_zip:
    with st.spinner("Gerando todos os documentos..."):
        try:
            import zipfile as _zf
            from io import BytesIO as _ZBio
            _zip_buf = _ZBio()
            _erros_zip = []
            _env_rel_z  = _env_relatorio()
            _env_val_z  = _env_valores()
            _res_doc_z  = normalizar_resultado(res)
            _df_ciclos_z = _res_doc_z.get("df_ciclos", res.get("df_ciclos"))
            _df_vu_z     = montar_df_valores_unitarios(_res_doc_z)
            with _zf.ZipFile(_zip_buf, 'w', _zf.ZIP_DEFLATED) as _zzip:
                _docs_zip = [
                    ("criar_pdf_relatorio",                  None,         "Relatorio_Executivo.pdf"),
                    ("gerar_minuta_apostilamento_docx",      None,         "Minuta_Apostilamento.docx"),
                    ("gerar_despacho_saneador_docx",         None,         "Despacho_Saneador.docx"),
                    ("gerar_previsao_orcamentaria_docx",     None,         "Previsao_Orcamentaria.docx"),
                    ("gerar_publicacao_dou_docx",            None,         "Extrato_DOU.docx"),
                ]
                for _fn, _, _fname in _docs_zip:
                    try:
                        _f = _env_rel_z.get(_fn)
                        if _f:
                            _zzip.writestr(_fname, _f(adm, _res_doc_z))
                    except Exception as _ez:
                        _erros_zip.append(f"{_fname}: {_ez}")
                # Sumário XLSX
                try:
                    _fp = _env_val_z.get("gerar_planilha_executiva")
                    if _fp:
                        _zzip.writestr("Sumario_Executivo.xlsx", _fp(_res_doc_z))
                except Exception as _ez:
                    _erros_zip.append(f"Sumario_Executivo.xlsx: {_ez}")
                # Itens XLSX
                try:
                    _fi = _env_val_z.get("gerar_excel_valores_unitarios_por_ciclo")
                    if _fi and isinstance(_df_vu_z, pd.DataFrame) and not _df_vu_z.empty and "Ciclo" in _df_vu_z.columns:
                        _zzip.writestr("Itens_Por_Ciclo.xlsx", _fi(_df_vu_z, _df_ciclos_z))
                except Exception as _ez:
                    _erros_zip.append(f"Itens_Por_Ciclo.xlsx: {_ez}")
            _zip_buf.seek(0)
            if _erros_zip:
                st.info("ZIP gerado com ressalvas: " + " | ".join(_erros_zip), icon="ℹ️")
            st.download_button(
                "Baixar ZIP completo",
                data=_zip_buf.getvalue(),
                file_name="cl8us_documentos_completos.zip",
                mime="application/zip",
                key="ac_dl_zip",
            )
        except Exception as exc:
            st.error(f"Erro ao gerar ZIP: {exc}")

elif _btn_pdf:
    _criar_pdf = _env_rel.get("criar_pdf_relatorio")
    if not _criar_pdf:
        st.error("Função de PDF não carregada. Verifique 04_Relatorio_Global.py.")
    else:
        with st.spinner("Gerando PDF..."):
            try:
                pdf_bytes = _criar_pdf(adm, res_doc)
                st.download_button(
                    "Baixar Relatório Executivo",
                    data=pdf_bytes,
                    file_name="Relatorio_Executivo.pdf",
                    mime="application/pdf",
                    key="ac_dl_pdf",
                )
            except Exception as exc:
                st.error(f"Erro ao gerar PDF: {exc}")

elif _btn_docx:
    _gerar_docx = _env_rel.get("gerar_minuta_apostilamento_docx")
    if not _gerar_docx:
        st.error("Função de DOCX não carregada. Verifique 04_Relatorio_Global.py.")
    else:
        with st.spinner("Gerando minuta..."):
            try:
                docx_bytes = _gerar_docx(adm, res_doc)
                st.download_button(
                    "Baixar Minuta de Apostilamento",
                    data=docx_bytes,
                    file_name="Minuta_Apostilamento.docx",
                    mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                    key="ac_dl_docx",
                )
            except Exception as exc:
                st.error(f"Erro ao gerar minuta: {exc}")

elif _btn_sum:
    _gerar_planilha = _env_valores().get("gerar_planilha_executiva")
    if not _gerar_planilha:
        st.error("Função não carregada. Verifique 03_Valor_Global.py.")
    else:
        with st.spinner("Gerando sumário..."):
            try:
                xlsx_bytes = _gerar_planilha(res_doc)
                # >>> AJUSTE_SUMARIO_EXECUTIVO_M20_V2
                try:
                    from _sumario_m20_utils import ajustar_sumario_executivo_m20
                    _res_base_sumario_m20 = res_doc if 'res_doc' in locals() else res
                    xlsx_bytes = ajustar_sumario_executivo_m20(xlsx_bytes, _res_base_sumario_m20)
                except Exception as _e_sumario_m20:
                    st.caption(f"Ajuste visual Matriz 2.0 do Sumário Executivo não aplicado: {_e_sumario_m20}")
                # <<< AJUSTE_SUMARIO_EXECUTIVO_M20_V2
                st.download_button(
                    "Baixar Sumário Executivo",
                    data=xlsx_bytes,
                    file_name="Sumario_Executivo.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="ac_dl_sumario",
                )
            except Exception as exc:
                st.error(f"Erro ao gerar sumário: {exc}")

elif _btn_itens:
    _env_val     = _env_valores()
    _gerar_itens = _env_val.get("gerar_excel_valores_unitarios_por_ciclo")
    if not _gerar_itens:
        st.error("Função não carregada. Verifique 03_Valor_Global.py.")
    else:
        with st.spinner("Gerando planilha de itens..."):
            try:
                df_ciclos = res_doc.get("df_ciclos", res.get("df_ciclos"))
                df_vu     = montar_df_valores_unitarios(res_doc)

                if not isinstance(df_vu, pd.DataFrame) or df_vu.empty:
                    st.info(
                        "Planilha de itens por ciclo indisponível: a Coleta processada não gerou dados "
                        "itemizados suficientes. Confira se a aba ITENS possui valor unitário original e "
                        "remanescente por ciclo."
                    )
                elif "Ciclo" not in df_vu.columns:
                    st.info(
                        "Planilha de itens por ciclo indisponível: a tabela de itens não contém a coluna "
                        "'Ciclo'. Reprocesse a coleta ou confira a aba ITENS."
                    )
                else:
                    xlsx_itens = _gerar_itens(df_vu, df_ciclos)
                    st.download_button(
                        "Baixar Itens por Ciclo",
                        data=xlsx_itens,
                        file_name="Itens_Por_Ciclo.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="ac_dl_itens",
                    )
            except Exception as exc:
                st.error(f"Erro ao gerar itens: {exc}")

elif _btn_saneador:
    _gerar_saneador = _env_rel.get("gerar_despacho_saneador_docx")
    if not _gerar_saneador:
        st.error("Função não carregada. Verifique 04_Relatorio_Global.py.")
    else:
        with st.spinner("Gerando despacho saneador..."):
            try:
                docx_bytes = _gerar_saneador(adm, res_doc)
                st.download_button(
                    "Baixar Despacho Saneador",
                    data=docx_bytes,
                    file_name="Despacho_Saneador.docx",
                    mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                    key="ac_dl_saneador",
                )
            except Exception as exc:
                st.error(f"Erro ao gerar despacho saneador: {exc}")

elif _btn_orc:
    st.session_state["_painel_ativo"] = "orc"
    st.rerun()

if st.session_state.get("_painel_ativo") == "orc" and res:
    import datetime as _dt_orc
    _vig_orc_default = _vig_default or _dt_orc.date.today().replace(year=_dt_orc.date.today().year + 1)
    with st.form(key="ac_form_orc"):
        st.caption("Informe a data de término da vigência contratual para calcular os meses futuros.")
        _vig_orc = st.date_input(
            "Data final da vigência",
            value=_vig_orc_default,
            format="DD/MM/YYYY",
            key="ac_vig_orc_input",
        )
        _submit_orc = st.form_submit_button("Gerar Previsão Orçamentária (DOCX)", type="primary")
    if _submit_orc:
        _gerar_orc = _env_rel.get("gerar_previsao_orcamentaria_docx")
        if not _gerar_orc:
            st.error("Função não carregada. Verifique 04_Relatorio_Global.py.")
        else:
            with st.spinner("Gerando previsão orçamentária..."):
                try:
                    # Injetar vigência escolhida no resultado
                    _res_orc = dict(res_doc)
                    if not _res_orc.get("params_v10"):
                        _res_orc["params_v10"] = {}
                    _res_orc["params_v10"]["vigencia_final"] = str(_vig_orc)
                    docx_bytes = _gerar_orc(adm, _res_orc)
                    st.download_button(
                        "Baixar Previsão Orçamentária",
                        data=docx_bytes,
                        file_name="Previsao_Orcamentaria.docx",
                        mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                        key="ac_dl_orc",
                    )
                except Exception as exc:
                    st.error(f"Erro ao gerar previsão orçamentária: {exc}")

elif _btn_dou:
    _gerar_dou = _env_rel.get("gerar_publicacao_dou_docx")
    if not _gerar_dou:
        st.error("Função não carregada. Verifique 04_Relatorio_Global.py.")
    else:
        with st.spinner("Gerando extrato DOU..."):
            try:
                docx_bytes = _gerar_dou(adm, res_doc)
                st.download_button(
                    "Baixar Extrato DOU",
                    data=docx_bytes,
                    file_name="Extrato_DOU.docx",
                    mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                    key="ac_dl_dou",
                )
            except Exception as exc:
                st.error(f"Erro ao gerar extrato DOU: {exc}")


elif _btn_garantia:
    st.session_state["_painel_ativo"] = "garantia"
    st.rerun()

if st.session_state.get("_painel_ativo") == "garantia" and res:
    # ── Garantia — Método Delta (cálculo automático + XLSX) ─────────
    import datetime as _dt_gar
    from dateutil.relativedelta import relativedelta as _reldt

    _val_total_gar  = float(res.get("valor_atualizado_contrato", res.get("valor_global_estoque", 0)) or 0)
    _params_gar     = res.get("params_v10") or res.get("params") or {}
    _val_orig_gar   = float(_params_gar.get("valor_original_do_contrato", res.get("valor_original_contrato", 0)) or 0)
    _val_form_ant   = float(_params_gar.get("valor_formalizado_antes_desta_analise", 0) or 0)
    _sugestao_base  = max(_val_total_gar - (_val_form_ant if _val_form_ant > 0 else _val_orig_gar), 0.0)
    _vig_raw_gar    = _params_gar.get("vigencia_final", "")
    _vig_def_gar    = _dt_gar.date.today().replace(year=_dt_gar.date.today().year + 1)
    if _vig_raw_gar:
        try:
            import pandas as _pd_gar
            _vig_def_gar = _pd_gar.to_datetime(str(_vig_raw_gar)).date()
        except Exception:
            pass

    _pct_gar = 5.0  # percentual padrão — raras vezes alterado

    with st.form(key="ac_form_garantia"):
        st.caption("Controle de garantia — valores pré-preenchidos. Informe apenas a garantia vigente e sua validade.")
        _gf1, _gf2 = st.columns(2)
        _gar_vigente  = _gf1.number_input("Garantia hoje vigente (R$)", min_value=0.0, value=0.0, step=100.0, format="%.2f")
        _gar_valid    = _gf2.date_input("Validade da garantia hoje", value=_vig_def_gar, format="DD/MM/YYYY")
        _submit_gar   = st.form_submit_button("Calcular e Baixar XLSX", type="primary")

    if _submit_gar:
        with st.spinner("Calculando garantia..."):
            try:
                import openpyxl as _oxl
                from openpyxl.styles import (
                    Font as _Fnt, PatternFill as _Fil,
                    Alignment as _Aln, Border as _Brd, Side as _Sid
                )
                from io import BytesIO as _Bio

                _pct_dec      = _pct_gar / 100.0
                _fat_gar      = float(res.get("fator_acumulado", 1.0) or 1.0)
                _var_gar      = float(res.get("variacao_acumulada", _fat_gar - 1.0) or 0)
                _vig_ini_raw  = _params_gar.get("vigencia_inicial", "")
                _vig_ini_gar  = None
                if _vig_ini_raw:
                    try:
                        import pandas as _pd_gi
                        _vig_ini_gar = _pd_gi.to_datetime(str(_vig_ini_raw)).date()
                    except Exception:
                        pass
                _ciclos_gar   = adm.get("ciclos", []) if adm else []
                _nomes_gar    = ", ".join(
                    str(c.get("ciclo","")).strip().upper() for c in _ciclos_gar if isinstance(c, dict)
                ) or "[campo a preencher]"
                _gar_exigida  = round(_val_total_gar * _pct_dec, 2)
                _complemento  = round(max(_gar_exigida - _gar_vigente, 0.0), 2)
                _val_min_gar  = _vig_def_gar + _reldt(months=3)
                _cobertura_ok = _gar_valid >= _val_min_gar
                if _complemento == 0 and _cobertura_ok:
                    _status = "OK"
                elif _complemento == 0:
                    _status = "ATUALIZAR VALIDADE"
                else:
                    _status = "ATUALIZAR / ENDOSSAR"

                def _m(v):
                    return f"R$ {float(v or 0):,.2f}".replace(",","X").replace(".",",").replace("X",".")

                _AZUL  = "1F4E78"; _BRAN = "FFFFFF"
                _CLAR  = "D9E2F3"; _CINZ = "F1F5F9"
                _VERD  = "E2EFDA"; _WARN = "FFF3CD"; _VERM = "FDECEA"
                _AZUL2 = "EBF3FB"

                def _brd_all():
                    s = _Sid(style="thin", color="D9E2F3")
                    return _Brd(left=s, right=s, top=s, bottom=s)

                def _hdr(ws, r, txt, ncols=6):
                    ws.cell(r,1).value = txt
                    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
                    for col in range(1, ncols+1):
                        c = ws.cell(r, col)
                        c.fill = _Fil("solid", fgColor=_AZUL)
                        c.font = _Fnt(color=_BRAN, bold=True, size=10)
                        c.alignment = _Aln(horizontal="center", vertical="center")
                        c.border = _brd_all()
                    ws.row_dimensions[r].height = 18

                def _cel(ws, r, c_idx, val, bold=False, fill=_BRAN, color="000000", fmt=None, align="left"):
                    c = ws.cell(r, c_idx, val)
                    c.font = _Fnt(bold=bold, size=10, color=color)
                    c.fill = _Fil("solid", fgColor=fill)
                    c.alignment = _Aln(horizontal=align, vertical="center")
                    c.border = _brd_all()
                    if fmt: c.number_format = fmt
                    ws.row_dimensions[r].height = 16

                wb = _oxl.Workbook()
                ws = wb.active
                ws.title = "Garantias"
                ws.sheet_view.showGridLines = False
                for col, w in [("A",38),("B",18),("C",4),("D",32),("E",4),("F",22)]:
                    ws.column_dimensions[col].width = w

                r = 1
                _hdr(ws, r, "cl8us · Controle de Garantia Contratual"); r+=1
                _cel(ws,r,1,"Preencha os campos em azul se necessário. Valores pré-preenchidos pelo sistema.",
                     fill=_CLAR, color=_AZUL, align="left"); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6); r+=1
                r+=1

                # Parâmetros | Resumo
                _cel(ws,r,1,"Parâmetros iniciais",bold=True,fill=_CLAR,color=_AZUL); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=2)
                _cel(ws,r,3,"",fill=_BRAN)
                _cel(ws,r,4,"Resumo automático",bold=True,fill=_CLAR,color=_AZUL); ws.merge_cells(start_row=r,start_column=4,end_row=r,end_column=6); r+=1

                _pairs = [
                    ("Valor original do contrato", _m(_val_orig_gar),   "Valor contratual atual",       _m(_val_total_gar)),
                    ("Percentual da garantia",     f"{_pct_gar:.2f}%".replace(".",","), "Garantia exigida atual", _m(_gar_exigida)),
                    ("Início da vigência",         _vig_ini_gar.strftime("%d/%m/%Y") if _vig_ini_gar else "[campo a preencher]", "Garantia vigente considerada", _m(_gar_vigente)),
                    ("Fim da vigência",            _vig_def_gar.strftime("%d/%m/%Y"), "Complemento/endosso necessário", _m(_complemento)),
                    ("Garantia hoje vigente (R$)", _m(_gar_vigente),    "Vigência contratual",          f"{_vig_ini_gar.strftime('%d/%m/%Y') if _vig_ini_gar else '?'} a {_vig_def_gar.strftime('%d/%m/%Y')}"),
                    ("Validade da garantia hoje",  _gar_valid.strftime("%d/%m/%Y"), "Cobertura mínima exigida", _val_min_gar.strftime("%d/%m/%Y")),
                    ("", "", "Status", _status),
                ]
                for lbl1, v1, lbl2, v2 in _pairs:
                    _fil_status = _VERD if _status=="OK" else (_WARN if _status=="ATUALIZAR VALIDADE" else _VERM)
                    _is_status  = lbl2 == "Status"
                    _cel(ws,r,1,lbl1,fill=_CINZ,color="334155")
                    _cel(ws,r,2,v1,bold=True,fill=_AZUL2,color=_AZUL,align="right")
                    _cel(ws,r,3,"",fill=_BRAN)
                    _cel(ws,r,4,lbl2,fill=_CINZ,color="334155")
                    _cel(ws,r,5,"",fill=_BRAN)
                    _cel(ws,r,6,v2,bold=_is_status,fill=_fil_status if _is_status else _AZUL2,
                         color=_AZUL if not _is_status else "000000",align="right")
                    r+=1
                r+=1

                # Tabela de eventos
                _hdr(ws, r, "Histórico de eventos"); r+=1
                for i, (hdr, col, w_) in enumerate([
                    ("Nº",1,None),("Data do evento",2,None),("Tipo",None,None),
                    ("Descrição",4,None),("Variação (R$)",5,None),("Valor após evento",6,None)
                ]):
                    _cel(ws,r,i+1,hdr,bold=True,fill=_CLAR,color=_AZUL,align="center")
                r+=1
                # Linha do reajuste pré-preenchida
                _var_abs = round(_val_total_gar - _val_orig_gar, 2)
                import datetime as _dt2
                _cel(ws,r,1,"1",align="center",fill=_AZUL2,color=_AZUL)
                _cel(ws,r,2,_dt2.date.today().strftime("%d/%m/%Y"),fill=_AZUL2,color=_AZUL)
                _cel(ws,r,3,"Reajuste",fill=_AZUL2,color=_AZUL)
                _cel(ws,r,4,_nomes_gar,fill=_AZUL2,color=_AZUL)
                _cel(ws,r,5,_m(_var_abs),fill=_AZUL2,color=_AZUL,align="right")
                _cel(ws,r,6,_m(_val_total_gar),fill=_AZUL2,color=_AZUL,align="right")
                r+=1
                # Linhas em branco
                for i in range(2, 11):
                    _cel(ws,r,1,str(i),align="center")
                    for col in range(2,7): _cel(ws,r,col,"")
                    r+=1

                buf = _Bio()
                wb.save(buf)
                buf.seek(0)
                st.download_button(
                    "Baixar Garantia (XLSX)",
                    data=buf.getvalue(),
                    file_name="Garantia_Contratual.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="ac_dl_garantia",
                )
            except Exception as exc:
                st.error(f"Erro ao gerar garantia: {exc}")


# ── Detalhes Técnicos (expansível) ────────────────────────────────────
st.divider()
with st.expander("▼ Detalhes técnicos da apuração", expanded=False):

    # Alertas informativos (rastreabilidade — não são erros)
    _info_list = res.get("_alertas_info", [])
    _ajst_n    = res.get("_ajustes_gcc", 0)
    if _info_list or _ajst_n:
        st.markdown("##### Rastreabilidade")
        for _inf in _info_list:
            # Remover emojis duplicados e mostrar como info
            _txt = str(_inf).replace("⚠️","").replace("ℹ️","").strip()
            st.info(_txt, icon="ℹ️")
        if _ajst_n and not any("ajuste" in str(x).lower() for x in _info_list):
            st.info(f"{_ajst_n} célula(s) com ajuste manual GCC registrado(s).", icon="ℹ️")

    st.markdown("##### Metodologia de corte")
    _env_rel2 = _env_relatorio()
    _aviso_corte = _env_rel2.get("aviso_metodologia_corte_html")
    if _aviso_corte:
        try:
            st.markdown(_aviso_corte(res), unsafe_allow_html=True)
        except Exception:
            pass

    st.markdown("##### Quadro executivo por ciclo")
    df_comp = res.get("df_comparativo", res.get("df_financeiro_por_ciclo"))
    if isinstance(df_comp, pd.DataFrame) and not df_comp.empty:
        _moeda_cols_exec = [c for c in [
            "Valor", "Antes do Reajuste", "Após Reajuste", "Diferença",
            "Valor pago efetivo", "Valor teórico calculado", "Delta do ciclo",
            "Valor pago/faturado", "Valor devido reajustado", "Delta acumulado",
        ] if c in df_comp.columns]
        st.dataframe(
            _df_visual(df_comp, moeda_cols=_moeda_cols_exec),
            use_container_width=True, hide_index=True,
        )
    else:
        st.info("Quadro executivo não disponível.")

    st.markdown("##### Financeiro por ciclo")
    df_fin_ciclo = res.get("df_financeiro_por_ciclo")
    if isinstance(df_fin_ciclo, pd.DataFrame) and not df_fin_ciclo.empty:
        _moeda_cols_fin = [c for c in [
            "Valor pago efetivo","Valor teórico calculado","Delta do ciclo",
            "Valor pago/faturado","Valor devido reajustado","Delta acumulado",
        ] if c in df_fin_ciclo.columns]
        st.dataframe(
            _df_visual(df_fin_ciclo, moeda_cols=_moeda_cols_fin),
            use_container_width=True, hide_index=True,
        )
    else:
        st.info("Financeiro por ciclo não disponível.")

    st.markdown("##### Composição do Valor Total Atualizado")
    df_composicao = res.get("df_composicao_valor_total")
    if isinstance(df_composicao, pd.DataFrame) and not df_composicao.empty:
        st.dataframe(
            _df_visual(df_composicao, moeda_cols=["Valor"]),
            use_container_width=True, hide_index=True,
        )
    else:
        st.info("Composição do valor total não disponível.")

    df_ad = res.get("df_aditivos_executivo", res.get("df_aditivos"))
    if isinstance(df_ad, pd.DataFrame) and not df_ad.empty:
        st.markdown("##### Aditivos e Supressões")
        st.caption("Quadro de controle formal. Valores não somados automaticamente ao total quando já incorporados.")
        cols_ad = [c for c in [
            "Aditivo","Ciclo/Marco","Tratamento do aditivo",
            "Valor do aditivo na assinatura","Fator aplicado","Valor do aditivo reajustado",
        ] if c in df_ad.columns]
        st.dataframe(
            _df_visual(df_ad[cols_ad] if cols_ad else df_ad,
                       moeda_cols=["Valor do aditivo na assinatura","Valor do aditivo reajustado"]),
            use_container_width=True, hide_index=True,
        )



