
from io import BytesIO

try:
    import pandas as pd
except Exception:
    pd = None

try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
except Exception:
    load_workbook = None


AZUL = "1F4E79"
VERDE = "D9EAD3"
BRANCO = "FFFFFF"
BORDA = Side(style="thin", color="B7B7B7") if "Side" in globals() else None


def _eh_df_valido(df):
    return pd is not None and isinstance(df, pd.DataFrame) and not df.empty


def _num(v, pad=0.0):
    try:
        if v is None:
            return pad
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v).strip().replace("R$", "").replace(" ", "")
        if not s:
            return pad
        if "," in s:
            s = s.replace(".", "").replace(",", ".")
        return float(s)
    except Exception:
        return pad


def _money(cell):
    cell.number_format = 'R$ #,##0.00'
    return cell


def _style_header(ws, row, c1, c2):
    for c in range(c1, c2 + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = PatternFill("solid", fgColor=AZUL)
        cell.font = Font(color=BRANCO, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if BORDA:
            cell.border = Border(top=BORDA, bottom=BORDA, left=BORDA, right=BORDA)


def _style_range(ws, r1, r2, c1, c2):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if BORDA:
                cell.border = Border(top=BORDA, bottom=BORDA, left=BORDA, right=BORDA)


def _col(df, candidatos):
    cols = list(df.columns)
    normal = {str(c).lower().strip(): c for c in cols}
    for cand in candidatos:
        got = normal.get(str(cand).lower().strip())
        if got is not None:
            return got
    for c in cols:
        cn = str(c).lower()
        if any(str(cand).lower() in cn for cand in candidatos):
            return c
    return None


def _linhas_composicao(res):
    linhas = []
    df = res.get("df_composicao_valor_total") if isinstance(res, dict) else None

    if _eh_df_valido(df):
        c_comp = _col(df, ["Componente", "Parcela", "Descrição", "Descricao", "Indicador"])
        c_orig = _col(df, ["Origem", "Fonte"])
        c_crit = _col(df, ["Critério", "Criterio"])
        c_val = _col(df, ["Valor", "valor"])
        c_status = _col(df, ["Status"])
        c_obs = _col(df, ["Observação", "Observacao"])

        if c_val:
            for _, row in df.iterrows():
                comp = str(row.get(c_comp, "") if c_comp else "").strip()
                if not comp:
                    continue
                linhas.append({
                    "componente": comp,
                    "origem": str(row.get(c_orig, "") if c_orig else "").strip(),
                    "criterio": str(row.get(c_crit, "") if c_crit else "").strip(),
                    "valor": _num(row.get(c_val, 0.0), 0.0),
                    "status": str(row.get(c_status, "") if c_status else "").strip(),
                    "obs": str(row.get(c_obs, "") if c_obs else "").strip(),
                })

    if not linhas and isinstance(res, dict):
        linhas.append({
            "componente": "TOTAL",
            "origem": "Resultado consolidado",
            "criterio": "Valor informado pelo processamento",
            "valor": _num(res.get("valor_atualizado_contrato", res.get("valor_global_estoque", 0)), 0.0),
            "status": "VTA",
            "obs": "Sem composição detalhada disponível.",
        })

    return linhas


def _totais(linhas, res):
    out = {"execucao": 0.0, "remanescente": 0.0, "aditivos": 0.0, "total": 0.0}
    for item in linhas:
        comp = item["componente"].lower()
        origem = item["origem"].lower()
        val = float(item["valor"] or 0)
        if "total" in comp and ("vta" in comp or "atualizado" in comp or comp == "total"):
            out["total"] = val
        elif "aditivo" in comp or "supress" in comp or "aditivo" in origem:
            out["aditivos"] += val
        elif "remanescente" in comp:
            out["remanescente"] += val
        elif comp.startswith("c") or "financeiro" in origem or "itens" in origem:
            out["execucao"] += val

    if abs(out["total"]) <= 0.004 and isinstance(res, dict):
        out["total"] = _num(res.get("valor_atualizado_contrato", res.get("valor_global_estoque", 0)), 0.0)
    return out


def _reset_sheet(wb, nome, index=0):
    if nome in wb.sheetnames:
        idx = wb.sheetnames.index(nome)
        wb.remove(wb[nome])
        return wb.create_sheet(nome, idx)
    return wb.create_sheet(nome, index)


def _aba_resumo(wb, res, linhas, totais):
    ws = _reset_sheet(wb, "RESUMO_EXECUTIVO", 0)
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Sumário Executivo da Apuração — Matriz 2.0"
    ws["A1"].font = Font(bold=True, size=14, color="17365D")
    ws["A2"] = "Síntese objetiva com a composição auditável do Valor Total Atualizado do Contrato."
    ws["A2"].font = Font(italic=True, color="666666")

    ws["A4"] = "Indicador"
    ws["B4"] = "Valor"
    _style_header(ws, 4, 1, 2)

    dados = [
        ("Valor executado atualizado por ciclos", totais["execucao"]),
        ("Saldo remanescente atualizado", totais["remanescente"]),
        ("Aditivos/supressões computáveis", totais["aditivos"]),
        ("Valor represado a pagar", _num(res.get("valor_represado_a_pagar", 0) if isinstance(res, dict) else 0)),
        ("Valor Total Atualizado do Contrato", totais["total"]),
    ]

    r = 5
    for k, v in dados:
        ws.cell(r, 1, k)
        _money(ws.cell(r, 2, float(v or 0)))
        if "Valor Total Atualizado" in k:
            ws.cell(r, 1).fill = PatternFill("solid", fgColor=VERDE)
            ws.cell(r, 2).fill = PatternFill("solid", fgColor=VERDE)
            ws.cell(r, 1).font = Font(bold=True)
            ws.cell(r, 2).font = Font(bold=True)
        r += 1
    _style_range(ws, 4, r - 1, 1, 2)

    ws.cell(r + 1, 1, "Composição do VTA")
    ws.cell(r + 1, 1).font = Font(bold=True, size=12, color="17365D")

    hr = r + 3
    headers = ["Componente", "Origem", "Critério", "Valor", "Status", "Observação"]
    for i, h in enumerate(headers, 1):
        ws.cell(hr, i, h)
    _style_header(ws, hr, 1, 6)

    rr = hr + 1
    for item in linhas:
        ws.cell(rr, 1, item["componente"])
        ws.cell(rr, 2, item["origem"])
        ws.cell(rr, 3, item["criterio"])
        _money(ws.cell(rr, 4, float(item["valor"] or 0)))
        ws.cell(rr, 5, item["status"])
        ws.cell(rr, 6, item["obs"])
        if item["componente"].strip().upper() == "TOTAL":
            for c in range(1, 7):
                ws.cell(rr, c).fill = PatternFill("solid", fgColor=VERDE)
                ws.cell(rr, c).font = Font(bold=True)
        rr += 1

    _style_range(ws, hr, rr - 1, 1, 6)

    for col, width in {"A": 36, "B": 24, "C": 46, "D": 18, "E": 18, "F": 54}.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A4"


def _aba_composicao(wb, linhas):
    ws = _reset_sheet(wb, "COMPOSICAO_VALOR_TOTAL", 1)
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Composição do Valor Total Atualizado do Contrato"
    ws["A1"].font = Font(bold=True, size=13, color="17365D")
    ws["A2"] = "Execução atualizada por ciclo + saldo remanescente atualizado + aditivos/supressões computáveis."
    ws["A2"].font = Font(italic=True, color="666666")

    headers = ["Componente", "Origem", "Critério", "Valor", "Status", "Observação"]
    for i, h in enumerate(headers, 1):
        ws.cell(4, i, h)
    _style_header(ws, 4, 1, 6)

    r = 5
    for item in linhas:
        ws.cell(r, 1, item["componente"])
        ws.cell(r, 2, item["origem"])
        ws.cell(r, 3, item["criterio"])
        _money(ws.cell(r, 4, float(item["valor"] or 0)))
        ws.cell(r, 5, item["status"])
        ws.cell(r, 6, item["obs"])
        if item["componente"].strip().upper() == "TOTAL":
            for c in range(1, 7):
                ws.cell(r, c).fill = PatternFill("solid", fgColor=VERDE)
                ws.cell(r, c).font = Font(bold=True)
        r += 1

    _style_range(ws, 4, r - 1, 1, 6)
    for col, width in {"A": 36, "B": 22, "C": 48, "D": 18, "E": 20, "F": 56}.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A5"


def _ocultar_abas_antigas(wb):
    manter = {"RESUMO_EXECUTIVO", "COMPOSICAO_VALOR_TOTAL"}
    for ws in wb.worksheets:
        if ws.title in manter:
            ws.sheet_state = "visible"
        else:
            ws.sheet_state = "hidden"


def ajustar_sumario_executivo_m20(xlsx_bytes, res):
    if load_workbook is None:
        return xlsx_bytes

    if hasattr(xlsx_bytes, "getvalue"):
        xlsx_bytes = xlsx_bytes.getvalue()

    wb = load_workbook(BytesIO(xlsx_bytes))
    linhas = _linhas_composicao(res if isinstance(res, dict) else {})
    totais = _totais(linhas, res if isinstance(res, dict) else {})

    _aba_resumo(wb, res if isinstance(res, dict) else {}, linhas, totais)
    _aba_composicao(wb, linhas)
    _ocultar_abas_antigas(wb)

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()

# >>> PATCH_SUMARIO_RESUMO_FINANCEIRO_VISIVEL_V4
# Pós-processamento para garantir que a aba RESUMO_FINANCEIRO exista e fique visível.
def _m21v4_sumario_num(v, default=0.0):
    try:
        if v is None or isinstance(v, bool):
            return default
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v).strip().replace("R$", "").replace(" ", "")
        if not s:
            return default
        if "," in s:
            s = s.replace(".", "").replace(",", ".")
        return float(s)
    except Exception:
        return default


try:
    _ajustar_sumario_executivo_m20_base_v4 = ajustar_sumario_executivo_m20

    def ajustar_sumario_executivo_m20(xlsx_bytes, res):
        data = _ajustar_sumario_executivo_m20_base_v4(xlsx_bytes, res)
        try:
            from io import BytesIO as _BytesIO_m21v4
            from openpyxl import load_workbook as _load_workbook_m21v4
            from openpyxl.styles import Font as _Font_m21v4, PatternFill as _PatternFill_m21v4
            wb = _load_workbook_m21v4(_BytesIO_m21v4(data))
            if "RESUMO_FINANCEIRO" not in wb.sheetnames:
                ws = wb.create_sheet("RESUMO_FINANCEIRO", min(2, len(wb.sheetnames)))
                ws["A1"] = "Resumo Financeiro da Análise"
                ws["A1"].font = _Font_m21v4(bold=True, size=13, color="17365D")
                ws["A3"] = "Indicador"
                ws["B3"] = "Valor"
                for cell in ws[3]:
                    cell.fill = _PatternFill_m21v4("solid", fgColor="1F4E79")
                    cell.font = _Font_m21v4(color="FFFFFF", bold=True)
                linhas = [
                    ("Valor Total Atualizado", _m21v4_sumario_num(res.get("valor_atualizado_contrato", res.get("valor_global_estoque", 0)) if isinstance(res, dict) else 0)),
                    ("Valor represado/retroativo", _m21v4_sumario_num(res.get("valor_represado_a_pagar", res.get("valor_retroativo", 0)) if isinstance(res, dict) else 0)),
                    ("Aditivos/supressões", _m21v4_sumario_num(res.get("total_aditivos_supressoes", res.get("valor_total_aditivos_supressoes", 0)) if isinstance(res, dict) else 0)),
                ]
                r = 4
                for k, v in linhas:
                    ws.cell(r, 1, k)
                    ws.cell(r, 2, v)
                    ws.cell(r, 2).number_format = 'R$ #,##0.00'
                    r += 1
                ws.column_dimensions["A"].width = 34
                ws.column_dimensions["B"].width = 18
            wb["RESUMO_FINANCEIRO"].sheet_state = "visible"
            out = _BytesIO_m21v4()
            wb.save(out)
            out.seek(0)
            return out.getvalue()
        except Exception:
            return data
except Exception:
    pass
# <<< PATCH_SUMARIO_RESUMO_FINANCEIRO_VISIVEL_V4

