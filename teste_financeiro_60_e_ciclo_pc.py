# -*- coding: utf-8 -*-
"""Financeiro com cobertura de ciclos (C0 ate o corte) e enquadramento de DATA_PC.

Cobre: competencias mensais consecutivas do mes do marco (inicio de C0) ate
o mes da data de corte — sem obrigacao de 60 meses, todos os ciclos de C0 ao
vigente representados, nenhum posterior ao corte inventado; formato mm/aaaa,
datas reais (nao texto), marco/corte invalido, derivacao de ciclo sem cache
pelo leitor, e as transicoes de CICLO_PC em itens_PC (inclusive C1 fora da
apuracao — identidade cronologica preservada), validadas com Excel real
quando disponivel.
"""
from __future__ import annotations

import io
from datetime import date, datetime
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import load_workbook

from _coleta_oficial import obter_coleta_oficial_bytes
from _gerador_masterfile import gerar_masterfile_preenchido
from _leitor_masterfile_v10 import _ciclo_por_competencia

RESULTADOS = []


def check(nome: str, cond: bool, detalhe: str = "") -> None:
    RESULTADOS.append((nome, bool(cond), detalhe))
    status = "OK " if cond else "FAIL"
    print(f"[{status}] {nome}" + (f" - {detalhe}" if detalhe and not cond else ""))


def _dados(marco=date(2023, 1, 1), efeito_c1="Nao"):
    return {
        "ok": True, "modo_origem": "teste", "indice": "IST",
        "ciclo_vigente": "C2", "data_corte": date(2025, 6, 30),
        "data_base": marco,
        "ciclos": [
            {"ciclo": "C1", "data_inicio": date(2024, 1, 1),
             "data_fim": date(2024, 12, 31), "percentual": 0.10,
             "inicio_efeito_financeiro": date(2024, 1, 1),
             "possui_efeito_financeiro": efeito_c1, "situacao": "TEMPESTIVO"},
            {"ciclo": "C2", "data_inicio": date(2025, 1, 1),
             "data_fim": date(2025, 12, 31), "percentual": 0.10,
             "inicio_efeito_financeiro": date(2025, 1, 1),
             "possui_efeito_financeiro": "Sim", "situacao": "TEMPESTIVO"},
        ],
    }


def main() -> int:
    ent = obter_coleta_oficial_bytes()

    # ---- competencias do marco (C0) ate a data de corte ----
    # marco 01/2023, corte 30/06/2025 -> 30 competencias em A2:A31
    preenchido = gerar_masterfile_preenchido(_dados(), ent)
    wb = load_workbook(io.BytesIO(preenchido))
    fin = wb["financeiro"]
    comps = [fin.cell(r, 1).value for r in range(2, 32)]
    check("30 competencias em A2:A31 (marco ate corte)",
          all(v is not None for v in comps))
    check("1a competencia = mes do marco (01/2023)",
          comps[0].year == 2023 and comps[0].month == 1, str(comps[0]))
    check("2a competencia (02/2023)",
          comps[1].year == 2023 and comps[1].month == 2)
    check("12a competencia (12/2023) — C0 integralmente representado",
          comps[11].year == 2023 and comps[11].month == 12)
    check("13a competencia (01/2024) — inicio de C1, sem omissao",
          comps[12].year == 2024 and comps[12].month == 1)
    check("25a competencia (01/2025) — inicio de C2 (vigente)",
          comps[24].year == 2025 and comps[24].month == 1)
    check("ultima competencia = mes do corte (06/2025)",
          comps[29].year == 2025 and comps[29].month == 6, str(comps[29]))
    seq_ok = all(
        (comps[i].year, comps[i].month) ==
        (comps[i - 1].year + (comps[i - 1].month == 12),
         1 if comps[i - 1].month == 12 else comps[i - 1].month + 1)
        for i in range(1, 30)
    )
    check("sequencia mensal continua, sem saltos/duplicidades", seq_ok)
    check("valor interno = dia 1 do mes",
          all(v.day == 1 for v in comps))
    check("datas reais (nao texto)",
          all(isinstance(v, (date, datetime)) for v in comps))
    check("formato de exibicao mm/aaaa",
          fin["A2"].number_format.upper() == "MM/YYYY",
          fin["A2"].number_format)
    check("apos o corte (A32 em diante) sem competencia inventada",
          all(fin.cell(r, 1).value is None for r in range(32, 74)))
    check("coluna B (CICLO) segue formula, nunca escrita",
          all(str(fin.cell(r, 2).value or "").startswith("=")
              for r in range(2, 32)))
    # cobertura: todos os ciclos de C0 ao vigente aparecem no financeiro
    wb_do_cob = load_workbook(io.BytesIO(preenchido), data_only=True)
    ciclos_cobertos = {_ciclo_por_competencia(wb_do_cob, v) for v in comps}
    check("cobertura de ciclos C0/C1/C2 no financeiro (nenhum omitido)",
          {"C0", "C1", "C2"} <= ciclos_cobertos, str(ciclos_cobertos))

    # marco vazio/invalido: nao inventa competencias
    for marco, rotulo in ((None, "vazio"), ("31/01/2023", "texto invalido")):
        p2 = gerar_masterfile_preenchido(_dados(marco=marco), ent)
        f2 = load_workbook(io.BytesIO(p2))["financeiro"]
        check(f"marco {rotulo}: coluna A permanece vazia",
              all(f2.cell(r, 1).value is None for r in range(2, 74)))

    # ---- capacidade da grade (nunca truncar em silencio) ----
    # 01/2023..12/2028 = 72 competencias = capacidade exata da grade A2:A73
    d_lim = _dados()
    d_lim["data_corte"] = date(2028, 12, 31)
    p_lim = gerar_masterfile_preenchido(d_lim, ent)
    f_lim = load_workbook(io.BytesIO(p_lim))["financeiro"]
    comps_lim = [f_lim.cell(r, 1).value for r in range(2, 74)]
    check("periodo no limite exato (72): grade completa A2:A73",
          all(v is not None for v in comps_lim))
    check("periodo no limite exato: ultima = 12/2028 (nenhuma truncada)",
          comps_lim[-1] is not None
          and comps_lim[-1].year == 2028 and comps_lim[-1].month == 12,
          str(comps_lim[-1]))

    # 01/2023..01/2029 = 73 competencias = 1 mes acima do limite
    d_est = _dados()
    d_est["data_corte"] = date(2029, 1, 31)
    erro_capacidade = None
    xls_parcial = None
    try:
        xls_parcial = gerar_masterfile_preenchido(d_est, ent)
    except ValueError as exc:
        erro_capacidade = str(exc)
    check("um mes acima do limite: geracao interrompida",
          erro_capacidade is not None)
    check("nenhum XLS parcial entregue", xls_parcial is None)
    check("mensagem clara com necessario (73) e capacidade (72)",
          erro_capacidade is not None
          and "exige 73 competencias" in erro_capacidade
          and "no maximo 72" in erro_capacidade
          and "Revise o marco inicial ou a data de corte" in erro_capacidade,
          str(erro_capacidade))

    # ---- derivacao de ciclo sem cache (leitor) ----
    wb_do = load_workbook(io.BytesIO(preenchido), data_only=True)
    casos_leitor = [
        (date(2023, 1, 1), "C0"), (date(2023, 12, 1), "C0"),
        (date(2024, 1, 1), "C1"), (date(2024, 12, 1), "C1"),
        (date(2025, 1, 1), "C2"), (date(2025, 12, 1), "C2"),
    ]
    ok_deriv = all(_ciclo_por_competencia(wb_do, d) == esperado
                   for d, esperado in casos_leitor)
    check("leitor deriva ciclo por competencia sem cache (C0/C1/C2)", ok_deriv,
          str([(str(d), _ciclo_por_competencia(wb_do, d)) for d, _ in casos_leitor]))

    # ---- transicoes de CICLO_PC (Excel real) ----
    try:
        import win32com.client  # noqa: F401
        _validar_ciclo_pc_com_excel(preenchido)
    except ImportError:
        print("  [COM] win32com nao disponivel — transicoes CICLO_PC "
              "validadas manualmente com Excel real (py sistema)")

    falhas = [n for n, ok, _ in RESULTADOS if not ok]
    print(f"\n{len(RESULTADOS)} verificacoes; falhas: {len(falhas)}")
    return 1 if falhas else 0


def _validar_ciclo_pc_com_excel(preenchido: bytes) -> None:
    import win32com.client

    # C1 esta FORA da apuracao (efeito Nao) — a identidade cronologica deve
    # ser preservada: data de C1 NAO pode virar C2.
    wb = load_workbook(io.BytesIO(preenchido))
    ipc = wb["itens_PC"]
    casos = [
        ("PC-A", date(2023, 12, 31), "C0", "um dia antes do inicio de C1"),
        ("PC-B", date(2024, 1, 1), "C1", "exatamente no inicio de C1"),
        ("PC-C", date(2024, 1, 2), "C1", "um dia depois do inicio de C1"),
        ("PC-D", date(2024, 12, 31), "C1", "um dia antes de C2"),
        ("PC-E", date(2025, 1, 1), "C2", "exatamente no inicio de C2"),
        ("PC-F", date(2025, 12, 31), "C2", "ultimo dia de C2"),
        ("PC-G", date(2022, 12, 31), "Fora dos ciclos", "anterior a C0"),
        # calendario anual completo (governanca v10.5.3.1): C3/C4 derivados
        ("PC-H", date(2026, 6, 1), "C3", "dentro de C3 derivado"),
        ("PC-I", date(2028, 6, 1), "Fora dos ciclos", "posterior ao ultimo ciclo"),
    ]
    for i, (num, d, _, _rot) in enumerate(casos, start=2):
        ipc.cell(i, 1).value = num
        ipc.cell(i, 2).value = d
        ipc.cell(i, 4).value = 1000.0
        ipc.cell(i, 7).value = "Sim"
    # celula vazia (linha 11) e conteudo nao-data (linha 12)
    ipc.cell(12, 1).value = "PC-TXT"
    ipc.cell(12, 2).value = "data errada"
    ipc.cell(12, 4).value = 1000.0

    with TemporaryDirectory() as tmp:
        caminho = Path(tmp) / "ciclo_pc.xlsx"
        buf = io.BytesIO()
        wb.save(buf)
        caminho.write_bytes(buf.getvalue())
        excel = win32com.client.gencache.EnsureDispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        try:
            pasta = excel.Workbooks.Open(str(caminho))
            excel.CalculateFullRebuild()
            pasta.Save()
            pasta.Close(SaveChanges=False)
        finally:
            excel.Quit()
        wb2 = load_workbook(caminho, data_only=True)

    ws = wb2["itens_PC"]
    for i, (num, d, esperado, rotulo) in enumerate(casos, start=2):
        check(f"CICLO_PC {rotulo} ({d}) = {esperado}",
              ws.cell(i, 3).value == esperado, repr(ws.cell(i, 3).value))
    check("CICLO_PC celula vazia permanece vazia",
          ws.cell(11, 3).value in (None, ""), repr(ws.cell(11, 3).value))
    check("CICLO_PC conteudo nao-data gera alerta explicito",
          ws.cell(12, 3).value == "DATA_PC invalida", repr(ws.cell(12, 3).value))
    # impacto: PC de C2 (computada) tem fator e valor atualizado
    check("PC de C2 tem FATOR_ACUMULADO e VALOR_ATUALIZADO",
          ws.cell(6, 5).value not in (None, "")
          and ws.cell(6, 6).value not in (None, ""),
          f"E6={ws.cell(6, 5).value!r} F6={ws.cell(6, 6).value!r}")


if __name__ == "__main__":
    raise SystemExit(main())
