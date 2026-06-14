
# -*- coding: utf-8 -*-
"""
Leitor experimental da Matriz 2.1 — cl8us/FARC/Streamlit

Objetivo:
- Ler a ColetaReajuste_Matriz21.xlsx sem depender das fórmulas do Excel.
- Recalcular a MEMORIA_VTA em Python.
- Validar C0, saldo remanescente em ciclo em execução, aditivos/supressões e conciliação.
- Não integra ainda ao app Streamlit.

Uso isolado:
    cd /d C:\_DesktopReal\15.ColetaUnica
    py _leitor_matriz_2_1.py "_TESTES_MATRIZ21_PADTEC\TESTE_01_CONCILIACAO_GMP.xlsx"
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
import json
import sys
from typing import Any, Dict, List, Optional, Tuple

try:
    import pandas as pd
except Exception:  # pragma: no cover
    pd = None

from openpyxl import load_workbook


CICLOS = ["C0", "C1", "C2", "C3", "C4"]


# =============================================================================
# Utilitários
# =============================================================================

def _strip(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _norm(v: Any) -> str:
    return _strip(v).lower().replace("ã", "a").replace("á", "a").replace("à", "a").replace("â", "a").replace("é", "e").replace("ê", "e").replace("í", "i").replace("ó", "o").replace("ô", "o").replace("õ", "o").replace("ú", "u").replace("ç", "c")


def _num(v: Any, padrao: float = 0.0) -> float:
    if v is None:
        return padrao
    if isinstance(v, bool):
        return padrao
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return padrao
    s = s.replace("R$", "").replace(" ", "")
    # Trata pt-BR: 1.234.567,89
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        return padrao


def _is_num(v: Any) -> bool:
    try:
        return abs(_num(v, None)) >= 0  # type: ignore[arg-type]
    except Exception:
        return False


def _date(v: Any) -> Optional[date]:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, (int, float)):
        # Excel serial date, base aproximada Windows.
        try:
            return (datetime(1899, 12, 30) + timedelta(days=int(v))).date()
        except Exception:
            return None

    s = str(v).strip()
    if not s:
        return None

    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d", "%m/%Y", "%m/%y"):
        try:
            dt = datetime.strptime(s, fmt)
            return dt.date()
        except Exception:
            pass

    return None


def _add_months(d: date, months: int) -> date:
    month = d.month - 1 + months
    year = d.year + month // 12
    month = month % 12 + 1
    day = min(d.day, [31, 29 if year % 4 == 0 and not year % 100 == 0 or year % 400 == 0 else 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31][month - 1])
    return date(year, month, day)


def _money(v: float) -> str:
    s = f"R$ {v:,.2f}"
    return s.replace(",", "X").replace(".", ",").replace("X", ".")


def _round2(v: float) -> float:
    return round(float(v or 0), 2)


def _df(rows: List[Dict[str, Any]]):
    if pd is None:
        return rows
    return pd.DataFrame(rows)


def _cycle_sort(ciclo: str) -> int:
    try:
        return int(str(ciclo).upper().replace("C", ""))
    except Exception:
        return 99


# =============================================================================
# Estruturas
# =============================================================================

@dataclass
class CicloInfo:
    ciclo: str
    inicio: date
    fim: date
    percentual: float
    fator_proprio: float
    fator_acumulado: float
    efeito_financeiro: str


# =============================================================================
# Leitura base
# =============================================================================

def eh_matriz_2_1(caminho_ou_abas: Any) -> bool:
    if isinstance(caminho_ou_abas, (list, tuple, set)):
        abas = set(caminho_ou_abas)
    else:
        try:
            wb = load_workbook(caminho_ou_abas, read_only=True, data_only=False)
            abas = set(wb.sheetnames)
            wb.close()
        except Exception:
            return False

    obrigatorias = {
        "ENTRADAS_CRITICAS",
        "BASE_CICLOS",
        "FINANCEIRO",
        "ITENS_SALDO",
        "ADITIVOS_SUPRESSOES",
        "MEMORIA_VTA",
        "VALIDACOES",
    }
    return obrigatorias.issubset(abas)


def _load_wb(path: str | Path):
    return load_workbook(path, data_only=True)


def _cell(ws, ref: str) -> Any:
    try:
        return ws[ref].value
    except Exception:
        return None


def _entradas(ws) -> Dict[str, Any]:
    """
    A aba ENTRADAS_CRITICAS usa os campos na coluna B e valores na C.
    Retorna dicionário campo -> valor.
    """
    out: Dict[str, Any] = {}
    for row in range(1, ws.max_row + 1):
        campo = _strip(ws.cell(row=row, column=2).value)
        if campo:
            out[campo] = ws.cell(row=row, column=3).value
    return out


def _ler_base_ciclos(ws_base, entradas: Dict[str, Any]) -> Dict[str, CicloInfo]:
    inicio_c0 = _date(entradas.get("Data início C0 / início contratual")) or _date(_cell(ws_base, "C5"))
    if not inicio_c0:
        # fallback conservador apenas para não quebrar leitura; validação acusará problema.
        inicio_c0 = date(1900, 1, 1)

    ciclos: Dict[str, CicloInfo] = {}
    fator_acum = 1.0

    for idx, ciclo in enumerate(CICLOS):
        row = 5 + idx
        inicio = _date(ws_base.cell(row=row, column=3).value) or _add_months(inicio_c0, 12 * idx)
        fim = _date(ws_base.cell(row=row, column=4).value) or (_add_months(inicio, 12) - timedelta(days=1))
        percentual = _num(ws_base.cell(row=row, column=5).value, 0.0)
        if idx == 0:
            percentual = 0.0
            fator_proprio = 1.0
            fator_acum = 1.0
        else:
            fator_proprio = 1.0 + percentual
            fator_acum *= fator_proprio

        efeito = _strip(ws_base.cell(row=row, column=8).value) or ("Não" if idx == 0 else "Sim")
        ciclos[ciclo] = CicloInfo(
            ciclo=ciclo,
            inicio=inicio,
            fim=fim,
            percentual=percentual,
            fator_proprio=fator_proprio,
            fator_acumulado=fator_acum,
            efeito_financeiro=efeito,
        )

    return ciclos


def _identificar_ciclo(dt: Optional[date], ciclos: Dict[str, CicloInfo]) -> str:
    if not dt:
        return ""
    for ciclo in sorted(ciclos.values(), key=lambda c: _cycle_sort(c.ciclo)):
        if ciclo.inicio <= dt <= ciclo.fim:
            return ciclo.ciclo
    # Se posterior ao C4, assume último ciclo como fallback operacional.
    ult = max(ciclos.values(), key=lambda c: c.fim)
    if dt > ult.fim:
        return ult.ciclo
    return ""


# =============================================================================
# Apurações
# =============================================================================

def _ler_financeiro(ws_fin, ciclos: Dict[str, CicloInfo]) -> Tuple[List[Dict[str, Any]], Dict[str, float]]:
    linhas: List[Dict[str, Any]] = []
    totais = {c: 0.0 for c in CICLOS}

    for row in range(5, ws_fin.max_row + 1):
        competencia = _date(ws_fin.cell(row=row, column=1).value)
        valor = _num(ws_fin.cell(row=row, column=2).value, 0.0)
        obs = _strip(ws_fin.cell(row=row, column=3).value)

        if not competencia and abs(valor) <= 0.0001 and not obs:
            continue

        ciclo = _identificar_ciclo(competencia, ciclos)
        fator = ciclos[ciclo].fator_acumulado if ciclo in ciclos else 1.0
        valor_atualizado = valor * fator

        if ciclo in totais:
            totais[ciclo] += valor_atualizado

        linhas.append({
            "linha": row,
            "competencia": competencia.isoformat() if competencia else "",
            "ciclo": ciclo,
            "valor_base": _round2(valor),
            "fator": fator,
            "valor_atualizado": _round2(valor_atualizado),
            "observacao": obs,
        })

    return linhas, {k: _round2(v) for k, v in totais.items()}


def _ler_itens(ws_itens, ciclos: Dict[str, CicloInfo], ciclo_saldo: str) -> Tuple[List[Dict[str, Any]], Dict[str, float]]:
    linhas: List[Dict[str, Any]] = []
    soma_c0_itens = 0.0
    soma_saldo_apos = 0.0
    soma_saldo_data_corte = 0.0

    fator_saldo = ciclos.get(ciclo_saldo, ciclos.get("C0")).fator_acumulado if ciclo_saldo in ciclos else 1.0

    for row in range(5, ws_itens.max_row + 1):
        item = _strip(ws_itens.cell(row=row, column=1).value)
        qtd_original = _num(ws_itens.cell(row=row, column=2).value, 0.0)
        vu = _num(ws_itens.cell(row=row, column=3).value, 0.0)
        rem_ini_c1 = _num(ws_itens.cell(row=row, column=4).value, 0.0)
        rem_ini_c2 = _num(ws_itens.cell(row=row, column=5).value, 0.0)
        rem_ini_c3 = _num(ws_itens.cell(row=row, column=6).value, 0.0)
        rem_ini_c4 = _num(ws_itens.cell(row=row, column=7).value, 0.0)
        saldo_apos = _num(ws_itens.cell(row=row, column=8).value, 0.0)
        saldo_corte = _num(ws_itens.cell(row=row, column=9).value, 0.0)
        obs = _strip(ws_itens.cell(row=row, column=10).value)

        if not item and abs(qtd_original) <= 0.0001 and abs(vu) <= 0.0001 and not obs:
            continue

        valor_c0_itens = max(qtd_original - rem_ini_c1, 0.0) * vu
        valor_saldo_apos = saldo_apos * vu * fator_saldo
        valor_saldo_corte = saldo_corte * vu * fator_saldo

        soma_c0_itens += valor_c0_itens
        soma_saldo_apos += valor_saldo_apos
        soma_saldo_data_corte += valor_saldo_corte

        linhas.append({
            "linha": row,
            "item": item,
            "qtd_original_c0": qtd_original,
            "vu_c0": vu,
            "rem_inicio_c1": rem_ini_c1,
            "rem_inicio_c2": rem_ini_c2,
            "rem_inicio_c3": rem_ini_c3,
            "rem_inicio_c4": rem_ini_c4,
            "saldo_apos_ultima_competencia": saldo_apos,
            "saldo_data_corte": saldo_corte,
            "valor_c0_itens": _round2(valor_c0_itens),
            "valor_saldo_apos_ultima_competencia": _round2(valor_saldo_apos),
            "valor_saldo_data_corte": _round2(valor_saldo_corte),
            "observacao": obs,
        })

    return linhas, {
        "c0_itens": _round2(soma_c0_itens),
        "saldo_apos_ultima_competencia": _round2(soma_saldo_apos),
        "saldo_data_corte": _round2(soma_saldo_data_corte),
    }


def _ler_aditivos(ws_adit, ciclos: Dict[str, CicloInfo]) -> Tuple[List[Dict[str, Any]], float]:
    linhas: List[Dict[str, Any]] = []
    total = 0.0

    for row in range(5, ws_adit.max_row + 1):
        ident = _strip(ws_adit.cell(row=row, column=1).value)
        tipo = _strip(ws_adit.cell(row=row, column=2).value)
        forma = _strip(ws_adit.cell(row=row, column=3).value)
        dt = _date(ws_adit.cell(row=row, column=4).value)
        ciclo = _identificar_ciclo(dt, ciclos)
        qtd = _num(ws_adit.cell(row=row, column=6).value, 0.0)
        vu = _num(ws_adit.cell(row=row, column=7).value, 0.0)
        valor_consolidado = _num(ws_adit.cell(row=row, column=8).value, 0.0)
        aplicar_fator = _strip(ws_adit.cell(row=row, column=10).value)
        tratamento = _strip(ws_adit.cell(row=row, column=13).value)
        obs = _strip(ws_adit.cell(row=row, column=15).value)

        if not ident and not tipo and not forma and not dt and abs(qtd) <= 0.0001 and abs(vu) <= 0.0001 and abs(valor_consolidado) <= 0.0001:
            continue

        if forma == "Valor consolidado":
            valor_original = valor_consolidado
        else:
            valor_original = qtd * vu

        fator = ciclos[ciclo].fator_acumulado if (aplicar_fator == "Sim" and ciclo in ciclos) else 1.0
        valor_atualizado = valor_original * fator

        if tratamento in ("Informativo / não computar", "Já embutido no saldo"):
            valor_computavel = 0.0
        else:
            base = valor_original if tratamento == "Computar nominal" else valor_atualizado
            sinal = -1.0 if tipo == "Supressão" else 1.0
            valor_computavel = sinal * base

        total += valor_computavel

        linhas.append({
            "linha": row,
            "identificacao": ident,
            "tipo": tipo,
            "forma": forma,
            "data": dt.isoformat() if dt else "",
            "ciclo": ciclo,
            "quantidade": qtd,
            "valor_unitario": _round2(vu),
            "valor_consolidado_informado": _round2(valor_consolidado),
            "valor_original": _round2(valor_original),
            "aplicar_fator": aplicar_fator,
            "fator": fator,
            "valor_atualizado": _round2(valor_atualizado),
            "tratamento": tratamento,
            "valor_computavel": _round2(valor_computavel),
            "observacao": obs,
        })

    return linhas, _round2(total)


def _calcular_memoria(
    entradas: Dict[str, Any],
    ciclos: Dict[str, CicloInfo],
    financeiro_totais: Dict[str, float],
    itens_totais: Dict[str, float],
    total_aditivos: float,
) -> List[Dict[str, Any]]:
    forma_c0 = _strip(entradas.get("Forma de informar C0"))
    valor_c0_consolidado = _num(entradas.get("Valor consolidado do C0"), 0.0)

    if forma_c0 == "Valor consolidado":
        c0 = valor_c0_consolidado
        fonte_c0 = "Valor consolidado"
        obs_c0 = "C0 informado em ENTRADAS_CRITICAS."
    elif forma_c0 == "Financeiro mensal":
        c0 = financeiro_totais.get("C0", 0.0)
        fonte_c0 = "Financeiro"
        obs_c0 = "C0 somado a partir da aba FINANCEIRO."
    elif forma_c0 == "Itens":
        c0 = itens_totais.get("c0_itens", 0.0)
        fonte_c0 = "Itens"
        obs_c0 = "C0 calculado por Qtd original C0 - Remanescente início C1."
    else:
        c0 = 0.0
        fonte_c0 = "Não disponível"
        obs_c0 = "C0 não informado."

    forma_saldo = _strip(entradas.get("Forma de informar saldo remanescente"))
    marco_saldo = _strip(entradas.get("Marco do saldo remanescente"))
    valor_saldo_consolidado = _num(entradas.get("Valor consolidado do saldo remanescente"), 0.0)

    if forma_saldo == "Valor consolidado":
        saldo = valor_saldo_consolidado
        fonte_saldo = f"Valor consolidado | {marco_saldo}"
    elif forma_saldo == "Itens" and marco_saldo == "Após última competência financeira":
        saldo = itens_totais.get("saldo_apos_ultima_competencia", 0.0)
        fonte_saldo = "Itens | Após última competência financeira"
    elif forma_saldo == "Itens" and marco_saldo == "Data de corte":
        saldo = itens_totais.get("saldo_data_corte", 0.0)
        fonte_saldo = "Itens | Data de corte"
    else:
        saldo = 0.0
        fonte_saldo = f"{forma_saldo or 'Não disponível'} | {marco_saldo or 'Sem marco'}"

    rows = [
        {
            "Componente": "C0",
            "Fonte": fonte_c0,
            "Valor": _round2(c0),
            "Status": "Incluído" if c0 > 0 else "Ausente",
            "Observação": obs_c0,
        },
        {
            "Componente": "C1",
            "Fonte": "Financeiro",
            "Valor": financeiro_totais.get("C1", 0.0),
            "Status": "Incluído" if financeiro_totais.get("C1", 0.0) > 0 else "Sem financeiro",
            "Observação": "",
        },
        {
            "Componente": "C2",
            "Fonte": "Financeiro",
            "Valor": financeiro_totais.get("C2", 0.0),
            "Status": "Incluído" if financeiro_totais.get("C2", 0.0) > 0 else "Sem financeiro",
            "Observação": "",
        },
        {
            "Componente": "C3 executado",
            "Fonte": "Financeiro",
            "Valor": financeiro_totais.get("C3", 0.0),
            "Status": "Incluído" if financeiro_totais.get("C3", 0.0) > 0 else "Sem financeiro",
            "Observação": "Executado até a última competência financeira informada, quando aplicável.",
        },
        {
            "Componente": "C4 executado",
            "Fonte": "Financeiro",
            "Valor": financeiro_totais.get("C4", 0.0),
            "Status": "Incluído" if financeiro_totais.get("C4", 0.0) > 0 else "Sem financeiro",
            "Observação": "",
        },
        {
            "Componente": "Saldo remanescente após última competência",
            "Fonte": fonte_saldo,
            "Valor": _round2(saldo),
            "Status": "Incluído" if saldo > 0 else "Ausente",
            "Observação": "Só deve somar com financeiro parcial se representar saldo após última competência ou data de corte equivalente.",
        },
        {
            "Componente": "Aditivos/supressões",
            "Fonte": "ADITIVOS_SUPRESSOES",
            "Valor": _round2(total_aditivos),
            "Status": "Incluído" if abs(total_aditivos) > 0.004 else "Sem valor computável",
            "Observação": "Supressões computáveis entram negativas.",
        },
    ]

    total = _round2(sum(r["Valor"] for r in rows))
    rows.append({
        "Componente": "TOTAL",
        "Fonte": "Consolidação",
        "Valor": total,
        "Status": "VTA",
        "Observação": "Valor Total Atualizado do Contrato.",
    })
    return rows


def _validacoes(
    entradas: Dict[str, Any],
    memoria: List[Dict[str, Any]],
    financeiro_totais: Dict[str, float],
) -> List[Dict[str, str]]:
    valor_por_comp = {r["Componente"]: _num(r["Valor"]) for r in memoria}
    c0 = valor_por_comp.get("C0", 0.0)
    saldo = valor_por_comp.get("Saldo remanescente após última competência", 0.0)
    aditivos = valor_por_comp.get("Aditivos/supressões", 0.0)
    total = valor_por_comp.get("TOTAL", 0.0)

    forma_c0 = _strip(entradas.get("Forma de informar C0"))
    ha_saldo = _strip(entradas.get("Há saldo remanescente após essa competência?"))
    ha_ciclo_execucao = _strip(entradas.get("Há ciclo em execução?"))
    ciclo_execucao = _strip(entradas.get("Ciclo em execução"))
    marco_saldo = _strip(entradas.get("Marco do saldo remanescente"))
    ha_aditivos = _strip(entradas.get("Há aditivos/supressões nesta análise?"))

    out: List[Dict[str, str]] = []

    def add(nome: str, resultado: str, detalhe: str):
        out.append({"Validação": nome, "Resultado": resultado, "Detalhe": detalhe})

    add(
        "C0 informado ou justificado",
        "OK" if forma_c0 == "Não disponível" or c0 > 0 else "ERRO",
        "Informar C0 consolidado, financeiro mensal ou itens."
    )
    add(
        "C0 usado no VTA quando informado",
        "ERRO" if forma_c0 != "Não disponível" and c0 <= 0 else "OK",
        "C0 deve aparecer na memória quando informado."
    )
    add(
        "Saldo remanescente tratado",
        "ERRO" if ha_saldo == "Sim" and saldo <= 0 else "OK",
        "Informar saldo por valor consolidado ou itens."
    )
    add(
        "Risco de dupla contagem no ciclo em execução",
        "ERRO" if (
            ha_ciclo_execucao == "Sim"
            and financeiro_totais.get(ciclo_execucao, 0.0) > 0
            and marco_saldo == "Início do ciclo em execução"
            and saldo > 0
        ) else "OK",
        "Se há financeiro parcial no ciclo, o saldo deve representar o que resta após a última competência."
    )
    add(
        "Saldo atual sem vínculo claro",
        "ALERTA" if ha_saldo == "Sim" and marco_saldo == "Saldo atual sem vínculo claro" else "OK",
        "Informar marco do saldo para evitar dupla contagem."
    )
    add(
        "Aditivos declarados e computados",
        "ALERTA" if ha_aditivos == "Sim" and abs(aditivos) <= 0.004 else "OK",
        "Verificar ADITIVOS_SUPRESSOES e tratamento."
    )
    add(
        "VTA calculado na memória",
        "OK" if total > 0 else "ERRO",
        "Verificar C0, financeiro, saldo e aditivos."
    )

    return out


def _ler_conciliacao(ws_conc, memoria: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    valor_mem = {r["Componente"]: _num(r["Valor"]) for r in memoria}
    mapa = {
        "C0 executado": "C0",
        "C1 atualizado": "C1",
        "C2 atualizado": "C2",
        "C3 executado até corte": "C3 executado",
        "Residual/remanescente após corte": "Saldo remanescente após última competência",
        "Aditivos/supressões computáveis": "Aditivos/supressões",
        "TOTAL": "TOTAL",
    }

    rows = []
    for row in range(5, 12):
        parcela = _strip(ws_conc.cell(row=row, column=1).value)
        if not parcela:
            continue
        ref = _num(ws_conc.cell(row=row, column=2).value, 0.0)
        comp = mapa.get(parcela, parcela)
        cl8us = valor_mem.get(comp, 0.0)
        diff = _round2(ref - cl8us)
        status = "OK" if ref > 0 and abs(diff) <= 1.0 else ("A JUSTIFICAR" if ref == 0 else "DIVERGENTE")
        rows.append({
            "Parcela": parcela,
            "Valor referência externa": _round2(ref),
            "Valor cl8us": _round2(cl8us),
            "Diferença": diff,
            "Status": status,
        })
    return rows


# =============================================================================
# API principal
# =============================================================================

def ler_matriz_2_1(caminho: str | Path) -> Dict[str, Any]:
    caminho = Path(caminho)
    wb = _load_wb(caminho)

    if not eh_matriz_2_1(wb.sheetnames):
        raise ValueError("O arquivo não parece ser uma Matriz 2.1 válida.")

    entradas = _entradas(wb["ENTRADAS_CRITICAS"])
    ciclos = _ler_base_ciclos(wb["BASE_CICLOS"], entradas)

    financeiro_linhas, financeiro_totais = _ler_financeiro(wb["FINANCEIRO"], ciclos)

    ciclo_saldo = _strip(entradas.get("Ciclo do saldo")) or _strip(entradas.get("Ciclo em execução")) or "C0"
    itens_linhas, itens_totais = _ler_itens(wb["ITENS_SALDO"], ciclos, ciclo_saldo)

    aditivos_linhas, total_aditivos = _ler_aditivos(wb["ADITIVOS_SUPRESSOES"], ciclos)

    memoria = _calcular_memoria(
        entradas=entradas,
        ciclos=ciclos,
        financeiro_totais=financeiro_totais,
        itens_totais=itens_totais,
        total_aditivos=total_aditivos,
    )

    validacoes = _validacoes(entradas, memoria, financeiro_totais)

    conciliacao = []
    if "CONCILIACAO_REFERENCIA" in wb.sheetnames:
        conciliacao = _ler_conciliacao(wb["CONCILIACAO_REFERENCIA"], memoria)

    total = next((_num(r["Valor"]) for r in memoria if r["Componente"] == "TOTAL"), 0.0)

    result = {
        "tipo": "Matriz 2.1",
        "arquivo": str(caminho),
        "entradas_criticas": entradas,
        "ciclos": {k: {
            "inicio": v.inicio.isoformat(),
            "fim": v.fim.isoformat(),
            "percentual": v.percentual,
            "fator_proprio": v.fator_proprio,
            "fator_acumulado": v.fator_acumulado,
            "efeito_financeiro": v.efeito_financeiro,
        } for k, v in ciclos.items()},
        "financeiro_linhas": financeiro_linhas,
        "financeiro_totais": financeiro_totais,
        "itens_linhas": itens_linhas,
        "itens_totais": itens_totais,
        "aditivos_linhas": aditivos_linhas,
        "total_aditivos": total_aditivos,
        "memoria_vta_m21": memoria,
        "validacoes_m21": validacoes,
        "conciliacao_referencia": conciliacao,
        "total_vta_m21": _round2(total),
        "valor_atualizado_contrato": _round2(total),
        "df_composicao_valor_total": _df(memoria),
        "df_validacoes_m21": _df(validacoes),
        "df_conciliacao_referencia": _df(conciliacao),
    }

    return result


# Alias para integração futura
ler_matriz21 = ler_matriz_2_1


def _json_safe(obj: Any):
    if isinstance(obj, (date, datetime)):
        return obj.isoformat()
    try:
        if pd is not None and isinstance(obj, pd.DataFrame):
            return obj.to_dict(orient="records")
    except Exception:
        pass
    return str(obj)


def _main():
    if len(sys.argv) < 2:
        print("Uso: py _leitor_matriz_2_1.py <arquivo.xlsx>")
        sys.exit(2)

    res = ler_matriz_2_1(sys.argv[1])

    print("")
    print("=== MATRIZ 2.1 — LEITURA PYTHON ===")
    print(f"Arquivo: {res['arquivo']}")
    print(f"Valor Total Atualizado: {_money(res['total_vta_m21'])}")
    print("")
    print("MEMORIA_VTA:")
    for row in res["memoria_vta_m21"]:
        print(f"- {row['Componente']}: {_money(_num(row['Valor']))} | {row['Fonte']} | {row['Status']}")

    print("")
    print("VALIDACOES:")
    for row in res["validacoes_m21"]:
        print(f"- {row['Validação']}: {row['Resultado']}")

    if res["conciliacao_referencia"]:
        print("")
        print("CONCILIACAO_REFERENCIA:")
        for row in res["conciliacao_referencia"]:
            print(f"- {row['Parcela']}: ref={_money(row['Valor referência externa'])} | cl8us={_money(row['Valor cl8us'])} | dif={_money(row['Diferença'])} | {row['Status']}")


if __name__ == "__main__":
    _main()
# >>> PATCH_M21_DETECCAO_ASSINATURA_V5
def eh_matriz_2_1(abas) -> bool:
    """Assinatura oficial robusta da Coleta Matriz 2.1."""
    try:
        nomes = {str(a).strip().upper() for a in abas}
    except Exception:
        return False
    obrigatorias = {"ENTRADAS_CRITICAS", "BASE_CICLOS"}
    evidencias = {"MEMORIA_VTA", "CONCILIACAO_REFERENCIA", "VALIDACOES"}
    return obrigatorias.issubset(nomes) and bool(nomes & evidencias)
# <<< PATCH_M21_DETECCAO_ASSINATURA_V5

# >>> PATCH_M21_INDICE_CARD_V6_READER
# Preenche o indice da Matriz 2.1 por varredura defensiva do XLSX.
def _m21_v6_normalizar_indice(valor):
    if valor is None:
        return ""
    texto = str(valor).strip()
    if not texto:
        return ""
    up = texto.upper()
    mapa = [
        ("ICTI", "ICTI"),
        ("IST", "IST"),
        ("IPCA", "IPCA"),
        ("INPC", "INPC"),
        ("IGP-M", "IGP-M"),
        ("IGPM", "IGP-M"),
    ]
    for chave, nome in mapa:
        if chave in up:
            return nome
    return ""


def _m21_v6_ler_indice_xlsx(caminho_xlsx):
    try:
        from openpyxl import load_workbook
    except Exception:
        return ""

    try:
        wb = load_workbook(caminho_xlsx, read_only=True, data_only=True)
    except Exception:
        return ""

    try:
        candidatos_abas = [
            "ENTRADAS_CRITICAS",
            "BASE_CICLOS",
            "VALIDACOES",
            "PARAMETROS",
            "PARAMETROS_REAJUSTE",
        ]
        nomes = [n for n in candidatos_abas if n in wb.sheetnames] + [n for n in wb.sheetnames if n not in candidatos_abas]
        for nome_aba in nomes:
            ws = wb[nome_aba]
            max_row = min(ws.max_row or 1, 80)
            max_col = min(ws.max_column or 1, 20)
            for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
                vals = [c.value for c in row]
                for idx, val in enumerate(vals):
                    norm = _m21_v6_normalizar_indice(val)
                    if norm:
                        # Evita capturar apenas texto instrucional muito genérico quando houver valor à direita.
                        texto = str(val).strip().lower()
                        if "indice" in texto or "índice" in texto:
                            for prox in vals[idx + 1: idx + 4]:
                                norm_prox = _m21_v6_normalizar_indice(prox)
                                if norm_prox:
                                    return norm_prox
                        return norm
    finally:
        try:
            wb.close()
        except Exception:
            pass
    return ""


try:
    _ler_matriz_2_1_base_v6 = ler_matriz_2_1
except NameError:
    _ler_matriz_2_1_base_v6 = None


def ler_matriz_2_1(caminho_xlsx):
    if _ler_matriz_2_1_base_v6 is None:
        raise RuntimeError("Função base ler_matriz_2_1 não localizada.")
    diag = _ler_matriz_2_1_base_v6(caminho_xlsx)
    if not isinstance(diag, dict):
        return diag

    indice = (
        diag.get("indice")
        or diag.get("indice_utilizado")
        or diag.get("indice_nome")
        or (diag.get("parametros") or {}).get("indice")
        or (diag.get("parametros") or {}).get("indice_utilizado")
        or ""
    )
    indice = _m21_v6_normalizar_indice(indice) or _m21_v6_ler_indice_xlsx(caminho_xlsx)

    if indice:
        diag["indice"] = indice
        diag["indice_utilizado"] = indice
        diag["indice_nome"] = indice
        params = diag.get("parametros")
        if not isinstance(params, dict):
            params = {}
        params["indice"] = indice
        params["indice_utilizado"] = indice
        diag["parametros"] = params

    return diag
# <<< PATCH_M21_INDICE_CARD_V6_READER

