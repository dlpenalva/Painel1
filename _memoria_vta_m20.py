
from __future__ import annotations

from io import BytesIO
from datetime import datetime
import re

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


def _parse_mm_yyyy(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return datetime(v.year, v.month, 1)
    if hasattr(v, "year") and hasattr(v, "month"):
        return datetime(v.year, v.month, 1)
    s = str(v).strip()
    m = re.search(r"(\d{1,2})[\/\-](\d{4})", s)
    if m:
        return datetime(int(m.group(2)), int(m.group(1)), 1)
    m = re.search(r"(\d{4})[\/\-](\d{1,2})", s)
    if m:
        return datetime(int(m.group(1)), int(m.group(2)), 1)
    return None


def _add_months(dt, months):
    y = dt.year + (dt.month - 1 + months) // 12
    m = (dt.month - 1 + months) % 12 + 1
    return datetime(y, m, 1)


def _fmt_periodo(ini, fim):
    if not ini or not fim:
        return ""
    return f"{ini.month:02d}/{ini.year} a {fim.month:02d}/{fim.year}"


def _ciclo_num(ciclo):
    try:
        return int(str(ciclo).upper().replace("C", "").strip())
    except Exception:
        return 0


def _find_current_cycle(ws_base):
    val = ws_base["K6"].value
    if val:
        return str(val).strip().upper()
    for row in range(1, min(ws_base.max_row, 30) + 1):
        for col in range(1, min(ws_base.max_column, 12) + 1):
            if str(ws_base.cell(row, col).value or "").strip().lower() in {"ciclo atual/corte", "ciclo atual", "corte"}:
                v = ws_base.cell(row, col + 1).value
                if v:
                    return str(v).strip().upper()
    return "C4"


def _fill_base_dates(wb):
    ws = wb["BASE"]
    c1_dt = _parse_mm_yyyy(ws["E7"].value)
    c0_dt = _parse_mm_yyyy(ws["E6"].value)
    if not c0_dt and c1_dt:
        c0_dt = _add_months(c1_dt, -12)
    if not c0_dt:
        first = None
        first_row = None
        for r in range(6, 11):
            dt = _parse_mm_yyyy(ws.cell(r, 5).value)
            if dt:
                first = dt
                first_row = r
                break
        if first and first_row:
            c0_dt = _add_months(first, -12 * (first_row - 6))
    if not c0_dt:
        return {}

    ciclo_datas = {}
    for i, r in enumerate(range(6, 11)):
        ciclo = f"C{i}"
        ini = _add_months(c0_dt, 12 * i)
        fim = _add_months(ini, 11)
        ciclo_datas[ciclo] = (ini, fim)
        ws.cell(r, 1).value = ciclo
        ws.cell(r, 2).value = _fmt_periodo(ini, fim)
        ws.cell(r, 3).value = ini
        ws.cell(r, 4).value = fim
        if not ws.cell(r, 5).value:
            ws.cell(r, 5).value = f"{ini.month:02d}/{ini.year}"
        ws.cell(r, 3).number_format = "dd/mm/yyyy"
        ws.cell(r, 4).number_format = "dd/mm/yyyy"
        ws.cell(r, 6).number_format = "0.00%"
        ws.cell(r, 7).number_format = "0.000000"
        ws.cell(r, 8).number_format = "0.000000"
    return ciclo_datas


def _fill_execucao(wb, ciclo_datas):
    if "EXECUCAO_FINANCEIRA" not in wb.sheetnames:
        return
    ws = wb["EXECUCAO_FINANCEIRA"]
    ws_base = wb["BASE"]
    ciclo_corte = _find_current_cycle(ws_base)
    corte_num = min(_ciclo_num(ciclo_corte), 4)

    for row in range(6, max(ws.max_row, 306) + 1):
        for col in range(1, 9):
            if col not in (2, 3):
                ws.cell(row, col).value = None

    row = 6
    for i in range(0, corte_num + 1):
        ciclo = f"C{i}"
        ini, fim = ciclo_datas.get(ciclo, (None, None))
        if not ini or not fim:
            continue
        dt = ini
        while dt <= fim:
            ws.cell(row, 1).value = dt
            ws.cell(row, 1).number_format = "mm/yyyy"
            if ws.cell(row, 3).value in (None, ""):
                ws.cell(row, 3).value = "Nao" if ciclo == "C0" else "Sim"
            ws.cell(row, 4).value = ciclo
            ws.cell(row, 5).value = f'=IFERROR(INDEX(BASE!$H$6:$H$10,MATCH(D{row},BASE!$A$6:$A$10,0)),1)'
            ws.cell(row, 6).value = f'=IF(B{row}="",0,B{row})'
            ws.cell(row, 7).value = f'=IF(UPPER(C{row})="SIM",F{row}*(E{row}-1),0)'
            ws.cell(row, 8).value = f'=F{row}+G{row}'
            for c in (2, 6, 7, 8):
                ws.cell(row, c).number_format = 'R$ #,##0.00'
            ws.cell(row, 5).number_format = "0.000000"
            row += 1
            dt = _add_months(dt, 1)

    for r in range(row, max(ws.max_row, row + 10) + 1):
        for c in (1, 4, 5, 6, 7, 8):
            ws.cell(r, c).value = None


def _fill_historico(wb):
    if "HISTORICO_CICLOS" not in wb.sheetnames:
        return
    ws = wb["HISTORICO_CICLOS"]
    for idx, r in enumerate(range(6, 11)):
        ciclo = f"C{idx}"
        ws.cell(r, 1).value = ciclo
        ws.cell(r, 2).value = f'=IFERROR(INDEX(BASE!$B$6:$B$10,MATCH(A{r},BASE!$A$6:$A$10,0)),"")'
        ws.cell(r, 5).value = f'=IFERROR(INDEX(BASE!$F$6:$F$10,MATCH(A{r},BASE!$A$6:$A$10,0)),0)'
        ws.cell(r, 7).value = f'=IF(F{r}="",0,F{r}*(1+E{r}))'
        ws.cell(r, 8).value = f'=G{r}-IF(F{r}="",0,F{r})'
        ws.cell(r, 9).value = f'=SUMIF(EXECUCAO_FINANCEIRA!$D:$D,A{r},EXECUCAO_FINANCEIRA!$B:$B)'
        ws.cell(r, 10).value = f'=SUMIF(EXECUCAO_FINANCEIRA!$D:$D,A{r},EXECUCAO_FINANCEIRA!$H:$H)'
        ws.cell(r, 11).value = f'=IFERROR(SUM(ITENS_CONTRATADOS!$K:$K),0)'
        ws.cell(r, 12).value = f'=IFERROR(SUM(ITENS_CONTRATADOS!$M:$M),0)'
        ws.cell(r, 5).number_format = "0.00%"
        for c in range(6, 13):
            ws.cell(r, c).number_format = 'R$ #,##0.00'


def _hide_itens_cols(wb):
    if "ITENS_CONTRATADOS" not in wb.sheetnames:
        return
    ws = wb["ITENS_CONTRATADOS"]
    corte = _find_current_cycle(wb["BASE"])
    n = min(_ciclo_num(corte), 4)

    for col in range(5, 19):
        ws.column_dimensions[get_column_letter(col)].hidden = False

    for ciclo_num, col in [(3, 8), (4, 9)]:
        if ciclo_num > n:
            ws.column_dimensions[get_column_letter(col)].hidden = True

    for ciclo_num, col in [(3, 17), (4, 18)]:
        if ciclo_num > n:
            ws.column_dimensions[get_column_letter(col)].hidden = True

    ws["J4"].value = f"Corte atual: {corte}"


def _rebuild_memoria(wb):
    if "MEMORIA_VTA" in wb.sheetnames:
        del wb["MEMORIA_VTA"]
    ws = wb.create_sheet("MEMORIA_VTA")
    ws_base = wb["BASE"]

    dark = PatternFill("solid", fgColor="1F4E78")
    total_fill = PatternFill("solid", fgColor="0F766E")
    pale = PatternFill("solid", fgColor="F8FAFC")
    yellow = PatternFill("solid", fgColor="FFF2CC")
    gray = PatternFill("solid", fgColor="E5E7EB")
    white = Font(color="FFFFFF", bold=True)
    bold = Font(bold=True)
    small = Font(size=9, italic=True, color="475569")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:I1")
    ws["A1"] = "MEMÓRIA DO VALOR TOTAL ATUALIZADO — MATRIZ 2.0"
    ws["A1"].fill = total_fill
    ws["A1"].font = white
    ws["A1"].alignment = Alignment(horizontal="left")

    ws.merge_cells("A2:I3")
    ws["A2"] = (
        "Regra: C0 sempre compõe o VTA. A coluna 'Possui Efeito Financeiro?' serve para delta/retroativo, não para excluir execução histórica. "
        "Fonte padrão: financeiro; fallback: itens; se não houver dados, registrar ausência. Em ciclo em execução, soma-se financeiro até a última competência informada + remanescente parcial por itens."
    )
    ws["A2"].font = small
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")

    headers = ["Ciclo/Componente", "Fonte usada", "Período/critério", "Valor financeiro atualizado", "Valor por itens (fallback)", "Fator aplicado", "Valor incluído no VTA", "Status", "Observação"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(4, col, h)
        cell.fill = dark
        cell.font = white
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    item_cols = {"C0": "N", "C1": "O", "C2": "P", "C3": "Q", "C4": "R"}
    for i, row in enumerate(range(5, 10)):
        ciclo = f"C{i}"
        col_item = item_cols[ciclo]
        ws.cell(row, 1, ciclo)
        ws.cell(row, 2).value = (
            f'=IF(--SUBSTITUTE(A{row},"C","")>--SUBSTITUTE(BASE!$K$6,"C",""),"Fora do corte",'
            f'IF(D{row}>0,"Financeiro",IF(E{row}>0,"Itens","Ausente")))'
        )
        ws.cell(row, 3).value = f'=IFERROR(INDEX(BASE!$B$6:$B$10,MATCH(A{row},BASE!$A$6:$A$10,0)),"")'
        ws.cell(row, 4).value = f'=SUMIF(EXECUCAO_FINANCEIRA!$D:$D,A{row},EXECUCAO_FINANCEIRA!$H:$H)'
        ws.cell(row, 5).value = f'=MAX(0,SUMPRODUCT(ITENS_CONTRATADOS!${col_item}$6:${col_item}$505,ITENS_CONTRATADOS!$C$6:$C$505,INDEX(BASE!$H$6:$H$10,MATCH(A{row},BASE!$A$6:$A$10,0))))'
        ws.cell(row, 6).value = f'=IFERROR(INDEX(BASE!$H$6:$H$10,MATCH(A{row},BASE!$A$6:$A$10,0)),1)'
        ws.cell(row, 7).value = f'=IF(B{row}="Financeiro",D{row},IF(B{row}="Itens",E{row},0))'
        ws.cell(row, 8).value = f'=IF(B{row}="Fora do corte","Fora do corte",IF(B{row}="Ausente","Não incluído por falta de dados","Incluído"))'
        ws.cell(row, 9).value = (
            f'=IF(B{row}="Financeiro","Fonte padrão: financeiro. Incluído no VTA mesmo se Possui Efeito Financeiro = Não.",'
            f'IF(B{row}="Itens","Financeiro ausente; calculado por itens.","Sem dados financeiros nem itemizados suficientes."))'
        )

    r = 10
    ws.cell(r, 1, "Remanescente atual")
    ws.cell(r, 2, "Itens")
    ws.cell(r, 3, "Saldo remanescente informado na aba ITENS_CONTRATADOS")
    ws.cell(r, 4, 0)
    ws.cell(r, 5, '=SUM(ITENS_CONTRATADOS!$M:$M)')
    ws.cell(r, 6, '=IFERROR(INDEX(BASE!$H$6:$H$10,MATCH(BASE!$K$6,BASE!$A$6:$A$10,0)),1)')
    ws.cell(r, 7, '=E10')
    ws.cell(r, 8, "Incluído se houver saldo")
    ws.cell(r, 9, "Parte ainda não executada. Em ciclo em execução, soma-se ao financeiro parcial já informado.")

    r = 11
    ws.cell(r, 1, "Aditivos/supressões computáveis")
    ws.cell(r, 2, "ADITIVOS")
    ws.cell(r, 3, "Tratamento = Computar nesta análise")
    ws.cell(r, 4, 0)
    ws.cell(r, 5, 0)
    ws.cell(r, 6, 0)
    ws.cell(r, 7, '=SUM(ADITIVOS!$L:$L)')
    ws.cell(r, 8, "Incluído")
    ws.cell(r, 9, "Supressões entram negativas na coluna Valor Computável.")

    r = 13
    ws.cell(r, 1, "VALOR TOTAL ATUALIZADO DO CONTRATO")
    ws.cell(r, 3, "Soma dos componentes incluídos no VTA")
    ws.cell(r, 7, '=SUM(G5:G11)')
    ws.cell(r, 8, "VTA")

    r = 15
    ws.cell(r, 1, "Conferência")
    ws.cell(r, 2, "Valor")
    ws.cell(r, 3, "Observação")
    ws.cell(r + 1, 1, "VTA da memória")
    ws.cell(r + 1, 2, "=G13")
    ws.cell(r + 1, 3, "Valor oficial auditável desta aba.")
    ws.cell(r + 2, 1, "VTA da BASE")
    ws.cell(r + 2, 2, "=BASE!B16")
    ws.cell(r + 2, 3, "Deve coincidir com a memória.")
    ws.cell(r + 3, 1, "Diferença")
    ws.cell(r + 3, 2, "=B16-B17")
    ws.cell(r + 3, 3, "Se diferente de zero, há divergência entre BASE e MEMORIA_VTA.")

    for row in range(4, 19):
        for col in range(1, 10):
            cell = ws.cell(row, col)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if row == 4:
                continue
            elif row == 10:
                cell.fill = yellow
            elif row == 11:
                cell.fill = gray
            elif row == 13:
                cell.fill = total_fill
                cell.font = white
            elif row >= 15:
                cell.fill = PatternFill("solid", fgColor="EFF6FF")
                if row == 15:
                    cell.font = bold
            else:
                cell.fill = pale
        for col in (4, 5, 7):
            ws.cell(row, col).number_format = 'R$ #,##0.00'
        ws.cell(row, 6).number_format = '0.000000'
        if row >= 16:
            ws.cell(row, 2).number_format = 'R$ #,##0.00'

    widths = {1:24, 2:18, 3:38, 4:24, 5:24, 6:14, 7:24, 8:18, 9:64}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A5"

    ws_base["B13"] = "=SUM(MEMORIA_VTA!G5:G9)"
    ws_base["B14"] = "=MEMORIA_VTA!G10"
    ws_base["B15"] = "=MEMORIA_VTA!G11"
    ws_base["B16"] = "=MEMORIA_VTA!G13"
    for cell in ("B13", "B14", "B15", "B16"):
        ws_base[cell].number_format = 'R$ #,##0.00'


def aplicar_memoria_vta_xlsx(xlsx_bytes: bytes) -> bytes:
    wb = load_workbook(BytesIO(xlsx_bytes))
    if "BASE" not in wb.sheetnames:
        return xlsx_bytes

    ciclo_datas = _fill_base_dates(wb)
    _fill_execucao(wb, ciclo_datas)
    _fill_historico(wb)
    _hide_itens_cols(wb)
    _rebuild_memoria(wb)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()
