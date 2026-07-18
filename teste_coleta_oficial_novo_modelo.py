# -*- coding: utf-8 -*-
"""Contrato estrutural do novo XLS Coleta oficial (aba posicao_contratual).

Cobre: geracao do template limpo, ordem das abas, ausencia de residuos,
integridade ZIP, preservacao de formulas no preenchimento pela Calculadora,
leitura da posicao_contratual e rejeicao controlada do modelo antigo.
"""
from __future__ import annotations

import io
import zipfile
from datetime import date

from openpyxl import Workbook, load_workbook

from _coleta_oficial import (
    ABAS_COLETA_OFICIAL,
    COLUNAS_POSICAO_CONTRATUAL,
    TEMPLATE_COLETA_OFICIAL,
    obter_coleta_oficial_bytes,
)
from _gerador_masterfile import gerar_masterfile_preenchido
from _leitor_masterfile_v10 import ler_masterfile_v10

RESULTADOS = []


def check(nome: str, cond: bool, detalhe: str = "") -> None:
    RESULTADOS.append((nome, bool(cond), detalhe))
    status = "OK " if cond else "FAIL"
    print(f"[{status}] {nome}" + (f" — {detalhe}" if detalhe and not cond else ""))


def _formulas(wb):
    return {
        f"{ws.title}!{c.coordinate}": c.value
        for ws in wb.worksheets
        for row in ws.iter_rows()
        for c in row
        if isinstance(c.value, str) and c.value.startswith("=")
    }


def _dados_calculadora():
    return {
        "ok": True, "modo_origem": "teste", "indice": "IST (Serie Local)",
        "ciclo_vigente": "C2", "data_corte": date(2025, 9, 30),
        "data_base": date(2022, 10, 1),
        "ciclos": [
            {"ciclo": "C1", "data_inicio": date(2023, 10, 1),
             "data_fim": date(2024, 9, 30), "percentual": 0.0157775,
             "possui_efeito_financeiro": "Sim", "situacao": "TEMPESTIVO"},
            {"ciclo": "C2", "data_inicio": date(2024, 10, 1),
             "data_fim": date(2025, 9, 30), "percentual": 0.0488346,
             "possui_efeito_financeiro": "Sim", "situacao": "TEMPESTIVO"},
        ],
    }


def main() -> int:
    check("template oficial existe", TEMPLATE_COLETA_OFICIAL.exists())

    ent = obter_coleta_oficial_bytes()

    # Integridade ZIP/PK
    with zipfile.ZipFile(io.BytesIO(ent)) as z:
        check("XLSX gerado integro (ZIP)", z.testzip() is None)

    wb = load_workbook(io.BytesIO(ent), data_only=False)
    check("abas e ordem oficiais", wb.sheetnames == ABAS_COLETA_OFICIAL,
          str(wb.sheetnames))
    check("posicao_contratual presente", "posicao_contratual" in wb.sheetnames)

    pos = wb["posicao_contratual"]
    headers = [pos.cell(1, i + 1).value for i in range(len(COLUNAS_POSICAO_CONTRATUAL))]
    check("cabecalhos posicao_contratual", headers == COLUNAS_POSICAO_CONTRATUAL,
          str(headers[:6]))
    check("posicao_contratual 100% formula (A2)",
          str(pos["A2"].value or "").startswith("="))

    # Ausencia de residuos demonstrativos nas celulas de entrada
    residuos = []
    for coord in ("B2", "B3", "B7", "B8"):
        if wb["CONTROLE"][coord].value is not None:
            residuos.append(f"CONTROLE!{coord}")
    for lin in range(2, 7):
        for col in ("A", "C", "D", "E", "G"):
            if wb["parametros"][f"{col}{lin}"].value is not None:
                residuos.append(f"parametros!{col}{lin}")
    for lin in range(2, 74):
        for col in ("A", "C", "G"):
            v = wb["financeiro"][f"{col}{lin}"].value
            if v is not None and not str(v).startswith("="):
                residuos.append(f"financeiro!{col}{lin}")
    for ws_nome, cols in (("itens_Remanesc", "ABCEGIK"), ("aditivos", "ABDEFHK")):
        ws = wb[ws_nome]
        for lin in range(2, 201):
            for col in cols:
                v = ws[f"{col}{lin}"].value
                if v is not None and not str(v).startswith("="):
                    residuos.append(f"{ws_nome}!{col}{lin}")
    check("sem dados residuais de template", not residuos, str(residuos[:8]))

    # Preenchimento pela Calculadora preserva formulas
    formulas_template = _formulas(wb)
    preenchido = gerar_masterfile_preenchido(_dados_calculadora(), ent)
    wb2 = load_workbook(io.BytesIO(preenchido), data_only=False)
    check("preenchido preserva todas as formulas",
          _formulas(wb2) == formulas_template)
    check("preenchido mantem abas oficiais", wb2.sheetnames == ABAS_COLETA_OFICIAL)
    check("parametros preenchidos (C1 Sim)",
          wb2["parametros"]["A3"].value == "Sim"
          and wb2["parametros"]["E3"].value == 0.0157775)
    check("financeiro competencia na linha 2",
          wb2["financeiro"]["A2"].value is not None
          and str(wb2["financeiro"]["B2"].value or "").startswith("="))
    check("CONTROLE ciclo vigente/data corte",
          wb2["CONTROLE"]["B2"].value == "C2"
          and wb2["CONTROLE"]["B3"].value is not None)

    with zipfile.ZipFile(io.BytesIO(preenchido)) as z:
        check("XLSX preenchido integro (ZIP)", z.testzip() is None)

    # Upload/leitura do novo modelo
    r = ler_masterfile_v10(preenchido)
    check("leitor aceita novo modelo gerado", not r["erro"], r["erro"])
    check("leitor expoe posicao_contratual ok",
          r["posicao_contratual"].get("ok") is True)
    check("leitor le ciclos C0-C4",
          [c["ciclo"] for c in r["parametros_v10"]["ciclos"]]
          == ["C0", "C1", "C2", "C3", "C4"])

    # Cache das formulas: arquivo gerado por openpyxl nunca foi recalculado
    # pelo Excel — posicao_contratual sem valores. O sistema deve sinalizar
    # cache_ausente e BLOQUEAR a formalizacao (nao apenas alertar).
    from _politica_entrega_segura import avaliar_entrega_segura
    check("cache ausente detectado (arquivo sem recalculo do Excel)",
          r["posicao_contratual"].get("cache_ausente") is True)
    check("alerta orienta abrir e salvar no Excel",
          any("recalculo do Excel" in a for a in r["avisos"]), str(r["avisos"]))
    seg = avaliar_entrega_segura(r)
    check("formalizacao bloqueada sem cache (pode_confirmar=False)",
          seg["pode_confirmar"] is False
          and any("posicao_contratual" in b for b in seg["bloqueios"]),
          str(seg["bloqueios"]))
    check("status BLOQUEADO_PARA_FORMALIZACAO sem cache",
          seg["status"] == "BLOQUEADO_PARA_FORMALIZACAO", seg["status"])

    # Etapa 6: RESULTADOS exposta como auditoria (nunca fonte de calculo)
    rx = r.get("resultados_xls") or {}
    check("leitor expoe nomes de RESULTADOS para reconciliacao",
          rx.get("disponivel") is True and "RETRO_OFICIAL" in rx.get("nomes_presentes", []),
          str(rx.get("nomes_presentes")))
    check("RESULTADOS sem recalculo marcada cache_ausente",
          rx.get("cache_ausente") is True)
    reconc = r.get("reconciliacao_xls_python") or {}
    check("reconciliacao XLS x Python classifica INDISPONIVEL_POR_CACHE",
          reconc.get("status_geral") == "RESULTADO_XLS_INDISPONIVEL_POR_CACHE",
          str(reconc.get("status_geral")))
    check("grade financeiro estendida ate a linha 73 (B73 formula)",
          str(wb["financeiro"]["B73"].value or "").startswith("="))
    check("itens_PC oficial comeca em NUMERO_PC",
          wb["itens_PC"]["A1"].value == "NUMERO_PC")

    # Rejeicao controlada do modelo antigo (sem posicao_contratual)
    velho = Workbook()
    velho.active.title = "CONTROLE"
    for aba in ("parametros", "financeiro", "itens_Remanesc", "itens_Consumidos",
                "itens_PC", "aditivos", "itens_RC", "historico_VU", "historico"):
        velho.create_sheet(aba)
    buf = io.BytesIO()
    velho.save(buf)
    ra = ler_masterfile_v10(buf.getvalue(), exigir_modelo_oficial=True)
    check("modelo antigo rejeitado com mensagem clara",
          "versão anterior do XLS Coleta" in str(ra["erro"])
          and "posicao_contratual" in str(ra["erro"]), str(ra["erro"]))

    falhas = [n for n, ok, _ in RESULTADOS if not ok]
    print(f"\n{len(RESULTADOS)} verificacoes; falhas: {len(falhas)}")
    return 1 if falhas else 0


if __name__ == "__main__":
    raise SystemExit(main())
