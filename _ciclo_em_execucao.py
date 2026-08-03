"""Posição itemizada do remanescente no ciclo em execução.

Este módulo é deliberadamente isolado do VTA oficial. Ele concentra:

* o motor puro da posição física por data;
* a criação idempotente da aba ``CICLO_EM_EXECUCAO``;
* a leitura compatível (ausente, legado agregado e layout itemizado);
* a validação e a reconciliação apenas diagnóstica com ``itens_Consumidos``.

Nenhuma função deste arquivo altera RESULTADOS, MEMORIA_RESULTADOS ou qualquer
parcela usada pelo cálculo homologado do VTA.
"""
from __future__ import annotations

import math
import re
from datetime import date, datetime
from typing import Any, Iterable, Mapping

from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName


ABA_CICLO_EM_EXECUCAO = "CICLO_EM_EXECUCAO"
MARCADOR_LAYOUT = "CICLO_EM_EXECUCAO_LAYOUT_VERSION"
VERSAO_LAYOUT_ITEMIZADO = "2_ITEMIZADO"

LINHA_CABECALHO = 12
PRIMEIRA_LINHA_ITEM = 13
ULTIMA_LINHA_ITEM = 211  # espelha itens_Remanesc!2:200
CELULA_CICLO = "C3"
CELULA_INICIO_CICLO = "F3"
CELULA_FIM_CICLO_TECNICA = "H3"
CELULA_DATA_POSICAO = "D5"
CELULA_TOTAL = "A9"
CELULA_VERSAO = "H2"

COLUNAS_VISIVEIS = (
    "ITEM (AUTO)",
    "QTD REMANESCENTE NO INÍCIO DO CICLO (AUTO)",
    "QTD REMANESCENTE NA DATA ATUAL (PREENCHER)",
    "QTD CONSUMIDA NO CICLO ATÉ A DATA (AUTO)",
    "VU ATUALIZADO (AUTO)",
    "VALOR CONSUMIDO (AUTO)",
    "VALOR REMANESCENTE ATUALIZADO (AUTO)",
)

COLUNAS_TECNICAS = (
    "ALTERACOES_LIQUIDAS_NO_PERIODO",
    "CHECK_FISICO",
    "STATUS_POSICAO",
    "CONSUMO_ITENS_CONFERENCIA",
    "DIVERGENCIA_CONSUMO",
    "QTD_REGISTROS_CONSUMO",
    "QTD_EVENTOS_POSITIVOS_ATE_DATA",
)

FORMATO_QTD = "#,##0.00"
FORMATO_MOEDA = "R$ #,##0.00"
FORMATO_DATA = "dd/mm/yyyy"
TOLERANCIA = 0.005


def _data(valor: Any) -> date | None:
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    texto = str(valor or "").strip()
    for formato in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(texto, formato).date()
        except ValueError:
            continue
    return None


def _numero(valor: Any) -> float | None:
    if valor in (None, "") or isinstance(valor, bool):
        return None
    try:
        numero = float(valor)
    except (TypeError, ValueError):
        return None
    return numero if math.isfinite(numero) else None


def _item(valor: Any) -> str:
    return str(valor or "").strip()


def _chave_item(valor: Any) -> str:
    return _item(valor).casefold()


def _delta_movimento(registro: Mapping[str, Any]) -> float | None:
    for chave in ("delta_quantidade", "delta_qtd", "delta", "quantidade"):
        numero = _numero(registro.get(chave))
        if numero is not None:
            if chave != "quantidade":
                return numero
            tipo = str(
                registro.get("tipo")
                or registro.get("tipo_movimento")
                or registro.get("tipo_alteracao")
                or ""
            ).strip().upper()
            if tipo.startswith(("SUPR", "DECR")):
                return -abs(numero)
            if tipo.startswith("ACR"):
                return abs(numero)
            return numero
    return None


def calcular_posicao_ciclo_por_data(
    *,
    ciclo: str,
    data_inicio: Any,
    data_fim: Any,
    data_posicao: Any,
    itens: Iterable[Mapping[str, Any]],
    movimentos: Iterable[Mapping[str, Any]] = (),
    consumo_registrado: Mapping[str, Any] | None = None,
    tolerancia: float = TOLERANCIA,
) -> dict[str, Any]:
    """Calcula a posição física/monetária na data ``D`` sem ler ou escrever XLS.

    O remanescente atual declarado é sempre autoritativo. Movimentos com
    ``data_efeito <= t0`` são considerados já refletidos na abertura; apenas
    ``t0 < data_efeito <= D`` integra o balanço do ciclo. Movimentos posteriores
    a ``D`` são ignorados nesta fotografia.
    """
    inicio = _data(data_inicio)
    fim = _data(data_fim)
    posicao = _data(data_posicao)
    erros: list[str] = []
    alertas: list[str] = []

    if inicio is None or fim is None or fim < inicio:
        erros.append("PERIODO_DO_CICLO_INVALIDO")
    if posicao is None:
        erros.append("DATA_DA_POSICAO_INVALIDA")
    elif inicio is not None and fim is not None and not (inicio <= posicao <= fim):
        erros.append("DATA_DA_POSICAO_FORA_DO_CICLO")

    movimentos_validos: list[dict[str, Any]] = []
    eventos_vistos: set[tuple[Any, ...]] = set()
    for bruto in movimentos or ():
        ident_item = _item(bruto.get("item"))
        efeito = _data(
            bruto.get("data_efeito")
            or bruto.get("data_aditivo")
            or bruto.get("data")
        )
        delta = _delta_movimento(bruto)
        if not ident_item or efeito is None or delta is None:
            erros.append("MOVIMENTO_QUANTITATIVO_INVALIDO")
            continue
        id_evento = _item(bruto.get("id_evento") or bruto.get("id"))
        chave_evento = (
            ("id", id_evento.casefold())
            if id_evento
            else ("conteudo", _chave_item(ident_item), efeito, round(delta, 8))
        )
        if chave_evento in eventos_vistos:
            erros.append(f"EVENTO_QUANTITATIVO_DUPLICADO:{ident_item}")
            continue
        eventos_vistos.add(chave_evento)
        movimentos_validos.append({
            "item": ident_item,
            "chave": _chave_item(ident_item),
            "data_efeito": efeito,
            "delta": delta,
            "vu_atualizado": _numero(
                bruto.get("vu_atualizado") or bruto.get("valor_unitario")
            ),
        })

    cadastros: dict[str, dict[str, Any]] = {}
    ordem: list[str] = []
    for bruto in itens or ():
        ident_item = _item(bruto.get("item"))
        chave = _chave_item(ident_item)
        if not ident_item:
            erros.append("ITEM_SEM_IDENTIFICADOR")
            continue
        if chave in cadastros:
            erros.append(f"ITEM_DUPLICADO:{ident_item}")
            continue
        cadastros[chave] = {
            "item": ident_item,
            "remanescente_inicio": _numero(
                bruto.get("remanescente_inicio")
                if "remanescente_inicio" in bruto
                else bruto.get("qtd_remanescente_inicio")
            ),
            "remanescente_atual": _numero(
                bruto.get("remanescente_atual")
                if "remanescente_atual" in bruto
                else bruto.get("qtd_remanescente_atual")
            ),
            "remanescente_atual_fornecido": (
                bruto.get("remanescente_atual") not in (None, "")
                if "remanescente_atual" in bruto
                else bruto.get("qtd_remanescente_atual") not in (None, "")
            ),
            "vu_atualizado": _numero(
                bruto.get("vu_atualizado") or bruto.get("vu")
            ),
            "novo_item": bool(bruto.get("novo_item")),
        }
        ordem.append(chave)

    # Movimento sem cadastro ainda aparece no motor puro, mas o VU ausente
    # permanece bloqueante. No XLS oficial, novos itens seguem cadastrados em
    # itens_Remanesc (N001, N002...) para manter a fonte oficial do VU.
    for mov in movimentos_validos:
        chave = mov["chave"]
        if chave in cadastros:
            continue
        cadastros[chave] = {
            "item": mov["item"],
            "remanescente_inicio": 0.0,
            "remanescente_atual": None,
            "remanescente_atual_fornecido": False,
            "vu_atualizado": mov.get("vu_atualizado"),
            "novo_item": True,
        }
        ordem.append(chave)

    if erros and (inicio is None or fim is None or posicao is None):
        return {
            "ciclo": str(ciclo or "").strip().upper(),
            "data_inicio": inicio,
            "data_fim": fim,
            "data_posicao": posicao,
            "itens": [],
            "completo": False,
            "valido": False,
            "erros": erros,
            "alertas": alertas,
            "total_valor_consumido": None,
            "total_valor_remanescente": None,
            "total_base_fisica_atualizada": None,
        }

    consumo_por_item = {
        _chave_item(k): _numero(v) for k, v in (consumo_registrado or {}).items()
    }
    linhas: list[dict[str, Any]] = []
    total_consumido = 0.0
    total_remanescente = 0.0
    total_base = 0.0
    completo = True

    for chave in ordem:
        cadastro = cadastros[chave]
        movs_item = [m for m in movimentos_validos if m["chave"] == chave]
        nascimentos = [m["data_efeito"] for m in movs_item if m["delta"] > 0]
        nascimento = min(nascimentos) if nascimentos else None
        novo = cadastro["novo_item"] or (
            cadastro["remanescente_inicio"] in (None, 0.0) and nascimento is not None
        )

        # Item novo posterior à fotografia ainda não compõe esta posição.
        if novo and nascimento is not None and nascimento > posicao:
            continue

        abertura = cadastro["remanescente_inicio"]
        if novo and nascimento is not None and nascimento > inicio:
            abertura = 0.0
        delta_periodo = round(sum(
            m["delta"] for m in movs_item if inicio < m["data_efeito"] <= posicao
        ), 2)
        disponivel = (
            round(abertura + delta_periodo, 2) if abertura is not None else None
        )
        atual = cadastro["remanescente_atual"]
        vu = cadastro["vu_atualizado"]
        erros_item: list[str] = []
        alertas_item: list[str] = []

        if abertura is None:
            erros_item.append("REMANESCENTE_INICIAL_AUSENTE")
        if not cadastro["remanescente_atual_fornecido"]:
            completo = False
            erros_item.append("REMANESCENTE_ATUAL_NAO_INFORMADO")
        elif atual is None or atual < 0:
            erros_item.append("REMANESCENTE_ATUAL_INVALIDO")
        if disponivel is not None and disponivel < -tolerancia:
            erros_item.append("BASE_FISICA_DISPONIVEL_NEGATIVA")
        if atual is not None and disponivel is not None and atual - disponivel > tolerancia:
            erros_item.append("REMANESCENTE_ATUAL_SUPERA_DISPONIVEL")
        if vu is None or vu < 0:
            erros_item.append("VU_ATUALIZADO_INVALIDO")

        consumida = None
        valor_consumido = None
        valor_remanescente = None
        valor_base = None
        check = None
        if atual is not None and disponivel is not None:
            consumida = round(disponivel - atual, 2)
            if consumida < -tolerancia:
                erros_item.append("CONSUMO_NEGATIVO")
            check = round(abertura + delta_periodo - consumida - atual, 2)
            if abs(check) > tolerancia:
                erros_item.append("CHECK_FISICO_DIVERGENTE")
        if consumida is not None and vu is not None:
            valor_consumido = round(consumida * vu, 2)
            valor_remanescente = round(atual * vu, 2)
            valor_base = round(disponivel * vu, 2)
            if abs(valor_consumido + valor_remanescente - valor_base) > tolerancia:
                erros_item.append("RECONCILIACAO_MONETARIA_DIVERGENTE")

        registrado = consumo_por_item.get(chave)
        divergencia = None
        if registrado is not None and consumida is not None:
            divergencia = round(consumida - registrado, 2)
            if abs(divergencia) > tolerancia:
                alertas_item.append("DIVERGENCIA_COM_ITENS_CONSUMIDOS")
                alertas.append(
                    f"{cadastro['item']}: consumo do balanço ({consumida:.2f}) "
                    f"diverge de itens_Consumidos ({registrado:.2f})."
                )

        if erros_item:
            erros.extend(f"{cadastro['item']}:{e}" for e in erros_item)
        elif valor_consumido is not None and valor_remanescente is not None:
            total_consumido += valor_consumido
            total_remanescente += valor_remanescente
            total_base += valor_base or 0.0

        linhas.append({
            "item": cadastro["item"],
            "remanescente_inicio": abertura,
            "alteracoes_liquidas_periodo": delta_periodo,
            "remanescente_atual": atual,
            "quantidade_consumida": consumida,
            "vu_atualizado": vu,
            "valor_consumido": valor_consumido,
            "valor_remanescente_atualizado": valor_remanescente,
            "valor_base_fisica_atualizada": valor_base,
            "check_fisico": check,
            "data_nascimento": nascimento,
            "suprimido_integralmente": (
                disponivel is not None and abs(disponivel) <= tolerancia
            ),
            "consumo_itens_conferencia": registrado,
            "divergencia_consumo": divergencia,
            "completo": not erros_item,
            "valido": not erros_item,
            "erros": erros_item,
            "alertas": alertas_item,
        })

    valido = completo and not erros and bool(linhas)
    return {
        "ciclo": str(ciclo or "").strip().upper(),
        "data_inicio": inicio,
        "data_fim": fim,
        "data_posicao": posicao,
        "itens": linhas,
        "completo": completo and all(i["completo"] for i in linhas),
        "valido": valido,
        "erros": list(dict.fromkeys(erros)),
        "alertas": list(dict.fromkeys(alertas)),
        "total_valor_consumido": round(total_consumido, 2) if valido else None,
        "total_valor_remanescente": round(total_remanescente, 2) if valido else None,
        "total_base_fisica_atualizada": round(total_base, 2) if valido else None,
    }


def identificar_layout_ciclo_em_execucao(wb) -> dict[str, Any]:
    """Distingue ausência, agregado legado e layout 2 itemizado."""
    if ABA_CICLO_EM_EXECUCAO not in wb.sheetnames:
        return {"tipo": "ausente", "versao": None}
    ws = wb[ABA_CICLO_EM_EXECUCAO]
    # ``ws[coord]`` materializa uma celula vazia no openpyxl. Consultar o mapa
    # existente evita qualquer mutacao estrutural ao apenas inspecionar uma aba
    # agregada legada.
    marcador_cell = ws._cells.get((2, 8))
    rotulo_cell = ws._cells.get((1, 8))
    marcador = str(marcador_cell.value if marcador_cell else "").strip()
    rotulo = str(rotulo_cell.value if rotulo_cell else "").strip()
    if marcador == VERSAO_LAYOUT_ITEMIZADO and rotulo == MARCADOR_LAYOUT:
        return {"tipo": "itemizado", "versao": marcador}
    return {"tipo": "legado", "versao": None}


def _formula_abertura(linha: int, origem: int) -> str:
    por_ciclo = {
        "C0": f"posicao_contratual!$F{origem}",
        "C1": f"posicao_contratual!$J{origem}",
        "C2": f"posicao_contratual!$N{origem}",
        "C3": f"posicao_contratual!$R{origem}",
        "C4": f"posicao_contratual!$V{origem}",
    }
    escolha = "\"\""
    for ciclo, ref in reversed(tuple(por_ciclo.items())):
        escolha = f'IF($C$3="{ciclo}",{ref},{escolha})'
    return (
        f'=IF(A{linha}="","",IF(AND(posicao_contratual!$Z{origem},'
        f'COUNTIFS(aditivos!$A$2:$A$200,itens_Remanesc!$A{origem},'
        f'aditivos!$B$2:$B$200,"<="&$F$3,'
        f'aditivos!$L$2:$L$200,">0")=0),0,{escolha}))'
    )


def _formula_vu(linha: int, origem: int) -> str:
    por_ciclo = {
        "C0": f"historico_VU!$C{origem}",
        "C1": f"historico_VU!$D{origem}",
        "C2": f"historico_VU!$E{origem}",
        "C3": f"historico_VU!$F{origem}",
        "C4": f"historico_VU!$G{origem}",
    }
    escolha = "\"\""
    for ciclo, ref in reversed(tuple(por_ciclo.items())):
        escolha = f'IF($C$3="{ciclo}",{ref},{escolha})'
    return f'=IF(A{linha}="","",{escolha})'


def _formula_conferencia(linha: int, *, contar: bool) -> str:
    colunas = {"C0": "E", "C1": "G", "C2": "I", "C3": "K", "C4": "M"}
    escolha = "\"\""
    for ciclo, coluna in reversed(tuple(colunas.items())):
        if contar:
            expr = (
                f'COUNTIFS(itens_Consumidos!$A$2:$A$200,A{linha},'
                f'itens_Consumidos!${coluna}$2:${coluna}$200,"<>")'
            )
        else:
            expr = (
                f'SUMIFS(itens_Consumidos!${coluna}$2:${coluna}$200,'
                f'itens_Consumidos!$A$2:$A$200,A{linha})'
            )
        escolha = f'IF($C$3="{ciclo}",{expr},{escolha})'
    return f'=IF(A{linha}="","",{escolha})'


def _registrar_nome_layout(wb) -> None:
    if MARCADOR_LAYOUT in wb.defined_names:
        del wb.defined_names[MARCADOR_LAYOUT]
    wb.defined_names.add(DefinedName(
        MARCADOR_LAYOUT,
        attr_text=f"'{ABA_CICLO_EM_EXECUCAO}'!${CELULA_VERSAO[0]}${CELULA_VERSAO[1:]}",
    ))


def _criar_aba_itemizada(wb):
    indice = (
        wb.sheetnames.index("posicao_referencia") + 1
        if "posicao_referencia" in wb.sheetnames
        else len(wb.sheetnames)
    )
    ws = wb.create_sheet(ABA_CICLO_EM_EXECUCAO, indice)
    ws.sheet_properties.tabColor = "FF0F5B50"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = f"A{PRIMEIRA_LINHA_ITEM}"
    ws.sheet_view.topLeftCell = "A1"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = f"A1:G{ULTIMA_LINHA_ITEM}"
    ws.print_title_rows = f"1:{LINHA_CABECALHO}"

    branco = "FFFFFFFF"
    verde = "FF0F5B50"
    verde_auto = "FFE8F5F1"
    verde_input = "FFC6EFCE"
    laranja = "FFE67E22"
    laranja_claro = "FFFCE4D6"
    azul = "FF1F4E78"
    azul_claro = "FFDDEBF7"
    cinza = "FF4F5B66"
    cinza_claro = "FFF2F2F2"
    amarelo = "FFFFC000"
    vermelho_claro = "FFFFC7CE"

    fina = Side(style="thin", color="FFB7B7B7")
    media_verde = Side(style="medium", color=verde)
    grossa_azul = Side(style="thick", color=azul)
    borda = Border(left=fina, right=fina, top=fina, bottom=fina)
    borda_input = Border(
        left=media_verde, right=media_verde, top=media_verde, bottom=media_verde
    )
    borda_total = Border(
        left=grossa_azul, right=grossa_azul, top=grossa_azul, bottom=grossa_azul
    )

    ws.merge_cells("A1:G1")
    ws["A1"] = "POSIÇÃO ATUAL DO REMANESCENTE NO CICLO EM EXECUÇÃO"
    ws["A1"].font = Font(name="Aptos Display", size=16, bold=True, color=branco)
    ws["A1"].fill = PatternFill("solid", fgColor=verde)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].border = borda_input
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A3:B3")
    ws["A3"] = "CICLO EM EXECUÇÃO (AUTO)"
    ws["C3"] = "=CONTROLE!$B$2"
    ws.merge_cells("D3:E3")
    ws["D3"] = "DATA DE INÍCIO DO CICLO (AUTO)"
    ws.merge_cells("F3:G3")
    ws["F3"] = (
        '=IFERROR(INDEX(parametros!$C$2:$C$6,'
        'MATCH($C$3,parametros!$B$2:$B$6,0)),"")'
    )
    ws["H3"] = (
        '=IFERROR(INDEX(parametros!$D$2:$D$6,'
        'MATCH($C$3,parametros!$B$2:$B$6,0)),"")'
    )
    ws["H3"].number_format = FORMATO_DATA
    for coord in ("A3", "D3"):
        ws[coord].fill = PatternFill("solid", fgColor=cinza)
        ws[coord].font = Font(name="Aptos", size=10, bold=True, color=branco)
        ws[coord].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws[coord].border = borda
    for coord in ("C3", "F3"):
        ws[coord].fill = PatternFill("solid", fgColor=cinza_claro)
        ws[coord].font = Font(name="Aptos", size=11, bold=True, color="FF1F1F1F")
        ws[coord].alignment = Alignment(horizontal="center", vertical="center")
        ws[coord].border = borda
    ws["F3"].number_format = FORMATO_DATA
    ws.row_dimensions[3].height = 28

    ws.merge_cells("A5:C5")
    ws["A5"] = "DATA DA POSIÇÃO ATUAL (PREENCHER)"
    ws.merge_cells("D5:G5")
    ws["D5"] = None
    for coord in ("A5", "D5"):
        ws[coord].fill = PatternFill("solid", fgColor=amarelo)
        ws[coord].font = Font(name="Aptos Display", size=14, bold=True, color="FF3F2A00")
        ws[coord].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws[coord].border = Border(
            left=Side(style="thick", color="FF9C6500"),
            right=Side(style="thick", color="FF9C6500"),
            top=Side(style="thick", color="FF9C6500"),
            bottom=Side(style="thick", color="FF9C6500"),
        )
    ws["D5"].number_format = FORMATO_DATA
    ws["D5"].protection = Protection(locked=False)
    ws.row_dimensions[5].height = 34

    ws.merge_cells("A7:G7")
    ws["A7"] = "VALOR TOTAL ATUALIZADO REMANESCENTE NA DATA DA POSIÇÃO (AUTO)"
    ws.merge_cells("A8:G8")
    ws["A8"] = (
        '=IF($D$5="","Data da posição: [PREENCHER]",'
        '"Data da posição: "&RIGHT("0"&DAY($D$5),2)&"/"&'
        'RIGHT("0"&MONTH($D$5),2)&"/"&YEAR($D$5))'
    )
    ws.merge_cells("A9:G10")
    ws["A9"] = (
        '=IF(OR($D$5="",COUNTIF($A$13:$A$211,"<>")=0,'
        'COUNTIF($K$13:$K$211,"ERRO:*")>0,'
        'COUNTIF($K$13:$K$211,"INCOMPLETO:*")>0),"",'
        'ROUND(SUM($G$13:$G$211),2))'
    )
    for coord in ("A7", "A8", "A9"):
        ws[coord].fill = PatternFill("solid", fgColor=azul)
        ws[coord].font = Font(
            name="Aptos Display",
            size=12 if coord != "A9" else 22,
            bold=True,
            color=branco,
        )
        ws[coord].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws[coord].border = borda_total
    ws["A9"].number_format = FORMATO_MOEDA
    ws.row_dimensions[7].height = 28
    ws.row_dimensions[8].height = 23
    ws.row_dimensions[9].height = 26
    ws.row_dimensions[10].height = 26

    cores_cabecalho = (cinza, verde, verde, laranja, azul, azul, azul)
    for coluna, (rotulo, cor) in enumerate(zip(COLUNAS_VISIVEIS, cores_cabecalho), start=1):
        celula = ws.cell(LINHA_CABECALHO, coluna, rotulo)
        celula.fill = PatternFill("solid", fgColor=cor)
        celula.font = Font(name="Aptos", size=10, bold=True, color=branco)
        celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celula.border = borda
    ws.row_dimensions[LINHA_CABECALHO].height = 58

    ws["H1"] = MARCADOR_LAYOUT
    ws["H2"] = VERSAO_LAYOUT_ITEMIZADO
    for idx, rotulo in enumerate(COLUNAS_TECNICAS, start=9):
        ws.cell(LINHA_CABECALHO, idx, rotulo)

    for linha in range(PRIMEIRA_LINHA_ITEM, ULTIMA_LINHA_ITEM + 1):
        origem = linha - PRIMEIRA_LINHA_ITEM + 2
        ws.cell(linha, 15).value = (
            f'=IF(itens_Remanesc!$A{origem}="","",'
            f'COUNTIFS(aditivos!$A$2:$A$200,itens_Remanesc!$A{origem},'
            f'aditivos!$B$2:$B$200,"<="&$D$5,'
            f'aditivos!$L$2:$L$200,">0"))'
        )
        ws.cell(linha, 1).value = (
            f'=IF(itens_Remanesc!$A{origem}="","",'
            f'IF(NOT(posicao_contratual!$Z{origem}),itens_Remanesc!$A{origem},'
            f'IF(AND(ISNUMBER($D$5),$O{linha}>0),'
            f'itens_Remanesc!$A{origem},"")))'
        )
        ws.cell(linha, 2).value = _formula_abertura(linha, origem)
        ws.cell(linha, 3).value = None  # vazio real; zero somente quando confirmado
        ws.cell(linha, 4).value = (
            f'=IF(OR(A{linha}="",$D$5="",C{linha}=""),"",'
            f'ROUND(B{linha}+I{linha}-C{linha},2))'
        )
        ws.cell(linha, 5).value = _formula_vu(linha, origem)
        ws.cell(linha, 6).value = (
            f'=IF(OR(D{linha}="",E{linha}=""),"",ROUND(D{linha}*E{linha},2))'
        )
        ws.cell(linha, 7).value = (
            f'=IF(OR(C{linha}="",E{linha}=""),"",ROUND(C{linha}*E{linha},2))'
        )
        ws.cell(linha, 9).value = (
            f'=IF(OR(A{linha}="",$D$5="",$F$3=""),"",'
            f'ROUND(SUMIFS(aditivos!$L$2:$L$200,'
            f'aditivos!$A$2:$A$200,A{linha},'
            f'aditivos!$B$2:$B$200,">"&$F$3,'
            f'aditivos!$B$2:$B$200,"<="&$D$5),2))'
        )
        ws.cell(linha, 10).value = (
            f'=IF(OR(A{linha}="",D{linha}="",C{linha}="",I{linha}=""),"",'
            f'ROUND(B{linha}+I{linha}-D{linha}-C{linha},2))'
        )
        ws.cell(linha, 11).value = (
            f'=IF(A{linha}="","",IF(OR(NOT(ISNUMBER($D$5)),'
            f'NOT(ISNUMBER($F$3)),NOT(ISNUMBER($H$3)),$D$5<$F$3,$D$5>$H$3),'
            f'"ERRO: DATA FORA DO CICLO",IF(COUNTIF($A$13:$A$211,A{linha})>1,'
            f'"ERRO: ITEM DUPLICADO",IF(NOT(ISNUMBER(B{linha})),'
            f'"INCOMPLETO: REMANESCENTE INICIAL",IF(C{linha}="",'
            f'"INCOMPLETO: QTD ATUAL",IF(OR(NOT(ISNUMBER(C{linha})),C{linha}<0),'
            f'"ERRO: QTD INVALIDA",IF(NOT(ISNUMBER(I{linha})),'
            f'"INCOMPLETO: ADITIVOS",IF(C{linha}>B{linha}+I{linha},'
            f'"ERRO: REMANESCENTE SUPERA DISPONIVEL",IF(D{linha}<0,'
            f'"ERRO: CONSUMO NEGATIVO",IF(NOT(ISNUMBER(E{linha})),'
            f'"INCOMPLETO: VU",IF(ABS(J{linha})>0.005,'
            f'"ERRO: CHECK FISICO","OK")))))))))))'
        )
        ws.cell(linha, 12).value = _formula_conferencia(linha, contar=False)
        ws.cell(linha, 13).value = (
            f'=IF(OR(A{linha}="",D{linha}="",N{linha}=0),"",'
            f'ROUND(D{linha}-L{linha},2))'
        )
        ws.cell(linha, 14).value = _formula_conferencia(linha, contar=True)

        for coluna in range(1, 8):
            celula = ws.cell(linha, coluna)
            celula.font = Font(name="Aptos", size=10, color="FF1F1F1F")
            celula.alignment = Alignment(
                horizontal="left" if coluna == 1 else "right",
                vertical="center",
            )
            celula.border = borda_input if coluna == 3 else borda
            if coluna == 1:
                celula.fill = PatternFill("solid", fgColor=cinza_claro)
            elif coluna == 2:
                celula.fill = PatternFill("solid", fgColor=verde_auto)
            elif coluna == 3:
                celula.fill = PatternFill("solid", fgColor=verde_input)
                celula.protection = Protection(locked=False)
            elif coluna == 4:
                celula.fill = PatternFill("solid", fgColor=laranja_claro)
            else:
                celula.fill = PatternFill("solid", fgColor=azul_claro)
            if coluna in (2, 3, 4):
                celula.number_format = FORMATO_QTD
            elif coluna in (5, 6, 7):
                celula.number_format = FORMATO_MOEDA
        ws.row_dimensions[linha].height = 20

    larguras = {"A": 18, "B": 24, "C": 24, "D": 23, "E": 18, "F": 20, "G": 25}
    for coluna, largura in larguras.items():
        ws.column_dimensions[coluna].width = largura
    for coluna in range(8, 16):
        ws.column_dimensions[get_column_letter(coluna)].hidden = True

    dv_data = DataValidation(
        type="custom",
        formula1='=AND(ISNUMBER($D$5),ISNUMBER($F$3),ISNUMBER($H$3),$D$5>=$F$3,$D$5<=$H$3)',
        allow_blank=True,
    )
    dv_data.errorTitle = "Data fora do ciclo"
    dv_data.error = "Informe uma data válida dentro do ciclo em execução."
    dv_data.promptTitle = "Data da posição atual"
    dv_data.prompt = "Única data de referência da nova fotografia itemizada."
    dv_data.showErrorMessage = True
    dv_data.showInputMessage = True
    ws.add_data_validation(dv_data)
    dv_data.add(ws[CELULA_DATA_POSICAO])

    dv_qtd = DataValidation(
        type="custom",
        formula1=(
            '=OR(C13="",AND(A13<>"",ISNUMBER(C13),C13>=0,'
            'C13<=ROUND(B13+I13,2)))'
        ),
        allow_blank=True,
    )
    dv_qtd.errorTitle = "Quantidade incompatível"
    dv_qtd.error = (
        "Informe zero ou uma quantidade não negativa que não supere o "
        "remanescente inicial acrescido das alterações do período."
    )
    dv_qtd.promptTitle = "Remanescente atual"
    dv_qtd.prompt = "Vazio = não confirmado; zero = saldo confirmado em zero."
    dv_qtd.showErrorMessage = True
    dv_qtd.showInputMessage = True
    ws.add_data_validation(dv_qtd)
    dv_qtd.add(f"C{PRIMEIRA_LINHA_ITEM}:C{ULTIMA_LINHA_ITEM}")

    ws.conditional_formatting.add(
        CELULA_DATA_POSICAO,
        FormulaRule(
            formula=['$D$5=""'],
            fill=PatternFill("solid", fgColor="FFFFE699"),
            font=Font(name="Aptos Display", size=14, bold=True, color="FF9C0006"),
        ),
    )
    ws.conditional_formatting.add(
        f"A{PRIMEIRA_LINHA_ITEM}:G{ULTIMA_LINHA_ITEM}",
        FormulaRule(
            formula=[
                f'AND($A{PRIMEIRA_LINHA_ITEM}<>"",ISNUMBER($B{PRIMEIRA_LINHA_ITEM}),'
                f'ISNUMBER($I{PRIMEIRA_LINHA_ITEM}),ROUND($B{PRIMEIRA_LINHA_ITEM}+'
                f'$I{PRIMEIRA_LINHA_ITEM},2)=0)'
            ],
            fill=PatternFill("solid", fgColor="FFD9D9D9"),
            font=Font(name="Aptos", size=10, color="FF7F7F7F"),
        ),
    )
    ws.conditional_formatting.add(
        f"C{PRIMEIRA_LINHA_ITEM}:C{ULTIMA_LINHA_ITEM}",
        FormulaRule(
            formula=[
                f'AND($A{PRIMEIRA_LINHA_ITEM}<>"",$C{PRIMEIRA_LINHA_ITEM}<>"",'
                f'OR(NOT(ISNUMBER($C{PRIMEIRA_LINHA_ITEM})),'
                f'$C{PRIMEIRA_LINHA_ITEM}<0,'
                f'$C{PRIMEIRA_LINHA_ITEM}>$B{PRIMEIRA_LINHA_ITEM}+$I{PRIMEIRA_LINHA_ITEM}))'
            ],
            fill=PatternFill("solid", fgColor=vermelho_claro),
        ),
    )

    ws.protection.sheet = True
    ws.protection.selectLockedCells = False
    ws.protection.selectUnlockedCells = True
    _registrar_nome_layout(wb)
    return ws


def garantir_aba_ciclo_em_execucao(wb, *, limpar_entradas: bool = False) -> dict[str, Any]:
    """Cria a aba itemizada sem substituir uma aba agregada legada homônima."""
    layout = identificar_layout_ciclo_em_execucao(wb)
    if layout["tipo"] == "legado":
        return {"criada": False, "legado_preservado": True, **layout}
    if layout["tipo"] == "ausente":
        ws = _criar_aba_itemizada(wb)
        criada = True
    else:
        ws = wb[ABA_CICLO_EM_EXECUCAO]
        _registrar_nome_layout(wb)
        criada = False
    if limpar_entradas:
        ws[CELULA_DATA_POSICAO] = None
        ws[CELULA_DATA_POSICAO].protection = Protection(locked=False)
        for linha in range(PRIMEIRA_LINHA_ITEM, ULTIMA_LINHA_ITEM + 1):
            ws.cell(linha, 3).value = None
            ws.cell(linha, 3).protection = Protection(locked=False)
    return {
        "criada": criada,
        "legado_preservado": False,
        "tipo": "itemizado",
        "versao": VERSAO_LAYOUT_ITEMIZADO,
    }


def _periodo_por_ciclo(wb, ciclo: str) -> tuple[date | None, date | None]:
    if "parametros" not in wb.sheetnames:
        return None, None
    ws = wb["parametros"]
    for linha in range(2, 7):
        if str(ws.cell(linha, 2).value or "").strip().upper() == ciclo:
            return _data(ws.cell(linha, 3).value), _data(ws.cell(linha, 4).value)
    return None, None


def validar_posicao_ciclo_lida(
    *,
    ciclo: str,
    data_inicio: Any,
    data_fim: Any,
    data_posicao: Any,
    itens: list[dict[str, Any]],
    tolerancia: float = TOLERANCIA,
) -> dict[str, Any]:
    """Validação Python independente dos resultados materializados pelo Excel."""
    inicio, fim, posicao = _data(data_inicio), _data(data_fim), _data(data_posicao)
    erros: list[str] = []
    alertas: list[str] = []
    if posicao is None:
        erros.append("DATA_DA_POSICAO_INVALIDA")
    elif inicio is None or fim is None or not (inicio <= posicao <= fim):
        erros.append("DATA_DA_POSICAO_FORA_DO_CICLO")

    chaves: set[str] = set()
    total_consumido = 0.0
    total_remanescente = 0.0
    total_base = 0.0
    completo = bool(itens)
    for registro in itens:
        ident = _item(registro.get("item"))
        chave = _chave_item(ident)
        if chave in chaves:
            erros.append(f"{ident}:ITEM_DUPLICADO")
        chaves.add(chave)
        abertura = _numero(registro.get("remanescente_inicio"))
        delta = _numero(registro.get("alteracoes_liquidas_periodo"))
        atual_fornecido = registro.get("remanescente_atual") not in (None, "")
        atual = _numero(registro.get("remanescente_atual"))
        consumida = _numero(registro.get("quantidade_consumida"))
        vu = _numero(registro.get("vu_atualizado"))
        check = _numero(registro.get("check_fisico"))
        erros_item: list[str] = []
        status_xls = str(registro.get("status_xls") or "").strip()
        status_xls_normalizado = status_xls.upper()
        if status_xls_normalizado.startswith("INCOMPLETO:"):
            completo = False
            erros_item.append(f"STATUS_XLS_{status_xls_normalizado}")
        elif status_xls_normalizado.startswith("ERRO:"):
            erros_item.append(f"STATUS_XLS_{status_xls_normalizado}")
        if abertura is None:
            erros_item.append("REMANESCENTE_INICIAL_INDISPONIVEL")
        if delta is None:
            erros_item.append("ALTERACOES_DO_PERIODO_INDISPONIVEIS")
        if not atual_fornecido:
            completo = False
            erros_item.append("REMANESCENTE_ATUAL_NAO_INFORMADO")
        elif atual is None or atual < 0:
            erros_item.append("REMANESCENTE_ATUAL_INVALIDO")
        if vu is None or vu < 0:
            erros_item.append("VU_ATUALIZADO_INVALIDO")
        if abertura is not None and delta is not None and atual is not None:
            disponivel = round(abertura + delta, 2)
            esperado = round(disponivel - atual, 2)
            if atual - disponivel > tolerancia:
                erros_item.append("REMANESCENTE_ATUAL_SUPERA_DISPONIVEL")
            if consumida is None or abs(consumida - esperado) > tolerancia:
                erros_item.append("CONSUMO_DIVERGENTE_DO_BALANCO")
            if esperado < -tolerancia:
                erros_item.append("CONSUMO_NEGATIVO")
            if check is None or abs(check) > tolerancia:
                erros_item.append("CHECK_FISICO_DIVERGENTE")
            if vu is not None and not erros_item:
                valor_consumido = round(esperado * vu, 2)
                valor_remanescente = round(atual * vu, 2)
                valor_base = round(disponivel * vu, 2)
                if abs(valor_consumido + valor_remanescente - valor_base) > tolerancia:
                    erros_item.append("RECONCILIACAO_MONETARIA_DIVERGENTE")
                else:
                    total_consumido += valor_consumido
                    total_remanescente += valor_remanescente
                    total_base += valor_base
        erros.extend(f"{ident}:{e}" for e in erros_item)

    valido = completo and bool(itens) and not erros
    return {
        "ciclo": ciclo,
        "data_inicio": inicio,
        "data_fim": fim,
        "data_posicao": posicao,
        "completo": completo and not any("NAO_INFORMADO" in e for e in erros),
        "valido": valido,
        "erros": list(dict.fromkeys(erros)),
        "alertas": alertas,
        "total_valor_consumido": round(total_consumido, 2) if valido else None,
        "total_valor_remanescente": round(total_remanescente, 2) if valido else None,
        "total_base_fisica_atualizada": round(total_base, 2) if valido else None,
    }


def reconciliar_com_itens_consumidos(
    posicao: Mapping[str, Any],
    itens_consumidos: Mapping[str, Any] | None,
    *,
    tolerancia: float = TOLERANCIA,
) -> list[dict[str, Any]]:
    """Compara consumos sem substituir a fotografia declarada pelo fiscal."""
    ciclo = str(posicao.get("ciclo") or "").strip().upper()
    registrados: dict[str, float] = {}
    for registro in (itens_consumidos or {}).get("itens") or []:
        quantidade = _numero(((registro.get("consumos") or {}).get(ciclo) or {}).get("qtd"))
        if quantidade is not None:
            registrados[_chave_item(registro.get("item"))] = quantidade
    divergencias: list[dict[str, Any]] = []
    for registro in posicao.get("itens") or []:
        chave = _chave_item(registro.get("item"))
        calculado = _numero(registro.get("quantidade_consumida"))
        if chave not in registrados or calculado is None:
            continue
        divergencia = round(calculado - registrados[chave], 2)
        if abs(divergencia) > tolerancia:
            divergencias.append({
                "item": registro.get("item"),
                "ciclo": ciclo,
                "consumo_balanco": calculado,
                "consumo_itens_consumidos": registrados[chave],
                "divergencia": divergencia,
                "natureza": "DIAGNOSTICO_NAO_SUBSTITUTIVO",
            })
    return divergencias


def ler_ciclo_em_execucao(
    wb,
    *,
    itens_consumidos: Mapping[str, Any] | None = None,
) -> dict[str, Any]:
    """Lê somente o layout marcado como 2_ITEMIZADO.

    Aba ausente, aba vazia e agregado legado não interferem nos cálculos atuais.
    """
    layout = identificar_layout_ciclo_em_execucao(wb)
    resultado: dict[str, Any] = {
        "disponivel": layout["tipo"] == "itemizado",
        "utilizado": False,
        "layout_tipo": layout["tipo"],
        "layout_version": layout.get("versao"),
        "legado_ignorado": layout["tipo"] == "legado",
        "itens": [],
        "diagnosticos": [],
        "alertas": [],
        "erros": [],
        "completo": False,
        "valido": False,
    }
    if layout["tipo"] != "itemizado":
        return resultado

    ws = wb[ABA_CICLO_EM_EXECUCAO]
    encontrados = tuple(ws.cell(LINHA_CABECALHO, c).value for c in range(1, 8))
    if encontrados != COLUNAS_VISIVEIS:
        resultado["erros"].append("CABECALHOS_DO_LAYOUT_ITEMIZADO_INVALIDOS")
        return resultado

    ciclo = str(ws[CELULA_CICLO].value or "").strip().upper()
    if not re.fullmatch(r"C[0-4]", ciclo) and "CONTROLE" in wb.sheetnames:
        ciclo = str(wb["CONTROLE"]["B2"].value or "").strip().upper()
    inicio = _data(ws[CELULA_INICIO_CICLO].value)
    fim = _data(ws[CELULA_FIM_CICLO_TECNICA].value)
    if inicio is None or fim is None:
        inicio_fallback, fim_fallback = _periodo_por_ciclo(wb, ciclo)
        inicio = inicio or inicio_fallback
        fim = fim or fim_fallback
    data_posicao = _data(ws[CELULA_DATA_POSICAO].value)

    itens_lidos: list[dict[str, Any]] = []
    entrada_quantitativa_presente = False
    for linha in range(PRIMEIRA_LINHA_ITEM, ULTIMA_LINHA_ITEM + 1):
        item_bruto = ws.cell(linha, 1).value
        # O leitor publico normalmente recebe data_only=True. Este guard torna
        # a funcao segura tambem em inspecoes estruturais data_only=False: uma
        # formula ainda sem cache nao e um identificador de item.
        ident = "" if isinstance(item_bruto, str) and item_bruto.startswith("=") else _item(item_bruto)
        atual_bruto = ws.cell(linha, 3).value
        if atual_bruto not in (None, ""):
            entrada_quantitativa_presente = True
        if not ident and atual_bruto not in (None, "") and "itens_Remanesc" in wb.sheetnames:
            origem = linha - PRIMEIRA_LINHA_ITEM + 2
            ident = _item(wb["itens_Remanesc"].cell(origem, 1).value)
        if not ident:
            continue
        itens_lidos.append({
            "origem_linha": linha,
            "item": ident,
            "remanescente_inicio": ws.cell(linha, 2).value,
            "remanescente_atual": atual_bruto,
            "quantidade_consumida": ws.cell(linha, 4).value,
            "vu_atualizado": ws.cell(linha, 5).value,
            "valor_consumido": ws.cell(linha, 6).value,
            "valor_remanescente_atualizado": ws.cell(linha, 7).value,
            "alteracoes_liquidas_periodo": ws.cell(linha, 9).value,
            "check_fisico": ws.cell(linha, 10).value,
            "status_xls": ws.cell(linha, 11).value,
            "consumo_itens_conferencia_xls": ws.cell(linha, 12).value,
            "divergencia_consumo_xls": ws.cell(linha, 13).value,
            "qtd_eventos_positivos_ate_data": ws.cell(linha, 15).value,
        })

    utilizado = data_posicao is not None or entrada_quantitativa_presente
    resultado.update({
        "utilizado": utilizado,
        "ciclo": ciclo,
        "data_inicio": inicio,
        "data_fim": fim,
        "data_posicao": data_posicao,
        "itens": itens_lidos,
        "total_xls": ws[CELULA_TOTAL].value,
    })
    if not utilizado:
        return resultado

    validacao = validar_posicao_ciclo_lida(
        ciclo=ciclo,
        data_inicio=inicio,
        data_fim=fim,
        data_posicao=data_posicao,
        itens=itens_lidos,
    )
    resultado.update({
        "completo": validacao["completo"],
        "valido": validacao["valido"],
        "erros": validacao["erros"],
        "alertas": validacao["alertas"],
        "total_valor_consumido": validacao["total_valor_consumido"],
        "total_valor_remanescente": validacao["total_valor_remanescente"],
        "total_base_fisica_atualizada": validacao["total_base_fisica_atualizada"],
    })
    divergencias = reconciliar_com_itens_consumidos(resultado, itens_consumidos)
    resultado["diagnosticos"] = divergencias
    for divergencia in divergencias:
        resultado["alertas"].append(
            f"CICLO_EM_EXECUCAO: item {divergencia['item']} diverge de "
            "itens_Consumidos; a declaração atual foi preservada."
        )
    return resultado
