from __future__ import annotations

import argparse
from copy import copy
from pathlib import Path
import re

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


INPUT = PatternFill("solid", fgColor="FFFFF2CC")
AUTO = PatternFill("solid", fgColor="FFEDEDED")
HEADER = PatternFill("solid", fgColor="FF1F4E79")
CHECK = PatternFill("solid", fgColor="FFE2EFDA")
HISTORICAL_REQUIRED = PatternFill("solid", fgColor="FFFCE4D6")
WHITE = "FFFFFFFF"
NAVY = "FF1F4E79"
GRAY = "FF595959"
MONEY = 'R$ #,##0.00'
QTY = '#,##0.00'
PERCENT = '0.00%'
FACTOR = '#,##0.0000'
THIN = Side(style="thin", color="FFD9D9D9")
GRID = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _assert_resultados_integra(wb, etapa: str) -> None:
    if "RESULTADOS" not in wb.sheetnames:
        raise ValueError(f"A aba RESULTADOS desapareceu na etapa {etapa}.")
    ws = wb["RESULTADOS"]
    formulas = sum(
        1
        for row in ws.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    )
    conteudo = sum(1 for row in ws.iter_rows() for cell in row if cell.value not in (None, ""))
    if ws.sheet_state != "visible":
        raise ValueError(f"A aba RESULTADOS não está visível na etapa {etapa}.")
    if ws["A1"].value != "RESULTADOS CONSOLIDADOS — REAJUSTE CONTRATUAL":
        raise ValueError(f"A aba RESULTADOS está vazia ou foi substituída na etapa {etapa}.")
    if formulas < 3000 or conteudo < 3300:
        raise ValueError(
            f"A aba RESULTADOS perdeu conteúdo na etapa {etapa}: "
            f"{formulas} fórmulas e {conteudo} células preenchidas."
        )


def _deslocar_referencia_parametros(match: re.Match[str]) -> str:
    def deslocar(coluna: str) -> str:
        return chr(ord(coluna.upper()) - 1) if coluna.upper() >= "D" else coluna.upper()

    prefixo, cifra1, coluna1, linha1, cifra2, coluna2, linha2 = match.groups()
    resultado = f"{prefixo}{cifra1}{deslocar(coluna1)}{linha1 or ''}"
    if coluna2:
        resultado += f":{cifra2}{deslocar(coluna2)}{linha2 or ''}"
    return resultado


REFERENCIA_PARAMETROS = re.compile(
    r"(parametros!)(\$?)([A-H])(\$?\d+)?(?:\:(\$?)([A-H])(\$?\d+)?)?",
    re.IGNORECASE,
)


def _excluir_campo_c_parametros(wb) -> None:
    """Exclui a antiga coluna C e desloca somente referências a PARAMETROS."""

    ws = wb["parametros"]
    cabecalho = str(ws["C1"].value or "").strip().upper()
    if cabecalho == "DATA_INICIO":
        return
    if cabecalho not in {"", "PERIODO_DO_CICLO"}:
        raise ValueError(f"Cabeçalho inesperado em parametros!C1: {ws['C1'].value!r}")
    ws.delete_cols(3, 1)

    for planilha in wb.worksheets:
        for row in planilha.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    cell.value = REFERENCIA_PARAMETROS.sub(_deslocar_referencia_parametros, cell.value)
        for regras in planilha.conditional_formatting._cf_rules.values():
            for regra in regras:
                if regra.formula:
                    regra.formula = [
                        REFERENCIA_PARAMETROS.sub(_deslocar_referencia_parametros, formula)
                        for formula in regra.formula
                    ]
        for validacao in planilha.data_validations.dataValidation:
            if validacao.formula1:
                validacao.formula1 = REFERENCIA_PARAMETROS.sub(
                    _deslocar_referencia_parametros, validacao.formula1
                )
            if validacao.formula2:
                validacao.formula2 = REFERENCIA_PARAMETROS.sub(
                    _deslocar_referencia_parametros, validacao.formula2
                )

    for nome in wb.defined_names.values():
        if nome.attr_text:
            nome.attr_text = REFERENCIA_PARAMETROS.sub(_deslocar_referencia_parametros, nome.attr_text)


def _style_formula_range(ws, row_start: int, row_end: int, formulas: dict[str, str]) -> None:
    for row in range(row_start, row_end + 1):
        for col, pattern in formulas.items():
            ws[f"{col}{row}"] = pattern.format(r=row)
            ws[f"{col}{row}"].fill = AUTO
            ws[f"{col}{row}"].font = Font(name="Calibri", size=10, color=GRAY)
            ws[f"{col}{row}"].border = GRID


def _reset_controle(wb) -> None:
    ws = wb["CONTROLE"]
    ws["B1"] = "Principal"
    ws["B2"] = "C0"
    ws["B3"] = None
    ws["B7"] = None
    ws["B8"] = None
    ws["A7"] = "Índice utilizado"
    ws["A8"] = "Data-base original"
    ws["A9"] = "Quantidade de ciclos desta análise"
    ws["A10"] = "Variação acumulada final"
    ws["A11"] = "Fator acumulado total"
    ws["A12"] = "Último ciclo considerado nesta apuração"
    ws["A13"] = "Início do último ciclo considerado"
    for ref in ("B3", "B8", "B13", "B14"):
        ws[ref].number_format = "mm/yyyy"
    for row in (10, 11, 12, 13):
        ws.row_dimensions[row].hidden = False
    ws.row_dimensions[14].hidden = True
    ws.sheet_view.showGridLines = False


def _reset_parametros(wb) -> None:
    ws = wb["parametros"]
    ws.conditional_formatting._cf_rules.clear()
    if ws.max_column > 7:
        ws.delete_cols(8, ws.max_column - 7)
    headers = [
        "COMPUTAR_NESTA_APURACAO", "CICLO", "DATA_INICIO", "DATA_FIM",
        "PERCENTUAL_DO_CICLO", "FATOR_ACUMULADO", "SITUACAO",
    ]
    for col, value in enumerate(headers, 1):
        cell = ws.cell(1, col)
        cell.value = value
        cell.fill = HEADER
        cell.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = GRID
    for row, cycle in enumerate(("C0", "C1", "C2", "C3", "C4"), 2):
        ws.cell(row, 1, "Nao").fill = INPUT
        ws.cell(row, 2, cycle)
        for col in (3, 4):
            ws.cell(row, col).value = None
            ws.cell(row, col).number_format = "mm/yyyy"
        ws.cell(row, 5).value = 0 if cycle == "C0" else None
        ws.cell(row, 5).fill = INPUT
        ws.cell(row, 5).number_format = PERCENT
        if cycle == "C0":
            ws.cell(row, 6, "=1")
        else:
            prev = row - 1
            ws.cell(row, 6, f'=IF(E{row}="","",IF(NOT(ISNUMBER(E{row})),"",IF(NOT(ISNUMBER(F{prev})),"",F{prev}*(1+E{row}))))')
        ws.cell(row, 6).fill = AUTO
        ws.cell(row, 6).number_format = FACTOR
        ws.cell(row, 7, "Base" if cycle == "C0" else "Fora desta apuracao")
        for col in range(1, 8):
            ws.cell(row, col).font = Font(name="Calibri", size=10, color=GRAY)
            ws.cell(row, col).border = GRID
            ws.cell(row, col).alignment = Alignment(vertical="center")

    if ws.merged_cells:
        for merged in list(ws.merged_cells.ranges):
            ws.unmerge_cells(str(merged))
    ws["A9"] = "MEMORIA DO FATOR APLICAVEL A APURACAO"
    ws.merge_cells("A9:F9")
    ws["A9"].font = Font(name="Calibri", size=10, bold=True, color=NAVY)
    for col, value in enumerate(("CICLO", "COMPUTA?", "PERCENTUAL", "FATOR", "FATOR_ACUMULADO_APURACAO", "STATUS"), 1):
        c = ws.cell(10, col, value)
        c.fill = HEADER
        c.font = Font(name="Calibri", size=9, bold=True, color=WHITE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = GRID
    for i, cycle in enumerate(("C0", "C1", "C2", "C3", "C4")):
        row = 11 + i
        src = 2 + i
        ws.cell(row, 1, cycle)
        ws.cell(row, 2, f"=A{src}")
        ws.cell(row, 3, f'=IF(B{row}="Sim",E{src},"")')
        ws.cell(row, 4, "=1" if cycle == "C0" else f'=IF(C{row}="","",1+C{row})')
        ws.cell(row, 5, "=1" if cycle == "C0" else f'=IF(B{row}="Sim",E{row-1}*D{row},E{row-1})')
        ws.cell(row, 6, "Base" if cycle == "C0" else f'=IF(B{row}="Sim","Aplicado","Fora da apuracao")')
        for col in range(1, 7):
            c = ws.cell(row, col)
            c.fill = CHECK if cycle != "C0" else AUTO
            c.font = Font(name="Calibri", size=9, color=GRAY)
            c.border = GRID
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row, 3).number_format = PERCENT
        ws.cell(row, 4).number_format = FACTOR
        ws.cell(row, 5).number_format = FACTOR
    for col in range(1, 7):
        ws.cell(16, col).fill = PatternFill("solid", fgColor="FFD9E1F2")
        ws.cell(16, col).font = Font(name="Calibri", size=9, bold=True, color=GRAY)
        ws.cell(16, col).border = GRID
    ws["A16"] = "TOTAL APURACAO"
    ws["E16"] = "=E15"
    ws["E16"].number_format = FACTOR
    ws["F16"] = "=E16-1"
    ws["F16"].number_format = PERCENT
    ws.data_validations.dataValidation.clear()
    dv = DataValidation(type="list", formula1='"Sim,Nao"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add("A2:A6")
    # F/G vazios em ciclo histórico anterior ao ciclo analisado exigem ação do fiscal.
    # A regra é reativa: ao informar o percentual, o destaque desaparece.
    ws.conditional_formatting.add(
        "E3:F6",
        FormulaRule(
            formula=['AND($A3="Nao",$C3<>"",$E3="")'],
            fill=HISTORICAL_REQUIRED,
            font=Font(color="FF9C5700", bold=True),
        ),
    )
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    widths = {"A": 30, "B": 10, "C": 16, "D": 16, "E": 23, "F": 20, "G": 28}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _reset_financeiro(wb) -> None:
    ws = wb["financeiro"]
    if ws.max_column > 7:
        ws.delete_cols(8, ws.max_column - 7)
    if ws.max_row > 61:
        ws.delete_rows(62, ws.max_row - 61)
    for col, value in enumerate(
        ("COMPETENCIA", "CICLO", "VALOR_PAGO", "FATOR_APLICAVEL", "VALOR_ATUALIZADO", "DELTA", "EFEITO_FINANCEIRO"),
        1,
    ):
        ws.cell(1, col, value)
    formulas = {
        "B": '=IF(A{r}="","",IF(OR(A{r}<MIN(parametros!$C$2:$C$6),A{r}>MAX(parametros!$D$2:$D$6)),"Fora dos ciclos",LOWER(LOOKUP(A{r},parametros!$C$2:$C$6,parametros!$B$2:$B$6))))',
        "D": '=IF(B{r}="","",IF(B{r}="c0",1,IF(B{r}="c1",IF(parametros!$A$3<>"Sim","",IF(NOT(ISNUMBER(parametros!$E$3)),"",1+parametros!$E$3)),IF(B{r}="c2",IF(parametros!$A$4<>"Sim","",IF(NOT(ISNUMBER(parametros!$E$4)),"",IF(parametros!$A$3="Sim",1+parametros!$E$3,1)*(1+parametros!$E$4))),IF(B{r}="c3",IF(parametros!$A$5<>"Sim","",IF(NOT(ISNUMBER(parametros!$E$5)),"",IF(parametros!$A$3="Sim",1+parametros!$E$3,1)*IF(parametros!$A$4="Sim",1+parametros!$E$4,1)*(1+parametros!$E$5))),IF(B{r}="c4",IF(parametros!$A$6<>"Sim","",IF(NOT(ISNUMBER(parametros!$E$6)),"",IF(parametros!$A$3="Sim",1+parametros!$E$3,1)*IF(parametros!$A$4="Sim",1+parametros!$E$4,1)*IF(parametros!$A$5="Sim",1+parametros!$E$5,1)*(1+parametros!$E$6))),""))))))',
        "E": '=IF(OR(C{r}="",NOT(ISNUMBER(D{r}))),"",ROUND(C{r}*D{r},2))',
        "F": '=IF(OR(C{r}="",E{r}=""),"",IF(G{r}<>"Sim",0,ROUND(E{r}-C{r},2)))',
    }
    for row in range(2, 62):
        for col in ("A", "C", "G"):
            ws[f"{col}{row}"] = None
            ws[f"{col}{row}"].border = GRID
            ws[f"{col}{row}"].font = Font(name="Calibri", size=10, color=GRAY)
        ws[f"A{row}"].number_format = "mm/yyyy"
        ws[f"B{row}"].fill = AUTO
        ws[f"C{row}"].number_format = MONEY
        ws[f"C{row}"].fill = INPUT
        ws[f"G{row}"].alignment = Alignment(horizontal="center")
    _style_formula_range(ws, 2, 61, formulas)
    for row in range(2, 62):
        ws[f"D{row}"].number_format = FACTOR
        ws[f"E{row}"].number_format = MONEY
        ws[f"F{row}"].number_format = MONEY
    ws.freeze_panes = None
    ws.sheet_view.showGridLines = False
    for col, width in {"A": 16, "B": 14, "C": 18, "D": 17, "E": 19, "F": 18, "G": 19}.items():
        ws.column_dimensions[col].width = width


def _reset_itens_remanesc(wb) -> None:
    ws = wb["itens_Remanesc"]
    ws.conditional_formatting._cf_rules.clear()
    for row in range(2, 201):
        for col in ("A", "B", "C", "E", "G", "I", "K", "S", "T"):
            ws[f"{col}{row}"] = None
    formulas = {
        "D": '=IF(OR(A{r}="",B{r}="",C{r}=""),"",ROUND(B{r}*C{r},2))',
        "F": '=IF(OR(A{r}="",E{r}="",C{r}="",NOT(ISNUMBER($Z$3))),"",ROUND(E{r}*C{r}*$Z$3,2))',
        "H": '=IF(OR(A{r}="",G{r}="",C{r}="",NOT(ISNUMBER($Z$4))),"",ROUND(G{r}*C{r}*$Z$4,2))',
        "J": '=IF(OR(A{r}="",I{r}="",C{r}="",NOT(ISNUMBER($Z$5))),"",ROUND(I{r}*C{r}*$Z$5,2))',
        "L": '=IF(OR(A{r}="",K{r}="",C{r}="",NOT(ISNUMBER($Z$6))),"",ROUND(K{r}*C{r}*$Z$6,2))',
        "M": '=IF(OR(E{r}="",G{r}=""),"",ROUND(MAX(E{r}-G{r},0),2))',
        "N": '=IF(OR(M{r}="",C{r}="",NOT(ISNUMBER($Z$3))),"",ROUND(M{r}*C{r}*$Z$3,2))',
        "O": '=IF(OR(G{r}="",I{r}=""),"",ROUND(MAX(G{r}-I{r},0),2))',
        "P": '=IF(OR(O{r}="",C{r}="",NOT(ISNUMBER($Z$4))),"",ROUND(O{r}*C{r}*$Z$4,2))',
        "Q": '=IF(OR(I{r}="",K{r}=""),"",ROUND(MAX(I{r}-K{r},0),2))',
        "R": '=IF(OR(Q{r}="",C{r}="",NOT(ISNUMBER($Z$5))),"",ROUND(Q{r}*C{r}*$Z$5,2))',
        "U": '=IF(A{r}="","",IF(AND(G{r}<>"",E{r}<>"",G{r}>E{r}),"ALERTA: REM_C2>REM_C1",IF(AND(I{r}<>"",G{r}<>"",I{r}>G{r}),"ALERTA: REM_C3>REM_C2",IF(AND(K{r}<>"",I{r}<>"",K{r}>I{r}),"ALERTA: REM_C4>REM_C3","OK"))))',
        "V": '=IF(A{r}="","",IF(B{r}="","",IF(OR(B{r}<0,E{r}>B{r},G{r}>B{r},I{r}>B{r},K{r}>B{r}),"ALERTA: QTD_INVALIDA","OK")))',
        "AB": '=IF(OR(A{r}="",E{r}=""),"",ROUND(B{r}-E{r},2))',
        "AC": '=IF(OR(A{r}="",E{r}=""),"",ROUND(D{r}-E{r}*C{r},2))',
    }
    _style_formula_range(ws, 2, 200, formulas)
    # Os valores totais aparecem na primeira linha livre depois do último item.
    # A coluna U identifica visualmente a linha dinâmica; os itens devem ser
    # preenchidos de forma contígua, como já exigido pelo fluxo de coleta.
    money_cols = ("D", "F", "H", "J", "L", "N", "P", "R", "T", "AC")
    base_formulas = dict(formulas)
    for row in range(2, 202):
        prev = row - 1
        following = row + 1
        is_total = (
            "FALSE"
            if row == 2
            else f'AND(A{row}="",A{prev}<>"",COUNTIF(A{following}:$A$200,"<>")=0)'
        )
        for col, pattern in base_formulas.items():
            if col in money_cols:
                total = f'ROUND(SUMIF($A$2:A{prev},"<>",${col}$2:{col}{prev}),2)'
                regular = pattern.format(r=row)
                formulas_row = f'=IF({is_total},{total},{regular[1:]})'
                ws[f"{col}{row}"] = formulas_row
                ws[f"{col}{row}"].fill = AUTO
                ws[f"{col}{row}"].font = Font(name="Calibri", size=10, color=GRAY)
                ws[f"{col}{row}"].border = GRID
        if row >= 3:
            ws[f"U{row}"] = (
                f'=IF({is_total},"TOTAL",'
                f'IF(A{row}="","",IF(AND(G{row}<>"",E{row}<>"",G{row}>E{row}),'
                '"ALERTA: REM_C2>REM_C1",IF(AND(I{r}<>"",G{r}<>"",I{r}>G{r}),'
                '"ALERTA: REM_C3>REM_C2",IF(AND(K{r}<>"",I{r}<>"",K{r}>I{r}),'
                '"ALERTA: REM_C4>REM_C3","OK")))))'.format(r=row)
            )
    ws.conditional_formatting.add(
        "A2:AC201",
        FormulaRule(
            formula=['$U2="TOTAL"'],
            fill=PatternFill("solid", fgColor="FFD9E1F2"),
            font=Font(color=NAVY, bold=True),
        ),
    )
    for row in range(2, 201):
        for col in ("A", "B", "C", "E", "G", "I", "K"):
            ws[f"{col}{row}"].fill = INPUT
            ws[f"{col}{row}"].border = GRID
        for col in ("B", "E", "G", "I", "K", "M", "O", "Q", "S", "AB"):
            ws[f"{col}{row}"].number_format = QTY
        for col in ("C", "D", "F", "H", "J", "L", "N", "P", "R", "T", "AC"):
            ws[f"{col}{row}"].number_format = MONEY
    for i, cycle in enumerate(("C0", "C1", "C2", "C3", "C4"), 2):
        ws[f"W{i}"] = cycle
        ws[f"X{i}"] = f"=parametros!$A${i}"
        ws[f"Y{i}"] = f'=IF(ISNUMBER(parametros!$E${i}),parametros!$E${i},"")'
        ws[f"Z{i}"] = f'=IF(ISNUMBER(parametros!$F${i}),parametros!$F${i},"")'
        for col in ("W", "X", "Y", "Z"):
            ws[f"{col}{i}"].fill = AUTO
            ws[f"{col}{i}"].border = GRID
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


def _reset_itens_consumidos(wb) -> None:
    ws = wb["itens_Consumidos"]
    ws["Q1"] = "CHECK"
    formulas = {
        "D": '=IF(OR(A{r}="",B{r}="",C{r}=""),"",ROUND(B{r}*C{r},2))',
        "F": '=IF(OR(A{r}="",E{r}="",C{r}="",NOT(ISNUMBER($U$2))),"",ROUND(E{r}*C{r}*$U$2,2))',
        "H": '=IF(OR(A{r}="",G{r}="",C{r}="",NOT(ISNUMBER($U$3))),"",ROUND(G{r}*C{r}*$U$3,2))',
        "J": '=IF(OR(A{r}="",I{r}="",C{r}="",NOT(ISNUMBER($U$4))),"",ROUND(I{r}*C{r}*$U$4,2))',
        "L": '=IF(OR(A{r}="",K{r}="",C{r}="",NOT(ISNUMBER($U$5))),"",ROUND(K{r}*C{r}*$U$5,2))',
        "N": '=IF(OR(A{r}="",M{r}="",C{r}="",NOT(ISNUMBER($U$6))),"",ROUND(M{r}*C{r}*$U$6,2))',
        "O": '=IF(A{r}="","",SUM(E{r},G{r},I{r},K{r},M{r}))',
        "P": '=IF(A{r}="","",SUM(F{r},H{r},J{r},L{r},N{r}))',
        "Q": '=IF(A{r}="","",IF(O{r}>B{r},"DIVERGENCIA: CONSUMO MAIOR QUE CONTRATADO","OK"))',
    }
    for row in range(2, 201):
        for col in ("A", "B", "C", "E", "G", "I", "K", "M"):
            ws[f"{col}{row}"] = None
            ws[f"{col}{row}"].fill = INPUT
            ws[f"{col}{row}"].border = GRID
    _style_formula_range(ws, 2, 200, formulas)
    for row in range(2, 201):
        for col in ("B", "E", "G", "I", "K", "M", "O"):
            ws[f"{col}{row}"].number_format = QTY
        for col in ("C", "D", "F", "H", "J", "L", "N", "P"):
            ws[f"{col}{row}"].number_format = MONEY
    for i, cycle in enumerate(("C0", "C1", "C2", "C3", "C4"), 2):
        ws[f"R{i}"] = cycle
        ws[f"S{i}"] = f"=parametros!$A${i}"
        ws[f"T{i}"] = f'=IF(ISNUMBER(parametros!$E${i}),parametros!$E${i},"")'
        ws[f"U{i}"] = f'=IF(ISNUMBER(parametros!$F${i}),parametros!$F${i},"")'
    widths = {"A": 14, "B": 20, "C": 18, "D": 18, "E": 16, "F": 17, "G": 16, "H": 17, "I": 16, "J": 17, "K": 16, "L": 17, "M": 16, "N": 17, "O": 18, "P": 20, "Q": 34}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for cell in ws[1]:
        if cell.value is not None:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
            font = copy(cell.font)
            font.name = "Calibri"
            font.sz = 10
            font.b = True
            cell.font = font
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "D2"
    ws.sheet_view.showGridLines = False


def _reset_itens_pc(wb) -> None:
    ws = wb["itens_PC"]
    while str(ws["A1"].value or "").strip().upper() in {"ITEM", "NUMERO_PC"}:
        ws.delete_cols(1, 1)
    ws.column_dimensions["A"].hidden = False
    for table_name in list(ws.tables):
        del ws.tables[table_name]
    for row in range(101, ws.max_row + 1):
        for col in range(1, 12):
            ws.cell(row, col).value = None
    for row in range(2, 101):
        for col in ("A", "C", "F"):
            ws[f"{col}{row}"] = None
            ws[f"{col}{row}"].fill = INPUT
            ws[f"{col}{row}"].border = GRID
        ws[f"A{row}"].number_format = "dd/mm/yyyy"
        ws[f"C{row}"].number_format = MONEY
    formulas = {
        "B": '=IF(A{r}="","",IF(OR(A{r}<MIN(parametros!$C$2:$C$6),A{r}>MAX(parametros!$D$2:$D$6)),"Fora dos ciclos",LOOKUP(A{r},parametros!$C$2:$C$6,parametros!$B$2:$B$6)))',
        "D": '=IF(B{r}="","",IFERROR(VLOOKUP(B{r},parametros!$A$11:$E$15,5,0),""))',
        "E": '=IF(OR(C{r}="",D{r}=""),"",ROUND(C{r}*D{r},2))',
        "G": '=IF(F{r}="Sim",IF(OR(E{r}="",C{r}=""),"",ROUND(E{r}-C{r},2)),0)',
        "H": '=IF(F{r}="Nao",IF(E{r}="","",E{r}),0)',
        "I": '=IF(F{r}="Nao",IF(OR(E{r}="",C{r}=""),"",ROUND(E{r}-C{r},2)),0)',
        "J": '=IF(COUNTA(A{r}:F{r})=0,"",IF(A{r}="","DATA_PC vazia",IF(OR(B{r}="",B{r}="Fora dos ciclos"),"CICLO_PC nao identificado",IF(OR(C{r}="",C{r}=0),"VALOR_PC vazio ou zero",IF(AND(F{r}<>"Sim",F{r}<>"Nao"),"PC_PAGO_A_CONTRATADA invalido","OK")))))',
    }
    _style_formula_range(ws, 2, 100, formulas)
    # Reconstrói o resumo lateral após a exclusão física de NUMERO_PC.
    for row, cycle in enumerate(("C0", "C1", "C2", "C3", "C4"), 2):
        ws[f"L{row}"] = cycle
        ws[f"M{row}"] = f'=COUNTIF($B$2:$B$100,L{row})'
        ws[f"N{row}"] = f'=SUMIF($B$2:$B$100,L{row},$C$2:$C$100)'
        ws[f"O{row}"] = f'=SUMIF($B$2:$B$100,L{row},$E$2:$E$100)'
        ws[f"P{row}"] = f'=SUMIF($B$2:$B$100,L{row},$G$2:$G$100)'
        ws[f"Q{row}"] = f'=SUMIF($B$2:$B$100,L{row},$H$2:$H$100)'
        ws[f"R{row}"] = f'=SUMIF($B$2:$B$100,L{row},$I$2:$I$100)'
        ws[f"S{row}"] = f'=COUNTIFS($B$2:$B$100,L{row},$J$2:$J$100,"<>OK",$J$2:$J$100,"<>")'
    ws["L7"] = "TOTAL"
    for col in ("M", "N", "O", "P", "Q", "R", "S"):
        ws[f"{col}7"] = f"=SUM({col}2:{col}6)"
    for row in range(2, 101):
        ws[f"D{row}"].number_format = FACTOR
        for col in ("E", "G", "H", "I"):
            ws[f"{col}{row}"].number_format = MONEY
    ws.data_validations.dataValidation.clear()
    dv = DataValidation(type="list", formula1='"Sim,Nao"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add("F2:F100")
    widths = {"A": 16, "B": 16, "C": 18, "D": 18, "E": 20, "F": 24, "G": 28, "H": 26, "I": 20, "J": 30}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


def _reset_aditivos(wb) -> None:
    ws = wb["aditivos"]
    formulas = {
        "C": '=IF(B{r}="","",IFERROR(IF(AND(B{r}>=parametros!$C$2,B{r}<=parametros!$D$2),"C0",IF(AND(B{r}>=parametros!$C$3,B{r}<=parametros!$D$3),"C1",IF(AND(B{r}>=parametros!$C$4,B{r}<=parametros!$D$4),"C2",IF(AND(B{r}>=parametros!$C$5,B{r}<=parametros!$D$5),"C3",IF(AND(B{r}>=parametros!$C$6,B{r}<=parametros!$D$6),"C4","Fora dos ciclos"))))),"Fora dos ciclos"))',
        "F": '=IF(A{r}="","",IFERROR(VLOOKUP(A{r},itens_Remanesc!$A:$C,3,0),""))',
        "G": '=IF(OR(E{r}="",F{r}=""),"",ROUND(E{r}*F{r},2))',
        "I": '=IFERROR(VLOOKUP(C{r},parametros!$B:$F,5,0),"")',
        "J": '=IF(G{r}="","",ROUND(IF(OR(UPPER(D{r})="SUPRESSAO",UPPER(D{r})="DECRESCIMO"),-1,1)*G{r}*IF(AND(UPPER(H{r})="SIM",ISNUMBER(I{r})),I{r},1),2))',
    }
    for row in range(2, 201):
        for col in ("A", "B", "D", "E", "H", "K"):
            ws[f"{col}{row}"] = None
            ws[f"{col}{row}"].fill = INPUT
            ws[f"{col}{row}"].border = GRID
        ws[f"B{row}"].number_format = "mm/yyyy"
        ws[f"E{row}"].number_format = QTY
    _style_formula_range(ws, 2, 200, formulas)
    for row in range(2, 201):
        for col in ("F", "G", "J"):
            ws[f"{col}{row}"].number_format = MONEY
        ws[f"I{row}"].number_format = FACTOR
    ws.data_validations.dataValidation.clear()
    for rng, values in (("D2:D200", "Acrescimo,Supressao"), ("H2:H200", "Sim,Nao"), ("K2:K200", "Sim,Nao")):
        dv = DataValidation(type="list", formula1=f'"{values}"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(rng)
    red_fill = PatternFill("solid", fgColor="FFFFC7CE")
    red_font = Font(color="FF9C0006")
    ws.conditional_formatting.add("A2:K200", FormulaRule(formula=['OR(LEFT(UPPER($D2),6)="SUPRES",LEFT(UPPER($D2),4)="DECR")'], fill=red_fill, font=red_font))
    widths = {"A": 12, "B": 18, "C": 16, "D": 30, "E": 25, "F": 22, "G": 24, "H": 24, "I": 22, "J": 25, "K": 25}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.row_dimensions[1].height = 40
    for cell in ws[1]:
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


def _reset_historico_vu(wb) -> None:
    ws = wb["historico_VU"]
    formulas = {
        "A": '=IF(itens_Remanesc!A{r}="","",itens_Remanesc!A{r})',
        "B": '=IF(itens_Remanesc!A{r}="","",itens_Remanesc!B{r})',
        "C": '=IF(itens_Remanesc!A{r}="","",itens_Remanesc!C{r})',
        "D": '=IF(OR(A{r}="",C{r}="",NOT(ISNUMBER($L$3))),"",ROUND(C{r}*$L$3,2))',
        "E": '=IF(OR(A{r}="",C{r}="",NOT(ISNUMBER($L$4))),"",ROUND(C{r}*$L$4,2))',
        "F": '=IF(OR(A{r}="",C{r}="",NOT(ISNUMBER($L$5))),"",ROUND(C{r}*$L$5,2))',
        "G": '=IF(OR(A{r}="",C{r}="",NOT(ISNUMBER($L$6))),"",ROUND(C{r}*$L$6,2))',
        "H": '=IF(OR(A{r}="",C{r}="",C{r}=0,NOT(ISNUMBER($L$6))),"",ROUND($L$6-1,4))',
    }
    _style_formula_range(ws, 2, 200, formulas)
    for row in range(2, 201):
        ws[f"B{row}"].number_format = QTY
        for col in ("C", "D", "E", "F", "G"):
            ws[f"{col}{row}"].number_format = MONEY
        ws[f"H{row}"].number_format = PERCENT
    for i, cycle in enumerate(("C0", "C1", "C2", "C3", "C4"), 2):
        ws[f"J{i}"] = cycle
        ws[f"K{i}"] = f'=IF(ISNUMBER(parametros!$E${i}),parametros!$E${i},"")'
        ws[f"L{i}"] = f'=IF(ISNUMBER(parametros!$F${i}),parametros!$F${i},"")'
        ws[f"K{i}"].number_format = PERCENT
        ws[f"L{i}"].number_format = FACTOR
        for col in ("J", "K", "L"):
            ws[f"{col}{i}"].fill = AUTO
            ws[f"{col}{i}"].border = GRID
    ws.freeze_panes = "B2"
    ws.sheet_view.showGridLines = False


def _reset_itens_rc(wb) -> None:
    ws = wb["itens_RC"]
    ws.conditional_formatting._cf_rules.clear()
    for row in range(3, 203):
        src = row - 1
        prev_src = src - 1
        tail_start = src + 1
        is_total = (
            "FALSE"
            if row == 3
            else f'AND(itens_Remanesc!A{src}="",itens_Remanesc!A{prev_src}<>"",COUNTIF(itens_Remanesc!A{tail_start}:$A$200,"<>")=0)'
        )
        formulas = {
            "A": f'=IF({is_total},"TOTAL",IF(itens_Remanesc!A{src}="","",itens_Remanesc!A{src}))',
            "B": f'=IF(OR(itens_Remanesc!A{src}="",NOT(ISNUMBER(parametros!$F$2))),"",ROUND(itens_Remanesc!C{src}*parametros!$F$2,2))',
            "C": f'=IF(itens_Remanesc!A{src}="","",itens_Remanesc!B{src})',
            "D": f'=IF(A{row}="TOTAL",ROUND(SUM($D$3:D{row-1}),2),IF(A{row}="","",ROUND(B{row}*C{row},2)))',
            "E": f'=IF(OR(itens_Remanesc!A{src}="",NOT(ISNUMBER(parametros!$F$3))),"",ROUND(itens_Remanesc!C{src}*parametros!$F$3,2))',
            "F": f'=IF(itens_Remanesc!A{src}="","",itens_Remanesc!E{src})',
            "G": f'=IF(A{row}="TOTAL",ROUND(SUM($G$3:G{row-1}),2),IF(OR(A{row}="",E{row}="",F{row}=""),"",ROUND(E{row}*F{row},2)))',
            "H": f'=IF(OR(itens_Remanesc!A{src}="",NOT(ISNUMBER(parametros!$F$4))),"",ROUND(itens_Remanesc!C{src}*parametros!$F$4,2))',
            "I": f'=IF(itens_Remanesc!A{src}="","",itens_Remanesc!G{src})',
            "J": f'=IF(A{row}="TOTAL",ROUND(SUM($J$3:J{row-1}),2),IF(OR(A{row}="",H{row}="",I{row}=""),"",ROUND(H{row}*I{row},2)))',
            "K": f'=IF(OR(itens_Remanesc!A{src}="",NOT(ISNUMBER(parametros!$F$5))),"",ROUND(itens_Remanesc!C{src}*parametros!$F$5,2))',
            "L": f'=IF(itens_Remanesc!A{src}="","",itens_Remanesc!I{src})',
            "M": f'=IF(A{row}="TOTAL",ROUND(SUM($M$3:M{row-1}),2),IF(OR(A{row}="",K{row}="",L{row}=""),"",ROUND(K{row}*L{row},2)))',
            "N": f'=IF(OR(itens_Remanesc!A{src}="",NOT(ISNUMBER(parametros!$F$6))),"",ROUND(itens_Remanesc!C{src}*parametros!$F$6,2))',
            "O": f'=IF(itens_Remanesc!A{src}="","",itens_Remanesc!K{src})',
            "P": f'=IF(A{row}="TOTAL",ROUND(SUM($P$3:P{row-1}),2),IF(OR(A{row}="",N{row}="",O{row}=""),"",ROUND(N{row}*O{row},2)))',
        }
        for col, formula in formulas.items():
            ws[f"{col}{row}"] = formula
            ws[f"{col}{row}"].border = GRID
    for row in range(3, 203):
        for col in ("B", "D", "E", "G", "H", "J", "K", "M", "N", "P"):
            ws[f"{col}{row}"].number_format = MONEY
        for col in ("C", "F", "I", "L", "O"):
            ws[f"{col}{row}"].number_format = QTY
    for col in range(1, 17):
        ws.cell(203, col).value = None
    ws.conditional_formatting.add(
        "A3:P202",
        FormulaRule(
            formula=['$A3="TOTAL"'],
            fill=PatternFill("solid", fgColor="FFD9E1F2"),
            font=Font(color=NAVY, bold=True),
        ),
    )
    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False


def _reset_resultados(wb) -> None:
    _assert_resultados_integra(wb, "antes do ajuste de referências")
    if "historico" in wb.sheetnames:
        wb.remove(wb["historico"])
    ws = wb["RESULTADOS"]
    wb._sheets.append(wb._sheets.pop(wb._sheets.index(ws)))
    for row, cycle in enumerate(("C0", "C1", "C2", "C3", "C4"), 10):
        ws[f"C{row}"] = (
            f'=IF(COUNTIFS(itens_PC!$B$2:$B$100,"{cycle}",itens_PC!$C$2:$C$100,">0",'
            f'itens_PC!$F$2:$F$100,"Sim")=0,"",ROUND(SUMIFS(itens_PC!$G$2:$G$100,'
            f'itens_PC!$B$2:$B$100,"{cycle}",itens_PC!$F$2:$F$100,"Sim"),2))'
        )
    ws["D44"] = '=SUMPRODUCT(--(((itens_PC!$A$2:$A$100<>"")+(itens_PC!$C$2:$C$100<>"")+(itens_PC!$F$2:$F$100<>""))>0),--(itens_PC!$J$2:$J$100<>"OK"))'
    ws["D45"] = '=ROUND(SUM(itens_PC!$I$2:$I$100),2)'
    ws["B20"] = '=IF(COUNTIF(itens_Remanesc!$A$2:$A$200,"<>")=0,"",ROUND(SUMIF(itens_Remanesc!$A$2:$A$200,"<>",itens_Remanesc!$D$2:$D$200),2))'
    ws["D32"] = '=IF(COUNTIF(itens_Remanesc!$A$2:$A$200,"<>")=0,"",ROUND(IF($F$4="C0",SUMIFS(itens_RC!$D$3:$D$202,itens_RC!$A$3:$A$202,"<>TOTAL",itens_RC!$A$3:$A$202,"<>"),IF($F$4="C1",SUMIFS(itens_RC!$G$3:$G$202,itens_RC!$A$3:$A$202,"<>TOTAL",itens_RC!$A$3:$A$202,"<>"),IF($F$4="C2",SUMIFS(itens_RC!$J$3:$J$202,itens_RC!$A$3:$A$202,"<>TOTAL",itens_RC!$A$3:$A$202,"<>"),IF($F$4="C3",SUMIFS(itens_RC!$M$3:$M$202,itens_RC!$A$3:$A$202,"<>TOTAL",itens_RC!$A$3:$A$202,"<>"),SUMIFS(itens_RC!$P$3:$P$202,itens_RC!$A$3:$A$202,"<>TOTAL",itens_RC!$A$3:$A$202,"<>"))))),2))'
    ws.sheet_view.showGridLines = False
    _assert_resultados_integra(wb, "depois do ajuste de referências")


def _normalize(wb) -> None:
    for obsolete in ("itens_Execucao_Saldo", "REGRA_NEGOCIO_CLAUS"):
        if obsolete in wb.sheetnames:
            wb.remove(wb[obsolete])
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                cell.comment = None
                if cell.value is not None and ws.title != "CONTROLE":
                    font = copy(cell.font)
                    font.name = "Calibri"
                    font.sz = 10
                    cell.font = font
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.active = 0
    view = wb.views[0]
    view.visibility = "visible"
    view.minimized = False
    view.xWindow = 0
    view.yWindow = 0
    view.windowWidth = 22000
    view.windowHeight = 12000
    view.firstSheet = 0
    view.activeTab = 0


def build(source: Path, destination: Path) -> None:
    wb = load_workbook(source, data_only=False)
    required = {
        "CONTROLE", "parametros", "financeiro", "itens_Remanesc",
        "itens_Consumidos", "itens_PC", "aditivos", "RESULTADOS",
        "itens_RC", "historico_VU",
    }
    missing = sorted(required.difference(wb.sheetnames))
    if missing:
        raise ValueError(f"Modelo de origem sem abas obrigatorias: {', '.join(missing)}")
    _assert_resultados_integra(wb, "logo após o carregamento do template")
    _excluir_campo_c_parametros(wb)
    _reset_controle(wb)
    _reset_parametros(wb)
    _reset_financeiro(wb)
    _reset_itens_remanesc(wb)
    _reset_itens_consumidos(wb)
    _reset_itens_pc(wb)
    _reset_aditivos(wb)
    _reset_resultados(wb)
    _reset_itens_rc(wb)
    _reset_historico_vu(wb)
    _normalize(wb)
    _assert_resultados_integra(wb, "imediatamente antes do salvamento")
    destination.parent.mkdir(parents=True, exist_ok=True)
    wb.save(destination)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument("destination", type=Path)
    args = parser.parse_args()
    build(args.source, args.destination)


if __name__ == "__main__":
    main()
