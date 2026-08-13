# -*- coding: utf-8 -*-
"""Correcao temporal definitiva dos itens incluidos por aditivos (DATA_EFEITO).

CAUSA RAIZ (Regra 2 sinalizada na etapa anterior)
-------------------------------------------------
`posicao_contratual!Y` (CICLO_NASCIMENTO) deriva da QUANTIDADE: e o primeiro
ciclo cuja QTD_CONTRATADA_Cn e positiva. Como `DELTA_Cn` agrega TODOS os
aditivos do ciclo (SUMIFS por rotulo de ciclo, sem recorte de data), um item
incluido no PRIMEIRO DIA do ciclo e outro incluido no MEIO do mesmo ciclo
recebem o mesmo Y e a mesma abertura. A fotografia de abertura passa entao a
conter quantidade que ainda nao existia naquela data.

REGRA CANONICA APLICADA
-----------------------
    QTD CONTRATUAL NA ABERTURA DE Cn
        = quantidade-base + deltas com DATA_EFEITO <= data de abertura de Cn

Implementada por DECOMPOSICAO do delta oficial de cada ciclo, sem tocar em
nenhuma formula da cadeia oficial do VTA:

    DELTA_POSTERIOR_ABERTURA_Cn = deltas do ciclo Cn com data > inicio de Cn
    QTD_CONTRATUAL_ABERTURA_Cn  = QTD_CONTRATADA_Cn - DELTA_POSTERIOR_ABERTURA_Cn
    CICLO_NASCIMENTO_DATA       = menor n com QTD_CONTRATUAL_ABERTURA_Cn > 0

Como os deltas de ciclos anteriores sempre tem data <= abertura de Cn, a
subtracao acima isola exatamente a parcela do proprio ciclo cujo efeito e
posterior a abertura. Um aditivo datado no ultimo dia de C0 continua valendo
para a abertura de C1; um aditivo datado no primeiro dia de C1 entra na
abertura de C1 uma unica vez; um aditivo do meio de C1 sai da abertura de C1 e
so reaparece na abertura de C2 (se houver saldo).

TRAVAS (Secao 5 do enunciado)
-----------------------------
* `MEMORIA_RESULTADOS!B26` (VTA_FINAL) e `T25` permanecem com a formula e o
  valor originais. Nenhuma coluna nova alimenta T21/T22/T23.
* A FORMA 2 auditavel (`W48`) e REDECOMPOSTA, nao reduzida: o remanescente da
  abertura passa a ser temporalmente correto (`W53`) e os aditivos posteriores
  a abertura entram por COMPONENTE PROPRIO (`W54`), considerados uma unica vez.
  Por construcao `W53 + W54 == SUM(AB)`, de modo que o total de W48 nao muda —
  o que muda e a leitura temporal de cada parcela. `W55` e a trava
  anti-dupla-contagem (deve ser 0,00).
* O tool aborta se a formula ou o valor de B26/T25 mudar.

Metodo de mutacao: openpyxl para autoria em massa -> gate Excel COM
(recalculo total, zero erros de formula, salvar, reabrir sem reparo) ->
formatacao condicional como ULTIMA escrita openpyxl (mantem CF padrao OOXML,
preservada pelos Coletas gerados pelo app).
"""
from __future__ import annotations

import argparse
import shutil
import tempfile
from pathlib import Path

import pythoncom
import win32com.client

XL_CALC_MANUAL = -4135
XL_CALC_AUTOMATIC = -4105
XL_CELLTYPE_FORMULAS = -4123
XL_CELLTYPE_CONSTANTS = 2
XL_ERRORS = 16

ABA_PC = "posicao_contratual"
ABA_MEM = "MEMORIA_RESULTADOS"
ABA_RC = "itens_RC"
ABA_REM = "itens_Remanesc"

# posicao_contratual: linhas de cadastro contratual (mesma faixa de X/Y/Z).
PC_INI, PC_FIM = 2, 200
# itens_RC: linha 3 espelha posicao_contratual linha 2 (deslocamento -1).
RC_INI, RC_FIM = 3, 202
# MEMORIA_RESULTADOS: colunas auxiliares por item.
MEM_INI, MEM_FIM = 2, 201
# itens_Remanesc: linhas de cadastro (1 = rotulo; 201 = total dinamico, com
# formula propria de SUMIF e portanto fora da guarda por item).
REM_INI, REM_FIM = 2, 200

# --- posicao_contratual: colunas novas (nenhuma consumida pela cadeia oficial)
COL_DATA_EFEITO = "AA"
COLS_DELTA_POSTERIOR = ("AB", "AC", "AD", "AE", "AF")   # C0..C4
COLS_QTD_ABERTURA = ("AG", "AH", "AI", "AJ", "AK")      # C0..C4
COL_NASCIMENTO_DATA = "AL"

# Coluna de QTD_CONTRATADA_Cn ja existente (cumulativa) por ciclo.
COLS_QTD_CONTRATADA = ("E", "I", "M", "Q", "U")
# parametros!C{linha} = janela mensal do ciclo Cn (guarda de existencia);
# parametros!I{linha} = DATA_ABERTURA_FISICA_EXATA (fronteira fisica).
LINHAS_PARAM_INICIO = (2, 3, 4, 5, 6)

CABECALHOS_PC = {
    COL_DATA_EFEITO: "DATA_EFEITO_INICIAL",
    "AB": "DELTA_POSTERIOR_ABERTURA_C0",
    "AC": "DELTA_POSTERIOR_ABERTURA_C1",
    "AD": "DELTA_POSTERIOR_ABERTURA_C2",
    "AE": "DELTA_POSTERIOR_ABERTURA_C3",
    "AF": "DELTA_POSTERIOR_ABERTURA_C4",
    "AG": "QTD_CONTRATUAL_ABERTURA_C0",
    "AH": "QTD_CONTRATUAL_ABERTURA_C1",
    "AI": "QTD_CONTRATUAL_ABERTURA_C2",
    "AJ": "QTD_CONTRATUAL_ABERTURA_C3",
    "AK": "QTD_CONTRATUAL_ABERTURA_C4",
    COL_NASCIMENTO_DATA: "CICLO_NASCIMENTO_DATA",
}

# --- MEMORIA_RESULTADOS: decomposicao temporal da FORMA 2 (por item)
MEM_COL_ABERTURA_TEMP = "AC"
MEM_COL_POSTERIOR = "AD"

# --- itens_RC: bloco de aplicabilidade temporal
RC_COLS = ("Z", "AA", "AB", "AC")
RC_CABECALHOS = (
    "DATA DE EFEITO DO ITEM (AUTO)",
    "CICLO DE NASCIMENTO POR DATA (AUTO)",
    "APLICAVEL NA ABERTURA ADOTADA? (AUTO)",
    "QTD CONTRATUAL NA ABERTURA ADOTADA (AUTO)",
)
RC_TITULO = (
    "APLICABILIDADE TEMPORAL NA ABERTURA "
    "(AUTO - origem: DATA DE EFEITO DO ADITIVO)"
)

# Bloco POSICAO ATUAL (AUTO) ja existente: coluna de itens_RC -> referencia
# INDIRECT para CICLO_EM_EXECUCAO. A forma `&(ROW()+10)` e a homologada (a linha
# 3 de itens_RC espelha a linha 13 da aba) e permanece inalterada aqui: so muda
# o tratamento de vazio x zero ao redor dela.
RC_POSICAO_ATUAL = (
    ("Q", 'INDIRECT("CICLO_EM_EXECUCAO!$D$5")'),
    ("S", 'INDIRECT("CICLO_EM_EXECUCAO!B"&(ROW()+10))'),
    ("T", 'INDIRECT("CICLO_EM_EXECUCAO!D"&(ROW()+10))'),
    ("U", 'INDIRECT("CICLO_EM_EXECUCAO!C"&(ROW()+10))'),
    ("V", 'INDIRECT("CICLO_EM_EXECUCAO!E"&(ROW()+10))'),
    ("W", 'INDIRECT("CICLO_EM_EXECUCAO!F"&(ROW()+10))'),
    ("X", 'INDIRECT("CICLO_EM_EXECUCAO!G"&(ROW()+10))'),
    ("Y", 'INDIRECT("CICLO_EM_EXECUCAO!K"&(ROW()+10))'),
)

# --- itens_Remanesc: VALOR_REM_INICIO_Cn e o indice do ciclo de abertura.
COLS_VALOR_ABERTURA = (("F", 1), ("H", 2), ("J", 3), ("L", 4))
# Colunas manuais QTD_REM_BASE_Cn (formatacao condicional dos 4 estados).
COLS_QTD_ABERTURA_MANUAL = (("E", 1), ("G", 2), ("I", 3), ("K", 4))
# Espelho local do CICLO_NASCIMENTO_DATA. A formatacao condicional PRECISA ser
# da propria aba: o Excel migra para a extensao x14 toda regra que referencie
# outra planilha, e a extensao x14 e invisivel para o openpyxl — a linhagem que
# gera cada Coleta. Com o espelho, a CF continua OOXML padrao.
COL_ESPELHO_NASCIMENTO = "BI"


def _criterios_aditivo(linha_pc: int) -> str:
    """Criterios *IFS: aditivos do item com delta numerico != 0 e data valida."""
    return (
        f'aditivos!$A$2:$A$200,$A{linha_pc},'
        f'aditivos!$L$2:$L$200,"<>",'
        f'aditivos!$L$2:$L$200,"<>0",'
        f'aditivos!$B$2:$B$200,">0"'
    )


def _formula_data_efeito(r: int) -> str:
    """Menor data de aditivo com efeito quantitativo para o item.

    Coluna INFORMATIVA: a classificacao canonica (AL) nao depende dela.
    `MINIFS` e gravada com o prefixo OOXML `_xlfn.` — a codificacao correta para
    uma funcao posterior a linha de base do formato; o Excel a exibe como
    MINIFS e nao reporta reparo. Em versoes anteriores a 2019 ela resolveria
    como #NOME?, e o IFERROR degrada para vazio: nunca para um valor errado.
    """
    criterios = _criterios_aditivo(r)
    return (
        f'=IF($A{r}="","",'
        f'IF(COUNTIFS({criterios})=0,"",'
        f'IFERROR(_xlfn.MINIFS(aditivos!$B$2:$B$200,{criterios}),"")))'
    )


def _formula_delta_posterior(r: int, ciclo: int) -> str:
    """Deltas do ciclo Cn com DATA_EFEITO posterior a abertura do proprio ciclo.

    A fronteira e por DIA, nao por instante: `>= abertura + 1` mantem no lado da
    abertura qualquer aditivo datado NO dia da abertura, mesmo que a celula
    carregue componente de hora (data digitada com horario, importacao, COM).

    ETAPA 48: a fronteira FISICA le parametros!I (DATA_ABERTURA_FISICA_EXATA);
    a guarda de existencia do ciclo permanece na janela mensal (parametros!C).
    ETAPA 48.3: a guarda tambem reconhece I vazio (ciclo futuro sem fotografia
    fisica) e devolve 0 sem executar INT sobre celula vazia, que viraria a
    fronteira ficticia 01/01/1900.
    """
    linha_param = LINHAS_PARAM_INICIO[ciclo]
    return (
        f'=IF($A{r}="","",'
        f'IF(OR(parametros!$C${linha_param}="",parametros!$I${linha_param}=""),0,'
        f'ROUND(SUMIFS(aditivos!$L$2:$L$200,'
        f'aditivos!$A$2:$A$200,$A{r},'
        f'aditivos!$C$2:$C$200,"C{ciclo}",'
        f'aditivos!$B$2:$B$200,">="&(INT(parametros!$I${linha_param})+1)),2)))'
    )


def _formula_qtd_abertura(r: int, ciclo: int) -> str:
    """QTD contratual existente EXATAMENTE na data de abertura do ciclo Cn."""
    contratada = COLS_QTD_CONTRATADA[ciclo]
    posterior = COLS_DELTA_POSTERIOR[ciclo]
    return (
        f'=IF($A{r}="","",'
        f'IF(OR(NOT(ISNUMBER(${contratada}{r})),NOT(ISNUMBER(${posterior}{r}))),"",'
        f'ROUND(${contratada}{r}-${posterior}{r},2)))'
    )


def _formula_nascimento_data(r: int) -> str:
    """Primeiro ciclo cuja ABERTURA ja contem quantidade contratual do item."""
    q = COLS_QTD_ABERTURA
    return (
        f'=IF($A{r}="","",'
        f'IF(N(${q[0]}{r})>0,0,'
        f'IF(N(${q[1]}{r})>0,1,'
        f'IF(N(${q[2]}{r})>0,2,'
        f'IF(N(${q[3]}{r})>0,3,'
        f'IF(N(${q[4]}{r})>0,4,""))))))'
    )


def _choose_por_ciclo(sel: str, aba: str, colunas: tuple[str, ...], r: int) -> str:
    refs = ",".join(f"{aba}!${c}{r}" for c in colunas)
    return f"CHOOSE({sel}+1,{refs})"


def _formula_mem_posterior(r: int) -> str:
    """Componente proprio: alteracoes contratuais posteriores a abertura adotada."""
    vu = _choose_por_ciclo("$W$46", "historico_VU", ("C", "D", "E", "F", "G"), r)
    post = _choose_por_ciclo("$W$46", ABA_PC, COLS_DELTA_POSTERIOR, r)
    return (
        f'=IF(OR({ABA_PC}!$A{r}="",$W$46=""),0,'
        f'IF($AB{r}="","",'
        f'IF(NOT(ISNUMBER({post})),0,'
        f'ROUND(ROUND({vu},2)*{post},2))))'
    )


def _formula_mem_abertura_temporal(r: int) -> str:
    """Remanescente da abertura adotada, ja sem os aditivos de efeito posterior."""
    return (
        f'=IF(OR({ABA_PC}!$A{r}="",$W$46=""),0,'
        f'IF(OR($AB{r}="",$AD{r}=""),"",'
        f'ROUND($AB{r}-$AD{r},2)))'
    )


LINHAS_W = {
    48: (
        None,
        '=IF($W$46="","",ROUND($T$21'
        '+SUMPRODUCT((ROW(itens_PC!$P$2:$P$6)-2>=1)'
        '*(ROW(itens_PC!$P$2:$P$6)-2<$W$46)*itens_PC!$P$2:$P$6)'
        '+$W$53+$W$54,2))',
    ),
    53: (
        "FORMA 2 - remanescente na abertura (data de efeito)",
        "=SUM($AC$2:$AC$201)",
    ),
    54: (
        "FORMA 2 - alteracoes contratuais posteriores a abertura",
        "=SUM($AD$2:$AD$201)",
    ),
    55: (
        "Trava anti-dupla-contagem (deve ser 0,00)",
        "=ROUND(SUM($AB$2:$AB$201)-($W$53+$W$54),2)",
    ),
    56: (
        "Status da decomposicao temporal",
        '=IF(ABS($W$55)<=$D$4,"SEM DUPLA CONTAGEM",'
        '"REVISE - DUPLA CONTAGEM NA FORMA 2")',
    ),
    57: (
        "Itens incluidos apos a abertura adotada (qtd)",
        '=IF($W$46="","",SUMPRODUCT((posicao_contratual!$A$2:$A$200<>"")'
        '*ISNUMBER(posicao_contratual!$AL$2:$AL$200)'
        '*(posicao_contratual!$AL$2:$AL$200>$W$46)))',
    ),
}


def _escrever_formulas(caminho: Path) -> dict[str, int]:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = load_workbook(caminho)
    contagem: dict[str, int] = {}

    # ------------------------------------------------ posicao_contratual
    pc = wb[ABA_PC]
    for col, rotulo in CABECALHOS_PC.items():
        pc[f"{col}1"] = rotulo
    novas = 0
    for r in range(PC_INI, PC_FIM + 1):
        pc[f"{COL_DATA_EFEITO}{r}"] = _formula_data_efeito(r)
        pc[f"{COL_DATA_EFEITO}{r}"].number_format = "dd/mm/yyyy;@"
        novas += 1
        for ciclo in range(5):
            pc[f"{COLS_DELTA_POSTERIOR[ciclo]}{r}"] = _formula_delta_posterior(
                r, ciclo
            )
            pc[f"{COLS_QTD_ABERTURA[ciclo]}{r}"] = _formula_qtd_abertura(r, ciclo)
            novas += 2
        pc[f"{COL_NASCIMENTO_DATA}{r}"] = _formula_nascimento_data(r)
        novas += 1
    ocultas = (
        (COL_DATA_EFEITO,)
        + COLS_DELTA_POSTERIOR
        + COLS_QTD_ABERTURA
        + (COL_NASCIMENTO_DATA,)
    )
    for col in ocultas:
        pc.column_dimensions[col].hidden = True
    contagem["posicao_contratual_novas"] = novas

    # ------------------------------------------------ MEMORIA_RESULTADOS
    mem = wb[ABA_MEM]
    novas = 0
    for r in range(MEM_INI, MEM_FIM + 1):
        mem[f"{MEM_COL_POSTERIOR}{r}"] = _formula_mem_posterior(r)
        mem[f"{MEM_COL_ABERTURA_TEMP}{r}"] = _formula_mem_abertura_temporal(r)
        novas += 2
    for linha, (rotulo, formula) in LINHAS_W.items():
        if rotulo is not None:
            mem[f"V{linha}"] = rotulo
            novas += 1
        mem[f"W{linha}"] = formula
    contagem["memoria_novas"] = novas

    # ------------------------------------------------ itens_RC
    rc = wb[ABA_RC]
    rc[f"{RC_COLS[0]}1"] = RC_TITULO
    rc.merge_cells(f"{RC_COLS[0]}1:{RC_COLS[-1]}1")
    titulo = rc[f"{RC_COLS[0]}1"]
    titulo.font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
    titulo.fill = PatternFill("solid", fgColor="7F7F7F")
    titulo.alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    for col, rotulo in zip(RC_COLS, RC_CABECALHOS):
        cel = rc[f"{col}2"]
        cel.value = rotulo
        cel.font = Font(name="Aptos", size=9, bold=True)
        cel.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        rc.column_dimensions[col].width = 20.0
    # Regra A: vazio e PENDENCIA, zero e saldo CONFIRMADO. `INDIRECT` sobre
    # celula vazia devolve 0, o que transformava "nao informado" em "zero
    # confirmado" no bloco POSICAO ATUAL. A forma abaixo distingue os dois
    # (`INDIRECT(x)=""` e verdadeiro so para vazio; `0=""` e falso) mantendo as
    # mesmas DUAS chamadas volateis por celula.
    for coluna, alvo in RC_POSICAO_ATUAL:
        queda = '"CICLO_EM_EXECUCAO AUSENTE"' if coluna == "Y" else '""'
        for r in range(RC_INI, RC_FIM + 1):
            rc[f"{coluna}{r}"] = (
                f'=IF($A{r}="","",IFERROR(IF({alvo}="","",{alvo}),{queda}))'
            )

    novas = 0
    for r in range(RC_INI, RC_FIM + 1):
        origem = r - 1
        rc[f"Z{r}"] = (
            f'=IF($A{r}="","",'
            f'IF({ABA_PC}!$AA{origem}="","",{ABA_PC}!$AA{origem}))'
        )
        rc[f"Z{r}"].number_format = "dd/mm/yyyy;@"
        rc[f"AA{r}"] = (
            f'=IF($A{r}="","",'
            f'IF({ABA_PC}!$AL{origem}="","","C"&{ABA_PC}!$AL{origem}))'
        )
        rc[f"AB{r}"] = (
            f'=IF($A{r}="","",'
            f'IF(OR({ABA_MEM}!$W$46="",{ABA_PC}!$AL{origem}=""),"",'
            f'IF({ABA_PC}!$AL{origem}<={ABA_MEM}!$W$46,"SIM","NAO APLICAVEL")))'
        )
        rc[f"AC{r}"] = (
            f'=IF($A{r}="","",IF({ABA_MEM}!$W$46="","",'
            + _choose_por_ciclo(
                f"{ABA_MEM}!$W$46", ABA_PC, COLS_QTD_ABERTURA, origem
            )
            + "))"
        )
        novas += 4
    contagem["itens_rc_novas"] = novas

    # ------------------------------------------------ itens_Remanesc
    # Regra C: item incluido depois da abertura nao tem VALOR de abertura.
    rem = wb[ABA_REM]
    rem[f"{COL_ESPELHO_NASCIMENTO}1"] = "CICLO_NASCIMENTO_DATA (AUTO)"
    espelhadas = 0
    for r in range(REM_INI, REM_FIM + 1):
        rem[f"{COL_ESPELHO_NASCIMENTO}{r}"] = (
            f'=IF($A{r}="","",'
            f'IF({ABA_PC}!${COL_NASCIMENTO_DATA}{r}="","",'
            f'{ABA_PC}!${COL_NASCIMENTO_DATA}{r}))'
        )
        espelhadas += 1
    rem.column_dimensions[COL_ESPELHO_NASCIMENTO].hidden = True
    contagem["itens_remanesc_espelho"] = espelhadas
    ajustadas = 0
    for col, idx in COLS_VALOR_ABERTURA:
        for r in range(REM_INI, REM_FIM + 1):
            atual = rem[f"{col}{r}"].value
            if not isinstance(atual, str) or not atual.startswith("="):
                continue
            guarda = f'AND(ISNUMBER({ABA_PC}!$AL{r}),{ABA_PC}!$AL{r}>{idx})'
            if guarda in atual:
                continue
            marca = f'IF(OR(A{r}="",'
            if marca not in atual:
                raise RuntimeError(
                    f"{ABA_REM}!{col}{r}: formula inesperada {atual!r}"
                )
            rem[f"{col}{r}"] = atual.replace(marca, marca + guarda + ",", 1)
            ajustadas += 1
    contagem["itens_remanesc_ajustadas"] = ajustadas

    wb.save(caminho)
    return contagem


def _reescrever_cf_remanesc(caminho: Path) -> None:
    """Regra 4 com base temporal: os 4 estados passam a usar CICLO_NASCIMENTO_DATA.

    Ultima escrita do processo (openpyxl), para manter a formatacao condicional
    como CF padrao OOXML — e nao a extensao x14 que o Excel gravaria — de modo
    que os Coletas gerados pelo app (linhagem openpyxl) a preservem.
    """
    from openpyxl import load_workbook
    from openpyxl.formatting.formatting import ConditionalFormattingList
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import PatternFill

    wb = load_workbook(caminho)
    ws = wb[ABA_REM]

    alvos = {f"{col}{REM_INI}:{col}201" for col, _ in COLS_QTD_ABERTURA_MANUAL}
    preservadas = ConditionalFormattingList()
    for cf in ws.conditional_formatting:
        if str(cf.sqref) in alvos:
            continue
        for regra in cf.rules:
            preservadas.add(str(cf.sqref), regra)
    ws.conditional_formatting = preservadas

    f_na = PatternFill("solid", fgColor="D9D9D9")
    f_pos = PatternFill("solid", fgColor="C6EFCE")
    f_zero = PatternFill("solid", fgColor="DDEBF7")
    f_pend = PatternFill("solid", fgColor="FFF2CC")
    for col, idx in COLS_QTD_ABERTURA_MANUAL:
        faixa = f"{col}{REM_INI}:{col}201"
        a = f"$A{REM_INI}"
        nasc = f"${COL_ESPELHO_NASCIMENTO}{REM_INI}"
        cel = f"{col}{REM_INI}"
        # NAO APLICAVEL primeiro (stopIfTrue): abertura anterior ao nascimento
        # POR DATA nunca e pintada como pendencia.
        ws.conditional_formatting.add(
            faixa,
            FormulaRule(
                formula=[f'AND({a}<>"",ISNUMBER({nasc}),{nasc}>{idx})'],
                fill=f_na,
                stopIfTrue=True,
            ),
        )
        ws.conditional_formatting.add(
            faixa,
            FormulaRule(
                formula=[f'AND({a}<>"",ISNUMBER({cel}),{cel}>0)'], fill=f_pos
            ),
        )
        ws.conditional_formatting.add(
            faixa,
            FormulaRule(formula=[f'AND({a}<>"",{cel}=0)'], fill=f_zero),
        )
        ws.conditional_formatting.add(
            faixa,
            FormulaRule(
                formula=[
                    f'AND({a}<>"",{cel}="",'
                    f'OR(NOT(ISNUMBER({nasc})),{nasc}<={idx}))'
                ],
                fill=f_pend,
            ),
        )
    wb.save(caminho)


def _conferir_cf_preservada(caminho: Path) -> None:
    """Aborta se o Excel tiver movido a CF dos 4 estados para a extensao x14.

    A CF e escrita por openpyxl (OOXML padrao) e o Excel salva depois; esta
    conferencia garante que ela continua legivel por openpyxl — a mesma
    linhagem usada pelo app para gerar cada Coleta.
    """
    from openpyxl import load_workbook

    wb = load_workbook(caminho)
    ws = wb[ABA_REM]
    por_faixa = {
        str(cf.sqref): list(cf.rules) for cf in ws.conditional_formatting
    }
    for col, idx in COLS_QTD_ABERTURA_MANUAL:
        faixa = f"{col}{REM_INI}:{col}201"
        regras = por_faixa.get(faixa)
        if not regras or len(regras) != 4:
            raise RuntimeError(f"CF perdida em {ABA_REM}!{faixa}: {regras!r}")
        juntas = " ".join(str(r.formula) for r in regras)
        if f"${COL_ESPELHO_NASCIMENTO}{REM_INI}>{idx}" not in juntas:
            raise RuntimeError(f"CF de {faixa} nao usa CICLO_NASCIMENTO_DATA.")


def _nomes_abas(wb) -> list[str]:
    return [ws.Name for ws in wb.Worksheets]


def _verificar_sem_erros(wb) -> None:
    import pywintypes

    problemas = []
    for ws in wb.Worksheets:
        for tipo in (XL_CELLTYPE_FORMULAS, XL_CELLTYPE_CONSTANTS):
            try:
                celulas = ws.UsedRange.SpecialCells(tipo, XL_ERRORS)
                problemas.append(f"{ws.Name}!{celulas.Address}")
            except pywintypes.com_error:
                continue
    if problemas:
        raise RuntimeError(f"Erros de formula apos recalculo: {problemas}")


def _snapshot_travas(wb) -> dict[str, object]:
    mem = wb.Worksheets(ABA_MEM)
    dados: dict[str, object] = {}
    for endereco in ("B26", "T25"):
        dados[f"{endereco}.formula"] = str(mem.Range(endereco).Formula)
        valor = mem.Range(endereco).Value
        dados[f"{endereco}.valor"] = (
            round(float(valor), 2) if isinstance(valor, (int, float)) else valor
        )
    return dados


def aplicar(origem: Path, destino: Path) -> None:
    origem = Path(origem).resolve()
    destino = Path(destino).resolve()
    if not origem.is_file():
        raise FileNotFoundError(origem)
    if origem == destino:
        raise ValueError("Origem e destino devem ser diferentes.")

    tmp_dir = Path(tempfile.mkdtemp(prefix="cl8us_temporal_"))
    tmp_xlsx = tmp_dir / origem.name
    shutil.copyfile(origem, tmp_xlsx)

    contagem = _escrever_formulas(tmp_xlsx)
    # A formatacao condicional e escrita ANTES do gate COM: o Excel e a ultima
    # escrita do processo, de modo que o arquivo promovido carrega os valores
    # recalculados em cache (leitores openpyxl usam `data_only=True`).
    _reescrever_cf_remanesc(tmp_xlsx)

    pythoncom.CoInitialize()
    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    wb = None
    salvo = False
    try:
        wb = excel.Workbooks.Open(
            str(tmp_xlsx), UpdateLinks=0, ReadOnly=False, CorruptLoad=0
        )
        excel.ScreenUpdating = False
        excel.Calculation = XL_CALC_MANUAL
        aba_ativa = wb.ActiveSheet.Name
        for aba in (ABA_PC, ABA_MEM, ABA_RC, ABA_REM):
            if aba not in _nomes_abas(wb):
                raise ValueError(f"Aba {aba} ausente; layout inesperado.")

        excel.Calculation = XL_CALC_AUTOMATIC
        excel.CalculateFullRebuild()
        travas = _snapshot_travas(wb)
        _verificar_sem_erros(wb)

        mem = wb.Worksheets(ABA_MEM)
        trava_dupla = mem.Range("W55").Value
        if trava_dupla is None or abs(float(trava_dupla)) > 0.005:
            raise RuntimeError(
                f"TRAVA ANTI-DUPLA-CONTAGEM VIOLADA: W55={trava_dupla!r}"
            )

        if aba_ativa in _nomes_abas(wb):
            wb.Worksheets(aba_ativa).Activate()
        wb.Save()
        salvo = True
        wb.Close(SaveChanges=False)
        wb = None

        wb = excel.Workbooks.Open(
            str(tmp_xlsx), UpdateLinks=0, ReadOnly=True, CorruptLoad=0
        )
        if _snapshot_travas(wb) != travas:
            raise RuntimeError("TRAVA VIOLADA B26/T25 apos reabertura.")
        pc = wb.Worksheets(ABA_PC)
        if str(pc.Range(f"{COL_NASCIMENTO_DATA}1").Value or "") != (
            "CICLO_NASCIMENTO_DATA"
        ):
            raise RuntimeError("Coluna CICLO_NASCIMENTO_DATA nao persistiu.")
        mem = wb.Worksheets(ABA_MEM)
        f48 = str(mem.Range("W48").Formula)
        f48_r1c1 = str(mem.Range("W48").FormulaR1C1)
        if "W53" not in f48 and "R53C23" not in f48_r1c1:
            raise RuntimeError(
                f"FORMA 2 nao foi redecomposta em W48: {f48!r} / {f48_r1c1!r}"
            )
        wb.Close(SaveChanges=False)
        wb = None
        print("travas B26/T25:", travas)
    finally:
        if wb is not None:
            wb.Close(SaveChanges=False)
        excel.Quit()
        del wb
        del excel
        pythoncom.CoUninitialize()

    if not salvo:
        shutil.rmtree(tmp_dir, ignore_errors=True)
        raise RuntimeError("Excel nao salvou; destino preservado.")
    _conferir_cf_preservada(tmp_xlsx)
    destino.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(tmp_xlsx, destino)
    shutil.rmtree(tmp_dir, ignore_errors=True)
    for chave, valor in sorted(contagem.items()):
        print(f"{chave}: {valor}")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("origem", type=Path)
    parser.add_argument("destino", type=Path)
    args = parser.parse_args()
    aplicar(args.origem, args.destino)
    print("temporalidade por DATA_EFEITO aplicada:", args.destino)


if __name__ == "__main__":
    main()
