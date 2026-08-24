# -*- coding: utf-8 -*-
"""Etapa VTA — 3 referencias na Tabela 1 (RESULTADOS) + fotografia atual em itens_RC.

Aplica, via Excel COM em copia temporaria (padrao zero-corrupcao de
`aplicar_resultados_consolidados_26f.py`):

* MEMORIA_RESULTADOS: bloco auxiliar V/W (linhas 40..52) + coluna AB (2..201),
  motor das tres referencias do VTA. NAO altera B26/T25.
    - W41:W45  completude da abertura por ciclo (n=0..4);
    - W46      ultima abertura completa <= vigente (fallback C4->C0);
    - W47      motivo de nao adocao das aberturas posteriores;
    - AB2:AB201 remanescente por item na abertura selecionada (espelha Y com W46);
    - W48      FORMA 2 (VTA pela ultima posicao de abertura disponivel);
    - W49      CICLO_EM_EXECUCAO disponivel? (INDIRECT+ISERROR);
    - W50      FORMA 1 (VTA pela posicao atual do contrato);
    - W51/W52  reconciliacao FORMA 1 x FORMA 2.
* RESULTADOS: Tabela 1 redesenhada (linhas 8..13) com as tres referencias
  nominais obrigatorias + linha de reconciliacao. Nao mexe em linhas >= 14.
  Repointa o nome VTA_ATUALIZACAO_CHEIA para B12.
* itens_RC: bloco "POSICAO ATUAL (AUTO)" nas colunas Q..Y (linhas 3..202),
  consumindo CICLO_EM_EXECUCAO por INDIRECT+ISERROR (bloco vazio quando ausente).

Invariantes garantidas por construcao: quando W46 = T20 (abertura vigente
completa), W48 == T25 (=VTA-PC oficial), sem qualquer ajuste para forcar valor.
"""
from __future__ import annotations

import argparse
import shutil
import tempfile
import warnings
from pathlib import Path

import pythoncom
import win32com.client

XL_CALC_MANUAL = -4135
XL_CALC_AUTOMATIC = -4105
XL_CELLTYPE_FORMULAS = -4123
XL_CELLTYPE_CONSTANTS = 2
XL_ERRORS = 16
XL_CONTINUOUS = 1
XL_THIN = 2
XL_CENTER = -4108

ABA_MEMORIA = "MEMORIA_RESULTADOS"
ABA_RESULTADOS = "RESULTADOS"
ABA_ITENS_RC = "itens_RC"

# Formula promovida pelo hotfix RESULTADOS posterior a esta migracao. Este
# aplicador preserva sua semantica historica, mas nao pode ser reaplicado sobre
# o template oficial atual porque sua etapa era proprietaria da formula antiga.
W48_CANONICA_ATUAL = '=IF(OR($W$46="",$W$67=""),"",ROUND($W$67+$W$53+$W$54,2))'
ERRO_W48_TEMPLATE_ATUAL = (
    "Aplicador historico incompatível com o template oficial atual: "
    "MEMORIA_RESULTADOS!W48 ja contem a formula canonica e foi preservada."
)

MOEDA_BR = "R$ #.##0,00"
QTD_BR = "#.##0,00"
DATA_BR = "dd/mm/aaaa"


def _cor(hex_rgb: str) -> int:
    texto = hex_rgb.strip().lstrip("#")
    r, g, b = (int(texto[i : i + 2], 16) for i in (0, 2, 4))
    return r + (g << 8) + (b << 16)


CORES = {
    "vinho": _cor("8A1538"),
    "azul": _cor("1F4E78"),
    "azul_muito_claro": _cor("EEF4F8"),
    "cinza": _cor("F2F2F2"),
    "cinza_texto": _cor("595959"),
    "branco": _cor("FFFFFF"),
    "borda": _cor("B0C4D8"),
}

_PROT_FLAGS = (
    "AllowFormattingCells",
    "AllowFormattingColumns",
    "AllowFormattingRows",
    "AllowInsertingColumns",
    "AllowInsertingRows",
    "AllowInsertingHyperlinks",
    "AllowDeletingColumns",
    "AllowDeletingRows",
    "AllowSorting",
    "AllowFiltering",
    "AllowUsingPivotTables",
)


def _nomes_abas(wb) -> list[str]:
    return [ws.Name for ws in wb.Worksheets]


def _nomes_definidos(wb) -> set[str]:
    return {str(nome.Name).split("!")[-1] for nome in wb.Names}


def _capturar_protecao(ws):
    if not bool(ws.ProtectContents):
        return None, None
    protecao = ws.Protection
    estado = {}
    for flag in _PROT_FLAGS:
        try:
            estado[flag] = getattr(protecao, flag)
        except Exception:
            estado[flag] = True
    try:
        selecao = ws.EnableSelection
    except Exception:
        selecao = None
    ws.Unprotect()
    return estado, selecao


def _restaurar_protecao(ws, estado, selecao) -> None:
    if estado is None:
        return
    ws.Protect(DrawingObjects=True, Contents=True, Scenarios=True, **estado)
    if selecao is not None:
        try:
            ws.EnableSelection = selecao
        except Exception:
            pass


def _bordas(rng) -> None:
    for indice in (7, 8, 9, 10, 11, 12):
        borda = rng.Borders(indice)
        borda.LineStyle = XL_CONTINUOUS
        borda.Weight = XL_THIN
        borda.Color = CORES["borda"]


def _validar_origem(wb) -> None:
    abas = _nomes_abas(wb)
    obrigatorias = {ABA_MEMORIA, ABA_RESULTADOS, ABA_ITENS_RC, "posicao_contratual",
                    "historico_VU", "itens_PC", "comparativo_VTA", "CONTROLE"}
    ausentes = sorted(obrigatorias.difference(abas))
    if ausentes:
        raise ValueError(f"Abas obrigatorias ausentes: {', '.join(ausentes)}")
    mem = wb.Worksheets(ABA_MEMORIA)
    if str(mem.Range("W48").Formula or "") == W48_CANONICA_ATUAL:
        raise ValueError(ERRO_W48_TEMPLATE_ATUAL)
    # Travas: B26/T25 homologados devem existir e permanecer intactos.
    for endereco in ("B26", "T20", "T21", "T22", "T23", "T25", "Y2", "X2"):
        if not str(mem.Range(endereco).Formula or ""):
            raise ValueError(f"MEMORIA_RESULTADOS!{endereco} ausente; layout inesperado.")
    if "VTA_FINAL" not in _nomes_definidos(wb):
        raise ValueError("Nome definido VTA_FINAL ausente.")


def _recusar_template_atual_w48(caminho: Path) -> None:
    """Falha antes de criar copia mutavel ou iniciar o Excel COM."""
    from openpyxl import load_workbook

    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        wb = load_workbook(caminho, data_only=False, read_only=True, keep_links=False)
    try:
        formula = wb[ABA_MEMORIA]["W48"].value if ABA_MEMORIA in wb.sheetnames else None
    finally:
        wb.close()
    if formula == W48_CANONICA_ATUAL:
        raise ValueError(ERRO_W48_TEMPLATE_ATUAL)


def _snapshot_travas(wb) -> dict[str, str]:
    mem = wb.Worksheets(ABA_MEMORIA)
    return {e: str(mem.Range(e).Formula) for e in ("B26", "T25", "T21", "T22", "T23")}


# --------------------------------------------------------------------------- #
# A. MEMORIA_RESULTADOS — bloco auxiliar V/W + AB                              #
# --------------------------------------------------------------------------- #
def _formula_completude(col_rem: str, limiar: int) -> str:
    return (
        '=IF(COUNTIF(posicao_contratual!$A$2:$A$201,"<>")=0,0,'
        'IF(SUMPRODUCT((posicao_contratual!$A$2:$A$201<>"")'
        f'*(posicao_contratual!$Y$2:$Y$201<={limiar})'
        f'*(1-ISNUMBER(posicao_contratual!${col_rem}$2:${col_rem}$201)))>0,0,1))'
    )


def _formula_ab() -> str:
    # Espelha MEMORIA!Y2, trocando $T$20 por $W$46 (abertura selecionada).
    return (
        '=IF(OR(posicao_contratual!$A2="",$W$46=""),0,'
        'IF(AND(ISNUMBER(CHOOSE($W$46+1,historico_VU!$C2,historico_VU!$D2,'
        'historico_VU!$E2,historico_VU!$F2,historico_VU!$G2)),'
        'ISNUMBER(CHOOSE($W$46+1,posicao_contratual!$G2,posicao_contratual!$K2,'
        'posicao_contratual!$O2,posicao_contratual!$S2,posicao_contratual!$W2))),'
        'ROUND(ROUND(CHOOSE($W$46+1,historico_VU!$C2,historico_VU!$D2,'
        'historico_VU!$E2,historico_VU!$F2,historico_VU!$G2),2)'
        '*CHOOSE($W$46+1,posicao_contratual!$G2,posicao_contratual!$K2,'
        'posicao_contratual!$O2,posicao_contratual!$S2,posicao_contratual!$W2),2),""))'
    )


def _aplicar_memoria(wb) -> None:
    mem = wb.Worksheets(ABA_MEMORIA)
    estado, selecao = _capturar_protecao(mem)

    mem.Range("V40").Value = "APOIO — DECOMPOSICOES DO VTA (3 REFERENCIAS)"
    mem.Range("V40").Font.Bold = True

    labels_w = {
        41: "Abertura C0 completa?",
        42: "Abertura C1 completa?",
        43: "Abertura C2 completa?",
        44: "Abertura C3 completa?",
        45: "Abertura C4 completa?",
        46: "Ciclo da ultima abertura completa (num)",
        47: "Motivo de nao adocao das aberturas posteriores",
        48: "FORMA 2 — VTA ultima abertura disponivel",
        49: "CICLO_EM_EXECUCAO disponivel/utilizado?",
        50: "FORMA 1 — VTA posicao atual",
        51: "Reconciliacao (Posicao atual − Ultima abertura)",
        52: "Status reconciliacao",
    }
    for linha, texto in labels_w.items():
        mem.Range(f"V{linha}").Value = texto

    # W41:W45 — completude por ciclo.
    for linha, col_rem, limiar in (
        (41, "G", 0), (42, "K", 1), (43, "O", 2), (44, "S", 3), (45, "W", 4)
    ):
        mem.Range(f"W{linha}").Formula = _formula_completude(col_rem, limiar)

    # W46 — ultima abertura completa <= vigente.
    mem.Range("W46").Formula = (
        '=IF($T$20="","",'
        'IF(AND($T$20>=4,INDEX($W$41:$W$45,5)=1),4,'
        'IF(AND($T$20>=3,INDEX($W$41:$W$45,4)=1),3,'
        'IF(AND($T$20>=2,INDEX($W$41:$W$45,3)=1),2,'
        'IF(AND($T$20>=1,INDEX($W$41:$W$45,2)=1),1,'
        'IF(INDEX($W$41:$W$45,1)=1,0,""))))))'
    )
    # W47 — motivo de nao adocao.
    mem.Range("W47").Formula = (
        '=IF(OR($W$46="",$W$46=$T$20),"",'
        '"Aberturas posteriores a C"&$W$46&" incompletas (ciclo vigente C"&$T$20&").")'
    )

    # AB2:AB201 — remanescente por item na abertura selecionada.
    mem.Range("AB1").Value = "aux: remanescente por item na abertura selecionada (W46)"
    mem.Range("AB2:AB201").Formula = _formula_ab()

    # W48 — FORMA 2.
    mem.Range("W48").Formula = (
        '=IF($W$46="","",'
        'ROUND($T$21'
        '+SUMPRODUCT((ROW(itens_PC!$P$2:$P$6)-2>=1)'
        '*(ROW(itens_PC!$P$2:$P$6)-2<$W$46)*itens_PC!$P$2:$P$6)'
        '+SUM($AB$2:$AB$201),2))'
    )
    # W49 — CICLO_EM_EXECUCAO disponivel?
    mem.Range("W49").Formula = (
        '=IF(ISERROR(INDIRECT("CICLO_EM_EXECUCAO!A9")),0,'
        'IF(INDIRECT("CICLO_EM_EXECUCAO!A9")="",0,1))'
    )
    # W50 — FORMA 1 (posicao atual). IFERROR defensivo nas somas INDIRECT.
    mem.Range("W50").Formula = (
        '=IF(OR($W$49=0,$T$21="",NOT(ISNUMBER($T$21))),"",'
        'ROUND($T$21+$T$22'
        '+IFERROR(SUM(INDIRECT("CICLO_EM_EXECUCAO!F13:F211")),0)'
        '+IFERROR(SUM(INDIRECT("CICLO_EM_EXECUCAO!G13:G211")),0),2))'
    )
    # W51/W52 — reconciliacao.
    mem.Range("W51").Formula = (
        '=IF(OR($W$50="",$W$48=""),"",ROUND($W$50-$W$48,2))'
    )
    mem.Range("W52").Formula = (
        '=IF($W$50="","POSICAO ATUAL NAO INFORMADA",'
        'IF($W$48="","ULTIMA ABERTURA INDISPONIVEL",'
        'IF(ABS($W$51)<=$D$4,"RECONCILIADO",'
        '"REVISE — DECOMPOSICOES DO VTA NAO RECONCILIADAS")))'
    )

    mem.Range("W48").NumberFormatLocal = MOEDA_BR
    mem.Range("W50").NumberFormatLocal = MOEDA_BR
    mem.Range("W51").NumberFormatLocal = MOEDA_BR
    mem.Range("V40:V52").Font.Color = CORES["cinza_texto"]
    _restaurar_protecao(mem, estado, selecao)


# --------------------------------------------------------------------------- #
# B. RESULTADOS — Tabela 1 (3 referencias + reconciliacao)                     #
# --------------------------------------------------------------------------- #
def _aplicar_resultados(wb) -> None:
    ws = wb.Worksheets(ABA_RESULTADOS)
    estado, selecao = _capturar_protecao(ws)

    # Desfaz merges antigos da Tabela 1 antes de reconstruir.
    for endereco in ("C8:H8", "C9:H9", "C10:H10", "C11:H11", "A8:H8"):
        try:
            ws.Range(endereco).UnMerge()
        except Exception:
            pass

    # Titulo (H8 preservado — alimenta STATUS GLOBAL B3).
    ws.Range("A8").Value = "1. VALOR TOTAL DO CONTRATO — TRES REFERENCIAS DO VTA"

    # Cabecalho linha 9.
    ws.Range("A9").Value = "Referencia"
    ws.Range("B9").Value = "Valor"
    ws.Range("C9:E9").Merge()
    ws.Range("C9").Value = "Composicao auditavel (aba!celula)"
    ws.Range("F9:G9").Merge()
    ws.Range("F9").Value = "Fontes"
    ws.Range("H9").Value = "Situacao"

    # Linha 10 — FORMA 1.
    ws.Range("A10").Value = "VTA PELA POSICAO ATUAL DO CONTRATO"
    ws.Range("B10").Formula = (
        f'=IF({ABA_MEMORIA}!$W$50="","",{ABA_MEMORIA}!$W$50)'
    )
    ws.Range("C10:E10").Merge()
    ws.Range("C10").Formula = (
        f'=IF({ABA_MEMORIA}!$W$50="",'
        '"POSICAO ATUAL NAO INFORMADA (CICLO_EM_EXECUCAO ausente/incompleta)",'
        '"C0 exec (MEMORIA!T21) + ciclos encerrados (MEMORIA!T22) + execucao ate a data'
        ' + remanescente atual (CICLO_EM_EXECUCAO!F/G)")'
    )
    ws.Range("F10:G10").Merge()
    ws.Range("F10").Value = "CICLO_EM_EXECUCAO (fiscal) + itens_PC (C0/encerrados)"
    ws.Range("H10").Formula = (
        f'=IF({ABA_MEMORIA}!$W$50="","INDISPONIVEL — POSICAO ATUAL NAO INFORMADA",'
        f'IF({ABA_MEMORIA}!$W$52="RECONCILIADO","DISPONIVEL PARA CONFERENCIA",'
        f'{ABA_MEMORIA}!$W$52))'
    )

    # Linha 11 — FORMA 2.
    ws.Range("A11").Value = "VTA PELA ULTIMA POSICAO DE ABERTURA DISPONIVEL"
    ws.Range("B11").Formula = (
        f'=IF({ABA_MEMORIA}!$W$48="","",{ABA_MEMORIA}!$W$48)'
    )
    ws.Range("C11:E11").Merge()
    ws.Range("C11").Formula = (
        '="C0 exec (MEMORIA!T21) + ciclos encerrados (MEMORIA!T22)'
        ' + remanescente na abertura C"&'
        f'IF({ABA_MEMORIA}!$W$46="","?",{ABA_MEMORIA}!$W$46)&" (MEMORIA!AB)"&'
        f'IF({ABA_MEMORIA}!$W$47=""," "," — "&{ABA_MEMORIA}!$W$47)'
    )
    ws.Range("F11:G11").Merge()
    ws.Range("F11").Value = "itens_PC!P + posicao_contratual (G/K/O/S/W) x historico_VU"
    ws.Range("H11").Formula = (
        f'=IF({ABA_MEMORIA}!$W$48="","INCOMPLETO",'
        f'IF({ABA_MEMORIA}!$W$46={ABA_MEMORIA}!$T$20,'
        f'"ADOTADO — ABERTURA DO CICLO VIGENTE C"&{ABA_MEMORIA}!$T$20,'
        f'"ADOTADO — ULTIMA ABERTURA COMPLETA: C"&{ABA_MEMORIA}!$W$46))'
    )

    # Linha 12 — FORMA 3.
    ws.Range("A12").Value = "CONTRATO ORIGINAL INTEGRALMENTE REAJUSTADO"
    ws.Range("B12").Formula = '=IFERROR(comparativo_VTA!$B$208,"")'
    ws.Range("C12:E12").Merge()
    ws.Range("C12").Value = (
        "SUMPRODUCT(posicao_contratual!B x C) x CONTROLE!B11 (fator historico integral)"
    )
    ws.Range("F12:G12").Merge()
    ws.Range("F12").Value = "comparativo_VTA!B208"
    ws.Range("H12").Value = "COMPARATIVO — NAO E VTA OFICIAL"

    # Linha 13 — Reconciliacao.
    ws.Range("A13").Value = "Reconciliacao (Posicao atual − Ultima abertura)"
    ws.Range("B13").Formula = (
        f'=IF({ABA_MEMORIA}!$W$51="","",{ABA_MEMORIA}!$W$51)'
    )
    ws.Range("C13:E13").Merge()
    ws.Range("C13").Value = "Diferenca entre FORMA 1 e FORMA 2 (0 = reconciliado)"
    ws.Range("F13:G13").Merge()
    ws.Range("F13").Value = "MEMORIA!W50 − MEMORIA!W48"
    ws.Range("H13").Formula = f'={ABA_MEMORIA}!$W$52'

    # Formatos e estilo.
    ws.Range("B10:B13").NumberFormatLocal = MOEDA_BR
    ws.Range("A8:H8").Interior.Color = CORES["vinho"]
    ws.Range("A8:H8").Font.Color = CORES["branco"]
    ws.Range("A8:H8").Font.Bold = True
    ws.Range("A9:H9").Interior.Color = CORES["azul"]
    ws.Range("A9:H9").Font.Color = CORES["branco"]
    ws.Range("A9:H9").Font.Bold = True
    ws.Range("A10:H13").Font.Color = CORES["cinza_texto"]
    ws.Range("A10:B12").Font.Bold = True
    ws.Range("A10:H10").Interior.Color = CORES["azul_muito_claro"]
    ws.Range("A11:H11").Interior.Color = CORES["branco"]
    ws.Range("A12:H12").Interior.Color = CORES["cinza"]
    ws.Range("A12:H12").Font.Italic = True
    ws.Range("A13:H13").Interior.Color = CORES["azul_muito_claro"]
    ws.Range("C10:H13").WrapText = True
    ws.Range("C9:H9").HorizontalAlignment = XL_CENTER
    _bordas(ws.Range("A8:H13"))

    # Repointa VTA_ATUALIZACAO_CHEIA para B12 (contrato integralmente reajustado).
    if "VTA_ATUALIZACAO_CHEIA" in _nomes_definidos(wb):
        wb.Names("VTA_ATUALIZACAO_CHEIA").Delete()
    wb.Names.Add(Name="VTA_ATUALIZACAO_CHEIA", RefersTo=f"={ABA_RESULTADOS}!$B$12")

    _restaurar_protecao(ws, estado, selecao)


# --------------------------------------------------------------------------- #
# C. itens_RC — bloco POSICAO ATUAL (AUTO) colunas Q..Y                        #
# --------------------------------------------------------------------------- #
def _formula_indirect_col(letra: str) -> str:
    return (
        '=IF($A3="","",'
        f'IF(ISERROR(INDIRECT("CICLO_EM_EXECUCAO!{letra}"&(ROW()+10))),"",'
        f'INDIRECT("CICLO_EM_EXECUCAO!{letra}"&(ROW()+10))))'
    )


def _aplicar_itens_rc(wb) -> None:
    ws = wb.Worksheets(ABA_ITENS_RC)
    estado, selecao = _capturar_protecao(ws)

    ws.Range("Q1:Y1").Merge()
    ws.Range("Q1").Value = "POSICAO ATUAL NA DATA (AUTO — origem: CICLO_EM_EXECUCAO)"
    ws.Range("Q1").Font.Bold = True
    ws.Range("Q1").HorizontalAlignment = XL_CENTER
    ws.Range("Q1:Y1").Interior.Color = CORES["azul"]
    ws.Range("Q1").Font.Color = CORES["branco"]

    subcab = {
        "Q2": "DATA DA POSICAO ATUAL (AUTO)",
        "R2": "CICLO DA ULTIMA ABERTURA (AUTO)",
        "S2": "QTD REMANESCENTE NA ABERTURA (AUTO)",
        "T2": "QTD CONSUMIDA DESDE ABERTURA (AUTO)",
        "U2": "QTD REMANESCENTE NA DATA ATUAL (AUTO)",
        "V2": "VU ATUALIZADO NA DATA (AUTO)",
        "W2": "VALOR CONSUMIDO ATUALIZADO (AUTO)",
        "X2": "VALOR REMANESCENTE ATUALIZADO (AUTO)",
        "Y2": "STATUS DA POSICAO ATUAL (AUTO)",
    }
    for cel, txt in subcab.items():
        ws.Range(cel).Value = txt
    ws.Range("Q2:Y2").Font.Bold = True
    ws.Range("Q2:Y2").WrapText = True
    ws.Range("Q2:Y2").Interior.Color = CORES["cinza"]

    # Q — data da posicao (D5, absoluto).
    ws.Range("Q3:Q202").Formula = (
        '=IF($A3="","",'
        'IF(ISERROR(INDIRECT("CICLO_EM_EXECUCAO!$D$5")),"",'
        'INDIRECT("CICLO_EM_EXECUCAO!$D$5")))'
    )
    # R — ciclo da ultima abertura (MEMORIA!W46).
    ws.Range("R3:R202").Formula = (
        '=IF($A3="","",'
        f'IF({ABA_MEMORIA}!$W$46="","","C"&{ABA_MEMORIA}!$W$46))'
    )
    # S..X — por item (offset +10), colunas B/D/C/E/F/G da aba.
    ws.Range("S3:S202").Formula = _formula_indirect_col("B")  # rem. na abertura
    ws.Range("T3:T202").Formula = _formula_indirect_col("D")  # consumida
    ws.Range("U3:U202").Formula = _formula_indirect_col("C")  # rem. atual
    ws.Range("V3:V202").Formula = _formula_indirect_col("E")  # VU atualizado
    ws.Range("W3:W202").Formula = _formula_indirect_col("F")  # valor consumido
    ws.Range("X3:X202").Formula = _formula_indirect_col("G")  # valor remanescente
    # Y — status (fallback explicito quando aba ausente).
    ws.Range("Y3:Y202").Formula = (
        '=IF($A3="","",'
        'IF(ISERROR(INDIRECT("CICLO_EM_EXECUCAO!K"&(ROW()+10))),'
        '"CICLO_EM_EXECUCAO AUSENTE",INDIRECT("CICLO_EM_EXECUCAO!K"&(ROW()+10))))'
    )

    ws.Range("Q3:Q202").NumberFormatLocal = DATA_BR
    ws.Range("S3:U202").NumberFormatLocal = QTD_BR
    ws.Range("V3:X202").NumberFormatLocal = MOEDA_BR
    ws.Columns("Q:Y").ColumnWidth = 18
    _restaurar_protecao(ws, estado, selecao)


def _verificar_sem_erros(wb) -> None:
    import pywintypes

    problemas = []
    for ws in wb.Worksheets:
        for tipo in (XL_CELLTYPE_FORMULAS, XL_CELLTYPE_CONSTANTS):
            try:
                celulas = ws.UsedRange.SpecialCells(tipo, XL_ERRORS)
                problemas.append(f"{ws.Name}!{celulas.Address}")
            except pywintypes.com_error:
                continue
    if problemas:
        raise RuntimeError(f"Erros de formula apos recalculo: {problemas}")


def aplicar(origem: Path, destino: Path) -> None:
    origem = Path(origem).resolve()
    destino = Path(destino).resolve()
    if not origem.is_file():
        raise FileNotFoundError(origem)
    if origem == destino:
        raise ValueError("Origem e destino devem ser diferentes.")
    _recusar_template_atual_w48(origem)

    tmp_dir = Path(tempfile.mkdtemp(prefix="cl8us_vta_pos_"))
    tmp_xlsx = tmp_dir / origem.name
    shutil.copyfile(origem, tmp_xlsx)

    pythoncom.CoInitialize()
    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    wb = None
    salvo = False
    try:
        wb = excel.Workbooks.Open(
            str(tmp_xlsx), UpdateLinks=0, ReadOnly=False, CorruptLoad=0
        )
        excel.ScreenUpdating = False
        excel.Calculation = XL_CALC_MANUAL
        aba_ativa = wb.ActiveSheet.Name
        _validar_origem(wb)
        travas_antes = _snapshot_travas(wb)

        _aplicar_memoria(wb)
        _aplicar_resultados(wb)
        _aplicar_itens_rc(wb)

        # Trava: B26/T25 e componentes NAO podem ter mudado de formula.
        travas_depois = _snapshot_travas(wb)
        if travas_antes != travas_depois:
            difs = {k: (travas_antes[k], travas_depois[k])
                    for k in travas_antes if travas_antes[k] != travas_depois[k]}
            raise RuntimeError(f"TRAVA VIOLADA: B26/T25 alterados: {difs}")

        excel.Calculation = XL_CALC_AUTOMATIC
        excel.CalculateFullRebuild()
        _verificar_sem_erros(wb)

        if aba_ativa in _nomes_abas(wb):
            wb.Worksheets(aba_ativa).Activate()
        wb.Save()
        salvo = True
        wb.Close(SaveChanges=False)
        wb = None

        # Reabre sem reparo para provar zero-corrupcao.
        wb = excel.Workbooks.Open(
            str(tmp_xlsx), UpdateLinks=0, ReadOnly=True, CorruptLoad=0
        )
        abas = _nomes_abas(wb)
        for obrig in (ABA_MEMORIA, ABA_RESULTADOS, ABA_ITENS_RC):
            if obrig not in abas:
                raise RuntimeError(f"Aba {obrig} ausente apos reabertura.")
        if "VTA_ATUALIZACAO_CHEIA" not in _nomes_definidos(wb):
            raise RuntimeError("Nome VTA_ATUALIZACAO_CHEIA ausente apos reabertura.")
        wb.Close(SaveChanges=False)
        wb = None
    finally:
        if wb is not None:
            wb.Close(SaveChanges=False)
        excel.Quit()
        del wb
        del excel
        pythoncom.CoUninitialize()

    if not salvo:
        shutil.rmtree(tmp_dir, ignore_errors=True)
        raise RuntimeError("Excel nao salvou; destino preservado.")
    destino.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(tmp_xlsx, destino)
    shutil.rmtree(tmp_dir, ignore_errors=True)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("origem", type=Path)
    parser.add_argument("destino", type=Path)
    args = parser.parse_args()
    aplicar(args.origem, args.destino)
    print("VTA posicoes/Tabela 1/itens_RC aplicada:", args.destino)


if __name__ == "__main__":
    main()
