# -*- coding: utf-8 -*-
"""Etapa UI — acabamento visual do template oficial da Coleta.

Alteracao de APRESENTACAO apenas. Nenhuma formula, referencia, validacao,
nome definido, ordem de abas ou protecao de planilha e tocada. Frentes:

* parametros!G2:G6 — padroniza fonte/cor/preenchimento/bordas da coluna
  SITUACAO com o padrao das proprias linhas de dados da tabela (A2:A6).
  G3/G4 vinham em negrito azul-escuro (FF123B63); as demais em cinza
  regular (FF595959).
* parametros!A17:G20 — retira as bordas residuais que davam aparencia de
  continuacao da tabela MEMORIA DO FATOR (encerrada na linha 16). Nao mexe
  em showGridLines.
* parametros!G10 — reescreve APENAS o rotulo do marcador por ciclo, que
  hoje pressupoe a existencia de reajuste anterior. As validacoes
  OPCOES_SIM_NAO (G13:G15) e OPCOES_SIM_NAO_NA (G12) ficam intactas.
* posicao_referencia!H9:I11 — o painel de marcos tem 11 linhas de rotulo,
  mas o acabamento (preenchimento + moldura) parava na linha 8. Estende o
  MESMO padrao (copiado de H2/I2) ate a linha 11 e desloca a borda inferior
  de H8:I8 para H11:I11.
* cobertura_temporal coluna C — largura moderadamente maior, quebra
  automatica nas notas longas (C8/C14/C17) e altura suficiente para o texto
  caber na propria celula.
* RESULTADOS!H8 — fonte branca sobre os preenchimentos claros da formatacao
  condicional (VALIDADO/ESTIMADO/REVISE) tornava o status ilegivel. Passa a
  usar o mesmo padrao de H14/H24/H33 (cinza FF595959).
* RESULTADOS!A8 — desliga a quebra de texto (como A14/A24/A33/A41), de modo
  que o titulo da secao 1 ocupe horizontalmente a faixa A8:D8. Sem mesclagem.
* RESULTADOS!A54:C66 — a tabela 5 (TOTAIS CANONICOS DE PCs) era a unica sem
  linhas de tabela. Aplica a borda fina FFB0C4D8 usada nas demais tabelas da
  aba, sem habilitar gridlines globais.

Gravador: openpyxl. O template nao possui VML, comentarios, drawings, tabelas
nem pivots (25 entradas no zip), e o proprio gerador de producao
(_gerador_masterfile.gerar_masterfile_preenchido) ja faz round-trip deste
arquivo por openpyxl a cada Coleta emitida — logo nao ha superficie de
corrupcao especifica de COM a evitar aqui. A prova de ausencia de reparo
continua sendo o Excel real (tools/verificar_template_sem_reparo.py).

FAIL-CLOSED: recusa reaplicacao (marcador em parametros!G10).

Uso: python tools/aplicar_acabamento_visual_coleta.py <origem> <destino>
"""
from __future__ import annotations

import argparse
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.styles.colors import Color

ABA_PARAMETROS = "parametros"
ABA_POSICAO_REFERENCIA = "posicao_referencia"
ABA_COBERTURA = "cobertura_temporal"
ABA_RESULTADOS = "RESULTADOS"

# --- parametros: padrao das linhas de dados da tabela de ciclos (A2:H6) ------
# Referencia = coluna F (FATOR_ACUMULADO): e a unica que mantem o mesmo
# acabamento regular/cinza nas cinco linhas de dados. As linhas 3 e 4 (C1/C2)
# vem destacadas em negrito FF123B63 em A:E e H — residuo fora deste escopo,
# apenas registrado.
COLUNA_REFERENCIA_DADOS = 6    # F
COLUNA_SITUACAO = 7            # G
COR_TEXTO_DADOS = "FF595959"
COR_FUNDO_DADOS = "FFEDEDED"
COR_BORDA_DADOS = "FFD9D9D9"
LINHAS_CICLOS = range(2, 7)

# --- parametros: faixa sem conteudo operacional -----------------------------
FAIXA_LIMPA_PARAMETROS = (1, 7, range(17, 21))   # colunas A..G, linhas 17..20

# --- parametros!G10: rotulo do marcador de reajuste anterior -----------------
ROTULO_G10_ANTIGO = "REAJUSTE ANTERIOR JA FORMALIZADO?"
MARCADOR_G10_NOVO = "EXISTE REAJUSTE ANTERIOR FORMALIZADO?"
ROTULO_G10_NOVO = f"{MARCADOR_G10_NOVO} (Sim/Não; vazio=não comprovado)"

# --- posicao_referencia: painel de marcos em H:I ----------------------------
LINHA_MODELO_PAINEL = 2       # linha ja vestida, usada como padrao
ULTIMA_LINHA_PAINEL_ANTIGA = 8
ULTIMA_LINHA_PAINEL_NOVA = 11

# --- cobertura_temporal: coluna C (notas) -----------------------------------
LARGURA_C_COBERTURA = 62.0
CHARS_POR_LINHA_C = 60
ALTURA_LINHA_TEXTO = 14.5
NOTAS_COBERTURA = (8, 14, 17)

# --- RESULTADOS -------------------------------------------------------------
CELULA_STATUS_TABELA1 = "H8"
CELULA_STATUS_MODELO = "H14"
CELULA_TITULO_TABELA1 = "A8"
COR_BORDA_TABELAS_RESULTADOS = "FFB0C4D8"
TABELA5_PRIMEIRA_LINHA = 54    # cabecalho Medida/Valor/Composicao auditavel
TABELA5_ULTIMA_LINHA = 66      # 12a medida
TABELA5_COLUNAS = range(1, 4)  # A:C


def _formulas(wb) -> dict[str, str]:
    """Fotografia de TODAS as formulas do arquivo (trava de escopo)."""
    return {
        f"{ws.title}!{cell.coordinate}": cell.value
        for ws in wb.worksheets
        for row in ws.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    }


def _com_fonte(celula, **atributos) -> None:
    fonte = copy(celula.font)
    for nome, valor in atributos.items():
        setattr(fonte, nome, valor)
    celula.font = fonte


def _com_alinhamento(celula, **atributos) -> None:
    alinhamento = copy(celula.alignment)
    for nome, valor in atributos.items():
        setattr(alinhamento, nome, valor)
    celula.alignment = alinhamento


def _altura_para_texto(texto: str, chars_por_linha: int) -> float:
    linhas = max(1, -(-len(texto) // chars_por_linha))
    return round(linhas * ALTURA_LINHA_TEXTO, 2)


def _estilo_lado(borda, lado: str):
    """Estilo da borda tolerando lados ausentes (Border() nasce com None)."""
    return getattr(getattr(borda, lado, None), "style", None)


# ---------------------------------------------------------------------------
# parametros
# ---------------------------------------------------------------------------

def _padronizar_situacao_parametros(ws) -> None:
    """G2:G6 com o mesmo acabamento das demais celulas de dados da tabela."""
    fina = Side(style="thin", color=COR_BORDA_DADOS)
    for linha in LINHAS_CICLOS:
        celula = ws.cell(linha, COLUNA_SITUACAO)
        _com_fonte(
            celula,
            name="Calibri",
            sz=10.0,
            b=False,
            i=False,
            color=Color(rgb=COR_TEXTO_DADOS),
        )
        celula.fill = PatternFill("solid", fgColor=COR_FUNDO_DADOS)
        celula.border = Border(left=fina, right=fina, top=fina, bottom=fina)
        _com_alinhamento(celula, vertical="center")


def _limpar_faixa_sem_conteudo(ws) -> None:
    """A17:G20 sem bordas: a tabela MEMORIA DO FATOR termina na linha 16."""
    primeira, ultima, linhas = FAIXA_LIMPA_PARAMETROS
    for linha in linhas:
        for coluna in range(primeira, ultima + 1):
            ws.cell(linha, coluna).border = Border()


def _reescrever_rotulo_reajuste_anterior(ws) -> None:
    ws["G10"].value = ROTULO_G10_NOVO


# ---------------------------------------------------------------------------
# posicao_referencia
# ---------------------------------------------------------------------------

def _estender_painel_marcos(ws) -> None:
    """Leva o acabamento do painel H:I da linha 8 para a linha 11."""
    modelos = {
        8: ws.cell(LINHA_MODELO_PAINEL, 8),
        9: ws.cell(LINHA_MODELO_PAINEL, 9),
    }
    lado_inferior = copy(ws.cell(ULTIMA_LINHA_PAINEL_ANTIGA, 8).border.bottom)

    # Linhas 9..11 recebem o padrao do painel; a linha 8 perde a moldura
    # inferior, que passa a fechar o painel na ultima linha real (11).
    for linha in range(ULTIMA_LINHA_PAINEL_ANTIGA, ULTIMA_LINHA_PAINEL_NOVA + 1):
        for coluna, modelo in modelos.items():
            celula = ws.cell(linha, coluna)
            celula.fill = copy(modelo.fill)
            celula.border = copy(modelo.border)

    for coluna in modelos:
        celula = ws.cell(ULTIMA_LINHA_PAINEL_NOVA, coluna)
        borda = copy(celula.border)
        borda.bottom = copy(lado_inferior)
        celula.border = borda


# ---------------------------------------------------------------------------
# cobertura_temporal
# ---------------------------------------------------------------------------

def _acomodar_notas_cobertura(ws) -> None:
    ws.column_dimensions["C"].width = LARGURA_C_COBERTURA
    for linha in NOTAS_COBERTURA:
        celula = ws.cell(linha, 3)
        _com_alinhamento(celula, wrap_text=True, vertical="top")
        altura = _altura_para_texto(str(celula.value or ""), CHARS_POR_LINHA_C)
        atual = ws.row_dimensions[linha].height or 0.0
        ws.row_dimensions[linha].height = max(atual, altura)


# ---------------------------------------------------------------------------
# RESULTADOS
# ---------------------------------------------------------------------------

def _corrigir_status_tabela1(ws) -> None:
    """H8 herda o padrao de fonte de H14 (legivel sob a formatacao condicional)."""
    modelo = ws[CELULA_STATUS_MODELO].font
    _com_fonte(
        ws[CELULA_STATUS_TABELA1],
        name=modelo.name,
        sz=modelo.sz,
        b=modelo.b,
        i=modelo.i,
        color=copy(modelo.color),
    )


def _horizontalizar_titulo_tabela1(ws) -> None:
    """Sem quebra, o titulo ocupa a faixa A8:D8 (B8:D8 estao vazias)."""
    ocupadas = [
        coord for coord in ("B8", "C8", "D8") if ws[coord].value not in (None, "")
    ]
    if ocupadas:
        raise RuntimeError(
            f"RESULTADOS!{'/'.join(ocupadas)} tem conteudo; "
            "revisar antes de horizontalizar A8."
        )
    _com_alinhamento(ws[CELULA_TITULO_TABELA1], wrap_text=False, vertical="center")


def _aplicar_bordas_tabela5(ws) -> None:
    fina = Side(style="thin", color=COR_BORDA_TABELAS_RESULTADOS)
    for linha in range(TABELA5_PRIMEIRA_LINHA, TABELA5_ULTIMA_LINHA + 1):
        for coluna in TABELA5_COLUNAS:
            ws.cell(linha, coluna).border = Border(
                left=fina, right=fina, top=fina, bottom=fina
            )


# ---------------------------------------------------------------------------
# Orquestracao
# ---------------------------------------------------------------------------

def _validar_origem(wb) -> None:
    faltantes = [
        aba
        for aba in (
            ABA_PARAMETROS,
            ABA_POSICAO_REFERENCIA,
            ABA_COBERTURA,
            ABA_RESULTADOS,
        )
        if aba not in wb.sheetnames
    ]
    if faltantes:
        raise RuntimeError(f"Abas ausentes no template: {', '.join(faltantes)}")

    g10 = str(wb[ABA_PARAMETROS]["G10"].value or "")
    if g10.startswith(MARCADOR_G10_NOVO):
        raise RuntimeError("Acabamento visual ja aplicado (parametros!G10). FAIL-CLOSED.")
    if ROTULO_G10_ANTIGO not in g10:
        raise RuntimeError(f"parametros!G10 inesperado: {g10!r} — nao aplicar as cegas.")


def _conferir_resultado(wb) -> None:
    par = wb[ABA_PARAMETROS]
    for linha in LINHAS_CICLOS:
        alvo = par.cell(linha, COLUNA_SITUACAO)
        referencia = par.cell(linha, COLUNA_REFERENCIA_DADOS)
        if (alvo.font.b, alvo.font.sz, alvo.font.color.rgb) != (
            referencia.font.b,
            referencia.font.sz,
            referencia.font.color.rgb,
        ):
            raise RuntimeError(f"parametros!G{linha} nao seguiu o padrao de F{linha}.")
        if alvo.fill.fgColor.rgb != referencia.fill.fgColor.rgb:
            raise RuntimeError(f"parametros!G{linha} com preenchimento divergente.")

    _, ultima_coluna, linhas = FAIXA_LIMPA_PARAMETROS
    for linha in linhas:
        for coluna in range(1, ultima_coluna + 1):
            celula = par.cell(linha, coluna)
            if any(
                _estilo_lado(celula.border, lado)
                for lado in ("left", "right", "top", "bottom")
            ):
                raise RuntimeError(f"parametros!{celula.coordinate} ainda tem borda.")
    if not str(par["G10"].value or "").startswith(MARCADOR_G10_NOVO):
        raise RuntimeError("parametros!G10 nao foi reescrita.")

    pos = wb[ABA_POSICAO_REFERENCIA]
    modelo = pos.cell(LINHA_MODELO_PAINEL, 8)
    for linha in range(9, ULTIMA_LINHA_PAINEL_NOVA + 1):
        if pos.cell(linha, 8).fill.fgColor.rgb != modelo.fill.fgColor.rgb:
            raise RuntimeError(
                f"posicao_referencia!H{linha} sem o acabamento do painel."
            )
    if _estilo_lado(pos.cell(ULTIMA_LINHA_PAINEL_ANTIGA, 8).border, "bottom"):
        raise RuntimeError("posicao_referencia!H8 ainda fecha o painel.")
    if not _estilo_lado(pos.cell(ULTIMA_LINHA_PAINEL_NOVA, 8).border, "bottom"):
        raise RuntimeError("posicao_referencia!H11 nao fecha o painel.")

    cob = wb[ABA_COBERTURA]
    if (cob.column_dimensions["C"].width or 0) < LARGURA_C_COBERTURA:
        raise RuntimeError("cobertura_temporal coluna C nao foi alargada.")
    for linha in NOTAS_COBERTURA:
        if not cob.cell(linha, 3).alignment.wrap_text:
            raise RuntimeError(f"cobertura_temporal!C{linha} sem quebra de texto.")

    res = wb[ABA_RESULTADOS]
    if (
        res[CELULA_STATUS_TABELA1].font.color.rgb
        != res[CELULA_STATUS_MODELO].font.color.rgb
    ):
        raise RuntimeError("RESULTADOS!H8 nao herdou a fonte de H14.")
    if res[CELULA_TITULO_TABELA1].alignment.wrap_text:
        raise RuntimeError("RESULTADOS!A8 ainda quebra o texto.")
    for linha in (TABELA5_PRIMEIRA_LINHA, TABELA5_ULTIMA_LINHA):
        for coluna in TABELA5_COLUNAS:
            celula = res.cell(linha, coluna)
            if _estilo_lado(celula.border, "left") != "thin":
                raise RuntimeError(f"RESULTADOS!{celula.coordinate} sem borda.")


def aplicar(origem: Path, destino: Path) -> None:
    origem = Path(origem).resolve()
    destino = Path(destino).resolve()
    if not origem.is_file():
        raise FileNotFoundError(origem)
    if origem == destino:
        raise ValueError("Origem e destino devem ser diferentes.")

    wb = load_workbook(origem, data_only=False)
    try:
        _validar_origem(wb)
        formulas_antes = _formulas(wb)

        par = wb[ABA_PARAMETROS]
        _padronizar_situacao_parametros(par)
        _limpar_faixa_sem_conteudo(par)
        _reescrever_rotulo_reajuste_anterior(par)

        _estender_painel_marcos(wb[ABA_POSICAO_REFERENCIA])
        _acomodar_notas_cobertura(wb[ABA_COBERTURA])

        res = wb[ABA_RESULTADOS]
        _corrigir_status_tabela1(res)
        _horizontalizar_titulo_tabela1(res)
        _aplicar_bordas_tabela5(res)

        formulas_depois = _formulas(wb)
        if formulas_depois != formulas_antes:
            divergentes = sorted(
                chave
                for chave in set(formulas_antes) | set(formulas_depois)
                if formulas_antes.get(chave) != formulas_depois.get(chave)
            )
            raise RuntimeError(f"TRAVA VIOLADA: formulas alteradas: {divergentes[:10]}")

        _conferir_resultado(wb)
        destino.parent.mkdir(parents=True, exist_ok=True)
        wb.save(destino)
    finally:
        wb.close()

    conferencia = load_workbook(destino, data_only=False)
    try:
        if _formulas(conferencia) != formulas_antes:
            raise RuntimeError("TRAVA VIOLADA: formulas alteradas na gravacao.")
        _conferir_resultado(conferencia)
    finally:
        conferencia.close()


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("origem", type=Path)
    parser.add_argument("destino", type=Path)
    args = parser.parse_args()
    aplicar(args.origem, args.destino)
    print("Acabamento visual da Coleta aplicado:", args.destino)


if __name__ == "__main__":
    main()
