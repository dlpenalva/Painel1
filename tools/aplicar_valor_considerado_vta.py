# -*- coding: utf-8 -*-
"""Alinha a camada de formulas do XLSX ao VALOR CONSIDERADO por PC.

Motivacao
---------
`itens_PC!P` (VALOR_ATUALIZADO_TOTAL por ciclo) e a soma de `F = D * E`, com
`E = FATOR_ACUMULADO` historico do ciclo. Essa medida permanece valida como
referencia tecnica (Etapa 26C), mas NAO e a que compoe resultado, retroativo,
VTA e documentos.

A medida que governa e o VALOR HISTORICO CONSIDERADO: base efetivamente
executada mais o retroativo reconhecido a pagar. O PC anterior ao inicio do
efeito financeiro do seu ciclo pertence ao ciclo, usa fator efetivo 1 e
permanece no valor original -- por isso nao pode entrar no VTA multiplicado
pelo fator integral.

Por ciclo, o valor considerado ja existe na propria aba:
``itens_PC!O`` (VALOR_PC_TOTAL) + ``itens_PC!Q`` (RETROATIVO_RECONHECIDO).

Este utilitario troca a fonte de `MEMORIA_RESULTADOS!T21`, `T22` e `W48` de
`P` para `O+Q`, publica o valor considerado por PC em `itens_PC!U` e atualiza
os rotulos correspondentes. `T23`, `T25`, `W50`, `W51`, `W52` e `B26` sao
travados: suas formulas nao podem mudar.

Uso:
    python tools/aplicar_valor_considerado_vta.py <origem.xlsx> <destino.xlsx>
"""
from __future__ import annotations

import argparse
import shutil
import tempfile
import warnings
from pathlib import Path

import pythoncom
import win32com.client

ABA_MEMORIA = "MEMORIA_RESULTADOS"
ABA_RESULTADOS = "RESULTADOS"
ABA_ITENS_PC = "itens_PC"

# Formula promovida pelo hotfix RESULTADOS posterior a esta migracao. A
# aplicacao completa deste script mantem a semantica historica e, portanto,
# deve recusar o template atual antes de tentar substituir W48.
W48_CANONICA_ATUAL = '=IF(OR($W$46="",$W$67=""),"",ROUND($W$67+$W$53+$W$54,2))'
ERRO_W48_TEMPLATE_ATUAL = (
    "Aplicador historico incompatível com o template oficial atual: "
    "MEMORIA_RESULTADOS!W48 ja contem a formula canonica e foi preservada."
)

XL_CALC_MANUAL = -4135
XL_CALC_AUTOMATIC = -4105
XL_CELLTYPE_FORMULAS = -4123
XL_CELLTYPE_CONSTANTS = 2
XL_ERRORS = 16

XL_PATTERN_NONE = -4142
XL_LINE_STYLE_NONE = -4142
# xlEdgeLeft, xlEdgeTop, xlEdgeBottom, xlEdgeRight.
ARESTAS = (7, 8, 9, 10)
# xlInsideHorizontal. Num intervalo de varias linhas, xlEdgeTop/xlEdgeBottom
# desenham apenas as bordas EXTERNAS do bloco; sem esta aresta, as 5.000 linhas
# de `U` ficariam sem as divisorias horizontais que `H` tem celula a celula.
XL_INSIDE_HORIZONTAL = 12

ULTIMA_LINHA_PC = 5001

# Coluna-modelo da apresentacao de `itens_PC!U`. `H`
# (RETROATIVO_RECONHECIDO_A_PAGAR) e uma das duas parcelas somadas por `U` e
# tem exatamente o mesmo perfil: derivada, por linha, monetaria e viva em
# 2:5001. O bloco de resumo por ciclo `M:T` NAO serve de modelo: vive apenas
# nas linhas 2:7 e `T` (QTD_COM_CHECK) e contagem, nao moeda.
COLUNA_MODELO_U = "H"
LARGURA_U = 24

# Soma, por ciclo, do VALOR CONSIDERADO (VALOR_PC_TOTAL + RETROATIVO).
_CONSIDERADO_POR_CICLO = "(itens_PC!$O$2:$O$6+itens_PC!$Q$2:$Q$6)"
_FAIXA_CICLOS = "(ROW(itens_PC!$O$2:$O$6)-2"

FORMULA_T21 = (
    "=IF(itens_PC!$O$2>0,ROUND(itens_PC!$O$2+itens_PC!$Q$2,2),"
    "SUM($X$2:$X$201))"
)
FORMULA_T22 = (
    "=SUMPRODUCT(" + _FAIXA_CICLOS + ">=1)*" + _FAIXA_CICLOS + "<$T$20)*"
    + _CONSIDERADO_POR_CICLO + ")"
)
FORMULA_W48 = (
    '=IF($W$46="","",ROUND($T$21+SUMPRODUCT('
    + _FAIXA_CICLOS + ">=1)*" + _FAIXA_CICLOS + "<$W$46)*"
    + _CONSIDERADO_POR_CICLO + ")+$W$53+$W$54,2))"
)
FORMULA_U_PC = '=IF(OR($D{linha}="",$H{linha}=""),"",ROUND($D{linha}+$H{linha},2))'

# C0 executado por item: movimentacao entre as ABERTURAS TEMPORALMENTE
# CORRETAS. A abertura de um ciclo e a quantidade declarada MENOS os deltas
# com data de efeito posterior a sua abertura (DELTA_POSTERIOR_ABERTURA_Cn):
# um aditivo do meio de C1 nao pode retroagir sobre a fotografia de abertura
# de C1 e, por consequencia, encolher a execucao de C0.
FORMULA_X_C0 = (
    '=IF(posicao_contratual!$A2="",0,'
    'IF(AND(ISNUMBER(posicao_contratual!$Y2),posicao_contratual!$Y2>0),0,'
    "IF(AND(ISNUMBER(posicao_contratual!$G2),ISNUMBER(posicao_contratual!$K2),"
    "ISNUMBER(historico_VU!$C2)),"
    "((posicao_contratual!$G2-N(posicao_contratual!$AB2))"
    "-(posicao_contratual!$K2-N(posicao_contratual!$AC2)))*historico_VU!$C2,"
    '"")))'
)
ULTIMA_LINHA_POSICAO = 201
ROTULO_X1 = (
    "VTA-PC aux: C0 fisico por item ((abertura C0 - abertura C1) x VU_C0)"
)

ROTULO_S21 = "C0 executado (valor considerado: PC C0 ou movimentacao fisica x VU_C0)"
ROTULO_S22 = "Execucao PC (ciclos 1..vigente-1, VALOR CONSIDERADO = base + retroativo)"
ROTULO_U1 = "VALOR_CONSIDERADO"
FONTE_F11 = (
    "itens_PC!O+Q (valor considerado) + posicao_contratual (G/K/O/S/W) x historico_VU"
)

# Formulas que NAO podem mudar nesta aplicacao.
TRAVAS = ("T23", "T25", "W50", "W51", "W52", "B26")

# --------------------------------------------------------------------------- #
# Posicao fisica do ciclo em execucao e status global
# --------------------------------------------------------------------------- #
# `MEMORIA_RESULTADOS!W49` vale 1 quando CICLO_EM_EXECUCAO!A9 nao esta vazia --
# e A9 so produz valor quando ha data de posicao, ha itens e nenhum item esta
# em ERRO ou INCOMPLETO. E, portanto, a prova de posicao fisica COMPLETA e
# VALIDA exigida para que ela prevaleca sobre a estimativa por PCs.
COND_FISICO = "MEMORIA_RESULTADOS!$W$49=1"
SOMA_EXECUCAO_FISICA = (
    'ROUND(IFERROR(SUM(INDIRECT("CICLO_EM_EXECUCAO!F13:F211")),0),2)'
)
MARCA_FISICO = "CICLO_EM_EXECUCAO!F13:F211"

FORMULA_H33 = (
    '=IF(OR($B$35="",$B$36="",$B$37="",$B$38="",$B$38<0),"REVISE",'
    "IF(OR(posicao_referencia!$I$2=TRUE," + COND_FISICO + '),"VALIDADO","ESTIMADO"))'
)
FORMULA_A39 = (
    "=IF(" + COND_FISICO + ',"Posicao fisica itemizada informada pelo fiscal '
    '(CICLO_EM_EXECUCAO): execucao e remanescente do ciclo atual sao medidos, '
    'nao estimados.",'
    'IF(AND($B$36<>"",posicao_referencia!$I$2<>TRUE),'
    '"Estimativa: assume execucao registrada pelo metodo oficial ("&$B$5&") '
    'aproximadamente igual ao consumo fisico do ciclo.",""))'
)

# Trava anti-dupla-contagem dos aditivos: o calculo manual so continua exigido
# enquanto NAO houver prova de que a base contratual ja incorporou o aditivo.
# A prova sao as duas travas independentes fechando em zero: a reconciliacao
# das decomposicoes do VTA (W51) e a trava anti-dupla-contagem (W55).
GUARDA_ADITIVOS_ANTIGA = "$B$45>0"
GUARDA_ADITIVOS_NOVA = (
    "AND($B$45>0,NOT(AND(ISNUMBER($W$51),ROUND($W$51,2)=0,"
    "ISNUMBER($W$55),ROUND($W$55,2)=0)))"
)
NOTA_A27 = (
    "Aditivos marcados como considerados exigem prova de que a base contratual "
    "ja os incorpora. A planilha so dispensa o calculo manual quando a "
    "reconciliacao das duas decomposicoes do VTA e a trava anti-dupla-contagem "
    "fecham ambas em 0,00; caso contrario, segue exigindo decisao humana."
)

# --------------------------------------------------------------------------- #
# DATA DE CORTE e TOTAIS CANONICOS DE PCs
# --------------------------------------------------------------------------- #
# CONTROLE!B3 e a data de corte unica do contrato. PC com DATA_PC posterior
# permanece no inventario do arquivo e sai de todo resultado "ate o corte".
# A exclusao exige PROVA de posterioridade: corte ausente ou data ilegivel
# nunca excluem -- por isso o limite cai para 31/12/9999 quando B3 nao e data,
# e os totais sao obtidos SUBTRAINDO os comprovadamente posteriores.
LIMITE_CORTE = "T31"
FORMULA_LIMITE_CORTE = (
    "=IF(ISNUMBER(CONTROLE!$B$3),CONTROLE!$B$3,DATE(9999,12,31))"
)
ROTULO_LIMITE_CORTE = "Limite da data de corte (CONTROLE!B3; sem corte = 31/12/9999)"

_D = "itens_PC!$D$2:$D$5001"
_B = "itens_PC!$B$2:$B$5001"
_C = "itens_PC!$C$2:$C$5001"
_L = "itens_PC!$L$2:$L$5001"
_POSTERIOR = f'{_B},">"&$T$31'


def _posteriores_por_ciclo() -> str:
    return "+".join(
        f'SUMIFS({_D},{_C},"C{n}",{_POSTERIOR})' for n in range(5)
    )


TOTAIS_CANONICOS = {
    # (celula, rotulo, formula)
    "T33": (
        "Total cadastrado de PCs (inventario integral do arquivo)",
        f"=ROUND(SUM({_D}),2)",
    ),
    "T34": (
        "Total considerado ate a data de corte",
        f'=ROUND($T$33-SUMIFS({_D},{_POSTERIOR}),2)',
    ),
    "T35": (
        "Total enquadrado nos ciclos (ate o corte)",
        "=ROUND(SUM(itens_PC!$O$2:$O$6)-(" + _posteriores_por_ciclo() + "),2)",
    ),
    "T36": (
        "Total com efeito financeiro (ate o corte)",
        f'=ROUND(SUMIFS({_D},{_L},"Sim")-SUMIFS({_D},{_L},"Sim",{_POSTERIOR}),2)',
    ),
    "T37": (
        "Total sem efeito financeiro (ate o corte)",
        "=ROUND($T$35-$T$36,2)",
    ),
}

# Retroativo por ciclo (metodo PC) tambem passa a respeitar a data de corte.
FORMULA_RETRO_CICLO = (
    '=IF(COUNTIFS({c},"{ciclo}",{d},">0")=0,"",'
    'ROUND(SUMIFS(itens_PC!$H$2:$H$5001,{c},"{ciclo}",'
    'itens_PC!$G$2:$G$5001,"Sim",{b},"<="&$T$31),2))'
)

# RESULTADOS: base considerada por ciclo (ramo PCs) filtrada pelo corte.
FILTRO_PC_ANTIGO = 'itens_PC!$G$2:$G$5001,"Sim"'
FILTRO_PC_NOVO = (
    'itens_PC!$G$2:$G$5001,"Sim",itens_PC!$B$2:$B$5001,'
    '"<="&MEMORIA_RESULTADOS!$T$31'
)

# Secao 5 de RESULTADOS: as doze medidas, com nomes claros.
LINHA_SECAO_TOTAIS = 53
MEDIDAS_CANONICAS = [
    ("Total cadastrado de PCs", "=IF($B$5<>\"PCs\",\"\",MEMORIA_RESULTADOS!$T$33)",
     "itens_PC!D (inventario integral, sem corte)"),
    ("Total considerado ate a data de corte",
     "=IF($B$5<>\"PCs\",\"\",MEMORIA_RESULTADOS!$T$34)",
     "MEMORIA!T34 = T33 - PCs com DATA_PC > CONTROLE!B3"),
    ("Total enquadrado nos ciclos",
     "=IF($B$5<>\"PCs\",\"\",MEMORIA_RESULTADOS!$T$35)",
     "itens_PC!O (por ciclo) - posteriores ao corte"),
    ("Total com efeito financeiro",
     "=IF($B$5<>\"PCs\",\"\",MEMORIA_RESULTADOS!$T$36)",
     "itens_PC!L=Sim, ate o corte"),
    ("Total sem efeito financeiro",
     "=IF($B$5<>\"PCs\",\"\",MEMORIA_RESULTADOS!$T$37)",
     "enquadrado - com efeito (competencias anteriores ao inicio do efeito)"),
    ("Retroativo reconhecido", "=$D$22", "MEMORIA!B16 (retroativo oficial)"),
    ("Execucao fisica do ciclo atual", "=$B$36",
     "CICLO_EM_EXECUCAO!F13:F211 quando a posicao fisica esta completa"),
    ("Remanescente fisico atual", "=$B$38",
     "abertura do ciclo vigente - execucao fisica"),
    ("VTA oficial", "=$B$10", "MEMORIA!W50 (FORMA 1 - posicao atual)"),
    ("Referencias auditaveis", '="Linhas 10 a 12 desta aba"',
     "posicao atual / ultima abertura / contrato integralmente reajustado"),
    ("Reconciliacao", "=$B$13", "MEMORIA!W51 (FORMA 1 - FORMA 2); 0,00 reconcilia"),
    ("Status", "=$B$3", "STATUS GLOBAL desta aba"),
]
TITULO_SECAO_TOTAIS = "5. TOTAIS CANONICOS DE PCs — MEDIDAS COM NOMES CLAROS"


def _nomes_abas(wb) -> list[str]:
    return [ws.Name for ws in wb.Worksheets]


def _validar_origem(wb, *, escreve_w48: bool = True) -> None:
    abas = _nomes_abas(wb)
    for obrig in (ABA_MEMORIA, ABA_RESULTADOS, ABA_ITENS_PC):
        if obrig not in abas:
            raise RuntimeError(f"Aba {obrig} ausente na origem.")
    mem = wb.Worksheets(ABA_MEMORIA)
    if escreve_w48 and str(mem.Range("W48").Formula or "") == W48_CANONICA_ATUAL:
        raise RuntimeError(ERRO_W48_TEMPLATE_ATUAL)
    for endereco in ("T21", "T22", "T23", "T25", "W48", "W50"):
        if not str(mem.Range(endereco).Formula or ""):
            raise RuntimeError(
                f"{ABA_MEMORIA}!{endereco} vazio: origem incompativel."
            )
    if str(wb.Worksheets(ABA_ITENS_PC).Range("O1").Value or "") != "VALOR_PC_TOTAL":
        raise RuntimeError("itens_PC!O1 nao e VALOR_PC_TOTAL; layout inesperado.")


def _recusar_template_atual_w48(caminho: Path) -> None:
    """Falha antes de criar copia mutavel ou iniciar o Excel COM."""
    from openpyxl import load_workbook

    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        wb = load_workbook(caminho, data_only=False, read_only=True, keep_links=False)
    try:
        formula = wb[ABA_MEMORIA]["W48"].value if ABA_MEMORIA in wb.sheetnames else None
    finally:
        wb.close()
    if formula == W48_CANONICA_ATUAL:
        raise RuntimeError(ERRO_W48_TEMPLATE_ATUAL)


def _snapshot_travas(wb) -> dict[str, str]:
    mem = wb.Worksheets(ABA_MEMORIA)
    return {e: str(mem.Range(e).Formula) for e in TRAVAS}


def _aplicar_memoria(wb) -> None:
    mem = wb.Worksheets(ABA_MEMORIA)
    mem.Range("T21").Formula = FORMULA_T21
    mem.Range("T22").Formula = FORMULA_T22
    mem.Range("W48").Formula = FORMULA_W48
    mem.Range("S21").Value = ROTULO_S21
    mem.Range("S22").Value = ROTULO_S22
    mem.Range("X1").Value = ROTULO_X1
    mem.Range(f"X2:X{ULTIMA_LINHA_POSICAO}").Formula = FORMULA_X_C0


def _copiar_borda(borda_origem, borda_destino) -> None:
    borda_destino.LineStyle = borda_origem.LineStyle
    if borda_origem.LineStyle != XL_LINE_STYLE_NONE:
        borda_destino.Weight = borda_origem.Weight
        borda_destino.Color = borda_origem.Color


def _copiar_apresentacao(origem, destino, divisorias_internas: bool = False) -> None:
    """Replica a apresentacao de `origem` em `destino`, propriedade a propriedade.

    Deliberadamente NAO usa `Copy()` + `PasteSpecial(xlPasteFormats)`: a colagem
    de formatos arrasta junto as regras de formatacao condicional da origem, e a
    coluna-modelo esta dentro de `A2:L5001`, que carrega o alerta de linha
    incompleta. Copiar por propriedade mantem `U` livre de formatacao
    condicional.
    """
    destino.NumberFormatLocal = origem.NumberFormatLocal
    destino.HorizontalAlignment = origem.HorizontalAlignment
    destino.VerticalAlignment = origem.VerticalAlignment
    destino.WrapText = origem.WrapText

    destino.Interior.Pattern = origem.Interior.Pattern
    if origem.Interior.Pattern != XL_PATTERN_NONE:
        destino.Interior.Color = origem.Interior.Color

    destino.Font.Name = origem.Font.Name
    destino.Font.Size = origem.Font.Size
    destino.Font.Bold = origem.Font.Bold
    destino.Font.Italic = origem.Font.Italic
    destino.Font.Color = origem.Font.Color

    for aresta in ARESTAS:
        _copiar_borda(origem.Borders(aresta), destino.Borders(aresta))
    if divisorias_internas:
        # A divisoria entre linhas espelha a borda inferior da celula-modelo.
        _copiar_borda(
            origem.Borders(9), destino.Borders(XL_INSIDE_HORIZONTAL)
        )

    destino.Locked = origem.Locked


def _aplicar_estilo_u(wb) -> None:
    """Integra `itens_PC!U` a tabela A:L, sem tocar valor, formula ou rotulo.

    `U` e uma coluna POR LINHA: a formula le `$D` (VALOR_PC) e `$H`
    (RETROATIVO_RECONHECIDO_A_PAGAR) da propria linha, em `U2:U5001`. Ela
    pertence, portanto, a tabela `A:L` -- e nao ao bloco de resumo por ciclo
    `M:T`, que so vive nas linhas 2:7 e usa formato de contagem `#,##0`. A
    coluna-modelo e `H`: derivada, por linha, monetaria e com o mesmo alcance.
    """
    ws = wb.Worksheets(ABA_ITENS_PC)

    rotulo_antes = ws.Range("U1").Value
    formulas_antes = (
        str(ws.Range("U2").Formula),
        str(ws.Range(f"U{ULTIMA_LINHA_PC}").Formula),
    )

    _copiar_apresentacao(ws.Range(f"{COLUNA_MODELO_U}1"), ws.Range("U1"))
    _copiar_apresentacao(
        ws.Range(f"{COLUNA_MODELO_U}2"),
        ws.Range(f"U2:U{ULTIMA_LINHA_PC}"),
        divisorias_internas=True,
    )
    ws.Columns("U").ColumnWidth = LARGURA_U

    if ws.Range("U1").Value != rotulo_antes:
        raise RuntimeError("itens_PC!U1 mudou de rotulo ao aplicar o estilo.")
    formulas_depois = (
        str(ws.Range("U2").Formula),
        str(ws.Range(f"U{ULTIMA_LINHA_PC}").Formula),
    )
    if formulas_depois != formulas_antes:
        raise RuntimeError(
            f"Formulas de itens_PC!U mudaram ao aplicar o estilo: "
            f"{formulas_antes!r} -> {formulas_depois!r}"
        )


def _aplicar_itens_pc(wb) -> None:
    ws = wb.Worksheets(ABA_ITENS_PC)
    ws.Range("U1").Value = ROTULO_U1
    ws.Range(f"U2:U{ULTIMA_LINHA_PC}").Formula = FORMULA_U_PC.format(linha=2)
    _aplicar_estilo_u(wb)


def _aplicar_resultados(wb) -> None:
    ws = wb.Worksheets(ABA_RESULTADOS)
    ws.Range("F11").Value = FONTE_F11

    # Ciclo atual: a execucao fisica itemizada PREVALECE sobre a estimativa
    # pelo metodo oficial. A formula estimativa e preservada integralmente como
    # ramo alternativo (fallback homologado quando nao ha posicao fisica).
    atual = str(ws.Range("B36").Formula or "")
    if MARCA_FISICO not in atual:
        if not atual.startswith("="):
            raise RuntimeError("RESULTADOS!B36 nao e formula; origem inesperada.")
        ws.Range("B36").Formula = (
            "=IF(" + COND_FISICO + "," + SOMA_EXECUCAO_FISICA
            + ",(" + atual[1:] + "))"
        )
    ws.Range("H33").Formula = FORMULA_H33
    ws.Range("A39").Formula = FORMULA_A39


def _aplicar_corte(wb) -> None:
    """Aplica CONTROLE!B3 aos resultados 'ate a data de corte'."""
    mem = wb.Worksheets(ABA_MEMORIA)
    mem.Range(f"S{LIMITE_CORTE[1:]}").Value = ROTULO_LIMITE_CORTE
    mem.Range(LIMITE_CORTE).Formula = FORMULA_LIMITE_CORTE

    for n in range(5):
        mem.Range(f"C{10 + n}").Formula = FORMULA_RETRO_CICLO.format(
            c=_C, d=_D, b=_B, ciclo=f"C{n}"
        )

    for celula, (rotulo, formula) in TOTAIS_CANONICOS.items():
        mem.Range("S" + celula[1:]).Value = rotulo
        mem.Range(celula).Formula = formula

    res = wb.Worksheets(ABA_RESULTADOS)
    for linha in range(16, 21):
        atual = str(res.Range(f"B{linha}").Formula or "")
        if FILTRO_PC_NOVO in atual:
            continue
        if FILTRO_PC_ANTIGO not in atual:
            raise RuntimeError(
                f"{ABA_RESULTADOS}!B{linha} sem o filtro de PC esperado."
            )
        res.Range(f"B{linha}").Formula = atual.replace(
            FILTRO_PC_ANTIGO, FILTRO_PC_NOVO
        )


def _aplicar_totais_canonicos(wb) -> None:
    """Publica as doze medidas com nomes claros, sem criar aba ou painel."""
    res = wb.Worksheets(ABA_RESULTADOS)
    linha = LINHA_SECAO_TOTAIS
    res.Range(f"A{linha}").Value = TITULO_SECAO_TOTAIS
    res.Range(f"A{linha}").Font.Bold = True
    res.Range(f"A{linha + 1}").Value = "Medida"
    res.Range(f"B{linha + 1}").Value = "Valor"
    res.Range(f"C{linha + 1}").Value = "Composicao auditavel"
    res.Range(f"A{linha + 1}:C{linha + 1}").Font.Bold = True
    for i, (rotulo, formula, fonte) in enumerate(MEDIDAS_CANONICAS, start=2):
        alvo = linha + i
        res.Range(f"A{alvo}").Value = f"{i - 1}. {rotulo}"
        res.Range(f"B{alvo}").Formula = formula
        res.Range(f"C{alvo}").Value = fonte


def _aplicar_guarda_aditivos(wb) -> None:
    mem = wb.Worksheets(ABA_MEMORIA)
    formula = str(mem.Range("E26").Formula or "")
    if GUARDA_ADITIVOS_NOVA in formula:
        return
    if GUARDA_ADITIVOS_ANTIGA not in formula:
        raise RuntimeError(
            f"{ABA_MEMORIA}!E26 sem a guarda {GUARDA_ADITIVOS_ANTIGA!r}; "
            "origem incompativel."
        )
    mem.Range("E26").Formula = formula.replace(
        GUARDA_ADITIVOS_ANTIGA, GUARDA_ADITIVOS_NOVA, 1
    )
    mem.Range("A27").Value = NOTA_A27


def _verificar_sem_erros(wb) -> None:
    import pywintypes

    problemas = []
    for ws in wb.Worksheets:
        for tipo in (XL_CELLTYPE_FORMULAS, XL_CELLTYPE_CONSTANTS):
            try:
                celulas = ws.UsedRange.SpecialCells(tipo, XL_ERRORS)
                problemas.append(f"{ws.Name}!{celulas.Address}")
            except pywintypes.com_error:
                continue
    if problemas:
        raise RuntimeError(f"Erros de formula apos recalculo: {problemas}")


def _fechar_silencioso(wb) -> None:
    """Fecha o workbook tolerando proxy COM ja desconectado.

    So deve ser usado DEPOIS de um `Save()` bem-sucedido: a desconexao pos-Save
    e um defeito da automacao, nao do arquivo gravado.
    """
    try:
        wb.Close(SaveChanges=False)
    except Exception:
        pass


def _reabrir_conferindo(caminho: Path) -> None:
    """Abre `caminho` numa instancia nova do Excel e exige arquivo integro."""
    pythoncom.CoInitialize()
    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    wb = None
    try:
        wb = excel.Workbooks.Open(
            str(caminho), UpdateLinks=0, ReadOnly=True, CorruptLoad=0
        )
        abas = _nomes_abas(wb)
        for obrig in (ABA_MEMORIA, ABA_RESULTADOS, ABA_ITENS_PC):
            if obrig not in abas:
                raise RuntimeError(f"Aba {obrig} ausente apos reabertura.")
        _verificar_sem_erros(wb)
    finally:
        if wb is not None:
            _fechar_silencioso(wb)
        try:
            excel.Quit()
        except Exception:
            pass
        del wb
        del excel
        pythoncom.CoUninitialize()


def aplicar(origem: Path, destino: Path, somente_estilo_u: bool = False) -> None:
    """Aplica o pacote do valor considerado; com `somente_estilo_u`, so o estilo.

    O modo `somente_estilo_u` existe para reparar arquivos que ja receberam a
    camada de formulas de `U` quando ela ainda nascia sem apresentacao. Ele nao
    escreve nenhuma formula: apenas veste `itens_PC!U`.
    """
    origem = Path(origem).resolve()
    destino = Path(destino).resolve()
    if not origem.is_file():
        raise FileNotFoundError(origem)
    if origem == destino:
        raise ValueError("Origem e destino devem ser diferentes.")
    if not somente_estilo_u:
        _recusar_template_atual_w48(origem)

    tmp_dir = Path(tempfile.mkdtemp(prefix="cl8us_valor_considerado_"))
    tmp_xlsx = tmp_dir / origem.name
    shutil.copyfile(origem, tmp_xlsx)

    pythoncom.CoInitialize()
    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    wb = None
    salvo = False
    try:
        wb = excel.Workbooks.Open(
            str(tmp_xlsx), UpdateLinks=0, ReadOnly=False, CorruptLoad=0
        )
        excel.ScreenUpdating = False
        excel.Calculation = XL_CALC_MANUAL
        aba_ativa = wb.ActiveSheet.Name
        _validar_origem(wb, escreve_w48=not somente_estilo_u)
        travas_antes = _snapshot_travas(wb)

        if somente_estilo_u:
            _aplicar_estilo_u(wb)
        else:
            _aplicar_memoria(wb)
            _aplicar_itens_pc(wb)
            _aplicar_resultados(wb)
            _aplicar_corte(wb)
            _aplicar_totais_canonicos(wb)
            _aplicar_guarda_aditivos(wb)

        travas_depois = _snapshot_travas(wb)
        if travas_antes != travas_depois:
            difs = {k: (travas_antes[k], travas_depois[k])
                    for k in travas_antes if travas_antes[k] != travas_depois[k]}
            raise RuntimeError(f"TRAVA VIOLADA: {difs}")

        excel.Calculation = XL_CALC_AUTOMATIC
        excel.CalculateFullRebuild()
        _verificar_sem_erros(wb)

        if aba_ativa in _nomes_abas(wb):
            wb.Worksheets(aba_ativa).Activate()
        wb.Save()
        salvo = True
        # Em arquivos grandes esta instancia costuma se desconectar logo apos o
        # Save ("O objeto chamado foi desconectado de seus clientes"). O arquivo
        # ja esta gravado; a prova de zero-corrupcao e feita adiante, numa
        # instancia NOVA do Excel, que e um teste mais forte que reusar esta.
        _fechar_silencioso(wb)
        wb = None
    finally:
        if wb is not None:
            _fechar_silencioso(wb)
        try:
            excel.Quit()
        except Exception:
            pass
        del wb
        del excel
        pythoncom.CoUninitialize()

    if not salvo:
        shutil.rmtree(tmp_dir, ignore_errors=True)
        raise RuntimeError("Excel nao salvou; destino preservado.")

    # Prova de zero-corrupcao: uma instancia NOVA precisa abrir o arquivo sem
    # reparo, com as abas obrigatorias e sem erro de formula. Se falhar, o
    # destino nao e tocado.
    _reabrir_conferindo(tmp_xlsx)

    destino.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(tmp_xlsx, destino)
    shutil.rmtree(tmp_dir, ignore_errors=True)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("origem", type=Path)
    parser.add_argument("destino", type=Path)
    parser.add_argument(
        "--somente-estilo-u",
        action="store_true",
        help=(
            "Aplica apenas a apresentacao de itens_PC!U (cabecalho, fonte, "
            "preenchimento, bordas, alinhamento, moeda, protecao e largura), "
            "sem escrever nenhuma formula."
        ),
    )
    args = parser.parse_args()
    aplicar(args.origem, args.destino, somente_estilo_u=args.somente_estilo_u)
    if args.somente_estilo_u:
        print("Estilo de itens_PC!U aplicado:", args.destino)
    else:
        print("Valor considerado aplicado ao VTA:", args.destino)


if __name__ == "__main__":
    main()
