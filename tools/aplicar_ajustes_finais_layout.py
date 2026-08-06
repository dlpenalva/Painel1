# -*- coding: utf-8 -*-
"""Ajustes finais de layout (Etapa VTA) — legibilidade e limpeza estrutural.

Aplica, via Excel COM (padrao zero-corrupcao), dois ajustes de APRESENTACAO que
NAO alteram nenhum calculo, nome definido, formula-fonte ou o VTA oficial:

1. RESULTADOS!Tabela 1 (linhas 8:13 — as tres referencias do VTA):
   * habilita quebra de texto (wrap) nos rotulos, composicoes, fontes e situacao;
   * alarga a coluna H (Situacao), que estava na largura padrao e cortava o
     texto ("DISPONIVEL PARA CONFERENCIA", "ADOTADO - ULTIMA ABERTURA...");
   * ajusta a altura das linhas ao conteudo (as celulas ja tinham wrap ligado
     mas altura fixa de 20 pt, o que cortava a composicao auditavel multilinha).
   Larguras e alturas recebem limites minimos/maximos razoaveis (clamp) para
   nao criar colunas desproporcionais. Excel AutoFit e usado como base para as
   colunas nao mescladas; a composicao mesclada (C:E) recebe altura calculada.

2. cobertura_temporal: remove a linha "Fonte temporal de conferencia"
   (diagnostico secundario, redundante com "Fonte temporal principal" e sem
   consumidor — nem formula, nem leitor por rotulo, nem documento, nem teste).
   A remocao e feita por Rows(n).Delete: o Excel reajusta automaticamente as
   referencias A1 das linhas seguintes; o leitor le por ROTULO (prefixo), nao
   por numero de linha, portanto nada quebra.

Travas: B26/T25 (formula) verificadas antes/depois; recalculo total sem erros;
reabertura sem reparo.
"""
from __future__ import annotations

import argparse
import math
import shutil
import tempfile
from pathlib import Path

import pythoncom
import win32com.client

XL_CALC_MANUAL = -4135
XL_CALC_AUTOMATIC = -4105
XL_CELLTYPE_FORMULAS = -4123
XL_CELLTYPE_CONSTANTS = 2
XL_ERRORS = 16
XL_SHIFT_UP = -4162

ABA_RES = "RESULTADOS"
ABA_COB = "cobertura_temporal"
ABA_REM = "itens_Remanesc"

ROTULO_CONFERENCIA = "fonte temporal de conferencia"

# Colunas manuais de QTD_REM_BASE por ciclo em itens_Remanesc e o indice do
# ciclo de abertura correspondente (usado contra posicao_contratual!Y).
COLS_ABERTURA = (("E", 1), ("G", 2), ("I", 3), ("K", 4))
LINHA_INI_REM = 2
LINHA_FIM_REM = 201

# Clamps de largura de coluna (unidade Excel).
LARG_MIN = 9.0
LARG_MAX = 60.0
# Coluna H (Situacao) precisa caber textos de estado longos.
LARG_H_ALVO = 24.0
# Clamps de altura de linha (pt).
ALT_MIN = 18.0
ALT_MAX = 96.0


def _linhas_para_texto(texto: str, largura_total: float) -> int:
    """Estima o numero de linhas de wrap para um texto numa largura Excel."""
    if not texto:
        return 1
    cpl = max(8.0, largura_total - 1.0)
    return max(1, int(math.ceil(len(texto) / cpl)))


def _nomes_abas(wb) -> list[str]:
    return [ws.Name for ws in wb.Worksheets]


def _verificar_sem_erros(wb) -> None:
    import pywintypes

    problemas = []
    for ws in wb.Worksheets:
        for tipo in (XL_CELLTYPE_FORMULAS, XL_CELLTYPE_CONSTANTS):
            try:
                celulas = ws.UsedRange.SpecialCells(tipo, XL_ERRORS)
                problemas.append(f"{ws.Name}!{celulas.Address}")
            except pywintypes.com_error:
                continue
    if problemas:
        raise RuntimeError(f"Erros de formula apos recalculo: {problemas}")


def _snapshot_travas(wb) -> dict[str, str]:
    mem = wb.Worksheets("MEMORIA_RESULTADOS")
    return {e: str(mem.Range(e).Formula) for e in ("B26", "T25")}


def _clamp(v: float, lo: float, hi: float) -> float:
    return max(lo, min(hi, v))


def _ajustar_tabela1(ws) -> None:
    """Legibilidade das linhas 8:13 (tres referencias do VTA)."""
    if bool(ws.ProtectContents):
        ws.Unprotect()

    # 1) Quebra de texto em toda a area da Tabela 1 (rotulos, valores,
    #    composicao mesclada, fontes mescladas e situacao).
    area = ws.Range("A8:H13")
    area.WrapText = True
    area.VerticalAlignment = -4108  # xlCenter

    # 2) Coluna H (Situacao) — alargar com clamp; colunas B/H passam por clamp
    #    de base (as demais ja tem largura adequada e sao mantidas).
    col_h = ws.Columns("H")
    if (col_h.ColumnWidth or 0) < LARG_H_ALVO:
        col_h.ColumnWidth = LARG_H_ALVO
    for letra in ("B", "H"):
        col = ws.Columns(letra)
        col.ColumnWidth = _clamp(float(col.ColumnWidth or 0), LARG_MIN, LARG_MAX)

    # 3) Altura das linhas ajustada ao conteudo mais longo de cada linha.
    #    A composicao (C:E mescladas) e o texto mais longo; AutoFit ignora
    #    celulas mescladas, entao a altura e calculada a partir do texto.
    larg_ce = sum(float(ws.Columns(c).ColumnWidth or 0) for c in ("C", "D", "E"))
    larg_fg = sum(float(ws.Columns(c).ColumnWidth or 0) for c in ("F", "G"))
    larg_h = float(ws.Columns("H").ColumnWidth or 0)
    larg_a = float(ws.Columns("A").ColumnWidth or 0)

    def altura_linha(r: int) -> float:
        def txt(cel: str) -> str:
            v = ws.Range(cel).Text
            return str(v) if v is not None else ""

        linhas = max(
            _linhas_para_texto(txt(f"A{r}"), larg_a),
            _linhas_para_texto(txt(f"C{r}"), larg_ce),
            _linhas_para_texto(txt(f"F{r}"), larg_fg),
            _linhas_para_texto(txt(f"H{r}"), larg_h),
        )
        return _clamp(linhas * 15.0 + 6.0, ALT_MIN, ALT_MAX)

    ws.Range("A8").RowHeight = _clamp(24.0, ALT_MIN, ALT_MAX)   # titulo
    ws.Range("A9").RowHeight = _clamp(30.0, ALT_MIN, ALT_MAX)   # cabecalho
    for r in (10, 11, 12, 13):
        # piso de 3 linhas nas linhas de dados: a composicao preenchida (com a
        # aba fiscal presente) e mais longa que a versao do template vazio.
        alt = max(altura_linha(r), 3 * 15.0 + 6.0)
        ws.Range(f"A{r}").RowHeight = _clamp(alt, ALT_MIN, ALT_MAX)


def _remover_conferencia(ws) -> int:
    """Remove a linha 'Fonte temporal de conferencia'. Retorna a linha removida."""
    if bool(ws.ProtectContents):
        ws.Unprotect()
    alvo = None
    ultima = int(ws.UsedRange.Row + ws.UsedRange.Rows.Count - 1)
    for r in range(1, ultima + 1):
        v = ws.Range(f"A{r}").Value
        if isinstance(v, str) and v.strip().lower() == ROTULO_CONFERENCIA:
            alvo = r
            break
    if alvo is None:
        return 0  # idempotente: ja removida em aplicacao anterior
    ws.Rows(alvo).Delete(Shift=XL_SHIFT_UP)
    return alvo


def _marcar_estados_remanesc(caminho: Path) -> None:
    """Regra 4: distingue visualmente os 4 estados do remanescente de abertura.

    Via openpyxl (formatacao condicional OOXML padrao — a linhagem do template e
    openpyxl; validado depois por reabertura COM sem reparo). Sobre as colunas
    manuais QTD_REM_BASE_Cn (E/G/I/K), por item (linha alinhada a
    posicao_contratual!Y = CICLO_NASCIMENTO):
      * NAO APLICAVEL (cinza) — item nasce depois do ciclo (Y>indice); stopIfTrue
        para nunca colorir como pendente uma abertura anterior ao nascimento;
      * saldo positivo (verde) — numero > 0;
      * saldo confirmado em zero (azul) — igual a 0;
      * aplicavel e pendente (ambar) — vazio e o item ja existia no ciclo.
    Apenas cor; nao altera valor nem formula. Idempotente (recria as regras).
    """
    from openpyxl import load_workbook
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import PatternFill

    wb = load_workbook(caminho)
    ws = wb[ABA_REM]
    f_na = PatternFill("solid", fgColor="D9D9D9")
    f_pos = PatternFill("solid", fgColor="C6EFCE")
    f_zero = PatternFill("solid", fgColor="DDEBF7")
    f_pend = PatternFill("solid", fgColor="FFF2CC")
    for col, idx in COLS_ABERTURA:
        faixa = f"{col}{LINHA_INI_REM}:{col}{LINHA_FIM_REM}"
        topo = LINHA_INI_REM
        a = f"$A{topo}"
        y = f"posicao_contratual!$Y{topo}"
        cel = f"{col}{topo}"
        # NAO APLICAVEL primeiro (maior prioridade), com stopIfTrue.
        ws.conditional_formatting.add(
            faixa,
            FormulaRule(
                formula=[f'AND({a}<>"",ISNUMBER({y}),{y}>{idx})'],
                fill=f_na, stopIfTrue=True,
            ),
        )
        ws.conditional_formatting.add(
            faixa,
            FormulaRule(formula=[f'AND({a}<>"",ISNUMBER({cel}),{cel}>0)'], fill=f_pos),
        )
        ws.conditional_formatting.add(
            faixa,
            FormulaRule(formula=[f'AND({a}<>"",{cel}=0)'], fill=f_zero),
        )
        ws.conditional_formatting.add(
            faixa,
            FormulaRule(
                formula=[f'AND({a}<>"",{cel}="",OR(NOT(ISNUMBER({y})),{y}<={idx}))'],
                fill=f_pend,
            ),
        )
    wb.save(caminho)


def aplicar(origem: Path, destino: Path) -> None:
    origem = Path(origem).resolve()
    destino = Path(destino).resolve()
    if not origem.is_file():
        raise FileNotFoundError(origem)
    if origem == destino:
        raise ValueError("Origem e destino devem ser diferentes.")

    tmp_dir = Path(tempfile.mkdtemp(prefix="cl8us_layout_"))
    tmp_xlsx = tmp_dir / origem.name
    shutil.copyfile(origem, tmp_xlsx)

    pythoncom.CoInitialize()
    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    wb = None
    salvo = False
    try:
        wb = excel.Workbooks.Open(
            str(tmp_xlsx), UpdateLinks=0, ReadOnly=False, CorruptLoad=0
        )
        excel.ScreenUpdating = False
        excel.Calculation = XL_CALC_MANUAL
        aba_ativa = wb.ActiveSheet.Name
        for aba in (ABA_RES, ABA_COB, ABA_REM):
            if aba not in _nomes_abas(wb):
                raise ValueError(f"Aba {aba} ausente; layout inesperado.")
        travas_antes = _snapshot_travas(wb)

        # Recalcula primeiro para que .Text reflita as formulas.
        excel.Calculation = XL_CALC_AUTOMATIC
        excel.CalculateFullRebuild()

        _ajustar_tabela1(wb.Worksheets(ABA_RES))
        removida = _remover_conferencia(wb.Worksheets(ABA_COB))

        travas_depois = _snapshot_travas(wb)
        if travas_antes != travas_depois:
            raise RuntimeError(
                f"TRAVA VIOLADA B26/T25: {travas_antes} -> {travas_depois}"
            )

        excel.CalculateFullRebuild()
        _verificar_sem_erros(wb)
        if aba_ativa in _nomes_abas(wb):
            wb.Worksheets(aba_ativa).Activate()
        wb.Save()
        salvo = True
        wb.Close(SaveChanges=False)
        wb = None

        # Regra 4 (cor dos estados) como ULTIMA escrita, via openpyxl: mantem a
        # formatacao condicional como CF padrao (sem extensao x14 do Excel), de
        # modo que os Coletas gerados pelo app (linhagem openpyxl) a preservem.
        _marcar_estados_remanesc(tmp_xlsx)

        # Reabre sem reparo (somente leitura: nao reescreve, preserva a CF) e
        # confere sentinelas.
        wb = excel.Workbooks.Open(
            str(tmp_xlsx), UpdateLinks=0, ReadOnly=True, CorruptLoad=0
        )
        cob = wb.Worksheets(ABA_COB)
        for r in range(1, int(cob.UsedRange.Rows.Count) + 2):
            v = cob.Range(f"A{r}").Value
            if isinstance(v, str) and v.strip().lower() == ROTULO_CONFERENCIA:
                raise RuntimeError("Conferencia persistiu apos remocao.")
        res = wb.Worksheets(ABA_RES)
        if float(res.Columns("H").ColumnWidth or 0) < LARG_H_ALVO - 0.5:
            raise RuntimeError("Coluna H nao persistiu o alargamento.")
        wb.Close(SaveChanges=False)
        wb = None
        print(f"linha conferencia removida: {removida}")
    finally:
        if wb is not None:
            wb.Close(SaveChanges=False)
        excel.Quit()
        del wb
        del excel
        pythoncom.CoUninitialize()

    if not salvo:
        shutil.rmtree(tmp_dir, ignore_errors=True)
        raise RuntimeError("Excel nao salvou; destino preservado.")
    destino.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(tmp_xlsx, destino)
    shutil.rmtree(tmp_dir, ignore_errors=True)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("origem", type=Path)
    parser.add_argument("destino", type=Path)
    args = parser.parse_args()
    aplicar(args.origem, args.destino)
    print("ajustes finais de layout aplicados:", args.destino)


if __name__ == "__main__":
    main()
