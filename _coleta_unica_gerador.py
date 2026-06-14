"""
_coleta_unica_gerador.py
------------------------
Gerador da Coleta Mestre v10 (cl8us v2.0).

Produz o XLSX ColetaMestre com os dados da Calculadora de Reajustes
já pré-preenchidos nas abas PARAMETROS_CONTRATO e CICLOS.

Uso:
    from _coleta_unica_gerador import gerar_coleta_unica_inteligente
    bytes_xlsx = gerar_coleta_unica_inteligente()

Compatível com o leitor _leitor_coleta_mestre.py.
"""

from io import BytesIO
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ── Paleta ──────────────────────────────────────────────────────
_C = {
    "calc_h":"1E3A8A","calc_b":"DBEAFE","calc_f":"1E3A8A",
    "gcc_h": "7C3AED","gcc_b": "EDE9FE","gcc_f": "4C1D95",
    "fisc_h":"065F46","fisc_b":"DCFCE7","fisc_f":"065F46",
    "auto_h":"374151","auto_b":"F1F5F9","auto_f":"475569",
    "ajst_h":"B45309","ajst_b":"FDE68A","ajst_f":"78350F",
    "hist_b":"E5E7EB","hist_h":"4B5563","hist_f":"1F2937",
    "alrt_b":"FED7AA","alrt_f":"7C2D12",
    "res_h": "065F46",
}
_MOEDA = 'R$ #,##0.00'
_PERC  = '0.00%'
_FATOR = '0.000000'
_QTD   = '#,##0.##'


def _fill(h): return PatternFill("solid", fgColor=h)
def _fonte(bold=False, color="0F172A", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Arial")
def _aln(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def _brd(cor="CBD5E1"):
    s = Side(style="thin", color=cor)
    return Border(left=s, right=s, top=s, bottom=s)
def _brd_m(cor="9CA3AF"):
    s = Side(style="medium", color=cor)
    return Border(left=s, right=s, top=s, bottom=s)

def _cel(ws, r, c, val="", bg="FFFFFF", fg="0F172A", bold=False,
         fmt=None, h="center", wrap=False, size=10, italic=False, brd_fn=None):
    cc = ws.cell(r, c, val)
    cc.fill = _fill(bg)
    cc.font = _fonte(bold, fg, size, italic)
    cc.alignment = _aln(h, "center", wrap)
    cc.border = (brd_fn or _brd)()
    if fmt: cc.number_format = fmt
    return cc

def _auto(ws, r, c, formula, fmt=None):
    cc = ws.cell(r, c, formula)
    cc.fill = _fill(_C["auto_b"]); cc.font = _fonte(color=_C["auto_f"])
    cc.alignment = _aln(); cc.border = _brd()
    if fmt: cc.number_format = fmt
    return cc

def _titulo(ws, row, c1, c2, txt, bg="1E3A8A", size=11):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    c = ws.cell(row, c1, txt)
    c.fill = _fill(bg); c.font = _fonte(True, "FFFFFF", size)
    c.alignment = _aln("left", "center"); c.border = _brd()
    ws.row_dimensions[row].height = 26

def _nota(ws, row, c1, c2, txt, bg=None, fg=None, size=9):
    bg = bg or _C["auto_b"]; fg = fg or _C["auto_f"]
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    c = ws.cell(row, c1, txt)
    c.fill = _fill(bg); c.font = _fonte(italic=True, color=fg, size=size)
    c.alignment = _aln("left", "center", True); c.border = _brd()
    ws.row_dimensions[row].height = 20

def _secao(ws, row, c1, c2, txt, bg, fg="FFFFFF", size=9):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    c = ws.cell(row, c1, f"  {txt}")
    c.fill = _fill(bg); c.font = _fonte(True, fg, size)
    c.alignment = _aln("left", "center"); c.border = _brd()
    ws.row_dimensions[row].height = 20

def _hdr(ws, row, col, txt, bg, w=None):
    c = ws.cell(row, col, txt)
    c.fill = _fill(bg); c.font = _fonte(True, "FFFFFF", 9)
    c.alignment = _aln("center", "center", True); c.border = _brd()
    ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 0, 42)
    if w: ws.column_dimensions[get_column_letter(col)].width = w
    return c

def _dv(ws, formula, ref):
    d = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(d); d.add(ref); return d

def _campo_param(ws, r, label, val, bg_l=None, bg_v=None, fmt=None, dvf=None):
    """Linha de campo em PARAMETROS: col A+B = label, col C = valor."""
    bg_l = bg_l or _C["auto_b"]
    bg_v = bg_v or _C["gcc_b"]
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    cl = ws.cell(r, 1, label)
    cl.fill = _fill(bg_l); cl.font = _fonte(False, _C["auto_f"], 9)
    cl.alignment = _aln("right", "center"); cl.border = _brd()
    ws.cell(r, 2).fill = _fill(bg_l); ws.cell(r, 2).border = _brd()
    vi = ws.cell(r, 3, val)
    vi.fill = _fill(bg_v); vi.font = _fonte(color=_C["gcc_f"])
    vi.alignment = _aln("left", "center"); vi.border = _brd()
    if fmt: vi.number_format = fmt
    ws.cell(r, 4).fill = _fill(bg_v); ws.cell(r, 4).border = _brd()
    ws.row_dimensions[r].height = 22
    if dvf:
        d = DataValidation(type="list", formula1=dvf, allow_blank=True)
        ws.add_data_validation(d); d.add(f"C{r}")
    return vi

def _campo_calc(ws, r, label, val, fmt=None):
    """Campo pré-preenchido pela calculadora — fundo azul."""
    _campo_param(ws, r, label, val, bg_l=_C["calc_b"], bg_v=_C["calc_b"], fmt=fmt)
    ws.cell(r, 3).font = _fonte(color=_C["calc_f"])


# ── Leitura do session_state ─────────────────────────────────────

def _dados_sessao():
    """Extrai dados da Calculadora de Reajustes do session_state."""
    try:
        adm = st.session_state.get('dados_admissibilidade') or {}
    except Exception:
        adm = {}
    ciclos_raw = adm.get('ciclos', []) or []
    return {
        "indice":      str(adm.get('indice', '')),
        "data_base":   str(adm.get('data_base_original', '')),
        "valor_orig":  adm.get('valor_original', 0) or 0,
        "ciclos":      ciclos_raw,
        "tem_dados":   bool(ciclos_raw),
    }




# ── Construtores de cada aba ─────────────────────────────────────

def _construir_parametros(wb, dados):
    ws = wb.create_sheet("PARAMETROS_CONTRATO")
    ws.sheet_view.showGridLines = False; ws.sheet_view.zoomScale = 90
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 2
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 2

    _titulo(ws, 1, 1, 4, "cl8us · Parâmetros do Contrato  (identificação + parâmetros de reajuste)")
    _nota(ws, 2, 1, 4,
          "  Campos AZUIS = pré-preenchidos pela Calculadora de Reajustes. "
          "Campos ROXOS = GCC preenche.")

    # Identificação
    _secao(ws, 3, 1, 4, "Identificação", _C["gcc_h"])
    _campo_param(ws, 4, "Vigência inicial", "")
    _campo_param(ws, 5, "Vigência final",   "")

    # Parâmetros Calculadora
    _secao(ws, 6, 1, 4, "Parâmetros de Reajuste  ← pré-preenchido pela Calculadora", _C["calc_h"])
    _campo_calc(ws, 7,  "Índice contratual",             dados["indice"])
    _campo_calc(ws, 8,  "Data-base original",            dados["data_base"])
    _campo_calc(ws, 9,  "Valor original do contrato",    dados["valor_orig"] or 0, _MOEDA)
    _campo_calc(ws, 10, "Último ciclo já formalizado",   "C0 / Nenhum")

    # Campos calculados — ciclo inicial/final/competência/variação/fator
    ciclo_ini = ""
    ciclo_atu = ""
    variacao  = 0.0
    fator_acum = 1.0
    if dados["ciclos"]:
        primeiro = dados["ciclos"][0]
        ultimo   = dados["ciclos"][-1]
        ciclo_ini  = str(primeiro.get('ciclo', ''))
        ciclo_atu  = str(ultimo.get('ciclo', ''))
        fator_acum = float(ultimo.get('fator_acumulado', 1.0) or 1.0)
        variacao   = round(fator_acum - 1.0, 8)

    _campo_calc(ws, 11, "Ciclo inicial da análise atual", ciclo_ini)
    _campo_calc(ws, 12, "Ciclo atual / ciclo de corte",   ciclo_atu)
    _campo_calc(ws, 13, "Competência de corte",           "")
    _campo_calc(ws, 14, "Variação acumulada apurada",     variacao,  _PERC)
    _campo_calc(ws, 15, "Fator acumulado efetivo",        fator_acum, _FATOR)

    # Controle
    _secao(ws, 16, 1, 4, "Controle  —  GCC preenche", _C["gcc_h"])
    _campo_param(ws, 17, "Há aditivos/supressões?",   "Não", dvf='"Sim,Não"')
    _campo_param(ws, 18, "Há corte operacional?",      "Não", dvf='"Sim,Não"')

    # Separador
    ws.row_dimensions[19].height = 6
    for c in range(1,5): ws.cell(19,c).border = _brd()

    # Modo detectado automaticamente pelo sistema — campo removido do XLS
    # (detecção automática no site substitui a declaração manual)

    # Ciclo em execução
    _secao(ws, 22, 1, 4, "Ciclo em Execução  —  GCC preenche se houver corte operacional", _C["gcc_h"])
    _campo_param(ws, 23, "Há ciclo em execução com dados parciais?", "Não", dvf='"Sim,Não"')
    _campo_param(ws, 24, "Se Sim — o fiscal informou:", "",
        dvf='"a. Só financeiro até o corte,b. Financeiro + saldo remanescente,c. Só saldo remanescente atual"')
    _campo_param(ws, 25, "Competência de corte", "")
    _campo_param(ws, 26, "C0 executado manual",                   0, fmt=_MOEDA)
    _campo_param(ws, 27, "Saldo remanescente ORIGINAL no corte",  0, fmt=_MOEDA)
    _campo_param(ws, 28, "Saldo remanescente ATUALIZADO no corte",0, fmt=_MOEDA)
    _campo_param(ws, 29, "O saldo inclui aditivos?", "Não", dvf='"Sim,Não"')


def _construir_ciclos(wb, dados):
    ws = wb.create_sheet("CICLOS")
    ws.sheet_view.showGridLines = False; ws.sheet_view.zoomScale = 90

    _titulo(ws, 1, 1, 12, "cl8us  ·  Ciclos Contratuais  ← pré-preenchido pela Calculadora de Reajustes")
    _nota(ws, 2, 1, 12,
          "  Campos CINZA = gerados pela Calculadora e protegidos. "
          "Colunas roxas = GCC confirma conforme o contrato.")

    # Cabeçalho duplo (linha 4 = nomes limpos, linha 5 = tags)
    hdrs_clean = ["Ciclo","Data-base","Início\nfinanceiro","Fim\nfinanceiro",
                  "Percentual\naplicado","Fator\ndo ciclo","Fator acumulado\nefetivo",
                  "Situação","É objeto da\nAnálise atual?","Tem efeito\nfinanceiro?",
                  "Entra no\nValor Total?","Observação"]
    hdrs_tags  = ["[CALC]\nCiclo","[CALC]\nData-base","[CALC]\nInício fin.",
                  "[CALC]\nFim fin.","[CALC]\nPercentual","[CALC]\nFator",
                  "[CALC]\nFator acum.\nefetivo","[CALC]\nSituação",
                  "[GCC]\nÉ objeto\nAnálise?","[GCC]\nTem efeito\nfinanceiro?",
                  "[GCC]\nEntra no\nValor Total?","[AUTO]\nObservação"]
    widths = [8,14,14,14,14,13,16,18,16,16,14,34]

    ws.row_dimensions[4].height = 38
    ws.row_dimensions[5].height = 52
    for ci, (tc, tt, w) in enumerate(zip(hdrs_clean, hdrs_tags, widths), start=1):
        _hdr(ws, 4, ci, tc, _C["calc_h"], w)
        bg = _C["gcc_h"] if ci in [9,10,11] else (_C["auto_h"] if ci==12 else _C["calc_h"])
        _hdr(ws, 5, ci, tt, bg)

    # C0 sempre presente
    ws.row_dimensions[6].height = 22
    c0_dados = [
        "C0", dados["data_base"], "", "",
        0.0, 1.0, 1.0,
        "Base sem reajuste", "Não", "Não", "Sim",
        "C0 não reajusta, mas integra o Valor Total com fator 1,0000."
    ]
    for ci, val in enumerate(c0_dados, start=1):
        fmt = _PERC if ci==5 else (_FATOR if ci in [6,7] else None)
        if ci <= 8:
            _cel(ws, 6, ci, val, _C["calc_b"], _C["calc_f"], fmt=fmt)
        elif ci <= 11:
            _cel(ws, 6, ci, val, _C["gcc_b"], _C["gcc_f"])
        else:
            _cel(ws, 6, ci, val, _C["auto_b"], _C["auto_f"], italic=True, h="left", wrap=True, size=9)

    # Ciclos calculados
    for i, ciclo in enumerate(dados["ciclos"]):
        r = 7 + i
        ws.row_dimensions[r].height = 22
        nome   = str(ciclo.get('ciclo', f'C{i+1}'))
        db     = str(ciclo.get('data_base_original', ciclo.get('data_base', '')))
        ini_f  = str(ciclo.get('inicio_financeiro', ciclo.get('data_inicio_financeiro', '')))
        fim_f  = str(ciclo.get('fim_financeiro', ciclo.get('data_fim_financeiro', '')))
        pct    = float(ciclo.get('percentual_apurado', ciclo.get('percentual', 0)) or 0)
        fat    = float(ciclo.get('fator_ciclo', ciclo.get('fator', 1)) or 1)
        fat_ac = float(ciclo.get('fator_acumulado', 1) or 1)
        sit    = str(ciclo.get('situacao_aplicada', ciclo.get('situacao', '')) or '')

        linha = [nome, db, ini_f, fim_f, pct, fat, fat_ac, sit, "Sim", "Sim", "Sim", ""]
        for ci, val in enumerate(linha, start=1):
            fmt = _PERC if ci==5 else (_FATOR if ci in [6,7] else None)
            if ci <= 8:
                _cel(ws, r, ci, val, _C["calc_b"], _C["calc_f"], fmt=fmt)
            elif ci <= 11:
                _cel(ws, r, ci, val, _C["gcc_b"], _C["gcc_f"])
            else:
                _cel(ws, r, ci, val, _C["auto_b"], _C["auto_f"], italic=True, h="left", size=9)

    _dv(ws, '"Tempestivo,Precluso,Adiantado,Admissível com ressalva,Base sem reajuste"', "H6:H11")
    _dv(ws, '"Sim,Não"', "I6:K11")


def _construir_financeiro(wb, dados):
    ws = wb.create_sheet("FINANCEIRO")
    ws.sheet_view.showGridLines = False; ws.sheet_view.zoomScale = 90

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 36

    _titulo(ws, 1, 1, 5, "cl8us  ·  Financeiro — Valores Informados por Competência")

    # Legenda
    for ini,fim,txt,bg,fg in [
        (1,1,"  ■ [AUTO] Cinza",_C["auto_b"],_C["auto_h"]),
        (2,3,"  ■ [FISCAL] Verde — fiscal preenche",_C["fisc_b"],_C["fisc_h"]),
        (4,4,"  ■ [GCC] Roxo",_C["gcc_b"],_C["gcc_h"]),
        (5,5,"  ■ [AJUSTE] Âmbar",_C["ajst_b"],_C["ajst_h"]),
    ]:
        if ini<fim: ws.merge_cells(start_row=2,start_column=ini,end_row=2,end_column=fim)
        c=ws.cell(2,ini,txt); c.fill=_fill(bg); c.font=_fonte(True,fg,8)
        c.alignment=_aln("left","center"); c.border=_brd()
    ws.row_dimensions[2].height=18

    _nota(ws,3,1,5,
        "  FISCAL: preencha SOMENTE col B (competência) e col C (valor informado). "
        "C0 e C1 são linhas fixas consolidadas — GCC informa o total.",
        _C["fisc_b"], _C["fisc_h"], size=9)
    ws.row_dimensions[3].height=22

    # Cabeçalho
    ws.row_dimensions[4].height=38
    _hdr(ws,4,1,"[AUTO]\nCiclo",        _C["auto_h"],20)
    _hdr(ws,4,2,"[FISCAL]\nCompetência",_C["fisc_h"],16)
    _hdr(ws,4,3,"[FISCAL]\nValor Informado",_C["fisc_h"],24)
    _hdr(ws,4,4,"[GCC]\nTem efeito\nfinanceiro?",_C["gcc_h"],22)
    _hdr(ws,4,5,"[AJUSTE]\nNota técnica / Obs. GCC",_C["ajst_h"],36)

    # Linhas consolidadas dinâmicas — C0 sempre; demais se não forem objeto da análise
    # Cor âmbar para ciclo precluso
    _COR_PRECLUSO_B = "FFF3CD"  # fundo âmbar claro
    _COR_PRECLUSO_F = "7D4E00"  # texto âmbar escuro

    _nao_objeto = [{"ciclo": "C0", "label": "C0 — Executado consolidado (pré-reajuste)",
                    "precluso": False, "valor_estimado": 0}]
    for _c in dados.get("ciclos", []):
        _nome_c = str(_c.get("ciclo", "")).strip().upper()
        if _nome_c == "C0":
            continue
        # Campos do calculador (02/01) usam objeto_analise_atual e ciclo_ja_concedido
        # Campos do leitor ColetaMestre usam objeto_analise e ja_formalizado
        _obj_atual = _c.get("objeto_analise_atual")  # bool ou None
        _obj_str   = str(_c.get("objeto_analise", "sim")).strip().lower()
        _ja        = _c.get("ciclo_ja_concedido", False)
        _ja_str    = str(_c.get("ja_formalizado", _c.get("ja_concedido", "não"))).strip().lower()
        _sit       = str(_c.get("situacao_aplicada", _c.get("situacao", ""))).strip().lower()

        # É não-objeto se: objeto_analise_atual=False, ou ciclo_ja_concedido=True,
        # ou campo string indicar não/nao, ou já formalizado
        if _obj_atual is not None:
            _eh_nao_objeto = not bool(_obj_atual) or bool(_ja)
        else:
            _eh_nao_objeto = (_obj_str in ("não", "nao", "false", "0")
                             or _ja_str in ("sim", "true", "1", "yes"))
        _eh_precluso = "precluso" in _sit or "preclus" in _sit or "preclu" in _sit
        if _eh_nao_objeto:
            # Estimar valor: valor_original × fator_acumulado_do_ciclo
            _fat_c = float(_c.get("fator_acumulado", 1) or 1)
            _val_orig_c = float(dados.get("valor_orig", 0) or 0)
            _val_est = 0  # fiscal informa — estimativa apenas na nota
            _nao_objeto.append({
                "ciclo":           _nome_c,
                "label":           f"{_nome_c} — Executado consolidado (não objeto da análise atual)",
                "precluso":        _eh_precluso,
                "valor_estimado":  _val_est,
                "fator":           _fat_c,
            })

    for _idx, _cn in enumerate(_nao_objeto):
        _r = 5 + _idx
        ws.row_dimensions[_r].height = 24
        _precluso = _cn.get("precluso", False)
        # Cores: precluso = âmbar, normal = cinza histórico
        _bg_ciclo = _COR_PRECLUSO_B if _precluso else _C["hist_b"]
        _fg_ciclo = _COR_PRECLUSO_F if _precluso else _C["hist_h"]
        _fg_label = _COR_PRECLUSO_F if _precluso else _C["hist_f"]
        _nota_e   = ("Precluso. Informe o valor total executado neste ciclo. "
                     "GCC mantém efeito = Não para fins de retroativo.") if _precluso else ""
        _cel(ws,_r,1,_cn["ciclo"],_bg_ciclo,_fg_ciclo,True,brd_fn=_brd_m)
        _cel(ws,_r,2,_cn["label"],_bg_ciclo,_fg_label,italic=True,h="left",brd_fn=_brd_m)
        _cel(ws,_r,3,_cn.get("valor_estimado",0),_C["gcc_b"],_C["gcc_f"],fmt=_MOEDA,brd_fn=_brd_m)
        _cel(ws,_r,4,"Não",_C["gcc_b"],_C["gcc_f"],brd_fn=_brd_m)
        _cel(ws,_r,5,_nota_e,_C["ajst_b"],_C["ajst_f"],italic=True,h="left",size=9,brd_fn=_brd_m)

    # Separador dinâmico
    _r_sep = 5 + len(_nao_objeto)
    ws.row_dimensions[_r_sep].height = 10
    ws.merge_cells(f"A{_r_sep}:E{_r_sep}")
    c = ws.cell(_r_sep, 1, "  ▼  Linhas mensais — ciclos objeto da análise de reajuste")
    c.fill=_fill(_C["calc_h"]); c.font=_fonte(True,"FFFFFF",8)
    c.alignment=_aln("left","center"); c.border=_brd()

    # Ciclos objeto: apenas os que NÃO foram para o bloco consolidado
    _nomes_nao_objeto = {_cn["ciclo"] for _cn in _nao_objeto}
    ciclos = [
        _c for _c in dados.get("ciclos", [])
        if str(_c.get("ciclo", "")).strip().upper() not in _nomes_nao_objeto
    ] or dados.get("ciclos", [])

    r = _r_sep + 1
    if ciclos:
        for ciclo in ciclos:
            nome = str(ciclo.get('ciclo', ''))
            ini  = str(ciclo.get('inicio_financeiro', ciclo.get('data_inicio_financeiro','')) or '')
            fim  = str(ciclo.get('fim_financeiro',    ciclo.get('data_fim_financeiro',''))    or '')
            # Estimar número de meses (aprox 12 por ciclo) — linhas em branco com ciclo pré-preenchido
            for _ in range(12):
                if r > 207: break
                ws.row_dimensions[r].height=18
                # Col A: ciclo pré-preenchido (auto-like — escuro mas editável)
                _cel(ws,r,1,nome,_C["calc_b"],_C["calc_f"])
                _cel(ws,r,2,"",_C["fisc_b"],_C["fisc_f"])  # B: competência fiscal
                _cel(ws,r,3,"",_C["fisc_b"],_C["fisc_f"],fmt=_MOEDA)  # C: valor fiscal
                _cel(ws,r,4,"Sim",_C["gcc_b"],_C["gcc_f"])  # D: efeito gcc
                _cel(ws,r,5,"",_C["ajst_b"],_C["ajst_f"],italic=True,h="left",size=9)
                r += 1
    # Preencher linhas restantes em branco
    while r <= 207:
        ws.row_dimensions[r].height=18
        _auto(ws,r,1,"")  # ciclo vazio
        _cel(ws,r,2,"",_C["fisc_b"],_C["fisc_f"])
        _cel(ws,r,3,"",_C["fisc_b"],_C["fisc_f"],fmt=_MOEDA)
        _cel(ws,r,4,"Sim",_C["gcc_b"],_C["gcc_f"])
        _cel(ws,r,5,"",_C["ajst_b"],_C["ajst_f"],italic=True,h="left",size=9)
        r += 1

    # Total
    ws.row_dimensions[208].height=22
    ws.merge_cells("A208:B208")
    c=ws.cell(208,1,"TOTAL — somente linhas com efeito financeiro (Sim)")
    c.fill=_fill(_C["auto_b"]); c.font=_fonte(True,_C["auto_h"],9)
    c.alignment=_aln("right","center"); c.border=_brd()
    ws.cell(208,2).fill=_fill(_C["auto_b"]); ws.cell(208,2).border=_brd()
    t=ws.cell(208,3,
        '=ROUND(SUMIF(D8:D207,"Sim",C8:C207)'
        '+IF(D5="Sim",C5,0)+IF(D6="Sim",C6,0),2)')
    t.fill=_fill(_C["auto_b"]); t.font=_fonte(True,_C["calc_h"],10)
    t.alignment=_aln(); t.border=_brd(); t.number_format=_MOEDA
    for ci in [4,5]:
        ws.cell(208,ci).fill=_fill(_C["auto_b"]); ws.cell(208,ci).border=_brd()

    _dv(ws,'"Sim,Não"',"D5:D207")


def _construir_itens(wb):
    ws = wb.create_sheet("ITENS")
    ws.sheet_view.showGridLines = False; ws.sheet_view.zoomScale = 90

    for l,w in {"A":12,"B":12,"C":16,"D":16,"E":14,"F":14,"G":14,"H":14,"I":14,"J":16,"K":16,"L":12,"M":16,"N":28}.items():
        ws.column_dimensions[l].width = w

    _titulo(ws,1,1,14,"cl8us  ·  Itens do Contrato — Base C0 + Remanescentes por Ciclo")

    # Legenda
    for ini,fim,txt,bg,fg in [
        (1,4,"  ■ [GCC] Roxo — você preenche",_C["gcc_b"],_C["gcc_h"]),
        (5,9,"  ■ [FISCAL] Verde — fiscal preenche somente quantidade remanescente",_C["fisc_b"],_C["fisc_h"]),
        (10,13,"  ■ [AUTO] Cinza — automático",_C["auto_b"],_C["auto_h"]),
        (14,14,"  ■ [AJUSTE] Âmbar",_C["ajst_b"],_C["ajst_h"]),
    ]:
        if ini<fim: ws.merge_cells(start_row=2,start_column=ini,end_row=2,end_column=fim)
        c=ws.cell(2,ini,txt); c.fill=_fill(bg); c.font=_fonte(True,fg,8)
        c.alignment=_aln("left","center"); c.border=_brd()
    ws.row_dimensions[2].height=18

    _nota(ws,3,1,14,
        "  FISCAL: preencha SOMENTE as colunas verdes (quantidade remanescente no início de cada ciclo). NÃO altere as demais colunas.",
        _C["fisc_b"],_C["fisc_h"])
    _nota(ws,4,1,14,
        "  GCC: preencha colunas A (Item), B (Qtd C0), C (VU C0). Col J = remanescente atual/corte. Col L = fator do ciclo atual. Col N = ajuste.",
        _C["gcc_b"],_C["gcc_f"])

    # Linha 5 — datas de início de cada ciclo (AUTO)
    ws.row_dimensions[5].height=20
    for ci in range(1,15):
        if 5 <= ci <= 9:
            ciclo_row = ci + 1  # C1→6, C2→7, C3→8, C4→9, C5→10
            cc=ws.cell(5,ci,f'=IFERROR(TEXT(CICLOS!C{ciclo_row},"MM/AAAA"),"")')
            cc.fill=_fill(_C["auto_b"]); cc.font=_fonte(italic=True,color=_C["auto_f"],size=9)
            cc.alignment=_aln(); cc.border=_brd()
        else:
            _cel(ws,5,ci,"",_C["auto_b"])

    # Linha 6 — grupos
    ws.row_dimensions[6].height=20
    ws.merge_cells("A6:D6"); c=ws.cell(6,1,"Base Contratual C0  —  [GCC] preenche")
    c.fill=_fill(_C["gcc_h"]); c.font=_fonte(True,"FFFFFF",9); c.alignment=_aln(); c.border=_brd()
    ws.merge_cells("E6:I6"); c=ws.cell(6,5,"Remanescentes por Ciclo  —  [FISCAL] preenche (quantidade)")
    c.fill=_fill(_C["fisc_h"]); c.font=_fonte(True,"FFFFFF",9); c.alignment=_aln(); c.border=_brd()
    ws.merge_cells("J6:M6"); c=ws.cell(6,10,"Corte/Atualização  —  [GCC] define + [AUTO] calcula")
    c.fill=_fill(_C["gcc_h"]); c.font=_fonte(True,"FFFFFF",9); c.alignment=_aln(); c.border=_brd()
    _cel(ws,6,14,"[AJUSTE]",_C["ajst_h"],"FFFFFF",True)

    # Linha 7 — cabeçalhos individuais
    ws.row_dimensions[7].height=42
    hdrs=[
        ("[GCC]\nItem",_C["gcc_h"]),("[GCC]\nQtd C0",_C["gcc_h"]),
        ("[GCC]\nVU C0\n(R$)",_C["gcc_h"]),("[AUTO]\nVT C0",_C["auto_h"]),
        ("[FISCAL]\nRem. C1\n(qtd)",_C["fisc_h"]),("[FISCAL]\nRem. C2\n(qtd)",_C["fisc_h"]),
        ("[FISCAL]\nRem. C3\n(qtd)",_C["fisc_h"]),("[FISCAL]\nRem. C4\n(qtd)",_C["fisc_h"]),
        ("[FISCAL]\nRem. C5\n(qtd)",_C["fisc_h"]),("[GCC]\nRem. atual\n(qtd)",_C["gcc_h"]),
        ("[AUTO]\nVT Rem.\natual (R$)",_C["auto_h"]),("[GCC]\nFator\naplicável",_C["gcc_h"]),
        ("[AUTO]\nVT Rem.\natualiz. (R$)",_C["auto_h"]),("[AJUSTE]\nNota técnica",_C["ajst_h"]),
    ]
    for ci,(txt,bg) in enumerate(hdrs,start=1):
        c=ws.cell(7,ci,txt); c.fill=_fill(bg); c.font=_fonte(True,"FFFFFF",9)
        c.alignment=_aln("center","center",True); c.border=_brd()

    # Dados 8-207
    for r in range(8,208):
        ws.row_dimensions[r].height=18
        _cel(ws,r,1,"",_C["gcc_b"],_C["gcc_f"])
        _cel(ws,r,2,0,_C["gcc_b"],_C["gcc_f"],fmt=_QTD)
        _cel(ws,r,3,0,_C["gcc_b"],_C["gcc_f"],fmt=_MOEDA)
        _auto(ws,r,4,f"=IF(OR(B{r}=0,C{r}=0),0,ROUND(B{r}*C{r},2))",_MOEDA)
        for ci in range(5,10): _cel(ws,r,ci,0,_C["fisc_b"],_C["fisc_f"],fmt=_QTD)
        _cel(ws,r,10,0,_C["gcc_b"],_C["gcc_f"],fmt=_QTD)
        _auto(ws,r,11,f"=IF(OR(J{r}=0,C{r}=0),0,ROUND(J{r}*C{r},2))",_MOEDA)
        _cel(ws,r,12,1.0,_C["gcc_b"],_C["gcc_f"],fmt=_FATOR)
        _auto(ws,r,13,f"=IF(OR(K{r}=0,L{r}=0),0,ROUND(K{r}*L{r},2))",_MOEDA)
        _cel(ws,r,14,"",_C["ajst_b"],_C["ajst_f"],italic=True,h="left",size=9)

    # Total
    ws.row_dimensions[208].height=22
    ws.merge_cells("A208:C208")
    c=ws.cell(208,1,"TOTAL"); c.fill=_fill(_C["auto_b"]); c.font=_fonte(True,_C["auto_h"])
    c.alignment=_aln("right","center"); c.border=_brd()
    for ci in [2,3,12,14]:
        ws.cell(208,ci).fill=_fill(_C["auto_b"]); ws.cell(208,ci).border=_brd()
    for ci,col_l in [(4,"D"),(5,"E"),(6,"F"),(7,"G"),(8,"H"),(9,"I"),(10,"J"),(11,"K"),(13,"M")]:
        t=ws.cell(208,ci,f"=ROUND(SUM({col_l}8:{col_l}207),2)")
        t.fill=_fill(_C["auto_b"]); t.font=_fonte(True,_C["auto_h"]); t.alignment=_aln(); t.border=_brd()
        t.number_format=_QTD if ci in [5,6,7,8,9,10] else _MOEDA

    _dv(ws,'"Remanescente,Consumido,Saldo no corte"',"D8:D207")


def _construir_aditivos(wb):
    ws = wb.create_sheet("ADITIVOS")
    ws.sheet_view.showGridLines = False; ws.sheet_view.zoomScale = 90

    _titulo(ws,1,1,11,"cl8us  ·  Aditivos e Supressões")
    for ini,fim,txt,bg,fg in [
        (1,4,"  ■ [GCC] preenche",_C["gcc_b"],_C["gcc_h"]),
        (5,7,"  ■ [AUTO] automático",_C["auto_b"],_C["auto_h"]),
        (8,10,"  ■ [AJUSTE] Âmbar",_C["ajst_b"],_C["ajst_h"]),
        (11,11,"  ■ TRATAMENTO — trava dupla contagem",_C["alrt_b"],_C["alrt_f"]),
    ]:
        if ini<fim: ws.merge_cells(start_row=2,start_column=ini,end_row=2,end_column=fim)
        c=ws.cell(2,ini,txt); c.fill=_fill(bg); c.font=_fonte(True,fg,8)
        c.alignment=_aln("left","center"); c.border=_brd()
    ws.row_dimensions[2].height=18

    _nota(ws,3,1,11,
        "  Informe a data do aditivo — o Ciclo/Marco é preenchido automaticamente pelas datas da aba CICLOS.",
        _C["auto_b"],_C["auto_f"])

    hdrs5=[
        ("[GCC]\nItem",_C["gcc_h"],14),("[GCC]\nData do aditivo",_C["gcc_h"],16),
        ("[AUTO]\nCiclo / Marco\n(pela data)",_C["auto_h"],18),
        ("[GCC]\nTipo de\nalteração",_C["gcc_h"],16),
        ("[GCC]\nQtd acrescida /\nsuprimida",_C["gcc_h"],16),
        ("[GCC]\nValor unitário\noriginal",_C["gcc_h"],18),
        ("[AUTO]\nValor original\nda alteração",_C["auto_h"],18),
        ("[GCC]\nAplicar reajuste\nacumulado?",_C["gcc_h"],16),
        ("[AUTO]\nFator acumulado\naplicável",_C["auto_h"],18),
        ("[AUTO]\nValor atualizado\nda alteração",_C["auto_h"],18),
        ("[TRAT]\nTratamento\ndo aditivo",_C["alrt_b"],26),
    ]
    ws.row_dimensions[5].height=52
    for ci,(txt,bg,w) in enumerate(hdrs5,start=1):
        fg_t = _C["alrt_f"] if ci==11 else "FFFFFF"
        c=ws.cell(5,ci,txt); c.fill=_fill(bg); c.font=_fonte(True,fg_t,9)
        c.alignment=_aln("center","center",True); c.border=_brd()
        ws.column_dimensions[get_column_letter(ci)].width=w

    def f_ciclo(r):
        return (f'=IFERROR(IF(B{r}="","",IFS('
                f'AND(B{r}>=CICLOS!C6,B{r}<=CICLOS!D6),CICLOS!A6,'
                f'AND(B{r}>=CICLOS!C7,B{r}<=CICLOS!D7),CICLOS!A7,'
                f'AND(B{r}>=CICLOS!C8,B{r}<=CICLOS!D8),CICLOS!A8,'
                f'AND(B{r}>=CICLOS!C9,B{r}<=CICLOS!D9),CICLOS!A9,'
                f'AND(B{r}>=CICLOS!C10,B{r}<=CICLOS!D10),CICLOS!A10,'
                f'TRUE,"Verificar")),"")')

    def f_fator(r):
        return (f'=IFERROR(IF(OR(C{r}="","—"),1,'
                f'IF(H{r}="Não",1,VLOOKUP(C{r},CICLOS!A6:G11,7,FALSE))),1)')

    _dv(ws,'"Acréscimo,Supressão"',"D6:D55")
    _dv(ws,'"Sim,Não"',"H6:H55")
    _dv(ws,'"Computar nesta análise,Apenas informativo,Já incorporado ao remanescente,Já incorporado à execução,Já incorporado ao valor formalizado"',"K6:K55")

    for r in range(6,56):
        ws.row_dimensions[r].height=20
        for ci in [1,2,3]: _cel(ws,r,ci,"",_C["gcc_b"],_C["gcc_f"])
        _cel(ws,r,3,f_ciclo(r),_C["auto_b"],_C["auto_f"])  # auto
        _cel(ws,r,4,"Acréscimo",_C["gcc_b"],_C["gcc_f"])
        for ci in [5,6]: _cel(ws,r,ci,0,_C["gcc_b"],_C["gcc_f"],fmt=_MOEDA if ci==6 else _QTD)
        _auto(ws,r,7,f"=IF(OR(E{r}=0,F{r}=0),0,ROUND(E{r}*F{r},2))",_MOEDA)
        _cel(ws,r,8,"Sim",_C["gcc_b"],_C["gcc_f"])
        _auto(ws,r,9,f_fator(r),_FATOR)
        _auto(ws,r,10,f"=IF(G{r}=0,0,ROUND(G{r}*I{r},2))",_MOEDA)
        _cel(ws,r,11,"Computar nesta análise",_C["alrt_b"],_C["alrt_f"])

    ws.row_dimensions[56].height=24
    ws.merge_cells("A56:I56")
    c=ws.cell(56,1,"TOTAL COMPUTÁVEL ATUALIZADO")
    c.fill=_fill(_C["res_h"]); c.font=_fonte(True,"FFFFFF",10)
    c.alignment=_aln("right","center"); c.border=_brd()
    for cc in range(2,10): ws.cell(56,cc).fill=_fill(_C["res_h"]); ws.cell(56,cc).border=_brd()
    t=ws.cell(56,10,'=ROUND(SUMIF(K6:K55,"Computar nesta análise",J6:J55),2)')
    t.fill=_fill(_C["res_h"]); t.font=_fonte(True,"FFFFFF",11); t.alignment=_aln(); t.border=_brd(); t.number_format=_MOEDA
    ws.cell(56,11).fill=_fill(_C["res_h"]); ws.cell(56,11).border=_brd()


def _construir_diagnostico(wb):
    ws = wb.create_sheet("DIAGNOSTICO")
    ws.sheet_view.showGridLines = False; ws.sheet_view.zoomScale = 90

    _titulo(ws,1,1,3,"cl8us  ·  Diagnóstico Automático da Base  ← calculado automaticamente")
    _nota(ws,2,1,3,"  Leitura automática das abas. Mostra o que está disponível, o modo detectado e alertas de suficiência. Consulte antes do upload.",_C["auto_b"],_C["auto_f"])

    ws.column_dimensions["A"].width=50; ws.column_dimensions["B"].width=32; ws.column_dimensions["C"].width=36

    _secao(ws,4,1,3,"Indicadores de Base",_C["calc_h"])

    indicadores=[
        ("Linhas financeiras preenchidas",      '=COUNTIF(FINANCEIRO!C8:C207,">0")',None),
        ("Total financeiro (com efeito Sim)",    '=ROUND(SUMIF(FINANCEIRO!D5:D207,"Sim",FINANCEIRO!C5:C207),2)',_MOEDA),
        ("Itens cadastrados em ITENS",           '=COUNTA(ITENS!A8:A207)-COUNTBLANK(ITENS!A8:A207)',None),
        ("Qtd Remanescentes em ITENS (C1)",     '=SUMPRODUCT((ITENS!E8:E207>0)*1)',None),
        ("Qtd Saldos no corte (col J)",         '=COUNTA(ITENS!J8:J207)-COUNTBLANK(ITENS!J8:J207)',None),
        ("Total rem. atual atualizado",          '=ROUND(SUM(ITENS!M8:M207),2)',_MOEDA),
        ("Aditivos a computar",                  '=COUNTIF(ADITIVOS!K6:K55,"Computar nesta análise")',None),
        ("Valor aditivos computáveis",           '=IFERROR(ADITIVOS!J56,0)',_MOEDA),
        ("Corte operacional declarado",          '=IF(PARAMETROS_CONTRATO!C23="Sim","Sim","Não")',None),
        ("C0 manual informado",                  '=IF(PARAMETROS_CONTRATO!C26>0,"Sim","Não")',None),
        ("C0 no financeiro consolidado",         '=IF(FINANCEIRO!C5>0,"Sim","Não")',None),
        ("Modo declarado por GCC",               '=PARAMETROS_CONTRATO!C21',None),
    ]
    for i,(desc,formula,fmt) in enumerate(indicadores):
        r=5+i; ws.row_dimensions[r].height=20
        _cel(ws,r,1,desc,_C["auto_b"],_C["auto_h"],True,h="right",size=9)
        v=ws.cell(r,2,formula); v.fill=_fill(_C["calc_b"]); v.font=_fonte(True,_C["calc_h"])
        v.alignment=_aln("left","center"); v.border=_brd()
        if fmt: v.number_format=fmt
        ws.cell(r,3).fill=_fill(_C["auto_b"]); ws.cell(r,3).border=_brd()

    _secao(ws,18,1,3,"Modo Detectado",_C["res_h"])
    ws.row_dimensions[19].height=28
    ws.merge_cells("A19:B19")
    modo=ws.cell(19,1,
        '=IF(AND(B5>0,B4>0),"✅ Financeiro + Remanescentes",'
        'IF(AND(B5>0,B5>0),"✅ Financeiro + Saldo no corte",'
        'IF(B5>0,"⚠️ Apenas Financeiro Histórico",'
        'IF(B4>0,"⚠️ Apenas Remanescentes",'
        'IF(B6>0,"⚠️ Apenas Saldo no Corte","❌ Base insuficiente")))))')
    modo.fill=_fill(_C["res_h"]); modo.font=_fonte(True,"FFFFFF",11)
    modo.alignment=_aln("left","center"); modo.border=_brd()
    ws.cell(19,3).fill=_fill(_C["res_h"]); ws.cell(19,3).border=_brd()

    _secao(ws,21,1,3,"Alertas de Consistência",_C["alrt_b"],_C["alrt_f"])
    alertas=[
        ("Risco dupla contagem",
         '=IF(AND(PARAMETROS_CONTRATO!C29="Sim",B8>0),"⚠️ Saldo inclui aditivo E há aditivo para computar","✅ OK")'),
        ("C0 sem cobertura",
         '=IF(AND(B5>0,B10="Não",B11="Não"),"⚠️ Financeiro não cobre C0 — informe C0 manual","✅ OK")'),
        ("Base insuficiente",
         '=IF(B5+B4+B6=0,"❌ Nenhuma base de execução ou saldo identificada","✅ OK")'),
        ("Suficiência declarada",
         '=IF(B12="Sim","✅ Definitivo",IF(B12="Parcial","⚠️ Parcial ou estimativo","❌ Base insuficiente para Valor Total definitivo"))'),
    ]
    for i,(desc,formula) in enumerate(alertas):
        r=22+i; ws.row_dimensions[r].height=24
        _cel(ws,r,1,desc,_C["auto_b"],_C["auto_h"],True,h="right",size=9)
        v=ws.cell(r,2,formula); v.fill=_fill(_C["auto_b"]); v.font=_fonte(True,_C["auto_h"],10)
        v.alignment=_aln("left","center"); v.border=_brd()
        ws.cell(r,3).fill=_fill(_C["auto_b"]); ws.cell(r,3).border=_brd()



# >>> ROTEIRO_INFO_FISCAIS_XLS_V5_PONTUAL
def _construir_roteiro_info_fiscais(wb):
    """Cria aba declaratória com o Roteiro das Informações dos Fiscais.

    A aba é apenas documental/orientativa. Não altera cálculos.
    """
    if "ROTEIRO_INFO_FISCAIS" in wb.sheetnames:
        del wb["ROTEIRO_INFO_FISCAIS"]

    try:
        roteiro = dict(st.session_state.get("roteiro_info_fiscais", {}) or {})
    except Exception:
        roteiro = {}

    def _v(chave, padrao="Não informado"):
        valor = roteiro.get(chave, "")
        texto = "" if valor is None else str(valor).strip()
        return texto if texto else padrao

    ws = wb.create_sheet("ROTEIRO_INFO_FISCAIS")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "0F766E"
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 62
    ws.column_dimensions["C"].width = 86

    def _st(cel, bg="FFFFFF", fg="0F172A", bold=False, size=10, wrap=True, h="left"):
        cel.fill = _fill(bg)
        cel.font = _fonte(bold=bold, color=fg, size=size)
        cel.alignment = _aln(h, "center", wrap)
        cel.border = _brd("99F6E4")
        return cel

    ws.merge_cells("B1:C1")
    c = ws["B1"]
    c.value = "Roteiro das Informações dos Fiscais"
    _st(c, bg="0F766E", fg="FFFFFF", bold=True, size=14)
    ws.row_dimensions[1].height = 30

    ws.merge_cells("B2:C4")
    c = ws["B2"]
    c.value = (
        "Esta aba registra a orientação prévia definida antes do envio da ColetaMestre ao fiscal. "
        "Ela não altera cálculos; documenta a base esperada, reduz ruído de preenchimento e apoia a leitura posterior pela ferramenta."
    )
    _st(c, bg="F0FDFA", fg="134E4A", wrap=True)
    ws.row_dimensions[2].height = 54

    row = 6
    _st(ws.cell(row, 2, "Resumo automático"), bg="115E59", fg="FFFFFF", bold=True, h="center")
    _st(ws.cell(row, 3, "Resultado"), bg="115E59", fg="FFFFFF", bold=True, h="center")
    row += 1

    resumo = [
        ("Modelo sugerido", _v("modelo_sugerido")),
        ("Base esperada", _v("base_esperada")),
        ("Observação sobre qualidade da base", _v("observacao_qualidade_base")),
        ("Contexto do roteiro", _v("contexto")),
        ("Data do roteiro", _v("data_roteiro")),
    ]
    for campo, valor in resumo:
        _st(ws.cell(row, 2, campo), bg="ECFDF5", fg="134E4A", bold=True)
        _st(ws.cell(row, 3, valor), bg="FFFFFF", fg="0F172A")
        ws.row_dimensions[row].height = 30 if len(str(valor)) < 95 else 54
        row += 1

    row += 1
    _st(ws.cell(row, 2, "Pergunta"), bg="115E59", fg="FFFFFF", bold=True, h="center")
    _st(ws.cell(row, 3, "Resposta"), bg="115E59", fg="FFFFFF", bold=True, h="center")
    row += 1

    perguntas = [
        ("1. O fiscal conseguirá informar valores finais reconhecidos para pagamento por competência mensal?", "tem_financeiro_mensal"),
        ("2. Esses valores estarão organizados pela competência de referência, e não apenas pela data do pagamento?", "financeiro_por_competencia"),
        ("3. Até qual competência/mês haverá informação financeira?", "ultima_competencia_financeira"),
        ("4. O fiscal conseguirá informar saldo remanescente por itens/quantidades?", "tem_saldo_remanescente"),
        ("5. Esse saldo remanescente representa a posição de qual competência/data de corte?", "competencia_data_saldo_remanescente"),
        ("6. Caso não haja saldo remanescente por itens, o fiscal conseguirá informar itens/quantidades consumidos ou executados por ciclo?", "tem_itens_consumidos"),
        ("7. Há casos em que o valor informado, pedido de compra ou preço utilizado foi formado em uma competência, mas a entrega, execução, medição ou atesto ocorreu em competência posterior ou em outro ciclo?", "ha_preco_origem_diferente"),
        ("8. O fiscal conseguirá informar os valores finais reconhecidos/pagos dos ciclos anteriores, como C0, C1 ou C2?", "tem_valores_ciclos_anteriores"),
        ("9. Caso não consiga informar os valores por ciclo anterior, existe valor formalizado/consolidado anterior que possa servir como ponto de partida da análise?", "tem_valor_formalizado_anterior"),
    ]
    for pergunta, chave in perguntas:
        _st(ws.cell(row, 2, pergunta), bg="ECFDF5", fg="134E4A", bold=True)
        resposta = _v(chave)
        _st(ws.cell(row, 3, resposta), bg="FFF2CC" if resposta == "Não informado" else "FFFFFF", fg="0F172A")
        ws.row_dimensions[row].height = 46 if len(pergunta) > 105 else 32
        row += 1

    row += 1
    _st(ws.cell(row, 2, "Orientação-base ao fiscal"), bg="D1FAE5", fg="134E4A", bold=True)
    _st(ws.cell(row, 3, (
        "Informe o valor final reconhecido pela fiscalização para aquela competência, já descontadas glosas ou descontos contratuais. "
        "A competência deve seguir a referência definida pela GCC para o caso."
    )), bg="F0FDFA", fg="134E4A")
    ws.row_dimensions[row].height = 60
    row += 1

    _st(ws.cell(row, 2, "Conceito de saldo remanescente"), bg="D1FAE5", fg="134E4A", bold=True)
    _st(ws.cell(row, 3, (
        "Saldo remanescente é aquilo que ainda falta executar, entregar, consumir ou faturar no contrato a partir de determinada data de corte. "
        "A data/competência do saldo é essencial para evitar dupla contagem no Valor Total Atualizado."
    )), bg="F0FDFA", fg="134E4A")
    ws.row_dimensions[row].height = 62
    row += 1

    _st(ws.cell(row, 2, "Efeito na análise"), bg="D1FAE5", fg="134E4A", bold=True)
    _st(ws.cell(row, 3, (
        "Retroativo: prioriza a base financeira reconhecida por competência. "
        "VTA: combina execução já realizada até a competência informada com saldo remanescente por itens a partir da competência seguinte, evitando dupla contagem."
    )), bg="F0FDFA", fg="134E4A")
    ws.row_dimensions[row].height = 66

    ws.freeze_panes = "B7"
# <<< ROTEIRO_INFO_FISCAIS_XLS_V5_PONTUAL

# ── Função pública ───────────────────────────────────────────────






def gerar_coleta_unica_inteligente():
    """
    Gera a Coleta Mestre v10 pré-preenchida com os dados da Calculadora.
    Mantém compatibilidade com chamadas existentes nos botões.

    Retorna: bytes do XLSX.
    """
    dados = _dados_sessao()

    wb = Workbook()
    wb.remove(wb.active)

    _construir_parametros(wb, dados)
    _construir_ciclos(wb, dados)
    _construir_financeiro(wb, dados)
    _construir_itens(wb)
    _construir_aditivos(wb)
    _construir_diagnostico(wb)
    # ROTEIRO_INFO_FISCAIS desabilitado no XLS inicial; equalização ocorrerá após upload.
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

# >>> PATCH_XLS_COLETAMESTRE_ENXUTA_V1
# Enxugamento visual da ColetaMestre: aba INICIO, abas técnicas ocultas e esquema de cores por responsabilidade.

def _aplicar_ux_coletamestre_enxuta_v1(bytes_xlsx):
    """Pós-processa a ColetaMestre para simplificar a experiência do fiscal/GCC.

    Não altera valores, fórmulas, leitor, cálculo ou documentos.
    Apenas cria uma aba INICIO, aplica cores e oculta abas técnicas.
    """
    try:
        from io import BytesIO
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    except Exception:
        return bytes_xlsx

    def _as_bytes(data):
        if hasattr(data, "getvalue"):
            return data.getvalue()
        return data

    try:
        bio = BytesIO(_as_bytes(bytes_xlsx))
        wb = load_workbook(bio)
    except Exception:
        return bytes_xlsx

    # Paleta por responsabilidade.
    COR_SISTEMA = "0F766E"       # teal escuro
    COR_SISTEMA_CLARO = "CCFBF1" # teal claro
    COR_AUTO = "E0F2FE"          # azul claro / automático
    COR_GCC = "EDE9FE"           # lilás claro
    COR_GCC_TXT = "4C1D95"
    COR_FISCAL = "FFF2CC"        # amarelo preenchimento
    COR_ALERTA = "FEF3C7"        # âmbar
    COR_TECNICA = "CBD5E1"       # cinza/azul técnico
    COR_CINZA = "F8FAFC"
    COR_BORDA = "CBD5E1"
    COR_TEXTO = "0F172A"
    COR_MUTED = "475569"
    BRANCO = "FFFFFF"

    thin = Side(style="thin", color=COR_BORDA)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def fill(cor):
        return PatternFill("solid", fgColor=cor)

    def set_cell(ws, ref, value=None, *, font=None, fill_color=None, border_on=True, align=None, num_format=None):
        cell = ws[ref]
        if value is not None:
            cell.value = value
        if font is not None:
            cell.font = font
        if fill_color is not None:
            cell.fill = fill(fill_color)
        if border_on:
            cell.border = border
        if align is not None:
            cell.alignment = align
        if num_format is not None:
            cell.number_format = num_format
        return cell

    def merge_set(ws, cell_range, value, *, font=None, fill_color=None, align=None):
        ws.merge_cells(cell_range)
        start = cell_range.split(":", 1)[0]
        set_cell(ws, start, value, font=font, fill_color=fill_color, align=align or Alignment(vertical="center", wrap_text=True))
        for row in ws[cell_range]:
            for cell in row:
                cell.border = border
                if fill_color:
                    cell.fill = fill(fill_color)
                if align:
                    cell.alignment = align

    def safe_text(v):
        if v is None:
            return ""
        return str(v).strip()

    def buscar_valor(nome_aba, rotulo):
        if nome_aba not in wb.sheetnames:
            return ""
        ws = wb[nome_aba]
        alvo = safe_text(rotulo).lower()
        for row in ws.iter_rows():
            for cell in row:
                if safe_text(cell.value).lower() == alvo:
                    # busca primeiro valor útil nas próximas 4 colunas da mesma linha
                    for col in range(cell.column + 1, min(cell.column + 5, ws.max_column) + 1):
                        val = ws.cell(cell.row, col).value
                        if safe_text(val):
                            return val
        return ""

    # Remove INICIO anterior para reconstruir limpo.
    if "INICIO" in wb.sheetnames:
        wb.remove(wb["INICIO"])
    ws = wb.create_sheet("INICIO", 0)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A9"
    ws.sheet_properties.tabColor = COR_SISTEMA

    # Larguras.
    widths = {"A": 4, "B": 24, "C": 32, "D": 34, "E": 22, "F": 18, "G": 18}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    title_font = Font(bold=True, color=BRANCO, size=14)
    subtitle_font = Font(color=COR_MUTED, size=10)
    section_font = Font(bold=True, color=BRANCO, size=11)
    header_font = Font(bold=True, color=BRANCO, size=10)
    normal_font = Font(color=COR_TEXTO, size=10)
    muted_font = Font(color=COR_MUTED, size=9)
    gcc_font = Font(color=COR_GCC_TXT, bold=True, size=10)
    fiscal_font = Font(color="7C2D12", bold=True, size=10)
    auto_font = Font(color="064E3B", bold=True, size=10)
    alert_font = Font(color="92400E", bold=True, size=10)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    merge_set(ws, "B2:G2", "cl8us · ColetaMestre simplificada", font=title_font, fill_color=COR_SISTEMA, align=left)
    merge_set(ws, "B3:G4", "Use esta planilha para coletar apenas as informações necessárias à apuração. O único dado específico do contrato a ser informado neste modelo é a vigência final; dados como nome, número do contrato ou fornecedor não são necessários nesta etapa.", font=subtitle_font, fill_color=COR_CINZA, align=left)

    merge_set(ws, "B6:G6", "1. Como usar", font=section_font, fill_color=COR_SISTEMA, align=left)
    uso = [
        ["Etapa", "O que fazer", "Responsável"],
        ["1", "Revise o roteiro abaixo e confirme o tipo de informação que será solicitada ao fiscal.", "GCC"],
        ["2", "Envie ao fiscal apenas as abas FINANCEIRO e/ou ITENS, conforme o caso. ADITIVOS é opcional.", "GCC"],
        ["3", "O fiscal preenche somente as células amarelas indicadas como FISCAL.", "Fiscal"],
        ["4", "As demais abas ficam ocultas porque são técnicas/automáticas e necessárias apenas ao sistema.", "Sistema"],
    ]
    for r_idx, row in enumerate(uso, start=7):
        for c_idx, val in enumerate(row, start=2):
            ref = ws.cell(r_idx, c_idx).coordinate
            if r_idx == 7:
                set_cell(ws, ref, val, font=header_font, fill_color=COR_SISTEMA, align=center)
            else:
                cor = COR_CINZA
                if val == "GCC": cor = COR_GCC
                elif val == "Fiscal": cor = COR_FISCAL
                elif val == "Sistema": cor = COR_AUTO
                set_cell(ws, ref, val, font=normal_font, fill_color=cor, align=left if c_idx == 3 else center)
        ws.row_dimensions[r_idx].height = 28

    merge_set(ws, "B13:G13", "2. Esquema de cores", font=section_font, fill_color=COR_SISTEMA, align=left)
    cores = [
        ["Categoria", "Cor", "Aplicação"],
        ["Sistema / automático", "Teal ou azul claro", "Dados calculados, ciclos, diagnóstico e abas técnicas. Não alterar."],
        ["GCC", "Lilás", "Orientação, ajustes, validação e campos de controle da GCC."],
        ["Fiscal", "Amarelo", "Campos que devem ser preenchidos pelo fiscal."],
        ["Atenção", "Âmbar", "Alertas, observações, exceções e pontos de conferência."],
    ]
    for r_idx, row in enumerate(cores, start=14):
        for c_idx, val in enumerate(row, start=2):
            ref = ws.cell(r_idx, c_idx).coordinate
            if r_idx == 14:
                set_cell(ws, ref, val, font=header_font, fill_color=COR_SISTEMA, align=center)
            else:
                categoria = row[0]
                cor = COR_AUTO
                font = auto_font
                if categoria == "GCC":
                    cor, font = COR_GCC, gcc_font
                elif categoria == "Fiscal":
                    cor, font = COR_FISCAL, fiscal_font
                elif categoria == "Atenção":
                    cor, font = COR_ALERTA, alert_font
                set_cell(ws, ref, val, font=font if c_idx == 2 else normal_font, fill_color=cor if c_idx == 3 else COR_CINZA, align=left)
        ws.row_dimensions[r_idx].height = 30

    # Roteiro / vigência.
    modelo = buscar_valor("ROTEIRO_INFO_FISCAIS", "Modelo sugerido")
    base = buscar_valor("ROTEIRO_INFO_FISCAIS", "Base esperada")
    obs = buscar_valor("ROTEIRO_INFO_FISCAIS", "Observação sobre qualidade da base")
    contexto = buscar_valor("ROTEIRO_INFO_FISCAIS", "Contexto do roteiro")
    vig_final = buscar_valor("PARAMETROS_CONTRATO", "Vigência final")

    merge_set(ws, "B21:G21", "3. Roteiro das Informações dos Fiscais", font=section_font, fill_color=COR_SISTEMA, align=left)
    resumo = [
        ["Item", "Resultado"],
        ["Modelo sugerido", modelo or "[não informado]"],
        ["Base esperada", base or "[não informado]"],
        ["Qualidade/ressalva", obs or "[não informado]"],
        ["Contexto do roteiro", contexto or "[não informado]"],
        ["Vigência final do contrato", vig_final or "[preencher, se necessário]"],
    ]
    for r_idx, row in enumerate(resumo, start=22):
        for c_idx, val in enumerate(row, start=2):
            ref = ws.cell(r_idx, c_idx).coordinate
            if r_idx == 22:
                set_cell(ws, ref, val, font=header_font, fill_color=COR_SISTEMA, align=center)
            else:
                fill_color = COR_GCC if row[0] == "Vigência final do contrato" else COR_CINZA
                set_cell(ws, ref, val, font=normal_font, fill_color=fill_color, align=left)
        # funde resultado C:G
        try:
            ws.merge_cells(start_row=r_idx, start_column=3, end_row=r_idx, end_column=7)
        except Exception:
            pass
        for col in range(3, 8):
            ws.cell(r_idx, col).border = border
            ws.cell(r_idx, col).alignment = left
            ws.cell(r_idx, col).fill = fill(COR_GCC if row[0] == "Vigência final do contrato" else COR_CINZA)
        ws.row_dimensions[r_idx].height = 30

    merge_set(ws, "B30:G30", "4. Abas da planilha", font=section_font, fill_color=COR_SISTEMA, align=left)
    abas = [
        ["Aba", "Status", "Finalidade"],
        ["FINANCEIRO", "Visível", "Base preferencial para retroativo: valor final reconhecido por competência."],
        ["ITENS", "Visível", "Saldo remanescente/itens para VTA e conferência do estoque contratual."],
        ["ADITIVOS", "Visível opcional", "Preencher somente se houver aditivo/supressão a considerar."],
        ["PARAMETROS_CONTRATO, CICLOS, DIAGNOSTICO, ROTEIRO_INFO_FISCAIS", "Ocultas", "Abas técnicas mantidas para leitura, rastreabilidade e diagnóstico do sistema."],
    ]
    for r_idx, row in enumerate(abas, start=31):
        for c_idx, val in enumerate(row, start=2):
            ref = ws.cell(r_idx, c_idx).coordinate
            if r_idx == 31:
                set_cell(ws, ref, val, font=header_font, fill_color=COR_SISTEMA, align=center)
            else:
                cor = COR_FISCAL if "Visível" in row[1] else COR_AUTO
                set_cell(ws, ref, val, font=normal_font, fill_color=cor if c_idx == 3 else COR_CINZA, align=left)
        try:
            ws.merge_cells(start_row=r_idx, start_column=4, end_row=r_idx, end_column=7)
        except Exception:
            pass
        for col in range(4, 8):
            ws.cell(r_idx, col).border = border
            ws.cell(r_idx, col).alignment = left
            ws.cell(r_idx, col).fill = fill(COR_CINZA)
        ws.row_dimensions[r_idx].height = 32

    merge_set(ws, "B39:G40", "Regra operacional: o valor financeiro solicitado deve ser o valor final reconhecido para pagamento na competência de referência, já considerando glosas/descontos contratuais, em base bruta econômica. A competência deve seguir a orientação da GCC para o caso.", font=Font(color="92400E", bold=True, size=10), fill_color=COR_ALERTA, align=left)

    # Estilo das abas principais.
    def style_financeiro():
        if "FINANCEIRO" not in wb.sheetnames:
            return
        sh = wb["FINANCEIRO"]
        sh.sheet_properties.tabColor = COR_FISCAL
        sh.sheet_view.showGridLines = False
        sh.freeze_panes = "A5"
        # Atualiza legenda para amarelo.
        for cell in sh[2]:
            if isinstance(cell.value, str):
                cell.value = cell.value.replace("Verde", "Amarelo")
        # Cabeçalhos e áreas.
        for row in range(4, min(sh.max_row, 208) + 1):
            if row == 4:
                for c in range(1, min(sh.max_column, 5) + 1):
                    sh.cell(row, c).font = Font(bold=True, color=BRANCO)
                    sh.cell(row, c).alignment = center
            # A automático
            sh.cell(row, 1).fill = fill(COR_AUTO if row > 4 else COR_SISTEMA)
            # B:C fiscal
            for c in [2, 3]:
                sh.cell(row, c).fill = fill(COR_FISCAL if row > 4 else COR_SISTEMA)
            # D GCC
            if sh.max_column >= 4:
                sh.cell(row, 4).fill = fill(COR_GCC if row > 4 else COR_SISTEMA)
            # E atenção/GCC
            if sh.max_column >= 5:
                sh.cell(row, 5).fill = fill(COR_ALERTA if row > 4 else COR_SISTEMA)
            for c in range(1, min(sh.max_column, 5) + 1):
                sh.cell(row, c).border = border
                sh.cell(row, c).alignment = Alignment(vertical="center", wrap_text=True)
        sh.column_dimensions["A"].width = 13
        sh.column_dimensions["B"].width = 32
        sh.column_dimensions["C"].width = 24
        sh.column_dimensions["D"].width = 22
        sh.column_dimensions["E"].width = 38

    def style_itens():
        if "ITENS" not in wb.sheetnames:
            return
        sh = wb["ITENS"]
        sh.sheet_properties.tabColor = COR_FISCAL
        sh.sheet_view.showGridLines = False
        sh.freeze_panes = "A8"
        for row in sh.iter_rows(min_row=1, max_row=min(sh.max_row, 8)):
            for cell in row:
                if isinstance(cell.value, str):
                    cell.value = cell.value.replace("Verde", "Amarelo").replace("você preenche", "GCC preenche")
        for row in range(7, min(sh.max_row, 208) + 1):
            # Headers
            if row == 7:
                for c in range(1, min(sh.max_column, 14) + 1):
                    sh.cell(row, c).font = Font(bold=True, color=BRANCO)
                    sh.cell(row, c).fill = fill(COR_SISTEMA)
                    sh.cell(row, c).alignment = center
            else:
                # A:C GCC; D/K/M automático; E:I fiscal; J/L/N GCC/atenção conforme estrutura atual.
                for c in range(1, min(sh.max_column, 14) + 1):
                    if c in [1, 2, 3, 10, 12, 14]:
                        cor = COR_GCC
                    elif c in [5, 6, 7, 8, 9]:
                        cor = COR_FISCAL
                    elif c in [4, 11, 13]:
                        cor = COR_AUTO
                    else:
                        cor = COR_CINZA
                    sh.cell(row, c).fill = fill(cor)
                    sh.cell(row, c).alignment = Alignment(vertical="center", wrap_text=True)
            for c in range(1, min(sh.max_column, 14) + 1):
                sh.cell(row, c).border = border
        for col, width in {"A": 18, "B": 15, "C": 18, "D": 18, "E": 16, "F": 16, "G": 16, "H": 16, "I": 16, "J": 16, "K": 18, "L": 16, "M": 18, "N": 28}.items():
            sh.column_dimensions[col].width = width

    def style_aditivos():
        if "ADITIVOS" not in wb.sheetnames:
            return
        sh = wb["ADITIVOS"]
        sh.sheet_properties.tabColor = COR_GCC
        sh.sheet_view.showGridLines = False
        sh.freeze_panes = "A5"
        for row in range(1, min(sh.max_row, 10) + 1):
            for cell in sh[row]:
                if cell.value is not None:
                    cell.fill = fill(COR_GCC if row <= 4 else COR_CINZA)
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                    cell.border = border

    style_financeiro()
    style_itens()
    style_aditivos()

    # Oculta abas técnicas sem excluir, para preservar leitor/rastreabilidade.
    for nome in ["PARAMETROS_CONTRATO", "CICLOS", "DIAGNOSTICO", "ROTEIRO_INFO_FISCAIS"]:
        if nome in wb.sheetnames:
            sh = wb[nome]
            sh.sheet_properties.tabColor = COR_TECNICA
            sh.sheet_state = "hidden"

    # Mantém abas operacionais visíveis.
    for nome in ["INICIO", "FINANCEIRO", "ITENS", "ADITIVOS"]:
        if nome in wb.sheetnames:
            wb[nome].sheet_state = "visible"

    # Garante que a aba ativa seja INICIO.
    try:
        wb.active = wb.sheetnames.index("INICIO")
    except Exception:
        pass

    saida = BytesIO()
    try:
        wb.save(saida)
        return saida.getvalue()
    except Exception:
        return bytes_xlsx


# Envolve a função central de geração sem alterar sua lógica interna.
try:
    _gerar_coleta_unica_inteligente_base_enxuta_v1 = gerar_coleta_unica_inteligente

    def gerar_coleta_unica_inteligente(*args, **kwargs):
        bytes_xlsx = _gerar_coleta_unica_inteligente_base_enxuta_v1(*args, **kwargs)
        return _aplicar_ux_coletamestre_enxuta_v1(bytes_xlsx)
except Exception:
    pass
# <<< PATCH_XLS_COLETAMESTRE_ENXUTA_V1

# >>> PATCH_VISUAL_ADITIVOS_PROFISSIONAL_V11
# Ajuste exclusivamente visual da aba ADITIVOS da ColetaMestre.
# Não altera valores, fórmulas, validações, nomes de abas ou lógica de cálculo.

def _aplicar_visual_aditivos_profissional_v11(bytes_xlsx):
    try:
        from io import BytesIO
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    except Exception:
        return bytes_xlsx

    try:
        entrada = BytesIO(bytes_xlsx)
        wb = load_workbook(entrada)

        if "ADITIVOS" not in wb.sheetnames:
            saida = BytesIO()
            wb.save(saida)
            saida.seek(0)
            return saida.getvalue()

        ws = wb["ADITIVOS"]

        # Paleta corporativa sóbria.
        COR_TITULO = "0F766E"       # verde-petróleo
        COR_TITULO2 = "115E59"
        COR_HEADER = "334155"       # azul ardósia escuro
        COR_HEADER_ALT = "475569"
        COR_GCC_BG = "FEF3C7"       # amarelo suave/preenchimento
        COR_GCC_FG = "78350F"
        COR_AUTO_BG = "F1F5F9"      # cinza azulado/automático
        COR_AUTO_FG = "1E3A5F"
        COR_TRAT_BG = "FFEDD5"      # âmbar claro/tratamento
        COR_TRAT_FG = "7C2D12"
        COR_OBS_BG = "F8FAFC"
        COR_OBS_FG = "334155"
        COR_BORDA = "CBD5E1"
        COR_BRANCO = "FFFFFF"

        thin = Side(style="thin", color=COR_BORDA)
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def fill(hex_color):
            return PatternFill("solid", fgColor=hex_color)

        def style_range(ref, *, bg=None, fg=None, bold=False, italic=False, size=None,
                        h="left", v="center", wrap=True):
            for row in ws[ref]:
                for cell in row:
                    if bg:
                        cell.fill = fill(bg)
                    cell.font = Font(
                        name="Aptos",
                        size=size or 10,
                        bold=bold,
                        italic=italic,
                        color=fg or "0F172A",
                    )
                    cell.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap)
                    cell.border = border

        max_row = max(ws.max_row or 1, 35)

        # Título, legenda e observação.
        style_range("A1:K1", bg=COR_TITULO, fg=COR_BRANCO, bold=True, size=13, h="left")
        ws.row_dimensions[1].height = 26

        style_range("A2:K2", bg="ECFDF5", fg="134E4A", bold=True, size=10, h="left")
        ws.row_dimensions[2].height = 24

        style_range("A3:K3", bg=COR_OBS_BG, fg=COR_OBS_FG, italic=True, size=10, h="left")
        ws.row_dimensions[3].height = 28

        # Linha branca de respiro, se existir.
        style_range("A4:K4", bg="FFFFFF", fg="0F172A", size=10, h="left")
        ws.row_dimensions[4].height = 18

        # Cabeçalho da tabela: contraste alto.
        style_range("A5:K5", bg=COR_HEADER, fg=COR_BRANCO, bold=True, size=10, h="center")
        ws.row_dimensions[5].height = 64

        # Corpo: campos de preenchimento, automáticos e tratamento.
        if max_row >= 6:
            # Campos GCC/preenchíveis.
            style_range(f"A6:B{max_row}", bg=COR_GCC_BG, fg=COR_GCC_FG, h="left")
            style_range(f"D6:F{max_row}", bg=COR_GCC_BG, fg=COR_GCC_FG, h="left")
            style_range(f"H6:H{max_row}", bg=COR_GCC_BG, fg=COR_GCC_FG, h="center")

            # Campos automáticos/calculados.
            style_range(f"C6:C{max_row}", bg=COR_AUTO_BG, fg=COR_AUTO_FG, h="center")
            style_range(f"G6:G{max_row}", bg=COR_AUTO_BG, fg=COR_AUTO_FG, h="right")
            style_range(f"I6:J{max_row}", bg=COR_AUTO_BG, fg=COR_AUTO_FG, h="right")

            # Tratamento do aditivo.
            style_range(f"K6:K{max_row}", bg=COR_TRAT_BG, fg=COR_TRAT_FG, bold=True, h="left")

            for r in range(6, max_row + 1):
                ws.row_dimensions[r].height = 24

        # Larguras ajustadas para evitar cortes e excesso horizontal.
        widths = {
            "A": 16, "B": 17, "C": 18, "D": 17, "E": 17,
            "F": 18, "G": 19, "H": 17, "I": 18, "J": 19, "K": 30,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        # Formatos numéricos, preservando fórmulas/valores.
        for row in range(6, max_row + 1):
            ws[f"E{row}"].number_format = '#,##0.00'
            ws[f"F{row}"].number_format = '"R$" #,##0.00'
            ws[f"G{row}"].number_format = '"R$" #,##0.00'
            ws[f"I{row}"].number_format = '0.000000'
            ws[f"J{row}"].number_format = '"R$" #,##0.00'

        # Guia da aba com cor própria do grupo aditivos.
        ws.sheet_properties.tabColor = "D97706"

        # Congelamento visual e filtro apenas no cabeçalho, sem alterar dados.
        ws.freeze_panes = "A6"
        try:
            ws.auto_filter.ref = f"A5:K{max_row}"
        except Exception:
            pass

        # Demais abas: tab colors coerentes, sem tocar em conteúdo.
        tab_colors = {
            "INICIO": "0F766E",
            "FINANCEIRO": "2563EB",
            "FINANCEIRO_MENSAL": "2563EB",
            "FINANCEIRO_HISTORICO": "2563EB",
            "ITENS": "16A34A",
            "ITENS_REMANESCENTES": "16A34A",
            "ITENS_CICLOS": "16A34A",
            "ADITIVOS": "D97706",
            "ADITIVOS_SUPRESSOES": "D97706",
            "PARAMETROS_CONTRATO": "64748B",
            "CICLOS": "64748B",
            "DIAGNOSTICO": "64748B",
        }
        for nome, cor in tab_colors.items():
            if nome in wb.sheetnames:
                try:
                    wb[nome].sheet_properties.tabColor = cor
                except Exception:
                    pass

        saida = BytesIO()
        wb.save(saida)
        saida.seek(0)
        return saida.getvalue()
    except Exception:
        return bytes_xlsx


try:
    _gerar_coleta_unica_inteligente_base_visual_aditivos_v11 = gerar_coleta_unica_inteligente

    def gerar_coleta_unica_inteligente(*args, **kwargs):
        bytes_xlsx = _gerar_coleta_unica_inteligente_base_visual_aditivos_v11(*args, **kwargs)
        return _aplicar_visual_aditivos_profissional_v11(bytes_xlsx)
except Exception:
    pass
# <<< PATCH_VISUAL_ADITIVOS_PROFISSIONAL_V11
