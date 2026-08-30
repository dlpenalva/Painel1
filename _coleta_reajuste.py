"""Motor canônico do XLS-first do Master 2.0.

O Excel é a fonte de verdade: este módulo apenas preenche os marcos já
apurados pela calculadora e valida a estrutura no retorno. Os resultados
financeiros permanecem fórmulas da própria planilha.
"""

from __future__ import annotations

from copy import copy
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
import re
import unicodedata
from typing import Any

from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, PatternFill

from _seguranca_xlsx import (
    ErroSegurancaXlsx,
    garantir_xlsx_validado,
    validar_geometria_workbook,
)

from _capacidade_pcs import CAPACIDADE_PCS, ULTIMA_LINHA_PCS
from _capacidades_apuracao import avaliar_capacidades_apuracao
from _efeitos_financeiros_pc import (
    efeito_financeiro_pc,
    reconciliar_inicios_efeito,
)


NOME_ARQUIVO_COLETA = "Coleta_Reajuste.xlsx"
CAMINHO_MODELO_COLETA = Path(__file__).resolve().parent / "templates" / NOME_ARQUIVO_COLETA

ABAS_CANONICAS = (
    "CONTROLE",
    "parametros",
    "financeiro",
    "itens_Remanesc",
    "itens_Consumidos",
    "itens_PC",
    "aditivos",
    "posicao_contratual",
    "itens_RC",
    "historico_VU",
    "RESULTADOS",
)

ABAS_OBRIGATORIAS_LEGADO = tuple(
    aba for aba in ABAS_CANONICAS if aba != "posicao_contratual"
)

NOMES_RESULTADOS_OBRIGATORIOS = {
    "METODO_RETROATIVO",
    "TOLERANCIA_DIVERGENCIA",
    "VALOR_MANUAL_RETRO",
    "JUSTIFICATIVA_RETRO",
    "RETRO_FIN",
    "RETRO_PC",
    "RETRO_ITENS",
    "RETRO_OFICIAL",
    "VTA_CALCULADO",
    "AJUSTE_MANUAL_VTA",
    "VTA_MANUAL_OFICIAL",
    "VTA_FINAL",
    "QTD_REM_OFICIAL",
    "REM_BASE_OFICIAL",
    "REM_ATUALIZADO_OFICIAL",
}

ABAS_PROIBIDAS = ("itens_Execucao_Saldo", "Itens_Execução", "REGRA_NEGOCIO_CLAUS", "Regra")
ERROS_EXCEL = {"#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#N/A", "#NUM!", "#NULL!"}

COR_TEXTO = "FF595959"
COR_MARINHO = "FF123B63"
PREENCHIMENTO_AUTOMATICO = PatternFill("solid", fgColor="FFEDEDED")
PREENCHIMENTO_ENTRADA = PatternFill("solid", fgColor="FFFFF2CC")


def _texto_sem_acento(valor: Any) -> str:
    texto = unicodedata.normalize("NFKD", str(valor or ""))
    return "".join(ch for ch in texto if not unicodedata.combining(ch))


def _data(valor: Any) -> datetime | None:
    if valor in (None, ""):
        return None
    if isinstance(valor, datetime):
        return valor.replace(tzinfo=None)
    if isinstance(valor, date):
        return datetime(valor.year, valor.month, valor.day)
    texto = str(valor).strip()
    for formato in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%Y"):
        try:
            return datetime.strptime(texto, formato)
        except ValueError:
            continue
    return None


def _primeiro_dia_mes(valor: datetime) -> datetime:
    return valor.replace(day=1, hour=0, minute=0, second=0, microsecond=0)


def _numero_ciclo(valor: Any) -> int | None:
    match = re.search(r"\bC\s*([0-4])\b", str(valor or "").upper())
    return int(match.group(1)) if match else None


def _numero(valor: Any) -> float | None:
    if valor in (None, "") or isinstance(valor, bool):
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    texto = str(valor).strip().replace("%", "").replace(" ", "")
    if "," in texto:
        texto = texto.replace(".", "").replace(",", ".")
    try:
        return float(texto)
    except ValueError:
        return None


def _lista_compacta_competencias(referencias: list[str], limite: int = 12) -> str:
    """Lista compacta de competencias afetadas para UMA unica mensagem.

    EF-G1: inconsistencias de `financeiro!G` sao agrupadas — nunca uma caixa
    por competencia. Acima do limite, o excedente vira contagem.
    """
    exibidas = referencias[:limite]
    texto = "; ".join(exibidas)
    restantes = len(referencias) - len(exibidas)
    if restantes > 0:
        texto += f"; e mais {restantes} competência(s)"
    return texto + "."


def _tabela_ciclos_financeiros(wb) -> list[tuple[str, date | None, date | None]]:
    ws = wb["parametros"]
    return [
        (
            str(ws[f"B{row}"].value or "").strip().upper(),
            _data(ws[f"C{row}"].value),
            _data(ws[f"D{row}"].value),
        )
        for row in range(2, 7)
    ]


def _ciclo_por_competencia_financeira(
    wb, competencia: Any, tabela_ciclos: list[tuple[str, date | None, date | None]] | None = None
) -> str:
    data_comp = _data(competencia)
    if data_comp is None:
        return ""
    if tabela_ciclos is None:
        tabela_ciclos = _tabela_ciclos_financeiros(wb)
    for ciclo, inicio, fim in tabela_ciclos:
        if ciclo and inicio and fim and inicio <= data_comp <= fim:
            return ciclo
    return ""


def _marco_grade_financeira(wb) -> datetime | None:
    """Primeira competencia da grade da aba financeiro (ancora cronologica).

    A grade e escrita pelo gerador como competencias mensais consecutivas a
    partir do marco (inicio cronologico de C0), e o CICLO de cada competencia e
    o bloco fixo de 12 meses contado desse marco. Ler a ancora do proprio
    arquivo mantem a identificacao de ciclo colada a execucao financeira.
    """
    ws = wb["financeiro"]
    for row in range(2, 74):
        data = _data(ws[f"A{row}"].value)
        if data is not None:
            return data
    return None


def _ciclo_cronologico_financeiro(marco: datetime | None, competencia: Any) -> str:
    """CICLO da competencia pela cronologia FISICA da execucao financeira.

    Mesma identificacao usada por `_gerador_masterfile._preencher_financeiro`
    para escrever `financeiro!G` (bloco fixo de 12 competencias contado do
    marco). NUNCA reconstroi o ciclo pela janela juridica `parametros!C:D`:
    ela se desloca depois de TEMPESTIVO*/ADIANTADO e passa a atribuir a
    competencia de fronteira a um ciclo diferente daquele que decidiu o efeito
    — origem dos avisos falsos de ajuste manual. `financeiro!B` tambem nao
    serve a este diagnostico: a formula do template resolve o ciclo pela mesma
    janela deslocavel de `parametros!C:D`.
    """
    data_comp = _data(competencia)
    if marco is None or data_comp is None:
        return ""
    meses = (data_comp.year - marco.year) * 12 + (data_comp.month - marco.month)
    if meses < 0:
        return ""
    indice = meses // 12
    return f"C{indice}" if indice <= 4 else ""


def _percentual_ciclo(ciclo: dict[str, Any]) -> float | None:
    for chave in ("percentual_aplicado", "percentual_indice", "variacao"):
        valor = _numero(ciclo.get(chave))
        if valor is not None:
            return valor / 100 if abs(valor) > 1 else valor
    fator = _numero(ciclo.get("fator"))
    if fator is not None:
        return fator - 1 if fator >= 0.5 else fator
    return None


def _ciclo_em_analise(ciclo: dict[str, Any]) -> bool:
    if "objeto_analise_atual" in ciclo:
        return bool(ciclo.get("objeto_analise_atual"))
    if "ciclo_ja_concedido" in ciclo:
        return not bool(ciclo.get("ciclo_ja_concedido"))
    return True


def _inicio_teorico(ciclo: dict[str, Any]) -> datetime | None:
    """Início do ciclo declarado explicitamente pela calculadora.

    Só o marco explícito (`inicio_ciclo`) serve de âncora do calendário
    contratual. `data_base`/`periodo_inicio` são a janela do ÍNDICE — conceito
    independente que jamais define DATA_INICIO/DATA_FIM nem desloca o ciclo
    seguinte (ver `_montar_ciclos`).
    """
    inicio_direto = _data(ciclo.get("inicio_ciclo"))
    if inicio_direto:
        return _primeiro_dia_mes(inicio_direto)
    return None


def _montar_ciclos(dados: dict[str, Any]) -> tuple[list[dict[str, Any]], set[int], list[str]]:
    ciclos_origem = dados.get("ciclos") or []
    fornecidos: dict[int, dict[str, Any]] = {}
    for ciclo in ciclos_origem:
        if not isinstance(ciclo, dict):
            continue
        numero = _numero_ciclo(ciclo.get("ciclo") or ciclo.get("Ciclo"))
        if numero is not None and numero > 0:
            fornecidos[numero] = ciclo
    if not fornecidos:
        raise ValueError("A calculadora não informou nenhum ciclo entre C1 e C4.")

    ultimo = max(fornecidos)
    alvos = {n for n, ciclo in fornecidos.items() if _ciclo_em_analise(ciclo)}
    if not alvos:
        raise ValueError("Nenhum ciclo foi marcado como objeto desta apuração.")

    # CALENDÁRIO CONTRATUAL — âncora única.
    # Cada ciclo tem exatamente 12 competências mensais consecutivas. A âncora
    # é tomada UMA vez (marco explícito mais antigo, senão a data-base original)
    # e todos os demais ciclos decorrem dela em blocos de 12 meses. A janela do
    # índice e o início do efeito financeiro não participam desta materialização.
    explicitos = {n: inicio for n, ciclo in sorted(fornecidos.items())
                  if (inicio := _inicio_teorico(ciclo))}
    if explicitos:
        ancora_numero = min(explicitos)
        ancora_inicio = explicitos[ancora_numero]
    else:
        # Sem marco explícito, a âncora é a janela do índice do ciclo MAIS
        # ANTIGO informado: só nele a janela coincide com o aniversário
        # contratual. A janela dos ciclos SEGUINTES jamais é consultada — é ela
        # que carrega o início do efeito financeiro e deslocava o calendário.
        ancora_numero = min(fornecidos)
        primeiro = fornecidos[ancora_numero]
        ancora = (
            _data(primeiro.get("data_base"))
            or _data(primeiro.get("periodo_inicio"))
            or _data(dados.get("data_base_original"))
        )
        if not ancora:
            raise ValueError("Não foi possível identificar a data-base dos ciclos.")
        ancora_inicio = _primeiro_dia_mes(ancora + relativedelta(months=12))

    inicios = {
        numero: ancora_inicio + relativedelta(months=12 * (numero - ancora_numero))
        for numero in range(0, ultimo + 1)
    }

    contexto = dados.get("contexto_contratual_anterior") or {}
    ultimo_contexto = _numero_ciclo(contexto.get("ultimo_ciclo_concedido"))
    percentual_contexto = _numero(contexto.get("percentual_ja_aplicado_pct"))
    if percentual_contexto is not None and abs(percentual_contexto) > 1:
        percentual_contexto /= 100

    alertas: list[str] = []
    saida: list[dict[str, Any]] = []
    for numero in range(0, ultimo + 1):
        origem = fornecidos.get(numero, {})
        percentual = 0.0 if numero == 0 else _percentual_ciclo(origem)
        if numero > 0 and percentual is None and numero == ultimo_contexto:
            percentual = percentual_contexto
        inicio = _primeiro_dia_mes(inicios[numero])
        fim = inicio + relativedelta(months=12) - relativedelta(days=1)
        alvo = numero in alvos
        if numero > 0 and percentual is None and numero <= max(alvos):
            alertas.append(
                f"C{numero}: percentual histórico não informado; resultados acumulados dependentes ficarão em branco."
            )
        situacao = (
            "Base"
            if numero == 0
            else str(origem.get("situacao_aplicada") or origem.get("situacao") or "").strip()
            or ("Em análise" if alvo else "Histórico fora desta apuração")
        )
        saida.append(
            {
                "numero": numero,
                "nome": f"C{numero}",
                "inicio": inicio,
                "fim": fim,
                "percentual": percentual,
                "situacao": situacao,
                "computar": alvo,
                "financeiro_inicio": _data(origem.get("financeiro_inicio")),
            }
        )
    return saida, alvos, alertas


def _formulas(wb) -> dict[str, str]:
    return {
        f"{ws.title}!{cell.coordinate}": cell.value
        for ws in wb.worksheets
        for row in ws.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    }


def _validar_resultados_integra(wb, etapa: str) -> dict[str, Any]:
    if "RESULTADOS" not in wb.sheetnames:
        raise ValueError(f"A aba RESULTADOS desapareceu na etapa {etapa}.")
    ws = wb["RESULTADOS"]
    formulas = sum(
        1
        for row in ws.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    )
    conteudo = sum(1 for row in ws.iter_rows() for cell in row if cell.value not in (None, ""))
    if ws.sheet_state != "visible":
        raise ValueError(f"A aba RESULTADOS não está visível na etapa {etapa}.")
    if ws["A1"].value != "RESULTADOS CONSOLIDADOS — REAJUSTE CONTRATUAL":
        raise ValueError(f"A aba RESULTADOS está vazia ou foi substituída na etapa {etapa}.")
    if "MEMORIA_RESULTADOS" in wb.sheetnames:
        memoria = wb["MEMORIA_RESULTADOS"]
        formulas_memoria = sum(
            1
            for row in memoria.iter_rows()
            for cell in row
            if isinstance(cell.value, str) and cell.value.startswith("=")
        )
        conteudo_memoria = sum(
            1
            for row in memoria.iter_rows()
            for cell in row
            if cell.value not in (None, "")
        )
        if memoria.sheet_state == "visible":
            raise ValueError(
                f"A MEMORIA_RESULTADOS deve permanecer oculta na etapa {etapa}."
            )
        if formulas < 40 or conteudo < 100:
            raise ValueError(
                f"A RESULTADOS executiva perdeu conteúdo na etapa {etapa}: "
                f"{formulas} fórmulas e {conteudo} células preenchidas."
            )
        if formulas_memoria < 3000 or conteudo_memoria < 3300:
            raise ValueError(
                f"A MEMORIA_RESULTADOS perdeu conteúdo na etapa {etapa}: "
                f"{formulas_memoria} fórmulas e {conteudo_memoria} células preenchidas."
            )
        return {
            "visivel": True,
            "formulas": formulas,
            "conteudo": conteudo,
            "memoria_formulas": formulas_memoria,
            "memoria_conteudo": conteudo_memoria,
        }
    if formulas < 3000 or conteudo < 3300:
        raise ValueError(
            f"A aba RESULTADOS perdeu conteúdo na etapa {etapa}: "
            f"{formulas} fórmulas e {conteudo} células preenchidas."
        )
    return {"visivel": True, "formulas": formulas, "conteudo": conteudo}


def _normalizar_arquivo(wb) -> None:
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                cell.comment = None
        ws.sheet_view.showGridLines = False
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.active = 0
    if wb.views:
        wb.views[0].activeTab = 0
        wb.views[0].firstSheet = 0


def gerar_coleta_reajuste(dados_admissibilidade: dict[str, Any]) -> bytes:
    """Preenche o modelo canônico sem substituir ou calcular suas fórmulas."""

    if not CAMINHO_MODELO_COLETA.exists():
        raise FileNotFoundError(f"Modelo canônico não encontrado: {CAMINHO_MODELO_COLETA}")
    dados = dados_admissibilidade or {}
    ciclos, alvos, _alertas = _montar_ciclos(dados)

    wb = load_workbook(CAMINHO_MODELO_COLETA, data_only=False)
    if tuple(wb.sheetnames) != ABAS_CANONICAS:
        raise ValueError("O modelo canônico possui abas inesperadas ou fora de ordem.")
    _validar_resultados_integra(wb, "logo após o carregamento do template")
    formulas_originais = _formulas(wb)

    ws = wb["CONTROLE"]
    ws["B1"] = "Principal"
    ws["B2"] = f"C{max(alvos)}"
    ws["B3"] = max(ciclo["fim"] for ciclo in ciclos if ciclo["numero"] <= max(alvos))
    ws["B3"].number_format = "mm/yyyy"
    ws["B7"] = str(dados.get("indice") or "").strip()
    ws["B8"] = ciclos[0]["inicio"]
    ws["B8"].number_format = "mm/yyyy"

    ws = wb["parametros"]
    for ciclo in ciclos:
        row = ciclo["numero"] + 2
        ws[f"A{row}"] = "Sim" if ciclo["computar"] else "Nao"
        ws[f"B{row}"] = ciclo["nome"]
        ws[f"C{row}"] = ciclo["inicio"]
        ws[f"D{row}"] = ciclo["fim"]
        ws[f"E{row}"] = ciclo["percentual"]
        ws[f"G{row}"] = ciclo["situacao"]
        for col in ("A", "B", "C", "D", "E", "G"):
            cell = ws[f"{col}{row}"]
            cell.fill = PREENCHIMENTO_AUTOMATICO
            font = copy(cell.font)
            font.color = COR_MARINHO if ciclo["computar"] else COR_TEXTO
            font.b = bool(ciclo["computar"])
            cell.font = font
        for col in ("C", "D"):
            ws[f"{col}{row}"].number_format = "mm/yyyy"
        ws[f"E{row}"].number_format = "0.00%"

    # C1-C4 ainda não alcançados permanecem estruturalmente presentes, mas vazios.
    ultimo = max(ciclo["numero"] for ciclo in ciclos)
    for numero in range(ultimo + 1, 5):
        row = numero + 2
        ws[f"A{row}"] = "Nao"
        ws[f"B{row}"] = f"C{numero}"
        for col in ("C", "D", "E"):
            ws[f"{col}{row}"] = None
        ws[f"G{row}"] = "Não aplicável"

    ws = wb["financeiro"]
    row = 2
    for ciclo in ciclos:
        for deslocamento in range(12):
            competencia = ciclo["inicio"] + relativedelta(months=deslocamento)
            ws[f"A{row}"] = competencia
            ws[f"A{row}"].number_format = "mm/yyyy"
            ws[f"C{row}"] = None
            financeiro_inicio = ciclo["financeiro_inicio"]
            efeito = bool(
                ciclo["computar"]
                and financeiro_inicio
                and _primeiro_dia_mes(competencia) >= _primeiro_dia_mes(financeiro_inicio)
            )
            ws[f"G{row}"] = "Sim" if efeito else "Nao"
            for col in ("A", "G"):
                cell = ws[f"{col}{row}"]
                cell.fill = PREENCHIMENTO_AUTOMATICO
                cell.font = Font(
                    name="Calibri",
                    size=10,
                    bold=bool(ciclo["computar"]),
                    color=COR_MARINHO if ciclo["computar"] else COR_TEXTO,
                )
            ws[f"C{row}"].fill = PREENCHIMENTO_ENTRADA
            row += 1
    while row <= 61:
        for col in ("A", "C", "G"):
            ws[f"{col}{row}"] = None
        ws[f"C{row}"].fill = PREENCHIMENTO_ENTRADA
        row += 1

    _normalizar_arquivo(wb)
    _validar_resultados_integra(wb, "imediatamente antes do salvamento")
    formulas_finais = _formulas(wb)
    if formulas_finais != formulas_originais:
        alteradas = sorted(set(formulas_originais) ^ set(formulas_finais))[:5]
        raise RuntimeError(f"A geração alterou a matriz de fórmulas: {alteradas}")

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def _celula_tem_observacao(valor: Any) -> bool:
    texto = _texto_sem_acento(valor).upper()
    return "OBSERVACAO" in texto or texto.strip() == "OBS"


def ler_coleta_reajuste(conteudo: bytes, *, contexto=None) -> dict[str, Any]:
    """Valida o XLS no upload sem recalcular nem substituir resultados do Excel.

    ``contexto`` (PERF-ARCH-1) e o contexto da execucao corrente do fluxo
    oficial: quando fornecido, as duas representacoes deste arquivo ja foram
    abertas e aprovadas na fronteira de geometria, e sao reaproveitadas em vez
    de repagar o parse. Sem contexto, a leitura isolada abre os proprios bytes
    exatamente como antes.
    """

    conteudo = garantir_xlsx_validado(conteudo)
    if contexto is not None:
        try:
            # O gate de geometria roda dentro do contexto, antes da entrega.
            wb = contexto.workbook_formulas
        except ErroSegurancaXlsx:
            # Erro de fronteira mantem a mensagem controlada da fronteira.
            raise
        except Exception as exc:
            raise ValueError("O arquivo não é um XLSX válido.") from exc
    else:
        try:
            wb = load_workbook(BytesIO(conteudo), data_only=False, read_only=False)
        except Exception as exc:
            raise ValueError("O arquivo não é um XLSX válido.") from exc

        # Fronteira de geometria: aprova o orçamento de varredura antes que
        # qualquer percurso de células aconteça, de modo que _formulas() e as
        # demais leituras trabalhem sempre dentro de um retângulo já validado.
        validar_geometria_workbook(wb)

    faltantes = [aba for aba in ABAS_OBRIGATORIAS_LEGADO if aba not in wb.sheetnames]
    proibidas = [aba for aba in ABAS_PROIBIDAS if aba in wb.sheetnames]
    bloqueios_estruturais: list[str] = []
    bloqueios_criticos: list[str] = []
    lacunas_apuracao: list[str] = []
    avisos: list[str] = []
    if faltantes:
        bloqueios_estruturais.append("Abas obrigatórias ausentes: " + ", ".join(faltantes))
    if proibidas:
        bloqueios_estruturais.append("Abas excluídas reapareceram: " + ", ".join(proibidas))
    if faltantes:
        capacidades = avaliar_capacidades_apuracao({}, {}, bloqueios_estruturais, [])
        return {
            "valido": False,
            "pronto_para_consolidar": False,
            "status_base": "ARQUIVO_ESTRUTURALMENTE_INVALIDO",
            "processamento_progressivo": True,
            "pendencias": bloqueios_estruturais,
            "bloqueios_estruturais": bloqueios_estruturais,
            "bloqueios_criticos": bloqueios_criticos,
            "inconsistencias": bloqueios_criticos,
            "lacunas_apuracao": [],
            "avisos": avisos,
            "contagens": {},
            "metadados": {},
            "capacidades": capacidades,
        }

    possui_posicao_contratual = "posicao_contratual" in wb.sheetnames
    if not possui_posicao_contratual:
        avisos.append(
            "Arquivo legado sem a camada posicao_contratual; quantidades por ciclo seguem o leiaute historico."
        )

    # Detecta modelo oficial: NUMERO_PC na coluna A de itens_PC desloca CICLO_PC para C2
    _ws_ipc = wb["itens_PC"] if "itens_PC" in wb.sheetnames else None
    _header_a1 = (_ws_ipc["A1"].value or "") if _ws_ipc is not None else ""
    _chave_ciclo_pc = "itens_PC!C2" if str(_header_a1).strip().upper() == "NUMERO_PC" else "itens_PC!B2"
    _modelo_pc_etapa3 = bool(
        _ws_ipc is not None
        and str(_ws_ipc["L1"].value or "").strip().upper()
        == "EFEITO_FINANCEIRO_PC"
    )

    formulas = _formulas(wb)
    if len(formulas) < 1000:
        bloqueios_estruturais.append("A matriz de fórmulas foi removida ou está incompleta.")
    aba_resultados_tecnicos = (
        "MEMORIA_RESULTADOS"
        if "MEMORIA_RESULTADOS" in wb.sheetnames
        else "RESULTADOS"
    )
    for chave in (
        "financeiro!D2",
        "itens_Remanesc!D2",
        "itens_Consumidos!O2",
        _chave_ciclo_pc,
        f"{aba_resultados_tecnicos}!B15",
        f"{aba_resultados_tecnicos}!B16",
        f"{aba_resultados_tecnicos}!B23",
        f"{aba_resultados_tecnicos}!B26",
        f"{aba_resultados_tecnicos}!B35",
        f"{aba_resultados_tecnicos}!C35",
        f"{aba_resultados_tecnicos}!D35",
        f"{aba_resultados_tecnicos}!F36",
    ):
        if chave not in formulas:
            bloqueios_estruturais.append(f"Fórmula estrutural ausente em {chave}.")
    if _modelo_pc_etapa3 and "itens_PC!L2" not in formulas:
        bloqueios_estruturais.append(
            "Formula estrutural ausente em itens_PC!L2."
        )
    if possui_posicao_contratual:
        for chave in (
            "aditivos!L2",
            "posicao_contratual!E2",
            "posicao_contratual!I2",
            "posicao_contratual!M2",
            "posicao_contratual!Q2",
            "posicao_contratual!U2",
            "posicao_contratual!X2",
            "itens_Remanesc!F2",
            "itens_RC!C3",
            "historico_VU!N2",
        ):
            if chave not in formulas:
                bloqueios_estruturais.append(f"Fórmula estrutural ausente em {chave}.")
    referencias_quebradas = [chave for chave, formula in formulas.items() if "#REF!" in formula.upper()]
    if referencias_quebradas:
        bloqueios_estruturais.append("Há fórmulas com referência quebrada: " + ", ".join(referencias_quebradas[:5]))

    nomes_definidos = set(wb.defined_names)
    nomes_ausentes = sorted(NOMES_RESULTADOS_OBRIGATORIOS - nomes_definidos)
    if nomes_ausentes:
        bloqueios_estruturais.append("Nomes estruturais da aba RESULTADOS ausentes: " + ", ".join(nomes_ausentes[:8]))
    if wb.sheetnames[-1] != "RESULTADOS":
        avisos.append("A aba RESULTADOS deve permanecer como a última aba do arquivo.")
    abas_coloridas = [ws.title for ws in wb.worksheets if ws.sheet_properties.tabColor is not None]
    if abas_coloridas != ["RESULTADOS"]:
        avisos.append("Somente a guia RESULTADOS deve possuir cor de aba.")

    comentarios = []
    observacoes = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                if cell.comment is not None:
                    comentarios.append(f"{ws.title}!{cell.coordinate}")
                if cell.row == 1 and _celula_tem_observacao(cell.value):
                    observacoes.append(f"{ws.title}!{cell.coordinate}")
    if comentarios:
        bloqueios_estruturais.append("Comentários/observações de célula não são admitidos: " + ", ".join(comentarios[:5]))
    if observacoes:
        bloqueios_estruturais.append("Campos de observação não são admitidos: " + ", ".join(observacoes[:5]))

    parametros = wb["parametros"]
    ativos = [numero for numero in range(1, 5) if str(parametros[f"A{numero + 2}"].value).strip().lower() == "sim"]
    if not ativos:
        lacunas_apuracao.append("Nenhum ciclo está marcado para computar nesta apuração.")
    else:
        for numero in range(1, max(ativos) + 1):
            if _numero(parametros[f"E{numero + 2}"].value) is None:
                lacunas_apuracao.append(f"C{numero}: percentual necessário ao acumulado está ausente.")

    # Etapa 3: fonte visivel e metadado sao copias da mesma data canonica.
    # Em legado, a inferencia Python so e aceita quando ha fonte confiavel.
    inicios_pc, erros_inicios_pc, tem_inicio_visivel, tem_inicio_metadado = (
        reconciliar_inicios_efeito(wb)
    )
    bloqueios_criticos.extend(
        f"Inicio dos efeitos inconsistente: {erro}" for erro in erros_inicios_pc
    )
    # Fonte visivel (parametros!INICIO_EFEITO_FINANCEIRO) e canonica. O metadado
    # CL8US_INICIO_EFEITO e apenas copia de integridade: se desaparecer apos o
    # fiscal abrir/salvar o XLS em editores legitimos, o arquivo continua valido.
    if tem_inicio_visivel and not tem_inicio_metadado:
        avisos.append(
            "Metadado tecnico CL8US_INICIO_EFEITO ausente; adotada a fonte "
            "visivel parametros!INICIO_EFEITO_FINANCEIRO como canonica."
        )
    ativos_nome = {f"C{numero}" for numero in ativos}
    ws_pc = wb["itens_PC"]
    layout_pc_numero = str(_header_a1).strip().upper() == "NUMERO_PC"
    col_numero_pc = "A" if layout_pc_numero else None
    col_data_pc = "B" if layout_pc_numero else "A"
    col_valor_pc = "D" if layout_pc_numero else "C"
    # 26G: capacidade excedida NUNCA e truncada em silencio — bloqueio explicito.
    if ws_pc.max_row > ULTIMA_LINHA_PCS:
        linhas_alem = sorted(
            {
                cell.row
                for row_alem in ws_pc.iter_rows(
                    min_row=ULTIMA_LINHA_PCS + 1, min_col=1, max_col=12
                )
                for cell in row_alem
                if cell.value not in (None, "")
            }
        )
        if linhas_alem:
            bloqueios_estruturais.append(
                f"Capacidade de PCs excedida: ha conteudo em itens_PC apos a "
                f"linha {ULTIMA_LINHA_PCS} (ex.: linha {linhas_alem[0]}). "
                f"Maximo suportado: {CAPACIDADE_PCS} PCs."
            )
    # 26G: duplicidade de NUMERO_PC verificada em TODA a base (TRIM+UPPER),
    # nao apenas nos primeiros 100 PCs.
    if col_numero_pc:
        _vistos: dict[str, int] = {}
        _duplicados: dict[str, list[int]] = {}
        for row in range(2, ULTIMA_LINHA_PCS + 1):
            bruto = ws_pc[f"{col_numero_pc}{row}"].value
            chave = str(bruto).strip().upper() if bruto not in (None, "") else ""
            if not chave:
                continue
            if chave in _vistos:
                _duplicados.setdefault(chave, [_vistos[chave]]).append(row)
            else:
                _vistos[chave] = row
        for chave, linhas in sorted(_duplicados.items())[:8]:
            bloqueios_criticos.append(
                f"NUMERO_PC duplicado: {chave} (linhas "
                + ", ".join(str(li) for li in linhas)
                + ")."
            )
        if len(_duplicados) > 8:
            bloqueios_criticos.append(
                f"Ha mais {len(_duplicados) - 8} NUMERO_PC duplicados em itens_PC."
            )
    tabela_ciclos_fin = _tabela_ciclos_financeiros(wb)
    for row in range(2, ULTIMA_LINHA_PCS + 1):
        if not any(
            ws_pc[f"{col}{row}"].value not in (None, "")
            for col in tuple(
                c for c in (col_numero_pc, col_data_pc, col_valor_pc) if c
            )
        ):
            continue
        numero_pc = str(
            (ws_pc[f"{col_numero_pc}{row}"].value if col_numero_pc else None)
            or f"linha {row}"
        ).strip()
        data_pc = _data(ws_pc[f"{col_data_pc}{row}"].value)
        if data_pc is None:
            bloqueios_criticos.append(
                f"DATA_PC vazia ou invalida: PC {numero_pc}, linha {row}."
            )
            continue
        ciclo_pc = _ciclo_por_competencia_financeira(wb, data_pc, tabela_ciclos_fin)
        if not ciclo_pc:
            bloqueios_criticos.append(
                f"Ciclo cronologico nao identificado: PC {numero_pc}, linha {row}."
            )
            continue
        linha_param = int(ciclo_pc[1:]) + 2
        reg_pc = {
            "computar_nesta_apuracao": parametros[f"A{linha_param}"].value,
            "inicio_efeito_financeiro": inicios_pc.get(ciclo_pc),
        }
        efeito_esperado = efeito_financeiro_pc(data_pc, ciclo_pc, reg_pc)
        ciclo_ativo = ciclo_pc in ativos_nome
        if ciclo_ativo and efeito_esperado is None:
            # Sem inicio de efeito para o ciclo: insuficiencia, nao inconsistencia.
            # O efeito daquele PC fica indeterminado; upload segue aceito e o
            # calculo dependente e barrado pela politica de capacidades/lacunas.
            lacunas_apuracao.append(
                "INICIO_EFEITO_FINANCEIRO ausente para ciclo ativo: "
                f"PC {numero_pc} - {ciclo_pc}. Efeito financeiro indeterminado "
                "ate a complementacao do inicio do efeito."
            )
            continue
        marcador = ws_pc[f"L{row}"].value if _modelo_pc_etapa3 else None
        marcador_formula = isinstance(marcador, str) and marcador.startswith("=")
        if _modelo_pc_etapa3 and not marcador_formula:
            bloqueios_criticos.append(
                f"Marcador EFEITO_FINANCEIRO_PC ausente: PC {numero_pc} - {ciclo_pc}."
            )
        elif marcador not in (None, "") and not marcador_formula:
            if str(marcador).strip() not in {"Sim", "Nao"}:
                bloqueios_criticos.append(
                    f"EFEITO_FINANCEIRO_PC invalido: PC {numero_pc} - {ciclo_pc}."
                )
            elif efeito_esperado is not None and str(marcador).strip() != efeito_esperado:
                bloqueios_criticos.append(
                    "EFEITO_FINANCEIRO_PC divergente da data canonica: "
                    f"PC {numero_pc} - {ciclo_pc}."
                )
        # NOTA: a exigencia historica de "fonte dupla" (visivel + metadado) foi
        # removida. O metadado tecnico pode ser legitimamente apagado por editores
        # de XLSX; a fonte visivel em parametros e suficiente e canonica.

    financeiro = wb["financeiro"]
    # EF-G1 — REGRA PETREA: `financeiro!G` e a decisao canonica do efeito
    # financeiro da competencia. "Sim" e "Nao" sao aceitos em silencio: o leitor
    # NAO reconstroi a marcacao "esperada" por parametros/metadados/cronologia e
    # NAO infere se a marcacao foi automatica ou manual. Restam inconsistencias
    # reais de preenchimento (vazio e valor invalido), agrupadas numa unica
    # mensagem por classe. G nao decide a existencia da competencia no VTA:
    # decide apenas se ela recebe o efeito financeiro do reajuste.
    marco_financeiro = _marco_grade_financeira(wb)
    efeitos_invalidos: list[str] = []
    efeitos_nao_informados: list[str] = []
    for row in range(2, 74):
        competencia = financeiro[f"A{row}"].value
        valor = _numero(financeiro[f"C{row}"].value)
        efeito = str(financeiro[f"G{row}"].value or "")
        ciclo = _ciclo_cronologico_financeiro(marco_financeiro, competencia)
        data_comp = _data(competencia)
        referencia = (
            f"{ciclo or 'ciclo nao identificado'} — "
            + (data_comp.strftime("%m/%Y") if data_comp else f"linha {row}")
        )
        if valor is not None and competencia not in (None, "") and data_comp is None:
            bloqueios_criticos.append(
                f"Competencia invalida na aba financeiro: linha {row}. "
                "Informe uma data mensal valida."
            )
            continue
        if efeito not in ("", "Sim", "Nao"):
            efeitos_invalidos.append(referencia)
            continue
        if valor is not None and not efeito:
            efeitos_nao_informados.append(referencia)
            continue
    if efeitos_nao_informados:
        bloqueios_criticos.append(
            "Há competências com o campo 'Efeito financeiro' não preenchido na "
            "aba financeiro. Corrija a coluna G antes de prosseguir: "
            + _lista_compacta_competencias(efeitos_nao_informados)
        )
    if efeitos_invalidos:
        bloqueios_criticos.append(
            "Há competências com valor inválido no campo 'Efeito financeiro' da "
            "aba financeiro. Use o dropdown da coluna G e selecione exatamente "
            "Sim ou Nao: " + _lista_compacta_competencias(efeitos_invalidos)
        )

    contagens = {
        "competencias_com_valor": sum(1 for row in range(2, 74) if _numero(wb["financeiro"][f"C{row}"].value) is not None),
        "itens_remanescentes": sum(1 for row in range(2, 201) if wb["itens_Remanesc"][f"A{row}"].value not in (None, "")),
        "itens_consumidos": sum(1 for row in range(2, 201) if wb["itens_Consumidos"][f"A{row}"].value not in (None, "")),
        "pedidos_de_compra": sum(1 for row in range(2, ULTIMA_LINHA_PCS + 1) if wb["itens_PC"][f"A{row}"].value not in (None, "")),
        "aditivos": sum(1 for row in range(2, 201) if wb["aditivos"][f"A{row}"].value not in (None, "")),
        "formulas": len(formulas),
        "posicao_contratual_itens": 0,
        "posicao_contratual_calculada": 0,
        "historico_vu_itens": 0,
        "historico_vu_calculado": 0,
    }
    if contagens["competencias_com_valor"] == 0 and contagens["itens_remanescentes"] == 0:
        avisos.append("Ainda não há valores mensais nem itens remanescentes preenchidos pelo fiscal.")

    status_resultados: dict[str, Any] = {}
    try:
        # PERF-ARCH-1: a representacao data_only deste mesmo arquivo ja existe
        # na execucao oficial (o leitor masterfile a abriu e o contexto a
        # guarda). Reabri-la aqui repagava o parse XML integral so para varrer
        # erros de calculo em cache.
        wb_valores = (
            contexto.workbook_valores
            if contexto is not None
            else load_workbook(BytesIO(conteudo), data_only=True, read_only=True)
        )
        erros = []
        for ws in wb_valores.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.upper() in ERROS_EXCEL:
                        erros.append(f"{ws.title}!{cell.coordinate}={cell.value}")
        if erros:
            lacunas_apuracao.append("O Excel salvou erros de cálculo: " + ", ".join(erros[:8]))
        if possui_posicao_contratual:
            contagens["posicao_contratual_itens"] = sum(
                1 for row in range(2, 201)
                if wb_valores["posicao_contratual"][f"A{row}"].value not in (None, "")
            )
            contagens["posicao_contratual_calculada"] = sum(
                1 for row in range(2, 201)
                if wb_valores["posicao_contratual"][f"A{row}"].value not in (None, "")
                and any(
                    _numero(wb_valores["posicao_contratual"][f"{col}{row}"].value) is not None
                    for col in ("E", "G", "I", "K", "M", "O", "Q", "S", "U", "W")
                )
            )
        contagens["historico_vu_itens"] = sum(
            1 for row in range(2, 201)
            if wb_valores["historico_VU"][f"A{row}"].value not in (None, "")
        )
        contagens["historico_vu_calculado"] = sum(
            1 for row in range(2, 201)
            if wb_valores["historico_VU"][f"A{row}"].value not in (None, "")
            and any(
                _numero(wb_valores["historico_VU"][f"{col}{row}"].value) is not None
                for col in ("B", "D", "F", "H", "J", "L", "N", "P")
            )
        )
        if possui_posicao_contratual:
            alertas_aditivos = [
                f"aditivos!M{row}={wb_valores['aditivos'][f'M{row}'].value}"
                for row in range(2, 201)
                if str(wb_valores["aditivos"][f"M{row}"].value or "").startswith("ALERTA:")
            ]
            alertas_posicao = [
                f"posicao_contratual!X{row}={wb_valores['posicao_contratual'][f'X{row}'].value}"
                for row in range(2, 201)
                if str(wb_valores["posicao_contratual"][f"X{row}"].value or "").startswith("ALERTA:")
            ]
            if alertas_aditivos:
                bloqueios_criticos.append(
                    "Aditivos quantitativos inconsistentes: " + ", ".join(alertas_aditivos[:5])
                )
            if alertas_posicao:
                bloqueios_criticos.append(
                    "Posição contratual inconsistente: " + ", ".join(alertas_posicao[:5])
                )
        resultados_valores = wb_valores[aba_resultados_tecnicos]
        resultados_executivos = wb_valores["RESULTADOS"]
        status_resultados = {
            "geral": (
                resultados_executivos["B3"].value
                if "MEMORIA_RESULTADOS" in wb_valores.sheetnames
                else resultados_valores["J4"].value
            ),
            "metodo_retroativo": resultados_valores["B4"].value,
            "origem_retroativo_oficial": resultados_valores["D16"].value,
            "retroativo": resultados_valores["F16"].value,
            "vta": resultados_valores["E26"].value,
            "remanescente": resultados_valores["F36"].value,
            "valores": {
                "retroativo_financeiro": resultados_valores["B15"].value,
                "retroativo_pc": resultados_valores["C15"].value,
                "retroativo_itens": resultados_valores["D15"].value,
                "retroativo_oficial": resultados_valores["B16"].value,
                "vta_base_contratual": resultados_valores["B20"].value,
                "vta_retroativo": resultados_valores["B21"].value,
                "vta_ajuste_remanescente": resultados_valores["B22"].value,
                "vta_calculado": resultados_valores["B23"].value,
                "vta_pc_execucao_anterior": resultados_valores["T21"].value,
                "vta_pc_parcelas_intermediarias": resultados_valores["T22"].value,
                "vta_pc_remanescente_corte": resultados_valores["T23"].value,
                "vta_pc_total": resultados_valores["T25"].value,
                "vta_ajuste_manual": resultados_valores["B24"].value,
                "vta_manual_oficial": resultados_valores["B25"].value,
                "vta_oficial": resultados_valores["B26"].value,
                "quantidade_remanescente": resultados_valores["B35"].value,
                "remanescente_original": resultados_valores["C35"].value,
                "remanescente_atualizado": resultados_valores["D35"].value,
            },
        }
        if not status_resultados["geral"]:
            avisos.append(
                "Os status de RESULTADOS não estão calculados em cache; abra, recalcule e salve o XLS no Excel."
            )
    except Exception:
        avisos.append("Não foi possível conferir os valores calculados em cache; abra e salve o arquivo no Excel.")

    metadados = {
        "indice": wb["CONTROLE"]["B7"].value,
        "ciclo_vigente": wb["CONTROLE"]["B2"].value,
        "data_corte": wb["CONTROLE"]["B3"].value,
        "ciclos_em_analise": [f"C{numero}" for numero in ativos],
        "status_resultados": status_resultados,
        "arquitetura_posicao_contratual": "canonica" if possui_posicao_contratual else "legada",
    }
    # HARD BLOCK: somente falhas estruturais tornam o arquivo invalido e barram
    # a leitura dos blocos. SOFT BLOCK: inconsistencias de negocio e lacunas de
    # informacao NAO rejeitam o upload; barram apenas o bloco dependente e a
    # formalizacao. Por isso somente bloqueios_estruturais alimentam capacidades.
    inconsistencias = bloqueios_criticos
    capacidades = avaliar_capacidades_apuracao(
        contagens,
        metadados,
        bloqueios_estruturais,
        lacunas_apuracao,
    )
    possui_base = capacidades["resumo"]["tem_alguma_evidencia"]
    resultados_seguros = capacidades["resumo"]["apuracao_integral"]
    estruturalmente_invalido = bool(bloqueios_estruturais)
    tem_inconsistencias = bool(inconsistencias)
    tem_insuficiencia = bool(lacunas_apuracao) or not resultados_seguros
    if estruturalmente_invalido:
        status_base = "ARQUIVO_ESTRUTURALMENTE_INVALIDO"
    elif tem_inconsistencias:
        status_base = "ANALISE_COM_INCONSISTENCIAS"
    elif tem_insuficiencia or not possui_base:
        status_base = "ANALISE_PARCIAL_INFORMACOES_INSUFICIENTES"
    else:
        status_base = "APTO_PARA_ANALISE"
    # Formalizacao permanece protegida: exige estrutura integra, ausencia de
    # inconsistencias, base presente e apuracao integral.
    formalizacao_liberada = (
        not estruturalmente_invalido
        and not tem_inconsistencias
        and possui_base
        and resultados_seguros
    )
    pendencias = bloqueios_estruturais + inconsistencias + lacunas_apuracao
    return {
        "valido": not estruturalmente_invalido,
        "pronto_para_consolidar": formalizacao_liberada,
        "status_base": status_base,
        "processamento_progressivo": True,
        "pendencias": pendencias,
        "bloqueios_estruturais": bloqueios_estruturais,
        "bloqueios_criticos": bloqueios_criticos,
        "inconsistencias": inconsistencias,
        "lacunas_apuracao": lacunas_apuracao,
        "avisos": avisos,
        "contagens": contagens,
        "metadados": metadados,
        "capacidades": capacidades,
    }


def eh_coleta_reajuste(conteudo: bytes) -> bool:
    try:
        conteudo = garantir_xlsx_validado(conteudo)
        wb = load_workbook(BytesIO(conteudo), read_only=True, data_only=False)
        nomes = set(wb.sheetnames)
        nucleares = {"CONTROLE", "parametros", "financeiro"}
        # Também reconhece uma coleta canônica danificada, para que o validador
        # possa explicar a aba ausente em vez de desviá-la ao leitor legado.
        return nucleares.issubset(nomes) and len(nomes.intersection(ABAS_OBRIGATORIAS_LEGADO)) >= 5
    except Exception:
        return False
