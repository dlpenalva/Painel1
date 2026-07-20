from __future__ import annotations

import re
import zipfile as _zipfile
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path
from typing import Any
from dateutil.relativedelta import relativedelta

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# v10.5.2: _FILL_GAP (rosa FFC7CE) REMOVIDO — rosa em financeiro A7:G e bug
# proibido (skills/masterfile_xls_rules.md). Gap financeiro segue marcado
# apenas por G="Não", sem destaque de cor.
# Laranja Ênfase 6 Mais Escuro 25% — F column para ciclos fora da apuração (v10.3.14)
_FILL_FORA = PatternFill("solid", fgColor="FFC55A11")

# ---------------------------------------------------------------------------
# Corretor de namespace VML
# ---------------------------------------------------------------------------
# openpyxl gera .vml com prefixos dinamicos (ns0, ns1...) redeclarados
# mid-document com URIs diferentes. Excel emite aviso de reparo nesse caso.
# Normaliza para v:/o:/x: com declaracao unica na raiz <xml>.

_VML_ROOT_NS = (
    '<xml xmlns:v="urn:schemas-microsoft-com:vml"'
    ' xmlns:o="urn:schemas-microsoft-com:office:office"'
    ' xmlns:x="urn:schemas-microsoft-com:office:excel">'
)
_URI_TO_STD = {
    "urn:schemas-microsoft-com:vml": "v",
    "urn:schemas-microsoft-com:office:office": "o",
    "urn:schemas-microsoft-com:office:excel": "x",
}


def _corrigir_vml(texto: str) -> str:
    ns_map = {m.group(1): m.group(2) for m in re.finditer(r'xmlns:(ns\d+)="([^"]+)"', texto)}
    prefix_map = {p: _URI_TO_STD[u] for p, u in ns_map.items() if u in _URI_TO_STD}
    if not prefix_map:
        return texto
    resultado = re.sub(r'\s+xmlns:ns\d+="[^"]+"', "", texto)
    for old, new in sorted(prefix_map.items(), key=lambda x: -len(x[0])):
        resultado = re.sub(rf"(?<![A-Za-z0-9_]){re.escape(old)}:", f"{new}:", resultado)
    resultado = re.sub(r"^<xml>", _VML_ROOT_NS, resultado, count=1)
    return resultado


def _corrigir_vml_xlsx(conteudo: bytes) -> bytes:
    """Corrige prefixos VML em todas as entradas .vml de um XLSX."""
    entrada = BytesIO(conteudo)
    saida = BytesIO()
    with _zipfile.ZipFile(entrada, "r") as zin, \
         _zipfile.ZipFile(saida, "w", compression=_zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            dados = zin.read(item.filename)
            if item.filename.endswith(".vml"):
                try:
                    dados = _corrigir_vml(dados.decode("utf-8")).encode("utf-8")
                except Exception:
                    pass
            zout.writestr(item, dados)
    return saida.getvalue()

from _masterfile_config import ABAS_OBRIGATORIAS

# Aliases históricos: nome antigo (lower) → nome canônico atual (lower)
# Permite aceitar workbooks com nomes legados sem quebrar a geração.
_ALIAS_ABAS_LOWER: dict[str, str] = {
    "itens_a": "itens_remanesc",
    "itens_b": "itens_consumidos",
    "aditivos": "aditivos",  # normaliza caixa (era ADITIVOS antes de jun/2026)
}

def _corrigir_selection_pane_xlsx(conteudo: bytes) -> bytes:
    # Bug openpyxl: freeze_panes=None remove pane elem mas deixa pane= em selection
    import xml.etree.ElementTree as _ET
    NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    TV = "{" + NS + "}sheetView"
    TP = "{" + NS + "}pane"
    TS = "{" + NS + "}selection"

    def _fix(dados):
        txt = dados.decode("utf-8")
        if "pane=" not in txt:
            return dados
        try:
            root = _ET.fromstring(txt)
            ok = False
            for sv in root.iter(TV):
                if sv.find(TP) is None:
                    for sel in sv.findall(TS):
                        if "pane" in sel.attrib:
                            del sel.attrib["pane"]
                            ok = True
            if not ok:
                return dados
            _ET.register_namespace("", NS)
            return _ET.tostring(root, encoding="unicode", xml_declaration=False).encode("utf-8")
        except Exception:
            return dados

    entrada = BytesIO(conteudo)
    saida = BytesIO()
    with _zipfile.ZipFile(entrada, "r") as zi, _zipfile.ZipFile(saida, "w", compression=_zipfile.ZIP_DEFLATED) as zo:
        for item in zi.infolist():
            d = zi.read(item.filename)
            if item.filename.startswith("xl/worksheets/") and item.filename.endswith(".xml"):
                d = _fix(d)
            zo.writestr(item, d)
    return saida.getvalue()



def _localizar_aba(wb, nome: str) -> str | None:
    """Localiza aba no workbook por nome exato, case-insensitive ou alias legado.

    Retorna o nome real da aba no workbook, ou None se não encontrada.
    Garante que renomeações físicas (ex.: ADITIVOS→aditivos) não quebrem a geração.
    """
    if nome in wb.sheetnames:
        return nome
    lower = nome.lower()
    for aba in wb.sheetnames:
        if aba.lower() == lower:
            return aba
    alvo_lower = _ALIAS_ABAS_LOWER.get(lower)
    if alvo_lower:
        for aba in wb.sheetnames:
            if aba.lower() == alvo_lower:
                return aba
    return None


def obter_masterfile_padrao(caminho: str | Path) -> bytes:
    caminho = Path(caminho)
    if not caminho.exists():
        raise FileNotFoundError(f"Template não encontrado: {caminho}")
    return caminho.read_bytes()


def _formulas(wb) -> dict[str, str]:
    return {
        f"{ws.title}!{cell.coordinate}": cell.value
        for ws in wb.worksheets
        for row in ws.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    }


def _escrever_entrada(ws, coordenada: str, valor: Any) -> None:
    cell = ws[coordenada]
    if isinstance(cell.value, str) and cell.value.startswith("="):
        raise ValueError(f"Tentativa de sobrescrever fórmula em {ws.title}!{coordenada}")
    if isinstance(valor, date) and not isinstance(valor, datetime):
        valor = datetime.combine(valor, datetime.min.time())
    cell.value = valor



def _completar_periodos_ciclos(ciclos: dict[str, Any], data_corte=None) -> dict[str, Any]:
    """Garante que C0-C4 existam temporalmente, sem pular nem compactar ciclo (v10.5.2).

    COMPUTAR_NESTA_APURACAO = Nao (ou ciclo ausente na Calculadora) NAO apaga o
    ciclo da linha do tempo: o periodo faltante e derivado dos vizinhos —
    data_inicio = data_fim do ciclo anterior + 1 dia; data_fim = data_inicio do
    proximo ciclo conhecido - 1 dia (ou +12 meses quando nao ha proximo).
    O ciclo seguinte nunca herda o periodo do ciclo nao computado.
    Ciclos derivados ficam com possui_efeito_financeiro="Não".
    """
    nomes = ("C0", "C1", "C2", "C3", "C4")
    completos: dict[str, Any] = {}
    for nome in nomes:
        c = dict(ciclos.get(nome) or {})
        c.setdefault("ciclo", nome)
        completos[nome] = c

    # forward: deriva data_inicio a partir do fim do anterior
    for i, nome in enumerate(nomes):
        c = completos[nome]
        if not c.get("data_inicio") and i > 0:
            fim_ant = completos[nomes[i - 1]].get("data_fim")
            if fim_ant:
                c["data_inicio"] = fim_ant + timedelta(days=1)
                c["_derivado"] = True
        if not c.get("data_fim") and c.get("data_inicio"):
            prox_ini = None
            for j in range(i + 1, len(nomes)):
                prox_ini = (ciclos.get(nomes[j]) or {}).get("data_inicio")
                if prox_ini:
                    break
            if prox_ini and prox_ini > c["data_inicio"]:
                c["data_fim"] = prox_ini - timedelta(days=1)
            else:
                c["data_fim"] = c["data_inicio"] + relativedelta(months=12) - timedelta(days=1)
            c["_derivado"] = True

    # backward: deriva data_inicio/fim de ciclos iniciais ausentes (ex.: C0)
    for i in range(len(nomes) - 2, -1, -1):
        c = completos[nomes[i]]
        prox = completos[nomes[i + 1]]
        if not c.get("data_fim") and prox.get("data_inicio"):
            c["data_fim"] = prox["data_inicio"] - timedelta(days=1)
            c["_derivado"] = True
        if not c.get("data_inicio") and c.get("data_fim"):
            c["data_inicio"] = c["data_fim"] - relativedelta(months=12) + timedelta(days=1)
            c["_derivado"] = True

    for nome in nomes:
        c = completos[nome]
        if c.get("_derivado") and not c.get("possui_efeito_financeiro"):
            c["possui_efeito_financeiro"] = "Não"
    return completos


def _capacidade_financeiro(ws) -> int:
    """Capacidade real da grade do financeiro no template carregado.

    Conta as linhas com formula de CICLO na coluna B (a grade util do
    template); a capacidade e o total de competencias que cabem sem truncar.
    Fallback conservador: 72 (grade oficial A2:A73).
    """
    ultima = 1
    for r in range(2, min(ws.max_row, 500) + 1):
        v = ws.cell(r, 2).value
        if isinstance(v, str) and v.startswith("="):
            ultima = r
    return (ultima - 1) if ultima > 1 else 72


def _preencher_financeiro(ws, ciclos: dict[str, Any], data_corte_fallback=None, v102: bool = False, novo: bool = False, marco_inicial=None) -> None:
    """Preenche a aba FINANCEIRO com ciclos e competências.

    Colunas A (Competência), B (Ciclo) e G (EFEITO_FINANCEIRO) são preenchidas
    automaticamente. Coluna C (Valor liquidado/pago) fica vazia para o usuário.
    Colunas D, E, F são preenchidas pelo usuário no Excel.

    Gap financeiro (v10.3.3): competências anteriores ao mês do pedido efetivo
    recebem G="Não" e destaque visual, mesmo que o ciclo seja computável.
    Isso impede que essas competências somem delta retroativo.

    data_corte_fallback: usada como data_fim quando o ciclo não informa data_fim.
    v102: se True, preserva linhas 2-6 (agregados + separador); competências
          mensais iniciam a partir da linha 7.
    """
    if novo:
        # Novo modelo oficial: competencias mensais consecutivas do MES do
        # marco inicial (inicio de C0) ate o MES da data de corte — todos os
        # ciclos existentes, de C0 ao vigente, ficam representados, nenhum
        # ciclo anterior e omitido e nenhuma competencia posterior ao corte
        # e inventada (nao ha obrigacao de 60 meses). Valor interno = dia 1,
        # exibicao mm/aaaa. Coluna B (CICLO) e formula do template — nunca
        # escrita. C (valor pago) e do fiscal; G recebe a decisao inicial da
        # Calculadora e permanece editavel pelo fiscal.
        marco = marco_inicial
        if isinstance(marco, datetime):
            marco = marco.date()
        corte = data_corte_fallback
        if isinstance(corte, datetime):
            corte = corte.date()
        for linha in range(2, 74):
            _escrever_entrada(ws, f"A{linha}", None)
            _escrever_entrada(ws, f"C{linha}", None)
            _escrever_entrada(ws, f"G{linha}", None)
        if not isinstance(marco, date) or not isinstance(corte, date):
            return  # marco/corte vazio ou invalido: nao inventa competencias
        competencia = marco.replace(day=1)
        limite = corte.replace(day=1)
        necessarias = (
            (limite.year - competencia.year) * 12
            + (limite.month - competencia.month) + 1
        )
        capacidade = _capacidade_financeiro(ws)
        if necessarias > capacidade:
            # Nunca truncar competencias em silencio nem entregar XLS parcial.
            raise ValueError(
                f"O periodo financeiro exige {necessarias} competencias, mas o "
                f"XLS comporta no maximo {capacidade}. Revise o marco inicial "
                "ou a data de corte antes de gerar o arquivo."
            )
        linha = 2
        while competencia <= limite:
            _escrever_entrada(ws, f"A{linha}", competencia)
            ws[f"A{linha}"].number_format = "MM/YYYY"
            ciclo_competencia = None
            for nome in ("C0", "C1", "C2", "C3", "C4"):
                candidato = ciclos.get(nome) or {}
                inicio = candidato.get("data_inicio")
                fim = candidato.get("data_fim")
                if isinstance(inicio, datetime):
                    inicio = inicio.date()
                if isinstance(fim, datetime):
                    fim = fim.date()
                if isinstance(inicio, date) and isinstance(fim, date) and inicio <= competencia <= fim:
                    ciclo_competencia = candidato
                    break
            efeito = "Nao"
            if ciclo_competencia:
                computavel = str(
                    ciclo_competencia.get("possui_efeito_financeiro") or ""
                ).upper() in ("SIM", "S")
                inicio_efeito = ciclo_competencia.get("inicio_efeito_financeiro")
                if isinstance(inicio_efeito, datetime):
                    inicio_efeito = inicio_efeito.date()
                if computavel and isinstance(inicio_efeito, date):
                    efeito = (
                        "Sim"
                        if (competencia.year, competencia.month)
                        >= (inicio_efeito.year, inicio_efeito.month)
                        else "Nao"
                    )
            _escrever_entrada(ws, f"G{linha}", efeito)
            competencia = competencia + relativedelta(months=1)
            linha += 1
        return

    linha_escrita = 7 if v102 else 2
    linha_max = 200
    meses_adicionados = set()

    for ciclo_nome in ("C1", "C2", "C3", "C4"):
        ciclo = ciclos.get(ciclo_nome)
        if not ciclo:
            continue

        data_inicio = ciclo.get("data_inicio")
        data_fim = ciclo.get("data_fim")

        if not data_fim and data_corte_fallback:
            data_fim = data_corte_fallback

        if not data_inicio or not data_fim:
            continue

        # Determina se o ciclo é computável (possui_efeito_financeiro = Sim)
        possui_efeito = str(ciclo.get("possui_efeito_financeiro") or "").upper() in ("SIM", "S")
        # Pedido efetivo: mês a partir do qual o efeito financeiro é válido
        data_pedido_ciclo = ciclo.get("data_pedido")

        data_atual = data_inicio
        while data_atual <= data_fim:
            chave_mes = (data_atual.year, data_atual.month)

            if chave_mes not in meses_adicionados and linha_escrita <= linha_max:
                _escrever_entrada(ws, f"A{linha_escrita}", data_atual)
                ws[f"A{linha_escrita}"].number_format = "MM/YYYY"
                if not novo:
                    _escrever_entrada(ws, f"B{linha_escrita}", ciclo_nome.lower())
                _escrever_entrada(ws, f"C{linha_escrita}", None)

                # --- Gap financeiro v10.3.3 ---
                # G = "Não" quando:
                #   1. Ciclo não computável (Precluso, Adiantado sem gate, etc.)
                #   2. Competência anterior ao mês do pedido (gap entre apta e pedido)
                if not possui_efeito:
                    efeito_g = "Não"
                    em_gap = False
                elif (
                    data_pedido_ciclo is not None
                    and (data_atual.year, data_atual.month)
                    < (data_pedido_ciclo.year, data_pedido_ciclo.month)
                ):
                    efeito_g = "Não"
                    em_gap = True
                else:
                    efeito_g = "Sim"
                    em_gap = False

                _escrever_entrada(ws, f"G{linha_escrita}", efeito_g)
                # v10.5.2: sem fill rosa no gap — apenas G="Não" (em_gap ja aplicado acima)

                meses_adicionados.add(chave_mes)
                linha_escrita += 1

            data_atual = data_atual.replace(day=1) + relativedelta(months=1)

    # Limpar linhas não utilizadas até o limite da grade (B só no layout antigo)
    for linha in range(linha_escrita, linha_max + 1):
        _escrever_entrada(ws, f"A{linha}", None)
        if not novo:
            _escrever_entrada(ws, f"B{linha}", None)
        _escrever_entrada(ws, f"C{linha}", None)
        _escrever_entrada(ws, f"G{linha}", None)


def _registrar_inicio_efeitos_financeiros(wb, ciclos: dict[str, Any]) -> None:
    """Persiste a mesma data na fonte visivel e na copia de integridade."""
    registros = []
    ws_parametros = wb["parametros"] if "parametros" in wb.sheetnames else None
    layout_oficial = bool(
        ws_parametros is not None
        and str(ws_parametros["B1"].value or "").strip().upper() == "CICLO"
    )
    if layout_oficial:
        ws_parametros["H1"] = "INICIO_EFEITO_FINANCEIRO"
    for linha, nome in enumerate(("C0", "C1", "C2", "C3", "C4"), start=2):
        inicio = (ciclos.get(nome) or {}).get("inicio_efeito_financeiro")
        if isinstance(inicio, datetime):
            inicio = inicio.date()
        if layout_oficial:
            _escrever_entrada(
                ws_parametros,
                f"H{linha}",
                inicio if isinstance(inicio, date) else None,
            )
            ws_parametros[f"H{linha}"].number_format = "dd/mm/yyyy"
        if isinstance(inicio, date):
            registros.append(f"{nome}={inicio.isoformat()}")
    anterior = str(wb.properties.keywords or "")
    anterior = ";".join(
        parte for parte in anterior.split(";")
        if parte and not parte.startswith("CL8US_INICIO_EFEITO:")
    )
    novo = "CL8US_INICIO_EFEITO:" + ",".join(registros)
    wb.properties.keywords = ";".join(parte for parte in (anterior, novo) if parte)


def _preencher_datas_fotografias_remanescentes(
    ws, ciclos: dict[str, Any], data_corte: Any, ciclo_vigente: str,
) -> None:
    """Carimba datas dos ciclos; o fiscal informa somente item e quantidades."""
    for r in range(2, 201):
        for nome, col in zip(("C0", "C1", "C2", "C3", "C4"), (2, 4, 6, 8, 10)):
            _escrever_entrada(ws, f"{get_column_letter(col)}{r}", (ciclos.get(nome) or {}).get("data_inicio"))
            ws.cell(r, col).number_format = "dd/mm/yyyy"
        _escrever_entrada(ws, f"L{r}", data_corte)
        ws.cell(r, 12).number_format = "dd/mm/yyyy"
        _escrever_entrada(ws, f"M{r}", ciclo_vigente)


def _preencher_memoria_fator(ws_par, ciclos: dict) -> None:
    """Tabela MEMORIA DO FATOR APLICAVEL A APURACAO abaixo da tabela principal (v10.3.14).

    Linha 9: título; linha 10: cabeçalhos; linhas 11-15: C0-C4; linha 16: total.
    Mostra apenas ciclos Sim no produto acumulado; ciclos Nao não entram.
    """
    NOMES = ("C0", "C1", "C2", "C3", "C4")
    PAR_ROW = {n: 2 + i for i, n in enumerate(NOMES)}

    def _computar(nome: str) -> str:
        return str(ws_par.cell(PAR_ROW[nome], 1).value or "")

    def _pct(nome: str):
        v = ws_par.cell(PAR_ROW[nome], 6).value
        return v if isinstance(v, (int, float)) else None

    R_TITLE = 9
    R_HDR = 10
    R_C0 = 11

    _fill_hdr  = PatternFill("solid", fgColor="FF1F4E79")
    _fill_sim  = PatternFill("solid", fgColor="FFE2EFDA")
    _fill_nao  = PatternFill("solid", fgColor="FFEDEDED")
    _fill_sum  = PatternFill("solid", fgColor="FFD9E1F2")
    _f_hdr     = Font(name="Calibri", size=9, bold=True, color="FFFFFFFF")
    _f_title   = Font(name="Calibri", size=10, bold=True, color="FF1F4E79")
    _f_data    = Font(name="Calibri", size=9)
    _f_sum     = Font(name="Calibri", size=9, bold=True)
    _ctr       = Alignment(horizontal="center", vertical="center")
    _thin      = Side(border_style="thin", color="FF000000")
    _borda     = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

    HDRS = ["CICLO", "COMPUTA?", "PERCENTUAL", "FATOR", "FATOR_ACUMULADO_APURACAO", "STATUS"]

    # Title
    ws_par.cell(R_TITLE, 1).value = "MEMORIA DO FATOR APLICAVEL A APURACAO"
    ws_par.cell(R_TITLE, 1).font = _f_title
    ws_par.merge_cells(f"A{R_TITLE}:F{R_TITLE}")

    # Headers
    for ci, h in enumerate(HDRS, 1):
        c = ws_par.cell(R_HDR, ci)
        c.value = h
        c.font = _f_hdr
        c.fill = _fill_hdr
        c.alignment = _ctr
        c.border = _borda

    # Data rows
    acum = 1.0
    acum_valido = True
    for i, nome in enumerate(NOMES):
        r = R_C0 + i
        computar = _computar(nome)
        pct = _pct(nome)
        is_sim = computar == "Sim"

        if nome == "C0":
            fator_val = 1.0
            acum = 1.0
            obs = "Base"
        elif is_sim and pct is not None:
            fator_val = round(1.0 + pct, 10)
            if acum_valido:
                acum = round(acum * fator_val, 10)
            obs = "Aplicado"
        elif is_sim:
            fator_val = None
            acum_valido = False
            obs = "Aguardando preenchimento manual"
        else:
            fator_val = None
            obs = "Fora da apuracao"

        row_fill = _fill_sim if is_sim else _fill_nao
        vals = [
            nome,
            computar or "Nao",
            pct if is_sim else None,
            fator_val,
            acum if acum_valido else None,
            obs,
        ]
        for ci, v in enumerate(vals, 1):
            c = ws_par.cell(r, ci)
            c.value = v
            c.font = _f_data
            c.fill = row_fill
            c.border = _borda
            c.alignment = _ctr
        ws_par.cell(r, 3).number_format = "0.00%"
        ws_par.cell(r, 4).number_format = "0.0000"
        ws_par.cell(r, 5).number_format = "0.0000"

    # Summary row
    r_sum = R_C0 + len(NOMES)
    variacao = (acum - 1.0) if acum_valido else None
    for ci in range(1, 7):
        c = ws_par.cell(r_sum, ci)
        c.fill = _fill_sum
        c.font = _f_sum
        c.border = _borda
        c.alignment = _ctr
    ws_par.cell(r_sum, 1).value = "TOTAL APURACAO"
    ws_par.cell(r_sum, 5).value = acum if acum_valido else None
    ws_par.cell(r_sum, 5).number_format = "0.0000"
    ws_par.cell(r_sum, 6).value = variacao
    ws_par.cell(r_sum, 6).number_format = "0.00%"


def gerar_masterfile_preenchido(
    dados_calculadora: dict[str, Any],
    template_path: str | Path | bytes,
) -> bytes:
    """Preenche entradas do Masterfile (v9/v10.1/v10.2) e preserva todas as fórmulas.

    template_path pode ser um caminho (str/Path) ou bytes do template já carregado.
    A versão v10.2 é detectada automaticamente via cabeçalho da aba parametros.
    """
    if not dados_calculadora or not dados_calculadora.get("ok"):
        faltantes = ", ".join(dados_calculadora.get("campos_ausentes", [])) if dados_calculadora else "dados"
        raise ValueError(f"Dados insuficientes para gerar o Masterfile preenchido: {faltantes or 'verifique os avisos'}")

    if isinstance(template_path, BytesIO):
        template_path.seek(0)
        wb = load_workbook(template_path, data_only=False)
    elif isinstance(template_path, (bytes, bytearray)):
        wb = load_workbook(BytesIO(template_path), data_only=False)
    elif isinstance(template_path, (str, Path)):
        template = Path(template_path)
        if not template.exists():
            raise FileNotFoundError(f"Template não encontrado: {template}")
        wb = load_workbook(template, data_only=False)
    else:
        raise TypeError(
            f"template_path deve ser bytes, bytearray, BytesIO, str ou Path; "
            f"recebido: {type(template_path).__name__}"
        )

    # Novo modelo oficial (Coleta com aba posicao_contratual): conjunto de
    # abas obrigatorias proprio — 'historico' e ENTRADA_XLS_* nao existem.
    from _coleta_oficial import ABAS_COLETA_OFICIAL, eh_layout_coleta_oficial
    _vnova = eh_layout_coleta_oficial(wb)
    abas_exigidas = ABAS_COLETA_OFICIAL if _vnova else ABAS_OBRIGATORIAS
    faltantes = [aba for aba in abas_exigidas if _localizar_aba(wb, aba) is None]
    if faltantes:
        raise ValueError(
            f"Template inválido; abas ausentes: {', '.join(faltantes)} "
            f"(abas encontradas: {', '.join(wb.sheetnames)})"
        )

    _v102 = str(wb["parametros"]["A1"].value or "").strip() == "COMPUTAR_NESTA_APURACAO"

    formulas_antes = _formulas(wb)
    controle = wb["CONTROLE"]
    parametros = wb["parametros"]
    agora = datetime.now().astimezone()

    _escrever_entrada(controle, "B1", "Principal")
    _escrever_entrada(controle, "B2", dados_calculadora.get("ciclo_vigente") or "C0")
    _escrever_entrada(controle, "B3", dados_calculadora.get("data_corte"))
    ciclos = {str(c.get("ciclo", "")).upper(): c for c in dados_calculadora.get("ciclos", [])}

    # v10.5.2: pos-calculadora, preencher B7 (Indice utilizado) e B8 (Data-base
    # original) do quadro RESUMO DO REAJUSTE — antes ficavam sempre vazios.
    if str(controle["A7"].value or "").strip().replace("Í", "I").startswith("Indice"):
        indice_utilizado = dados_calculadora.get("indice")
        if indice_utilizado:
            _escrever_entrada(controle, "B7", indice_utilizado)
        data_base_original = dados_calculadora.get("data_base")
        if not data_base_original:
            c0 = ciclos.get("C0") or {}
            data_base_original = c0.get("data_base") or c0.get("data_inicio")
        if data_base_original:
            _escrever_entrada(controle, "B8", data_base_original)
            controle["B8"].number_format = "dd/mm/yyyy"

    # v10.5.2: C0-C4 sempre existem temporalmente — ciclo fora da apuracao nao
    # some da linha do tempo nem tem seu periodo herdado pelo ciclo seguinte.
    ciclos_completos = _completar_periodos_ciclos(ciclos, dados_calculadora.get("data_corte"))
    _registrar_inicio_efeitos_financeiros(wb, ciclos_completos)

    for linha, nome in enumerate(("C0", "C1", "C2", "C3", "C4"), start=2):
        ciclo = ciclos.get(nome)
        if _vnova:
            # Novo modelo oficial: A=COMPUTAR_NESTA_APURACAO, B=CICLO,
            # C=DATA_INICIO, D=DATA_FIM, E=PERCENTUAL_DO_CICLO,
            # F=FATOR_ACUMULADO (formula — nao tocar), G=SITUACAO
            for coluna in ("A", "C", "D", "E", "G"):
                _escrever_entrada(parametros, f"{coluna}{linha}", None)
            _escrever_entrada(parametros, f"B{linha}", nome)
            for coluna in ("A", "E", "G"):
                parametros[f"{coluna}{linha}"].comment = None
            derivado = ciclos_completos.get(nome) or {}
            fonte = ciclo or derivado
            if fonte.get("data_inicio"):
                _escrever_entrada(parametros, f"C{linha}", fonte["data_inicio"])
                parametros[f"C{linha}"].number_format = "dd/mm/yyyy"
            if fonte.get("data_fim"):
                _escrever_entrada(parametros, f"D{linha}", fonte["data_fim"])
                parametros[f"D{linha}"].number_format = "dd/mm/yyyy"
            if not ciclo:
                if nome == "C0":
                    _escrever_entrada(parametros, f"E{linha}", 0.0)
                else:
                    parametros[f"E{linha}"].fill = _FILL_FORA
                _escrever_entrada(parametros, f"A{linha}", "Nao")
                _escrever_entrada(
                    parametros, f"G{linha}",
                    "Base" if nome == "C0" else "Fora desta apuracao",
                )
                continue
            _escrever_entrada(
                parametros, f"E{linha}",
                0.0 if nome == "C0" else ciclo.get("percentual"),
            )
            efeito_raw = str(ciclo.get("possui_efeito_financeiro") or "")
            computar = "Sim" if efeito_raw.upper() in ("SIM", "S") else "Nao"
            _escrever_entrada(parametros, f"A{linha}", computar)
            _escrever_entrada(parametros, f"G{linha}", ciclo.get("situacao"))
        elif _v102:
            # v10.2: A=COMPUTAR_NESTA_APURACAO, B=CICLO, C=PERIODO, D=DATA_INICIO,
            #         E=DATA_FIM, F=FATOR_PROPRIO, G=formula (preservar), H=SITUACAO
            for coluna in ("A", "C", "D", "E", "F", "H"):
                _escrever_entrada(parametros, f"{coluna}{linha}", None)
            _escrever_entrada(parametros, f"B{linha}", nome)
            parametros[f"A{linha}"].comment = None
            parametros[f"F{linha}"].comment = None
            parametros[f"G{linha}"].comment = None
            if not ciclo:
                if nome == "C0":
                    _escrever_entrada(parametros, f"F{linha}", 0.0)
                else:
                    parametros[f"F{linha}"].fill = _FILL_FORA
                _escrever_entrada(parametros, f"A{linha}", "Nao")
                _escrever_entrada(parametros, f"H{linha}", "Base" if nome == "C0" else "Fora desta apuracao")
                # v10.5.2: ciclo fora da apuracao mantem periodo real na linha
                # do tempo (datas derivadas dos vizinhos) — nao pula nem compacta.
                derivado = ciclos_completos.get(nome) or {}
                if derivado.get("data_inicio"):
                    _escrever_entrada(parametros, f"D{linha}", derivado["data_inicio"])
                    parametros[f"D{linha}"].number_format = "dd/mm/yyyy"
                if derivado.get("data_fim"):
                    _escrever_entrada(parametros, f"E{linha}", derivado["data_fim"])
                    parametros[f"E{linha}"].number_format = "dd/mm/yyyy"
                if derivado.get("data_inicio"):
                    # col C = data real dd/mm/yyyy (regra v10.5.0), nunca texto
                    _escrever_entrada(parametros, f"C{linha}", derivado["data_inicio"])
                    parametros[f"C{linha}"].number_format = "dd/mm/yyyy"
                continue
            # v10.5.2: col C = data real dd/mm/yyyy (regra v10.5.0 do auditor),
            # nunca texto de periodo
            if ciclo.get("data_inicio"):
                _escrever_entrada(parametros, f"C{linha}", ciclo.get("data_inicio"))
                parametros[f"C{linha}"].number_format = "dd/mm/yyyy"
            _escrever_entrada(parametros, f"D{linha}", ciclo.get("data_inicio"))
            _escrever_entrada(parametros, f"E{linha}", ciclo.get("data_fim"))
            _escrever_entrada(parametros, f"F{linha}", 0.0 if nome == "C0" else ciclo.get("percentual"))
            efeito_raw = str(ciclo.get("possui_efeito_financeiro") or "")
            computar = "Sim" if efeito_raw.upper() in ("SIM", "S") else "Nao"
            _escrever_entrada(parametros, f"A{linha}", computar)
            _escrever_entrada(parametros, f"H{linha}", ciclo.get("situacao"))
        else:
            # v10.1/v9: A=CICLO, B=PERIODO, C=DATA_INICIO, D=DATA_FIM, E=PERCENTUAL,
            #            H=EFEITO_FIN, I=SITUACAO
            _escrever_entrada(parametros, f"A{linha}", nome)
            for coluna in ("B", "C", "D", "E", "H", "I"):
                _escrever_entrada(parametros, f"{coluna}{linha}", None)
            parametros[f"B{linha}"].comment = None
            parametros[f"F{linha}"].comment = None
            parametros[f"G{linha}"].comment = None
            if not ciclo:
                _escrever_entrada(parametros, f"E{linha}", 0.0)
                _escrever_entrada(parametros, f"H{linha}", "Não")
                _escrever_entrada(parametros, f"I{linha}", "Fora desta apuração")
                continue
            _escrever_entrada(parametros, f"B{linha}", ciclo.get("periodo"))
            _escrever_entrada(parametros, f"C{linha}", ciclo.get("data_inicio"))
            _escrever_entrada(parametros, f"D{linha}", ciclo.get("data_fim"))
            _escrever_entrada(parametros, f"E{linha}", ciclo.get("percentual"))
            _escrever_entrada(parametros, f"H{linha}", ciclo.get("possui_efeito_financeiro"))
            _escrever_entrada(parametros, f"I{linha}", ciclo.get("situacao"))

    if _vnova:
        # Novo modelo: quadro MEMORIA DO FATOR (linhas 9-16) e 100% formula
        # no template — nada a preencher via Python.
        # Etapa 4: bloco MEMORIA DE CALCULO (parametros!J2:R80) — valores da
        # memoria mensal ja calculada pelas Calculadoras; template legado sem
        # os cabecalhos J1:R1 e ignorado sem erro.
        from _memoria_calculo import escrever_memoria_calculo
        escrever_memoria_calculo(parametros, ciclos)
    elif _v102 and "parametros" in wb.sheetnames:
        _preencher_memoria_fator(parametros, ciclos)

    if "financeiro" in wb.sheetnames:
        financeiro = wb["financeiro"]
        data_corte = dados_calculadora.get("data_corte")
        # v10.5.2: usa ciclos com periodos completos — financeiro lista C1-C4
        # com datas corretas, sem pular ciclo nao computado (G="Não" nesses meses).
        # Etapa 4: a grade financeira comeca no inicio de C0 (linha temporal
        # financeira). data_base pode ser a data-base do indice, ate 12 meses
        # antes de C0 — esse periodo pertence a memoria de calculo, nao a grade.
        marco_c0 = (ciclos_completos.get("C0") or {}).get("data_inicio")
        _preencher_financeiro(
            financeiro, ciclos_completos, data_corte_fallback=data_corte,
            v102=_v102, novo=_vnova,
            marco_inicial=marco_c0 or dados_calculadora.get("data_base"),
        )

    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.properties.creator = wb.properties.creator or "cl8us"
    wb.properties.lastModifiedBy = "cl8us / Calculadora 2.1"
    wb.properties.modified = agora.replace(tzinfo=None)
    wb.properties.description = (
        f"MASTERFILE_{'v10.2' if _v102 else 'v9'} preenchido pela Calculadora 2.1; "
        f"modo={dados_calculadora.get('modo_origem', '')}; índice={dados_calculadora.get('indice', '')}"
    )

    # Remove comentarios antes de salvar (evita VML/legacyDrawing)
    for _ws in wb.worksheets:
        for _row in _ws.iter_rows():
            for _cell in _row:
                if _cell.comment is not None:
                    _cell.comment = None
    # Limpa pane orfao em sheetViews nativamente
    for _ws in wb.worksheets:
        sv = _ws.sheet_view
        if sv.pane is None:
            for _sel in list(sv.selection):
                p = getattr(_sel, "pane", None)
                if p is not None and p != "topLeft":
                    _sel.pane = None

    if _formulas(wb) != formulas_antes:
        raise AssertionError("A geração alterou fórmulas do template antes da gravação.")

    saida = BytesIO()
    wb.save(saida)
    conteudo = _corrigir_selection_pane_xlsx(_corrigir_vml_xlsx(saida.getvalue()))

    verificacao = load_workbook(BytesIO(conteudo), data_only=False, read_only=False)
    if _formulas(verificacao) != formulas_antes:
        raise AssertionError("A geração alterou fórmulas do template durante a gravação.")
    verificacao.close()
    return conteudo
