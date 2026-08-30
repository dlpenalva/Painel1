"""
_leitor_masterfile_v10.py
-------------------------
Leitor do MASTERFILE v10 RC.

Evolucoes em relacao ao v9:
  - Reconhece MASTERFILE_VERSION "v10-rc"
  - Aceita v9 como legado (delegado ao leitor v9 se necessario)
  - Le aba historico_VU por cabecalho (nao por posicao)
  - Le itens_Consumidos no novo formato Qtd+Valor por ciclo
  - Le itens_PC com TIPO_PC: Unitario e Global/Multi-item
  - Retorna alertas estruturados quando campos obrigatorios ausentes
  - Nao recalcula nada; le valores ja calculados pelo Excel (data_only=True)
"""

from __future__ import annotations

import unicodedata
from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from _seguranca_xlsx import (
    ErroSegurancaXlsx,
    garantir_xlsx_validado,
    validar_geometria_workbook,
)

from _masterfile_config_v10 import (
    ABAS_OBRIGATORIAS_V10,
    ABA_HISTORICO_VU,
    ABA_ITENS_PC,
    ABA_EXECUCAO_SALDO,
    MODOS_VALIDOS_V10,
)
from _capacidade_pcs import ULTIMA_LINHA_PCS
from _motor_vta_sombra import calcular_vta_sombra
from _motor_composicao_vta import montar_composicao_vta
from _motor_reconciliacao import reconciliar_execucoes
from _reconciliacao_evidencias import reconciliar_evidencias
from _log_decisoes_gcc import carregar_decisoes, hash_entrada
from _estado_contratual_sombra import (
    estado_contratual_para_dict,
    montar_event_log_sombra,
    reconstruir_estado_contratual,
)
from _posicao_contratual_sombra import montar_posicao_contratual_sombra
from _objeto_processo_reajuste import montar_objeto_processo_reajuste
from _motor_temporal import (
    classificar_enquadramento_pc,
    enquadrar_data_pc,
    ENQ_CICLO,
    ENQ_INDETERMINADO,
    ENQ_INTERVALO_PRECLUSO,
)


# ---------------------------------------------------------------------------
# Utilitarios
# ---------------------------------------------------------------------------

def _norm(txt: Any) -> str:
    if txt is None:
        return ""
    t = str(txt).strip().lower()
    t = unicodedata.normalize("NFKD", t)
    return "".join(ch for ch in t if not unicodedata.combining(ch))


def _achar_valor(ws, rotulo_norm: str, max_lin: int = 60) -> Any:
    for r in range(1, max_lin + 1):
        if _norm(ws.cell(r, 1).value) == _norm(rotulo_norm):
            return ws.cell(r, 2).value
    return None


def _valor_por_prefixo(ws, prefixo: str, max_lin: int = 40) -> Any:
    """Valor em col B da 1a linha cujo rotulo (col A) COMECA com `prefixo`.

    Leitura semantica por rotulo (tolerante a sufixos como "(GCC, opcional)"),
    sem depender de coordenada fixa.
    """
    alvo = _norm(prefixo)
    for r in range(1, max_lin + 1):
        rot = _norm(ws.cell(r, 1).value)
        if rot and rot.startswith(alvo):
            return ws.cell(r, 2).value
    return None


def _norm_data(valor: Any):
    """Normaliza um valor de celula para datetime.date; caso contrario None.

    NUNCA transforma vazio, numero solto ou texto em data (fail-closed): apenas
    datas reais (date/datetime) sao aceitas.
    """
    from datetime import date as _date, datetime as _dt
    if isinstance(valor, _dt):
        return valor.date()
    if isinstance(valor, _date):
        return valor
    return None


# Rotulos canonicos da aba cobertura_temporal (leitura por prefixo).
# Cada chave aceita o rotulo atual e os legados: arquivos gerados antes da
# padronizacao das datas continuam sendo lidos sem conversao.
_ROTULOS_COBERTURA_TEMPORAL = {
    "data_analise": ("data de geracao/analise", "data da analise"),
    "financeiro_confirmado_completo_ate": (
        "financeiro conferido ate", "financeiro confirmado completo ate",
    ),
    "pc_confirmado_completo_ate": (
        "pcs conferidos ate", "pc confirmado completo ate",
    ),
}


def _ler_cobertura_temporal(wb) -> dict[str, Any]:
    """Le a aba de diagnostico cobertura_temporal (GCC) por rotulo.

    Le SOMENTE as declaracoes GCC (Data da analise, Financeiro/PC confirmado
    completo ate). NAO recalcula formula e NAO le a ultima evidencia (MAX): a
    ultima evidencia continua sendo derivada pelo motor a partir de
    financeiro/itens_PC. Aba ausente => estrutura vazia compativel (sem
    excecao, sem bloquear arquivos oficiais anteriores).
    """
    vazio = {
        "ok": False,
        "data_analise": None,
        "financeiro_confirmado_completo_ate": None,
        "pc_confirmado_completo_ate": None,
    }
    if "cobertura_temporal" not in wb.sheetnames:
        return vazio
    ws = wb["cobertura_temporal"]
    saida: dict[str, Any] = {"ok": True}
    for chave, prefixos in _ROTULOS_COBERTURA_TEMPORAL.items():
        valor = None
        for prefixo in prefixos:
            valor = _norm_data(_valor_por_prefixo(ws, prefixo))
            if valor is not None:
                break
        saida[chave] = valor
    return saida


def _ler_posicao_fisica_fiscal(wb):
    """Itens da base (itens_Remanesc!A) e a fotografia recente do fiscal.

    Fornece ao motor de cobertura temporal os dois insumos da posicao fisica
    ATUAL: a lista de itens e o remanescente declarado pelo fiscal.

    FONTE UNICA: a quantidade vem de CICLO_EM_EXECUCAO!C13:C211 (entrada manual
    real, imune a cache de formula). posicao_referencia!B e apenas a memoria
    automatica derivada dela e so e consultada em arquivos legados, gerados
    antes da unificacao, nos quais aquela coluna ainda era digitada.

    Zero conta como informado; celula vazia nao entra (a fotografia so e
    completa quando TODOS os itens tem quantidade). Nao le formula nem inventa.
    """
    if "itens_Remanesc" not in wb.sheetnames:
        return [], {}
    ir = wb["itens_Remanesc"]
    cee = wb["CICLO_EM_EXECUCAO"] if "CICLO_EM_EXECUCAO" in wb.sheetnames else None
    pr = wb["posicao_referencia"] if "posicao_referencia" in wb.sheetnames else None
    itens_base: list[str] = []
    remanescente: dict[str, float] = {}
    # 26H.2: capacidade canonica de cadastro = linhas 2:200 (linha 201 e do
    # total dinamico e NAO e linha valida de item — fronteira da auditoria).
    for r in range(2, 201):
        item = ir.cell(r, 1).value
        cod = str(item).strip() if item not in (None, "") else ""
        if not cod:
            continue
        itens_base.append(cod)
        q = None
        if cee is not None:
            # itens_Remanesc!A{r} corresponde a CICLO_EM_EXECUCAO!C{r+11}.
            q = cee.cell(r + 11, 3).value
        if not isinstance(q, (int, float)) or isinstance(q, bool):
            q = pr.cell(r, 2).value if pr is not None else None
        if isinstance(q, (int, float)) and not isinstance(q, bool):
            remanescente[cod] = float(q)
    return itens_base, remanescente


def _ler_financeiro_competencias(wb) -> list[dict[str, Any]]:
    """Competencias financeiras (mensais) para o motor de cobertura temporal.

    Fonte mensal: COMPETENCIA (col A), CICLO (col B), VALOR_PAGO (col C). Apenas
    linhas cuja competencia e uma DATA real entram; zero conta como competencia
    informada, a AUSENCIA de linha e lacuna. Nao interpreta valor nem infere
    continuidade aqui — o motor decide. Compat: aba ausente => lista vazia.
    """
    if "financeiro" not in wb.sheetnames:
        return []
    # Semantica homologada da Adequacao Rev.2 (fonte unica): ZERO != VAZIO.
    from _adequacao_orcamentaria import valor_original_foi_informado
    ws = wb["financeiro"]
    linhas: list[dict[str, Any]] = []
    for r in range(2, min(ws.max_row or 2, 200) + 1):
        comp = _norm_data(ws.cell(r, 1).value)
        if comp is None:
            continue
        valor_bruto = ws.cell(r, 3).value
        # informado=True inclui zero explicito; None/vazio/NaN/traco => False.
        # Linha vazia PERMANECE no payload (diagnostico da lacuna), mas sinalizada.
        linhas.append({
            "competencia": comp,
            "ciclo": str(ws.cell(r, 2).value or "").strip().upper(),
            "valor": valor_bruto,
            "informado": valor_original_foi_informado(valor_bruto),
        })
    return linhas


def _mapear_colunas_por_cabecalho(ws, linha_header: int = 1) -> dict[str, int]:
    mapa: dict[str, int] = {}
    for cell in ws[linha_header]:
        if cell.value:
            # Primeira ocorrencia vence: o layout oficial fica a esquerda e o
            # bloco MEMORIA DE CALCULO (J:R) repete nomes como CICLO e
            # FATOR_ACUMULADO sem sombrear as colunas principais.
            mapa.setdefault(_norm(cell.value), cell.column)
    return mapa


def _col(mapa: dict[str, int], *nomes: str) -> int | None:
    for nome in nomes:
        achada = mapa.get(_norm(nome))
        if achada:
            return achada
    return None


def _normalizar_data(valor: Any) -> Any:
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    return valor


def _normalizar_modo(valor_bruto: str) -> str:
    # Fonte unica de verdade do metodo (CONTROLE!B1): normalizador canonico
    # compartilhado, com aliases legados e novos rotulos do dropdown.
    from _metodo_apuracao import normalizar_metodo
    if not valor_bruto:
        return ""
    return normalizar_metodo(valor_bruto)


# ---------------------------------------------------------------------------
# Leitores especificos v10
# ---------------------------------------------------------------------------

def _ler_historico_vu(wb) -> dict[str, Any]:
    """Le a aba historico_VU por cabecalho."""
    resultado: dict[str, Any] = {"itens": [], "alertas": [], "ok": False}

    if ABA_HISTORICO_VU not in wb.sheetnames:
        resultado["alertas"].append(
            f"Aba '{ABA_HISTORICO_VU}' ausente — necessaria no modo v10."
        )
        return resultado

    ws = wb[ABA_HISTORICO_VU]
    mapa = _mapear_colunas_por_cabecalho(ws)

    col_item = mapa.get("item")
    if not col_item:
        resultado["alertas"].append("historico_VU: coluna ITEM nao encontrada.")
        col_item = 1
    elif col_item != 1:
        resultado["alertas"].append(
            f"historico_VU: ITEM nao esta na coluna A (esta em {col_item})."
        )

    col_vu_orig = mapa.get("vu_original")
    col_desc    = mapa.get("descricao")
    col_vu_vig  = mapa.get("vu_vigente_ultimo_ciclo")
    col_fator   = mapa.get("fator_acumulado_ultimo_ciclo")
    col_var     = mapa.get("variacao_acumulada")
    col_fonte   = mapa.get("fonte")
    col_obs     = mapa.get("observacao")

    cols_vu_ciclo = {f"VU_C{i}": mapa.get(f"vu_c{i}") for i in range(5)}
    cols_qtd_vigente = {
        f"C{i}": _col(mapa, f"QTD_VIGENTE_C{i}") for i in range(5)
    }
    cols_qtd_rem_ajustada = {
        f"C{i}": _col(mapa, f"QTD_REM_AJUSTADA_C{i}") for i in range(5)
    }

    for r in range(2, ws.max_row + 1):
        item = ws.cell(r, col_item).value
        if not item:
            continue

        vu_orig = ws.cell(r, col_vu_orig).value if col_vu_orig else None
        if vu_orig is None:
            resultado["alertas"].append(
                f"historico_VU: item '{item}' sem VU original."
            )

        vu_ciclos: dict[str, Any] = {}
        for chave, col in cols_vu_ciclo.items():
            vu_ciclos[chave] = ws.cell(r, col).value if col else None

        resultado["itens"].append({
            "item":            item,
            "descricao":       ws.cell(r, col_desc).value   if col_desc  else None,
            "vu_original":     vu_orig,
            "vu_ciclos":       vu_ciclos,
            "qtd_vigente_ciclos": {
                ciclo: ws.cell(r, col).value if col else None
                for ciclo, col in cols_qtd_vigente.items()
            },
            "qtd_rem_ajustada_ciclos": {
                ciclo: ws.cell(r, col).value if col else None
                for ciclo, col in cols_qtd_rem_ajustada.items()
            },
            "vu_vigente":      ws.cell(r, col_vu_vig).value if col_vu_vig else None,
            "fator_acumulado": ws.cell(r, col_fator).value  if col_fator  else None,
            "variacao":        ws.cell(r, col_var).value    if col_var    else None,
            "fonte":           ws.cell(r, col_fonte).value  if col_fonte  else None,
            "observacao":      ws.cell(r, col_obs).value    if col_obs    else None,
        })

    resultado["ok"] = bool(resultado["itens"])
    return resultado


def _ler_parametros_v10(wb) -> dict[str, Any]:
    """Le a aba parametros por cabecalho, sem depender do layout A:H/A:I."""
    resultado: dict[str, Any] = {
        "ciclos": [], "por_ciclo": {}, "alertas": [], "ok": False
    }

    if "parametros" not in wb.sheetnames:
        resultado["alertas"].append("Aba 'parametros' ausente.")
        return resultado

    ws = wb["parametros"]
    mapa = _mapear_colunas_por_cabecalho(ws)

    col_ciclo = _col(mapa, "CICLO")
    if not col_ciclo:
        resultado["alertas"].append("parametros: coluna CICLO nao encontrada.")
        return resultado

    col_computar = _col(mapa, "COMPUTAR_NESTA_APURACAO", "COMPUTAR")
    col_periodo = _col(mapa, "PERIODO", "Período")
    col_data_inicio = _col(mapa, "DATA_INICIO", "Data início", "Data inicio")
    col_data_fim = _col(mapa, "DATA_FIM", "Data fim")
    col_percentual = _col(
        mapa, "PERCENTUAL_REAJUSTE", "PERCENTUAL_DO_CICLO",
        "% reajuste ciclo", "% REAJUSTE CICLO",
        "PCT_REAJUSTE", "PERCENTUAL"
    )
    col_fator_proprio = _col(mapa, "FATOR_PROPRIO", "Fator próprio", "Fator proprio")
    col_fator_acumulado = _col(mapa, "FATOR_ACUMULADO", "Fator acumulado")
    col_efeito_fin = _col(mapa, "EFEITO_FINANCEIRO", "Efeito financeiro?")
    col_situacao = _col(mapa, "SITUACAO", "Situação")
    col_inicio_efeito = _col(
        mapa, "INICIO_EFEITO_FINANCEIRO", "INICIO DO EFEITO FINANCEIRO"
    )
    # DATA_PEDIDO (parametros!U): data em que a CONTRATADA apresentou o pedido,
    # gravada pelo gerador. Arquivo anterior a essa gravacao nao tem a coluna e
    # devolve None — ausencia continua sendo ausencia, nunca vira data presumida.
    col_data_pedido = _col(mapa, "DATA_PEDIDO", "DATA DO PEDIDO")
    from _efeitos_financeiros_pc import reconciliar_inicios_efeito
    inicios_reconciliados, erros_inicio, tem_inicio_visivel, tem_inicio_metadado = (
        reconciliar_inicios_efeito(wb)
    )
    resultado["fonte_inicio_efeito_visivel"] = tem_inicio_visivel
    resultado["fonte_inicio_efeito_metadado"] = tem_inicio_metadado
    resultado["inicio_efeito_consistente"] = not erros_inicio
    resultado["alertas"].extend(erros_inicio)

    # No layout v10.2 real, FATOR_PROPRIO guarda o percentual do ciclo.
    # No legado A:I, o percentual tem coluna propria e FATOR_PROPRIO pode ser 1+pct.
    percentual_fallback = col_fator_proprio if not col_percentual else None

    for r in range(2, ws.max_row + 1):
        ciclo = str(ws.cell(r, col_ciclo).value or "").strip().upper()
        if not ciclo:
            continue
        if ciclo not in {f"C{i}" for i in range(5)}:
            resultado["alertas"].append(
                f"parametros: ciclo fora do catalogo C0-C4 ignorado em linha {r}: {ciclo!r}."
            )
            continue

        percentual_col = col_percentual or percentual_fallback
        registro = {
            "ciclo": ciclo,
            "computar_nesta_apuracao": (
                ws.cell(r, col_computar).value if col_computar else None
            ),
            "periodo": ws.cell(r, col_periodo).value if col_periodo else None,
            "data_inicio": (
                _normalizar_data(ws.cell(r, col_data_inicio).value)
                if col_data_inicio else None
            ),
            "data_fim": (
                _normalizar_data(ws.cell(r, col_data_fim).value)
                if col_data_fim else None
            ),
            "percentual_reajuste": (
                ws.cell(r, percentual_col).value if percentual_col else None
            ),
            "fator_proprio": (
                ws.cell(r, col_fator_proprio).value if col_fator_proprio else None
            ),
            "fator_acumulado": (
                ws.cell(r, col_fator_acumulado).value if col_fator_acumulado else None
            ),
            "efeito_financeiro": (
                ws.cell(r, col_efeito_fin).value if col_efeito_fin else None
            ),
            "situacao": ws.cell(r, col_situacao).value if col_situacao else None,
            "data_pedido": (
                _normalizar_data(ws.cell(r, col_data_pedido).value)
                if col_data_pedido else None
            ),
            "inicio_efeito_financeiro": inicios_reconciliados.get(ciclo),
            "inicio_efeito_financeiro_parametros": (
                _normalizar_data(ws.cell(r, col_inicio_efeito).value)
                if col_inicio_efeito else None
            ),
            "origem_aba": "parametros",
            "origem_linha": r,
        }
        resultado["ciclos"].append(registro)
        resultado["por_ciclo"][ciclo] = registro

    if not resultado["ciclos"]:
        resultado["alertas"].append("parametros: nenhum ciclo C0-C4 preenchido.")

    resultado["ok"] = bool(resultado["ciclos"])
    return resultado


def _ler_itens_consumidos_v10(wb) -> dict[str, Any]:
    """Le itens_Consumidos v10 por cabecalho: QTD_CONS_Cn / VALOR_CONS_Cn."""
    resultado: dict[str, Any] = {
        "itens": [], "totais": {}, "alertas": [], "ok": False
    }

    aba_fisica = next((n for n in ("itens_Consumidos", "itens_B") if n in wb.sheetnames), None)
    if not aba_fisica:
        resultado["alertas"].append("Aba itens_Consumidos nao encontrada.")
        return resultado

    ws = wb[aba_fisica]
    mapa = _mapear_colunas_por_cabecalho(ws)

    if mapa.get("ciclo_pc") or mapa.get("numero_pc"):
        resultado["alertas"].append(
            "ATENCAO: itens_Consumidos parece ter layout de itens_PC. Verificar template."
        )

    col_item  = mapa.get("item") or 1
    col_qtd_c = mapa.get("qtd_contratada")
    col_vu    = mapa.get("vu_original") or mapa.get("vu")

    ciclos_cols: dict[str, dict[str, int | None]] = {}
    for i in range(5):
        ciclos_cols[f"C{i}"] = {
            "qtd":   (mapa.get(f"qtd_cons_c{i}")
                      or mapa.get(f"qtd cons. c{i}")
                      or mapa.get(f"qtd cons c{i}")),
            "valor": (mapa.get(f"valor_cons_c{i}")
                      or mapa.get(f"valor cons. c{i}")
                      or mapa.get(f"valor cons c{i}")),
        }

    col_qtd_total   = mapa.get("cons_qtd_total") or mapa.get("cons. qtd. total")
    col_valor_total = mapa.get("cons_valor_total") or mapa.get("cons. valor total")
    col_check       = mapa.get("check")

    if col_valor_total is None:
        resultado["alertas"].append(
            "Coluna CONS_VALOR_TOTAL nao encontrada — template v10 desatualizado."
        )

    for r in range(2, ws.max_row + 1):
        item = ws.cell(r, col_item).value
        if not item:
            continue
        consumos: dict[str, dict[str, Any]] = {}
        for ciclo_key, cols in ciclos_cols.items():
            qtd_col   = cols.get("qtd")
            valor_col = cols.get("valor")
            consumos[ciclo_key] = {
                "qtd":   ws.cell(r, qtd_col).value   if qtd_col   else None,
                "valor": ws.cell(r, valor_col).value if valor_col else None,
            }
        resultado["itens"].append({
            "item":           item,
            "qtd_contratada": ws.cell(r, col_qtd_c).value if col_qtd_c else None,
            "vu_original":    ws.cell(r, col_vu).value    if col_vu    else None,
            "consumos":       consumos,
            "qtd_total":      ws.cell(r, col_qtd_total).value   if col_qtd_total   else None,
            "valor_total":    ws.cell(r, col_valor_total).value if col_valor_total else None,
            "check":          ws.cell(r, col_check).value       if col_check       else None,
        })

    resultado["totais"] = {
        "qtd_total":   sum(i["qtd_total"]   or 0 for i in resultado["itens"]
                           if isinstance(i["qtd_total"],   (int, float))),
        "valor_total": sum(i["valor_total"] or 0 for i in resultado["itens"]
                           if isinstance(i["valor_total"], (int, float))),
    }
    resultado["ok"] = bool(resultado["itens"])
    return resultado


_VTA_DEFAULTS = {
    "computa_vta": "Nao",
    "tipo_parcela": "PC em analise",
    "origem_dado": "Pedido de Compra",
    "tipo_financeiro": "Impacto Potencial",
    "fonte_parcela": "PC",
    "ja_refletido_em": "Nao",
    "status_consolidacao": "NAO_COMPUTADO",
    "justificativa_vta": "Default seguro do leitor: sem impacto no VTA oficial",
}

_VTA_OPCOES = {
    "computa_vta": {
        "sim": "Sim",
        "nao": "Nao",
    },
    "tipo_parcela": {
        "execucao realizada": "Execucao realizada",
        "saldo remanescente": "Saldo remanescente",
        "pc pago": "PC pago",
        "pc em analise": "PC em analise",
        "aditivo": "Aditivo",
        "item consumido": "Item consumido",
        "item remanescente": "Item remanescente",
    },
    "origem_dado": {
        "financeiro": "Financeiro",
        "itens consumidos": "Itens Consumidos",
        "itens remanescentes": "Itens Remanescentes",
        "pedido de compra": "Pedido de Compra",
        "aditivo": "Aditivo",
    },
    "tipo_financeiro": {
        "execucao atualizada": "Execucao Atualizada",
        "saldo remanescente": "Saldo Remanescente",
        "retroativo reconhecido": "Retroativo Reconhecido",
        "impacto potencial": "Impacto Potencial",
        "aditivo computavel": "Aditivo Computavel",
        "informativo": "Informativo",
    },
    "fonte_parcela": {
        "financeiro": "Financeiro",
        "itens consumidos": "Itens consumidos",
        "itens remanescentes": "Itens remanescentes",
        "pc": "PC",
        "aditivo": "Aditivo",
        "mista": "Mista",
    },
    "ja_refletido_em": {
        "nao": "Nao",
        "financeiro": "Financeiro",
        "itens": "Itens",
        "pc": "PC",
        "aditivo": "Aditivo",
        "historico": "Historico",
    },
    "status_consolidacao": {
        "computado": "COMPUTADO",
        "nao_computado": "NAO_COMPUTADO",
        "nao computado": "NAO_COMPUTADO",
        "descartado_duplicidade": "DESCARTADO_DUPLICIDADE",
        "descartado duplicidade": "DESCARTADO_DUPLICIDADE",
        "em_analise": "EM_ANALISE",
        "em analise": "EM_ANALISE",
        "inconsistente": "INCONSISTENTE",
    },
}

_VTA_CAMPOS = [
    "computa_vta",
    "tipo_parcela",
    "origem_dado",
    "tipo_financeiro",
    "fonte_parcela",
    "ja_refletido_em",
    "status_consolidacao",
    "justificativa_vta",
]

_PC_INTELIGENTE_OBRIGATORIOS = [
    "NUMERO_PC", "DATA_PC", "VALOR_PC", "STATUS_PAGAMENTO_PC"
]
_PC_INTELIGENTE_OPCIONAIS = [
    "PC_PAGO_INTEGRALMENTE",
    "VALOR_EFETIVAMENTE_PAGO",
    "DATA_PAGAMENTO_PC",
    "JA_REFLETIDO_EM",
    "OBSERVACAO",
]
_PC_INTELIGENTE_INFERIDOS = [
    "CICLO_PC",
    "FATOR_ACUMULADO",
    "VALOR_ATUALIZADO",
    "PC_PAGO_A_CONTRATADA",
    "COMPUTA_VTA",
    "TIPO_PARCELA",
    "ORIGEM_DADO",
    "TIPO_FINANCEIRO",
    "FONTE_PARCELA",
    "STATUS_CONSOLIDACAO",
    "JUSTIFICATIVA_VTA",
    "metodologia recomendada",
]
_PC_LAYOUT_ATUAL_MANUAIS_OU_HERDADOS = [
    "NUMERO_PC",
    "DATA_PC",
    "CICLO_PC",
    "VALOR_PC",
    "FATOR_ACUMULADO",
    "VALOR_ATUALIZADO",
    "PC_PAGO_A_CONTRATADA",
    "COMPUTA_VTA",
    "TIPO_PARCELA",
    "ORIGEM_DADO",
    "TIPO_FINANCEIRO",
    "FONTE_PARCELA",
    "JA_REFLETIDO_EM",
    "STATUS_CONSOLIDACAO",
    "JUSTIFICATIVA_VTA",
]

ABA_FISCAL_PCS_V2 = "ENTRADA_XLS_PCS"
ABA_FISCAL_FINANCEIRO_V2 = "ENTRADA_XLS_FINANCEIRO"
ABA_FISCAL_CONSUMIDOS_V2 = "ENTRADA_XLS_CONSUMIDOS"
ABA_FISCAL_REMANESCENTES_V2 = "ENTRADA_XLS_REMANESCENTES"
ABA_FISCAL_OBSERVACOES_V2 = "ENTRADA_XLS_OBSERVACOES"
ABA_FISCAL_RESSALVAS_DEFINITIVO = "ENTRADA_XLS_RESSALVAS"
ABA_FISCAL_HISTORICO_FINANCEIRO = "CICLOS_PASSADOS"
ABA_FISCAL_ITENS_CONTRATO = "ITENS_CONTRATO"
ABA_FISCAL_ADITIVOS = "ENTRADA_XLS_ADITIVOS"

_ABAS_LEGADAS_FISCAL = {
    ABA_FISCAL_PCS_V2: "FISCAL_PCS",
    ABA_FISCAL_FINANCEIRO_V2: "FISCAL_FINANCEIRO",
    ABA_FISCAL_CONSUMIDOS_V2: "FISCAL_CONSUMIDOS",
    ABA_FISCAL_REMANESCENTES_V2: "FISCAL_REMANESCENTES",
    ABA_FISCAL_OBSERVACOES_V2: "FISCAL_OBSERVACOES",
    ABA_FISCAL_RESSALVAS_DEFINITIVO: "FISCAL_RESSALVAS",
    ABA_FISCAL_HISTORICO_FINANCEIRO: "ENTRADA_XLS_HISTORICO_FINANCEIRO",
}

_ABAS_ALTERNATIVAS_FISCAL = {
    ABA_FISCAL_HISTORICO_FINANCEIRO: ["FISCAL_HISTORICO_FINANCEIRO"],
}

_FISCAL_V2_ABAS = [
    ABA_FISCAL_PCS_V2,
    ABA_FISCAL_FINANCEIRO_V2,
    ABA_FISCAL_CONSUMIDOS_V2,
    ABA_FISCAL_REMANESCENTES_V2,
    ABA_FISCAL_OBSERVACOES_V2,
]

_FISCAL_DEFINITIVO_ABAS = [
    ABA_FISCAL_PCS_V2,
    ABA_FISCAL_FINANCEIRO_V2,
    ABA_FISCAL_CONSUMIDOS_V2,
    ABA_FISCAL_REMANESCENTES_V2,
    ABA_FISCAL_HISTORICO_FINANCEIRO,
    ABA_FISCAL_ADITIVOS,
    ABA_FISCAL_OBSERVACOES_V2,
    ABA_FISCAL_RESSALVAS_DEFINITIVO,
]


def _aba_entrada_existente(wb, nome: str) -> str | None:
    if nome in wb.sheetnames:
        return nome
    legado = _ABAS_LEGADAS_FISCAL.get(nome)
    if legado in wb.sheetnames:
        return legado
    for alternativo in _ABAS_ALTERNATIVAS_FISCAL.get(nome, []):
        if alternativo in wb.sheetnames:
            return alternativo
    return None


def _ws_entrada(wb, nome: str):
    existente = _aba_entrada_existente(wb, nome)
    return wb[existente] if existente else None

_FISCAL_V2_CAMPOS_FISCAL = [
    "financeiro executado/pago",
    "PCs",
    "DATA_PC",
    "VALOR_PC",
    "consumidos",
    "remanescentes",
    "historico financeiro",
]

_FISCAL_V2_NAO_PEDIR = [
    "VTA",
    "retroativo",
    "metodologia",
    "ciclo calculado",
    "fator acumulado",
    "valor atualizado",
    "decisao de reconhecimento",
    "campos tecnicos do Claus New",
]

_FISCAL_DEFINITIVO_CLASSIFICACAO = {
    "CONTROLE": {
        "gcc_calculadora": ["MODO DE LEITURA", "ciclo vigente", "data de corte"],
        "entrada_xls": [],
        "claus_new_infere": [],
        "sai_do_preenchimento_xls": ["parametros operacionais do upload"],
        "vira_alerta": ["modo de leitura ausente ou incompativel"],
    },
    "parametros": {
        "gcc_calculadora": ["CICLO", "PERIODO", "DATA_INICIO", "DATA_FIM", "FATOR_PROPRIO"],
        "entrada_xls": [],
        "claus_new_infere": ["FATOR_ACUMULADO", "reajuste aplicavel por ciclo"],
        "sai_do_preenchimento_xls": ["COMPUTAR_NESTA_APURACAO", "SITUACAO"],
        "vira_alerta": ["ciclo sem datas", "fator ausente"],
    },
    "financeiro": {
        "gcc_calculadora": [],
        "entrada_xls": ["competencia executada ou paga", "valor executado ou pago", "fonte", "observacao"],
        "claus_new_infere": ["estado pago/reconhecido/em analise quando houver evidencia"],
        "sai_do_preenchimento_xls": ["FATOR_APLICAVEL", "VALOR_ATUALIZADO", "DELTA", "EFEITO_FINANCEIRO"],
        "vira_alerta": ["pagamento citado sem comprovante", "valor financeiro divergente"],
    },
    "itens_Remanesc": {
        "gcc_calculadora": ["ITEM", "QTD_CONTRATADA", "VU_ORIGINAL", "VALOR_TOTAL"],
        "entrada_xls": ["data de corte", "tipo de corte", "QTD_REMANESCENTE", "observacao"],
        "claus_new_infere": ["ciclo aplicavel", "saldo por ciclo quando houver linha temporal suficiente"],
        "sai_do_preenchimento_xls": ["QTD_REM_INICIO_C*", "VALOR_REM_INICIO_C*", "CHECK", "FATOR_*"],
        "vira_alerta": ["saldo sem data de referencia", "saldo inconsistente com contratado"],
    },
    "itens_Consumidos": {
        "gcc_calculadora": ["ITEM", "QTD_CONTRATADA", "VU_ORIGINAL", "VALOR_TOTAL"],
        "entrada_xls": ["QTD_CONSUMIDA", "DATA_REFERENCIA", "OBSERVACAO"],
        "claus_new_infere": ["ciclo do consumo pela data", "valor consumido quando houver VU"],
        "sai_do_preenchimento_xls": ["QTD_CONS_C*", "VALOR_CONS_C*", "CONS_*", "FATOR_*"],
        "vira_alerta": ["consumo sem data", "consumo maior que contratado"],
    },
    "itens_PC": {
        "gcc_calculadora": [],
        "entrada_xls": ["NUMERO_PC", "DATA_PC", "VALOR_PC", "STATUS_PAGAMENTO_PC", "VALOR_EFETIVAMENTE_PAGO", "DATA_PAGAMENTO_PC", "OBSERVACAO"],
        "claus_new_infere": ["CICLO_PC", "FATOR_ACUMULADO", "VALOR_ATUALIZADO", "PC_PAGO_A_CONTRATADA"],
        "sai_do_preenchimento_xls": ["VTA", "retroativo", "metodologia", "decisao de reconhecimento"],
        "vira_alerta": ["PC sem DATA_PC", "PC sem VALOR_PC", "pagamento sem evidencia"],
    },
    "aditivos": {
        "gcc_calculadora": ["dados formais do aditivo quando existirem"],
        "entrada_xls": ["ressalva factual sobre alteracao formal"],
        "claus_new_infere": ["impacto por ciclo no motor proprio"],
        "sai_do_preenchimento_xls": ["fator acumulado", "valor atualizado", "considerado no calculo"],
        "vira_alerta": ["PC informado como aditivo", "aditivo sem dado formal"],
    },
    "historico": {
        "gcc_calculadora": ["resumo oficial", "VTA oficial", "historico!B51"],
        "entrada_xls": [],
        "claus_new_infere": ["conferencia sombra sem alterar oficial"],
        "sai_do_preenchimento_xls": ["todos os indicadores consolidados"],
        "vira_alerta": ["valor oficial ausente ou inconsistente"],
    },
    "historico_VU": {
        "gcc_calculadora": ["ITEM", "VU_ORIGINAL"],
        "entrada_xls": ["observacao sobre VU original quando houver ressalva"],
        "claus_new_infere": ["VU por ciclo", "variacao acumulada"],
        "sai_do_preenchimento_xls": ["VU_C*", "FATOR_*", "VARIACAO_ACUMULADA"],
        "vira_alerta": ["VU original ausente", "VU divergente"],
    },
    "historico_financeiro": {
        "gcc_calculadora": [],
        "entrada_xls": ["data/competencia passada", "valor executado ou pago passado", "fonte", "observacao"],
        "claus_new_infere": ["ciclo aplicavel quando houver parametro temporal"],
        "sai_do_preenchimento_xls": ["VTA", "retroativo", "delta"],
        "vira_alerta": ["historico passado fora da linha temporal parametrizada"],
    },
    "itens_RC": {
        "gcc_calculadora": ["itens referenciais do contrato"],
        "entrada_xls": [],
        "claus_new_infere": ["valores por ciclo"],
        "sai_do_preenchimento_xls": ["C0", "C1", "C2", "C3", "C4"],
        "vira_alerta": ["item sem correspondencia"],
    },
    "itens_Execucao_Saldo": {
        "gcc_calculadora": ["ITEM", "QTD_CONTRATADA", "VU_ORIGINAL"],
        "entrada_xls": ["PC", "QTD_EMITIDA", "VALOR_EMITIDO", "QTD_SALDO", "VALOR_SALDO"],
        "claus_new_infere": ["sanidade financeira", "saldo de referencia"],
        "sai_do_preenchimento_xls": ["CHECK_FISICO"],
        "vira_alerta": ["emitido maior que contratado", "saldo negativo"],
    },
}


def _normalizar_campo_vta(
    campo: str,
    valor: Any,
    ident: str,
    linha: int,
    alertas: list[str],
) -> str:
    if campo == "justificativa_vta":
        texto = str(valor or "").strip()
        return texto or _VTA_DEFAULTS[campo]

    if valor in (None, ""):
        return _VTA_DEFAULTS[campo]

    opcoes = _VTA_OPCOES[campo]
    chave = _norm(valor).replace("-", " ").replace("__", "_")
    if chave in opcoes:
        return opcoes[chave]

    alertas.append(
        f"PC '{ident}' linha {linha}: {campo.upper()} invalido "
        f"({valor!r}); aplicado default seguro {_VTA_DEFAULTS[campo]!r}."
    )
    return _VTA_DEFAULTS[campo]


def _aba_tem_dados(ws, linha_inicial: int = 2) -> bool:
    for r in range(linha_inicial, ws.max_row + 1):
        if any(ws.cell(r, c).value not in (None, "") for c in range(1, ws.max_column + 1)):
            return True
    return False


def _resumo_masterfile_fiscal_v2(wb) -> dict[str, Any]:
    abas_presentes = [aba for aba in _FISCAL_V2_ABAS if _aba_entrada_existente(wb, aba)]
    linhas_por_aba: dict[str, int] = {}
    for aba in abas_presentes:
        ws = _ws_entrada(wb, aba)
        linhas_por_aba[aba] = sum(
            1 for r in range(2, ws.max_row + 1)
            if any(ws.cell(r, c).value not in (None, "") for c in range(1, ws.max_column + 1))
        )
    return {
        "layout": "fiscal_v2" if abas_presentes else "",
        "layout_fiscal_v2": bool(abas_presentes),
        "abas_presentes": abas_presentes,
        "linhas_por_aba": linhas_por_aba,
        "papel_gcc": [
            "usa a Calculadora",
            "gera o MasterFile customizado",
            "faz upload no Claus New",
            "valida metodologia, VTA, retroativo, pendencias e dossie na Mesa da GCC",
        ],
        "entrada_xls": list(_FISCAL_V2_CAMPOS_FISCAL),
        "calculado_pelo_claus_new": list(_PC_INTELIGENTE_INFERIDOS),
        "nao_exigir_no_xls": list(_FISCAL_V2_NAO_PEDIR),
        "compatibilidade": (
            "Coexiste com o MasterFile atual: se a aba de PCs de entrada existir, o leitor usa "
            "os fatos enxutos; caso contrario, segue lendo itens_PC atual/legado."
        ),
    }


def _linha_exemplo_fiscal(ws, r: int, max_col: int | None = None) -> bool:
    """Linha ilustrativa do Arquivo 3.0: qualquer celula iniciando por EXEMPLO.

    O gerador do XLS de entrada planta linhas de exemplo prefixadas; se o
    fiscal nao as apagar, elas nao podem contaminar VTA nem retroativo.
    """
    if max_col is None:
        max_col = ws.max_column
    for c in range(1, max_col + 1):
        valor = ws.cell(r, c).value
        if isinstance(valor, str) and valor.strip().upper().startswith("EXEMPLO"):
            return True
    return False


def _mapear_fatos_por_aba(ws) -> list[dict[str, Any]]:
    mapa = _mapear_colunas_por_cabecalho(ws)
    headers = [
        str(cell.value).strip()
        for cell in ws[1]
        if str(cell.value or "").strip()
    ]
    fatos: list[dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        if not any(ws.cell(r, c).value not in (None, "") for c in range(1, ws.max_column + 1)):
            continue
        if _linha_exemplo_fiscal(ws, r):
            continue
        fatos.append({
            header: ws.cell(r, mapa[_norm(header)]).value
            for header in headers
            if mapa.get(_norm(header))
        })
    return fatos


def _resumo_masterfile_fiscal_definitivo(wb) -> dict[str, Any]:
    abas_presentes = [aba for aba in _FISCAL_DEFINITIVO_ABAS if _aba_entrada_existente(wb, aba)]
    definitivo = bool(
        _aba_entrada_existente(wb, ABA_FISCAL_PCS_V2)
        and _aba_entrada_existente(wb, ABA_FISCAL_FINANCEIRO_V2)
        and _aba_entrada_existente(wb, ABA_FISCAL_REMANESCENTES_V2)
    )
    fatos = {
        aba: _mapear_fatos_por_aba(_ws_entrada(wb, aba))
        for aba in abas_presentes
    }
    linhas_por_aba = {
        aba: len(linhas)
        for aba, linhas in fatos.items()
    }
    campos_atuais_classificados = sum(
        len(campos)
        for grupo in _FISCAL_DEFINITIVO_CLASSIFICACAO.values()
        for campos in grupo.values()
    )
    campos_fiscais_definitivos = sum(
        len([cell for cell in wb[aba][1] if str(cell.value or "").strip()])
        for aba in abas_presentes
    )
    return {
        "layout": "fiscal_definitivo" if definitivo else "",
        "layout_fiscal_definitivo": definitivo,
        "abas_visiveis_fiscal": abas_presentes,
        "abas_calculo_preservadas": [
            aba for aba in wb.sheetnames
            if aba not in abas_presentes
        ],
        "linhas_por_aba": linhas_por_aba,
        "tem_observacoes_ressalvas": bool(
            linhas_por_aba.get(ABA_FISCAL_OBSERVACOES_V2)
            or linhas_por_aba.get(ABA_FISCAL_RESSALVAS_DEFINITIVO)
        ),
        "tem_historico_financeiro": bool(
            linhas_por_aba.get(ABA_FISCAL_HISTORICO_FINANCEIRO)
        ),
        "classificacao_abas": _FISCAL_DEFINITIVO_CLASSIFICACAO,
        "campos_atuais_classificados": campos_atuais_classificados,
        "campos_fiscais_definitivos": campos_fiscais_definitivos,
        "reducao_estimada_campos": max(campos_atuais_classificados - campos_fiscais_definitivos, 0),
        "campos_obrigatorios": {
            "ENTRADA_XLS_PCS": ["NUMERO_PC", "DATA_PC", "VALOR_PC", "STATUS_PAGAMENTO_PC"],
        },
        "nao_exigir_no_xls": list(_FISCAL_V2_NAO_PEDIR),
        "mensagem_produto": (
            "Este XLS pede somente os fatos necessarios. "
            "Todo calculo e decisao ficam com o Claus New."
        ),
        "compatibilidade": (
            "Layout de entrada definitivo coexiste com o MasterFile atual; abas oficiais "
            "sao preservadas para upload e conferencia da GCC."
        ),
    }


def _ciclo_por_data_fiscal(data_ref: Any, parametros: dict[str, Any]) -> str | None:
    return enquadrar_data_pc(data_ref, parametros.get("por_ciclo") or {})


def _ler_itens_contrato(wb) -> dict[str, Any]:
    """Le a aba visivel ITENS_CONTRATO (cadastro de item, qtd e VU original).

    Fallback de VU/QTD para consumidos e remanescentes quando as abas ocultas
    (historico_VU/execucao_saldo) nao vierem populadas pela Calculadora.
    """
    resultado: dict[str, Any] = {"itens": [], "por_item": {}, "ok": False}
    ws = _ws_entrada(wb, ABA_FISCAL_ITENS_CONTRATO)
    if ws is None:
        return resultado
    mapa = _mapear_colunas_por_cabecalho(ws)
    col_item = _col(mapa, "ITEM")
    col_qtd = _col(mapa, "QTD_CONTRATADA", "QUANTIDADE_CONTRATADA")
    col_vu = _col(mapa, "VU_ORIGINAL", "VALOR_UNITARIO_ORIGINAL", "VU")
    col_desc = _col(mapa, "DESCRICAO")
    col_vt = _col(mapa, "VALOR_TOTAL_ORIGINAL")
    cols_vu_ciclo = {c: _col(mapa, f"VU_{c}") for c in ("C1", "C2", "C3", "C4")}
    if not col_item:
        return resultado
    for r in range(2, ws.max_row + 1):
        if _linha_exemplo_fiscal(ws, r):
            continue
        item = str(ws.cell(r, col_item).value or "").strip()
        if not item:
            continue
        qtd_raw = ws.cell(r, col_qtd).value if col_qtd else None
        vu_raw = ws.cell(r, col_vu).value if col_vu else None
        desc_raw = ws.cell(r, col_desc).value if col_desc else None
        # Compatibilidade: linha A:D do layout anterior colada sob o header novo.
        if (not isinstance(vu_raw, (int, float))) and isinstance(desc_raw, (int, float)) and isinstance(qtd_raw, (int, float)):
            qtd_raw, vu_raw, desc_raw = desc_raw, qtd_raw, None
        registro = {
            "item": item,
            "qtd_contratada": _to_float_sombra(
                qtd_raw, default=0.0
            ),
            "vu_original": _to_float_sombra(
                vu_raw, default=0.0
            ),
            "descricao": desc_raw,
            "valor_total_original": _to_float_sombra(
                ws.cell(r, col_vt).value if col_vt else None, default=0.0
            ),
            "vu_ciclos": {
                ciclo: _to_float_sombra(ws.cell(r, col).value, default=0.0)
                for ciclo, col in cols_vu_ciclo.items() if col
            },
            "origem": ABA_FISCAL_ITENS_CONTRATO,
        }
        resultado["itens"].append(registro)
        resultado["por_item"][item] = registro
    resultado["ok"] = bool(resultado["itens"])
    return resultado


def _vu_por_item(res: dict[str, Any], item: Any) -> float | None:
    chave = str(item or "").strip()
    if not chave:
        return None
    for reg in (res.get("historico_vu") or {}).get("itens") or []:
        if str(reg.get("item") or "").strip() != chave:
            continue
        for valor in [reg.get("vu_original"), *((reg.get("vu_ciclos") or {}).values())]:
            val = _to_float_sombra(valor, default=0.0)
            if val:
                return val
    for reg in (res.get("execucao_saldo") or {}).get("itens") or []:
        if str(reg.get("item") or "").strip() == chave:
            val = _to_float_sombra(reg.get("vu_original"), default=0.0)
            if val:
                return val
    registro = ((res.get("itens_contrato") or {}).get("por_item") or {}).get(chave)
    if registro:
        val = _to_float_sombra(registro.get("vu_original"), default=0.0)
        if val:
            return val
    return None


def _contrato_por_item(res: dict[str, Any], item: Any) -> dict[str, Any]:
    chave = str(item or "").strip()
    for reg in (res.get("execucao_saldo") or {}).get("itens") or []:
        if str(reg.get("item") or "").strip() == chave:
            return dict(reg)
    registro = ((res.get("itens_contrato") or {}).get("por_item") or {}).get(chave)
    if registro:
        return dict(registro)
    return {}


def _materializar_valor_atualizado(
    parcelas: list[dict[str, Any]], parametros: dict[str, Any]
) -> None:
    """Grava fator do ciclo e valor atualizado (devido) em cada parcela.

    O VTA por composicao usa o executado ATUALIZADO (base x fator acumulado
    do ciclo), como nas apostilas; a base segue intacta em ``valor`` para
    retroativo e reconciliacao.
    """
    por_ciclo = (parametros or {}).get("por_ciclo") or {}
    for parcela in parcelas:
        ciclo = str(parcela.get("ciclo") or "").strip().upper()
        fator = _to_float_sombra(
            (por_ciclo.get(ciclo) or {}).get("fator_acumulado"), default=0.0
        )
        parcela["fator_acumulado"] = fator or None
        parcela["valor_atualizado"] = (
            round(_to_float_sombra(parcela.get("valor")) * fator, 2)
            if fator else None
        )


def _normalizar_fiscal_financeiro(wb, parametros: dict[str, Any]) -> tuple[list[dict[str, Any]], list[str]]:
    ws = _ws_entrada(wb, ABA_FISCAL_FINANCEIRO_V2)
    if ws is None:
        return [], []
    mapa = _mapear_colunas_por_cabecalho(ws)
    col_data = _col(mapa, "COMPETENCIA_MES", "DATA_EXECUCAO_OU_PAGAMENTO", "DATA_REFERENCIA")
    col_valor = _col(mapa, "VALOR_EXECUTADO_OU_PAGO", "VALOR")
    col_situacao = _col(mapa, "SITUACAO", "STATUS")
    col_fonte = _col(mapa, "FONTE")
    col_obs = _col(mapa, "OBSERVACAO")
    parcelas: list[dict[str, Any]] = []
    alertas: list[str] = []
    for r in range(2, ws.max_row + 1):
        if _linha_exemplo_fiscal(ws, r):
            continue
        valor = _to_float_sombra(ws.cell(r, col_valor).value if col_valor else None)
        data_ref = ws.cell(r, col_data).value if col_data else None
        if not valor and data_ref in (None, ""):
            continue
        ciclo = _ciclo_por_data_fiscal(data_ref, parametros)
        if not ciclo:
            alertas.append(
                f"ENTRADA_XLS_FINANCEIRO linha {r}: data ausente ou fora da linha temporal; "
                "evidencia financeira mantida como pendente."
            )
            continue
        situacao = str(ws.cell(r, col_situacao).value if col_situacao else "").strip().lower()
        tipo_fin = (
            "Retroativo Reconhecido"
            if "retro" in situacao or "reconhec" in situacao
            else "Execucao Atualizada"
        )
        parcelas.append({
            "linha": r,
            "identificador": f"entrada_xls_financeiro:{ciclo}:{r}",
            "origem_dado": "Financeiro",
            "tipo_parcela": "Execucao realizada",
            "tipo_financeiro": tipo_fin,
            "fonte_parcela": "Financeiro",
            "valor": valor,
            "justificativa_vta": (
                "Evidencia financeira informada no XLS definitivo; "
                "ciclo inferido pela data na linha temporal."
            ),
            "ciclo": ciclo,
            "data_referencia": data_ref,
            "fonte": ws.cell(r, col_fonte).value if col_fonte else None,
            "observacao": ws.cell(r, col_obs).value if col_obs else None,
        })
    return parcelas, alertas


_ORDEM_CICLOS = ["C0", "C1", "C2", "C3", "C4"]


def _montar_potencial_futuro(res: dict[str, Any]) -> dict[str, Any]:
    """Potencial futuro: saldo remanescente valorado pelo fator concedido.

    Nao projeta indice de ciclo ainda nao concedido — ciclos restantes sem
    fator parametrizado aparecem listados sem valor, como pendencia.
    """
    saida: dict[str, Any] = {
        "disponivel": False,
        "motivo": "",
        "saldo_remanescente_base": None,
        "ciclo_vigente": "",
        "fator_vigente": None,
        "valor_atualizado_vigente": None,
        "ciclos_restantes": [],
        "projecao_por_ciclo": {},
    }
    resumo = res.get("resumo") or {}
    saldo = _to_float_sombra(resumo.get("saldo_remanescente"))
    if not saldo:
        saida["motivo"] = "Sem saldo remanescente informado na data de corte."
        return saida
    ciclo_vigente = str((res.get("controle") or {}).get("ciclo_vigente") or "").strip().upper()
    por_ciclo = (res.get("parametros_v10") or {}).get("por_ciclo") or {}
    if ciclo_vigente not in por_ciclo:
        saida["motivo"] = "Ciclo vigente nao parametrizado."
        saida["saldo_remanescente_base"] = saldo
        return saida
    fator_vigente = _to_float_sombra(
        por_ciclo[ciclo_vigente].get("fator_acumulado"), default=0.0
    )
    if not fator_vigente:
        saida["motivo"] = f"Fator acumulado do ciclo vigente {ciclo_vigente} ausente."
        saida["saldo_remanescente_base"] = saldo
        saida["ciclo_vigente"] = ciclo_vigente
        return saida

    restantes = [
        c for c in _ORDEM_CICLOS
        if c in por_ciclo and _ORDEM_CICLOS.index(c) > _ORDEM_CICLOS.index(ciclo_vigente)
    ]
    projecao: dict[str, Any] = {}
    for ciclo in restantes:
        fator = _to_float_sombra(por_ciclo[ciclo].get("fator_acumulado"), default=0.0)
        projecao[ciclo] = {
            "fator_acumulado": fator or None,
            "valor_projetado": round(saldo * fator, 2) if fator else None,
            "situacao": (
                "fator concedido" if fator else "fator ainda nao concedido — sem projecao"
            ),
        }
    saida.update({
        "disponivel": True,
        "saldo_remanescente_base": saldo,
        "ciclo_vigente": ciclo_vigente,
        "fator_vigente": fator_vigente,
        "valor_atualizado_vigente": round(saldo * fator_vigente, 2),
        "ciclos_restantes": restantes,
        "projecao_por_ciclo": projecao,
    })
    return saida


def _meses_do_periodo(inicio, fim) -> list[Any]:
    """Dia 15 de cada competencia mensal do periodo (evita bordas de ciclo)."""
    from datetime import date as _date
    meses = []
    ano, mes = inicio.year, inicio.month
    while (ano, mes) <= (fim.year, fim.month):
        meses.append(_date(ano, mes, 15))
        mes += 1
        if mes > 12:
            mes, ano = 1, ano + 1
    return meses


def _parcela_historico(r: int, ciclo: str, valor: float, data_ref: Any,
                       fonte: Any, obs: Any, confianca: str,
                       justificativa: str, sufixo: str = "") -> dict[str, Any]:
    return {
        "linha": r,
        "identificador": f"ciclos_passados:{ciclo}:{r}{sufixo}",
        "origem_dado": "Financeiro",
        "tipo_parcela": "Execucao realizada",
        "tipo_financeiro": "Execucao Atualizada",
        "fonte_parcela": "Historico financeiro",
        "valor": valor,
        "justificativa_vta": justificativa,
        "ciclo": ciclo,
        "data_referencia": data_ref,
        "confianca": confianca,
        "fonte": fonte,
        "observacao": obs,
    }


def _normalizar_fiscal_historico_financeiro(wb, parametros: dict[str, Any]) -> tuple[list[dict[str, Any]], list[str]]:
    ws = _ws_entrada(wb, ABA_FISCAL_HISTORICO_FINANCEIRO)
    if ws is None:
        return [], []
    mapa = _mapear_colunas_por_cabecalho(ws)
    col_data = _col(
        mapa,
        "DATA_EXECUCAO_OU_PAGAMENTO_PASSADOS",
        "COMPETENCIA_PASSADA",
        "DATA_REFERENCIA",
    )
    col_fim = _col(mapa, "PERIODO_FIM", "DATA_FIM_PERIODO")
    col_valor = _col(
        mapa,
        "VALOR_EXECUTADO_OU_PAGO_PASSADOS",
        "VALOR_EXECUTADO_OU_PAGO",
        "VALOR",
    )
    col_conf = _col(mapa, "CONFIANCA", "NIVEL_CONFIANCA")
    col_fonte = _col(mapa, "FONTE")
    col_obs = _col(mapa, "OBSERVACAO")
    parcelas: list[dict[str, Any]] = []
    alertas: list[str] = []
    for r in range(2, ws.max_row + 1):
        if _linha_exemplo_fiscal(ws, r):
            continue
        valor = _to_float_sombra(ws.cell(r, col_valor).value if col_valor else None)
        data_ref = ws.cell(r, col_data).value if col_data else None
        if not valor and data_ref in (None, ""):
            continue
        fonte = ws.cell(r, col_fonte).value if col_fonte else None
        obs = ws.cell(r, col_obs).value if col_obs else None
        confianca = str(ws.cell(r, col_conf).value or "").strip() if col_conf else ""
        confianca = confianca or "ciclo inferido pela data"

        data_fim = ws.cell(r, col_fim).value if col_fim else None
        ini_norm = _normalizar_data(data_ref)
        fim_norm = _normalizar_data(data_fim)
        if not hasattr(ini_norm, "year"):
            ini_norm = None
        if not hasattr(fim_norm, "year"):
            if fim_norm not in (None, ""):
                alertas.append(
                    f"CICLOS_PASSADOS linha {r}: PERIODO_FIM invalido "
                    f"({data_fim!r}); linha tratada como pontual."
                )
            fim_norm = None

        # Linha pontual ou de periodo dentro de um unico ciclo: sem rateio.
        if not fim_norm or not ini_norm or fim_norm <= ini_norm:
            ciclo = _ciclo_por_data_fiscal(data_ref, parametros)
            if not ciclo:
                alertas.append(
                    f"CICLOS_PASSADOS linha {r}: data ausente ou fora da linha temporal; "
                    "historico mantido como evidencia pendente."
                )
                continue
            parcelas.append(_parcela_historico(
                r, ciclo, valor, data_ref, fonte, obs, confianca,
                "Historico financeiro passado informado no XLS definitivo; "
                "ciclo inferido pela data na linha temporal.",
            ))
            continue

        # Periodo com inicio e fim: rateio pro-rata por meses de cada ciclo.
        meses = _meses_do_periodo(ini_norm, fim_norm)
        meses_por_ciclo: dict[str, int] = {}
        meses_fora = 0
        for mes_ref in meses:
            ciclo_mes = _ciclo_por_data_fiscal(mes_ref, parametros)
            if ciclo_mes:
                meses_por_ciclo[ciclo_mes] = meses_por_ciclo.get(ciclo_mes, 0) + 1
            else:
                meses_fora += 1
        if not meses_por_ciclo:
            alertas.append(
                f"CICLOS_PASSADOS linha {r}: periodo inteiro fora da linha temporal; "
                "historico mantido como evidencia pendente."
            )
            continue
        if len(meses_por_ciclo) == 1:
            ciclo = next(iter(meses_por_ciclo))
            parcelas.append(_parcela_historico(
                r, ciclo, valor, data_ref, fonte, obs, confianca,
                "Historico financeiro de periodo dentro de um unico ciclo; "
                "sem perda de precisao (fator do ciclo e unico).",
            ))
        else:
            total_meses = len(meses)
            acumulado = 0.0
            itens_ciclo = sorted(meses_por_ciclo.items())
            for idx, (ciclo, qtd_meses) in enumerate(itens_ciclo):
                if idx == len(itens_ciclo) - 1 and not meses_fora:
                    parte = round(valor - acumulado, 2)
                else:
                    parte = round(valor * qtd_meses / total_meses, 2)
                acumulado += parte
                parcelas.append(_parcela_historico(
                    r, ciclo, parte, data_ref, fonte, obs, "estimado por rateio",
                    f"Rateio pro-rata: {qtd_meses} de {total_meses} meses do "
                    f"periodo caem em {ciclo}. Valor estimado; confirmar com "
                    "financeiro mensal quando disponivel.",
                    sufixo=f":{ciclo.lower()}",
                ))
            alertas.append(
                f"CICLOS_PASSADOS linha {r}: valor de periodo rateado pro-rata "
                f"entre {', '.join(c for c, _ in itens_ciclo)}; parcelas marcadas "
                "como estimadas."
            )
        if meses_fora:
            valor_fora = round(valor * meses_fora / len(meses), 2)
            alertas.append(
                f"CICLOS_PASSADOS linha {r}: {meses_fora} mes(es) do periodo fora "
                f"da linha temporal (~R$ {valor_fora:,.2f} nao enquadrado)."
            )
    return parcelas, alertas


def _normalizar_fiscal_aditivos(wb, parametros: dict[str, Any]) -> dict[str, Any]:
    """Le a aba visivel ENTRADA_XLS_ADITIVOS (evento, assinatura, valor).

    O ciclo-marco e inferido pela DATA_ASSINATURA na linha temporal e o valor
    atualizado usa o fator acumulado do ciclo-marco, como nas apostilas
    (supressao = valor negativo). JA_REFLETIDO_EM diferente de "Nao" marca o
    evento como ja refletido em outra fonte (nao soma ao VTA).
    """
    resultado: dict[str, Any] = {"itens": [], "totais": {}, "alertas": [], "ok": False}
    ws = _ws_entrada(wb, ABA_FISCAL_ADITIVOS)
    if ws is None:
        return resultado
    mapa = _mapear_colunas_por_cabecalho(ws)
    col_evento = _col(mapa, "EVENTO", "ADITIVO", "DESCRICAO")
    col_data = _col(mapa, "DATA_ASSINATURA", "DATA_DO_EVENTO", "DATA")
    col_valor = _col(mapa, "VALOR_NA_ASSINATURA", "VALOR_ASSINATURA", "VALOR")
    col_fator = _col(mapa, "FATOR_APLICADO", "FATOR_FORMALIZADO")
    col_refletido = _col(mapa, "JA_REFLETIDO_EM")
    col_obs = _col(mapa, "OBSERVACAO")
    por_ciclo = (parametros or {}).get("por_ciclo") or {}
    total_computavel = 0.0
    for r in range(2, ws.max_row + 1):
        if _linha_exemplo_fiscal(ws, r):
            continue
        evento = str(ws.cell(r, col_evento).value or "").strip() if col_evento else ""
        valor = _to_float_sombra(ws.cell(r, col_valor).value if col_valor else None)
        data_ass = ws.cell(r, col_data).value if col_data else None
        if not evento and not valor:
            continue
        if not valor:
            resultado["alertas"].append(
                f"ENTRADA_XLS_ADITIVOS linha {r}: evento sem VALOR_NA_ASSINATURA; "
                "registrado como pendencia, nao compoe o VTA."
            )
            continue
        ciclo_marco = _ciclo_por_data_fiscal(data_ass, parametros)
        fator = _to_float_sombra(
            (por_ciclo.get(ciclo_marco) or {}).get("fator_acumulado"), default=0.0
        ) if ciclo_marco else 0.0
        fator_origem = "parametros (ciclo-marco)"
        fator_informado = _to_float_sombra(
            ws.cell(r, col_fator).value if col_fator else None
        )
        if fator_informado > 0:
            # Ato formal pode ter fixado fator proprio (ex.: percentual
            # publicado arredondado); prevalece sobre o parametrico.
            fator = fator_informado
            fator_origem = "FATOR_APLICADO informado no XLS"
        if not ciclo_marco and not fator_informado:
            resultado["alertas"].append(
                f"ENTRADA_XLS_ADITIVOS linha {r}: DATA_ASSINATURA ausente ou fora "
                "da linha temporal; aditivo composto pelo valor da assinatura, "
                "sem atualizacao."
            )
        refletido = str(
            ws.cell(r, col_refletido).value if col_refletido else ""
        ).strip() or "Nao"
        if _norm(refletido) in ("nao", "não", "no"):
            refletido = "Nao"
        valor_atualizado = round(valor * fator, 2) if fator else round(valor, 2)
        if refletido == "Nao":
            total_computavel += valor_atualizado
        resultado["itens"].append({
            "linha": r,
            "evento": evento or f"Aditivo linha {r}",
            "data_assinatura": data_ass,
            "ciclo_marco": ciclo_marco,
            "valor_assinatura": round(valor, 2),
            "fator_acumulado": fator or None,
            "fator_origem": fator_origem,
            "valor_atualizado": valor_atualizado,
            "ja_refletido_em": refletido,
            "natureza": "Supressao" if valor < 0 else "Aditivo",
            "observacao": ws.cell(r, col_obs).value if col_obs else None,
            "origem": ABA_FISCAL_ADITIVOS,
        })
    resultado["totais"] = {
        "count_itens": len(resultado["itens"]),
        "total_valor_assinatura": round(
            sum(i["valor_assinatura"] for i in resultado["itens"]), 2
        ),
        "total_valor_atualizado_computavel": round(total_computavel, 2),
    }
    resultado["ok"] = bool(resultado["itens"])
    return resultado


def _normalizar_fiscal_consumidos(wb, res: dict[str, Any]) -> dict[str, Any]:
    resultado: dict[str, Any] = {"itens": [], "totais": {}, "alertas": [], "ok": False}
    ws = _ws_entrada(wb, ABA_FISCAL_CONSUMIDOS_V2)
    if ws is None:
        return resultado
    mapa = _mapear_colunas_por_cabecalho(ws)
    col_item = _col(mapa, "ITEM")
    col_qtd = _col(mapa, "QTD_CONSUMIDA", "QUANTIDADE_CONSUMIDA")
    col_data = _col(mapa, "DATA_REFERENCIA", "DATA_CONSUMO")
    col_obs = _col(mapa, "OBSERVACAO")
    parametros = res.get("parametros_v10") or {}
    for r in range(2, ws.max_row + 1):
        if _linha_exemplo_fiscal(ws, r):
            continue
        item = ws.cell(r, col_item).value if col_item else None
        qtd = _to_float_sombra(ws.cell(r, col_qtd).value if col_qtd else None)
        data_ref = ws.cell(r, col_data).value if col_data else None
        if not item and not qtd:
            continue
        ciclo = _ciclo_por_data_fiscal(data_ref, parametros)
        if not ciclo:
            resultado["alertas"].append(
                f"ENTRADA_XLS_CONSUMIDOS linha {r}: DATA_REFERENCIA ausente ou fora dos ciclos; "
                "consumo nao enquadrado."
            )
        vu = _vu_por_item(res, item)
        valor = round(qtd * vu, 2) if qtd and vu else None
        if valor is None:
            resultado["alertas"].append(
                f"ENTRADA_XLS_CONSUMIDOS linha {r}: sem VU de referencia para calcular valor; "
                "quantidade mantida como fato de entrada XLS."
            )
        consumos = {f"C{i}": {"qtd": None, "valor": None} for i in range(5)}
        if ciclo:
            consumos[ciclo] = {"qtd": qtd, "valor": valor}
        resultado["itens"].append({
            "item": item,
            "qtd_contratada": None,
            "vu_original": vu,
            "consumos": consumos,
            "qtd_total": qtd,
            "valor_total": valor,
            "check": "",
            "data_referencia": data_ref,
            "ciclo_inferido": ciclo,
            "observacao": ws.cell(r, col_obs).value if col_obs else None,
            "origem": "ENTRADA_XLS_CONSUMIDOS",
        })
    resultado["totais"] = {
        "qtd_total": sum(_to_float_sombra(i.get("qtd_total")) for i in resultado["itens"]),
        "valor_total": sum(_to_float_sombra(i.get("valor_total")) for i in resultado["itens"]),
    }
    resultado["ok"] = bool(resultado["itens"])
    return resultado


def _normalizar_fiscal_remanescentes(wb, res: dict[str, Any]) -> dict[str, Any]:
    resultado: dict[str, Any] = {
        "itens": [], "totais": {}, "alertas": [], "ok": False,
        "aba_presente": bool(_aba_entrada_existente(wb, ABA_FISCAL_REMANESCENTES_V2)),
    }
    ws = _ws_entrada(wb, ABA_FISCAL_REMANESCENTES_V2)
    if ws is None:
        return resultado
    mapa = _mapear_colunas_por_cabecalho(ws)
    if _col(mapa, "QTD_REM_C0"):
        resultado["fotografias_ciclo"] = []
        parametros = res.get("parametros_v10") or {}
        col_item_largo = _col(mapa, "ITEM")
        col_atual = _col(mapa, "QTD_REM_ATUAL")
        col_data_atual = _col(mapa, "DATA_CORTE_ATUAL")
        col_ciclo_atual = _col(mapa, "CICLO_ATUAL")
        col_obs_largo = _col(mapa, "OBSERVACAO")
        for r in range(2, ws.max_row + 1):
            if _linha_exemplo_fiscal(ws, r):
                continue
            item = ws.cell(r, col_item_largo).value if col_item_largo else None
            if not item:
                continue
            if _normalizar_data(item) is not None and isinstance(ws.cell(r, 3).value, str):
                # Compatibilidade: linha do layout longo anterior anexada ao arquivo novo.
                data_legado = item
                item_legado = ws.cell(r, 3).value
                qtd_legado = _to_float_sombra(ws.cell(r, 4).value)
                contrato = _contrato_por_item(res, item_legado)
                vu = _to_float_sombra(contrato.get("vu_original"), default=0.0) or _vu_por_item(res, item_legado)
                ciclo_legado = _ciclo_por_data_fiscal(data_legado, parametros)
                foto = {
                    "item": item_legado, "ciclo": ciclo_legado,
                    "data_referencia": data_legado, "qtd_remanescente": qtd_legado,
                    "vu_original": vu, "valor_original": round(qtd_legado * vu, 2) if vu else None,
                    "tipo_fotografia": "CICLO_ATUAL_EM_EXECUCAO", "origem_linha": r,
                }
                resultado["fotografias_ciclo"].append(foto)
                qtd_contratada = _to_float_sombra(contrato.get("qtd_contratada"), default=0.0)
                resultado["itens"].append({
                    "item": item_legado, "descricao": contrato.get("descricao"),
                    "qtd_contratada": qtd_contratada, "vu_original": vu,
                    "valor_total_original": _to_float_sombra(contrato.get("valor_total_original"), default=0.0),
                    "qtd_emitida": max(qtd_contratada - qtd_legado, 0.0) if qtd_contratada else 0.0,
                    "valor_emitido": 0.0, "qtd_saldo": qtd_legado,
                    "valor_saldo": foto["valor_original"] or 0.0,
                    "data_referencia": data_legado, "ciclo_inferido": ciclo_legado,
                    "tipo_corte": "CICLO_ATUAL_EM_EXECUCAO", "observacao": ws.cell(r, 5).value,
                    "origem": "ENTRADA_XLS_REMANESCENTES",
                })
                continue
            contrato = _contrato_por_item(res, item)
            vu = _to_float_sombra(contrato.get("vu_original"), default=0.0) or _vu_por_item(res, item)
            fotos_item = []
            for ciclo in ("C0", "C1", "C2", "C3", "C4"):
                col_q = _col(mapa, f"QTD_REM_{ciclo}")
                col_d = _col(mapa, f"DATA_INICIO_{ciclo}")
                bruto = ws.cell(r, col_q).value if col_q else None
                if bruto in (None, ""):
                    continue
                qtd = _to_float_sombra(bruto)
                data_ciclo = ws.cell(r, col_d).value if col_d else None
                if data_ciclo in (None, ""):
                    data_ciclo = (parametros.get("por_ciclo") or {}).get(ciclo, {}).get("data_inicio")
                foto = {
                    "item": item, "ciclo": ciclo,
                    "data_referencia": data_ciclo,
                    "qtd_remanescente": qtd,
                    "vu_original": vu,
                    "valor_original": round(qtd * vu, 2) if vu else None,
                    "tipo_fotografia": "INICIO_CICLO", "origem_linha": r,
                }
                fotos_item.append(foto)
                resultado["fotografias_ciclo"].append(foto)
            bruto_atual = ws.cell(r, col_atual).value if col_atual else None
            if bruto_atual not in (None, ""):
                ciclo_atual = str(
                    ws.cell(r, col_ciclo_atual).value if col_ciclo_atual else ""
                ).strip().upper()
                if not ciclo_atual:
                    ciclo_atual = str((res.get("controle") or {}).get("ciclo_vigente") or "").strip().upper()
                data_atual = ws.cell(r, col_data_atual).value if col_data_atual else None
                if data_atual in (None, ""):
                    data_atual = (res.get("controle") or {}).get("data_corte")
                qtd_atual = _to_float_sombra(bruto_atual)
                foto = {
                    "item": item, "ciclo": ciclo_atual,
                    "data_referencia": data_atual,
                    "qtd_remanescente": qtd_atual, "vu_original": vu,
                    "valor_original": round(qtd_atual * vu, 2) if vu else None,
                    "tipo_fotografia": "CICLO_ATUAL_EM_EXECUCAO", "origem_linha": r,
                }
                fotos_item.append(foto)
                resultado["fotografias_ciclo"].append(foto)
            if not fotos_item:
                continue
            atual = next((f for f in reversed(fotos_item) if f["tipo_fotografia"] == "CICLO_ATUAL_EM_EXECUCAO"), fotos_item[-1])
            qtd_contratada = _to_float_sombra(contrato.get("qtd_contratada"), default=0.0)
            resultado["itens"].append({
                "item": item, "descricao": contrato.get("descricao"),
                "qtd_contratada": qtd_contratada, "vu_original": vu,
                "valor_total_original": _to_float_sombra(contrato.get("valor_total_original"), default=0.0),
                "qtd_emitida": max(qtd_contratada - atual["qtd_remanescente"], 0.0) if qtd_contratada else 0.0,
                "valor_emitido": 0.0, "qtd_saldo": atual["qtd_remanescente"],
                "valor_saldo": atual["valor_original"] or 0.0,
                "data_referencia": atual["data_referencia"],
                "ciclo_inferido": atual["ciclo"], "tipo_corte": atual["tipo_fotografia"],
                "observacao": ws.cell(r, col_obs_largo).value if col_obs_largo else None,
                "origem": "ENTRADA_XLS_REMANESCENTES",
            })
        # Uma fotografia corrente por item alimenta saldo/VTA; as demais ficam
        # preservadas exclusivamente na memoria historica por ciclo.
        atuais_por_item: dict[str, dict[str, Any]] = {}
        for reg in resultado["itens"]:
            chave = str(reg.get("item") or "")
            anterior = atuais_por_item.get(chave)
            data_reg = _normalizar_data(reg.get("data_referencia"))
            data_ant = _normalizar_data((anterior or {}).get("data_referencia"))
            if anterior is None or (data_reg is not None and (data_ant is None or data_reg >= data_ant)):
                atuais_por_item[chave] = reg
        resultado["itens"] = list(atuais_por_item.values())
        resultado["totais"] = {
            "count_itens": len(resultado["itens"]),
            "total_qtd_saldo": sum(i["qtd_saldo"] for i in resultado["itens"]),
            "total_valor_saldo": sum(i["valor_saldo"] for i in resultado["itens"]),
            "fotografias_ciclo": len(resultado["fotografias_ciclo"]),
        }
        resultado["ok"] = bool(resultado["itens"])
        return resultado
    col_item = _col(mapa, "ITEM")
    col_qtd = _col(mapa, "QTD_REMANESCENTE")
    col_valor = _col(mapa, "VALOR_REMANESCENTE")
    col_data = _col(mapa, "DATA_CORTE", "DATA_REFERENCIA")
    col_tipo_corte = _col(mapa, "TIPO_CORTE")
    col_obs = _col(mapa, "OBSERVACAO")
    parametros = res.get("parametros_v10") or {}
    total_valor_saldo = 0.0

    # A aba pode trazer varias fotografias por item (inicio de cada ciclo +
    # corte). O saldo do contrato usa somente a fotografia do corte de cada
    # item — somar fotografias de datas diferentes multiplica o estoque
    # (achado PADTEC: 4 fotografias somadas).
    data_corte = _normalizar_data((res.get("controle") or {}).get("data_corte"))
    if not hasattr(data_corte, "year"):
        data_corte = None
    candidatas: dict[str, list[tuple[Any, int]]] = {}
    for r in range(2, ws.max_row + 1):
        if _linha_exemplo_fiscal(ws, r):
            continue
        item = ws.cell(r, col_item).value if col_item else None
        qtd_saldo = _to_float_sombra(ws.cell(r, col_qtd).value if col_qtd else None)
        valor_saldo = _to_float_sombra(ws.cell(r, col_valor).value if col_valor else None)
        if not item and not qtd_saldo and not valor_saldo:
            continue
        data_norm = _normalizar_data(ws.cell(r, col_data).value if col_data else None)
        if not hasattr(data_norm, "year"):
            data_norm = None
        chave = str(item or f"__linha_{r}").strip()
        candidatas.setdefault(chave, []).append((data_norm, r))

    linhas_selecionadas: list[int] = []
    fotografias_descartadas = 0
    for chave, fotos in candidatas.items():
        datas = [d for d, _ in fotos if d is not None]
        if not datas:
            linhas_selecionadas.extend(r for _, r in fotos)
            continue
        no_corte = [d for d in datas if data_corte is None or d <= data_corte]
        data_escolhida = max(no_corte) if no_corte else min(datas)
        for data_norm, r in fotos:
            if data_norm == data_escolhida or data_norm is None:
                linhas_selecionadas.append(r)
            else:
                fotografias_descartadas += 1
    if fotografias_descartadas:
        resultado["alertas"].append(
            f"ENTRADA_XLS_REMANESCENTES: {fotografias_descartadas} fotografia(s) "
            "de outras datas mantidas como historico; o saldo usa apenas a "
            "fotografia do corte de cada item."
        )

    for r in sorted(linhas_selecionadas):
        item = ws.cell(r, col_item).value if col_item else None
        qtd_saldo = _to_float_sombra(ws.cell(r, col_qtd).value if col_qtd else None)
        valor_saldo = _to_float_sombra(ws.cell(r, col_valor).value if col_valor else None)
        data_ref = ws.cell(r, col_data).value if col_data else None
        ciclo = _ciclo_por_data_fiscal(data_ref, parametros)
        if not ciclo:
            resultado["alertas"].append(
                f"ENTRADA_XLS_REMANESCENTES linha {r}: DATA_REFERENCIA ausente ou fora dos ciclos; "
                "remanescente mantido como fato sem enquadramento temporal."
            )
        contrato = _contrato_por_item(res, item)
        vu = _to_float_sombra(contrato.get("vu_original"), default=0.0) or _vu_por_item(res, item)
        qtd_contratada = _to_float_sombra(contrato.get("qtd_contratada"), default=0.0)
        if not valor_saldo and qtd_saldo and vu:
            valor_saldo = round(qtd_saldo * vu, 2)
        if not qtd_contratada and qtd_saldo:
            qtd_contratada = qtd_saldo
            resultado["alertas"].append(
                f"ENTRADA_XLS_REMANESCENTES linha {r}: QTD_CONTRATADA nao localizada; "
                "saldo usado apenas como evidencia de entrada XLS."
            )
        qtd_emitida = max(qtd_contratada - qtd_saldo, 0.0) if qtd_contratada else 0.0
        valor_emitido = _to_float_sombra(contrato.get("valor_emitido"), default=0.0)
        total_valor_saldo += valor_saldo
        resultado["itens"].append({
            "item": item,
            "descricao": contrato.get("descricao"),
            "qtd_contratada": qtd_contratada,
            "vu_original": vu,
            "valor_total_original": _to_float_sombra(contrato.get("valor_total_original"), default=0.0),
            "pc": contrato.get("pc"),
            "requisicao_sap": contrato.get("requisicao_sap"),
            "qtd_emitida": qtd_emitida,
            "valor_emitido": valor_emitido,
            "qtd_saldo": qtd_saldo,
            "valor_saldo": valor_saldo,
            "check_fisico": "",
            "valor_total_calculado": round(qtd_contratada * vu, 2) if qtd_contratada and vu else None,
            "check_fisico_calculado": "",
            "data_referencia": data_ref,
            "ciclo_inferido": ciclo,
            "tipo_corte": ws.cell(r, col_tipo_corte).value if col_tipo_corte else None,
            "observacao": ws.cell(r, col_obs).value if col_obs else None,
            "origem": "ENTRADA_XLS_REMANESCENTES",
        })
    resultado["totais"] = {
        "count_itens": len(resultado["itens"]),
        "total_qtd_contratada": sum(i["qtd_contratada"] for i in resultado["itens"]),
        "total_valor_emitido": sum(i["valor_emitido"] for i in resultado["itens"]),
        "total_qtd_saldo": sum(i["qtd_saldo"] for i in resultado["itens"]),
        "total_valor_saldo": total_valor_saldo,
    }
    resultado["ok"] = bool(resultado["itens"])
    return resultado


def _selecionar_aba_pc(wb):
    ws_fiscal = _ws_entrada(wb, ABA_FISCAL_PCS_V2)
    if ws_fiscal is not None:
        mapa_fiscal = _mapear_colunas_por_cabecalho(ws_fiscal)
        if (
            _col(mapa_fiscal, "NUMERO_PC")
            and _col(mapa_fiscal, "DATA_PC")
            and _col(mapa_fiscal, "VALOR_PC")
            and _aba_tem_dados(ws_fiscal)
        ):
            return ws_fiscal, True
    if ABA_ITENS_PC in wb.sheetnames:
        return wb[ABA_ITENS_PC], False
    return None, False


def _validar_campos_vta(registro_vta: dict[str, str], ident: str,
                        linha: int, alertas: list[str]) -> None:
    computa = registro_vta["computa_vta"]
    tipo_fin = registro_vta["tipo_financeiro"]
    refletido = registro_vta["ja_refletido_em"]
    status = registro_vta["status_consolidacao"]

    if computa == "Sim" and refletido != "Nao":
        alertas.append(
            f"PC '{ident}' linha {linha}: COMPUTA_VTA=Sim com "
            f"JA_REFLETIDO_EM={refletido}; potencial duplicidade."
        )

    if status == "COMPUTADO" and computa != "Sim":
        alertas.append(
            f"PC '{ident}' linha {linha}: STATUS_CONSOLIDACAO=COMPUTADO "
            "mas COMPUTA_VTA nao e Sim."
        )

    if status == "COMPUTADO" and tipo_fin in ("Impacto Potencial", "Informativo"):
        alertas.append(
            f"PC '{ident}' linha {linha}: TIPO_FINANCEIRO={tipo_fin} "
            "incompativel com STATUS_CONSOLIDACAO=COMPUTADO."
        )


def _atualizar_resumo_vta(resumo: dict[str, Any],
                          registro_vta: dict[str, str]) -> None:
    resumo["linhas_total"] += 1
    status = registro_vta["status_consolidacao"]
    computa = registro_vta["computa_vta"]
    refletido = registro_vta["ja_refletido_em"]

    if status == "INCONSISTENTE":
        resumo["inconsistentes"] += 1
    elif status == "EM_ANALISE":
        resumo["em_analise"] += 1
    elif computa == "Sim" and refletido == "Nao" and status == "COMPUTADO":
        resumo["computaveis"] += 1
    else:
        resumo["nao_computaveis"] += 1

    if status == "DESCARTADO_DUPLICIDADE" or refletido != "Nao":
        resumo["descartadas_duplicidade"] += 1


def _to_float_sombra(valor: Any, default: float = 0.0) -> float:
    try:
        if valor in (None, ""):
            return default
        return float(valor)
    except (TypeError, ValueError):
        return default


def _ciclo_por_competencia(wb, competencia: Any) -> str:
    """Deriva o rotulo do ciclo (C0..C4) da competencia pela CRONOLOGIA FIXA.

    ETAPA 31 — espelha a formula de financeiro!B: bloco fixo de 12 meses
    contado do inicio cronologico do ciclo mais antigo com DATA_INICIO real
    em parametros (nao a janela de reajuste C:D, que pode conter intervalo
    intencional). Antes da ancora, C0; alem de C4, o ultimo bloco nao e
    extrapolado (devolve ""). Usado quando financeiro!B e formula sem cache.
    """
    if not isinstance(competencia, (date, datetime)):
        return ""
    comp = competencia.date() if isinstance(competencia, datetime) else competencia
    try:
        params = wb["parametros"]
    except KeyError:
        return ""
    ancora = None
    indice_ancora = 0
    for r in range(2, 7):
        rotulo = str(params.cell(r, 2).value or "").strip().upper()
        inicio = params.cell(r, 3).value
        if isinstance(inicio, datetime):
            inicio = inicio.date()
        if rotulo in ("C0", "C1", "C2", "C3", "C4") and isinstance(inicio, date):
            ancora = inicio.replace(day=1)
            indice_ancora = int(rotulo[1])
            break
    if ancora is None:
        return ""
    meses = (comp.year - ancora.year) * 12 + (comp.month - ancora.month)
    if meses >= 0:
        indice = indice_ancora + meses // 12
    else:
        indice = indice_ancora - ((-meses + 11) // 12)
    if indice < 0:
        return "C0"
    if indice > 4:
        return ""
    return f"C{indice}"


def _ler_parcelas_sombra_financeiro(wb) -> list[dict[str, Any]]:
    """Materializa parcelas de financeiro sem recalcular nem escrever no XLS.

    Layouts suportados (campos localizados por cabecalho, com fallback
    posicional):
    - novo modelo oficial: historico mensal a partir da LINHA 2 (A e data);
      B (CICLO) e formula do template — sem cache, o ciclo e derivado da
      competencia pela aba parametros;
    - layout v10.2 legado: agregados nas linhas 2-5 ou mensal a partir da 7.
    Toda parcela carrega a chave "ciclo" (rotulo em maiusculas) para a
    memoria por ciclo do objeto_processo.
    """
    if "financeiro" not in wb.sheetnames:
        return []

    ws = wb["financeiro"]
    mapa = _mapear_colunas_por_cabecalho(ws)
    col_comp = _col(mapa, "COMPETENCIA") or 1
    col_ciclo = _col(mapa, "CICLO") or 2
    col_pago = _col(mapa, "VALOR_PAGO") or 3
    col_delta = _col(mapa, "DELTA") or 6
    col_efeito = _col(mapa, "EFEITO_FINANCEIRO") or 7

    parcelas: list[dict[str, Any]] = []
    novo_mensal = isinstance(ws.cell(2, col_comp).value, (date, datetime))
    if novo_mensal:
        linhas: Any = range(2, min(ws.max_row, 200) + 1)
    else:
        rows_agregadas = range(2, 6)
        tem_agregado = any(
            _to_float_sombra(ws.cell(r, col_pago).value)
            or _to_float_sombra(ws.cell(r, col_delta).value)
            for r in rows_agregadas
        )
        linhas = rows_agregadas if tem_agregado else range(7, min(ws.max_row, 200) + 1)

    for r in linhas:
        ciclo = str(ws.cell(r, col_ciclo).value or "").strip()
        if not ciclo and novo_mensal:
            ciclo = _ciclo_por_competencia(wb, ws.cell(r, col_comp).value)
        if not ciclo:
            continue

        efeito = str(ws.cell(r, col_efeito).value or "").strip()
        delta = _to_float_sombra(ws.cell(r, col_delta).value)

        base = _to_float_sombra(ws.cell(r, col_pago).value)
        if base:
            # Novo layout mensal: valor atualizado REAL da competencia
            # (base + delta com EFEITO_FINANCEIRO=Sim; sem efeito, base pura)
            # — evita retroativo inventado por fator quando G="Nao" e evita
            # dupla contagem com a parcela-delta na memoria por ciclo.
            atualizado = None
            if novo_mensal:
                atualizado = (
                    round(base + delta, 2)
                    if (efeito == "Sim" and delta) else base
                )
            parcelas.append({
                "linha": r,
                "identificador": f"financeiro:{ciclo}:base:{r}",
                "origem_dado": "Financeiro",
                "tipo_parcela": "Execucao realizada",
                "tipo_financeiro": "Execucao Atualizada",
                "fonte_parcela": "Financeiro",
                "ciclo": ciclo.upper(),
                "valor": base,
                "valor_atualizado": atualizado,
                "justificativa_vta": "Base executada do financeiro; G nao exclui base.",
            })
        if delta and efeito == "Sim":
            parcelas.append({
                "linha": r,
                "identificador": f"financeiro:{ciclo}:delta:{r}",
                "origem_dado": "Financeiro",
                "tipo_parcela": "Execucao realizada",
                "tipo_financeiro": "Retroativo Reconhecido",
                "fonte_parcela": "Financeiro",
                "ciclo": ciclo.upper(),
                "valor": delta,
                "justificativa_vta": "Delta financeiro computado apenas com EFEITO_FINANCEIRO=Sim.",
            })

    return parcelas


def _ler_parcelas_sombra_saldo(wb) -> list[dict[str, Any]]:
    """Materializa saldo remanescente como parcela agregada auditavel.

    Novo modelo oficial: nao existe a aba historico — o remanescente por
    item/ciclo passa a vir de posicao_contratual (memoria por ciclo do
    objeto_processo); nenhuma parcela agregada e criada aqui e nenhum
    dado sintetico substitui historico!B5.
    """
    if "historico" not in wb.sheetnames:
        return []

    ws = wb["historico"]
    valor = _to_float_sombra(ws["B5"].value)
    if not valor:
        return []
    return [{
        "linha": 5,
        "identificador": "historico:B5:saldo_remanescente",
        "origem_dado": "Itens Remanescentes",
        "tipo_parcela": "Saldo remanescente",
        "tipo_financeiro": "Saldo Remanescente",
        "fonte_parcela": "Itens remanescentes",
        "valor": valor,
        "justificativa_vta": "Saldo remanescente agregado; sem recomposicao item a item na Fase 3.5.",
    }]


def _ler_parcelas_sombra_aditivos(wb) -> list[dict[str, Any]]:
    """Materializa aditivos como parcelas auditaveis.

    Novo modelo oficial (sem aba historico): le diretamente a aba aditivos
    — linhas com ITEM preenchido e K=CONSIDERADO NO CALCULO FINANCEIRO=Sim
    geram parcela com o valor atualizado da alteracao (J, cache do Excel)
    e ciclo (C, cache; sem cache, derivado da data do aditivo em B).
    Layout legado: consolidado homologado de historico!D23:D27 (ou D28).
    """
    if "historico" not in wb.sheetnames:
        if "aditivos" not in wb.sheetnames:
            return []
        ws_adt = wb["aditivos"]
        parcelas_novo: list[dict[str, Any]] = []
        for r in range(2, min(ws_adt.max_row, 200) + 1):
            item = ws_adt.cell(r, 1).value
            if item in (None, ""):
                continue
            considerado = str(ws_adt.cell(r, 11).value or "").strip().lower()
            if considerado != "sim":
                continue
            valor = _to_float_sombra(ws_adt.cell(r, 10).value)  # J: valor atualizado
            if not valor:
                # Formula sem cache do Excel — sem valor confiavel; o upload
                # sem cache ja e bloqueado pela politica (posicao_contratual).
                continue
            ciclo = str(ws_adt.cell(r, 3).value or "").strip()
            if not ciclo:
                ciclo = _ciclo_por_competencia(wb, ws_adt.cell(r, 2).value)
            parcelas_novo.append({
                "linha": r,
                "identificador": f"aditivos:{ciclo or 'sem_ciclo'}:{r}",
                "item": item,
                "data_aditivo": ws_adt.cell(r, 2).value,
                "tipo_alteracao": ws_adt.cell(r, 4).value,
                "quantidade": ws_adt.cell(r, 5).value,
                "vu_original": ws_adt.cell(r, 6).value,
                "valor_original": ws_adt.cell(r, 7).value,
                "origem_dado": "Aditivo",
                "tipo_parcela": "Aditivo",
                "tipo_financeiro": "Aditivo Computavel",
                "fonte_parcela": "Aditivo",
                "ciclo": ciclo.upper() if ciclo else "",
                "valor": valor,
                "justificativa_vta": (
                    "Aditivo computavel lido da aba aditivos (novo modelo, "
                    "K=Sim; valor atualizado J)."
                ),
            })
        return parcelas_novo

    ws = wb["historico"]
    parcelas: list[dict[str, Any]] = []
    for r in range(23, 28):
        valor = _to_float_sombra(ws.cell(r, 4).value)
        if not valor:
            continue
        ciclo = str(ws.cell(r, 1).value or f"linha {r}").strip()
        parcelas.append({
            "linha": r,
            "identificador": f"historico:D{r}:aditivo:{ciclo}",
            "origem_dado": "Aditivo",
            "tipo_parcela": "Aditivo",
            "tipo_financeiro": "Aditivo Computavel",
            "fonte_parcela": "Aditivo",
            "valor": valor,
            "justificativa_vta": "Aditivo computavel lido do bloco homologado D23:D27.",
        })

    if not parcelas:
        valor_total = _to_float_sombra(ws["D28"].value)
        if valor_total:
            parcelas.append({
                "linha": 28,
                "identificador": "historico:D28:aditivos_total",
                "origem_dado": "Aditivo",
                "tipo_parcela": "Aditivo",
                "tipo_financeiro": "Aditivo Computavel",
                "fonte_parcela": "Aditivo",
                "valor": valor_total,
                "justificativa_vta": "Total consolidado de aditivos lido de D28.",
            })
    return parcelas


def _apenas_data(valor: Any) -> Any:
    """Reduz datetime a date para comparacao por dia (nunca por instante)."""
    return valor.date() if hasattr(valor, "date") and hasattr(valor, "hour") else valor


def _data_corte_valida(data_corte: Any) -> Any:
    """Data de corte utilizavel (CONTROLE!B3) ou ``None`` quando ilegivel."""
    corte = _normalizar_data(data_corte)
    return _apenas_data(corte) if hasattr(corte, "year") else None


def _totais_canonicos_pc(
    itens: list[dict[str, Any]],
    data_corte: Any = None,
) -> dict[str, Any]:
    """Medidas canonicas dos PCs, separadas por significado.

    Distingue explicitamente o que a homologacao provou serem coisas diferentes:

      * informado          — inventario integral do arquivo (todo PC valido);
      * ate_o_corte        — subconjunto alcancado pela data de corte unica;
      * enquadrado_ciclos  — PCs cuja DATA_PC cai dentro de um ciclo (C0..C4);
      * com_efeito         — PCs que alcancaram o inicio do efeito financeiro;
      * sem_efeito_ciclo   — dentro de ciclo, antes do inicio do efeito;
      * intervalo_precluso — entre dois ciclos conhecidos, execucao sem reajuste;
      * indeterminado      — sem faixa temporal justificavel (exige revisao).

    Todos os valores "considerado" usam VALOR_HISTORICO_CONSIDERADO; nenhuma
    medida aplica fator a PC sem efeito financeiro.
    """
    def _bloco() -> dict[str, float]:
        return {
            "quantidade": 0,
            "valor_pc": 0.0,
            "valor_considerado": 0.0,
            "retroativo": 0.0,
            "valor_atualizado_em_analise": 0.0,
            "delta_potencial": 0.0,
        }

    chaves = (
        "informado", "ate_o_corte", "enquadrado_ciclos", "com_efeito",
        "sem_efeito_ciclo", "intervalo_precluso", "indeterminado",
        "posterior_ao_corte",
    )
    blocos = {chave: _bloco() for chave in chaves}
    por_ciclo: dict[str, dict[str, float]] = {}

    def _somar(
        bloco: dict[str, float],
        valor: float,
        considerado: float,
        retro: float,
        em_analise: float,
        potencial: float,
    ) -> None:
        bloco["quantidade"] += 1
        bloco["valor_pc"] += valor
        bloco["valor_considerado"] += considerado
        bloco["retroativo"] += retro
        bloco["valor_atualizado_em_analise"] += em_analise
        bloco["delta_potencial"] += potencial

    for item in itens:
        if str(item.get("entra_no_calculo") or "Sim").strip().lower() not in (
            "sim", "s", "true", "1"
        ):
            continue
        if (item.get("campos_vta") or {}).get(
            "status_consolidacao"
        ) == "DESCARTADO_DUPLICIDADE":
            continue
        valor = float(item.get("valor_pc") or 0.0)
        considerado = float(
            item.get("valor_historico_considerado")
            if item.get("valor_historico_considerado") is not None else valor
        )
        retro = float(item.get("retroativo_reconhecido_a_pagar") or 0.0)
        em_analise = float(item.get("valor_atualizado_em_analise") or 0.0)
        potencial = float(item.get("delta_potencial") or 0.0)
        enq = str(item.get("enquadramento") or ENQ_CICLO)
        no_corte = bool(item.get("dentro_do_corte", True))

        _somar(
            blocos["informado"], valor, considerado, retro, em_analise, potencial
        )
        if not no_corte:
            _somar(
                blocos["posterior_ao_corte"],
                valor,
                considerado,
                retro,
                em_analise,
                potencial,
            )
            continue
        _somar(
            blocos["ate_o_corte"], valor, considerado, retro, em_analise, potencial
        )

        if enq == ENQ_INTERVALO_PRECLUSO:
            _somar(
                blocos["intervalo_precluso"],
                valor,
                considerado,
                retro,
                em_analise,
                potencial,
            )
            continue
        if enq == ENQ_INDETERMINADO or not item.get("ciclo"):
            _somar(
                blocos["indeterminado"],
                valor,
                considerado,
                retro,
                em_analise,
                potencial,
            )
            continue

        _somar(
            blocos["enquadrado_ciclos"],
            valor,
            considerado,
            retro,
            em_analise,
            potencial,
        )
        destino = (
            "com_efeito" if str(item.get("efeito_financeiro_pc") or "").strip()
            == "Sim" else "sem_efeito_ciclo"
        )
        _somar(blocos[destino], valor, considerado, retro, em_analise, potencial)
        ciclo = str(item["ciclo"]).strip().upper()
        _somar(
            por_ciclo.setdefault(ciclo, _bloco()),
            valor,
            considerado,
            retro,
            em_analise,
            potencial,
        )

    for bloco in list(blocos.values()) + list(por_ciclo.values()):
        for campo in (
            "valor_pc",
            "valor_considerado",
            "retroativo",
            "valor_atualizado_em_analise",
            "delta_potencial",
        ):
            bloco[campo] = round(bloco[campo], 2)

    return {
        **blocos,
        "por_ciclo": por_ciclo,
        "data_corte": _data_corte_valida(data_corte),
        "corte_aplicado": _data_corte_valida(data_corte) is not None,
    }


def _pc_dentro_do_corte(data_pc: Any, data_corte: Any) -> bool:
    """PC alcancado pela data de corte unica do contrato (CONTROLE!B3).

    Sem data de corte informada, todo PC e alcancado (comportamento anterior).
    Data de PC ilegivel tambem nao e excluida: exclusao exige prova de que a
    data e posterior ao corte, nunca ausencia de informacao.
    """
    corte = _data_corte_valida(data_corte)
    if corte is None:
        return True
    data = _normalizar_data(data_pc)
    if not hasattr(data, "year"):
        return True
    return _apenas_data(data) <= corte


def _ler_itens_pc_v10(
    wb,
    ciclo_vigente: str | None = None,
    data_corte: Any = None,
) -> dict[str, Any]:
    """Le itens_PC v10: registro financeiro SAP por cabecalho.
    Aceita PC de item especifico e PC global/agregado.
    Sem TIPO_PC, QTD ou VU — usa VALOR_PC como base financeira.

    Etapa 27B: a participacao do PC no VTA (computa_vta) e ESTRUTURAL —
    valor valido + ciclo valido + ciclo anterior a linha de corte
    (ciclo_vigente, quando informado). O pagamento (coluna G) governa
    somente a classificacao do retroativo: Sim -> reconhecido; Nao ou
    vazio -> potencial (vazio = pagamento nao confirmado).
    """
    resultado: dict[str, Any] = {
        "itens": [], "totais": {}, "alertas": [],
        "ok": False, "pc_unitarios": [], "pc_globais": [],
        "resumo_vta": {
            "campos_presentes": False,
            "campos_ausentes": [],
            "linhas_total": 0,
            "computaveis": 0,
            "nao_computaveis": 0,
            "em_analise": 0,
            "inconsistentes": 0,
            "descartadas_duplicidade": 0,
        },
    }

    ciclo_vigente_norm = str(ciclo_vigente or "").strip().upper()
    n_vig_corte = (
        int(ciclo_vigente_norm[1:])
        if ciclo_vigente_norm.startswith("C") and ciclo_vigente_norm[1:].isdigit()
        else None
    )
    pagamentos_nao_confirmados = 0

    ws, layout_fiscal_v2 = _selecionar_aba_pc(wb)
    layout_fiscal_definitivo = (
        layout_fiscal_v2
        and bool(_aba_entrada_existente(wb, ABA_FISCAL_FINANCEIRO_V2))
        and bool(_aba_entrada_existente(wb, ABA_FISCAL_REMANESCENTES_V2))
    )
    if ws is None:
        resultado["alertas"].append(
            f"Aba '{ABA_ITENS_PC}' ou '{ABA_FISCAL_PCS_V2}' ausente — necessaria para modo PC."
        )
        return resultado

    mapa = _mapear_colunas_por_cabecalho(ws)

    # Alerta se layout parece itens_Consumidos (nao deve ter qtd_cons_c0)
    if mapa.get("qtd_cons_c0") or mapa.get("cons_qtd_total"):
        resultado["alertas"].append(
            "ATENCAO: itens_PC parece ter layout de itens_Consumidos. Verificar template."
        )

    col_num    = mapa.get("numero_pc")
    # Modelo oficial (NUMERO_PC presente): NUMERO_PC e o identificador
    # documental; sem fallback posicional — nenhuma outra coluna (DATA_PC,
    # primeira coluna etc.) vira identificador. O fallback da coluna 1
    # permanece apenas para estruturas legadas sem NUMERO_PC.
    col_item   = mapa.get("item_ou_grupo") or mapa.get("item") or (
        None if col_num else 1
    )
    col_data   = mapa.get("data_pc")
    col_ciclo  = mapa.get("ciclo_pc") or mapa.get("ciclo")
    col_valor  = mapa.get("valor_pc")
    col_fator  = mapa.get("fator_acumulado") or mapa.get("fator_aplicavel")
    col_vatual = mapa.get("valor_atualizado")
    # Claus New (integracao real): marcador de pagamento do PC (coluna G do
    # template v10.4.1+). Sem ele o Motor Temporal nao responde Q5/Q8.
    col_pago   = mapa.get("pc_pago_a_contratada")
    col_entra  = mapa.get("entra_no_calculo")
    col_obs    = mapa.get("observacao_itens_abrangidos") or mapa.get("observacao")
    col_check  = mapa.get("check")
    col_efeito_pc = mapa.get("efeito_financeiro_pc")
    col_retro = _col(
        mapa, "RETROATIVO_RECONHECIDO_A_PAGAR", "RETROATIVO RECONHECIDO"
    )
    col_analise = mapa.get("valor_atualizado_em_analise")
    col_delta = _col(mapa, "DELTA_POTENCIAL", "RETROATIVO POTENCIAL")
    col_pagto_smart = _col(
        mapa,
        "STATUS_PAGAMENTO_PC",
        "PC_PAGO_INTEGRALMENTE",
        "PAGAMENTO_COMPROVADO",
        "PAGAMENTO_PC",
    )
    col_data_pagto_smart = _col(
        mapa, "DATA_PAGAMENTO_PC", "DATA_PAGAMENTO", "DATA_PAGAMENTO_CONTRATADA"
    )
    col_valor_pago_smart = _col(
        mapa, "VALOR_EFETIVAMENTE_PAGO", "VALOR_PAGO_PC", "VALOR_PAGO", "VALOR_PAGO_A_CONTRATADA"
    )
    col_refletido_smart = _col(mapa, "JA_REFLETIDO_EM", "JA_REFLETIDO")

    vta_derivados = {
        "computa_vta", "tipo_parcela", "origem_dado", "tipo_financeiro",
        "fonte_parcela", "status_consolidacao", "justificativa_vta",
    }
    campos_derivados_presentes = any(
        c for c in (col_ciclo, col_fator, col_vatual, col_pago, col_entra, col_check)
    ) or any(mapa.get(campo) for campo in vta_derivados)
    layout_inteligente = bool(
        col_num and col_data and col_valor and not campos_derivados_presentes
    )

    cols_vta = (
        {campo: None for campo in _VTA_CAMPOS}
        if layout_inteligente else {campo: mapa.get(campo) for campo in _VTA_CAMPOS}
    )
    campos_ausentes = [campo.upper() for campo, col in cols_vta.items() if not col]
    resultado["resumo_vta"]["campos_ausentes"] = campos_ausentes
    resultado["resumo_vta"]["campos_presentes"] = not campos_ausentes
    if cols_vta and any(cols_vta.values()) and campos_ausentes:
        resultado["alertas"].append(
            "Campos estruturais PC/VTA parcialmente ausentes em itens_PC: "
            + ", ".join(campos_ausentes)
            + ". Defaults seguros aplicados."
        )

    campos_detectados = [
        str(cell.value).strip() for cell in ws[1] if str(cell.value or "").strip()
    ]
    campos_eliminados = [
        c for c in _PC_LAYOUT_ATUAL_MANUAIS_OU_HERDADOS
        if c not in _PC_INTELIGENTE_OBRIGATORIOS
    ]
    resultado["masterfile_inteligente"] = {
        "layout": "fiscal_definitivo" if layout_fiscal_definitivo else "fiscal_v2" if layout_fiscal_v2 else (
            "inteligente" if layout_inteligente else "atual_ou_legado"
        ),
        "layout_inteligente": layout_inteligente,
        "layout_fiscal_v2": layout_fiscal_v2,
        "layout_fiscal_definitivo": layout_fiscal_definitivo,
        "aba_origem": ws.title,
        "campos_detectados": campos_detectados,
        "campos_obrigatorios": (
            list(_PC_INTELIGENTE_OBRIGATORIOS) if layout_inteligente else []
        ),
        "quantidade_campos_obrigatorios": (
            len(_PC_INTELIGENTE_OBRIGATORIOS) if layout_inteligente else 0
        ),
        "campos_opcionais": (
            list(_PC_INTELIGENTE_OPCIONAIS) if layout_inteligente else []
        ),
        "campos_inferidos": (
            list(_PC_INTELIGENTE_INFERIDOS) if layout_inteligente else []
        ),
        "campos_deixaram_de_ser_preenchidos": (
            campos_eliminados if layout_inteligente else []
        ),
        "quantidade_campos_layout_atual": len(_PC_LAYOUT_ATUAL_MANUAIS_OU_HERDADOS),
        "quantidade_campos_layout_inteligente_minimo": len(_PC_INTELIGENTE_OBRIGATORIOS),
        "reducao_estimada_campos": len(campos_eliminados) if layout_inteligente else 0,
        "reducao_percentual_estimada": (
            round(len(campos_eliminados) / len(_PC_LAYOUT_ATUAL_MANUAIS_OU_HERDADOS), 4)
            if layout_inteligente else 0.0
        ),
        "separacao_responsabilidades": {
            "gcc": [
                "Calculadora gera MasterFile customizado",
                "GCC faz upload",
                "Mesa da GCC valida metodologia, VTA, retroativo, pendencias e dossie",
            ],
            "entrada_xls": list(_FISCAL_V2_CAMPOS_FISCAL),
            "claus_new": list(_PC_INTELIGENTE_INFERIDOS),
        },
        "nao_exigir_no_xls": list(_FISCAL_V2_NAO_PEDIR),
        "compatibilidade": (
            "Leitor aceita layout de entrada, inteligente, atual e legado deslocado."
        ),
    }

    def _tofl(v, default=0.0):
        try:
            return float(v) if v not in (None, "") else default
        except (TypeError, ValueError):
            return default

    parametros_pc = _ler_parametros_v10(wb)
    por_ciclo_pc = parametros_pc.get("por_ciclo") or {}

    # Guarda fail-closed do calendario: cada ciclo com exatamente 12
    # competencias consecutivas e ciclos contiguos. Uma lacuna aqui significa
    # que DATA_INICIO/DATA_FIM foram materializados a partir da janela do
    # indice ou do INICIO_EFEITO_FINANCEIRO — nenhuma competencia pode ficar
    # fora dos ciclos entre ciclos consecutivos.
    from _motor_posicao_contratual import divergencias_calendario_canonico
    for problema in divergencias_calendario_canonico([
        {"ciclo": nome, **(reg or {})} for nome, reg in por_ciclo_pc.items()
    ]):
        resultado["alertas"].append(f"Calendario de ciclos invalido: {problema}")

    def _inferir_ciclo_fator_valor(data_pc: Any, valor_pc: Any) -> tuple[Any, Any, Any]:
        data_norm = _normalizar_data(data_pc)
        valor_num = _tofl(valor_pc, default=None)
        if data_norm is None:
            return None, None, None
        ciclo_inferido = enquadrar_data_pc(data_norm, por_ciclo_pc)
        if not ciclo_inferido:
            return None, None, None
        fator = _tofl((por_ciclo_pc.get(ciclo_inferido) or {}).get("fator_acumulado"),
                      default=None)
        if fator is None and ciclo_inferido == "C0":
            fator = 1.0
        valor_atualizado = (
            round(valor_num * fator, 2)
            if valor_num is not None and fator is not None else None
        )
        return ciclo_inferido, fator, valor_atualizado

    def _fator_pc_na_apuracao(ciclo_alvo: Any) -> float | None:
        try:
            limite = int(str(ciclo_alvo or "").strip().upper().removeprefix("C"))
        except (TypeError, ValueError):
            return None
        fator = 1.0
        for numero in range(1, limite + 1):
            reg = por_ciclo_pc.get(f"C{numero}") or {}
            if not _sim_local(reg.get("computar_nesta_apuracao")):
                continue
            percentual = _tofl(reg.get("percentual_reajuste"), default=None)
            if percentual is None:
                return None
            if abs(percentual) > 1.0:
                percentual /= 100.0
            fator *= 1.0 + percentual
        return round(fator, 12)

    def _sim_local(valor: Any) -> bool:
        return _norm(valor) in {"sim", "s", "true", "1", "yes"}

    def _nao_local(valor: Any) -> bool:
        return _norm(valor) in {"nao", "n", "false", "0", "no"}

    def _nao_sei_local(valor: Any) -> bool:
        return _norm(valor) in {"nao sei", "nao_sei", "desconhecido", "ignorado"}

    def _classificar_status_pagamento(valor: Any) -> str:
        status = _norm(valor)
        if _sim_local(valor) or status in {"pago", "pago integral", "integralmente pago"}:
            return "PAGO_INTEGRAL"
        if "pago" in status and "glosa" in status:
            return "PAGO_COM_GLOSA"
        if _nao_local(valor) or status in {"nao pago", "sem pagamento"}:
            return "NAO_PAGO"
        if "pendent" in status or "analise" in status:
            return "PAGAMENTO_PENDENTE"
        if _nao_sei_local(valor) or not status:
            return "NAO_CONFIRMADO"
        return "INVALIDO"

    def _inferir_pagamento_e_vta(
        r: int, ident: str, valor_pc: Any
    ) -> tuple[str, dict[str, Any], str]:
        declarado = ws.cell(r, col_pagto_smart).value if col_pagto_smart else None
        data_pagto = ws.cell(r, col_data_pagto_smart).value if col_data_pagto_smart else None
        valor_pago = ws.cell(r, col_valor_pago_smart).value if col_valor_pago_smart else None
        refletido = (
            ws.cell(r, col_refletido_smart).value
            if col_refletido_smart and col_refletido_smart not in (col_num, col_data, col_valor)
            else None
        )

        status_pagamento = _classificar_status_pagamento(declarado)
        valor_pago_num = _tofl(valor_pago, default=None)
        valor_pc_num = _tofl(valor_pc, default=None)
        if status_pagamento == "INVALIDO":
            resultado["alertas"].append(
                f"PC '{ident}' linha {r}: STATUS_PAGAMENTO_PC invalido; "
                "tratado como pagamento nao confirmado."
            )
            status_pagamento = "NAO_CONFIRMADO"

        if status_pagamento == "PAGO_INTEGRAL" and valor_pago_num is None:
            # Compatibilidade: a declaracao expressa de pagamento integral
            # permite usar VALOR_PC. Para glosa, o valor final e obrigatorio.
            valor_pago_num = valor_pc_num
        if status_pagamento == "PAGO_COM_GLOSA" and valor_pago_num is None:
            resultado["alertas"].append(
                f"PC '{ident}' linha {r}: Pago com glosa exige "
                "VALOR_EFETIVAMENTE_PAGO; PC excluido do retroativo."
            )
        if status_pagamento not in {"PAGO_INTEGRAL", "PAGO_COM_GLOSA"} and (
            data_pagto not in (None, "") or valor_pago_num is not None
        ):
            resultado["alertas"].append(
                f"PC '{ident}' linha {r}: data/valor de pagamento sem status pago; "
                "PC mantido fora do retroativo."
            )

        tem_evidencia_pagamento = (
            status_pagamento in {"PAGO_INTEGRAL", "PAGO_COM_GLOSA"}
            and valor_pago_num is not None
            and valor_pago_num > 0.0
        )

        ja_refletido = _normalizar_campo_vta(
            "ja_refletido_em", refletido, ident, r, resultado["alertas"]
        )
        if tem_evidencia_pagamento:
            computa = "Sim"
            status = "COMPUTADO"
            justificativa = (
                "Inferido pelo MasterFile Inteligente: ha evidencia de pagamento "
                "do PC; parcela tratada como execucao, nunca como aditivo."
            )
            if ja_refletido != "Nao":
                computa = "Nao"
                status = "DESCARTADO_DUPLICIDADE"
                justificativa = (
                    "Inferido pelo MasterFile Inteligente: PC pago, mas ja "
                    "refletido em outra fonte; descartado para evitar dupla contagem."
                )
            campos_vta = {
                "computa_vta": computa,
                "tipo_parcela": "PC pago",
                "origem_dado": "Pedido de Compra",
                "tipo_financeiro": "Execucao Atualizada",
                "fonte_parcela": "PC",
                "ja_refletido_em": ja_refletido,
                "status_consolidacao": status,
                "justificativa_vta": justificativa,
                "valor_pago": valor_pago_num,
                "status_pagamento_pc": status_pagamento,
                "data_pagamento_pc": data_pagto,
                "elegivel_retroativo_pc": computa == "Sim",
            }
            return "Sim", campos_vta, "pagamento_definitivo_comprovado"

        campos_vta = {
            "computa_vta": "Nao",
            "tipo_parcela": "PC em analise",
            "origem_dado": "Pedido de Compra",
            "tipo_financeiro": "Impacto Potencial",
            "fonte_parcela": "PC",
            "ja_refletido_em": "Nao",
            "status_consolidacao": "EM_ANALISE",
            "justificativa_vta": (
                "Sem pagamento definitivo elegivel; PC historico, pendente ou "
                "sem valor final nao compoe retroativo nem VTA."
            ),
            "valor_pago": valor_pago_num,
            "status_pagamento_pc": status_pagamento,
            "data_pagamento_pc": data_pagto,
            "elegivel_retroativo_pc": False,
        }
        flag_pagamento = "Nao sei" if status_pagamento == "NAO_CONFIRMADO" else "Nao"
        return flag_pagamento, campos_vta, "pagamento_nao_elegivel"

    def _parece_data(v: Any) -> bool:
        return isinstance(v, (date, datetime))

    def _parece_ciclo(v: Any) -> bool:
        return str(v or "").strip().upper() in {f"C{i}" for i in range(5)}

    def _linha_pc_legado_deslocada(r: int) -> bool:
        """Detecta preenchimento A:H legado sob cabecalho novo A:G.

        Em alguns arquivos reais, os valores foram colados como:
        ITEM, NUMERO_PC, DATA_PC, CICLO_PC, VALOR_PC, FATOR, VALOR_ATUALIZADO,
        PC_PAGO_A_CONTRATADA. O cabecalho atual comeca em NUMERO_PC; sem esta
        correcao, a data vira ciclo informado e o valor chega zerado ao motor.
        """
        if not (
            col_num == 1 and col_data == 2 and col_ciclo == 3
            and col_valor == 4 and col_fator == 5 and col_vatual == 6
        ):
            return False
        bruto_b = ws.cell(r, 2).value
        bruto_c = ws.cell(r, 3).value
        bruto_d = ws.cell(r, 4).value
        bruto_e = ws.cell(r, 5).value
        valor_na_coluna_atual = _tofl(bruto_d, default=None)
        valor_na_coluna_legada = _tofl(bruto_e, default=None)
        return (
            bruto_b not in (None, "")
            and valor_na_coluna_legada is not None
            and (
                (_parece_data(bruto_c) and _parece_ciclo(bruto_d))
                or valor_na_coluna_atual is None
            )
        )

    def _campos_vta_legado_deslocados(r: int) -> bool:
        col_computa = cols_vta.get("computa_vta")
        col_tipo = cols_vta.get("tipo_parcela")
        if not (col_computa and col_tipo):
            return False
        valor_atual = ws.cell(r, col_computa).value
        valor_deslocado = ws.cell(r, col_computa + 1).value
        tipo_deslocado = ws.cell(r, col_tipo + 1).value
        return (
            valor_atual in (None, "", "Nao")
            and _norm(valor_deslocado) in {"sim", "nao"}
        )

    # Etapa 26H: o teto operacional de itens_PC segue a capacidade canonica
    # (_capacidade_pcs). O teto legado de 100 linhas truncava silenciosamente
    # os PCs a partir da linha 101 (C0 parcial e C1 inteiro fora da memoria
    # por ciclo e do Sumario Executivo).
    limite_operacional = (
        min(ws.max_row, ULTIMA_LINHA_PCS) if ws.title == "itens_PC" else ws.max_row
    )
    max_col_itens_pc = ws.max_column
    for r in range(2, limite_operacional + 1):
        if _linha_exemplo_fiscal(ws, r, max_col_itens_pc):
            continue
        item_ou_grupo = ws.cell(r, col_item).value if col_item else None
        num_pc        = ws.cell(r, col_num).value  if col_num  else None

        data_pc = ws.cell(r, col_data).value  if col_data  else None
        valor   = ws.cell(r, col_valor).value if col_valor else None
        fator   = ws.cell(r, col_fator).value if col_fator else None
        vatual  = ws.cell(r, col_vatual).value if col_vatual else None
        ciclo   = ws.cell(r, col_ciclo).value  if col_ciclo  else None
        entra   = ws.cell(r, col_entra).value  if col_entra  else "Sim"
        obs     = ws.cell(r, col_obs).value    if col_obs    else None
        pago    = ws.cell(r, col_pago).value   if col_pago   else None
        efeito_pc = ws.cell(r, col_efeito_pc).value if col_efeito_pc else None
        layout_corrigido = False

        if _linha_pc_legado_deslocada(r):
            item_ou_grupo = ws.cell(r, 1).value
            num_pc = ws.cell(r, 2).value
            data_pc = ws.cell(r, 3).value
            ciclo = ws.cell(r, 4).value
            valor = ws.cell(r, 5).value
            fator = ws.cell(r, 6).value
            vatual = ws.cell(r, 7).value
            pago = ws.cell(r, 8).value
            layout_corrigido = True

        if not item_ou_grupo and not num_pc:
            if col_num and (data_pc not in (None, "") or valor not in (None, "")):
                resultado["alertas"].append(
                    f"PC linha {r}: NUMERO_PC vazio — linha lida, mas sem "
                    "controle de duplicidade para este registro."
                )
            else:
                continue

        ident = str(num_pc or item_ou_grupo or f"linha {r}").strip()
        inferencias_linha: list[str] = []
        pagamento_inferido = ""

        if layout_inteligente:
            ciclo_inf, fator_inf, vatual_inf = _inferir_ciclo_fator_valor(data_pc, valor)
            ciclo = ciclo_inf
            fator = fator_inf
            vatual = vatual_inf
            pago, campos_vta, pagamento_inferido = _inferir_pagamento_e_vta(
                r, ident, valor
            )
            inferencias_linha.extend([
                "CICLO_PC pela DATA_PC na linha temporal",
                "FATOR_ACUMULADO pela aba parametros",
                "VALOR_ATUALIZADO por VALOR_PC x fator inferido",
                "PC_PAGO_A_CONTRATADA por evidencia de pagamento",
                "campos PC/VTA por estado de pagamento e antiduplicidade",
            ])
            if ciclo is None and data_pc not in (None, ""):
                resultado["alertas"].append(
                    f"PC '{ident}': DATA_PC fora dos ciclos ou ciclo nao calculado."
                )
            if fator is None and ciclo not in (None, ""):
                resultado["alertas"].append(
                    f"PC '{ident}': FATOR_ACUMULADO nao inferido na aba parametros."
                )

        # Compatibilidade segura quando o cache das formulas ainda nao existe:
        # o ciclo continua vindo da mesma linha temporal, e o efeito usa somente
        # a fonte canonica reconciliada. Nunca interpretar L vazia como Sim.
        ciclo_norm = str(ciclo or "").strip().upper()
        if ciclo_norm not in {f"C{i}" for i in range(5)}:
            ciclo_inferido = enquadrar_data_pc(_normalizar_data(data_pc), por_ciclo_pc)
            if ciclo_inferido:
                ciclo = ciclo_inferido
                ciclo_norm = ciclo_inferido
                inferencias_linha.append("CICLO_PC recomposto pela linha temporal")

        # Enquadramento canonico da DATA_PC: ciclo, INTERVALO PRECLUSO entre
        # dois ciclos conhecidos, ou indeterminado. O intervalo precluso e
        # execucao VALIDA sem ciclo de reajuste — nunca erro, nunca atribuido
        # ao ciclo anterior nem ao seguinte.
        enquadramento = classificar_enquadramento_pc(data_pc, por_ciclo_pc)
        if enquadramento.e_precluso:
            ciclo = None
            ciclo_norm = ""
            inferencias_linha.append(
                "DATA_PC entre ciclos conhecidos: " + enquadramento.rotulo
            )

        from _efeitos_financeiros_pc import efeito_financeiro_pc
        # Sem ciclo de reajuste nao ha inicio de efeito financeiro a alcancar:
        # o efeito e "Nao" por definicao, sem indeterminacao.
        efeito_inferido = (
            "Nao" if enquadramento.e_precluso
            else efeito_financeiro_pc(data_pc, ciclo, por_ciclo_pc.get(ciclo_norm))
        )
        if efeito_pc in (None, ""):
            efeito_pc = efeito_inferido
            if efeito_pc is not None:
                inferencias_linha.append(
                    "EFEITO_FINANCEIRO_PC pela DATA_PC e inicio canonico do ciclo"
                )
        elif str(efeito_pc).strip() not in {"Sim", "Nao"}:
            resultado["alertas"].append(
                f"PC '{ident}' linha {r}: EFEITO_FINANCEIRO_PC invalido."
            )
            efeito_pc = None
        elif efeito_inferido is None:
            resultado["alertas"].append(
                f"PC '{ident}' linha {r}: marcador de efeito descartado; "
                "a fonte canonica esta ausente ou inconsistente."
            )
            efeito_pc = None
        elif str(efeito_pc).strip() != efeito_inferido:
            resultado["alertas"].append(
                f"PC '{ident}' linha {r}: EFEITO_FINANCEIRO_PC diverge da "
                f"DATA_PC e da fonte canonica; prevalece {efeito_inferido}."
            )
            efeito_pc = efeito_inferido
        if efeito_pc is None and str(ciclo or "").strip().upper() not in {"", "C0"}:
            resultado["alertas"].append(
                f"PC '{ident}' linha {r}: efeito financeiro indeterminado; "
                "INICIO_EFEITO_FINANCEIRO ausente ou inconsistente."
            )
        # Etapa 26C: VALOR HISTORICO ATUALIZADO e desacoplado do EFEITO
        # FINANCEIRO. O fator historico e o da memoria da apuracao
        # (parametros!A11:E15 no XLS; _fator_pc_na_apuracao aqui), aplicavel ao
        # CICLO_PC independentemente de efeito Sim/Nao. O efeito continua
        # governando somente retroativo/status/cor.
        valor_num = _tofl(valor, default=None)
        vatual_cache = _tofl(vatual, default=None)
        # Etapa 26C (preservada): VALOR_ATUALIZADO e o valor-base x fator
        # historico integral do CICLO_PC, desacoplado do efeito financeiro. Quem
        # governa resultado, retroativo, VTA e documentos e o VALOR HISTORICO
        # CONSIDERADO (base + retroativo reconhecido), calculado adiante — nele,
        # o PC anterior ao inicio do efeito usa fator efetivo 1 e permanece no
        # valor original.
        fator_historico = _fator_pc_na_apuracao(ciclo_norm)
        fator = fator_historico
        vatual = (
            round(valor_num * fator_historico, 2)
            if valor_num is not None and fator_historico is not None else None
        )
        if (
            vatual is not None
            and vatual_cache is not None
            and abs(vatual - vatual_cache) > 0.01
        ):
            resultado["alertas"].append(
                f"PC '{ident}' linha {r}: VALOR_ATUALIZADO do XLS "
                f"({vatual_cache:.2f}) diverge do valor historico recomputado "
                f"({vatual:.2f}); adotado o recomputado (compatibilidade com "
                "XLS anterior ao desacoplamento historico x efeito)."
            )

        if layout_corrigido:
            resultado["alertas"].append(
                f"PC '{ident}' linha {r}: preenchimento parece estar no layout "
                "anterior de itens_PC; leitura corrigida sem alterar o arquivo."
            )
        is_global = _norm(str(item_ou_grupo or "")) in ("global",) or (
            obs and "," in str(obs)
        )

        if data_pc in (None, ""):
            resultado["alertas"].append(f"PC '{ident}': DATA_PC nao informada.")
        if valor in (None, ""):
            resultado["alertas"].append(f"PC '{ident}': VALOR_PC nao informado.")
        if (
            not layout_inteligente
            and (fator in (None, "") or _tofl(fator, default=-1) == -1)
        ):
            resultado["alertas"].append(
                f"PC '{ident}': FATOR_ACUMULADO ausente ou invalido."
            )
        if (
            not layout_inteligente
            and not enquadramento.e_precluso
            and ciclo in (None, "", "Fora dos ciclos")
        ):
            resultado["alertas"].append(
                f"PC '{ident}': DATA_PC fora dos ciclos ou ciclo nao calculado."
            )

        desloc_vta = 1 if _campos_vta_legado_deslocados(r) else 0
        if desloc_vta:
            resultado["alertas"].append(
                f"PC '{ident}' linha {r}: campos PC/VTA parecem estar em W:AD; "
                "leitura corrigida sem alterar o arquivo."
            )

        if not layout_inteligente:
            campos_vta = {
                campo: _normalizar_campo_vta(
                    campo,
                    (
                        ws.cell(r, cols_vta[campo] + desloc_vta).value
                        if cols_vta.get(campo) else None
                    ),
                    ident,
                    r,
                    resultado["alertas"],
                )
                for campo in _VTA_CAMPOS
            }
            # Modelo oficial: PC_PAGO_A_CONTRATADA (coluna G) é a decisão
            # fiscal de pagamento. Os campos V:AC são técnicos/derivados e
            # não podem anular silenciosamente esse dado visível.
            # Etapa 27B: G classifica o retroativo (Sim -> reconhecido;
            # Nao/vazio -> potencial); a participacao no VTA e estrutural
            # e definida no ajuste unificado abaixo, sem condicao em G.
            if col_pago and not col_pagto_smart:
                valor_pago_oficial = _tofl(valor, default=None)
                pago_norm_oficial = _norm(pago)
                pago_oficial = pago_norm_oficial in {"sim", "s", "yes", "true", "1"}
                nao_pago_oficial = pago_norm_oficial in {"nao", "n", "no", "false", "0"}
                if pago_norm_oficial and not pago_oficial and not nao_pago_oficial:
                    resultado["alertas"].append(
                        f"PC '{ident}' linha {r}: PC_PAGO_A_CONTRATADA invalido "
                        f"({pago!r}); esperado Sim, Nao ou vazio."
                    )
                elif not pago_norm_oficial and _tofl(valor, default=0.0) > 0.0:
                    pagamentos_nao_confirmados += 1
                elegivel_oficial = bool(
                    pago_oficial and valor_pago_oficial is not None
                    and valor_pago_oficial > 0.0
                )
                campos_vta.update({
                    "tipo_parcela": "PC pago" if elegivel_oficial else "PC em analise",
                    "origem_dado": "Pedido de Compra",
                    "fonte_parcela": "PC",
                    "ja_refletido_em": "Nao",
                    "valor_pago": valor_pago_oficial if elegivel_oficial else None,
                    "status_pagamento_pc": (
                        "PAGO_INTEGRAL" if elegivel_oficial
                        else "NAO_PAGO" if nao_pago_oficial
                        else "NAO_CONFIRMADO"
                    ),
                    "data_pagamento_pc": None,
                    "elegivel_retroativo_pc": elegivel_oficial,
                })

        # Etapa 27B — ajuste unificado (ambos os layouts): a participacao do
        # PC no VTA e ESTRUTURAL (valor valido + ciclo valido + ciclo anterior
        # a linha de corte). O pagamento NAO entra nessa decisao; segue
        # governando apenas retroativo (elegivel_retroativo_pc/valor_pago).
        # Duplicidade ja resolvida (DESCARTADO_DUPLICIDADE) e preservada.
        if campos_vta.get("status_consolidacao") != "DESCARTADO_DUPLICIDADE":
            valor_estrutural = _tofl(valor, default=None)
            n_pc_corte = (
                int(ciclo_norm[1:])
                if ciclo_norm in {f"C{i}" for i in range(5)} else None
            )
            estrutura_ok = bool(
                valor_estrutural is not None and valor_estrutural > 0.0
                and n_pc_corte is not None
            )
            dentro_corte = bool(
                n_pc_corte is not None
                and (n_vig_corte is None or n_pc_corte < n_vig_corte)
            )
            precluso_ok = bool(
                enquadramento.e_precluso
                and valor_estrutural is not None and valor_estrutural > 0.0
            )
            if precluso_ok:
                # Execucao valida entre dois ciclos conhecidos: compoe o total de
                # PCs e a execucao historica do VTA pelo valor original, sem
                # reajuste e sem retroativo. Nao integra nenhum ciclo.
                campos_vta.update({
                    "computa_vta": "Sim",
                    "tipo_financeiro": "Execucao Sem Reajuste",
                    "status_consolidacao": "COMPUTADO",
                    "justificativa_vta": (
                        f"{enquadramento.rotulo}: execucao valida entre ciclos "
                        "conhecidos; compoe o VTA pelo valor original, sem "
                        "reajuste e sem retroativo."
                    ),
                })
            elif estrutura_ok and dentro_corte:
                campos_vta.update({
                    "computa_vta": "Sim",
                    "tipo_financeiro": "Execucao Atualizada",
                    "status_consolidacao": "COMPUTADO",
                    "justificativa_vta": (
                        "PC estruturalmente valido anterior a linha de corte; "
                        "participa do VTA independentemente do pagamento."
                    ),
                })
            elif estrutura_ok:
                campos_vta.update({
                    "computa_vta": "Nao",
                    "tipo_financeiro": "Impacto Potencial",
                    "status_consolidacao": "NAO_COMPUTADO",
                    "justificativa_vta": (
                        "PC no ciclo vigente ou posterior a linha de corte; "
                        "nao entra na execucao historica (mesmo corte temporal)."
                    ),
                })
            else:
                campos_vta.update({
                    "computa_vta": "Nao",
                    "tipo_financeiro": "Impacto Potencial",
                    "status_consolidacao": "EM_ANALISE",
                    "justificativa_vta": (
                        "PC sem valor ou ciclo validos; nao participa do VTA "
                        "ate a complementacao dos dados."
                    ),
                })
        _validar_campos_vta(campos_vta, ident, r, resultado["alertas"])
        _atualizar_resumo_vta(resultado["resumo_vta"], campos_vta)

        pago_norm = _norm(pago)
        retro_lido = _tofl(ws.cell(r, col_retro).value, default=None) if col_retro else None
        analise_lida = _tofl(ws.cell(r, col_analise).value, default=None) if col_analise else None
        delta_lido = _tofl(ws.cell(r, col_delta).value, default=None) if col_delta else None
        incremento = (
            round(vatual - valor_num, 2)
            if vatual is not None and valor_num is not None else None
        )
        # Retroativo/delta continuam GOVERNADOS PELO EFEITO (espelha itens_PC!H/J):
        # efeito "Nao" -> retroativo/delta 0; indeterminado -> None. O valor
        # historico (vatual) NAO participa desse gate.
        incremento_retroativo = (
            incremento if efeito_pc == "Sim"
            else 0.0 if efeito_pc == "Nao" else None
        )
        # Etapa 27B: G="Sim" -> reconhecido; G="Nao" OU vazio -> potencial
        # (vazio = pagamento nao confirmado; nunca zera silenciosamente).
        pago_confirmado_fb = pago_norm in {"sim", "s", "true", "1", "yes"}
        if retro_lido is None and pago_confirmado_fb:
            retro_lido = incremento_retroativo
        if analise_lida is None and not pago_confirmado_fb and valor not in (None, ""):
            analise_lida = vatual
        if delta_lido is None and not pago_confirmado_fb and valor not in (None, ""):
            delta_lido = incremento_retroativo

        # Medida canonica do valor historico EFETIVAMENTE considerado. Nao
        # substitui VALOR_ATUALIZADO (que segue sendo o valor-base x fator
        # historico integral, desacoplado do efeito): e o valor que a apuracao
        # de fato reconhece — base paga mais o retroativo reconhecido a pagar.
        # PC sem efeito financeiro e PC do intervalo precluso ficam no valor
        # original; PC nao pago/em analise nao antecipa retroativo potencial.
        retro_reconhecido = _tofl(retro_lido, default=0.0)
        valor_hist_considerado = (
            round(valor_num + retro_reconhecido, 2)
            if valor_num is not None else None
        )
        check_lido = ws.cell(r, col_check).value if col_check else None
        if enquadramento.e_precluso:
            # O CHECK do XLS foi calculado antes de existir a classe do
            # intervalo; aqui ele vira INFO valida, nunca erro.
            check_lido = enquadramento.rotulo

        registro = {
            "linha":            r,
            "item_ou_grupo":    item_ou_grupo,
            "numero_pc":        num_pc,
            "data_pc":          data_pc,
            "ciclo":            ciclo,
            "enquadramento":    enquadramento.tipo,
            "enquadramento_rotulo": enquadramento.rotulo,
            "ciclo_anterior_precluso": enquadramento.ciclo_anterior,
            "ciclo_seguinte_precluso": enquadramento.ciclo_seguinte,
            "valor_pc":         _tofl(valor),
            "fator_acumulado":  _tofl(fator, default=None),
            "fator_efetivo_considerado": (
                1.0 if enquadramento.e_precluso or efeito_pc == "Nao"
                else _tofl(fator, default=None)
            ),
            "valor_atualizado": _tofl(vatual, default=None),
            "valor_historico_considerado": valor_hist_considerado,
            "dentro_do_corte":  _pc_dentro_do_corte(data_pc, data_corte),
            "efeito_financeiro_pc": efeito_pc,
            "retroativo_reconhecido_a_pagar": retro_lido,
            "valor_atualizado_em_analise": analise_lida,
            "delta_potencial": delta_lido,
            "entra_no_calculo": str(entra or "Sim"),
            "pc_pago_a_contratada": pago,
            "observacao":       obs,
            "check":            ws.cell(r, col_check).value if col_check else None,
            "is_global":        is_global,
            "layout_corrigido":  "legado_deslocado" if layout_corrigido else "",
            "layout_inteligente": layout_inteligente,
            "inferencias":       inferencias_linha,
            "pagamento_inferido": pagamento_inferido,
            "status_pagamento_pc": campos_vta.get("status_pagamento_pc"),
            "valor_efetivamente_pago": campos_vta.get("valor_pago"),
            "data_pagamento_pc": campos_vta.get("data_pagamento_pc"),
            "elegivel_retroativo_pc": bool(campos_vta.get("elegivel_retroativo_pc")),
            "layout_vta_corrigido": "W:AD" if desloc_vta else "",
            "campos_vta":       campos_vta,
        }
        resultado["itens"].append(registro)
        if is_global:
            resultado["pc_globais"].append(registro)
        else:
            resultado["pc_unitarios"].append(registro)

    # Controle de duplicidade por NUMERO_PC. Normalizacao APENAS para
    # comparacao (espacos externos removidos, sem distincao de caixa);
    # o valor original informado no arquivo e preservado nos registros.
    if col_num:
        vistos: dict[str, dict[str, Any]] = {}
        for reg in resultado["itens"]:
            bruto = reg.get("numero_pc")
            if bruto in (None, ""):
                continue
            chave = str(bruto).strip().casefold()
            if not chave:
                continue
            info = vistos.setdefault(chave, {"original": bruto, "linhas": []})
            info["linhas"].append(reg["linha"])
        for info in vistos.values():
            if len(info["linhas"]) > 1:
                resultado["alertas"].append(
                    f"NUMERO_PC duplicado: '{info['original']}' nas linhas "
                    f"{info['linhas']}."
                )

    if pagamentos_nao_confirmados:
        resultado["alertas"].append(
            f"{pagamentos_nao_confirmados} PC(s) sem PC_PAGO_A_CONTRATADA "
            "informado: pagamento nao confirmado; retroativo classificado "
            "como potencial."
        )

    resultado["totais"] = {
        "total_original":   sum((i["valor_pc"] or 0.0) for i in resultado["itens"]),
        "total_atualizado": sum((i["valor_atualizado"] or 0.0) for i in resultado["itens"]),
        "count_pcs":        len(resultado["itens"]),
        "count_unitarios":  len(resultado["pc_unitarios"]),
        "count_globais":    len(resultado["pc_globais"]),
    }
    resultado["totais_canonicos"] = _totais_canonicos_pc(
        resultado["itens"], data_corte
    )
    resultado["masterfile_inteligente"]["linhas_lidas"] = len(resultado["itens"])
    resultado["masterfile_inteligente"]["linhas_com_inferencia"] = sum(
        1 for i in resultado["itens"] if i.get("layout_inteligente")
    )
    resultado["ok"] = bool(resultado["itens"])
    return resultado


def _promover_resumo_pc(res: dict[str, Any]) -> None:
    """Deriva resumo operacional quando modo PC tem dados validos em itens_PC."""
    itens_pc = res.get("itens_pc_v10") or {}
    if not itens_pc.get("ok"):
        return

    totais = itens_pc.get("totais") or {}
    total_original = totais.get("total_original") or 0
    total_atualizado = totais.get("total_atualizado") or 0
    if not total_atualizado:
        total_atualizado = total_original

    resumo = dict(res.get("resumo") or {})
    for chave, valor in {
        "execucao_atualizada": total_atualizado,
        "valor_total_atualizado": total_atualizado,
        "valor_total_original": total_original,
        "retroativo": 0,
        "saldo_remanescente": 0,
    }.items():
        if resumo.get(chave) in (None, ""):
            resumo[chave] = valor
    resumo["modo_efetivo"] = "Pedido de Compra (itens_PC)"
    resumo["fonte_dados"] = "itens_PC"
    resumo["status_saldo"] = "PC reconhecido"
    res["resumo"] = resumo
    res["itens_pc"] = itens_pc


def _ler_execucao_saldo(wb) -> dict[str, Any]:
    """Le aba itens_Execucao_Saldo (v10.1) por cabecalho.

    Tolera ausencia da aba — retorna estrutura vazia com aviso informativo.
    Gera alertas V018-V021 quando detecta inconsistencias.
    Nao bloqueia processamento.
    """
    resultado: dict[str, Any] = {
        "itens": [], "totais": {}, "alertas": [], "ok": False,
        "aba_presente": False,
    }

    if ABA_EXECUCAO_SALDO not in wb.sheetnames:
        resultado["alertas"].append(
            f"Aba '{ABA_EXECUCAO_SALDO}' nao presente — entrada XLS consolidada nao preenchida "
            f"(opcional em v10.1; ciclos e indices devem ser fornecidos pela GCC)."
        )
        return resultado

    resultado["aba_presente"] = True
    ws = wb[ABA_EXECUCAO_SALDO]
    mapa = _mapear_colunas_por_cabecalho(ws)

    col_item   = mapa.get("item") or 1
    col_desc   = mapa.get("descricao_referencial") or mapa.get("descricao")
    col_qtdc   = mapa.get("qtd_contratada")
    col_vu     = mapa.get("vu_original") or mapa.get("vu")
    col_valtot = mapa.get("valor_total_original") or mapa.get("valor_total")
    # v10.1: coluna renomeada para PC; aceita legado requisicao_sap
    col_sap    = mapa.get("pc") or mapa.get("requisicao_sap") or mapa.get("requisicao")
    col_qtde   = mapa.get("qtd_emitida")
    col_vale   = mapa.get("valor_emitido")
    col_qtds   = mapa.get("qtd_saldo")
    col_vals   = mapa.get("valor_saldo")
    col_check  = mapa.get("check_fisico") or mapa.get("check")
    col_obs    = mapa.get("observacao")

    def _tofl(v: Any, default: float = 0.0) -> float:
        try:
            return float(v) if v not in (None, "") else default
        except (TypeError, ValueError):
            return default

    total_emitido = 0.0

    for r in range(2, ws.max_row + 1):
        item = ws.cell(r, col_item).value
        if not item:
            continue

        qtd_c  = _tofl(ws.cell(r, col_qtdc).value  if col_qtdc  else None)
        vu_ori = _tofl(ws.cell(r, col_vu).value    if col_vu    else None)
        qtd_e  = _tofl(ws.cell(r, col_qtde).value  if col_qtde  else None)
        val_e  = _tofl(ws.cell(r, col_vale).value  if col_vale  else None)
        qtd_s  = _tofl(ws.cell(r, col_qtds).value  if col_qtds  else None)
        val_s  = _tofl(ws.cell(r, col_vals).value  if col_vals  else None)

        # V018: QTD_SALDO negativo
        if qtd_s < -0.001:
            resultado["alertas"].append(
                f"V018 [{item}]: QTD_SALDO={qtd_s:.2f} negativo — execucao acima do contratado."
            )

        # V019: CHECK_FISICO — QTD_CONT <> QTD_EMIT + QTD_SALDO
        if qtd_c and (abs(qtd_c - qtd_e - qtd_s) > 0.011):
            resultado["alertas"].append(
                f"V019 [{item}]: QTD_CONTRATADA({qtd_c:.2f}) <> "
                f"QTD_EMITIDA({qtd_e:.2f}) + QTD_SALDO({qtd_s:.2f}) — divergencia fisica."
            )

        # V020: preco unitario implicito difere >20% do VU_ORIGINAL
        if qtd_e > 0.001 and vu_ori > 0.001:
            vu_impl = val_e / qtd_e
            if abs(vu_impl - vu_ori) / vu_ori > 0.20:
                resultado["alertas"].append(
                    f"V020 [{item}]: VU implicito (VALOR_EMITIDO/QTD_EMITIDA={vu_impl:.2f}) "
                    f"diverge >20% de VU_ORIGINAL({vu_ori:.2f})."
                )

        # Calculos oficiais Python (sem dependencia de formula Excel)
        valor_total_calc = round(qtd_c * vu_ori, 2) if (qtd_c and vu_ori) else None
        _tem_fisico = (qtd_c > 0.001 and qtd_e >= 0 and qtd_s >= -0.001)
        check_fisico_calc = ("OK" if abs(qtd_c - qtd_e - qtd_s) <= 0.011 else "ALERTA") if _tem_fisico else ""
        total_emitido += val_e

        resultado["itens"].append({
            "item":                item,
            "descricao":           ws.cell(r, col_desc).value  if col_desc  else None,
            "qtd_contratada":      qtd_c,
            "vu_original":         vu_ori,
            "valor_total_original": _tofl(ws.cell(r, col_valtot).value if col_valtot else None),
            "pc":                  ws.cell(r, col_sap).value   if col_sap   else None,
            "requisicao_sap":      ws.cell(r, col_sap).value   if col_sap   else None,  # compat legado
            "qtd_emitida":         qtd_e,
            "valor_emitido":       val_e,
            "qtd_saldo":           qtd_s,
            "valor_saldo":         val_s,
            "check_fisico":        ws.cell(r, col_check).value if col_check else None,
            "valor_total_calculado":  valor_total_calc,
            "check_fisico_calculado": check_fisico_calc,
            "observacao":          ws.cell(r, col_obs).value   if col_obs   else None,
        })

    resultado["totais"] = {
        "count_itens":          len(resultado["itens"]),
        "total_qtd_contratada": sum(i["qtd_contratada"] for i in resultado["itens"]),
        "total_valor_emitido":  total_emitido,
        "total_qtd_saldo":      sum(i["qtd_saldo"]      for i in resultado["itens"]),
        "total_valor_saldo":    sum(i["valor_saldo"]     for i in resultado["itens"]),
    }
    resultado["ok"] = bool(resultado["itens"])
    return resultado


def _materializar_aberturas_temporais(registro: dict[str, Any]) -> None:
    """Deriva a fotografia da abertura de cada ciclo com temporalidade correta.

    ``QTD_REM_AJUSTADA_Cn`` soma ao remanescente informado pelo fiscal TODOS os
    deltas rotulados naquele ciclo, inclusive os de DATA_EFEITO posterior a
    abertura — o que retroage o aditivo e altera a fotografia declarada. A
    abertura canonica desconta essa parcela:

        QTD_REM_ABERTURA_Cn = QTD_REM_AJUSTADA_Cn - DELTA_POSTERIOR_ABERTURA_Cn

    Deltas de ciclos anteriores sempre tem efeito <= abertura de Cn, entao a
    subtracao isola exatamente o que ainda nao existia naquela data. Sem a
    coluna auxiliar (arquivos anteriores a essa camada), a abertura permanece
    igual a QTD_REM_AJUSTADA_Cn: nenhuma regressao.
    """
    for n in range(5):
        ajustada = registro.get(f"QTD_REM_AJUSTADA_C{n}")
        posterior = registro.get(f"DELTA_POSTERIOR_ABERTURA_C{n}")
        if isinstance(ajustada, (int, float)) and not isinstance(ajustada, bool) \
                and isinstance(posterior, (int, float)) \
                and not isinstance(posterior, bool):
            registro[f"QTD_REM_ABERTURA_C{n}"] = round(
                float(ajustada) - float(posterior), 2
            )
            registro[f"ALTERACAO_POSTERIOR_ABERTURA_C{n}"] = float(posterior)
        else:
            registro[f"QTD_REM_ABERTURA_C{n}"] = ajustada
            registro[f"ALTERACAO_POSTERIOR_ABERTURA_C{n}"] = 0.0


def _ler_posicao_contratual(
    wb,
    origem: bytes | str | None = None,
    *,
    validar_geometria: bool = False,
    contexto=None,
) -> dict[str, Any]:
    """Le a aba posicao_contratual do novo modelo oficial (por cabecalho).

    A aba e 100% calculada (formulas). Com data_only=True, arquivos gerados
    por openpyxl sem recalculo do Excel retornam None nas formulas — nesse
    caso o resultado marca cache_ausente=True e o processamento nao deve
    tratar os campos como validos. A deteccao compara a leitura data_only
    (cache) com a presenca de formula no arquivo original (``origem``):
    formula presente + valor None = arquivo nunca recalculado pelo Excel.
    """
    from _coleta_oficial import ABA_POSICAO_CONTRATUAL, COLUNAS_POSICAO_CONTRATUAL

    resultado: dict[str, Any] = {
        "ok": False, "itens": [], "alertas": [],
        "cabecalhos_ausentes": [], "cache_ausente": False,
    }
    if ABA_POSICAO_CONTRATUAL not in wb.sheetnames:
        resultado["alertas"].append("Aba posicao_contratual ausente.")
        return resultado

    ws = wb[ABA_POSICAO_CONTRATUAL]
    mapa = _mapear_colunas_por_cabecalho(ws)
    resultado["cabecalhos_ausentes"] = [
        c for c in COLUNAS_POSICAO_CONTRATUAL if not _col(mapa, c)
    ]
    if resultado["cabecalhos_ausentes"]:
        resultado["alertas"].append(
            "posicao_contratual: cabecalhos ausentes: "
            + ", ".join(resultado["cabecalhos_ausentes"])
        )
        return resultado

    col_item = _col(mapa, "ITEM")
    linhas_com_item = 0
    valores_presentes = 0
    for r in range(2, ws.max_row + 1):
        item = ws.cell(r, col_item).value
        if item is None or str(item).strip() == "":
            continue
        linhas_com_item += 1
        registro: dict[str, Any] = {"origem_linha": r}
        for cab in COLUNAS_POSICAO_CONTRATUAL:
            v = ws.cell(r, _col(mapa, cab)).value
            registro[cab] = v
            if v is not None:
                valores_presentes += 1
        # Camada temporal (colunas auxiliares, quando presentes): deltas com
        # DATA_EFEITO POSTERIOR a abertura de cada ciclo. Ausentes em arquivos
        # anteriores a essa camada — nesse caso a abertura cai no comportamento
        # anterior, sem regressao.
        for n in range(5):
            for cab in (f"DELTA_POSTERIOR_ABERTURA_C{n}",
                        f"QTD_CONTRATUAL_ABERTURA_C{n}"):
                col_aux = _col(mapa, cab)
                registro[cab] = ws.cell(r, col_aux).value if col_aux else None
        _materializar_aberturas_temporais(registro)
        resultado["itens"].append(registro)

    if linhas_com_item and valores_presentes <= linhas_com_item:
        # Apenas a coluna ITEM tem valor: formulas sem cache de recalculo.
        resultado["cache_ausente"] = True
    elif not linhas_com_item and origem is not None:
        # Sem nenhum valor lido: pode ser aba vazia legitima (sem itens) ou
        # arquivo sem cache — a propria coluna ITEM e formula e vem None.
        # Distingue reabrindo o arquivo com data_only=False: se ha formula
        # onde a leitura data_only retornou None, o cache esta ausente.
        try:
            if contexto is not None:
                # PERF-ARCH-1: no fluxo oficial este caminho raro deixa de ser
                # uma materializacao extra. O contexto ja abre (ou vai abrir)
                # a representacao de formulas para ler_coleta_reajuste — a
                # pergunta daqui viaja de carona nela, e o gate de geometria
                # corre dentro do proprio contexto, antes da entrega.
                wb_formulas = contexto.workbook_formulas
            else:
                fonte2 = BytesIO(origem) if isinstance(origem, (bytes, bytearray)) else origem
                # PERF-TEST-1: esta reabertura existe para UMA pergunta — a
                # celula tem formula? Responde-la materializava o workbook
                # inteiro (~2 milhoes de celulas, 1,69 s) para consultar duas
                # celulas. Em modo read_only o mesmo valor sai por streaming
                # em 0,04 s. data_only continua False: e a formula, nao o
                # cache, que responde a pergunta.
                wb_formulas = load_workbook(fonte2, data_only=False, read_only=True)
                # XSEC-09: segunda materializacao dos MESMOS bytes. A geometria
                # e propriedade do arquivo, entao aqui o gate so confirma o que
                # ja foi aprovado na abertura principal — mas confirma, para
                # que a regra "todo workbook aberto passa pelo gate" nao
                # dependa de quem chamou.
                if validar_geometria:
                    validar_geometria_workbook(wb_formulas)
            ws_f = wb_formulas[ABA_POSICAO_CONTRATUAL]
            a2_formula = str(ws_f.cell(2, col_item).value or "").startswith("=")
            a2_cache = ws.cell(2, col_item).value
            if a2_formula and a2_cache is None:
                resultado["cache_ausente"] = True
        except ErroSegurancaXlsx:
            # Erro de fronteira nao pode ser engolido pelo fail-safe de cache.
            raise
        except Exception:
            pass
    if resultado["cache_ausente"]:
        resultado["alertas"].append(
            "posicao_contratual: valores de formula indisponiveis (arquivo "
            "sem recalculo do Excel); abra e salve o arquivo no Excel antes "
            "do upload para materializar os valores."
        )
    resultado["ok"] = not resultado["cabecalhos_ausentes"]
    return resultado


_NOMES_RESULTADOS_XLS = (
    "METODO_RETROATIVO", "TOLERANCIA_DIVERGENCIA",
    "RETRO_FIN", "RETRO_PC", "RETRO_ITENS", "RETRO_OFICIAL",
    "VTA_CALCULADO", "AJUSTE_MANUAL_VTA", "VTA_MANUAL_OFICIAL", "VTA_FINAL",
    "QTD_REM_OFICIAL", "REM_BASE_OFICIAL", "REM_ATUALIZADO_OFICIAL",
)


def _ler_resultados_xls(wb) -> dict[str, Any]:
    """Le os intervalos nomeados dos resultados do modelo oficial.

    Desde a Etapa 26F, os nomes apontam para MEMORIA_RESULTADOS e a aba
    RESULTADOS e somente executiva. A leitura continua orientada pelos nomes,
    sem depender de coordenadas. Uso exclusivo de AUDITORIA/reconciliacao com
    o motor Python — nunca substitui o calculo. Valores vem do cache do Excel
    (data_only=True); sem recalculo, tudo retorna None e cache_ausente=True.
    """
    resultado: dict[str, Any] = {
        "disponivel": False, "nomes_presentes": [], "valores": {},
        "cache_ausente": False,
    }
    if "RESULTADOS" not in wb.sheetnames:
        return resultado
    for nome in _NOMES_RESULTADOS_XLS:
        try:
            dn = wb.defined_names[nome]
        except KeyError:
            continue
        resultado["nomes_presentes"].append(nome)
        try:
            aba, ref = list(dn.destinations)[0]
            resultado["valores"][nome] = wb[aba][ref.replace("$", "")].value
        except Exception:
            resultado["valores"][nome] = None
    resultado["disponivel"] = bool(resultado["nomes_presentes"])
    # Cache avaliado apenas sobre os nomes CALCULADOS (formulas); nomes de
    # entrada manual (ex.: TOLERANCIA_DIVERGENCIA) tem valor no proprio
    # template e nao indicam recalculo.
    calculados = (
        "RETRO_FIN", "RETRO_PC", "RETRO_ITENS", "RETRO_OFICIAL",
        "VTA_CALCULADO", "VTA_FINAL",
        "QTD_REM_OFICIAL", "REM_BASE_OFICIAL", "REM_ATUALIZADO_OFICIAL",
    )
    presentes_calc = [n for n in calculados if n in resultado["nomes_presentes"]]
    resultado["cache_ausente"] = bool(presentes_calc) and all(
        resultado["valores"].get(n) is None for n in presentes_calc
    )
    return resultado


def _num_ou_none(v: Any) -> float | None:
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    return None


def _ler_referencias_vta(wb) -> dict[str, Any]:
    """Le as TRES REFERENCIAS do VTA da Tabela 1 (RESULTADOS!8:13) para AUDITORIA.

    Superficie de leitura = a propria Tabela 1 executiva (rotulos + valores +
    situacao), espelho de MEMORIA_RESULTADOS!W48/W50/W51/W52. Uso EXCLUSIVO de
    apresentacao/auditoria nos documentos: NUNCA substitui VTA_FINAL/B26. Valores
    vem do cache (data_only): sem recalculo, tudo None (fail-closed, sem inventar).

    Classificacao fixa:
      * forma1 (posicao atual)        -> REFERENCIA AUDITAVEL (vazia sem CICLO)
      * forma2 (ultima abertura)      -> REFERENCIA AUDITAVEL
      * forma3 (integral reajustado)  -> COMPARATIVO (nunca oficial)
    """
    out: dict[str, Any] = {
        "disponivel": False,
        "forma1_posicao_atual": None,
        "forma2_ultima_abertura": None,
        "forma3_integral_reajustado": None,
        "reconciliacao_valor": None,
        "reconciliacao_status": None,
        "forma1_situacao": None,
        "forma2_situacao": None,
        "ciclo_ultima_abertura": None,
        "fallback_razao": None,
        "ciclo_vigente": None,
        "data_posicao_atual": None,
        "posicao_atual_disponivel": False,
        "fontes": {
            "forma1": "MEMORIA_RESULTADOS!W50 (RESULTADOS!B10)",
            "forma2": "MEMORIA_RESULTADOS!W48 (RESULTADOS!B11)",
            "forma3": "comparativo_VTA!B208 (RESULTADOS!B12)",
            "reconciliacao": "MEMORIA_RESULTADOS!W51/W52 (RESULTADOS!B13/H13)",
            "data_posicao": "CICLO_EM_EXECUCAO!D5",
            "ciclo_vigente": "CONTROLE!B2",
        },
    }
    if "RESULTADOS" not in wb.sheetnames:
        return out
    res = wb["RESULTADOS"]
    out["disponivel"] = True
    out["forma1_posicao_atual"] = _num_ou_none(res["B10"].value)
    out["forma2_ultima_abertura"] = _num_ou_none(res["B11"].value)
    out["forma3_integral_reajustado"] = _num_ou_none(res["B12"].value)
    out["reconciliacao_valor"] = _num_ou_none(res["B13"].value)
    st10, st11, st13 = res["H10"].value, res["H11"].value, res["H13"].value
    out["forma1_situacao"] = str(st10).strip() if st10 not in (None, "") else None
    out["forma2_situacao"] = str(st11).strip() if st11 not in (None, "") else None
    out["reconciliacao_status"] = str(st13).strip() if st13 not in (None, "") else None
    if "MEMORIA_RESULTADOS" in wb.sheetnames:
        mem = wb["MEMORIA_RESULTADOS"]
        w46 = mem["W46"].value
        if isinstance(w46, (int, float)) and not isinstance(w46, bool):
            out["ciclo_ultima_abertura"] = int(w46)
        w47 = mem["W47"].value
        out["fallback_razao"] = str(w47).strip() if w47 not in (None, "") else None
    if "CONTROLE" in wb.sheetnames:
        b2 = wb["CONTROLE"]["B2"].value
        out["ciclo_vigente"] = str(b2).strip() if b2 not in (None, "") else None
    if "CICLO_EM_EXECUCAO" in wb.sheetnames:
        ciclo = wb["CICLO_EM_EXECUCAO"]
        d5 = ciclo["D5"].value
        out["data_posicao_atual"] = _norm_data(d5) if d5 not in (None, "") else None
        a9 = ciclo["A9"].value  # total: nao-vazio => posicao completa e valida
        out["posicao_atual_disponivel"] = isinstance(a9, (int, float)) and not isinstance(a9, bool)
    return out


def _detectar_versao(wb) -> str:
    if "CONTROLE" in wb.sheetnames:
        val = wb["CONTROLE"]["B10"].value
        if val and "v10" in str(val).lower():
            return "v10-rc"
    if ABA_HISTORICO_VU in wb.sheetnames:
        return "v10-rc"
    return "v9"


# ---------------------------------------------------------------------------
# Entrada publica
# ---------------------------------------------------------------------------

def ler_masterfile_v10(
    origem: bytes | str,
    *,
    exigir_modelo_oficial: bool = False,
    contexto=None,
) -> dict[str, Any]:
    """
    Le Masterfile v10 RC. Aceita bytes (upload) ou caminho.
    Retorna dict com ok, erro, avisos, controle, resumo,
    historico_vu, itens_consumidos_v10, itens_pc_v10, posicao_contratual,
    versao_detectada.

    exigir_modelo_oficial=True (fluxo de upload): rejeita arquivos sem a aba
    posicao_contratual com mensagem controlada — o novo XLS Coleta e o modelo
    oficial. False (padrao): compatibilidade temporaria com estruturas
    anteriores, amparada na politica de versionamento existente
    (_detectar_versao), para consumidores internos e testes.

    contexto (PERF-ARCH-1): contexto da execucao corrente do fluxo oficial.
    Quando fornecido, o workbook data_only=True deste arquivo ja foi aberto e
    ja passou pelo gate de geometria dentro do contexto, e e reaproveitado em
    vez de repagar o parse. Sem contexto, a leitura isolada abre os proprios
    bytes exatamente como antes — inclusive quanto a quando o gate roda.
    """
    res: dict[str, Any] = {
        "ok": False, "erro": "", "avisos": [], "abas_ausentes": [],
        "controle": {}, "resumo": {}, "matrizes": {},
        "parametros_v10": {}, "historico_vu": {},
        "itens_consumidos_v10": {}, "itens_pc_v10": {},
        "masterfile_fiscal_v2": {}, "masterfile_fiscal_definitivo": {},
        "vta_sombra": {},
        "event_log_sombra": {}, "estado_contratual_sombra": {},
        "posicao_contratual_sombra": {},
        "posicao_contratual": {},
        "ciclo_em_execucao": {},
        "resultados_xls": {},
        "objeto_processo": {},
        "execucao_saldo": {},          # v10.1 — entrada fiscal consolidada
        "versao_detectada": "",
        # Etapa 4 — memoria de calculo persistida (parametros!J2:R80);
        # leitura opcional: arquivos legados sem o bloco retornam {}.
        "memoria_calculo": {},
        # Compatibilidade com _documentos_masterfile (usa .get())
        "itens_d": {}, "itens_pc": {},
    }

    try:
        if isinstance(origem, (bytes, bytearray, memoryview)):
            origem = garantir_xlsx_validado(origem)
        if contexto is not None:
            # PERF-ARCH-1: o contexto da execucao ja materializou ESTES bytes
            # em data_only=True e ja aplicou validar_geometria_workbook antes
            # de entregar o objeto — o gate XSEC-09 continua correndo antes do
            # primeiro acesso a celula, so que uma vez por upload em vez de
            # uma vez por leitor.
            wb = contexto.workbook_valores
        else:
            fonte = BytesIO(origem) if isinstance(origem, (bytes, bytearray)) else origem
            wb = load_workbook(fonte, data_only=True)
            # XSEC-09: em chamada isolada este e o PRIMEIRO parser vivo. O gate
            # de geometria precisa correr aqui, antes de qualquer percurso de
            # celulas — as varreduras deste modulo usam ws.cell(), que
            # MATERIALIZA a celula ausente, entao um max_row inflado vira custo
            # real de memoria.
            #
            # So no modelo oficial: a allowlist de abas descreve a Coleta, e o
            # caminho legado (exigir_modelo_oficial=False) le Masterfiles com
            # abas "historico"/"validacoes", que estao fora dela. O upload real
            # e o unico consumidor nao confiavel e sempre passa
            # exigir_modelo_oficial.
            if exigir_modelo_oficial:
                validar_geometria_workbook(wb)
    except ErroSegurancaXlsx as exc:
        res["erro"] = str(exc)
        return res
    except Exception:
        res["erro"] = "O arquivo enviado não é um XLSX válido ou está corrompido."
        return res

    res["versao_detectada"] = _detectar_versao(wb)

    abas_lower = {a.lower() for a in wb.sheetnames}

    # Novo modelo oficial do XLS Coleta: aba posicao_contratual obrigatoria
    # no fluxo de upload (exigir_modelo_oficial=True) — modelo antigo e
    # rejeitado de forma controlada. Consumidores internos e testes de
    # estruturas anteriores seguem a politica de versionamento existente.
    from _coleta_oficial import ABAS_COLETA_OFICIAL, ABA_POSICAO_CONTRATUAL
    modelo_oficial = ABA_POSICAO_CONTRATUAL.lower() in abas_lower
    if exigir_modelo_oficial and not modelo_oficial:
        res["erro"] = (
            "O arquivo enviado utiliza uma versão anterior do XLS Coleta. "
            "Gere ou utilize o novo modelo oficial, que contém a aba "
            "posicao_contratual."
        )
        return res

    if modelo_oficial:
        # Etapa 3: posicao_referencia e OPCIONAL para leitura — Coletas oficiais
        # geradas antes da Etapa 3 nao a possuem e devem continuar sendo lidas
        # (compatibilidade retroativa; a aba pertence ao layout novo, mas sua
        # ausencia nao descaracteriza o modelo oficial). cobertura_temporal
        # (diagnostico desta etapa) segue a mesma politica de opcionalidade.
        # comparativo_VTA e aba de saida/referencia (nao aparece na web e nao
        # e necessaria para leitura); Coletas geradas antes dela devem seguir
        # sendo lidas (compatibilidade retroativa).
        ABAS_OPCIONAIS_COMPAT = {
            "posicao_referencia",
            "cobertura_temporal",
            "comparativo_VTA",
            "MEMORIA_RESULTADOS",
            "CICLO_EM_EXECUCAO",
        }
        res["abas_ausentes"] = [
            a for a in ABAS_COLETA_OFICIAL
            if a.lower() not in abas_lower and a not in ABAS_OPCIONAIS_COMPAT
        ]
        if res["abas_ausentes"]:
            res["erro"] = (
                "Arquivo nao segue o modelo oficial do XLS Coleta. "
                "Abas ausentes: " + ", ".join(res["abas_ausentes"])
            )
            return res
        # No modelo oficial NUMERO_PC e obrigatorio em itens_PC: e o
        # identificador documental do Pedido de Compra e a chave do
        # controle de duplicidade. Arquivo oficial sem a coluna e
        # template incompativel — rejeitado no fluxo de upload.
        if exigir_modelo_oficial and ABA_ITENS_PC in wb.sheetnames:
            mapa_pc_oficial = _mapear_colunas_por_cabecalho(wb[ABA_ITENS_PC])
            if not mapa_pc_oficial.get("numero_pc"):
                res["erro"] = (
                    "Template incompativel: a aba itens_PC do modelo oficial "
                    "exige a coluna NUMERO_PC (identificador documental do "
                    "Pedido de Compra). Gere o arquivo a partir do template "
                    "oficial corrigido."
                )
                return res
    else:
        res["abas_ausentes"] = [
            a for a in ABAS_OBRIGATORIAS_V10 if a.lower() not in abas_lower
        ]
        abas_estruturais_ausentes = [
            a for a in ABAS_OBRIGATORIAS_V10
            if a not in (ABA_HISTORICO_VU, "validacoes")
            and a.lower() not in abas_lower
        ]
        if abas_estruturais_ausentes:
            res["erro"] = (
                "Arquivo nao e um Masterfile valido. Abas ausentes: "
                + ", ".join(abas_estruturais_ausentes)
            )
            return res
        if res["abas_ausentes"]:
            res["avisos"].append(
                f"Abas novas v10 ausentes: {res['abas_ausentes']} "
                f"— arquivo pode ser v9 legado."
            )

    # Etapa 4 — bloco MEMORIA DE CALCULO (opcional; nunca bloqueia a leitura).
    if "parametros" in wb.sheetnames:
        try:
            from _memoria_calculo import ler_memoria_calculo
            res["memoria_calculo"] = ler_memoria_calculo(wb["parametros"])
        except Exception:
            res["memoria_calculo"] = {}

    c = wb["CONTROLE"]
    modo_bruto = str(_achar_valor(c, "modo de leitura") or "").strip()
    ciclo      = _achar_valor(c, "ciclo vigente (em execucao)")
    # Rotulo padronizado ("DATA DE CORTE DA APURACAO") e o legado
    # ("Data de corte (unica p/ contrato)") apontam para a MESMA celula
    # (CONTROLE!B3): a leitura por prefixo aceita os dois sem conversao.
    corte      = _valor_por_prefixo(c, "data de corte")
    indice     = _achar_valor(c, "indice utilizado")
    res["controle"] = {
        "modo":          _normalizar_modo(modo_bruto),
        "modo_bruto":    modo_bruto,
        "ciclo_vigente": str(ciclo or "").strip(),
        "data_corte":    corte,
        # Etapa 26H: indice contratual (CONTROLE!B7) propagado pela cadeia
        # canonica ate o Sumario Executivo (sem hardcode no renderer).
        "indice":        str(indice or "").strip(),
        "versao":        res["versao_detectada"],
    }

    # Integracao cobertura_temporal (aba de diagnostico GCC). Expoe a fonte
    # canonica para o motor: data_analise em controle e confirmacoes GCC em
    # confirmacao_gcc. Ausencia da aba => campos None (compativel com arquivos
    # oficiais anteriores; ver ABAS_OPCIONAIS_COMPAT). MAX nunca vira aqui.
    cobertura = _ler_cobertura_temporal(wb)
    res["cobertura_temporal"] = cobertura
    res["controle"]["data_analise"] = cobertura.get("data_analise")
    res["confirmacao_gcc"] = {
        "financeiro_ate": cobertura.get("financeiro_confirmado_completo_ate"),
        "pc_ate":         cobertura.get("pc_confirmado_completo_ate"),
    }
    # Competencias financeiras (mensais) para o motor de cobertura temporal.
    res["financeiro"] = _ler_financeiro_competencias(wb)
    # Posicao fisica do fiscal (itens + fotografia recente posicao_referencia!B).
    res["itens_base"], res["remanescente_atual"] = _ler_posicao_fisica_fiscal(wb)

    if "historico" in wb.sheetnames:
        try:
            from _leitor_masterfile import INDICADORES
            h = wb["historico"]
            resumo: dict[str, Any] = {}
            for rotulo, chave in INDICADORES.items():
                resumo[chave] = _achar_valor(h, rotulo, max_lin=60)
            res["resumo"] = resumo
        except ImportError:
            res["avisos"].append(
                "Nao foi possivel importar INDICADORES de _leitor_masterfile."
            )

    res["parametros_v10"] = _ler_parametros_v10(wb)
    for alerta in res["parametros_v10"].get("alertas", []):
        res["avisos"].append(f"parametros: {alerta}")

    res["masterfile_fiscal_v2"] = _resumo_masterfile_fiscal_v2(wb)
    res["masterfile_fiscal_definitivo"] = _resumo_masterfile_fiscal_definitivo(wb)
    layout_fiscal_definitivo = bool(
        res["masterfile_fiscal_definitivo"].get("layout_fiscal_definitivo")
    )
    normalizacoes_fiscais: dict[str, Any] = {
        "financeiro_parcelas": 0,
        "historico_financeiro_parcelas": 0,
        "consumidos_itens": 0,
        "remanescentes_itens": 0,
    }

    res["historico_vu"] = _ler_historico_vu(wb)
    for alerta in res["historico_vu"].get("alertas", []):
        res["avisos"].append(f"historico_VU: {alerta}")

    res["posicao_contratual"] = _ler_posicao_contratual(
        wb, origem, validar_geometria=exigir_modelo_oficial, contexto=contexto
    )
    for alerta in res["posicao_contratual"].get("alertas", []):
        res["avisos"].append(alerta)

    # RESULTADOS (novo modelo): valores nomeados para reconciliacao/auditoria
    # XLS x Python — nunca fonte exclusiva de calculo.
    res["resultados_xls"] = _ler_resultados_xls(wb)
    # Tres referencias do VTA (auditoria/apresentacao): posicao atual, ultima
    # abertura e integral reajustado. Nunca substitui o VTA oficial (B26).
    res["referencias_vta"] = _ler_referencias_vta(wb)

    # Cadastro visivel de itens do contrato (fallback de VU/QTD_CONTRATADA
    # quando as abas ocultas nao vierem populadas pela Calculadora).
    res["itens_contrato"] = _ler_itens_contrato(wb)

    # v10.1 — execucao_saldo (aba opcional)
    res["execucao_saldo"] = _ler_execucao_saldo(wb)
    for alerta in res["execucao_saldo"].get("alertas", []):
        nivel = "V018" if "V018" in alerta else (
                "V019" if "V019" in alerta else (
                "V020" if "V020" in alerta else (
                "V021" if "V021" in alerta else "INFO")))
        if nivel.startswith("V"):
            res["avisos"].append(f"execucao_saldo [{nivel}]: {alerta}")
        else:
            res["avisos"].append(f"execucao_saldo: {alerta}")

    # Pre-populacao logica: se historico_VU nao tem VU_ORIGINAL para um item
    # e execucao_saldo tem, usar como default. Nao sobrescreve dado manual.
    if res["execucao_saldo"].get("ok") and res["historico_vu"].get("itens"):
        _vu_fiscal: dict[str, float] = {
            str(i["item"]).strip(): i["vu_original"]
            for i in res["execucao_saldo"]["itens"]
            if i.get("vu_original")
        }
        for item_hvu in res["historico_vu"]["itens"]:
            chave = str(item_hvu.get("item", "")).strip()
            if item_hvu.get("vu_original") is None and chave in _vu_fiscal:
                item_hvu["vu_original"] = _vu_fiscal[chave]
                item_hvu["_vu_origem"] = "execucao_saldo"
            elif item_hvu.get("vu_original") is not None and chave in _vu_fiscal:
                diff = abs(item_hvu["vu_original"] - _vu_fiscal[chave])
                base = abs(_vu_fiscal[chave]) or 1.0
                if diff / base > 0.05:
                    res["avisos"].append(
                        f"historico_VU x execucao_saldo: item '{chave}' — "
                        f"VU_ORIGINAL diverge >5% "
                        f"(hist_VU={item_hvu['vu_original']:.2f}, "
                        f"exec_saldo={_vu_fiscal[chave]:.2f})."
                    )

    if layout_fiscal_definitivo:
        fiscal_remanescentes = _normalizar_fiscal_remanescentes(wb, res)
        if fiscal_remanescentes.get("ok"):
            res["execucao_saldo"] = fiscal_remanescentes
            normalizacoes_fiscais["remanescentes_itens"] = len(
                fiscal_remanescentes.get("itens") or []
            )
            for alerta in fiscal_remanescentes.get("alertas", []):
                res["avisos"].append(f"ENTRADA_XLS_REMANESCENTES: {alerta}")
            resumo = dict(res.get("resumo") or {})
            saldo = fiscal_remanescentes.get("totais", {}).get("total_valor_saldo")
            if saldo and not _to_float_sombra(resumo.get("saldo_remanescente")):
                resumo["saldo_remanescente"] = saldo
                res["resumo"] = resumo

    # V021: VALOR_EMITIDO de execucao_saldo vs total de financeiro
    if res["execucao_saldo"].get("ok") and "financeiro" in wb.sheetnames:
        ws_fin = wb["financeiro"]
        total_fin = 0.0
        for r in range(2, ws_fin.max_row + 1):
            val = ws_fin.cell(r, 3).value   # col C = valor
            try:
                total_fin += float(val) if val not in (None, "") else 0.0
            except (TypeError, ValueError):
                pass
        if total_fin > 0.001:
            total_emitido = res["execucao_saldo"]["totais"].get("total_valor_emitido", 0.0)
            if total_emitido > 0.001:
                diff_pct = abs(total_emitido - total_fin) / total_fin
                if diff_pct > 0.05:
                    res["avisos"].append(
                        f"V021: VALOR_EMITIDO total de execucao_saldo "
                        f"(R$ {total_emitido:,.2f}) diverge {diff_pct:.1%} "
                        f"do total do financeiro (R$ {total_fin:,.2f}). "
                        f"Verificar qual fonte e mais completa."
                    )

    # Aditivos/supressoes informados pelo fiscal na aba visivel (Arquivo 3.0).
    res["aditivos_visiveis"] = _normalizar_fiscal_aditivos(wb, res["parametros_v10"])
    for alerta in res["aditivos_visiveis"].get("alertas", []):
        res["avisos"].append(f"ENTRADA_XLS_ADITIVOS: {alerta}")

    res["itens_consumidos_v10"] = _ler_itens_consumidos_v10(wb)
    if layout_fiscal_definitivo:
        fiscal_consumidos = _normalizar_fiscal_consumidos(wb, res)
        if fiscal_consumidos.get("ok"):
            res["itens_consumidos_v10"] = fiscal_consumidos
            normalizacoes_fiscais["consumidos_itens"] = len(
                fiscal_consumidos.get("itens") or []
            )
    for alerta in res["itens_consumidos_v10"].get("alertas", []):
        origem = "ENTRADA_XLS_CONSUMIDOS" if layout_fiscal_definitivo else "itens_Consumidos"
        res["avisos"].append(f"{origem}: {alerta}")

    # 29C.1: fotografia itemizada durante o ciclo. Leitura isolada e apenas
    # diagnostica; nao substitui a posicao_contratual nem alimenta o VTA.
    from _ciclo_em_execucao import ler_ciclo_em_execucao
    res["ciclo_em_execucao"] = ler_ciclo_em_execucao(
        wb,
        itens_consumidos=res["itens_consumidos_v10"],
    )
    if res["ciclo_em_execucao"].get("utilizado"):
        for alerta in res["ciclo_em_execucao"].get("alertas", []):
            res["avisos"].append(alerta)

    parcelas_sombra = []
    parcelas_sombra.extend(_ler_parcelas_sombra_financeiro(wb))
    if layout_fiscal_definitivo:
        parcelas_fiscal, alertas_fiscal_fin = _normalizar_fiscal_financeiro(
            wb, res["parametros_v10"]
        )
        _materializar_valor_atualizado(parcelas_fiscal, res["parametros_v10"])
        parcelas_sombra.extend(parcelas_fiscal)
        normalizacoes_fiscais["financeiro_parcelas"] = len(parcelas_fiscal)
        for alerta in alertas_fiscal_fin:
            res["avisos"].append(f"ENTRADA_XLS_FINANCEIRO: {alerta}")
        parcelas_hist, alertas_hist = _normalizar_fiscal_historico_financeiro(
            wb, res["parametros_v10"]
        )
        _materializar_valor_atualizado(parcelas_hist, res["parametros_v10"])
        parcelas_sombra.extend(parcelas_hist)
        normalizacoes_fiscais["historico_financeiro_parcelas"] = len(parcelas_hist)
        for alerta in alertas_hist:
            res["avisos"].append(f"CICLOS_PASSADOS: {alerta}")
    parcelas_sombra.extend(_ler_parcelas_sombra_saldo(wb))
    parcelas_sombra.extend(_ler_parcelas_sombra_aditivos(wb))
    if layout_fiscal_definitivo:
        res["masterfile_fiscal_definitivo"]["normalizacoes"] = normalizacoes_fiscais

    modo = _norm(res["controle"].get("modo", ""))
    if modo == "pc":
        res["itens_pc_v10"] = _ler_itens_pc_v10(
            wb,
            ciclo_vigente=res["controle"].get("ciclo_vigente"),
            data_corte=res["controle"].get("data_corte"),
        )
        for alerta in res["itens_pc_v10"].get("alertas", []):
            res["avisos"].append(f"itens_PC: {alerta}")
        res["vta_sombra"] = calcular_vta_sombra(
            res.get("resumo"), res.get("itens_pc_v10"), parcelas_sombra
        )
        for alerta in res["vta_sombra"].get("alertas", []):
            res["avisos"].append(f"vta_sombra: {alerta}")
        _promover_resumo_pc(res)
    else:
        res["vta_sombra"] = calcular_vta_sombra(res.get("resumo"), {}, parcelas_sombra)

    # Potencial futuro: saldo remanescente x fator concedido (aditivo).
    res["potencial_futuro"] = _montar_potencial_futuro(res)

    # Fase 4.2 — Event Log + Estado Contratual sombra (aditivo, paralelo ao
    # retorno oficial; nao altera VTA, historico!B51, formulas nem parsing).
    event_log = montar_event_log_sombra(parcelas_sombra, res.get("itens_pc_v10"))
    marco = res["controle"].get("ciclo_vigente") or ""
    estado = reconstruir_estado_contratual(event_log, marco)
    res["event_log_sombra"] = {
        "eventos_total": len(event_log.eventos),
        "eventos": [estado_contratual_para_dict(evento) for evento in event_log.eventos],
    }
    res["estado_contratual_sombra"] = estado_contratual_para_dict(estado)
    res["posicao_contratual_sombra"] = montar_posicao_contratual_sombra(res)

    # RFC Motor de Reconciliacao: camada paralela por ciclo, com decisoes
    # manuais vigentes carregadas do log apartado (JSONL por hash da entrada).
    _hash_arquivo = hash_entrada(origem if isinstance(origem, (bytes, bytearray)) else None)
    res["hash_entrada"] = _hash_arquivo
    res["reconciliacao"] = reconciliar_execucoes(
        res, carregar_decisoes(_hash_arquivo)
    )
    for alerta in res["reconciliacao"].get("alertas", []):
        res["avisos"].append(f"reconciliacao: {alerta}")

    # RFC v10.5.4 — Reconciliacao de Evidencias (grao evidencia/execucao,
    # slice PC + Remanescentes). Sombra aditiva: nao altera VTA, vta_sombra,
    # historico!B51 nem o motor de reconciliacao por ciclo acima.
    try:
        res["reconciliacao_evidencias_sombra"] = reconciliar_evidencias(
            res, decisoes=carregar_decisoes(_hash_arquivo)
        )
        for alerta in res["reconciliacao_evidencias_sombra"].get("alertas", []):
            res["avisos"].append(f"reconciliacao_evidencias: {alerta}")
    except Exception as exc:  # sombra nunca bloqueia o upload homologado
        res["reconciliacao_evidencias_sombra"] = {
            "ok": False,
            "alertas": [f"Falha tecnica controlada na reconciliacao: {exc}"],
        }
        res["avisos"].append(
            f"reconciliacao_evidencias: falha tecnica controlada: {exc}"
        )

    # Motor de Composicao do VTA: quadro auditavel que espelha a memoria
    # fiscal das apostilas (executado x fator por ciclo + saldo do corte
    # atualizado + aditivos pelo fator do ciclo-marco).
    res["composicao_vta"] = montar_composicao_vta(res)
    for alerta in res["composicao_vta"].get("alertas", []):
        res["avisos"].append(f"composicao_vta: {alerta}")

    if modo and modo not in MODOS_VALIDOS_V10:
        modos_str = ", ".join(f"'{m}'" for m in sorted(MODOS_VALIDOS_V10))
        res["avisos"].append(
            f"Modo nao reconhecido: '{res['controle']['modo']}' (esperados: {modos_str})."
        )

    res["ok"] = True
    res["objeto_processo"] = montar_objeto_processo_reajuste(res)

    # Etapa 6 — reconciliacao XLS (RESULTADOS) x motor Python: auditoria,
    # nunca fonte de calculo; divergencia relevante bloqueia na politica.
    from _reconciliacao_xls_python import reconciliar_xls_python
    res["reconciliacao_xls_python"] = reconciliar_xls_python(res)
    if isinstance(origem, (bytes, bytearray)):
        res["_bytes_arquivo"] = bytes(origem)
    return res


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Uso: py _leitor_masterfile_v10.py caminho\\do\\arquivo.xlsx")
        sys.exit(1)
    resultado = ler_masterfile_v10(sys.argv[1])
    print("ok:", resultado["ok"])
    print("versao:", resultado["versao_detectada"])
    if resultado["erro"]:
        print("ERRO:", resultado["erro"])
    for a in resultado["avisos"]:
        print("AVISO:", a)
    print("historico_VU itens:", len(resultado.get("historico_vu", {}).get("itens", [])))
