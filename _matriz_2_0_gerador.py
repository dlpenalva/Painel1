"""
_matriz_2_0_gerador.py
-----------------------
Gerador da Matriz 2.0 — cl8us FARC.
5 abas: BASE, EXECUCAO_FINANCEIRA, HISTORICO_CICLOS, ITENS_CONTRATADOS, ADITIVOS.
"""

from io import BytesIO
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Paleta
C = {
    "titulo":    "0F172A", "titulo_bg": "0F766E",
    "auto_bg":   "F1F5F9", "auto_fg":   "374151",
    "fiscal_bg": "FFF9C4", "fiscal_fg": "78350F",
    "gcc_bg":    "DBEAFE", "gcc_fg":    "1E3A8A",
    "resumo_bg": "D1FAE5", "resumo_fg": "065F46",
    "alrt_bg":   "FEE2E2", "alrt_fg":   "991B1B",
    "hdr_bg":    "1E293B", "hdr_fg":    "FFFFFF",
    "borda":     "CBD5E1", "branco":    "FFFFFF",
    "sup_bg":    "FFCCCC", "sup_fg":    "7F0000",
}

FMT_MOEDA  = '"R$" #,##0.00'
FMT_QTD    = '#,##0.##'
FMT_PCT    = '0.000%'
FMT_FATOR  = '0.000000'
FMT_DATA   = 'MM/YYYY'


def _f(hex_c): return PatternFill("solid", fgColor=hex_c)
def _ft(bold=False, color="0F172A", size=9, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Arial")
def _al(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def _bd():
    s = Side(style="thin", color=C["borda"])
    return Border(left=s, right=s, top=s, bottom=s)

def _c(ws, r, c, val="", bg="FFFFFF", fg="0F172A", bold=False,
       fmt=None, h="center", wrap=False, italic=False, size=9):
    cc = ws.cell(r, c, val)
    cc.fill = _f(bg); cc.font = _ft(bold, fg, size, italic)
    cc.alignment = _al(h, "center", wrap); cc.border = _bd()
    if fmt: cc.number_format = fmt
    return cc

def _titulo(ws, row, c1, c2, txt):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    c = ws.cell(row, c1, txt)
    c.fill = _f(C["titulo_bg"]); c.font = _ft(True, "FFFFFF", 11)
    c.alignment = _al("left"); c.border = _bd()
    ws.row_dimensions[row].height = 24

def _nota(ws, row, c1, c2, txt, bg=None, fg=None):
    bg = bg or C["auto_bg"]; fg = fg or C["auto_fg"]
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    c = ws.cell(row, c1, txt)
    c.fill = _f(bg); c.font = _ft(italic=True, color=fg, size=8)
    c.alignment = _al("left", "center", True); c.border = _bd()
    ws.row_dimensions[row].height = 18

def _hdr(ws, r, col, txt, bg=None, w=None):
    bg = bg or C["hdr_bg"]
    c = ws.cell(r, col, txt)
    c.fill = _f(bg); c.font = _ft(True, "FFFFFF", 8)
    c.alignment = _al("center", "center", True); c.border = _bd()
    ws.row_dimensions[r].height = max(ws.row_dimensions[r].height or 0, 48)
    if w: ws.column_dimensions[get_column_letter(col)].width = w
    return c

def _dv(ws, formula, ref):
    d = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(d); d.add(ref)

def _fmt_mmaaaa(valor):
    """Normaliza qualquer formato de data para MM/AAAA."""
    if not valor:
        return ""
    s = str(valor).strip().split(" ")[0]  # remove hora
    # yyyy-mm-dd
    if len(s) == 10 and s[4] == "-":
        try: return f"{int(s[5:7]):02d}/{s[0:4]}"
        except: pass
    # dd/mm/yyyy
    if len(s) == 10 and s[2] == "/":
        try: return f"{s[3:5]}/{s[6:10]}"
        except: pass
    # mm/yyyy (ja no formato certo)
    if len(s) == 7 and s[2] == "/":
        return s
    # mm/yyyy com mes 1 digito: m/yyyy
    if len(s) == 6 and s[1] == "/":
        try: return f"0{s[0]}/{s[2:]}"
        except: pass
    return s  # fallback — retorna como veio


def _fmt_ini_fin(valor):
    """Normaliza data para uso em formula DATE(). Retorna (ano,mes,dia) ou None."""
    if not valor:
        return None
    s = str(valor).strip().split(" ")[0]
    # yyyy-mm-dd
    if len(s) == 10 and s[4] == "-":
        try: return (int(s[0:4]), int(s[5:7]), int(s[8:10]))
        except: pass
    # dd/mm/yyyy
    if len(s) == 10 and s[2] == "/":
        try: return (int(s[6:10]), int(s[3:5]), int(s[0:2]))
        except: pass
    # mm/yyyy -> primeiro dia
    if len(s) == 7 and s[2] == "/":
        try: return (int(s[3:7]), int(s[0:2]), 1)
        except: pass
    return None


def _sessao():
    try:
        adm = st.session_state.get("dados_admissibilidade") or {}
    except Exception:
        adm = {}
    ciclos_raw = adm.get("ciclos", []) or []
    ciclos = []
    for i, c in enumerate(ciclos_raw[:4]):
        nome = str(c.get("ciclo", f"C{i+1}")).strip().upper()
        ini_raw = c.get("inicio_financeiro", c.get("data_inicio_financeiro",""))
        fim_raw = c.get("fim_financeiro",    c.get("data_fim_financeiro",""))
        db_raw  = c.get("data_base_original", c.get("data_base",""))
        ini_fmt = _fmt_mmaaaa(ini_raw)
        fim_fmt = _fmt_mmaaaa(fim_raw)
        db_fmt  = _fmt_mmaaaa(db_raw)
        ini_tup = _fmt_ini_fin(ini_raw)
        fim_tup = _fmt_ini_fin(fim_raw)
        ciclos.append({
            "ciclo":      nome,
            "periodo":    f"{ini_fmt} a {fim_fmt}" if ini_fmt and fim_fmt else "",
            "ini_fmt":    ini_fmt,
            "fim_fmt":    fim_fmt,
            "ini_fin":    ini_raw,  # mantido para _date_f nas formulas
            "fim_fin":    fim_raw,
            "ini_tup":    ini_tup,  # (ano,mes,dia) para EDATE/DATE
            "fim_tup":    fim_tup,
            "data_base":  db_fmt,
            "pct":        float(c.get("percentual_apurado", c.get("percentual",0)) or 0),
            "fator":      float(c.get("fator_ciclo", c.get("fator",1)) or 1),
            "fat_acum":   float(c.get("fator_acumulado",1) or 1),
            "efeito":     "Sim" if c.get("objeto_analise_atual") else "Nao",
        })
    c0 = {
        "ciclo": "C0", "periodo": "vigencia inicial",
        "ini_fmt": "", "fim_fmt": "", "ini_fin": "", "fim_fin": "",
        "ini_tup": None, "fim_tup": None, "data_base": "",
        "pct": 0, "fator": 1.0, "fat_acum": 1.0, "efeito": "Nao",
    }
    ciclo_corte = ciclos[-1]["ciclo"] if ciclos else "C0"
    return c0, ciclos, ciclo_corte, adm



def _parse_ym(val):
    s = str(val).strip().split(" ")[0]
    if "-" in s:
        p = s.split("-")
        try: return (int(p[0]), int(p[1]))
        except: pass
    if "/" in s:
        p = s.split("/")
        try:
            if len(p) == 2:
                a, b = p[0].strip(), p[1].strip()
                if len(b) == 4: return (int(b), int(a))
                if len(a) == 4: return (int(a), int(b))
            if len(p) == 3:
                return (int(p[2]), int(p[1]))
        except: pass
    return None


def _mapear_competencias(ciclos):
    mapa = []
    for cd in ciclos:
        ini = _parse_ym(cd.get("ini_fin","") or cd.get("inicio_financeiro",""))
        fim = _parse_ym(cd.get("fim_fin","") or cd.get("fim_financeiro",""))
        if not ini or not fim:
            continue
        ano, mes = ini[0], ini[1]
        while (ano, mes) <= (fim[0], fim[1]):
            mapa.append({"comp": f"{mes:02d}/{ano}",
                         "ciclo": cd["ciclo"],
                         "fat":   float(cd["fat_acum"])})
            mes += 1
            if mes > 12: mes = 1; ano += 1
    return mapa


def _construir_mapa_ciclos(wb, mapa):
    ws = wb.create_sheet("MAPA_CICLOS")
    ws.sheet_state = "hidden"
    for ci, txt in enumerate(["Competencia","Ciclo","Fator"], 1):
        ws.cell(1, ci, txt)
    for i, entry in enumerate(mapa, start=2):
        ws.cell(i, 1, entry["comp"])
        ws.cell(i, 2, entry["ciclo"])
        ws.cell(i, 3, entry["fat"])
    return len(mapa)


# ── ABA BASE ────────────────────────────────────────────────────

def _base(wb, c0, ciclos, ciclo_corte):
    ws = wb.create_sheet("BASE")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C["titulo_bg"]

    for col, w in [("A",34),("B",18),("C",16),("D",16),("E",16),
                   ("F",12),("G",12),("H",18),("I",14),("J",22),("K",22)]:
        ws.column_dimensions[col].width = w

    _titulo(ws, 1, 1, 11, "MATRIZ 2.0 — BASE")
    _nota(ws, 2, 1, 11,
          "  Aba 100% automatica: tabela-verdade dos ciclos, percentuais, fatores e parametros globais. "
          "Gerada pela Calculadora. Nao alterar.", C["auto_bg"], C["auto_fg"])
    for r in [3, 4]: ws.row_dimensions[r].height = 8

    ws.row_dimensions[5].height = 36
    hdrs5 = ["ID Ciclo","Periodo do Ciclo","Inicio Financeiro","Fim Financeiro",
             "Data-Base do Ciclo","% Reajuste do Ciclo","Fator do Ciclo",
             "Fator Acumulado Aplicavel","Possui Efeito Financeiro?","Parametro","Valor"]
    for ci, txt in enumerate(hdrs5, 1):
        _hdr(ws, 5, ci, txt)

    todos = [c0] + ciclos

    # Parametros — so na linha 6 (C0)
    params = [
        (6,  "Ciclo atual/corte",   ciclo_corte),
        (7,  "Maximo de ciclos",    f"C{len(ciclos)}"),
        (8,  "Observacao",          "Gerado automaticamente pela Calculadora cl8us."),
    ]

    for i, cd in enumerate(todos):
        r = 6 + i
        ws.row_dimensions[r].height = 18
        _c(ws, r, 1, cd["ciclo"],     C["auto_bg"], C["auto_fg"], bold=True)
        _c(ws, r, 2, cd.get("periodo",""),   C["auto_bg"], C["auto_fg"])
        _c(ws, r, 3, cd["ini_fin"],   C["auto_bg"], C["auto_fg"])
        _c(ws, r, 4, cd["fim_fin"],   C["auto_bg"], C["auto_fg"])
        _c(ws, r, 5, cd["data_base"], C["auto_bg"], C["auto_fg"])
        _c(ws, r, 6, cd["pct"],       C["auto_bg"], C["auto_fg"], fmt=FMT_PCT)
        _c(ws, r, 7, cd["fator"],     C["auto_bg"], C["auto_fg"], fmt=FMT_FATOR)
        _c(ws, r, 8, cd["fat_acum"],  C["auto_bg"], C["auto_fg"], fmt=FMT_FATOR)
        _c(ws, r, 9, cd["efeito"],    C["auto_bg"], C["auto_fg"])
        # Parametros SOMENTE na linha C0 (r=6)
        if r == 6:
            _c(ws, r, 10, params[0][1], C["gcc_bg"], C["gcc_fg"])
            _c(ws, r, 11, params[0][2], C["gcc_bg"], C["gcc_fg"], bold=True)
        else:
            _c(ws, r, 10, "", C["auto_bg"], C["auto_fg"])
            _c(ws, r, 11, "", C["auto_bg"], C["auto_fg"])

    # Linhas vazias ate C4
    for i in range(len(todos), 5):
        r = 6 + i
        ws.row_dimensions[r].height = 18
        _c(ws, r, 1, f"C{i}", C["auto_bg"], C["auto_fg"])
        for ci in range(2, 12):
            _c(ws, r, ci, "", C["auto_bg"], C["auto_fg"])

    # Separador linha 11
    ws.row_dimensions[11].height = 10
    for ci in range(1, 12):
        ws.cell(11, ci).fill = _f(C["titulo_bg"]); ws.cell(11, ci).border = _bd()

    # Resumo executivo linhas 12-18
    ws.row_dimensions[12].height = 20
    ws.merge_cells("A12:K12")
    t = ws.cell(12, 1, "RESUMO EXECUTIVO — VALOR TOTAL ATUALIZADO")
    t.fill = _f(C["resumo_bg"]); t.font = _ft(True, C["resumo_fg"], 10)
    t.alignment = _al("left"); t.border = _bd()
    for ci in range(2, 12): ws.cell(12, ci).fill = _f(C["resumo_bg"]); ws.cell(12, ci).border = _bd()

    resumo = [
        (13, "Execucao financeira considerada (Sim = atualizado; Nao = nominal)",
         '=IFERROR(ROUND(SUMIF(EXECUCAO_FINANCEIRA!D6:D305,"Sim",EXECUCAO_FINANCEIRA!J6:J305),2),0)'),
        (14, "Remanescente atual atualizado dos itens",
         '=IFERROR(ROUND(SUM(ITENS_CONTRATADOS!M6:M105),2),0)'),
        (15, "Aditivos/supressoes computaveis",
         '=IFERROR(ROUND(SUMIF(ADITIVOS!K6:K205,"Computar no VTA",ADITIVOS!L6:L205),2),0)'),
        (16, "VTA_TOTAL", "=B13+B14+B15"),
    ]
    for r, label, formula in resumo:
        ws.row_dimensions[r].height = 20
        bg = C["titulo_bg"] if r == 16 else C["resumo_bg"]
        fg = "FFFFFF" if r == 16 else C["resumo_fg"]
        bold = r == 16
        _c(ws, r, 1, label, bg, fg, bold=bold, h="left")
        v = ws.cell(r, 2, formula)
        v.fill = _f(bg); v.font = _ft(bold, fg, 10 if r==16 else 9)
        v.alignment = _al(); v.border = _bd(); v.number_format = FMT_MOEDA
        for ci in range(3, 12):
            ws.cell(r, ci).fill = _f(bg); ws.cell(r, ci).border = _bd()

    _nota(ws, 17, 1, 11,
          "  Nota: linhas marcadas como Nao na execucao permanecem com valor nominal (sem reajuste aplicado).",
          C["auto_bg"], C["auto_fg"])
    _nota(ws, 18, 1, 11,
          "  Nota: o Remanescente Atual da aba ITENS_CONTRATADOS deve corresponder a data de corte informada pelo fiscal.",
          C["auto_bg"], C["auto_fg"])


def _execucao(wb, c0, ciclos, mapa, n_mapa):
    ws = wb.create_sheet("EXECUCAO_FINANCEIRA")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C["fiscal_fg"]
    ws.freeze_panes = "A6"
    for col, w in [("A",16),("B",20),("C",14),("D",12),("E",14),
                   ("F",18),("G",16),("H",20)]:
        ws.column_dimensions[col].width = w
    _titulo(ws, 1, 1, 8, "MATRIZ 2.0 — EXECUCAO FINANCEIRA")
    _nota(ws, 2, 1, 8,
          "  Fiscal preenche APENAS: B (valor nominal consumido) e C (tem efeito financeiro?). "
          "Col A e Cols D:H sao totalmente automaticas.",
          C["fiscal_bg"], C["fiscal_fg"])
    _nota(ws, 3, 1, 8,
          "  Col A = competencia MM/AAAA (calculada pelo sistema). "
          "Col D = ciclo via PROCV. Col E = fator. Col F = base. Col G = delta. Col H = reajustado.",
          C["auto_bg"], C["auto_fg"])
    ws.row_dimensions[4].height = 8
    ws.row_dimensions[5].height = 52
    hdrs = [
        ("[AUTO]\nCompetencia\n(MM/AAAA)", C["hdr_bg"]),
        ("Valor Nominal\nConsumido (R$)",   C["fiscal_fg"]),
        ("Possui Efeito\nFinanceiro?",      C["fiscal_fg"]),
        ("[AUTO]\nCiclo",                   C["hdr_bg"]),
        ("[AUTO]\nFator\nAplicavel",       C["hdr_bg"]),
        ("[AUTO] Valor Base\nConsiderado (R$)", C["hdr_bg"]),
        ("[AUTO] Delta do\nReajuste (R$)",  C["hdr_bg"]),
        ("[AUTO] Valor\nReajustado (R$)",   C["hdr_bg"]),
    ]
    for ci, (txt, bg) in enumerate(hdrs, 1):
        _hdr(ws, 5, ci, txt, bg)
    lkp2 = f"MAPA_CICLOS!$A$2:$B${n_mapa+1}"
    lkp3 = f"MAPA_CICLOS!$A$2:$C${n_mapa+1}"
    for r in range(6, 306):
        ws.row_dimensions[r].height = 16
        idx = r - 6
        cv = mapa[idx]["comp"] if idx < len(mapa) else ""
        ws.cell(r,1,cv); ws.cell(r,1).fill=_f(C["auto_bg"]); ws.cell(r,1).font=_ft(color=C["auto_fg"],bold=bool(cv)); ws.cell(r,1).alignment=_al(); ws.cell(r,1).border=_bd()
        _c(ws,r,2,"",C["fiscal_bg"],C["fiscal_fg"],fmt=FMT_MOEDA)
        _c(ws,r,3,"Sim",C["fiscal_bg"],C["fiscal_fg"])
        ws.cell(r,4,f'=IFERROR(VLOOKUP(A{r},{lkp2},2,0),"")'); ws.cell(r,4).fill=_f(C["auto_bg"]); ws.cell(r,4).font=_ft(color=C["auto_fg"]); ws.cell(r,4).alignment=_al(); ws.cell(r,4).border=_bd()
        ws.cell(r,5,f'=IFERROR(VLOOKUP(A{r},{lkp3},3,0),1)'); ws.cell(r,5).fill=_f(C["auto_bg"]); ws.cell(r,5).font=_ft(color=C["auto_fg"]); ws.cell(r,5).alignment=_al(); ws.cell(r,5).border=_bd(); ws.cell(r,5).number_format=FMT_FATOR
        ws.cell(r,6,f'=IF(OR(B{r}="",B{r}=0),0,IF(C{r}="Sim",B{r},0))'); ws.cell(r,6).fill=_f(C["auto_bg"]); ws.cell(r,6).font=_ft(color=C["auto_fg"]); ws.cell(r,6).alignment=_al(); ws.cell(r,6).border=_bd(); ws.cell(r,6).number_format=FMT_MOEDA
        ws.cell(r,7,f'=IF(F{r}=0,0,ROUND(F{r}*E{r},2)-F{r})'); ws.cell(r,7).fill=_f(C["auto_bg"]); ws.cell(r,7).font=_ft(color=C["auto_fg"]); ws.cell(r,7).alignment=_al(); ws.cell(r,7).border=_bd(); ws.cell(r,7).number_format=FMT_MOEDA
        ws.cell(r,8,f'=IF(F{r}=0,0,ROUND(F{r}*E{r},2))'); ws.cell(r,8).fill=_f(C["auto_bg"]); ws.cell(r,8).font=_ft(color=C["auto_fg"]); ws.cell(r,8).alignment=_al(); ws.cell(r,8).border=_bd(); ws.cell(r,8).number_format=FMT_MOEDA
    _dv(ws,'"Sim,Nao"',"C6:C305")
    ws.row_dimensions[306].height=22
    ws.merge_cells("A306:C306")
    tc=ws.cell(306,1,"TOTAL — valor reajustado (Sim)"); tc.fill=_f(C["resumo_bg"]); tc.font=_ft(True,C["resumo_fg"],9); tc.alignment=_al("right"); tc.border=_bd()
    for ci in [2,3]: ws.cell(306,ci).fill=_f(C["resumo_bg"]); ws.cell(306,ci).border=_bd()
    for ci,f2,fmt in [(4,'=COUNTIF(C6:C305,"Sim")','#,##0'),(6,'=ROUND(SUM(F6:F305),2)',FMT_MOEDA),(7,'=ROUND(SUM(G6:G305),2)',FMT_MOEDA),(8,'=ROUND(SUM(H6:H305),2)',FMT_MOEDA)]:
        t=ws.cell(306,ci,f2); t.fill=_f(C["resumo_bg"]); t.font=_ft(True,C["resumo_fg"]); t.alignment=_al(); t.border=_bd(); t.number_format=fmt
    ws.cell(306,5).fill=_f(C["resumo_bg"]); ws.cell(306,5).border=_bd()


def _historico(wb, c0, ciclos):
    ws = wb.create_sheet("HISTORICO_CICLOS")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C["gcc_fg"]
    ws.freeze_panes = "A6"

    larguras = [("A",10),("B",22),("C",20),("D",18),("E",14),("F",18),
                ("G",18),("H",16),("I",22),("J",22),("K",22),("L",22),("M",28)]
    for col, w in larguras: ws.column_dimensions[col].width = w

    _titulo(ws, 1, 1, 13, "MATRIZ 2.0 — HISTORICO DE CICLOS")
    _nota(ws, 2, 1, 13,
          "  Aba gerencial C0 a C4. Registre quanto foi pago/informado em cada ciclo, "
          "inclusive ciclos sem reajuste. Campos C:F sao do fiscal/GCC.",
          C["gcc_bg"], C["gcc_fg"])
    _nota(ws, 3, 1, 13,
          "  Campos I:L sao automaticos: I/J consolidam EXECUCAO_FINANCEIRA por ciclo; "
          "K/L calculam remanescente a partir de ITENS_CONTRATADOS. Nao alterar.",
          C["auto_bg"], C["auto_fg"])
    ws.row_dimensions[4].height = 8

    # Cabecalho linha 5
    ws.row_dimensions[5].height = 52
    hdrs = [
        ("ID Ciclo",                                          C["hdr_bg"]),
        ("Periodo do Ciclo",                                  C["hdr_bg"]),
        ("Situacao do Ciclo",                                 C["gcc_fg"]),
        ("Houve Reajuste Concedido?",                         C["gcc_fg"]),
        ("% Concedido/Formalizado",                           C["gcc_fg"]),
        ("Valor Pago/Informado no Ciclo (R$)",                C["fiscal_fg"]),
        ("Valor Pago Atualizado/Reajustado (R$)",             C["hdr_bg"]),
        ("Delta do Ciclo (R$)",                               C["hdr_bg"]),
        ("Consumido Financeiro do Ciclo - Nominal (R$)",      C["hdr_bg"]),
        ("Consumido Financeiro do Ciclo - Reajustado (R$)",   C["hdr_bg"]),
        ("Remanescente ao Fim do Ciclo - Nominal (R$)",       C["hdr_bg"]),
        ("Remanescente ao Fim do Ciclo - Atualizado (R$)",    C["hdr_bg"]),
        ("Observacao",                                        C["gcc_fg"]),
    ]
    for ci, (txt, bg) in enumerate(hdrs, 1):
        _hdr(ws, 5, ci, txt, bg)

    todos = [c0] + ciclos
    situacoes_ciclo = '"Tempestivo,Precluso,Formalizado,Em analise,Pendente,Nao aplicavel"'

    for i, cd in enumerate(todos):
        r = 6 + i
        ws.row_dimensions[r].height = 20
        nome = cd["ciclo"]
        # A: ID auto
        _c(ws, r, 1, nome, C["auto_bg"], C["auto_fg"], bold=True)
        # B: periodo auto
        _c(ws, r, 2, cd.get("periodo",""), C["auto_bg"], C["auto_fg"])
        # C: situacao GCC dropdown
        _c(ws, r, 3, "", C["gcc_bg"], C["gcc_fg"])
        # D: houve reajuste GCC dropdown
        _c(ws, r, 4, "Nao" if cd["pct"] == 0 else "Sim", C["gcc_bg"], C["gcc_fg"])
        # E: pct concedido GCC
        _c(ws, r, 5, cd["pct"], C["gcc_bg"], C["gcc_fg"], fmt=FMT_PCT)
        # F: valor pago fiscal
        _c(ws, r, 6, "", C["fiscal_bg"], C["fiscal_fg"], fmt=FMT_MOEDA)
        # G: valor pago atualizado auto = F * fator
        ws.cell(r, 7, f'=IF(F{r}="",0,ROUND(F{r}*{cd["fat_acum"]},2))')
        ws.cell(r, 7).fill = _f(C["auto_bg"]); ws.cell(r, 7).font = _ft(color=C["auto_fg"])
        ws.cell(r, 7).alignment = _al(); ws.cell(r, 7).border = _bd(); ws.cell(r, 7).number_format = FMT_MOEDA
        # H: delta auto
        ws.cell(r, 8, f'=IF(F{r}="",0,ROUND(G{r}-F{r},2))')
        ws.cell(r, 8).fill = _f(C["auto_bg"]); ws.cell(r, 8).font = _ft(color=C["auto_fg"])
        ws.cell(r, 8).alignment = _al(); ws.cell(r, 8).border = _bd(); ws.cell(r, 8).number_format = FMT_MOEDA
        # I: consumido nominal auto (SOMASE por ciclo na execucao)
        ws.cell(r, 9, f'=IFERROR(ROUND(SUMIF(EXECUCAO_FINANCEIRA!D6:D305,"{nome}",EXECUCAO_FINANCEIRA!B6:B305),2),0)')
        ws.cell(r, 9).fill = _f(C["auto_bg"]); ws.cell(r, 9).font = _ft(color=C["auto_fg"])
        ws.cell(r, 9).alignment = _al(); ws.cell(r, 9).border = _bd(); ws.cell(r, 9).number_format = FMT_MOEDA
        # J: consumido reajustado auto
        ws.cell(r, 10, f'=IFERROR(ROUND(SUMIF(EXECUCAO_FINANCEIRA!D6:D305,"{nome}",EXECUCAO_FINANCEIRA!I6:I305),2),0)')
        ws.cell(r, 10).fill = _f(C["auto_bg"]); ws.cell(r, 10).font = _ft(color=C["auto_fg"])
        ws.cell(r, 10).alignment = _al(); ws.cell(r, 10).border = _bd(); ws.cell(r, 10).number_format = FMT_MOEDA
        # K: remanescente nominal ao fim do ciclo
        rem_col = ["E","F","G","H","I"][i] if i < 5 else "J"
        ws.cell(r, 11, f'=IFERROR(ROUND(SUMPRODUCT(ITENS_CONTRATADOS!{rem_col}6:ITENS_CONTRATADOS!{rem_col}105*ITENS_CONTRATADOS!C6:ITENS_CONTRATADOS!C105),2),0)')
        ws.cell(r, 11).fill = _f(C["auto_bg"]); ws.cell(r, 11).font = _ft(color=C["auto_fg"])
        ws.cell(r, 11).alignment = _al(); ws.cell(r, 11).border = _bd(); ws.cell(r, 11).number_format = FMT_MOEDA
        # L: remanescente atualizado
        ws.cell(r, 12, f'=IFERROR(ROUND(K{r}*{cd["fat_acum"]},2),0)')
        ws.cell(r, 12).fill = _f(C["auto_bg"]); ws.cell(r, 12).font = _ft(color=C["auto_fg"])
        ws.cell(r, 12).alignment = _al(); ws.cell(r, 12).border = _bd(); ws.cell(r, 12).number_format = FMT_MOEDA
        # M: observacao GCC
        _c(ws, r, 13, "", C["gcc_bg"], C["gcc_fg"], h="left")

    # Preencher linhas vazias ate C4
    for i in range(len(todos), 5):
        r = 6 + i
        nome = f"C{i}"
        ws.row_dimensions[r].height = 20
        _c(ws, r, 1, nome, C["auto_bg"], C["auto_fg"], bold=True)
        for ci in [2]: _c(ws, r, ci, "", C["auto_bg"], C["auto_fg"])
        for ci in [3, 4, 13]: _c(ws, r, ci, "", C["gcc_bg"], C["gcc_fg"])
        _c(ws, r, 5, 0, C["gcc_bg"], C["gcc_fg"], fmt=FMT_PCT)
        _c(ws, r, 6, "", C["fiscal_bg"], C["fiscal_fg"], fmt=FMT_MOEDA)
        for ci in [7, 8, 9, 10, 11, 12]:
            _c(ws, r, ci, 0, C["auto_bg"], C["auto_fg"], fmt=FMT_MOEDA)

    _dv(ws, situacoes_ciclo, "C6:C10")
    _dv(ws, '"Sim,Nao"', "D6:D10")


# ── ABA ITENS_CONTRATADOS ────────────────────────────────────────

def _itens(wb, c0, ciclos):
    ws = wb.create_sheet("ITENS_CONTRATADOS")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C["gcc_fg"]
    ws.freeze_panes = "A6"

    larguras = [("A",12),("B",12),("C",16),("D",16),
                ("E",14),("F",14),("G",14),("H",14),("I",14),
                ("J",16),("K",18),("L",14),("M",18),
                ("N",14),("O",14),("P",14),("Q",14),("R",14)]
    for col, w in larguras: ws.column_dimensions[col].width = w

    _titulo(ws, 1, 1, 18, "MATRIZ 2.0 — ITENS CONTRATADOS")
    _nota(ws, 2, 1, 18,
          "  Controle fisico-financeiro sem descricao e sem unidade. "
          "GCC preenche A:C (item, qtd, VU). Fiscal preenche E:J (remanescentes por ciclo).",
          C["gcc_bg"], C["gcc_fg"])
    _nota(ws, 3, 1, 18,
          "  Fiscal/GCC: trabalhe principalmente com A:M. "
          "Colunas N:R (consumos) sao automaticas e ficam ocultas para o fiscal.",
          C["auto_bg"], C["auto_fg"])

    todos = [c0] + ciclos
    fat_corte = ciclos[-1]["fat_acum"] if ciclos else 1.0
    ciclo_corte = ciclos[-1]["ciclo"] if ciclos else "C0"

    # Linha 4 — grupo e corte
    ws.row_dimensions[4].height = 16
    for ci in range(1, 19): ws.cell(4, ci).fill = _f(C["auto_bg"]); ws.cell(4, ci).border = _bd()
    ws.merge_cells("J4:J4")
    ws.cell(4, 10, f"Corte atual: {ciclo_corte}").fill = _f(C["gcc_bg"])
    ws.cell(4, 10).font = _ft(True, C["gcc_fg"], 8)
    ws.cell(4, 10).alignment = _al(); ws.cell(4, 10).border = _bd()

    # Cabecalho linha 5
    ws.row_dimensions[5].height = 52
    hdrs5 = [
        ("ID Item",                            C["gcc_fg"]),
        ("Qtd C0",                             C["gcc_fg"]),
        ("VU C0 (R$)",                         C["gcc_fg"]),
        ("VT C0 (R$)",                         C["hdr_bg"]),
        ("Remanescente Inicio C0 (Qtd)",       C["fiscal_fg"]),
        ("Remanescente Inicio C1 (Qtd)",       C["fiscal_fg"]),
        ("Remanescente Inicio C2 (Qtd)",       C["fiscal_fg"]),
        ("Remanescente Inicio C3 (Qtd)",       C["fiscal_fg"]),
        ("Remanescente Inicio C4 (Qtd)",       C["fiscal_fg"]),
        ("Remanescente Atual (Qtd)",           C["fiscal_fg"]),
        ("VT Remanescente Atual Nominal (R$)", C["hdr_bg"]),
        ("Fator Aplicavel",                    C["gcc_fg"]),
        ("VT Remanescente Atualizado (R$)",    C["hdr_bg"]),
        ("Consumo Qtd C0",                     C["hdr_bg"]),
        ("Consumo Qtd C1",                     C["hdr_bg"]),
        ("Consumo Qtd C2",                     C["hdr_bg"]),
        ("Consumo Qtd C3",                     C["hdr_bg"]),
        ("Consumo Qtd C4/Atual",               C["hdr_bg"]),
    ]
    for ci, (txt, bg) in enumerate(hdrs5, 1):
        _hdr(ws, 5, ci, txt, bg)

    # Dados linhas 6-105
    for r in range(6, 106):
        ws.row_dimensions[r].height = 16
        _c(ws, r, 1, "", C["gcc_bg"], C["gcc_fg"])              # A item
        _c(ws, r, 2, 0,  C["gcc_bg"], C["gcc_fg"], fmt=FMT_QTD) # B qtd c0
        _c(ws, r, 3, 0,  C["gcc_bg"], C["gcc_fg"], fmt=FMT_MOEDA) # C VU
        # D VT C0 auto
        ws.cell(r, 4, f"=IF(OR(B{r}=0,C{r}=0),0,ROUND(B{r}*C{r},2))")
        ws.cell(r, 4).fill = _f(C["auto_bg"]); ws.cell(r, 4).font = _ft(color=C["auto_fg"])
        ws.cell(r, 4).alignment = _al(); ws.cell(r, 4).border = _bd(); ws.cell(r, 4).number_format = FMT_MOEDA
        # E-J remanescentes fiscal
        for ci in range(5, 11):
            _c(ws, r, ci, 0, C["fiscal_bg"], C["fiscal_fg"], fmt=FMT_QTD)
        # K VT rem atual nominal auto = J * C
        ws.cell(r, 11, f"=IF(OR(J{r}=0,C{r}=0),0,ROUND(J{r}*C{r},2))")
        ws.cell(r, 11).fill = _f(C["auto_bg"]); ws.cell(r, 11).font = _ft(color=C["auto_fg"])
        ws.cell(r, 11).alignment = _al(); ws.cell(r, 11).border = _bd(); ws.cell(r, 11).number_format = FMT_MOEDA
        # L fator GCC
        _c(ws, r, 12, fat_corte, C["gcc_bg"], C["gcc_fg"], fmt=FMT_FATOR)
        # M VT rem atualizado auto = K * L
        ws.cell(r, 13, f"=IF(K{r}=0,0,ROUND(K{r}*L{r},2))")
        ws.cell(r, 13).fill = _f(C["auto_bg"]); ws.cell(r, 13).font = _ft(color=C["auto_fg"])
        ws.cell(r, 13).alignment = _al(); ws.cell(r, 13).border = _bd(); ws.cell(r, 13).number_format = FMT_MOEDA
        # N-R consumos auto (diferenca entre remanescentes sucessivos)
        consumos = [
            f"=IF(E{r}=0,0,E{r}-F{r})",  # N: C0 = remC0 - remC1
            f"=IF(F{r}=0,0,F{r}-G{r})",  # O: C1 = remC1 - remC2
            f"=IF(G{r}=0,0,G{r}-H{r})",  # P: C2 = remC2 - remC3
            f"=IF(H{r}=0,0,H{r}-I{r})",  # Q: C3 = remC3 - remC4
            f"=IF(I{r}=0,0,I{r}-J{r})",  # R: C4 = remC4 - remAtual
        ]
        for ci, formula in enumerate(consumos, 14):
            ws.cell(r, ci, formula)
            ws.cell(r, ci).fill = _f(C["auto_bg"]); ws.cell(r, ci).font = _ft(color=C["auto_fg"])
            ws.cell(r, ci).alignment = _al(); ws.cell(r, ci).border = _bd(); ws.cell(r, ci).number_format = FMT_QTD

    # Linha de total 106
    ws.row_dimensions[106].height = 22
    _c(ws, 106, 1, "TOTAL", C["resumo_bg"], C["resumo_fg"], bold=True, h="right")
    for ci in [2, 5, 6, 7, 8, 9, 10]:
        t = ws.cell(106, ci, f"=ROUND(SUM({get_column_letter(ci)}6:{get_column_letter(ci)}105),2)")
        t.fill = _f(C["resumo_bg"]); t.font = _ft(True, C["resumo_fg"])
        t.alignment = _al(); t.border = _bd(); t.number_format = FMT_QTD
    for ci in [3, 4, 11, 13]:
        t = ws.cell(106, ci, f"=ROUND(SUM({get_column_letter(ci)}6:{get_column_letter(ci)}105),2)")
        t.fill = _f(C["resumo_bg"]); t.font = _ft(True, C["resumo_fg"])
        t.alignment = _al(); t.border = _bd(); t.number_format = FMT_MOEDA
    for ci in [12, 14, 15, 16, 17, 18]:
        ws.cell(106, ci).fill = _f(C["resumo_bg"]); ws.cell(106, ci).border = _bd()

    # Ocultar colunas N:R (consumos — para GCC apenas)
    for col in ["N", "O", "P", "Q", "R"]:
        ws.column_dimensions[col].hidden = True


# ── ABA ADITIVOS ─────────────────────────────────────────────────

def _aditivos(wb, c0, ciclos):
    from openpyxl.formatting.rule import CellIsRule, FormulaRule
    ws = wb.create_sheet("ADITIVOS")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C["alrt_fg"]
    ws.freeze_panes = "A6"

    larguras = [("A",12),("B",14),("C",12),("D",18),("E",16),
                ("F",18),("G",18),("H",16),("I",14),("J",18),("K",22),("L",20)]
    for col, w in larguras: ws.column_dimensions[col].width = w

    _titulo(ws, 1, 1, 12, "MATRIZ 2.0 — ADITIVOS")
    _nota(ws, 2, 1, 12,
          "  GCC preenche: A (item), B (data), D (tipo), E (qtd), F (VU original), H (aplicar reajuste?), K (tratamento). "
          "Col C (ciclo) e G (valor) sao automaticos. Supressoes ficam vermelhas e valor L fica negativo.",
          C["alrt_bg"], C["alrt_fg"])
    ws.row_dimensions[3].height = 8; ws.row_dimensions[4].height = 8

    # Cabecalho linha 5
    ws.row_dimensions[5].height = 52
    hdrs_ad = [
        ("Item",                               C["gcc_fg"]),
        ("Data do Aditivo",                    C["gcc_fg"]),
        ("[AUTO] Ciclo / Marco",               C["hdr_bg"]),
        ("Tipo de Alteracao",                  C["gcc_fg"]),
        ("Qtd Acrescida / Suprimida",          C["gcc_fg"]),
        ("Valor Unitario Original (R$)",       C["gcc_fg"]),
        ("[AUTO] Valor Original da Alteracao", C["hdr_bg"]),
        ("Aplicar Reajuste Acumulado?",        C["gcc_fg"]),
        ("[AUTO] Fator Acumulado Aplicavel",   C["hdr_bg"]),
        ("[AUTO] Valor Atualizado (R$)",       C["hdr_bg"]),
        ("Tratamento do Aditivo",              C["gcc_fg"]),
        ("[AUTO] Valor Computavel (R$)",       C["hdr_bg"]),
    ]
    for ci, (txt, bg) in enumerate(hdrs_ad, 1):
        _hdr(ws, 5, ci, txt, bg)

    # Dropdowns
    _dv(ws, '"Acrescimo,Supressao"', "D6:D205")
    _dv(ws, '"Sim,Nao"', "H6:H205")
    _dv(ws, '"Computar no VTA,Ja refletido no financeiro/itens,Apenas informativo,Nao sei"', "K6:K205")

    # Montar IFS de ciclo por data (col B -> col C)
    todos = [c0] + ciclos
    # Formula ciclo: IFS(AND(B>=ini,B<=fim),"C0", AND(B>=ini1,B<=fim1),"C1",...)
    partes_ciclo = []
    for cd in todos:
        ini = cd.get("ini_fin","")
        fim = cd.get("fim_fin","")
        if ini and fim and ini not in ("","nan") and fim not in ("","nan"):
            # Usar DATE() com os valores numericos extraidos
            try:
                ini_parts = str(ini).split(" ")[0].split("-")
                fim_parts = str(fim).split(" ")[0].split("-")
                if len(ini_parts) == 3 and len(fim_parts) == 3:
                    ini_date = f"DATE({ini_parts[0]},{ini_parts[1]},{ini_parts[2]})"
                    fim_date = f"DATE({fim_parts[0]},{fim_parts[1]},{fim_parts[2]})"
                    partes_ciclo.append((cd["ciclo"], ini_date, fim_date))
            except Exception:
                pass

    # Cores alternadas para col C por ciclo
    cores_ciclo = {
        "C0": C["auto_bg"],
        "C1": "E8F4FD",
        "C2": "FEF9E7",
        "C3": "F0FFF0",
        "C4": "FDF0FF",
    }

    # Fator IFS por ciclo para col I
    fat_ifs_parts = ",".join([f'C{{r}}="{cd["ciclo"]}",{cd["fat_acum"]}' for cd in todos])

    for r in range(6, 206):
        ws.row_dimensions[r].height = 16

        # A: item GCC
        _c(ws, r, 1, "", C["gcc_bg"], C["gcc_fg"])
        # B: data GCC
        _c(ws, r, 2, "", C["gcc_bg"], C["gcc_fg"])
        ws.cell(r, 2).number_format = "DD/MM/YYYY"

        # C: ciclo AUTO por data
        if partes_ciclo:
            ifs_str = ",".join([f'AND(B{r}>={ini},B{r}<={fim}),"{nome}"'
                                for nome, ini, fim in partes_ciclo])
            formula_ciclo = f'=IF(B{r}="","",IFERROR(IFS({ifs_str}),"?"))'
        else:
            formula_ciclo = f'=IF(B{r}="","","")'
        ws.cell(r, 3, formula_ciclo)
        ws.cell(r, 3).fill = _f(C["auto_bg"])
        ws.cell(r, 3).font = _ft(True, C["auto_fg"])
        ws.cell(r, 3).alignment = _al(); ws.cell(r, 3).border = _bd()

        # D: tipo GCC
        _c(ws, r, 4, "Acrescimo", C["gcc_bg"], C["gcc_fg"])
        # E: qtd GCC
        _c(ws, r, 5, "", C["gcc_bg"], C["gcc_fg"], fmt=FMT_QTD)
        # F: VU original GCC
        _c(ws, r, 6, "", C["gcc_bg"], C["gcc_fg"], fmt=FMT_MOEDA)

        # G: valor original AUTO = ABS(E) * F
        ws.cell(r, 7, f"=IF(OR(E{r}=0,F{r}=0),0,ROUND(ABS(E{r})*F{r},2))")
        ws.cell(r, 7).fill = _f(C["auto_bg"]); ws.cell(r, 7).font = _ft(color=C["auto_fg"])
        ws.cell(r, 7).alignment = _al(); ws.cell(r, 7).border = _bd(); ws.cell(r, 7).number_format = FMT_MOEDA

        # H: aplicar reajuste GCC
        _c(ws, r, 8, "Nao", C["gcc_bg"], C["gcc_fg"])

        # I: fator AUTO por ciclo
        fat_formula = '=IFERROR(IFS(' + ','.join([f'C{r}="{cd["ciclo"]}",{cd["fat_acum"]}' for cd in todos]) + f'),1)'
        ws.cell(r, 9, fat_formula)
        ws.cell(r, 9).fill = _f(C["auto_bg"]); ws.cell(r, 9).font = _ft(color=C["auto_fg"])
        ws.cell(r, 9).alignment = _al(); ws.cell(r, 9).border = _bd(); ws.cell(r, 9).number_format = FMT_FATOR

        # J: valor atualizado AUTO
        ws.cell(r, 10, f'=IF(G{r}=0,0,IF(H{r}="Sim",ROUND(G{r}*I{r},2),G{r}))')
        ws.cell(r, 10).fill = _f(C["auto_bg"]); ws.cell(r, 10).font = _ft(color=C["auto_fg"])
        ws.cell(r, 10).alignment = _al(); ws.cell(r, 10).border = _bd(); ws.cell(r, 10).number_format = FMT_MOEDA

        # K: tratamento GCC
        _c(ws, r, 11, "Nao sei", C["gcc_bg"], C["gcc_fg"])

        # L: valor computavel AUTO — negativo se supressao
        ws.cell(r, 12, f'=IF(J{r}=0,0,IF(D{r}="Supressao",-ABS(J{r}),J{r}))')
        ws.cell(r, 12).fill = _f(C["auto_bg"]); ws.cell(r, 12).font = _ft(color=C["auto_fg"])
        ws.cell(r, 12).alignment = _al(); ws.cell(r, 12).border = _bd(); ws.cell(r, 12).number_format = FMT_MOEDA

    # Formatacao condicional: linha inteira vermelha se col D = "Supressao"
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import Font as OFont, PatternFill as OFill
    red_fill = OFill("solid", fgColor=C["sup_bg"])
    red_font = OFont(color=C["sup_fg"], bold=True, name="Arial", size=9)
    ws.conditional_formatting.add(
        f"A6:L205",
        FormulaRule(formula=[f'$D6="Supressao"'], fill=red_fill, font=red_font)
    )

    # Formatacao condicional col C: cor alternada por ciclo
    cores_regras = [
        ("C0", C["auto_bg"]),
        ("C1", "E8F4FD"),
        ("C2", "FEF9E7"),
        ("C3", "F0FFF0"),
        ("C4", "FDF0FF"),
    ]
    for nome_c, cor in cores_regras:
        fill_c = OFill("solid", fgColor=cor)
        ws.conditional_formatting.add(
            f"C6:C205",
            FormulaRule(formula=[f'$C6="{nome_c}"'], fill=fill_c)
        )

    # Total linha 206
    ws.row_dimensions[206].height = 22
    ws.merge_cells("A206:K206")
    tc = ws.cell(206, 1, "TOTAL COMPUTAVEL NO VTA (Supressoes = negativo)")
    tc.fill = _f(C["resumo_bg"]); tc.font = _ft(True, C["resumo_fg"], 9)
    tc.alignment = _al("right"); tc.border = _bd()
    for ci in range(2, 12): ws.cell(206, ci).fill = _f(C["resumo_bg"]); ws.cell(206, ci).border = _bd()
    tv = ws.cell(206, 12, '=ROUND(SUMIF(K6:K205,"Computar no VTA",L6:L205),2)')
    tv.fill = _f(C["resumo_bg"]); tv.font = _ft(True, C["resumo_fg"], 10)
    tv.alignment = _al(); tv.border = _bd(); tv.number_format = FMT_MOEDA



# ── FUNCAO PUBLICA ───────────────────────────────────────────────

def _gerar_matriz_2_0_sem_memoria_vta():
    """Gera a Matriz 2.0 e retorna bytes XLSX."""
    c0, ciclos, ciclo_corte, adm = _sessao()

    wb = Workbook()
    wb.remove(wb.active)

    mapa   = _mapear_competencias(ciclos)
    n_mapa = _construir_mapa_ciclos(wb, mapa)
    _base(wb, c0, ciclos, ciclo_corte)
    _execucao(wb, c0, ciclos, mapa, n_mapa)
    _historico(wb, c0, ciclos)
    _itens(wb, c0, ciclos)
    _aditivos(wb, c0, ciclos)

    try:
        wb.active = wb.index(wb["BASE"])
    except Exception:
        pass

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()

# >>> WRAPPER_MEMORIA_VTA_M20
# Regra: o XLS baixado pelo botão deve nascer com MEMORIA_VTA e BASE!B16 alinhado.
def gerar_matriz_2_0(*args, **kwargs):
    from _memoria_vta_m20 import aplicar_memoria_vta_xlsx
    _bytes = _gerar_matriz_2_0_sem_memoria_vta(*args, **kwargs)
    return aplicar_memoria_vta_xlsx(_bytes)
# <<< WRAPPER_MEMORIA_VTA_M20
