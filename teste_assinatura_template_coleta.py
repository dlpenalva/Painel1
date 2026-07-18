# -*- coding: utf-8 -*-
"""Assinatura SHA-256 do template oficial como chave de cache do download.

Cobre: assinaturas distintas para conteudo distinto (mesmo nome/versao e
mesmo tamanho), assinatura estavel para template inalterado (cache
reutilizado), alteracao do template produzindo novo download (chave nova
regenera os bytes) e itens_PC integra no arquivo entregue pelo fluxo
cacheado. A semantica do st.cache_data (mesma chave = reuso; chave nova =
reexecucao) e emulada com um dicionario chaveado por (versao, assinatura),
que e exatamente o contrato usado em app._obter_masterfile_entrada_enxuto.
"""
from __future__ import annotations

import io
import shutil
import sys
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import load_workbook

import _coleta_oficial
from _coleta_oficial import (
    COLUNAS_ITENS_PC_OFICIAL,
    TEMPLATE_COLETA_OFICIAL,
    assinatura_template_coleta,
    obter_coleta_oficial_bytes,
)

RESULTADOS = []


def check(nome: str, cond: bool, detalhe: str = "") -> None:
    RESULTADOS.append((nome, bool(cond), detalhe))
    status = "OK " if cond else "FAIL"
    print(f"[{status}] {nome}" + (f" - {detalhe}" if detalhe and not cond else ""))


def main() -> int:
    with TemporaryDirectory() as tmp:
        base = Path(tmp)

        # mesmo nome e mesma versao da aplicacao, conteudo diferente
        a = base / "TPL" / "COLETA_REAJUSTE_OFICIAL.xlsx"
        b = base / "TPL2" / "COLETA_REAJUSTE_OFICIAL.xlsx"
        a.parent.mkdir()
        b.parent.mkdir()
        shutil.copy2(TEMPLATE_COLETA_OFICIAL, a)
        conteudo = bytearray(a.read_bytes())
        conteudo[len(conteudo) // 2] ^= 0xFF  # muda 1 byte, tamanho identico
        b.write_bytes(bytes(conteudo))
        check("mesmo nome/versao com conteudo diferente: assinaturas distintas",
              assinatura_template_coleta(a) != assinatura_template_coleta(b))
        check("mesmo tamanho com conteudo diferente: assinaturas distintas",
              a.stat().st_size == b.stat().st_size
              and assinatura_template_coleta(a) != assinatura_template_coleta(b))

        # template inalterado: assinatura estavel -> mesma chave de cache
        s1 = assinatura_template_coleta(a)
        s2 = assinatura_template_coleta(a)
        check("template inalterado: assinatura estavel (cache reutilizado)",
              s1 == s2 and len(s1) == 64)

        # alteracao do template produz novo download (chave nova reexecuta)
        copia = base / "templates_ativos" / "COLETA_REAJUSTE_OFICIAL.xlsx"
        copia.parent.mkdir()
        shutil.copy2(TEMPLATE_COLETA_OFICIAL, copia)
        original = _coleta_oficial.TEMPLATE_COLETA_OFICIAL
        try:
            _coleta_oficial.TEMPLATE_COLETA_OFICIAL = copia
            cache: dict[tuple[str, str], bytes] = {}
            geracoes = 0

            def download() -> bytes:
                nonlocal geracoes
                chave = ("v-teste", assinatura_template_coleta(copia))
                if chave not in cache:
                    geracoes += 1
                    cache[chave] = obter_coleta_oficial_bytes()
                return cache[chave]

            d1 = download()
            d2 = download()
            check("template inalterado: download reutiliza o cache",
                  geracoes == 1 and d1 is d2)

            # alteracao real de conteudo (regravar o XLSX muda os bytes,
            # preservando a estrutura homologada para o guard de itens_PC)
            s_antes = assinatura_template_coleta(copia)
            wb = load_workbook(copia)
            wb.properties.description = "assinatura-teste"
            wb.save(copia)
            check("alteracao do template: assinatura muda",
                  assinatura_template_coleta(copia) != s_antes)
            d3 = download()
            check("alteracao do template: novo download gerado",
                  geracoes == 2 and d3 is not d1)

            # itens_PC integra no arquivo entregue pelo fluxo cacheado
            ws = load_workbook(io.BytesIO(d3))["itens_PC"]
            cab = [ws.cell(1, c).value for c in range(1, 12)]
            formulas = sum(
                1
                for row in ws.iter_rows(min_row=2, max_row=100, min_col=3,
                                        max_col=11)
                for cell in row
                if isinstance(cell.value, str) and cell.value.startswith("=")
            )
            check("itens_PC integra no arquivo entregue (cabecalhos A1:K1)",
                  cab == COLUNAS_ITENS_PC_OFICIAL, str(cab))
            check("itens_PC integra no arquivo entregue (formulas C2:K100)",
                  formulas >= 600, str(formulas))
        finally:
            _coleta_oficial.TEMPLATE_COLETA_OFICIAL = original

    falhas = [n for n, ok, _ in RESULTADOS if not ok]
    print(f"\n{len(RESULTADOS)} verificacoes; falhas: {len(falhas)}")
    return 1 if falhas else 0


if __name__ == "__main__":
    sys.exit(main())
